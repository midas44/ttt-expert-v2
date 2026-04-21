#!/usr/bin/env python3
"""
Process a Curated Test Collection — resolve references, inject tags, report missing specs.

A collection is an XLSX workbook at test-docs/collections/<name>/<name>.xlsx with a
COL-<name> sheet that references test cases from other modules. Processing:
  1. Parses the collection XLSX
  2. Resolves each reference against the main manifest (test-cases.json)
  3. For specs that exist: injects @col-<name> tag into the test title
  4. For specs that don't exist: records them as needing generation
  5. Outputs a report JSON at autotests/manifest/collection-<name>.json

Usage:
    python3 autotests/scripts/process_collection.py --collection <name> [--dry-run]
    python3 autotests/scripts/process_collection.py --collection absences
    python3 autotests/scripts/process_collection.py --collection absences --dry-run
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Module name → spec filename prefix mapping
MODULE_SPEC_PREFIX = {
    "vacation": "vacation",
    "day-off": "dayoff",
    "cross-service": "cross-service",
    "planner": "planner",
    "reports": "reports",
    "sick-leave": "sick-leave",
    "accounting": "accounting",
    "statistics": "statistics",
    "admin": "admin",
    "security": "security",
}

# test_id pattern: TC-XXX-NNN → extract the numeric part
TC_NUM_RE = re.compile(r"TC-[A-Z0-9]+-(\d+)")


def parse_collection_xlsx(xlsx_path: Path) -> list[dict]:
    """Parse a collection XLSX and return list of reference entries.

    Locates the header row dynamically by scanning for a row whose column A
    contains the literal "test_id". This keeps the parser stable across:
      - absences-style layout (6 cols, headers row 1, data row 2+)
      - cron-style layout    (7 cols, metadata rows 1-3, headers row 4, data row 5+)
      - digest-style layout  (7 cols, metadata row 1, headers row 2, data row 3+)
    Data columns are looked up by header name, so the optional `source_workbook`
    column slots in without breaking older 6-column collections.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    col_sheet = None
    for name in wb.sheetnames:
        if name.startswith("COL-"):
            col_sheet = wb[name]
            break

    if col_sheet is None:
        print(f"ERROR: No COL-* sheet found in {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Locate header row
    header_row_idx = None
    header_map: dict[str, int] = {}
    for row_idx, row in enumerate(col_sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue
        first = str(row[0] or "").strip().lower()
        if first == "test_id":
            header_row_idx = row_idx
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                header_map[str(cell).strip().lower()] = col_idx
            break

    if header_row_idx is None:
        print(
            f"ERROR: No header row (row starting with 'test_id') found in COL-* of {xlsx_path}",
            file=sys.stderr,
        )
        sys.exit(1)

    def lookup(row: tuple, key: str, default: str = "") -> str:
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        return str(val).strip() if val is not None else default

    entries = []
    for row_idx, row in enumerate(
        col_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row:
            continue
        test_id = lookup(row, "test_id")
        if not test_id.startswith("TC-"):
            continue
        entries.append({
            "test_id": test_id,
            "source_module": lookup(row, "source_module"),
            "source_workbook": lookup(row, "source_workbook"),
            "source_suite": lookup(row, "source_suite"),
            "title": lookup(row, "title"),
            "inclusion_reason": lookup(row, "inclusion_reason"),
            "priority_override": lookup(row, "priority_override"),
        })

    wb.close()
    return entries


def load_manifest(manifest_path: Path) -> dict:
    """Load the test-cases.json manifest."""
    if not manifest_path.exists():
        print(f"ERROR: Manifest not found: {manifest_path}", file=sys.stderr)
        print("Run: python3 autotests/scripts/parse_xlsx.py", file=sys.stderr)
        sys.exit(1)
    with open(manifest_path) as f:
        return json.load(f)


def find_spec_path(
    module: str, test_id: str, collection: str | None = None
) -> Path | None:
    """Find the spec file path for a test case. Returns None if not found.

    If `collection` is given, search `tests/<collection>/` first with prefix
    derived from the collection name — so collection-scoped TCs land in a
    dedicated directory (e.g., TC-DIGEST-001 → tests/digest/digest-tc001.spec.ts)
    rather than polluting the source module's test directory.

    Falls back to the module-scoped location (tests/<module>/) if the
    collection path doesn't resolve — this keeps per-module TCs referenced by
    a collection (like absences referencing cross-service TCs) working as
    before.
    """
    match = TC_NUM_RE.search(test_id)
    if not match:
        return None
    num = match.group(1)

    candidates: list[Path] = []

    if collection:
        coll_prefix = collection
        coll_dir = PROJECT_ROOT / "autotests" / "e2e" / "tests" / collection
        candidates += [
            coll_dir / f"{coll_prefix}-tc{num}.spec.ts",
            coll_dir / f"{coll_prefix}-tc{num.lstrip('0') or '0'}.spec.ts",
            coll_dir / f"{coll_prefix}-tc{num.zfill(3)}.spec.ts",
        ]

    prefix = MODULE_SPEC_PREFIX.get(module)
    if prefix:
        tests_dir = PROJECT_ROOT / "autotests" / "e2e" / "tests" / module
        candidates += [
            tests_dir / f"{prefix}-tc{num}.spec.ts",
            tests_dir / f"{prefix}-tc{num.lstrip('0') or '0'}.spec.ts",
            tests_dir / f"{prefix}-tc{num.zfill(3)}.spec.ts",
        ]

    # A spec file's numeric suffix alone doesn't guarantee it's the right TC —
    # different test_id families can share the numeric part (TC-VAC-001 and
    # TC-DIGEST-001 both extract num=001). Verify the spec actually references
    # the requested test_id before returning a match.
    for candidate in candidates:
        if not candidate.exists():
            continue
        try:
            content = candidate.read_text(encoding="utf-8")
        except OSError:
            continue
        if test_id in content:
            return candidate
    return None


def add_tag_to_spec(spec_path: Path, tag: str, dry_run: bool = False) -> bool:
    """Add a collection tag to the test title in a spec file. Returns True if modified."""
    content = spec_path.read_text(encoding="utf-8")

    if tag in content:
        return False  # Already has the tag

    # Match the test() title line: test("TC-XXX-NNN: ... @tag1 @tag2", async
    # Insert the new tag before the closing quote
    pattern = re.compile(r'(test\("TC-[A-Z0-9]+-\d+:.*?@\w[\w-]*)(")')
    match = pattern.search(content)
    if not match:
        return False  # Could not find the test title pattern

    new_content = content[: match.end(1)] + " " + tag + content[match.start(2) :]

    if not dry_run:
        spec_path.write_text(new_content, encoding="utf-8")
    return True


def resolve_in_manifest(
    manifest: dict, test_id: str, module: str
) -> tuple[dict | None, str | None]:
    """Find a test case in the manifest by test_id.

    Returns (tc_dict, resolved_module_key). The module key is the manifest key
    where the TC was actually found — may differ from the requested `module`
    when the TC was registered under a collection pseudo-module (e.g. COL-digest
    references source_module='vacation' but the TC lives under manifest module
    'digest' because parse_xlsx.py scans test-docs/collections/digest/ under
    the collection name).

    Resolution order:
      1. Look in the requested module (preserves existing behavior for
         per-module TCs referenced by collections like absences).
      2. Fall back to any module with source_type=collection that contains
         the test_id (the common case for collection-scoped TCs).
      3. Fall back to any module at all — catches mismatches like TC-DIGEST-*
         stored under a collection pseudo-module when source_module in COL
         still names a canonical module.
    """
    modules = manifest.get("modules", {})

    mod = modules.get(module, {})
    for suite in mod.get("suites", {}).values():
        for tc in suite.get("test_cases", []):
            if tc.get("test_id") == test_id:
                return tc, module

    for mod_name, mod_data in modules.items():
        if mod_data.get("source_type") != "collection":
            continue
        for suite in mod_data.get("suites", {}).values():
            for tc in suite.get("test_cases", []):
                if tc.get("test_id") == test_id:
                    return tc, mod_name

    for mod_name, mod_data in modules.items():
        for suite in mod_data.get("suites", {}).values():
            for tc in suite.get("test_cases", []):
                if tc.get("test_id") == test_id:
                    return tc, mod_name

    return None, None


def process_collection(collection_name: str, dry_run: bool = False) -> dict:
    """Process a collection: parse XLSX, inject tags, produce report."""
    xlsx_path = PROJECT_ROOT / "test-docs" / "collections" / collection_name / f"{collection_name}.xlsx"
    if not xlsx_path.exists():
        print(f"ERROR: Collection XLSX not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    manifest_path = PROJECT_ROOT / "autotests" / "manifest" / "test-cases.json"
    manifest = load_manifest(manifest_path)

    entries = parse_collection_xlsx(xlsx_path)
    tag = f"@col-{collection_name}"

    results = []
    tagged_count = 0
    already_tagged = 0
    needs_generation = 0
    not_in_manifest = 0

    for entry in entries:
        test_id = entry["test_id"]
        module = entry["source_module"]

        # Resolve in manifest (falls back to collection pseudo-modules)
        manifest_tc, resolved_module = resolve_in_manifest(manifest, test_id, module)
        if manifest_tc is None:
            results.append({
                "test_id": test_id,
                "module": module,
                "spec_file": None,
                "action": "not_in_manifest",
                "detail": f"Test case {test_id} not found in manifest for module {module}",
            })
            not_in_manifest += 1
            continue

        # Check if spec exists. Pass collection_name so collection-scoped specs
        # land in tests/<collection>/, independent of source_module.
        spec_path = find_spec_path(module, test_id, collection=collection_name)

        if spec_path is not None:
            modified = add_tag_to_spec(spec_path, tag, dry_run=dry_run)
            if modified:
                action = "tag_added" if not dry_run else "tag_would_add"
                tagged_count += 1
            else:
                action = "tag_already_present"
                already_tagged += 1

            results.append({
                "test_id": test_id,
                "module": module,
                "resolved_module": resolved_module,
                "spec_file": str(spec_path.relative_to(PROJECT_ROOT)),
                "action": action,
            })
        else:
            # Default target for collection-scoped generation:
            # tests/<collection>/<collection>-tc<num>.spec.ts.
            num_match = TC_NUM_RE.search(test_id)
            target_hint = None
            if num_match:
                num = num_match.group(1)
                target_hint = (
                    f"autotests/e2e/tests/{collection_name}/"
                    f"{collection_name}-tc{num}.spec.ts"
                )
            results.append({
                "test_id": test_id,
                "module": module,
                "resolved_module": resolved_module,
                "spec_file": None,
                "target_spec_path": target_hint,
                "action": "needs_generation",
                "manifest_title": manifest_tc.get("title", ""),
                "manifest_type": manifest_tc.get("classified_type", manifest_tc.get("type", "")),
                "manifest_priority": entry.get("priority_override") or manifest_tc.get("priority", ""),
                "source_suite": entry.get("source_suite", ""),
                "source_workbook": entry.get("source_workbook", ""),
            })
            needs_generation += 1

    report = {
        "collection_name": collection_name,
        "tag": tag,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "dry_run": dry_run,
        "total_cases": len(entries),
        "summary": {
            "tags_added": tagged_count,
            "tags_already_present": already_tagged,
            "needs_generation": needs_generation,
            "not_in_manifest": not_in_manifest,
        },
        "cases": results,
    }

    return report


def main():
    parser = argparse.ArgumentParser(description="Process a Curated Test Collection")
    parser.add_argument("--collection", required=True, help="Collection name (e.g., absences)")
    parser.add_argument("--dry-run", action="store_true", help="Report without modifying files")
    args = parser.parse_args()

    print(f"Processing collection: {args.collection}" + (" [DRY RUN]" if args.dry_run else ""))

    report = process_collection(args.collection, dry_run=args.dry_run)

    # Save report
    report_path = PROJECT_ROOT / "autotests" / "manifest" / f"collection-{args.collection}.json"
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2)

    # Print summary
    s = report["summary"]
    print(f"\n{'=' * 60}")
    print(f"Collection: {report['collection_name']}")
    print(f"Tag:        {report['tag']}")
    print(f"Total:      {report['total_cases']} test cases")
    print(f"{'=' * 60}")
    print(f"  Tags added:           {s['tags_added']}")
    print(f"  Already tagged:       {s['tags_already_present']}")
    print(f"  Needs generation:     {s['needs_generation']}")
    print(f"  Not in manifest:      {s['not_in_manifest']}")
    print(f"{'=' * 60}")
    print(f"Report saved: {report_path}")

    if s["needs_generation"] > 0:
        print(f"\nTest cases needing generation:")
        for case in report["cases"]:
            if case["action"] == "needs_generation":
                print(f"  {case['test_id']} ({case['module']}) — {case.get('manifest_title', '')[:60]}")

    if s["not_in_manifest"] > 0:
        print(f"\nWARNING: Test cases not found in manifest:")
        for case in report["cases"]:
            if case["action"] == "not_in_manifest":
                print(f"  {case['test_id']} ({case['module']}) — {case.get('detail', '')}")


if __name__ == "__main__":
    main()
