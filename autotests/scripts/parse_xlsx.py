#!/usr/bin/env python3
"""
Parse XLSX test documentation from test-docs/ into a JSON manifest.

Reads all XLSX workbooks from test-docs/<module>/<module>.xlsx, extracts test cases
from TS-* tabs, classifies them by type (UI/API/hybrid), and produces a unified
manifest at autotests/manifest/test-cases.json.

Usage:
    python3 autotests/scripts/parse_xlsx.py [--output autotests/manifest/test-cases.json]
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# Project root (parent of autotests/)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# Expected column order in TS-* tabs (0-indexed)
COLUMNS = [
    "test_id",        # A
    "title",          # B
    "preconditions",  # C
    "steps",          # D
    "expected_result", # E
    "priority",       # F
    "type",           # G
    "requirement_ref", # H
    "module_component", # I
    "notes",          # J
]

# Patterns to classify test type from Steps column
API_PATTERNS = re.compile(
    r"(POST|GET|PUT|PATCH|DELETE)\s+/api/|"
    r"curl\s|"
    r"API\s+(request|call|endpoint)|"
    r"request\.(get|post|put|patch|delete)\(",
    re.IGNORECASE
)

UI_PATTERNS = re.compile(
    r"Navigate\s+to|"
    r"Click\s|"
    r"Fill\s|"
    r"Open\s+(the\s+)?dialog|"
    r"Open\s+(the\s+)?page|"
    r"Select\s+from|"
    r"Check\s+the\s+checkbox|"
    r"Log\s*in\s|"
    r"Verify\s+(the\s+)?(UI|page|form|dialog|table|button|field)",
    re.IGNORECASE
)


def classify_test_type(steps: str) -> str:
    """Classify test as UI, API, or hybrid based on Steps text."""
    has_api = bool(API_PATTERNS.search(steps))
    has_ui = bool(UI_PATTERNS.search(steps))
    if has_api and has_ui:
        return "hybrid"
    if has_api:
        return "API"
    return "UI"


def parse_workbook(xlsx_path: Path, module: str) -> dict:
    """Parse a single XLSX workbook, extracting test cases from TS-* tabs."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    suites = {}
    total = 0

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("TS-"):
            continue

        ws = wb[sheet_name]
        test_cases = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            # Skip back-link rows (first row often has "← Back to Plan")
            cell_val = str(row[0]).strip()
            if cell_val.startswith("←") or cell_val.startswith("Back"):
                continue

            # Extract columns (pad with empty strings if row is short)
            values = list(row) + [""] * max(0, len(COLUMNS) - len(row))
            tc = {}
            for i, col_name in enumerate(COLUMNS):
                val = values[i]
                tc[col_name] = str(val).strip() if val is not None else ""

            # Skip rows without a valid test ID
            if not tc["test_id"] or not tc["test_id"].startswith("TC-"):
                continue

            # Classify type from steps
            tc["classified_type"] = classify_test_type(tc["steps"])
            tc["automation_status"] = "pending"

            test_cases.append(tc)

        if test_cases:
            suites[sheet_name] = {"test_cases": test_cases}
            total += len(test_cases)

    wb.close()
    return {
        "xlsx_path": str(xlsx_path.relative_to(PROJECT_ROOT)),
        "total": total,
        "suites": suites,
    }


def parse_all(output_dir: Path) -> dict:
    """Parse all XLSX workbooks from output directory."""
    modules = {}
    total_cases = 0

    # Scan test-docs/ for module directories containing .xlsx files
    for entry in sorted(output_dir.iterdir()):
        if not entry.is_dir():
            continue
        xlsx_files = list(entry.glob("*.xlsx"))
        if not xlsx_files:
            continue

        module_name = entry.name
        xlsx_path = xlsx_files[0]  # One workbook per module

        print(f"  Parsing {module_name}: {xlsx_path.name}...", end=" ")
        module_data = parse_workbook(xlsx_path, module_name)
        modules[module_name] = module_data
        total_cases += module_data["total"]
        suite_count = len(module_data["suites"])
        print(f"{module_data['total']} cases in {suite_count} suites")

    return {
        "generated": datetime.now(timezone.utc).isoformat(),
        "total_cases": total_cases,
        "modules": modules,
    }


def main():
    parser = argparse.ArgumentParser(description="Parse XLSX test docs into JSON manifest")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "autotests" / "manifest" / "test-cases.json"),
        help="Output JSON file path",
    )
    args = parser.parse_args()

    output_dir = PROJECT_ROOT / "test-docs"
    if not output_dir.exists():
        print(f"ERROR: test-docs directory not found: {output_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing XLSX workbooks from {output_dir}...")
    manifest = parse_all(output_dir)

    # Write manifest
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"\nManifest written to {out_path}")
    print(f"Total: {manifest['total_cases']} test cases across {len(manifest['modules'])} modules")

    # Summary by module
    for module_name, data in manifest["modules"].items():
        types = {}
        for suite in data["suites"].values():
            for tc in suite["test_cases"]:
                t = tc["classified_type"]
                types[t] = types.get(t, 0) + 1
        type_str = ", ".join(f"{k}: {v}" for k, v in sorted(types.items()))
        print(f"  {module_name}: {data['total']} cases ({type_str})")


if __name__ == "__main__":
    main()
