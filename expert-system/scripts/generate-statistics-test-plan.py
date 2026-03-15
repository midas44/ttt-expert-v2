#!/usr/bin/env python3
"""Generate test-plan-statistics.xlsx for TTT Statistics module."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/test-plan-statistics.xlsx"

# Shared styles
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ROW_FILL_ALT = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


def style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        fill = ROW_FILL_ALT if (r - start_row) % 2 == 1 else PatternFill()
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if fill.start_color and fill.start_color.rgb != "00000000":
                cell.fill = fill


def auto_width(ws, col_widths):
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Overview ──────────────────────────────────────────────
def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"

    sections = [
        ("Module", "Statistics (General Statistics + Employee Reports)"),
        ("Version", "release/2.1 (Sprint 15)"),
        ("Date", "2026-03-15"),
        ("Author", "Expert System (Claude Code)"),
        ("", ""),
        ("SCOPE", ""),
        (
            "In Scope",
            "1. General Statistics page — 13 permission-gated tabs, search filters, date range, view options (Tree/Flat, Hours/Days), data display, export (CSV, clipboard, Google Sheets link)\n"
            "2. Employee Reports page — employee search (Latin/Cyrillic/keyboard layout), month picker, over-limit toggle, deviation display, norm calculation, comment CRUD, absence icons, project breakdown\n"
            "3. Statistics API — 12+ endpoints across /v1/reports/* and /v1/statistic/* families, mixed unit handling (HOURS vs MINUTES)\n"
            "4. Norm calculation — personal norm vs budget norm, deviation formula, ExcessStatus (HIGH/LOW/NEUTRAL/NA), edge cases (zero norm, partial month)\n"
            "5. Statistic report cache — pre-computed table, 3 update paths (nightly cron, task report event, RabbitMQ), race conditions\n"
            "6. Access control — role-based tab visibility, Employee Reports page access, API permissions\n"
            "7. Export — CSV download, clipboard copy, Google Sheets link, individual norm export (#3400)",
        ),
        (
            "Out of Scope",
            "- Report CRUD (covered in Reports test plan)\n"
            "- Confirmation flow (covered in Confirmation test plan)\n"
            "- Period management (covered in Accounting test plan)\n"
            "- Admin settings (threshold configuration — covered in Admin test plan)",
        ),
        ("", ""),
        ("OBJECTIVES", ""),
        (
            "Primary",
            "1. Verify correct data display across all statistics views for all roles\n"
            "2. Validate norm calculation accuracy (personal, budget, deviation)\n"
            "3. Confirm role-based access control and tab visibility\n"
            "4. Test API endpoint correctness, error handling, and unit consistency\n"
            "5. Verify export functionality (CSV format, data accuracy)\n"
            "6. Validate statistic report cache synchronization",
        ),
        (
            "Secondary",
            "1. Document known bugs and their impact on test execution\n"
            "2. Cross-environment consistency (timemachine vs stage)\n"
            "3. Edge cases: zero norm, terminated employees, partial months",
        ),
        ("", ""),
        ("APPROACH", ""),
        (
            "Test Types",
            "- Functional UI: Playwright browser automation against timemachine/stage\n"
            "- API: Direct endpoint testing via Swagger MCP or curl\n"
            "- Data validation: PostgreSQL SELECT queries to verify calculations\n"
            "- Cross-environment: Compare timemachine (dev) vs stage (prod-like)\n"
            "- Negative/boundary: Missing params, invalid values, permission violations",
        ),
        (
            "Test Data Strategy",
            "- Use existing employee accounts on timemachine: alsmirnov (EMPLOYEE-only, 1 tab), pvaynmaster (multi-role, 8 tabs), dergachev (manager)\n"
            "- Mine test data via SQL: SELECT DISTINCT employee_login, reported_effort, month_norm, budget_norm FROM ttt_backend.statistic_report WHERE report_date >= '2026-01-01'\n"
            "- For norm edge cases: identify employees with admin vacations via SQL: SELECT e.login FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id=e.id WHERE v.type='ADMINISTRATIVE' AND v.status='APPROVED'\n"
            "- For zero-norm scenarios: find contractors or part-time employees with budget_norm=0\n"
            "- For terminated employees: SELECT login, end_date FROM ttt_backend.employee WHERE end_date IS NOT NULL AND end_date >= '2026-01-01'",
        ),
        ("", ""),
        ("ENVIRONMENT", ""),
        (
            "Requirements",
            "- Primary: timemachine (ttt-timemachine.noveogroup.com) — full dev environment with statistic_report cache table\n"
            "- Secondary: stage (ttt-stage.noveogroup.com) — production-like, for cross-env comparison\n"
            "- Comparison: qa-1 (ttt-qa-1.noveogroup.com) — no statistic_report table (on-the-fly computation)\n"
            "- Browser: Chrome/Chromium via Playwright\n"
            "- Database: PostgreSQL 5433 (ttt_backend schema for statistic_report, task_report; ttt_vacation for absences)\n"
            "- API: Swagger MCP tools or curl with API_SECRET_TOKEN header\n"
            "- VPN: Required for all environments (*.noveogroup.com)",
        ),
        (
            "Known Limitations",
            "- #3400 individual norm CSV export is marked 'Production Ready' but code not found in release/2.1 — may need separate branch or manual verification\n"
            "- Customer tabs (VIEW_CUSTOMER permission) never observed live — may require specific test account setup\n"
            "- Google Docs statistics spec is inaccessible (401) — testing based on Confluence requirements + live behavior\n"
            "- statistic_report nightly sync at 4:00 AM — timing-sensitive tests may need manual trigger via test API",
        ),
    ]

    auto_width(ws, [20, 120])
    for i, (label, value) in enumerate(sections, 1):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=value)
        c1.font = BOLD_FONT if label else BODY_FONT
        c2.font = BODY_FONT
        c1.alignment = WRAP
        c2.alignment = WRAP
        if label in ("SCOPE", "OBJECTIVES", "APPROACH", "ENVIRONMENT"):
            c1.font = SECTION_FONT


# ── Sheet 2: Feature Matrix ────────────────────────────────────────
def build_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    headers = [
        "Feature Area",
        "Functional UI",
        "API",
        "Data/DB",
        "Security/Access",
        "Negative/Boundary",
        "Cross-Env",
        "Total Cases",
        "Qase Existing",
        "Notes",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    features = [
        ["General Statistics — Tab Visibility", "5", "1", "-", "3", "1", "-", "10", "0", "13 code-defined tabs, max 8 observed live"],
        ["General Statistics — Search & Filters", "5", "-", "-", "-", "2", "-", "7", "0", "4 filter types: project, employee, task, customer"],
        ["General Statistics — Date & View Controls", "5", "-", "-", "-", "1", "-", "6", "0", "Date range, presets, Tree/Flat, Hours/Days"],
        ["General Statistics — Data Display", "4", "2", "2", "-", "1", "1", "10", "0", "Expandable rows, CS links, report links"],
        ["General Statistics — Export", "4", "1", "-", "-", "1", "1", "7", "0", "CSV, clipboard, Google Sheets, #3400 norm export"],
        ["Employee Reports — Access & Layout", "3", "1", "-", "3", "1", "-", "8", "0", "Role-gated page, 403 for EMPLOYEE"],
        ["Employee Reports — Search & Filters", "5", "1", "-", "-", "2", "-", "8", "0", "Latin/Cyrillic/login, keyboard layout detection"],
        ["Employee Reports — Norm & Deviation", "5", "2", "3", "-", "4", "1", "15", "0", "Personal vs budget, deviation formula, N/A%"],
        ["Employee Reports — Comments", "4", "2", "1", "-", "1", "-", "8", "0", "Inline edit, per-month storage, blur save"],
        ["Employee Reports — Absence Icons", "3", "1", "1", "-", "1", "-", "6", "0", "Vacation/sick leave icons with tooltips"],
        ["Statistics API — Reports Family", "-", "8", "2", "1", "3", "1", "15", "0", "Mixed HOURS/MINUTES units, 500-on-missing-params bug"],
        ["Statistics API — Statistic Family", "-", "6", "1", "1", "2", "1", "11", "0", "Tree endpoints, export, permissions"],
        ["Statistic Report Cache", "-", "2", "5", "-", "2", "1", "10", "0", "3 update paths, race condition, 2-month window"],
        ["TOTAL", "43", "27", "15", "8", "22", "6", "121", "0", ""],
    ]

    for r, row in enumerate(features, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 2, len(features) + 1, len(headers))
    # Bold the total row
    total_row = len(features) + 1
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).font = BOLD_FONT

    auto_width(ws, [40, 14, 8, 10, 16, 18, 12, 13, 14, 50])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(features)+1}"


# ── Sheet 3: Risk Assessment ──────────────────────────────────────
def build_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    risks = [
        [
            "Mixed API Units (HOURS/MINUTES)",
            "Frontend may display incorrect values if unit conversion is wrong or inconsistent across endpoints",
            "High",
            "High",
            "Critical",
            "Cross-validate API response units against DB values and UI display for same employee/period. Compare /v1/reports/summary (HOURS) vs /v1/reports/total (MINUTES) for same data.",
        ],
        [
            "Norm Calculation — Budget vs Personal",
            "Incorrect deviation % if wrong norm type used, or admin vacation handling error. Excess uses budgetNorm not personalNorm (design issue).",
            "Medium",
            "High",
            "High",
            "Test employees with admin vacations: verify {individual} ({budget}) display. Cross-reference statistic_report.month_norm vs budget_norm. Verify deviation formula uses budgetNorm.",
        ],
        [
            "Deviation N/A% Edge Case",
            "Zero-norm employees (norm=0, reported>0) should show '+N/A%' and sort to top. Incorrect handling could hide over-reporters.",
            "Medium",
            "Medium",
            "High",
            "Find zero-norm employees via SQL: WHERE budget_norm=0 AND reported_effort>0. Verify N/A% display and sort order.",
        ],
        [
            "Statistic Report Race Condition",
            "Concurrent MQ event + task report event for same employee/month. No pessimistic locking — last-write-wins could corrupt data.",
            "Low",
            "High",
            "Medium",
            "Submit task report while triggering vacation change for same employee/month. Verify statistic_report consistency after both complete.",
        ],
        [
            "Tab Permission Gating",
            "Wrong tabs visible for role, or missing tabs. 13 code-defined tabs but max 8 observed. Customer tabs (2) never triggered.",
            "Medium",
            "Medium",
            "Medium",
            "Test with specific role accounts. Attempt to identify/create user with VIEW_CUSTOMER permission. Verify exact tab set per role.",
        ],
        [
            "Employee Reports 403 for EMPLOYEE",
            "EMPLOYEE-only users should get 403 on /statistics/employee-reports. If not enforced, data exposure risk.",
            "Low",
            "High",
            "Medium",
            "Test with alsmirnov (EMPLOYEE-only). Verify 403 response both in UI and direct API call.",
        ],
        [
            "500 Error on Missing Params",
            "Several endpoints return 500 instead of 400 when required @RequestParam params missing (MissingServletRequestParameterException).",
            "High",
            "Low",
            "Medium",
            "Test all @RequestParam endpoints without required params. Document expected vs actual HTTP status. Verify no stack trace leakage.",
        ],
        [
            "Comment Save on Blur Only",
            "Comment changes lost if user navigates away without blur event (tab close, browser back). No auto-save, no unsaved-changes warning.",
            "Medium",
            "Low",
            "Low",
            "Edit comment, then navigate away without clicking outside. Verify data loss behavior. Document as known limitation.",
        ],
        [
            "Stale Filter Persistence (redux-persist)",
            "General Statistics filters persisted across sessions. User sees stale data from previous session, causing confusion.",
            "Medium",
            "Low",
            "Low",
            "Set specific filters, close browser, reopen. Verify filters persist. Test 'Reset all filters' button clears persisted state.",
        ],
        [
            "Cross-Env Field Differences",
            "Timemachine and Stage have different field sets (15 vs 17 fields in statistic API responses). May cause UI errors on one env.",
            "Low",
            "Medium",
            "Low",
            "Run same API calls on both environments. Document field differences. Verify UI handles missing/extra fields gracefully.",
        ],
        [
            "#3400 Individual Norm Export",
            "Ticket marked 'Production Ready' but code not found in release/2.1 codebase. Export may not exist or be in unmerged branch.",
            "High",
            "Low",
            "Low",
            "Check if export endpoint exists on timemachine/stage. If missing, document as implementation gap. If present, verify CSV columns match spec.",
        ],
        [
            "Vacation Statistic 403 (BUG-STAT-UI-2)",
            "Vacation statistics endpoint returns 403 for some users, causing absence data to fail silently with error banner.",
            "Medium",
            "Medium",
            "Medium",
            "Identify affected roles/users. Verify absence icons presence/absence correlates with API 403. Test fallback behavior.",
        ],
    ]

    for r, row in enumerate(risks, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    style_body(ws, 2, len(risks) + 1, len(headers))
    auto_width(ws, [35, 55, 12, 10, 10, 65])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(risks)+1}"


# ── Main ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
build_overview(wb)
build_feature_matrix(wb)
build_risk_assessment(wb)
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
