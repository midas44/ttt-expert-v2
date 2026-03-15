#!/usr/bin/env python3
"""Generate test-cases-statistics.xlsx for TTT Statistics module.

~120 test cases across 7 feature sheets. Each case includes:
Test ID, Title, Preconditions, Steps, Expected Result, Priority, Type,
Requirement Ref, Module/Component, Notes (incl. test data generation).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/home/v/Dev/ttt-expert-v1/expert-system/output/test-cases-statistics.xlsx"

# ── Styles ─────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BODY_FONT = Font(name="Arial", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ROW_ALT = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")

HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps",
    "Expected Result", "Priority", "Type",
    "Requirement Ref", "Module/Component", "Notes",
]
COL_WIDTHS = [14, 38, 38, 55, 45, 10, 14, 18, 22, 45]


def write_sheet(wb, sheet_name, cases):
    ws = wb.create_sheet(sheet_name)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_C
        cell.border = THIN_BORDER
    for r, case in enumerate(cases, 2):
        fill = ROW_ALT if r % 2 == 0 else PatternFill()
        for c, val in enumerate(case, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if fill.start_color and fill.start_color.rgb != "00000000":
                cell.fill = fill
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(cases)+1}"


# ═══════════════════════════════════════════════════════════════════
#  SHEET 1 — General Statistics UI
# ═══════════════════════════════════════════════════════════════════
general_stats_ui = [
    # --- Tab Visibility ---
    [
        "TC-STAT-001", "Tab visibility — EMPLOYEE-only user sees 1 tab",
        "Logged in as alsmirnov (EMPLOYEE role only) on timemachine.",
        "1. Navigate to /statistics/general\n2. Count visible tabs",
        "Only 1 tab visible: 'My tasks'. No other tabs displayed.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: alsmirnov has EMPLOYEE role only. Verify via GET /v1/statistic/permissions — should return minimal set.",
    ],
    [
        "TC-STAT-002", "Tab visibility — multi-role user sees 8 tabs",
        "Logged in as pvaynmaster (7+ roles including PM, DM, etc.) on timemachine.",
        "1. Navigate to /statistics/general\n2. Count visible tabs\n3. Note exact tab names",
        "8 tabs visible: My tasks, My projects, Employees on my projects, Department projects, Department employees, Office projects, Office employees, Tasks by employees.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Verified in session 29. Tab visibility driven by VIEW_MY_TASKS, VIEW_MY_PROJECTS, VIEW_MY_DEPARTMENT, VIEW_MY_OFFICE, VIEW_PROJECT permissions.",
    ],
    [
        "TC-STAT-003", "Tab visibility — maximum 13 tabs with all permissions",
        "User with ALL statistics permissions (including VIEW_CUSTOMER). May need API key or admin with full permissions.",
        "1. Navigate to /statistics/general with a user having VIEW_CUSTOMER permission\n2. Count visible tabs\n3. Verify Customer projects and Customer employees tabs",
        "All 13 code-defined tabs visible including: Customer projects, Customer employees.",
        "Medium", "UI",
        "Confluence 119244531; code: TabsWithFiltersContainer", "frontend-statistics-module",
        "Customer tabs (VIEW_CUSTOMER) never observed live in session 29. May require specific permission setup. SQL: SELECT DISTINCT role FROM ttt_backend.employee_role WHERE role LIKE '%CUSTOMER%'",
    ],
    [
        "TC-STAT-004", "Tab visibility — DEPARTMENT_MANAGER sees department tabs",
        "Logged in as a DEPARTMENT_MANAGER on timemachine.",
        "1. Navigate to /statistics/general\n2. Verify 'Department projects' and 'Department employees' tabs are visible\n3. Click each tab and verify data scoped to department",
        "Department tabs visible. Data shows only employees/projects belonging to user's department.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: find DM users via SQL: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role r ON e.id=r.employee_id WHERE r.role='DEPARTMENT_MANAGER'",
    ],
    [
        "TC-STAT-005", "Tab switching preserves date range (redux-persist)",
        "Logged in as multi-role user. General Statistics page loaded.",
        "1. Set custom date range (e.g., 2026-01-15 to 2026-02-15)\n2. Switch from 'My tasks' to 'My projects' tab\n3. Verify date range preserved\n4. Close browser tab, reopen /statistics/general\n5. Verify date range still persisted",
        "Date range preserved across tab switches AND browser sessions (redux-persist).",
        "Medium", "UI",
        "Code: redux-persist on state.statistics", "frontend-statistics-module",
        "Known tech debt: stale filter persistence can confuse users. Related: TC-STAT-017.",
    ],
    # --- Search & Filters ---
    [
        "TC-STAT-006", "Search filter — by project name",
        "Logged in as multi-role user on timemachine. General Statistics page loaded.",
        "1. Click 'project' filter button in search bar\n2. Type a known project name (partial match)\n3. Select from suggestions\n4. Verify table filters to show only selected project's data",
        "Table shows data only for the selected project. Other projects hidden.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: pick active project from SQL: SELECT name FROM ttt_backend.project WHERE status='ACTIVE' LIMIT 5",
    ],
    [
        "TC-STAT-007", "Search filter — by employee name",
        "Logged in as multi-role user. General Statistics page loaded.",
        "1. Click 'employee' filter button\n2. Type employee name (e.g., 'dergachev')\n3. Select from suggestions\n4. Verify data scoped to that employee",
        "Table shows data only for the selected employee.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: use known employees: alsmirnov, pvaynmaster, dergachev.",
    ],
    [
        "TC-STAT-008", "Search filter — by task name",
        "Logged in as multi-role user. General Statistics page loaded.",
        "1. Click 'task' filter button\n2. Type a known task name\n3. Select from suggestions\n4. Verify table shows only matching task data",
        "Table filtered to selected task only.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: SQL: SELECT DISTINCT name FROM ttt_backend.task WHERE deleted=false LIMIT 10",
    ],
    [
        "TC-STAT-009", "Search filter — by customer (VIEW_CUSTOMER required)",
        "Logged in as user with VIEW_CUSTOMER permission.",
        "1. Verify 'customer' filter button appears in search bar\n2. Click customer filter\n3. Search and select a customer\n4. Verify data scoped to customer's projects",
        "Customer filter button visible (4th filter). Data filtered to customer's projects.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "EMPLOYEE-only users see only 3 filters (no customer). Requires VIEW_CUSTOMER permission.",
    ],
    [
        "TC-STAT-010", "Reset all filters clears all active filters",
        "Logged in with filters applied (project + date range set).",
        "1. Apply project filter and custom date range\n2. Verify 'Reset all filters' button is enabled\n3. Click 'Reset all filters'\n4. Verify all filters cleared, default data restored",
        "'Reset all filters' button clears all active filters. Date range resets to default (Jan 1 – Dec 31 current year). Table shows unfiltered data.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Button is disabled when no filters active. Also clears redux-persisted state.",
    ],
    # --- Date & View Controls ---
    [
        "TC-STAT-011", "Date range — custom date selection",
        "General Statistics page loaded.",
        "1. Click start date picker, select 2026-02-01\n2. Click end date picker, select 2026-02-28\n3. Verify table data updates to show only February data",
        "Table displays data only for the February 2026 period. Period/total columns reflect Feb range.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Default date range is Jan 1 – Dec 31 of current year. Calendar icon next to each picker.",
    ],
    [
        "TC-STAT-012", "Date range — period presets dropdown",
        "General Statistics page loaded.",
        "1. Click period preset dropdown ('Please select a date range')\n2. Select a preset (e.g., 'This month', 'Last month', 'This quarter')\n3. Verify date pickers update to match preset\n4. Verify table data refreshes",
        "Date pickers reflect preset range. Table data matches the selected preset period.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Available presets: verify exact list of options in dropdown.",
    ],
    [
        "TC-STAT-013", "View toggle — Tree vs Flat mode",
        "General Statistics page loaded with data.",
        "1. Select 'Tree' radio button\n2. Verify rows are hierarchical (employee → project → task) with expand/collapse\n3. Select 'Flat' radio button\n4. Verify all rows displayed flat without hierarchy",
        "Tree: hierarchical with expand arrows. Flat: all rows at same level. Default is Flat.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Tree mode uses rc-table with lazy child expansion. Flat mode shows all data at once.",
    ],
    [
        "TC-STAT-014", "Time format toggle — Hours vs Days",
        "General Statistics page loaded with data.",
        "1. Select 'Hours' radio button, note effort values\n2. Select 'Days' radio button, note effort values\n3. Verify conversion: hours / 8 = days (approx.)",
        "Hours: values in hours (e.g., 152.0). Days: values in working days (e.g., 19.0). Conversion: hours ÷ 8.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "effortDisplayType toggle affects all effort columns. Units from effortTypeRadioGroupContainer.",
    ],
    [
        "TC-STAT-015", "Refresh data button reloads current view",
        "General Statistics page loaded with filters applied.",
        "1. Note current table data\n2. Click blue 'Refresh data' button\n3. Verify data reloads (loading spinner shown briefly)\n4. Verify filters preserved after refresh",
        "Data reloads. Filters and view mode preserved. New data reflected if changes occurred.",
        "Low", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "",
    ],
    # --- Data Display ---
    [
        "TC-STAT-016", "Tree mode — expand employee row to see projects and tasks",
        "General Statistics page in Tree mode with data.",
        "1. Click expand arrow on an employee row\n2. Verify project rows appear below\n3. Click expand on a project row\n4. Verify task rows appear below project",
        "Employee → project → task hierarchy displayed. Each level shows For the period and Total columns.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Lazy child expansion — API call made on expand. processStatisticsData may overwrite last absence record (tech debt #5).",
    ],
    [
        "TC-STAT-017", "Employee row — CompanyStaff profile link",
        "General Statistics page with data. Employee rows visible.",
        "1. Find an employee row\n2. Click CS profile link icon\n3. Verify opens CompanyStaff profile in new tab",
        "CS profile link opens https://companystaff.noveogroup.com/profile/{login} in new tab.",
        "Low", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Hardcoded CS URL (tech debt #4). Verify link target is correct for the employee's login.",
    ],
    [
        "TC-STAT-018", "Employee row — Report page navigation link",
        "General Statistics page with data. Employee rows visible.",
        "1. Find an employee row\n2. Click report page link icon\n3. Verify navigates to /report page for that employee",
        "Report page link navigates to employee's report page.",
        "Low", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-019", "Column sorting — ascending and descending",
        "General Statistics page with multiple data rows.",
        "1. Click 'For the period' column header\n2. Verify rows sorted ascending (↑)\n3. Click again\n4. Verify rows sorted descending (↓)\n5. Repeat for 'Total' column",
        "Sorting toggles ascending ↔ descending on each click. Sort indicator (↑/↓) displayed.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-020", "Sick leave absence icon and tooltip",
        "General Statistics page, employee with sick leave in current period.",
        "1. Navigate to a tab showing employees\n2. Find employee with sick leave icon (medical cross icon)\n3. Hover over icon to see tooltip\n4. Verify tooltip shows: total hours + individual period dates, status, type",
        "Sick leave icon (IconSick) visible. Tooltip shows breakdown: total hours, each period with date range and status. Hours/days unit matches current effortDisplayType setting.",
        "Medium", "UI",
        "Confluence 119244531; #2435", "frontend-statistics-module",
        "Test data: SQL: SELECT sl.employee, sl.start_date, sl.end_date, sl.status FROM ttt_vacation.sick_leave sl WHERE sl.start_date >= '2026-01-01' AND sl.status != 'DELETED'",
    ],
    [
        "TC-STAT-021", "Vacation absence icon and tooltip",
        "General Statistics page, employee with approved vacation in current period.",
        "1. Find employee with vacation icon\n2. Hover to see tooltip\n3. Verify tooltip shows vacation periods within date range, using N-dash for date ranges",
        "Vacation icon (IconVacation) visible. Tooltip shows dates with N-dash (–), not hyphen (-).",
        "Medium", "UI",
        "Confluence 119244531; #2435", "frontend-statistics-module",
        "Test data: SQL: SELECT v.employee_id, v.start_date, v.end_date FROM ttt_vacation.vacation v WHERE v.status='APPROVED' AND v.start_date >= '2026-01-01'",
    ],
    [
        "TC-STAT-022", "Permission denied message — typo verification",
        "Logged in as EMPLOYEE-only user. Navigate to tab requiring higher permission.",
        "1. Log in as alsmirnov\n2. Navigate to /statistics/general\n3. Attempt to access a restricted tab via direct URL manipulation\n4. Observe error message",
        "Error message displayed. Known BUG-STAT-UI-1: message contains typo 'You do'nt have access' instead of 'You don't have access'.",
        "Low", "UI",
        "BUG-STAT-UI-1", "frontend-statistics-module",
        "Verify typo is still present. Document for future fix.",
    ],
    # --- Export ---
    [
        "TC-STAT-023", "Export — Download CSV",
        "General Statistics page with data loaded. Date range set.",
        "1. Click Export dropdown\n2. Select 'Download CSV'\n3. Verify CSV file downloads\n4. Open CSV, verify column headers and data match UI display",
        "CSV file downloaded. Columns match table headers. Data in HOURS (export unit). Values match UI display.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Export endpoints use HOURS unit. Verify conversion from MINUTES (API native) to HOURS in export.",
    ],
    [
        "TC-STAT-024", "Export — Copy the table to clipboard",
        "General Statistics page with data loaded.",
        "1. Click Export dropdown\n2. Select 'Copy the table'\n3. Paste into a text editor\n4. Verify tab-separated data matches table content",
        "Table data copied to clipboard in tab-separated format. Data matches current view.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-025", "Export — Copy link for Google tables",
        "General Statistics page with data loaded.",
        "1. Click Export dropdown\n2. Select 'Copy link for Google tables'\n3. Paste link into Google Sheets via IMPORTDATA\n4. Verify data loads in Google Sheets",
        "Link copied to clipboard. When used in Google Sheets IMPORTDATA, correct data displayed.",
        "Medium", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Link format is a direct API URL with auth token. Verify format and data accuracy.",
    ],
    [
        "TC-STAT-026", "Export — employees-largest-customers CSV",
        "API access via Swagger or curl.",
        "1. GET /v1/statistic/export/employees-largest-customers?startDate=2026-01-01&endDate=2026-03-31\n2. Verify response is CSV format\n3. Verify unit is HOURS\n4. Cross-reference values with /v1/statistic/employees response (MINUTES ÷ 60)",
        "CSV response with correct headers. Values in HOURS. Match: statistic_employee_effort_MINUTES / 60 = CSV_HOURS.",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "timeUnit param may be configurable. Test with default and explicit HOURS/MINUTES.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 2 — Employee Reports UI
# ═══════════════════════════════════════════════════════════════════
employee_reports_ui = [
    # --- Access & Layout ---
    [
        "TC-STAT-027", "Employee Reports — access granted for ADMIN",
        "Logged in as user with ADMIN role on timemachine.",
        "1. Navigate to /statistics/employee-reports\n2. Verify page loads with employee table\n3. Verify all employees visible (not scoped to office/department)",
        "Page loads successfully. Table shows all employees. Menu shows 'General Statistics' + 'Employee Reports'.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "ADMIN and CHIEF_ACCOUNTANT see all employees.",
    ],
    [
        "TC-STAT-028", "Employee Reports — access granted for CHIEF_ACCOUNTANT",
        "Logged in as CHIEF_ACCOUNTANT on timemachine.",
        "1. Navigate to /statistics/employee-reports\n2. Verify page loads\n3. Verify all employees visible",
        "Page loads with full employee list (same as ADMIN scope).",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Test data: find CHIEF_ACCOUNTANT via SQL: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role r ON e.id=r.employee_id WHERE r.role='CHIEF_ACCOUNTANT'",
    ],
    [
        "TC-STAT-029", "Employee Reports — OFFICE_ACCOUNTANT sees own office only",
        "Logged in as OFFICE_ACCOUNTANT (specific salary office).",
        "1. Navigate to /statistics/employee-reports\n2. Verify page loads\n3. Verify only employees from user's salary office(s) are shown",
        "Employee list scoped to user's salary office(s). Other office employees not visible.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "Cross-verify with SQL: SELECT COUNT(*) FROM ttt_backend.employee WHERE salary_office_id IN (SELECT salary_office_id FROM ... WHERE login='accountant_login')",
    ],
    [
        "TC-STAT-030", "Employee Reports — DEPARTMENT_MANAGER sees subordinates only",
        "Logged in as DEPARTMENT_MANAGER.",
        "1. Navigate to /statistics/employee-reports\n2. Verify only department subordinates listed",
        "Employee list scoped to user's department. Other department employees not visible.",
        "High", "UI",
        "Confluence 119244531", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-031", "Employee Reports — 403 for EMPLOYEE-only user",
        "Logged in as alsmirnov (EMPLOYEE role only).",
        "1. Navigate to /statistics/employee-reports\n2. Observe response",
        "Page shows 403 Forbidden or access denied. No employee data visible.",
        "High", "UI/Security",
        "Confluence 119244531", "frontend-statistics-module",
        "Also test direct API: GET /v1/statistic/report/employees — should return 403 for EMPLOYEE role.",
    ],
    # --- Search & Filters ---
    [
        "TC-STAT-032", "Employee search — by Latin first/last name",
        "Employee Reports page loaded as privileged user.",
        "1. Type Latin name fragment (e.g., 'Derg') in search field\n2. Observe table filtering in real-time\n3. Verify matching employees shown",
        "Table filters as user types. Only employees matching Latin name fragment displayed.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-033", "Employee search — by Cyrillic name",
        "Employee Reports page loaded. UI in EN or RU mode.",
        "1. Type Cyrillic name fragment (e.g., 'Дерг')\n2. Verify table filters to matching employees",
        "Search works with Cyrillic characters. Matching employees displayed.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Application supports bilingual names (Latin + Cyrillic).",
    ],
    [
        "TC-STAT-034", "Employee search — by login",
        "Employee Reports page loaded.",
        "1. Type employee login (e.g., 'dergachev')\n2. Verify table shows matching employee",
        "Search by login works. Exact login match displayed.",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-035", "Employee search — wrong keyboard layout detection",
        "Employee Reports page loaded.",
        "1. Switch keyboard to Russian layout\n2. Type English name using Russian keys (e.g., 'вукпфсрум' for 'dergachev')\n3. Verify search detects wrong layout and shows correct results",
        "Search detects wrong keyboard layout. Shows results matching the intended Latin name.",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Implementation: keyboard layout detection in search component. Verify both RU→EN and EN→RU directions.",
    ],
    [
        "TC-STAT-036", "Month picker — default value",
        "Employee Reports page loaded for the first time.",
        "1. Observe month picker default value\n2. Verify it shows last month open for confirmation in user's salary office",
        "Default month = last month open for confirmation in user's salary office (approve period start month).",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Cross-verify: GET /v1/periods/approve for user's office → the period start month should match default.",
    ],
    [
        "TC-STAT-037", "Month picker — change month",
        "Employee Reports page loaded.",
        "1. Click month picker\n2. Select a different month (e.g., January 2026)\n3. Verify table data refreshes for selected month\n4. Verify Reported, Norm, Deviation values update",
        "Table shows data for newly selected month. All columns reflect recalculated values for that month.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-038", "Over-limit toggle — filter employees above/below thresholds",
        "Employee Reports page loaded.",
        "1. Enable 'Only over-limit deviations' toggle\n2. Verify table shows only employees exceeding over/under thresholds\n3. Disable toggle\n4. Verify all employees restored",
        "Toggle on: only employees with |deviation%| exceeding admin thresholds shown. Toggle off: all employees visible. Thresholds from TTT Parameters (default: over=10%, under=30%).",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Threshold values: GET admin settings → notification.reporting.over, notification.reporting.under. The toggle triggers server-side filtering via exceedingLimit param.",
    ],
    # --- Norm & Deviation ---
    [
        "TC-STAT-039", "Reported column — over-report indicator (red arrow up)",
        "Employee Reports page with employee who has reported > budgetNorm.",
        "1. Find employee with over-reported hours\n2. Verify red ↑ arrow next to Reported value\n3. Verify text color is red when deviation % > over threshold",
        "Red arrow up (↑) displayed. Text turns red when deviation exceeds notification.reporting.over threshold.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Test data: SQL: SELECT employee_login, reported_effort, budget_norm FROM ttt_backend.statistic_report WHERE reported_effort > budget_norm AND report_date = '2026-02-01'",
    ],
    [
        "TC-STAT-040", "Reported column — under-report indicator (purple arrow down)",
        "Employee Reports page with employee who has reported < budgetNorm.",
        "1. Find employee with under-reported hours\n2. Verify purple ↓ arrow next to Reported value\n3. Verify text color is purple when |deviation %| > under threshold",
        "Purple arrow down (↓) displayed. Text turns purple when |deviation| exceeds notification.reporting.under threshold.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Test data: SQL: SELECT employee_login, reported_effort, budget_norm FROM ttt_backend.statistic_report WHERE reported_effort < budget_norm AND budget_norm > 0 AND report_date = '2026-02-01'",
    ],
    [
        "TC-STAT-041", "Norm display — with admin vacation ({individual} ({budget}))",
        "Employee Reports page. Employee with approved administrative (unpaid) vacation in selected month.",
        "1. Find employee with admin vacation\n2. Verify Norm column shows two values: '{individual} ({budget})'\n3. Verify individual < budget (admin vacation subtracted from personal but not budget)",
        "Norm displays as '{individual} ({budget})' format. Individual norm is lower than budget norm. Example: '120 (152)' means 120h personal, 152h budget.",
        "High", "UI",
        "Confluence 119244531 #3381", "frontend-statistics-module",
        "Test data: SQL: SELECT e.login FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id=e.id WHERE v.type='ADMINISTRATIVE' AND v.status='APPROVED' AND v.start_date >= '2026-01-01'\nBudget norm excludes admin vacation from off-periods (employee still 'counted').",
    ],
    [
        "TC-STAT-042", "Norm display — without admin vacation (single budget value)",
        "Employee Reports page. Employee with regular vacation or sick leave (no admin vacation).",
        "1. Find employee with only regular vacation/sick leave\n2. Verify Norm column shows single value (budget norm only)\n3. Verify value equals individual norm (they're the same)",
        "Norm displays single value: '{budget}'. No parenthetical. Individual == budget when no admin vacation.",
        "High", "UI",
        "Confluence 119244531 #3381", "frontend-statistics-module",
        "Case 2 from requirements: regular vacation/sick leave reduces both norms equally.",
    ],
    [
        "TC-STAT-043", "Norm display — no absences (general norm)",
        "Employee Reports page. Employee with no absences in selected month.",
        "1. Find employee with no vacations, sick leaves, or day-offs\n2. Verify Norm column shows single value = general office norm",
        "Single value displayed. Equals office calendar working hours for the month.",
        "Medium", "UI",
        "Confluence 119244531 #3381", "frontend-statistics-module",
        "Test data: SQL: SELECT employee_login FROM ttt_backend.statistic_report WHERE month_norm = budget_norm AND report_date = '2026-02-01'",
    ],
    [
        "TC-STAT-044", "Deviation display — integer value",
        "Employee Reports page. Employee with |deviation| >= 1%.",
        "1. Find employee with deviation >= 1% or <= -1%\n2. Verify deviation displayed as integer (e.g., '+15%', '-25%')\n3. No decimal places shown",
        "Deviation shown as integer percentage. Positive: '+15%'. Negative: '-25%'. No decimal places for |value| >= 1.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Formula: (reported - budgetNorm) / budgetNorm × 100%. Integer display for values outside (-1, +1) range.",
    ],
    [
        "TC-STAT-045", "Deviation display — decimal in (-1, +1) range",
        "Employee Reports page. Employee with deviation between -1% and +1%.",
        "1. Find employee with very small deviation (e.g., reported=152.5, norm=152)\n2. Verify deviation shows 1 decimal place (e.g., '+0.3%')",
        "Deviation displayed with 1 decimal place when value is in (-1, +1) range. Example: '+0.3%'.",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Test data: need employees with reported_effort very close to budget_norm. SQL: WHERE ABS(reported_effort - budget_norm/60.0) < budget_norm/60.0 * 0.01",
    ],
    [
        "TC-STAT-046", "Deviation display — N/A% for zero-norm employee",
        "Employee Reports page. Employee with budgetNorm=0 and reported_effort > 0.",
        "1. Find employee with zero norm but non-zero reported hours\n2. Verify deviation shows '+N/A%'\n3. Verify this employee sorts to TOP in deviation descending order",
        "Deviation shows '+N/A%'. In default sort (deviation descending), N/A% employees appear first (treated as maximum value).",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Test data: SQL: SELECT employee_login, reported_effort, budget_norm FROM ttt_backend.statistic_report WHERE budget_norm = 0 AND reported_effort > 0. ExcessStatus: NA.",
    ],
    [
        "TC-STAT-047", "Deviation display — 0% for zero-norm and zero-reported",
        "Employee Reports page. Employee with budgetNorm=0 AND reported_effort=0.",
        "1. Find employee with both zero norm and zero reported\n2. Verify deviation shows '0%'",
        "Deviation shows '0%'. Not N/A% (that requires reported > 0).",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Edge case: norm=0, reported=0 → ExcessStatus: NEUTRAL → 0%.",
    ],
    [
        "TC-STAT-048", "Default sort — by deviation descending",
        "Employee Reports page loaded with multiple employees.",
        "1. Observe default sort order\n2. Verify employees sorted by deviation % descending\n3. N/A% employees at top, then highest over-reporters, then normal, then under-reporters",
        "Default sort: deviation descending. Order: N/A% → highest positive → 0% → most negative.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    # --- Comments ---
    [
        "TC-STAT-049", "Comment field — create new comment",
        "Employee Reports page. Employee row without existing comment.",
        "1. Hover over comment cell of an employee\n2. Verify edit frame appears\n3. Click to enter edit mode\n4. Type comment text\n5. Click outside (blur) to save\n6. Refresh page, verify comment persisted",
        "Comment saved on blur. Persisted after page refresh. Cursor appears on click.",
        "High", "UI",
        "Confluence 119244531 #3309", "frontend-statistics-module",
        "API: POST /v1/statistic/report with body {employeeLogin, reportDate, comment}.",
    ],
    [
        "TC-STAT-050", "Comment field — edit existing comment",
        "Employee Reports page. Employee row with existing comment.",
        "1. Click on existing comment to enter edit mode\n2. Modify text\n3. Click outside to save\n4. Verify updated text shown",
        "Comment updated on blur. Previous text replaced. Change persisted.",
        "Medium", "UI",
        "Confluence 119244531 #3309", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-051", "Comment field — per-month storage",
        "Employee Reports page. Comment exists for one month.",
        "1. Add comment for employee in February\n2. Switch to March via month picker\n3. Verify February comment NOT shown (empty for March)\n4. Switch back to February\n5. Verify original comment still present",
        "Comments stored per-month. Switching months shows that month's comments only.",
        "High", "UI",
        "Confluence 119244531 #3309", "frontend-statistics-module",
        "DB: statistic_report stores comment per (employee_login, report_date) where report_date is 1st of month.",
    ],
    [
        "TC-STAT-052", "Comment field — data loss on navigate away without blur",
        "Employee Reports page.",
        "1. Click on comment cell, start typing\n2. WITHOUT clicking outside (no blur), navigate to another page\n3. Return to Employee Reports\n4. Verify comment was NOT saved",
        "Comment lost if user navigates without blur event. No unsaved-changes warning. Known limitation: saves on blur only (tech debt #9).",
        "Medium", "UI",
        "Tech debt #9", "frontend-statistics-module",
        "Document as known limitation. No auto-save, no dirty-state warning.",
    ],
    # --- Absence Icons & Expand ---
    [
        "TC-STAT-053", "Absence icons — vacation icon with tooltip",
        "Employee Reports page. Employee with approved vacation in selected month.",
        "1. Find employee row with vacation icon (small calendar icon)\n2. Hover to see tooltip\n3. Verify tooltip shows: total hours + individual periods with dates, status, payment type",
        "Vacation icon displayed inline. Tooltip shows: total hours, each period (date range with N-dash, status). Hours/days unit matches effortDisplayType.",
        "Medium", "UI",
        "Confluence 119244531 #2435", "frontend-statistics-module",
        "Icons: IconSick (medical cross), IconVacation (calendar). Both fetch data from vacation service.",
    ],
    [
        "TC-STAT-054", "Absence icons — sick leave icon with tooltip",
        "Employee Reports page. Employee with active/closed sick leave in selected month.",
        "1. Find employee row with sick leave icon\n2. Hover for tooltip\n3. Verify tooltip shows sick leave periods",
        "Sick leave icon displayed. Tooltip shows period dates, status, total hours.",
        "Medium", "UI",
        "Confluence 119244531 #2435", "frontend-statistics-module",
        "BUG-STAT-UI-2: If vacation statistic endpoint returns 403, absence data fails silently with error banner.",
    ],
    [
        "TC-STAT-055", "Expand project breakdown",
        "Employee Reports page with employees listed.",
        "1. Click on employee row (anywhere except name link)\n2. Verify accordion expands showing project breakdown\n3. Verify projects sorted by hours descending\n4. Verify individual project effort sums to total reported",
        "Accordion expands below employee row. Projects listed with hours, sorted descending. Sum of project hours = Reported column total.",
        "High", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "API: GET /v1/statistic/report/projects?employeeLogin=X&startDate=Y&endDate=Z. projectBreakdown stored in local state (re-fetched on every navigation — tech debt #10).",
    ],
    [
        "TC-STAT-056", "Manager filter column",
        "Employee Reports page with employees listed.",
        "1. Click Manager column header filter\n2. Select a specific manager from filter options\n3. Verify table filters to show only employees under selected manager\n4. Verify manager name links to CS profile",
        "Filter works. Only employees with selected manager shown. Manager name is clickable link to CS profile.",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-057", "Employee name — CS profile link",
        "Employee Reports page.",
        "1. Click on employee name link\n2. Verify opens CompanyStaff profile page",
        "Name is clickable, opens CS profile (https://companystaff.noveogroup.com/profile/{login}).",
        "Low", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-058", "Report page link on row hover",
        "Employee Reports page.",
        "1. Hover over employee row\n2. Verify report icon appears\n3. Hover over icon to see tooltip 'Employee report page'\n4. Click icon\n5. Verify navigates to employee's report page",
        "Report icon visible on hover. Tooltip: 'Employee report page'. Click navigates to /report page filtered for that employee.",
        "Low", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "",
    ],
    [
        "TC-STAT-059", "Terminated employee — visible in last active month",
        "Employee Reports page. Select month when an employee was terminated.",
        "1. Identify terminated employee via SQL\n2. Set month picker to their last active month\n3. Verify employee appears in list\n4. Change to next month\n5. Verify employee NOT in list",
        "Terminated employee visible in their last active month. Not visible in subsequent months.",
        "Medium", "UI",
        "Confluence 119244531 #3195", "frontend-statistics-module",
        "Test data: SQL: SELECT login, end_date FROM ttt_backend.employee WHERE end_date IS NOT NULL AND end_date >= '2025-01-01' ORDER BY end_date DESC LIMIT 5",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 3 — Statistics API
# ═══════════════════════════════════════════════════════════════════
statistics_api = [
    [
        "TC-STAT-060", "GET /v1/reports/summary — valid request",
        "API access to timemachine. Valid employee login and date.",
        "1. GET /v1/reports/summary?login=dergachev&date=2026-03-01\n2. Verify 200 response\n3. Verify response contains: week {reported, personalNorm, norm, personalNormForDate, normForDate} + month {same}",
        "200 OK. Response has week + month objects. Units: HOURS. normForDate = 0 for closed months, incremental for current.",
        "High", "API",
        "API surface map", "statistics-service-implementation",
        "Units: HOURS (not MINUTES). normForDate/personalNormForDate are progressive values.",
    ],
    [
        "TC-STAT-061", "GET /v1/reports/summary — missing login param (500 bug)",
        "API access to timemachine.",
        "1. GET /v1/reports/summary?date=2026-03-01 (omit login)\n2. Verify response status",
        "Expected: 400 Bad Request. Actual: 500 Internal Server Error (MissingServletRequestParameterException). BUG: should be 400.",
        "Medium", "API/Negative",
        "BUG session 11", "statistics-service-implementation",
        "Known bug: @RequestParam-required params return 500 when missing. Affects summary, effort endpoints.",
    ],
    [
        "TC-STAT-062", "GET /v1/reports/total — EMPLOYEE type, MONTH period",
        "API access to timemachine.",
        "1. GET /v1/reports/total?type=EMPLOYEE&startDate=2026-03-01&endDate=2026-03-31&periodType=MONTH\n2. Verify response items with employee objects\n3. Verify units are MINUTES",
        "200 OK. Items array with periodStartDate, employee object, statuses (NOTHING_APPROVE, APPROVED, WAITING_APPROVAL), effort. Units: MINUTES (300/day = 5 hours).",
        "High", "API",
        "API surface map", "statistics-service-implementation",
        "Units: MINUTES. Polymorphic: type=EMPLOYEE returns employee objects, type=PROJECT returns project objects.",
    ],
    [
        "TC-STAT-063", "GET /v1/reports/total — PROJECT type",
        "API access to timemachine.",
        "1. GET /v1/reports/total?type=PROJECT&startDate=2026-03-01&endDate=2026-03-31&periodType=MONTH\n2. Verify response items with project objects",
        "200 OK. Items contain project objects instead of employee objects. Same effort/status structure.",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-064", "GET /v1/reports/total — DAY and WEEK period types",
        "API access to timemachine.",
        "1. GET /v1/reports/total?type=EMPLOYEE&startDate=2026-03-01&endDate=2026-03-07&periodType=DAY\n2. Verify daily breakdown\n3. GET same with periodType=WEEK\n4. Verify weekly breakdown",
        "DAY: items per day. WEEK: items per week. Effort values aggregate accordingly.",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-065", "GET /v1/reports/effort — valid request",
        "API access to timemachine. Known taskId.",
        "1. GET /v1/reports/effort?taskId=<valid_task_id>\n2. Verify 200 response\n3. Verify cumulative all-time effort in MINUTES",
        "200 OK. Effort value in MINUTES. Cumulative all-time total for the task.",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "Test data: SQL: SELECT id FROM ttt_backend.task WHERE deleted=false LIMIT 1. No auth required (BUG-REPORT-6).",
    ],
    [
        "TC-STAT-066", "GET /v1/reports/effort — missing taskId (500 bug)",
        "API access to timemachine.",
        "1. GET /v1/reports/effort (omit taskId)\n2. Verify response",
        "Expected: 400. Actual: 500 (MissingServletRequestParameterException). Known bug.",
        "Low", "API/Negative",
        "BUG session 11", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-067", "GET /v1/reports/employees-over-reported — isPersonalNorm toggle",
        "API access to timemachine.",
        "1. GET /v1/reports/employees-over-reported?date=2026-02-01&isPersonalNorm=true\n2. Note employee count\n3. GET same with isPersonalNorm=false\n4. Compare counts",
        "isPersonalNorm=true: more employees (e.g., 91). isPersonalNorm=false: fewer (e.g., 77). Difference: employees where personalNorm < budgetNorm (part-time, admin vacation).",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "Units: HOURS. Response includes bilingual names but NO login field. 14-employee difference observed in session 11.",
    ],
    [
        "TC-STAT-068", "GET /v1/statistic/employees — tree data",
        "API access to timemachine.",
        "1. GET /v1/statistic/employees?startDate=2026-01-01&endDate=2026-03-31\n2. Verify tree-structured response\n3. Verify effort values in MINUTES\n4. Cross-reference: effort_MINUTES / 60 = effort_HOURS from summary endpoint",
        "Tree-structured response with expandable nodes. Units: MINUTES. 9120 MINUTES = 152 HOURS.",
        "High", "API",
        "API surface map", "statistics-service-implementation",
        "18 optional filter params. Test with and without filters. Verify tree-flat difference.",
    ],
    [
        "TC-STAT-069", "GET /v1/statistic/permissions — permission list",
        "API access with JWT token or API key.",
        "1. GET /v1/statistic/permissions\n2. Verify response is array of permission strings\n3. With API key: verify all 5 permissions returned\n4. With JWT (EMPLOYEE): verify limited permissions",
        "API key: all permissions (EMPLOYEES_VIEW, OFFICES_VIEW, STATISTICS_VIEW, etc.). JWT EMPLOYEE: limited set matching VIEW_MY_TASKS only.",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "22 permissions total observed in session 29. 5 statistics-specific permissions drive tab visibility.",
    ],
    [
        "TC-STAT-070", "GET /v1/statistic/report/employees — Employee Reports data",
        "API access as privileged user.",
        "1. GET /v1/statistic/report/employees?startDate=2026-02-01&endDate=2026-02-28\n2. Verify response contains employee list with: reported, monthNorm, budgetNorm, reportedNotificationStatus, comment\n3. Verify units",
        "Response with employee reports data. reportedNotificationStatus: HIGH/LOW/NEUTRAL/NA. monthNorm and budgetNorm in response.",
        "High", "API",
        "API surface map; Confluence #3195", "statistics-service-implementation",
        "This is the cache-backed endpoint reading from statistic_report table.",
    ],
    [
        "TC-STAT-071", "POST /v1/statistic/report — save comment",
        "API access as privileged user. ALLOW_API_MUTATIONS=true.",
        "1. POST /v1/statistic/report with body: {employeeLogin: 'dergachev', reportDate: '2026-02-01', comment: 'Test comment'}\n2. Verify 200 response\n3. GET /v1/statistic/report/employees for same month\n4. Verify comment appears for that employee",
        "Comment saved. Visible in subsequent GET request for the same employee/month.",
        "Medium", "API",
        "Confluence #3309", "statistics-service-implementation",
        "Mutation endpoint — requires explicit approval. Comment stored in statistic_report table.",
    ],
    [
        "TC-STAT-072", "Mixed unit consistency — cross-endpoint verification",
        "API access to timemachine. Same employee and period.",
        "1. GET /v1/reports/summary?login=X&date=2026-02-01 → note reported (HOURS)\n2. GET /v1/reports/total?type=EMPLOYEE&startDate=2026-02-01&endDate=2026-02-28&periodType=MONTH → find same employee → note effort (MINUTES)\n3. Verify: summary_hours × 60 = total_minutes\n4. Query DB: SELECT reported_effort FROM ttt_backend.statistic_report WHERE employee_login='X' AND report_date='2026-02-01'\n5. Verify DB value (HOURS) matches summary",
        "Consistent values across endpoints: summary(HOURS) × 60 = total(MINUTES). DB statistic_report.reported_effort (HOURS) = summary.reported.",
        "High", "API/Data",
        "Session 11 finding", "statistics-service-implementation",
        "Critical cross-cutting test. Mixed units are a known design issue. DB task_report.actual_efforts in MINUTES, DB statistic_report.reported_effort in HOURS.",
    ],
    [
        "TC-STAT-073", "Cross-environment field comparison — TM vs Stage",
        "API access to both timemachine and stage.",
        "1. GET /v1/statistic/report/employees on timemachine\n2. GET same on stage\n3. Compare response field sets\n4. Document differences",
        "Field sets may differ: 15 fields on TM vs 17 on Stage (session 31 finding). Name format may differ. Document exact differences.",
        "Medium", "API",
        "Session 31 finding", "statistics-service-implementation",
        "Known cross-env differences. Verify UI handles both gracefully.",
    ],
    [
        "TC-STAT-074", "GET /v1/reports/effort — no auth required (security bug)",
        "Unauthenticated API access (no JWT, no API key).",
        "1. GET /v1/reports/effort?taskId=<valid_id> WITHOUT any auth header\n2. Verify response",
        "Expected: 401 Unauthorized. Actual: 200 OK with data (BUG-REPORT-6: missing @PreAuthorize). Security vulnerability.",
        "High", "API/Security",
        "BUG-REPORT-6", "statistics-service-implementation",
        "Same bug on /v1/reports/employee-projects endpoint.",
    ],
    [
        "TC-STAT-075", "Statistics export endpoint — CSV format verification",
        "API access to timemachine.",
        "1. GET /v1/statistic/export/employees?startDate=2026-01-01&endDate=2026-03-31\n2. Verify Content-Type: text/csv or similar\n3. Parse CSV headers: EmployeeLogin, EmployeeName, Contractor, DepartmentManagerLogin, DepartmentManagerName, EffortForPeriod, EffortTotal, BeginDate, EndDate, NodeType\n4. Verify data accuracy",
        "CSV response with expected headers. Data in HOURS. Values match tree endpoint data (after MINUTES→HOURS conversion).",
        "Medium", "API",
        "API surface map", "statistics-service-implementation",
        "10 export endpoints under /v1/statistic/export/. Test at least employees and employees-largest-customers.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 4 — Norm Calculation
# ═══════════════════════════════════════════════════════════════════
norm_calculation = [
    [
        "TC-STAT-076", "Personal norm — employee with no absences",
        "Database access to timemachine. Employee with no time-offs in selected month.",
        "1. Find employee with no vacations/sick leaves/day-offs in Feb 2026\n   SQL: SELECT e.login FROM ttt_backend.employee e WHERE e.login NOT IN (SELECT DISTINCT sl.employee FROM ttt_vacation.sick_leave sl WHERE ...) AND ...\n2. GET /v1/reports/summary?login=X&date=2026-02-01\n3. Verify personalNorm == norm (office calendar hours)\n4. Query calendar service for office working hours in Feb",
        "personalNorm = norm = office calendar working hours. No reduction for absences.",
        "High", "Data",
        "Backend: BaseStatistic.getReportingNorm()", "statistics-service-implementation",
        "Norm clamped to employee work period (effectiveBounds). For full-month employee, personalNorm = office total.",
    ],
    [
        "TC-STAT-077", "Personal norm — employee with regular vacation",
        "Database access. Employee with approved regular vacation in selected month.",
        "1. Find employee with approved vacation in Feb 2026\n   SQL: SELECT e.login, v.start_date, v.end_date FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id=e.id WHERE v.status='APPROVED' AND v.type='REGULAR' AND v.start_date <= '2026-02-28' AND v.end_date >= '2026-02-01'\n2. GET /v1/reports/summary → personalNorm\n3. Calculate expected: officeNorm - vacationWorkingHours\n4. Verify match",
        "personalNorm = officeNorm - vacationWorkingHours. Reduction proportional to vacation working days in month.",
        "High", "Data",
        "Backend: BaseStatistic.getReportingNorm()", "statistics-service-implementation",
        "Regular vacation reduces BOTH personal and budget norm equally.",
    ],
    [
        "TC-STAT-078", "Personal norm — employee with admin (unpaid) vacation",
        "Employee with approved ADMINISTRATIVE vacation in selected month.",
        "1. Find employee with admin vacation\n   SQL: SELECT e.login FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id=e.id WHERE v.type='ADMINISTRATIVE' AND v.status='APPROVED'\n2. GET /v1/reports/summary → personalNorm, norm\n3. Verify personalNorm < norm (admin vacation reduces personal only)\n4. GET /v1/statistic/report/employees → verify monthNorm != budgetNorm",
        "personalNorm < budgetNorm. Admin vacation subtracted from personal norm but NOT from budget norm. budgetNorm = office norm minus regular absences only.",
        "High", "Data",
        "Confluence #3381; Backend: BaseStatistic.getBudgetNorm()", "statistics-service-implementation",
        "Key design issue: deviation formula uses budgetNorm, so admin vacation employees may appear as under-reporters even if they reported correctly against their personal norm.",
    ],
    [
        "TC-STAT-079", "Personal norm — employee with sick leave",
        "Employee with active/closed sick leave in selected month.",
        "1. Find employee with sick leave in Feb 2026\n   SQL: SELECT e.login, sl.start_date, sl.end_date, sl.work_days FROM ttt_vacation.sick_leave sl JOIN ttt_vacation.employee e ON sl.employee=e.id WHERE sl.status IN ('OPEN','CLOSED') AND sl.start_date <= '2026-02-28' AND sl.end_date >= '2026-02-01'\n2. Verify personalNorm reduced by sick leave working hours",
        "personalNorm reduced by sick leave working hours (work_days × 8 hours/day, clamped to month).",
        "Medium", "Data",
        "Backend: BaseStatistic.getReportingNorm()", "statistics-service-implementation",
        "Sick leave hours merged with other time-offs. Overlapping periods are de-duplicated.",
    ],
    [
        "TC-STAT-080", "Personal norm — overlapping absences merged",
        "Employee with overlapping vacation + sick leave in same period.",
        "1. Find employee with concurrent vacation and sick leave\n   SQL: complex join to find overlaps\n2. Verify norm reduction is NOT double-counted\n3. Overlapping days counted once",
        "Overlapping off-periods merged (union). Total reduction = merged period hours, not sum of individual reductions.",
        "Medium", "Data",
        "Backend: BaseStatistic — merge overlapping off-periods", "statistics-service-implementation",
        "Backend merges all off-periods before subtracting. Sick leave crossing vacation triggers separate notification.",
    ],
    [
        "TC-STAT-081", "Personal norm — clamped to employee work period",
        "Employee who started mid-month or terminated mid-month.",
        "1. Find employee with mid-month start\n   SQL: SELECT login, start_date FROM ttt_backend.employee WHERE EXTRACT(DAY FROM start_date) > 1 AND start_date >= '2025-01-01'\n2. Verify personalNorm reflects only working days from start_date to month end",
        "personalNorm clamped to [start_date, end_date] intersection with [monthStart, monthEnd]. Employee starting on 15th has ~half-month norm.",
        "High", "Data",
        "Backend: effectiveBounds; #3353, #3356", "statistics-service-implementation",
        "Related Sprint 15 tickets: #3353 (exclude pre-employment), #3356 (part-month).",
    ],
    [
        "TC-STAT-082", "Budget norm — admin vacation excluded from off-periods",
        "Employee with ADMINISTRATIVE vacation.",
        "1. Calculate personal norm (includes admin vacation subtraction)\n2. Calculate budget norm (excludes admin vacation)\n3. Verify budget_norm > month_norm for this employee in statistic_report",
        "budget_norm ≥ month_norm always. budget_norm = office norm minus non-admin absences. Admin vacation not subtracted from budget.",
        "High", "Data",
        "Backend: BaseStatistic.getBudgetNorm()", "statistics-service-implementation",
        "Cross-ref: 2.7% of employees have month_norm != budget_norm (part-time or admin vacation). SQL: SELECT employee_login, month_norm, budget_norm FROM ttt_backend.statistic_report WHERE month_norm != budget_norm",
    ],
    [
        "TC-STAT-083", "Deviation formula verification",
        "Database access. Multiple employees with varying reported/norm values.",
        "1. SELECT employee_login, reported_effort, budget_norm FROM ttt_backend.statistic_report WHERE report_date='2026-02-01' LIMIT 10\n2. For each: calculate (reported_effort*60 - budget_norm) / budget_norm * 100\n   Note: reported_effort in HOURS, budget_norm in MINUTES\n3. GET /v1/statistic/report/employees and compare deviation values",
        "API deviation % matches: (reported_hours × 60 - budget_norm_minutes) / budget_norm_minutes × 100. Handle budgetNorm=0 → N/A.",
        "High", "Data",
        "Confluence #3195", "statistics-service-implementation",
        "CRITICAL: reported_effort is in HOURS, budget_norm is in MINUTES in the DB. Conversion needed for formula.",
    ],
    [
        "TC-STAT-084", "budgetNorm null fallback to monthNorm",
        "Employee where budget_norm is NULL in statistic_report.",
        "1. SQL: SELECT employee_login FROM ttt_backend.statistic_report WHERE budget_norm IS NULL\n2. If found, verify API uses month_norm as fallback\n3. Verify deviation formula uses the fallback value",
        "When budget_norm is NULL, system falls back to month_norm for deviation calculation. Verify no NPE/error.",
        "Medium", "Data",
        "Code: budgetNorm null fallback", "statistics-service-implementation",
        "Known design issue. May produce incorrect deviation for employees with admin vacation if fallback triggers.",
    ],
    [
        "TC-STAT-085", "Norm cache — Caffeine 5min TTL",
        "API access to timemachine. Employee with a known norm value.",
        "1. GET /v1/reports/summary?login=X&date=2026-03-01 → note personalNorm\n2. [If mutations allowed] Create a vacation for employee X in March\n3. Immediately GET /v1/reports/summary again\n4. Verify personalNorm may NOT reflect change (cached)\n5. Wait 5 minutes, GET again\n6. Verify updated personalNorm",
        "Norm values cached for 5 minutes (Caffeine, max 1000 entries). Changes to absences may not reflect immediately.",
        "Low", "Data",
        "Backend: reportingNorm cache (TTL 5min)", "statistics-service-implementation",
        "Cache key includes employee login + period. Mutation-dependent test.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 5 — Access Control
# ═══════════════════════════════════════════════════════════════════
access_control = [
    [
        "TC-STAT-086", "ADMIN — full access to all statistics views",
        "Logged in as user with ADMIN role on timemachine.",
        "1. Navigate to /statistics/general → verify all available tabs\n2. Navigate to /statistics/employee-reports → verify page loads\n3. Verify all employees visible in Employee Reports\n4. GET /v1/statistic/permissions → verify full permission set",
        "ADMIN has access to General Statistics (all tabs) and Employee Reports (all employees). Full permission set returned by API.",
        "High", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-087", "CHIEF_ACCOUNTANT — full access scope",
        "Logged in as CHIEF_ACCOUNTANT.",
        "1. Verify General Statistics tabs available\n2. Verify Employee Reports shows all employees\n3. Compare employee count with ADMIN view",
        "Same scope as ADMIN: all employees visible in Employee Reports.",
        "Medium", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "Test data: identify CHIEF_ACCOUNTANT user in DB.",
    ],
    [
        "TC-STAT-088", "ACCOUNTANT — office-scoped access",
        "Logged in as ACCOUNTANT (specific salary office).",
        "1. Navigate to Employee Reports\n2. Count visible employees\n3. Verify all shown employees belong to user's salary office\n4. SQL: verify no employees from other offices appear",
        "Only employees from ACCOUNTANT's salary office(s) visible. Other office employees excluded.",
        "High", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "Test data: compare API result count with SQL: SELECT COUNT(*) FROM ttt_backend.employee WHERE salary_office_id IN (...)",
    ],
    [
        "TC-STAT-089", "DEPARTMENT_MANAGER — department-scoped access",
        "Logged in as DEPARTMENT_MANAGER.",
        "1. Navigate to Employee Reports\n2. Verify only subordinates visible\n3. Attempt to search for employees outside department\n4. Verify they don't appear",
        "Only department subordinates visible. Out-of-department employees not shown even when searched.",
        "High", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-090", "TECH_LEAD — subordinate-scoped access",
        "Logged in as TECH_LEAD.",
        "1. Navigate to Employee Reports\n2. Verify only assigned employees visible\n3. Count matches expected subordinate count",
        "Scope = own employees (possibly same as DM since DM = CS.user.Employees > 0 per requirements).",
        "Medium", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "Requirements note: TechLead scope 'possibly = DM' — verify exact behavior.",
    ],
    [
        "TC-STAT-091", "EMPLOYEE — restricted to 'My tasks' tab only",
        "Logged in as alsmirnov (EMPLOYEE role only).",
        "1. Navigate to /statistics/general\n2. Verify only 'My tasks' tab visible\n3. Verify no other tabs accessible\n4. GET /v1/statistic/permissions → verify limited permissions",
        "1 tab only: 'My tasks'. Permissions: minimal set (VIEW_MY_TASKS only).",
        "High", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "Verified in session 29 with alsmirnov: exactly 1 tab.",
    ],
    [
        "TC-STAT-092", "EMPLOYEE — cannot access Employee Reports via URL",
        "Logged in as alsmirnov. Directly navigate to /statistics/employee-reports.",
        "1. Navigate directly to /statistics/employee-reports\n2. Verify access denied (403 or redirect)\n3. GET /v1/statistic/report/employees via API\n4. Verify 403 response",
        "403 Forbidden or access denied both in UI and API.",
        "High", "Security",
        "Confluence 119244531", "statistics-service-implementation",
        "Both UI and API should enforce access control independently.",
    ],
    [
        "TC-STAT-093", "API key authentication — returns full permissions",
        "API access using API_SECRET_TOKEN header (no JWT).",
        "1. GET /v1/statistic/permissions with API_SECRET_TOKEN header\n2. Verify all permissions returned (not scoped to any role)",
        "All 5+ statistics permissions returned. API key grants full access.",
        "Medium", "Security",
        "Session 29 finding", "statistics-service-implementation",
        "API key auth bypasses role-based scoping. Verify this is intended behavior.",
    ],
    [
        "TC-STAT-094", "Employee Reports — OFFICE_HR access scope",
        "Logged in as OFFICE_HR user.",
        "1. Navigate to Employee Reports\n2. Verify scope matches assigned HR employees",
        "Scope: assigned HR employees only. Not all office employees.",
        "Medium", "Security",
        "Backend: access control switch-case", "statistics-service-implementation",
        "OFFICE_HR scope different from ACCOUNTANT (office-wide). HR has specific employee assignments.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 6 — Data & Cache
# ═══════════════════════════════════════════════════════════════════
data_cache = [
    [
        "TC-STAT-095", "statistic_report — nightly sync creates/updates records",
        "Database access to timemachine. Check sync timing.",
        "1. Note current statistic_report row count: SELECT COUNT(*) FROM ttt_backend.statistic_report\n2. Check last_updated_time for recent records: SELECT MAX(last_updated_time) FROM ttt_backend.statistic_report\n3. Verify sync runs at 4:00 AM (check ShedLock: SELECT * FROM ttt_backend.shedlock WHERE name LIKE '%statistic%')\n4. After next sync: compare row counts and timestamps",
        "Nightly sync at 4:00 AM updates current + previous month records. ShedLock prevents concurrent runs. New employees get new rows; removed employees get deleted rows.",
        "Medium", "Data",
        "Backend: StatisticReportScheduler", "statistics-service-implementation",
        "Test data: SELECT COUNT(*), report_date FROM ttt_backend.statistic_report GROUP BY report_date ORDER BY report_date DESC LIMIT 5. Expected ~2 months of data.",
    ],
    [
        "TC-STAT-096", "statistic_report — task report event triggers update",
        "API mutation access. Submit a task report and observe cache update.",
        "1. SELECT reported_effort FROM ttt_backend.statistic_report WHERE employee_login='X' AND report_date='2026-03-01'\n2. POST /v1/reports to add effort for employee X in March\n3. Wait 2-3 seconds (@Async, @TransactionalEventListener)\n4. Re-query statistic_report\n5. Verify reported_effort increased",
        "statistic_report.reported_effort updated within seconds of task report submission. Creates record if missing.",
        "Medium", "Data",
        "Backend: @TransactionalEventListener", "statistics-service-implementation",
        "Mutation test — requires allow_api_mutations=true. Event fires after transaction commit.",
    ],
    [
        "TC-STAT-097", "statistic_report — RabbitMQ vacation change event",
        "Vacation service change triggers norm recalculation.",
        "1. Note current month_norm for employee X\n2. Create/modify vacation for employee X (via vacation API)\n3. Monitor RabbitMQ delivery (or wait for async processing)\n4. Re-query statistic_report\n5. Verify month_norm updated",
        "month_norm and budget_norm recalculated after vacation service fires VACATION_CHANGES event. Updated values reflect new absence.",
        "Medium", "Data",
        "Backend: RabbitMQ listener", "statistics-service-implementation",
        "MQ event types: INITIAL_SYNC (deletes extras), VACATION_CHANGES, SICK_LEAVE_CHANGES (upsert only). Test at least VACATION_CHANGES.",
    ],
    [
        "TC-STAT-098", "statistic_report — data consistency with task_report",
        "Database access. Compare aggregated task_report with statistic_report.",
        "1. For a specific employee/month:\n   SQL: SELECT SUM(actual_efforts)/60.0 FROM ttt_backend.task_report WHERE executor_login='X' AND report_date BETWEEN '2026-02-01' AND '2026-02-28'\n2. SELECT reported_effort FROM ttt_backend.statistic_report WHERE employee_login='X' AND report_date='2026-02-01'\n3. Verify values match",
        "statistic_report.reported_effort (HOURS) = SUM(task_report.actual_efforts) (MINUTES) / 60. Values should match within rounding tolerance.",
        "High", "Data",
        "Session 11 verification", "statistics-service-implementation",
        "Verified in session 11: 152 hours = 9120 minutes. Run for multiple employees to confirm consistency.",
    ],
    [
        "TC-STAT-099", "statistic_report — 2-month sync window limitation",
        "Database access. Check historical data freshness.",
        "1. SELECT report_date, MAX(last_updated_time) FROM ttt_backend.statistic_report GROUP BY report_date ORDER BY report_date\n2. Verify only current and previous month have recent last_updated_time\n3. Older months have stale timestamps",
        "Only current + previous month updated by nightly sync. Older months retain historical values (no back-fill).",
        "Low", "Data",
        "Backend: 2-month sync window", "statistics-service-implementation",
        "Known design issue. If historical data is corrected (e.g., late report for 3 months ago), statistic_report won't reflect it until manual intervention.",
    ],
    [
        "TC-STAT-100", "statistic_report — terminated employee cleanup",
        "Employee who was terminated during a synced month.",
        "1. Find terminated employee: SELECT login, end_date FROM ttt_backend.employee WHERE end_date IS NOT NULL AND end_date >= '2026-01-01'\n2. Check statistic_report: SELECT * FROM ttt_backend.statistic_report WHERE employee_login='terminated_login'\n3. Verify: record exists for last active month, deleted for months after termination",
        "Nightly sync deletes statistic_report rows for removed employees. Record remains for historical months before termination.",
        "Medium", "Data",
        "Backend: INITIAL_SYNC MQ type", "statistics-service-implementation",
        "INITIAL_SYNC deletes extras. Verify cleanup happens correctly.",
    ],
    [
        "TC-STAT-101", "Race condition — concurrent MQ + task report event",
        "Requires timing control. Submit task report while vacation event processes.",
        "1. Identify employee X with existing statistic_report row\n2. Trigger vacation change for X (MQ event)\n3. Simultaneously submit task report for X\n4. After both complete, verify statistic_report consistency\n5. Check reported_effort and month_norm are both correct",
        "Both updates should succeed. However, no pessimistic locking — last-write-wins. Verify final values are consistent (not partially updated).",
        "Low", "Data/Negative",
        "Known design issue; BUG-STATS-1", "statistics-service-implementation",
        "Difficult to reproduce reliably. Race condition between @Async event handler and MQ consumer. Document actual behavior.",
    ],
    [
        "TC-STAT-102", "QA-1 environment — no statistic_report table (on-the-fly)",
        "API access to qa-1.",
        "1. GET /v1/statistic/report/employees on qa-1\n2. Verify data returned (computed on-the-fly from task_report)\n3. Compare with timemachine response for same employee/month\n4. Verify values match",
        "QA-1 computes statistics on-the-fly (no cache table). Results should match timemachine's cached values for same data.",
        "Low", "Data",
        "Session 11 finding", "statistics-service-implementation",
        "Performance difference only — no functional difference expected. May be slower on qa-1.",
    ],
    [
        "TC-STAT-103", "Hardcoded CEO login — special handling",
        "Database access. Check CEO_LOGIN = 'ilnitsky' behavior.",
        "1. Check if employee 'ilnitsky' exists: SELECT * FROM ttt_backend.employee WHERE login='ilnitsky'\n2. If exists, check statistics for CEO\n3. Verify no special treatment visible in UI/API (hardcoded login is internal optimization)",
        "CEO login 'ilnitsky' used internally in BaseStatistic for access control shortcuts. Should not affect displayed data.",
        "Low", "Data",
        "Code: CEO_LOGIN constant", "statistics-service-implementation",
        "Known hardcoded value. Verify behavior if CEO account is missing or renamed.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  SHEET 7 — Export & Individual Norm
# ═══════════════════════════════════════════════════════════════════
export_norm = [
    [
        "TC-STAT-104", "CSV export — column headers match spec",
        "API access. Any export endpoint.",
        "1. GET /v1/statistic/export/employees?startDate=2026-01-01&endDate=2026-03-31\n2. Parse first line (headers)\n3. Verify: EmployeeLogin, EmployeeName, Contractor, DepartmentManagerLogin, DepartmentManagerName, EffortForPeriod, EffortTotal, BeginDate, EndDate, NodeType",
        "CSV headers match ExportEmployeeNode structure. 10 columns in correct order.",
        "Medium", "API",
        "Code: StatisticExportController", "statistics-service-implementation",
        "10 export endpoints exist. Test at least /employees and /employees-largest-customers.",
    ],
    [
        "TC-STAT-105", "CSV export — data accuracy vs UI display",
        "API access + UI access. Same date range.",
        "1. In General Statistics UI, set date range 2026-01-01 to 2026-03-31\n2. Note values for 3 employees\n3. Download CSV via export\n4. Find same employees in CSV\n5. Compare EffortForPeriod/EffortTotal values",
        "CSV values match UI display. Both should be in HOURS. No conversion errors.",
        "High", "API/UI",
        "API surface map", "statistics-service-implementation",
        "Export uses HOURS by default. UI may show hours or days depending on toggle — compare in hours mode.",
    ],
    [
        "TC-STAT-106", "CSV export — Contractor flag accuracy",
        "API access. Mix of regular employees and contractors.",
        "1. GET /v1/statistic/export/employees?...\n2. Find Contractor column\n3. Verify true/false matches employee type\n4. SQL: SELECT login, is_contractor FROM ttt_backend.employee WHERE login IN ('...')",
        "Contractor column correctly reflects employee type from DB.",
        "Low", "API/Data",
        "Code: ExportEmployeeNode", "statistics-service-implementation",
        "",
    ],
    [
        "TC-STAT-107", "Google Sheets link — IMPORTDATA compatibility",
        "General Statistics UI. Google Sheets access.",
        "1. In General Statistics, click Export → 'Copy link for Google tables'\n2. In Google Sheets, use =IMPORTDATA(\"<pasted_link>\")\n3. Verify data loads correctly\n4. Verify column structure matches CSV",
        "Google Sheets IMPORTDATA successfully loads data from the link. Columns and data match CSV export.",
        "Medium", "UI/Integration",
        "Confluence 119244531", "frontend-statistics-module",
        "Link includes auth token — verify security implications. Link may expire.",
    ],
    [
        "TC-STAT-108", "#3400 Individual norm CSV export — endpoint exists",
        "API access to timemachine.",
        "1. Search for individual norm export endpoint: try GET /v1/statistic/export/individual-norm or similar\n2. If found, verify response format\n3. If NOT found, document as implementation gap",
        "If endpoint exists: CSV with columns login, name, surname, department_manager_login, salary_office, individual_norm. If missing: document gap (ticket marked Production Ready but code not in release/2.1).",
        "Medium", "API",
        "Ticket #3400", "statistics-service-implementation",
        "Known gap: code not found in release/2.1 codebase (session 26 analysis). May be in unmerged branch.",
    ],
    [
        "TC-STAT-109", "#3400 Individual norm CSV — column accuracy",
        "API access. If #3400 endpoint exists.",
        "1. GET individual norm export\n2. Verify columns: login, name, surname, department_manager_login, salary_office, individual_norm\n3. Cross-reference individual_norm with personalNorm from /v1/reports/summary\n4. Verify salary_office matches employee's actual office",
        "individual_norm = personalNorm (all absence types subtracted). salary_office matches DB employee.salary_office_id resolved to name.",
        "Medium", "API/Data",
        "Ticket #3400; Confluence #3381", "statistics-service-implementation",
        "Conditional: only if endpoint exists. individual_norm accounts for: vacations, sick leaves, transferred weekends/holidays.",
    ],
    [
        "TC-STAT-110", "Export — all 10 export endpoints return valid CSV",
        "API access to timemachine.",
        "1. Test each export endpoint:\n   /employees, /employees/projects, /employees/tasks, /departments, /tasks, /tasks/employees, /projects, /projects/employees, /employees-largest-customers, /task-bound-employees/tasks\n2. Verify each returns valid CSV with appropriate headers\n3. Verify no 500 errors",
        "All 10 endpoints return valid CSV responses. No server errors. Headers match endpoint-specific node types.",
        "Medium", "API",
        "Code: StatisticExportController (10 endpoints)", "statistics-service-implementation",
        "Bulk smoke test. Required params vary per endpoint — check controller for @RequestParam requirements.",
    ],
    [
        "TC-STAT-111", "Export — timeUnit parameter for largest-customers",
        "API access.",
        "1. GET /v1/statistic/export/employees-largest-customers?...&timeUnit=HOURS → note values\n2. GET same with timeUnit=MINUTES\n3. Verify MINUTES values = HOURS values × 60",
        "timeUnit param correctly converts output. HOURS default. MINUTES = HOURS × 60.",
        "Low", "API",
        "Session 11 finding", "statistics-service-implementation",
        "Configurable via timeUnit param. Verify on employees-largest-customers endpoint.",
    ],
]


# ═══════════════════════════════════════════════════════════════════
#  Generate workbook
# ═══════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

sheets = [
    ("General Statistics UI", general_stats_ui),
    ("Employee Reports UI", employee_reports_ui),
    ("Statistics API", statistics_api),
    ("Norm Calculation", norm_calculation),
    ("Access Control", access_control),
    ("Data & Cache", data_cache),
    ("Export & Norm CSV", export_norm),
]

total_cases = 0
for name, cases in sheets:
    write_sheet(wb, name, cases)
    total_cases += len(cases)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Total test cases: {total_cases}")
print("Sheets:", ", ".join(f"{name} ({len(cases)})" for name, cases in sheets))
