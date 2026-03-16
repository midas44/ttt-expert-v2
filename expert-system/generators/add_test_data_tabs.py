#!/usr/bin/env python3
"""Add 'Test Data' tab to all 10 XLSX workbooks.

Addresses Mission Directive requirement:
  "Must include description how to generate input test data
   (by database mining with criteria, random generation in given range,
   timestamp addition, static values etc.)"

Adds a structured "Test Data" tab to each workbook with:
  1. Test user matrix (login, roles, use case)
  2. Module-specific SQL queries for finding/creating test data
  3. Data generation strategies (timestamps, boundaries, random values)
  4. API-based test data setup sequences

Safe: only appends a new tab — never modifies existing tabs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling constants ─────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_CODE = Font(name="Consolas", size=9)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_CODE_BG = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_DATA = "BF8F00"  # Gold for test data tabs


# ── Common test user matrix ───────────────────────────────────

COMMON_USERS = [
    ("perekrest", "ACCOUNTANT, ADMIN, CHIEF_ACCOUNTANT, DEPT_MGR, EMPLOYEE, HR, PM",
     "Best multi-role user (7 roles). Use for cross-role permission testing."),
    ("pvaynmaster", "ACCOUNTANT, ADMIN, CHIEF_OFFICER, DEPT_MGR, EMPLOYEE, HR, PM",
     "Multi-role with CHIEF_OFFICER. Use for highest-privilege scenarios."),
    ("ilnitsky", "ACCOUNTANT, ADMIN, DEPT_MGR, EMPLOYEE, HR, PM",
     "Admin+accountant combo. Use for accounting and admin tests."),
    ("ann", "ACCOUNTANT, ADMIN, DEPT_MGR, EMPLOYEE, HR, PM",
     "Same role set as ilnitsky. Use as second admin/accountant."),
    ("juliet_nekhor", "DEPT_MGR, EMPLOYEE, HR, PM, VIEW_ALL",
     "Has VIEW_ALL. Use for read-only permission tests."),
    ("kcherenkov", "DEPT_MGR, EMPLOYEE, PM, TECH_LEAD",
     "Tech Lead + PM. Use for planner/project management tests."),
    ("alsmirnov", "EMPLOYEE",
     "EMPLOYEE-only. Use as minimum-privilege baseline user."),
]

# SQL to find single-role employees dynamically
COMMON_SQL = {
    "find_employee_only": (
        "Find EMPLOYEE-only users (minimum privilege)",
        "SELECT e.login, e.full_name\n"
        "FROM employee e\n"
        "JOIN employee_global_roles r ON e.id = r.employee_id\n"
        "WHERE e.enabled = true\n"
        "GROUP BY e.id\n"
        "HAVING COUNT(*) = 1 AND MAX(r.role) = 'ROLE_EMPLOYEE'\n"
        "LIMIT 5;"
    ),
    "find_contractors": (
        "Find active contractors",
        "SELECT e.login, e.full_name\n"
        "FROM employee e\n"
        "JOIN employee_global_roles r ON e.id = r.employee_id\n"
        "WHERE e.enabled = true AND r.role = 'ROLE_CONTRACTOR'\n"
        "LIMIT 5;"
    ),
    "find_managers": (
        "Find active project managers",
        "SELECT e.login, e.full_name\n"
        "FROM employee e\n"
        "JOIN employee_global_roles r ON e.id = r.employee_id\n"
        "WHERE e.enabled = true AND r.role = 'ROLE_PROJECT_MANAGER'\n"
        "LIMIT 5;"
    ),
    "find_accountants": (
        "Find active accountants",
        "SELECT e.login, e.full_name\n"
        "FROM employee e\n"
        "JOIN employee_global_roles r ON e.id = r.employee_id\n"
        "WHERE e.enabled = true AND r.role = 'ROLE_ACCOUNTANT'\n"
        "LIMIT 5;"
    ),
    "find_admins": (
        "Find active admins",
        "SELECT e.login, e.full_name\n"
        "FROM employee e\n"
        "JOIN employee_global_roles r ON e.id = r.employee_id\n"
        "WHERE e.enabled = true AND r.role = 'ROLE_ADMIN'\n"
        "LIMIT 5;"
    ),
    "current_periods": (
        "Get current report/approve periods per office",
        "SELECT o.id as office_id, o.name,\n"
        "       rp.start as report_period_start,\n"
        "       ap.start as approve_period_start\n"
        "FROM office o\n"
        "JOIN office_period rp ON o.report_period_id = rp.id\n"
        "JOIN office_period ap ON o.approve_period_id = ap.id\n"
        "ORDER BY o.id;"
    ),
}


# ── Module-specific test data ─────────────────────────────────

MODULE_DATA = {
    "vacation": {
        "title": "Vacation Module Test Data",
        "users": [
            ("Use perekrest or ilnitsky", "ACCOUNTANT+ADMIN", "For vacation payment and admin actions"),
            ("Use alsmirnov", "EMPLOYEE", "For basic vacation creation/view"),
            ("Use kcherenkov", "PM+DEPT_MGR", "For approval workflow testing"),
        ],
        "sql_queries": [
            ("Find employees with sufficient accrued days (AV=false offices)",
             "SELECT e.login, ev.days as available_days, o.name as office\n"
             "FROM ttt_vacation.employee e\n"
             "JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee_id\n"
             "JOIN ttt_vacation.office o ON e.office_id = o.id\n"
             "WHERE ev.days >= 5 AND o.advance_vacation = false\n"
             "AND e.login IN (SELECT login FROM ttt_backend.employee WHERE enabled = true)\n"
             "ORDER BY ev.days DESC LIMIT 10;"),
            ("Find employees in AV=true offices (advance vacation)",
             "SELECT e.login, o.name as office\n"
             "FROM ttt_vacation.employee e\n"
             "JOIN ttt_vacation.office o ON e.office_id = o.id\n"
             "WHERE o.advance_vacation = true\n"
             "AND e.login IN (SELECT login FROM ttt_backend.employee WHERE enabled = true)\n"
             "LIMIT 10;"),
            ("Find vacations by status for testing transitions",
             "SELECT v.id, v.login, v.status, v.start_date, v.end_date, v.payment_type\n"
             "FROM ttt_vacation.vacation v\n"
             "WHERE v.status IN ('NEW', 'APPROVED')\n"
             "AND v.start_date > CURRENT_DATE\n"
             "ORDER BY v.status, v.start_date LIMIT 20;"),
            ("Find employees with their approvers (for approval flow)",
             "SELECT e.login, m.login as manager_login\n"
             "FROM ttt_backend.employee e\n"
             "JOIN ttt_backend.employee m ON e.manager_id = m.id\n"
             "WHERE e.enabled = true AND m.enabled = true\n"
             "LIMIT 10;"),
            ("Find vacation days grouped by year (for balance testing)",
             "SELECT ev.employee_id, e.login, ev.year, ev.days\n"
             "FROM ttt_vacation.employee_vacation ev\n"
             "JOIN ttt_vacation.employee e ON ev.employee_id = e.id\n"
             "WHERE ev.year >= 2025\n"
             "ORDER BY e.login, ev.year;"),
        ],
        "strategies": [
            ("Start date generation", "Use CURRENT_DATE + 14 days (next Monday after 2-week buffer). "
             "Format: YYYY-MM-DD. Always pick a Monday for 5-day vacation."),
            ("End date generation", "Start date + 4 days (Friday) for standard 5-day vacation. "
             "For boundary: use 1-day (same as start), cross-month, cross-year."),
            ("Payment month", "First of the month containing startDate. Format: YYYY-MM-01."),
            ("Boundary values", "0 days (expect error), 1 day, max allowed days, days > available balance."),
        ],
    },

    "sick-leave": {
        "title": "Sick Leave Module Test Data",
        "users": [
            ("Use alsmirnov", "EMPLOYEE", "Create sick leaves as regular employee"),
            ("Use kcherenkov", "PM", "View team sick leaves as manager"),
            ("Use ilnitsky", "ACCOUNTANT", "Process sick leave accounting status"),
        ],
        "sql_queries": [
            ("Find employees without active sick leaves (for create tests)",
             "SELECT e.login\nFROM ttt_backend.employee e\n"
             "WHERE e.enabled = true\n"
             "AND e.login NOT IN (\n"
             "  SELECT sl.login FROM ttt_vacation.sick_leave sl\n"
             "  WHERE sl.status = 'OPEN' AND sl.end_date >= CURRENT_DATE\n"
             ")\nLIMIT 10;"),
            ("Find OPEN sick leaves (for close/accounting tests)",
             "SELECT sl.id, sl.login, sl.start_date, sl.end_date, sl.status, sl.accounting_status\n"
             "FROM ttt_vacation.sick_leave sl\n"
             "WHERE sl.status = 'OPEN'\n"
             "ORDER BY sl.start_date DESC;"),
            ("Find sick leaves by accounting status (for status transition tests)",
             "SELECT sl.id, sl.login, sl.status, sl.accounting_status, sl.number\n"
             "FROM ttt_vacation.sick_leave sl\n"
             "WHERE sl.accounting_status IN ('NEW', 'PAID')\n"
             "ORDER BY sl.accounting_status, sl.id;"),
            ("Count sick leaves with attachments",
             "SELECT sl.id, sl.login, COUNT(a.id) as attachment_count\n"
             "FROM ttt_vacation.sick_leave sl\n"
             "LEFT JOIN ttt_vacation.sick_leave_attachment a ON sl.id = a.sick_leave_id\n"
             "GROUP BY sl.id, sl.login\n"
             "HAVING COUNT(a.id) > 0\n"
             "LIMIT 10;"),
        ],
        "strategies": [
            ("Date range", "startDate: CURRENT_DATE - 3 days, endDate: CURRENT_DATE. "
             "Sick leaves are typically retroactive."),
            ("Number field", "Required for CLOSE. Use format: 'SL-YYYY-NNN' or any unique string."),
            ("Force flag", "force=false for normal path; force=true to bypass overlap checks."),
            ("Boundary values", "Same day (1-day), cross-month, overlapping existing sick leave."),
        ],
    },

    "day-off": {
        "title": "Day-Off Module Test Data",
        "users": [
            ("Use alsmirnov", "EMPLOYEE", "Create day-off requests"),
            ("Use kcherenkov", "PM+DEPT_MGR", "Approve/reject day-off requests"),
            ("Use ilnitsky", "HR", "View all day-offs as HR"),
        ],
        "sql_queries": [
            ("Find upcoming public holidays for day-off source dates",
             "SELECT cd.date, cd.day_type, c.name as calendar_name\n"
             "FROM ttt_calendar.calendar_day cd\n"
             "JOIN ttt_calendar.calendar c ON cd.calendar_id = c.id\n"
             "WHERE cd.day_type = 'PUBLIC_HOLIDAY' AND cd.date > CURRENT_DATE\n"
             "ORDER BY cd.date LIMIT 20;"),
            ("Find employees with day-off requests in NEW status",
             "SELECT edr.id, edr.employee_login, edr.status, edr.original_date, edr.personal_date\n"
             "FROM ttt_vacation.employee_dayoff_request edr\n"
             "WHERE edr.status = 'NEW'\n"
             "ORDER BY edr.original_date;"),
            ("Find employees with calendar-based day-offs",
             "SELECT ed.employee_id, e.login, ed.date, ed.duration\n"
             "FROM ttt_vacation.employee_dayoff ed\n"
             "JOIN ttt_vacation.employee e ON ed.employee_id = e.id\n"
             "WHERE ed.date > CURRENT_DATE\n"
             "ORDER BY ed.date LIMIT 20;"),
        ],
        "strategies": [
            ("Public date (originalDate)", "Must be a PUBLIC_HOLIDAY in employee's office calendar. "
             "Query calendar_day table to find valid dates."),
            ("Personal date", "Must be a WORKDAY in employee's calendar. "
             "Pick a weekday within same month as public date."),
            ("Duration", "8 (full day) or 4 (half day). Half-day support depends on office config."),
            ("Boundary values", "Same-day reschedule, cross-month, date in closed period, "
             "conflict with existing vacation."),
        ],
    },

    "reports": {
        "title": "Reports Module Test Data",
        "users": [
            ("Use alsmirnov", "EMPLOYEE", "Create/edit reports as regular employee"),
            ("Use kcherenkov", "PM", "Approve/reject reports as project manager"),
            ("Use juliet_nekhor", "VIEW_ALL", "View all reports read-only"),
        ],
        "sql_queries": [
            ("Find employees with active project/task assignments",
             "SELECT e.login, p.name as project_name, t.name as task_name, t.id as task_id\n"
             "FROM ttt_backend.employee e\n"
             "JOIN ttt_backend.employee_project ep ON e.id = ep.employee_id\n"
             "JOIN ttt_backend.project p ON ep.project_id = p.id\n"
             "JOIN ttt_backend.task t ON t.project_id = p.id\n"
             "WHERE e.enabled = true AND p.status = 'ACTIVE'\n"
             "AND e.login = 'alsmirnov'\n"
             "LIMIT 10;"),
            ("Find valid task IDs for report creation",
             "SELECT t.id, t.name, p.name as project_name\n"
             "FROM ttt_backend.task t\n"
             "JOIN ttt_backend.project p ON t.project_id = p.id\n"
             "JOIN ttt_backend.employee_project ep ON p.id = ep.project_id\n"
             "JOIN ttt_backend.employee e ON ep.employee_id = e.id\n"
             "WHERE e.enabled = true AND p.status = 'ACTIVE'\n"
             "AND e.login = '<target_login>'\n"
             "LIMIT 10;"),
            ("Find reports in REPORTED state (for approve/reject tests)",
             "SELECT tr.id, tr.executor_login, tr.task_id, tr.report_date, tr.effort, tr.state\n"
             "FROM ttt_backend.task_report tr\n"
             "WHERE tr.state = 'REPORTED'\n"
             "AND tr.report_date >= '2026-03-01'\n"
             "ORDER BY tr.report_date DESC LIMIT 20;"),
            ("Check if report period is open for target date",
             "SELECT o.id as office_id, o.name, rp.start as report_period_start\n"
             "FROM ttt_backend.office o\n"
             "JOIN ttt_backend.office_period rp ON o.report_period_id = rp.id\n"
             "WHERE rp.start <= CURRENT_DATE;"),
            ("Find employees with over-reported hours",
             "SELECT tr.executor_login, SUM(tr.effort) as total_minutes,\n"
             "       ROUND(SUM(tr.effort) / 60.0, 1) as total_hours\n"
             "FROM ttt_backend.task_report tr\n"
             "WHERE tr.report_date = CURRENT_DATE AND tr.state != 'DELETED'\n"
             "GROUP BY tr.executor_login\n"
             "HAVING SUM(tr.effort) > 480\n"
             "ORDER BY total_minutes DESC;"),
        ],
        "strategies": [
            ("reportDate", "Must fall within OPEN report period. Use CURRENT_DATE or dates "
             "within current period. Format: YYYY-MM-DD."),
            ("effort", "In MINUTES. Standard: 60 (1h), 480 (8h). Boundary: 1 (min), "
             "1440 (24h — system max), 0 (expect error)."),
            ("taskId", "Must be a valid task ID in a project the employee is assigned to. "
             "Query employee_project + task tables."),
            ("Timestamp generation", "Use ISO format: YYYY-MM-DDTHH:mm:ss. "
             "Reports use date-only (YYYY-MM-DD)."),
        ],
    },

    "accounting": {
        "title": "Accounting Module Test Data",
        "users": [
            ("Use ilnitsky or ann", "ACCOUNTANT+ADMIN", "Period management and vacation payment"),
            ("Use perekrest", "CHIEF_ACCOUNTANT", "Chief accountant operations"),
        ],
        "sql_queries": [
            ("Find offices with current period state",
             "SELECT o.id, o.name,\n"
             "       rp.start as report_start, ap.start as approve_start\n"
             "FROM ttt_backend.office o\n"
             "JOIN ttt_backend.office_period rp ON o.report_period_id = rp.id\n"
             "JOIN ttt_backend.office_period ap ON o.approve_period_id = ap.id\n"
             "ORDER BY o.id;"),
            ("Find APPROVED vacations available for payment",
             "SELECT v.id, v.login, v.start_date, v.end_date, v.payment_type, v.days\n"
             "FROM ttt_vacation.vacation v\n"
             "WHERE v.status = 'APPROVED'\n"
             "ORDER BY v.start_date LIMIT 20;"),
            ("Find employees with vacation day corrections",
             "SELECT vdc.id, vdc.employee_login, vdc.days, vdc.year, vdc.reason\n"
             "FROM ttt_vacation.vacation_day_correction vdc\n"
             "ORDER BY vdc.id DESC LIMIT 20;"),
            ("Find office norm data for period validation",
             "SELECT sr.employee_login, sr.date, sr.norm, sr.reported\n"
             "FROM ttt_backend.statistic_report sr\n"
             "WHERE sr.date >= '2026-03-01'\n"
             "ORDER BY sr.employee_login, sr.date LIMIT 30;"),
        ],
        "strategies": [
            ("Period dates", "Format: YYYY-MM-01 (always 1st of month). "
             "Advance: current start + 1 month. Revert: current start - 1 month."),
            ("officeId", "Use offices 2 (Saturn), 10 (Venus), 4 (Jupiter) — largest offices. "
             "Avoid office 9 (stuck at 2020-03-01)."),
            ("Payment validation", "Payment requires: status=APPROVED, startDate in approved period, "
             "5 business checks pass. Use APPROVED vacations from query above."),
            ("Day correction", "days field accepts positive (add) and negative (subtract). "
             "Boundary: 0 (no-op), large negative > available (expect error)."),
        ],
    },

    "admin": {
        "title": "Admin Panel Test Data",
        "users": [
            ("Use perekrest or ilnitsky", "ADMIN", "Full admin panel access"),
            ("Use alsmirnov", "EMPLOYEE", "Verify no admin access (negative test)"),
        ],
        "sql_queries": [
            ("Find active projects (for project admin tests)",
             "SELECT p.id, p.name, p.status, p.customer_name, p.type\n"
             "FROM ttt_backend.project p\n"
             "WHERE p.status = 'ACTIVE'\n"
             "ORDER BY p.name LIMIT 20;"),
            ("Find employees by enabled status",
             "SELECT e.login, e.full_name, e.enabled, e.salary_office\n"
             "FROM ttt_backend.employee e\n"
             "ORDER BY e.enabled DESC, e.login LIMIT 20;"),
            ("Find production calendars and their day counts",
             "SELECT c.id, c.name, c.country_code, COUNT(cd.id) as day_count\n"
             "FROM ttt_calendar.calendar c\n"
             "LEFT JOIN ttt_calendar.calendar_day cd ON c.id = cd.calendar_id\n"
             "GROUP BY c.id, c.name, c.country_code\n"
             "ORDER BY c.name;"),
            ("Find projects with PM Tool links",
             "SELECT p.id, p.name, p.pm_tool_id\n"
             "FROM ttt_backend.project p\n"
             "WHERE p.pm_tool_id IS NOT NULL AND p.status = 'ACTIVE'\n"
             "LIMIT 10;"),
            ("List office-to-calendar mappings",
             "SELECT o.id as office_id, o.name as office_name, c.name as calendar_name\n"
             "FROM ttt_vacation.office o\n"
             "LEFT JOIN ttt_calendar.calendar c ON o.calendar_id = c.id\n"
             "ORDER BY o.id;"),
        ],
        "strategies": [
            ("Project creation", "Name: 'Test-' + timestamp (e.g. Test-20260315). "
             "Type: COMMERCIAL. Status: defaults to ACTIVE."),
            ("Employee search", "Use login prefix for search filter. "
             "Active employees: ~410. Disabled: ~1431."),
            ("Calendar events", "Calendar day types: WORKDAY, PUBLIC_HOLIDAY, SHORTENED_DAY, WEEKEND. "
             "Date format: YYYY-MM-DD."),
        ],
    },

    "statistics": {
        "title": "Statistics Module Test Data",
        "users": [
            ("Use alsmirnov", "EMPLOYEE", "See only 'My Tasks' tab (1 tab)"),
            ("Use kcherenkov", "PM+DEPT_MGR", "See My Tasks + My Projects + Department (3 tabs)"),
            ("Use perekrest", "ADMIN+ACCOUNTANT", "See all 13 tabs including Office/Accounting"),
            ("Use juliet_nekhor", "VIEW_ALL", "See all tabs in read-only mode"),
        ],
        "sql_queries": [
            ("Find statistic_report data for a given month",
             "SELECT sr.employee_login, sr.date, sr.norm, sr.reported,\n"
             "       sr.reported - sr.norm as deviation\n"
             "FROM ttt_backend.statistic_report sr\n"
             "WHERE sr.date = '2026-03-01'\n"
             "ORDER BY sr.employee_login LIMIT 30;"),
            ("Find employees with over/under-reporting",
             "SELECT sr.employee_login, sr.norm, sr.reported,\n"
             "       CASE WHEN sr.reported > sr.norm THEN 'OVER'\n"
             "            WHEN sr.reported < sr.norm THEN 'UNDER'\n"
             "            ELSE 'EXACT' END as deviation_type\n"
             "FROM ttt_backend.statistic_report sr\n"
             "WHERE sr.date = '2026-03-01' AND sr.reported != sr.norm\n"
             "ORDER BY ABS(sr.reported - sr.norm) DESC LIMIT 20;"),
            ("Find offices with report period data (for office tab)",
             "SELECT o.id, o.name, rp.start as current_period\n"
             "FROM ttt_backend.office o\n"
             "JOIN ttt_backend.office_period rp ON o.report_period_id = rp.id\n"
             "ORDER BY o.name;"),
            ("Tab visibility by role (reference)",
             "-- Role → visible tabs mapping:\n"
             "-- EMPLOYEE: My Tasks (1 tab)\n"
             "-- PM: My Tasks, My Projects (2 tabs)\n"
             "-- DEPT_MGR: + Department (3 tabs)\n"
             "-- HR: + Department (varies)\n"
             "-- ACCOUNTANT: + Office, Accounting tabs\n"
             "-- ADMIN/VIEW_ALL: All 13 tabs"),
        ],
        "strategies": [
            ("Date filter", "Format: YYYY-MM-01 (always 1st of month). "
             "Statistics aggregate by month."),
            ("Employee filter", "Use login or name substring. API: searchText query param."),
            ("Export testing", "CSV export available for Employee Reports tab. "
             "Verify column headers, decimal separators, encoding (UTF-8 BOM)."),
        ],
    },

    "security": {
        "title": "Security Module Test Data",
        "users": [
            ("Use perekrest", "7 roles", "Maximum privilege — JWT contains all authorities"),
            ("Use alsmirnov", "EMPLOYEE only", "Minimum privilege baseline"),
            ("Use any contractor", "CONTRACTOR", "Special permissions: cannot create vacations"),
        ],
        "sql_queries": [
            ("Find users by role count (for authorization boundary tests)",
             "SELECT e.login, COUNT(r.role) as role_count,\n"
             "       STRING_AGG(r.role, ', ') as roles\n"
             "FROM ttt_backend.employee e\n"
             "JOIN ttt_backend.employee_global_roles r ON e.id = r.employee_id\n"
             "WHERE e.enabled = true\n"
             "GROUP BY e.login\n"
             "ORDER BY role_count DESC LIMIT 15;"),
            ("Find endpoints missing @PreAuthorize (code-level)",
             "-- Grep codebase for controllers without @PreAuthorize:\n"
             "-- grep -rn '@RestController' --include='*.java' | \\\n"
             "--   xargs -I{} grep -L '@PreAuthorize' {}\n"
             "-- Known gaps: EmployeeController in admin module"),
            ("Find API token configuration",
             "-- API token is configured in application properties:\n"
             "-- ttt.api.secret.token=<value>\n"
             "-- Header: API_SECRET_TOKEN\n"
             "-- Grants ROLE_API_USER scope (limited endpoints)"),
            ("Active session check",
             "-- JWT tokens are stateless (no server-side session table)\n"
             "-- Token expiry: check 'exp' claim in decoded JWT\n"
             "-- Decode: echo '<jwt_token>' | cut -d. -f2 | base64 -d | jq ."),
        ],
        "strategies": [
            ("JWT token extraction", "Login via CAS SSO. Extract TTT_JWT_TOKEN cookie "
             "from browser dev tools or Playwright. Decode with base64."),
            ("API token", "Set header API_SECRET_TOKEN with value from env config. "
             "Test both valid and invalid/expired tokens."),
            ("Permission boundary", "For each endpoint: test as authorized role (expect 200), "
             "test as unauthorized role (expect 403), test as unauthenticated (expect 401)."),
            ("IDOR testing", "Replace {id} in URL with another user's resource ID. "
             "Expect 403 or filtered response (not 200 with other user's data)."),
        ],
    },

    "cross-service": {
        "title": "Cross-Service Integration Test Data",
        "users": [
            ("Use ilnitsky", "ADMIN+ACCOUNTANT", "Trigger cross-service operations"),
            ("Use perekrest", "CHIEF_ACCOUNTANT", "Period management that cascades"),
        ],
        "sql_queries": [
            ("Compare employee office across services (divergence check)",
             "SELECT be.login, be.salary_office as ttt_backend_office,\n"
             "       ve.office_id as ttt_vacation_office\n"
             "FROM ttt_backend.employee be\n"
             "JOIN ttt_vacation.employee ve ON be.login = ve.login\n"
             "WHERE be.salary_office != ve.office_id\n"
             "AND be.enabled = true\n"
             "LIMIT 20;"),
            ("Check RabbitMQ exchange bindings (requires admin API)",
             "-- RabbitMQ Management: http://<rmq-host>:15672\n"
             "-- Key exchanges: calendar.event, office.period.changed,\n"
             "-- vacation.event, employee.sync.event\n"
             "-- No DLQ configured for any exchange"),
            ("Verify CS sync timestamp",
             "SELECT e.login, e.cs_updated_at, e.salary_office\n"
             "FROM ttt_backend.employee e\n"
             "WHERE e.enabled = true\n"
             "ORDER BY e.cs_updated_at DESC LIMIT 10;"),
        ],
        "strategies": [
            ("CS sync trigger", "POST /api/ttt/test/sync/employees — triggers full employee sync. "
             "Monitor employee table for changes."),
            ("Calendar event cascade", "Modify calendar day via API, observe vacation recalculation "
             "in ttt_vacation schema."),
            ("Period cascade", "Advance report period → verify statistic_report update "
             "and vacation day recalculation."),
        ],
    },

    "planner": {
        "title": "Planner Module Test Data",
        "users": [
            ("Use kcherenkov", "PM+TECH_LEAD", "Create/edit assignments as PM"),
            ("Use perekrest", "ADMIN+PM", "Full planner access with admin privileges"),
            ("Use alsmirnov", "EMPLOYEE", "View own assignments only"),
        ],
        "sql_queries": [
            ("Find employees with generated assignments",
             "SELECT ta.employee_login, COUNT(*) as assignment_count,\n"
             "       MIN(ta.date) as earliest, MAX(ta.date) as latest\n"
             "FROM ttt_backend.task_assignment ta\n"
             "GROUP BY ta.employee_login\n"
             "ORDER BY assignment_count DESC LIMIT 10;"),
            ("Find assignment details for specific employee",
             "SELECT ta.id, ta.employee_login, ta.task_id, ta.date,\n"
             "       ta.position, ta.next_assignment_id, ta.is_locked\n"
             "FROM ttt_backend.task_assignment ta\n"
             "WHERE ta.employee_login = '<target_login>'\n"
             "AND ta.date >= CURRENT_DATE\n"
             "ORDER BY ta.date, ta.position;"),
            ("Find projects with close tags",
             "SELECT ct.id, ct.name, ct.project_id, p.name as project_name\n"
             "FROM ttt_backend.close_tag ct\n"
             "JOIN ttt_backend.project p ON ct.project_id = p.id\n"
             "WHERE p.status = 'ACTIVE'\n"
             "LIMIT 10;"),
            ("Find active projects for generate endpoint",
             "SELECT p.id, p.name, COUNT(ep.employee_id) as member_count\n"
             "FROM ttt_backend.project p\n"
             "JOIN ttt_backend.employee_project ep ON p.id = ep.project_id\n"
             "JOIN ttt_backend.employee e ON ep.employee_id = e.id\n"
             "WHERE p.status = 'ACTIVE' AND e.enabled = true\n"
             "GROUP BY p.id, p.name\n"
             "HAVING COUNT(ep.employee_id) > 1\n"
             "ORDER BY member_count DESC LIMIT 10;"),
        ],
        "strategies": [
            ("Assignment dates", "Format: YYYY-MM-DD. Use CURRENT_DATE for today, "
             "+7 days for next week. Assignments are per-day."),
            ("Generate endpoint", "POST /v1/assignments/generate with projectId + date range. "
             "Creates assignments for all project members."),
            ("Position/ordering", "Integer position field. Lower = higher priority. "
             "Use 0, 1, 2... for ordered assignments."),
            ("WebSocket monitoring", "Subscribe to /topic/assignments after generate/reorder "
             "to verify real-time updates."),
        ],
    },
}


# ── Main logic ────────────────────────────────────────────────

def write_section_header(ws, row, title, num_cols):
    """Write a section header spanning all columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = ALIGN_LEFT
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def write_table_header(ws, row, headers, fill=None):
    """Write a table header row."""
    f = fill or FILL_HEADER
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = f
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    return row + 1


def write_data_row(ws, row, values, is_even=False):
    """Write a data row with alternating colors."""
    fill = FILL_ROW_EVEN if is_even else FILL_ROW_ODD
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_BODY
        cell.fill = fill
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
    return row + 1


def write_code_row(ws, row, label, code, num_cols):
    """Write a label + code block row."""
    ws.cell(row=row, column=1, value=label).font = FONT_BODY
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=2, value=code)
    cell.font = FONT_CODE
    cell.fill = FILL_CODE_BG
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for col in range(2, num_cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def add_test_data_tab(wb, module_key):
    """Add a Test Data tab to the workbook."""
    data = MODULE_DATA[module_key]
    num_cols = 3

    ws = wb.create_sheet(title="Test Data", index=0)  # Insert as first tab, will be moved
    ws.sheet_properties.tabColor = TAB_COLOR_DATA

    # Title
    ws.merge_cells("A1:C1")
    cell = ws.cell(row=1, column=1, value=data["title"])
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT

    # Back link
    cell = ws.cell(row=2, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"

    row = 4

    # ── Section 1: Recommended Test Users ──
    row = write_section_header(ws, row, "1. Recommended Test Users (Module-Specific)", num_cols)
    row = write_table_header(ws, row, ["User / Query", "Roles", "Use Case"])
    for i, (user, roles, use_case) in enumerate(data["users"]):
        row = write_data_row(ws, row, [user, roles, use_case], is_even=i % 2 == 0)
    row += 1

    # ── Section 2: Common Test Users (All Modules) ──
    row = write_section_header(ws, row, "2. Common Multi-Role Test Users (All Modules)", num_cols)
    row = write_table_header(ws, row, ["Login", "Roles", "Best For"])
    for i, (login, roles, best_for) in enumerate(COMMON_USERS):
        row = write_data_row(ws, row, [login, roles, best_for], is_even=i % 2 == 0)
    row += 1

    # ── Section 3: Common SQL Queries ──
    row = write_section_header(ws, row, "3. Common SQL Queries (User Discovery)", num_cols)
    row = write_table_header(ws, row, ["Purpose", "SQL Query", "Database"])
    for i, (key, (purpose, sql)) in enumerate(COMMON_SQL.items()):
        db = "ttt_backend" if "employee" in sql.lower() else "ttt"
        row = write_data_row(ws, row, [purpose, sql, db], is_even=i % 2 == 0)
        # Make SQL column use code font
        ws.cell(row=row - 1, column=2).font = FONT_CODE
        ws.cell(row=row - 1, column=2).fill = FILL_CODE_BG
    row += 1

    # ── Section 4: Module-Specific SQL Queries ──
    row = write_section_header(ws, row, "4. Module-Specific SQL Queries", num_cols)
    row = write_table_header(ws, row, ["Purpose", "SQL Query", "Notes"])
    for i, (purpose, sql) in enumerate(data["sql_queries"]):
        row = write_data_row(ws, row, [purpose, sql, "Run on timemachine/qa-1 DB"], is_even=i % 2 == 0)
        ws.cell(row=row - 1, column=2).font = FONT_CODE
        ws.cell(row=row - 1, column=2).fill = FILL_CODE_BG
    row += 1

    # ── Section 5: Data Generation Strategies ──
    row = write_section_header(ws, row, "5. Data Generation Strategies", num_cols)
    row = write_table_header(ws, row, ["Parameter", "Strategy", "Notes"])
    for i, (param, strategy) in enumerate(data["strategies"]):
        row = write_data_row(ws, row, [param, strategy, ""], is_even=i % 2 == 0)
    row += 1

    # ── Section 6: Environment Reference ──
    row = write_section_header(ws, row, "6. Environment Reference", num_cols)
    envs = [
        ("timemachine", "https://ttt-timemachine.noveogroup.com", "Primary dev env — has test clock (PATCH to advance time)"),
        ("qa-1", "https://ttt-qa-1.noveogroup.com", "Secondary dev env — real clock"),
        ("stage", "https://ttt-stage.noveogroup.com", "Production-like env — read-only testing preferred"),
    ]
    row = write_table_header(ws, row, ["Environment", "URL", "Notes"])
    for i, (name, url, notes) in enumerate(envs):
        row = write_data_row(ws, row, [name, url, notes], is_even=i % 2 == 0)
    row += 1

    # ── Section 7: API Authentication ──
    row = write_section_header(ws, row, "7. API Authentication", num_cols)
    auth_info = [
        ("JWT (user context)", "Login via CAS SSO at /login, extract TTT_JWT_TOKEN cookie. "
         "Pass as Authorization: Bearer <token> or Cookie header.", "Most endpoints"),
        ("API token", "Set header API_SECRET_TOKEN: <token_value>. "
         "Value from env config file.", "Test/sync endpoints, limited scope"),
    ]
    row = write_table_header(ws, row, ["Method", "How to Obtain/Use", "Scope"])
    for i, (method, how, scope) in enumerate(auth_info):
        row = write_data_row(ws, row, [method, how, scope], is_even=i % 2 == 0)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 40

    # Move Test Data tab to end (after all TS- tabs)
    sheets = wb.sheetnames
    wb.move_sheet("Test Data", offset=len(sheets) - 1)

    return ws


def process_workbook(area, xlsx_path):
    """Add Test Data tab to a workbook."""
    if area not in MODULE_DATA:
        print(f"  SKIP {area} — no module data defined")
        return False

    print(f"  Processing {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path)

    # Remove existing Test Data tab if present (for re-runs)
    if "Test Data" in wb.sheetnames:
        del wb["Test Data"]

    add_test_data_tab(wb, area)

    # Add hyperlink from Plan Overview to Test Data tab
    if "Plan Overview" in wb.sheetnames:
        plan_ws = wb["Plan Overview"]
        # Find last row with content
        last_row = plan_ws.max_row + 2
        cell = plan_ws.cell(row=last_row, column=1,
                           value="Test Data Reference")
        cell.font = FONT_LINK_BOLD
        cell.hyperlink = "#'Test Data'!A1"

    wb.save(xlsx_path)
    print(f"  DONE: {area} — Test Data tab added ({len(wb.sheetnames)} tabs total)")
    return True


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    workbooks = {
        "vacation": os.path.join(base_dir, "vacation", "vacation.xlsx"),
        "sick-leave": os.path.join(base_dir, "sick-leave", "sick-leave.xlsx"),
        "day-off": os.path.join(base_dir, "day-off", "day-off.xlsx"),
        "reports": os.path.join(base_dir, "reports", "reports.xlsx"),
        "accounting": os.path.join(base_dir, "accounting", "accounting.xlsx"),
        "admin": os.path.join(base_dir, "admin", "admin.xlsx"),
        "statistics": os.path.join(base_dir, "statistics", "statistics.xlsx"),
        "security": os.path.join(base_dir, "security", "security.xlsx"),
        "cross-service": os.path.join(base_dir, "cross-service", "cross-service.xlsx"),
        "planner": os.path.join(base_dir, "planner", "planner.xlsx"),
    }

    print("Adding Test Data tabs to all workbooks...")
    success = 0
    for area, path in workbooks.items():
        if not os.path.exists(path):
            print(f"  SKIP {area} — file not found: {path}")
            continue
        if process_workbook(area, path):
            success += 1

    print(f"\nComplete: {success}/{len(workbooks)} workbooks updated.")


if __name__ == "__main__":
    main()
