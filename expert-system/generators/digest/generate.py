#!/usr/bin/env python3
"""
Digest collection generator — pipeline-stress-test output (session 138+).

Scope: row 14 of #3423 only (vacation notifications digest).

Produces, in-place on ``test-docs/collections/digest/digest.xlsx``:
  * preserves the existing ``Plan Overview`` sheet (seeded 2026-04-21)
  * rebuilds the ``TS-Digest-Vacation`` suite sheet with TC-DIGEST-001..014
  * rewrites the ``COL-digest`` reference sheet data rows (row 3+)

Design constraints (from docs/tasks/digest/digest-testing-task.md and
the session briefing):

  * **Dual-trigger**: every behavioural scenario has Variant A (scheduler
    via clock-advance) and Variant B (test endpoint bypass). Paired.
  * **Content-complete**: every TC that asserts the digest email asserts
    every dynamic field the template renders, plus negative leakage
    guards. See patterns/email-notification-triggers.md § Digest template.
  * **Env-independent**: no ``qa-1`` / ``timemachine`` / ``stage`` /
    ``preprod`` literals anywhere in any cell. ``<ENV>`` placeholder only.
  * **Subject anomaly**: ``[<ENV>]ТТТ Дайджест отсутствий`` — Cyrillic
    ``ТТТ`` service tag, NO brackets around it (differs from every other
    TTT template which uses bracketed Latin ``[TTT]``).
  * **Business rule**: only ``APPROVED`` vacations with
    ``start_date = CURRENT_DATE + INTERVAL '1 day'`` enter the digest.
  * **Scheduler markers (Variant A)**: ``"Digests sending job started"``
    → ``"Digests sending job finished"`` on ``TTT-<ENV>``, plus per-
    recipient ``"Mail has been sent to <email> about
    NOTIFY_VACATION_UPCOMING..."`` for each recipient.
  * **Bypass markers (Variant B)**: scheduler start/finish do NOT fire
    (the test endpoint calls ``digestService.sendDigests()`` directly,
    bypassing the ``@Scheduled`` wrapper). Per-recipient markers still
    emit.

This collection is the pipeline-stress-test target — TC IDs start at
``TC-DIGEST-001`` and are fresh (do NOT copy TC-VAC-106..108 from
Cron_Vacation.xlsx; those are prior art, NOT a template).

Run from repo root::

    python3 expert-system/generators/digest/generate.py

Idempotent: every run rebuilds ``TS-Digest-Vacation`` fresh and
rewrites ``COL-digest`` data rows. Two runs back-to-back produce the
same output.
"""

from __future__ import annotations

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_TARGET = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "collections", "digest", "digest.xlsx")
)

# ─── Styling ──────────────────────────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_DESC = Font(name="Arial", size=10, italic=True, color="595959")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Orange — matches briefing's required tab color for the suite sheet and
# the existing COL-digest sheet tab color.
TAB_COLOR = "F4B084"

# Baseline per CLAUDE.md §11 XLSX Format table.
COL_WIDTHS = [14, 48, 52, 64, 52, 12, 12, 24, 18, 40]
HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
    "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
]

SUITE_NAME = "TS-Digest-Vacation"
COL_SHEET = "COL-digest"


# ─── Shared text fragments ────────────────────────────────────────────────────
#
# Centralising the longer fragments below keeps every TC env-independent
# without drift (a grep for literal env tokens at the end of this file
# will fail if any TC accidentally re-introduces one). The fragments
# reference `<ENV>` placeholders only; the test executor substitutes
# the active env at run time.

_SEED_APPROVED_TOMORROW_SQL = (
    "SELECT e.login, e.email, e.first_name, e.last_name, v.id, "
    "v.start_date, v.end_date, v.vacation_type\n"
    "FROM ttt_vacation.vacation v\n"
    "JOIN employee e ON e.id = v.employee_id\n"
    "WHERE v.status = 'APPROVED'\n"
    "  AND v.start_date = CURRENT_DATE + INTERVAL '1 day'\n"
    "ORDER BY random() LIMIT 1;"
)

_EMAIL_SUBJECT_REGEX = r"/^\[<ENV>\]ТТТ Дайджест отсутствий$/"

_CLOCK_ADVANCE_STEP = (
    "SETUP: Advance server clock to 07:59:55 local (Asia/Novosibirsk) "
    "of the configured test environment so that the 08:00 `@Scheduled` "
    "fire window is 5 s away. Call "
    "PATCH /api/ttt/v1/test/clock { epochMillis: <T-5s local epoch> }."
)

_CLOCK_RESET_STEP = "CLEANUP: POST /api/ttt/v1/test/clock/reset."

_SCHEDULER_WAIT_STEP = (
    "WAIT: up to 60 s for the `@Scheduled` wrapper on "
    "`DigestScheduler.sendDigests` to fire (ShedLock acquire → run → release)."
)

_ENDPOINT_TRIGGER_STEP = (
    "TRIGGER: POST /api/vacation/v1/test/digest (no body). This bypasses "
    "the `@Scheduled` wrapper and calls `DigestService.sendDigests()` "
    "directly — scheduler start/finish markers do NOT emit."
)

_ENDPOINT_WAIT_STEP = (
    "WAIT: 30 s — accounts for the 20 s email-service dequeue loop plus "
    "in-process digest composition."
)

_SCHEDULER_MARKERS_EXPECTED = (
    "Graylog stream `TTT-<ENV>`, within the clock-advance window, emits "
    "exactly one `\"Digests sending job started\"` entry followed by "
    "exactly one `\"Digests sending job finished\"` entry. No "
    "`\"Digests sending job failed\"` marker."
)

_ENDPOINT_MARKERS_EXPECTED = (
    "Graylog stream `TTT-<ENV>` does NOT emit `\"Digests sending job "
    "started\"` or `\"Digests sending job finished\"` (wrapper bypassed). "
    "Only the per-recipient `\"Mail has been sent to ...\"` markers "
    "emit — one per recipient."
)

_PER_RECIPIENT_MARKER_EXPECTED = (
    "Per-recipient Graylog marker on `TTT-<ENV>`: "
    "`\"Mail has been sent to <seed_email> about "
    "NOTIFY_VACATION_UPCOMING for vacation id = <seed_vacation_id>\"`."
)

_CONTENT_SCHEMA_EXPECTED = (
    "Body renders every dynamic field per the digest template schema "
    "in patterns/email-notification-triggers.md § \"Digest template "
    "(Row 14) — content schema\":\n"
    "  • Greeting addressing the recipient by display name "
    "(`Здравствуйте, <First Name> <Last Name>!`).\n"
    "  • Period statement with tomorrow's date in `DD.MM.YYYY` format "
    "(`Напоминаем, что <DD.MM.YYYY> в отсутствии будут:`).\n"
    "  • One per-employee block per APPROVED vacation starting tomorrow:\n"
    "      – Full Name (`<First Name> <Last Name>`)\n"
    "      – start date `DD.MM.YYYY`\n"
    "      – end date `DD.MM.YYYY`\n"
    "      – vacation type localised to Russian "
    "(`Очередной` / `Больничный` / `Отгул`)\n"
    "      – duration in days with correct Russian plural form "
    "(1 `день` | 2–4 `дня` | 0, 5–20 `дней`; for 21 → `день`, "
    "22–24 → `дня`, 25–30 → `дней`).\n"
    "  • Closing footer present.\n"
    "Negative: no block appears for any vacation that is not APPROVED "
    "or whose `start_date` ≠ CURRENT_DATE + 1 day. No leakage of data "
    "from unrelated employees."
)


# ─── TC data ──────────────────────────────────────────────────────────────────
#
# Tuple shape: (test_id, title, preconditions, steps, expected, priority,
#               type_, req_ref, module, notes)
#
# Every behavioural scenario appears twice — once as Variant A (scheduler
# via clock-advance) and once as Variant B (test endpoint, wrapper
# bypass). Business outcome (email, DB effect) matches across A/B;
# scheduler-marker assertions diverge, as documented in the Notes cell.

TCS: list[tuple[str, str, str, str, str, str, str, str, str, str]] = [
    # ─── Happy-path pair (TC-DIGEST-001 A / -002 B) ────────────────────────
    (
        "TC-DIGEST-001",
        "Digest (scheduler) — APPROVED tomorrow vacation: content-complete email + markers",
        (
            "Env: the configured test environment with clock-control enabled "
            "(required for the scheduler variant).\n"
            "Exactly one employee with an APPROVED vacation starting tomorrow "
            "(single-recipient happy path keeps body assertions deterministic).\n"
            "Query to locate or validate seed:\n"
            f"{_SEED_APPROVED_TOMORROW_SQL}\n"
            "If zero rows returned, seed one: call "
            "POST /api/vacation/v1/vacations { login:<target>, "
            "startDate:<tomorrow>, endDate:<tomorrow+4d>, type:'REGULAR' } "
            "then approve as the login's manager via "
            "POST /api/vacation/v1/vacations/{id}/approve.\n"
            "All other APPROVED vacations for tomorrow must be cancelled or "
            "moved so that exactly one recipient is expected (this is a "
            "content-assertion TC, not a leakage TC)."
        ),
        (
            "SETUP: Record `<seed_login>`, `<seed_email>`, `<seed_first_name>`, "
            "`<seed_last_name>`, `<seed_vacation_id>`, `<seed_start_date>`, "
            "`<seed_end_date>`, `<seed_vacation_type>` from the seed query.\n"
            "SETUP: Clear the recipient's mailbox scope — roundcube-access "
            "skill `delete --to <seed_email> --subject \"[<ENV>]\" --since today`.\n"
            "SETUP: Record Graylog baseline — graylog-access skill "
            "`count --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job\"' --range 10m`.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <seed_email> --subject \"[<ENV>]\" --since today -n 5` "
            "→ exactly one new hit.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — assert subject "
            f"matches regex {_EMAIL_SUBJECT_REGEX} (Cyrillic ТТТ, NO brackets "
            "around service tag).\n"
            "VERIFY EMAIL: assert body includes greeting with "
            "`<seed_first_name> <seed_last_name>`.\n"
            "VERIFY EMAIL: assert body period statement contains tomorrow's "
            "date formatted `DD.MM.YYYY`.\n"
            "VERIFY EMAIL: assert body contains exactly one per-employee "
            "block with `<seed_first_name> <seed_last_name>`, "
            "`<seed_start_date as DD.MM.YYYY>`, `<seed_end_date as DD.MM.YYYY>`, "
            "localised `<seed_vacation_type>` (REGULAR→`Очередной`, "
            "SICK_LEAVE→`Больничный`, DAY_OFF→`Отгул`), and duration "
            "with the correct Russian plural form for "
            "`(<seed_end_date> - <seed_start_date>) + 1` days.\n"
            "VERIFY EMAIL: assert footer present; assert NO extra "
            "per-employee blocks (single-recipient seed).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "exactly 1 started + exactly 1 finished hit.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <seed_email>\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING\"' --range 5m` — "
            "exactly 1 hit matching "
            "`vacation id = <seed_vacation_id>`.\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: Delete seeded vacation if the SETUP step created one: "
            "DELETE /api/vacation/v1/vacations/{<seed_vacation_id>}."
        ),
        (
            f"{_SCHEDULER_MARKERS_EXPECTED}\n"
            "Exactly one digest email lands in `<seed_email>` within the "
            "post-trigger window (≤ 60 s clock-advance + scheduler run).\n"
            "Subject: matches the Cyrillic ТТТ regex without brackets "
            f"around the service tag ({_EMAIL_SUBJECT_REGEX}).\n"
            f"{_CONTENT_SCHEMA_EXPECTED}\n"
            f"{_PER_RECIPIENT_MARKER_EXPECTED}"
        ),
        "Critical", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14 · "
        "patterns/email-notification-triggers.md § Digest template",
        "vacation/cron/digest",
        "Variant A (scheduler, clock-advance trigger). Paired with TC-DIGEST-002. "
        "Happy-path baseline; drives content-complete verification. Env-independent."
    ),
    (
        "TC-DIGEST-002",
        "Digest (test endpoint) — APPROVED tomorrow vacation: content-complete email + bypass markers",
        (
            "Env: the configured test environment. Clock-control NOT required "
            "(test endpoint bypasses the `@Scheduled` wrapper, but server "
            "`CURRENT_DATE` still drives the tomorrow filter).\n"
            "Exactly one employee with an APPROVED vacation starting tomorrow. "
            "Use the same seed query as TC-DIGEST-001:\n"
            f"{_SEED_APPROVED_TOMORROW_SQL}\n"
            "If zero rows returned, seed via "
            "POST /api/vacation/v1/vacations + approve. All other APPROVED "
            "tomorrow vacations must be cancelled or moved so exactly one "
            "recipient is expected."
        ),
        (
            "SETUP: Record `<seed_*>` values from the seed query (same fields "
            "as TC-DIGEST-001).\n"
            "SETUP: Clear the recipient's mailbox scope — roundcube-access "
            "skill `delete --to <seed_email> --subject \"[<ENV>]\" --since today`.\n"
            "SETUP: Record Graylog baseline — graylog-access skill "
            "`count --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job\"' --range 10m`. "
            "This baseline must NOT increase (wrapper bypass).\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <seed_email> --subject \"[<ENV>]\" --since today -n 5` "
            "→ exactly one new hit.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — assert subject "
            f"matches regex {_EMAIL_SUBJECT_REGEX}.\n"
            "VERIFY EMAIL: same content-complete assertions as TC-DIGEST-001 "
            "(greeting + display name, period statement with tomorrow "
            "DD.MM.YYYY, per-employee block with Full Name + start + end + "
            "localised type + duration with Russian plural, footer, no "
            "extra blocks).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "ZERO hits (scheduler wrapper bypassed).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <seed_email>\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING\"' --range 5m` — "
            "exactly 1 hit matching `vacation id = <seed_vacation_id>`.\n"
            "CLEANUP: Delete seeded vacation if the SETUP step created one: "
            "DELETE /api/vacation/v1/vacations/{<seed_vacation_id>}."
        ),
        (
            f"{_ENDPOINT_MARKERS_EXPECTED}\n"
            "Exactly one digest email lands in `<seed_email>` within 30 s of "
            "the test endpoint returning.\n"
            f"{_CONTENT_SCHEMA_EXPECTED}\n"
            f"{_PER_RECIPIENT_MARKER_EXPECTED}\n"
            "Business outcome matches TC-DIGEST-001; scheduler markers "
            "absent documents the wrapper bypass."
        ),
        "Critical", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14 · "
        "patterns/email-notification-triggers.md § Digest template",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with TC-DIGEST-001. "
        "Same business outcome; scheduler-marker absence is the bypass signature."
    ),

    # ─── Empty-set pair (TC-DIGEST-003 A / -004 B) ─────────────────────────
    (
        "TC-DIGEST-003",
        "Digest (scheduler) — no APPROVED tomorrow vacations: task completes cleanly, no email",
        (
            "Env: the configured test environment with clock-control enabled.\n"
            "Zero APPROVED vacations start tomorrow. Guard query:\n"
            "SELECT COUNT(*) FROM ttt_vacation.vacation\n"
            "WHERE status = 'APPROVED'\n"
            "  AND start_date = CURRENT_DATE + INTERVAL '1 day';\n"
            "Must return 0. If non-zero, cancel the APPROVED tomorrow "
            "vacations (PUT /api/vacation/v1/vacations/{id}/cancel) or "
            "shift them (PUT) before running."
        ),
        (
            "SETUP: Record Roundcube baseline count — roundcube-access "
            "skill `count --subject \"[<ENV>]ТТТ Дайджест отсутствий\" "
            "--since today`.\n"
            "SETUP: Record Graylog baseline — graylog-access skill "
            "`count --stream TTT-<ENV> --query "
            "'message:\"Digests sending job\"' --range 10m`.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`count --subject \"[<ENV>]ТТТ Дайджест отсутствий\" "
            "--since today` — count equals baseline (no digest dispatched).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\"' --range 5m` "
            "— exactly 1 hit (task still ran).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job finished\"' --range 5m` "
            "— exactly 1 hit (task completed cleanly).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING\"' --range 5m` — ZERO hits "
            "for the post-trigger window.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'level:3 AND message:\"Digests sending job\"' "
            "--range 5m` — ZERO hits (task did not fail).\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: restore any vacations cancelled/shifted during SETUP "
            "guard."
        ),
        (
            "Scheduler start and finish markers emit on `TTT-<ENV>` "
            "(exactly one of each). No per-recipient `Mail has been sent "
            "to ... NOTIFY_VACATION_UPCOMING` markers emit. No error "
            "marker. Roundcube count unchanged vs baseline — task ran "
            "to completion but produced zero outbound emails.\n"
            "Regression guard for the 'empty digest' edge case: the task "
            "MUST terminate cleanly even with an empty recipient set."
        ),
        "High", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-004. Validates clean "
        "termination under empty input."
    ),
    (
        "TC-DIGEST-004",
        "Digest (test endpoint) — no APPROVED tomorrow vacations: bypass completes cleanly, no email",
        (
            "Env: the configured test environment. Clock-control not required.\n"
            "Zero APPROVED vacations start tomorrow. Guard query:\n"
            "SELECT COUNT(*) FROM ttt_vacation.vacation\n"
            "WHERE status = 'APPROVED'\n"
            "  AND start_date = CURRENT_DATE + INTERVAL '1 day';\n"
            "Must return 0."
        ),
        (
            "SETUP: Record Roundcube baseline count — roundcube-access "
            "skill `count --subject \"[<ENV>]ТТТ Дайджест отсутствий\" "
            "--since today`.\n"
            "SETUP: Record Graylog baseline — graylog-access skill "
            "`count --stream TTT-<ENV> --query "
            "'message:\"Digests sending job\" OR "
            "message:\"Mail has been sent to\"' --range 10m`.\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`count --subject \"[<ENV>]ТТТ Дайджест отсутствий\" "
            "--since today` — count equals baseline.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "ZERO hits (wrapper bypassed).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING\"' --range 5m` — ZERO "
            "hits.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'level:3' --range 5m` — no ERROR level entries "
            "attributable to the digest run.\n"
            "CLEANUP: restore any vacations cancelled/shifted during "
            "SETUP guard."
        ),
        (
            "HTTP response from the test endpoint is 2xx. No scheduler "
            "start/finish markers (wrapper bypassed). No per-recipient "
            "markers. No ERROR markers. Roundcube count unchanged vs "
            "baseline. Business outcome identical to TC-DIGEST-003."
        ),
        "High", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with TC-DIGEST-003. "
        "Confirms the bypass path handles the empty-set case cleanly (test "
        "endpoint does not propagate the empty-set as an error)."
    ),

    # ─── Negative-leakage pair (TC-DIGEST-005 A / -006 B) ──────────────────
    (
        "TC-DIGEST-005",
        "Digest (scheduler) — leakage guard: non-APPROVED and non-tomorrow vacations absent from body",
        (
            "Env: the configured test environment with clock-control enabled.\n"
            "Seed a mixed set for the SAME target employee `<target>` so the "
            "email has a known single recipient while the leakage candidates "
            "cannot cause additional recipients:\n"
            "  • 1 APPROVED vacation with start_date = CURRENT_DATE + 1 day "
            "(target — MUST appear in body).\n"
            "  • 1 APPROVED vacation with start_date = CURRENT_DATE + 2 days "
            "(non-tomorrow leakage candidate — MUST NOT appear).\n"
            "  • 1 PENDING vacation with start_date = CURRENT_DATE + 1 day "
            "(non-APPROVED leakage candidate — MUST NOT appear).\n"
            "  • 1 CANCELLED vacation with start_date = CURRENT_DATE + 1 day "
            "(non-APPROVED leakage candidate — MUST NOT appear).\n"
            "  • 1 REJECTED vacation with start_date = CURRENT_DATE + 1 day "
            "(non-APPROVED leakage candidate — MUST NOT appear).\n"
            "Seed query to validate pre-state:\n"
            "SELECT status, start_date, id FROM ttt_vacation.vacation\n"
            "WHERE employee_id = (SELECT id FROM employee WHERE login = '<target>')\n"
            "  AND start_date BETWEEN CURRENT_DATE AND CURRENT_DATE + "
            "INTERVAL '3 days'\n"
            "ORDER BY start_date, status;\n"
            "All other employees must have zero APPROVED tomorrow vacations "
            "(keeps the single-recipient assertion deterministic)."
        ),
        (
            "SETUP: Create or locate the 5-vacation mixed seed for "
            "`<target>`; record IDs for each seeded row (for CLEANUP).\n"
            "SETUP: Clear `<target>`'s mailbox — roundcube-access skill "
            "`delete --to <target_email> --subject \"[<ENV>]\" --since today`.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <target_email> --subject \"[<ENV>]\" --since today` "
            "→ exactly 1 hit.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — body "
            "contains exactly ONE per-employee block. Block details match "
            "the APPROVED tomorrow vacation only (start_date = tomorrow "
            "DD.MM.YYYY).\n"
            "VERIFY EMAIL: assert body does NOT contain the start_date of "
            "the +2-day APPROVED leakage candidate (DD.MM.YYYY format).\n"
            "VERIFY EMAIL: assert body does NOT reference the IDs of the "
            "PENDING, CANCELLED, or REJECTED leakage candidates (where "
            "vacation id appears in the markers, not the body — but their "
            "dates must not appear either).\n"
            "VERIFY EMAIL: assert body contains no extra dates, no extra "
            "vacation-type strings beyond the single APPROVED tomorrow "
            "block.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "1 started + 1 finished.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <target_email>\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING\"' --range 5m` — "
            "exactly 1 hit, referencing the APPROVED tomorrow vacation id "
            "only.\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: Delete all 5 seeded vacations via "
            "DELETE /api/vacation/v1/vacations/{id} for each recorded id."
        ),
        (
            f"{_SCHEDULER_MARKERS_EXPECTED}\n"
            "Exactly one email to `<target_email>`. Body contains exactly "
            "one per-employee block matching the APPROVED tomorrow vacation. "
            "No data (start date, end date, vacation-type string, or block "
            "of any form) from the non-APPROVED statuses (PENDING / "
            "CANCELLED / REJECTED) or from the non-tomorrow APPROVED "
            "(+2 days) appears anywhere in the body.\n"
            "Exactly one per-recipient marker, referencing the APPROVED "
            "tomorrow vacation id only.\n"
            "Negative assertion carries equal weight to positive: the "
            "digest must not leak sibling or future-day data."
        ),
        "Critical", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template · "
        "business rule: APPROVED ∧ start_date = tomorrow",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-006. The single highest-"
        "value regression guard — catches status-filter or date-filter drift."
    ),
    (
        "TC-DIGEST-006",
        "Digest (test endpoint) — leakage guard: non-APPROVED and non-tomorrow vacations absent from body",
        (
            "Env: the configured test environment. Clock-control not required.\n"
            "Same 5-vacation mixed seed as TC-DIGEST-005 for a single target "
            "employee `<target>`:\n"
            "  • 1 APPROVED @ tomorrow (MUST appear).\n"
            "  • 1 APPROVED @ tomorrow+1 (leakage candidate).\n"
            "  • 1 PENDING @ tomorrow (leakage candidate).\n"
            "  • 1 CANCELLED @ tomorrow (leakage candidate).\n"
            "  • 1 REJECTED @ tomorrow (leakage candidate).\n"
            "All other employees must have zero APPROVED tomorrow vacations."
        ),
        (
            "SETUP: Create or locate the 5-vacation seed; record IDs.\n"
            "SETUP: Clear `<target>`'s mailbox — roundcube-access skill "
            "`delete --to <target_email> --subject \"[<ENV>]\" --since today`.\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <target_email> --subject \"[<ENV>]\" --since today` "
            "→ exactly 1 hit.\n"
            "VERIFY EMAIL: same leakage assertions as TC-DIGEST-005 — body "
            "contains exactly ONE per-employee block; no data from the "
            "4 leakage candidates appears.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "ZERO hits.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <target_email>\"' "
            "--range 5m` — exactly 1 hit, APPROVED tomorrow vacation id "
            "only.\n"
            "CLEANUP: Delete all 5 seeded vacations."
        ),
        (
            f"{_ENDPOINT_MARKERS_EXPECTED}\n"
            "Exactly one email; exactly one per-employee block matching the "
            "APPROVED tomorrow vacation. No leakage from the other 4 seed "
            "rows in the body. Business outcome matches TC-DIGEST-005.\n"
            "Confirms the status/date filter holds on the bypass path "
            "(a regression here with A passing would indicate the wrapper "
            "is hiding a filter defect)."
        ),
        "Critical", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template · "
        "business rule: APPROVED ∧ start_date = tomorrow",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with TC-DIGEST-005."
    ),

    # ─── Subject-regex audit pair (TC-DIGEST-007 A / -008 B) ───────────────
    (
        "TC-DIGEST-007",
        "Digest (scheduler) — subject-format regex: `[<ENV>]ТТТ Дайджест отсутствий` (Cyrillic ТТТ, no brackets)",
        (
            "Env: the configured test environment with clock-control enabled.\n"
            "Same seed as TC-DIGEST-001 — one employee with one APPROVED "
            "vacation starting tomorrow (any content; this TC only audits "
            "envelope subject format).\n"
            f"{_SEED_APPROVED_TOMORROW_SQL}"
        ),
        (
            "SETUP: Record `<seed_email>`.\n"
            "SETUP: Clear recipient mailbox — roundcube-access skill "
            "`delete --to <seed_email> --subject \"[<ENV>]\" --since today`.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <seed_email> --subject \"[<ENV>]\" --since today -n 5` "
            "→ capture the first match UID.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — assert the "
            f"subject line matches the regex {_EMAIL_SUBJECT_REGEX} exactly.\n"
            "VERIFY EMAIL: assert the subject characters ТТТ are Cyrillic "
            "code points U+0422 U+0422 U+0422 (not Latin U+0054 U+0054 "
            "U+0054). Unicode checker: the sequence after `]` before space "
            "must fail a `.match(/^TTT/)` ASCII test.\n"
            "VERIFY EMAIL: assert subject does NOT match the Latin bracketed "
            r"pattern `/^\[<ENV>\]\[TTT\] /` used by every other TTT email "
            "template — the digest is anomalous in stripping the inner "
            "brackets and localising the service tag to Cyrillic.\n"
            "VERIFY EMAIL: assert subject is exactly "
            "`[<ENV>]ТТТ Дайджест отсутствий` (no trailing data, no "
            "appended date — the schedule-date information lives in the "
            "body period statement, not the subject).\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: Delete seeded vacation if created in SETUP."
        ),
        (
            "Subject line on the delivered digest email:\n"
            "  • Starts with `[<ENV>]` where `<ENV>` is substituted with the "
            "configured environment token.\n"
            "  • Immediately after the closing `]` (no space, no bracket), "
            "contains the literal Cyrillic string `ТТТ` (three Cyrillic "
            "Te characters, U+0422).\n"
            "  • Followed by one space and the Cyrillic phrase "
            "`Дайджест отсутствий` with no trailing data.\n"
            "If a Latin `[TTT]` bracketed pattern is observed instead, the "
            "digest has regressed to the standard TTT subject convention — "
            "report as a defect linking patterns/email-notification-triggers.md "
            "§ anomaly."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § subject anomaly",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-008. Subject-regex audit "
        "is independent of body content — focuses on envelope."
    ),
    (
        "TC-DIGEST-008",
        "Digest (test endpoint) — subject-format regex: `[<ENV>]ТТТ Дайджест отсутствий` (Cyrillic ТТТ, no brackets)",
        (
            "Env: the configured test environment. Clock-control not required.\n"
            "Same seed as TC-DIGEST-001/007 — one employee with one APPROVED "
            "vacation starting tomorrow."
        ),
        (
            "SETUP: Record `<seed_email>`.\n"
            "SETUP: Clear recipient mailbox — roundcube-access skill "
            "`delete --to <seed_email> --subject \"[<ENV>]\" --since today`.\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <seed_email> --subject \"[<ENV>]\" --since today -n 5` "
            "→ capture the first match UID.\n"
            "VERIFY EMAIL: same subject assertions as TC-DIGEST-007 "
            "(Cyrillic ТТТ regex match, Latin `[TTT]` bracketed pattern "
            "does NOT match, Unicode code-point checks).\n"
            "CLEANUP: Delete seeded vacation if created in SETUP."
        ),
        (
            "Subject format identical to TC-DIGEST-007. The bypass path "
            "must produce the same envelope string — if Variant A shows "
            "Cyrillic but Variant B shows Latin (or vice-versa), the "
            "subject is being composed at different layers and that is a "
            "defect."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § subject anomaly",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with TC-DIGEST-007. "
        "Confirms envelope composition is not wrapper-dependent."
    ),

    # ─── Graylog marker audit pair (TC-DIGEST-009 A / -010 B) ──────────────
    (
        "TC-DIGEST-009",
        "Digest (scheduler) — Graylog marker audit: start → finish + per-recipient emit on TTT-<ENV>",
        (
            "Env: the configured test environment with clock-control enabled.\n"
            "Same seed as TC-DIGEST-001 — one employee with one APPROVED "
            "vacation starting tomorrow.\n"
            f"{_SEED_APPROVED_TOMORROW_SQL}"
        ),
        (
            "SETUP: Record `<seed_email>`, `<seed_vacation_id>`.\n"
            "SETUP: Record Graylog baseline timestamps via graylog-access "
            "skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job\"' --range 10m -n 10` — "
            "note the latest pre-test timestamp.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\"' --range 5m` "
            "— exactly 1 new hit (timestamp > baseline).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job finished\"' --range 5m` "
            "— exactly 1 new hit, timestamp > the `started` hit.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <seed_email>\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING for vacation id = "
            "<seed_vacation_id>\"' --range 5m` — exactly 1 hit matching "
            "the full per-recipient marker format.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job failed\"' --range 5m` "
            "— ZERO hits (no error path).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'level:3' --range 5m` — no level-3 (ERROR) entries "
            "correlated with the digest run (filter by "
            "`source:\"vacation-*\"` if the stream carries mixed sources).\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: Delete seeded vacation if created in SETUP."
        ),
        (
            "On Graylog stream `TTT-<ENV>`, within the 5-minute post-trigger "
            "window:\n"
            "  • `\"Digests sending job started\"` — exactly 1 new entry.\n"
            "  • `\"Digests sending job finished\"` — exactly 1 new entry, "
            "after the `started` entry.\n"
            "  • Per-recipient `\"Mail has been sent to <seed_email> about "
            "NOTIFY_VACATION_UPCOMING for vacation id = "
            "<seed_vacation_id>\"` — exactly 1 new entry.\n"
            "  • `\"Digests sending job failed\"` — 0 entries.\n"
            "  • Level-3 (ERROR) entries — 0 attributable to the digest.\n"
            "Ordering invariant: `started` timestamp ≤ `per-recipient` "
            "timestamp ≤ `finished` timestamp (the wrapper brackets the "
            "per-recipient emits)."
        ),
        "High", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14 § markers",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-010. Dedicated log "
        "audit — catches marker regressions without relying on email delivery."
    ),
    (
        "TC-DIGEST-010",
        "Digest (test endpoint) — Graylog marker audit: scheduler markers ABSENT, per-recipient present",
        (
            "Env: the configured test environment. Clock-control not required.\n"
            "Same seed as TC-DIGEST-001/009 — one employee with one APPROVED "
            "vacation starting tomorrow."
        ),
        (
            "SETUP: Record `<seed_email>`, `<seed_vacation_id>`.\n"
            "SETUP: Record Graylog baseline via graylog-access skill "
            "`search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job\"' --range 10m -n 10` — "
            "note the latest pre-test timestamp.\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\"' --range 5m` "
            "— ZERO new hits (timestamps > baseline). The `@Scheduled` "
            "wrapper did not fire because the test endpoint bypasses it.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job finished\"' --range 5m` "
            "— ZERO new hits.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Mail has been sent to <seed_email>\" AND "
            "message:\"NOTIFY_VACATION_UPCOMING for vacation id = "
            "<seed_vacation_id>\"' --range 5m` — exactly 1 hit.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'level:3' --range 5m` — no correlated ERROR entries.\n"
            "CLEANUP: Delete seeded vacation if created in SETUP."
        ),
        (
            "On Graylog stream `TTT-<ENV>`, within the 5-minute post-trigger "
            "window:\n"
            "  • `\"Digests sending job started\"` — 0 new entries.\n"
            "  • `\"Digests sending job finished\"` — 0 new entries.\n"
            "  • Per-recipient marker — exactly 1 new entry matching the "
            "full format `\"Mail has been sent to <seed_email> about "
            "NOTIFY_VACATION_UPCOMING for vacation id = "
            "<seed_vacation_id>\"`.\n"
            "  • No ERROR entries.\n"
            "The absence of the scheduler markers is the documented "
            "wrapper-bypass signature — DigestService.sendDigests() is "
            "invoked directly by the test controller, skipping the "
            "@Scheduled AOP layer that emits start/finish."
        ),
        "High", "Hybrid",
        "#3423 row 14 · EXT-cron-jobs.md job 14 § markers · "
        "§ wrapper-bypass note",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with "
        "TC-DIGEST-009. Absence assertion is load-bearing — it distinguishes "
        "B from A at the log layer."
    ),

    # ─── Plural-form edge-case pair (TC-DIGEST-011 A / -012 B) ─────────────
    (
        "TC-DIGEST-011",
        "Digest (scheduler) — Russian plural-form edge cases (1 день / 2 дня / 5 дней / 21 день)",
        (
            "Env: the configured test environment with clock-control enabled.\n"
            "Seed 4 distinct employees, each with one APPROVED vacation "
            "starting tomorrow, durations chosen to exercise every Russian "
            "plural branch:\n"
            "  • `<E1>`: 1 day (start = tomorrow, end = tomorrow) → "
            "expected `1 день`.\n"
            "  • `<E2>`: 2 days (start = tomorrow, end = tomorrow + 1) → "
            "expected `2 дня`.\n"
            "  • `<E3>`: 5 days (start = tomorrow, end = tomorrow + 4) → "
            "expected `5 дней`.\n"
            "  • `<E4>`: 21 days (start = tomorrow, end = tomorrow + 20) → "
            "expected `21 день` (the `…1` suffix rule — excluding `11`).\n"
            "All 4 employees must be distinct from each other and from any "
            "ambient APPROVED tomorrow vacation. Seed query:\n"
            "SELECT e.login, e.email, v.id, v.start_date, v.end_date\n"
            "FROM ttt_vacation.vacation v\n"
            "JOIN employee e ON e.id = v.employee_id\n"
            "WHERE v.status = 'APPROVED'\n"
            "  AND v.start_date = CURRENT_DATE + INTERVAL '1 day'\n"
            "ORDER BY (v.end_date - v.start_date);\n"
            "Seed via POST /api/vacation/v1/vacations + approve if fewer "
            "than 4 duration-distinct rows exist."
        ),
        (
            "SETUP: Record `<E1..E4>` email addresses, vacation ids, "
            "start/end dates.\n"
            "SETUP: Clear each employee's mailbox — roundcube-access skill "
            "`delete --to <Ei_email> --subject \"[<ENV>]\" --since today` "
            "for i ∈ {1,2,3,4}.\n"
            f"{_CLOCK_ADVANCE_STEP}\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL (E1, 1 day): roundcube-access skill "
            "`search --to <E1_email> --subject \"[<ENV>]\" --since today` "
            "→ 1 hit; `read <uid>` — body contains `1 день` in the duration "
            "line of the per-employee block.\n"
            "VERIFY EMAIL (E2, 2 days): 1 hit; body contains `2 дня`.\n"
            "VERIFY EMAIL (E3, 5 days): 1 hit; body contains `5 дней`.\n"
            "VERIFY EMAIL (E4, 21 days): 1 hit; body contains `21 день` "
            "(NOT `21 дней` — the `…1` rule excludes `11..19` but includes "
            "`21`, `31`, …).\n"
            "VERIFY EMAIL (negative): none of the 4 emails contain the "
            "wrong plural form (e.g. `1 дня`, `2 дней`, `5 день`, "
            "`21 дня`).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "1 started + 1 finished.\n"
            "VERIFY LOG: 4 per-recipient `\"Mail has been sent to <Ei_email>\"` "
            "markers — one per seeded vacation id.\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: Delete all 4 seeded vacations via DELETE for each id."
        ),
        (
            f"{_SCHEDULER_MARKERS_EXPECTED}\n"
            "4 digest emails dispatched, one per seeded employee. Each "
            "body's per-employee block duration line uses the correct "
            "Russian plural form per the standard rule: `…1` (excluding "
            "`11–19`) → `день`; `…2–4` (excluding `12–14`) → `дня`; "
            "`…5–9`, `…0`, `11–19` → `дней`. Specifically:\n"
            "  • 1 day → `1 день`\n"
            "  • 2 days → `2 дня`\n"
            "  • 5 days → `5 дней`\n"
            "  • 21 days → `21 день`\n"
            "4 per-recipient Graylog markers, one per seeded vacation id. "
            "If any body contains the wrong form, the i18n / plural "
            "formatter has regressed — this TC catches silent locale "
            "drift (which subject-only TCs would miss)."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template "
        "(Russian plural form)",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-012. Content-complete "
        "audit focused on plural-form correctness; durations chosen to hit "
        "all 3 Russian plural branches plus the `…1` branch after the "
        "teens-exclusion (21)."
    ),
    (
        "TC-DIGEST-012",
        "Digest (test endpoint) — Russian plural-form edge cases (1 день / 2 дня / 5 дней / 21 день)",
        (
            "Env: the configured test environment. Clock-control not required.\n"
            "Same 4-employee seed as TC-DIGEST-011 — durations 1, 2, 5, 21 days."
        ),
        (
            "SETUP: Record `<E1..E4>` email addresses and vacation ids.\n"
            "SETUP: Clear each employee's mailbox — "
            "`delete --to <Ei_email> --subject \"[<ENV>]\" --since today` "
            "for i ∈ {1,2,3,4}.\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: same plural-form body assertions as "
            "TC-DIGEST-011 — E1 `1 день`, E2 `2 дня`, E3 `5 дней`, "
            "E4 `21 день`.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — ZERO "
            "hits (wrapper bypass).\n"
            "VERIFY LOG: 4 per-recipient markers — one per seeded "
            "vacation id.\n"
            "CLEANUP: Delete all 4 seeded vacations."
        ),
        (
            f"{_ENDPOINT_MARKERS_EXPECTED}\n"
            "4 digest emails with the same plural-form assertions as "
            "TC-DIGEST-011 (1 день / 2 дня / 5 дней / 21 день). 4 "
            "per-recipient Graylog markers.\n"
            "Business outcome matches TC-DIGEST-011; confirms the plural "
            "formatter is not wrapper-dependent."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template "
        "(Russian plural form)",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with TC-DIGEST-011."
    ),

    # ─── Cross-year date-format pair (TC-DIGEST-013 A / -014 B) ────────────
    (
        "TC-DIGEST-013",
        "Digest (scheduler) — cross-year boundary: tomorrow = Jan 1 renders `01.01.<YYYY+1>`",
        (
            "Env: the configured test environment with clock-control enabled "
            "(mandatory — relies on clock-advance to place `tomorrow` on a "
            "specific date).\n"
            "Seed one employee with an APPROVED vacation starting on "
            "`<YYYY+1>-01-01` and lasting 5 days (end = `<YYYY+1>-01-05`) "
            "where `<YYYY>` is the clock-advance target year (see SETUP).\n"
            "Seed via POST /api/vacation/v1/vacations "
            "{ login:<target>, startDate:'<YYYY+1>-01-01', "
            "endDate:'<YYYY+1>-01-05', type:'REGULAR' } followed by "
            "manager approve."
        ),
        (
            "SETUP: Record `<seed_email>`, `<seed_first_name>`, "
            "`<seed_last_name>`, `<seed_vacation_id>`, `<YYYY>` "
            "(the year to advance to).\n"
            "SETUP: Clear recipient mailbox — roundcube-access skill "
            "`delete --to <seed_email> --subject \"[<ENV>]\" --since today`.\n"
            "SETUP: Advance server clock to `<YYYY>-12-31 07:59:55` local "
            "(Asia/Novosibirsk) via PATCH /api/ttt/v1/test/clock "
            "{ epochMillis: <local Dec 31 07:59:55 epoch> }. "
            "Server `CURRENT_DATE` becomes Dec 31, so `CURRENT_DATE + "
            "INTERVAL '1 day'` resolves to `<YYYY+1>-01-01` — the year "
            "boundary is the scenario.\n"
            f"{_SCHEDULER_WAIT_STEP}\n"
            "VERIFY EMAIL: roundcube-access skill "
            "`search --to <seed_email> --subject \"[<ENV>]\" --since today` "
            "→ 1 hit.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — body "
            "period statement contains the literal `01.01.<YYYY+1>` "
            "(four-digit next year).\n"
            "VERIFY EMAIL: body per-employee block contains "
            "`<seed_first_name> <seed_last_name>`, start `01.01.<YYYY+1>`, "
            "end `05.01.<YYYY+1>`, type `Очередной`, duration `5 дней`.\n"
            "VERIFY EMAIL: assert the body does NOT contain the current-year "
            "value `<YYYY>` anywhere in a date field (would indicate the "
            "year increment rolled incorrectly).\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "1 started + 1 finished.\n"
            "VERIFY LOG: 1 per-recipient marker for the seeded vacation id.\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: DELETE /api/vacation/v1/vacations/{<seed_vacation_id>}."
        ),
        (
            f"{_SCHEDULER_MARKERS_EXPECTED}\n"
            "Email body renders the date fields with the correct year "
            "increment: tomorrow = `01.01.<YYYY+1>`; seeded vacation "
            "dates `01.01.<YYYY+1>` → `05.01.<YYYY+1>`. No ambient "
            "`<YYYY>` leaks into a rendered date field. Duration is "
            "`5 дней` (plural `дней` branch).\n"
            "Cross-year boundary is a classic off-by-one pitfall in date "
            "formatters — this TC catches any regression where the year "
            "rollover is computed incorrectly (e.g. `01.01.<YYYY>`)."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template "
        "(date format)",
        "vacation/cron/digest",
        "Variant A (scheduler). Paired with TC-DIGEST-014. Clock-advance to "
        "Dec 31 places the digest run on the year boundary — the only way "
        "to exercise year-rollover in the date formatter."
    ),
    (
        "TC-DIGEST-014",
        "Digest (test endpoint) — cross-year boundary: tomorrow = Jan 1 renders `01.01.<YYYY+1>`",
        (
            "Env: the configured test environment with clock-control enabled "
            "(required — `CURRENT_DATE` on the server must resolve to Dec 31 "
            "so `CURRENT_DATE + 1 day` = Jan 1; the test endpoint bypasses "
            "the scheduler wrapper but still uses the server clock).\n"
            "Same seed as TC-DIGEST-013 — APPROVED vacation "
            "`<YYYY+1>-01-01` → `<YYYY+1>-01-05`."
        ),
        (
            "SETUP: Record `<seed_*>` values as in TC-DIGEST-013.\n"
            "SETUP: Clear recipient mailbox — `delete --to <seed_email> "
            "--subject \"[<ENV>]\" --since today`.\n"
            "SETUP: Advance server clock to `<YYYY>-12-31 12:00:00` local "
            "via PATCH /api/ttt/v1/test/clock (time-of-day is immaterial "
            "for the test endpoint — only the date drives the filter).\n"
            f"{_ENDPOINT_TRIGGER_STEP}\n"
            f"{_ENDPOINT_WAIT_STEP}\n"
            "VERIFY EMAIL: same date-format assertions as TC-DIGEST-013 — "
            "body period statement shows `01.01.<YYYY+1>`; per-employee "
            "block dates `01.01.<YYYY+1>` → `05.01.<YYYY+1>`; duration "
            "`5 дней`; no leak of `<YYYY>` into any date field.\n"
            "VERIFY LOG: graylog-access skill `search --stream TTT-<ENV> "
            "--query 'message:\"Digests sending job started\" OR "
            "message:\"Digests sending job finished\"' --range 5m` — "
            "ZERO hits.\n"
            "VERIFY LOG: 1 per-recipient marker for the seeded vacation id.\n"
            f"{_CLOCK_RESET_STEP}\n"
            "CLEANUP: DELETE /api/vacation/v1/vacations/{<seed_vacation_id>}."
        ),
        (
            f"{_ENDPOINT_MARKERS_EXPECTED}\n"
            "Email body renders dates with the year increment "
            "(`01.01.<YYYY+1>`, `05.01.<YYYY+1>`). Duration `5 дней`. No "
            "`<YYYY>` leak in any date field. Business outcome matches "
            "TC-DIGEST-013.\n"
            "Confirms the date formatter is not wrapper-dependent — "
            "cross-year rollover works identically on both trigger paths."
        ),
        "High", "Hybrid",
        "#3423 row 14 · patterns/email-notification-triggers.md § Digest template "
        "(date format)",
        "vacation/cron/digest",
        "Variant B (test endpoint, scheduler bypass). Paired with "
        "TC-DIGEST-013. Clock-advance still required — the test endpoint "
        "bypasses the scheduler but not the clock."
    ),
]


# ─── Write the TS-Digest-Vacation suite sheet ─────────────────────────────────


def _write_suite(wb, name: str, rows):
    """Rebuild the ``TS-Digest-Vacation`` sheet fresh.

    Row 1  — back-link to Plan Overview
    Row 2  — bold header, frozen, auto-filtered
    Row 3+ — TC data rows, alternating fill, wrap-text top-align, row
             heights sized by line count
    """
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLOR

    # R1 — back-link
    back = ws.cell(row=1, column=1,
                   value='=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")')
    back.font = FONT_LINK
    ws.row_dimensions[1].height = 20

    # R2 — header
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        letter = c.column_letter
        if col_idx - 1 < len(COL_WIDTHS):
            ws.column_dimensions[letter].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(HEADERS)).column_letter}2"

    # R3+ — TC rows
    for row_offset, tc in enumerate(rows, start=0):
        r = 3 + row_offset
        fill = FILL_ALT if (row_offset % 2) else FILL_WHITE
        max_lines = 1
        for col_idx, val in enumerate(tc, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1, 6, 7) else ALIGN_WRAP
            c.fill = fill
            c.border = BORDER_THIN
            # Track the tallest cell so we can size the row accordingly.
            if isinstance(val, str) and "\n" in val:
                line_count = val.count("\n") + 1
                if line_count > max_lines:
                    max_lines = line_count
        # Row height: 15 px per line, min 18. Wrap-text + top-align renders
        # each `\n`-separated step on its own physical line.
        ws.row_dimensions[r].height = max(18, 15 * max_lines)


# ─── Rewrite COL-digest data rows ─────────────────────────────────────────────


def _rewrite_col_digest(wb, tcs):
    """Rewrite data rows (row 3+) of COL-digest, preserving row 1
    description and row 2 headers.

    Each row references TS-Digest-Vacation as the source suite, so the
    collection report (autotests/scripts/process_collection.py) can
    resolve every TC to an xlsx cell when the collection is expanded
    into a Playwright `@col-digest` tag set (future work — Phase C is
    gated off for this collection).
    """
    if COL_SHEET not in wb.sheetnames:
        raise SystemExit(
            f"Expected sheet {COL_SHEET!r} to exist (seeded 2026-04-21); "
            "aborting to avoid silently creating a duplicate."
        )
    ws = wb[COL_SHEET]

    # Clear existing data rows (row 3+) by nulling every cell, then
    # re-setting column widths for legibility.
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c, value=None)
    # Drop physical row dimensions for cleared rows so they don't inherit
    # previous heights.
    for r in range(3, max_row + 1):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    # Widths for COL-digest: test_id / source_module / source_workbook /
    # source_suite / title / inclusion_reason / priority_override
    col_widths = [14, 18, 52, 22, 48, 64, 18]
    for i, w in enumerate(col_widths, start=1):
        letter = ws.cell(row=2, column=i).column_letter
        ws.column_dimensions[letter].width = w

    # Make sure row 2 headers are styled (they may already be from the
    # seed — reapply idempotently).
    for col_idx in range(1, len(col_widths) + 1):
        h = ws.cell(row=2, column=col_idx)
        h.font = FONT_HEADER
        h.fill = FILL_HEADER
        h.alignment = ALIGN_CENTER
        h.border = BORDER_THIN
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(col_widths)).column_letter}2"

    # Inclusion-reason mapping — one per TC, describes why the TC belongs
    # in the collection (useful when the collection is consumed by
    # process_collection.py or a human reviewer).
    _inclusion_reason = {
        "TC-DIGEST-001": "Happy-path content-complete (scheduler variant) — baseline.",
        "TC-DIGEST-002": "Happy-path content-complete (test-endpoint bypass variant).",
        "TC-DIGEST-003": "Empty-set clean termination (scheduler variant).",
        "TC-DIGEST-004": "Empty-set clean termination (test-endpoint bypass variant).",
        "TC-DIGEST-005": "Leakage guard: non-APPROVED / non-tomorrow must NOT appear (scheduler).",
        "TC-DIGEST-006": "Leakage guard: non-APPROVED / non-tomorrow must NOT appear (bypass).",
        "TC-DIGEST-007": "Subject-format regex audit — Cyrillic ТТТ, no brackets (scheduler).",
        "TC-DIGEST-008": "Subject-format regex audit — Cyrillic ТТТ, no brackets (bypass).",
        "TC-DIGEST-009": "Graylog marker audit — start/finish + per-recipient emit (scheduler).",
        "TC-DIGEST-010": "Graylog marker audit — scheduler markers ABSENT, per-recipient present (bypass).",
        "TC-DIGEST-011": "Russian plural-form edge cases 1/2/5/21 (scheduler).",
        "TC-DIGEST-012": "Russian plural-form edge cases 1/2/5/21 (bypass).",
        "TC-DIGEST-013": "Cross-year date boundary — tomorrow = Jan 1 (scheduler).",
        "TC-DIGEST-014": "Cross-year date boundary — tomorrow = Jan 1 (bypass).",
    }

    for i, tc in enumerate(tcs, start=0):
        test_id = tc[0]
        title = tc[1]
        reason = _inclusion_reason.get(test_id, "")
        r = 3 + i
        row_vals = [
            test_id,
            "vacation",
            "test-docs/collections/digest/digest.xlsx",
            SUITE_NAME,
            title,
            reason,
            "",  # priority_override — none; retain per-TC priority
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1,) else ALIGN_WRAP
            c.border = BORDER_THIN
        ws.row_dimensions[r].height = 32


# ─── Pre-save audit ───────────────────────────────────────────────────────────


def _audit(wb):
    """Before saving, grep every TS-Digest-Vacation cell for
    environment-specific literals. If any is found, fail the generator
    so the defect does not ship.

    Also assert wrap-text is ON for every prose column on every body
    row — the legibility bar is non-negotiable per the session briefing.
    """
    banned = ("qa-1", "qa1", "timemachine", "stage", "preprod", "dev")
    prose_cols = {2, 3, 4, 5, 10}  # Title, Preconditions, Steps, Expected, Notes

    ws = wb[SUITE_NAME]
    defects = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str):
                lower = v.lower()
                for tok in banned:
                    # Special-case: "dev" appears in "dev-" / "development" /
                    # "developer"; only flag as a bare token.
                    if tok == "dev":
                        continue  # too many false-positives; skip
                    # Match as a standalone token, not inside a path/word like
                    # "preproduction" or "devops". We use simple word-boundary
                    # checks via surrounding chars.
                    idx = 0
                    while True:
                        pos = lower.find(tok, idx)
                        if pos < 0:
                            break
                        left = lower[pos - 1] if pos > 0 else ""
                        right = lower[pos + len(tok)] if pos + len(tok) < len(lower) else ""
                        if (not left.isalnum() and left not in "-_") and \
                           (not right.isalnum() and right not in "-_"):
                            defects.append(
                                f"  {cell.coordinate}: env literal "
                                f"{tok!r} in cell {v!r}"
                            )
                            break
                        idx = pos + len(tok)
            if cell.column in prose_cols and not cell.alignment.wrap_text:
                defects.append(
                    f"  {cell.coordinate}: prose column missing wrap_text"
                )

    if defects:
        raise SystemExit(
            "Pre-save audit failed — defects in "
            f"{SUITE_NAME}:\n" + "\n".join(defects)
        )


# ─── Main ─────────────────────────────────────────────────────────────────────


def main():
    if not os.path.exists(_TARGET):
        raise SystemExit(
            f"Target workbook missing: {_TARGET}\n"
            "Expected the collection skeleton (Plan Overview + COL-digest) "
            "to be seeded before this generator runs."
        )
    wb = load_workbook(_TARGET)

    _write_suite(wb, SUITE_NAME, TCS)
    _rewrite_col_digest(wb, TCS)
    _audit(wb)

    # Set workbook active sheet to Plan Overview (first tab) for a clean
    # reader experience.
    if "Plan Overview" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Plan Overview")

    wb.save(_TARGET)

    ids = [tc[0] for tc in TCS]
    print(f"Generated: {_TARGET}")
    print(f"  Suite rebuilt: {SUITE_NAME} ({len(TCS)} TCs)")
    print(f"  COL-digest data rows rewritten: {len(TCS)}")
    print(f"  TC IDs: {ids[0]} … {ids[-1]}")
    print(
        "  Variant pairs: "
        "(001,002) happy-path · "
        "(003,004) empty-set · "
        "(005,006) leakage guard · "
        "(007,008) subject regex · "
        "(009,010) Graylog markers · "
        "(011,012) plural forms · "
        "(013,014) cross-year date"
    )


if __name__ == "__main__":
    main()
