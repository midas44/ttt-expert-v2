#!/usr/bin/env python3
"""
Accounting Test Documentation Generator — Phase B (Session 93)
Generates test-docs/accounting/accounting.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 6 TS- test suite tabs.

Based on vault knowledge: accounting-service-deep-dive (17.1KB, 13 design issues),
accounting-ticket-findings (12KB, 12 tickets deep-read from 62+ searched),
accounting-api-testing, accounting-pages (UI flows), frontend-accounting-module,
office-period-model, vacation-day-calculation-architecture.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "accounting")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "accounting.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_periods_cases():
    """TS-Accounting-Periods: Salary page, period management for offices."""
    return [
        {
            "id": "TC-ACC-001", "title": "View Salary page — salary office list",
            "preconditions": "Accountant or admin with access to Accounting.\nQuery: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_ACCOUNTANT', 'ROLE_CHIEF_ACCOUNTANT', 'ROLE_ADMIN') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Salary page\n3. Verify a list of salary offices is displayed\n4. Verify each office shows: name, report period, approve period\n5. Verify a salary office filter/dropdown is present\n6. Select a specific salary office from the filter\n7. Verify the employee list updates to show only employees from that office\n8. Verify 'All' option shows employees from all accessible offices",
            "expected": "Salary page shows all accessible salary offices with period info. Filter works correctly.",
            "priority": "Critical", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Salary Page", "module": "accounting/salary",
            "notes": "Filter bug #2841: localStorage with stale data can cause empty filter. ACCOUNTANT+ADMIN role combination may show zero SO."
        },
        {
            "id": "TC-ACC-002", "title": "Salary office filter — all accessible offices visible (#2841)",
            "preconditions": "User with ACCOUNTANT + ADMIN dual roles.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ACCOUNTANT' AND e.enabled = true AND EXISTS (SELECT 1 FROM ttt_backend.employee_role er2 JOIN ttt_backend.role r2 ON er2.role_id = r2.id WHERE er2.employee_id = e.id AND r2.name = 'ROLE_ADMIN') ORDER BY random() LIMIT 1",
            "steps": "1. Login as the ACCOUNTANT+ADMIN user\n2. Navigate to Accounting > Salary page\n3. Open the salary office filter dropdown\n4. Count the number of offices in the dropdown\n5. Compare with expected number (all offices for ADMIN role)\n6. Verify no offices are missing\n7. Clear browser localStorage and reload\n8. Verify filter shows correct offices after localStorage clear\n9. Test with >20 salary offices — verify pagination works",
            "expected": "All salary offices visible in filter. No missing offices due to localStorage corruption.",
            "priority": "High", "type": "UI",
            "req_ref": "#2841 SO filter bug", "module": "accounting/salary",
            "notes": "Bug #2841: filter showed subset of offices. Also: >20 SO pagination bug, filter breaks with nameless SO, names don't update on language switch."
        },
        {
            "id": "TC-ACC-003", "title": "Salary office filter — language switch updates names",
            "preconditions": "Accountant on Salary page with salary offices that have both EN and RU names.",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Salary page\n3. Note salary office names in current language (e.g., English)\n4. Click language switcher → switch to Russian\n5. Verify salary office names update to Russian translations\n6. Switch back to English\n7. Verify names revert to English",
            "expected": "Salary office names update immediately on language switch. No stale cached names.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2841 language switch", "module": "accounting/salary",
            "notes": "Bug from #2841 comments: SO names don't update on language switch."
        },
        {
            "id": "TC-ACC-004", "title": "Archived salary offices hidden from dropdowns (#3323)",
            "preconditions": "At least one archived salary office exists in the system.",
            "steps": "1. Login as accountant/admin\n2. Navigate to Accounting > Salary page\n3. Open the salary office dropdown\n4. Verify archived salary offices are NOT listed\n5. Navigate to Accounting > Changing periods\n6. Open the salary office dropdown\n7. Verify archived offices are NOT listed here either\n8. Navigate to Admin > Salary offices\n9. Verify archived offices ARE visible but dimmed/grayed\n10. Verify archived offices are at the bottom of the list\n11. Verify archived offices cannot be edited",
            "expected": "Archived SO hidden from Accounting dropdowns. Visible but non-editable and dimmed in Admin panel.",
            "priority": "High", "type": "UI",
            "req_ref": "#3323 archived SO", "module": "accounting/salary",
            "notes": "#3323 (OPEN, Sprint 15): archived SO hidden from Accounting, dimmed in Admin."
        },
        {
            "id": "TC-ACC-005", "title": "Changing periods page — view current periods",
            "preconditions": "Accountant/admin.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Verify Report period displayed (month + year)\n5. Verify Approve period displayed (month + year)\n6. Verify Approve period ≤ Report period (invariant)\n7. Verify 'Edit' buttons are present for both periods",
            "expected": "Both periods displayed correctly. Approve period is at most equal to report period.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Dual Period Model", "module": "accounting/periods",
            "notes": "Invariant: approve ≤ report. Non-salary offices return 404 on PATCH."
        },
        {
            "id": "TC-ACC-006", "title": "Advance report period +1 month",
            "preconditions": "Accountant with a salary office where report period can advance.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Note current report period\n5. Click 'Edit' for Report period\n6. In the picker, select +1 month\n7. Click 'Save'\n8. Verify period advances\n9. Verify employees in that SO can no longer report in the old period\n10. Verify success notification",
            "expected": "Report period advances by 1 month. Old period locked for employee reporting.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3350 1-month rule", "module": "accounting/periods",
            "notes": "Max 1-month delta from saved value. Must be first of month."
        },
        {
            "id": "TC-ACC-007", "title": "Advance approve period +1 month (triggers side effects)",
            "preconditions": "Accountant. Approve period at least 1 month behind report period.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Changing periods\n3. Select salary office\n4. Click 'Edit' for Approve period\n5. Select +1 month\n6. Click 'Save'\n7. Verify period advances\n8. Verify PeriodChangedEvent triggers:\n   a. Auto-reject of unapproved hours (if enabled)\n   b. Vacation day recalculation\n9. Check: employees in the closed approve period can no longer be confirmed",
            "expected": "Approve period advances. Side effects fire: auto-reject + vacation recalc.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Approve Period", "module": "accounting/periods",
            "notes": "PeriodChangedEvent vs PeriodReopenedEvent: advance triggers changed, revert triggers reopened."
        },
        {
            "id": "TC-ACC-008", "title": "Revert approve period -1 month (reopen)",
            "preconditions": "Accountant. Approve period can be moved back (within 2-month lower bound).\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Changing periods\n3. Select salary office\n4. Note current approve period\n5. Click 'Edit' for Approve period\n6. Select -1 month (previous month)\n7. Click 'Save'\n8. Verify period reverts\n9. Verify PeriodReopenedEvent fires:\n   a. Previously closed month is now open for confirmation\n   b. Auto-rejected hours (if any) can be re-confirmed\n10. Login as manager, verify confirmation possible for the reopened month",
            "expected": "Approve period reverted. Closed month reopened for confirmation. No data loss.",
            "priority": "High", "type": "UI",
            "req_ref": "#3367 reopen after auto-reject", "module": "accounting/periods",
            "notes": "Lower bound: cannot go back >2 months from today. PeriodReopenedEvent fires on revert."
        },
        {
            "id": "TC-ACC-009", "title": "Period — 2+ month jump blocked in UI (#3350)",
            "preconditions": "Accountant on Changing periods page.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Changing periods\n3. Select salary office\n4. Click 'Edit' for Report period\n5. Verify months >1 ahead of saved value are DISABLED in picker\n6. Click 'Edit' for Approve period\n7. Verify months >1 ahead or >1 behind saved value are DISABLED\n8. Attempt to force submit via browser console (bypass UI)\n9. Verify 400 error from backend validation",
            "expected": "Months beyond 1-month delta are disabled in UI. Backend enforces the same rule.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3350 max 1-month delta", "module": "accounting/periods",
            "notes": "All offices now follow same rule — Persei special rule removed (#3264). Sequential saves needed for multi-month changes."
        },
        {
            "id": "TC-ACC-010", "title": "Period — confirmation month ahead of report period disabled (#3350 bug 3)",
            "preconditions": "Accountant on Changing periods.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Click 'Edit' for Approve period\n5. Verify months AHEAD of the saved Report period are DISABLED\n6. Attempt to select a month after the report period\n7. If selectable: save → verify 400 error\n8. Verify error notification appears explaining the constraint",
            "expected": "Approve period months after report period are disabled. Constraint enforced in UI and backend.",
            "priority": "High", "type": "UI",
            "req_ref": "#3350 bug 3", "module": "accounting/periods",
            "notes": "Bug #3350 finding: confirmation period month AHEAD of saved report period not disabled in UI."
        },
        {
            "id": "TC-ACC-011", "title": "Period — report month before confirmation disabled (#3365)",
            "preconditions": "Accountant on Changing periods.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Changing periods\n3. Select salary office\n4. Click 'Edit' for Report period\n5. Verify months BEFORE the saved Approve period are DISABLED\n6. Attempt to select such a month\n7. Verify it cannot be selected or 400 error on save",
            "expected": "Report period cannot precede approve period. Constraint enforced.",
            "priority": "High", "type": "UI",
            "req_ref": "#3365", "module": "accounting/periods",
            "notes": "Bug #3365 (OPEN): month preceding saved confirmation period not disabled. Partially fixed."
        },
        {
            "id": "TC-ACC-012", "title": "Period — partial save prevention (#3350 bug 4)",
            "preconditions": "Accountant on Changing periods with both periods editable.",
            "steps": "1. Login as accountant\n2. Navigate to Changing periods\n3. Select salary office\n4. Edit Report period: select +1 month\n5. Edit Approve period: select a value that creates constraint violation with the new report period\n6. Save\n7. Verify NEITHER period saves (both-or-nothing)\n8. If one saves and other fails: document as partial save bug",
            "expected": "No partial saves. Both periods validated as a pair before committing.",
            "priority": "High", "type": "UI",
            "req_ref": "#3350 bug 4", "module": "accounting/periods",
            "notes": "Known issue: one period may save while other fails validation → inconsistent state."
        },
        {
            "id": "TC-ACC-013", "title": "Period — labels and translations (#3351)",
            "preconditions": "Accountant on Changing periods in both EN and RU.",
            "steps": "1. Login as accountant\n2. Navigate to Changing periods (English)\n3. Verify page title matches spec\n4. Verify 'Edit' button label\n5. Verify tooltip/modal labels\n6. Verify datepicker month names\n7. Switch to Russian\n8. Repeat verification for all labels\n9. Note any missing or incorrect translations",
            "expected": "All labels, titles, tooltips, and datepicker content correctly translated in both languages.",
            "priority": "Low", "type": "UI",
            "req_ref": "#3351 UI cosmetics", "module": "accounting/periods",
            "notes": "#3351 (OPEN): page title rename, tooltip/modal/datepicker translation fixes."
        },
    ]


def get_payment_cases():
    """TS-Accounting-Payment: Vacation payment operations."""
    return [
        {
            "id": "TC-ACC-014", "title": "Pay single APPROVED vacation — happy path",
            "preconditions": "Accountant/admin. Employee with an APPROVED+EXACT vacation with payment month in an open period.\nQuery: SELECT v.id, e.login, v.start_date, v.end_date FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id = e.id WHERE v.status = 'APPROVED' AND v.period_type = 'EXACT' AND e.enabled = true AND EXISTS (SELECT 1 FROM ttt_vacation.employee_vacation ev WHERE ev.employee = e.id AND ev.available_vacation_days >= 0 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE)) ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant/admin\n2. Navigate to Accounting > Vacation payment page\n3. Select the salary office containing the employee\n4. Find the APPROVED vacation in the payment table\n5. Verify payment details: employee name, dates, days count, payment month\n6. Click the checkbox to select the vacation\n7. Click 'Pay' button\n8. Confirm payment in the confirmation dialog\n9. Verify success notification\n10. Verify vacation status changes to 'Paid' in the table\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — verify 'PAID'",
            "expected": "Vacation status changes to PAID. Payment recorded. Vacation disappears from pending payment list.",
            "priority": "Critical", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Payment", "module": "accounting/payment",
            "notes": "5-check payment guard: APPROVED + EXACT + payment month in approve period + office not archived + employee in office. All must pass."
        },
        {
            "id": "TC-ACC-015", "title": "Bulk pay multiple vacations",
            "preconditions": "Accountant. Multiple APPROVED+EXACT vacations in the same salary office.\nSETUP: May need to create vacations via API and approve them.",
            "steps": "SETUP: Ensure 2+ APPROVED vacations exist for different employees in same SO\n1. Login as accountant\n2. Navigate to Vacation payment page\n3. Select the salary office\n4. Select the 'Select all' checkbox (or individually check 2+ vacations)\n5. Click 'Pay' button\n6. Confirm payment\n7. Verify all selected vacations change to 'Paid'\n8. Verify vacation count decreases in the table",
            "expected": "Multiple vacations paid in one action. All selected move to PAID status.",
            "priority": "High", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Bulk Payment", "module": "accounting/payment",
            "notes": "Bulk payment uses same 5-check guard per vacation. Any failing check skips that vacation."
        },
        {
            "id": "TC-ACC-016", "title": "Payment blocked — vacation not in EXACT period type",
            "preconditions": "APPROVED vacation with period_type != EXACT (e.g., NON_EXACT).\nQuery: SELECT v.id, e.login FROM ttt_vacation.vacation v JOIN ttt_vacation.employee e ON v.employee_id = e.id WHERE v.status = 'APPROVED' AND v.period_type != 'EXACT' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant\n2. Navigate to Vacation payment page\n3. Attempt to find the non-EXACT vacation in the payment table\n4. If visible: attempt to pay\n5. Verify payment is BLOCKED with appropriate error\n6. Or: verify the vacation is not shown in the payment list at all",
            "expected": "Non-EXACT vacations cannot be paid. Either not shown or blocked with error.",
            "priority": "High", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §5-check guard", "module": "accounting/payment",
            "notes": "5-check: only EXACT period type eligible for payment. NON_EXACT created when payment month adjusted post-approval."
        },
        {
            "id": "TC-ACC-017", "title": "Payment with negative combined balance — 500 error (#3363)",
            "preconditions": "AV=true office. Employee with combined (current + next year) vacation balance < 0.\nQuery: SELECT e.login, SUM(ev.available_vacation_days) AS total FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee JOIN ttt_vacation.office o ON e.office_id = o.id WHERE o.advance_vacation = true AND e.enabled = true GROUP BY e.login HAVING SUM(ev.available_vacation_days) < 0 LIMIT 1",
            "steps": "1. Identify an AV=true employee with negative combined balance\n2. If no such employee exists: create the scenario via day corrections\n3. SETUP: Create an APPROVED vacation for this employee via API\n4. Login as accountant\n5. Navigate to Vacation payment page\n6. Find the vacation\n7. Attempt to pay\n8. Verify: if bug exists → HTTP 500 error\n9. Verify: if fixed → appropriate error message (not 500)\n10. Check: payment with current year negative but sum positive → should succeed",
            "expected": "Payment with combined negative balance shows proper error (not 500). Payment succeeds when sum is non-negative.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3363 negative balance 500", "module": "accounting/payment",
            "notes": "Bug #3363 (OPEN, Sprint 15): AV=true, combined balance <0 → 500. Works when sum ≥0. Not on production."
        },
        {
            "id": "TC-ACC-018", "title": "Auto-payment of expired approved vacations",
            "preconditions": "APPROVED vacation older than 2 months from current date (expired threshold).\nSETUP: Via test endpoint or clock manipulation.",
            "steps": "SETUP: Via test API — trigger auto-payment scheduler: POST /api/vacation/test/v1/pay-expired-approved\n1. Verify endpoint returns 200\n2. Check if the expired APPROVED vacation's status changed to PAID\n3. If auto-payment triggered: verify correct payment date and amount\nDB-CHECK: SELECT status, payment_date FROM ttt_vacation.vacation WHERE id = <id>",
            "expected": "Expired APPROVED vacations auto-paid by scheduler. Status changes to PAID with correct payment details.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "accounting-service-deep-dive.md §Auto-payment", "module": "accounting/payment",
            "notes": "2-month threshold for auto-payment of unpaid approved vacations."
        },
        {
            "id": "TC-ACC-019", "title": "Payment page — salary office filter completeness (#2841)",
            "preconditions": "Accountant with access to multiple salary offices.",
            "steps": "1. Login as accountant\n2. Navigate to Vacation payment page\n3. Open salary office filter dropdown\n4. Count offices and compare with expected count\n5. Verify ALL accessible offices are present (not subset)\n6. Navigate to Sick leave accounting page\n7. Open salary office filter\n8. Verify same completeness (bug #2841 also affects sick leave page)",
            "expected": "All accessible offices in filter on both Payment and Sick leave accounting pages.",
            "priority": "High", "type": "UI",
            "req_ref": "#2841 SO filter", "module": "accounting/payment",
            "notes": "Same filter bug as Salary page (#2841). Also test >20 SO pagination and nameless SO."
        },
    ]


def get_correction_cases():
    """TS-Accounting-DayCorrection: Vacation day manual corrections."""
    return [
        {
            "id": "TC-ACC-020", "title": "Manual day correction — add days (AV=No office)",
            "preconditions": "Accountant. Employee in AV=false (Russia) office with known balance.\nQuery: SELECT e.login, ev.available_vacation_days, o.name FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee JOIN ttt_vacation.office o ON e.office_id = o.id WHERE o.advance_vacation = false AND e.enabled = true AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant/admin\n2. Navigate to Accounting > Day corrections page (or Salary page inline edit)\n3. Find the employee from the AV=false office\n4. Note current available vacation days\n5. Click the days value to edit inline\n6. Enter a positive correction: +3\n7. Enter a comment: 'Annual adjustment'\n8. Save\n9. Verify available days increased by 3\n10. Verify the correction is logged in the events feed\nDB-CHECK: SELECT available_vacation_days FROM ttt_vacation.employee_vacation WHERE employee = <id> AND year = EXTRACT(YEAR FROM CURRENT_DATE)",
            "expected": "Vacation days increased by 3. Comment saved. Events feed shows the correction.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3283 AV-dependent correction", "module": "accounting/correction",
            "notes": "AV=No: overwork/underwork don't affect vacation days. Available days CANNOT go negative."
        },
        {
            "id": "TC-ACC-021", "title": "Manual day correction — prevent negative input for AV=No (#3283)",
            "preconditions": "Accountant. Employee in AV=false office.\nQuery: (same as TC-ACC-020 but with available_vacation_days > 0)",
            "steps": "1. Login as accountant\n2. Navigate to Day corrections\n3. Find employee in AV=false office with balance > 0 (e.g., 10)\n4. Click to edit days\n5. Attempt to enter -15 (which would make balance negative)\n6. Verify the minus sign is PREVENTED in the input field\n7. If minus accepted: save and verify DB stores ≥0 (not negative)\n8. Check events feed — verify no negative delta shown",
            "expected": "AV=No offices: minus sign prevented in correction input. Balance cannot go negative.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3283 prevent negative for AV=No", "module": "accounting/correction",
            "notes": "Bug #3283 AS-IS: system pretends to accept negative, DB stores ≥0, feed shows negative delta (inconsistency). TO-BE: prevent minus sign entirely."
        },
        {
            "id": "TC-ACC-022", "title": "Manual day correction — negative allowed for AV=Yes",
            "preconditions": "Accountant. Employee in AV=true office (Cyprus, Germany).\nQuery: SELECT e.login, ev.available_vacation_days FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee JOIN ttt_vacation.office o ON e.office_id = o.id WHERE o.advance_vacation = true AND e.enabled = true AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant\n2. Navigate to Day corrections\n3. Find employee in AV=true office\n4. Note current balance (e.g., 5)\n5. Enter correction: -8 (which makes balance negative: -3)\n6. Enter comment: 'Underwork correction'\n7. Save\n8. Verify balance shows -3 (negative displayed correctly)\n9. Verify events feed shows -8 delta\nDB-CHECK: SELECT available_vacation_days FROM ttt_vacation.employee_vacation WHERE employee = <id> AND year = EXTRACT(YEAR FROM CURRENT_DATE) — verify negative value stored",
            "expected": "AV=Yes offices: negative corrections allowed. Negative balance displayed correctly. DB stores negative value.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3283 AV=Yes allows negative", "module": "accounting/correction",
            "notes": "AV=Yes: overwork/underwork affect vacation days, negative balance CAN result."
        },
        {
            "id": "TC-ACC-023", "title": "Day correction — year-to-year distribution",
            "preconditions": "Employee with vacation days across multiple years.\nQuery: SELECT e.login, ev.year, ev.available_vacation_days FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE e.enabled = true GROUP BY e.login, ev.year, ev.available_vacation_days HAVING COUNT(*) > 1 ORDER BY random() LIMIT 5",
            "steps": "1. Login as accountant\n2. Find employee with multi-year balances (e.g., 2025: 5 days, 2026: 10 days)\n3. Apply a negative correction of -7\n4. Verify deduction follows FIFO: earliest year first\n5. Expected: 2025 → 0 days, 2026 → 8 days (5 from 2025 + 2 from 2026 = 7 deducted)\n6. Verify events feed shows correct per-year breakdown\n7. Apply a positive correction of +3\n8. Verify days added to current year",
            "expected": "Negative corrections follow FIFO (earliest year first). Positive corrections add to current year.",
            "priority": "High", "type": "UI",
            "req_ref": "#3204 year distribution", "module": "accounting/correction",
            "notes": "Bug #3204: VACATION_MONTHLY_RECALCULATION logic differs from DAYS_ADJUSTMENT distribution. Negative accrued shown as zeros."
        },
        {
            "id": "TC-ACC-024", "title": "Day correction — balance operates on sum not accrued days (architecture bug)",
            "preconditions": "Employee in AV=false office with correction history.\nQuery: (same as TC-ACC-020)",
            "steps": "1. Login as accountant\n2. Find employee with known accrued and balance values\n3. Apply a manual correction\n4. Verify: does the correction operate on 'balance sum' or 'accrued days'?\n5. If balance sum: correction may allow payment for non-accrued days\n6. Verify: manual correction event triggers redistribution between years\n7. If no redistribution: document as known architecture bug\nDB-CHECK: SELECT available_vacation_days, accrued_days FROM ttt_vacation.employee_vacation WHERE employee = <id>",
            "expected": "Correction should operate on accrued days, not balance sum. Document actual behavior.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3283 architecture findings", "module": "accounting/correction",
            "notes": "QA architecture finding: both corrections operate on balance sum instead of accrued days — fundamentally incorrect per #3283."
        },
        {
            "id": "TC-ACC-025", "title": "Day correction — comment field validation",
            "preconditions": "Accountant on correction page.",
            "steps": "1. Login as accountant\n2. Navigate to Day corrections\n3. Select an employee\n4. Enter correction value\n5. Leave comment empty → attempt save\n6. Verify if comment is required (should be — audit trail)\n7. Enter a very long comment (>255 chars)\n8. Verify backend validation (BigDecimal + comment 255 char limit)\n9. Enter valid comment and save",
            "expected": "Comment field enforces 255 char max. Verify if comment is required for corrections.",
            "priority": "Medium", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Day Correction Validation", "module": "accounting/correction",
            "notes": "Backend validation: BigDecimal for days, 255 chars max for comment."
        },
        {
            "id": "TC-ACC-026", "title": "AV=false — balance recalculation after day-off change (#3339)",
            "preconditions": "AV=false office. Employee with an active vacation that overlaps with a day-off period.\nThis is an edge case requiring specific calendar setup.",
            "steps": "1. Identify or create scenario: AV=false employee with vacation covering a date that also has a day-off\n2. Delete the day-off (via calendar admin)\n3. Verify vacation type conversion occurs (if applicable)\n4. Verify balance recalculation result is CORRECT (not 0)\n5. Check: does balance drop to 0 incorrectly? (#3339 bug)\n6. Transfer a day-off within the vacation period\n7. Verify recalculation is correct after transfer",
            "expected": "Balance recalculation after day-off changes produces correct values. No incorrect drop to 0.",
            "priority": "High", "type": "UI",
            "req_ref": "#3339 AV=false recalculation", "module": "accounting/correction",
            "notes": "Bug #3339: delete/transfer day-off within vacation → conversion results in 0 balance instead of correct value."
        },
    ]


def get_views_cases():
    """TS-Accounting-Views: Employee lists, employee details, departing employee tooltip."""
    return [
        {
            "id": "TC-ACC-027", "title": "Salary page — employee list with correct data",
            "preconditions": "Accountant viewing Salary page.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Salary page\n3. Select a salary office\n4. Verify employee list shows: name, vacation days, report status, payment info\n5. Verify data matches DB values\n6. Click an employee name\n7. Verify employee details open (either popup or separate page)\n8. Verify details include: current vacation balance, pending requests, payment history",
            "expected": "Employee list accurate. Details accessible with correct vacation/payment data.",
            "priority": "High", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Salary Page", "module": "accounting/views",
            "notes": "Salary page is the primary accounting interface for managing vacation payments and corrections."
        },
        {
            "id": "TC-ACC-028", "title": "Departing employee tooltip (#3336)",
            "preconditions": "Employee who is in the process of being dismissed (beingDismissed=true from CS).\nQuery: SELECT e.login, e.last_date FROM ttt_backend.employee e WHERE e.last_date IS NOT NULL AND e.last_date > CURRENT_DATE AND e.enabled = true ORDER BY e.last_date ASC LIMIT 1",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Salary page\n3. Find the departing employee in the list\n4. Verify a tooltip or indicator shows 'Date of leave: {date}'\n5. Verify the tooltip appears stably (not flicker/disappear)\n6. Check after CS sync removes the employee: tooltip should disappear\n7. Note timing: tooltip shows while beingDismissed=true, disappears when employee removed from CS list",
            "expected": "Departing employee shows leave date tooltip. Tooltip stable while beingDismissed flag is true.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3336 leave date tooltip", "module": "accounting/views",
            "notes": "Bug #3336: tooltip appears then disappears when CS sync removes employee. Timing issue between dismissal and salary calculation."
        },
        {
            "id": "TC-ACC-029", "title": "Vacation days — per-year grouped view",
            "preconditions": "Employee with vacation days across multiple years.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE e.enabled = true GROUP BY e.login HAVING COUNT(DISTINCT ev.year) > 1 ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant\n2. Navigate to Salary page, find the multi-year employee\n3. View their vacation day balance\n4. Verify per-year breakdown is shown\n5. Verify total matches sum of all years\n6. Compare with GET /api/vacation/v1/vacationDays/{login}/grouped API response\n7. Verify accrued vs available distinction (if shown)",
            "expected": "Per-year vacation day breakdown visible. Total matches DB sum. Accrued/available correctly distinguished.",
            "priority": "Medium", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Day Calculation", "module": "accounting/views",
            "notes": "API endpoint: GET /v1/vacationDays/{login}/grouped. Bug #3204: negative accrued shown as zeros."
        },
        {
            "id": "TC-ACC-030", "title": "Accounting page — regular employee access blocked (#3012)",
            "preconditions": "Regular employee with ROLE_EMPLOYEE only.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_EMPLOYEE' AND e.enabled = true AND NOT EXISTS (SELECT 1 FROM ttt_backend.employee_role er2 JOIN ttt_backend.role r2 ON er2.role_id = r2.id WHERE er2.employee_id = e.id AND r2.name NOT IN ('ROLE_EMPLOYEE', 'ROLE_CONTRACTOR')) ORDER BY random() LIMIT 1",
            "steps": "1. Login as regular employee\n2. Navigate directly to /accounting/sick-leaves\n3. Verify access is DENIED (redirect or error page)\n4. Navigate to /accounting/salary\n5. Verify access denied\n6. Navigate to /accounting/periods\n7. Verify access denied\n8. Verify the Accounting menu item is NOT visible in the navigation",
            "expected": "Regular employee cannot access any accounting routes. Menu item hidden. Direct URL access blocked.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3012 accounting route exposed", "module": "accounting/views",
            "notes": "Bug #3012 (CLOSED): /accounting/sick-leaves was accessible to regular employees using VACATIONS:VIEW instead of SICK_LEAVE_ACCOUNTING_VIEW."
        },
        {
            "id": "TC-ACC-031", "title": "Sick leave accounting page — list and filter",
            "preconditions": "Accountant with access to sick leave accounting.\nQuery: (same as TC-ACC-001)",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Sick leave accounting\n3. Verify page loads with employee sick leave list\n4. Verify salary office filter works\n5. Verify columns: employee name, sick leave dates, status, duration\n6. Verify filter shows ALL accessible offices (#2841)\n7. Select an office and verify correct employee list",
            "expected": "Sick leave accounting page loads with correct data. SO filter works completely.",
            "priority": "High", "type": "UI",
            "req_ref": "#2841 same filter bug", "module": "accounting/views",
            "notes": "Same SO filter bug as Salary page (#2841). SICK_LEAVE_ACCOUNTING_VIEW permission required."
        },
    ]


def get_sickleave_accounting_cases():
    """TS-Accounting-SickLeave: Sick leave accounting operations."""
    return [
        {
            "id": "TC-ACC-032", "title": "Sick leave — close (mark as closed by accountant)",
            "preconditions": "Accountant. Employee with an open sick leave.\nQuery: SELECT sl.id, e.login FROM ttt_vacation.sick_leave sl JOIN ttt_vacation.employee e ON sl.employee_id = e.id WHERE sl.status IN ('OPEN', 'ISSUED') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant\n2. Navigate to Accounting > Sick leave accounting\n3. Find the employee with open sick leave\n4. Click on the sick leave row\n5. Click 'Close' or change status to 'Closed'\n6. Confirm the action\n7. Verify status changes to CLOSED\n8. Verify the sick leave duration is finalized\nDB-CHECK: SELECT status, end_date FROM ttt_vacation.sick_leave WHERE id = <id>",
            "expected": "Sick leave closed by accountant. Status = CLOSED. End date finalized.",
            "priority": "High", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Sick Leave Accounting", "module": "accounting/sick-leave",
            "notes": "Sick leave has dual status model (internal + external). Accountant operates on external status."
        },
        {
            "id": "TC-ACC-033", "title": "Sick leave — overdue notification count",
            "preconditions": "Accountant. Sick leaves past expected end date without closure.\nQuery: SELECT COUNT(*) FROM ttt_vacation.sick_leave WHERE status IN ('OPEN', 'ISSUED') AND start_date < CURRENT_DATE - INTERVAL '14 days'",
            "steps": "1. Login as accountant\n2. Navigate to Sick leave accounting page\n3. Verify overdue counter/badge shows count of overdue sick leaves\n4. Compare with DB query result\n5. Click on overdue filter\n6. Verify only overdue sick leaves displayed",
            "expected": "Overdue count matches DB. Filter shows only overdue sick leaves.",
            "priority": "Medium", "type": "UI",
            "req_ref": "accounting-service-deep-dive.md §Overdue", "module": "accounting/sick-leave",
            "notes": "API endpoint: GET /v1/sick-leaves/overdue-count."
        },
    ]


def get_api_cases():
    """TS-Accounting-API: API-level tests for accounting endpoints."""
    return [
        {
            "id": "TC-ACC-034", "title": "Period API — missing first-day-of-month validation on approve period",
            "preconditions": "API access with accountant/admin token.",
            "steps": "1. Via API — PATCH /api/ttt/v1/{officeId}/periods/approve with start=2026-03-15 (not first of month)\n2. Verify: does the request succeed? (BUG: missing validation)\n3. If succeeds: verify DB has 2026-03-15 as approve period start\n4. If fails: verify proper 400 error with clear message\n5. Compare with report period: PATCH /api/ttt/v1/{officeId}/periods/report with start=2026-03-15\n6. Verify report period correctly validates first-of-month (getDayOfMonth() != 1 check)\nCLEANUP: If bug confirmed, restore valid period via PATCH",
            "expected": "Approve period should reject non-first-of-month dates. Report period already validates correctly.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "BUG-PERIOD-1 missing validation", "module": "accounting/api",
            "notes": "BUG-PERIOD-1 (HIGH): Approve period missing first-day-of-month validation at line 104. Report period has the check."
        },
        {
            "id": "TC-ACC-035", "title": "Period API — NPE on null start (BUG-PERIOD-2)",
            "preconditions": "API access with accountant/admin token.",
            "steps": "1. Via API — PATCH /api/ttt/v1/{officeId}/periods/approve with body: {}\n2. Verify: does the request return 500 NullPointerException?\n3. If 500: confirm BUG-PERIOD-2 (missing null check on start field)\n4. If 400: bug is fixed with proper validation\n5. Repeat for report period with empty body\nCLEANUP: No changes made (read-only test)",
            "expected": "Empty body should return 400 (validation error), not 500 (NPE).",
            "priority": "High", "type": "Hybrid",
            "req_ref": "BUG-PERIOD-2 NPE", "module": "accounting/api",
            "notes": "BUG-PERIOD-2 (HIGH): PATCH with {} → 500 NullPointerException. Missing null check on request.getStart()."
        },
        {
            "id": "TC-ACC-036", "title": "Period API — stack trace leakage (BUG-PERIOD-3)",
            "preconditions": "API access.",
            "steps": "1. Via API — PATCH /api/ttt/v1/{officeId}/periods/approve with start='invalid-date'\n2. Check response body for Java stack traces\n3. If stack trace present: confirm BUG-PERIOD-3\n4. Verify: error response should be a clean JSON error, not raw exception details\n5. Repeat with other invalid inputs: start='null', start=12345",
            "expected": "Invalid date format returns clean 400 error without exposing internal stack traces.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "BUG-PERIOD-3 stack trace", "module": "accounting/api",
            "notes": "BUG-PERIOD-3: Invalid date format returns full Java trace. Security: stack traces expose internal implementation details."
        },
        {
            "id": "TC-ACC-037", "title": "Period API — permission inconsistency (BUG-PERIOD-4)",
            "preconditions": "API access with multiple token types.",
            "steps": "1. Via API — GET /api/ttt/v1/periods/report/offices/min — note: requires JWT only\n2. Via API — GET /api/ttt/v1/periods/approve/offices/min — note: accepts JWT AND API token\n3. Verify the permission difference: report min/max requires JWT only; approve min/max accepts both\n4. Document the inconsistency\n5. Test with API token only (no JWT): verify report min/max fails, approve min/max succeeds",
            "expected": "Document permission inconsistency between report and approve period endpoints.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "BUG-PERIOD-4 permission inconsistency", "module": "accounting/api",
            "notes": "BUG-PERIOD-4: inconsistent auth requirements between report and approve period min/max endpoints."
        },
        {
            "id": "TC-ACC-038", "title": "Vacation day correction API — negative value handling (AV-dependent)",
            "preconditions": "API access. Employees in both AV=true and AV=false offices.",
            "steps": "1. For AV=false employee: POST /api/vacation/v1/vacationDays/{login}/correction with days=-5\n2. Verify: DB stores ≥0 (negative part lost) — AS-IS bug from #3283\n3. For AV=true employee: POST correction with days=-5\n4. Verify: DB correctly stores negative value\n5. Compare: manual correction vs automatic correction handling of negatives\n6. Check events feed: does the feed show the correct delta (including negative)?",
            "expected": "AV=false: DB stores ≥0 (negative ignored). AV=true: DB correctly stores negative. Feed shows actual delta.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "#3283 negative handling", "module": "accounting/api",
            "notes": "#3283: DB stores ≥0 for manual, feed preserves negative. Automatic corrections DO write negatives to DB."
        },
    ]


# ─── Feature Matrix ──────────────────────────────────────────────────────────

FEATURES = [
    {"feature": "Salary Page & Office Filter", "suites": ["TS-Accounting-Periods"], "ui": 4, "hybrid": 0, "api": 0},
    {"feature": "Period Management", "suites": ["TS-Accounting-Periods"], "ui": 9, "hybrid": 0, "api": 0},
    {"feature": "Vacation Payment", "suites": ["TS-Accounting-Payment"], "ui": 5, "hybrid": 1, "api": 0},
    {"feature": "Vacation Day Corrections", "suites": ["TS-Accounting-DayCorrection"], "ui": 7, "hybrid": 0, "api": 0},
    {"feature": "Accounting Views & Access", "suites": ["TS-Accounting-Views"], "ui": 5, "hybrid": 0, "api": 0},
    {"feature": "Sick Leave Accounting", "suites": ["TS-Accounting-SickLeave"], "ui": 2, "hybrid": 0, "api": 0},
    {"feature": "API Security & Validation", "suites": ["TS-Accounting-API"], "ui": 0, "hybrid": 5, "api": 0},
]


# ─── Risk Assessment ─────────────────────────────────────────────────────────

RISKS = [
    {"feature": "Period 1-month delta rule", "risk": "6 QA bugs found in #3350 alone. UI uses dynamic (unsaved) values for validation, enabling partial saves and constraint violations.", "likelihood": "High", "impact": "Critical", "severity": "Critical", "mitigation": "Comprehensive saved-vs-dynamic testing. Both periods validated as pair. Sequential change testing."},
    {"feature": "AV-dependent correction logic", "risk": "Fundamental architecture bug: corrections operate on balance sum instead of accrued days (#3283). Manual vs automatic corrections behave differently for negatives.", "likelihood": "High", "impact": "Critical", "severity": "Critical", "mitigation": "Test both AV=true and AV=false offices. Verify DB values vs UI display. Check year distribution."},
    {"feature": "Negative balance payment", "risk": "500 error when paying vacation with combined negative balance in AV=true offices (#3363).", "likelihood": "Medium", "impact": "High", "severity": "High", "mitigation": "Test payment with various balance states. Verify error handling for negative combined balance."},
    {"feature": "Salary office filter", "risk": "localStorage corruption shows subset of offices (#2841). Affects Salary, Payment, and Sick Leave pages.", "likelihood": "High", "impact": "High", "severity": "High", "mitigation": "Test with cleared localStorage. Verify >20 SO pagination. Test ACCOUNTANT+ADMIN dual role."},
    {"feature": "Archived SO handling", "risk": "New feature (#3323): archived SO must be hidden from Accounting but visible in Admin.", "likelihood": "Medium", "impact": "Medium", "severity": "Medium", "mitigation": "Verify archived SO visibility across all dropdowns. Test Admin panel display (dimmed, non-editable)."},
    {"feature": "Period API validation gaps", "risk": "Missing first-of-month validation on approve period (BUG-PERIOD-1), NPE on null start (BUG-PERIOD-2), stack trace leakage.", "likelihood": "High", "impact": "Medium", "severity": "Medium", "mitigation": "API boundary testing with invalid inputs. Verify no stack traces in responses."},
    {"feature": "Auto-payment scheduler", "risk": "Expired APPROVED vacations auto-paid after 2-month threshold. May affect accounting calculations if triggered unexpectedly.", "likelihood": "Low", "impact": "Medium", "severity": "Medium", "mitigation": "Verify auto-payment threshold. Test with expired vacations. Check payment date correctness."},
    {"feature": "Employee access to accounting routes", "risk": "Bug #3012: accounting/sick-leaves was accessible to regular employees due to wrong permission check.", "likelihood": "Low", "impact": "High", "severity": "Medium", "mitigation": "Verify all accounting routes blocked for regular employees. Direct URL and menu check."},
    {"feature": "Departing employee display", "risk": "Tooltip flicker due to CS sync timing (#3336). Read-only flag not set during dismissal.", "likelihood": "Low", "impact": "Low", "severity": "Low", "mitigation": "Check tooltip stability. Verify timing between dismissal and CS sync."},
    {"feature": "Day correction comment validation", "risk": "255 char limit on comment. No validation for empty comment despite audit trail need.", "likelihood": "Low", "impact": "Low", "severity": "Low", "mitigation": "Test max length and empty comment. Verify audit trail completeness."},
]


# ─── Helper Functions (same as reports generator) ─────────────────────────────

def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, value, is_alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if is_alt else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    return cell


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_link(ws):
    cell = ws.cell(row=1, column=1, value="← Back to Plan Overview")
    cell.font = FONT_BACK_LINK
    cell.hyperlink = "#'Plan Overview'!A1"


# ─── Sheet Generators ────────────────────────────────────────────────────────

def create_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="Accounting — Test Plan").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Phase B Session 93").font = FONT_BODY
    ws.cell(row=3, column=1, value="Module: Accounting (Periods, Payment, Corrections, Views) | Branch: release/2.1").font = FONT_BODY

    row = 5
    ws.cell(row=row, column=1, value="Scope & Objectives").font = FONT_SUBTITLE
    row += 1
    objectives = [
        "Period management: report period, approve period, 1-month max delta, saved vs dynamic validation",
        "Vacation payment: 5-check guard, bulk payment, negative balance handling, auto-payment",
        "Vacation day corrections: AV-dependent logic (AV=Yes allows negative, AV=No prevents it)",
        "Accounting views: Salary page, employee details, departing employee tooltip, per-year balance",
        "Sick leave accounting: close sick leaves, overdue tracking",
        "API security: period validation gaps, permission inconsistencies, stack trace leakage",
        "Cross-cutting: salary office filter completeness, archived SO hiding, access control",
    ]
    for obj in objectives:
        ws.cell(row=row, column=1, value=f"• {obj}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Knowledge Sources").font = FONT_SUBTITLE
    row += 1
    sources = [
        "Code analysis: accounting-service-deep-dive.md (17.1KB, 13 design issues)",
        "GitLab tickets: 62+ searched, 12 deep-read — 5 theme clusters",
        "API testing: accounting-api-testing (5 bugs found), period-api-live-testing",
        "UI testing: accounting-pages (5 pages explored), payment/correction live testing",
        "DB analysis: vacation-day-calculation-architecture, office-period-model",
        "Business rules: vacation-business-rules-reference (AV-dependent logic), office-period-model",
    ]
    for src in sources:
        ws.cell(row=row, column=1, value=f"• {src}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1

    suite_data = [
        ("TS-Accounting-Periods", "Salary & Period Management", 13),
        ("TS-Accounting-Payment", "Vacation Payment", 6),
        ("TS-Accounting-DayCorrection", "Day Corrections", 7),
        ("TS-Accounting-Views", "Views & Access Control", 5),
        ("TS-Accounting-SickLeave", "Sick Leave Accounting", 2),
        ("TS-Accounting-API", "API Security & Validation", 5),
    ]

    for suite_id, suite_name, count in suite_data:
        cell = ws.cell(row=row, column=1, value=f"{suite_id}: {suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{suite_id}'!A1"
        row += 1

    row += 1
    total = sum(s[2] for s in suite_data)
    ws.cell(row=row, column=1, value=f"Total Test Cases: {total}").font = FONT_SUBTITLE

    ws.column_dimensions["A"].width = 100


def create_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Test Suites", "UI Tests", "Hybrid Tests", "API Tests", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, f in enumerate(FEATURES):
        row = i + 2
        is_alt = i % 2 == 1
        total = f["ui"] + f["hybrid"] + f["api"]
        style_body_cell(ws, row, 1, f["feature"], is_alt)
        cell = style_body_cell(ws, row, 2, ", ".join(f["suites"]), is_alt)
        if f["suites"]:
            cell.font = FONT_LINK
            cell.hyperlink = f"#'{f['suites'][0]}'!A1"
        style_body_cell(ws, row, 3, f["ui"], is_alt)
        style_body_cell(ws, row, 4, f["hybrid"], is_alt)
        style_body_cell(ws, row, 5, f["api"], is_alt)
        style_body_cell(ws, row, 6, total, is_alt)

    set_column_widths(ws, [35, 28, 12, 14, 12, 10])


def create_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_HIGH, "Medium": FILL_RISK_MED, "Low": FILL_RISK_LOW}

    for i, r in enumerate(RISKS):
        row = i + 2
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, r["feature"], is_alt)
        style_body_cell(ws, row, 2, r["risk"], is_alt)
        style_body_cell(ws, row, 3, r["likelihood"], is_alt)
        style_body_cell(ws, row, 4, r["impact"], is_alt)
        sev_cell = style_body_cell(ws, row, 5, r["severity"], is_alt)
        sev_cell.fill = severity_fills.get(r["severity"], FILL_ROW_WHITE)
        style_body_cell(ws, row, 6, r["mitigation"], is_alt)

    set_column_widths(ws, [28, 55, 12, 12, 12, 55])


def create_suite_sheet(wb, suite_id, cases):
    ws = wb.create_sheet(suite_id)
    ws.sheet_properties.tabColor = TAB_COLOR_SUITE

    add_back_link(ws)

    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result", "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, tc in enumerate(cases):
        row = i + 3
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, tc["id"], is_alt)
        style_body_cell(ws, row, 2, tc["title"], is_alt)
        style_body_cell(ws, row, 3, tc["preconditions"], is_alt)
        style_body_cell(ws, row, 4, tc["steps"], is_alt)
        style_body_cell(ws, row, 5, tc["expected"], is_alt)
        prio_cell = style_body_cell(ws, row, 6, tc["priority"], is_alt)
        if tc["priority"] == "Critical":
            prio_cell.fill = FILL_RISK_HIGH
        style_body_cell(ws, row, 7, tc["type"], is_alt)
        style_body_cell(ws, row, 8, tc["req_ref"], is_alt)
        style_body_cell(ws, row, 9, tc["module"], is_alt)
        style_body_cell(ws, row, 10, tc.get("notes", ""), is_alt)

    set_column_widths(ws, [14, 40, 50, 70, 45, 10, 8, 35, 22, 40])
    ws.auto_filter.ref = f"A2:J{len(cases) + 2}"


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    create_plan_overview(wb)
    create_feature_matrix(wb)
    create_risk_assessment(wb)

    suite_generators = [
        ("TS-Accounting-Periods", get_periods_cases),
        ("TS-Accounting-Payment", get_payment_cases),
        ("TS-Accounting-DayCorrection", get_correction_cases),
        ("TS-Accounting-Views", get_views_cases),
        ("TS-Accounting-SickLeave", get_sickleave_accounting_cases),
        ("TS-Accounting-API", get_api_cases),
    ]

    total = 0
    for suite_id, gen_fn in suite_generators:
        cases = gen_fn()
        create_suite_sheet(wb, suite_id, cases)
        total += len(cases)
        print(f"  {suite_id}: {len(cases)} cases")

    wb.save(OUTPUT_FILE)
    print(f"\nGenerated {OUTPUT_FILE}")
    print(f"Total: {total} test cases across {len(suite_generators)} suites")


if __name__ == "__main__":
    main()
