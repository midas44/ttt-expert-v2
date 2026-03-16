#!/usr/bin/env python3
"""Generate supplementary test cases from Session 67 investigations.

Covers:
1. Maternity leave lifecycle (10 cases → vacation.xlsx TS-Vac-Maternity)
2. CS office unimplemented settings (4 cases → vacation.xlsx TS-Vac-CSSettings)
3. Calendar migration edge cases (4 cases → vacation.xlsx TS-Vac-CalendarMigr)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling constants (match existing workbooks) ──────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ───────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    hr = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, t in enumerate(test_cases):
        row = hr + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            t["id"], t["title"], t["preconditions"], t["steps"],
            t["expected"], t["priority"], t["type"],
            t["req_ref"], t["module"], t.get("notes", "")
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = THIN_BORDER
            c.fill = fill

    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{ws.max_row}"
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# MATERNITY LIFECYCLE TEST CASES (TC-VAC-143 to TC-VAC-152)
# ══════════════════════════════════════════════════════════════════

MATERNITY_TESTS = [
    tc("TC-VAC-143",
       "Maternity begin — proportional current year day reduction",
       "Employee in AV=false office (norm=28 days/year).\n"
       "Employee has 28 available days for current year.\n"
       "Current date: September 15 (3.5 months remaining in year).",
       "1. Record employee_vacation.available_vacation_days before maternity.\n"
       "2. Trigger CS sync with maternity=true for the employee.\n"
       "3. DB: Check timeline for MATERNITY_BEGIN event.\n"
       "4. DB: Check employee_vacation for current year.\n"
       "5. Verify formula: subtract = 28 × (108 / 365) = 8.3 → HALF_UP → 8.\n"
       "6. Expected current year days: 28 - 8 = 20.",
       "Current year days reduced from 28 to 20.\n"
       "Timeline event MATERNITY_BEGIN with days_accrued = -8.\n"
       "Formula: annualDays × remainingDays / totalYearDays, HALF_UP rounding.\n"
       "Remaining days counted from today to Dec 31 inclusive.",
       "High", "Functional",
       "#3370", "Vacation / Maternity",
       "EmployeeMaternityBeginEventListener. Real data: afanaseva maternity Sep 15 → "
       "accrued=-7 (28 × 107/365 = 8.2 → 8? Verify rounding)."),

    tc("TC-VAC-144",
       "Maternity begin — next year days zeroed",
       "Employee has employee_vacation record for next year with 28 days.\n"
       "Employee enters maternity leave.",
       "1. DB: Verify employee_vacation for next year = 28 before maternity.\n"
       "2. Trigger maternity begin event.\n"
       "3. DB: Check employee_vacation for next year.\n"
       "4. Verify next year days = 0.",
       "Next year available_vacation_days set to exactly 0.\n"
       "Timeline event records this change.",
       "High", "Functional",
       "#3370", "Vacation / Maternity",
       "Production bug V2.1.25: race condition with annual accrual job re-set "
       "next year to 24 after zeroing. Verify no race."),

    tc("TC-VAC-145",
       "Maternity end — proportional current year day restoration",
       "Employee on maternity leave (maternity=true).\n"
       "Office norm = 24 days/year.\n"
       "Maternity ends October 28 (~2 months remaining).",
       "1. DB: Record employee_vacation for current year before maternity end.\n"
       "2. Trigger CS sync with maternity=false.\n"
       "3. DB: Check timeline for MATERNITY_END event.\n"
       "4. DB: Check employee_vacation for current year.\n"
       "5. Verify: restored days = vacationDaysHelpService.calculate(officeId, Oct28).\n"
       "6. Expected: ~24 × (64/365) ≈ 4 days added.",
       "Current year days increased by ~4.\n"
       "Timeline event MATERNITY_END with days_accrued = +4.\n"
       "Calculation uses VacationDaysHelpService.calculate(officeId, endDate).\n"
       "Proportional accrual from maternity end date to year end.",
       "High", "Functional",
       "#3370", "Vacation / Maternity",
       "Real data: apliskina maternity end Oct 28 → accrued=+4. Confirmed."),

    tc("TC-VAC-146",
       "Maternity end — next year restored to full office allocation",
       "Employee on maternity with next year days = 0.\n"
       "Office annual leave for next year = 24.",
       "1. DB: Verify employee_vacation[next_year] = 0.\n"
       "2. Trigger maternity end event.\n"
       "3. DB: Check employee_vacation[next_year].\n"
       "4. Verify = office annual leave norm (24).",
       "Next year days restored to full office annual leave allocation (24).\n"
       "Uses officeAnnualLeaveService.getDays(officeId, nextYear).",
       "High", "Functional",
       "#3370", "Vacation / Maternity",
       "Real data: dshipacheva maternity end Dec 27 → 2026 days = 24 (full restore)."),

    tc("TC-VAC-147",
       "Maternity end near year-end — 0 days restored for current year",
       "Employee on maternity leave.\n"
       "Maternity ends December 27 (4 days before year end).",
       "1. Set clock to December 27.\n"
       "2. Trigger maternity end event.\n"
       "3. DB: Check timeline MATERNITY_END event.\n"
       "4. DB: Check employee_vacation for current year.\n"
       "5. Verify: calculate(officeId, Dec27) ≈ 24 × 4/365 ≈ 0.\n"
       "6. DB: Check employee_vacation for next year = full allocation.",
       "Current year: days_accrued = 0 (too few remaining days).\n"
       "Next year: full allocation restored (24).\n"
       "Edge case: year-end maternity end adds 0 current-year days but full next-year.",
       "Medium", "Boundary",
       "#3370", "Vacation / Maternity",
       "Real data: dshipacheva maternity end 2025-12-27, accrued=0, 2026=24."),

    tc("TC-VAC-148",
       "Multi-year maternity — vacation day check across years",
       "Employee on maternity for 2+ years (e.g., 2023 to 2025).\n"
       "Has accumulated vacation days from before maternity.\n"
       "employee.maternity = true.",
       "1. DB: SELECT year, available_vacation_days FROM employee_vacation WHERE employee = X.\n"
       "2. GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "3. Verify available = SUM of all years' available_vacation_days.\n"
       "4. Verify per-year view shows individual year balances.\n"
       "5. Create a vacation request — verify availability check uses total.",
       "Available days = SUM across ALL years (not single-year check).\n"
       "VacationAvailabilityChecker.hasEnoughDaysForMaternityEmployee() sums all years.\n"
       "Employee can take vacation if total > 0 even if current year = 0.",
       "Medium", "Functional",
       "#3370", "Vacation / Maternity",
       "VacationAvailabilityChecker: maternity employee uses total across all years "
       "instead of per-year strategy. Real data: iromanenko maternity since 2024-06."),

    tc("TC-VAC-149",
       "Maternity — negative vacation days edge case",
       "Employee enters maternity with negative vacation balance.\n"
       "Current year available_vacation_days = -1.\n"
       "(Used more vacation than allocated before maternity)",
       "1. DB: Verify employee_vacation shows negative balance.\n"
       "2. Trigger maternity begin event.\n"
       "3. DB: Check current year days after reduction.\n"
       "4. Verify: balance goes further negative (current - subtractDays).\n"
       "5. DB: Check next year zeroed.\n"
       "6. Verify no error or exception from negative balance.",
       "Maternity begin proceeds normally even with negative balance.\n"
       "Current year days become more negative.\n"
       "Next year zeroed to 0.\n"
       "No validation error — system allows negative days during maternity.",
       "Medium", "Boundary",
       "#3370", "Vacation / Maternity",
       "Real data: mgrishkevich has -1 days while on maternity. "
       "zmustafina also at -1. System does not prevent this."),

    tc("TC-VAC-150",
       "Same-day maternity begin and end (CS sync correction)",
       "CS sync first sets maternity=true, then a correction sync sets maternity=false.\n"
       "Both events occur within minutes (same-day correction).",
       "1. Trigger CS sync with maternity=true.\n"
       "2. Wait for MATERNITY_BEGIN event processing.\n"
       "3. Immediately trigger CS sync with maternity=false.\n"
       "4. Wait for MATERNITY_END event processing.\n"
       "5. DB: Check employee_vacation for current and next year.\n"
       "6. DB: Check timeline for both events.",
       "Both events processed: BEGIN then END.\n"
       "Days reduced then restored — should approximately return to original.\n"
       "Possible rounding discrepancy: HALF_UP on reduction, different formula on restoration.\n"
       "Timeline shows both events with opposite accrued_days values.",
       "Low", "Edge Case",
       "#3370", "Vacation / Maternity",
       "Real data: dprotopopova had BEGIN + END on 2024-09-25 (1.25 hours apart). "
       "accrued=-6 then +6. Same-day correction scenario."),

    tc("TC-VAC-151",
       "Maternity begin — annual accrual race condition (V2.1.25 bug)",
       "Employee enters maternity in November.\n"
       "Next year days zeroed by MATERNITY_BEGIN event.\n"
       "Annual accrual job runs shortly after.",
       "1. Set clock to November.\n"
       "2. Trigger CS sync with maternity=true.\n"
       "3. Verify next year days = 0.\n"
       "4. Trigger annual accrual job (POST /v1/test/vacation/annual-accruals).\n"
       "5. DB: Check employee_vacation for next year.\n"
       "6. Verify next year days still = 0 (not re-set to 24).",
       "After annual accrual job, next year days remain 0.\n"
       "Accrual job must skip maternity employees' next year.\n"
       "If days reset to 24: PRODUCTION BUG (V2.1.25 regression).",
       "High", "Regression",
       "V2.1.25, #3370", "Vacation / Maternity",
       "Production bug: mgrishkevich, edobrovlyanskaya, afanaseva had 2026 days "
       "incorrectly set to 24 after maternity begin. Fixed by manual migration."),

    tc("TC-VAC-152",
       "Office change during active maternity",
       "Employee on maternity leave (maternity=true).\n"
       "CS sync changes employee's salary office from Office A (norm=28) to Office B (norm=24).",
       "1. DB: Record current employee_vacation days.\n"
       "2. Trigger CS sync with new office (different norm).\n"
       "3. DB: Check employee.office_id updated.\n"
       "4. DB: Check employee_vacation days — are they recalculated?\n"
       "5. GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "6. Check: which office norm would be used if maternity ends now?",
       "Office change processed (employee.office_id updated).\n"
       "Current vacation days may NOT be recalculated during maternity.\n"
       "When maternity ends: NEW office norm used for restoration.\n"
       "Potential discrepancy: days reduced with old norm, restored with new norm.",
       "Medium", "Integration",
       "#2876, #3370", "Vacation / Maternity + Office Sync",
       "Interaction between EmployeeOfficeChangedProcessor and maternity state. "
       "If norm differs: days were subtracted at old rate, restored at new rate."),
]


# ══════════════════════════════════════════════════════════════════
# CS OFFICE SETTINGS TEST CASES (TC-VAC-153 to TC-VAC-156)
# ══════════════════════════════════════════════════════════════════

CS_SETTINGS_TESTS = [
    tc("TC-VAC-153",
       "First vacation restriction — verify 3-month hardcoded waiting period",
       "New employee hired within the last 3 months.\n"
       "Employee has vacation days available (e.g., AV=true office).",
       "1. Find or create employee with hire date < 3 months ago.\n"
       "2. POST /api/vacation/v1/vacation/create with valid vacation request.\n"
       "3. Check response for error about first vacation restriction.\n"
       "4. Set clock forward to 3 months + 1 day after hire.\n"
       "5. POST same vacation request.\n"
       "6. Verify success.",
       "Before 3 months: vacation request rejected (if restriction enforced).\n"
       "After 3 months: vacation request accepted.\n"
       "NOTE: Verify if this restriction actually exists in code — #3026 says hardcoded to 3 "
       "but the actual enforcement point was not found in code analysis.",
       "Medium", "Functional",
       "#3026", "Vacation / CS Office Settings",
       "CSSalaryOfficeVacationData.firstVacation exists in CS model but is NOT used in TTT. "
       "The 3-month restriction may be enforced differently or not at all."),

    tc("TC-VAC-154",
       "Vacation days carry-over — verify no expiration (burnOff unused)",
       "Employee with unused vacation days from 2+ years ago.\n"
       "Office has burnOff value set in CS but NOT used by TTT.",
       "1. DB: Find employee with employee_vacation records for year Y-2.\n"
       "2. DB: Verify available_vacation_days > 0 for year Y-2.\n"
       "3. GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "4. Verify: pastYearDays includes balance from Y-2 (no expiration).\n"
       "5. Create vacation consuming Y-2 days.\n"
       "6. Verify success — old days accepted.",
       "Vacation days from 2+ years ago are still available.\n"
       "No expiration logic applied — days carry over indefinitely.\n"
       "burnOff value from CS is completely ignored.\n"
       "FIFO consumption may apply (oldest year consumed first for AV=true).",
       "Medium", "Functional",
       "#3026", "Vacation / CS Office Settings",
       "CSSalaryOfficeVacationData.burnOff is defined but NOT synced or used. "
       "When implemented: test that days expire after configured months."),

    tc("TC-VAC-155",
       "CS sync — unused fields not causing errors",
       "CS returns all vacation data fields including firstVacation, burnOff, sickLeave.\n"
       "TTT only uses days, advanceVacation, overworkUnderwork.",
       "1. Trigger CS sync: POST /v1/test/employee/sync.\n"
       "2. Verify sync completes without errors.\n"
       "3. DB: Check office table — no new columns for unused fields.\n"
       "4. DB: Check office.advance_vacation correctly updated.\n"
       "5. Check application logs for no warnings about unused fields.\n"
       "6. Verify office_annual_leave.days correctly synced.",
       "Sync completes successfully.\n"
       "Unused CS fields (firstVacation, burnOff, sickLeave) silently ignored.\n"
       "Used fields (days, advanceVacation, overworkUnderwork) correctly synced.\n"
       "No errors or warnings in application logs.",
       "Low", "Integration",
       "#3026", "Vacation / CS Office Sync",
       "CSSalaryOfficeSynchronizer only processes 3 of 7 CS vacation fields. "
       "Remaining 4 ignored without error."),

    tc("TC-VAC-156",
       "Contractor sick leave eligibility — sickLeave setting unused",
       "Employee with contractor type in an office where CS has sickLeave=true.\n"
       "TTT does not use this setting.",
       "1. Find contractor employee in the system.\n"
       "2. POST /api/vacation/v1/sickleave/create for the contractor.\n"
       "3. Check response — is contractor allowed to create sick leave?\n"
       "4. Compare with non-contractor in same office.\n"
       "5. Verify behavior is NOT affected by CS sickLeave setting.",
       "Contractor sick leave behavior is uniform regardless of CS sickLeave value.\n"
       "Either all contractors can create sick leaves or none can.\n"
       "The CS sickLeave boolean has no effect on TTT behavior.\n"
       "Document actual contractor sick leave eligibility for future reference.",
       "Low", "Functional",
       "#3026", "Sick Leave / CS Office Settings",
       "CSSalaryOfficeVacationData.sickLeave is boolean in CS model. "
       "NOT used in TTT sick leave service. Verify contractor access rules."),
]


# ══════════════════════════════════════════════════════════════════
# CALENDAR MIGRATION TEST CASES (TC-VAC-157 to TC-VAC-160)
# ══════════════════════════════════════════════════════════════════

CALENDAR_MIGRATION_TESTS = [
    tc("TC-VAC-157",
       "Office calendar migration — working day norm change verification",
       "Office migrated from Russia to Cyprus calendar in 2024.\n"
       "Example: Венера (id=10), Нептун (id=11), Уран (id=12).\n"
       "Different holidays between Russia and Cyprus calendars.",
       "1. DB (ttt_calendar): SELECT * FROM office_calendar WHERE office_id = 10.\n"
       "2. Verify: since_year=2023 → Russia, since_year=2024 → Cyprus.\n"
       "3. DB: Compare calendar_days for Russia vs Cyprus for same month.\n"
       "4. GET /api/calendar/v1/calendar/office/{officeId}?year=2023 (Russia).\n"
       "5. GET /api/calendar/v1/calendar/office/{officeId}?year=2024 (Cyprus).\n"
       "6. Compare working days count — different holidays should produce different norms.",
       "2023 calendar uses Russia holidays (e.g., Jan 1-8 = New Year break, May 9).\n"
       "2024 calendar uses Cyprus holidays (different holidays, different norms).\n"
       "Monthly working day counts differ between calendars.\n"
       "Vacation duration calculation uses the correct calendar for each year.",
       "High", "Data Integrity",
       "#2876", "Calendar / Migration",
       "11 offices migrated in 2024. Russia→Cyprus is the most common migration path."),

    tc("TC-VAC-158",
       "Employee vacation spanning calendar boundary year",
       "Employee in office that switched calendar (e.g., Венера: Russia→Cyprus in 2024).\n"
       "Vacation spans December 2023 → January 2024.",
       "1. Find or create vacation spanning Dec 2023 to Jan 2024 for office 10.\n"
       "2. Check: how are working days calculated?\n"
       "3. December 2023 days should use Russia calendar.\n"
       "4. January 2024 days should use Cyprus calendar.\n"
       "5. Verify total vacation duration = sum of both periods.\n"
       "6. Check if calendar boundary is handled correctly.",
       "December portion: working days counted using Russia calendar holidays.\n"
       "January portion: working days counted using Cyprus calendar holidays.\n"
       "Total duration correctly combines both calendars.\n"
       "If not: vacation may count wrong number of working days.",
       "High", "Boundary",
       "#2876", "Vacation / Calendar Integration",
       "The since_year boundary in office_calendar determines which calendar applies. "
       "Cross-boundary vacations are a key risk area."),

    tc("TC-VAC-159",
       "Employee office transfer — same calendar offices (immediate update)",
       "Employee moves from Сатурн (id=2, Russia cal) to Юпитер (id=4, Russia cal).\n"
       "Both offices use the same production calendar.",
       "1. DB: Record employee_office for current year.\n"
       "2. Trigger CS sync with new office (Юпитер).\n"
       "3. DB: Check employee_office for current year — should update immediately.\n"
       "4. DB: Check employee.office_id — should match new office.\n"
       "5. Verify vacation day calculation uses new office annual leave norm.\n"
       "6. Check: no DEFERRED update since calendars are equal.",
       "employee_office updated immediately (same calendar = safe path).\n"
       "EmployeeOfficeChangedProcessor.isCalendarsAreEqual() returns true.\n"
       "Both employee.office_id and employee_office.office match new office.\n"
       "Vacation days recalculated with new office norm if different.",
       "Medium", "Functional",
       "#2876", "Vacation / Office Transfer",
       "Same-calendar transfer is the 'safe' path in EmployeeOfficeChangedProcessor."),

    tc("TC-VAC-160",
       "Employee office transfer — different calendar (deferred update)",
       "Employee moves from Сатурн (id=2, Russia) to Венера (id=10, Cyprus) mid-year.\n"
       "Different production calendars.",
       "1. Set clock to June (mid-year).\n"
       "2. DB: Record employee_office for current year.\n"
       "3. Trigger CS sync with new office (Венера, Cyprus calendar).\n"
       "4. DB: Check employee_office for CURRENT year — should NOT update.\n"
       "5. DB: Check employee_office for NEXT year — should update to new office.\n"
       "6. DB: Check employee.office_id — should update to new office.\n"
       "7. Verify: current year vacation uses OLD office norm.\n"
       "8. Verify: next year vacation will use NEW office norm.",
       "Current year employee_office NOT updated (deferred — different calendar mid-year).\n"
       "Next year employee_office IS updated to new office.\n"
       "employee.office_id IS updated to new office.\n"
       "Current year: old calendar holidays used for vacation duration.\n"
       "Next year: new calendar holidays used.\n"
       "This is an intentional architectural decision.",
       "High", "Functional",
       "#2876", "Vacation / Office Transfer",
       "EmployeeOfficeChangedProcessor defers mid-year cross-calendar changes. "
       "This is the documented architectural decision — test the deferred path."),
]


def add_supplement_tab_to_workbook(workbook_path, tab_name, suite_name, test_cases):
    """Add a supplementary test suite tab to an existing workbook."""
    wb = openpyxl.load_workbook(workbook_path)
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    count = write_ts_tab(ws, suite_name, test_cases)
    wb.save(workbook_path)
    return count


def main():
    import os

    base = os.path.dirname(os.path.abspath(__file__))
    vac_path = os.path.join(base, "vacation", "vacation.xlsx")

    # 1. Add Maternity lifecycle tests
    mat_count = add_supplement_tab_to_workbook(
        vac_path, "TS-Vac-Maternity",
        "Maternity Leave Lifecycle — Begin/End Events, Day Adjustment, Edge Cases",
        MATERNITY_TESTS
    )
    print(f"Maternity lifecycle: {mat_count} cases added")

    # 2. Add CS office settings tests
    cs_count = add_supplement_tab_to_workbook(
        vac_path, "TS-Vac-CSSettings",
        "CS Office Settings — Unimplemented Features, Hardcoded Behavior",
        CS_SETTINGS_TESTS
    )
    print(f"CS office settings: {cs_count} cases added")

    # 3. Add Calendar migration tests
    cal_count = add_supplement_tab_to_workbook(
        vac_path, "TS-Vac-CalendarMigr",
        "Calendar Migration — Office Calendar Switch, Cross-Calendar Transfers",
        CALENDAR_MIGRATION_TESTS
    )
    print(f"Calendar migration: {cal_count} cases added")

    total = mat_count + cs_count + cal_count
    print(f"\nTotal S67 supplementary cases: {total}")
    print(f"Vacation workbook now has 3 new tabs at: {vac_path}")


if __name__ == "__main__":
    main()
