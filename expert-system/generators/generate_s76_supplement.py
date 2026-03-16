#!/usr/bin/env python3
"""Session 76 — Supplementary test cases for Sprint 15 PM Tool cluster,
Confirmation notification bug (#3368), and Planner closed filter (#3386).

Adds:
- TS-ADM-PMTool-S15 tab to admin/admin.xlsx (12 cases: TC-ADM-081 to TC-ADM-092)
- TS-RPT-Confirmation tab to reports/reports.xlsx (5 cases: TC-RPT-111 to TC-RPT-115)
- TS-PLN-ClosedFilter tab to planner/planner.xlsx (3 cases: TC-PLN-106 to TC-PLN-108)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling constants (match existing workbooks) ──────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    hr = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, t in enumerate(test_cases):
        row = hr + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            t["id"], t["title"], t["preconditions"], t["steps"],
            t["expected"], t["priority"], t["type"],
            t["req_ref"], t["module"], t.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{ws.max_row}"
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# ADMIN SUPPLEMENT — PM Tool Sprint 15 Cluster
# ══════════════════════════════════════════════════════════════════

ADMIN_CASES = [
    tc("TC-ADM-081",
       "PM Tool sync — API pagination handles >100 projects",
       "PM Tool has >100 active projects; TTT sync enabled; timemachine env",
       "1. Trigger full PM Tool sync via POST /v1/test/project/sync\n"
       "2. Monitor sync status — verify paginated fetch with PAGE_SIZE=100\n"
       "3. Check pm_sync_status table for completion record\n"
       "4. Verify all projects synced (compare PM Tool count vs TTT project count)",
       "Sync completes successfully with multiple page fetches.\n"
       "pm_sync_status records show start/end timestamps.\n"
       "All PM Tool projects present in TTT project table.",
       "High", "Functional",
       "#3382", "PmToolEntitySyncLauncher",
       "Rate limiter applies per-page: each page fetch calls fetchRateLimiter.acquire()"),

    tc("TC-ADM-082",
       "PM Tool sync — id parameter workaround (no ?id= in requests)",
       "PM Tool sync enabled; timemachine env",
       "1. Trigger sync via POST /v1/test/project/sync\n"
       "2. Check application logs for PM Tool API requests\n"
       "3. Verify no request uses ?id=N query parameter\n"
       "4. Verify PmToolPageRequest is used for all fetches",
       "No PM Tool API call uses the id query parameter.\n"
       "All requests use the paginated PmToolPageRequest model.\n"
       "No 422 Unprocessable Content errors.",
       "Medium", "Regression",
       "#3383", "PmToolClient",
       "PM Tool API docs say id param works but it returns 422"),

    tc("TC-ADM-083",
       "PM Tool sync — pmtId persisted and exposed in API",
       "PM Tool project has pmtId; TTT sync runs; timemachine env",
       "1. Trigger sync via POST /v1/test/project/sync\n"
       "2. GET /v1/projects — find a synced project\n"
       "3. Verify response contains pmtId field (non-null Long)\n"
       "4. Query DB: SELECT pmt_id FROM project WHERE pm_tool_id IS NOT NULL\n"
       "5. Verify pmt_id column populated for synced projects",
       "API response includes pmtId field.\n"
       "DB pmt_id column populated for all PM Tool synced projects.\n"
       "pmtId matches PM Tool's internal project ID.",
       "High", "Functional",
       "#3387", "Project / PmToolProjectSynchronizer",
       "DB migration V2_1_26 adds pmt_id column"),

    tc("TC-ADM-084",
       "PM Tool sync — CSToolEntityReference type parsing",
       "PM Tool API returns employee refs as {id: N, type: 'employee'|'sales'|'contractor'}",
       "1. Trigger sync with project having mixed employee types\n"
       "2. Verify sales-type employees are filtered out from manager, owner, supervisor\n"
       "3. Verify sales-type watchers removed from watchersIds list\n"
       "4. Verify employee-type and contractor-type employees are processed normally\n"
       "5. Check project after sync: no sales employees in any role",
       "Sales-type employees excluded from all project roles.\n"
       "Employee-type and contractor-type processed normally.\n"
       "CSToolEntityReference.isSales() correctly identifies sales type (case-insensitive).",
       "Critical", "Functional",
       "#3389", "PmToolProjectSynchronizer / CSToolEntityReference",
       "New PM Tool API format: plain integer IDs replaced with {id, type} objects"),

    tc("TC-ADM-085",
       "Admin UI — Project name links to PM Tool via pmtId",
       "Admin > Projects page; projects synced from PM Tool; timemachine env",
       "1. Navigate to Admin > All Projects\n"
       "2. Find a project with pmtId (synced from PM Tool)\n"
       "3. Verify project name is a clickable hyperlink\n"
       "4. Verify link URL: https://pm.noveogroup.com/projects/{pmtId}/profile/general\n"
       "5. Find a project WITHOUT pmtId (legacy)\n"
       "6. Verify project name is plain text (not a link)",
       "Synced projects: name is blue hyperlink to PM Tool project page.\n"
       "Legacy projects without pmtId: name is plain text.\n"
       "Dev env uses pm-dev2 URL; prod uses pm.noveogroup.com.",
       "High", "UI",
       "#3093", "Admin / tableHelpers.js",
       "isDevelopmentEnv check distinguishes pm-dev2 vs pm.noveogroup.com URLs"),

    tc("TC-ADM-086",
       "Admin UI — Create Project button removed",
       "Admin > Projects page; any user with PROJECTS_ALL permission",
       "1. Navigate to Admin > All Projects\n"
       "2. Check toolbar for 'Create Project' button\n"
       "3. Navigate to Admin > My Projects\n"
       "4. Check toolbar for 'Create Project' button\n"
       "5. Verify no project creation UI exists anywhere in Admin",
       "No 'Create Project' button on All Projects tab.\n"
       "No 'Create Project' button on My Projects tab.\n"
       "Projects can only be created via PM Tool.",
       "High", "UI",
       "#3093", "Admin / TableProjectsButtons.js",
       "Project creation moved entirely to PM Tool"),

    tc("TC-ADM-087",
       "Admin UI — Edit Project shows only 3 tracker fields",
       "Admin > Projects; user with EDIT permission; timemachine env",
       "1. Navigate to Admin > All Projects\n"
       "2. Click Edit button (pencil icon) on a project\n"
       "3. Verify dialog title: 'Edit Tracker Data'\n"
       "4. Verify exactly 3 fields: sync script URL, tracker URL, proxy URL\n"
       "5. Verify NO fields for: name, customer, country, type, status, model, owner, manager, observers\n"
       "6. Edit tracker URL and save — verify change persisted",
       "Edit dialog shows only 3 link fields.\n"
       "All project metadata fields removed from edit.\n"
       "Save persists tracker field changes.\n"
       "Proxy field includes link to Google Docs setup guide.",
       "High", "UI",
       "#3093", "Admin / EditProjectForm.js",
       "Previous edit form had 10+ fields; now reduced to 3 tracker-related only"),

    tc("TC-ADM-088",
       "Admin UI — Renamed columns: Supervisor and Manager",
       "Admin > All Projects and My Projects tabs",
       "1. Navigate to Admin > All Projects\n"
       "2. Verify column header: 'Supervisor' (was 'Senior Manager')\n"
       "3. Navigate to Admin > My Projects\n"
       "4. Verify column header: 'Manager' (was 'Current Manager')\n"
       "5. Open Project Info modal for any project\n"
       "6. Verify label: 'Supervisor' (was 'Senior Manager')",
       "All Projects tab: 'Supervisor' column header.\n"
       "My Projects tab: 'Manager' column header.\n"
       "Project Info modal: 'Supervisor' label.\n"
       "Sorting and filtering work correctly on renamed columns.",
       "Medium", "UI",
       "#3093", "Admin / tableHelpers.js / InfoProjectModal.js",
       "Cosmetic rename but affects test case references and user documentation"),

    tc("TC-ADM-089",
       "Admin UI — Inline editing removed from Type/Status columns",
       "Admin > All Projects; user with EDIT permission",
       "1. Navigate to Admin > All Projects\n"
       "2. Click on Type column value for any project\n"
       "3. Verify NO inline edit dropdown appears\n"
       "4. Click on Status column value\n"
       "5. Verify NO inline edit dropdown appears\n"
       "6. Click on Transfer/Return Project action\n"
       "7. Verify action button does NOT exist",
       "Type column: click does not open inline editor.\n"
       "Status column: click does not open inline editor.\n"
       "No Transfer/Return Project action in action menu.\n"
       "These fields are read-only, managed by PM Tool.",
       "Medium", "UI",
       "#3093", "Admin / ProjectButton components removed",
       "Inline editing components (ProjectButton) removed from codebase"),

    tc("TC-ADM-090",
       "PM Tool sync — startup full sync on application boot",
       "TTT application restart; PM Tool sync enabled; timemachine env",
       "1. Restart TTT application (or deploy new build)\n"
       "2. Check logs for TttStartupApplicationListener activity\n"
       "3. Verify full PM Tool sync triggered on startup\n"
       "4. Check pm_sync_status for sync record at startup time\n"
       "5. Verify all PM Tool projects synced correctly",
       "Full PM Tool sync runs on application startup.\n"
       "TttStartupApplicationListener triggers sync.\n"
       "pm_sync_status shows startup sync record.\n"
       "Projects match PM Tool data after restart.",
       "Medium", "Functional",
       "#3083 / !5130", "TttStartupApplicationListener",
       "Added in !5130: ensures data is current after deployment"),

    tc("TC-ADM-091",
       "PM Tool sync — default sync script auto-assignment",
       "New project created in PM Tool; TTT sync runs; timemachine env",
       "1. Create a new project in PM Tool (no tracker config)\n"
       "2. Trigger TTT sync or wait for cron\n"
       "3. GET /v1/projects — find the new project\n"
       "4. Verify scriptUrl set to default: 'ttt.noveogroup.com/api/ttt/resource/defaultTaskInfoScript.js'\n"
       "5. Check project change history\n"
       "6. Verify default script assignment NOT logged as history event",
       "New projects get default sync script URL.\n"
       "Default script: ttt.noveogroup.com/api/ttt/resource/defaultTaskInfoScript.js.\n"
       "No change history entry for default script assignment.",
       "Medium", "Functional",
       "#3083 / !5270", "PmToolProjectSynchronizer",
       "Default script only for new projects; existing projects keep their scripts"),

    tc("TC-ADM-092",
       "PM Tool sync — accounting name set once, never updated",
       "Project synced from PM Tool; PM Tool project renamed; TTT sync runs",
       "1. Verify existing synced project has accountingName set\n"
       "2. Rename the project in PM Tool\n"
       "3. Trigger TTT sync\n"
       "4. GET /v1/projects — check the project\n"
       "5. Verify 'name' field updated to new PM Tool name\n"
       "6. Verify 'accountingName' still has ORIGINAL name (not updated)\n"
       "7. For a brand new project: verify accountingName = name on first sync",
       "accountingName set from PM Tool name on first sync only.\n"
       "Subsequent PM Tool name changes update 'name' but NOT 'accountingName'.\n"
       "New projects: accountingName = name on first sync.",
       "High", "Functional",
       "#3083", "PmToolProjectSynchronizer / Project",
       "accountingName used for billing/invoicing — must remain stable"),
]


# ══════════════════════════════════════════════════════════════════
# REPORTS SUPPLEMENT — Confirmation Notification Bug #3368
# ══════════════════════════════════════════════════════════════════

REPORTS_CASES = [
    tc("TC-RPT-111",
       "By Employee tab shows over/under report notification",
       "Employee has over-reported or under-reported hours for approve period;\n"
       "Confirmation page open; timemachine env",
       "1. Navigate to Confirmation > By Employee\n"
       "2. Select an employee with over/under reported hours\n"
       "3. Check for notification banner showing hours deviation\n"
       "4. Navigate to Confirmation > By Projects\n"
       "5. Select the same employee\n"
       "6. Compare notification content between both tabs",
       "Both tabs show identical over/under report notification.\n"
       "Notification displays: reported hours, norm hours, deviation.\n"
       "By Employee tab fetches statistics via employee approve period API.",
       "Critical", "Functional",
       "#3368", "Confirmation / employeeTabSagas.ts",
       "Root cause was By Employee tab never calling statistics endpoint"),

    tc("TC-RPT-112",
       "By Employee — norm display uses normForDate with fallback",
       "Employee has statistics data; some months have date-specific norm, others don't",
       "1. Navigate to Confirmation > By Employee\n"
       "2. Select employee with date-specific norm (part-month employment)\n"
       "3. Verify notification shows normForDate value\n"
       "4. Select employee with standard full-month norm\n"
       "5. Verify notification shows general norm when normForDate is null\n"
       "6. Verify 'Monthly norm' and 'Norm as of {date}' labels displayed correctly",
       "normForDate displayed when available.\n"
       "Fallback to general norm when normForDate is null.\n"
       "Labels 'Monthly norm' and 'Norm as of {date}' shown correctly.",
       "High", "Functional",
       "#3368 / !5132", "Confirmation / TEmployeeStatistics",
       "Fix: normForDate || norm fallback in saga"),

    tc("TC-RPT-113",
       "By Employee — approve period determines notification month",
       "Office approve period is for a past month (not current month);\n"
       "Employee has overwork in both past and current month",
       "1. Set office approve period to a past month (e.g., February when current is March)\n"
       "2. Ensure employee has over-reported hours in BOTH months\n"
       "3. Navigate to Confirmation > By Employee\n"
       "4. Select the employee\n"
       "5. Verify notification shows deviation for APPROVE MONTH (past), not current month\n"
       "6. Navigate to Confirmation > By Projects\n"
       "7. Verify SAME month shown in notification",
       "By Employee shows notification for approve period month, not current month.\n"
       "Both tabs show identical month context.\n"
       "When approve month < current month: notification uses approvePeriod.periodStart.",
       "Critical", "Functional",
       "#3368 / !5205", "Confirmation / employeeTabSagas.ts",
       "Bug: By Employee showed current month; By Projects showed approve month"),

    tc("TC-RPT-114",
       "By Employee — officeId resolution fallback",
       "Employee DTOs may have officeId at different paths: direct .officeId or .office.id",
       "1. Navigate to Confirmation > By Employee\n"
       "2. Select an employee (check that approve period loads correctly)\n"
       "3. Verify no JavaScript errors in browser console\n"
       "4. Test with employees from different offices\n"
       "5. Verify approve period fetched correctly for each office",
       "Approve period loaded correctly for all employees.\n"
       "No errors from undefined officeId.\n"
       "Fallback: uses employee.officeId || employee.office.id.",
       "Medium", "Functional",
       "#3368 / !5256", "Confirmation / employeeTabSagas.ts",
       "Different DTO shapes: some have officeId directly, others at office.id"),

    tc("TC-RPT-115",
       "By Employee — approve button switches to next user correctly",
       "Multiple employees with pending approvals; By Employee tab open",
       "1. Navigate to Confirmation > By Employee\n"
       "2. Select first employee in list with pending approval\n"
       "3. Click Approve button\n"
       "4. Verify: a) Approval succeeds b) View switches to next employee\n"
       "5. Verify notification updates for the new employee\n"
       "6. Verify employee list refreshes with updated status",
       "After approve: view auto-switches to next pending employee.\n"
       "Notification refreshes with new employee's statistics.\n"
       "Employee list updates (approved employee removed from pending).\n"
       "updateEmployeesWeeksAndMonths called with employee parameter.",
       "High", "Functional",
       "#3368 / !5132", "Confirmation / employeeTabSagas.ts",
       "Fix changed spawn to call for updateEmployeesWeeksAndMonths"),
]


# ══════════════════════════════════════════════════════════════════
# PLANNER SUPPLEMENT — Closed Filter Bug #3386
# ══════════════════════════════════════════════════════════════════

PLANNER_CASES = [
    tc("TC-PLN-106",
       "Planner search API — closed parameter filters assignments",
       "Planner has both open and closed assignments for a project; timemachine env",
       "1. GET /v1/assignments?projectId={id}&startDate=2026-03-01&endDate=2026-03-31\n"
       "   (no closed param) — count total assignments\n"
       "2. GET /v1/assignments?projectId={id}&startDate=2026-03-01&endDate=2026-03-31&closed=false\n"
       "   — count open-only assignments\n"
       "3. GET /v1/assignments?projectId={id}&startDate=2026-03-01&endDate=2026-03-31&closed=true\n"
       "   — count closed-only assignments\n"
       "4. Verify: open + closed = total",
       "closed=null → returns ALL assignments (open + closed).\n"
       "closed=false → returns only open assignments.\n"
       "closed=true → returns only closed assignments.\n"
       "Sum of open + closed equals total.",
       "High", "Functional",
       "#3386 / !5242", "TaskAssignmentGroupByEmployeeServiceImpl",
       "Filter applied in-memory after DB fetch, not at SQL level"),

    tc("TC-PLN-107",
       "Planner UI — Copy table excludes closed/deleted tasks",
       "Planner open for a project; at least one task has been closed/deleted",
       "1. Open Planner for a project with assigned tasks\n"
       "2. Enter edit mode\n"
       "3. Delete a task with 0 hours (close it)\n"
       "4. Verify deleted task disappears from the grid\n"
       "5. Click 'Copy the table' button\n"
       "6. Paste clipboard content into a text editor\n"
       "7. Verify: deleted/closed task is NOT in the pasted output",
       "Closed/deleted tasks excluded from planner grid.\n"
       "Copy table output does NOT contain closed tasks.\n"
       "Frontend passes closed=false to GET /v1/assignments.",
       "Critical", "Functional",
       "#3386 / !5282", "Planner / assignment sagas",
       "All 3 planner sagas (project, current, add-new) pass closed: false"),

    tc("TC-PLN-108",
       "Planner UI — auto-refresh after assignment changes",
       "Planner project view open; user makes assignment changes",
       "1. Open Planner for a project\n"
       "2. Add a new assignment to an employee\n"
       "3. Verify project view refreshes automatically\n"
       "4. Close an assignment (delete task)\n"
       "5. Verify project view refreshes automatically\n"
       "6. Verify no stale data shown after any change",
       "After adding assignment: project view refreshes with new data.\n"
       "After closing assignment: project view refreshes, closed task gone.\n"
       "fetchProjectAssignments() dispatched after handleManageAssignments.",
       "Medium", "Functional",
       "#3386 / !5282", "Planner / assignment sagas",
       "Added yield put(fetchProjectAssignments()) after manage"),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Write supplementary tabs to existing workbooks
# ══════════════════════════════════════════════════════════════════

BASE = os.path.dirname(os.path.abspath(__file__))


def add_supplement(xlsx_path, tab_name, suite_name, cases):
    wb = openpyxl.load_workbook(xlsx_path)
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    write_ts_tab(ws, suite_name, cases)
    wb.save(xlsx_path)
    print(f"  Added {tab_name} ({len(cases)} cases) to {os.path.basename(xlsx_path)}")


def main():
    print("Session 76 — Generating supplementary test cases...\n")

    # Admin
    admin_path = os.path.join(BASE, "admin", "admin.xlsx")
    add_supplement(admin_path, "TS-ADM-PMTool-S15",
                   "PM Tool Sprint 15 Integration Cluster", ADMIN_CASES)

    # Reports
    reports_path = os.path.join(BASE, "reports", "reports.xlsx")
    add_supplement(reports_path, "TS-RPT-Confirmation",
                   "Confirmation Notification Bug #3368", REPORTS_CASES)

    # Planner
    planner_path = os.path.join(BASE, "planner", "planner.xlsx")
    add_supplement(planner_path, "TS-PLN-ClosedFilter",
                   "Planner Closed Assignment Filter #3386", PLANNER_CASES)

    total = len(ADMIN_CASES) + len(REPORTS_CASES) + len(PLANNER_CASES)
    print(f"\nTotal: {total} new test cases across 3 workbooks")
    print("  Admin:   TC-ADM-081 to TC-ADM-092 (12 cases)")
    print("  Reports: TC-RPT-111 to TC-RPT-115 (5 cases)")
    print("  Planner: TC-PLN-106 to TC-PLN-108 (3 cases)")


if __name__ == "__main__":
    main()
