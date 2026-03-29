#!/usr/bin/env python3
"""
Reports & Confirmation Test Documentation Generator — Phase B (Session 93)
Generates test-docs/reports/reports.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 7 TS- test suite tabs.

Based on vault knowledge: ttt-report-service-deep-dive (9.8KB), reports-business-rules-reference
(4000+ words), reports-ticket-findings (22 tickets deep-read from 166+ searched),
confirmation-flow-live-testing, frontend-report-module, period-api-testing, statistics-service.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "reports")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "reports.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Test Suites ─────────────────────────────────────────────────────────────

SUITES = [
    ("TS-Reports-CRUD", "My Tasks CRUD", get_crud_cases := lambda: None),
    ("TS-Reports-Confirmation", "Confirmation Flow", None),
    ("TS-Reports-Periods", "Period Management", None),
    ("TS-Reports-AutoReject", "Auto-Rejection", None),
    ("TS-Reports-Statistics", "Statistics & Norms", None),
    ("TS-Reports-Notifications", "Notifications", None),
    ("TS-Reports-Permissions", "Permissions & API", None),
]


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_crud_cases():
    """TS-Reports-CRUD: My Tasks page — report creation, editing, deletion, batch operations."""
    return [
        {
            "id": "TC-RPT-001", "title": "Create a time report — happy path",
            "preconditions": "Employee with at least one active project and assigned task.\nQuery: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_project ep ON e.id = ep.employee_id JOIN ttt_backend.project p ON ep.project_id = p.id JOIN ttt_backend.task t ON t.project_id = p.id WHERE e.enabled = true AND p.status = 'ACTIVE' AND e.login != 'pvaynmaster' ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks page (/report)\n3. Verify the weekly grid is displayed with task rows and day columns\n4. Click on an empty cell in the current week for an assigned task\n5. Verify an input field appears in the cell\n6. Type '4' (hours)\n7. Press Enter or click outside the cell\n8. Verify the cell displays '4' and the weekly/daily totals update\n9. Verify the cell background is white (REPORTED state)\nDB-CHECK: SELECT effort FROM ttt_backend.task_report WHERE executor_login = '<login>' AND task_id = <task_id> AND report_date = '<date>' — verify effort = 240 (minutes)",
            "expected": "Report created with 4 hours (240 minutes). Cell shows '4'. Weekly row total and daily column total both increase by 4. Cell state is REPORTED (white background).",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Create Rules", "module": "reports/crud",
            "notes": "Effort stored in minutes in DB (×60). LockService prevents concurrent edits — cell shows lock icon if already being edited by another user."
        },
        {
            "id": "TC-RPT-002", "title": "Edit existing report — change hours",
            "preconditions": "Employee with an existing report in REPORTED state.\nSETUP: Via API — POST /api/ttt/v1/reports with effort=120 (2 hours).\nQuery: (same as TC-RPT-001)",
            "steps": "SETUP: Via API — create a report for 2 hours on a weekday in the current open period\n1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Navigate to the week containing the setup report\n4. Find the cell showing '2'\n5. Click the cell to activate edit mode\n6. Clear and type '6'\n7. Press Enter\n8. Verify the cell now shows '6'\n9. Verify weekly/daily totals updated (+4)\n10. Verify cell background is white (edit resets to REPORTED state)\nCLEANUP: Via API — DELETE /api/ttt/v1/reports with the report ID",
            "expected": "Report updated to 6 hours. State reset to REPORTED (even if was APPROVED before). Totals reflect change.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Update Rules", "module": "reports/crud",
            "notes": "Any effort change resets state to REPORTED and clears approver. This is a key regression test (state should always reset)."
        },
        {
            "id": "TC-RPT-003", "title": "Delete report by setting hours to 0",
            "preconditions": "Employee with an existing report.\nSETUP: Via API — create a report for 3 hours.",
            "steps": "SETUP: Via API — POST /api/ttt/v1/reports with effort=180\n1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Find the cell showing '3'\n4. Click the cell, clear content, type '0'\n5. Press Enter\n6. Verify the cell is now empty (report deleted)\n7. Verify weekly/daily totals decreased by 3\nDB-CHECK: SELECT COUNT(*) FROM ttt_backend.task_report WHERE executor_login = '<login>' AND report_date = '<date>' AND task_id = <task_id> — verify 0 (deleted)",
            "expected": "Setting effort to 0 deletes the report. Cell empty, totals decreased. No DB record exists.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Effort=0", "module": "reports/crud",
            "notes": "Effort=0 triggers deletion, not update to zero. This is an important behavioral detail."
        },
        {
            "id": "TC-RPT-004", "title": "Report in closed period — blocked",
            "preconditions": "Employee in a salary office where report period is ahead of the target date.\nQuery: SELECT e.login, op.start AS period_start FROM ttt_backend.employee e JOIN ttt_backend.office_period op ON e.salary_office_id = op.office_id WHERE op.type = 'REPORT' AND e.enabled = true AND e.login != 'pvaynmaster' ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Navigate to a week before the office report period start date\n4. Verify cells for dates before the period start are NOT editable (grayed out or no input)\n5. Attempt to click a closed-period cell\n6. Verify no input field appears or input is rejected",
            "expected": "Cells before report period start date are read-only. No editing possible for closed periods.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Report Period Rules", "module": "reports/crud",
            "notes": "Period check: reportDate < office report period start → blocked. Extended period can override for individual employees."
        },
        {
            "id": "TC-RPT-005", "title": "Add new task on My Tasks page",
            "preconditions": "Employee assigned to a project with available (unselected) tasks.\nQuery: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_project ep ON e.id = ep.employee_id JOIN ttt_backend.project p ON ep.project_id = p.id JOIN ttt_backend.task t ON t.project_id = p.id WHERE e.enabled = true AND p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Click 'Add task' button (or '+' icon)\n4. In the task selection dropdown/dialog, search for an available task\n5. Select the task\n6. Verify the task appears as a new row in the weekly grid\n7. Verify the row has empty cells ready for reporting",
            "expected": "New task row added to the grid. Task name displayed. Empty cells available for time entry.",
            "priority": "High", "type": "UI",
            "req_ref": "modules/frontend-report-module.md", "module": "reports/crud",
            "notes": "Task visibility depends on employee-project assignment. Pinned tasks always appear."
        },
        {
            "id": "TC-RPT-006", "title": "Pin/unpin task",
            "preconditions": "Employee with at least one visible task.\nQuery: (same as TC-RPT-005)",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Locate a task row in the grid\n4. Click the pin icon (📌) on the task row\n5. Verify pin icon becomes active/highlighted\n6. Navigate to a different week\n7. Verify the pinned task still appears even if no reports exist in that week\n8. Navigate back, click the pin icon again to unpin\n9. Verify task disappears from weeks where it has no reports\nDB-CHECK: SELECT * FROM ttt_backend.pinned_employee_task WHERE employee_login = '<login>' AND task_id = <task_id>",
            "expected": "Pinned task visible on all weeks. Unpinned task only visible on weeks with existing reports.",
            "priority": "Medium", "type": "UI",
            "req_ref": "reports-business-rules-reference.md Pin endpoints", "module": "reports/crud",
            "notes": "POST /v1/reports/pin/{taskId}/employees/{login} to pin, DELETE to unpin."
        },
        {
            "id": "TC-RPT-007", "title": "Rename task on My Tasks page",
            "preconditions": "Employee with a task that they can rename.\nSETUP: Via API — ensure employee has at least one task.",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Right-click (or click context menu) on a task name\n4. Select 'Rename' option\n5. In the rename popup, enter a new name\n6. Select 'for all time' option\n7. Click 'Save'\n8. Verify task name updates in the grid",
            "expected": "Task name updated. All employees see the new name.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3278 rename bugs", "module": "reports/crud",
            "notes": "Bug #3278: rename in Confirmation popup triggers 400 'Employee login not found'. Verify rename works on My Tasks (should work)."
        },
        {
            "id": "TC-RPT-008", "title": "Week navigation — arrow buttons",
            "preconditions": "Logged-in employee with tasks.",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Note the current displayed week (Mon-Sun dates in header)\n4. Click the right arrow (→) to advance one week\n5. Verify the header updates to next week's dates\n6. Verify task rows update with that week's data\n7. Click the left arrow (←) twice\n8. Verify navigating to the previous week's data\n9. Verify totals recalculate per week",
            "expected": "Week navigation arrows advance/retreat by one week. Header dates, cell data, and totals all update correctly.",
            "priority": "High", "type": "UI",
            "req_ref": "modules/frontend-report-module.md", "module": "reports/crud",
            "notes": "6 week tabs visible. Navigation beyond visible range shifts the tab set."
        },
        {
            "id": "TC-RPT-009", "title": "Batch create reports — multiple cells in one week",
            "preconditions": "Employee with multiple tasks.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_project ep ON e.id = ep.employee_id JOIN ttt_backend.project p ON ep.project_id = p.id JOIN ttt_backend.task t ON t.project_id = p.id WHERE e.enabled = true AND p.status = 'ACTIVE' GROUP BY e.login HAVING COUNT(DISTINCT t.id) >= 2 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Click empty cell on Task 1, Monday — type '4', press Tab\n4. Verify cursor moves to next cell (Tuesday)\n5. Type '4', press Tab\n6. Navigate to Task 2, Monday — type '2'\n7. Press Enter\n8. Verify all three cells saved correctly\n9. Verify daily totals for Monday show 6 (4+2)\n10. Verify weekly total for Task 1 shows 8\nCLEANUP: Via API — DELETE all created reports",
            "expected": "Multiple reports created efficiently. Tab key navigates between cells. All totals calculate correctly.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Batch Operations", "module": "reports/crud",
            "notes": "PUT /v1/reports batch endpoint used for multi-cell saves. Tab navigation is a key UX feature."
        },
        {
            "id": "TC-RPT-010", "title": "Report with decimal hours (e.g., 1.5)",
            "preconditions": "Employee with a task and open period.\nQuery: (same as TC-RPT-001)",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Click empty cell, type '1.5'\n4. Press Enter\n5. Verify cell displays '1.5'\n6. Verify totals include 1.5 in calculation\nDB-CHECK: SELECT effort FROM ttt_backend.task_report — verify effort = 90 (minutes)\nCLEANUP: Via API — DELETE",
            "expected": "Decimal hours accepted. 1.5 hours stored as 90 minutes. Display shows '1.5'. Totals correct.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1", "module": "reports/crud",
            "notes": "Decimal separator depends on user's office settings (dot or comma). Check config: decimalSeparator."
        },
        {
            "id": "TC-RPT-011", "title": "TAB key stacking bug (regression #3398)",
            "preconditions": "Employee on My Tasks page.\nQuery: (same as TC-RPT-001)",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Click on an empty cell\n4. Verify single input field appears\n5. Press TAB key without entering a value\n6. Verify cursor moves to next cell cleanly\n7. Repeat TAB 3-4 more times rapidly\n8. Verify NO duplicate/stacked input fields appear\n9. Verify the grid remains visually clean",
            "expected": "Tab navigation between cells does not create stacked/duplicate input fields. Each cell has at most one input.",
            "priority": "High", "type": "UI",
            "req_ref": "#3398 stacked inputs", "module": "reports/crud",
            "notes": "Bug #3398: production report of duplicate input fields on TAB. If still present, document as known issue."
        },
        {
            "id": "TC-RPT-012", "title": "Report comment — add and view",
            "preconditions": "Employee with an existing report.\nSETUP: Via API — create a report.",
            "steps": "SETUP: Via API — create report for 4 hours\n1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Find the cell with the report\n4. Hover over or click the cell to see options\n5. Click comment icon (or right-click context menu)\n6. Enter a comment: 'Code review for feature X'\n7. Save the comment\n8. Hover over the cell\n9. Verify comment tooltip shows 'Code review for feature X'\nCLEANUP: Via API — DELETE the report",
            "expected": "Comment saved and visible on hover. Comment persists across page reloads.",
            "priority": "Medium", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 PATCH", "module": "reports/crud",
            "notes": "Comments are stored via PATCH on the report entity. stateComment field on the task_report."
        },
        {
            "id": "TC-RPT-013", "title": "Cell locking — concurrent edit prevention",
            "preconditions": "Two accounts: employee and a manager who can report on their behalf.\nQuery: SELECT e.login AS employee, m.login AS manager FROM ttt_backend.employee e JOIN ttt_backend.employee m ON m.enabled = true WHERE e.enabled = true AND e.login != m.login AND EXISTS (SELECT 1 FROM ttt_backend.employee_project ep WHERE ep.employee_id = e.id) ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee in Browser 1\n2. Navigate to My Tasks, click a cell to start editing\n3. Login as the manager in Browser 2\n4. Navigate to the employee's report page (/report/<employee_login>)\n5. Attempt to click the same cell\n6. Verify a lock icon or 'cell locked' message appears\n7. Verify the manager cannot edit the locked cell\n8. In Browser 1, save or cancel the edit\n9. In Browser 2, verify the cell becomes editable",
            "expected": "Cell shows lock icon when another user is editing. HTTP 423 Locked returned for concurrent edit attempts. Lock released after save/cancel.",
            "priority": "Medium", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Cell locking", "module": "reports/crud",
            "notes": "LockService uses in-memory map. Batch PUT skips locking. Lock timeout not documented."
        },
        {
            "id": "TC-RPT-014", "title": "View another employee's report page (manager)",
            "preconditions": "Manager with REPORTS_VIEW permission and a subordinate.\nQuery: SELECT m.login AS manager, e.login AS employee FROM ttt_backend.employee m JOIN ttt_backend.employee e ON e.enabled = true WHERE m.enabled = true AND EXISTS (SELECT 1 FROM ttt_backend.employee_role er JOIN ttt_backend.role r ON er.role_id = r.id WHERE er.employee_id = m.id AND r.name IN ('ROLE_PROJECT_MANAGER', 'ROLE_ADMIN')) AND m.login != e.login ORDER BY random() LIMIT 1",
            "steps": "1. Login as the manager\n2. Navigate to /report/<employee_login>\n3. Verify the employee's report page loads\n4. Verify the header shows the employee's name\n5. Verify the weekly grid shows the employee's tasks and hours\n6. Verify the manager can edit cells (if REPORTS_EDIT permission exists)",
            "expected": "Manager can view and potentially edit employee's report page. Same grid layout with employee's data.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §8", "module": "reports/crud",
            "notes": "Contractor report page has known bug #3150: infinite spinner when opened from Statistics/Admin links."
        },
        {
            "id": "TC-RPT-015", "title": "Contractor report page — spinner bug (regression #3150)",
            "preconditions": "A contractor employee and a manager/admin.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_CONTRACTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as an admin/manager\n2. Navigate to Statistics > Employee reports\n3. Find the contractor in the table\n4. Click the contractor's name to open their report page\n5. Verify the page loads without infinite spinner\n6. If spinner occurs: note the error in console (global error expected)\n7. Try direct navigation to /report/<contractor_login>\n8. Note if the behavior differs between Statistics link and direct URL",
            "expected": "Contractor's report page loads successfully from all entry points. No infinite spinner or global error.",
            "priority": "High", "type": "UI",
            "req_ref": "#3150 contractor spinner", "module": "reports/crud",
            "notes": "Bug #3150 (OPEN since Nov 2024): Global error and spinner on contractor report from Statistics/Admin. Does NOT occur on contractor's own My Tasks page."
        },
    ]


def get_confirmation_cases():
    """TS-Reports-Confirmation: Confirmation page — By Employee and By Projects tabs."""
    return [
        {
            "id": "TC-RPT-016", "title": "Approve hours — By Projects tab, single task",
            "preconditions": "Manager with REPORTS_APPROVE permission. Employee with REPORTED hours in current approval period.\nSETUP: Via API — create a report for the employee.\nQuery: SELECT m.login AS manager, e.login AS employee, t.id AS task_id FROM ttt_backend.employee m JOIN ttt_backend.employee_project mp ON m.id = mp.employee_id JOIN ttt_backend.project p ON mp.project_id = p.id JOIN ttt_backend.task t ON t.project_id = p.id JOIN ttt_backend.employee_project ep ON ep.project_id = p.id JOIN ttt_backend.employee e ON ep.employee_id = e.id WHERE m.enabled = true AND e.enabled = true AND m.login != e.login AND EXISTS (SELECT 1 FROM ttt_backend.employee_role er JOIN ttt_backend.role r ON er.role_id = r.id WHERE er.employee_id = m.id AND r.name IN ('ROLE_PROJECT_MANAGER', 'ROLE_SENIOR_MANAGER', 'ROLE_ADMIN')) ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — POST /api/ttt/v1/reports with executor=employee, task=task_id, effort=480 (8h), date=today\n1. Login as the manager\n2. Navigate to Confirmation page (/approve)\n3. Click 'By Projects' tab\n4. Select the project from the Project dropdown\n5. Navigate to the week containing the report\n6. Find the employee's row and the task with 8 hours\n7. Click the approve (✓) button on the task row for the specific week\n8. Verify success notification or visual state change (green background)\n9. Verify the cell shows green/approved background color\nDB-CHECK: SELECT state FROM ttt_backend.task_report WHERE id = <report_id> — verify 'APPROVED'\nCLEANUP: Via API — PATCH report state back to REPORTED, then DELETE",
            "expected": "Hours approved. Cell shows green (APPROVED) background. DB state is APPROVED. Approver login recorded.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §2 Approve Rules", "module": "reports/confirmation",
            "notes": "Approve sets state=APPROVED and records approverLogin. PROJECT_MANAGER type projects cannot be approved."
        },
        {
            "id": "TC-RPT-017", "title": "Approve hours — By Employee tab",
            "preconditions": "Same as TC-RPT-016.\nSETUP: Via API — create a report.",
            "steps": "SETUP: Via API — create a report for the employee\n1. Login as the manager\n2. Navigate to Confirmation page (/approve)\n3. Click 'By Employee' tab (default)\n4. Select the employee from the Employee dropdown\n5. Navigate to the week containing the report\n6. Find the task row with hours\n7. Click the approve button\n8. Verify approval succeeds\n9. Verify Approve button behavior after approval (should not advance to next user — #3368 bug)\nCLEANUP: Via API — DELETE the report",
            "expected": "Hours approved via By Employee tab. Cell shows green background.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3368 By Employee bugs", "module": "reports/confirmation",
            "notes": "Bug #3368: Approve button on 'By Employee' doesn't switch to next user or change work week (unlike By Projects)."
        },
        {
            "id": "TC-RPT-018", "title": "Reject hours with comment",
            "preconditions": "Manager with a report to approve.\nSETUP: Via API — create a report.",
            "steps": "SETUP: Via API — create a report for the employee\n1. Login as the manager\n2. Navigate to Confirmation page (/approve)\n3. Select 'By Projects' tab, choose the project\n4. Find the employee's report cell\n5. Click the reject (✗) button\n6. Verify rejection tooltip/popup appears with a textarea\n7. Type a comment: 'Please split by subtasks'\n8. Click confirm/reject button in the tooltip\n9. Verify cell shows red/rejected background\n10. Verify the comment is stored\nDB-CHECK: SELECT r.description FROM ttt_backend.reject r JOIN ttt_backend.task_report tr ON tr.reject_id = r.id WHERE tr.id = <report_id>\nCLEANUP: Via API — DELETE the report",
            "expected": "Hours rejected with comment. Cell shows red (REJECTED) background. Reject record created in DB with the comment.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §2 Reject Rules", "module": "reports/confirmation",
            "notes": "Comment optional in API but UI shows textarea. Reject creates a separate 'reject' table record."
        },
        {
            "id": "TC-RPT-019", "title": "Re-report after rejection — clears rejected state",
            "preconditions": "Employee with a REJECTED report.\nSETUP: Via API — create report, then reject it via PATCH state=REJECTED.",
            "steps": "SETUP: Via API — create report (4h), then PATCH state to REJECTED with stateComment='Fix'\n1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Find the cell with rejected report (red background)\n4. Verify rejection indicator visible (red background or icon)\n5. Click the cell, change hours from 4 to 6\n6. Press Enter\n7. Verify cell reverts to white background (REPORTED state)\n8. Verify rejection indicator disappears\nDB-CHECK: SELECT state, reject_id FROM ttt_backend.task_report WHERE id = <report_id> — verify state='REPORTED', reject_id=NULL\nDB-CHECK: SELECT COUNT(*) FROM ttt_backend.reject WHERE id = <old_reject_id> — verify 0 (deleted)\nCLEANUP: Via API — DELETE the report",
            "expected": "Re-reporting deletes the reject record and resets state to REPORTED. No rejection history preserved.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §1 Re-reporting", "module": "reports/confirmation",
            "notes": "Critical: reject record is DELETED (not archived). No history of the previous rejection remains."
        },
        {
            "id": "TC-RPT-020", "title": "Bulk approve — 'Approve all' header button",
            "preconditions": "Manager with multiple employees having REPORTED hours.\nSETUP: Via API — create reports for 2+ employees.",
            "steps": "SETUP: Via API — create 2h reports for Employee A and Employee B on same day/task\n1. Login as the manager\n2. Navigate to Confirmation page (/approve)\n3. Select 'By Projects', choose the project\n4. Navigate to the week with the setup reports\n5. Click the 'Approve all' button in the week header\n6. Verify confirmation dialog appears ('Approve all active tasks for this week?')\n7. Confirm\n8. Verify all REPORTED cells turn green (APPROVED)\n9. Verify no APPROVED cells were double-processed\nCLEANUP: Via API — DELETE all created reports",
            "expected": "All REPORTED hours in the week approved in one action. Already APPROVED hours unaffected.",
            "priority": "High", "type": "UI",
            "req_ref": "#3354 approval granularity", "module": "reports/confirmation",
            "notes": "Granularity levels per #3354: per-cell, per-line (task), all active tasks, 'Approve all' header. Each should handle manual and auto-rejected hours identically."
        },
        {
            "id": "TC-RPT-021", "title": "Confirmation — 'By Employee' vs 'By Projects' data parity",
            "preconditions": "Manager with employees having various report states.\nSETUP: Via API — create reports with mixed states (REPORTED, APPROVED, REJECTED).",
            "steps": "SETUP: Via API — create 3 reports for same employee: one REPORTED, one APPROVED, one REJECTED\n1. Login as the manager\n2. Navigate to Confirmation (/approve)\n3. On 'By Projects' tab: select the project, note all cells/states for the employee\n4. Switch to 'By Employee' tab: select the same employee\n5. Compare: same tasks, same hours, same states visible?\n6. Verify both tabs show identical data for the same employee/week\n7. Check over/under-reporting notification appears on both tabs (#3368)\n8. Note any discrepancies",
            "expected": "Both tabs show identical report data for the same employee. Over/under notifications consistent.",
            "priority": "High", "type": "UI",
            "req_ref": "#3368 notification parity", "module": "reports/confirmation",
            "notes": "Known issue #3368: 'By Employee' doesn't call statistic/report/employees endpoint → missing over/under notification. Multiple sub-bugs documented."
        },
        {
            "id": "TC-RPT-022", "title": "Confirmation — orange dot on week tabs with pending items",
            "preconditions": "Manager with reports to approve in specific weeks.\nSETUP: Via API — create reports in different weeks.",
            "steps": "SETUP: Via API — create REPORTED hours in Week 1 and Week 3 (skip Week 2)\n1. Login as the manager\n2. Navigate to Confirmation (/approve)\n3. Select 'By Projects' tab, choose the project\n4. Observe the 6 week tabs at top\n5. Verify Week 1 and Week 3 tabs show an orange dot indicator\n6. Verify Week 2 tab has NO orange dot\n7. Click Week 1 tab → verify pending reports visible\n8. Click Week 2 tab → verify no pending reports\nCLEANUP: Via API — DELETE the reports",
            "expected": "Orange dots appear only on week tabs with pending (REPORTED) items. Empty weeks have no indicator.",
            "priority": "Medium", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §2 UI Confirmation Page", "module": "reports/confirmation",
            "notes": "Orange dot is the visual indicator for weeks needing attention."
        },
        {
            "id": "TC-RPT-023", "title": "Rename task in Confirmation — bug regression (#3278)",
            "preconditions": "Manager on Confirmation page with a task to rename.\nSETUP: Via API — create a report.",
            "steps": "SETUP: Via API — create a report for the employee\n1. Login as the manager\n2. Navigate to Confirmation (/approve) > By Projects\n3. Find a task name in the confirmation grid\n4. Click the task name or right-click for context menu\n5. Select 'Rename'\n6. In the rename popup, enter a new name\n7. For scope, select 'For all employees' (all checked)\n8. Click 'Save'\n9. Verify rename succeeds WITHOUT HTTP 400 error\n10. Verify task name updated in the grid\n11. Repeat with 'For this employee only' (uncheck 'all')\n12. On 'By Projects' tab: verify rename works (was broken per #3278)\nCLEANUP: Via API — DELETE the report",
            "expected": "Rename succeeds without errors on both tabs and for both scopes.",
            "priority": "High", "type": "UI",
            "req_ref": "#3278 rename 400 error", "module": "reports/confirmation",
            "notes": "Bug #3278: rename in Confirmation triggered 400 'Employee login not found'. Also: rename for single employee on By Projects closes popup without effect."
        },
        {
            "id": "TC-RPT-024", "title": "Confirmation — employee dropdown update after approve all (#3267)",
            "preconditions": "Manager with employee having REPORTED hours.\nSETUP: Via API — create reports for the employee.",
            "steps": "SETUP: Via API — create multiple REPORTED hours for the employee across the week\n1. Login as the manager\n2. Navigate to Confirmation (/approve) > By Projects\n3. Verify employee appears in the employee dropdown\n4. Approve ALL hours for the employee (via Approve All or individual cells)\n5. Via API (in separate tab): create a new REPORTED report for the same employee in the same period\n6. Return to Confirmation page\n7. Refresh or re-select the project\n8. Verify the employee still appears in the dropdown\n9. Verify the new REPORTED hours are visible\n10. Verify the 'Confirm' button is active\nCLEANUP: Via API — DELETE all created reports",
            "expected": "Employee remains in dropdown after new reports submitted. Red dot returns. Confirm button active.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3267 dropdown disappear", "module": "reports/confirmation",
            "notes": "Bug #3267: after confirming ALL hours, new reports caused employee to disappear from dropdown. Was attributed to backend glitch."
        },
        {
            "id": "TC-RPT-025", "title": "Over/under-reporting notification on Confirmation page",
            "preconditions": "Employee who has reported significantly more or less than their norm.\nSETUP: Via API — create reports totaling >110% of monthly norm.\nQuery: SELECT e.login, sr.month_norm FROM ttt_backend.statistic_report sr JOIN ttt_backend.employee e ON sr.employee_login = e.login WHERE sr.report_date >= DATE_TRUNC('month', CURRENT_DATE) AND sr.month_norm > 0 ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create enough REPORTED hours to exceed 110% of the employee's monthly norm\n1. Login as a manager with REPORTS_APPROVE permission\n2. Navigate to Confirmation (/approve) > By Projects\n3. Select the project containing the over-reporting employee\n4. Verify a non-dismissible banner appears at the top\n5. Verify the employee's name is highlighted in red (over-reporting)\n6. Click or hover on the clock icon next to the name\n7. Verify tooltip shows: deviation %, month, DM, projects with PM names\n8. Switch to 'By Employee' tab\n9. Verify the same over-reporting notification appears (#3368 known issue)\nCLEANUP: Via API — DELETE the excess reports",
            "expected": "Over-reporting banner visible on both tabs. Employee name highlighted red. Tooltip shows detailed deviation info.",
            "priority": "High", "type": "UI",
            "req_ref": "#3368, reports-business-rules-reference.md §6", "module": "reports/confirmation",
            "notes": "Bug #3368: 'By Employee' missing over/under notification. Also: period-dependent bug — approve month < current month AND report month = current month → missing on By Employee."
        },
        {
            "id": "TC-RPT-026", "title": "Reject notification — rejected hours banner on My Tasks (#3321, #3268)",
            "preconditions": "Employee with REJECTED reports.\nSETUP: Via API — create report, then reject it.",
            "steps": "SETUP: Via API — create report (4h), PATCH state=REJECTED with stateComment='Incorrect task'\n1. Login as the employee\n2. Navigate to My Tasks (/report)\n3. Verify a rejection notification banner appears at the top of the page\n4. Verify banner text indicates which task has rejected hours\n5. Verify the rejection comment is visible (hover or click)\n6. Click the link/button in the banner to navigate to the rejected week\n7. Verify the rejected cell is highlighted (red background)\n8. Verify no HTTP 400 'Period should be less than two months' error in console\nCLEANUP: Via API — DELETE the report",
            "expected": "Rejection banner visible. Comment accessible. Navigation link works. No API errors.",
            "priority": "High", "type": "UI",
            "req_ref": "#3321, #3268 rejected notification", "module": "reports/confirmation",
            "notes": "Bug #3268: backend 2-month sampling restriction caused 400 on notification API. Bug #3321: no notification when report month closed but confirmation open."
        },
        {
            "id": "TC-RPT-027", "title": "Approve hours in closed report period (confirmation still open)",
            "preconditions": "Manager approving hours where report period is closed but approve period is still open.\nQuery: SELECT op_r.start AS report_start, op_a.start AS approve_start, op_r.office_id FROM ttt_backend.office_period op_r JOIN ttt_backend.office_period op_a ON op_r.office_id = op_a.office_id WHERE op_r.type = 'REPORT' AND op_a.type = 'APPROVE' AND op_a.start < op_r.start",
            "steps": "1. Login as the manager\n2. Navigate to Confirmation (/approve)\n3. Select the salary office/project where report period is ahead of approve period\n4. Navigate to a week in the gap (between approve period start and report period start)\n5. Find REPORTED hours in this period\n6. Approve them\n7. Verify approval succeeds (report period closed doesn't block approval)\nCLEANUP: Via API — revert state if needed",
            "expected": "Approval succeeds for hours in the approve period even when report period is closed. The two periods are independent.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Dual Period Model", "module": "reports/confirmation",
            "notes": "Key distinction: approve period controls confirmation, report period controls data entry. They are NOT the same."
        },
    ]


def get_period_cases():
    """TS-Reports-Periods: Period management — report period, approve period, extended period."""
    return [
        {
            "id": "TC-RPT-028", "title": "Advance report period by 1 month — happy path",
            "preconditions": "Accountant or admin with access to Accounting > Changing periods page.\nQuery: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_ACCOUNTANT', 'ROLE_CHIEF_ACCOUNTANT', 'ROLE_ADMIN') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods page\n3. Select a salary office from the dropdown\n4. Note the current Report period month\n5. Click 'Edit' button for the Report period\n6. In the period selection popup, select the next month (+1)\n7. Click 'Save'\n8. Verify the report period advances by 1 month\n9. Verify success notification appears\n10. Verify the new period is displayed",
            "expected": "Report period advanced by 1 month. New period saved and displayed. Old period dates become read-only on My Tasks.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Report Period Rules", "module": "reports/periods",
            "notes": "Report period must be first of month. Must be >= approve period."
        },
        {
            "id": "TC-RPT-029", "title": "Advance approve period by 1 month",
            "preconditions": "Accountant or admin. Approve period is at least 1 month behind report period.\nQuery: (same as TC-RPT-028)",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Note current Approve period and Report period\n5. Click 'Edit' for the Approve period\n6. Select next month (+1) — must still be <= report period\n7. Click 'Save'\n8. Verify approve period advances\n9. Verify PeriodChangedEvent side effects:\n   - Auto-reject of unapproved hours (if feature enabled)\n   - Vacation recalculation triggered",
            "expected": "Approve period advanced. Side effects triggered (auto-reject, vacation recalc). Cannot exceed report period.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3350 1-month max delta", "module": "reports/periods",
            "notes": "Bug cluster #3350: max 1-month delta enforced (both directions). Validation uses SAVED values, not dynamic."
        },
        {
            "id": "TC-RPT-030", "title": "Period — 2+ month jump blocked (#3350)",
            "preconditions": "Accountant on Changing periods page.\nQuery: (same as TC-RPT-028)",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Click 'Edit' for the Report period\n5. Attempt to select a month that is 2+ months ahead of current saved value\n6. Verify the month is DISABLED in the calendar picker (not selectable)\n7. If selectable: click 'Save' and verify 400 error from backend\n8. Repeat for Approve period: attempt 2+ month jump\n9. Verify both periods enforce 1-month max delta",
            "expected": "Months more than 1 ahead/behind the SAVED value are disabled in the UI picker. If somehow submitted, backend returns 400.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3350 max 1-month delta", "module": "reports/periods",
            "notes": "Bug #3350 QA bugs: UI used DYNAMIC values instead of SAVED values. Partial saves possible. Sequential saves required for multi-month changes."
        },
        {
            "id": "TC-RPT-031", "title": "Period — saved vs dynamic value validation (#3350 bug 4)",
            "preconditions": "Accountant on Changing periods page.\nQuery: (same as TC-RPT-028)",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Click 'Edit' for Report period\n5. Select 1 month forward — do NOT save yet\n6. Click 'Edit' for Approve period\n7. Verify: available months for Approve period are calculated from the SAVED report period, not the unsaved selection\n8. If dynamic values are used (bug): the approve period allows months it shouldn't\n9. Save Report period first, then verify Approve period options update",
            "expected": "Period validation uses SAVED values. Unsaved changes to one period do not affect the other period's validation.",
            "priority": "High", "type": "UI",
            "req_ref": "#3350 bug 4 — dynamic vs saved", "module": "reports/periods",
            "notes": "Key QA finding from #3350: popup used DYNAMIC (unsaved) values instead of SAVED values for available months."
        },
        {
            "id": "TC-RPT-032", "title": "Report period preceding approve period — disabled (#3365)",
            "preconditions": "Accountant on Changing periods page.\nQuery: (same as TC-RPT-028)",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Click 'Edit' for the Report period\n5. Verify months BEFORE the saved approve period are DISABLED\n6. Attempt to select a month before approve period\n7. Verify it cannot be selected\n8. If selectable: save and verify 400 error\n9. Close without saving",
            "expected": "Report period months before approve period are disabled. Constraint: report period >= approve period.",
            "priority": "High", "type": "UI",
            "req_ref": "#3365 report month not disabled", "module": "reports/periods",
            "notes": "Bug #3365 (OPEN): Month preceding saved confirmation period not disabled for report period selection. Partially fixed."
        },
        {
            "id": "TC-RPT-033", "title": "Extended report period — grant to individual employee",
            "preconditions": "Accountant/admin. Employee whose office report period is already closed for a specific month.",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select the salary office\n4. Find the employee who needs an extension\n5. Click 'Extend' or 'Grant extended period' for the employee\n6. Select the month to extend\n7. Save the extension\n8. Login as the employee\n9. Navigate to My Tasks (/report)\n10. Navigate to the extended period month\n11. Verify cells in the extended period are EDITABLE (not locked)\n12. Enter a report value\n13. Verify report saved successfully",
            "expected": "Extended period granted. Employee can report in the otherwise-closed period. Extension visible in admin view.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Extended Report Period", "module": "reports/periods",
            "notes": "PUT /v1/periods/report/employees/{login}. Auto-cleaned by ExtendedPeriodScheduler every 5 minutes. Blocks approve period advancement."
        },
        {
            "id": "TC-RPT-034", "title": "Extended period blocks approve period advancement",
            "preconditions": "Accountant/admin. At least one employee with active extended period in the salary office.",
            "steps": "SETUP: Via API — PUT /v1/periods/report/employees/{login} to grant extended period\n1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select the salary office with the extended period\n4. Attempt to advance the Approve period\n5. Verify the advance is BLOCKED with an error message\n6. Verify the error indicates which employee has the extended period\n7. Remove the extended period (via admin or wait for auto-cleanup)\n8. Re-attempt approve period advancement\n9. Verify it now succeeds",
            "expected": "Approve period cannot advance while extended periods exist. Error message identifies the blocking employee.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §4 Extended periods block", "module": "reports/periods",
            "notes": "ExtendedPeriodScheduler auto-cleans every 5 minutes. Manual removal also possible."
        },
        {
            "id": "TC-RPT-035", "title": "Period — partial save prevention (#3350 bug 4)",
            "preconditions": "Accountant on Changing periods page.\nQuery: (same as TC-RPT-028)",
            "steps": "1. Login as the accountant/admin\n2. Navigate to Accounting > Changing periods\n3. Select a salary office\n4. Click 'Edit' for Report period — select +1 month\n5. Click 'Edit' for Approve period — select a value that would become invalid if report period saves first\n6. Click 'Save' (whichever save button is available)\n7. Verify BOTH periods are validated together\n8. If one saves and the other fails: this is a partial save bug\n9. Verify that a failed save for one period does NOT save the other",
            "expected": "Both periods validated as a pair. No partial saves — either both succeed or both fail.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3350 bug 4 — partial saves", "module": "reports/periods",
            "notes": "Known issue: one period may save while the other's validation fails. Result: inconsistent state."
        },
    ]


def get_autoreject_cases():
    """TS-Reports-AutoReject: Auto-rejection of unapproved hours on period closure."""
    return [
        {
            "id": "TC-RPT-036", "title": "Auto-reject — unapproved hours rejected on approve period close",
            "preconditions": "Accountant. Employee with REPORTED (unapproved) hours in the month being closed.\nSETUP: Via API — create REPORTED hours, verify approve period is 1 month behind.\nNote: Auto-reject feature was DISABLED (reverted) — this test verifies current behavior.",
            "steps": "SETUP: Via API — create REPORTED hours for the employee in the approve period month\n1. Login as the accountant\n2. Navigate to Accounting > Changing periods\n3. Select the salary office\n4. Note REPORTED hours exist for the closing month (verify via Confirmation page)\n5. Advance the Approve period by 1 month (closing the current month)\n6. After save, navigate to Confirmation page\n7. Check if REPORTED hours in the closed month are now REJECTED (auto-reject triggered)\n8. Or: check if they remain REPORTED (auto-reject disabled)\n9. Login as the employee\n10. Navigate to My Tasks — check for auto-rejection notification banner\nDB-CHECK: SELECT state FROM ttt_backend.task_report WHERE executor_login = '<login>' AND report_date IN (closed period)",
            "expected": "If auto-reject is active: REPORTED hours become REJECTED with 'auto.reject.state' comment. Notification banner on My Tasks. If disabled: hours remain REPORTED.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#2698 auto-reject, reports-business-rules-reference.md §3", "module": "reports/autoreject",
            "notes": "Feature reverted from release/2.1 (#3285). Code exists but may be disabled. Test to verify current state. If active, triggers on approve period change (not report period)."
        },
        {
            "id": "TC-RPT-037", "title": "Auto-reject ONLY triggers on approve period change (not report period)",
            "preconditions": "Reports exist in REPORTED state. Both periods can be changed.\nSETUP: Via API — create REPORTED hours.",
            "steps": "SETUP: Via API — create REPORTED hours in the month between approve and report period\n1. Login as accountant\n2. Advance the REPORT period by 1 month\n3. Check: REPORTED hours in the gap → should remain REPORTED (not auto-rejected)\n4. Verify no auto-rejection notification sent to employee\n5. Now advance the APPROVE period by 1 month\n6. Check: if auto-reject is active, REPORTED hours in the newly closed approve month → REJECTED\n7. Confirm the trigger is approve period, NOT report period",
            "expected": "Report period closure does NOT trigger auto-reject. Only approve period closure triggers it.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3285 wrong trigger", "module": "reports/autoreject",
            "notes": "Bug #3285: auto-reject was incorrectly triggered on REPORT period closure. The bug was reverted. This test verifies it's fixed."
        },
        {
            "id": "TC-RPT-038", "title": "Auto-reject — notification banner on My Tasks",
            "preconditions": "Employee with auto-rejected hours (if feature is active).\nSETUP: Must trigger auto-reject first (requires approve period advancement).",
            "steps": "1. Trigger auto-reject by advancing approve period (see TC-RPT-036)\n2. Login as the affected employee\n3. Navigate to My Tasks (/report)\n4. Verify notification banner at top: 'Unconfirmed hours for task {name} were automatically rejected upon month closure'\n5. Click 'Go to the report page' link in the banner\n6. Verify navigation to the first week with rejected hours\n7. Click the X button to dismiss the banner\n8. Navigate away and return\n9. Verify banner does NOT reappear (dismissed via localStorage)\n10. Clear browser data and reload\n11. Verify banner reappears (#2698 known edge case)",
            "expected": "Auto-reject banner visible with task name. 'Go to report' link navigates correctly. Dismiss persists via localStorage.",
            "priority": "High", "type": "UI",
            "req_ref": "#2698 banner behavior", "module": "reports/autoreject",
            "notes": "AutoRejectedReportsContainer on ReportPage. localStorage key: 'hiddenAutoRejectWarnings'. Banner reappears after site data clear."
        },
        {
            "id": "TC-RPT-039", "title": "Auto-reject — reopen period and re-confirm/edit (#3367)",
            "preconditions": "Approve period was advanced (auto-reject triggered), then accountant reopens period.\nSETUP: Requires full cycle: create reports → close period → auto-reject → reopen.",
            "steps": "SETUP: Create REPORTED hours → advance approve period (triggers auto-reject) → revert approve period (reopen)\n1. Login as the manager\n2. Navigate to Confirmation page\n3. Find auto-rejected hours in the reopened period\n4. Attempt to approve the auto-rejected hours\n5. Verify approval succeeds WITHOUT HTTP 500 error\n6. Login as the employee\n7. Navigate to My Tasks for the reopened period\n8. Attempt to edit an auto-rejected cell\n9. Verify editing works WITHOUT HTTP 500 error\n10. Verify re-reporting clears the auto-reject state",
            "expected": "After period reopen: confirmation and editing of auto-rejected hours works. No 500 errors.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3367 500 on reopen", "module": "reports/autoreject",
            "notes": "Bug #3367 (CLOSED): after period close→auto-reject→reopen, confirmation failed with 500. Fix in MR !5095."
        },
        {
            "id": "TC-RPT-040", "title": "Auto-reject vs manual reject — identical behavior (#3354)",
            "preconditions": "Reports in both auto-rejected and manually rejected states.\nSETUP: Create reports, auto-reject some via period close, manually reject others.",
            "steps": "SETUP: Create 4 reports. Close approve period (auto-rejects 2). Manually reject 2 others via Confirmation.\n1. Reopen the approve period\n2. Login as manager\n3. On Confirmation page, attempt to approve an auto-rejected report\n4. Verify behavior identical to approving a manually rejected report\n5. Login as employee\n6. On My Tasks, attempt to re-report on an auto-rejected cell\n7. Verify behavior identical to re-reporting on a manually rejected cell\n8. Compare: both types clear the reject record, reset to REPORTED state",
            "expected": "Auto-rejected and manually rejected reports behave identically for all operations: approve, re-report, edit.",
            "priority": "High", "type": "UI",
            "req_ref": "#3354 share settings", "module": "reports/autoreject",
            "notes": "Requirement from #3354: auto-rejected hours MUST behave identically to manually rejected ones after period reopening."
        },
    ]


def get_statistics_cases():
    """TS-Reports-Statistics: Statistics page, norm calculation, over/under reporting."""
    return [
        {
            "id": "TC-RPT-041", "title": "Statistics — Employee reports table basic view",
            "preconditions": "User with access to Statistics (ADMIN, CHIEF_ACCOUNTANT, OFFICE_DIRECTOR, ACCOUNTANT, DM, or TECH_LEAD).\nQuery: SELECT DISTINCT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_ADMIN', 'ROLE_CHIEF_ACCOUNTANT', 'ROLE_OFFICE_DIRECTOR', 'ROLE_ACCOUNTANT') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as a user with statistics access\n2. Navigate to Statistics > Employee reports\n3. Verify the table loads with columns: Employee, Reported hours, Monthly norm, Personal norm, Budget norm, Deviation %, Manager, Comment\n4. Verify data is populated for the current month\n5. Verify pagination works if >20 employees\n6. Verify sorting by clicking column headers",
            "expected": "Statistics table displays employee report data for current month. All columns present and populated.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §6", "module": "reports/statistics",
            "notes": "Access matrix: ADMIN/CHIEF_ACCOUNTANT see all, OFFICE_DIRECTOR/ACCOUNTANT see office, DM/TECH_LEAD see subordinates, EMPLOYEE cannot access."
        },
        {
            "id": "TC-RPT-042", "title": "Budget norm display — with/without admin vacation (#3381)",
            "preconditions": "Employee with administrative vacation in the current month AND employee without.\nQuery: SELECT sr.employee_login, sr.month_norm, sr.budget_norm FROM ttt_backend.statistic_report sr WHERE sr.report_date >= DATE_TRUNC('month', CURRENT_DATE) AND sr.budget_norm IS NOT NULL AND sr.budget_norm != sr.month_norm LIMIT 5",
            "steps": "1. Login as admin/accountant\n2. Navigate to Statistics > Employee reports\n3. Find an employee with admin vacation: norm column shows '{individual} ({budget})'\n4. Find an employee without admin vacation: norm column shows just '{budget}'\n5. Verify the format matches #3381 spec\n6. Hover over the 'Norm' column header\n7. Verify info icon tooltip explains budget norm concept",
            "expected": "Budget norm displayed differently based on admin vacation presence. Format: '{individual} ({budget})' when different, just '{budget}' when equal.",
            "priority": "High", "type": "UI",
            "req_ref": "#3381 budget norm", "module": "reports/statistics",
            "notes": "Budget norm = individual norm + administrative vacation hours. Deviation % calculated from budget norm."
        },
        {
            "id": "TC-RPT-043", "title": "Individual norm — partial-month employee (#3353)",
            "preconditions": "Employee whose first working day is NOT the 1st of a month (or has a last working day mid-month).\nQuery: SELECT e.login, e.first_working_day FROM ttt_backend.employee e WHERE e.first_working_day IS NOT NULL AND EXTRACT(DAY FROM e.first_working_day) > 1 AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Find the partial-month employee\n4. Note their 'Personal norm' value\n5. Compare with a full-month employee in the same office\n6. Verify partial-month employee's personal norm is LESS than full-month (prorated)\n7. Via API — GET /api/ttt/v1/employees/{login}/work-periods\n8. Verify work period dates match employee's start/end dates\n9. On My Tasks page: verify non-working days before employee start date are marked orange",
            "expected": "Partial-month employee has prorated personal norm. Work periods endpoint returns correct date ranges.",
            "priority": "High", "type": "UI",
            "req_ref": "#3353 individual norm", "module": "reports/statistics",
            "notes": "#3353: new /work-periods endpoint. Non-working days orange on My Tasks, Confirmation, Planner. Re-hired employees have gap periods."
        },
        {
            "id": "TC-RPT-044", "title": "Vacation impact on personal norm (#3380)",
            "preconditions": "Employee with a vacation (paid or unpaid) in the current month.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.vacation v ON e.id = v.employee_id WHERE v.start_date >= DATE_TRUNC('month', CURRENT_DATE) AND v.start_date < DATE_TRUNC('month', CURRENT_DATE) + INTERVAL '1 month' AND v.status IN ('APPROVED', 'NEW') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Find the employee with a vacation this month\n4. Note their 'Personal norm' value\n5. Calculate expected norm: office monthly norm − vacation hours\n6. Verify personal norm reflects the vacation deduction\n7. Compare with My Tasks page norm display for the same employee\n8. If discrepancy exists: note bug #3380",
            "expected": "Personal norm reduced by vacation hours. Consistent between Statistics and My Tasks pages.",
            "priority": "High", "type": "UI",
            "req_ref": "#3380 vacation norm impact", "module": "reports/statistics",
            "notes": "Bug #3380 (OPEN): vacation doesn't reduce personal monthly norm on My Tasks or Employee Reports. Discrepancy between qa-2 and stage."
        },
        {
            "id": "TC-RPT-045", "title": "Over-the-limit toggle (#3306)",
            "preconditions": "Statistics page with employees having various deviation percentages.\nParameters: notification.reporting.over and notification.reporting.under in TTT Parameters.",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Note all employees currently displayed\n4. Toggle 'Only over the limit' switcher ON\n5. Verify the table filters to show only employees exceeding the over/under thresholds\n6. Verify the threshold values match TTT Parameters configuration\n7. Toggle OFF\n8. Verify all employees return to the table\n9. Verify the switcher has immediate effect (no page reload needed)",
            "expected": "Toggle filters employees by reporting deviation thresholds. ON: only over-limit shown. OFF: all shown.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3306 toggle bug", "module": "reports/statistics",
            "notes": "Bug #3306 (CLOSED): toggle had no effect. Should now work. Default thresholds: over=10%, under=30%."
        },
        {
            "id": "TC-RPT-046", "title": "DM field and Comment field on statistics (#3309)",
            "preconditions": "Statistics page with employee data.\nQuery: (same as TC-RPT-041)",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Verify 'Manager' column exists with DM names\n4. Click a manager name — verify it links to CS profile\n5. Verify 'Comment' column exists\n6. Click the comment field for an employee\n7. Type 'Monthly review note'\n8. Click outside to save (or press Enter)\n9. Verify comment is saved\n10. Change month in the date filter\n11. Verify the comment field shows the comment for THAT specific month (per-employee per-month)\n12. Return to original month — verify original comment still there",
            "expected": "Manager column with CS links. Comment field inline-editable, stored per employee per month.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3309 DM and Comment fields", "module": "reports/statistics",
            "notes": "QA bugs from #3309: font size should be 13px, spelling error in column title, comments aggregated incorrectly."
        },
        {
            "id": "TC-RPT-047", "title": "Future/dismissed employee exclusion (#3320)",
            "preconditions": "Statistics page. Ideally an employee with a future start date or recent dismissal exists.",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Verify NO employees with future start dates appear in the table\n4. Verify employees with future start dates don't show negative deviation %\n5. Check the employee search dropdown\n6. Verify dismissed employees are NOT in the search dropdown\n7. If dismissed employees appear in the table: verify they are NOT searchable in dropdown (#3320)\nDB-CHECK: SELECT login, first_working_day FROM ttt_backend.employee WHERE first_working_day > CURRENT_DATE AND enabled = true",
            "expected": "Future employees excluded from statistics. Dismissed employees not searchable. No false under-reports.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3320 future employees", "module": "reports/statistics",
            "notes": "Bug #3320: future employees shown with false negative deviation. Fix: switched to vacation statistic endpoint."
        },
        {
            "id": "TC-RPT-048", "title": "Deviation percentage calculation and display",
            "preconditions": "Employee with known reported hours and budget norm.\nQuery: SELECT sr.employee_login, sr.reported_effort, sr.budget_norm, CASE WHEN sr.budget_norm > 0 THEN ROUND((sr.reported_effort * 60 - sr.budget_norm)::numeric / sr.budget_norm * 100, 1) ELSE NULL END AS deviation FROM ttt_backend.statistic_report sr WHERE sr.report_date >= DATE_TRUNC('month', CURRENT_DATE) AND sr.budget_norm > 0 ORDER BY ABS(sr.reported_effort * 60 - sr.budget_norm) DESC LIMIT 5",
            "steps": "1. Login as admin\n2. Navigate to Statistics > Employee reports\n3. Find the employee from the query\n4. Verify the Deviation % column shows the calculated value\n5. Check formatting: integer except (-1,+1) range → 1 decimal place\n6. Check color coding: positive (red/over), negative (purple/under), zero (neutral)\n7. Check N/A display: employees with 0 budget norm → '+N/A%'\n8. Verify N/A sorts to top of the table",
            "expected": "Deviation % formula: (reported - budgetNorm) / budgetNorm × 100%. Display formatting matches spec.",
            "priority": "High", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §6 Deviation Formula", "module": "reports/statistics",
            "notes": "ExcessStatus enum: HIGH (>0%), LOW (<0%), NEUTRAL (==0%), NA (budgetNorm=0). N/A sorts to top."
        },
    ]


def get_notification_cases():
    """TS-Reports-Notifications: Scheduled notifications for reports."""
    return [
        {
            "id": "TC-RPT-049", "title": "Forgotten reports notification trigger",
            "preconditions": "Employee with less than 90% of monthly norm reported.\nSETUP: Via test API — trigger forgotten notifications.\nQuery: SELECT sr.employee_login FROM ttt_backend.statistic_report sr WHERE sr.report_date >= DATE_TRUNC('month', CURRENT_DATE) AND sr.budget_norm > 0 AND sr.reported_effort * 60 < sr.budget_norm * 0.9 LIMIT 5",
            "steps": "SETUP: Via test endpoint — POST /api/ttt/test/v1/notifications/reports-forgotten to trigger\n1. Verify the endpoint returns 200\n2. Check email inbox of under-reporting employees\n3. Verify notification email received with correct details\n4. Verify email contains: employee name, reported hours, norm, deficit\nDB-CHECK: SELECT * FROM ttt_backend.notification WHERE type = 'REPORTS_FORGOTTEN' ORDER BY created_time DESC LIMIT 5",
            "expected": "Forgotten report notifications sent to employees below 90% norm threshold.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "reports-business-rules-reference.md §7 sendReportsForgottenNotifications", "module": "reports/notifications",
            "notes": "Normal schedule: Mon/Fri 16:00. Test endpoint triggers immediately. Threshold: 90% of personal norm."
        },
        {
            "id": "TC-RPT-050", "title": "Reports changed notification trigger",
            "preconditions": "Manager who has reported on behalf of an employee.\nSETUP: Via API — create report as API token owner for another employee.",
            "steps": "SETUP: Via API — POST /api/ttt/v1/reports with executor=<employee_login> (as API token owner)\nSETUP: Via test endpoint — POST /api/ttt/test/v1/notifications/reports-changed to trigger\n1. Verify the endpoint returns 200\n2. Verify employee receives notification that someone reported on their behalf\n3. Check email content: reporter name, task, hours, date",
            "expected": "Employee notified when manager/other user reports on their behalf.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "reports-business-rules-reference.md §7 sendReportsChangedNotifications", "module": "reports/notifications",
            "notes": "Normal schedule: Daily 07:50. Cross-employee reporting is a known permission gap (BUG-REPORT-4)."
        },
        {
            "id": "TC-RPT-051", "title": "Reject notification — 5-minute scheduler",
            "preconditions": "Employee with a recently rejected report.\nSETUP: Via API — create report and reject it.",
            "steps": "SETUP: Via API — create report, PATCH state=REJECTED with comment\nSETUP: Via test endpoint — POST /api/ttt/test/v1/notifications/reject to trigger\n1. Verify the endpoint returns 200\n2. Check employee email for rejection notification\n3. Verify email contains: task name, rejection comment, link to report\n4. Verify reject record marked as notified\nDB-CHECK: SELECT executor_notified FROM ttt_backend.reject WHERE id = <reject_id> — verify true",
            "expected": "Reject notification email sent to employee. Reject record marked executor_notified=true.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "reports-business-rules-reference.md §7 sendRejectNotifications", "module": "reports/notifications",
            "notes": "Runs every 5 minutes normally. executor_notified=false until sent. Bug #3321: no notification when report month closed."
        },
        {
            "id": "TC-RPT-052", "title": "Accounting report notifications — budget exceeded",
            "preconditions": "Accountant. Employee with over-reported hours exceeding budget norm.\nSETUP: Reports exceeding budget norm must exist.",
            "steps": "1. Login as accountant\n2. Navigate to Statistics > Employee reports or accounting notifications area\n3. Verify over-reporting employees are flagged\n4. Via API — POST /api/ttt/v1/reports/accounting/notifications to send accounting-level notifications\n5. Verify notification sent to relevant accountants/managers\n6. Verify notification content: employee name, excess hours, projects",
            "expected": "Accounting notifications triggered for budget exceedance. Relevant stakeholders notified.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "reports-business-rules-reference.md §8 POST /accounting/notifications", "module": "reports/notifications",
            "notes": "Only AUTH_USER required for this endpoint (no specific permission check)."
        },
    ]


def get_permission_cases():
    """TS-Reports-Permissions: Role-based access and API security."""
    return [
        {
            "id": "TC-RPT-053", "title": "Regular employee — My Tasks access only",
            "preconditions": "Regular employee with ROLE_EMPLOYEE only.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_EMPLOYEE' AND e.enabled = true AND NOT EXISTS (SELECT 1 FROM ttt_backend.employee_role er2 JOIN ttt_backend.role r2 ON er2.role_id = r2.id WHERE er2.employee_id = e.id AND r2.name != 'ROLE_EMPLOYEE') ORDER BY random() LIMIT 1",
            "steps": "1. Login as the regular employee\n2. Navigate to My Tasks (/report) — verify access\n3. Attempt to navigate to Confirmation (/approve)\n4. Verify access denied or redirect\n5. Attempt to navigate to Statistics (/statistics)\n6. Verify access denied or redirect\n7. Attempt to navigate to Accounting routes\n8. Verify access denied or redirect",
            "expected": "Regular employee can access My Tasks only. No access to Confirmation, Statistics, or Accounting.",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §8 Permission Matrix", "module": "reports/permissions",
            "notes": "Permission: REPORTS_VIEW allows GET /reports. REPORTS_EDIT allows POST/PATCH/DELETE. REPORTS_APPROVE for confirmation."
        },
        {
            "id": "TC-RPT-054", "title": "Manager — Confirmation page access",
            "preconditions": "Employee with PROJECT_MANAGER or SENIOR_MANAGER role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_PROJECT_MANAGER', 'ROLE_SENIOR_MANAGER') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the manager\n2. Navigate to Confirmation (/approve)\n3. Verify page loads with By Employee and By Projects tabs\n4. Verify the manager can see employees from their projects\n5. Verify the manager can approve/reject hours\n6. Navigate to Statistics — verify access\n7. Verify Statistics scope: only their subordinates/projects visible",
            "expected": "Manager has full Confirmation access. Statistics shows only their scope (subordinates/projects).",
            "priority": "Critical", "type": "UI",
            "req_ref": "reports-business-rules-reference.md §6 Access Matrix", "module": "reports/permissions",
            "notes": "REPORTS_APPROVE permission grants confirmation page access."
        },
        {
            "id": "TC-RPT-055", "title": "Self-approval via API — no check (BUG-REPORT-2)",
            "preconditions": "Employee who is also a manager on a project.\nQuery: (same as TC-RPT-054)",
            "steps": "SETUP: Via API — create a report as the manager (on their own task)\n1. Via API — PATCH /api/ttt/v1/reports/{id} with state=APPROVED (using same user's token)\n2. Verify the request succeeds (HTTP 200)\n3. Verify the report is now APPROVED with approverLogin = executorLogin\n4. Note: there is NO executor≠approver validation — self-approval is possible\nDB-CHECK: SELECT executor_login, approver_login FROM ttt_backend.task_report WHERE id = <id> — verify both are the same",
            "expected": "Self-approval succeeds via API. No validation preventing approver = executor. This is a known bug.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "BUG-REPORT-2 self-approval", "module": "reports/permissions",
            "notes": "BUG-REPORT-2 (HIGH): No executor≠approver check in API. UI may prevent this by not showing own reports in Confirmation."
        },
        {
            "id": "TC-RPT-056", "title": "Direct create-as-APPROVED bypass (BUG-REPORT-3)",
            "preconditions": "User with REPORTS_EDIT and REPORTS_APPROVE permissions.\nQuery: (same as TC-RPT-054)",
            "steps": "1. Via API — POST /api/ttt/v1/reports with state=APPROVED in the request body\n2. Verify the request succeeds\n3. Verify the report is created directly in APPROVED state\n4. Note: this bypasses the REPORTED → APPROVED workflow entirely\nDB-CHECK: SELECT state FROM ttt_backend.task_report WHERE id = <id> — verify 'APPROVED'\nCLEANUP: Via API — DELETE the report",
            "expected": "Report created directly as APPROVED, bypassing normal workflow. No validation prevents this.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "BUG-REPORT-3 direct approve", "module": "reports/permissions",
            "notes": "BUG-REPORT-3 (HIGH): POST accepts state=APPROVED. Should only allow REPORTED on create."
        },
        {
            "id": "TC-RPT-057", "title": "No upper bound on effort (BUG-REPORT-1)",
            "preconditions": "Employee with a task.\nQuery: (same as TC-RPT-001)",
            "steps": "1. Via API — POST /api/ttt/v1/reports with effort=90000 (1500 hours = 25 hours/day × 60 days)\n2. Verify the request succeeds (no max validation)\n3. Check: does the UI display this correctly on My Tasks?\n4. Via API — try effort=999999\n5. Verify response (should accept due to no upper bound)\nCLEANUP: Via API — DELETE the report",
            "expected": "Unreasonably high effort values accepted by API. No upper bound validation exists.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "BUG-REPORT-1 no max effort", "module": "reports/permissions",
            "notes": "BUG-REPORT-1 (HIGH): Only minimum (≥1 minute) validated. 25h (1500 min) accepted. No daily/weekly cap."
        },
        {
            "id": "TC-RPT-058", "title": "Unauthenticated access to /effort and /employee-projects (BUG-REPORT-6)",
            "preconditions": "No authentication required for this test.",
            "steps": "1. Via curl (no auth header) — GET /api/ttt/v1/reports/effort\n2. Verify the response — if 200 with data, confirms missing @PreAuthorize\n3. Via curl (no auth header) — GET /api/ttt/v1/reports/employee-projects\n4. Verify the response — if 200, confirms missing @PreAuthorize\n5. Compare with authenticated access — verify same data returned",
            "expected": "If bug exists: both endpoints return data without authentication. Should require at minimum AUTH_USER.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "BUG-REPORT-6 missing auth", "module": "reports/permissions",
            "notes": "BUG-REPORT-6 (LOW): Missing @PreAuthorize annotation on /effort and /employee-projects endpoints."
        },
        {
            "id": "TC-RPT-059", "title": "Cross-employee reporting via API (BUG-REPORT-4)",
            "preconditions": "API token owner.\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.enabled = true AND e.login != 'pvaynmaster' ORDER BY random() LIMIT 1",
            "steps": "1. Via API — POST /api/ttt/v1/reports with executorLogin = <another_employee_login>\n2. Verify the report is created under the OTHER employee's name\n3. Check: does this require any specific permission beyond AUTH_USER?\n4. Verify the 'Reports changed' notification is generated for the target employee\nCLEANUP: Via API — DELETE the report",
            "expected": "API token owner can report for any employee. No executor validation beyond auth check.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "BUG-REPORT-4 cross-employee", "module": "reports/permissions",
            "notes": "BUG-REPORT-4: any authenticated user can create reports for any other employee via API."
        },
        {
            "id": "TC-RPT-060", "title": "Approved status removed after tracker import (#3296)",
            "preconditions": "PM with an employee having APPROVED hours. External tracker configured.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_project ep ON e.id = ep.employee_id JOIN ttt_backend.project p ON ep.project_id = p.id WHERE p.status = 'ACTIVE' AND e.enabled = true AND EXISTS (SELECT 1 FROM ttt_backend.employee_role er JOIN ttt_backend.role r ON er.role_id = r.id WHERE er.employee_id = e.id AND r.name = 'ROLE_PROJECT_MANAGER') ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create report, approve it\n1. Login as the PM\n2. Navigate to Planner > Projects or Tasks\n3. Find the approved report cell (green background)\n4. Import from external tracker (if tracker configured)\n5. After import: verify the cell value changed\n6. Verify the green APPROVED background is REMOVED\n7. Verify state reset to REPORTED (white background)\n8. If background persists: document as #3296 regression\nCLEANUP: Via API — DELETE the report",
            "expected": "Any hour change after approval removes APPROVED status. Green background removed. State = REPORTED.",
            "priority": "High", "type": "UI",
            "req_ref": "#3296 approved status persist", "module": "reports/permissions",
            "notes": "Bug #3296 (OPEN): after tracker import, value changes but green APPROVED background persists on Planner."
        },
    ]


# ─── Feature Matrix ──────────────────────────────────────────────────────────

FEATURES = [
    {"feature": "Report CRUD (My Tasks)", "suites": ["TS-Reports-CRUD"], "ui": 15, "hybrid": 0, "api": 0},
    {"feature": "Confirmation Flow (Approve/Reject)", "suites": ["TS-Reports-Confirmation"], "ui": 12, "hybrid": 0, "api": 0},
    {"feature": "Period Management", "suites": ["TS-Reports-Periods"], "ui": 8, "hybrid": 0, "api": 0},
    {"feature": "Auto-Rejection Lifecycle", "suites": ["TS-Reports-AutoReject"], "ui": 5, "hybrid": 0, "api": 0},
    {"feature": "Statistics & Norms", "suites": ["TS-Reports-Statistics"], "ui": 8, "hybrid": 0, "api": 0},
    {"feature": "Notifications", "suites": ["TS-Reports-Notifications"], "ui": 0, "hybrid": 4, "api": 0},
    {"feature": "Permissions & API Security", "suites": ["TS-Reports-Permissions"], "ui": 2, "hybrid": 6, "api": 0},
]


# ─── Risk Assessment ─────────────────────────────────────────────────────────

RISKS = [
    {"feature": "Auto-rejection lifecycle", "risk": "Feature was reverted due to severe implementation bugs (#2698, #3285, #3367). Re-enabling may reintroduce data corruption.", "likelihood": "High", "impact": "Critical", "severity": "Critical", "mitigation": "Full cycle testing: close→reject→notify→reopen→confirm. Verify trigger is approve period only."},
    {"feature": "Confirmation tab parity", "risk": "'By Employee' and 'By Projects' tabs show different data and have different functionality (#3368, #3278).", "likelihood": "High", "impact": "High", "severity": "High", "mitigation": "Side-by-side comparison testing. Verify notifications, dropdowns, and approval behavior on both tabs."},
    {"feature": "Period validation (saved vs dynamic)", "risk": "UI uses unsaved values for validation, allowing invalid states (#3350 bugs 2-5).", "likelihood": "High", "impact": "High", "severity": "High", "mitigation": "Test all combination of period changes. Verify saved values drive validation. Check partial save prevention."},
    {"feature": "Norm calculation chain", "risk": "Three-tier norm (general→individual→budget) with multiple inputs. Vacation/sick-leave impact may be inconsistent (#3380).", "likelihood": "Medium", "impact": "High", "severity": "High", "mitigation": "Verify norm values with manual calculation. Compare across pages (My Tasks, Statistics, Confirmation)."},
    {"feature": "Self-approval & direct-approve API", "risk": "No server-side validation prevents self-approval (BUG-REPORT-2) or direct APPROVED creation (BUG-REPORT-3).", "likelihood": "High", "impact": "High", "severity": "High", "mitigation": "Verify API accepts self-approval. Document as known security gap. Test UI mitigation."},
    {"feature": "Cell locking & concurrency", "risk": "LockService is in-memory only. Batch operations skip locking. Race conditions possible.", "likelihood": "Medium", "impact": "Medium", "severity": "Medium", "mitigation": "Test concurrent edits from two sessions. Verify lock icon. Test batch operations."},
    {"feature": "TAB key stacking (production bug)", "risk": "Input field stacking on rapid TAB presses reported by production users (#3398).", "likelihood": "Medium", "impact": "Medium", "severity": "Medium", "mitigation": "Rapid TAB navigation test. Document reproduction steps if reproducible."},
    {"feature": "Contractor report page", "risk": "Infinite spinner when opening contractor report from Statistics/Admin (#3150). Open since Nov 2024.", "likelihood": "High", "impact": "Medium", "severity": "Medium", "mitigation": "Test contractor report from all entry points. Compare direct URL vs Statistics link."},
    {"feature": "Reject record deletion", "risk": "Re-reporting DELETES reject record permanently — no audit trail for rejected hours.", "likelihood": "Low", "impact": "Medium", "severity": "Medium", "mitigation": "Verify reject record deletion on re-report. Document as design issue for audit concerns."},
    {"feature": "Session expiry data loss", "risk": "CAS session expiry leaves UI available but saves may fail silently (#3331).", "likelihood": "Low", "impact": "Critical", "severity": "Medium", "mitigation": "Test reporting after session expiry. Verify redirect to login. Check error handling."},
    {"feature": "Future/dismissed employee display", "risk": "Statistics shows future employees with false under-reports, dismissed employees unsearchable (#3320).", "likelihood": "Medium", "impact": "Low", "severity": "Low", "mitigation": "Verify employee filtering in Statistics. Check search dropdown excludes dismissed."},
    {"feature": "Notification scheduling", "risk": "Multiple schedulers (5-min, daily, Mon/Fri) with different triggers. Forgotten notification threshold (90%) may not account for partial months.", "likelihood": "Low", "impact": "Medium", "severity": "Low", "mitigation": "Trigger each notification type via test endpoints. Verify email content and DB state."},
]


# ─── Helper Functions ─────────────────────────────────────────────────────────

def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, value, is_alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if is_alt else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    return cell


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_link(ws):
    cell = ws.cell(row=1, column=1, value="← Back to Plan Overview")
    cell.font = FONT_BACK_LINK
    cell.hyperlink = "#'Plan Overview'!A1"


# ─── Sheet Generators ────────────────────────────────────────────────────────

def create_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="Reports & Confirmation — Test Plan").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Phase B Session 93").font = FONT_BODY
    ws.cell(row=3, column=1, value="Module: Reports, Confirmation, Periods, Statistics | Branch: release/2.1").font = FONT_BODY

    row = 5
    ws.cell(row=row, column=1, value="Scope & Objectives").font = FONT_SUBTITLE
    row += 1
    objectives = [
        "Comprehensive testing of time report CRUD operations on My Tasks page",
        "Full confirmation workflow: approve/reject on By Employee and By Projects tabs",
        "Period management: report period, approve period, extended periods, 1-month delta rule",
        "Auto-rejection lifecycle: period close → auto-reject → notification → reopen → re-confirm",
        "Statistics: norm calculation chain (general→individual→budget), deviation display, filtering",
        "Notifications: forgotten reports, reject notifications, reports-changed, accounting",
        "Security: role-based access, API vulnerabilities (self-approval, cross-employee, no-auth endpoints)",
    ]
    for obj in objectives:
        ws.cell(row=row, column=1, value=f"• {obj}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Knowledge Sources").font = FONT_SUBTITLE
    row += 1
    sources = [
        "Code analysis: ttt-report-service-deep-dive.md (9.8KB), frontend-report-module.md",
        "Business rules: reports-business-rules-reference.md (4000+ words, 14 bugs documented)",
        "GitLab tickets: 166+ searched, 22 deep-read (descriptions + comments) — 6 theme clusters",
        "API testing: report-crud-api-testing, period-api-live-testing, reject-with-comment-e2e",
        "UI testing: confirmation-flow-live-testing, accounting-pages (for periods)",
        "DB analysis: ttt-backend-schema-deep-dive, office-period-model",
    ]
    for src in sources:
        ws.cell(row=row, column=1, value=f"• {src}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1

    suite_data = [
        ("TS-Reports-CRUD", "My Tasks CRUD", 15),
        ("TS-Reports-Confirmation", "Confirmation Flow", 12),
        ("TS-Reports-Periods", "Period Management", 8),
        ("TS-Reports-AutoReject", "Auto-Rejection", 5),
        ("TS-Reports-Statistics", "Statistics & Norms", 8),
        ("TS-Reports-Notifications", "Notifications", 4),
        ("TS-Reports-Permissions", "Permissions & API", 8),
    ]

    for suite_id, suite_name, count in suite_data:
        cell = ws.cell(row=row, column=1, value=f"{suite_id}: {suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{suite_id}'!A1"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Environment Requirements").font = FONT_SUBTITLE
    row += 1
    envs = [
        "Primary: qa-1 (QA testing environment)",
        "Secondary: timemachine (time manipulation for period tests)",
        "Test clock: PATCH /api/ttt/test/v1/clock for time-sensitive scenarios",
        "Test notifications: POST /api/ttt/test/v1/notifications/* for scheduler triggers",
        "Browser: Chrome (headless + headed modes)",
    ]
    for env in envs:
        ws.cell(row=row, column=1, value=f"• {env}").font = FONT_BODY
        row += 1

    row += 1
    total = sum(s[2] for s in suite_data)
    ws.cell(row=row, column=1, value=f"Total Test Cases: {total}").font = FONT_SUBTITLE

    ws.column_dimensions["A"].width = 100


def create_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Test Suites", "UI Tests", "Hybrid Tests", "API Tests", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, f in enumerate(FEATURES):
        row = i + 2
        is_alt = i % 2 == 1
        total = f["ui"] + f["hybrid"] + f["api"]
        style_body_cell(ws, row, 1, f["feature"], is_alt)
        cell = style_body_cell(ws, row, 2, ", ".join(f["suites"]), is_alt)
        if f["suites"]:
            cell.font = FONT_LINK
            cell.hyperlink = f"#'{f['suites'][0]}'!A1"
        style_body_cell(ws, row, 3, f["ui"], is_alt)
        style_body_cell(ws, row, 4, f["hybrid"], is_alt)
        style_body_cell(ws, row, 5, f["api"], is_alt)
        style_body_cell(ws, row, 6, total, is_alt)

    set_column_widths(ws, [35, 28, 12, 14, 12, 10])


def create_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_HIGH, "Medium": FILL_RISK_MED, "Low": FILL_RISK_LOW}

    for i, r in enumerate(RISKS):
        row = i + 2
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, r["feature"], is_alt)
        style_body_cell(ws, row, 2, r["risk"], is_alt)
        style_body_cell(ws, row, 3, r["likelihood"], is_alt)
        style_body_cell(ws, row, 4, r["impact"], is_alt)
        sev_cell = style_body_cell(ws, row, 5, r["severity"], is_alt)
        sev_cell.fill = severity_fills.get(r["severity"], FILL_ROW_WHITE)
        style_body_cell(ws, row, 6, r["mitigation"], is_alt)

    set_column_widths(ws, [28, 55, 12, 12, 12, 55])


def create_suite_sheet(wb, suite_id, cases):
    ws = wb.create_sheet(suite_id)
    ws.sheet_properties.tabColor = TAB_COLOR_SUITE

    add_back_link(ws)

    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result", "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, tc in enumerate(cases):
        row = i + 3
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, tc["id"], is_alt)
        style_body_cell(ws, row, 2, tc["title"], is_alt)
        style_body_cell(ws, row, 3, tc["preconditions"], is_alt)
        style_body_cell(ws, row, 4, tc["steps"], is_alt)
        style_body_cell(ws, row, 5, tc["expected"], is_alt)
        prio_cell = style_body_cell(ws, row, 6, tc["priority"], is_alt)
        if tc["priority"] == "Critical":
            prio_cell.fill = FILL_RISK_HIGH
        style_body_cell(ws, row, 7, tc["type"], is_alt)
        style_body_cell(ws, row, 8, tc["req_ref"], is_alt)
        style_body_cell(ws, row, 9, tc["module"], is_alt)
        style_body_cell(ws, row, 10, tc.get("notes", ""), is_alt)

    set_column_widths(ws, [14, 40, 50, 70, 45, 10, 8, 35, 22, 40])
    ws.auto_filter.ref = f"A2:J{len(cases) + 2}"


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    create_plan_overview(wb)
    create_feature_matrix(wb)
    create_risk_assessment(wb)

    suite_generators = [
        ("TS-Reports-CRUD", get_crud_cases),
        ("TS-Reports-Confirmation", get_confirmation_cases),
        ("TS-Reports-Periods", get_period_cases),
        ("TS-Reports-AutoReject", get_autoreject_cases),
        ("TS-Reports-Statistics", get_statistics_cases),
        ("TS-Reports-Notifications", get_notification_cases),
        ("TS-Reports-Permissions", get_permission_cases),
    ]

    total = 0
    for suite_id, gen_fn in suite_generators:
        cases = gen_fn()
        create_suite_sheet(wb, suite_id, cases)
        total += len(cases)
        print(f"  {suite_id}: {len(cases)} cases")

    wb.save(OUTPUT_FILE)
    print(f"\nGenerated {OUTPUT_FILE}")
    print(f"Total: {total} test cases across {len(suite_generators)} suites")


if __name__ == "__main__":
    main()
