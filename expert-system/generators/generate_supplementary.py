#!/usr/bin/env python3
"""Generate supplementary test cases for identified gaps.

Session 66 — adds TS-Vac-Supplement tab to vacation.xlsx,
TS-ADM-PMTool-Edge tab to admin.xlsx,
and creates cross-service/cross-service.xlsx for office sync consistency tests.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ── Styling constants (match existing workbooks) ──────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ───────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab."""
    # Back link
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    hr = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, t in enumerate(test_cases):
        row = hr + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            t["id"], t["title"], t["preconditions"], t["steps"],
            t["expected"], t["priority"], t["type"],
            t["req_ref"], t["module"], t.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{ws.max_row}"
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# VACATION SUPPLEMENTARY TEST CASES
# ══════════════════════════════════════════════════════════════════

VACATION_SUPPLEMENT = [
    # SO-transfer double accrual (#2789)
    tc("TC-VAC-131",
       "SO transfer mid-year — available days recalculation (AV=false)",
       "Employee in AV=false office (Russia, norm=28).\n"
       "Has accrued 14 days (6 months worked).\n"
       "Transfer to another AV=false office with norm=24.",
       "1. Note employee's available vacation days before transfer.\n"
       "2. Trigger CS sync with new salary office (norm=24).\n"
       "3. Wait for vacation service sync to complete.\n"
       "4. Query GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "5. Verify formula: accruedDays = 6 × (24/12) = 12.\n"
       "6. Check: availableDays = 12 + currentYearDays + pastYearDays - 24 + futureDays.",
       "Available days correctly recalculated using new office norm (24).\n"
       "No double counting: -normDays uses new office norm.\n"
       "If currentYearDays was initialized with old norm (28), the formula produces:\n"
       "12 + 28 - 24 = 16 (incorrect, should be 12 + 24 - 24 = 12).\n"
       "Flag if result differs from expected.",
       "High", "Functional",
       "#2789", "Vacation / Day Calculation",
       "KEY RISK: RegularCalculationStrategy uses current office norm for both accrual "
       "AND compensation (-normDays), but currentYearDays may have been initialized "
       "with the old office norm. Verify DB: employee_days.days for current year."),

    tc("TC-VAC-132",
       "SO transfer mid-year — AV=false to AV=true office switch",
       "Employee in AV=false office (Russia, norm=28), 6 months worked.\n"
       "Some days already used (e.g., 5 days vacation taken).\n"
       "Transfer to AV=true office (Cyprus, norm=24).",
       "1. Record available days before transfer (AV=false mode: accrual-based).\n"
       "2. Trigger CS sync with new AV=true salary office.\n"
       "3. Wait for vacation service sync.\n"
       "4. GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "5. Verify calculation strategy switched to AdvanceCalculationStrategy.\n"
       "6. Check: availableDays = currentYearDays + pastYearDays + futureDays (no accrual).\n"
       "7. Verify norm deviation recalculation is now active.",
       "Calculation strategy switches from Regular to Advance.\n"
       "Available days reflect full-year balance minus used days.\n"
       "Negative balance allowed (AV=true).\n"
       "Norm deviation recalculation activates for the employee.",
       "High", "Functional",
       "#2789, #3092", "Vacation / Day Calculation",
       "Cross-strategy transfer is high risk — verify employee_days table "
       "is properly adjusted for the strategy change."),

    tc("TC-VAC-133",
       "SO transfer — employee_office year record update verification",
       "Employee in Office A (e.g., Saturn).\n"
       "employee_office table has record for current year with Office A.\n"
       "Transfer to Office B (e.g., Jupiter) with SAME production calendar.",
       "1. DB: SELECT * FROM employee_office WHERE employee_id = X AND year = current_year.\n"
       "2. Trigger CS sync with new office (same calendar).\n"
       "3. DB: Verify employee_office.office_id updated to Office B.\n"
       "4. DB: Verify employee.office_id updated to Office B.\n"
       "5. Verify vacation day calculation uses new office's annual leave norm.",
       "employee_office record updated immediately (same calendar = safe update).\n"
       "Both employee.office_id and employee_office.office_id match new office.\n"
       "Annual leave norm from new office used in calculations.",
       "High", "Integration",
       "#2876", "Vacation / Office Sync",
       "EmployeeOfficeChangedProcessor: isCalendarsAreEqual() allows immediate "
       "update when calendars match. Tests the 'safe path'."),

    tc("TC-VAC-134",
       "SO transfer — different calendar mid-year (deferred update)",
       "Employee in Office A with Russian production calendar.\n"
       "employee_office record exists for current year.\n"
       "Transfer to Office B with Cyprus production calendar.\n"
       "Current date is NOT January 1 and year is current year.",
       "1. DB: Record employee_office for current year (Office A).\n"
       "2. Trigger CS sync with new office (different calendar).\n"
       "3. DB: Check employee_office for current year.\n"
       "4. DB: Check employee_office for next year.\n"
       "5. DB: Check employee.office_id (main record).\n"
       "6. Verify vacation calculations for current year.",
       "employee_office for CURRENT year NOT updated (deferred — different calendar mid-year).\n"
       "employee_office for NEXT year IS updated to new office.\n"
       "employee.office_id IS updated to new office.\n"
       "Current year vacation calculations use OLD office norm.\n"
       "Next year vacation calculations use NEW office norm.",
       "High", "Integration",
       "#2876", "Vacation / Office Sync",
       "This is the known architectural decision: mid-year changes with different "
       "calendars are deferred. Verify the divergence is intentional and documented."),

    # AV=true next-year corner cases (#3347)
    tc("TC-VAC-135",
       "AV=true — next year vacation before nextYearAvailableFromMonth",
       "Employee in AV=true office (Cyprus).\n"
       "Current date: January 15 (before February 1 default).\n"
       "nextYearAvailableFromMonth = 2 (default).",
       "1. POST /api/vacation/v1/vacation/create with startDate in next year.\n"
       "2. Verify response status and error message.\n"
       "3. Change clock to February 1.\n"
       "4. POST same vacation request.\n"
       "5. Verify success.",
       "Step 2: 400 error with NEXT_YEAR_VACATION_NOT_AVAILABLE on startDate.\n"
       "Step 5: Vacation created successfully — next year booking allowed after Feb 1.",
       "Medium", "Functional",
       "#3347, #3322", "Vacation / Create Validation",
       "Config-driven: nextYearAvailableFromMonth. Default Feb, but verify actual "
       "config value on test env. Ticket #3322 mentions Dec-01 which conflicts with code."),

    tc("TC-VAC-136",
       "AV=true — negative balance carry-over to next year",
       "Employee in AV=true office with negative current year balance (-3 days).\n"
       "Next year has just started.",
       "1. DB: Verify employee_days shows negative balance for previous year.\n"
       "2. GET /api/vacation/v1/vacation/days/{employeeId} — check available days.\n"
       "3. Verify AdvanceCalculationStrategy formula: currentYearDays + pastYearDays.\n"
       "4. Check if negative pastYearDays reduces current year available.\n"
       "5. Create a vacation request for current year.\n"
       "6. Verify available days account for the negative carry-over.",
       "Negative pastYearDays from previous year correctly reduces current year available.\n"
       "Available = currentYearDays + pastYearDays (negative) + futureDays.\n"
       "Employee can still create vacations if total is positive.\n"
       "If total is negative, vacation creation should still be allowed (AV=true).",
       "Medium", "Functional",
       "#3361", "Vacation / Day Calculation",
       "AV=true allows negative balances. Verify negative carry-over arithmetic."),

    tc("TC-VAC-137",
       "AV=true — multi-year accumulated balance (3+ years)",
       "Employee in AV=true office for 3+ years.\n"
       "Has unused days from year Y-2, Y-1, and current year Y.\n"
       "Some norm deviation adjustments in past years.",
       "1. DB: SELECT * FROM employee_days WHERE employee_id = X ORDER BY year.\n"
       "2. GET /api/vacation/v1/vacation/days/{employeeId}.\n"
       "3. Verify formula sums all year balances: pastYearDays includes Y-2 + Y-1.\n"
       "4. Create a vacation consuming more than current year balance.\n"
       "5. Verify FIFO consumption: oldest year (Y-2) consumed first.\n"
       "6. Check remaining balances per year after vacation.",
       "Available days = sum of all year balances (currentYearDays + pastYearDays).\n"
       "FIFO consumption: Y-2 depleted first, then Y-1, then Y.\n"
       "Norm deviation adjustments correctly reflected in each year's balance.\n"
       "No year is 'forgotten' or double-counted.",
       "Medium", "Functional",
       "#3361", "Vacation / Day Calculation",
       "Test with employee who has 3+ years of data. Check employee_days for each year."),

    tc("TC-VAC-138",
       "AV=true — norm deviation recalculation after overtime",
       "Employee in AV=true office.\n"
       "Has reported more hours than personal norm for current month.\n"
       "Norm deviation type = Include both (overtime adds, undertime deducts).",
       "1. GET /api/vacation/v1/vacation/days/{employeeId} — record balance.\n"
       "2. Submit time reports exceeding personal norm by 16 hours (2 days).\n"
       "3. Trigger statistics recalculation.\n"
       "4. GET /api/vacation/v1/vacation/days/{employeeId} — check updated balance.\n"
       "5. Verify: daysDelta = (reported - personalNorm) / 8 = 2 days added.\n"
       "6. Check which year's balance was incremented (earliest year with capacity).",
       "Available days increased by 2.\n"
       "daysDelta correctly calculated: (reported - personalNorm) / REPORTING_NORM(8).\n"
       "Days added to earliest year with remaining capacity.\n"
       "If all years full, current year balance increases (can go above norm).",
       "Medium", "Functional",
       "#3092", "Vacation / Norm Deviation",
       "Hardcoded REPORTING_NORM=8. Verify with non-standard work hours."),

    # Maternity edge cases
    tc("TC-VAC-139",
       "Maternity begin — reject NEW vacations only, keep APPROVED",
       "Employee with:\n"
       "- 1 vacation in NEW status (future date)\n"
       "- 1 vacation in APPROVED status (future date)\n"
       "- 1 vacation in PAID status (past date)",
       "1. Record all vacation statuses before maternity.\n"
       "2. Trigger EmployeeMaternityBeginEvent for the employee.\n"
       "3. GET /api/vacation/v1/vacation?employeeId={id} — list all vacations.\n"
       "4. Verify each vacation's status.\n"
       "5. Check employee_days for proportional reduction.\n"
       "6. Check next year days zeroed.",
       "NEW vacation → REJECTED.\n"
       "APPROVED vacation → unchanged (still APPROVED).\n"
       "PAID vacation → unchanged (still PAID).\n"
       "Current year days reduced proportionally.\n"
       "Next year days set to 0.",
       "Medium", "Functional",
       "#3370", "Vacation / Maternity",
       "EmployeeMaternityBeginEventListener.onMaternityBegin() only rejects NEW."),

    # nextYearAvailableFromMonth configuration
    tc("TC-VAC-140",
       "Next year vacation — boundary date (last day before unlock)",
       "nextYearAvailableFromMonth = 2 (February).\n"
       "Clock set to January 31 (last day before unlock).",
       "1. Set clock to Jan 31.\n"
       "2. POST vacation with startDate = next year Jan 5.\n"
       "3. Verify rejection.\n"
       "4. Set clock to Feb 1.\n"
       "5. POST same vacation.\n"
       "6. Verify acceptance.",
       "Jan 31: Rejected with NEXT_YEAR_VACATION_NOT_AVAILABLE.\n"
       "Feb 1: Accepted — next year booking now available.\n"
       "Boundary is exact: Feb 1 00:00:00 enables next year.",
       "Medium", "Boundary",
       "#3322", "Vacation / Create Validation",
       "Uses TimeUtils.today() — test with clock at 23:59 Jan 31 and 00:00 Feb 1."),

    # Calendar change + existing vacation interaction
    tc("TC-VAC-141",
       "Calendar change — existing vacation duration recalculation",
       "Employee in Office A with vacation request spanning Dec 22 - Jan 9.\n"
       "Office calendar currently uses Russian production calendar.\n"
       "Admin changes office to use Cyprus calendar (different holidays).",
       "1. Record vacation duration (working days) under Russian calendar.\n"
       "2. Change production calendar for the office.\n"
       "3. Verify vacation duration recalculated under new calendar.\n"
       "4. Check event feed for DAYS_PER_YEAR_CHANGED event.\n"
       "5. Check available vacation days recalculated.\n"
       "6. Try editing the vacation — verify no 'Calculation error'.",
       "Vacation duration recalculated if calendar change event is properly handled.\n"
       "If #2876 Bug 2 is present: editing/deleting may produce 'Calculation error'.\n"
       "Event feed should log the calendar change (if DAYS_PER_YEAR_CHANGED implemented).\n"
       "Available days should reflect new calendar's working days count.",
       "High", "Integration",
       "#2876", "Vacation / Calendar Integration",
       "Bug 2 from #2876: type change or deletion after calendar switch triggers "
       "'Calculation error'. Check if fixed in current build."),

    tc("TC-VAC-142",
       "VacationStatusUpdateJob — APPROVED to PAID transition timing",
       "Employee has APPROVED vacation ending yesterday.\n"
       "VacationStatusUpdateJob runs every 10 minutes.",
       "1. Create and approve a vacation ending today.\n"
       "2. Set clock to tomorrow.\n"
       "3. Wait up to 10 minutes for job execution.\n"
       "4. GET vacation status.\n"
       "5. Verify transition to PAID.",
       "Vacation transitions from APPROVED to PAID after job runs.\n"
       "Maximum orphan window: 10 minutes.\n"
       "Status is PAID, not still APPROVED.",
       "Low", "Functional",
       "", "Vacation / Status Job",
       "VacationStatusUpdateJob fixedDelay=600000ms. Test the timing window."),
]


# ══════════════════════════════════════════════════════════════════
# ADMIN SUPPLEMENTARY TEST CASES — PM Tool Edge Cases
# ══════════════════════════════════════════════════════════════════

ADMIN_PMTOOL_SUPPLEMENT = [
    tc("TC-ADM-071",
       "PM Tool sync — rate limiter boundary (50 req/min)",
       "PM Tool sync enabled (feature toggle PM_TOOL_SYNC-{env} = ON).\n"
       "PM Tool has 60+ projects (more than rate limit allows per minute).",
       "1. Trigger manual sync: POST /v1/test/project/sync.\n"
       "2. Monitor sync duration.\n"
       "3. Verify all projects synced (not just first 50).\n"
       "4. DB: SELECT COUNT(*) FROM project WHERE pm_tool_id IS NOT NULL.\n"
       "5. Check pm_sync_status for completion record.\n"
       "6. Verify no projects dropped due to rate limiting.",
       "All projects synced despite rate limit — sync takes longer but completes.\n"
       "RateLimiter.acquire() blocks thread until permit available (no timeout).\n"
       "pm_sync_status shows successful completion.\n"
       "No projects missing from sync.",
       "Medium", "Performance",
       "#3399", "Admin / PM Tool Sync",
       "Guava RateLimiter at 50 req/min. acquire() blocks — verify no OOM or timeout."),

    tc("TC-ADM-072",
       "PM Tool sync — API unavailable (connection timeout)",
       "PM Tool sync enabled.\n"
       "PM Tool API is unreachable (simulate by changing pmTool.url to invalid host).",
       "1. Update PM Tool URL config to unreachable host.\n"
       "2. Trigger manual sync: POST /v1/test/project/sync.\n"
       "3. Check application logs for error handling.\n"
       "4. DB: Verify pm_tool_sync_failed_entity table has entries.\n"
       "5. Verify existing project data unchanged.\n"
       "6. Check retry batch mechanism activates.",
       "Sync fails gracefully — no crash or data corruption.\n"
       "Failed entities recorded in pm_tool_sync_failed_entity.\n"
       "Retry mechanism attempts to re-sync failed entities.\n"
       "Existing project data preserved (no deletions on sync failure).",
       "Medium", "Error Handling",
       "", "Admin / PM Tool Sync",
       "Feign client timeout behavior. Check for proper error logging."),

    tc("TC-ADM-073",
       "PM Tool sync — partial failure (some projects fail, others succeed)",
       "PM Tool has 10 projects.\n"
       "1 project references an employee not in TTT DB.",
       "1. Trigger sync: POST /v1/test/project/sync.\n"
       "2. Check application logs for IllegalStateException.\n"
       "3. DB: Verify successfully synced projects are updated.\n"
       "4. DB: Check pm_tool_sync_failed_entity for the failing project.\n"
       "5. Verify failure is isolated — doesn't abort entire sync.\n"
       "6. Check retry: failed project retried in next batch.",
       "9 of 10 projects sync successfully.\n"
       "1 project fails with 500 (missing employee → IllegalStateException).\n"
       "Failed project recorded for retry.\n"
       "Other projects not affected by the single failure.",
       "High", "Error Handling",
       "#3384", "Admin / PM Tool Sync",
       "PmToolProjectSynchronizer validates employee existence — throws "
       "IllegalStateException if missing. Should be business error, not 500."),

    tc("TC-ADM-074",
       "PM Tool sync — concurrent sync attempts (ShedLock contention)",
       "PM Tool sync enabled.\n"
       "Cron schedule: every 15 minutes.\n"
       "ShedLock configured for sync job.",
       "1. Trigger manual sync: POST /v1/test/project/sync.\n"
       "2. Immediately trigger second manual sync.\n"
       "3. Check application logs for both sync attempts.\n"
       "4. Verify ShedLock prevents concurrent execution.\n"
       "5. DB: Check shedlock table for lock status.\n"
       "6. Verify data consistency — no duplicate projects.",
       "First sync acquires ShedLock and runs.\n"
       "Second sync either waits or skips (ShedLock behavior).\n"
       "No concurrent execution — data integrity maintained.\n"
       "No duplicate project entries created.",
       "Medium", "Concurrency",
       "", "Admin / PM Tool Sync",
       "ShedLock prevents concurrent cron execution. Manual trigger may bypass."),

    tc("TC-ADM-075",
       "PM Tool sync — draft project silent ACTIVE conversion",
       "PM Tool has a project in 'draft' status.\n"
       "Project does not exist in TTT yet.",
       "1. Verify project not in TTT: GET /api/ttt/v1/project?name=DraftProject.\n"
       "2. Trigger sync: POST /v1/test/project/sync.\n"
       "3. GET /api/ttt/v1/project?name=DraftProject.\n"
       "4. Check project status in TTT.\n"
       "5. Verify no notification sent about status conversion.",
       "Draft project synced as ACTIVE in TTT.\n"
       "No notification to PM/owner about draft→ACTIVE conversion.\n"
       "Project appears in employee project lists immediately.\n"
       "Design issue: draft projects should not be visible to employees.",
       "Medium", "Functional",
       "", "Admin / PM Tool Sync",
       "Design issue: PmToolProjectSynchronizer silently converts draft→ACTIVE. "
       "No feature flag to exclude drafts."),

    tc("TC-ADM-076",
       "PM Tool sync — pmtId null edge case in UI",
       "Project synced from PM Tool but pmtId field is null.\n"
       "(Legacy project created before pmtId was added)",
       "1. DB: Find project with pm_tool_id IS NOT NULL AND pmt_id IS NULL.\n"
       "2. Navigate to Admin → Projects page.\n"
       "3. Find the project in the table.\n"
       "4. Check project name link behavior.\n"
       "5. Verify Info modal shows correct data.",
       "Project name shows as plain text (no link) when pmtId is null.\n"
       "No broken link or JavaScript error.\n"
       "Info modal displays all available fields correctly.\n"
       "href='#' fallback should not trigger page navigation.",
       "Low", "UI",
       "", "Admin / PM Tool UI",
       "tableHelpers.js: pmtId null → falls back to plain text or href='#'. "
       "Verify no broken behavior."),

    tc("TC-ADM-077",
       "PM Tool sync — observer batch sync post-processing",
       "Project has 5 observers (watchers) in PM Tool.\n"
       "TTT project has 2 observers.",
       "1. Record current observers: GET /api/ttt/v1/project/{id}.\n"
       "2. Trigger sync: POST /v1/test/project/sync.\n"
       "3. GET /api/ttt/v1/project/{id} — check observers.\n"
       "4. Verify observers match PM Tool watchers.\n"
       "5. DB: Verify project_observer table entries.",
       "Observers updated to match PM Tool watchers (5 total).\n"
       "Old observers not in PM Tool are removed.\n"
       "New observers added with correct roles.\n"
       "Cache evicted after observer update.",
       "Medium", "Integration",
       "", "Admin / PM Tool Sync",
       "InternalProjectObserverService handles watcher sync in postProcess()."),

    tc("TC-ADM-078",
       "PM Tool sync — sales employee filtering",
       "PM Tool project has:\n"
       "- PM: employee type\n"
       "- Owner: sales type\n"
       "- Supervisor: employee type",
       "1. Trigger sync: POST /v1/test/project/sync.\n"
       "2. GET /api/ttt/v1/project/{id}.\n"
       "3. Check managerId, ownerId, seniorManagerId.\n"
       "4. Verify sales person not set as owner in TTT.\n"
       "5. Check log for 'removeSalesFromProject' action.",
       "Sales-type CSToolEntityReference filtered out.\n"
       "ownerId = null or previous value (sales person excluded).\n"
       "managerId and seniorManagerId set correctly (employee type).\n"
       "No error for sales filtering — silent exclusion.",
       "Medium", "Functional",
       "", "Admin / PM Tool Sync",
       "PmToolProjectSynchronizer.removeSalesFromProject() filters by isSales()."),
]


# ══════════════════════════════════════════════════════════════════
# CROSS-SERVICE OFFICE CONSISTENCY TEST CASES
# ══════════════════════════════════════════════════════════════════

CROSS_SERVICE_TESTS = [
    tc("TC-XSV-001",
       "Cross-service office consistency — verify after CS sync",
       "Employee exists in both ttt_backend and ttt_vacation.\n"
       "CS sync has run recently.",
       "1. Pick 5 employees from different offices.\n"
       "2. DB (ttt_backend): SELECT login, salary_office FROM employee WHERE login IN (...).\n"
       "3. DB (ttt_vacation): SELECT login, office_id FROM employee WHERE login IN (...).\n"
       "4. Compare salary_office vs office_id for each employee.\n"
       "5. If mismatch found, check employee_office table for year records.",
       "salary_office (ttt_backend) matches office_id (ttt_vacation) for all employees.\n"
       "If mismatch: document which service has correct value.\n"
       "NOTE: 62% historical mismatch rate documented — test newly synced employees.",
       "High", "Data Integrity",
       "#2876", "Cross-Service / Office Sync",
       "Known issue: 736/1190 employees mismatched. Focus on employees synced "
       "AFTER the fix was deployed."),

    tc("TC-XSV-002",
       "Cross-service office — impact on vacation day calculation",
       "Employee with different offices in ttt_backend vs ttt_vacation.\n"
       "Offices have different annual leave norms (e.g., 28 vs 24 days).",
       "1. Find mismatched employee in DB.\n"
       "2. GET /api/vacation/v1/vacation/days/{employeeId} — available days.\n"
       "3. GET /api/ttt/v1/statistic/report?employeeId={id} — check norm used.\n"
       "4. Compare: vacation uses ttt_vacation office norm, statistics uses ttt_backend office norm.\n"
       "5. Verify which norm is displayed to the user on the main page.",
       "Vacation service uses its own office_id for day calculation.\n"
       "Statistics service uses ttt_backend salary_office for norm calculation.\n"
       "If offices differ, employee sees inconsistent norms across pages.\n"
       "Document the discrepancy for each environment.",
       "High", "Data Integrity",
       "#2876", "Cross-Service / Calculation Impact",
       "The practical impact: employee might see '28 days norm' in statistics "
       "but get '24 days accrual' in vacation."),

    tc("TC-XSV-003",
       "Cross-service office — accounting period assignment consistency",
       "Employee with mismatched office between services.\n"
       "Accountant opens/closes periods per salary office.",
       "1. Find employee with office mismatch.\n"
       "2. Check accountant's period management page.\n"
       "3. Verify which office the employee appears under.\n"
       "4. Open period for one office, close for the other.\n"
       "5. Check if employee can submit reports / create vacations.\n"
       "6. Verify period constraints are consistent.",
       "Employee may appear in wrong office for period management.\n"
       "If accounting uses ttt_backend office → different from vacation office.\n"
       "Period open/close may affect wrong employee group.\n"
       "Document which service drives period assignment.",
       "Medium", "Integration",
       "#2876", "Cross-Service / Accounting",
       "Accounting period management uses ttt_backend data. Vacation service "
       "uses its own office_id. Potential for period conflicts."),

    tc("TC-XSV-004",
       "Cross-service office — new employee first sync consistency",
       "New employee created in CompanyStaff.\n"
       "First CS sync runs for both services.",
       "1. Identify a recently created employee (or create test employee in CS).\n"
       "2. Trigger CS sync for ttt_backend.\n"
       "3. Trigger CS sync for ttt_vacation.\n"
       "4. DB (ttt_backend): SELECT salary_office FROM employee WHERE login = 'new_emp'.\n"
       "5. DB (ttt_vacation): SELECT office_id FROM employee WHERE login = 'new_emp'.\n"
       "6. Compare — should match for newly synced employee.",
       "Both services assign the same office from CS data.\n"
       "employee_office year record created in ttt_vacation.\n"
       "No divergence for fresh sync.",
       "High", "Integration",
       "#2876", "Cross-Service / Initial Sync",
       "If first sync produces a match, the divergence is caused by subsequent "
       "office changes — confirms the EmployeeOfficeChangedProcessor is the fix point."),

    tc("TC-XSV-005",
       "Cross-service office — production calendar norm consistency",
       "Employee with mismatched office between services.\n"
       "Each office uses a different production calendar.",
       "1. Find employee with office mismatch where calendars differ.\n"
       "2. GET /api/ttt/v1/employee/{id}/report-period — check working hours norm.\n"
       "3. GET /api/vacation/v1/vacation/days/{employeeId} — check vacation norm.\n"
       "4. Compare production calendar used for working hours vs vacation duration.\n"
       "5. Verify holidays differ between calendars.\n"
       "6. Document impact on vacation spanning holiday periods.",
       "Working hours norm uses ttt_backend office → calendar A.\n"
       "Vacation duration uses ttt_vacation office → calendar B.\n"
       "Different holidays may cause vacation spanning a calendar A holiday "
       "to count differently than calendar B holiday.\n"
       "Employee sees inconsistent 'norm' across UI sections.",
       "Medium", "Data Integrity",
       "#2876", "Cross-Service / Calendar Impact",
       "Russian vs Cyprus calendars have different holidays. A vacation over "
       "Orthodox Christmas (Jan 7) counts differently."),

    tc("TC-XSV-006",
       "Cross-service office — employee_office year record vs employee.office_id",
       "Employee in ttt_vacation service.\n"
       "Multiple years of employee_office records exist.",
       "1. DB: SELECT year, office_id FROM employee_office WHERE employee_id = X ORDER BY year.\n"
       "2. DB: SELECT office_id FROM employee WHERE id = X.\n"
       "3. Compare: employee.office_id should match the latest employee_office year.\n"
       "4. Check for gaps (missing year records).\n"
       "5. Verify all years have valid office references.",
       "employee.office_id matches employee_office[current_year].office_id.\n"
       "No gaps in year records.\n"
       "All office_id values reference valid offices.\n"
       "Within ttt_vacation: internal consistency confirmed.",
       "Low", "Data Integrity",
       "#2876", "Vacation / Internal Consistency",
       "Vault note confirms internal consistency within vacation service. "
       "The divergence is cross-service only."),
]


# ══════════════════════════════════════════════════════════════════
# GENERATION
# ══════════════════════════════════════════════════════════════════

def add_supplement_tab_to_workbook(workbook_path, tab_name, suite_name, test_cases):
    """Add a supplementary test suite tab to an existing workbook."""
    wb = openpyxl.load_workbook(workbook_path)

    # Remove existing tab if present (for re-runs)
    if tab_name in wb.sheetnames:
        del wb[tab_name]

    ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    count = write_ts_tab(ws, suite_name, test_cases)
    wb.save(workbook_path)
    return count


def create_cross_service_workbook(output_path, test_cases):
    """Create a standalone cross-service test workbook."""
    wb = openpyxl.Workbook()

    # ── Plan Overview ──
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = "548235"

    ws_plan.cell(row=1, column=1, value="Cross-Service Office Consistency — Test Plan").font = FONT_TITLE
    ws_plan.cell(row=3, column=1, value="Scope").font = FONT_SECTION
    ws_plan.cell(row=4, column=1,
                 value="Test cases verifying data consistency between ttt_backend and ttt_vacation services, "
                       "focusing on salary office synchronization divergence discovered in Session 65 "
                       "(62% employee mismatch across all environments).").font = FONT_BODY

    ws_plan.cell(row=6, column=1, value="Objectives").font = FONT_SECTION
    objectives = [
        "1. Verify cross-service office consistency after CS sync",
        "2. Assess impact of office divergence on vacation calculations",
        "3. Verify impact on accounting period management",
        "4. Confirm new employee first-sync produces consistent data",
        "5. Document production calendar norm inconsistencies",
        "6. Verify internal consistency within vacation service",
    ]
    for i, obj in enumerate(objectives):
        ws_plan.cell(row=7 + i, column=1, value=obj).font = FONT_BODY

    ws_plan.cell(row=14, column=1, value="Environment Requirements").font = FONT_SECTION
    ws_plan.cell(row=15, column=1,
                 value="All three environments (qa-1, timemachine, stage). "
                       "Requires DB access to both ttt_backend and ttt_vacation schemas.").font = FONT_BODY

    ws_plan.cell(row=17, column=1, value="Test Suites").font = FONT_SECTION
    link_cell = ws_plan.cell(row=18, column=1,
                             value=f"TS-XService-Office — {len(test_cases)} cases")
    link_cell.font = FONT_LINK_BOLD
    link_cell.hyperlink = "#'TS-XService-Office'!A1"

    ws_plan.cell(row=20, column=1, value="Related Tickets").font = FONT_SECTION
    ws_plan.cell(row=21, column=1, value="#2876 — Vacation event feed: office/calendar sync bugs").font = FONT_BODY
    ws_plan.cell(row=22, column=1, value="#2789 — SO transfer double accrual").font = FONT_BODY

    ws_plan.column_dimensions["A"].width = 80

    # ── Risk Assessment ──
    ws_risk = wb.create_sheet("Risk Assessment")
    ws_risk.sheet_properties.tabColor = "548235"
    risk_headers = ["Feature", "Risk", "Likelihood", "Impact", "Severity", "Mitigation/Test Focus"]
    for col, h in enumerate(risk_headers, 1):
        c = ws_risk.cell(row=1, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_GREEN_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    risks = [
        ["Office Sync", "62% employee mismatch between services", "Confirmed", "High",
         "Critical", "TC-XSV-001, TC-XSV-002: verify consistency after fix deployment"],
        ["Vacation Calculation", "Wrong office norm used for accrual", "High", "High",
         "Critical", "TC-XSV-002: compare vacation days vs statistics norms"],
        ["Accounting Periods", "Employee in wrong office for period management", "Medium", "Medium",
         "High", "TC-XSV-003: verify period assignment consistency"],
        ["Calendar Norms", "Different calendars produce different working day counts", "Medium", "Medium",
         "High", "TC-XSV-005: compare holiday handling across services"],
    ]
    for i, risk in enumerate(risks):
        row = 2 + i
        fill = FILL_RISK_HIGH if risk[4] == "Critical" else FILL_RISK_MED if risk[4] == "High" else FILL_RISK_LOW
        for col, val in enumerate(risk, 1):
            c = ws_risk.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = THIN_BORDER
            c.fill = fill

    for i, w in enumerate([20, 45, 12, 10, 10, 50], 1):
        ws_risk.column_dimensions[get_column_letter(i)].width = w

    # ── Test Suite ──
    ws_ts = wb.create_sheet("TS-XService-Office")
    ws_ts.sheet_properties.tabColor = TAB_COLOR_TS
    write_ts_tab(ws_ts, "Cross-Service Office Consistency", test_cases)

    wb.save(output_path)
    return len(test_cases)


def main():
    import os

    base = os.path.dirname(os.path.abspath(__file__))

    # 1. Add TS-Vac-Supplement to vacation.xlsx
    vac_path = os.path.join(base, "vacation", "vacation.xlsx")
    vac_count = add_supplement_tab_to_workbook(
        vac_path, "TS-Vac-Supplement",
        "Vacation Supplementary — SO Transfer, AV Corners, Office Sync",
        VACATION_SUPPLEMENT
    )
    print(f"Vacation supplement: {vac_count} cases added to {vac_path}")

    # 2. Add TS-ADM-PMTool-Edge to admin.xlsx
    adm_path = os.path.join(base, "admin", "admin.xlsx")
    adm_count = add_supplement_tab_to_workbook(
        adm_path, "TS-ADM-PMTool-Edge",
        "PM Tool Edge Cases — Rate Limiting, Failures, Concurrency",
        ADMIN_PMTOOL_SUPPLEMENT
    )
    print(f"Admin PM Tool supplement: {adm_count} cases added to {adm_path}")

    # 3. Create cross-service workbook
    xsv_dir = os.path.join(base, "cross-service")
    os.makedirs(xsv_dir, exist_ok=True)
    xsv_path = os.path.join(xsv_dir, "cross-service.xlsx")
    xsv_count = create_cross_service_workbook(xsv_path, CROSS_SERVICE_TESTS)
    print(f"Cross-service workbook: {xsv_count} cases created at {xsv_path}")

    print(f"\nTotal supplementary cases: {vac_count + adm_count + xsv_count}")


if __name__ == "__main__":
    main()
