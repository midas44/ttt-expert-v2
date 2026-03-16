#!/usr/bin/env python3
"""Generate expanded cross-service integration test workbook.

Session 70 — replaces cross-service/cross-service.xlsx with comprehensive coverage:
  - TS-XService-Office (6 cases, existing — office sync divergence)
  - TS-XService-RabbitMQ (10 cases, NEW — event-driven integration)
  - TS-XService-CSSync (9 cases, NEW — CS sync and lifecycle events)
  - TS-XService-WebSocket (5 cases, NEW — STOMP real-time events)
Total: 30 cases across 4 suites (was 6 cases / 1 suite)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling constants (match existing workbooks) ──────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_BODY_BOLD = Font(name="Arial", bold=True, size=10)

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"
TAB_COLOR_PLAN = "548235"


# ── Helpers ───────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab with back-link, headers, and test case rows."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    hr = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, t in enumerate(test_cases):
        row = hr + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            t["id"], t["title"], t["preconditions"], t["steps"],
            t["expected"], t["priority"], t["type"],
            t["req_ref"], t["module"], t.get("notes", "")
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = THIN_BORDER
            c.fill = fill

    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{ws.max_row}"
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# SUITE 1: OFFICE SYNC DIVERGENCE (existing 6 cases)
# ══════════════════════════════════════════════════════════════════

OFFICE_SYNC_TESTS = [
    tc("TC-XSV-001",
       "Cross-service office consistency — verify after CS sync",
       "Employee exists in both ttt_backend and ttt_vacation.\n"
       "CS sync has run recently.",
       "1. Pick 5 employees from different offices.\n"
       "2. DB (ttt_backend): SELECT login, salary_office FROM employee WHERE login IN (...).\n"
       "3. DB (ttt_vacation): SELECT login, office_id FROM employee WHERE login IN (...).\n"
       "4. Compare salary_office vs office_id for each employee.\n"
       "5. If mismatch found, check employee_office table for year records.",
       "salary_office (ttt_backend) matches office_id (ttt_vacation) for all employees.\n"
       "If mismatch: document which service has correct value.\n"
       "NOTE: 62% historical mismatch rate documented — test newly synced employees.",
       "High", "Data Integrity",
       "#2876", "Cross-Service / Office Sync",
       "Known issue: 736/1190 employees mismatched. Focus on employees synced "
       "AFTER the fix was deployed."),

    tc("TC-XSV-002",
       "Cross-service office — impact on vacation day calculation",
       "Employee with different offices in ttt_backend vs ttt_vacation.\n"
       "Offices have different annual leave norms (e.g., 28 vs 24 days).",
       "1. Find mismatched employee in DB.\n"
       "2. GET /api/vacation/v1/vacation/days/{employeeId} — available days.\n"
       "3. GET /api/ttt/v1/statistic/report?employeeId={id} — check norm used.\n"
       "4. Compare: vacation uses ttt_vacation office norm, statistics uses ttt_backend office norm.\n"
       "5. Verify which norm is displayed to the user on the main page.",
       "Vacation service uses its own office_id for day calculation.\n"
       "Statistics service uses ttt_backend salary_office for norm calculation.\n"
       "If offices differ, employee sees inconsistent norms across pages.\n"
       "Document the discrepancy for each environment.",
       "High", "Data Integrity",
       "#2876", "Cross-Service / Calculation Impact",
       "The practical impact: employee might see '28 days norm' in statistics "
       "but get '24 days accrual' in vacation."),

    tc("TC-XSV-003",
       "Cross-service office — accounting period assignment consistency",
       "Employee with mismatched office between services.\n"
       "Accountant opens/closes periods per salary office.",
       "1. Find employee with office mismatch.\n"
       "2. Check accountant's period management page.\n"
       "3. Verify which office the employee appears under.\n"
       "4. Open period for one office, close for the other.\n"
       "5. Check if employee can submit reports / create vacations.\n"
       "6. Verify period constraints are consistent.",
       "Employee may appear in wrong office for period management.\n"
       "If accounting uses ttt_backend office → different from vacation office.\n"
       "Period open/close may affect wrong employee group.\n"
       "Document which service drives period assignment.",
       "Medium", "Integration",
       "#2876", "Cross-Service / Accounting",
       "Accounting period management uses ttt_backend data. Vacation service "
       "uses its own office_id. Potential for period conflicts."),

    tc("TC-XSV-004",
       "Cross-service office — new employee first sync consistency",
       "New employee created in CompanyStaff.\n"
       "First CS sync runs for both services.",
       "1. Identify a recently created employee (or create test employee in CS).\n"
       "2. Trigger CS sync for ttt_backend.\n"
       "3. Trigger CS sync for ttt_vacation.\n"
       "4. DB (ttt_backend): SELECT salary_office FROM employee WHERE login = 'new_emp'.\n"
       "5. DB (ttt_vacation): SELECT office_id FROM employee WHERE login = 'new_emp'.\n"
       "6. Compare — should match for newly synced employee.",
       "Both services assign the same office from CS data.\n"
       "employee_office year record created in ttt_vacation.\n"
       "No divergence for fresh sync.",
       "High", "Integration",
       "#2876", "Cross-Service / Initial Sync",
       "If first sync produces a match, the divergence is caused by subsequent "
       "office changes — confirms the EmployeeOfficeChangedProcessor is the fix point."),

    tc("TC-XSV-005",
       "Cross-service office — production calendar norm consistency",
       "Employee with mismatched office between services.\n"
       "Each office uses a different production calendar.",
       "1. Find employee with office mismatch where calendars differ.\n"
       "2. GET /api/ttt/v1/employee/{id}/report-period — check working hours norm.\n"
       "3. GET /api/vacation/v1/vacation/days/{employeeId} — check vacation norm.\n"
       "4. Compare production calendar used for working hours vs vacation duration.\n"
       "5. Verify holidays differ between calendars.\n"
       "6. Document impact on vacation spanning holiday periods.",
       "Working hours norm uses ttt_backend office → calendar A.\n"
       "Vacation duration uses ttt_vacation office → calendar B.\n"
       "Different holidays may cause vacation spanning a calendar A holiday "
       "to count differently than calendar B holiday.\n"
       "Employee sees inconsistent 'norm' across UI sections.",
       "Medium", "Data Integrity",
       "#2876", "Cross-Service / Calendar Impact",
       "Russian vs Cyprus calendars have different holidays. A vacation over "
       "Orthodox Christmas (Jan 7) counts differently."),

    tc("TC-XSV-006",
       "Cross-service office — employee_office year record vs employee.office_id",
       "Employee in ttt_vacation service.\n"
       "Multiple years of employee_office records exist.",
       "1. DB: SELECT year, office_id FROM employee_office WHERE employee_id = X ORDER BY year.\n"
       "2. DB: SELECT office_id FROM employee WHERE id = X.\n"
       "3. Compare: employee.office_id should match the latest employee_office year.\n"
       "4. Check for gaps (missing year records).\n"
       "5. Verify all years have valid office references.",
       "employee.office_id matches employee_office[current_year].office_id.\n"
       "No gaps in year records.\n"
       "All office_id values reference valid offices.\n"
       "Within ttt_vacation: internal consistency confirmed.",
       "Low", "Data Integrity",
       "#2876", "Vacation / Internal Consistency",
       "Vault note confirms internal consistency within vacation service. "
       "The divergence is cross-service only."),
]


# ══════════════════════════════════════════════════════════════════
# SUITE 2: RABBITMQ EVENT-DRIVEN INTEGRATION (10 new cases)
# ══════════════════════════════════════════════════════════════════

RABBITMQ_TESTS = [
    tc("TC-XSV-007",
       "Calendar change → vacation working day recalculation",
       "Office with active production calendar.\n"
       "Employees have approved vacations spanning a working day.\n"
       "Admin access to Calendar service.",
       "1. DB (ttt_vacation): Note vacation working_days for employee with vacation spanning target date.\n"
       "2. Admin: Change a working day to a holiday in the production calendar (Calendar service).\n"
       "3. Wait 5s for RabbitMQ CalendarChangedEvent to propagate.\n"
       "4. DB (ttt_vacation): Re-query vacation working_days for the same employee.\n"
       "5. Verify working_days decreased by 1 (holiday no longer counted).\n"
       "6. GET /api/vacation/v1/vacation/{id} — verify API reflects updated working days.",
       "CalendarChangedEvent published to ttt.calendar.topic with routing key 'calendar-changed'.\n"
       "Vacation service CalendarChangedEventHandler processes the event.\n"
       "CalendarUpdateProcessor detects working day → holiday change.\n"
       "VacationCalendarUpdateService.recalculateVacations() updates affected vacations.\n"
       "working_days count decreases by 1 for vacations spanning that date.\n"
       "API response reflects the recalculated value.",
       "High", "Integration",
       "", "RabbitMQ / Calendar → Vacation",
       "Exchange: ttt.calendar.topic. Handler: CalendarChangedEventHandler. "
       "Key flow: CalendarUpdateProcessor.process() with day.diff check."),

    tc("TC-XSV-008",
       "Calendar day deletion → day-off cascade delete and notification",
       "Office calendar has a non-working day (holiday).\n"
       "Employee(s) have NEW or APPROVED day-offs on that date.",
       "1. DB (ttt_vacation): Identify day-off requests with status NEW or APPROVED on the holiday date.\n"
       "2. Note day-off IDs and associated parent vacation IDs.\n"
       "3. Admin: Delete the calendar day entry (remove the holiday) via Calendar service.\n"
       "4. Wait 5s for RabbitMQ CalendarDeletedEvent to propagate.\n"
       "5. DB (ttt_vacation): Verify day-off entities deleted, request statuses updated.\n"
       "6. DB (ttt_vacation): Verify parent vacations recalculated (+1 working day).\n"
       "7. DB (ttt_email): Check for notification email to affected employees.",
       "CalendarDeletedEvent published to ttt.calendar.deleted.topic.\n"
       "EmployeeDayOffCalendarUpdateService.deleteDayOffs() executes:\n"
       "  - Day-off entities with status NEW/APPROVED for that date are deleted.\n"
       "  - Parent vacations recalculated (working days +1 per deleted day-off).\n"
       "  - EmployeeDayOffDeletedFromCalendarEvent published → notification sent.\n"
       "Employee receives notification email about the deleted day-off.",
       "High", "Integration",
       "", "RabbitMQ / Calendar → Vacation (delete)",
       "Exchange: ttt.calendar.deleted.topic. Handler: CalendarDeletedEventHandler. "
       "Cascade: delete day-offs → recalc parent vacation → notify employee."),

    tc("TC-XSV-009",
       "Period advance → vacation payment marking and day-off rejection",
       "Office with open APPROVE period.\n"
       "Employees have approved vacations in the current period.\n"
       "Employee(s) have day-offs past the new period boundary.",
       "1. Note current APPROVE period dates for the office.\n"
       "2. DB (ttt_vacation): Identify approved day-offs past the period boundary.\n"
       "3. Accountant: Advance the APPROVE period via TTT Backend.\n"
       "4. Wait 5s for PeriodChangedEvent to propagate via RabbitMQ.\n"
       "5. DB (ttt_vacation): Check vacation_status_updates table for NEW_FOR_PAID entries.\n"
       "6. DB (ttt_vacation): Verify day-offs past boundary have status REJECTED_BY_SYSTEM.\n"
       "7. GET /api/vacation/v1/vacation/days/{employeeId} — verify available days recalculated.",
       "PeriodChangedEvent published to ttt.backend.officePeriod.topic.\n"
       "Vacation PeriodChangedEventHandler processes (APPROVE type only):\n"
       "  - AvailableDaysRecalculationService.recalculate() runs for the office.\n"
       "  - EmployeeDayOffService.rejectedBySystem() rejects day-offs past boundary.\n"
       "  - VacationStatusUpdatesRepository inserts NEW_FOR_PAID entries.\n"
       "Available vacation days recalculated. Day-offs past boundary rejected.",
       "High", "Integration",
       "", "RabbitMQ / TTT → Vacation (period)",
       "Exchange: ttt.backend.officePeriod.topic. Only APPROVE type triggers processing. "
       "REPORT period changes do NOT trigger vacation recalculation."),

    tc("TC-XSV-010",
       "Period reopen → reverse available days recalculation",
       "Office APPROVE period was recently advanced.\n"
       "Available vacation days were recalculated during advance.",
       "1. DB (ttt_vacation): Note available vacation days for employees in the office.\n"
       "2. Accountant: Reopen the previously advanced APPROVE period.\n"
       "3. Wait 5s for PeriodReopenedEvent to propagate via RabbitMQ.\n"
       "4. DB (ttt_vacation): Verify available vacation days reverted to pre-advance values.\n"
       "5. GET /api/vacation/v1/vacation/days/{employeeId} — confirm API reflects reversal.",
       "PeriodReopenedEvent published to ttt.backend.officePeriod.reopened.topic.\n"
       "PeriodReopenedEventHandler processes (APPROVE type only):\n"
       "  - AvailableDaysRecalculationService.recalculationReverse() runs.\n"
       "  - Vacation day changes from the advance are rolled back.\n"
       "Available days return to pre-advance values.",
       "Medium", "Integration",
       "", "RabbitMQ / TTT → Vacation (reopen)",
       "Exchange: ttt.backend.officePeriod.reopened.topic. Reverse of TC-XSV-009."),

    tc("TC-XSV-011",
       "Employee change → cache invalidation in vacation service",
       "Employee exists in both services.\n"
       "Employee data (e.g., name, department) changes in CS or admin panel.",
       "1. GET /api/vacation/v1/employee/{login} — note cached employee data.\n"
       "2. DB (ttt_backend): UPDATE employee SET name = 'Test Name' WHERE login = '{login}'.\n"
       "3. Trigger EmployeeChangedEvent via CS sync or manual edit.\n"
       "4. Wait 3s for RabbitMQ propagation.\n"
       "5. GET /api/vacation/v1/employee/{login} — verify cache invalidated and fresh data returned.",
       "EmployeeChangedEvent published to ttt.backend.employee.topic.\n"
       "Vacation EmployeeChangedEventHandler evicts employee cache by login.\n"
       "Next API call to vacation service returns fresh employee data.\n"
       "Note: Event carries routing key 'employee-changed'.",
       "Medium", "Integration",
       "", "RabbitMQ / TTT → Vacation (cache)",
       "Exchange: ttt.backend.employee.topic. Handler: EmployeeChangedEventHandler. "
       "Cache eviction only — no data mutation."),

    tc("TC-XSV-012",
       "Month norm recalculation — Vacation → TTT Backend propagation",
       "Calendar change or office sync triggers month norm recalculation.\n"
       "Employee has reports in the affected month.",
       "1. DB (ttt_backend): SELECT month_norm, reported_effort FROM statistic_report "
       "WHERE employee_id = X AND period = '{affected_month}'.\n"
       "2. Trigger a calendar change that affects the working day count for a month.\n"
       "3. Wait 5s for the cascade: CalendarChanged → Vacation recalc → "
       "EmployeeMonthNormContextCalculatedEvent → TTT Backend.\n"
       "4. DB (ttt_backend): Re-query statistic_report — verify month_norm updated.\n"
       "5. GET /api/ttt/v1/statistic/report?employeeId={id}&period={month} — verify API reflects new norm.",
       "EmployeeMonthNormContextCalculatedEvent published by Vacation to ttt.backend.employee.topic.\n"
       "TTT EmployeeMonthNormContextCalculatedEventHandler processes it.\n"
       "StatisticReportSyncService.saveMonthNormAndReportedEffortForEmployees() updates:\n"
       "  - statistic_report.month_norm recalculated\n"
       "  - statistic_report.reported_effort preserved\n"
       "API reflects the new norm value for the affected period.",
       "High", "Integration",
       "", "RabbitMQ / Vacation → TTT (norm)",
       "Reverse direction flow: Vacation publishes back to TTT Backend. "
       "Triggered by MonthNormContextCalculator.prepareMonthNormCalculationContextForEmployees()."),

    tc("TC-XSV-013",
       "Async email event — TTT/Vacation → Email service delivery",
       "Email async feature toggle enabled (email-async).\n"
       "Action triggers notification (e.g., vacation approval).",
       "1. DB (ttt_email): Note current email count.\n"
       "2. Perform a notification-triggering action (e.g., approve a vacation request).\n"
       "3. Wait 5s for SendEmailEvent to propagate via RabbitMQ.\n"
       "4. DB (ttt_email): Verify new email record created with correct template and recipient.\n"
       "5. GET /api/email/v1/email?to={recipient} — verify email in queue.\n"
       "6. Check email subject and body match expected template.",
       "SendEmailEvent published to ttt.email.topic with routing key 'send-email'.\n"
       "Email SendEmailEventHandler processes:\n"
       "  - Converts SendEmailMQ → EmailBO\n"
       "  - Saves email record for batch sending\n"
       "Email record appears in ttt_email database with correct template, recipient, subject.",
       "Medium", "Integration",
       "", "RabbitMQ / TTT & Vacation → Email",
       "Exchange: ttt.email.topic. Feature toggle 'email-async' must be enabled. "
       "Without toggle, email is sent synchronously (no RabbitMQ)."),

    tc("TC-XSV-014",
       "No DLQ — message loss on handler failure (negative test)",
       "RabbitMQ configured with no Dead Letter Queue.\n"
       "Handler will fail (e.g., database constraint violation during processing).",
       "1. Identify a scenario that causes handler failure:\n"
       "   - Option A: Corrupt data that fails deserialization.\n"
       "   - Option B: Calendar event referencing non-existent office.\n"
       "2. Trigger the event.\n"
       "3. Monitor RabbitMQ management UI or logs for the message.\n"
       "4. Verify the message is rejected after exception (ack_mode=AUTO).\n"
       "5. Verify no DLQ queue exists — message is permanently lost.\n"
       "6. DB: Verify no state change occurred (handler failed before commit).",
       "Message is rejected by the handler (IllegalStateException or similar).\n"
       "No DLQ configured — message is silently dropped.\n"
       "No retry occurs (requeue_on_rejection=false).\n"
       "Database state remains unchanged.\n"
       "EXPECTED: Message is permanently lost — this is a KNOWN DESIGN GAP.\n"
       "No alerting or logging of the lost message beyond the exception stack trace.",
       "Medium", "Negative / Resilience",
       "", "RabbitMQ / Error Handling",
       "KNOWN DESIGN ISSUE: No DLQ. Spring AMQP config: max_attempts=3, "
       "exponential backoff 2s initial, requeue_on_rejection=false. "
       "After 3 failed attempts, message is silently lost."),

    tc("TC-XSV-015",
       "Duplicate event processing — no idempotency guard (negative test)",
       "RabbitMQ delivers the same event twice (e.g., network glitch, requeue).\n"
       "Handler has no idempotency check.",
       "1. Trigger a CalendarChangedEvent that adds a holiday.\n"
       "2. DB (ttt_vacation): Note vacation working_days count.\n"
       "3. Manually publish the same CalendarChangedEvent again to the queue.\n"
       "4. Wait 5s for second processing.\n"
       "5. DB (ttt_vacation): Check if working_days was decremented twice.\n"
       "6. Compare expected (1 decrement) vs actual (potentially 2 decrements).",
       "EXPECTED BEHAVIOR (bug): Handler processes the event again.\n"
       "working_days may be decremented twice (double processing).\n"
       "No idempotency key or event deduplication exists.\n"
       "KNOWN DESIGN GAP: All handlers process every message exactly as received,\n"
       "with no check for prior processing of the same event.\n"
       "Note: Actual impact depends on whether recalculation is idempotent by nature\n"
       "(if vacation.recalculate() recomputes from source data, result may be correct).",
       "Low", "Negative / Resilience",
       "", "RabbitMQ / Idempotency",
       "KNOWN DESIGN ISSUE: No idempotency guards. Some handlers may be naturally "
       "idempotent if they recompute from source data rather than incrementing."),

    tc("TC-XSV-016",
       "SystemClockChangedEvent broadcast to all services",
       "Timemachine environment with clock manipulation enabled.\n"
       "All 3 services connected to the same RabbitMQ.",
       "1. GET /api/ttt/v1/test/clock — note current clock across services.\n"
       "2. PATCH clock to a new date/time (e.g., advance 1 day).\n"
       "3. Wait 3s for SystemClockChangedEvent via ttt.fanout exchange.\n"
       "4. GET /api/ttt/v1/test/clock — verify TTT Backend clock updated.\n"
       "5. DB queries to vacation/calendar: verify operations use the new clock.\n"
       "6. Reset clock to original time.",
       "SystemClockChangedEvent published to ttt.fanout (fanout exchange).\n"
       "All 3 services receive via their dedicated queues:\n"
       "  - ttt.fanout.ttt-queue → TTT Backend\n"
       "  - ttt.fanout.email-queue → Email\n"
       "  - ttt.fanout.calendar-queue → Calendar\n"
       "  - ttt.fanout.vacation-queue → Vacation\n"
       "All services apply the new clock. Cron jobs and date-based logic use updated time.",
       "Medium", "Integration",
       "", "RabbitMQ / System Clock (fanout)",
       "Fanout exchange: ttt.fanout. Timemachine-only. "
       "CrudMessageListener processes the clock event in all services."),
]


# ══════════════════════════════════════════════════════════════════
# SUITE 3: CS SYNC AND LIFECYCLE EVENTS (9 new cases)
# ══════════════════════════════════════════════════════════════════

CS_SYNC_TESTS = [
    tc("TC-XSV-017",
       "CS sync — new employee creation across both services",
       "New employee created in CompanyStaff (working=true, active=true).\n"
       "CS sync not yet run.",
       "1. Trigger CS sync for TTT Backend: POST /api/ttt/v1/test/cs/sync.\n"
       "2. Trigger CS sync for Vacation: POST /api/vacation/v1/test/cs/sync.\n"
       "3. DB (ttt_backend): Verify employee record created with correct salary_office.\n"
       "4. DB (ttt_vacation): Verify employee record created with correct office_id.\n"
       "5. Compare salary_office vs office_id — should match for new employee.\n"
       "6. DB (ttt_vacation): Verify employee_office year record created.\n"
       "7. Verify EmployeeHiredEvent was published by vacation service.",
       "Employee created in both services with matching office assignments.\n"
       "TTT Backend: employee record + 7 role post-processors execute:\n"
       "  - DeptMgrRole, ProjMgrRole, Cache, Token, OfficeDirector, Accountant, OfficeHR.\n"
       "Vacation: employee record + EmployeeHiredEvent published.\n"
       "  - employee_office year record created for current year.\n"
       "  - Initial vacation day accrual setup triggered.\n"
       "Office assignments match between services.",
       "High", "Integration",
       "", "CS Sync / Employee Creation",
       "CSEmployeeSynchronizer detection: isHired = (employee null or !working) "
       "AND csEmployee.working AND csEmployee.active."),

    tc("TC-XSV-018",
       "CS sync — employee fired lifecycle event",
       "Active employee in both services.\n"
       "Employee marked as not working (working=false) in CompanyStaff.",
       "1. DB (ttt_vacation): Note employee status (working=true).\n"
       "2. Note any pending vacation requests for the employee.\n"
       "3. Update employee in CS: set working=false.\n"
       "4. Trigger CS sync for vacation service.\n"
       "5. DB (ttt_vacation): Verify employee.working = false.\n"
       "6. Verify EmployeeFiredEvent was published.\n"
       "7. Check downstream effects: are pending vacations canceled? Day-offs rejected?",
       "Vacation CSEmployeeSynchronizer detects: employee.working AND !csEmployee.working.\n"
       "EmployeeFiredEvent published.\n"
       "Employee marked as not working in vacation service.\n"
       "Downstream behavior: document actual impact on pending requests.\n"
       "Note: Contractor termination (#2842) is a separate unimplemented flow.",
       "High", "Integration",
       "#2842", "CS Sync / Employee Lifecycle",
       "Detection: isFired = employee.working AND !csEmployee.working. "
       "Separate from contractor blocking flow."),

    tc("TC-XSV-019",
       "CS sync — maternity start/end event processing",
       "Active employee in vacation service.\n"
       "Employee's maternity flag will be toggled in CompanyStaff.",
       "1. DB (ttt_vacation): Verify employee.maternity = false.\n"
       "2. Note employee's current vacation day balance.\n"
       "3. Update CS: set maternity = true for the employee.\n"
       "4. Trigger CS sync for vacation service.\n"
       "5. DB: Verify employee.maternity = true. Verify EmployeeMaternityBeginEvent published.\n"
       "6. Check: proportional vacation days recalculated? Accrual paused?\n"
       "7. Reverse: set maternity = false in CS, trigger sync.\n"
       "8. Verify EmployeeMaternityEndEvent published. Accrual resumes.",
       "Maternity start:\n"
       "  - Detection: !employee.maternity AND csEmployee.maternity\n"
       "  - EmployeeMaternityBeginEvent published\n"
       "  - Vacation day accrual calculation may be affected (proportional days)\n"
       "Maternity end:\n"
       "  - Detection: employee.maternity AND !csEmployee.maternity\n"
       "  - EmployeeMaternityEndEvent published\n"
       "  - Accrual resumes from the end date\n"
       "V2.1.25 known bug: double accrual possible during maternity period.",
       "High", "Integration",
       "", "CS Sync / Maternity Lifecycle",
       "See vault: maternity-leave-lifecycle.md. V2.1.25 bug: "
       "proportional days can double-count if maternity spans year boundary."),

    tc("TC-XSV-020",
       "CS sync — office change with same production calendar",
       "Employee transfers to new office that uses the SAME production calendar.\n"
       "EmployeeOfficeChangedProcessor handles the change.",
       "1. DB (ttt_vacation): Note employee's current office_id and office_calendar_id.\n"
       "2. Identify target office with the same calendar.\n"
       "3. Change employee's office in CS to the target office.\n"
       "4. Trigger CS sync for vacation service.\n"
       "5. DB: Verify office_id updated to new office.\n"
       "6. DB: Verify employee_office year record updated for current year.\n"
       "7. Verify vacation day balance unchanged (same calendar = same norms).",
       "EmployeeOfficeChangedProcessor.createOrUpdate() processes immediately:\n"
       "  - Same calendar → immediate update of employee_office year record.\n"
       "  - office_id updated in employee table.\n"
       "  - Vacation days unchanged (production calendar norms identical).\n"
       "  - No deferred processing needed.",
       "Medium", "Integration",
       "#2876", "CS Sync / Office Change (same calendar)",
       "EmployeeOfficeChangedProcessor: when old and new offices share the same "
       "calendar, update is immediate. Contrast with TC-XSV-021."),

    tc("TC-XSV-021",
       "CS sync — office change with different calendar (deferred mid-year)",
       "Employee transfers to new office that uses a DIFFERENT production calendar.\n"
       "Transfer happens mid-year (not Jan 1).",
       "1. DB (ttt_vacation): Note employee's current office_id and calendar.\n"
       "2. Identify target office with a different calendar (e.g., Russian → Cyprus).\n"
       "3. Change employee's office in CS to the target office.\n"
       "4. Trigger CS sync for vacation service.\n"
       "5. DB: Check if office_id was updated or deferred.\n"
       "6. DB: Check employee_office year records — was current year updated?\n"
       "7. Verify vacation day calculation uses the correct calendar.",
       "EmployeeOfficeChangedProcessor.createOrUpdate() detects different calendars:\n"
       "  - Mid-year change: conditionally skips year record update.\n"
       "  - BUG: createOrUpdate() may silently skip the update when calendars differ.\n"
       "  - This is the ROOT CAUSE of the 62% office divergence.\n"
       "  - employee.office_id may not match employee_office[current_year].office_id.\n"
       "Expected: Document whether the deferred update is applied or skipped.",
       "High", "Integration",
       "#2876", "CS Sync / Office Change (different calendar)",
       "ROOT CAUSE of 62% divergence. EmployeeOfficeChangedProcessor.createOrUpdate() "
       "conditionally skips mid-year changes when calendars differ. See vault: "
       "cross-service-office-sync-divergence.md."),

    tc("TC-XSV-022",
       "ProjectManagerRolePostProcessor — wrong role removed on demotion (CRITICAL BUG)",
       "Employee has ROLE_PROJECT_MANAGER in TTT.\n"
       "Employee is demoted from project manager in CompanyStaff.",
       "1. DB (ttt_backend): Verify employee has ROLE_PROJECT_MANAGER.\n"
       "2. DB (ttt_backend): Note if employee also has ROLE_DEPARTMENT_MANAGER.\n"
       "3. Demote employee from PM in CS (remove project manager flag).\n"
       "4. Trigger CS sync for TTT Backend.\n"
       "5. DB: Check which role was removed.\n"
       "6. Expected BUG: ROLE_DEPARTMENT_MANAGER removed instead of ROLE_PROJECT_MANAGER.",
       "KNOWN CRITICAL BUG (confirmed in code, line 39):\n"
       "ProjectManagerRolePostProcessor.removeRole() calls:\n"
       "  employeeRoleService.removeRole(employee.getId(), Role.DEPARTMENT_MANAGER)\n"
       "instead of:\n"
       "  employeeRoleService.removeRole(employee.getId(), Role.PROJECT_MANAGER)\n"
       "\n"
       "Result: PM demotion removes DEPARTMENT_MANAGER role.\n"
       "If employee was both PM and DM, they lose DM role but keep PM role.\n"
       "If employee was PM only, the removeRole call has no effect (no DM to remove).",
       "Critical", "Bug Verification",
       "", "CS Sync / Role Assignment",
       "CRITICAL: Copy-paste bug in ProjectManagerRolePostProcessor line 39. "
       "Wrong enum constant used. Affects ALL PM demotions."),

    tc("TC-XSV-023",
       "OfficeHRRolePostProcessor — sticky role never removed",
       "Employee has ROLE_OFFICE_HR assigned via CS sync.\n"
       "Employee is removed from HR duties in CompanyStaff.",
       "1. DB (ttt_backend): Verify employee has ROLE_OFFICE_HR.\n"
       "2. Remove HR assignment in CS.\n"
       "3. Trigger CS sync for TTT Backend.\n"
       "4. DB: Check if ROLE_OFFICE_HR was removed.\n"
       "5. Expected BUG: ROLE_OFFICE_HR remains — only add logic exists, no removal.",
       "KNOWN BUG: OfficeHRRolePostProcessor only implements role assignment.\n"
       "No removal logic exists — role is 'sticky'.\n"
       "Once assigned, ROLE_OFFICE_HR persists even after CS HR flag is removed.\n"
       "Manual DB intervention required to remove the role.\n"
       "Compare with DepartmentManagerRolePostProcessor which has both add AND remove.",
       "High", "Bug Verification",
       "", "CS Sync / Role Assignment",
       "Contrast: DeptMgrRole has add+remove. OfficeDirectorRole has add+remove. "
       "OfficeHRRole has add ONLY. Sticky role accumulation over time."),

    tc("TC-XSV-024",
       "CS sync feature toggle gate — disabled sync behavior",
       "CS sync feature toggle cs-sync-{env} exists in Unleash.\n"
       "Toggle can be disabled.",
       "1. Check current Unleash toggle state for cs-sync-{env}.\n"
       "2. If enabled: trigger CS sync, verify it processes employees.\n"
       "3. Disable the cs-sync-{env} toggle.\n"
       "4. Trigger CS sync again.\n"
       "5. Verify sync silently does nothing — no employees processed.\n"
       "6. Check logs: is a warning logged about disabled sync?\n"
       "7. Re-enable toggle after test.",
       "CSSyncLauncher.sync() checks feature toggle first.\n"
       "If toggle disabled: sync exits immediately with no processing.\n"
       "DESIGN ISSUE: No warning or alert logged when sync is disabled.\n"
       "Silent no-op — could mask configuration errors where sync is accidentally disabled.\n"
       "Employees/offices will progressively drift from CS data.",
       "Medium", "Functional",
       "", "CS Sync / Feature Toggle",
       "Unleash toggle: cs-sync-{env} with env-qualified naming. "
       "See vault: feature-toggles-unleash.md. 6 toggles total, all infrastructure."),

    tc("TC-XSV-025",
       "CS sync — concurrent execution across services (ShedLock isolation)",
       "All 3 services have CS sync scheduled every 15 minutes.\n"
       "ShedLock prevents concurrent execution within a service.",
       "1. Trigger CS sync for all 3 services simultaneously:\n"
       "   - POST /api/ttt/v1/test/cs/sync\n"
       "   - POST /api/vacation/v1/test/cs/sync\n"
       "   - POST /api/calendar/v1/test/cs/sync (if available)\n"
       "2. Monitor all 3 services for concurrent processing.\n"
       "3. DB: Verify ShedLock entries — each service has its own lock.\n"
       "4. Verify no cross-service lock contention.\n"
       "5. Verify each service processes its own entity scope independently.\n"
       "6. Check: do all 3 finish successfully without interfering?",
       "Each service uses ShedLock with independent lock names.\n"
       "All 3 syncs can execute concurrently without interference:\n"
       "  - TTT Backend: employees + contractors + offices (7 post-processors)\n"
       "  - Vacation: employees + offices (2 post-processors)\n"
       "  - Calendar: offices only (startup sync)\n"
       "No cross-service lock contention. Each service processes independently.\n"
       "Thread pool: 10s timeout per entity in all services.",
       "Low", "Integration",
       "", "CS Sync / Concurrency",
       "ShedLock ensures single execution per service. Cross-service concurrency is "
       "expected and safe. Page size: 50 entities per request."),
]


# ══════════════════════════════════════════════════════════════════
# SUITE 4: WEBSOCKET / STOMP REAL-TIME EVENTS (5 new cases)
# ══════════════════════════════════════════════════════════════════

WEBSOCKET_TESTS = [
    tc("TC-XSV-026",
       "Report CRUD → WebSocket event delivery to subscribed clients",
       "User logged in with WebSocket connection established.\n"
       "Subscribed to /topic/employees/{login}/reports/{period}.",
       "1. Open TTT in browser with DevTools Network/WS tab.\n"
       "2. Navigate to My Tasks / Planner (establishes WebSocket connection).\n"
       "3. Create a new time report for the current period.\n"
       "4. Monitor WebSocket frames for ADD event on reports topic.\n"
       "5. Edit the report (change hours).\n"
       "6. Monitor for PATCH event.\n"
       "7. Delete the report.\n"
       "8. Monitor for DELETE event.\n"
       "9. Verify each event contains: EventType, emitterLogin, timestamp, payload.",
       "Each CRUD operation triggers a @TransactionalEventListener (fires after commit).\n"
       "TaskReportEventService sends to /topic/employees/{executorLogin}/reports/{period}:\n"
       "  - ADD: Event<TaskReportEvent> with report data\n"
       "  - PATCH: Event<TaskReportEvent> with updated fields\n"
       "  - DELETE: Event<TaskReportEvent> with deleted report ID\n"
       "Event envelope: {type, emitterLogin, timestamp, value}.\n"
       "All events delivered via SimpMessagingTemplate (STOMP).",
       "High", "Integration",
       "", "WebSocket / Report Events",
       "Topic: /topic/employees/{login}/reports/{period}. "
       "Listener: WsTaskReportEventListener (@Async @TransactionalEventListener)."),

    tc("TC-XSV-027",
       "Task rename → cascade to 3 WebSocket channels",
       "Task exists in planner with reports and assignments.\n"
       "Multiple users subscribed to affected topics.",
       "1. Open browser with WS monitoring.\n"
       "2. Subscribe to:\n"
       "   - /topic/projects/{projectId}/tasks\n"
       "   - /topic/employees/{login}/reports/{period}\n"
       "   - /topic/employees/{login}/assignments/{period}\n"
       "3. Rename a task in the planner.\n"
       "4. Monitor all 3 channels for TASK_RENAME events.\n"
       "5. Verify the task topic receives TASK_RENAME with new name.\n"
       "6. Verify reports topic receives TASK_RENAME (task name displayed in reports).\n"
       "7. Verify assignments topic receives TASK_RENAME (task name displayed in assignments).",
       "WsTaskEventListener handles task rename:\n"
       "  1. Sends TASK_RENAME to /topic/projects/{projectId}/tasks\n"
       "  2. Extracts sub-events for affected employees\n"
       "  3. Re-publishes to /topic/employees/{login}/reports/{period}\n"
       "  4. Re-publishes to /topic/employees/{login}/assignments/{period}\n"
       "All 3 channels receive the rename event. Payload includes old and new task names.\n"
       "Event type: TASK_RENAME in all 3 messages.",
       "Medium", "Integration",
       "", "WebSocket / Task Rename Cascade",
       "Complex event: single rename triggers events on 3 different topic patterns. "
       "WsTaskEventListener extracts sub-events and re-publishes."),

    tc("TC-XSV-028",
       "Cell lock/unlock → STOMP event propagation between users",
       "Two users viewing the same employee's planner.\n"
       "Both subscribed to /topic/employees/{login}/locks.",
       "1. User A: Open planner for employee X. Subscribe to locks topic.\n"
       "2. User B: Open planner for employee X in another browser/session.\n"
       "3. User A: Click on a cell to lock it (start editing).\n"
       "4. User B: Monitor for LOCK event on locks topic.\n"
       "5. Verify LOCK event contains: locked cell coordinates, locker login.\n"
       "6. User A: Release the cell (finish editing or click elsewhere).\n"
       "7. User B: Monitor for UNLOCK event.\n"
       "8. Verify UNLOCK event received.",
       "WsLockEventListener (@EventListener — fires immediately, not transactional):\n"
       "  - LOCK event sent to /topic/employees/{employeeLogin}/locks\n"
       "  - UNLOCK event sent to same topic\n"
       "  - Payload: cell coordinates, emitterLogin, timestamp\n"
       "User B sees the lock/unlock in real time.\n"
       "DESIGN NOTE: No heartbeat-based cleanup — stale locks persist if user disconnects.\n"
       "Disconnected user's locks are NOT automatically released.",
       "Medium", "Integration",
       "", "WebSocket / Cell Locking",
       "Lock uses @EventListener (immediate fire), not @TransactionalEventListener. "
       "Known issue: stale lock risk on client disconnect — no cleanup mechanism."),

    tc("TC-XSV-029",
       "Assignment generation → WebSocket event on assignment channel",
       "Planner page open. Manager generates assignments for an employee/period.",
       "1. Open planner with WS monitoring.\n"
       "2. Subscribe to /topic/employees/{login}/assignments/{period}.\n"
       "3. Trigger assignment generation (Planner → Generate button).\n"
       "4. Monitor for GENERATE event on the assignments topic.\n"
       "5. Verify event payload contains generated assignment data.\n"
       "6. Verify the UI updates in real-time without manual refresh.",
       "TaskAssignmentEventService sends GENERATE event to\n"
       "/topic/employees/{assigneeLogin}/assignments/{period}.\n"
       "@TransactionalEventListener: fires after generation transaction commits.\n"
       "@Async: handled in thread pool.\n"
       "Event payload: generated assignment details.\n"
       "Frontend should process the event and update the planner grid.",
       "Medium", "Integration",
       "", "WebSocket / Planner Generation",
       "GENERATE event is planner-specific. Related to bugs #3332 and #3314 "
       "where frontend state management issues cause stale data after generation."),

    tc("TC-XSV-030",
       "Unauthorized topic subscription — cross-user data leakage test",
       "User A authenticated via WebSocket (JWT).\n"
       "User B has a different login.",
       "1. User A: Establish WebSocket connection, authenticate with JWT.\n"
       "2. User A: Subscribe to /topic/employees/{userB_login}/reports/{period}.\n"
       "3. User B: Create/edit a time report.\n"
       "4. Monitor User A's subscription for events from User B's topic.\n"
       "5. Check: does User A receive User B's report events?\n"
       "6. Verify: is STOMP subscription validated against the authenticated user?\n"
       "7. Document whether topic access control exists.",
       "KNOWN DESIGN GAP: No STOMP subscription authorization.\n"
       "WsChannelInterceptorAdapter validates authentication at CONNECT time,\n"
       "but does NOT validate topic subscriptions against the authenticated user.\n"
       "EXPECTED: User A CAN subscribe to User B's topics and receive events.\n"
       "This is an information leakage vulnerability.\n"
       "Mitigation: STOMP subscription interceptor needed to validate {login} in topic\n"
       "matches the authenticated user's login.",
       "High", "Security",
       "", "WebSocket / Authorization Gap",
       "KNOWN DESIGN ISSUE: WebSocket auth validates CONNECT only, not SUBSCRIBE. "
       "Topic pattern /topic/employees/{login}/* leaks data across users. "
       "See security workbook for related cases."),
]


# ══════════════════════════════════════════════════════════════════
# WORKBOOK GENERATION
# ══════════════════════════════════════════════════════════════════

SUITES = [
    ("TS-XService-Office", "Cross-Service Office Sync Divergence", OFFICE_SYNC_TESTS),
    ("TS-XService-RabbitMQ", "RabbitMQ Event-Driven Integration", RABBITMQ_TESTS),
    ("TS-XService-CSSync", "CompanyStaff Sync & Lifecycle Events", CS_SYNC_TESTS),
    ("TS-XService-WebSocket", "WebSocket/STOMP Real-Time Events", WEBSOCKET_TESTS),
]


def create_workbook(output_path):
    wb = openpyxl.Workbook()
    total_cases = sum(len(s[2]) for s in SUITES)

    # ── Plan Overview ──
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws_plan.cell(row=1, column=1,
                 value="Cross-Service Integration — Test Plan").font = FONT_TITLE
    ws_plan.cell(row=2, column=1,
                 value=f"Total: {total_cases} test cases across {len(SUITES)} suites").font = FONT_BODY

    ws_plan.cell(row=4, column=1, value="Scope").font = FONT_SECTION
    ws_plan.cell(row=5, column=1,
                 value="Comprehensive cross-service integration testing covering:\n"
                       "1. Office sync divergence between ttt_backend and ttt_vacation (62% mismatch)\n"
                       "2. RabbitMQ event-driven messaging: calendar cascades, period events, email, clock sync\n"
                       "3. CompanyStaff synchronization: lifecycle events, role assignment bugs, feature toggles\n"
                       "4. WebSocket/STOMP real-time events: planner coordination, authorization gap").font = FONT_BODY

    ws_plan.cell(row=7, column=1, value="Objectives").font = FONT_SECTION
    objectives = [
        "1. Verify cross-service data consistency after CS synchronization",
        "2. Validate RabbitMQ event propagation across all exchange/queue pairs",
        "3. Confirm CalendarChanged/Deleted cascades correctly update vacations and day-offs",
        "4. Verify period advance/reopen events trigger correct vacation recalculations",
        "5. Test CS sync lifecycle events: hired, fired, maternity, office change",
        "6. Confirm known role assignment bugs (PM demotion, sticky HR role)",
        "7. Verify WebSocket event delivery for report/assignment/lock CRUD",
        "8. Test WebSocket authorization gap (cross-user topic subscription)",
        "9. Document resilience gaps: no DLQ, no idempotency, no subscription auth",
    ]
    for i, obj in enumerate(objectives):
        ws_plan.cell(row=8 + i, column=1, value=obj).font = FONT_BODY

    ws_plan.cell(row=18, column=1, value="Environment Requirements").font = FONT_SECTION
    ws_plan.cell(row=19, column=1,
                 value="Primary: timemachine (clock manipulation, full test API).\n"
                       "Secondary: qa-1 (validation). Production: stage (read-only verification).\n"
                       "Requires: DB access to ttt_backend, ttt_vacation, ttt_calendar, ttt_email schemas.\n"
                       "Requires: RabbitMQ management access for message monitoring.\n"
                       "Requires: Browser DevTools for WebSocket frame inspection.").font = FONT_BODY

    ws_plan.cell(row=21, column=1, value="Test Suites").font = FONT_SECTION
    for i, (tab_name, suite_name, cases) in enumerate(SUITES):
        link_cell = ws_plan.cell(row=22 + i, column=1,
                                 value=f"{tab_name} — {suite_name} ({len(cases)} cases)")
        link_cell.font = FONT_LINK_BOLD
        link_cell.hyperlink = f"#'{tab_name}'!A1"

    ws_plan.cell(row=27, column=1, value="Related Tickets").font = FONT_SECTION
    tickets = [
        "#2876 — Vacation event feed: office/calendar sync bugs",
        "#2842 — Contractor termination (lifecycle, not yet implemented)",
        "#3026 — CS office settings: 3 unimplemented settings",
        "#2789 — SO transfer double accrual",
    ]
    for i, t in enumerate(tickets):
        ws_plan.cell(row=28 + i, column=1, value=t).font = FONT_BODY

    ws_plan.cell(row=33, column=1, value="Knowledge Base References").font = FONT_SECTION
    refs = [
        "analysis/cross-service-integration.md — Architecture, event flows, design issues",
        "architecture/rabbitmq-messaging.md — 8 exchanges, queue topology",
        "architecture/websocket-events.md — 12 event types, 7 STOMP topics",
        "modules/companystaff-integration.md — CS sync, 9 post-processors, 7 bugs",
        "data-findings/cross-service-office-sync-divergence.md — 62% employee mismatch",
        "investigations/maternity-leave-lifecycle.md — Maternity event flow",
    ]
    for i, r in enumerate(refs):
        ws_plan.cell(row=34 + i, column=1, value=r).font = FONT_BODY

    ws_plan.column_dimensions["A"].width = 90

    # ── Feature Matrix ──
    ws_matrix = wb.create_sheet("Feature Matrix")
    ws_matrix.sheet_properties.tabColor = TAB_COLOR_PLAN

    matrix_headers = ["Feature Area", "Integration", "Data Integrity",
                      "Bug Verification", "Negative/Resilience", "Security", "Total", "Suite"]
    for col, h in enumerate(matrix_headers, 1):
        c = ws_matrix.cell(row=1, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_GREEN_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    matrix_data = [
        ["Office Sync Divergence", 2, 4, 0, 0, 0, 6, "TS-XService-Office"],
        ["RabbitMQ Events", 7, 0, 0, 2, 0, 10, "TS-XService-RabbitMQ"],  # TC-XSV-014,015 are negative
        ["CS Sync Lifecycle", 7, 0, 2, 0, 0, 9, "TS-XService-CSSync"],
        ["WebSocket/STOMP", 4, 0, 0, 0, 1, 5, "TS-XService-WebSocket"],
    ]
    for i, row_data in enumerate(matrix_data):
        row = 2 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        for col, val in enumerate(row_data, 1):
            c = ws_matrix.cell(row=row, column=col, value=val)
            c.font = FONT_BODY if col < len(row_data) else FONT_LINK
            c.alignment = ALIGN_LEFT if col == 1 else ALIGN_CENTER
            c.border = THIN_BORDER
            c.fill = fill
            if col == len(row_data):
                c.hyperlink = f"#'{val}'!A1"

    # Totals row
    total_row = 2 + len(matrix_data)
    totals = ["TOTAL", 20, 4, 2, 2, 1, total_cases, ""]
    for col, val in enumerate(totals, 1):
        c = ws_matrix.cell(row=total_row, column=col, value=val)
        c.font = FONT_BODY_BOLD
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        c.fill = FILL_GREEN_HEADER
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    for i, w in enumerate([25, 12, 14, 16, 18, 10, 8, 22], 1):
        ws_matrix.column_dimensions[get_column_letter(i)].width = w

    # ── Risk Assessment ──
    ws_risk = wb.create_sheet("Risk Assessment")
    ws_risk.sheet_properties.tabColor = TAB_COLOR_PLAN

    risk_headers = ["Feature", "Risk", "Likelihood", "Impact", "Severity", "Mitigation/Test Focus"]
    for col, h in enumerate(risk_headers, 1):
        c = ws_risk.cell(row=1, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_GREEN_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    risks = [
        ["Office Sync", "62% employee mismatch between services",
         "Confirmed", "High", "Critical",
         "TC-XSV-001–006: verify consistency, impact on calculations and periods"],
        ["RabbitMQ: No DLQ", "Failed messages silently lost, no retry",
         "High", "High", "Critical",
         "TC-XSV-014: confirm message loss. No mitigation available without code change"],
        ["CS Sync: PM Role Bug", "PM demotion removes wrong role (DEPARTMENT_MANAGER)",
         "Confirmed", "High", "Critical",
         "TC-XSV-022: verify bug. Affects ALL project manager demotions"],
        ["Calendar Cascade", "Calendar change triggers cascading vacation recalc",
         "Medium", "High", "High",
         "TC-XSV-007–008: verify cascade correctness and notification delivery"],
        ["Period Events", "Period advance/reopen triggers vacation day recalculation",
         "Medium", "High", "High",
         "TC-XSV-009–010: verify advance marks payment, reopen reverses"],
        ["CS Sync: Sticky HR", "OfficeHRRole never removed on demotion",
         "Confirmed", "Medium", "High",
         "TC-XSV-023: verify role persists after CS removal"],
        ["WebSocket Auth Gap", "Cross-user topic subscription leaks data",
         "Confirmed", "Medium", "High",
         "TC-XSV-030: verify unauthorized subscription receives events"],
        ["Duplicate Events", "No idempotency — duplicate events may cause double processing",
         "Low", "Medium", "Medium",
         "TC-XSV-015: verify handler behavior on duplicate. May be naturally idempotent"],
        ["CS Sync Toggle", "Disabled toggle silently stops sync with no warning",
         "Low", "Medium", "Medium",
         "TC-XSV-024: verify silent no-op behavior"],
        ["WebSocket Stale Locks", "Disconnected client locks persist indefinitely",
         "Medium", "Low", "Medium",
         "TC-XSV-028: verify lock persistence on disconnect"],
    ]
    for i, risk in enumerate(risks):
        row = 2 + i
        sev = risk[4]
        fill = FILL_RISK_HIGH if sev == "Critical" else FILL_RISK_MED if sev == "High" else FILL_RISK_LOW
        for col, val in enumerate(risk, 1):
            c = ws_risk.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = THIN_BORDER
            c.fill = fill

    for i, w in enumerate([22, 48, 12, 10, 10, 55], 1):
        ws_risk.column_dimensions[get_column_letter(i)].width = w

    # ── Test Suite Tabs ──
    for tab_name, suite_name, cases in SUITES:
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_TS
        write_ts_tab(ws, suite_name, cases)

    wb.save(output_path)
    return total_cases


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    xsv_dir = os.path.join(base, "cross-service")
    os.makedirs(xsv_dir, exist_ok=True)
    output_path = os.path.join(xsv_dir, "cross-service.xlsx")

    total = create_workbook(output_path)
    print(f"Cross-service workbook: {total} cases created at {output_path}")
    print(f"  Suites: {len(SUITES)}")
    for tab_name, suite_name, cases in SUITES:
        print(f"    {tab_name}: {len(cases)} cases — {suite_name}")


if __name__ == "__main__":
    main()
