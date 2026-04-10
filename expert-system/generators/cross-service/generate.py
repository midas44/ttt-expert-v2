#!/usr/bin/env python3
"""
Cross-Service Integration Test Documentation Generator — Phase B
Generates test-docs/cross-service/cross-service.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 8 TS- test suite tabs.

Based on vault knowledge: cross-service-integration.md (RabbitMQ events, CS sync, WebSocket,
design issues), cross-service-office-sync-divergence.md (62% office mismatch), cross-service-
ticket-findings.md (~75 tickets), rabbitmq-messaging.md (8 exchanges, 9 queues),
companystaff-integration.md (sync across 3 services, 7 bugs), pm-tool-sync-implementation.md,
email-notification-deep-dive.md, websocket-events.md, feature-toggles-unleash.md.

8 Suites, ~78 test cases:
  TS-CrossService-CSSync        — CompanyStaff sync: employee, office, roles, failures
  TS-CrossService-EventProp     — RabbitMQ event propagation: calendar→vacation, period→vacation
  TS-CrossService-DataIntegrity — Cross-service data consistency and divergence
  TS-CrossService-WebSocket     — STOMP real-time planner events, JWT expiry
  TS-CrossService-TrackerSync   — External tracker integration (Jira, ClickUp, YouTrack)
  TS-CrossService-PMToolSync    — PM Tool project synchronization
  TS-CrossService-EmailDelivery — Async/sync email path, templates, batch processing
  TS-CrossService-Regression    — Ticket-derived regression tests
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Constants ---------------------------------------------------------------

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "cross-service")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "cross-service.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"

# --- Plan Overview -----------------------------------------------------------

PLAN_OVERVIEW = {
    "title": "Cross-Service Integration — Test Plan",
    "scope": (
        "Cross-service integration testing across all TTT microservices: CompanyStaff (CS) "
        "synchronization across 3 services, RabbitMQ event propagation (calendar→vacation, "
        "period→vacation, email delivery), cross-service data consistency (62% office divergence), "
        "WebSocket (STOMP) real-time planner events, external tracker integrations (Jira, ClickUp, "
        "YouTrack, Asana), PM Tool project synchronization, email notification delivery via async "
        "and sync paths, feature toggle behavior, and regression tests for 75+ integration tickets."
    ),
    "objectives": [
        "Verify CS sync propagates employee/office/role changes correctly across TTT, Vacation, and Calendar services",
        "Test RabbitMQ event cascade: calendar change → vacation recalculation, period advance → payment marking",
        "Detect cross-service data divergence: office assignments, employee fields, last_date, contractor sync gaps",
        "Validate WebSocket (STOMP) real-time events in Planner: task rename cascade, lock/unlock, JWT expiry reconnection",
        "Test external tracker integrations: Jira (PAT auth), ClickUp (API), YouTrack, task import correctness",
        "Verify PM Tool project sync: rate limiting, validation cascade, failed project retry, employee filtering",
        "Test email delivery across async (RabbitMQ) and sync (Feign) paths, template rendering, batch processing",
        "Regression coverage for 75+ cross-service integration tickets across CS sync, RabbitMQ, trackers, PM Tool",
    ],
    "environments": [
        "QA-1: https://ttt-qa-1.noveogroup.com (primary — current sprint build)",
        "Timemachine: https://ttt-timemachine.noveogroup.com (clock manipulation, sync trigger testing)",
        "Stage: https://ttt-stage.noveogroup.com (production baseline for divergence comparison)",
    ],
    "approach": (
        "Hybrid UI + API testing. CS sync and event propagation triggered via test API endpoints, "
        "results verified in UI (employee pages, vacation lists, planner). Data divergence tested "
        "via DB queries comparing ttt_backend and ttt_vacation schemas. WebSocket events tested via "
        "Planner UI with concurrent browser sessions. Tracker integrations tested via Admin UI "
        "configuration and Planner task import. Email delivery verified via email service API and "
        "DB status checks. Feature toggles tested by comparing behavior with toggle ON vs OFF."
    ),
    "dependencies": [
        "Test API access for CS sync trigger (POST /api/ttt/test/v1/cs-sync/employees)",
        "Test API access for clock manipulation (PATCH /api/ttt/test/v1/clock) on Timemachine",
        "Database access (SELECT) to ttt_backend, ttt_vacation, ttt_calendar, ttt_email schemas",
        "Multiple test accounts: ADMIN, EMPLOYEE, PM, DM, ACCOUNTANT for role-specific sync tests",
        "External tracker access (Jira, ClickUp) for integration tests or pre-configured tracker projects",
        "PM Tool API availability on the target environment",
        "RabbitMQ Management UI access (optional — for queue inspection)",
    ],
}

# --- Risk Assessment ---------------------------------------------------------

RISK_ASSESSMENT = [
    ("CS sync office divergence (62%)",
     "Independent CS sync paths in TTT and Vacation cause 736/1190 employees (62%) to have different "
     "salary office assignments. Affects vacation day calculation, statistics norms, and accounting periods.",
     "High", "Critical", "Critical",
     "Compare office_id between ttt_backend.employee and ttt_vacation.employee. Verify new employee sync "
     "results in consistent office across services."),
    ("RabbitMQ event loss — no DLQ",
     "No Dead Letter Queue configured. Failed messages after 3 retries are silently dropped. Calendar "
     "change could be committed but RabbitMQ message never delivered to Vacation service.",
     "Medium", "Critical", "Critical",
     "Test event propagation end-to-end: admin changes calendar → verify vacation recalculated. "
     "Accepted risk (#3262): 3 AM re-sync is the safety net."),
    ("CS ProjectManagerRolePostProcessor bug",
     "Post-processor removes ROLE_DEPARTMENT_MANAGER instead of ROLE_PROJECT_MANAGER on PM demotion. "
     "Wrong role removed from employee.",
     "Medium", "Critical", "Critical",
     "Verify role changes after CS sync for PM promotion/demotion. Check employee_global_roles table."),
    ("PM Tool validation cascade — 84% project sync failure",
     "Any missing watcher in a PM Tool project causes entire project sync to fail (not just skip "
     "the watcher). Results in 3132 vs 501 project count mismatch.",
     "High", "High", "High",
     "Trigger PM Tool sync, verify project count. Check pm_sync_status and pm_tool_sync_failed_entity tables."),
    ("Email async path — hours-long delays (#2518)",
     "After migration to RabbitMQ, email delivery experienced hours-long delays. Chronological "
     "ordering broken. Affects time-sensitive notifications.",
     "Medium", "High", "High",
     "Trigger notification (e.g., vacation creation), verify email arrives in email service DB within "
     "reasonable time. Check status progression NEW → SENT."),
    ("WebSocket JWT expiry — polling fallback (#2270)",
     "Expired JWT breaks WebSocket reconnection. Fix added polling /v1/authentication/check every 5s. "
     "Still fragile for overnight Planner sessions.",
     "Medium", "High", "High",
     "Test Planner open during JWT expiry. Verify no infinite reconnection loop, graceful fallback."),
    ("CS sync crash under load (#3023)",
     "Multiple sequential full CS syncs crash vacation service. DB connection pool exhaustion with "
     ">30 concurrent threads.",
     "Low", "Critical", "High",
     "Trigger CS sync and monitor service health. Not a standard test case — document as known limitation."),
    ("Tracker auth fragility (Jira PAT broken #2511)",
     "Jira PAT authentication broken — 'wrong server' error. ClickUp export was completely broken (#3148). "
     "Each tracker has unique auth with different failure modes.",
     "High", "Medium", "High",
     "Test tracker configuration and task import for each supported tracker type."),
    ("Feature toggle silently disables sync",
     "cs-sync toggle set to false = sync silently stops. No warning in logs or monitoring. "
     "Data drifts without any indication.",
     "Medium", "Medium", "Medium",
     "Verify sync behavior with toggle ON and OFF. Check that no misleading success messages appear."),
    ("Test clock corrupts sync state (#2629)",
     "Moving time backward creates duplicates in cs_sync_status table. Vacation service crashed "
     "after manual table cleanup.",
     "Medium", "Medium", "Medium",
     "On Timemachine: advance clock, trigger sync, revert clock, trigger sync. Verify no corruption."),
]

# --- Feature Matrix ----------------------------------------------------------
# [feature, CSSync, EventProp, DataIntegrity, WebSocket, TrackerSync, PMToolSync, EmailDelivery, Regression, Total]

FEATURE_MATRIX = [
    ["CS employee sync (new/update/deactivate)",     4, 0, 0, 0, 0, 0, 0, 0, 4],
    ["CS office sync (create/transfer)",             3, 0, 0, 0, 0, 0, 0, 0, 3],
    ["CS role post-processors",                      3, 0, 0, 0, 0, 0, 0, 0, 3],
    ["CS sync failure/recovery",                     2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["Calendar change → vacation recalculation",     0, 3, 0, 0, 0, 0, 0, 0, 3],
    ["Period advance → vacation payment",            0, 3, 0, 0, 0, 0, 0, 0, 3],
    ["Calendar deleted → day-off cascade",           0, 2, 0, 0, 0, 0, 0, 0, 2],
    ["Employee change → cache invalidation",         0, 2, 0, 0, 0, 0, 0, 0, 2],
    ["Office divergence (backend vs vacation)",      0, 0, 3, 0, 0, 0, 0, 0, 3],
    ["Employee field consistency",                   0, 0, 3, 0, 0, 0, 0, 0, 3],
    ["Contractor sync gap",                          0, 0, 2, 0, 0, 0, 0, 0, 2],
    ["STOMP planner events (rename/lock)",           0, 0, 0, 4, 0, 0, 0, 0, 4],
    ["WebSocket JWT expiry reconnection",            0, 0, 0, 2, 0, 0, 0, 0, 2],
    ["Concurrent editing conflict",                  0, 0, 0, 2, 0, 0, 0, 0, 2],
    ["Tracker auth and task import",                 0, 0, 0, 0, 5, 0, 0, 0, 5],
    ["Tracker multi-project conflict",               0, 0, 0, 0, 2, 0, 0, 0, 2],
    ["Tracker approved status stale",                0, 0, 0, 0, 1, 0, 0, 0, 1],
    ["PM Tool project sync lifecycle",               0, 0, 0, 0, 0, 4, 0, 0, 4],
    ["PM Tool rate limiting",                        0, 0, 0, 0, 0, 2, 0, 0, 2],
    ["PM Tool failed project retry",                 0, 0, 0, 0, 0, 2, 0, 0, 2],
    ["Email async (RabbitMQ) delivery",              0, 0, 0, 0, 0, 0, 3, 0, 3],
    ["Email template rendering",                     0, 0, 0, 0, 0, 0, 2, 0, 2],
    ["Digest compilation",                           0, 0, 0, 0, 0, 0, 2, 0, 2],
    ["Regression — CS sync bugs",                    0, 0, 0, 0, 0, 0, 0, 4, 4],
    ["Regression — RabbitMQ/event bugs",             0, 0, 0, 0, 0, 0, 0, 3, 3],
    ["Regression — PM Tool/tracker bugs",            0, 0, 0, 0, 0, 0, 0, 3, 3],
]

# --- Test Suites -------------------------------------------------------------

SUITES = {
    "TS-CrossService-CSSync": [
        ("TC-CS-001", "New employee appears in all services after CS sync",
         "CompanyStaff has a recently added employee not yet in TTT. Test API endpoint available: POST /api/ttt/test/v1/cs-sync/employees. "
         "Query: SELECT login FROM ttt_backend.employee WHERE enabled = true ORDER BY first_date DESC LIMIT 5 (to identify recently synced employees).",
         "SETUP: Via test API — trigger CS employee sync: POST /api/ttt/test/v1/cs-sync/employees\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for the recently synced employee by name\n"
         "4. Verify the employee appears in the employee list\n"
         "5. Navigate to Calendar of absences\n"
         "6. Search for the same employee — verify they appear (vacation service has the record)\n"
         "DB-CHECK: SELECT login, salary_office FROM ttt_backend.employee WHERE login = '<login>'\n"
         "DB-CHECK: SELECT login, office_id FROM ttt_vacation.employee WHERE login = '<login>'\n"
         "DB-CHECK: Verify salary_office and office_id point to the same office",
         "New employee visible in Admin employee list AND Calendar of absences. Office assignment consistent between ttt_backend and ttt_vacation.",
         "Critical", "Hybrid", "companystaff-integration.md, #2474", "cross-service",
         "CS sync runs every 15 min. Use test API to trigger immediately. Vacation service only syncs employees (not contractors)."),

        ("TC-CS-002", "Office transfer propagates to both services after CS sync",
         "Employee whose salary office was recently changed in CompanyStaff. Test API for sync trigger. "
         "Query: SELECT e.login, e.salary_office FROM ttt_backend.employee e WHERE e.enabled = true AND e.salary_office IS NOT NULL ORDER BY random() LIMIT 1",
         "SETUP: Note the employee's current office in both services before sync\n"
         "SETUP: Via test API — trigger CS employee sync: POST /api/ttt/test/v1/cs-sync/employees\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Find the employee and check their salary office\n"
         "4. Navigate to Calendar of absences\n"
         "5. Check the same employee's office context (reflected in absence calculations)\n"
         "DB-CHECK: SELECT salary_office FROM ttt_backend.employee WHERE login = '<login>'\n"
         "DB-CHECK: SELECT office_id FROM ttt_vacation.employee WHERE login = '<login>'\n"
         "DB-CHECK: SELECT office FROM ttt_vacation.employee_office WHERE employee_id = <id> AND year = EXTRACT(YEAR FROM CURRENT_DATE)",
         "After sync, employee's office is updated in both ttt_backend.employee.salary_office and ttt_vacation.employee.office_id. "
         "employee_office yearly record also reflects the new office.",
         "Critical", "Hybrid", "cross-service-office-sync-divergence.md, #2969", "cross-service",
         "Mid-year office changes with different calendars may silently skip the year record update (EmployeeOfficeChangedProcessor conditional logic)."),

        ("TC-CS-003", "Employee deactivation reflected across services",
         "Enabled employee that exists in both services. "
         "Query: SELECT b.login FROM ttt_backend.employee b JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.enabled = true AND b.deactivated = false LIMIT 5",
         "SETUP: Note employee status in both services\n"
         "SETUP: Via test API — trigger CS sync (simulates deactivation from CS)\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for the deactivated employee\n"
         "4. Verify the employee is no longer shown in active employee list (or marked as deactivated)\n"
         "5. Navigate to Calendar of absences\n"
         "6. Verify the employee no longer appears in active absence calendar\n"
         "DB-CHECK: SELECT enabled, deactivated, being_dismissed FROM ttt_backend.employee WHERE login = '<login>'\n"
         "DB-CHECK: SELECT working FROM ttt_vacation.employee WHERE login = '<login>'",
         "Deactivated employee marked as not active in both ttt_backend (enabled=false or deactivated=true) and ttt_vacation (working=false). "
         "Employee hidden from active lists in UI.",
         "High", "Hybrid", "#1275, #2953", "cross-service",
         "Ticket #2953: employee returned from maternity but TTT still showed 'in maternity' — sync stopped for months."),

        ("TC-CS-004", "Maternity status change detected and propagated",
         "Employee whose maternity status changes in CS. "
         "Query: SELECT login FROM ttt_vacation.employee WHERE maternity = true AND working = true LIMIT 3",
         "SETUP: Via test API — trigger CS employee sync\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for the employee — verify maternity status shown correctly\n"
         "4. Navigate to Calendar of absences\n"
         "5. Check the employee's availability (maternity employees may have different display)\n"
         "DB-CHECK: SELECT login, maternity, working FROM ttt_vacation.employee WHERE login = '<login>'\n"
         "DB-CHECK: Verify EmployeeMaternityBeginEvent or EmployeeMaternityEndEvent was published (check application logs)",
         "Maternity status change from CS is detected during sync. Vacation service publishes appropriate EmployeeMaternityBeginEvent or EmployeeMaternityEndEvent. "
         "Employee's maternity flag updated in ttt_vacation.employee.",
         "High", "Hybrid", "#2953, companystaff-integration.md §Vacation Service Differences", "cross-service",
         "Vacation CSEmployeeSynchronizer detects: !employee.maternity AND csEmployee.maternity → EmployeeMaternityBeginEvent."),

        ("TC-CS-005", "New salary office created from CS appears in TTT",
         "CompanyStaff has a new salary office not yet in TTT. "
         "Query: SELECT id, name FROM ttt_backend.office ORDER BY id DESC LIMIT 5",
         "SETUP: Via test API — trigger CS office sync\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Production calendars\n"
         "3. Verify the new salary office appears in the office dropdown/list\n"
         "4. Navigate to Accounting section\n"
         "5. Verify the new office appears in office selection filters\n"
         "DB-CHECK: SELECT id, name FROM ttt_backend.office WHERE name LIKE '%<new_office_name>%'\n"
         "DB-CHECK: SELECT id, name FROM ttt_vacation.office WHERE name LIKE '%<new_office_name>%'\n"
         "DB-CHECK: SELECT id, name FROM ttt_calendar.office WHERE name LIKE '%<new_office_name>%'",
         "New salary office appears in all 3 service databases and is accessible in Admin UI (Production calendars, Accounting).",
         "High", "Hybrid", "#3241, companystaff-integration.md §Office Sync", "cross-service",
         "Ticket #3241: new SO created in CS didn't appear in ttt_backend.office — only creation was broken."),

        ("TC-CS-006", "DM role assigned via CS post-processor after sync",
         "Employee who is a department head in CS but does not yet have ROLE_DEPARTMENT_MANAGER in TTT. "
         "Query: SELECT e.login FROM ttt_backend.employee e LEFT JOIN employee_global_roles r ON e.id = r.employee_id AND r.role = 'ROLE_DEPARTMENT_MANAGER' WHERE r.role IS NULL AND e.enabled = true LIMIT 5",
         "SETUP: Via test API — trigger CS sync\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for the employee\n"
         "4. Check the employee's roles displayed on their profile\n"
         "5. Verify DEPARTMENT_MANAGER role is present\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM ttt_backend.employee WHERE login = '<login>')",
         "After CS sync, DepartmentManagerRolePostProcessor assigns ROLE_DEPARTMENT_MANAGER to employees who are department heads in CS. "
         "Role visible in Admin employee profile.",
         "Critical", "Hybrid", "companystaff-integration.md §Post-Processors", "cross-service",
         "DepartmentManagerRolePostProcessor checks CS department head status. This test verifies role assignment on promotion."),

        ("TC-CS-007", "PM role demotion does NOT remove DM role (known bug)",
         "Employee with both ROLE_PROJECT_MANAGER and ROLE_DEPARTMENT_MANAGER. "
         "Query: SELECT e.login FROM ttt_backend.employee e "
         "JOIN employee_global_roles r1 ON e.id = r1.employee_id AND r1.role = 'ROLE_PROJECT_MANAGER' "
         "JOIN employee_global_roles r2 ON e.id = r2.employee_id AND r2.role = 'ROLE_DEPARTMENT_MANAGER' "
         "WHERE e.enabled = true LIMIT 3",
         "SETUP: Via test API — trigger CS sync (simulating PM demotion)\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Find the employee who was demoted from PM\n"
         "4. Check their roles list\n"
         "5. Verify ROLE_DEPARTMENT_MANAGER status\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM ttt_backend.employee WHERE login = '<login>')\n"
         "DB-CHECK: Verify whether ROLE_DEPARTMENT_MANAGER was incorrectly removed instead of ROLE_PROJECT_MANAGER",
         "KNOWN BUG: ProjectManagerRolePostProcessor removes ROLE_DEPARTMENT_MANAGER instead of ROLE_PROJECT_MANAGER on PM demotion. "
         "After sync, DM role may be incorrectly removed while PM role remains.",
         "Critical", "Hybrid", "companystaff-integration.md §Bug #1 (CRITICAL)", "cross-service",
         "Bug in ProjectManagerRolePostProcessor line 39: removes wrong enum value. Test documents current (buggy) behavior."),

        ("TC-CS-008", "Former HR retains ROLE_OFFICE_HR (known bug — no removal)",
         "Employee who is no longer HR in CompanyStaff but still has ROLE_OFFICE_HR in TTT. "
         "Query: SELECT e.login FROM ttt_backend.employee e "
         "JOIN employee_global_roles r ON e.id = r.employee_id AND r.role = 'ROLE_OFFICE_HR' "
         "WHERE e.enabled = true",
         "SETUP: Via test API — trigger CS office sync\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Find the former HR employee\n"
         "4. Check their roles — verify ROLE_OFFICE_HR is still present\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM ttt_backend.employee WHERE login = '<login>')",
         "KNOWN BUG: OfficeHRRolePostProcessor only adds ROLE_OFFICE_HR, never removes it. Former HR employees retain elevated permissions indefinitely.",
         "High", "Hybrid", "companystaff-integration.md §Bug #3", "cross-service",
         "OfficeHRRolePostProcessor has no removal logic. Needs manual role cleanup or code fix."),

        ("TC-CS-009", "CS sync failure — failed entity tracked and retried",
         "Test environment with CS sync functional. "
         "Query: SELECT * FROM ttt_backend.cs_sync_failed_entity ORDER BY id DESC LIMIT 5",
         "SETUP: Note current cs_sync_failed_entity count\n"
         "SETUP: Via test API — trigger CS employee sync\n"
         "1. Login as Admin (no direct UI for sync monitoring)\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.cs_sync_failed_entity\n"
         "DB-CHECK: SELECT entity_type, entity_id, error_message FROM ttt_backend.cs_sync_failed_entity ORDER BY id DESC LIMIT 10\n"
         "DB-CHECK: SELECT sync_type, last_success_date, status FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 5\n"
         "2. If failed entities exist, trigger sync again\n"
         "DB-CHECK: Verify previously failed entities are retried (count decreases or error changes)",
         "After sync, failed entities are recorded in cs_sync_failed_entity table with error details. Next sync cycle retries them. "
         "Successful sync recorded in cs_sync_status.",
         "High", "Hybrid", "#2496, #2381", "cross-service",
         "Ticket #2496: cron-based sync silently stopped. Ticket #2381: CS sync endpoint returning 500."),

        ("TC-CS-010", "CS sync incremental mode — only updated entities fetched",
         "Environment with recent successful CS sync. "
         "Query: SELECT sync_type, last_success_date FROM ttt_backend.cs_sync_status WHERE last_success_date IS NOT NULL ORDER BY last_success_date DESC LIMIT 3",
         "SETUP: Via test API — trigger CS sync, note the sync duration\n"
         "SETUP: Immediately trigger sync again\n"
         "DB-CHECK: Compare cs_sync_status.last_success_date before and after\n"
         "DB-CHECK: Check that second sync was faster (fewer entities fetched — incremental uses updatedAfter param)\n"
         "1. Login as Admin\n"
         "2. No direct UI verification — this is infrastructure behavior\n"
         "DB-CHECK: Verify cs_sync_status records updated_after timestamp matching previous sync success",
         "Incremental sync only fetches entities updated since last successful sync. Second sync should be significantly faster. "
         "cs_sync_status.last_success_date updated after each successful run.",
         "Medium", "Hybrid", "#2989, #3303", "cross-service",
         "Full sync scheduler is COMMENTED OUT — no automatic full re-sync. Delta sync may miss changes (#2989)."),

        ("TC-CS-011", "Feature toggle cs-sync OFF disables sync silently",
         "cs-sync feature toggle accessible. "
         "Query: SELECT * FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 3",
         "SETUP: Note current cs_sync_status timestamps\n"
         "SETUP: If possible, set cs-sync toggle to OFF (requires Unleash access)\n"
         "SETUP: Via test API — trigger CS sync\n"
         "DB-CHECK: Verify cs_sync_status.last_success_date did NOT change (sync was skipped)\n"
         "1. Login as Admin — verify no error messages or warnings in UI\n"
         "2. Check application logs for any indication that sync was skipped\n"
         "CLEANUP: Re-enable cs-sync toggle",
         "With cs-sync toggle OFF, sync trigger does nothing silently. No warning in logs or UI. "
         "cs_sync_status timestamps remain unchanged. This is a known design concern.",
         "Medium", "Hybrid", "feature-toggles-unleash.md, design issue #141", "cross-service",
         "Feature toggle gate: CS_SYNC-{env}. If disabled, CSSyncLauncherImpl.sync() silently returns null."),

        ("TC-CS-012", "Contractor sync only in TTT backend — absent from vacation service",
         "Contractor employee in the system. "
         "Query: SELECT login FROM ttt_backend.employee WHERE contractor = true AND enabled = true LIMIT 3",
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for a contractor — verify they appear in the list\n"
         "4. Navigate to Calendar of absences\n"
         "5. Search for the same contractor — verify they do NOT appear in absence calendar\n"
         "DB-CHECK: SELECT login, contractor FROM ttt_backend.employee WHERE login = '<contractor_login>'\n"
         "DB-CHECK: SELECT login FROM ttt_vacation.employee WHERE login = '<contractor_login>' — expect NULL or contractor=false",
         "Contractor appears in Admin employee list (TTT backend) but NOT in Calendar of absences (vacation service). "
         "Vacation service does not sync contractors — they are excluded from all active employee queries.",
         "High", "Hybrid", "companystaff-integration.md §Vacation Service Differences, exploration_finding #200", "cross-service",
         "Vacation CSEmployeeSynchronizer sets contractor=false for all synced employees. No CSContractorSynchronizer in vacation service."),
    ],

    "TS-CrossService-EventProp": [
        ("TC-CS-013", "Calendar change triggers vacation recalculation via RabbitMQ",
         "Admin user with calendar editing permissions. Office with existing vacations spanning a date that will be changed. "
         "Query: SELECT o.id, o.name FROM ttt_backend.office o LIMIT 5",
         "SETUP: Note a vacation's working days count for an employee in the target office\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Production calendars\n"
         "3. Select an office's production calendar\n"
         "4. Change a working day to a non-working day (e.g., add a holiday)\n"
         "5. Save the calendar change\n"
         "6. Wait 10-15 seconds for RabbitMQ event propagation\n"
         "7. Navigate to Calendar of absences\n"
         "8. Find an employee with a vacation spanning the changed date\n"
         "9. Verify the vacation's working days count has been recalculated\n"
         "DB-CHECK: SELECT id, working_days, admin_days FROM ttt_vacation.vacation WHERE employee_id = <id> AND start_date <= '<changed_date>' AND end_date >= '<changed_date>'",
         "Calendar change published as CalendarChangedEvent via RabbitMQ (ttt.calendar.topic). Vacation service receives and recalculates: "
         "working day→non-working reduces working_days count. Vacation visible with updated duration in UI.",
         "Critical", "Hybrid", "cross-service-integration.md §Calendar-Vacation Interaction, #2364", "cross-service",
         "CalendarChangedEventHandler → CalendarUpdateProcessor.process() → VacationCalendarUpdateService.recalculateVacations(). "
         "Also triggers month norm recalculation via EmployeeMonthNormContextCalculatedEvent."),

        ("TC-CS-014", "Calendar day deleted triggers day-off cascade deletion",
         "Admin with calendar permissions. Office with day-offs on a specific calendar day. "
         "Query: SELECT o.id, o.name FROM ttt_backend.office o WHERE id IN (SELECT DISTINCT office_id FROM ttt_vacation.employee_day_off) LIMIT 3",
         "SETUP: Via API — create a day-off for an employee on a date that is a non-working day in their office calendar\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Production calendars\n"
         "3. Select the office's production calendar\n"
         "4. Delete the non-working day entry (making it a regular working day)\n"
         "5. Save the change\n"
         "6. Wait 10-15 seconds for event propagation\n"
         "7. Navigate to Calendar of absences\n"
         "8. Check the employee's day-off — verify it was deleted or status changed\n"
         "DB-CHECK: SELECT status FROM ttt_vacation.employee_day_off_request WHERE employee_id = <id> AND date = '<deleted_date>'",
         "Calendar day deletion publishes CalendarDeletedEvent. Vacation service deletes NEW/APPROVED day-offs for the deleted date. "
         "Employee notified via EmployeeDayOffDeletedFromCalendarSendNotificationEventHandler.",
         "High", "Hybrid", "cross-service-integration.md §Calendar Deleted → Day-Off Cascade", "cross-service",
         "CalendarDeletedEventHandler → EmployeeDayOffCalendarUpdateService.deleteDayOffs(). Parent vacation recalculated."),

        ("TC-CS-015", "Period advance triggers vacation payment marking",
         "Accountant user. Office with APPROVED vacations in a period about to be advanced. "
         "Query: SELECT e.login FROM ttt_backend.employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_ACCOUNTANT' AND e.enabled = true LIMIT 1",
         "SETUP: Ensure there is an APPROVED vacation in the current approve period for the target office\n"
         "1. Login as Accountant\n"
         "2. Navigate to Accounting > Periods\n"
         "3. Select the target office\n"
         "4. Advance the approval period by one month\n"
         "5. Wait 10-15 seconds for RabbitMQ event propagation\n"
         "6. Navigate to Accounting > Vacation payment\n"
         "7. Verify the APPROVED vacation now appears as ready for payment\n"
         "DB-CHECK: SELECT status FROM ttt_vacation.vacation_status_updates WHERE office_id = <office_id> AND type = 'NEW_FOR_PAID'",
         "Period advance publishes PeriodChangedEvent (APPROVE type). Vacation service recalculates available days, rejects out-of-date day-offs, "
         "and inserts VacationStatusUpdate(NEW_FOR_PAID) marking vacations for payment cron.",
         "Critical", "Hybrid", "cross-service-integration.md §Period Changed → Vacation Impact", "cross-service",
         "PeriodChangedEventHandler → AvailableDaysRecalculationService.recalculate(). Only processes APPROVE period type."),

        ("TC-CS-016", "Period reopen reverses vacation day recalculation",
         "Accountant user. Office where period was recently advanced. "
         "Same preconditions as TC-CS-015 but starting from advanced state.",
         "SETUP: Ensure period was recently advanced (TC-CS-015 must have been executed)\n"
         "1. Login as Accountant\n"
         "2. Navigate to Accounting > Periods\n"
         "3. Select the target office\n"
         "4. Reopen (revert) the previously advanced period\n"
         "5. Wait 10-15 seconds for event propagation\n"
         "6. Navigate to Calendar of absences\n"
         "7. Verify vacation day balances are restored to pre-advance values\n"
         "DB-CHECK: Verify vacation available_days reverted to previous value",
         "Period reopen publishes PeriodReopenedEvent. Vacation service calls AvailableDaysRecalculationService.recalculationReverse() "
         "to undo the vacation day changes made during period advance.",
         "High", "Hybrid", "cross-service-integration.md §Period Reopened", "cross-service",
         "PeriodReopenedEventHandler only processes APPROVE period type."),

        ("TC-CS-017", "Employee change event triggers cache invalidation",
         "Employee whose data was recently modified. "
         "Query: SELECT login FROM ttt_backend.employee WHERE enabled = true ORDER BY random() LIMIT 1",
         "SETUP: Via test API — trigger CS sync to simulate employee data change\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Find the employee and note their current details\n"
         "4. Verify employee data is fresh (not stale cached version)\n"
         "5. Navigate to Calendar of absences\n"
         "6. Search for the same employee — verify data consistency\n"
         "DB-CHECK: Check ttt_vacation.employee cache state (EmployeeCachePostProcessor should have invalidated)",
         "EmployeeChangedEvent published via RabbitMQ. Vacation service evicts employee cache by login. "
         "TTT backend post-processors invalidate office and employee caches.",
         "Medium", "Hybrid", "cross-service-integration.md §Event Types", "cross-service",
         "EmployeeChangedEventHandler in vacation service calls cache.evict(login). TTT uses EmployeeCachePostProcessor."),

        ("TC-CS-018", "Month norm recalculation propagated from vacation to TTT backend",
         "Calendar change or vacation recalculation triggers norm update. "
         "Query: SELECT e.login FROM ttt_backend.employee e WHERE e.enabled = true AND e.salary_office IS NOT NULL ORDER BY random() LIMIT 1",
         "SETUP: Trigger a calendar change (as in TC-CS-013)\n"
         "1. Login as employee affected by the calendar change\n"
         "2. Navigate to My tasks page\n"
         "3. Check the monthly working hours norm displayed\n"
         "4. Navigate to Statistics page\n"
         "5. Verify the norm matches the updated production calendar\n"
         "DB-CHECK: Check statistic_report table for updated norm values after the cascade event",
         "After calendar change, vacation service publishes EmployeeMonthNormContextCalculatedEvent back to TTT backend. "
         "TTT receives it via EmployeeMonthNormContextCalculatedEventHandler → StatisticReportSyncService.saveMonthNormAndReportedEffortForEmployees(). "
         "Statistics norm updated.",
         "High", "Hybrid", "cross-service-integration.md §Calendar Changed, rabbitmq-messaging.md", "cross-service",
         "Bidirectional event flow: TTT→Vacation (employee changes) and Vacation→TTT (month norm context). Uses ttt.backend.employee.topic exchange."),

        ("TC-CS-019", "Email sent via async RabbitMQ path after vacation creation",
         "Employee with a manager (for approval notification). "
         "Query: SELECT e.login FROM ttt_vacation.employee e WHERE e.working = true AND e.manager_id IS NOT NULL AND e.maternity = false ORDER BY random() LIMIT 1",
         "SETUP: Via API — create a vacation for the employee (POST /api/vacation/v1/vacations)\n"
         "1. Wait 15-30 seconds for email processing\n"
         "DB-CHECK: SELECT id, subject, status, to_field, add_time FROM ttt_email.email WHERE subject LIKE '%<employee_name>%' ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify email status is SENT (not NEW or FAILED)\n"
         "DB-CHECK: Verify the email was sent to the manager's email address\n"
         "CLEANUP: Via API — delete the created vacation",
         "Vacation creation triggers notification email to manager. With EMAIL_ASYNC toggle ON, email published via RabbitMQ → email service → batch SMTP delivery. "
         "Email status should progress: NEW → SENT. Manager receives approval request notification.",
         "High", "Hybrid", "email-notification-deep-dive.md §Async Email Path, #2518", "cross-service",
         "Async path: Spring ApplicationEvent → RabbitMQ (ttt.email.topic) → SendEmailEventHandler → EmailService.save() → EmailSendScheduler → SMTP."),

        ("TC-CS-020", "Feature toggle email-async OFF routes email via sync Feign call",
         "EMAIL_ASYNC toggle can be verified via API. "
         "Query: GET /api/ttt/v1/feature-toggles/email-async",
         "1. Via API — check current EMAIL_ASYNC toggle state: GET /api/ttt/v1/feature-toggles/email-async\n"
         "2. Note whether it's ON (async/RabbitMQ) or OFF (sync/Feign)\n"
         "3. Via API — create a vacation to trigger notification: POST /api/vacation/v1/vacations\n"
         "4. Wait 15-30 seconds\n"
         "DB-CHECK: SELECT id, status, add_time FROM ttt_email.email ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify email arrived in DB regardless of toggle state\n"
         "CLEANUP: Via API — delete the vacation",
         "Email arrives in email service DB via either path. With async OFF, Feign EmailClient.send() is used directly (synchronous). "
         "Email status and delivery time may differ between paths.",
         "Medium", "Hybrid", "feature-toggles-unleash.md, email-notification-deep-dive.md §Feature Toggle EMAIL_ASYNC", "cross-service",
         "TTT InternalEmailService swallows all exceptions and returns false in sync path — callers cannot distinguish failure types."),

        ("TC-CS-021", "System clock change broadcast via ttt.fanout to all services",
         "Timemachine environment with clock manipulation API. "
         "Only available on timemachine env: PATCH /api/ttt/test/v1/clock",
         "SETUP: Note current time in all services\n"
         "1. Via test API — advance system clock by 1 day: PATCH /api/ttt/test/v1/clock\n"
         "2. Login to the application\n"
         "3. Navigate to My Tasks page\n"
         "4. Verify the displayed date matches the advanced clock\n"
         "5. Navigate to Calendar of absences\n"
         "6. Verify the calendar also shows the advanced date\n"
         "7. Navigate to Accounting > Periods\n"
         "8. Verify period dates reflect the clock change\n"
         "CLEANUP: Via test API — reset clock to current time",
         "SystemClockChangedEvent broadcast via ttt.fanout exchange to ALL services (TTT, Vacation, Email, Calendar). "
         "All 4 services update their internal clock. UI shows consistent advanced date across all pages.",
         "High", "Hybrid", "cross-service-integration.md §SystemClockChangedEvent, rabbitmq-messaging.md", "cross-service",
         "Timemachine-only test. ttt.fanout is the only fanout exchange — all other exchanges are topic-based."),

        ("TC-CS-022", "Employee period change event processed by vacation service",
         "Employee with reports in the current period. "
         "Query: SELECT DISTINCT e.login FROM ttt_backend.task_report tr JOIN ttt_backend.employee e ON tr.employee_id = e.id WHERE tr.report_date >= DATE_TRUNC('month', CURRENT_DATE) LIMIT 3",
         "SETUP: Via test API — trigger an employee period change event\n"
         "1. Login as the affected employee\n"
         "2. Navigate to My Tasks page\n"
         "3. Verify report period dates displayed correctly\n"
         "4. Navigate to Calendar of absences\n"
         "5. Check that vacation period aligns with the new employee period\n"
         "DB-CHECK: Check ttt_vacation tables for period-related updates for the employee",
         "EmployeePeriodChangedEvent published via ttt.backend.employeePeriod.topic. Vacation service processes it via EmployeePeriodChangedListener. "
         "Employee-specific period updates reflected in vacation calculations.",
         "Medium", "Hybrid", "cross-service-integration.md §Event Types", "cross-service",
         "Employee period events are distinct from office period events. Published per-employee, not per-office."),
    ],

    "TS-CrossService-DataIntegrity": [
        ("TC-CS-023", "Office assignment consistent between TTT backend and vacation service",
         "Random sample of employees from both services. "
         "Query: SELECT b.login, b.salary_office as backend_office, v.office_id as vacation_office FROM ttt_backend.employee b "
         "JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.enabled = true AND b.salary_office IS NOT NULL AND v.office_id IS NOT NULL LIMIT 20",
         "DB-CHECK: Run the comparison query above for a sample of 20 employees\n"
         "DB-CHECK: Count mismatches: SELECT COUNT(*) FROM ttt_backend.employee b JOIN ttt_vacation.employee v ON b.login = v.login "
         "WHERE b.salary_office != v.office_id AND b.enabled = true\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Pick 3 employees from the mismatch list\n"
         "4. For each, note the office shown in Admin\n"
         "5. Navigate to Calendar of absences, find same employees\n"
         "6. Compare the office context for vacation calculations\n"
         "7. Document all mismatches found",
         "KNOWN BUG: ~62% of employees have different office assignments between services (736/1190 on all envs). "
         "Test documents the current divergence rate and specific examples. Mismatched employees may get wrong vacation day calculations.",
         "Critical", "Hybrid", "cross-service-office-sync-divergence.md, #2876", "cross-service",
         "Root cause: independent CS sync paths with no cross-service reconciliation. Ticket #2876 explicitly documents this."),

        ("TC-CS-024", "Employee yearly office record matches vacation service office",
         "Employee with yearly office records. "
         "Query: SELECT e.login, e.office_id, eo.office as yearly_office, eo.year FROM ttt_vacation.employee e "
         "JOIN ttt_vacation.employee_office eo ON e.id = eo.employee_id WHERE eo.year = EXTRACT(YEAR FROM CURRENT_DATE) LIMIT 10",
         "DB-CHECK: Compare employee.office_id with employee_office[current_year].office for 10 employees\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_vacation.employee e JOIN ttt_vacation.employee_office eo ON e.id = eo.employee_id "
         "WHERE eo.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.office_id != eo.office\n"
         "1. Login as Admin\n"
         "2. Navigate to Statistics or Calendar of absences\n"
         "3. Check that employee's displayed office is consistent\n"
         "4. Document any internal divergence within vacation service",
         "Within the vacation service, employee.office_id should match employee_office[year].office for the current year. "
         "This is expected to be consistent (divergence is between services, not within).",
         "High", "Hybrid", "cross-service-office-sync-divergence.md §Year Records Consistent", "cross-service",
         "EmployeeOfficeChangedProcessor has conditional logic that can skip mid-year updates when calendars differ."),

        ("TC-CS-025", "last_date field synced between backend and vacation (#3374)",
         "Employees where last_date might differ between services. "
         "Query: SELECT b.login, b.last_date as backend_last, v.last_date as vacation_last FROM ttt_backend.employee b "
         "JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.last_date IS NOT NULL AND (v.last_date IS NULL OR b.last_date != v.last_date) LIMIT 10",
         "DB-CHECK: Run the comparison query above\n"
         "DB-CHECK: Count mismatches specifically for last_date\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Find a leaving employee (has last_date set)\n"
         "4. Verify their end date is displayed correctly\n"
         "5. Navigate to Calendar of absences\n"
         "6. Check the same employee's availability (should show leaving date)\n"
         "7. Document any discrepancy",
         "KNOWN BUG (#3374): last_date not updated during CS sync. Employees with leaving dates in ttt_backend may have NULL in ttt_vacation. "
         "This causes vacation service to be unaware of employee departure, potentially allowing vacation creation past employment end date.",
         "High", "Hybrid", "#3374, exploration_finding #240", "cross-service",
         "7 employees with mismatches found. 3 critical cases where backend has last_date but vacation is NULL."),

        ("TC-CS-026", "Contractor excluded from vacation service employee queries",
         "Contractor employees in both services. "
         "Query: SELECT b.login, b.contractor as backend_contractor, v.contractor as vacation_contractor FROM ttt_backend.employee b "
         "LEFT JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.contractor = true AND b.enabled = true LIMIT 10",
         "DB-CHECK: Run query above — verify contractors either absent from ttt_vacation or have contractor=false\n"
         "1. Login as employee with access to Calendar of absences\n"
         "2. Navigate to Calendar of absences\n"
         "3. Search for a known contractor by name\n"
         "4. Verify they do NOT appear in the absence calendar\n"
         "5. Navigate to Statistics\n"
         "6. Verify contractor is excluded from statistics views",
         "Contractors exist in ttt_backend but are excluded from ttt_vacation active queries. Vacation service sets contractor=false for all synced employees. "
         "No CSContractorSynchronizer exists in vacation service.",
         "Medium", "Hybrid", "companystaff-integration.md §Vacation Service Differences, exploration_finding #200", "cross-service",
         "beingDismissed field NOT set during contractor sync — only employee sync sets it. May cause edge cases with readOnly/dismissal flow."),

        ("TC-CS-027", "Statistics norm uses correct office calendar despite divergence",
         "Employee with different offices in backend vs vacation. "
         "Query: SELECT b.login, b.salary_office, v.office_id FROM ttt_backend.employee b "
         "JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.salary_office != v.office_id AND b.enabled = true LIMIT 5",
         "1. Login as the mismatched employee (or Admin viewing their stats)\n"
         "2. Navigate to Statistics page\n"
         "3. Check the monthly working hours norm displayed\n"
         "4. Navigate to Calendar of absences\n"
         "5. Check the vacation day balance\n"
         "DB-CHECK: Determine which office's production calendar is used for norm: ttt_backend.salary_office\n"
         "DB-CHECK: Determine which office's annual_leave is used for vacation days: ttt_vacation.office_id\n"
         "6. Document which service's office assignment drives each calculation",
         "For employees with divergent offices: statistics norm uses ttt_backend.salary_office calendar, while vacation day accrual uses ttt_vacation.office_id. "
         "These may yield inconsistent results if the offices have different production calendars.",
         "Critical", "Hybrid", "cross-service-office-sync-divergence.md §Impact Assessment", "cross-service",
         "Different offices may have different production calendars (Russian vs non-Russian). Employee sees office A's norm in reports but gets office B's vacation days."),

        ("TC-CS-028", "Null office handling during CS sync",
         "Employees with NULL salary office. "
         "Query: SELECT login FROM ttt_backend.employee WHERE salary_office IS NULL AND enabled = true LIMIT 5",
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.employee WHERE salary_office IS NULL AND enabled = true\n"
         "DB-CHECK: SELECT login, salary_office FROM ttt_backend.employee WHERE salary_office IS NULL AND enabled = true LIMIT 10\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for an employee with NULL office\n"
         "4. Verify their office field display (should show 'Not specified' / 'Не указано')\n"
         "5. Navigate to Statistics\n"
         "6. Check whether these employees are included in statistics calculations",
         "When CompanyStaff returns no accounting data, salaryOfficeId is set to NULL. Office factory creates stubs with empty names. "
         "Employees with NULL office may be excluded from or cause errors in calculations.",
         "Medium", "Hybrid", "cross-service-office-sync-divergence.md §Null Office Handling", "cross-service",
         "SalaryOfficeFactory.getOrCreate() creates offices with empty name (StringUtils.EMPTY) — never backfilled."),

        ("TC-CS-029", "Office 'Not specified' (id=9) divergence between services",
         "Employees with 'Not specified' office in one service but real office in the other. "
         "Query: SELECT b.login, b.salary_office, v.office_id FROM ttt_backend.employee b "
         "JOIN ttt_vacation.employee v ON b.login = v.login WHERE (b.salary_office = 9 AND v.office_id != 9) OR (b.salary_office != 9 AND v.office_id = 9) LIMIT 10",
         "DB-CHECK: Run the query above\n"
         "DB-CHECK: Count: SELECT COUNT(*) ... with the same WHERE clause\n"
         "1. Login as Admin\n"
         "2. For 3 employees from the result, check their office in Admin > Employees\n"
         "3. Navigate to Calendar of absences and compare\n"
         "4. Document the office shown in each view",
         "108 employees have 'Not specified' (id=9) in one service but a real office in the other (99 backend=Not specified, 9 vacation=Not specified). "
         "These employees are in a particularly broken state — one service doesn't know their office.",
         "High", "Hybrid", "cross-service-office-sync-divergence.md §Mismatch Categories", "cross-service",
         "Backend 'Not specified' typically means CS returned no accounting data. Vacation may have gotten it from a different sync cycle."),

        ("TC-CS-030", "Terminated employees retain office assignments in both services",
         "Terminated (not working) employees with office data still present. "
         "Query: SELECT b.login, b.salary_office, v.office_id FROM ttt_backend.employee b "
         "JOIN ttt_vacation.employee v ON b.login = v.login WHERE b.enabled = false AND (b.salary_office IS NOT NULL OR v.office_id IS NOT NULL) LIMIT 10",
         "DB-CHECK: Run query above\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.employee WHERE enabled = false AND salary_office IS NOT NULL\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_vacation.employee_office eo JOIN ttt_vacation.employee e ON eo.employee_id = e.id WHERE e.working = false\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Search for a terminated employee — verify they don't appear in active list\n"
         "4. Document whether their historical office data is retained",
         "Terminated employees retain office assignments in both services. No cleanup listener for deactivation events removes employee_office records. "
         "This is informational — may affect aggregated statistics if terminated employees are counted.",
         "Low", "Hybrid", "cross-service-office-sync-divergence.md §No Termination Cleanup", "cross-service",
         "No listener for deactivation events cleans up employee_office records. Historical data preserved."),
    ],

    "TS-CrossService-WebSocket": [
        ("TC-CS-031", "Task rename in Planner propagates to all connected clients",
         "Two browser sessions logged in as users on the same project. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' AND p.report = true ORDER BY random() LIMIT 1",
         "1. Open browser session 1 — login as PM of the project\n"
         "2. Navigate to Planner for the project\n"
         "3. Open browser session 2 — login as another team member on the same project\n"
         "4. Navigate to Planner for the same project in session 2\n"
         "5. In session 1, click on a task name to rename it\n"
         "6. Enter a new task name and save\n"
         "7. In session 2 (without refreshing), verify the task name updates automatically\n"
         "8. Navigate to My Tasks in session 2 — verify the renamed task shows the new name",
         "Task rename publishes TASK_RENAME event to 3 STOMP topics: /topic/projects/{id}/tasks, /topic/employees/{login}/reports, "
         "/topic/employees/{login}/assignments. All connected clients see the update without page refresh.",
         "High", "UI", "websocket-events.md §Task rename, cross-service-integration.md §WebSocket", "cross-service",
         "WsTaskEventListener cascades task rename to report and assignment channels. Extracts sub-events and re-publishes."),

        ("TC-CS-032", "Cell lock in Planner visible to other users",
         "Two users on the same project with Planner access. Same setup as TC-CS-031.",
         "1. Open two browser sessions on the same project's Planner\n"
         "2. In session 1, click on a report cell to start editing (lock it)\n"
         "3. In session 2, verify the cell shows as locked (visual indicator — e.g., colored border or lock icon)\n"
         "4. In session 1, finish editing and move to another cell (unlock)\n"
         "5. In session 2, verify the lock indicator disappears\n"
         "6. Try to click the same cell in session 2 — verify it's now editable",
         "LOCK event sent to /topic/employees/{login}/locks when cell is clicked. UNLOCK sent when editing ends. "
         "Other users see real-time lock/unlock status. Prevents concurrent editing conflicts.",
         "High", "UI", "websocket-events.md §Lock/Unlock", "cross-service",
         "Stale lock risk on disconnect — no heartbeat-based cleanup found. If user closes browser without unlocking, lock may persist."),

        ("TC-CS-033", "Stale lock persists after user disconnects without unlocking",
         "Two users on the same project Planner. Session 1 user will disconnect abruptly.",
         "1. Open two browser sessions on the same project's Planner\n"
         "2. In session 1, click on a report cell (lock it)\n"
         "3. In session 2, verify the cell is locked\n"
         "4. Close session 1's browser tab abruptly (without unlocking the cell)\n"
         "5. In session 2, check if the lock indicator persists\n"
         "6. Wait 30-60 seconds\n"
         "7. Check again — does the stale lock eventually clear?\n"
         "8. Try to edit the locked cell in session 2",
         "KNOWN ISSUE: No heartbeat-based lock cleanup mechanism found in code. After abrupt disconnect, the lock may persist until the next "
         "session or server-side timeout. Document observed behavior.",
         "Medium", "UI", "websocket-events.md §Test Implications", "cross-service",
         "No WebSocket reconnection awareness. Server doesn't track client connection state."),

        ("TC-CS-034", "WebSocket events for report CRUD in Planner",
         "Employee with Planner access and tasks in current period. "
         "Query: SELECT e.login FROM ttt_backend.employee e WHERE e.enabled = true AND e.deactivated = false ORDER BY random() LIMIT 1",
         "1. Open two browser sessions: same employee (different tabs) or PM monitoring + employee editing\n"
         "2. In session 1, navigate to Planner\n"
         "3. In session 2, navigate to the same Planner view\n"
         "4. In session 1, add a report entry (hours for a task on a specific day)\n"
         "5. In session 2, verify the report entry appears without refreshing\n"
         "6. In session 1, edit the report entry (change hours)\n"
         "7. In session 2, verify the change is reflected\n"
         "8. In session 1, delete the report entry\n"
         "9. In session 2, verify the deletion is reflected",
         "Report CRUD operations (ADD, PATCH, DELETE) publish events to /topic/employees/{login}/reports/{period}. "
         "All connected sessions see updates in real time. Each event carries EventType, emitterLogin, timestamp, and payload.",
         "High", "UI", "websocket-events.md §WsTaskReportEventListener", "cross-service",
         "Uses @TransactionalEventListener — fires after transaction commit. Thread pool handles async delivery."),

        ("TC-CS-035", "Assignment generation event triggers UI update in Planner",
         "PM with access to generate assignments. Project with tasks. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' AND p.report = true ORDER BY random() LIMIT 1",
         "1. Login as PM of the project\n"
         "2. Navigate to Planner for the project\n"
         "3. Open another browser tab with the same Planner view\n"
         "4. In tab 1, trigger 'Generate assignments' action\n"
         "5. In tab 2, verify GENERATE event is received (assignments appear without refresh)\n"
         "6. Verify TASK_REFRESH_START and TASK_REFRESH_FINISH events bracket the operation",
         "GENERATE event published to /topic/employees/{assigneeLogin}/assignments/{period}. "
         "TASK_REFRESH_START/FINISH events bracket the generation operation on /topic/projects/{id}/tasks.",
         "Medium", "UI", "websocket-events.md §Event Types", "cross-service",
         "Frontend state management bugs (#3332, #3314) may interfere with assignment display after generation."),

        ("TC-CS-036", "Tracker sync events shown in Planner",
         "Project with an external tracker configured (Jira, ClickUp, etc.). "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.tracker_type IS NOT NULL AND p.status = 'ACTIVE' LIMIT 3",
         "1. Login as PM of the project with a tracker\n"
         "2. Navigate to Planner for the project\n"
         "3. Trigger tracker sync (usually via a sync button in Planner)\n"
         "4. Verify TRACKER_SYNC_START event is received (UI shows sync indicator)\n"
         "5. Wait for sync to complete\n"
         "6. Verify TRACKER_SYNC_FINISH event is received (sync indicator disappears)\n"
         "7. Verify imported tasks appear in the Planner task list",
         "TRACKER_SYNC_START and TRACKER_SYNC_FINISH events published to /topic/projects/{id}/tracker-work-log. "
         "UI shows a loading/sync indicator during the operation. Imported tasks appear after completion.",
         "Medium", "UI", "websocket-events.md §Tracker Sync", "cross-service",
         "WsTrackerSyncEventListener uses @EventListener (not @TransactionalEventListener) — fires immediately."),

        ("TC-CS-037", "Project member CRUD events propagated via WebSocket",
         "PM with permission to add/remove project members. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
         "1. Login as PM of the project\n"
         "2. Navigate to Planner for the project\n"
         "3. Open another browser session on the same project\n"
         "4. In session 1, add a new member to the project\n"
         "5. In session 2, verify the new member appears without refresh\n"
         "6. In session 1, modify the member's role\n"
         "7. In session 2, verify the role change is reflected\n"
         "8. In session 1, remove the member\n"
         "9. In session 2, verify the member disappears",
         "Member CRUD events (ADD, PATCH, DELETE) published to /topic/projects/{id}/members. "
         "WsProjectMemberEventListener handles all member changes. Connected clients see real-time updates.",
         "Medium", "UI", "websocket-events.md §WsProjectMemberEventListener", "cross-service",
         "Uses @TransactionalEventListener — fires after DB transaction commits."),

        ("TC-CS-038", "Cell selection cursor visible to other users",
         "Two users editing the same project's Planner simultaneously.",
         "1. Open two browser sessions on the same project's Planner\n"
         "2. In session 1, click on a cell (select it)\n"
         "3. In session 2, check if a cursor/highlight indicator shows for session 1's selection\n"
         "4. In session 1, move to a different cell\n"
         "5. In session 2, verify the cursor indicator moves accordingly\n"
         "6. Document observed behavior (cursor visibility may be limited)",
         "SELECT event sent to /topic/employees/{login}/selections when a cell is clicked. Other users may see a cursor indicator "
         "showing which cell the other user has selected. Actual UI implementation may vary.",
         "Low", "UI", "websocket-events.md §SELECT event", "cross-service",
         "SELECT events are UI state only — not persisted. Used for collaborative awareness in real-time editing."),
    ],

    "TS-CrossService-TrackerSync": [
        ("TC-CS-039", "Jira Cloud tracker configured and tasks imported",
         "Project with Jira Cloud tracker. Admin/PM access. "
         "Query: SELECT p.id, p.name, p.tracker_type, p.tracker_url FROM ttt_backend.project p WHERE p.tracker_type = 'JIRA' AND p.status = 'ACTIVE' LIMIT 3",
         "1. Login as Admin or PM of the project\n"
         "2. Navigate to Admin > Projects\n"
         "3. Find the project with Jira integration\n"
         "4. Open the project settings — verify tracker configuration (URL, credentials)\n"
         "5. Navigate to Planner for the project\n"
         "6. Trigger task sync from Jira\n"
         "7. Verify tasks from Jira appear in the Planner task list\n"
         "8. Verify task names match Jira issue summaries\n"
         "9. Verify assignees are mapped to TTT employees",
         "Jira Cloud integration imports tasks via REST API. Tasks appear in Planner with Jira issue key as reference. "
         "Assignees mapped to TTT employees by login/email.",
         "High", "UI", "#2072, admin-ticket-findings.md", "cross-service",
         "REST API must not be blocked on Jira side (#2072). Tracker must be configured by project SPM."),

        ("TC-CS-040", "Jira PAT authentication — verify error handling (#2511)",
         "Project with Jira Server tracker using PAT auth. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.tracker_type = 'JIRA' AND p.tracker_url LIKE '%server%' LIMIT 3",
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Find a project with Jira Server integration\n"
         "4. Open tracker settings\n"
         "5. Attempt to configure PAT authentication\n"
         "6. Verify whether PAT auth works or shows 'wrong server' error\n"
         "7. If error, document the exact error message and behavior\n"
         "8. Test with login/password auth as fallback",
         "KNOWN BUG (#2511): Jira PAT auth broken — 'wrong server' error. Blocks Seagate projects. "
         "Login/password auth may still work as fallback. Document current behavior.",
         "High", "UI", "#2511, #2571", "cross-service",
         "Need PAT auth support for Jira Server (#2571). Current workaround: use login/password."),

        ("TC-CS-041", "ClickUp tracker configured and tasks imported",
         "Project with ClickUp tracker. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.tracker_type = 'CLICKUP' AND p.status = 'ACTIVE' LIMIT 3",
         "1. Login as PM of the project\n"
         "2. Navigate to Planner for the project with ClickUp integration\n"
         "3. Trigger task sync from ClickUp\n"
         "4. Verify tasks from ClickUp appear in the Planner\n"
         "5. Verify task hierarchy is preserved (spaces/folders/lists → tasks)\n"
         "6. Check for any error notifications during sync\n"
         "7. Verify task assignees are mapped correctly",
         "ClickUp integration imports tasks via API. Tasks mapped from ClickUp spaces/lists to TTT Planner. "
         "Space-specific tasks should be included (#3341).",
         "High", "UI", "#2397, #3341", "cross-service",
         "Ticket #3148: ClickUp export was completely broken on prod. #3237: API domain change required. #3341: space-specific tasks missing."),

        ("TC-CS-042", "YouTrack tracker configured and tasks imported",
         "Project with YouTrack tracker. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.tracker_type = 'YOUTRACK' AND p.status = 'ACTIVE' LIMIT 3",
         "1. Login as PM of the project\n"
         "2. Navigate to Planner for the project with YouTrack integration\n"
         "3. Trigger task sync from YouTrack\n"
         "4. Verify tasks from YouTrack appear in the Planner\n"
         "5. Verify task IDs and names match YouTrack issues\n"
         "6. Check sync status indicators",
         "YouTrack integration added in Sprint 14 (#3145). Tasks imported via REST API. "
         "18 comments on the implementation ticket indicate significant complexity.",
         "Medium", "UI", "#3145", "cross-service",
         "YouTrack integration is relatively new — may have undiscovered edge cases."),

        ("TC-CS-043", "Multiple projects with same tracker cause task search issues (#3198)",
         "Two projects configured with the same tracker instance. "
         "Query: SELECT tracker_url, COUNT(*) as cnt FROM ttt_backend.project WHERE tracker_url IS NOT NULL GROUP BY tracker_url HAVING COUNT(*) > 1",
         "1. Login as PM\n"
         "2. Navigate to Planner for project A (uses tracker instance X)\n"
         "3. Sync tasks\n"
         "4. Navigate to Planner for project B (uses same tracker instance X)\n"
         "5. Sync tasks\n"
         "6. Verify tasks are correctly separated — project A shows only its tasks, project B shows only its tasks\n"
         "7. Search for a task by name — verify it shows in the correct project only",
         "KNOWN BUG (#3198): Multiple projects with same tracker → task search breaks. Tasks may appear in wrong project or "
         "search returns results from all projects sharing the tracker.",
         "High", "UI", "#3198", "cross-service",
         "Open ticket. Impact: task assignments could go to wrong project, time reporting misattributed."),

        ("TC-CS-044", "Approved status not cleared after tracker import (#3296)",
         "Project with tracker where reports were previously approved, then new tasks imported. "
         "Query: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.tracker_type IS NOT NULL AND p.status = 'ACTIVE' LIMIT 3",
         "1. Login as PM\n"
         "2. Navigate to Confirmation page\n"
         "3. Approve some reports for the project\n"
         "4. Navigate to Planner\n"
         "5. Trigger tracker sync (import new tasks)\n"
         "6. Navigate back to Confirmation page\n"
         "7. Check if the previously approved reports still show as approved\n"
         "8. Verify no stale approval status on newly imported tasks",
         "KNOWN BUG (#3296): Approved status not cleared after tracker import. Stale approvals may remain on reports "
         "that should have been reset after task structure changed.",
         "High", "UI", "#3296", "cross-service",
         "Open ticket. Stale approvals can cause incorrect confirmation state."),

        ("TC-CS-045", "Invalid tracker credentials show clear error message (#2468)",
         "Admin access to project tracker settings. A project with tracker configured.",
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Open a project with tracker integration\n"
         "4. Edit tracker settings — enter invalid credentials (wrong password/token)\n"
         "5. Save and attempt to sync\n"
         "6. Verify a clear error message is shown (not a generic 500)\n"
         "7. Verify the error distinguishes between 'wrong credentials' and 'server unreachable'\n"
         "8. Restore valid credentials",
         "Invalid credentials should show a clear error message, not a cascading 500 error. The error should distinguish "
         "authentication failure from connection failure. (#2468: inadequate error messages.)",
         "Medium", "UI", "#2468", "cross-service",
         "Ticket #2468: invalid credentials cause cascading errors, inadequate error messages."),

        ("TC-CS-046", "Tracker HTTP proxy configuration (#1174)",
         "Project using tracker that requires HTTP proxy. Admin access.",
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Open a project that uses a tracker behind a proxy\n"
         "4. Check proxy configuration fields in tracker settings\n"
         "5. Verify HTTP proxy option is available and configurable\n"
         "6. Note: HTTPS proxy is NOT implemented (#1305 — open)\n"
         "7. Attempt sync with proxy configured\n"
         "8. Verify connection goes through the specified proxy",
         "HTTP proxy for trackers is supported (#1174). HTTPS proxy is NOT implemented (#1305 — SSL cert handling needed). "
         "Security concern: prohibit HTTPS proxy as keys would be in cleartext.",
         "Medium", "UI", "#1174, #1305", "cross-service",
         "HTTPS proxy (#1305) is an open ticket. HTTP proxy works but has security implications."),
    ],

    "TS-CrossService-PMToolSync": [
        ("TC-CS-047", "PM Tool sync imports projects successfully",
         "PM Tool integration enabled on environment. Feature toggle pmtool-sync ON. "
         "Query: SELECT COUNT(*) FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL",
         "SETUP: Via test API — trigger PM Tool sync (if endpoint available)\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Search for projects with PM Tool IDs\n"
         "4. Verify project names and details match PM Tool data\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.pm_tool_sync_failed_entity\n"
         "DB-CHECK: SELECT sync_type, last_success_date FROM ttt_backend.pm_sync_status ORDER BY last_success_date DESC LIMIT 3",
         "PM Tool sync imports projects from PM Tool REST API. Projects appear in Admin with pm_tool_id set. "
         "Sync status recorded in pm_sync_status table.",
         "High", "Hybrid", "pm-tool-sync-implementation.md", "cross-service",
         "PmToolSyncScheduler runs every 15 min with ShedLock. Rate limited to 50 fetches/minute."),

        ("TC-CS-048", "PM Tool project count mismatch due to validation cascade",
         "PM Tool sync has run. Compare synced vs available project counts. "
         "Query: SELECT COUNT(*) as total_projects FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL",
         "DB-CHECK: SELECT COUNT(*) as synced FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL\n"
         "DB-CHECK: SELECT COUNT(*) as failed FROM ttt_backend.pm_tool_sync_failed_entity\n"
         "DB-CHECK: SELECT entity_id, error_message FROM ttt_backend.pm_tool_sync_failed_entity LIMIT 10\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Count PM Tool projects visible\n"
         "4. Compare with total available in PM Tool API\n"
         "5. Document the mismatch ratio",
         "KNOWN ISSUE: ~84% of PM Tool projects fail sync due to validation cascade. Any project with a missing watcher "
         "(employee not yet in TTT DB) causes the entire project to fail validation. Approximately 501 out of 3132 projects sync successfully.",
         "High", "Hybrid", "pm-tool-sync-implementation.md §Root Cause, #3382, #3387", "cross-service",
         "PmToolProjectSynchronizer.validateEmployeesExist() throws IllegalStateException for any missing watcher."),

        ("TC-CS-049", "Failed PM Tool projects retry indefinitely (no max limit)",
         "PM Tool sync with failed projects. "
         "Query: SELECT COUNT(*) FROM ttt_backend.pm_tool_sync_failed_entity",
         "DB-CHECK: SELECT COUNT(*) as failed_count FROM ttt_backend.pm_tool_sync_failed_entity\n"
         "DB-CHECK: Note the IDs of specific failed projects\n"
         "SETUP: Trigger PM Tool sync twice with 15-minute gap\n"
         "DB-CHECK: SELECT entity_id, error_message FROM ttt_backend.pm_tool_sync_failed_entity WHERE entity_id IN (<ids_from_before>)\n"
         "DB-CHECK: Verify the same projects are still failing (no max retry limit)\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Verify failed projects are not visible in the project list",
         "KNOWN BUG: No max retry limit for failed PM Tool project syncs. Projects with permanently missing employee references "
         "retry every 15 minutes forever, generating log noise and wasted API calls.",
         "Medium", "Hybrid", "pm-tool-sync-implementation.md §Failed Project Retry", "cross-service",
         "Failed projects tracked in pm_tool_sync_failed_entity table. Retried in batches of 10."),

        ("TC-CS-050", "PM Tool rate limiting — 429 Too Many Requests handling",
         "PM Tool sync active. Rate limiter configured at 50 fetches/min. "
         "Query: Check application.yml for pmTool.sync.fetch-rate-per-minute setting.",
         "SETUP: Trigger PM Tool sync\n"
         "DB-CHECK: Check pm_sync_status for sync timing and duration\n"
         "DB-CHECK: SELECT sync_type, last_success_date FROM ttt_backend.pm_sync_status ORDER BY last_success_date DESC LIMIT 5\n"
         "1. Login as Admin (no direct UI for rate limit monitoring)\n"
         "2. Check application logs for 429 responses from PM Tool API\n"
         "3. Verify sync completes despite rate limiting (with appropriate backoff)\n"
         "4. Document whether 429 errors cause sync failure or graceful retry",
         "PM Tool has a 60 RPM limit (#3399). TTT uses Guava RateLimiter at 50 fetches/minute to stay under the limit. "
         "If rate limit exceeded, verify graceful handling rather than sync failure.",
         "Medium", "Hybrid", "#3399, #3401, pm-tool-sync-implementation.md §Rate Limiting", "cross-service",
         "Rate limiter needed on TTT side (#3401). Guava RateLimiter is per-instance — may not cover multiple replicas."),

        ("TC-CS-051", "PM Tool sync with feature toggle pmtool-sync OFF",
         "pmtool-sync feature toggle accessible. "
         "Query: GET /api/ttt/v1/feature-toggles/pmtool-sync",
         "1. Via API — check pmtool-sync toggle: GET /api/ttt/v1/feature-toggles/pmtool-sync\n"
         "2. Note current toggle state\n"
         "3. If toggle is ON, note current pm_sync_status timestamps\n"
         "DB-CHECK: SELECT sync_type, last_success_date FROM ttt_backend.pm_sync_status ORDER BY last_success_date DESC LIMIT 3\n"
         "4. If possible, toggle OFF and trigger sync\n"
         "5. Verify sync does not execute\n"
         "6. Verify no error messages — just silent skip\n"
         "CLEANUP: Restore toggle to previous state",
         "With pmtool-sync-{env} toggle OFF, PmToolSyncLauncherImpl silently skips sync. No warning or indication "
         "that sync is disabled. pm_sync_status timestamps remain unchanged.",
         "Medium", "Hybrid", "feature-toggles-unleash.md §PM Tool", "cross-service",
         "Environment-qualified toggle name: pmtool-sync-{env}. Uses dedicated thread pool (pmToolSyncPool)."),

        ("TC-CS-052", "PM Tool dual ID fields — pm_tool_id vs pmt_id (#3387)",
         "Projects with PM Tool integration. "
         "Query: SELECT id, name, pm_tool_id FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL LIMIT 10",
         "DB-CHECK: SELECT id, name, pm_tool_id FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL LIMIT 10\n"
         "DB-CHECK: Verify whether a separate pmt_id column exists or is planned\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Find PM Tool projects\n"
         "4. Check whether both IDs are displayed or only one\n"
         "5. Document the ID mapping: pm_tool_id (TTT→PM Tool) vs pmt_id (PM Tool's internal PK)",
         "Two ID fields: pm_tool_id (TTT's reference to PM Tool) vs pmt_id (PM Tool's own PK). 501 vs 3137 project count "
         "mismatch may be partly due to ID confusion. Document which ID is used where.",
         "Medium", "Hybrid", "#3387", "cross-service",
         "Open ticket. The two IDs serve different purposes and their relationship needs clarification."),

        ("TC-CS-053", "PM Tool employee type filtering (#3389)",
         "PM Tool sync active. Some projects have 'sales' type employees. "
         "Query: SELECT COUNT(*) FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL",
         "SETUP: Trigger PM Tool sync\n"
         "DB-CHECK: Check if sales-type employees are filtered during sync\n"
         "DB-CHECK: SELECT employee_id, role FROM ttt_backend.project_employee pe "
         "JOIN ttt_backend.project p ON pe.project_id = p.id WHERE p.pm_tool_id IS NOT NULL LIMIT 20\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Projects\n"
         "3. Open a PM Tool project\n"
         "4. Check member list — verify no 'sales' type employees\n"
         "5. Compare with PM Tool API response to verify filtering",
         "Ticket #3389: Filter employees with 'sales' type during sync. removeSalesFromProject() nullifies employee references "
         "marked as 'sales'. If a critical role (owner/PM) is sales-typed, reference becomes null → may trigger validation failure.",
         "Medium", "Hybrid", "#3389, pm-tool-sync-implementation.md §Sales Filtering", "cross-service",
         "Sales filtering can cascade to validation failures if a project's PM/owner is sales-typed."),
    ],

    "TS-CrossService-EmailDelivery": [
        ("TC-CS-054", "Vacation notification email reaches email service DB",
         "Employee with manager. "
         "Query: SELECT e.login FROM ttt_vacation.employee e WHERE e.working = true AND e.manager_id IS NOT NULL AND e.maternity = false ORDER BY random() LIMIT 1",
         "SETUP: Via API — create a vacation: POST /api/vacation/v1/vacations\n"
         "1. Wait 15-30 seconds for email processing\n"
         "DB-CHECK: SELECT id, subject, status, to_field, add_time FROM ttt_email.email WHERE subject LIKE '%vacation%' ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify status is 'SENT' (not 'NEW' or 'FAILED')\n"
         "DB-CHECK: Verify the email recipient matches the employee's manager email\n"
         "CLEANUP: Via API — delete the created vacation",
         "Vacation creation triggers NEW_VACATION_PM and/or APPROVE_REQUEST email template. Email appears in ttt_email.email "
         "with status NEW, then batch processor changes to SENT after SMTP delivery.",
         "High", "Hybrid", "email-notification-deep-dive.md §Template Categories", "cross-service",
         "EmailSendScheduler runs on cron. Check email.scheduler.send.cron for frequency."),

        ("TC-CS-055", "Report changed notification email delivered",
         "Employee with reports in current period. Manager configured. "
         "Query: SELECT DISTINCT e.login FROM ttt_backend.task_report tr JOIN ttt_backend.employee e ON tr.employee_id = e.id "
         "WHERE tr.report_date >= DATE_TRUNC('month', CURRENT_DATE) AND e.enabled = true LIMIT 3",
         "SETUP: Via API — modify an existing report (PATCH /api/ttt/v1/reports)\n"
         "SETUP: Via test API — trigger report change notifications: POST /api/ttt/test/v1/notifications/reports-changed\n"
         "DB-CHECK: SELECT id, subject, status, to_field FROM ttt_email.email WHERE subject LIKE '%report%' ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify REPORT_SHEET_CHANGED template was used\n"
         "1. Verify email status progresses to SENT",
         "Report modification triggers REPORT_SHEET_CHANGED notification to the report's manager/confirmer. "
         "TaskReportNotificationScheduler.sendReportsChangedNotifications() processes these in batch.",
         "Medium", "Hybrid", "email-notification-deep-dive.md §Scheduled Notification Jobs", "cross-service",
         "Report change notifications are scheduled, not immediate. Use test API to trigger for faster testing."),

        ("TC-CS-056", "Email template rendering with Mustache variables",
         "Valid email template in DB. "
         "Query: SELECT code, subject FROM ttt_email.email_template LIMIT 10",
         "DB-CHECK: SELECT code, subject, LEFT(body, 200) as body_preview FROM ttt_email.email_template WHERE code = 'NEW_VACATION_PM'\n"
         "DB-CHECK: Verify template contains Mustache variables: {{to_name}}, {{employee}}, {{period}}, etc.\n"
         "SETUP: Via API — create a vacation to trigger the template\n"
         "DB-CHECK: SELECT subject, LEFT(body, 500) as body_preview FROM ttt_email.email WHERE subject LIKE '%vacation%' ORDER BY add_time DESC LIMIT 1\n"
         "DB-CHECK: Verify Mustache variables were rendered (no raw {{variable}} in output)\n"
         "CLEANUP: Delete the test vacation",
         "Mustache template rendering replaces {{to_name}}, {{employee}}, {{period}} etc. with actual values. "
         "Rendered email should have no raw Mustache tags. Subject and body should contain the employee's name and dates.",
         "Medium", "Hybrid", "email-notification-deep-dive.md §Mustache Syntax", "cross-service",
         "35+ templates in ttt_email.email_template. Largest is DIGEST (21KB). Variables are JSON data from the caller."),

        ("TC-CS-057", "Failed email stays in FAILED status permanently",
         "Email records with FAILED status. "
         "Query: SELECT id, subject, status, error_message FROM ttt_email.email WHERE status = 'FAILED' LIMIT 5",
         "DB-CHECK: SELECT id, subject, status, error_message, add_time FROM ttt_email.email WHERE status = 'FAILED' ORDER BY add_time DESC LIMIT 10\n"
         "DB-CHECK: Verify failed emails have error_message populated\n"
         "DB-CHECK: Wait for next batch cycle — verify failed emails are NOT retried (status stays FAILED)\n"
         "1. Login as Admin (no direct UI for email monitoring)\n"
         "2. Document the failure reasons and whether they are recoverable",
         "KNOWN DESIGN ISSUE: Failed emails stay in FAILED status permanently. No retry mechanism from FAILED → NEW. "
         "Only manual DB update (SET status = 'NEW') can retry a failed email. No retry count tracking.",
         "Medium", "Hybrid", "email-notification-deep-dive.md §Design Issues #4", "cross-service",
         "EmailWriter: FAILED emails permanent unless manual intervention. MailAuthenticationException leaves emails as NEW forever."),

        ("TC-CS-058", "Email with INVALID status due to bad email address",
         "Email sent to an invalid email address. "
         "Query: SELECT id, subject, status, to_field, error_message FROM ttt_email.email WHERE status = 'INVALID' LIMIT 5",
         "DB-CHECK: SELECT id, subject, status, to_field, error_message FROM ttt_email.email WHERE status = 'INVALID' ORDER BY add_time DESC LIMIT 10\n"
         "DB-CHECK: Verify INVALID status is only for SendFailedException with invalid addresses\n"
         "DB-CHECK: Compare with FAILED status — different error types\n"
         "1. Document the distinction: INVALID = bad email address, FAILED = SMTP error",
         "INVALID status set when SMTP rejects the recipient address (SendFailedException). Distinct from FAILED which is for "
         "connection/auth/generic SMTP errors. Neither status is retried automatically.",
         "Low", "Hybrid", "email-notification-deep-dive.md §StatusType State Machine", "cross-service",
         "Status machine: NEW → SENT (success), NEW → FAILED (SMTP error), NEW → INVALID (bad address). No transitions out of terminal states."),

        ("TC-CS-059", "Digest email compilation — multiple event types",
         "Digest scheduler configured. Recent absence events exist. "
         "Query: SELECT COUNT(*) FROM ttt_vacation.timeline_event WHERE created_date >= CURRENT_DATE - INTERVAL '7 days'",
         "SETUP: Via test API — trigger digest: POST /api/vacation/test/v1/digests\n"
         "DB-CHECK: SELECT id, subject, status, to_field FROM ttt_email.email WHERE subject LIKE '%digest%' OR subject LIKE '%Digest%' ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify digest email body contains multiple sections (vacation, day-off, sick leave events)\n"
         "1. Check that the digest groups events by receiver (manager, approver, etc.)\n"
         "2. Verify the DIGEST template (21KB) rendered correctly with all event types",
         "Digest compilation aggregates timeline events (vacation created/approved/rejected, day-off events, sick leave events) "
         "into a single email per receiver. Uses 28+ TimelineEventProcessor implementations to process each event type.",
         "Medium", "Hybrid", "email-notification-deep-dive.md §Digest System", "cross-service",
         "DigestScheduler → DigestService.sendDigests() → TimelineEventProcessor(s). Receivers calculated by ReceiverRole."),

        ("TC-CS-060", "Hardcoded notification URL — broken on test environments",
         "Vacation notification email in the DB. "
         "Query: SELECT id, body FROM ttt_email.email WHERE body LIKE '%ttt.noveogroup.com%' ORDER BY add_time DESC LIMIT 3",
         "DB-CHECK: SELECT id, LEFT(body, 2000) FROM ttt_email.email WHERE body LIKE '%ttt.noveogroup.com%' ORDER BY add_time DESC LIMIT 3\n"
         "DB-CHECK: Search for hardcoded URL: SELECT COUNT(*) FROM ttt_email.email WHERE body LIKE '%ttt.noveogroup.com/vacation/request%'\n"
         "1. Verify the confirmation URL in notification emails points to ttt.noveogroup.com (production)\n"
         "2. On test environments, this URL is incorrect — should point to ttt-qa-1.noveogroup.com\n"
         "3. Document that clicking the link from a test notification will redirect to production",
         "KNOWN DESIGN ISSUE: AbstractVacationNotificationHelper hardcodes confirmUrl to 'https://ttt.noveogroup.com/vacation/request'. "
         "On test environments (qa-1, timemachine), notification emails contain production URLs. Should be configurable.",
         "Medium", "Hybrid", "email-notification-deep-dive.md §Design Issues #2", "cross-service",
         "Hardcoded in AbstractVacationNotificationHelper.fillBaseInfo(). Not environment-aware."),
    ],

    "TS-CrossService-Regression": [
        ("TC-CS-061", "CS sync silently stopped on environments (#2496)",
         "Test API for CS sync trigger. "
         "Query: SELECT sync_type, last_success_date, status FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 5",
         "DB-CHECK: Check cs_sync_status for recent sync timestamps\n"
         "DB-CHECK: Verify sync has run within the expected interval (15 min)\n"
         "SETUP: Via test API — trigger CS sync manually\n"
         "DB-CHECK: Verify cs_sync_status.last_success_date updated after manual trigger\n"
         "1. Login as Admin\n"
         "2. Check if any employee data appears stale (wrong office, missing roles)\n"
         "3. Document the gap between expected and actual last sync time",
         "Regression for #2496: Cron-based sync silently stopped on dev/stage/qa-1. Manual trigger worked. "
         "Verify cron scheduler is running and last_success_date is recent.",
         "High", "Hybrid", "#2496", "cross-service",
         "ShedLock may prevent cron execution if lock is held. Check shedlock table for stale locks."),

        ("TC-CS-062", "Self-referencing manager in CS breaks sync (#1892)",
         "Employee who is their own manager. "
         "Query: SELECT login FROM ttt_backend.employee WHERE manager_id = id AND enabled = true LIMIT 3",
         "DB-CHECK: SELECT login, manager_id, id FROM ttt_backend.employee WHERE manager_id = id AND enabled = true\n"
         "1. Login as the self-referencing employee\n"
         "2. Navigate to Calendar of absences\n"
         "3. Try to create a vacation\n"
         "4. Verify the approval flow works (self-referencing manager should still allow approval)\n"
         "5. Check Confirmation page — verify no infinite loop in approval chain\n"
         "6. Document observed behavior",
         "Regression for #1892: Self-referencing manager in CS broke TTT sync completely — reports/approval stopped working. "
         "After fix, self-referencing manager should not break sync or approval workflows.",
         "Critical", "Hybrid", "#1892", "cross-service",
         "Self-referencing manager: employee.manager_id = employee.id. May cause infinite loops in approval chain traversal."),

        ("TC-CS-063", "Delta sync data loss — acknowledged risk (#2989, #3262)",
         "Environment with active CS sync. "
         "Query: SELECT sync_type, last_success_date FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 5",
         "DB-CHECK: Check last successful sync time\n"
         "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.cs_sync_failed_entity — any lingering failures?\n"
         "1. Login as Admin\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Compare a sample of employees against CompanyStaff (if accessible)\n"
         "4. Document any discrepancies found\n"
         "5. Check if 3 AM full re-sync has run (compare cs_sync_status timestamps around 03:00 UTC)",
         "KNOWN RISK (#2989, #3262): Delta sync doesn't guarantee consistency — events can be lost, data changes without "
         "timestamp update. Full sync removed March 2023. 3 AM re-sync is the safety net. Between midnight and 3 AM, stale data may exist.",
         "High", "Hybrid", "#2989, #3262, #3303", "cross-service",
         "Acknowledged risk: 'cannot guarantee all events processed correctly, MQ doesn't store after ack.' Periodic re-sync at 3 AM as fallback."),

        ("TC-CS-064", "Email delays via RabbitMQ — hours-long delivery (#2518)",
         "Notification trigger available. Email service DB accessible. "
         "Query: SELECT id, add_time, sent_time, status FROM ttt_email.email WHERE status = 'SENT' ORDER BY add_time DESC LIMIT 10",
         "SETUP: Via API — create a vacation to trigger notification\n"
         "SETUP: Note the exact time of creation\n"
         "DB-CHECK: SELECT id, add_time, sent_time, EXTRACT(EPOCH FROM (sent_time - add_time)) as delay_seconds FROM ttt_email.email ORDER BY add_time DESC LIMIT 5\n"
         "DB-CHECK: Verify delay between add_time and sent_time is reasonable (<5 minutes)\n"
         "1. If delay is >5 minutes, document the actual delay\n"
         "CLEANUP: Delete the test vacation",
         "Regression for #2518: After RabbitMQ migration, email delivery had hours-long delays. Chronological ordering broken. "
         "Measure actual delay between email creation (add_time) and delivery (sent_time). Expected: <5 min.",
         "High", "Hybrid", "#2518", "cross-service",
         "Email batch processor runs on cron schedule. Delay = time between add_time (RabbitMQ save) and sent_time (SMTP delivery)."),

        ("TC-CS-065", "Test clock corruption of cs_sync_status (#2629)",
         "Timemachine environment with clock manipulation API. "
         "Only on timemachine: PATCH /api/ttt/test/v1/clock",
         "SETUP: On Timemachine env only\n"
         "SETUP: Note current cs_sync_status state\n"
         "1. Via test API — advance clock by 2 days\n"
         "2. Via test API — trigger CS sync\n"
         "DB-CHECK: SELECT * FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 5\n"
         "3. Via test API — revert clock to current time\n"
         "4. Via test API — trigger CS sync again\n"
         "DB-CHECK: Check cs_sync_status for duplicate entries or corrupted timestamps\n"
         "5. Verify the vacation service is still responsive (not crashed)\n"
         "CLEANUP: Reset clock to actual current time",
         "Regression for #2629: Time travel corrupts cs_sync_status table. Moving time backward creates duplicates. "
         "Vacation service crashed after table was manually cleared. Verify sync recovery after clock manipulation.",
         "High", "Hybrid", "#2629", "cross-service",
         "Timemachine-only test. After fix, clock revert should not create duplicate sync status entries."),

        ("TC-CS-066", "Application stability under concurrent sync load (#3023)",
         "Environment with CS sync and PM Tool sync active. DB connection pool monitoring. "
         "Query: SELECT sync_type, last_success_date FROM ttt_backend.cs_sync_status ORDER BY last_success_date DESC LIMIT 3",
         "SETUP: Via test API — trigger CS employee sync\n"
         "SETUP: Simultaneously trigger CS office sync and PM Tool sync (if endpoints available)\n"
         "1. Login to the application immediately after triggering syncs\n"
         "2. Navigate to several pages: My Tasks, Calendar, Planner, Admin\n"
         "3. Verify the application remains responsive during sync\n"
         "4. Check for 500 errors or timeouts\n"
         "5. Check application health endpoint if available\n"
         "DB-CHECK: Monitor active DB connections: SELECT count(*) FROM pg_stat_activity WHERE datname = 'ttt_backend'",
         "KNOWN ISSUE (#3023): App hangs with >30 concurrent threads. Multiple sequential full CS syncs crash vacation service. "
         "DB connection pool exhaustion. Test verifies application remains responsive during concurrent sync operations.",
         "Critical", "Hybrid", "#3023", "cross-service",
         "DB connection pool exhaustion is the root cause. JMeter showed failure at >30 threads. Don't push to that limit in manual testing."),

        ("TC-CS-067", "CS sync returns 200 on internal failure (#2969)",
         "CS sync trigger available. "
         "Query: SELECT COUNT(*) FROM ttt_backend.cs_sync_failed_entity",
         "SETUP: Via test API — trigger CS sync\n"
         "SETUP: Note the HTTP response status code from the sync trigger\n"
         "DB-CHECK: SELECT COUNT(*) as failures_before FROM ttt_backend.cs_sync_failed_entity\n"
         "DB-CHECK: After sync — SELECT COUNT(*) as failures_after FROM ttt_backend.cs_sync_failed_entity\n"
         "DB-CHECK: If failures_after > failures_before, some entities failed\n"
         "1. Verify the sync trigger returned 200 even if entities failed\n"
         "2. Document: sync returns 200 even when individual entity sync fails internally",
         "KNOWN BUG (#2969): CS sync returns 200 on internal failure. Individual entity failures are tracked in "
         "cs_sync_failed_entity but the overall sync endpoint returns success. No way to detect partial failures from the HTTP response.",
         "Medium", "Hybrid", "#2969", "cross-service",
         "Contractor salary office transfer was specifically not synced in the original ticket."),

        ("TC-CS-068", "WebSocket JWT expiry — polling fallback prevents storm (#2270)",
         "User with Planner open in browser. Timemachine env for clock advancement.",
         "SETUP: Via test API — advance clock by 25 hours (Timemachine only)\n"
         "1. Login as employee with Planner access\n"
         "2. Navigate to Planner\n"
         "3. Open browser DevTools > Network tab\n"
         "4. Wait for JWT expiry to trigger\n"
         "5. Monitor network requests — verify /v1/authentication/check polling every 5s\n"
         "6. Verify NO infinite WebSocket reconnection attempts (should see polling instead)\n"
         "7. Verify app eventually shows login prompt\n"
         "CLEANUP: Reset clock",
         "Regression for #2270: JWT expiry was causing WebSocket reconnection loops. Fix: poll /v1/authentication/check every 5s. "
         "After fix, no WebSocket storm. App gracefully detects expired auth and prompts for re-login.",
         "High", "Hybrid", "#2270", "cross-service",
         "Before fix, expired JWT caused infinite WebSocket reconnection storms. Polling fallback added as fix."),

        ("TC-CS-069", "RabbitMQ @Async publisher failure after transaction commit",
         "Calendar change that triggers vacation recalculation (same as TC-CS-013). "
         "Query: SELECT o.id, o.name FROM ttt_backend.office o LIMIT 3",
         "1. Login as Admin\n"
         "2. Navigate to Admin > Production calendars\n"
         "3. Make a calendar change (add a non-working day)\n"
         "4. Save the change\n"
         "5. Immediately check: did the calendar change persist in the DB?\n"
         "DB-CHECK: Verify calendar change saved in ttt_calendar tables\n"
         "6. Wait 30 seconds\n"
         "DB-CHECK: Check if vacation recalculation happened in ttt_vacation\n"
         "7. If recalculation did NOT happen, document: this confirms @Async publisher failure scenario",
         "KNOWN DESIGN ISSUE: @Async @EventListener publishers can fail silently after source transaction commits. "
         "Calendar change is saved (committed) but RabbitMQ message may never be sent. Vacation service never notified. "
         "Safety net: 3 AM re-sync. This test may not reliably reproduce the failure — it's a race condition.",
         "High", "Hybrid", "cross-service-integration.md §Design Issues #3, #3262", "cross-service",
         "Race condition: @Async event fires after tx commit. If RabbitMQ is slow/down at that moment, event lost with no DLQ."),

        ("TC-CS-070", "Browser memory leak on project assignments tab (#2865)",
         "Project with many assignments/members. "
         "Query: SELECT p.id, p.name, COUNT(*) as member_count FROM ttt_backend.project p "
         "JOIN ttt_backend.project_employee pe ON p.id = pe.project_id GROUP BY p.id, p.name ORDER BY member_count DESC LIMIT 5",
         "1. Login as PM of a large project (many members/assignments)\n"
         "2. Navigate to Planner for the project\n"
         "3. Switch to the Assignments tab\n"
         "4. Open browser Task Manager (Chrome: Shift+Esc)\n"
         "5. Monitor memory usage of the TTT tab\n"
         "6. Scroll through assignments, switch pages\n"
         "7. Document memory usage over 5 minutes of interaction\n"
         "8. If memory exceeds 2-3 GB, document as regression",
         "Regression for #2865: Browser runs out of memory (5-8 GB) on project assignments tab. "
         "1000 simultaneous requests observed. Monitor memory usage during Planner interaction with large projects.",
         "Medium", "UI", "#2865", "cross-service",
         "Root cause: too many simultaneous API requests and DOM nodes. May be related to WebSocket event flooding."),
    ],
}


# --- Helper functions --------------------------------------------------------

def style_header_row(ws, row_idx, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row_idx, max_col, alt=False):
    fill = FILL_ROW_ALT if alt else FILL_ROW_WHITE
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_BODY
        cell.fill = fill
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_link(ws):
    ws.cell(row=1, column=1, value='← Back to Plan')
    ws.cell(row=1, column=1).font = FONT_BACK_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"


# --- Sheet builders ----------------------------------------------------------

def build_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN
    set_col_widths(ws, [20, 90])

    ws.cell(row=1, column=1, value=PLAN_OVERVIEW["title"]).font = FONT_TITLE
    ws.merge_cells("A1:B1")

    row = 3
    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP

    row = 5
    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    for i, obj in enumerate(PLAN_OVERVIEW["objectives"]):
        ws.cell(row=row + i, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row + i, column=2).alignment = ALIGN_WRAP
    row += len(PLAN_OVERVIEW["objectives"]) + 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    for i, env in enumerate(PLAN_OVERVIEW["environments"]):
        ws.cell(row=row + i, column=2, value=env).font = FONT_BODY
    row += len(PLAN_OVERVIEW["environments"]) + 1

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    for i, dep in enumerate(PLAN_OVERVIEW["dependencies"]):
        ws.cell(row=row + i, column=2, value=f"• {dep}").font = FONT_BODY
    row += len(PLAN_OVERVIEW["dependencies"]) + 1

    # Suite hyperlinks
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, cases in SUITES.items():
        label = f"{suite_name} — {len(cases)} cases"
        ws.cell(row=row, column=2, value=label).font = FONT_LINK
        ws.cell(row=row, column=2).hyperlink = f"#'{suite_name}'!A1"
        row += 1

    return ws


def build_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "CSSync", "EventProp", "DataIntegrity", "WebSocket",
               "TrackerSync", "PMToolSync", "EmailDelivery", "Regression", "Total"]
    set_col_widths(ws, [42, 10, 10, 12, 10, 11, 11, 12, 10, 8])

    row = 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    suite_names = list(SUITES.keys())
    for i, feat_row in enumerate(FEATURE_MATRIX):
        row = i + 2
        for col, val in enumerate(feat_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 1:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_CENTER
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

        for col_idx in range(2, len(headers)):
            val = feat_row[col_idx]
            if val and val > 0:
                suite_idx = col_idx - 2
                if suite_idx < len(suite_names):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.font = FONT_LINK
                    cell.hyperlink = f"#'{suite_names[suite_idx]}'!A1"

    total_row = len(FEATURE_MATRIX) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col in range(2, len(headers) + 1):
        total = sum(r[col - 1] for r in FEATURE_MATRIX)
        ws.cell(row=total_row, column=col, value=total).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=total_row, column=col).alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(FEATURE_MATRIX) + 1}"
    return ws


def build_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    set_col_widths(ws, [30, 50, 12, 12, 12, 50])

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for i, (feat, risk, like, impact, sev, mit) in enumerate(RISK_ASSESSMENT):
        row = i + 2
        ws.cell(row=row, column=1, value=feat)
        ws.cell(row=row, column=2, value=risk)
        ws.cell(row=row, column=3, value=like)
        ws.cell(row=row, column=4, value=impact)
        ws.cell(row=row, column=5, value=sev)
        ws.cell(row=row, column=6, value=mit)
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))
        if sev in severity_fills:
            ws.cell(row=row, column=5).fill = severity_fills[sev]

    ws.auto_filter.ref = f"A1:F{len(RISK_ASSESSMENT) + 1}"
    return ws


def build_suite_sheet(wb, suite_name, cases):
    ws = wb.create_sheet(suite_name)
    ws.sheet_properties.tabColor = TAB_COLOR_SUITE

    add_back_link(ws)

    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    set_col_widths(ws, [14, 40, 40, 55, 40, 10, 10, 30, 14, 35])

    header_row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(headers))

    for i, case in enumerate(cases):
        row = header_row + 1 + i
        for col, val in enumerate(case, 1):
            ws.cell(row=row, column=col, value=val)
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(cases)}"
    return ws


# --- Main --------------------------------------------------------------------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    build_plan_overview(wb)
    build_feature_matrix(wb)
    build_risk_assessment(wb)

    for suite_name, cases in SUITES.items():
        build_suite_sheet(wb, suite_name, cases)

    wb.save(OUTPUT_FILE)

    total_cases = sum(len(cases) for cases in SUITES.values())
    print(f"Generated {OUTPUT_FILE}")
    print(f"  Suites: {len(SUITES)}")
    print(f"  Total test cases: {total_cases}")
    print(f"  Tabs: Plan Overview, Feature Matrix, Risk Assessment, {', '.join(SUITES.keys())}")


if __name__ == "__main__":
    main()
