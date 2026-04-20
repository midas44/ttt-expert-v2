"""Shared helpers for the per-domain Cron_<Domain>.xlsx generators under
``expert-system/generators/t3423/``.

Each ``generate_cron_<domain>.py`` creates a fresh workbook at
``test-docs/collections/cron/Cron_<Domain>.xlsx`` containing:

  * a Plan Overview sheet (authored by :func:`author_plan_overview`)
  * the domain's ``TS-*-Cron-*`` suite sheets with the TCs

The Plan Overview template is kept in this shared module so that all
five generators produce identical metadata. The one-shot migration script
(``migrate_to_cron_workbooks.py``) also uses this helper.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill


def author_plan_overview(
    wb,
    *,
    domain: str,  # PascalCase: "Vacation", "Reports", "CrossService", "Statistics", "Email"
    home_subdir: str,  # slug used under test-docs/: "vacation", "reports", "cross-service", "statistics", "email"
    rows_covered: str,
    tc_count: int,
    suite_row_map: list[tuple[str, str, int]],  # (suite_name, cron_rows_label, tc_count)
    is_email: bool = False,
):
    """Author the Plan Overview sheet as the first tab of `wb`.

    `is_email=True` signals the "entire workbook is cron" variant, which gets a
    different "What's NOT here" blurb — the email service is 100% cron-driven.
    """
    ws = wb.create_sheet(title="Plan Overview", index=0)
    ws.sheet_properties.tabColor = "0070AD47"  # green (matches home-workbook meta tabs)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    bold = Font(bold=True, size=12)
    h1 = Font(bold=True, size=16, color="1F4E78")
    h2 = Font(bold=True, size=13, color="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    top = Alignment(vertical="top", wrap_text=True)

    def put(row, col, val, *, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align

    r = 1
    put(r, 1, f"Cron_{domain} — Test Plan Overview", font=h1)
    r += 2

    put(r, 1, "Ticket", font=bold)
    put(r, 2, "#3423 — Cron & Startup Jobs Testing Collection (epic: #3402, linked via relates_to)")
    r += 1
    put(r, 1, "Domain", font=bold)
    put(r, 2, domain)
    r += 1
    put(r, 1, "Collection", font=bold)
    put(r, 2, "cron (87 TCs total across 5 domains · see COL-cron in cron.xlsx)")
    r += 1
    put(r, 1, "Cron rows covered", font=bold)
    put(r, 2, rows_covered, align=top)
    r += 1
    put(r, 1, "TC count", font=bold)
    put(r, 2, str(tc_count))
    r += 2

    put(r, 1, "Test Suites", font=h2)
    r += 1
    put(r, 1, "Suite", font=bold, fill=header_fill)
    put(r, 2, "Cron row(s) / TC count", font=bold, fill=header_fill)
    r += 1
    for suite_name, rows_label, tcs in suite_row_map:
        put(r, 1, suite_name)
        put(r, 2, f"row(s) {rows_label} · {tcs} TC" + ("s" if tcs != 1 else ""))
        r += 1
    r += 1

    put(r, 1, "What's NOT here", font=h2)
    r += 1
    if is_email:
        not_here_msg = (
            "This workbook contains the entire email-service test documentation. "
            "The email service is a dispatcher/retention layer — all of its behaviour is cron-driven "
            "(dispatch every 20s, prune daily 00:00). There is no separate non-cron email workbook; "
            "non-cron email flows (report notifications, vacation digest, budget alerts) live in the "
            "business-module workbooks that trigger them."
        )
    else:
        not_here_msg = (
            f"Non-cron TCs for the {domain} domain live in the home workbook "
            f"test-docs/{home_subdir}/{home_subdir}.xlsx. This file holds only cron-scheduled behaviour."
        )
    put(r, 1, not_here_msg, align=top)
    ws.row_dimensions[r].height = 60
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    put(r, 1, "Related files", font=h2)
    r += 1
    put(r, 1, "Task definition")
    put(r, 2, "docs/tasks/cron/cron-testing-task.md")
    r += 1
    put(r, 1, "Execution report (Phases A + B)")
    put(r, 2, "docs/tasks/cron/execution-phases-a-b.md")
    r += 1
    put(r, 1, "Collection reference")
    put(r, 2, "test-docs/collections/cron/cron.xlsx (COL-cron sheet · 87 TC refs)")
    r += 1
    put(r, 1, "Traceability matrix")
    put(r, 2, "test-docs/collections/cron/coverage.md")
    r += 1
    put(r, 1, "Prior-art vault notes")
    put(
        r,
        2,
        "expert-system/vault/external/EXT-cron-jobs.md, "
        "exploration/api-findings/cron-job-live-verification.md, "
        "patterns/email-notification-triggers.md",
        align=top,
    )
    ws.row_dimensions[r].height = 45
    r += 2

    put(r, 1, "Generator script", font=bold)
    put(
        r,
        2,
        f"expert-system/generators/t3423/generate_cron_{_snake(domain)}.py  (idempotent wipe-and-rewrite)",
    )
    r += 1
    put(r, 1, "Last regenerated", font=bold)
    put(r, 2, "Run this generator to refresh — a wipe-and-rewrite Workbook() is produced on every run.")


def _snake(pascal: str) -> str:
    """CrossService → crossservice; Vacation → vacation."""
    return pascal.lower()
