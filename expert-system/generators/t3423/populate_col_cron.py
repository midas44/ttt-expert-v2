"""Populate COL-cron sheet in cron.xlsx with home-module TC references.

Idempotent: wipes all rows below the header (row 4) and rewrites from scratch.
Rows are grouped by cron-row number for traceability.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[3]
CRON_XLSX = ROOT / "test-docs" / "collections" / "cron" / "cron.xlsx"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (test_id, source_module, source_suite, title, inclusion_reason, priority_override)
# Ordered by cron-row number for cross-reference clarity.
ROWS = [
    # Row 11 — AnnualAccruals
    ("TC-VAC-101", "vacation", "TS-Vac-Cron-AnnualAccruals", "Annual accruals — clock at 01 Jan advances balances for all enabled employees", "Row 11 happy-path; baseline accrual run", "Critical"),
    ("TC-VAC-102", "vacation", "TS-Vac-Cron-AnnualAccruals", "Annual accruals — disabled employees excluded from accrual run", "Row 11 negative; ensures enabled=true filter", "High"),
    ("TC-VAC-103", "vacation", "TS-Vac-Cron-AnnualAccruals", "Annual accruals — idempotent on second trigger within same year", "Row 11 idempotency; silent-failure guard", "High"),
    # Row 12, 13 — NotImpl
    ("TC-VAC-104", "vacation", "TS-Vac-Cron-NotImpl", "Row 12 stub endpoint returns 200 / 204 with no side effects", "Row 12 dead-config confirmation", "Low"),
    ("TC-VAC-105", "vacation", "TS-Vac-Cron-NotImpl", "Row 13 stub endpoint returns 200 / 204 with no side effects", "Row 13 dead-config confirmation", "Low"),
    # Row 14 — Digest
    ("TC-VAC-106", "vacation", "TS-Vac-Cron-Digest", "Vacation digest — all employees with upcoming vacation receive single digest email", "Row 14 happy-path; batch dispatch via email service", "Critical"),
    ("TC-VAC-107", "vacation", "TS-Vac-Cron-Digest", "Vacation digest — employees without upcoming vacation do not receive digest", "Row 14 filter; ensures non-empty-only recipients", "High"),
    ("TC-VAC-108", "vacation", "TS-Vac-Cron-Digest", "Vacation digest — idempotent on second trigger within same day", "Row 14 idempotency; ShedLock window guard", "Medium"),
    # Row 15 — CalendarReminder
    ("TC-VAC-109", "vacation", "TS-Vac-Cron-CalendarReminder", "Production-calendar annual-first reminder — per-recipient mail to target audience", "Row 15 happy-path via test endpoint (scheduler-wrapper bypass)", "Critical"),
    ("TC-VAC-110", "vacation", "TS-Vac-Cron-CalendarReminder", "Production-calendar reminder — email subject matches NOTIFY_VACATION_CALENDAR_NEXT_YEAR_FIRST template", "Row 15 template verification per delta #5", "High"),
    ("TC-VAC-111", "vacation", "TS-Vac-Cron-CalendarReminder", "Production-calendar reminder — inactive employees excluded from recipient list", "Row 15 negative; recipient filter", "High"),
    # Row 16 — AutoPay
    ("TC-VAC-112", "vacation", "TS-Vac-Cron-AutoPay", "Auto-pay — vacations with payment_date in past and status APPROVED transition to PAID", "Row 16 happy-path; baseline auto-pay scan", "Critical"),
    ("TC-VAC-113", "vacation", "TS-Vac-Cron-AutoPay", "Auto-pay — vacations with payment_date in future are not transitioned", "Row 16 negative; date-window filter", "High"),
    ("TC-VAC-114", "vacation", "TS-Vac-Cron-AutoPay", "Auto-pay — idempotent on repeated trigger within same day", "Row 16 idempotency", "Medium"),
    # Row 17 — ApprovedToPaid
    ("TC-VAC-115", "vacation", "TS-Vac-Cron-ApprovedToPaid", "Pay expired approved vacations — APPROVED vacations past payment_date transition to PAID", "Row 17 happy-path; DB-only assertion (no SchedulerLock, no markers)", "Critical"),
    ("TC-VAC-116", "vacation", "TS-Vac-Cron-ApprovedToPaid", "Pay expired approved — only expired APPROVED (payment_date ≤ today) are transitioned", "Row 17 filter; ensures date-window logic", "High"),
    ("TC-VAC-117", "vacation", "TS-Vac-Cron-ApprovedToPaid", "Pay expired approved — REJECTED / CANCELED / NEW vacations untouched", "Row 17 status-filter negative", "High"),
    # Rows 18, 19 — EmpProjectSync
    ("TC-VAC-118", "vacation", "TS-Vac-Cron-EmpProjectSync", "Periodic employee-project sync — TTT→Vacation table mirrors employee_project_assignments", "Row 18 happy-path", "Critical"),
    ("TC-VAC-119", "vacation", "TS-Vac-Cron-EmpProjectSync", "Periodic employee-project sync — new assignment within window is propagated", "Row 18 delta-detect", "High"),
    ("TC-VAC-120", "vacation", "TS-Vac-Cron-EmpProjectSync", "Periodic employee-project sync — removed assignment within window is propagated", "Row 18 delete-detect", "High"),
    ("TC-VAC-121", "vacation", "TS-Vac-Cron-EmpProjectSync", "Periodic employee-project sync — data-loss in sync window (WON'T FIX) reproduces and documents risk", "Row 18 known risk; regression guard for documented WON'T FIX bug", "Medium"),
    ("TC-VAC-122", "vacation", "TS-Vac-Cron-EmpProjectSync", "Periodic employee-project sync — idempotent on repeated trigger", "Row 18 idempotency", "Medium"),
    ("TC-VAC-123", "vacation", "TS-Vac-Cron-EmpProjectSync", "Startup employee-project sync — runs on first boot after EMPLOYEE_PROJECT_INITIAL_SYNC marker absent", "Row 19 startup happy-path; feature-toggle gated", "Critical"),
    ("TC-VAC-124", "vacation", "TS-Vac-Cron-EmpProjectSync", "Startup employee-project sync — does not rerun when java_migration marker present", "Row 19 idempotency gating", "High"),
    ("TC-VAC-125", "vacation", "TS-Vac-Cron-EmpProjectSync", "Startup employee-project sync — populates marker on successful completion", "Row 19 marker-writes confirmation", "High"),
    # Row 21 — StatReportInit
    ("TC-VAC-126", "vacation", "TS-Vac-Cron-StatReportInit", "Startup statistic-report sync — runs on first boot after STATISTIC_REPORT_INITIAL_SYNC marker absent", "Row 21 startup happy-path; feature-toggle gated", "Critical"),
    ("TC-VAC-127", "vacation", "TS-Vac-Cron-StatReportInit", "Startup statistic-report sync — does not rerun when java_migration marker present", "Row 21 idempotency gating", "High"),
    # Row 1 — ForgottenReport (daily reminder)
    ("TC-RPT-101", "reports", "TS-Reports-CronNotifications", "Forgotten-report reminder — employees with missing prior-day reports receive FORGOTTEN_REPORT email", "Row 1 happy-path; baseline daily notification", "Critical"),
    ("TC-RPT-102", "reports", "TS-Reports-CronNotifications", "Forgotten-report reminder — employees with fully-reported prior day not notified", "Row 1 filter; missing-reports-only recipients", "High"),
    ("TC-RPT-103", "reports", "TS-Reports-CronNotifications", "Forgotten-report reminder — idempotent on repeated trigger within same day", "Row 1 idempotency; ShedLock window guard", "Medium"),
    # Row 2 — ForgottenReportDelayed (next-day escalation)
    ("TC-RPT-104", "reports", "TS-Reports-CronNotifications", "Delayed forgotten-report reminder — two-days-prior employees without reports escalated", "Row 2 happy-path; next-day escalation", "Critical"),
    ("TC-RPT-105", "reports", "TS-Reports-CronNotifications", "Delayed forgotten-report reminder — inactive employees excluded from recipient list", "Row 2 negative; recipient filter", "High"),
    # Row 3 — ReportSheetChanged
    ("TC-RPT-106", "reports", "TS-Reports-CronNotifications", "Report-sheet-changed — employees with modified reports receive REPORT_SHEET_CHANGED email", "Row 3 happy-path; delta #1 folded (template key REPORT_SHEET_CHANGED, not TASK_REPORT_CHANGED)", "Critical"),
    ("TC-RPT-107", "reports", "TS-Reports-CronNotifications", "Report-sheet-changed — email subject matches REPORT_SHEET_CHANGED template per delta #1", "Row 3 template verification per delta #1", "High"),
    ("TC-RPT-108", "reports", "TS-Reports-CronNotifications", "Report-sheet-changed — idempotent on repeated trigger within same window", "Row 3 idempotency", "Medium"),
    # Row 4 — ReportReject (debounce-protected, zero log markers per delta #2)
    ("TC-RPT-109", "reports", "TS-Reports-CronNotifications", "Report-rejection notification — rejected reports trigger APPROVE_REJECT email to employee", "Row 4 happy-path; email-only verification per delta #2 (zero log markers)", "Critical"),
    ("TC-RPT-110", "reports", "TS-Reports-CronNotifications", "Report-rejection notification — DEBOUNCE_INTERVAL_MINUTES=5 suppresses repeat emails within window", "Row 4 debounce guard; timing-sensitive", "High"),
    ("TC-RPT-111", "reports", "TS-Reports-CronNotifications", "Report-rejection notification — closed report month + open confirmation period still dispatches reject email (#3321)", "Row 4 regression; #3321 folded", "High"),
    ("TC-RPT-112", "reports", "TS-Reports-CronNotifications", "Report-rejection notification — officeId=9 employees excluded per #685", "Row 4 negative; office-filter per #685", "Medium"),
    # Row 7 — CleanupExtendedPeriod (accountant-exclusion per #2289)
    ("TC-RPT-113", "reports", "TS-Reports-CronNotifications", "Cleanup-extended-period — expired extended-period approvals cleared for non-accountant employees", "Row 7 happy-path; accountant-exclusion folded per #2289", "Critical"),
    ("TC-RPT-114", "reports", "TS-Reports-CronNotifications", "Cleanup-extended-period — accountant-role employees not cleaned up per #2289", "Row 7 negative; accountant-exclusion regression per #2289", "High"),
    ("TC-RPT-115", "reports", "TS-Reports-CronNotifications", "Cleanup-extended-period — idempotent on repeated trigger within same day", "Row 7 idempotency", "Medium"),
    # Row 5 — BudgetNotification (three templates per delta #3)
    ("TC-RPT-116", "reports", "TS-Reports-BudgetNotifications", "Budget-notification — BUDGET_NOTIFICATION_EXCEEDED dispatched when project budget crossed above threshold", "Row 5 EXCEEDED path; delta #3 folded (3 templates, not 1)", "Critical"),
    ("TC-RPT-117", "reports", "TS-Reports-BudgetNotifications", "Budget-notification — BUDGET_NOTIFICATION_NOT_REACHED dispatched when budget not met by deadline", "Row 5 NOT_REACHED path; delta #3", "Critical"),
    ("TC-RPT-118", "reports", "TS-Reports-BudgetNotifications", "Budget-notification — BUDGET_NOTIFICATION_DATE_UPDATED dispatched when budget deadline moved", "Row 5 DATE_UPDATED path; delta #3", "High"),
    ("TC-RPT-119", "reports", "TS-Reports-BudgetNotifications", "Budget-notification — SAFETY_INTERVAL_SECONDS=10 suppresses rapid-fire dispatch", "Row 5 safety-interval guard per #892", "High"),
    ("TC-RPT-120", "reports", "TS-Reports-BudgetNotifications", "Budget-notification — idempotent on repeated trigger within same window", "Row 5 idempotency", "Medium"),
    # Row 6 — CS sync (ttt-service)
    ("TC-CS-101", "cross-service", "TS-CrossService-CronCSSync", "CS sync (ttt) — cron fires every 15 min and emits started/finished markers", "Row 6 happy-path; baseline partial sync", "Critical"),
    ("TC-CS-102", "cross-service", "TS-CrossService-CronCSSync", "CS sync (ttt) — Unleash flag CS_SYNC-qa-1 OFF makes launcher skip silently", "Row 6 negative; feature-toggle gated silence", "High"),
    ("TC-CS-103", "cross-service", "TS-CrossService-CronCSSync", "CS sync (ttt) — full sync fires at application startup only (no cron path)", "Row 6 startup-only full sync; delta #10 folded", "High"),
    # Row 10 — CS sync (vacation-service)
    ("TC-CS-104", "cross-service", "TS-CrossService-CronCSSync", "CS sync (vacation) — cron fires and emits distinct `CS sync started/finished` markers", "Row 10 happy-path; design asymmetry (different marker text vs rows 6/20)", "Critical"),
    ("TC-CS-105", "cross-service", "TS-CrossService-CronCSSync", "CS sync (vacation) — failure path logs at WARN level (not ERROR)", "Row 10 design issue; WARN-level regression guard", "Medium"),
    # Row 20 — CS sync (calendar-service)
    ("TC-CS-106", "cross-service", "TS-CrossService-CronCSSync", "CS sync (calendar) — fullSync=false via v2 test endpoint runs partial sync", "Row 20 happy-path partial; delta #7 endpoint folded", "Critical"),
    ("TC-CS-107", "cross-service", "TS-CrossService-CronCSSync", "CS sync (calendar) — fullSync=true via v2 test endpoint runs full sync", "Row 20 full sync via v2 endpoint; delta #7", "High"),
    # Rows 6+20 shared — marker collision + ShedLock + parallel execution
    ("TC-CS-108", "cross-service", "TS-CrossService-CronCSSync", "CS sync — marker collision: rows 6 and 20 both emit same text (disambiguate by stream)", "Shared rows 6+20; Graylog stream field disambiguation", "High"),
    ("TC-CS-109", "cross-service", "TS-CrossService-CronCSSync", "CS sync — ShedLock: concurrent pods do not double-execute within lock window", "Shared rows 6/10/20; ShedLock distributed-lock guard", "High"),
    ("TC-CS-110", "cross-service", "TS-CrossService-CronCSSync", "CS sync — parallel execution across services: ttt + vacation + calendar run concurrently", "Shared rows 6/10/20; per-service independence", "Medium"),
    ("TC-CS-111", "cross-service", "TS-CrossService-CronCSSync", "CS sync — idempotency: re-trigger within lock window has no effect", "Shared rows 6/10/20 idempotency", "Medium"),
    # Row 23 — PM Tool sync
    ("TC-CS-112", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — new project in PMT propagates to TTT with all synced fields", "Row 23 happy-path; #3083 contract baseline", "Critical"),
    ("TC-CS-113", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — existing project update: name/customer/status/supervisor fields mirrored", "Row 23 update flow; #3083 field contract", "Critical"),
    ("TC-CS-114", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — presales merge is APPEND-ONLY (existing presales preserved)", "Row 23 regression; #3382 append-only merge", "Critical"),
    ("TC-CS-115", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — accounting_name is IMMUTABLE (PMT rename ignored)", "Row 23 regression; #3286 accounting-name immutability", "High"),
    ("TC-CS-116", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — default tracker-script auto-populated without history event", "Row 23 regression; #3083 note on silent default populate", "High"),
    ("TC-CS-117", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — PMT-owned field edits do NOT produce project_event rows", "Row 23 event-history contract; PMT-owned silent writes", "High"),
    ("TC-CS-118", "cross-service", "TS-CrossService-CronPMToolSync", "TTT-owned field edit (tracker script) via UI produces project_event row (#3083 note 4)", "Row 23 regression; snavrockiy fix 2026-02-25", "High"),
    ("TC-CS-119", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — cs-id validation failure: unknown cs-id logged and skipped", "Row 23 negative; cs-id FK guard", "High"),
    ("TC-CS-120", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — Unleash `PM_TOOL_SYNC-qa-1` OFF: scheduler silent", "Row 23 negative; feature-toggle gated silence", "High"),
    ("TC-CS-121", "cross-service", "TS-CrossService-CronPMToolSync", "PM Tool sync — startup full sync via TttStartupApplicationListener", "Row 23 startup-only full sync; #3399 startup listener", "High"),
    # Row 22 — Statistic Report periodic sync (ttt-service)
    ("TC-STAT-077", "statistics", "TS-Stat-CronStatReportSync", "StatisticReportScheduler — cron fires at 04:00 NSK, ShedLock acquired/released, start/finish markers emit", "Row 22 happy-path; baseline periodic sync with ShedLock", "Critical"),
    ("TC-STAT-078", "statistics", "TS-Stat-CronStatReportSync", "Delta #8 regression — statistic-report sync failure logged at INFO (level:6), NOT ERROR (level:3)", "Row 22 design issue; delta #8 log-level regression guard", "High"),
    ("TC-STAT-079", "statistics", "TS-Stat-CronStatReportSync", "#3345 Bug 1 regression — employment-period filter excludes pre-hire and post-leave months", "Row 22 regression; #3345 MR !5101 fix", "Critical"),
    ("TC-STAT-080", "statistics", "TS-Stat-CronStatReportSync", "#3345 Bug 2 regression — day-off reschedule triggers month_norm recalc via RabbitMQ event", "Row 22 regression; #3345 event-fan-out fix", "High"),
    ("TC-STAT-081", "statistics", "TS-Stat-CronStatReportSync", "#3337 regression — sick-leave creation updates month_norm and reported_effort via SICK_LEAVE_CHANGES event", "Row 22 regression; #3337 event-enum broadening", "High"),
    ("TC-STAT-082", "statistics", "TS-Stat-CronStatReportSync", "#3337 regression — scoped event for employee A does NOT delete statistic_report rows for employee B", "Row 22 regression; #3337 scoped-delete fix", "High"),
    ("TC-STAT-083", "statistics", "TS-Stat-CronStatReportSync", "#3346 regression (bug #895498) — manual statistic_report row delete is restored on next optimized sync", "Row 22 regression; #3346 scheduler-wiring fix", "Critical"),
    ("TC-STAT-084", "statistics", "TS-Stat-CronStatReportSync", "Full vs optimized sync contract — full refreshes previous + current year; optimized refreshes previous + current month only", "Row 22 endpoint-contract; #3345 note 894873", "Medium"),
    # Row 8 — EmailSendScheduler (20s dispatcher)
    ("TC-EMAIL-001", "email", "TS-Email-CronDispatch", "EmailSendScheduler — cron fires every 20 s, ShedLock acquired/released, start/finish markers emit", "Row 8 happy-path; baseline dispatcher cadence + lock", "Critical"),
    ("TC-EMAIL-002", "email", "TS-Email-CronDispatch", "Dispatch — NEW email transitions to SENT after SMTP success (happy path)", "Row 8 SMTP-success status transition", "Critical"),
    ("TC-EMAIL-003", "email", "TS-Email-CronDispatch", "Dispatch — invalid recipient transitions NEW → INVALID (SendFailedException path)", "Row 8 invalid-recipient status transition", "High"),
    ("TC-EMAIL-004", "email", "TS-Email-CronDispatch", "Dispatch — partial SMTP failure sets FAILED for rejected messages, SENT for accepted", "Row 8 mixed-batch status transitions; DI-EMAIL-DISPATCH-RETRY context", "High"),
    ("TC-EMAIL-005", "email", "TS-Email-CronDispatch", "DI-EMAIL-DISPATCH-AUTH regression — MailAuthenticationException caught, status NOT updated, infinite retry loop", "Row 8 design-issue regression guard; stuck-as-NEW infinite loop", "High"),
    ("TC-EMAIL-006", "email", "TS-Email-CronDispatch", "Dispatch — pageSize = 300 caps per-batch dispatch; overflow drains in subsequent ticks", "Row 8 batch-size contract; pagination window", "Medium"),
    # Row 9 — EmailPruneScheduler (daily retention)
    ("TC-EMAIL-007", "email", "TS-Email-CronPrune", "EmailPruneScheduler — cron fires daily at 00:00, ShedLock acquired/released, start/finish markers emit", "Row 9 happy-path; baseline retention scheduler", "Critical"),
    ("TC-EMAIL-008", "email", "TS-Email-CronPrune", "Retention — emails older than 30 days deleted, emails newer than 30 days preserved", "Row 9 retention baseline; PT30D cutoff", "Critical"),
    ("TC-EMAIL-009", "email", "TS-Email-CronPrune", "Retention boundary — email exactly at 30-day cutoff behavior (strict less-than: <30d preserved, =30d deleted)", "Row 9 off-by-one guard on strict less-than ADD_TIME.lessThan", "High"),
    ("TC-EMAIL-010", "email", "TS-Email-CronPrune", "Retention — attachments deleted with parent email (no FK orphans)", "Row 9 cascade-delete contract; zero orphans", "High"),
    ("TC-EMAIL-011", "email", "TS-Email-CronPrune", "Retention no-op — all emails within retention window; finish marker reports 0 deleted", "Row 9 no-op idempotency; zero-delete path", "Medium"),
]


def _clear_data_rows(ws, header_row: int) -> None:
    """Wipe all rows below header to support idempotent re-runs."""
    last = ws.max_row
    if last > header_row:
        ws.delete_rows(header_row + 1, last - header_row)


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER


def _style_data(cell, *, zebra: bool) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
    if zebra:
        cell.fill = PatternFill("solid", fgColor="F2F2F2")


def populate() -> None:
    wb = load_workbook(CRON_XLSX)
    ws = wb["COL-cron"]

    # Header is at row 4 (row 1 = hyperlink, row 2 = title, row 3 = empty, row 4 = header).
    header_row = 4
    _clear_data_rows(ws, header_row)

    # Refresh header styling.
    for col, label in enumerate(
        ["test_id", "source_module", "source_suite", "title", "inclusion_reason", "priority_override"],
        start=1,
    ):
        c = ws.cell(row=header_row, column=col, value=label)
        _style_header(c)

    # Row 2 title update — flip from scaffold note to populated summary.
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = (
        f"COL-cron — Curated Test Collection (complete; {len(ROWS)} TCs referenced; "
        "vacation s135 + reports s136 + cross-service s137 + statistics + email s138)"
    )
    title_cell.font = Font(bold=True, size=12)

    # Data rows.
    widths = [14, 16, 28, 70, 48, 14]
    for idx, col_w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = col_w

    for i, row in enumerate(ROWS, start=1):
        r = header_row + i
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            _style_data(c, zebra=(i % 2 == 0))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:F{header_row + len(ROWS)}"

    wb.save(CRON_XLSX)
    print(f"Populated: {CRON_XLSX}")
    print(f"Rows written: {len(ROWS)}")


if __name__ == "__main__":
    populate()
