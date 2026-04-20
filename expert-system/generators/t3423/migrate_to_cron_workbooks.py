#!/usr/bin/env python3
"""One-shot migration: extract cron suites from home workbooks into dedicated
per-domain Cron_<Domain>.xlsx files under test-docs/collections/cron/.

Produces:
  test-docs/collections/cron/Cron_Vacation.xlsx      (8 suites · 27 TCs)
  test-docs/collections/cron/Cron_Reports.xlsx       (2 suites · 20 TCs)
  test-docs/collections/cron/Cron_CrossService.xlsx  (2 suites · 21 TCs)
  test-docs/collections/cron/Cron_Statistics.xlsx    (1 suite  ·  8 TCs)
  test-docs/collections/cron/Cron_Email.xlsx         (2 suites · 11 TCs + lifted Plan/Matrix/Risk)

Mutates:
  test-docs/{vacation,reports,cross-service,statistics}/<domain>.xlsx
    — TS-*-Cron-* sheets deleted.

Deletes:
  test-docs/email/                                   (whole dir — all content was cron)

Run from repo root:
  python3 expert-system/generators/t3423/migrate_to_cron_workbooks.py
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
COLLECTION_DIR = REPO_ROOT / "test-docs" / "collections" / "cron"

# domain -> (home workbook path, list of cron sheet names, dst file name, cron-row-number list, TC count)
DOMAINS = {
    "vacation": {
        "home": REPO_ROOT / "test-docs" / "vacation" / "vacation.xlsx",
        "dst": COLLECTION_DIR / "Cron_Vacation.xlsx",
        "cron_sheets": [
            "TS-Vac-Cron-AnnualAccruals",
            "TS-Vac-Cron-NotImpl",
            "TS-Vac-Cron-Digest",
            "TS-Vac-Cron-CalendarReminder",
            "TS-Vac-Cron-AutoPay",
            "TS-Vac-Cron-ApprovedToPaid",
            "TS-Vac-Cron-EmpProjectSync",
            "TS-Vac-Cron-StatReportInit",
        ],
        "rows_covered": "11, 12, 13, 14, 15, 16, 17, 18, 19, 21",
        "tc_count": 27,
        "suite_row_map": [
            ("TS-Vac-Cron-AnnualAccruals", "11", 3),
            ("TS-Vac-Cron-NotImpl", "12, 13 (NOT_IMPLEMENTED — dead config)", 2),
            ("TS-Vac-Cron-Digest", "14", 3),
            ("TS-Vac-Cron-CalendarReminder", "15", 3),
            ("TS-Vac-Cron-AutoPay", "16", 3),
            ("TS-Vac-Cron-ApprovedToPaid", "17", 3),
            ("TS-Vac-Cron-EmpProjectSync", "18, 19", 8),
            ("TS-Vac-Cron-StatReportInit", "21", 2),
        ],
    },
    "reports": {
        "home": REPO_ROOT / "test-docs" / "reports" / "reports.xlsx",
        "dst": COLLECTION_DIR / "Cron_Reports.xlsx",
        "cron_sheets": [
            "TS-Reports-CronNotifications",
            "TS-Reports-BudgetNotifications",
        ],
        "rows_covered": "1, 2, 3, 4, 5, 7",
        "tc_count": 20,
        "suite_row_map": [
            ("TS-Reports-CronNotifications", "1, 2, 3, 4, 7", 15),
            ("TS-Reports-BudgetNotifications", "5", 5),
        ],
    },
    "cross-service": {
        "home": REPO_ROOT / "test-docs" / "cross-service" / "cross-service.xlsx",
        "dst": COLLECTION_DIR / "Cron_CrossService.xlsx",
        "cron_sheets": [
            "TS-CrossService-CronCSSync",
            "TS-CrossService-CronPMToolSync",
        ],
        "rows_covered": "6, 10, 20, 23",
        "tc_count": 21,
        "suite_row_map": [
            ("TS-CrossService-CronCSSync", "6, 10, 20", 11),
            ("TS-CrossService-CronPMToolSync", "23", 10),
        ],
    },
    "statistics": {
        "home": REPO_ROOT / "test-docs" / "statistics" / "statistics.xlsx",
        "dst": COLLECTION_DIR / "Cron_Statistics.xlsx",
        "cron_sheets": [
            "TS-Stat-CronStatReportSync",
        ],
        "rows_covered": "22",
        "tc_count": 8,
        "suite_row_map": [
            ("TS-Stat-CronStatReportSync", "22", 8),
        ],
    },
    "email": {
        "home": REPO_ROOT / "test-docs" / "email" / "email.xlsx",
        "dst": COLLECTION_DIR / "Cron_Email.xlsx",
        "cron_sheets": [
            "TS-Email-CronDispatch",
            "TS-Email-CronPrune",
        ],
        "rows_covered": "8, 9",
        "tc_count": 11,
        "suite_row_map": [
            ("TS-Email-CronDispatch", "8", 6),
            ("TS-Email-CronPrune", "9", 5),
        ],
        # email is the only workbook that was 100% cron — lift all meta-sheets as-is.
        "lift_meta_sheets": True,
    },
}


def copy_cells(src_ws, dst_ws):
    """Copy all cells, including styles / hyperlinks / comments, from src_ws to dst_ws."""
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
            if cell.hyperlink is not None:
                new_cell.hyperlink = copy(cell.hyperlink)
            if cell.comment is not None:
                new_cell.comment = copy(cell.comment)


def copy_sheet(src_ws, dst_wb, new_title=None):
    """Full cross-workbook sheet copy — cells, styles, tab color, column widths, row heights, merged cells, freeze panes."""
    title = new_title or src_ws.title
    dst_ws = dst_wb.create_sheet(title=title)

    # Tab color
    if src_ws.sheet_properties.tabColor is not None:
        dst_ws.sheet_properties.tabColor = copy(src_ws.sheet_properties.tabColor)

    copy_cells(src_ws, dst_ws)

    # Column dimensions
    for letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[letter].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[letter].hidden = dim.hidden

    # Row dimensions
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[row_num].height = dim.height

    # Merged cells
    for mr in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(mr))

    # Freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    return dst_ws


def author_plan_overview(dst_wb, domain: str, info: dict):
    """Create the Plan Overview sheet in a fresh Cron_<Domain> workbook."""
    ws = dst_wb.create_sheet(title="Plan Overview")
    ws.sheet_properties.tabColor = "0070AD47"  # green (matches home-workbook meta tabs)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    bold = Font(bold=True, size=12)
    h1 = Font(bold=True, size=16, color="1F4E78")
    h2 = Font(bold=True, size=13, color="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    top = Alignment(vertical="top", wrap_text=True)

    def put(row, col, val, *, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align

    r = 1
    put(r, 1, f"Cron_{_camel(domain)} — Test Plan Overview", font=h1)
    r += 2

    put(r, 1, "Ticket", font=bold)
    put(r, 2, "#3423 — Cron & Startup Jobs Testing Collection (epic: #3402, linked via relates_to)")
    r += 1
    put(r, 1, "Domain", font=bold)
    put(r, 2, _camel(domain))
    r += 1
    put(r, 1, "Collection", font=bold)
    put(r, 2, "cron (87 TCs total across 5 domains · see COL-cron in cron.xlsx)")
    r += 1
    put(r, 1, "Cron rows covered", font=bold)
    put(r, 2, info["rows_covered"], align=top)
    r += 1
    put(r, 1, "TC count", font=bold)
    put(r, 2, str(info["tc_count"]))
    r += 2

    put(r, 1, "Test Suites", font=h2)
    r += 1
    put(r, 1, "Suite", font=bold, fill=header_fill)
    put(r, 2, "Cron row(s) / TC count", font=bold, fill=header_fill)
    r += 1
    for suite_name, rows, tcs in info["suite_row_map"]:
        put(r, 1, suite_name)
        put(r, 2, f"row(s) {rows} · {tcs} TC" + ("s" if tcs != 1 else ""))
        r += 1
    r += 1

    put(r, 1, "What's NOT here", font=h2)
    r += 1
    not_here_msg = (
        f"Non-cron TCs for the {_camel(domain)} domain live in the home workbook "
        f"test-docs/{domain}/{domain}.xlsx. This file holds only cron-scheduled behaviour."
    )
    if domain == "email":
        not_here_msg = (
            "This workbook contains the entire email-service test documentation. "
            "The email service is a dispatcher/retention layer — all of its behaviour is cron-driven "
            "(dispatch every 20s, prune daily 00:00). There is no separate non-cron email workbook; "
            "non-cron email flows (report notifications, vacation digest, budget alerts) live in the "
            "business-module workbooks that trigger them."
        )
    put(r, 1, not_here_msg, align=top)
    ws.row_dimensions[r].height = 60
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    put(r, 1, "Related files", font=h2)
    r += 1
    put(r, 1, "Task definition")
    put(r, 2, "docs/tasks/cron/cron-testing-task.md")
    r += 1
    put(r, 1, "Execution report (Phases A + B)")
    put(r, 2, "docs/tasks/cron/execution-phases-a-b.md")
    r += 1
    put(r, 1, "Collection reference")
    put(r, 2, "test-docs/collections/cron/cron.xlsx (COL-cron sheet · 87 TC refs)")
    r += 1
    put(r, 1, "Traceability matrix")
    put(r, 2, "test-docs/collections/cron/coverage.md")
    r += 1
    put(r, 1, "Prior-art vault notes")
    put(r, 2, "expert-system/vault/external/EXT-cron-jobs.md, exploration/api-findings/cron-job-live-verification.md, patterns/email-notification-triggers.md", align=top)
    ws.row_dimensions[r].height = 45
    r += 2

    put(r, 1, "Generator script", font=bold)
    put(r, 2, f"expert-system/generators/t3423/generate_cron_{_snake(domain)}.py  (idempotent wipe-and-rewrite)")
    r += 1
    put(r, 1, "Last migrated", font=bold)
    put(r, 2, "2026-04-20 by migrate_to_cron_workbooks.py (one-shot extraction from home workbook)")


def _camel(domain: str) -> str:
    if domain == "cross-service":
        return "CrossService"
    return domain.capitalize()


def _snake(domain: str) -> str:
    return domain.replace("-", "")


def migrate_domain(domain: str, info: dict) -> dict:
    src_path = info["home"]
    dst_path = info["dst"]

    if not src_path.exists():
        return {"domain": domain, "status": "skipped", "reason": f"home workbook missing: {src_path}"}

    src_wb = load_workbook(src_path)
    src_sheet_count_before = len(src_wb.sheetnames)

    # Build destination workbook
    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)  # drop default blank sheet

    if info.get("lift_meta_sheets"):
        # Email: lift Plan Overview + Feature Matrix + Risk Assessment as-is
        for meta in ("Plan Overview", "Feature Matrix", "Risk Assessment"):
            if meta in src_wb.sheetnames:
                copy_sheet(src_wb[meta], dst_wb, meta)
    else:
        author_plan_overview(dst_wb, domain, info)

    # Copy cron sheets
    missing = [s for s in info["cron_sheets"] if s not in src_wb.sheetnames]
    if missing:
        return {"domain": domain, "status": "error", "reason": f"cron sheets missing from home workbook: {missing}"}

    for cron_sheet_name in info["cron_sheets"]:
        copy_sheet(src_wb[cron_sheet_name], dst_wb, cron_sheet_name)

    dst_wb.save(dst_path)
    dst_wb_verify = load_workbook(dst_path, read_only=True)
    dst_sheet_count = len(dst_wb_verify.sheetnames)
    dst_wb_verify.close()

    # Strip cron sheets from home workbook (except email — whole dir gets deleted later)
    if domain != "email":
        for cron_sheet_name in info["cron_sheets"]:
            del src_wb[cron_sheet_name]
        src_wb.save(src_path)

    src_sheet_count_after = len(src_wb.sheetnames) if domain != "email" else src_sheet_count_before

    return {
        "domain": domain,
        "status": "ok",
        "src_sheets_before": src_sheet_count_before,
        "src_sheets_after": src_sheet_count_after,
        "dst_path": str(dst_path.relative_to(REPO_ROOT)),
        "dst_sheets": dst_sheet_count,
    }


def main():
    print("=== Cron-workbook migration ===\n")
    COLLECTION_DIR.mkdir(parents=True, exist_ok=True)

    results = []
    for domain in ["vacation", "reports", "cross-service", "statistics", "email"]:
        info = DOMAINS[domain]
        result = migrate_domain(domain, info)
        results.append(result)
        if result["status"] == "ok":
            print(
                f"[{domain}] OK — home {result['src_sheets_before']}→{result['src_sheets_after']} sheets, "
                f"dst {result['dst_path']} ({result['dst_sheets']} sheets)"
            )
        else:
            print(f"[{domain}] {result['status'].upper()} — {result.get('reason', '')}")

    # Delete test-docs/email/ only if Cron_Email.xlsx was written successfully
    email_result = next(r for r in results if r["domain"] == "email")
    cron_email_path = COLLECTION_DIR / "Cron_Email.xlsx"
    if email_result["status"] == "ok" and cron_email_path.exists():
        email_dir = REPO_ROOT / "test-docs" / "email"
        if email_dir.exists():
            shutil.rmtree(email_dir)
            print(f"\nDeleted {email_dir.relative_to(REPO_ROOT)}/ (all content migrated to Cron_Email.xlsx)")

    print("\n=== Done ===")
    return 0 if all(r["status"] == "ok" for r in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
