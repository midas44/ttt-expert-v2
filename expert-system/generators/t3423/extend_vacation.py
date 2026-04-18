#!/usr/bin/env python3
"""
Ticket #3423 — Vacation Cluster Cron Test Cases (Phase B, session 135+).

Extends ``test-docs/vacation/vacation.xlsx`` with 8 new cron-focused suites
covering jobs 11, 12, 13, 14, 15, 16, 17, 18, 19, 21 of the cron-testing
collection. Test IDs continue the home-module sequence — existing vacation
workbook ends at TC-VAC-100; new cron TCs start at TC-VAC-101 and run through
TC-VAC-127.

Idempotent: existing ``TS-Vac-Cron-*`` sheets are removed before re-adding.
This script does NOT touch the Plan Overview / Feature Matrix / Risk
Assessment tabs — only the cron cluster suites. (Those tabs still accurately
describe the core module; cron notifications are additive.)

Run from repo root:
    python3 expert-system/generators/t3423/extend_vacation.py

Exit 0 on success. Prints a summary of the test IDs written.

Canonical references:
    - expert-system/vault/exploration/tickets/t3423-investigation.md (preamble)
    - expert-system/vault/exploration/tickets/3262-ticket-findings.md (seed TCs for jobs 18/19/21/22)
    - expert-system/vault/external/EXT-cron-jobs.md (markers, YAML properties)
    - expert-system/vault/patterns/email-notification-triggers.md (subject formats)
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_VAC_XLSX = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "vacation", "vacation.xlsx")
)


# ─── Styling (matches existing vacation.xlsx rows) ────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR = "F4B084"  # Orange — distinguishes cron suites from core-module suites

COL_WIDTHS = [14, 40, 45, 70, 45, 10, 10, 35, 22, 35]
HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
    "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
]


# ─── Test case data ───────────────────────────────────────────────────────────
#
# Tuples: (test_id, title, preconditions, steps, expected, priority, type_,
#          req_ref, module, notes)
#
# Steps follow the cron verification shape from t3423-investigation.md §Verification shape:
#   SETUP → Clock → Trigger → Wait → Verify (DB/UI/Email/Log) → CLEANUP
# API-triggered steps are the exception per CLAUDE.md §11 (test endpoints have
# no UI equivalent); UI steps appear where a user-visible effect exists.

TS_ANNUAL_ACCRUALS = [
    (
        "TC-VAC-101",
        "Annual accruals cron — happy path on Jan 1",
        (
            "Env: ttt-timemachine (test-clock available).\n"
            "At least 3 enabled employees with open employment periods and "
            "ttt_vacation.employee_vacation_days rows for the PREVIOUS year.\n"
            "Query: SELECT login FROM employee WHERE enabled=true "
            "AND EXISTS (SELECT 1 FROM employee_period WHERE employee_id = employee.id "
            "AND start_date <= CURRENT_DATE AND (end_date IS NULL OR end_date > CURRENT_DATE)) "
            "ORDER BY random() LIMIT 3;"
        ),
        (
            "SETUP: Record baseline days-totals — "
            "SELECT employee_id, year, days FROM ttt_vacation.employee_vacation_days "
            "WHERE year = EXTRACT(year FROM CURRENT_DATE) + 1 ORDER BY employee_id.\n"
            "SETUP: Advance test clock to Jan 1, 00:05 NSK via "
            "PATCH /api/ttt/v1/test/clock { epochMillis: <Jan 1 00:05 local epoch> }.\n"
            "TRIGGER: POST /api/vacation/v1/test/annual-accruals (no body).\n"
            "WAIT: 5 s — AnnualAccrualsTask is synchronous; no fan-out.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE stream, query "
            "`message:\"Starting AnnualAccrualsTask\"` within last 1 min — exactly 1 hit.\n"
            "DB-CHECK: SELECT year, COUNT(*) FROM ttt_vacation.employee_vacation_days "
            "WHERE year = EXTRACT(year FROM CURRENT_DATE) GROUP BY year — row count "
            "matches enabled-employee count.\n"
            "DB-CHECK: Compare per-employee days delta vs employment-period coefficient "
            "(standard accrual formula).\n"
            "CLEANUP: POST /api/ttt/v1/test/clock/reset. Delete inserted rows only if "
            "they exceed the env's historical year coverage (usually skip — newly "
            "accrued year is reusable)."
        ),
        (
            "Marker `\"Starting AnnualAccrualsTask\"` fires once. New rows inserted into "
            "ttt_vacation.employee_vacation_days for the new year, one row per enabled "
            "employee. Days values match the accrual formula against employment period. "
            "No error markers."
        ),
        "High", "Hybrid",
        "#3423 row 11; EXT-cron-jobs.md §jobs-10-19/21",
        "vacation/cron/accruals",
        "Silent-failure pattern documented (Bug — no finish/error marker). Success is inferred from DB delta, not logs."
    ),
    (
        "TC-VAC-102",
        "Annual accruals idempotency — no double-accrual on same year",
        (
            "Env: ttt-timemachine. Annual accrual already ran for current year "
            "(e.g. TC-VAC-101 executed).\n"
            "Query: SELECT year, COUNT(*) FROM ttt_vacation.employee_vacation_days "
            "WHERE year = EXTRACT(year FROM CURRENT_DATE) GROUP BY year; "
            "rows > 0 → idempotency preconditions satisfied."
        ),
        (
            "SETUP: Capture row count for the current year in "
            "ttt_vacation.employee_vacation_days.\n"
            "TRIGGER: POST /api/vacation/v1/test/annual-accruals a second time.\n"
            "WAIT: 5 s.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — marker fires again "
            "(AnnualAccrualsTask has no guard).\n"
            "DB-CHECK: SELECT COUNT(*) FROM ttt_vacation.employee_vacation_days "
            "WHERE year = EXTRACT(year FROM CURRENT_DATE); row count unchanged vs baseline.\n"
            "DB-CHECK: Per-employee `days` column unchanged (no double-increment).\n"
            "CLEANUP: None."
        ),
        (
            "Second trigger logs the start marker but does not insert duplicate rows or "
            "increment days values. The task's upsert / row-existence check prevents "
            "double accrual within the same year even though the marker fires again."
        ),
        "High", "Hybrid",
        "#3423 row 11",
        "vacation/cron/accruals",
        "Idempotency is data-driven, not marker-driven."
    ),
    (
        "TC-VAC-103",
        "Annual accruals silent-failure regression — no finish/error marker",
        (
            "Env: ttt-timemachine. One employee row is deliberately corrupted "
            "(e.g. NULL employment_type) to force a per-row accrual failure.\n"
            "SETUP SQL: UPDATE employee SET employment_type = NULL "
            "WHERE login = '<seed_login>' — OR similar row-level constraint break.\n"
            "(Alternative: inject failure at accrual-service level via a feature "
            "toggle if available on env.)"
        ),
        (
            "SETUP: Break one employee row (see preconditions).\n"
            "TRIGGER: POST /api/vacation/v1/test/annual-accruals.\n"
            "WAIT: 5 s.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — `\"Starting AnnualAccrualsTask\"` "
            "fires; NO `\"finished\"` or `\"failed\"` marker (confirmed Bug — silent "
            "failure).\n"
            "DB-CHECK: ttt_vacation.employee_vacation_days has rows for all OTHER "
            "employees (partial success), but no row for the broken employee.\n"
            "CLEANUP: Restore employee row (e.g. UPDATE employee SET employment_type='FULL' "
            "WHERE login='<seed_login>')."
        ),
        (
            "Task logs start, silently swallows the per-row exception, inserts rows "
            "for the successful employees, does not log a finish or error marker, and "
            "leaves the broken employee's row absent. This is documented behaviour "
            "(AnnualAccrualsTask has no try/catch → finish-marker path)."
        ),
        "Medium", "Hybrid",
        "#3423 row 11 — known silent-failure limitation",
        "vacation/cron/accruals",
        "Regression ensures the silent-failure behaviour does not regress to a crash."
    ),
]

TS_NOT_IMPL = [
    (
        "TC-VAC-104",
        "Preliminary-vacation delete-expired endpoint is a no-op (dead config)",
        (
            "Env: qa-1. No preliminary vacations required.\n"
            "Job 12 is NOT_IMPLEMENTED — application.yml declares "
            "`preliminary-outdated.cron` but no Java code subscribes."
        ),
        (
            "SETUP: Snapshot vacation rows — "
            "SELECT COUNT(*) FROM ttt_vacation.vacation WHERE status = 'PRELIMINARY';\n"
            "TRIGGER: POST /api/vacation/v1/test/vacations/delete-expired-preliminary.\n"
            "WAIT: 2 s.\n"
            "DB-CHECK: Row count unchanged.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — no markers referencing "
            "`preliminary-outdated`, no errors.\n"
            "CLEANUP: None."
        ),
        (
            "Endpoint returns HTTP 200 with no side effects. No vacation rows are "
            "removed. No log markers are emitted (no scheduler exists). Confirms dead "
            "YAML config."
        ),
        "Low", "API",
        "#3423 row 12 — NOT_IMPLEMENTED",
        "vacation/cron/not-implemented",
        "Single no-op stub documenting dead config. Re-run if `preliminary-outdated.cron` is ever wired up."
    ),
    (
        "TC-VAC-105",
        "Preliminary-vacation close-outdated endpoint is a no-op (dead config)",
        (
            "Env: qa-1. No preliminary vacations required.\n"
            "Job 13 is NOT_IMPLEMENTED — application.yml declares "
            "`close-outdated.cron` but no Java code subscribes."
        ),
        (
            "SETUP: Snapshot vacation rows — "
            "SELECT COUNT(*) FROM ttt_vacation.vacation WHERE status = 'APPROVED';\n"
            "TRIGGER: POST /api/vacation/v1/test/vacations/close-outdated.\n"
            "WAIT: 2 s.\n"
            "DB-CHECK: Row count unchanged.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — no `close-outdated` markers, no errors.\n"
            "CLEANUP: None."
        ),
        (
            "Endpoint returns HTTP 200 with no side effects. Confirms dead YAML config. "
            "Note: lock name `CloseOutdatedTask.run` is REUSED by job 16 "
            "(auto-pay) via legacy naming — do not confuse."
        ),
        "Low", "API",
        "#3423 row 13 — NOT_IMPLEMENTED",
        "vacation/cron/not-implemented",
        "Bug #3423-documentation: scope table mislabels this as an active cron."
    ),
]

TS_DIGEST = [
    (
        "TC-VAC-106",
        "Vacation digest — employee with tomorrow's vacation receives email",
        (
            "Env: qa-1. One employee with an APPROVED vacation starting the next "
            "business day.\n"
            "Query: SELECT e.login, v.id FROM ttt_vacation.vacation v "
            "JOIN employee e ON e.id = v.employee_id WHERE v.status = 'APPROVED' "
            "AND v.start_date = CURRENT_DATE + INTERVAL '1 day' "
            "ORDER BY random() LIMIT 1; "
            "If no row, seed via POST /api/vacation/v1/vacations with "
            "startDate=tomorrow, endDate=tomorrow+4d, type=REGULAR, "
            "then approve as manager."
        ),
        (
            "SETUP: Via API, create or identify an APPROVED vacation starting tomorrow.\n"
            "SETUP: Clear mailbox target — roundcube-access skill `delete --to <employee_email> "
            "--subject \"[QA1][TTT]\" --since today`.\n"
            "TRIGGER: POST /api/vacation/v1/test/digest.\n"
            "WAIT: 30 s — digest dispatch enqueues on Email batch (20 s dequeue).\n"
            "VERIFY EMAIL: roundcube-access skill `search --to <employee_email> "
            "--subject \"[QA1][TTT]\" --since today -n 5` → exactly 1 hit matching "
            "the vacation digest subject.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — `\"Vacation digest notification started\"` "
            "→ `\"...finished\"`.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — per-recipient `\"Mail has been sent to "
            "<employee_email> about NOTIFY_VACATION_UPCOMING...\"` marker.\n"
            "CLEANUP: Delete seeded vacation if created by this test. Restore mailbox."
        ),
        (
            "Exactly one digest email lands in the employee's mailbox within 30 s. "
            "Subject uses `[QA1][TTT]` prefix. Scheduler start & finish markers fire. "
            "Per-recipient mail-sent marker identifies the target login."
        ),
        "High", "Hybrid",
        "#3423 row 14; EXT-cron-jobs.md job 14",
        "vacation/cron/digest",
        "DELTA: endpoint path is /digest (not /vacations/notify). Scope-table bug."
    ),
    (
        "TC-VAC-107",
        "Vacation digest — no applicable vacations → no emails dispatched",
        (
            "Env: qa-1. No APPROVED vacations start within the digest's query window "
            "(tomorrow + 1 business day).\n"
            "SETUP SQL: DELETE or CANCEL any APPROVED vacations with "
            "start_date = CURRENT_DATE + INTERVAL '1 day' (or move the test clock)."
        ),
        (
            "SETUP: Ensure no APPROVED vacations start tomorrow (query above).\n"
            "SETUP: Capture baseline email count — roundcube-access skill "
            "`count --subject \"[QA1][TTT]\" --since today`.\n"
            "TRIGGER: POST /api/vacation/v1/test/digest.\n"
            "WAIT: 30 s.\n"
            "VERIFY EMAIL: email count unchanged (baseline == post-trigger count).\n"
            "VERIFY LOG: Graylog TTT-QA-1 — start marker fires; finish marker fires; "
            "no per-recipient `Mail has been sent to ...` markers for digest topic.\n"
            "CLEANUP: None."
        ),
        (
            "Start and finish markers emit (the task runs to completion) but no email "
            "is dispatched. Mailbox count is unchanged."
        ),
        "Medium", "Hybrid",
        "#3423 row 14",
        "vacation/cron/digest",
        "Regression for the 'empty digest' edge case — the task must still finish cleanly."
    ),
    (
        "TC-VAC-108",
        "Vacation digest Cyrillic env-prefix subject format",
        (
            "Env: qa-1. At least one applicable employee with tomorrow vacation (same "
            "seed as TC-VAC-106)."
        ),
        (
            "SETUP: Use the seed from TC-VAC-106 (or rerun it); do not clear mailbox.\n"
            "TRIGGER: POST /api/vacation/v1/test/digest.\n"
            "WAIT: 30 s.\n"
            "VERIFY EMAIL: roundcube-access skill `search --to <employee_email> "
            "--subject \"[QA1][TTT]\" --since today`; capture first match.\n"
            "VERIFY EMAIL: roundcube-access skill `read <uid>` — assert subject starts "
            "with literal ASCII `[QA1][TTT]` and NOT the Cyrillic variant reported in "
            "patterns/email-notification-triggers.md anomaly notes.\n"
            "CLEANUP: None."
        ),
        (
            "Subject prefix is the ASCII `[QA1][TTT]` form consistently. If Cyrillic "
            "prefix is observed, that is a known anomaly tracked in the email-triggers "
            "pattern note and should be reported as a defect."
        ),
        "Medium", "Hybrid",
        "#3423 row 14; patterns/email-notification-triggers.md (Cyrillic anomaly)",
        "vacation/cron/digest",
        "If anomaly observed, file a defect in Qase and link this TC."
    ),
]

TS_CAL_REMINDER = [
    (
        "TC-VAC-109",
        "Production-calendar annual reminder fires to chief accountants on Nov 1",
        (
            "Env: ttt-timemachine. At least 2 active ROLE_CHIEF_ACCOUNTANT employees.\n"
            "Query: SELECT e.login, e.email FROM employee e "
            "JOIN employee_role er ON er.employee_id = e.id "
            "JOIN role r ON r.id = er.role_id "
            "WHERE r.name = 'ROLE_CHIEF_ACCOUNTANT' AND e.enabled = true;"
        ),
        (
            "SETUP: Advance test clock to Oct 31, 23:59 NSK via "
            "PATCH /api/ttt/v1/test/clock { epochMillis: <Oct 31 23:59 local epoch> }.\n"
            "SETUP: Clear mailbox for each chief-accountant target.\n"
            "TRIGGER: POST /api/vacation/v1/test/production-calendars/send-first-reminder.\n"
            "WAIT: 30 s.\n"
            "VERIFY EMAIL: For each chief accountant, roundcube-access skill "
            "`search --to <email> --subject \"[TIMEMACHINE][TTT] Производственный календарь\" "
            "--since today` → 1 hit.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — per-recipient "
            "`\"Mail has been sent to <email> about "
            "NOTIFY_VACATION_CALENDAR_NEXT_YEAR_FIRST for vacation id = ...\"` "
            "marker for each chief accountant.\n"
            "CLEANUP: POST /api/ttt/v1/test/clock/reset."
        ),
        (
            "Each chief accountant receives exactly one email with the Cyrillic "
            "subject tail `Производственный календарь`. Per-recipient mail-sent "
            "markers exist. No other recipients are targeted."
        ),
        "High", "Hybrid",
        "#3423 row 15; historical UIDs 609812/609813",
        "vacation/cron/calendar-reminder",
        "DELTA: cron property key is `production-calendar-annual-first.cron` (not `-first-notification.cron`)."
    ),
    (
        "TC-VAC-110",
        "Calendar reminder test-endpoint bypasses scheduler marker",
        (
            "Env: ttt-timemachine. Any chief-accountant recipient (reuse TC-VAC-109 "
            "seed)."
        ),
        (
            "SETUP: Clear mailbox as in TC-VAC-109.\n"
            "TRIGGER: POST /api/vacation/v1/test/production-calendars/send-first-reminder.\n"
            "WAIT: 30 s.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — query "
            "`message:\"1st october\"` or "
            "`message:\"Starting AnnualProductionCalendarTask\"` → ZERO hits "
            "(scheduler-wrapper is bypassed by the test endpoint).\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — per-recipient "
            "`\"Mail has been sent to <email> about "
            "NOTIFY_VACATION_CALENDAR_NEXT_YEAR_FIRST...\"` markers still emit.\n"
            "CLEANUP: None."
        ),
        (
            "Scheduler-level marker does not fire when the job is triggered via the "
            "test endpoint — tests must assert on per-recipient mail-sent markers "
            "instead. This is expected behaviour documented in the scope-table deltas."
        ),
        "Medium", "Hybrid",
        "#3423 row 15 DELTA — scheduler-wrapper bypass",
        "vacation/cron/calendar-reminder",
        "Regression ensures tests do not begin depending on the bypassed scheduler marker."
    ),
    (
        "TC-VAC-111",
        "Calendar reminder with zero chief-accountant recipients completes cleanly",
        (
            "Env: ttt-timemachine. No active ROLE_CHIEF_ACCOUNTANT employees in scope.\n"
            "SETUP SQL: UPDATE employee_role SET <disable> for each ROLE_CHIEF_ACCOUNTANT "
            "binding — OR use a temporary env where no such role exists.\n"
            "(Alternative: run on a short-lived timemachine snapshot with the role "
            "cleared for the duration of the test.)"
        ),
        (
            "SETUP: Disable all ROLE_CHIEF_ACCOUNTANT role bindings for the test window.\n"
            "SETUP: Capture baseline Graylog error count.\n"
            "TRIGGER: POST /api/vacation/v1/test/production-calendars/send-first-reminder.\n"
            "WAIT: 15 s.\n"
            "VERIFY LOG: No new ERROR-level markers emitted.\n"
            "VERIFY EMAIL: No new emails with `NOTIFY_VACATION_CALENDAR_NEXT_YEAR_FIRST` "
            "template dispatched.\n"
            "CLEANUP: Restore role bindings."
        ),
        (
            "Endpoint returns HTTP 200, no errors are logged, no emails are sent. The "
            "job handles an empty recipient list without raising an exception."
        ),
        "Low", "Hybrid",
        "#3423 row 15 — empty-recipients edge case",
        "vacation/cron/calendar-reminder",
        "Negative test — guards against NPE on empty recipient list."
    ),
]

TS_AUTO_PAY = [
    (
        "TC-VAC-112",
        "Auto-pay expired approved — APPROVED vacation past pay date → PAID",
        (
            "Env: ttt-timemachine. One employee with APPROVED vacation whose "
            "pay_date_expected is in the past.\n"
            "Query: SELECT v.id, e.login FROM ttt_vacation.vacation v "
            "JOIN employee e ON e.id = v.employee_id "
            "WHERE v.status = 'APPROVED' AND v.pay_date_expected < CURRENT_DATE "
            "ORDER BY random() LIMIT 1; "
            "If no row, seed via API — create vacation → approve → "
            "UPDATE ttt_vacation.vacation SET pay_date_expected = CURRENT_DATE - 1 "
            "WHERE id = <id>."
        ),
        (
            "SETUP: Via API, create or identify an APPROVED vacation with past "
            "pay_date_expected (see preconditions).\n"
            "TRIGGER: POST /api/vacation/v1/test/vacations/pay-expired-approved.\n"
            "WAIT: 10 s.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — "
            "`\"Automatically pay approved task started\"` → "
            "`\"...finished\"`.\n"
            "DB-CHECK: SELECT status, paid_date FROM ttt_vacation.vacation "
            "WHERE id = <seed_id>; status = 'PAID', paid_date ≈ NOW().\n"
            "CLEANUP: If seeded, DELETE the vacation after capture."
        ),
        (
            "Vacation transitions APPROVED → PAID. `paid_date` is stamped. Scheduler "
            "start & finish markers fire. Notification email (PAID) is dispatched to "
            "the employee via the downstream Email batch."
        ),
        "High", "Hybrid",
        "#3423 row 16; EXT-cron-jobs.md job 16",
        "vacation/cron/auto-pay",
        "Regression ensures legacy lock name `CloseOutdatedTask.run` still applies "
        "to class AutomaticallyPayApprovedTask."
    ),
    (
        "TC-VAC-113",
        "Auto-pay boundary — APPROVED vacation within grace period is NOT paid",
        (
            "Env: ttt-timemachine. One APPROVED vacation with "
            "pay_date_expected = CURRENT_DATE + 1 (future).\n"
            "Seed via API (create → approve → set pay_date_expected to tomorrow)."
        ),
        (
            "SETUP: Seed or identify an APPROVED vacation with pay_date_expected "
            "in the near future.\n"
            "DB-CHECK: Capture pre-run status — APPROVED.\n"
            "TRIGGER: POST /api/vacation/v1/test/vacations/pay-expired-approved.\n"
            "WAIT: 10 s.\n"
            "DB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <seed_id>; "
            "status remains 'APPROVED', paid_date remains NULL.\n"
            "CLEANUP: Delete seeded vacation."
        ),
        (
            "Vacation is NOT promoted to PAID because pay_date_expected has not yet "
            "elapsed. Task scoping query correctly excludes future-dated rows."
        ),
        "Medium", "Hybrid",
        "#3423 row 16",
        "vacation/cron/auto-pay",
        "Edge-case protection against premature payment."
    ),
    (
        "TC-VAC-114",
        "Auto-pay legacy lock-name regression (CloseOutdatedTask.run)",
        (
            "Env: qa-1. Two parallel triggers issued within the ShedLock window."
        ),
        (
            "SETUP: Ensure at least one APPROVED vacation qualifies for pay-expired-approved "
            "(reuse seed from TC-VAC-112).\n"
            "TRIGGER A: POST /api/vacation/v1/test/vacations/pay-expired-approved.\n"
            "TRIGGER B: Immediately (<1 s later) POST the same endpoint again.\n"
            "WAIT: 10 s.\n"
            "DB-CHECK: SELECT name, locked_at FROM ttt_vacation.shedlock "
            "WHERE name = 'CloseOutdatedTask.run' — exactly 1 active-lock row.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — only 1 set of start/finish markers, not 2.\n"
            "CLEANUP: None (lock self-releases)."
        ),
        (
            "ShedLock name `CloseOutdatedTask.run` (legacy — class renamed to "
            "AutomaticallyPayApprovedTask) serialises concurrent triggers. Only one run "
            "executes; the second is skipped at the lock boundary."
        ),
        "Medium", "Hybrid",
        "#3423 row 16 — legacy lock name; EXT-cron-jobs.md §ShedLock inventory",
        "vacation/cron/auto-pay",
        "Regression for the lock-name consistency between old and new scheduler class."
    ),
]

TS_APPROVED_TO_PAID = [
    (
        "TC-VAC-115",
        "Period close triggers APPROVED → PAID via updateVacations",
        (
            "Env: ttt-timemachine. One APPROVED vacation whose payment_month is the "
            "current open accounting period.\n"
            "Query: SELECT v.id, v.payment_month FROM ttt_vacation.vacation v "
            "WHERE v.status = 'APPROVED' AND v.payment_month = "
            "(SELECT start_date FROM ttt_backend.report_period "
            "WHERE approved = false ORDER BY start_date DESC LIMIT 1) "
            "ORDER BY random() LIMIT 1."
        ),
        (
            "SETUP: Via API, create or identify an APPROVED vacation for the current "
            "open accounting period.\n"
            "TRIGGER: PATCH /api/ttt/v1/report-periods/<id> { approved: true } "
            "(ptch-report-period) to close the period.\n"
            "WAIT: 15 min (periodic cron `0 */10 * * * *` — updateVacations); "
            "OR force via polling every 30 s for up to 15 min, asserting DB state.\n"
            "DB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <seed_id>; "
            "status = 'PAID'.\n"
            "CLEANUP: Re-open the accounting period via PATCH if the env does not "
            "auto-reset; delete seeded vacation."
        ),
        (
            "Within 15 min of period closure, the `VacationStatusUpdateJob.updateVacations` "
            "cron sweeps eligible APPROVED vacations in closed periods and promotes them "
            "to PAID. No log markers are emitted (Bug #2 — job has no scheduler lock or "
            "logging); DB assertion is the only verification."
        ),
        "High", "Hybrid",
        "#3423 row 17; EXT-cron-jobs.md Bug #2",
        "vacation/cron/approved-to-paid",
        "Regression for Bug #2 — no markers available; DB-state assertion only."
    ),
    (
        "TC-VAC-116",
        "Day-off update triggers checkVacationDaysAfterCalendarUpdate recomputation",
        (
            "Env: ttt-timemachine. One employee with APPROVED vacation in current "
            "quarter.\n"
            "Query: SELECT v.id, e.login, v.start_date, v.end_date "
            "FROM ttt_vacation.vacation v JOIN employee e ON e.id = v.employee_id "
            "WHERE v.status = 'APPROVED' AND v.start_date >= CURRENT_DATE "
            "ORDER BY random() LIMIT 1."
        ),
        (
            "SETUP: Capture pre-update vacation days remaining — "
            "SELECT days FROM ttt_vacation.employee_vacation_days "
            "WHERE employee_id = <id> AND year = EXTRACT(year FROM v.start_date).\n"
            "TRIGGER: Via calendar service API, add a non-working day WITHIN the "
            "vacation date range — POST /api/calendar/v2/production-calendar/days "
            "{ date: <vacation_mid_date>, type: HOLIDAY }.\n"
            "WAIT: Up to 5 min for `checkVacationDaysAfterCalendarUpdate` "
            "(`0 */5 * * * *`) to fire.\n"
            "DB-CHECK: Vacation days increased by 1 for the affected employee "
            "(holiday removes a workday from vacation consumption).\n"
            "CLEANUP: Remove the holiday via "
            "DELETE /api/calendar/v2/production-calendar/days?date=<date>."
        ),
        (
            "Calendar change recalculates vacation-days consumption within 5 min. "
            "Employee gains back 1 vacation day because the holiday overrides a "
            "previously counted workday."
        ),
        "Medium", "Hybrid",
        "#3423 row 17 — checkVacationDaysAfterCalendarUpdate schedule",
        "vacation/cron/approved-to-paid",
        "Cross-service: calendar change propagates to vacation via cron poll, not event."
    ),
    (
        "TC-VAC-117",
        "APPROVED → PAID has no log markers (Bug #2 regression)",
        (
            "Env: qa-1. Reuse TC-VAC-115 seed vacation."
        ),
        (
            "SETUP: Reuse TC-VAC-115 state (APPROVED vacation, period closed).\n"
            "SETUP: Capture Graylog baseline by tailing TTT-QA-1 for 5 min before "
            "the cron fires.\n"
            "WAIT: For the next `updateVacations` fire (up to 10 min, cron "
            "`0 */10 * * * *`).\n"
            "VERIFY LOG: Graylog TTT-QA-1 — query "
            "`message:\"VacationStatusUpdateJob\" OR message:\"updateVacations\"` "
            "→ ZERO hits (confirmed Bug #2).\n"
            "DB-CHECK: Vacation status transitioned to PAID (confirms the job ran, "
            "just without logging).\n"
            "VERIFY LOG: No ERROR markers during the window.\n"
            "CLEANUP: None."
        ),
        (
            "Job executes (DB transition confirms) but emits no log markers. Tests "
            "cannot use Graylog for job-15 verification; DB assertion is the only "
            "mechanism. Regression tests re-confirm Bug #2 remains."
        ),
        "Medium", "Hybrid",
        "#3423 row 17 — Bug #2 (no markers)",
        "vacation/cron/approved-to-paid",
        "If markers appear in future, file an observability-improvement ticket and update this TC."
    ),
]

TS_EMP_PROJECT_SYNC = [
    # Seed TCs from 3262-ticket-findings.md §3
    (
        "TC-VAC-118",
        "Employee-projects sync — endpoint triggers full marker chain + DB upsert",
        (
            "Env: qa-1. Fresh state — "
            "SELECT COUNT(*) FROM ttt_vacation.employee_projects "
            "WHERE updated_at > NOW() - INTERVAL '1 hour'; any value fine.\n"
            "At least 5 enabled employees with recent task_reports (to guarantee "
            "sync has work to do).\n"
            "Query: SELECT DISTINCT e.login FROM employee e "
            "JOIN ttt_backend.task_report tr ON tr.reporter_id = e.id "
            "WHERE e.enabled = true AND tr.report_date > CURRENT_DATE - INTERVAL '7 day' "
            "LIMIT 5;"
        ),
        (
            "SETUP: Capture baseline row count + max(updated_at) in "
            "ttt_vacation.employee_projects.\n"
            "TRIGGER: POST /api/vacation/v1/test/employee-projects.\n"
            "WAIT: 30 s (periodic sync processes pages sequentially; no fan-out).\n"
            "VERIFY LOG: Graylog TTT-QA-1 — all four markers in order: "
            "`\"Employee Projects sync started...\"`, "
            "`\"Employee Projects sync page {N} started\"` (may repeat per page), "
            "`\"Employee Projects sync deleted {M} records\"`, "
            "`\"Employee Projects sync finished!\"`.\n"
            "DB-CHECK: max(updated_at) advanced; row count within expected delta "
            "(± <= 10% of employee count).\n"
            "CLEANUP: None."
        ),
        (
            "Full marker chain fires in order. Cache is refreshed — max(updated_at) "
            "in ttt_vacation.employee_projects is within the last 30 s. Deletion "
            "tracking emits the delete-count marker (zero-count still emits)."
        ),
        "High", "Hybrid",
        "#3423 row 18; #3178; #3262",
        "vacation/cron/employee-project-sync",
        "DELTA: endpoint path has /vacation prefix (scope-table omitted it)."
    ),
    (
        "TC-VAC-119",
        "Employee-projects sync — deleted task_report propagates to cache deletion",
        (
            "Env: qa-1. One employee-project cache row with exactly one backing "
            "task_report.\n"
            "Query: SELECT ep.employee_id, ep.project_id, "
            "(SELECT COUNT(*) FROM ttt_backend.task_report tr "
            "WHERE tr.reporter_id = ep.employee_id AND tr.project_id = ep.project_id) "
            "AS tr_count FROM ttt_vacation.employee_projects ep "
            "WHERE (SELECT COUNT(*) FROM ttt_backend.task_report tr "
            "WHERE tr.reporter_id = ep.employee_id AND tr.project_id = ep.project_id) = 1 "
            "LIMIT 1."
        ),
        (
            "SETUP: Identify a cache row with exactly one backing task_report.\n"
            "SETUP: Via API, DELETE /api/ttt/v1/task-reports/<id> — remove the "
            "backing report.\n"
            "TRIGGER: POST /api/vacation/v1/test/employee-projects.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: SELECT COUNT(*) FROM ttt_vacation.employee_projects "
            "WHERE employee_id = <id> AND project_id = <id>; returns 0 "
            "(cache row removed).\n"
            "VERIFY LOG: `\"Employee Projects sync deleted 1 records\"` or "
            "similar non-zero delete-count marker.\n"
            "CLEANUP: Restore task_report if test requires repeatability."
        ),
        (
            "Cache row is deleted by the sync when its last backing task_report is "
            "removed. Delete-count marker reflects at least 1 deletion."
        ),
        "High", "Hybrid",
        "#3423 row 18; #3262 Bug 1",
        "vacation/cron/employee-project-sync",
        "Regression for the deletion-propagation fix in #3262."
    ),
    (
        "TC-VAC-120",
        "Employee-projects sync — earlier task_report rolls back first_report_date",
        (
            "Env: qa-1. One employee-project row whose first_report_date is NOT the "
            "minimum possible for that employee+project.\n"
            "Query: SELECT ep.employee_id, ep.project_id, ep.first_report_date "
            "FROM ttt_vacation.employee_projects ep "
            "ORDER BY ep.first_report_date DESC LIMIT 1; "
            "(Pick any — we'll insert an EARLIER task_report below.)"
        ),
        (
            "SETUP: Capture current first_report_date.\n"
            "SETUP: Via API, POST /api/ttt/v1/task-reports with "
            "report_date = first_report_date - 7 days, "
            "reporter_id = <id>, project_id = <id>.\n"
            "TRIGGER: POST /api/vacation/v1/test/employee-projects.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: SELECT first_report_date FROM ttt_vacation.employee_projects "
            "WHERE employee_id = <id> AND project_id = <id>; "
            "value = inserted report_date (rolled back).\n"
            "CLEANUP: Delete the inserted task_report to restore original cache state."
        ),
        (
            "Cache row's first_report_date is updated to the earliest matching "
            "task_report date. Sync correctly recomputes boundary values."
        ),
        "Medium", "Hybrid",
        "#3423 row 18; #3178",
        "vacation/cron/employee-project-sync",
        "Boundary recomputation is #3178 foundational fix."
    ),
    (
        "TC-VAC-121",
        "Employee-projects sync — employee with zero reports → cache row removed",
        (
            "Env: qa-1. One employee-project cache row with exactly one backing "
            "task_report (same seed as TC-VAC-119 if available)."
        ),
        (
            "SETUP: Identify a cache row whose employee has only ONE task_report for "
            "that project.\n"
            "SETUP: Via API, DELETE /api/ttt/v1/task-reports/<id>.\n"
            "TRIGGER: POST /api/vacation/v1/test/employee-projects.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: SELECT COUNT(*) FROM ttt_vacation.employee_projects "
            "WHERE employee_id = <id>; if 0 other projects, row count = 0.\n"
            "DB-CHECK: No orphan cache rows for that employee+project combo.\n"
            "CLEANUP: Restore task_report."
        ),
        (
            "Cache entirely cleared when the employee has no remaining reports for "
            "any project. Sync handles full-delete correctly."
        ),
        "Medium", "Hybrid",
        "#3423 row 18; #3178",
        "vacation/cron/employee-project-sync",
        "Edge case — zero-row state must be reached cleanly."
    ),
    (
        "TC-VAC-122",
        "Employee-projects sync race window — WON'T FIX documentation test",
        (
            "Env: qa-1. This test DOCUMENTS an accepted limitation rather than "
            "asserting correctness. Runs on demand; NOT part of nightly regression."
        ),
        (
            "SETUP: Identify an employee+project cache row. Prepare to concurrently "
            "insert a task_report and trigger the sync.\n"
            "TRIGGER A: POST /api/vacation/v1/test/employee-projects.\n"
            "TRIGGER B (within 500 ms of A): POST /api/ttt/v1/task-reports "
            "{ report_date: today, reporter_id: <id>, project_id: <id> }.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: SELECT report_date FROM ttt_backend.task_report WHERE id = <new_id>; "
            "report exists.\n"
            "DB-CHECK: SELECT last_report_date FROM ttt_vacation.employee_projects "
            "WHERE employee_id = <id> AND project_id = <id>; "
            "value MAY or MAY NOT reflect the new report (race — accepted loss).\n"
            "ANNOTATION: Record outcome — pass/fail on cache state does not block "
            "release.\n"
            "CLEANUP: None."
        ),
        (
            "Cache row may not reflect task_reports written during the sync window. "
            "This is documented as WON'T FIX (#3262 note 866870). The 03:00 NSK "
            "schedule minimises exposure. Test exists solely to confirm the behaviour "
            "has not regressed into a crash or data corruption."
        ),
        "Low", "Hybrid",
        "#3423 row 18; #3262 WON'T FIX limitation",
        "vacation/cron/employee-project-sync",
        "Regression documents the accepted trade-off; run manually when #3262 fixes revisit."
    ),
    (
        "TC-VAC-123",
        "Employee-projects startup sync — cold env runs + java_migration marker inserted",
        (
            "Env: ttt-timemachine OR stage. DB pre-state — "
            "SELECT * FROM ttt_vacation.java_migration "
            "WHERE name = 'EMPLOYEE_PROJECT_INITIAL_SYNC'; row must NOT exist.\n"
            "SETUP: DELETE FROM ttt_vacation.java_migration "
            "WHERE name = 'EMPLOYEE_PROJECT_INITIAL_SYNC' (to simulate cold env)."
        ),
        (
            "SETUP: Delete the java_migration row (see preconditions).\n"
            "SETUP: Snapshot employee_projects table size.\n"
            "TRIGGER: Via gitlab-access skill — trigger the `restart-<env>` CI job on "
            "the `release/2.1` pipeline for the target env.\n"
            "WAIT: 3 min for service startup + sync completion.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE (or target) — full marker chain as "
            "TC-VAC-118.\n"
            "DB-CHECK: SELECT * FROM ttt_vacation.java_migration "
            "WHERE name = 'EMPLOYEE_PROJECT_INITIAL_SYNC'; exactly 1 row with "
            "created_at near NOW().\n"
            "DB-CHECK: employee_projects row count matches expected post-sync size.\n"
            "CLEANUP: None (marker row is intentional post-condition)."
        ),
        (
            "Cold startup runs the initial sync, emits the marker chain, and inserts "
            "the `EMPLOYEE_PROJECT_INITIAL_SYNC` row into `java_migration` so future "
            "startups no-op."
        ),
        "High", "Hybrid",
        "#3423 row 19; #3303",
        "vacation/cron/employee-project-sync",
        "Startup-only path; CI restart is the only way to trigger."
    ),
    (
        "TC-VAC-124",
        "Employee-projects startup sync — warm env skips on re-start",
        (
            "Env: ttt-timemachine. DB pre-state — "
            "SELECT * FROM ttt_vacation.java_migration "
            "WHERE name = 'EMPLOYEE_PROJECT_INITIAL_SYNC'; row MUST exist "
            "(post TC-VAC-123 or existing historical run)."
        ),
        (
            "SETUP: Confirm the marker row exists.\n"
            "TRIGGER: Via gitlab-access skill — trigger restart-<env> again.\n"
            "WAIT: 3 min for service startup.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — search for "
            "`message:\"Employee Projects sync started\"` WITHIN the restart's "
            "time window → ZERO hits.\n"
            "DB-CHECK: ttt_vacation.java_migration row for "
            "`EMPLOYEE_PROJECT_INITIAL_SYNC` unchanged (same created_at as before).\n"
            "CLEANUP: None."
        ),
        (
            "Startup sync is skipped because the java_migration guard row is present. "
            "`migrationExecutor.executeOnce(...)` returns without invoking the sync; "
            "no markers emit."
        ),
        "High", "Hybrid",
        "#3423 row 19; #3303",
        "vacation/cron/employee-project-sync",
        "One-shot idempotency guard must not drift."
    ),
    (
        "TC-VAC-125",
        "Employee-projects startup — delete marker to re-run (dev backdoor)",
        (
            "Env: ttt-timemachine. DB pre-state — marker row exists.\n"
            "Test validates the manual re-sync backdoor documented in #3303 note 876432."
        ),
        (
            "SETUP: DELETE FROM ttt_vacation.java_migration "
            "WHERE name = 'EMPLOYEE_PROJECT_INITIAL_SYNC' RETURNING *;\n"
            "SETUP: Capture employee_projects row count.\n"
            "TRIGGER: Via gitlab-access skill — trigger restart-<env>.\n"
            "WAIT: 3 min.\n"
            "VERIFY LOG: Marker chain fires again (cold-env behaviour).\n"
            "DB-CHECK: java_migration row re-inserted.\n"
            "CLEANUP: None."
        ),
        (
            "Removing the java_migration guard row re-enables the one-shot; the next "
            "startup performs a fresh initial sync and re-inserts the marker."
        ),
        "Low", "Hybrid",
        "#3423 row 19; #3303 note 876432",
        "vacation/cron/employee-project-sync",
        "Dev backdoor; avoid on stage unless approved."
    ),
]

TS_STAT_REPORT_INIT = [
    (
        "TC-VAC-126",
        "Statistic-report startup sync — cold env inserts marker + populates cache",
        (
            "Env: ttt-timemachine OR stage.\n"
            "SETUP SQL: DELETE FROM ttt_vacation.java_migration "
            "WHERE name = 'STATISTIC_REPORT_INITIAL_SYNC';"
        ),
        (
            "SETUP: Delete the STATISTIC_REPORT_INITIAL_SYNC row.\n"
            "SETUP: TRUNCATE or snapshot ttt_backend.statistic_report for comparison.\n"
            "TRIGGER: Via gitlab-access skill — restart the vacation service on the "
            "target env.\n"
            "WAIT: 3–5 min (initial sync is event-driven via RabbitMQ; can take 1–2 "
            "min post-startup).\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — "
            "`message:\"statistic report sync\"` markers fire within 5 min of restart.\n"
            "DB-CHECK: SELECT * FROM ttt_vacation.java_migration "
            "WHERE name = 'STATISTIC_REPORT_INITIAL_SYNC'; exactly 1 row.\n"
            "DB-CHECK: SELECT COUNT(*) FROM ttt_backend.statistic_report "
            "WHERE updated_at > NOW() - INTERVAL '10 min'; row count > 0.\n"
            "CLEANUP: None."
        ),
        (
            "Cold startup runs the statistic-report initial sync, populates "
            "ttt_backend.statistic_report, and inserts the marker row so future "
            "startups no-op."
        ),
        "High", "Hybrid",
        "#3423 row 21; #3346",
        "vacation/cron/statistic-report-init",
        "Async — use polling DB-CHECK to avoid fixed sleep."
    ),
    (
        "TC-VAC-127",
        "Statistic-report startup sync — warm env skips on restart",
        (
            "Env: ttt-timemachine. DB pre-state — marker row exists."
        ),
        (
            "SETUP: Confirm the marker row exists.\n"
            "SETUP: Capture current ttt_backend.statistic_report row count + "
            "max(updated_at).\n"
            "TRIGGER: Via gitlab-access skill — restart the vacation service.\n"
            "WAIT: 5 min.\n"
            "VERIFY LOG: Graylog TTT-TIMEMACHINE — no `\"statistic report sync started\"` "
            "markers within the restart's time window.\n"
            "DB-CHECK: max(updated_at) unchanged in ttt_backend.statistic_report "
            "(no bulk upsert).\n"
            "CLEANUP: None."
        ),
        (
            "Startup sync skipped because the java_migration guard is present. "
            "Cache is not touched."
        ),
        "High", "Hybrid",
        "#3423 row 21; #3346",
        "vacation/cron/statistic-report-init",
        "One-shot idempotency mirror of TC-VAC-124."
    ),
]


SUITES = [
    ("TS-Vac-Cron-AnnualAccruals",  TS_ANNUAL_ACCRUALS),
    ("TS-Vac-Cron-NotImpl",         TS_NOT_IMPL),
    ("TS-Vac-Cron-Digest",          TS_DIGEST),
    ("TS-Vac-Cron-CalendarReminder", TS_CAL_REMINDER),
    ("TS-Vac-Cron-AutoPay",         TS_AUTO_PAY),
    ("TS-Vac-Cron-ApprovedToPaid",  TS_APPROVED_TO_PAID),
    ("TS-Vac-Cron-EmpProjectSync",  TS_EMP_PROJECT_SYNC),
    ("TS-Vac-Cron-StatReportInit",  TS_STAT_REPORT_INIT),
]


# ─── Sheet writer ─────────────────────────────────────────────────────────────

def _write_suite(wb, name, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLOR

    # R1 — back-link
    back = ws.cell(row=1, column=1,
                   value='=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")')
    back.font = FONT_LINK

    # R2 — header
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        letter = c.column_letter
        if col_idx - 1 < len(COL_WIDTHS):
            ws.column_dimensions[letter].width = COL_WIDTHS[col_idx - 1]
    ws.freeze_panes = ws.cell(row=3, column=1)

    # R3+ — TC rows
    for row_offset, tc in enumerate(rows, start=0):
        r = 3 + row_offset
        fill = FILL_ALT if (row_offset % 2) else FILL_WHITE
        for col_idx, val in enumerate(tc, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1, 6, 7) else ALIGN_WRAP
            c.fill = fill
            c.border = BORDER_THIN


def main():
    if not os.path.isfile(_VAC_XLSX):
        raise SystemExit(f"vacation.xlsx not found at {_VAC_XLSX}")

    wb = load_workbook(_VAC_XLSX)

    all_ids = []
    for suite_name, rows in SUITES:
        _write_suite(wb, suite_name, rows)
        all_ids.extend(r[0] for r in rows)

    wb.save(_VAC_XLSX)

    print(f"Extended: {_VAC_XLSX}")
    print(f"Suites added/updated: {len(SUITES)}")
    print(f"TCs written: {len(all_ids)} ({all_ids[0]} ... {all_ids[-1]})")
    for name, rows in SUITES:
        print(f"  {name}: {len(rows)} TCs")


if __name__ == "__main__":
    main()
