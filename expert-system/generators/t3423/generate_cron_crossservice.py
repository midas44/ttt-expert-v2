#!/usr/bin/env python3
"""
Ticket #3423 — Cross-Service Cluster Cron Test Cases (Phase B, session 137).

Generates ``test-docs/collections/cron/Cron_CrossService.xlsx`` from scratch with:
  * a Plan Overview sheet (authored via ``_common.author_plan_overview``)
  * 2 cron-focused suites covering jobs 6, 10, 20, 23:
        - TS-CrossService-CronCSSync     (rows 6, 10, 20 — 11 TCs: TC-CS-101..111)
        - TS-CrossService-CronPMToolSync (row 23        — 10 TCs: TC-CS-112..121)

Test IDs TC-CS-101…121 are preserved from the pre-migration state (cron
suites formerly lived inside ``test-docs/cross-service/cross-service.xlsx``;
extracted on 2026-04-20 by ``migrate_to_cron_workbooks.py``).

Idempotent: every run rebuilds a fresh workbook.

Tab color ``F4B084`` (orange) marks every cron suite — matches the
home-workbook convention established in Phase B.

Run from repo root:
    python3 expert-system/generators/t3423/generate_cron_crossservice.py

Canonical references:
    - expert-system/vault/exploration/tickets/t3423-investigation.md (scope)
    - expert-system/vault/external/EXT-cron-jobs.md §§Session 130-132 (markers)
    - expert-system/vault/exploration/tickets/3083-ticket-findings.md (row 23 contract)
    - Phase-A deltas folded into TCs:
        Delta #7 (row 20 path): POST /api/calendar/v2/test/salary-office/sync
                                 ?fullSync={true|false} (v2, not v1).
        Delta #10: "Daily 00:00 full CS sync" wording is incorrect —
                   companyStaff.full-sync YAML key is dead config;
                   full sync fires only at startup via TttStartupApplicationListener.
    - GitLab regression tickets folded into TCs:
        #3083 / #3286 — PM Tool field contract (append-only presales, immutable
                         accounting_name, default script-url no history event,
                         PMT-owned fields silent, cs-id validation).
        #3382 — Presales append-only merge regression.
        #3399 — PM Tool startup full sync on restart.
    - Marker collision: rows 6 + 20 both emit
      "Company staff synchronization started/finished". Graylog stream filter
      (TTT-QA-1 vs calendar-backend stream) disambiguates. Row 10 uses a
      DIFFERENT marker ("CS sync started/finished") in the vacation service.
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _common import author_plan_overview  # noqa: E402

# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_TARGET = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "collections", "cron", "Cron_CrossService.xlsx")
)


# ─── Styling (matches existing cross-service.xlsx rows + orange cron tab) ─────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR = "F4B084"  # Orange — distinguishes cron suites from 8 pre-existing TS-CrossService-* suites

COL_WIDTHS = [14, 40, 40, 55, 40, 10, 10, 30, 14, 35]
HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
    "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
]


# ─── Test case data ───────────────────────────────────────────────────────────
#
# Tuples: (test_id, title, preconditions, steps, expected, priority, type_,
#          req_ref, module, notes)
#
# Shape per t3423-investigation.md §Verification shape:
#   SETUP → Clock → Trigger → Wait → Verify (DB/UI/Email/Log) → CLEANUP
# UI-first rule: UI steps are used where a user-visible effect exists (project
# form edits for row 23 tracker-script history test, CS admin login audit for
# row 6/10/20). Cron and startup triggers use API test endpoints because there
# is no UI equivalent (CLAUDE.md §11 exception).
#
# Marker collision between rows 6 (ttt) and 20 (calendar): both emit
# "Company staff synchronization started/finished". Row 10 (vacation) emits a
# DIFFERENT marker ("CS sync started/finished") — this is a useful design
# asymmetry to cover in tests.


TS_CRON_CS_SYNC = [
    # ─── Row 6 — ttt-service CS sync (every 15 min) ──────────────────────────
    (
        "TC-CS-101",
        "CS sync (ttt) — cron fires every 15 min and emits start/finish markers",
        (
            "Env: ttt-qa-1. Unleash flag `CS_SYNC-qa-1` = ON (default).\n"
            "ShedLock table `ttt_backend.shedlock` accessible; no held lock "
            "named `CSSyncScheduler.doCsSynchronization` from another pod.\n"
            "Graylog API token for stream `TTT-QA-1`."
        ),
        (
            "SETUP: Confirm Unleash `CS_SYNC-qa-1` is enabled (admin UI at "
            "unleash.noveogroup.com).\n"
            "SETUP: Snapshot latest timestamp on `ttt_backend.shedlock "
            "WHERE name = 'CSSyncScheduler.doCsSynchronization'` via postgres MCP.\n"
            "WAIT: up to 16 min for the next cron boundary (*/15). "
            "Alternative: trigger via `POST /api/ttt/v1/test/company-staff/sync` "
            "if available (verify test endpoint; row 6 has no dedicated test "
            "endpoint in release/2.1 — cron wait is canonical).\n"
            "VERIFY LOG: Graylog `stream:TTT-QA-1` — `message:\"Company staff "
            "synchronization started\"` within last 16 min.\n"
            "VERIFY LOG: same stream — `message:\"Company staff synchronization "
            "finished\"` within last 16 min, paired with the start marker.\n"
            "DB-CHECK: `shedlock.lock_until` updated past snapshot (lock was "
            "acquired and released during the run).\n"
            "CLEANUP: None."
        ),
        (
            "Scheduler runs on its 15-min cadence; start/finish markers appear "
            "on TTT-QA-1 stream; shedlock row refreshed."
        ),
        "Critical", "Functional",
        "#3423 row 6; EXT-cron-jobs §Session 131.1",
        "ttt-service/cs-sync/cron",
        "Marker text identical to row 20 (calendar) — disambiguate via stream filter."
    ),
    (
        "TC-CS-102",
        "CS sync (ttt) — Unleash flag CS_SYNC-qa-1 OFF makes sync body a silent no-op",
        (
            "Env: ttt-qa-1. Ability to toggle Unleash flag `CS_SYNC-qa-1`."
        ),
        (
            "SETUP: Via Unleash admin UI — disable `CS_SYNC-qa-1`.\n"
            "SETUP: Pick a CS employee whose `ttt_backend.employee` row could "
            "legitimately be updated by the sync (e.g., freshly added in CS).\n"
            "SETUP: Snapshot target employee's `updated_at` in "
            "`ttt_backend.employee`.\n"
            "WAIT: up to 16 min for the next cron boundary.\n"
            "VERIFY LOG: Graylog `stream:TTT-QA-1` — start/finish markers still "
            "fire (scheduler runs), but no body-level markers (e.g., "
            "`\"Employee {} synched\"`).\n"
            "DB-CHECK: target employee's `updated_at` UNCHANGED — no CS payload "
            "was applied.\n"
            "CLEANUP: Re-enable `CS_SYNC-qa-1` in Unleash."
        ),
        (
            "Feature-toggle gate stops CSSyncLauncherImpl.sync(false) before "
            "CSSyncServiceV2.sync(). Scheduler marker pair still emits because "
            "the guard is inside the launcher, not at the scheduler level."
        ),
        "High", "Negative",
        "#3423 row 6; EXT-cron-jobs §Feature-toggle gates",
        "ttt-service/cs-sync/feature-toggle",
        "Requires Unleash admin permission. Reference: session 131.1 launcher gate note."
    ),
    (
        "TC-CS-103",
        "CS sync (ttt) — full sync fires at application startup via TttStartupApplicationListener",
        (
            "Env: stage (CI restart permission required via GitLab pipeline). "
            "CS_SYNC flag ON for the target env."
        ),
        (
            "SETUP: Trigger a ttt-service restart via GitLab pipeline "
            "(`restart-stage` job on `stage` pipeline).\n"
            "WAIT: ~60-90 s for service startup; tail Graylog stream "
            "`TTT-STAGE` via `graylog-access tail --stream TTT-STAGE -n 200`.\n"
            "VERIFY LOG: marker `message:\"Company staff synchronization started\"` "
            "AND a FULL sync indicator (e.g., paging markers showing ALL CS "
            "offices / employees — not delta). Look for a long run-time "
            "between start and finish markers (order of minutes, not seconds).\n"
            "DB-CHECK: `ttt_backend.shedlock WHERE name = "
            "'CSSyncScheduler.doCsSynchronization'` is NOT the source — "
            "startup bypasses shedlock (fires from "
            "`TttStartupApplicationListener.onApplicationEvent` calling "
            "`csSyncLauncher.sync(true)` directly).\n"
            "CLEANUP: None (restart itself is the operational action)."
        ),
        (
            "Full CS sync (`sync(true)`) runs once per service startup; "
            "delta cron continues at */15 thereafter."
        ),
        "High", "Functional",
        "#3423 row 6; delta #10 (full CS sync is startup-only)",
        "ttt-service/cs-sync/startup",
        "Delta #10 folded — scope-table 'Daily 00:00 full CS sync' is incorrect; YAML key `companyStaff.full-sync` is DEAD CONFIG. Only startup fires full sync."
    ),
    # ─── Row 10 — vacation-service CS sync (different marker text!) ──────────
    (
        "TC-CS-104",
        "CS sync (vacation) — cron fires and emits distinct `CS sync started/finished` markers",
        (
            "Env: ttt-qa-1. Vacation service is up; Graylog stream for vacation "
            "backend accessible (via `graylog-access streams` — likely "
            "`VACATION-QA-1` or similar; verify stream name per session 130 "
            "findings).\n"
            "Unleash flag `CS_SYNC-qa-1` ON (vacation shares the flag with ttt)."
        ),
        (
            "SETUP: Confirm vacation-backend Graylog stream name (session 130 "
            "used `tail --stream <vacation-stream>`).\n"
            "WAIT: up to 16 min for the vacation cron cycle.\n"
            "VERIFY LOG: on VACATION stream — `message:\"CS sync started\"` "
            "followed by `message:\"CS sync finished\"`.\n"
            "VERIFY LOG: these markers are DIFFERENT from ttt (row 6) and "
            "calendar (row 20), which emit \"Company staff synchronization "
            "started/finished\". A single Graylog query "
            "`message:\"Company staff synchronization\" OR message:\"CS sync started\"` "
            "would return both families; the text difference is the "
            "disambiguator.\n"
            "DB-CHECK: `ttt_vacation.shedlock WHERE name = "
            "'CSSyncScheduler.sync'` has `lock_until` bumped (method name `sync`, "
            "not `doCsSynchronization` — per EXT-cron-jobs §row 10).\n"
            "CLEANUP: None."
        ),
        (
            "Vacation service's CS sync cron fires every 15 min; lock name and "
            "log marker differ from ttt/calendar services — noise isolation."
        ),
        "Critical", "Functional",
        "#3423 row 10; EXT-cron-jobs §Session 130 row 10",
        "vacation-service/cs-sync/cron",
        "Marker text differs across services; do NOT assume all three use 'Company staff synchronization'."
    ),
    (
        "TC-CS-105",
        "CS sync (vacation) — failure path logs at WARN level, not ERROR (caution for Graylog level:3 filters)",
        (
            "Env: ttt-qa-1. Ability to induce a CS API error (blocked network, "
            "CS service down, or mock an invalid response). Practically: "
            "observe during a real incident, or request infra to briefly "
            "block CS outbound from vacation pod."
        ),
        (
            "SETUP: Induce a CS API failure (see preconditions). Alternative: "
            "mine Graylog history for a prior CS sync failure on the vacation "
            "stream (search `message:\"CS sync failed\"`).\n"
            "VERIFY LOG: Graylog VACATION stream — marker "
            "`message:\"CS sync failed\"` appears at level WARN (level:4), "
            "NOT ERROR (level:3).\n"
            "VERIFY LOG: a Graylog query filtered to `level:3` (ERROR only) "
            "would NOT find this event — file as design issue. Tests that "
            "want to catch CS-sync failures must search by message pattern, "
            "not by level.\n"
            "CLEANUP: Restore CS connectivity."
        ),
        (
            "Failure surfaces as WARN-level log only; monitoring rules filtered "
            "to ERROR will miss it. Document as design issue."
        ),
        "Medium", "Negative",
        "#3423 row 10; EXT-cron-jobs §row 10 (warn not error)",
        "vacation-service/cs-sync/error-handling",
        "Design issue candidate — log-level mis-classification. Link with row 22 (StatisticReportScheduler also logs failure at INFO, not ERROR) — a pattern."
    ),
    # ─── Row 20 — calendar-service CS sync (marker collision with row 6) ────
    (
        "TC-CS-106",
        "CS sync (calendar) — fullSync=false via v2 test endpoint runs delta sync",
        (
            "Env: ttt-qa-1. API_SECRET_TOKEN with calendar service access.\n"
            "Unleash `CS_SYNC-qa-1` ON.\n"
            "Graylog stream for calendar backend accessible."
        ),
        (
            "SETUP: Identify a calendar entity (e.g., a salary office) in CS "
            "that has a recent update (or update one via CS admin if possible).\n"
            "TRIGGER: `POST /api/calendar/v2/test/salary-office/sync"
            "?fullSync=false` with API_SECRET_TOKEN.\n"
            "WAIT: 30 s.\n"
            "VERIFY LOG: calendar stream — `message:\"Company staff "
            "synchronization started\"` and `\"...finished\"` within last 2 min.\n"
            "VERIFY LOG: The same markers on TTT-QA-1 stream are UNRELATED "
            "to this invocation — use stream filter.\n"
            "DB-CHECK: `ttt_calendar.shedlock WHERE name = "
            "'CSSyncScheduler.doCsSynchronization'` updated (SAME lock name "
            "as ttt, distinct shedlock table — no cross-service contention).\n"
            "DB-CHECK: target salary_office row in `ttt_calendar` schema "
            "reflects the CS update (e.g., name/address change applied).\n"
            "CLEANUP: None."
        ),
        (
            "Calendar's delta CS sync applies incremental updates; shedlock in "
            "separate schema prevents contention with ttt's identically-named "
            "lock."
        ),
        "Critical", "Functional",
        "#3423 row 20; delta #7 (v2 endpoint path)",
        "calendar-service/cs-sync/test-endpoint",
        "Delta #7 folded — scope-table says /api/calendar/v1/salary-offices/sync; actual is v2 + /test/ + singular 'salary-office' + query param fullSync."
    ),
    (
        "TC-CS-107",
        "CS sync (calendar) — fullSync=true via v2 test endpoint forces full reconciliation",
        (
            "Env: ttt-qa-1. Same preconditions as TC-CS-106."
        ),
        (
            "SETUP: Identify a salary office that exists in calendar DB but "
            "was deleted in CS (an orphaned row), or snapshot the row count "
            "before the trigger.\n"
            "TRIGGER: `POST /api/calendar/v2/test/salary-office/sync"
            "?fullSync=true` with API_SECRET_TOKEN.\n"
            "WAIT: 60 s (full sync scans entire CS catalog).\n"
            "VERIFY LOG: calendar stream — same start/finish markers, but the "
            "time between them is noticeably longer than delta (order of "
            "tens of seconds vs sub-second).\n"
            "DB-CHECK: orphaned salary_office rows are reconciled (deleted or "
            "marked inactive per business rules).\n"
            "DB-CHECK: total salary_office count aligns with CS source of "
            "truth.\n"
            "CLEANUP: None."
        ),
        (
            "fullSync=true re-fetches full CS catalog and reconciles deletions. "
            "No cron path triggers fullSync=true — only test endpoint and "
            "CalendarStartupApplicationListener at service startup."
        ),
        "High", "Functional",
        "#3423 row 20; delta #7; delta #10",
        "calendar-service/cs-sync/full-sync",
        "fullSync=true not reachable via cron — test endpoint or restart are the only paths."
    ),
    (
        "TC-CS-108",
        "CS sync — marker collision: rows 6 and 20 both emit 'Company staff synchronization' — disambiguate by Graylog stream",
        (
            "Env: ttt-qa-1. Unleash `CS_SYNC-qa-1` ON for BOTH ttt and "
            "calendar services.\n"
            "Graylog API access with ability to filter by `stream` field."
        ),
        (
            "WAIT: For a natural */15 cron boundary — both rows 6 and 20 fire "
            "within seconds of each other.\n"
            "VERIFY LOG: Graylog query `message:\"Company staff synchronization "
            "started\" AND timestamp:[NOW-2m TO NOW]` returns TWO hits (one "
            "per service).\n"
            "VERIFY LOG: adding `stream:TTT-QA-1` narrows to ttt only; "
            "adding `stream:<calendar-stream>` narrows to calendar only. The "
            "`source` field (pod name) is a secondary disambiguator when "
            "stream names are ambiguous.\n"
            "VERIFY LOG: confirm that row 10 (vacation) does NOT appear in "
            "either filter — it uses `\"CS sync started\"` instead."
        ),
        (
            "Identical marker text across rows 6 and 20 requires stream or "
            "source filter; monitoring rules based on message-only matching "
            "double-count unless filtered."
        ),
        "High", "Verification",
        "#3423 rows 6 & 20; EXT-cron-jobs §Marker collision",
        "ttt+calendar/cs-sync/observability",
        "Use `stream:<name>` filter in Graylog. Session 131 documented this collision; include in monitoring-rule review."
    ),
    # ─── Infra-level concerns shared across rows 6, 10, 20 ───────────────────
    (
        "TC-CS-109",
        "CS sync — ShedLock: concurrent pods do not double-run the same cron",
        (
            "Env: stage or qa-1 with 2+ ttt-service pods. Postgres access to "
            "`ttt_backend.shedlock`."
        ),
        (
            "SETUP: Identify two running ttt-service pods via GitLab/k8s. "
            "Confirm `shedlock.lock_until` column is populated for "
            "`CSSyncScheduler.doCsSynchronization`.\n"
            "WAIT: for the */15 cron boundary.\n"
            "VERIFY LOG: Graylog `stream:TTT-STAGE` — exactly ONE "
            "\"Company staff synchronization started\" marker within 5 s of "
            "the boundary, from one pod source. The second pod does NOT log "
            "the marker (lock held).\n"
            "DB-CHECK: `SELECT name, locked_by, locked_at, lock_until FROM "
            "ttt_backend.shedlock WHERE name = "
            "'CSSyncScheduler.doCsSynchronization'` — `locked_by` identifies "
            "the pod that held the lock.\n"
            "CLEANUP: None."
        ),
        (
            "Only one pod executes the cron body; the other skips silently "
            "due to ShedLock."
        ),
        "High", "Idempotency",
        "#3423 rows 6/10/20; ShedLock design",
        "ttt/vacation/calendar/shedlock",
        "HA guard. If both pods run, duplicate side-effects (emails, CS payload re-apply) would result."
    ),
    (
        "TC-CS-110",
        "CS sync — parallel execution across services: ttt (row 6) and calendar (row 20) lock names identical but schemas distinct",
        (
            "Env: ttt-qa-1 OR stage. Postgres access to BOTH "
            "`ttt_backend.shedlock` and `ttt_calendar.shedlock`."
        ),
        (
            "WAIT: for the */15 cron boundary.\n"
            "DB-CHECK: `SELECT schemaname, locked_by, locked_at FROM (SELECT "
            "'ttt_backend' AS schemaname, locked_by, locked_at FROM "
            "ttt_backend.shedlock WHERE name = "
            "'CSSyncScheduler.doCsSynchronization' UNION ALL SELECT "
            "'ttt_calendar', locked_by, locked_at FROM ttt_calendar.shedlock "
            "WHERE name = 'CSSyncScheduler.doCsSynchronization') t` — both "
            "rows fresh, different `locked_by` pods.\n"
            "VERIFY LOG: both services log start/finish within the same "
            "minute (parallel execution), confirmed by timestamps across "
            "their streams."
        ),
        (
            "Separate schemas (ttt_backend vs ttt_calendar) give each service "
            "its own shedlock table — identical lock name across services is "
            "NOT a collision at the DB level."
        ),
        "Medium", "Verification",
        "#3423 rows 6 & 20; EXT-cron-jobs §ShedLock name collision",
        "ttt+calendar/shedlock",
        "Design note: lock-name reuse across services is acceptable here only because of schema isolation. A deployment merging schemas would create a cross-service contention."
    ),
    (
        "TC-CS-111",
        "CS sync — idempotency: re-trigger within lock window yields zero DB deltas",
        (
            "Env: ttt-qa-1. Unleash `CS_SYNC-qa-1` ON. Access to calendar v2 "
            "test endpoint (simplest to re-trigger)."
        ),
        (
            "SETUP: Snapshot `SELECT max(updated_at) FROM "
            "ttt_calendar.salary_office` (or equivalent target table).\n"
            "TRIGGER 1: `POST /api/calendar/v2/test/salary-office/sync"
            "?fullSync=false`.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: note the new `max(updated_at)` after trigger 1.\n"
            "TRIGGER 2: re-fire the same endpoint immediately.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: `max(updated_at)` is UNCHANGED between trigger 1 and "
            "trigger 2 (assuming no new CS changes between calls).\n"
            "VERIFY LOG: both triggers emit start/finish markers; the second "
            "run's body logs 0 changes applied (verify body-level log "
            "`message:\"Applied {} updates\"` or similar; capture exact text "
            "in an autotest follow-up)."
        ),
        (
            "Sync is a query → diff → apply pipeline; with no new CS state "
            "between runs, the diff is empty and no writes occur."
        ),
        "Medium", "Idempotency",
        "#3423 rows 6/10/20",
        "cs-sync/idempotency",
        "Edge case: if CS emits a tombstone or manual edit lands between triggers, this assertion fails by design. Timing window < 1 s in practice."
    ),
]


TS_CRON_PMTOOL_SYNC = [
    # ─── Row 23 — PM Tool sync (every 15 min) — 10 TCs ────────────────────────
    (
        "TC-CS-112",
        "PM Tool sync — new project in PMT propagates to TTT on next cron cycle",
        (
            "Env: ttt-qa-1. Unleash `PM_TOOL_SYNC-qa-1` ON. API_SECRET_TOKEN.\n"
            "A PM Tool project created in the last hour that is NOT yet in "
            "`ttt_backend.project` (query `WHERE pm_id IS NULL` excludes "
            "pre-existing TTT projects).\n"
            "Project manager's cs_id exists in `ttt_backend.employee.cs_id` "
            "(per #3083 cs-id validation rule)."
        ),
        (
            "SETUP: Identify the new PM Tool project (e.g., from PMT admin "
            "UI or PMT API). Note its `pm_id`, name, manager cs_id.\n"
            "SETUP: Confirm `SELECT 1 FROM ttt_backend.project WHERE pm_id = "
            "<pm_id>` returns 0 rows.\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync` with "
            "API_SECRET_TOKEN.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — chain fires: "
            "`\"Pm tool synchronization started\"` → `\"PmTool Sync PROJECT "
            "started (fullSync=false)\"` → `\"PROJECT <pm_id> synched\"` → "
            "`\"PmTool Sync PROJECT finished (fullSync=false), result = "
            "SyncResult(success=true, successCount>=1, ...)\"` → "
            "`\"Pm tool synchronization finished\"`.\n"
            "DB-CHECK: `SELECT pm_id, name, status, accounting_name FROM "
            "ttt_backend.project WHERE pm_id = <pm_id>` returns one row; "
            "`accounting_name` equals the first-sync `name` value.\n"
            "CLEANUP: None (new project is legitimate state)."
        ),
        (
            "Full PMT-to-TTT propagation chain works end-to-end; pm_id column "
            "populated; accounting_name initialized from first-sync name "
            "(per #3083 §Accounting name rules)."
        ),
        "Critical", "Functional",
        "#3423 row 23; #3083; #3286",
        "ttt-service/pm-tool-sync/cron",
        "Happy path for full propagation chain. Per #3083: cs-ids validated, default script-url auto-populated (no history event — see TC-CS-116)."
    ),
    (
        "TC-CS-113",
        "PM Tool sync — existing project update: name/customer/status reflect PMT changes silently",
        (
            "Env: ttt-qa-1. A project that exists in BOTH PMT and TTT "
            "(SELECT pm_id, name FROM ttt_backend.project WHERE pm_id IS NOT "
            "NULL LIMIT 1)."
        ),
        (
            "SETUP: Pick target (pm_id, ttt.project.id). Snapshot current "
            "`name`, `status`, `customer_id`, and `accounting_name`.\n"
            "SETUP: Via PMT admin UI — rename the project, change customer, "
            "change status (e.g., DRAFT→ACTIVE).\n"
            "SETUP: Snapshot count of `ttt_backend.project_event WHERE "
            "project_id = <ttt_project_id>` BEFORE trigger.\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `name`, `customer_id`, `status` reflect PMT values.\n"
            "DB-CHECK: `accounting_name` UNCHANGED (immutable, see TC-CS-115).\n"
            "DB-CHECK: `project_event` row count UNCHANGED — PMT-owned field "
            "changes emit NO history event (per #3083 §Event history rules).\n"
            "VERIFY LOG: `\"PROJECT <pm_id> synched\"` marker fires."
        ),
        (
            "Non-immutable fields update silently; history table untouched. "
            "Status mapping: PMT 'draft' → TTT ACTIVE (per #3083 note 1)."
        ),
        "Critical", "Functional",
        "#3423 row 23; #3083",
        "ttt-service/pm-tool-sync/update",
        "Confirm `project_event` delta = 0 for PMT-owned field changes."
    ),
    (
        "TC-CS-114",
        "PM Tool sync — presales merge is APPEND-ONLY (existing TTT-only IDs preserved)",
        (
            "Env: ttt-qa-1. Ability to edit `ttt_backend.project.pre_sales_ids` "
            "AND PMT-side presales array (admin access)."
        ),
        (
            "SETUP: Pick a project synced from PMT with pm_id = X.\n"
            "SETUP: Via PMT admin — set presales IDs array to [A, B].\n"
            "SETUP: Trigger sync; confirm TTT `pre_sales_ids = [A, B]`.\n"
            "SETUP: Via direct DB write (or TTT admin if UI exists) — ADD "
            "TTT-only presale ID C: UPDATE project SET pre_sales_ids = "
            "[A, B, C].\n"
            "SETUP: Via PMT admin — change PMT presales array to [A, D] "
            "(remove B, add D).\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `SELECT pre_sales_ids FROM ttt_backend.project WHERE "
            "pm_id = X` — union [A, B, C, D] (order may vary). "
            "B is NOT removed (append-only); C (TTT-only) is preserved; "
            "D (new on PMT) is added.\n"
            "CLEANUP: Restore PMT + TTT presales arrays to pre-test state."
        ),
        (
            "Presales merge behaves as SET union — PMT can only add, never "
            "delete. TTT-only IDs (entered manually) survive every sync cycle."
        ),
        "Critical", "Regression",
        "#3423 row 23; #3083 §pre_sales_ids; #3382",
        "ttt-service/pm-tool-sync/presales",
        "Append-only merge — #3083 description 'existing IDs never deleted'. #3382 regression fix preserved this behavior after a separate refactor."
    ),
    (
        "TC-CS-115",
        "PM Tool sync — accounting_name is IMMUTABLE (PMT rename does NOT update it)",
        (
            "Env: ttt-qa-1. A project synced from PMT with a known "
            "`accounting_name`."
        ),
        (
            "SETUP: Pick project with pm_id = X. Snapshot current `name` and "
            "`accounting_name` (should be equal if project was created via "
            "sync — #3083 §Accounting name rules).\n"
            "SETUP: Via PMT admin — rename the project to a NEW name.\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `name` is the new PMT value.\n"
            "DB-CHECK: `accounting_name` IS THE ORIGINAL (unchanged).\n"
            "STEP (UI): Login as ADMIN, navigate to Admin → Projects → find "
            "the renamed project. Verify display shows updated `name` and "
            "retained `accounting_name` — both fields rendered separately on "
            "the edit form.\n"
            "CLEANUP: Restore PMT name."
        ),
        (
            "accounting_name is a one-time assignment at project creation; "
            "never updated from PMT. Critical for downstream API contracts "
            "(TTT exports accounting_name to external systems per #3083)."
        ),
        "High", "Regression",
        "#3423 row 23; #3083 §Accounting name",
        "ttt-service/pm-tool-sync/immutable-fields",
        "One-time-write field — test protects backwards-compat for downstream consumers."
    ),
    (
        "TC-CS-116",
        "PM Tool sync — default tracker-script auto-populated for new project emits NO history event",
        (
            "Env: ttt-qa-1. A PMT project without a defined tracker-sync "
            "script (empty `taskInfoScript` in PMT payload)."
        ),
        (
            "SETUP: Identify a new PMT project with empty script-url. "
            "Alternatively, create one via PMT admin and leave script empty.\n"
            "SETUP: Confirm `SELECT 1 FROM ttt_backend.project WHERE pm_id = "
            "<pm_id>` returns 0 rows.\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `SELECT task_info_script FROM ttt_backend.project "
            "WHERE pm_id = <pm_id>` returns the TTT default URL "
            "(`ttt.noveogroup.com/api/ttt/resource/defaultTaskInfoScript.js`).\n"
            "DB-CHECK: `SELECT count(*) FROM ttt_backend.project_event WHERE "
            "project_id = <ttt_project_id>` returns 0. The default-script "
            "auto-set MUST NOT emit an event (per #3083 §Default script-url "
            "rule).\n"
            "CLEANUP: None."
        ),
        (
            "Auto-populate happens silently at create-time; differs from a "
            "manual tracker-script edit (TC-CS-118), which DOES emit an event."
        ),
        "High", "Regression",
        "#3423 row 23; #3083 §Default script-url; #3286",
        "ttt-service/pm-tool-sync/default-script",
        "Per #3083: 'does NOT create an event in the project history' — important for test-case design to not assert event count on new-project syncs."
    ),
    (
        "TC-CS-117",
        "PM Tool sync — PMT-owned field edits do NOT produce project_event rows (silent update)",
        (
            "Env: ttt-qa-1. Existing PMT-synced project with a stable event "
            "history count."
        ),
        (
            "SETUP: Pick project with pm_id = X. Snapshot `SELECT count(*) "
            "FROM ttt_backend.project_event WHERE project_id = <ttt_id>` as "
            "N.\n"
            "SETUP: Via PMT admin — change PMT-OWNED fields: Supervisor, "
            "Observers, Type, Model. DO NOT change Status or Name (those are "
            "already covered by TC-CS-113; this test focuses on the "
            "secondary PMT-owned fields).\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `SELECT supervisor_id, ... FROM ttt_backend.project` "
            "reflects new PMT values.\n"
            "DB-CHECK: `SELECT count(*) FROM ttt_backend.project_event WHERE "
            "project_id = <ttt_id>` STILL = N (no new event).\n"
            "CLEANUP: Restore PMT values."
        ),
        (
            "Per #3083 §Event history rules: only TTT-owned field changes "
            "(tracker script, issue tracker, proxy) create events. All "
            "PMT-owned updates are silent."
        ),
        "High", "Regression",
        "#3423 row 23; #3083 §Event history",
        "ttt-service/pm-tool-sync/events",
        "Negative complement to TC-CS-118 (which asserts TTT-owned edits DO emit events)."
    ),
    (
        "TC-CS-118",
        "TTT-owned field edit (tracker script) via UI produces a project_event entry",
        (
            "Env: ttt-qa-1. Login as a user with EDIT permission on a PMT-"
            "synced project: Supervisor, Manager, Owner, or ADMIN."
        ),
        (
            "SETUP: Pick a PMT-synced project where user has EDIT permission. "
            "Snapshot `SELECT count(*) FROM ttt_backend.project_event WHERE "
            "project_id = <ttt_id>` as N.\n"
            "STEP 1: Login via browser.\n"
            "STEP 2: Navigate to Admin → Projects → find the project.\n"
            "STEP 3: Edit the `Скрипт синхронизации с трекерами` (tracker "
            "sync script) field — append a test comment like `// snavrockiy "
            "regression 2026-02-25`.\n"
            "STEP 4: Save.\n"
            "STEP 5: Verify success toast.\n"
            "WAIT: No cron needed — event is emitted synchronously by the "
            "edit endpoint.\n"
            "DB-CHECK: `SELECT count(*) FROM ttt_backend.project_event WHERE "
            "project_id = <ttt_id>` returns N+1.\n"
            "DB-CHECK: `GET /api/ttt/v1/projects/<id>/events` (or DB query on "
            "`project_event` ORDER BY date DESC LIMIT 1) shows a new entry "
            "with the tracker-script field in the delta.\n"
            "CLEANUP: Revert the tracker-script change (another UI edit or DB)."
        ),
        (
            "TTT-owned field edits via UI emit `project_event` entries; the "
            "sync cron does not suppress this path."
        ),
        "High", "Regression",
        "#3423 row 23; #3083 note 4 (snavrockiy regression)",
        "ttt-service/projects/events",
        "#3083 note 4 regression fix (2026-02-25 @omaksimova/@snavrockiy). UI-first; protects manual-edit history."
    ),
    (
        "TC-CS-119",
        "PM Tool sync — cs-id validation failure: unknown cs_id in payload logs ERROR + parks pm_id in pm_tool_sync_failed_project",
        (
            "Env: ttt-qa-1. Ability to construct (or observe) a PMT project "
            "whose manager cs_id is NOT present in `ttt_backend.employee.cs_id`."
        ),
        (
            "SETUP: Identify a PMT project whose manager's cs_id doesn't "
            "exist in TTT (e.g., freshly-added PMT manager not yet synced by "
            "CS sync on ttt-service). Alternatively: observe a historical "
            "ERROR in Graylog matching `\"Unable to sync PROJECT \"`.\n"
            "TRIGGER: `POST /api/ttt/v1/test/project/sync`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — `message:\"Unable to sync PROJECT "
            "<pm_id>\" AND level:3` (ERROR).\n"
            "DB-CHECK: `SELECT pm_id, retry_count, last_attempt_at FROM "
            "ttt_backend.pm_tool_sync_failed_project WHERE pm_id = <pm_id>` "
            "returns one row with `retry_count >= 1`.\n"
            "VERIFY LOG: next cron cycle includes retry markers: "
            "`\"PmTool Sync failed PROJECT ids count = N start\"` → "
            "`\"retry batch ...\"` → `\"count = N finished\"`.\n"
            "CLEANUP: Once CS sync brings the missing employee into TTT, "
            "next PM Tool sync retry succeeds and removes the row from "
            "pm_tool_sync_failed_project."
        ),
        (
            "cs-id validation rejects the project payload; pm_id parked for "
            "retry. Log marker `\"Unable to sync PROJECT\"` at ERROR level; "
            "retry runs next cycle per #3083 §cs-ids validation."
        ),
        "High", "Negative",
        "#3423 row 23; #3083 §IDs are cs-ids",
        "ttt-service/pm-tool-sync/cs-id-validation",
        "Failure path — integration between CS sync (row 6) ordering and PM Tool sync (row 23) on pod startup. Out-of-order start = transient failures until CS sync catches up."
    ),
    (
        "TC-CS-120",
        "PM Tool sync — Unleash `PM_TOOL_SYNC-qa-1` OFF: scheduler start/finish emit but no launcher markers",
        (
            "Env: ttt-qa-1. Unleash admin access to toggle `PM_TOOL_SYNC-qa-1`."
        ),
        (
            "SETUP: Via Unleash admin — disable `PM_TOOL_SYNC-qa-1`.\n"
            "SETUP: Snapshot last N PMT-sync log entries.\n"
            "WAIT: up to 16 min for the */15 cron boundary.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — ONLY scheduler-level markers "
            "fire: `\"Pm tool synchronization started\"` and `\"Pm tool "
            "synchronization finished\"`.\n"
            "VERIFY LOG: NO launcher markers — `\"PmTool Sync PROJECT "
            "started\"`, `\"PROJECT <pm_id> synched\"`, `\"PmTool Sync "
            "PROJECT finished\"` are ABSENT in the window.\n"
            "DB-CHECK: no writes to `ttt_backend.project` (verify by "
            "snapshotting `max(updated_at)` before and after).\n"
            "CLEANUP: Re-enable `PM_TOOL_SYNC-qa-1`."
        ),
        (
            "Launcher-level gate — per EXT-cron-jobs §Feature toggle gate: "
            "when flag OFF, only scheduler wrap emits, no launcher body."
        ),
        "High", "Negative",
        "#3423 row 23; EXT-cron-jobs §Feature-toggle gate row 23",
        "ttt-service/pm-tool-sync/feature-toggle",
        "Flag OFF → silent no-op at launcher level. Useful for emergency kill-switch without disabling the scheduler."
    ),
    (
        "TC-CS-121",
        "PM Tool sync — startup full sync via TttStartupApplicationListener re-fetches entire PMT catalog",
        (
            "Env: stage (CI restart permission). Unleash `PM_TOOL_SYNC-stage` "
            "ON."
        ),
        (
            "SETUP: Count existing PMT-synced projects: `SELECT count(*) "
            "FROM ttt_backend.project WHERE pm_id IS NOT NULL;` as N.\n"
            "SETUP: Trigger GitLab pipeline `restart-stage` for ttt-service.\n"
            "WAIT: 90-180 s for startup + full PMT sync (full sync is "
            "slower than delta — expect minutes).\n"
            "VERIFY LOG: Graylog TTT-STAGE — `\"PmTool Sync PROJECT started "
            "(fullSync=true)\"` (note the `true` variant — never emitted by "
            "cron path).\n"
            "VERIFY LOG: chain: scheduler start → entity start `(fullSync=true)` "
            "→ N × `PROJECT <pm_id> synched` → entity finish with "
            "`SyncResult(successCount=N, ...)` → scheduler finish.\n"
            "DB-CHECK: `max(updated_at)` on `ttt_backend.project` with "
            "`pm_id IS NOT NULL` is within the last 5 min (all rows "
            "re-touched).\n"
            "CLEANUP: None (restart is operational)."
        ),
        (
            "Startup-only full-sync path: `fullSync=true` token never "
            "appears in cron-triggered logs; only at application boot "
            "(per #3083 §No end-to-end 'full sync' via API)."
        ),
        "High", "Functional",
        "#3423 row 23; #3083 §No full sync via API; #3399",
        "ttt-service/pm-tool-sync/startup",
        "Per #3083 description: the scheduler always calls `sync(false)`. Only `TttStartupApplicationListener` calls `sync(true)`. Restart = only way to re-fetch entire catalog."
    ),
]


SUITES = [
    ("TS-CrossService-CronCSSync",     TS_CRON_CS_SYNC),
    ("TS-CrossService-CronPMToolSync", TS_CRON_PMTOOL_SYNC),
]


# ─── Sheet writer ─────────────────────────────────────────────────────────────

def _write_suite(wb, name, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLOR

    # R1 — back-link
    back = ws.cell(row=1, column=1,
                   value='=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")')
    back.font = FONT_LINK

    # R2 — blank (matches existing cross-service suite layout)

    # R3 — header
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        letter = c.column_letter
        if col_idx - 1 < len(COL_WIDTHS):
            ws.column_dimensions[letter].width = COL_WIDTHS[col_idx - 1]
    ws.freeze_panes = ws.cell(row=4, column=1)

    # R4+ — TC rows
    for row_offset, tc in enumerate(rows, start=0):
        r = 4 + row_offset
        fill = FILL_ALT if (row_offset % 2) else FILL_WHITE
        for col_idx, val in enumerate(tc, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1, 6, 7) else ALIGN_WRAP
            c.fill = fill
            c.border = BORDER_THIN


def main():
    wb = Workbook()
    wb.remove(wb.active)

    suite_row_labels = {
        "TS-CrossService-CronCSSync": "6, 10, 20",
        "TS-CrossService-CronPMToolSync": "23",
    }
    suite_row_map = [(name, suite_row_labels[name], len(rows)) for name, rows in SUITES]
    tc_count = sum(len(rows) for _, rows in SUITES)

    author_plan_overview(
        wb,
        domain="CrossService",
        home_subdir="cross-service",
        rows_covered="6, 10, 20, 23",
        tc_count=tc_count,
        suite_row_map=suite_row_map,
    )

    all_ids = []
    for suite_name, rows in SUITES:
        _write_suite(wb, suite_name, rows)
        all_ids.extend(r[0] for r in rows)

    wb.save(_TARGET)

    print(f"Generated: {_TARGET}")
    print(f"Suites written: {len(SUITES)}")
    print(f"TCs written: {len(all_ids)} ({all_ids[0]} ... {all_ids[-1]})")
    for name, rows in SUITES:
        print(f"  {name}: {len(rows)} TCs")


if __name__ == "__main__":
    main()
