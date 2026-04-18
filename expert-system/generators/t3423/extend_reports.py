#!/usr/bin/env python3
"""
Ticket #3423 — Reports Cluster Cron Test Cases (Phase B, session 136).

Extends ``test-docs/reports/reports.xlsx`` with 2 new cron-focused suites
covering jobs 1, 2, 3, 4, 5, 7 of the cron-testing collection:
    - TS-Reports-CronNotifications  (rows 1, 2, 3, 4, 7 — 15 TCs)
    - TS-Reports-BudgetNotifications (row 5 — 5 TCs)

Test IDs continue the home-module sequence — existing reports workbook ends
at TC-RPT-060; new cron TCs start at TC-RPT-101 and run through TC-RPT-120.

Idempotent: existing ``TS-Reports-CronNotifications`` / ``TS-Reports-BudgetNotifications``
sheets are removed before re-adding. Does NOT touch Plan Overview / Feature
Matrix / Risk Assessment / existing TS-Reports-* suites.

Run from repo root:
    python3 expert-system/generators/t3423/extend_reports.py

Canonical references:
    - expert-system/vault/exploration/tickets/t3423-investigation.md (scope)
    - expert-system/vault/external/EXT-cron-jobs.md §§Session 132 (markers)
    - expert-system/vault/patterns/email-notification-triggers.md (subject predicates)
    - Phase-A deltas (folded into rows 3, 4, 5 TC preconditions):
        Delta #1 (row 3): Template is REPORT_SHEET_CHANGED (not TASK_REPORT_CHANGED).
        Delta #2 (row 4): ZERO log markers — DB assertion reject.executor_notified=true
                          is the only non-email signal; DEBOUNCE_INTERVAL_MINUTES=5.
        Delta #3 (row 5): Three templates (EXCEEDED / NOT_REACHED / DATE_UPDATED);
                          SAFETY_INTERVAL_SECONDS=10.
    - GitLab bug regressions folded into TCs:
        #685 — officeId=9 ("Office not assigned") excluded from FORGOTTEN_REPORT.
        #559 — base FORGOTTEN_REPORT logic (underreported → send).
        #570 — confirms three budget templates + REPORT_SHEET_CHANGED exist.
        #892 — budget templates regression (all three paths).
        #3321 — report month closed + confirmation period open must still reject-notify.
        #3252 — deduplication across consecutive runs (debounce).
        #2289 — EXTENDED_PERIOD_REPORT must not reach office accountants.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_RPT_XLSX = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "reports", "reports.xlsx")
)


# ─── Styling (matches existing reports.xlsx rows + orange cron tab) ───────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR = "F4B084"  # Orange — distinguishes cron suites from core-module suites

COL_WIDTHS = [14, 40, 45, 72, 45, 10, 10, 35, 26, 35]
HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
    "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
]


# ─── Test case data ───────────────────────────────────────────────────────────
#
# Tuples: (test_id, title, preconditions, steps, expected, priority, type_,
#          req_ref, module, notes)
#
# Shape per t3423-investigation.md §Verification shape:
#   SETUP → Clock → Trigger → Wait → Verify (DB/UI/Email/Log) → CLEANUP
# API-trigger steps appear because cron test endpoints have no UI equivalent
# (CLAUDE.md §11 exception). UI-first rule still applies wherever there is a
# visible effect — e.g., manager editing hours in TC-RPT-106 is a UI action,
# and the verification side (email arrival + DB flag) uses the appropriate
# out-of-band predicates.

TS_CRON_NOTIFICATIONS = [
    # ─── Row 1 — Forgotten report (Mon, Fri 16:00 NSK) ────────────────────────
    (
        "TC-RPT-101",
        "Forgotten-report notification — underreported employee receives weekly email",
        (
            "Env: ttt-timemachine (clock controllable) or ttt-qa-1.\n"
            "Employee must be enabled, NOT in officeId=9 (per regression #685), "
            "and must have reported less than the weekly norm for the previous "
            "report period (taking vacations/absences into account).\n"
            "Query: SELECT e.login FROM ttt_backend.employee e "
            "JOIN ttt_backend.office o ON e.salary_office_id = o.id "
            "WHERE e.enabled = true AND o.id <> 9 "
            "AND NOT EXISTS ("
            "  SELECT 1 FROM ttt_backend.task_report tr "
            "  WHERE tr.executor_id = e.id "
            "  AND tr.report_date BETWEEN (current_date - INTERVAL '7 days') AND (current_date - INTERVAL '1 day')"
            ") ORDER BY random() LIMIT 1;"
        ),
        (
            "SETUP: Pick an underreported employee from the query above.\n"
            "SETUP: Clear any historical FORGOTTEN_REPORT emails via "
            "`roundcube-access --delete --subject '[QA1][TTT] Скорее всего вы забыли'` for that mailbox (or note baseline UID).\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-forgotten` (empty body, requires API_SECRET_TOKEN).\n"
            "WAIT: ~60 s for EmailSendScheduler (every 20 s) to dispatch.\n"
            "VERIFY LOG: Graylog TTT-QA-1 — `message:\"Report forgotten notification started\"` within last 3 min (debug level).\n"
            "VERIFY LOG: Graylog TTT-QA-1 — `message:\"Employees not reached reporting norm count = \"` with count ≥ 1.\n"
            "EMAIL-CHECK: `roundcube-access search --since today --subject '[QA1][TTT]' --subject 'Скорее всего вы забыли зарепортиться за прошлую неделю' --to <employee-email>` returns ≥ 1 message.\n"
            "CLEANUP: None (idempotent email noise; next run will re-evaluate)."
        ),
        (
            "Employee's mailbox receives FORGOTTEN_REPORT email (template key "
            "`FORGOTTEN_REPORT`). Graylog shows the scheduler marker and the "
            "\"Employees not reached reporting norm\" count reflects the test "
            "subject employee."
        ),
        "Critical", "Hybrid",
        "#3423 row 1; #559",
        "reports/cron/forgotten",
        "Subject per email-notification-triggers §Row 1."
    ),
    (
        "TC-RPT-102",
        "Forgotten-report — employees with salary_office_id=9 (Office not assigned) excluded",
        (
            "Env: ttt-qa-1.\n"
            "At least one enabled employee with `salary_office_id = 9` who is "
            "underreported for the previous report period.\n"
            "Query: SELECT e.login FROM ttt_backend.employee e "
            "WHERE e.enabled = true AND e.salary_office_id = 9 "
            "LIMIT 1;"
        ),
        (
            "SETUP: Identify target employee with officeId=9.\n"
            "SETUP: Clear historical FORGOTTEN_REPORT emails for this mailbox.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-forgotten`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog shows scheduler marker fired (global — "
            "`message:\"Report forgotten notification started\"`).\n"
            "VERIFY LOG: Verify `\"Sending notification to = \"` does NOT mention "
            "this employee's login in last 5 min.\n"
            "EMAIL-CHECK: `roundcube-access search --since today --to <employee-email> --subject 'Скорее всего'` returns 0 messages.\n"
            "DB-CHECK: No `notification_log` insert (if applicable) for this employee in last 5 min.\n"
            "CLEANUP: None."
        ),
        (
            "Even when the employee meets the underreport predicate, the office "
            "filter (officeId != 9) excludes them. Other employees may still be "
            "notified — verify the scope table via the service-level count marker."
        ),
        "High", "Hybrid",
        "#3423 row 1; #685 regression",
        "reports/cron/forgotten",
        "#685 fix guard — 'Office not assigned' employees never get weekly reminder."
    ),
    (
        "TC-RPT-103",
        "Forgotten-report — fully-reported employee excluded",
        (
            "Env: ttt-qa-1.\n"
            "Enabled employee with all workdays of the previous report period "
            "fully reported (effort_sum ≥ norm, ignoring vacations/sick leave).\n"
            "Query: SELECT e.login FROM ttt_backend.employee e "
            "JOIN (SELECT executor_id, SUM(effort) AS total_min "
            "  FROM ttt_backend.task_report "
            "  WHERE report_date BETWEEN (current_date - INTERVAL '7 days') "
            "  AND (current_date - INTERVAL '1 day') "
            "  GROUP BY executor_id) tr ON tr.executor_id = e.id "
            "WHERE e.enabled = true AND tr.total_min >= 40 * 60 "
            "ORDER BY random() LIMIT 1;"
        ),
        (
            "SETUP: Pick a fully-reported employee.\n"
            "SETUP: Clear historical FORGOTTEN_REPORT emails for this mailbox.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-forgotten`.\n"
            "WAIT: 60 s.\n"
            "EMAIL-CHECK: `roundcube-access search --since today --to <employee-email> --subject 'Скорее всего'` returns 0 messages.\n"
            "VERIFY LOG: `\"Sending notification to = \"` does NOT mention this employee.\n"
            "CLEANUP: None."
        ),
        (
            "Query predicate `not_reached_norm = true` filters out this "
            "employee; no email dispatched despite scheduler firing."
        ),
        "High", "Hybrid",
        "#3423 row 1; #559",
        "reports/cron/forgotten",
        "Norm-reached negative path."
    ),
    # ─── Row 2 — Forgotten-report delayed (Daily 16:30 NSK) ───────────────────
    (
        "TC-RPT-104",
        "Forgotten-delayed notification — employee who remained underreported receives catch-up email",
        (
            "Env: ttt-timemachine (clock control recommended).\n"
            "Enabled employee who received the first FORGOTTEN_REPORT on Mon/Fri "
            "16:00 NSK and has STILL not reported by 16:30 NSK the same day.\n"
            "Query: same base as TC-RPT-101."
        ),
        (
            "SETUP: Pick underreported employee (query from TC-RPT-101).\n"
            "SETUP: Clear historical FORGOTTEN_REPORT emails.\n"
            "TRIGGER 1: `POST /api/ttt/v1/test/reports/notify-forgotten` "
            "(simulate the 16:00 first pass).\n"
            "WAIT: 60 s; verify one FORGOTTEN_REPORT email dispatched.\n"
            "TRIGGER 2: `POST /api/ttt/v1/test/reports/notify-forgotten-delayed` "
            "(simulate the 16:30 delayed pass).\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog — `message:\"Report forgotten delayed notification started\"` (debug).\n"
            "VERIFY LOG: `message:\"Sending delayed notification to = \"` mentions the employee.\n"
            "EMAIL-CHECK: Roundcube — employee's mailbox now shows TWO messages matching the FORGOTTEN_REPORT subject, sent ~30 min apart.\n"
            "CLEANUP: None."
        ),
        (
            "Delayed path sends a second FORGOTTEN_REPORT email (same template, "
            "different invocation) for employees who remain underreported after "
            "the first pass."
        ),
        "High", "Hybrid",
        "#3423 row 2; #559",
        "reports/cron/forgotten-delayed",
        "Subject predicate identical to row 1 — disambiguate via send-time (~16:30) or via log marker ‘delayed notification to ='."
    ),
    (
        "TC-RPT-105",
        "Forgotten-delayed — employee who reported after the 16:00 pass is excluded",
        (
            "Env: ttt-timemachine.\n"
            "Employee received FORGOTTEN_REPORT at 16:00 NSK pass, then submitted "
            "reports to close the deficit before 16:30 NSK."
        ),
        (
            "SETUP: Pick underreported employee; dispatch first FORGOTTEN_REPORT via `/notify-forgotten`.\n"
            "SETUP: Via API (as that employee) — POST missing task_reports to reach weekly norm.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-forgotten-delayed`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: `message:\"Employee {} already reported hours for period from {} to {}\"` mentions this employee.\n"
            "EMAIL-CHECK: Roundcube — only ONE FORGOTTEN_REPORT email for this employee (from the first pass), not two.\n"
            "CLEANUP: Delete the seeded task_reports if test data hygiene requires."
        ),
        (
            "Delayed-pass skips employees whose norm gap was closed between "
            "16:00 and 16:30 (service-level log marker `already reported` confirms the skip)."
        ),
        "High", "Hybrid",
        "#3423 row 2; #559",
        "reports/cron/forgotten-delayed",
        "Negative path using the `already reported` marker as witness."
    ),
    # ─── Row 3 — REPORT_SHEET_CHANGED (every 5 min) ───────────────────────────
    (
        "TC-RPT-106",
        "Reports-changed notification — manager edits employee hours → REPORT_SHEET_CHANGED email",
        (
            "Env: ttt-qa-1.\n"
            "Manager with a direct report. Manager and employee must be different "
            "users (predicate `reporter_id != executor_id`).\n"
            "Query for manager+employee pair: SELECT mgr.login AS mgr_login, emp.login AS emp_login "
            "FROM ttt_backend.employee emp "
            "JOIN ttt_backend.employee mgr ON emp.manager_id = mgr.id "
            "WHERE emp.enabled = true AND mgr.enabled = true "
            "AND mgr.login <> 'pvaynmaster' ORDER BY random() LIMIT 1;"
        ),
        (
            "SETUP: Resolve the (manager, employee) pair.\n"
            "SETUP: Ensure the employee has at least one task on an active project "
            "shared with the manager (join via employee_project).\n"
            "SETUP: Clear historical REPORT_SHEET_CHANGED emails for the employee.\n"
            "STEP 1: Login as MANAGER via browser.\n"
            "STEP 2: Navigate to Employees & Contractors → find the subordinate.\n"
            "STEP 3: Open the weekly report sheet for YESTERDAY.\n"
            "STEP 4: Click an empty cell for an assigned task, enter 2 (hours), press Enter.\n"
            "STEP 5: Verify the cell saves with manager as reporter (reporter_id = manager.id, executor_id = employee.id).\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-changed`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog — `message:\"Start notification process\"` AND `message:\"Find: \"` with count ≥ 1.\n"
            "EMAIL-CHECK: Roundcube — employee mailbox receives email with template data referencing the manager's name and the edited date. Subject must match the REPORT_SHEET_CHANGED predicate (captured in session 136 during first real seeding run — record the exact subject as a P2 follow-up for patterns/email-notification-triggers.md row 3).\n"
            "DB-CHECK: `SELECT reporter_id, executor_id FROM ttt_backend.task_report WHERE report_date = current_date - 1 AND reporter_id = <mgr_id> AND executor_id = <emp_id>` returns ≥ 1 row.\n"
            "CLEANUP: Via API — DELETE the manager-reported task_report rows."
        ),
        (
            "Template `REPORT_SHEET_CHANGED` (NOT TASK_REPORT_CHANGED — per "
            "delta #1) fires for yesterday's manager-reported rows. Email "
            "arrives grouped by (manager, report_date)."
        ),
        "Critical", "Hybrid",
        "#3423 row 3; delta #1 (template key); #570",
        "reports/cron/changed",
        "Delta #1 folded: template key is REPORT_SHEET_CHANGED — scope table said TASK_REPORT_CHANGED (incorrect). Predicate: reporter_id ≠ executor_id AND report_date = yesterday."
    ),
    (
        "TC-RPT-107",
        "Reports-changed — manager edits OWN hours does NOT fire REPORT_SHEET_CHANGED",
        (
            "Env: ttt-qa-1.\n"
            "Any manager with assigned tasks (self-report scenario).\n"
            "Query: same as TC-RPT-106 mgr_login."
        ),
        (
            "SETUP: Pick a manager.\n"
            "SETUP: Clear historical REPORT_SHEET_CHANGED emails.\n"
            "STEP 1: Login as MANAGER via browser.\n"
            "STEP 2: Navigate to My Tasks → report 2 hours for yesterday on an assigned task.\n"
            "STEP 3: Verify `reporter_id = executor_id = manager.id` in DB.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-changed`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: `message:\"Find: 0 task reported by manager today\"` (or count excludes this row).\n"
            "EMAIL-CHECK: Roundcube — no REPORT_SHEET_CHANGED email for this manager.\n"
            "DB-CHECK: No row matches `reporter_id != executor_id AND report_date = current_date - 1` from this manager's edits.\n"
            "CLEANUP: Via API — DELETE the seeded row."
        ),
        (
            "Query predicate `reporter_id != executor_id` excludes self-reports; "
            "no email dispatched."
        ),
        "High", "Hybrid",
        "#3423 row 3",
        "reports/cron/changed",
        "Negative path — executor-equals-reporter filter."
    ),
    (
        "TC-RPT-108",
        "Reports-changed — idempotency on repeated trigger within the same day",
        (
            "Env: ttt-qa-1.\n"
            "Pre-state: TC-RPT-106 ran today and dispatched a REPORT_SHEET_CHANGED email."
        ),
        (
            "SETUP: Capture baseline email count for the affected employee "
            "(`roundcube-access count --since today --to <emp-email> --subject 'REPORT_SHEET_CHANGED-subject'`).\n"
            "TRIGGER 1: `POST /api/ttt/v1/test/reports/notify-changed` (within 5–10 min of the TC-RPT-106 run).\n"
            "WAIT: 60 s.\n"
            "TRIGGER 2: `POST /api/ttt/v1/test/reports/notify-changed` (second invocation).\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Both invocations log `Start notification process`; `Find: N` count is the SAME on both runs (the service re-selects yesterday's rows every time).\n"
            "EMAIL-CHECK: Roundcube — post-trigger email count equals baseline + N (where N = rows found first time). Second trigger does NOT add emails UNLESS new manager-reported rows landed in between.\n"
            "CLEANUP: None."
        ),
        (
            "Current code sends on EVERY invocation — idempotency is enforced by "
            "an `already_notified` DB flag on task_report (if present) or by the "
            "5-minute scheduler cadence, not by the test endpoint. Document "
            "observed behavior; file a design-issue note if duplicates appear."
        ),
        "Medium", "Hybrid",
        "#3423 row 3; #3252 (duplicate-suppression pattern)",
        "reports/cron/changed",
        "Design-issue probe — if two consecutive triggers send two emails, flag as bug and link to #3252."
    ),
    # ─── Row 4 — APPROVE_REJECT (every 5 min, ZERO log markers — delta #2) ────
    (
        "TC-RPT-109",
        "Reject notification — rejected period produces APPROVE_REJECT email and flips reject.executor_notified=true",
        (
            "Env: ttt-qa-1.\n"
            "Manager with a direct report who has approvable hours in an open "
            "confirmation period. Reject must have been created at least 5 min "
            "ago (DEBOUNCE_INTERVAL_MINUTES=5).\n"
            "Query (same manager+employee as TC-RPT-106)."
        ),
        (
            "SETUP: Via API (as manager) — POST approvable task_reports for the employee in the current confirmation period, then POST a reject for those hours via `POST /api/ttt/v1/rejects` with `cause = 'Session 136 test-reject'`.\n"
            "SETUP: Wait ≥ 5 min (or advance clock 5 min on timemachine) so the reject is past DEBOUNCE_INTERVAL_MINUTES.\n"
            "SETUP: Clear historical APPROVE_REJECT emails for this employee.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-rejected`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog — **no markers expected from the reject service** (delta #2). Absence of error markers (`\"notification FAILED\"`) is the success indicator.\n"
            "DB-CHECK: `SELECT executor_notified FROM ttt_backend.reject WHERE id = <reject_id>` returns `true`.\n"
            "EMAIL-CHECK: Roundcube — employee mailbox receives email with subject `[QA1][TTT] Ваши часы за период <start>-<end> были отклонены менеджером <Manager Russian Name>`.\n"
            "CLEANUP: Via API — DELETE the seeded reject and the task_reports."
        ),
        (
            "APPROVE_REJECT email dispatched; DB flag `reject.executor_notified` "
            "transitions false → true in the same run. No service-level log "
            "marker — assertion relies on DB state + email arrival only."
        ),
        "Critical", "Hybrid",
        "#3423 row 4; delta #2 (no markers)",
        "reports/cron/reject",
        "Delta #2 folded — zero log markers across the reject flow; DB flag + Roundcube subject are the only verification vectors. Subject per patterns/email-notification-triggers §Row 4."
    ),
    (
        "TC-RPT-110",
        "Reject notification — debounce: reject created <5 min ago is skipped by next run",
        (
            "Env: ttt-qa-1 or timemachine.\n"
            "Manager+employee pair (same as TC-RPT-109)."
        ),
        (
            "SETUP: Via API — POST approvable task_reports + a reject (cause='debounce test') **immediately before trigger** (i.e., 0 min ago).\n"
            "SETUP: Capture `reject.executor_notified = false`.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-rejected` within 60 s of reject creation.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `SELECT executor_notified FROM ttt_backend.reject WHERE id = <reject_id>` returns `false` (still, skipped due to debounce).\n"
            "EMAIL-CHECK: No APPROVE_REJECT email for this employee yet.\n"
            "WAIT: 5–6 min (or advance clock 5 min).\n"
            "TRIGGER 2: `POST /api/ttt/v1/test/reports/notify-rejected`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `executor_notified = true`.\n"
            "EMAIL-CHECK: Email now delivered.\n"
            "CLEANUP: Via API — DELETE reject + task_reports."
        ),
        (
            "First trigger: debounce skips reject (age < DEBOUNCE_INTERVAL_MINUTES). "
            "Second trigger (after 5 min): reject now past debounce → email + "
            "flag flip."
        ),
        "High", "Hybrid",
        "#3423 row 4; DEBOUNCE_INTERVAL_MINUTES=5",
        "reports/cron/reject",
        "Timing-sensitive. On non-timemachine envs, use real 5-min wait."
    ),
    (
        "TC-RPT-111",
        "Reject notification — regression #3321: report month closed + confirmation period open still dispatches email",
        (
            "Env: ttt-qa-1 or timemachine (requires accountant-level setup).\n"
            "Report month = closed (via Accounting → Reporting periods). "
            "Confirmation period = OPEN for the same interval.\n"
            "Manager+employee pair (same as TC-RPT-109)."
        ),
        (
            "SETUP: Via API (as accountant) — close the report month covering "
            "yesterday but keep the confirmation period open.\n"
            "SETUP: Via API (as manager) — POST approvable task_reports for the "
            "employee on yesterday, then POST a reject with `cause = 'regression #3321'`. Wait ≥ 5 min.\n"
            "SETUP: Clear historical APPROVE_REJECT emails.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-rejected`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `reject.executor_notified = true`.\n"
            "EMAIL-CHECK: Employee mailbox receives the APPROVE_REJECT email.\n"
            "CLEANUP: Via API — DELETE reject; reopen report month if this test polluted the env's period state."
        ),
        (
            "Regression #3321 — even with the report month closed, the reject "
            "notification still dispatches when the confirmation period is open. "
            "Pre-fix, executor_notified stayed false and no email arrived."
        ),
        "High", "Hybrid",
        "#3423 row 4; #3321 regression",
        "reports/cron/reject",
        "#3321 Frontend+backend interplay — verify both DB flag AND email arrival."
    ),
    (
        "TC-RPT-112",
        "Reject notification — idempotency: already-notified reject not re-sent on subsequent runs",
        (
            "Env: ttt-qa-1.\n"
            "Pre-state: TC-RPT-109 or TC-RPT-111 ran and set `reject.executor_notified = true`."
        ),
        (
            "SETUP: Confirm pre-state — `SELECT executor_notified FROM ttt_backend.reject WHERE id = <reject_id>` returns `true`; baseline email count for the employee.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/notify-rejected`.\n"
            "WAIT: 60 s.\n"
            "DB-CHECK: `executor_notified` stays `true` (unchanged).\n"
            "EMAIL-CHECK: Roundcube — email count unchanged (no new APPROVE_REJECT for this reject_id).\n"
            "CLEANUP: None."
        ),
        (
            "Rejects filtered by `executor_notified = false` — once true, the "
            "query excludes them on every subsequent run."
        ),
        "High", "Hybrid",
        "#3423 row 4",
        "reports/cron/reject",
        "Idempotency via DB-flag predicate."
    ),
    # ─── Row 7 — Extended-period cleanup (every 5 min) ────────────────────────
    (
        "TC-RPT-113",
        "Extended-period cleanup — expired extended report period transitions to closed",
        (
            "Env: ttt-qa-1 or timemachine.\n"
            "An `extended_report_period` row exists with `end_date < now()` and "
            "`closed = false`.\n"
            "Query: SELECT id, employee_id, start_date, end_date "
            "FROM ttt_backend.extended_report_period "
            "WHERE closed = false AND end_date < NOW() - INTERVAL '1 minute' "
            "LIMIT 1;\n"
            "If none exists, SETUP must create one."
        ),
        (
            "SETUP: If no expired extended period exists, create one via accountant UI/API — "
            "`POST /api/ttt/v1/extended-report-periods` with `end_date = NOW() - INTERVAL '1 min'` "
            "(or set end_date in the past via DB for timemachine envs).\n"
            "SETUP: Capture `closed` value before trigger.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/cleanup-extended`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog — `message:\"Extended period clean up started\"` → `message:\"Extended period clean up finished\"` pair within 3 min (both debug level).\n"
            "DB-CHECK: `SELECT closed FROM ttt_backend.extended_report_period WHERE id = <row_id>` returns `true`.\n"
            "DB-CHECK: No error markers — query `message:\"Unable to clean up timed out report extended periods\"` is empty in the same window.\n"
            "CLEANUP: Delete the seeded extended_report_period row if it was created for this test."
        ),
        (
            "Scheduler marks expired extended periods as closed; log markers "
            "show start/finish pair; no error marker."
        ),
        "Critical", "Hybrid",
        "#3423 row 7; #603",
        "reports/cron/extended-cleanup",
        "Happy-path for extended-period scheduler."
    ),
    (
        "TC-RPT-114",
        "Extended-period — regression #2289: EXTENDED_PERIOD_REPORT email not sent to office accountants",
        (
            "Env: ttt-qa-1 or timemachine.\n"
            "Env has at least one office accountant and at least one main "
            "accountant (role distinction is critical).\n"
            "Query: SELECT e.login, r.name FROM ttt_backend.employee e "
            "JOIN ttt_backend.employee_role er ON er.employee_id = e.id "
            "JOIN ttt_backend.role r ON r.id = er.role_id "
            "WHERE e.enabled = true "
            "AND r.name IN ('ROLE_OFFICE_ACCOUNTANT', 'ROLE_HEAD_ACCOUNTANT', 'ROLE_DEPARTMENT_MANAGER');"
        ),
        (
            "SETUP: Reproduce the full EXTENDED_PERIOD_REPORT trigger flow (per #2289 comment by @mpotter):\n"
            "  a. Close the report period (distinct from approve period).\n"
            "  b. As accountant — reopen the individual report period for a target employee.\n"
            "  c. As that employee — report hours on a project.\n"
            "  d. As accountant — close the individual report period (or wait 1 hour for auto-close).\n"
            "SETUP: Identify office accountant, main accountant, and project manager of the project above.\n"
            "SETUP: Clear historical EXTENDED_PERIOD_REPORT emails from ALL three mailboxes (office accountant, main accountant, PM).\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/cleanup-extended` (or wait for scheduler).\n"
            "WAIT: 60 s.\n"
            "EMAIL-CHECK: Roundcube — MAIN accountant receives EXTENDED_PERIOD_REPORT; PM receives EXTENDED_PERIOD_REPORT; OFFICE accountant does NOT receive it.\n"
            "VERIFY LOG: Graylog — extended-cleanup markers fire; `\"Sending notification to = \"` mentions only main accountant + PM logins, NOT the office accountant login.\n"
            "CLEANUP: Revert period state; DELETE seeded task_reports."
        ),
        (
            "Regression #2289 — office accountants must NOT receive "
            "EXTENDED_PERIOD_REPORT; only project managers + main accountants do."
        ),
        "High", "Hybrid",
        "#3423 row 7; #2289 regression",
        "reports/cron/extended-cleanup",
        "#2289 fix — tighten recipient audience. Setup flow is complex — reference comment thread on ticket."
    ),
    (
        "TC-RPT-115",
        "Extended-period cleanup — idempotency: already-closed periods remain untouched",
        (
            "Env: ttt-qa-1.\n"
            "Pre-state: TC-RPT-113 already closed the target extended-period row."
        ),
        (
            "SETUP: Confirm `closed = true` in `ttt_backend.extended_report_period WHERE id = <row_id>`; snapshot `updated_at`.\n"
            "TRIGGER: `POST /api/ttt/v1/test/reports/cleanup-extended`.\n"
            "WAIT: 60 s.\n"
            "VERIFY LOG: Graylog — start/finish pair fires, but service should touch 0 rows.\n"
            "DB-CHECK: `SELECT updated_at FROM ttt_backend.extended_report_period WHERE id = <row_id>` unchanged from snapshot.\n"
            "CLEANUP: None."
        ),
        (
            "Query filter `closed = false` excludes this row; cleanup is a no-op."
        ),
        "Medium", "Hybrid",
        "#3423 row 7",
        "reports/cron/extended-cleanup",
        "Idempotency via closed-flag predicate."
    ),
]


TS_BUDGET_NOTIFICATIONS = [
    # ─── Row 5 — BudgetNotification (every 10 s, three templates, delta #3) ───
    (
        "TC-RPT-116",
        "Budget notification — EXCEEDED: first crossing of budget_limit dispatches BUDGET_NOTIFICATION_EXCEEDED",
        (
            "Env: ttt-qa-1.\n"
            "A `BudgetNotification` entity seeded for a (project, employee, "
            "date range) with budget_limit = 4 hours. Target employee has NOT "
            "yet reported hours in the window.\n"
            "Query to pick a project+employee: SELECT p.id, p.name, e.login "
            "FROM ttt_backend.project p "
            "JOIN ttt_backend.employee_project ep ON ep.project_id = p.id "
            "JOIN ttt_backend.employee e ON e.id = ep.employee_id "
            "WHERE p.status = 'ACTIVE' AND e.enabled = true "
            "AND e.login <> 'pvaynmaster' ORDER BY random() LIMIT 1;"
        ),
        (
            "SETUP: Pick an active (project, employee, manager-watcher) triple.\n"
            "SETUP: Via API — `POST /api/ttt/v1/budget-notifications` with watcher_id = <mgr_id>, "
            "project_id = <project_id>, employee_id = <emp_id>, budget_limit = 240 "
            "(4 hours in minutes), start_date = today, end_date = today.\n"
            "SETUP: Clear historical BUDGET_NOTIFICATION_EXCEEDED emails from watcher's mailbox.\n"
            "STEP (UI): Login as the employee; navigate to My Tasks; report 5 hours on a task of <project_id> for today.\n"
            "STEP: Wait ≥ 10 s (SAFETY_INTERVAL_SECONDS) so the task_report is considered stable.\n"
            "TRIGGER: `POST /api/ttt/v1/test/budgets/notify`.\n"
            "WAIT: 30 s (scheduler cadence + email-send cadence).\n"
            "VERIFY LOG: Graylog — `message:\"Budget notification job is done\"` fires (info).\n"
            "DB-CHECK: `SELECT limit_reached_date FROM ttt_backend.budget_notification WHERE id = <notif_id>` is non-null and equals today.\n"
            "EMAIL-CHECK: Roundcube — watcher mailbox receives email whose template key maps to `BUDGET_NOTIFICATION_EXCEEDED`.\n"
            "CLEANUP: Via API — DELETE the seeded BudgetNotification and the task_report."
        ),
        (
            "First budget crossing — BudgetServiceImpl detects prev=null → "
            "current=today transition; dispatches EXCEEDED template to watcher."
        ),
        "Critical", "Hybrid",
        "#3423 row 5; delta #3 (three templates); #892 regression",
        "reports/cron/budget",
        "Delta #3 folded — template is BUDGET_NOTIFICATION_EXCEEDED. SAFETY_INTERVAL_SECONDS=10 means recent task_reports are skipped."
    ),
    (
        "TC-RPT-117",
        "Budget notification — NOT_REACHED: exceeded budget then reduced dispatches BUDGET_NOTIFICATION_NOT_REACHED",
        (
            "Env: ttt-qa-1.\n"
            "Pre-state: TC-RPT-116 dispatched EXCEEDED; `limit_reached_date = today`."
        ),
        (
            "SETUP: Confirm pre-state — `limit_reached_date` non-null for <notif_id>.\n"
            "STEP (UI): Login as the employee; navigate to My Tasks; edit the 5h entry down to 2h for today.\n"
            "STEP: Wait ≥ 10 s.\n"
            "TRIGGER: `POST /api/ttt/v1/test/budgets/notify`.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: `limit_reached_date` now `null` (prev=today → current=null transition).\n"
            "EMAIL-CHECK: Watcher mailbox receives email matching BUDGET_NOTIFICATION_NOT_REACHED.\n"
            "VERIFY LOG: `\"Budget notification job is done\"` fires.\n"
            "CLEANUP: Delete seeded entities."
        ),
        (
            "Budget fell below limit — NOT_REACHED template dispatched. "
            "limit_reached_date cleared back to null."
        ),
        "High", "Hybrid",
        "#3423 row 5; delta #3; #892 regression",
        "reports/cron/budget",
        "Template key BUDGET_NOTIFICATION_NOT_REACHED."
    ),
    (
        "TC-RPT-118",
        "Budget notification — DATE_UPDATED: hours moved to a different day keeping budget exceeded",
        (
            "Env: ttt-qa-1.\n"
            "Pre-state: TC-RPT-116 dispatched EXCEEDED for today's report."
        ),
        (
            "SETUP: Confirm `limit_reached_date = today` in DB.\n"
            "STEP (UI): Login as employee; move today's 5h entry to yesterday (delete today's report, create an identical one for yesterday with 5h).\n"
            "SETUP: Via API (if UI does not allow yesterday edit) — DELETE today's task_report; POST the same hours on yesterday.\n"
            "STEP: Wait ≥ 10 s.\n"
            "TRIGGER: `POST /api/ttt/v1/test/budgets/notify`.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: `limit_reached_date` changed from today → yesterday.\n"
            "EMAIL-CHECK: Watcher receives email matching BUDGET_NOTIFICATION_DATE_UPDATED.\n"
            "CLEANUP: Delete seeded entities."
        ),
        (
            "Budget still exceeded but on a different date — DATE_UPDATED "
            "template dispatched."
        ),
        "High", "Hybrid",
        "#3423 row 5; delta #3; #892 regression",
        "reports/cron/budget",
        "Template key BUDGET_NOTIFICATION_DATE_UPDATED."
    ),
    (
        "TC-RPT-119",
        "Budget notification — safety window: task_report younger than 10 s is skipped by next run",
        (
            "Env: ttt-qa-1.\n"
            "Seeded BudgetNotification (same shape as TC-RPT-116)."
        ),
        (
            "SETUP: Seed BudgetNotification as in TC-RPT-116; confirm `limit_reached_date = null`.\n"
            "STEP (UI): Report 5h on the project for today.\n"
            "TRIGGER: `POST /api/ttt/v1/test/budgets/notify` within 5 s of the report (definitely < SAFETY_INTERVAL_SECONDS=10).\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: `limit_reached_date` still `null` (skipped — too recent).\n"
            "EMAIL-CHECK: No BUDGET_NOTIFICATION_EXCEEDED email yet.\n"
            "WAIT: additional 15 s (total ~20 s past the report).\n"
            "TRIGGER 2: `POST /api/ttt/v1/test/budgets/notify`.\n"
            "WAIT: 30 s.\n"
            "DB-CHECK: `limit_reached_date = today` now.\n"
            "EMAIL-CHECK: Email arrives.\n"
            "CLEANUP: Delete seeded entities."
        ),
        (
            "SAFETY_INTERVAL_SECONDS=10 guards against flapping; first trigger "
            "skips, second trigger (after safety window) dispatches."
        ),
        "Medium", "Hybrid",
        "#3423 row 5; delta #3 (SAFETY_INTERVAL_SECONDS=10)",
        "reports/cron/budget",
        "Timing-sensitive. Real-time env; cannot advance clock easily here."
    ),
    (
        "TC-RPT-120",
        "Budget notification — recipient: email goes to watcher_id, not employee",
        (
            "Env: ttt-qa-1.\n"
            "Seeded BudgetNotification with watcher_id = <mgr_id>, employee_id "
            "= <emp_id>, and mgr_id ≠ emp_id."
        ),
        (
            "SETUP: Seed as in TC-RPT-116 with distinct watcher and employee.\n"
            "SETUP: Clear BUDGET_NOTIFICATION_EXCEEDED emails from BOTH mailboxes.\n"
            "STEP (UI): Employee reports 5h for today on the project.\n"
            "STEP: Wait 15 s (past SAFETY_INTERVAL).\n"
            "TRIGGER: `POST /api/ttt/v1/test/budgets/notify`.\n"
            "WAIT: 30 s.\n"
            "EMAIL-CHECK: Watcher's mailbox has ONE new EXCEEDED email; employee's mailbox has ZERO new BUDGET_* emails.\n"
            "VERIFY LOG: `\"Budget notification job is done\"` fires.\n"
            "CLEANUP: Delete seeded entities."
        ),
        (
            "Email sent only to `notification.watcherId` resolved via "
            "employeeService.find(watcherId); employee never receives a copy."
        ),
        "High", "Hybrid",
        "#3423 row 5",
        "reports/cron/budget",
        "Recipient-correctness guard."
    ),
]


SUITES = [
    ("TS-Reports-CronNotifications",   TS_CRON_NOTIFICATIONS),
    ("TS-Reports-BudgetNotifications", TS_BUDGET_NOTIFICATIONS),
]


# ─── Sheet writer ─────────────────────────────────────────────────────────────

def _write_suite(wb, name, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLOR

    # R1 — back-link
    back = ws.cell(row=1, column=1,
                   value='=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")')
    back.font = FONT_LINK

    # R2 — header
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        letter = c.column_letter
        if col_idx - 1 < len(COL_WIDTHS):
            ws.column_dimensions[letter].width = COL_WIDTHS[col_idx - 1]
    ws.freeze_panes = ws.cell(row=3, column=1)

    # R3+ — TC rows
    for row_offset, tc in enumerate(rows, start=0):
        r = 3 + row_offset
        fill = FILL_ALT if (row_offset % 2) else FILL_WHITE
        for col_idx, val in enumerate(tc, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1, 6, 7) else ALIGN_WRAP
            c.fill = fill
            c.border = BORDER_THIN


def main():
    if not os.path.isfile(_RPT_XLSX):
        raise SystemExit(f"reports.xlsx not found at {_RPT_XLSX}")

    wb = load_workbook(_RPT_XLSX)

    all_ids = []
    for suite_name, rows in SUITES:
        _write_suite(wb, suite_name, rows)
        all_ids.extend(r[0] for r in rows)

    wb.save(_RPT_XLSX)

    print(f"Extended: {_RPT_XLSX}")
    print(f"Suites added/updated: {len(SUITES)}")
    print(f"TCs written: {len(all_ids)} ({all_ids[0]} ... {all_ids[-1]})")
    for name, rows in SUITES:
        print(f"  {name}: {len(rows)} TCs")


if __name__ == "__main__":
    main()
