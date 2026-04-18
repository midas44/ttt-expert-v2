#!/usr/bin/env python3
"""
Ticket #3423 — Statistics Cluster Cron Test Cases (Phase B, session 138).

Extends ``test-docs/statistics/statistics.xlsx`` with 1 new cron-focused suite
covering row 22 of the cron-testing collection:
    - TS-Stat-CronStatReportSync (row 22 — 8 TCs: TC-STAT-077..084)

Test IDs continue the home-module sequence — the pre-existing statistics
workbook ended at TC-STAT-076 (TS-Stat-Regression). TC-STAT-077 starts a new
contiguous block for cron coverage. Keeping the numbering tight rather than
jumping to TC-STAT-100 so the workbook reads naturally in tab order.

Idempotent: existing ``TS-Stat-CronStatReportSync`` sheet is removed before
re-adding. Does NOT touch Plan Overview / Feature Matrix / Risk Assessment /
the 8 pre-existing TS-Stat-* suites (including TS-Stat-CacheSync which is
UI cache consistency, NOT cron-focused — the two suites intentionally coexist).

Tab color is ``F4B084`` (orange) — matches the convention established by the
cross-service cron suites (session 137), distinguishing cron-focused suites
from general UI/regression suites in the same workbook.

Run from repo root:
    python3 expert-system/generators/t3423/extend_statistics.py

Canonical references:
    - expert-system/vault/exploration/tickets/t3423-investigation.md (scope)
    - expert-system/vault/external/EXT-cron-jobs.md §Session 131.4 (row 22 markers)
    - expert-system/vault/exploration/tickets/3262-ticket-findings.md §3 Job 22 seeds
    - expert-system/vault/investigations/statistics-caffeine-caching-performance-3337.md
    - Phase-A deltas folded into TCs:
        Delta #8 (row 22 log level): Failure logged at INFO (level:6), NOT
                ERROR (level:3). Catch block uses log.info(...) for failure —
                Graylog level:3 filters MISS failures from this job.
    - GitLab regression tickets folded into TCs:
        #3345 Bug 1 — Employment-period filtering (pre-hire and post-leave
                        rows were incorrectly persisted).
        #3345 Bug 2 — Day-off reschedule did not trigger month_norm update.
        #3346 — Periodic cron not firing at 04:00 NSK (bug #895498).
        #3337 — Caffeine L1 + statistic_report L2 cache architecture;
                 sick-leave month_norm propagation; scoped delete regression.
    - Cron schedule: "0 0 4 * * *" (04:00 NSK daily — moved from 03:00 post-#3346).
    - Test endpoints (bypass scheduler, direct service call):
        POST /api/ttt/v1/test/statistic-reports/full-sync     — full refresh
        POST /api/ttt/v1/test/statistic-reports/sync          — optimized refresh
        POST /api/ttt/v1/test/statistic-reports               — single row (generic)
    - Event-driven fan-out: topic `TTT_BACKEND_EMPLOYEE_TOPIC`, routing
      `employee-month-norm-context-calculated`. StatisticReportUpdateEventType:
      INITIAL_SYNC, VACATION_CHANGES, SICK_LEAVE_CHANGES. Wait 1-2 min after
      trigger for RabbitMQ consumer to apply changes (async contract).
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_STAT_XLSX = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "statistics", "statistics.xlsx")
)


# ─── Styling (matches existing statistics.xlsx + orange cron tab) ─────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR = "F4B084"  # Orange — cron suites (matches session 137 cross-service convention)

COL_WIDTHS = [14, 40, 40, 55, 40, 10, 10, 30, 14, 35]
HEADERS = [
    "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
    "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
]


# ─── Test case data ───────────────────────────────────────────────────────────
#
# Tuples: (test_id, title, preconditions, steps, expected, priority, type_,
#          req_ref, module, notes)
#
# Shape per t3423-investigation.md §Verification shape:
#   SETUP → (Trigger or Wait) → Verify (DB/Log) → CLEANUP
#
# UI-first rule: row 22 is a cron/API job with NO UI trigger. All 8 TCs use
# API test endpoints for trigger + DB/Graylog for verification. This is the
# CLAUDE.md §11 "test endpoints" exception — no UI equivalent exists for
# force-triggering the statistic-report sync.
#
# TC slate:
#   TC-STAT-077 — Cron fires at 04:00 NSK + ShedLock + start/finish markers
#   TC-STAT-078 — Delta #8: failure logged at INFO (level:6), NOT ERROR
#   TC-STAT-079 — #3345 Bug 1: employment-period filter (pre-hire + post-leave)
#   TC-STAT-080 — #3345 Bug 2: day-off reschedule → RabbitMQ → month_norm update
#   TC-STAT-081 — #3337: sick-leave create → event → month_norm+reported_effort
#   TC-STAT-082 — #3337: scoped event → unrelated employees' rows preserved
#   TC-STAT-083 — #3346: manual row delete + optimized sync restores row
#   TC-STAT-084 — Contract: full (prev+curr year) vs optimized (prev+curr month)


TS_CRON_STAT_REPORT_SYNC = [
    # ─── TC-STAT-077 — Scheduled cron fires + ShedLock + markers ───────────────
    (
        "TC-STAT-077",
        "StatisticReportScheduler — cron fires at 04:00 NSK, ShedLock acquired/released, start/finish markers emit",
        (
            "Env: ttt-qa-1 (primary) or any env where `ttt.statistic-report.cron` "
            "= `0 0 4 * * *` in `application.yml` and ttt-service is deployed.\n"
            "ShedLock table accessible via postgres MCP: "
            "`SELECT name, lock_until, locked_at, locked_by FROM ttt_backend.shedlock "
            "WHERE name = 'StatisticReportScheduler.sync';`.\n"
            "Graylog API token for stream `TTT-QA-1` (or TTT-STAGE for stage env).\n"
            "No other pod is actively holding the ShedLock row at assertion time."
        ),
        (
            "SETUP: Snapshot current `shedlock` row for name "
            "`StatisticReportScheduler.sync` — record `lock_until`, `locked_at`.\n"
            "SETUP: Snapshot latest Graylog event for marker "
            "`\"Periodic statistic report sync started...\"` on the env's stream "
            "(record max timestamp).\n"
            "WAIT: Up to 16 h for the next 04:00 NSK boundary. Alternative "
            "(preferred for QA): trigger manually via "
            "`POST /api/ttt/v1/test/statistic-reports/sync` — the scheduler "
            "and the test endpoint both delegate to "
            "`StatisticReportScheduler.sync()` / service method, emitting the "
            "same markers.\n"
            "VERIFY LOG: Graylog `stream:TTT-QA-1 AND message:\"Periodic "
            "statistic report sync started\"` — new event after snapshot "
            "timestamp (match exact string including trailing ellipsis `...`).\n"
            "VERIFY LOG: Graylog same stream — "
            "`message:\"Periodic statistic report sync finished.\"` — paired "
            "with the start marker (matching request correlation id if "
            "present).\n"
            "DB-CHECK: `ttt_backend.shedlock` row `lock_until` advanced past "
            "snapshot (lock acquired), and after finish marker, `lock_until` "
            "is in the past (lock released — ready for next run).\n"
            "CLEANUP: None."
        ),
        (
            "Start and finish INFO markers appear in sequence on the env's "
            "Graylog stream. ShedLock row refreshed (lock_until advanced) "
            "during the sync window and released afterward. No concurrent "
            "pod re-entered the method while locked."
        ),
        "Critical", "Functional",
        "#3423 row 22; EXT-cron-jobs §Session 131.4",
        "ttt-service/statistic-report/cron",
        "Scheduler and test endpoint share the same service method, so "
        "markers are identical. Verify both marker strings EXACTLY — "
        "trailing ellipsis on start, trailing period on finish. Matches "
        "StatisticReportScheduler.sync() log.info calls verbatim."
    ),

    # ─── TC-STAT-078 — Delta #8: failure logged at INFO (level:6), NOT ERROR ──
    (
        "TC-STAT-078",
        "Delta #8 regression — statistic-report sync failure is logged at INFO level (level:6), NOT ERROR (level:3)",
        (
            "Env: ttt-qa-1 with ability to induce a controlled failure in "
            "the sync path. One practical route: momentarily break the "
            "RabbitMQ connection (drop the `TTT_BACKEND_EMPLOYEE_TOPIC` "
            "exchange binding) or exhaust a DB connection pool. Requires "
            "infra-team coordination in shared envs — prefer timemachine "
            "or a disposable dev env for this test.\n"
            "Graylog API token; ability to filter by level (level:3 = ERROR, "
            "level:6 = INFO in Graylog's Syslog level mapping)."
        ),
        (
            "SETUP: Confirm `ttt.statistic-report.cron` active. "
            "Snapshot latest `level:3` (ERROR) events on stream.\n"
            "SETUP: Induce controlled failure — e.g., revoke DB write "
            "permission for the ttt-service DB user on "
            "`ttt_backend.statistic_report` for ~60 s, or stop the RabbitMQ "
            "broker briefly (infra coordination required).\n"
            "TRIGGER: `POST /api/ttt/v1/test/statistic-reports/sync` "
            "(expect non-200 OR silent swallow — scheduler catches and logs).\n"
            "WAIT: 5 s for log propagation.\n"
            "VERIFY LOG (negative): Graylog `stream:TTT-QA-1 AND level:3 "
            "AND message:\"statistic report\"` — NO new ERROR events. "
            "This is the regression assertion — if ERROR events appear, the "
            "bug was accidentally fixed and this TC must be re-triaged.\n"
            "VERIFY LOG (positive): Graylog same stream "
            "`level:6 AND message:\"Periodic statistic report sync failed "
            "with cause\"` — new INFO event captured after trigger with the "
            "exception detail in the cause payload.\n"
            "CLEANUP: Restore DB permissions / RabbitMQ broker. Trigger "
            "the endpoint again to confirm success marker returns."
        ),
        (
            "Failure event is emitted at INFO level (level:6) with the "
            "literal prefix `Periodic statistic report sync failed with "
            "cause: `. NO event at ERROR level (level:3) for the same "
            "failure. Alerting rules watching `level:3 AND message:\""
            "statistic report\"` would not fire — this is the design "
            "issue, and this TC acts as a regression guard in case a "
            "future patch \"fixes\" it by elevating to ERROR (which would "
            "require updating alerting rules too)."
        ),
        "High", "Regression",
        "#3423 row 22; Delta #8; EXT-cron-jobs §Session 131.4 Bug",
        "ttt-service/statistic-report/logging",
        "Ticket body flags this as a design issue, not a defect to be "
        "fixed right now. The TC exists to detect silent behavior changes "
        "on either side (INFO→ERROR or ERROR→INFO). Needs infra support "
        "to reliably induce failure — if skipped in CI, run manually "
        "during regression passes."
    ),

    # ─── TC-STAT-079 — #3345 Bug 1: employment-period filter ───────────────────
    (
        "TC-STAT-079",
        "#3345 Bug 1 regression — employment-period filter excludes pre-hire and post-leave months",
        (
            "Env: ttt-qa-1 with at least 2 employees having non-trivial "
            "`employee_period` ranges (one hired mid-year, one dismissed "
            "mid-year). Reference data from #3345 QA: omaksimova hired "
            "2025-09-01; jsaidov dismissed 2025-11-12. If those specific "
            "employees are not on the env, pick equivalents:\n"
            "  SELECT e.login, ep.start_date, ep.end_date\n"
            "  FROM ttt_backend.employee e\n"
            "  JOIN ttt_vacation.employee_period ep ON ep.employee_id = e.id\n"
            "  WHERE ep.start_date > (CURRENT_DATE - INTERVAL '12 months')\n"
            "    OR ep.end_date BETWEEN (CURRENT_DATE - INTERVAL '12 months') AND CURRENT_DATE\n"
            "  LIMIT 10;\n"
            "Pick one pre-hire and one post-leave candidate."
        ),
        (
            "SETUP: Record candidate employees' login, id, start_date, "
            "end_date.\n"
            "TRIGGER: `POST /api/ttt/v1/test/statistic-reports/full-sync` "
            "(full sync — covers previous + current year scope).\n"
            "WAIT: 120 s for RabbitMQ event fan-out to complete (async "
            "contract per [[3262-ticket-findings]] §3D note 5).\n"
            "DB-CHECK (pre-hire employee): `SELECT count(*) FROM "
            "ttt_backend.statistic_report sr WHERE sr.employee_id = <id> "
            "AND sr.month_start < '<start_date-month>';` → must be 0. "
            "Confirms no rows exist for months before the employee was "
            "hired.\n"
            "DB-CHECK (post-leave employee): `SELECT count(*) FROM "
            "ttt_backend.statistic_report sr WHERE sr.employee_id = <id> "
            "AND sr.month_start > '<end_date-month>';` → must be 0. "
            "Confirms no rows exist for months after the employee left.\n"
            "DB-CHECK (sanity — in-period rows present): "
            "`SELECT count(*) FROM ttt_backend.statistic_report sr "
            "WHERE sr.employee_id = <id> AND sr.month_start BETWEEN "
            "'<start_date-month>' AND COALESCE('<end_date-month>', "
            "CURRENT_DATE);` → must be > 0 (sync produced in-period "
            "rows — otherwise the whole sync failed silently).\n"
            "CLEANUP: None."
        ),
        (
            "For pre-hire employee: statistic_report has zero rows before "
            "start_date month. For post-leave employee: zero rows after "
            "end_date month. In-period rows exist (sync worked). "
            "Regression for bug fixed in !5101 — before the fix, pre-hire "
            "and post-leave rows were incorrectly persisted."
        ),
        "Critical", "Regression",
        "#3345 Bug 1 (MR !5101); #3423 row 22",
        "ttt-service/statistic-report/sync",
        "Mid-month hire/leave edge case remains open as #3356 — currently "
        "any month in which the employee was employed ANY day is included. "
        "This TC is month-granularity, not day-granularity, so it does "
        "not exercise the #3356 boundary. Do NOT tighten assertions to "
        "include/exclude partial months without coordinating with #3356."
    ),

    # ─── TC-STAT-080 — #3345 Bug 2: day-off reschedule → month_norm update ────
    (
        "TC-STAT-080",
        "#3345 Bug 2 regression — day-off reschedule triggers month_norm recalculation via RabbitMQ event",
        (
            "Env: ttt-qa-1. Active employee with at least one existing "
            "day-off in the current or next month whose reschedule can be "
            "tested without affecting live data.\n"
            "Query:\n"
            "  SELECT d.id, d.employee_id, e.login, d.day, d.day_off_type\n"
            "  FROM ttt_vacation.day_off d\n"
            "  JOIN ttt_backend.employee e ON e.id = d.employee_id\n"
            "  WHERE d.day >= CURRENT_DATE\n"
            "    AND d.day < CURRENT_DATE + INTERVAL '60 days'\n"
            "    AND e.enabled = true\n"
            "  ORDER BY random() LIMIT 1;\n"
            "RabbitMQ `TTT_BACKEND_EMPLOYEE_TOPIC` healthy and consumer "
            "bound on routing `employee-month-norm-context-calculated`."
        ),
        (
            "SETUP: Record chosen day-off id, employee_id, login, original "
            "day, day_off_type.\n"
            "SETUP: Snapshot `month_norm` for the affected month: "
            "`SELECT month_norm FROM ttt_backend.statistic_report WHERE "
            "employee_id = <emp_id> AND month_start = "
            "date_trunc('month', '<original_day>'::date);` → record as "
            "`norm_before`.\n"
            "TRIGGER (UI-first): Log into TTT as the employee or a "
            "permitted manager. Navigate to Calendar → Day-off page. "
            "Reschedule the day-off to a date within the same month (e.g., "
            "move from the 10th to the 11th). Confirm reschedule via UI.\n"
            "  Alternative (API): `PATCH /api/vacation/v1/day-offs/<id>` "
            "with new `day` field.\n"
            "WAIT: 60 s for the RabbitMQ event to propagate (routing "
            "`employee-month-norm-context-calculated`; "
            "StatisticReportUpdateEventType.VACATION_CHANGES fires).\n"
            "DB-CHECK: `SELECT month_norm FROM ttt_backend.statistic_report "
            "WHERE employee_id = <emp_id> AND month_start = "
            "date_trunc('month', '<new_day>'::date);` → record as "
            "`norm_after`. Value must be recomputed (can be equal to "
            "`norm_before` if both days are workdays with the same hours; "
            "in that case, verify the `updated_at` column advanced to "
            "confirm the row was touched).\n"
            "VERIFY LOG: Graylog — event with routing "
            "`employee-month-norm-context-calculated` emitted within 60 s "
            "window after trigger.\n"
            "CLEANUP: Restore day-off to original date via UI or API."
        ),
        (
            "`statistic_report.month_norm` OR `updated_at` reflects the "
            "day-off reschedule. RabbitMQ event fired on the employee-norm "
            "routing within 60 s. Before the #3345 fix, this did nothing "
            "(event was not wired); the TC guards against that regression."
        ),
        "High", "Regression",
        "#3345 Bug 2; #3423 row 22",
        "ttt-service/statistic-report/event-handler",
        "If both original and new day are workdays with identical "
        "norm-hour semantics, month_norm might not change numerically — "
        "in that case, asserting on updated_at advancing is the primary "
        "signal. Use a cross-month reschedule (e.g., from Sept 30 to "
        "Oct 1) for stronger numerical assertions."
    ),

    # ─── TC-STAT-081 — #3337: sick-leave create → event → updates ─────────────
    (
        "TC-STAT-081",
        "#3337 regression — sick-leave creation updates month_norm and reported_effort via SICK_LEAVE_CHANGES event",
        (
            "Env: ttt-qa-1. Employee without a sick-leave in the target "
            "month, with at least 3 working days available in that month "
            "for a synthetic sick-leave.\n"
            "Query:\n"
            "  SELECT e.id, e.login\n"
            "  FROM ttt_backend.employee e\n"
            "  WHERE e.enabled = true\n"
            "    AND NOT EXISTS (\n"
            "      SELECT 1 FROM ttt_vacation.sick_leave sl\n"
            "      WHERE sl.employee_id = e.id\n"
            "        AND sl.start_date >= date_trunc('month', CURRENT_DATE)\n"
            "        AND sl.start_date < date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'\n"
            "    )\n"
            "  ORDER BY random() LIMIT 1;\n"
            "RabbitMQ healthy."
        ),
        (
            "SETUP: Record employee id, login, target month.\n"
            "SETUP: Snapshot stat-report row: "
            "`SELECT month_norm, reported_effort, updated_at FROM "
            "ttt_backend.statistic_report WHERE employee_id = <id> AND "
            "month_start = date_trunc('month', CURRENT_DATE);`.\n"
            "SETUP (create sick-leave via API): "
            "`POST /api/vacation/v1/sick-leaves` body `{\"login\": \"<login>\", "
            "\"startDate\": \"<today>\", \"endDate\": \"<today+2>\", "
            "\"type\": \"PERSONAL\"}`. Record new sick-leave id.\n"
            "WAIT: 60 s for RabbitMQ fan-out "
            "(StatisticReportUpdateEventType.SICK_LEAVE_CHANGES).\n"
            "DB-CHECK: Re-query the stat-report row — `updated_at` must "
            "advance past snapshot. `month_norm` should decrease by the "
            "working-day-hours covered by the sick-leave (typically 3 "
            "days × 8 h = 24 h, but depends on office production "
            "calendar). `reported_effort` may be unchanged (sick-leave "
            "itself is not a task report) or reduced per business rule.\n"
            "VERIFY LOG: Graylog — sick-leave service logs the event emit; "
            "routing `employee-month-norm-context-calculated` fired.\n"
            "CLEANUP (via API): "
            "`DELETE /api/vacation/v1/sick-leaves/<new_id>`. Re-check "
            "stat-report row — `updated_at` advances again, `month_norm` "
            "restored to snapshot value."
        ),
        (
            "Sick-leave creation triggers the stat-report row refresh "
            "within 60 s; `updated_at` advances; `month_norm` decreases "
            "by the expected working hours. Regression for the #3337 "
            "pre-fix state where sick-leave events did not propagate."
        ),
        "High", "Regression",
        "#3337 (sick-leave month_norm propagation); #3423 row 22",
        "ttt-service/statistic-report/event-handler",
        "If allow_api_mutations is false in config.yaml, this TC is "
        "blocked — document as SKIPPED in the run. Requires either a "
        "dev env with mutations enabled or full manual execution."
    ),

    # ─── TC-STAT-082 — #3337: scoped event → unrelated employees preserved ───
    (
        "TC-STAT-082",
        "#3337 regression — scoped event for employee A does NOT delete statistic_report rows for employee B",
        (
            "Env: ttt-qa-1. Two distinct active employees A and B, each "
            "with existing `statistic_report` rows for current and "
            "previous month. Employees must be unrelated (different "
            "offices, different projects, different managers) to ensure "
            "no shared event fan-out by coincidence.\n"
            "Query:\n"
            "  SELECT e.id, e.login, e.office_id, e.manager_id\n"
            "  FROM ttt_backend.employee e\n"
            "  WHERE e.enabled = true\n"
            "    AND EXISTS (\n"
            "      SELECT 1 FROM ttt_backend.statistic_report sr\n"
            "      WHERE sr.employee_id = e.id\n"
            "        AND sr.month_start >= date_trunc('month', CURRENT_DATE - INTERVAL '1 month')\n"
            "    )\n"
            "  ORDER BY random() LIMIT 5;\n"
            "Pick A and B with different office_id and different "
            "manager_id."
        ),
        (
            "SETUP: Record employee A and B details. "
            "Snapshot employee B's stat-report rows: "
            "`SELECT month_start, month_norm, reported_effort, updated_at "
            "FROM ttt_backend.statistic_report WHERE employee_id = <B_id> "
            "ORDER BY month_start;` — save full resultset as "
            "`B_snapshot`.\n"
            "TRIGGER: Cause a scoped event for employee A ONLY. Options:\n"
            "  (a) Create a task_report for A via UI or "
            "`POST /api/ttt/v1/task-reports`.\n"
            "  (b) Update A's day-off (see TC-STAT-080 steps).\n"
            "  (c) Add a sick-leave for A (see TC-STAT-081 steps).\n"
            "WAIT: 90 s for full event propagation and fan-out.\n"
            "DB-CHECK (primary): Re-query employee B's stat-report rows "
            "with identical SQL — must equal `B_snapshot` exactly "
            "(same count of rows, same month_norm, same reported_effort, "
            "same updated_at timestamps — B's rows were NOT touched).\n"
            "DB-CHECK (sanity): Employee A's stat-report row for the "
            "affected month was refreshed (updated_at advanced) — "
            "confirms the event delivery path works; it's the SCOPING "
            "that's under test, not fan-out itself.\n"
            "CLEANUP: Revert the trigger mutation (delete the test "
            "task_report / sick-leave, restore day-off)."
        ),
        (
            "Employee B's statistic_report rows are IDENTICAL before and "
            "after the trigger (updated_at unchanged). Employee A's row "
            "for the affected month was refreshed. Regression for the "
            "#3337 bug where a bulk recalc incorrectly deleted unrelated "
            "rows — the fix scoped the DELETE to `employee_id IN "
            "(changed_set)`, and this TC guards against that scoping "
            "regressing to an unscoped delete."
        ),
        "High", "Regression",
        "#3337 (scoped delete fix); #3423 row 22",
        "ttt-service/statistic-report/event-handler",
        "The more employee B differs from A (different office, manager, "
        "department), the stronger this TC's coverage. Picking two "
        "employees from the same project weakens the test because their "
        "events may legitimately fan out together."
    ),

    # ─── TC-STAT-083 — #3346: manual delete + sync restores row ──────────────
    (
        "TC-STAT-083",
        "#3346 regression (bug #895498) — manual statistic_report row delete is restored on next optimized sync",
        (
            "Env: ttt-qa-1. Active employee with a statistic_report row "
            "for current month that can be temporarily deleted and "
            "restored. Choose an employee unlikely to view their "
            "statistics during the test window (avoid disrupting QA's "
            "own reports page).\n"
            "DB write access to `ttt_backend.statistic_report` required.\n"
            "Query candidate:\n"
            "  SELECT sr.id, sr.employee_id, e.login, sr.month_start\n"
            "  FROM ttt_backend.statistic_report sr\n"
            "  JOIN ttt_backend.employee e ON e.id = sr.employee_id\n"
            "  WHERE sr.month_start = date_trunc('month', CURRENT_DATE)\n"
            "    AND e.enabled = true\n"
            "  ORDER BY random() LIMIT 1;"
        ),
        (
            "SETUP: Record candidate row — full SELECT of all columns. "
            "Save as `deleted_snapshot`.\n"
            "TRIGGER (destructive): "
            "`DELETE FROM ttt_backend.statistic_report WHERE id = <row_id>;` "
            "via postgres MCP.\n"
            "DB-CHECK (precondition): Re-query for the row — must be "
            "absent (delete succeeded).\n"
            "TRIGGER (sync): "
            "`POST /api/ttt/v1/test/statistic-reports/sync` (optimized "
            "sync — previous + current month scope).\n"
            "WAIT: 120 s for RabbitMQ fan-out "
            "(StatisticReportUpdateEventType.INITIAL_SYNC on the affected "
            "employee).\n"
            "DB-CHECK (primary): Row reappears — "
            "`SELECT * FROM ttt_backend.statistic_report WHERE "
            "employee_id = <emp_id> AND month_start = "
            "'<current_month_start>'::date;` returns exactly one row. "
            "Columns match `deleted_snapshot` (id may differ — verify "
            "month_norm, reported_effort, employee_id).\n"
            "VERIFY LOG: Graylog — `\"Periodic statistic report sync "
            "finished.\"` marker fired between trigger and row "
            "reappearance.\n"
            "CLEANUP: No additional cleanup needed — the row is fully "
            "regenerated. If numerical values differ from snapshot, "
            "investigate (they should match within the event-consistency "
            "window)."
        ),
        (
            "Periodic sync restores the deleted row. Regression for bug "
            "#895498 (pre-#3346 fix) where the scheduler was not wired, "
            "so deleted rows stayed deleted until a restart. Also "
            "validates the idempotency contract — sync should be safe "
            "to rerun."
        ),
        "Critical", "Regression",
        "#3346 bug #895498; #3423 row 22",
        "ttt-service/statistic-report/sync",
        "The full-sync endpoint "
        "`/api/ttt/v1/test/statistic-reports/full-sync` would also "
        "restore the row but covers a wider year scope; use optimized "
        "sync here to exercise the minimum-coverage path. Requires DB "
        "write access — runs on non-shared envs preferred."
    ),

    # ─── TC-STAT-084 — Contract: full vs optimized sync scope ────────────────
    (
        "TC-STAT-084",
        "Full vs optimized sync contract — full refreshes previous + current year; optimized refreshes previous + current month only",
        (
            "Env: ttt-qa-1 with `statistic_report` populated for at least "
            "the past 12 months across a sample of employees. Access to "
            "the two test endpoints:\n"
            "  POST /api/ttt/v1/test/statistic-reports/full-sync     (full)\n"
            "  POST /api/ttt/v1/test/statistic-reports/sync          (optimized)"
        ),
        (
            "SETUP: Snapshot `max(updated_at)` per month for an employee "
            "with a full year of data:\n"
            "  SELECT month_start, max(updated_at) as last_updated\n"
            "  FROM ttt_backend.statistic_report\n"
            "  WHERE employee_id = <chosen_emp_id>\n"
            "  GROUP BY month_start ORDER BY month_start;\n"
            "Record as `snapshot_before_optimized`.\n"
            "TRIGGER (optimized): "
            "`POST /api/ttt/v1/test/statistic-reports/sync`.\n"
            "WAIT: 120 s.\n"
            "DB-CHECK (optimized scope): Re-query same SQL. Only rows for "
            "`date_trunc('month', CURRENT_DATE)` and "
            "`date_trunc('month', CURRENT_DATE - INTERVAL '1 month')` "
            "should have `updated_at` advanced; all OLDER month rows "
            "unchanged.\n"
            "SETUP (reset snapshot): Re-snapshot as "
            "`snapshot_before_full`.\n"
            "TRIGGER (full): "
            "`POST /api/ttt/v1/test/statistic-reports/full-sync`.\n"
            "WAIT: 300 s (full sync is slower — year scope).\n"
            "DB-CHECK (full scope): Re-query. ALL rows for "
            "`date_trunc('year', CURRENT_DATE)` and "
            "`date_trunc('year', CURRENT_DATE - INTERVAL '1 year')` "
            "should have `updated_at` advanced; rows older than 2 years "
            "unchanged.\n"
            "CLEANUP: None."
        ),
        (
            "Optimized sync updates previous + current month rows only. "
            "Full sync updates previous + current year rows (all 12–24 "
            "months in the year scope). Rows outside each scope remain "
            "untouched (`updated_at` preserved). Confirms the two "
            "endpoints have distinct, documented effects — critical for "
            "QA's data-setup playbooks (which endpoint to call to refresh "
            "what)."
        ),
        "Medium", "Functional",
        "#3345 note 894873; #3423 row 22; EXT-cron-jobs §Session 131.4",
        "ttt-service/statistic-report/sync",
        "Per #3345 note 894895 (omaksimova 2026-01-07), full-sync "
        "'NOT-A-BUG' observation: dismissed employees briefly appeared "
        "in current-year records after full sync because the scope "
        "includes them by definition. Expected behavior — do NOT assert "
        "dismissed-employee absence in full-sync scope. That assertion "
        "belongs in TC-STAT-079 (post-leave filter) only."
    ),
]


SUITES = [
    ("TS-Stat-CronStatReportSync", TS_CRON_STAT_REPORT_SYNC),
]


# ─── Sheet writer ─────────────────────────────────────────────────────────────

def _write_suite(wb, name, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = TAB_COLOR

    # R1 — back-link
    back = ws.cell(row=1, column=1,
                   value='=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")')
    back.font = FONT_LINK

    # R2 — blank (matches existing statistics suite layout)

    # R3 — header
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        letter = c.column_letter
        if col_idx - 1 < len(COL_WIDTHS):
            ws.column_dimensions[letter].width = COL_WIDTHS[col_idx - 1]
    ws.freeze_panes = ws.cell(row=4, column=1)

    # R4+ — TC rows
    for row_offset, tc in enumerate(rows, start=0):
        r = 4 + row_offset
        fill = FILL_ALT if (row_offset % 2) else FILL_WHITE
        for col_idx, val in enumerate(tc, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_CENTER if col_idx in (1, 6, 7) else ALIGN_WRAP
            c.fill = fill
            c.border = BORDER_THIN


def main():
    if not os.path.isfile(_STAT_XLSX):
        raise SystemExit(f"statistics.xlsx not found at {_STAT_XLSX}")

    wb = load_workbook(_STAT_XLSX)

    all_ids = []
    for suite_name, rows in SUITES:
        _write_suite(wb, suite_name, rows)
        all_ids.extend(r[0] for r in rows)

    wb.save(_STAT_XLSX)

    print(f"Extended: {_STAT_XLSX}")
    print(f"Suites added/updated: {len(SUITES)}")
    print(f"TCs written: {len(all_ids)} ({all_ids[0]} ... {all_ids[-1]})")
    for name, rows in SUITES:
        print(f"  {name}: {len(rows)} TCs")


if __name__ == "__main__":
    main()
