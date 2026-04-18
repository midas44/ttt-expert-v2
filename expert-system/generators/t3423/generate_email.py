#!/usr/bin/env python3
"""
Ticket #3423 — Email Cluster Cron Test Cases (Phase B, session 138).

Creates a NEW home-module workbook at ``test-docs/email/email.xlsx`` covering
rows 8 and 9 of the cron-testing collection:

    - TS-Email-CronDispatch (row 8 — 6 TCs: TC-EMAIL-001..006)
        Scheduler:  EmailSendScheduler.sendEmails()
        Cron:       ${email.scheduler.send.cron} = "*/20 * * * * *" (every 20 s)
        PageSize:   ${email.scheduler.send.page-size} = 300
        ShedLock:   EmailSendScheduler.sendEmails
        Markers:    "sendEmails: started" / "sendEmails: finished, sent {} emails"
        Endpoint:   POST /api/email/v1/test/emails/send
        Status:     NEW -> SENT | FAILED | INVALID

    - TS-Email-CronPrune (row 9 — 5 TCs: TC-EMAIL-007..011)
        Scheduler:  EmailPruneScheduler.pruneEmails()
        Cron:       ${email.scheduler.prune.cron} = "0 0 0 * * *" (daily 00:00)
        Retention:  ${email.scheduler.prune.older} = PT30D (30 days)
        ShedLock:   EmailPruneScheduler.pruneEmails
        Markers:    "pruneEmails: started" / "pruneEmails: finished, removed {} emails"
        Endpoint:   POST /api/email/v1/test/emails/delete
        Cutoff:     email.add_time < now - retention  (attachments + email rows)
        No feature-toggle gate. Runs unconditionally.

Context / rationale for a dedicated email workbook:
    The email service is the dispatcher layer — non-cron email flows (vacation
    digest, report forgotten, reject notification, budget) live in the
    business-module workbooks that trigger the queued emails. This workbook
    documents the dispatcher and retention schedulers themselves, which no
    business module owns. Test IDs are module-local (TC-EMAIL-NNN) per the
    home-module-TC convention established in session 135, and the COL-cron
    sheet in test-docs/collections/cron/cron.xlsx carries traceability via
    source_module = "email".

Idempotent: overwrites the workbook on every run. Creates ``test-docs/email/``
if it does not exist.

Tab colors:
    - Plan / Matrix / Risk: green 70AD47 (plan convention)
    - TS-Email-Cron* suites: orange F4B084 (cron suite convention, matching
      cross-service / statistics cron suites from sessions 137 / 138)

Run from repo root:
    python3 expert-system/generators/t3423/generate_email.py

Canonical references:
    - expert-system/vault/external/EXT-cron-jobs.md §Row 8, §Job 9 markers
    - expert-system/vault/modules/email-notification-deep-dive.md §4 Batch
        (EmailSendScheduler + EmailBatchServiceImpl.send + status transitions)
    - expert-system/repos/project/email/service/service-impl/src/main/java/
        com/noveogroup/ttt/email/service/batch/EmailSendScheduler.java
        com/noveogroup/ttt/email/service/prune/EmailPruneScheduler.java
        com/noveogroup/ttt/email/service/impl/EmailBatchServiceImpl.java
    - expert-system/repos/project/email/rest/src/main/java/com/noveogroup/ttt/
        email/rest/controller/v1/test/TestEmailController.java  (test endpoints)

Design issues folded into TCs (from email-notification-deep-dive.md §4):
    DI-EMAIL-DISPATCH-AUTH: MailAuthenticationException caught and logged;
        email status NOT updated (stuck as NEW). Scheduler fires every 20 s -
        infinite retry loop. TC-EMAIL-005 is the regression guard.
    DI-EMAIL-DISPATCH-RETRY:  No retry count tracking. FAILED stays FAILED
        forever; manual DB update to NEW is the only recovery path. Informs
        TC-EMAIL-004 notes (no retry assertion).
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- Paths --------------------------------------------------------------------

_HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "email")
)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "email.xlsx")


# --- Styling (matches reports/statistics/cross-service generators) ----------

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_CRON = "F4B084"  # orange — cron-focused suite convention


# --- Helper functions --------------------------------------------------------

def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, value, is_alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if is_alt else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    return cell


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_link(ws):
    cell = ws.cell(row=1, column=1, value="\u2190 Back to Plan Overview")
    cell.font = FONT_BACK_LINK
    cell.hyperlink = "#'Plan Overview'!A1"


# --- Test Case Data ----------------------------------------------------------

# NOTE: Preconditions call out the environment requirement first, then any
# data/seed requirement. Steps use SETUP:/CLEANUP:/VERIFY LOG:/DB-CHECK:
# prefixes consistent with vacation/reports/cross-service/statistics cron
# suites (sessions 135-138).


def get_dispatch_cases():
    """TS-Email-CronDispatch: EmailSendScheduler.sendEmails() cron + states."""
    return [
        {
            "id": "TC-EMAIL-001",
            "title": "EmailSendScheduler - cron fires every 20 s, ShedLock acquired/released, start/finish markers emit",
            "preconditions": (
                "Env: ttt-qa-1 (primary). Any env where the email service is running.\n"
                "Graylog access to the email/gotenberg backend stream (stream list "
                "exposes it - name differs from TTT-* streams; confirm via "
                "graylog-access `streams` subcommand before running).\n"
                "At least one email with status = NEW and no attachments in "
                "`email_service.email` so the batch has something to dispatch "
                "(otherwise finish marker reads `sent 0 emails` which is also a "
                "valid pass state - see TC-EMAIL-011 for no-op verification)."
            ),
            "steps": (
                "OBSERVE (no action): Tail Graylog stream for 25 s and watch for "
                "the marker pair. Scheduler fires every 20 s so the pair MUST "
                "appear at least once in the window.\n"
                "VERIFY LOG: `source:\"ttt-email-*\" AND "
                "message:\"sendEmails: started\"` - INFO event every 20 s.\n"
                "VERIFY LOG: `source:\"ttt-email-*\" AND "
                "message:\"sendEmails: finished, sent\"` - INFO event paired "
                "with each start, suffix `sent N emails` where N = 0 when queue "
                "empty, 1..300 when populated.\n"
                "VERIFY LOCK: query shedlock table for lock "
                "`EmailSendScheduler.sendEmails` during a run - `lock_until` is "
                "in the future. Immediately after finish marker, lock row either "
                "removed or `lock_until` now in the past.\n"
                "VERIFY TIMING: start/finish pair completes in < 20 s so the "
                "next scheduler tick never overlaps the previous run. If finish "
                "marker is missing 20 s after start, scheduler is stuck or lock "
                "contention occurred - escalate."
            ),
            "expected": (
                "Scheduler fires every 20 s. `sendEmails: started` and "
                "`sendEmails: finished, sent {N} emails` INFO markers appear in "
                "pairs in the email backend Graylog stream. ShedLock row "
                "`EmailSendScheduler.sendEmails` acquired during the run and "
                "released afterward. No overlap - each run fits inside its 20 s "
                "window."
            ),
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "#3423 row 8; EXT-cron-jobs §Row 8; EmailSendScheduler.java",
            "module": "email-service/batch/cron",
            "notes": (
                "20 s cadence is unusually tight compared to other schedulers "
                "(most fire every 5 min or daily). If multiple pods run the "
                "email service, ShedLock must serialize across pods - a second "
                "start marker within 20 s without a prior finish indicates a "
                "bug or lock-table split. Test endpoint "
                "`POST /api/email/v1/test/emails/send` calls the same batch "
                "method (bypassing the scheduler wrapper) - use when you need "
                "a deterministic trigger."
            ),
        },
        {
            "id": "TC-EMAIL-002",
            "title": "Dispatch - NEW email transitions to SENT after SMTP success (happy path)",
            "preconditions": (
                "Env: ttt-qa-1 with SMTP relay configured (verify via Graylog - "
                "past `sent N emails` markers where N > 0 prove delivery works).\n"
                "Test mailbox reachable at `vulyanov@office.local` via the "
                "`roundcube-access` skill.\n"
                "Ability to seed an email via either the email-api "
                "`POST /api/email/v1/emails` (authenticated) or direct DB insert "
                "into `email_service.email` with status = NEW. Skill: "
                "`swagger-qa1-email-api` (`send-using-pst`)."
            ),
            "steps": (
                "SETUP: Snapshot the latest sentTime in "
                "`email_service.email` so we can spot the new row after "
                "dispatch.\n"
                "SETUP: Seed one NEW email - body / subject deterministic (e.g., "
                "`[QA1][t3423-email-cron-test]` subject + freshly generated "
                "UUID in body). Prefer the REST endpoint over raw DB insert "
                "since it exercises the upstream path.\n"
                "OPTION A (wait for cron): Sleep 25 s - guarantees one "
                "scheduler tick.\n"
                "OPTION B (deterministic): "
                "`POST /api/email/v1/test/emails/send` - triggers the same "
                "batch path immediately.\n"
                "VERIFY LOG: `source:\"ttt-email-*\" AND "
                "message:\"sendEmails: finished, sent\"` - INFO event after "
                "trigger with N >= 1.\n"
                "DB-CHECK: `SELECT status, sent_time FROM email_service.email "
                "WHERE id = '<seeded-id>'` - status = 'SENT', sent_time is "
                "within the last 60 s.\n"
                "EMAIL-CHECK: `roundcube-access search --from timereporting "
                "--subject \"[QA1][t3423-email-cron-test]\" --since <today>` - "
                "one hit with matching UUID in body.\n"
                "CLEANUP: `POST /api/email/v1/test/emails/delete` (retention "
                "prune - will not delete this fresh row, see TC-EMAIL-008). "
                "Explicit cleanup: `DELETE FROM email_service.email "
                "WHERE id = '<seeded-id>'`."
            ),
            "expected": (
                "Seeded NEW email reaches status = SENT with sent_time set. "
                "SMTP delivery produces one Roundcube message with the "
                "expected subject and UUID. Graylog finish marker reports the "
                "send (N >= 1 after trigger)."
            ),
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "email-notification-deep-dive.md §4; EmailBatchServiceImpl.send; EmailWriter.write",
            "module": "email-service/batch/send",
            "notes": (
                "This is the load-bearing dispatch contract - every other "
                "business module (vacation digest, report forgotten, budget "
                "notification) ultimately depends on this NEW -> SENT "
                "transition. Breaking it kills all email notifications. Keep "
                "the seeded subject tagged so future cleanup scripts can "
                "identify orphaned test emails."
            ),
        },
        {
            "id": "TC-EMAIL-003",
            "title": "Dispatch - invalid recipient transitions NEW -> INVALID (SendFailedException path)",
            "preconditions": (
                "Env: ttt-qa-1 with SMTP relay configured.\n"
                "Ability to seed an email with a deliberately malformed "
                "recipient address (e.g., `not-a-real-address@@invalid..`) via "
                "DB insert (REST endpoint may reject malformed addresses "
                "upstream - fall back to direct DB insert in that case)."
            ),
            "steps": (
                "SETUP: `INSERT INTO email_service.email (id, subject, body, "
                "to_list, status, add_time) VALUES (gen_random_uuid(), "
                "'[QA1][t3423-invalid-recipient]', 'body', "
                "'not-a-real-address@@invalid..', 'NEW', now())`.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/send`.\n"
                "WAIT: 15 s for SMTP timeout and status update to land.\n"
                "DB-CHECK: `SELECT status FROM email_service.email "
                "WHERE id = '<seeded-id>'` - status = 'INVALID' (SendFailed "
                "with invalid addresses branch per "
                "email-notification-deep-dive.md §4 EmailWriter.write).\n"
                "VERIFY LOG: Graylog email stream - an exception trace or "
                "warning referencing the bad address. Finish marker `sent N "
                "emails` reflects successful sends in the same batch only.\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE id = "
                "'<seeded-id>'`."
            ),
            "expected": (
                "Email reaches status = INVALID. Scheduler keeps running; "
                "other NEW emails in the same batch dispatch normally. SMTP "
                "error is logged but does not crash the batch."
            ),
            "priority": "High", "type": "Hybrid",
            "req_ref": "email-notification-deep-dive.md §4 EmailWriter.write; SendFailedException",
            "module": "email-service/batch/error-paths",
            "notes": (
                "INVALID is terminal - email will not retry (unlike FAILED, "
                "which is also terminal but for transient causes). If a real "
                "employee gets a typo in their email, their notifications are "
                "permanently lost until the typo is fixed AND the email is "
                "manually re-queued (status reset to NEW via DB update)."
            ),
        },
        {
            "id": "TC-EMAIL-004",
            "title": "Dispatch - partial SMTP failure sets FAILED for rejected messages, SENT for succeeded",
            "preconditions": (
                "Env: ttt-qa-1. Best exercised on timemachine or a dev env "
                "where the SMTP relay can be coaxed into a transient failure "
                "for a specific recipient without affecting others (infra "
                "coordination required).\n"
                "Alternative: insert two NEW emails in the same batch, one "
                "with a recipient domain that triggers a transient bounce "
                "(e.g., a deliberately throttled test domain) and one normal."
            ),
            "steps": (
                "SETUP: Seed exactly 2 NEW emails in the batch:\n"
                "  (a) id = A, recipient = `vulyanov@office.local` (known good)\n"
                "  (b) id = B, recipient = `<domain that transiently rejects>` "
                "(coordinate with infra - simplest route is to point at a "
                "nonroutable IP via MX config for the test domain).\n"
                "TRIGGER: `POST /api/email/v1/test/emails/send`.\n"
                "WAIT: 20 s for SMTP responses.\n"
                "DB-CHECK: email A -> status = 'SENT', sent_time set.\n"
                "DB-CHECK: email B -> status = 'FAILED' (MailSendException "
                "partial-failure branch; NOT INVALID because the address "
                "parses correctly - only delivery fails).\n"
                "VERIFY LOG: Graylog email stream - exception details for "
                "email B; finish marker `sent 1 emails` (A only).\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE id IN (A, B)`."
            ),
            "expected": (
                "Good email reaches SENT. Transient-failure email reaches "
                "FAILED. Scheduler treats the two independently inside the "
                "single batch - no cross-contamination."
            ),
            "priority": "High", "type": "Hybrid",
            "req_ref": "email-notification-deep-dive.md §4 EmailWriter.write MailSendException",
            "module": "email-service/batch/error-paths",
            "notes": (
                "DI-EMAIL-DISPATCH-RETRY: FAILED is terminal - there is no "
                "retry-count tracking. The email will NOT be re-attempted "
                "even after the transient cause clears. Only manual "
                "recovery (DB update status -> NEW) re-queues. This TC does "
                "not assert retry; if a future patch adds retry, author a "
                "companion TC against the new contract."
            ),
        },
        {
            "id": "TC-EMAIL-005",
            "title": "DI-EMAIL-DISPATCH-AUTH regression - MailAuthenticationException caught; email stays NEW and re-dispatches every 20 s",
            "preconditions": (
                "Env: ttt-qa-1 or (preferred) a disposable dev env because "
                "this test deliberately breaks SMTP authentication. Requires "
                "infra coordination - do NOT run on shared envs without "
                "scheduling.\n"
                "Ability to change SMTP credentials in email-service config "
                "(or stop the SMTP relay briefly so authentication cannot "
                "complete)."
            ),
            "steps": (
                "SETUP: Snapshot current `email_service.email` counts by "
                "status.\n"
                "SETUP: Seed one NEW email with a recognizable subject.\n"
                "INDUCE: Change SMTP password to an invalid value "
                "(application.yaml `spring.mail.password`) and restart the "
                "email-service pod - or swap credentials at the SMTP relay "
                "side so authentication fails.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/send` OR wait 25 s "
                "for the natural tick.\n"
                "VERIFY LOG: Graylog email stream shows "
                "`MailAuthenticationException` trace at ERROR (DEEP-DIVE §4 "
                "documents the ERROR log; no retry-count, no status update).\n"
                "DB-CHECK: `SELECT status FROM email_service.email WHERE id = "
                "'<seeded-id>'` - status is STILL 'NEW' (the regression; "
                "status not updated on auth failure).\n"
                "WAIT: 25 s for the next scheduler tick.\n"
                "VERIFY: Graylog shows a SECOND MailAuthenticationException "
                "for the SAME email (infinite-loop hot pattern).\n"
                "CLEANUP: Restore SMTP credentials and restart the pod. "
                "Trigger once more - email finally transitions to SENT "
                "(confirming the recovery path).\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE id = "
                "'<seeded-id>'` if it survived the retry window."
            ),
            "expected": (
                "MailAuthenticationException is caught and logged at ERROR "
                "every tick. The seeded email stays as NEW and retries every "
                "20 s until auth is restored. After credentials are fixed, "
                "the next tick delivers successfully (status -> SENT). This "
                "TC is a regression guard: if a future patch updates status "
                "to FAILED on auth errors, the test must be re-triaged "
                "because the 20 s retry loop is replaced by a one-shot "
                "failure."
            ),
            "priority": "High", "type": "Regression",
            "req_ref": "email-notification-deep-dive.md §4 Design issue; DI-EMAIL-DISPATCH-AUTH",
            "module": "email-service/batch/error-paths",
            "notes": (
                "The current behavior is explicitly documented as a design "
                "issue - retried forever, no operator signal besides Graylog "
                "ERROR noise. In a prolonged outage this can produce "
                "thousands of exception traces per hour. Alerting rules "
                "should filter on `MailAuthenticationException` pattern at "
                "the email backend stream."
            ),
        },
        {
            "id": "TC-EMAIL-006",
            "title": "Dispatch - pageSize = 300 caps per-batch dispatch; overflow drains in subsequent ticks",
            "preconditions": (
                "Env: timemachine or a disposable dev env - test seeds 301 "
                "NEW emails which will all send via SMTP (keep recipients "
                "internal test inboxes to avoid external noise).\n"
                "Access to `vulyanov@office.local` mailbox via "
                "`roundcube-access` for sampling.\n"
                "Script capable of seeding 301 rows (psql + VALUES list or "
                "email-api in a loop)."
            ),
            "steps": (
                "SETUP: Snapshot latest `email_service.email` id ordering by "
                "add_time.\n"
                "SETUP: Seed 301 NEW emails with deterministic subjects "
                "`[QA1][t3423-pagesize-{N}]` where N = 1..301, all recipients "
                "`vulyanov@office.local`, bodies carry the same tag.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/send` (first tick - "
                "deterministic).\n"
                "WAIT: 10 s for first batch to complete.\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email "
                "WHERE subject LIKE '[QA1][t3423-pagesize-%' AND status = "
                "'SENT'` - exactly 300 (page-size cap enforced).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email "
                "WHERE subject LIKE '[QA1][t3423-pagesize-%' AND status = "
                "'NEW'` - exactly 1 (the overflow).\n"
                "VERIFY LOG: First `sendEmails: finished, sent 300 emails` "
                "INFO marker after the trigger.\n"
                "WAIT: 25 s for the next scheduler tick.\n"
                "DB-CHECK: Now 301 SENT, 0 NEW for the tagged emails.\n"
                "VERIFY LOG: Second finish marker `sent 1 emails`.\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE subject "
                "LIKE '[QA1][t3423-pagesize-%'`. Consider also "
                "bulk-deleting from Roundcube via IMAP (skill `save` then "
                "manual mailbox cleanup) to keep the test inbox tidy."
            ),
            "expected": (
                "First batch sends exactly 300 emails (page-size cap). "
                "Remainder drains on the next 20 s tick. No emails lost, no "
                "emails duplicated."
            ),
            "priority": "Medium", "type": "Boundary",
            "req_ref": "email-notification-deep-dive.md §4 ${email.scheduler.send.page-size}",
            "module": "email-service/batch/config",
            "notes": (
                "Page-size ceiling matters under bursty load (e.g., budget "
                "notification fan-out to hundreds of watchers). With a 20 s "
                "cadence and 300 cap, the dispatcher can drain ~54k emails/h - "
                "plenty for TTT's current scale. If the queue grows faster "
                "than that (future scaling), either page-size or cadence "
                "must change. Document any config change in this TC."
            ),
        },
    ]


def get_prune_cases():
    """TS-Email-CronPrune: EmailPruneScheduler.pruneEmails() 30-day retention."""
    return [
        {
            "id": "TC-EMAIL-007",
            "title": "EmailPruneScheduler - cron fires daily at 00:00, ShedLock acquired/released, start/finish markers emit",
            "preconditions": (
                "Env: timemachine (primary for this TC - time-machine control "
                "lets us cross midnight on demand) OR ttt-qa-1 if tolerant of "
                "waiting for natural 00:00 NSK firing.\n"
                "Graylog access to the email/gotenberg backend stream.\n"
                "At least one row in `email_service.email` older than 30 days "
                "so the prune has something to delete (otherwise finish marker "
                "reads `removed 0 emails` - also valid, see TC-EMAIL-011)."
            ),
            "steps": (
                "OPTION A (timemachine): Advance clock to 23:59:55 NSK and "
                "wait 10 s - scheduler fires at 00:00.\n"
                "OPTION B (qa-1 natural): Verify past midnight runs in Graylog "
                "history (`graylog-access search --stream <email-stream> "
                "--query 'message:\"pruneEmails: started\"' --range 24h`).\n"
                "OPTION C (deterministic): `POST /api/email/v1/test/emails/delete` "
                "- triggers the same batch method (bypasses scheduler "
                "wrapper; see note below).\n"
                "VERIFY LOG (start): `source:\"ttt-email-*\" AND "
                "message:\"pruneEmails: started\"` - INFO event at 00:00 NSK "
                "(Option A/B) OR within 5 s of the test-endpoint call "
                "(Option C).\n"
                "VERIFY LOG (finish): `source:\"ttt-email-*\" AND "
                "message:\"pruneEmails: finished, removed\"` - INFO event "
                "paired with start, suffix `removed N emails`.\n"
                "VERIFY LOCK: query shedlock for "
                "`EmailPruneScheduler.pruneEmails` during the run - "
                "`lock_until` in the future. After finish marker, lock row "
                "removed or expired.\n"
                "VERIFY TIMING: Start/finish pair completes within seconds "
                "(deletion of 30+ day rows is fast - no network, just DB)."
            ),
            "expected": (
                "Scheduler fires at 00:00 NSK daily. `pruneEmails: started` "
                "and `pruneEmails: finished, removed {N} emails` INFO markers "
                "appear in pairs in the email backend Graylog stream. "
                "ShedLock row acquired and released."
            ),
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "#3423 row 9; EXT-cron-jobs §Job 9; EmailPruneScheduler.java",
            "module": "email-service/prune/cron",
            "notes": (
                "SCHEDULER-WRAPPER DELTA: The test endpoint "
                "`POST /api/email/v1/test/emails/delete` calls the service "
                "method directly (`emailBatchService.pruneEmails()`), "
                "bypassing the scheduler bean. This means the endpoint WILL "
                "emit the pruneEmails start/finish markers because those are "
                "logged at the service level - confirm via code "
                "(EmailPruneScheduler.pruneEmails logs both, "
                "EmailBatchServiceImpl.pruneEmails does not log; so the "
                "scheduler wrapper IS the source of the markers). "
                "Consequence: Option C (test endpoint) does NOT produce the "
                "markers. Use Options A/B for marker verification; use "
                "Option C only for deletion verification."
            ),
        },
        {
            "id": "TC-EMAIL-008",
            "title": "Retention - emails older than 30 days are deleted, emails newer than 30 days are preserved",
            "preconditions": (
                "Env: ttt-qa-1. Direct DB write access to `email_service.email` "
                "for seeding backdated rows (the REST endpoint always sets "
                "add_time = now(); cannot seed old emails via API).\n"
                "A known `vulyanov@office.local` inbox for sampling if needed "
                "(though delete is DB-level only - no IMAP side effect "
                "on Roundcube-stored mail)."
            ),
            "steps": (
                "SETUP: Seed three emails with precise add_time values:\n"
                "  (a) id = A, subject = '[QA1][t3423-retention-old]', "
                "add_time = now() - interval '31 days'.\n"
                "  (b) id = B, subject = '[QA1][t3423-retention-boundary]', "
                "add_time = now() - interval '29 days'.\n"
                "  (c) id = C, subject = '[QA1][t3423-retention-fresh]', "
                "add_time = now().\n"
                "Set status = 'SENT' for all three so prune operates on "
                "realistic final-state rows.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/delete` "
                "(deterministic - no need to wait for 00:00 NSK).\n"
                "WAIT: 5 s for DB commit.\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email WHERE id "
                "= '<A>'` - 0 (old row deleted, retention = 30 d enforced).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email WHERE id "
                "= '<B>'` - 1 (within retention window, preserved).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email WHERE id "
                "= '<C>'` - 1 (fresh row preserved).\n"
                "VERIFY LOG: (if triggered via scheduler) `removed 1 emails` "
                "in the finish marker.\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE subject "
                "LIKE '[QA1][t3423-retention-%'`."
            ),
            "expected": (
                "Rows with add_time < now() - 30 days are deleted. Rows "
                "within the 30-day window are preserved. Cutoff is computed "
                "at trigger time (`LocalDateTime.now().minus(oldDuration)` "
                "where `oldDuration` = PT30D)."
            ),
            "priority": "High", "type": "Functional",
            "req_ref": "EmailBatchServiceImpl.pruneEmails; email.scheduler.prune.older = PT30D",
            "module": "email-service/prune/retention",
            "notes": (
                "The retention value is configurable via "
                "`email.scheduler.prune.older` (Spring Duration). Current "
                "value PT30D. Any env override changes the cutoff - always "
                "confirm the effective value from env config or actuator "
                "before running boundary tests."
            ),
        },
        {
            "id": "TC-EMAIL-009",
            "title": "Retention boundary - email exactly at 30-day cutoff behavior (strict less-than semantics)",
            "preconditions": (
                "Env: ttt-qa-1. Direct DB write access for seeding with "
                "sub-second-precise add_time values.\n"
                "Understanding of the cutoff predicate: "
                "`email.add_time < timeThreshold` (LESS-THAN, not "
                "less-than-or-equal) - this TC verifies the strict "
                "less-than semantics."
            ),
            "steps": (
                "SETUP: Compute a reference timestamp `T = now()` at test "
                "start. Record `T` for assertion math.\n"
                "SETUP: Seed two emails:\n"
                "  (a) id = E1, add_time = T - interval '30 days' - "
                "interval '1 second' (older than cutoff by 1 s).\n"
                "  (b) id = E2, add_time = T - interval '30 days' + "
                "interval '1 second' (newer than cutoff by 1 s).\n"
                "Set status = 'SENT' for both.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/delete` within "
                "~1 second of T (otherwise drift may ambiguate the "
                "boundary; retry if >5 s elapsed).\n"
                "DB-CHECK: E1 is deleted (add_time < threshold; threshold "
                "approx T - 30 d).\n"
                "DB-CHECK: E2 is preserved (add_time > threshold).\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE id IN "
                "('<E1>', '<E2>')`.\n"
                "DOCUMENT: If both deleted or both preserved, the cutoff "
                "predicate is NOT what the deep-dive documents - re-read "
                "EmailBatchServiceImpl.pruneEmails source and update the "
                "note here. The current code uses "
                "`Tables.EMAIL.ADD_TIME.lessThan(time)` (strict less-than)."
            ),
            "expected": (
                "Strict less-than semantics: email older than exactly 30 d "
                "by >= 1 s is deleted; email younger than exactly 30 d by "
                ">= 1 s is preserved. No ambiguity at sub-second level."
            ),
            "priority": "Medium", "type": "Boundary",
            "req_ref": "EmailRepositoryImpl.deleteBefore; Tables.EMAIL.ADD_TIME.lessThan",
            "module": "email-service/prune/retention",
            "notes": (
                "Clock skew between the test client and the DB server "
                "matters for this TC. If the DB is on a different host / "
                "timezone, adjust the add_time injection to use DB `now()` "
                "as the reference (e.g., `now() - interval '30 days 1 "
                "second'`) rather than the client clock. Always write "
                "add_time in UTC or server-local to match TimeUtils.now()."
            ),
        },
        {
            "id": "TC-EMAIL-010",
            "title": "Retention - attachments are deleted together with their parent email (no orphan attachments)",
            "preconditions": (
                "Env: ttt-qa-1. Direct DB write to both "
                "`email_service.email` and `email_service.attachment`.\n"
                "Understanding of the attachment schema: attachments are "
                "linked to emails by email_id; retention cut on "
                "attachments uses the email's add_time threshold "
                "(`attachmentRepository.deleteAll(timeThreshold)` in "
                "EmailBatchServiceImpl.pruneEmails)."
            ),
            "steps": (
                "SETUP: Seed one email `E_old` with add_time = now() - "
                "interval '31 days', status = 'SENT', and two attachments "
                "`A1`, `A2` linked to `E_old` in "
                "`email_service.attachment` (with their own add_time "
                "values matching `E_old` add_time).\n"
                "SETUP: Seed one email `E_new` with add_time = now(), "
                "status = 'SENT', and one attachment `A3` linked to it.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/delete`.\n"
                "WAIT: 5 s.\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email WHERE "
                "id = '<E_old>'` - 0 (deleted).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.attachment "
                "WHERE id IN ('<A1>', '<A2>')` - 0 (deleted together).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email WHERE "
                "id = '<E_new>'` - 1 (preserved).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.attachment "
                "WHERE id = '<A3>'` - 1 (preserved).\n"
                "DB-CHECK (orphan scan): `SELECT COUNT(*) FROM "
                "email_service.attachment a LEFT JOIN email_service.email "
                "e ON a.email_id = e.id WHERE e.id IS NULL` - 0 "
                "(no orphans left by the prune).\n"
                "CLEANUP: `DELETE FROM email_service.email WHERE id = "
                "'<E_new>'` and `DELETE FROM email_service.attachment "
                "WHERE id = '<A3>'`."
            ),
            "expected": (
                "Attachments are deleted atomically with their parent "
                "emails when the email is pruned. No orphan attachments "
                "remain after the scheduler run. Attachments within the "
                "retention window are preserved alongside their parent "
                "emails."
            ),
            "priority": "Medium", "type": "Functional",
            "req_ref": "EmailBatchServiceImpl.pruneEmails; AttachmentRepository.deleteAll",
            "module": "email-service/prune/attachments",
            "notes": (
                "Attachment cleanup uses the same time threshold as email "
                "cleanup but operates on a separate repository call. If "
                "ordering matters (e.g., FK constraint forcing attachment "
                "delete before email), verify transaction semantics in a "
                "follow-up TC. Current code calls "
                "attachmentRepository.deleteAll(timeThreshold) BEFORE "
                "emailRepository.deleteBefore(timeThreshold) - "
                "intentional ordering for FK safety."
            ),
        },
        {
            "id": "TC-EMAIL-011",
            "title": "Retention no-op - all emails within retention window; finish marker reads `removed 0 emails`",
            "preconditions": (
                "Env: ttt-qa-1 (or any env). Precondition is that NO row in "
                "`email_service.email` has add_time < now() - 30 days at the "
                "moment of trigger - verify with "
                "`SELECT COUNT(*) FROM email_service.email WHERE add_time "
                "< now() - interval '30 days'` = 0 before proceeding. If "
                "non-zero, either wait for the next daily run to drain them "
                "or exclude this TC run."
            ),
            "steps": (
                "PRECHECK: `SELECT COUNT(*) FROM email_service.email WHERE "
                "add_time < now() - interval '30 days'` = 0.\n"
                "TRIGGER: `POST /api/email/v1/test/emails/delete` OR wait for "
                "the natural 00:00 NSK run and observe Graylog.\n"
                "VERIFY LOG: `source:\"ttt-email-*\" AND message:\"pruneEmails: "
                "finished, removed 0 emails\"` - INFO event with the literal "
                "`0` suffix (SLF4J placeholder populated from "
                "`deletedCount = emailBatchService.pruneEmails()`).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.email` is "
                "unchanged (snapshot before/after).\n"
                "DB-CHECK: `SELECT COUNT(*) FROM email_service.attachment` "
                "is unchanged."
            ),
            "expected": (
                "Scheduler runs successfully with a no-op result. Finish "
                "marker reports `removed 0 emails`. No rows deleted, no "
                "exceptions thrown, no orphans created."
            ),
            "priority": "Medium", "type": "Functional",
            "req_ref": "EmailPruneScheduler.pruneEmails; EmailBatchServiceImpl.pruneEmails",
            "module": "email-service/prune/no-op",
            "notes": (
                "Freshly provisioned envs (like disposable dev envs) always "
                "start in this state - no emails old enough to prune. Over "
                "time, envs accumulate rows and the no-op condition no "
                "longer holds unless the env is reset. This TC is the "
                "baseline smoke test for the prune path on clean envs."
            ),
        },
    ]


# --- Feature Matrix (cron-scope only) -----------------------------------------

FEATURES = [
    {
        "feature": "Email dispatch scheduler (job 8 / row 8)",
        "suites": ["TS-Email-CronDispatch"],
        "ui": 0, "hybrid": 6, "api": 0,
    },
    {
        "feature": "Email retention prune scheduler (job 9 / row 9)",
        "suites": ["TS-Email-CronPrune"],
        "ui": 0, "hybrid": 5, "api": 0,
    },
]


# --- Risk Assessment ---------------------------------------------------------

RISKS = [
    {
        "feature": "Dispatch cadence (20 s)",
        "risk": (
            "Unusually tight 20-second scheduler cadence. If a batch cycle "
            "exceeds 20 s (SMTP slowness, oversized attachments, DB "
            "contention), subsequent ticks are blocked by ShedLock; queue "
            "grows."
        ),
        "likelihood": "Low", "impact": "Medium", "severity": "Medium",
        "mitigation": (
            "Monitor finish-marker timing in Graylog. Alert if start/finish "
            "pair > 15 s (approaching the cadence window). Profile batch "
            "duration under realistic attachment sizes."
        ),
    },
    {
        "feature": "MailAuthenticationException loop (DI-EMAIL-DISPATCH-AUTH)",
        "risk": (
            "When SMTP credentials are invalid, caught exception does NOT "
            "update email status. Scheduler re-dispatches the same email "
            "every 20 s, producing ERROR log noise and never failing "
            "forward."
        ),
        "likelihood": "Medium", "impact": "High", "severity": "High",
        "mitigation": (
            "Alert on MailAuthenticationException in the email backend "
            "Graylog stream. TC-EMAIL-005 regression-guards the behavior; "
            "any change (status transition to FAILED / retry-count) "
            "requires TC update. Operationally, rotate SMTP creds "
            "carefully - this is the worst error class because it "
            "affects ALL queued emails until fixed."
        ),
    },
    {
        "feature": "FAILED terminal state (DI-EMAIL-DISPATCH-RETRY)",
        "risk": (
            "FAILED emails never retry automatically. Transient SMTP "
            "issues (relay hiccup, recipient temporary-defer) become "
            "permanent delivery failures. No retry-count tracking means "
            "no self-healing."
        ),
        "likelihood": "Low", "impact": "Medium", "severity": "Medium",
        "mitigation": (
            "TC-EMAIL-004 documents the partial-failure path. Operational "
            "recovery is a manual DB update (`UPDATE email SET status = "
            "'NEW' WHERE id = ...`). Consider adding a retry tool / "
            "scheduled retry budget in a follow-up ticket."
        ),
    },
    {
        "feature": "Retention window (30 d) hardcoded via config",
        "risk": (
            "`email.scheduler.prune.older` is set at startup and NOT "
            "dynamic. Env-specific overrides may diverge from production. "
            "Shorter windows may delete emails before investigation is "
            "complete (audit / bug reproduction)."
        ),
        "likelihood": "Low", "impact": "Low", "severity": "Low",
        "mitigation": (
            "Verify the effective retention value via actuator on each "
            "env before running boundary tests. Document env-specific "
            "overrides in the env-config vault notes."
        ),
    },
    {
        "feature": "Attachment orphaning if prune order breaks",
        "risk": (
            "Attachments are deleted first, then emails. If either step "
            "partially fails (e.g., attachment deleteAll throws before "
            "emailRepository.deleteBefore), attachments may be orphaned "
            "or emails may reference deleted attachments."
        ),
        "likelihood": "Low", "impact": "Medium", "severity": "Low",
        "mitigation": (
            "TC-EMAIL-010 verifies zero orphans after a normal run. "
            "Transaction boundaries in pruneEmails should be audited "
            "if orphans ever appear in production (follow-up ticket)."
        ),
    },
    {
        "feature": "No feature-toggle gate on prune",
        "risk": (
            "If an operational issue requires temporarily halting "
            "retention deletion (e.g., during incident investigation "
            "where emails are evidence), there is no runtime kill switch. "
            "Only option is to stop the email-service pod or edit "
            "`email.scheduler.prune.cron` and restart."
        ),
        "likelihood": "Low", "impact": "Medium", "severity": "Low",
        "mitigation": (
            "Document the no-gate behavior. Escalate as a feature "
            "request if incident forensics ever need it. Short-term "
            "workaround: set the cron expression to a far-future value "
            "and restart."
        ),
    },
]


# --- Sheet generators --------------------------------------------------------

def create_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="Email Service - Cron Schedulers (#3423 rows 8, 9)").font = FONT_TITLE
    ws.cell(
        row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Phase B session 138 | Ticket #3423 collection-shape",
    ).font = FONT_BODY
    ws.cell(
        row=3, column=1,
        value="Module: email-service (dispatch + prune) | Branch: release/2.1",
    ).font = FONT_BODY

    row = 5
    ws.cell(row=row, column=1, value="Scope & Objectives").font = FONT_SUBTITLE
    row += 1
    objectives = [
        "Regression + functional coverage of the two cron schedulers in email-service.",
        "TS-Email-CronDispatch: EmailSendScheduler fires every 20 s; NEW -> SENT | FAILED | INVALID.",
        "TS-Email-CronPrune: EmailPruneScheduler fires daily at 00:00 NSK; 30-day retention of email + attachments.",
        "Regression guard on DI-EMAIL-DISPATCH-AUTH (MailAuthenticationException infinite-loop design issue).",
        "Boundary + no-op coverage on retention cutoff and page-size cap.",
        "Non-cron email flows (vacation digest, report notifications, budget alerts) are OUT OF SCOPE - covered in the per-business-module workbooks that queue emails which this dispatcher sends.",
    ]
    for obj in objectives:
        ws.cell(row=row, column=1, value=f"\u2022 {obj}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Knowledge Sources").font = FONT_SUBTITLE
    row += 1
    sources = [
        "Vault: external/EXT-cron-jobs.md (§Row 8 dispatch; §Job 9 prune)",
        "Vault: modules/email-notification-deep-dive.md (§4 Batch Email Processing)",
        "Code: email/service/service-impl/.../batch/EmailSendScheduler.java",
        "Code: email/service/service-impl/.../prune/EmailPruneScheduler.java",
        "Code: email/service/service-impl/.../impl/EmailBatchServiceImpl.java",
        "Code: email/db/db-impl/.../EmailRepositoryImpl.java (deleteBefore)",
        "Code: email/rest/.../v1/test/TestEmailController.java (test endpoints)",
    ]
    for src in sources:
        ws.cell(row=row, column=1, value=f"\u2022 {src}").font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1

    suite_data = [
        ("TS-Email-CronDispatch", "EmailSendScheduler.sendEmails - dispatch cron (row 8)", 6),
        ("TS-Email-CronPrune", "EmailPruneScheduler.pruneEmails - retention prune (row 9)", 5),
    ]
    for suite_id, suite_name, count in suite_data:
        cell = ws.cell(row=row, column=1, value=f"{suite_id}: {suite_name} - {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{suite_id}'!A1"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Environment Requirements").font = FONT_SUBTITLE
    row += 1
    envs = [
        "Primary: ttt-qa-1 (dispatch + prune behavior; seed via email-api + DB).",
        "Secondary: timemachine (for midnight-crossing prune verification without waiting).",
        "Disposable dev env (recommended for TC-EMAIL-005 DI-EMAIL-DISPATCH-AUTH - requires breaking SMTP auth).",
        "Test endpoints: POST /api/email/v1/test/emails/send, POST /api/email/v1/test/emails/delete (TestEmailController, @Profile('!production')).",
        "Graylog: email backend stream (name differs from TTT-* streams; confirm via graylog-access `streams`).",
        "Roundcube: vulyanov@office.local for SENT verification.",
        "Direct DB: email_service.email, email_service.attachment - required for seeding backdated rows (REST endpoint always sets add_time = now()).",
    ]
    for env in envs:
        ws.cell(row=row, column=1, value=f"\u2022 {env}").font = FONT_BODY
        row += 1

    row += 1
    total = sum(s[2] for s in suite_data)
    ws.cell(row=row, column=1, value=f"Total Test Cases: {total}").font = FONT_SUBTITLE

    row += 2
    ws.cell(row=row, column=1, value="Traceability").font = FONT_SUBTITLE
    row += 1
    trace = [
        "Ticket: #3423 - Cron & Startup Jobs Testing Collection",
        "Collection: test-docs/collections/cron/cron.xlsx (COL-cron sheet rows for TC-EMAIL-001..011, source_module = email)",
        "Investigation: expert-system/vault/exploration/tickets/t3423-investigation.md",
        "Coverage: test-docs/collections/cron/coverage.md (rows 8 and 9)",
    ]
    for t in trace:
        ws.cell(row=row, column=1, value=f"\u2022 {t}").font = FONT_BODY
        row += 1

    ws.column_dimensions["A"].width = 110


def create_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Test Suites", "UI Tests", "Hybrid Tests", "API Tests", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, f in enumerate(FEATURES):
        row = i + 2
        is_alt = i % 2 == 1
        total = f["ui"] + f["hybrid"] + f["api"]
        style_body_cell(ws, row, 1, f["feature"], is_alt)
        cell = style_body_cell(ws, row, 2, ", ".join(f["suites"]), is_alt)
        if f["suites"]:
            cell.font = FONT_LINK
            cell.hyperlink = f"#'{f['suites'][0]}'!A1"
        style_body_cell(ws, row, 3, f["ui"], is_alt)
        style_body_cell(ws, row, 4, f["hybrid"], is_alt)
        style_body_cell(ws, row, 5, f["api"], is_alt)
        style_body_cell(ws, row, 6, total, is_alt)

    set_column_widths(ws, [48, 28, 12, 14, 12, 10])


def create_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    severity_fills = {
        "Critical": FILL_RISK_HIGH,
        "High": FILL_RISK_HIGH,
        "Medium": FILL_RISK_MED,
        "Low": FILL_RISK_LOW,
    }

    for i, r in enumerate(RISKS):
        row = i + 2
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, r["feature"], is_alt)
        style_body_cell(ws, row, 2, r["risk"], is_alt)
        style_body_cell(ws, row, 3, r["likelihood"], is_alt)
        style_body_cell(ws, row, 4, r["impact"], is_alt)
        sev_cell = style_body_cell(ws, row, 5, r["severity"], is_alt)
        sev_cell.fill = severity_fills.get(r["severity"], FILL_ROW_WHITE)
        style_body_cell(ws, row, 6, r["mitigation"], is_alt)

    set_column_widths(ws, [36, 60, 12, 12, 12, 55])


def create_suite_sheet(wb, suite_id, cases):
    ws = wb.create_sheet(suite_id)
    ws.sheet_properties.tabColor = TAB_COLOR_CRON

    add_back_link(ws)

    headers = [
        "Test ID", "Title", "Preconditions", "Steps", "Expected Result",
        "Priority", "Type", "Requirement Ref", "Module/Component", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, tc in enumerate(cases):
        row = i + 3
        is_alt = i % 2 == 1
        style_body_cell(ws, row, 1, tc["id"], is_alt)
        style_body_cell(ws, row, 2, tc["title"], is_alt)
        style_body_cell(ws, row, 3, tc["preconditions"], is_alt)
        style_body_cell(ws, row, 4, tc["steps"], is_alt)
        style_body_cell(ws, row, 5, tc["expected"], is_alt)
        prio_cell = style_body_cell(ws, row, 6, tc["priority"], is_alt)
        if tc["priority"] == "Critical":
            prio_cell.fill = FILL_RISK_HIGH
        style_body_cell(ws, row, 7, tc["type"], is_alt)
        style_body_cell(ws, row, 8, tc["req_ref"], is_alt)
        style_body_cell(ws, row, 9, tc["module"], is_alt)
        style_body_cell(ws, row, 10, tc.get("notes", ""), is_alt)

    set_column_widths(ws, [14, 40, 45, 70, 42, 10, 10, 45, 28, 42])
    ws.auto_filter.ref = f"A2:J{len(cases) + 2}"


# --- Main ---------------------------------------------------------------------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    create_plan_overview(wb)
    create_feature_matrix(wb)
    create_risk_assessment(wb)

    suite_generators = [
        ("TS-Email-CronDispatch", get_dispatch_cases),
        ("TS-Email-CronPrune", get_prune_cases),
    ]

    total = 0
    for suite_id, gen_fn in suite_generators:
        cases = gen_fn()
        create_suite_sheet(wb, suite_id, cases)
        total += len(cases)
        print(f"  {suite_id}: {len(cases)} cases")

    wb.save(OUTPUT_FILE)
    print(f"\nGenerated {OUTPUT_FILE}")
    print(f"Total: {total} test cases across {len(suite_generators)} suites")


if __name__ == "__main__":
    main()
