#!/usr/bin/env python3
"""
Ticket #3423 — Cron & Startup Jobs Testing Collection — Phase B Generator (SCAFFOLD).

Produces three collection deliverables:
  test-docs/collections/cron/test-plan.md   — human-readable plan
  test-docs/collections/cron/cron.xlsx      — Plan Overview + COL-cron reference sheet
  test-docs/collections/cron/coverage.md    — traceability matrix (cron row -> TC IDs -> home spec)

This is a SCAFFOLD: the 23-row scope table is embedded, but COL-cron test-case rows
are empty and coverage cells are TBD. Phase B sessions populate both as home-module
TCs land in reports.xlsx / vacation.xlsx / cross-service.xlsx / statistics.xlsx.

Design conventions (from exploration/tickets/t3423-investigation.md preamble):
- Output path is test-docs/collections/cron/ — NOT test-docs/t3423/. The ticket is
  scope-shaped as collection, not default ticket-scope.
- Test IDs inside COL-cron follow TC-CRON-### (per ticket body); source_module /
  source_suite columns point back to where the TC definition canonically lives.
- COL- sheet prefix ensures autotests/scripts/parse_xlsx.py skips this workbook.
- Phase C is gated off (autotest.enabled: false). This script does NOT touch
  fixtures or specs.

Idempotent: re-running overwrites the three deliverable files.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Paths ────────────────────────────────────────────────────────────────────

_HERE = os.path.dirname(os.path.abspath(__file__))
_OUT_DIR = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "test-docs", "collections", "cron")
)
_XLSX_PATH = os.path.join(_OUT_DIR, "cron.xlsx")
_PLAN_PATH = os.path.join(_OUT_DIR, "test-plan.md")
_COVERAGE_PATH = os.path.join(_OUT_DIR, "coverage.md")


# ─── Scope table (23 rows) ────────────────────────────────────────────────────
# Mirrors docs/tasks/cron/cron-testing-task.md and the t3423-investigation preamble,
# with Phase-A corrections already folded in (see Phase-A deltas below).

SCOPE_ROWS = [
    # (row#, service, job, schedule, trigger, channels, home_module, notes_for_phase_b)
    (1,  "TTT",      "Forgotten-report notification (weekly)",              "Mon, Fri 16:00",           "POST /api/ttt/v1/test/reports/notify-forgotten",                           "E, L",     "reports",       "Template FORGOTTEN_REPORT; previousWeekMonday..previousWeekSunday window."),
    (2,  "TTT",      "Forgotten-report delayed notification",               "Daily 16:30",              "POST /api/ttt/v1/test/reports/notify-forgotten-delayed",                   "E, L",     "reports",       "Same template as row 1; subject sample shared by design."),
    (3,  "TTT",      "Report-changed notification",                         "Daily 07:50",              "POST /api/ttt/v1/test/reports/notify-changed",                             "E, L",     "reports",       "DELTA: template is REPORT_SHEET_CHANGED (not TASK_REPORT_CHANGED). Seed: yesterday task_reports where reporter_id != executor_id."),
    (4,  "TTT",      "Report-reject notification",                          "Every 5 min (code)",       "POST /api/ttt/v1/test/reports/notify-rejected",                            "E, L",     "reports",       "DELTA: ZERO log markers; DB assertion via reject.executor_notified. Historical UID 565319. DEBOUNCE_INTERVAL_MINUTES=5."),
    (5,  "TTT",      "Budget-overrun notification",                         "Every 30 min",             "POST /api/ttt/v1/test/budgets/notify",                                     "E, L",     "reports",       "DELTA: THREE templates: BUDGET_NOTIFICATION_{EXCEEDED,NOT_REACHED,DATE_UPDATED}. SAFETY_INTERVAL_SECONDS=10."),
    (6,  "TTT",      "CS sync (partial / startup-full)",                    "Every 15 min; startup",    "POST /api/ttt/v1/test/employees/sync?fullSync={true,false}",               "CS, DB, L","cross-service", "DELTA: Marker collision with row 20. Feature toggle CS_SYNC-{env}. Full sync only at startup."),
    (7,  "TTT",      "Extended report-period cleanup",                      "Every 5 min",              "POST /api/ttt/v1/test/reports/cleanup-extended",                           "DB, L",    "reports",       "Cleanup of timed-out extended periods."),
    (8,  "Email",    "Email dispatch batch",                                "Every 20 sec",             "POST /api/email/v1/test/emails/send",                                      "E, L",     "email",         "Highest-throughput cron; every 20s dequeue loop."),
    (9,  "Email",    "Email retention prune (> 30 days)",                   "Daily 00:00",              "POST /api/email/v1/test/emails/delete",                                    "DB, L",    "email",         "Retention policy > 30 days."),
    (10, "Vacation", "CS sync (partial / startup-full)",                    "Every 15 min; startup",    "POST /api/vacation/v1/test/employees/sync?fullSync={true,false}",          "CS, DB, L","vacation",      "Same pattern as row 6. Feature toggle CS_SYNC-{env}."),
    (11, "Vacation", "Annual vacation accruals",                            "Jan 1, 00:00",             "POST /api/vacation/v1/test/annual-accruals",                               "DB, L",    "vacation",      "Only start marker; no finish/error marker. Silent failure."),
    (12, "Vacation", "Preliminary-vacation outdated removal",               "Hourly",                   "POST /api/vacation/v1/test/vacations/delete-expired-preliminary",          "DB, L",    "vacation",      "NOT_IMPLEMENTED — dead YAML config. Single no-op stub TC."),
    (13, "Vacation", "Preliminary-vacation close-outdated",                 "Hourly",                   "POST /api/vacation/v1/test/vacations/close-outdated",                      "DB, L",    "vacation",      "NOT_IMPLEMENTED — dead YAML config. Single no-op stub TC."),
    (14, "Vacation", "Vacation notifications (digest)",                     "Daily 08:00",              "POST /api/vacation/v1/test/digest",                                        "E, L",     "vacation",      "DELTA: path is /digest (scope-table /vacations/notify is wrong). Cyrillic env-prefix anomaly."),
    (15, "Vacation", "Production-calendar annual reminder",                 "Nov 1, 00:01",             "POST /api/vacation/v1/test/production-calendars/send-first-reminder",      "E, L",     "vacation",      "DELTA: scheduler-wrapper bypass — test endpoint skips marker; assert per-recipient mail-sent markers. Template NOTIFY_VACATION_CALENDAR_NEXT_YEAR_FIRST. UIDs 609812/609813."),
    (16, "Vacation", "Auto-pay expired approved vacations",                 "Daily 00:00",              "POST /api/vacation/v1/test/vacations/pay-expired-approved",                "DB, L",    "vacation",      "Lock name legacy: CloseOutdatedTask.run (class renamed AutomaticallyPayApprovedTask)."),
    (17, "Vacation", "APPROVED -> PAID after period close",                 "Every 10 min",             "(no dedicated endpoint — trigger via ptch-report-period)",                 "DB, L",    "vacation",      "NO @SchedulerLock, NO log markers (Bug #2). DB-only verification on vacation.status transition."),
    (18, "Vacation", "Employee-project periodic sync",                      "Daily 03:00",              "POST /api/vacation/v1/test/employee-projects",                             "DB, L",    "vacation",      "DELTA: path has /vacation prefix (scope-table missing it). Sync-window data loss WON'T FIX."),
    (19, "Vacation", "Employee-project initial sync (startup-only)",        "Application startup",      "CI restart (release/2.1 or stage pipeline)",                               "DB, L",    "vacation",      "java_migration row EMPLOYEE_PROJECT_INITIAL_SYNC gates re-run. Delete row to re-execute."),
    (20, "Calendar", "CS sync (partial / startup-full)",                    "Every 15 min; startup",    "POST /api/calendar/v2/test/salary-office/sync?fullSync={true|false}",      "CS, DB, L","cross-service", "DELTA: path is v2/test/salary-office. Marker collision with row 6."),
    (21, "Vacation", "Statistic-report full sync (startup-only)",           "Application startup",      "CI restart (release/2.1 or stage pipeline)",                               "DB, L",    "vacation",      "java_migration row STATISTIC_REPORT_INITIAL_SYNC gates re-run."),
    (22, "TTT",      "Statistic-report optimized sync",                     "Daily 04:00",              "POST /api/v1/test/statistic-reports",                                      "DB, L",    "statistics",    "DELTA: failure logged at INFO not ERROR — assert by message pattern, not level. Post-#3346: cron moved to 04:00 NSK."),
    (23, "TTT",      "PM Tool project sync (partial / startup-full)",       "Every 15 min; startup",    "POST /api/ttt/v1/test/project/sync",                                       "PM, DB, L","cross-service", "Feature toggle PM_TOOL_SYNC-{env}. 8 seed TCs in 3083-ticket-findings."),
]

NOT_IMPLEMENTED_ROWS = {12, 13}

PHASE_A_DELTAS = [
    ("Row 3",   "Template key", "`TASK_REPORT_CHANGED`",                                  "`REPORT_SHEET_CHANGED`"),
    ("Row 4",   "Log markers",  "Implied presence",                                       "**Zero** log markers; DB assertion via `reject.executor_notified`"),
    ("Row 5",   "Template key", "Generic `BUDGET_*`",                                     "`BUDGET_NOTIFICATION_{EXCEEDED,NOT_REACHED,DATE_UPDATED}`"),
    ("Row 14",  "Endpoint path","`POST /api/vacation/v1/test/vacations/notify`",           "`POST /api/vacation/v1/test/digest`"),
    ("Row 15",  "Cron property","`production-calendar-first-notification.cron`",          "`production-calendar-annual-first.cron`"),
    ("Row 15",  "Log marker",   "Scheduler marker always present",                         "Test endpoint bypasses scheduler — assert per-recipient mail-sent markers"),
    ("Row 18",  "Endpoint path","`POST /api/v1/test/employee-projects`",                   "`POST /api/vacation/v1/test/employee-projects`"),
    ("Row 20",  "Endpoint path","`POST /api/calendar/v1/salary-offices/sync`",             "`POST /api/calendar/v2/test/salary-office/sync?fullSync={true|false}`"),
    ("Row 22",  "Failure log level", "Implied ERROR",                                      "INFO (tests match by message pattern, not level)"),
    ("Global",  "Full CS sync schedule", "\"Daily 00:00 full CS sync\"",                  "Startup-only; `companyStaff.full-sync` YAML key is **dead config**"),
]


# ─── Styling ──────────────────────────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY   = Font(name="Arial", size=10)
FONT_LINK   = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE  = Font(name="Arial", bold=True, size=14)
FONT_SUB    = Font(name="Arial", bold=True, size=11)

FILL_HEADER    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ALT       = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_WHITE     = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_NOTIMPL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ALIGN_WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN  = "70AD47"
TAB_COLOR_COL   = "4472C4"


# ─── XLSX generators ──────────────────────────────────────────────────────────

def _apply_header(ws, row_idx, headers, widths=None):
    for col_idx, name in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=name)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        if widths and col_idx <= len(widths):
            ws.column_dimensions[c.column_letter].width = widths[col_idx - 1]
    ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)


def _write_plan_overview(wb):
    ws = wb.create_sheet("Plan Overview")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws["A1"] = "Cron & Startup Jobs Testing Collection — Plan Overview"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:F1")

    ws["A3"] = "Ticket"
    ws["B3"] = "#3423"
    ws["A4"] = "Epic"
    ws["B4"] = "#3402"
    ws["A5"] = "Collection name"
    ws["B5"] = "cron"
    ws["A6"] = "Scope rows"
    ws["B6"] = len(SCOPE_ROWS)
    ws["A7"] = "Generated"
    ws["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    ws["A8"] = "Canonical preamble"
    ws["B8"] = "expert-system/vault/exploration/tickets/t3423-investigation.md"

    for r in range(3, 9):
        ws.cell(row=r, column=1).font = FONT_SUB
        ws.cell(row=r, column=1).alignment = ALIGN_WRAP
        ws.cell(row=r, column=2).font = FONT_BODY
        ws.cell(row=r, column=2).alignment = ALIGN_WRAP

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    ws["A10"] = "Scope (23 cron & startup jobs)"
    ws["A10"].font = FONT_SUB
    ws.merge_cells("A10:F10")

    headers = ["#", "Service", "Job", "Schedule", "Trigger", "Channels"]
    _apply_header(ws, 11, headers, widths=[5, 10, 52, 28, 68, 12])

    for i, (num, svc, job, sched, trig, ch, _hm, _nts) in enumerate(SCOPE_ROWS, start=12):
        fill = FILL_NOTIMPL if num in NOT_IMPLEMENTED_ROWS else (FILL_ALT if (num % 2 == 0) else FILL_WHITE)
        for col_idx, val in enumerate([num, svc, job, sched, trig, ch], start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_WRAP if col_idx > 1 else ALIGN_CENTER
            c.fill = fill
            c.border = BORDER_THIN

    next_row = 11 + len(SCOPE_ROWS) + 2
    ws.cell(row=next_row, column=1, value="Navigate").font = FONT_SUB
    nav = ws.cell(row=next_row + 1, column=1,
                  value='=HYPERLINK("#\'COL-cron\'!A1", "COL-cron reference sheet →")')
    nav.font = FONT_LINK


def _write_col_cron(wb):
    ws = wb.create_sheet("COL-cron")
    ws.sheet_properties.tabColor = TAB_COLOR_COL

    ws["A1"] = '=HYPERLINK("#\'Plan Overview\'!A1", "← Back to Plan Overview")'
    ws["A1"].font = FONT_LINK
    ws["A2"] = "COL-cron — Curated Test Collection (scaffold; populated as home-module TCs land)"
    ws["A2"].font = FONT_TITLE
    ws.merge_cells("A2:F2")

    headers = ["test_id", "source_module", "source_suite", "title", "inclusion_reason", "priority_override"]
    _apply_header(ws, 4, headers, widths=[14, 20, 30, 60, 40, 16])

    # No TC rows yet — Phase B sessions populate. Leave a placeholder legend:
    ws.cell(row=5, column=1, value="(empty — populate from home-module workbooks)").font = FONT_BODY
    ws.cell(row=5, column=1).alignment = ALIGN_WRAP
    ws.merge_cells("A5:F5")


def build_xlsx():
    wb = Workbook()
    wb.remove(wb.active)
    _write_plan_overview(wb)
    _write_col_cron(wb)
    wb.save(_XLSX_PATH)


# ─── test-plan.md generator ───────────────────────────────────────────────────

def build_test_plan_md():
    lines = []
    a = lines.append

    a("# Test Plan — Cron & Startup Jobs Testing Collection")
    a("")
    a("**Ticket:** #3423  ")
    a("**Epic:** #3402  ")
    a("**Collection:** `cron` (XLSX at `test-docs/collections/cron/cron.xlsx`, sheet `COL-cron`)  ")
    a(f"**Generated:** {datetime.now().strftime('%Y-%m-%d')}  ")
    a("**Status:** SCAFFOLD — populated iteratively by Phase B sessions.")
    a("")
    a("## 1. Overview")
    a("")
    a("Curated test collection consolidating end-to-end regression tests for every "
      "scheduled job in TTT and its integrated services (Vacation, Calendar, Email). "
      "Scope covers 23 cron & startup jobs across four backend services, with "
      "observability via DB, Roundcube (email), Graylog (logs), and cross-system writes "
      "to CS and PM Tool.")
    a("")
    a("Canonical conventions for this ticket live in "
      "`expert-system/vault/exploration/tickets/t3423-investigation.md`. Knowledge base "
      "highlights: `expert-system/vault/external/EXT-cron-jobs.md`, "
      "`expert-system/vault/exploration/tickets/3262-ticket-findings.md`, "
      "`expert-system/vault/exploration/tickets/3083-ticket-findings.md`, "
      "`expert-system/vault/patterns/email-notification-triggers.md`.")
    a("")
    a("## 2. Environment matrix")
    a("")
    a("| Env              | Primary use                              | Test-clock | API token  | Roundcube subject prefix |")
    a("|------------------|------------------------------------------|:---:|:---:|--------------------------|")
    a("| qa-1             | Default target — most TCs run here       | ✅ | ✅ | `[QA1]` or `[QA1][TTT]` |")
    a("| ttt-timemachine  | Clock-sensitive jobs (11, 14-16, 18, 22) | ✅ | ✅ | `[TIMEMACHINE]` or `[TIMEMACHINE][TTT]` |")
    a("| stage            | Startup-only jobs (19, 21, 23-full)      | ✅ | ✅ | `[STAGE]` or `[STAGE][TTT]` |")
    a("")
    a("Note: startup-only jobs are triggered by restarting the service via GitLab CI "
      "(`release/2.1` pipeline for qa-1 & ttt-timemachine; `stage` pipeline for stage). "
      "Feature toggles (`EMPLOYEE_PROJECT_INITIAL_SYNC`, `STATISTIC_REPORT_INITIAL_SYNC` in "
      "`ttt_vacation.java_migration`; `PM_TOOL_SYNC-{env}` in Unleash) gate re-execution.")
    a("")
    a("## 3. Risk areas by cluster")
    a("")
    a("| Cluster | Rows | Key risks |")
    a("|---|---|---|")
    a("| Notifications (ttt) | 1, 2, 3, 4, 5 | Template key drift (row 3); zero-markers (row 4); three-template fan-out (row 5); DEBOUNCE/SAFETY intervals gate email emission |")
    a("| Cleanup (ttt)       | 7             | DB-only; easy to assert |")
    a("| Email (email)       | 8, 9          | High-throughput dispatch (20s loop); retention prune every 30 days boundary |")
    a("| CS sync             | 6, 10, 20     | Marker collision between 6 & 20; Unleash toggle gating; RabbitMQ fan-out with 1-3 settle loops; full sync only at startup |")
    a("| Vacation time jobs  | 11, 14-17     | Clock manipulation; accruals silent failure (11); scheduler-wrapper bypass (15); no markers (17) |")
    a("| Employee-project    | 18, 19        | Data loss in sync window WON'T FIX; startup idempotency gated by feature-toggle |")
    a("| Statistic-report    | 21, 22        | INFO-level failure logging (22); Caffeine + DB cache invalidation; mid-month business rule deferred to #3356 |")
    a("| PM Tool             | 23            | Feature contract (11 fields); append-only presales merge; immutable accounting_name; startup-full vs partial behavior |")
    a("| Dead config         | 12, 13        | NOT_IMPLEMENTED — single no-op TC each confirms endpoint returns success |")
    a("")
    a("## 4. Verification recipe (applied to every TC)")
    a("")
    a("1. **SETUP** — seed minimal state via TTT / Vacation / Calendar Swagger API.")
    a("2. **Clock** — if time-sensitive, `ptch-using-ptch-11` or `reset-using-pst`.")
    a("3. **Trigger** — test endpoint for cron jobs; CI restart for startup-only (19, 21, 23-full).")
    a("4. **Wait** — respect async: email 20s dequeue; RabbitMQ fan-out = 1–3 settle loops.")
    a("5. **Verify** — DB (postgres MCP) ∪ UI (Playwright) ∪ Email (Roundcube) ∪ Log (Graylog), whichever channels apply per row.")
    a("6. **CLEANUP** — delete seeded data; reset clock if advanced.")
    a("")
    a("**Policy:** Roundcube is **mandatory** for E-channel rows; Graylog is **mandatory** for rows with only server-side side-effects.")
    a("")
    a("## 5. Phase-A-discovered scope-table deltas (fold into TC preconditions)")
    a("")
    a("| Location | Field | Scope table says | Actual (release/2.1) |")
    a("|---|---|---|---|")
    for loc, field, before, after in PHASE_A_DELTAS:
        a(f"| {loc} | {field} | {before} | {after} |")
    a("")
    a("## 6. Entry criteria")
    a("")
    a("- VPN active for logs.noveogroup.com, dev.noveogroup.com/mail, gitlab.noveogroup.com, ttt-{env}.noveogroup.com")
    a("- Roundcube IMAP credentials at `config/roundcube/envs/*.yaml`")
    a("- Graylog API token at `config/graylog/envs/secret.yaml`")
    a("- Test-clock permissions on target env (`reset-using-pst`, `ptch-using-ptch-11`)")
    a("- CS preprod UI access (shared Admin SSO) for CS-sync TC assertions")
    a("- PM Tool preprod UI access for PM Tool sync TC assertions")
    a("- GitLab CI permission on `ttt-spring` to trigger `restart-<env>` job (for 19, 21, 23-startup-full)")
    a("")
    a("## 7. Exit criteria")
    a("")
    a("- 23 / 23 scope rows have ≥ 1 TC in `COL-cron` (see `coverage.md`)")
    a("- `coverage.md` has no TBD cells — every cron maps to at least one TC ID and home-module workbook path")
    a("- All 8 Phase-A deltas reflected in TC preconditions (not in the ticket-body scope table — optional)")
    a("- SQLite `test_case_tracking` populated with every TC ID")
    a("")
    a("## 8. Open questions")
    a("")
    a("- Email cluster home — extend `reports.xlsx` or create dedicated `email.xlsx`? (Decision expected in session 135 when first email cluster TC lands.)")
    a("- Do rows 12, 13 (NOT_IMPLEMENTED) need more than a single no-op stub TC each? (Current recommendation: no.)")
    a("- `graylog-access search` subcommand regression — resolved or do TCs fall back to `tail ... | grep` permanently? (Skill-maintenance item; workaround is canonical for now.)")
    a("")
    a("---")
    a("*This plan is a scaffold. As home-module TCs land in `reports.xlsx`, `vacation.xlsx`, "
      "`cross-service.xlsx`, `statistics.xlsx` (and possibly `email.xlsx`), this document "
      "accumulates references and the `COL-cron` sheet becomes the single entry point.*")
    a("")

    with open(_PLAN_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─── coverage.md generator ────────────────────────────────────────────────────

def build_coverage_md():
    lines = []
    a = lines.append

    a("# Traceability — Cron & Startup Jobs Testing Collection")
    a("")
    a(f"**Generated:** {datetime.now().strftime('%Y-%m-%d')}  ")
    a("**Status:** SCAFFOLD — all TC cells are TBD. Phase B sessions populate.")
    a("")
    a("Each row maps a cron / startup job to the home-module TCs that exercise it and "
      "to the final spec path once Phase C resumes. Until TCs land, cells read `TBD`; "
      "until Phase C runs, spec paths read `n/a (Phase C gated off)`.")
    a("")
    a("| # | Job | Home module workbook | TC IDs | Spec file(s) |")
    a("|---:|-----|----------------------|--------|--------------|")
    for num, svc, job, sched, trig, ch, home_mod, _nts in SCOPE_ROWS:
        tag = " **[NOT_IMPLEMENTED]**" if num in NOT_IMPLEMENTED_ROWS else ""
        a(f"| {num} | {job}{tag} | `test-docs/{home_mod}/{home_mod}.xlsx` | TBD | n/a (Phase C gated off) |")
    a("")
    a("## Status legend")
    a("")
    a("- `TBD` — TC not yet authored; Phase B session to deliver")
    a("- **[NOT_IMPLEMENTED]** — cron code is dead YAML config; single no-op stub TC sufficient")
    a("- `n/a (Phase C gated off)` — ticket Stage D — spec paths populated when `autotest.enabled: true`")
    a("")
    a("## Verification channels per row (shortcut)")
    a("")
    a("`E` = Roundcube email (mandatory for E-rows)  |  `L` = Graylog log (mandatory for server-only rows)  |  `CS` / `PM` / `DB` = cross-system write")
    a("")
    a("| # | Channels |")
    a("|---:|---|")
    for num, _svc, _job, _sched, _trig, ch, _hm, _nts in SCOPE_ROWS:
        a(f"| {num} | {ch} |")
    a("")

    with open(_COVERAGE_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(_OUT_DIR, exist_ok=True)
    build_xlsx()
    build_test_plan_md()
    build_coverage_md()

    print(f"Generated:")
    print(f"  {os.path.relpath(_XLSX_PATH)}")
    print(f"  {os.path.relpath(_PLAN_PATH)}")
    print(f"  {os.path.relpath(_COVERAGE_PATH)}")
    print(f"Scope rows: {len(SCOPE_ROWS)} ({len(NOT_IMPLEMENTED_ROWS)} NOT_IMPLEMENTED)")


if __name__ == "__main__":
    main()
