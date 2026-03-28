#!/usr/bin/env python3
"""
t2724 Close-by-Tag Test Documentation Generator — Phase B
Generates test-docs/t2724/t2724.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 3 TS- test suite tabs.

Based on vault knowledge: t2724-investigation.md (23K+ words), planner-close-by-tag-implementation.md
(23K+ words), 33 ticket comments analyzed, 8 QA bugs documented, 6 MRs reviewed,
API tested on qa-1 (16 tests), Confluence + Figma specs verified.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "t2724")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "t2724.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Helpers ─────────────────────────────────────────────────────────────────

def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 1 else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER


# ─── Test Case Data ────────���─────────────────────────────────────────────────

def get_crud_cases():
    """TS-T2724-CRUD: Close-by-tag CRUD operations via UI and API."""
    return [
        {
            "id": "TC-T2724-001", "title": "Create a close tag — happy path (UI)",
            "preconditions": "Project with tracker integration. User is PM or SPM of the project.\nQuery: SELECT p.id, p.name FROM ttt_backend.project p JOIN ttt_backend.employee e ON p.manager_id = e.id WHERE e.enabled = true AND p.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the project manager\n2. Navigate to Planner > Projects tab\n3. Select the project from dropdown\n4. Click the Project Settings icon (gear icon at far right of table header)\n5. In the Project Settings modal, click 'Tasks closing' tab\n6. In the 'Add a tag' input field, type '[closed]'\n7. Click the '+' button\n8. Verify 'Changes saved' notification appears\n9. Verify '[closed]' appears in the tags table\n10. Verify the new tag row has green left-border highlight\n11. Verify the table auto-scrolled to the new tag",
            "expected": "Tag '[closed]' created and visible in tags table. Green highlight on new row. 'Changes saved' notification displayed. Auto-scroll to new item.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#2724 §7.4.3-7.4.5, Confluence §7.4", "module": "planner/close-by-tag",
            "notes": "POST /v1/projects/{projectId}/close-tags returns HTTP 200 (not 201). Immediate save, no confirmation needed."
        },
        {
            "id": "TC-T2724-002", "title": "Create duplicate tag — idempotent behavior",
            "preconditions": "Project with existing tag '[closed]'.\nSETUP: Via API — POST /v1/projects/{projectId}/close-tags with {\"tag\":\"[closed]\"}",
            "steps": "SETUP: Via API — create tag '[closed]' for the project\n1. Login as PM\n2. Navigate to Planner > Projects > select project > Project Settings > Tasks closing\n3. Type '[closed]' in the input field (same as existing tag)\n4. Click '+' button\n5. Verify behavior: either silent no-op or 'Changes saved' notification\n6. Verify only ONE '[closed]' tag exists in the table (no duplicate)\nCLEANUP: Via API — DELETE /v1/projects/{projectId}/close-tags/{tagId}",
            "expected": "Duplicate create is idempotent — returns existing tag. No duplicate row in table. Backend catches DataIntegrityViolationException and returns existing entity.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724, planner-close-by-tag-implementation.md §create()", "module": "planner/close-by-tag",
            "notes": "Backend uses REQUIRES_NEW transaction; on unique constraint violation, catches and returns existing tag."
        },
        {
            "id": "TC-T2724-003", "title": "Create blank/whitespace tag — validation error",
            "preconditions": "User is PM of a project.",
            "steps": "1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Leave the tag input empty and click '+'\n4. Verify validation error or button disabled\n5. Type '   ' (spaces only) and click '+'\n6. Verify validation error: 'Tag must not be blank'",
            "expected": "Blank and whitespace-only tags are rejected. @NotBlank validation enforced. HTTP 400 with error message.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724, planner-close-by-tag-implementation.md §Validation", "module": "planner/close-by-tag",
            "notes": "Backend validates @NotBlank on tag field. Frontend may also have client-side validation."
        },
        {
            "id": "TC-T2724-004", "title": "Inline edit a tag — happy path",
            "preconditions": "Project with existing tag 'test-tag'.\nSETUP: Via API — create tag",
            "steps": "SETUP: Via API — POST /v1/projects/{projectId}/close-tags with {\"tag\":\"test-tag\"}\n1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Click on the tag text 'test-tag' in the table (cursor should be pointer/hand)\n4. Verify field becomes editable (inline edit mode)\n5. Change text to 'updated-tag'\n6. Press Enter to save\n7. Verify 'Changes saved' notification appears\n8. Verify tag text now shows 'updated-tag'\nCLEANUP: Via API — DELETE the tag",
            "expected": "Inline editing works: click to edit, Enter to save. PATCH /close-tags/{tagId} called. Tag updated in table.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724 §7.4.4, Figma design", "module": "planner/close-by-tag",
            "notes": "Inline editing uses PlannerTag.js component. Escape should cancel without saving."
        },
        {
            "id": "TC-T2724-005", "title": "Inline edit — Escape cancels without saving",
            "preconditions": "Project with existing tag.\nSETUP: Via API — create tag",
            "steps": "SETUP: Via API — create tag 'original-tag'\n1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Click on 'original-tag' to enter inline edit mode\n4. Change text to 'modified'\n5. Press Escape key\n6. Verify tag text reverts to 'original-tag'\n7. Verify no 'Changes saved' notification appeared\nCLEANUP: Via API — DELETE the tag",
            "expected": "Escape cancels inline edit. No PATCH request sent. Original text preserved.",
            "priority": "Medium", "type": "UI",
            "req_ref": "Figma design, PlannerTag.js", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-006", "title": "Edit tag to duplicate value — validation error",
            "preconditions": "Project with two tags: 'tag-a' and 'tag-b'.\nSETUP: Via API — create both tags",
            "steps": "SETUP: Via API — create 'tag-a' and 'tag-b'\n1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Click on 'tag-b' to edit inline\n4. Change text to 'tag-a' (duplicate of existing)\n5. Press Enter\n6. Verify error message: 'Planner close tag already exists for project'\n7. Verify 'tag-b' remains unchanged\nCLEANUP: Via API — DELETE both tags",
            "expected": "Duplicate update rejected. ValidationException thrown. Tag not modified.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §update()", "module": "planner/close-by-tag",
            "notes": "No-op optimization: if new tag equals existing tag, returns without saving."
        },
        {
            "id": "TC-T2724-007", "title": "Delete a tag — happy path",
            "preconditions": "Project with existing tag.\nSETUP: Via API — create tag",
            "steps": "SETUP: Via API — create tag 'to-delete'\n1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Find 'to-delete' in the tags table\n4. Click the delete (trash) icon on that row\n5. Verify 'Changes saved' notification appears\n6. Verify 'to-delete' is removed from the tags table",
            "expected": "Tag deleted from table. DELETE /close-tags/{tagId} called. Immediate removal, no confirmation dialog.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724 §7.4.5", "module": "planner/close-by-tag",
            "notes": "Deletion is immediate — no confirmation prompt (consistent with tag creation)."
        },
        {
            "id": "TC-T2724-008", "title": "List tags — empty state",
            "preconditions": "Project with no close tags configured.",
            "steps": "1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Verify the tags table shows 'No data' message\n4. Verify explanatory text is visible above the table: 'Project tickets containing added values in the Info column will be automatically removed from the list on days when there are no more reports for them'",
            "expected": "Empty state shows 'No data'. Informational text about auto-closing behavior is displayed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2724 §7.4.1, planner-project-settings-pages.md", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-009", "title": "Permission — employee can list tags but cannot create",
            "preconditions": "Plain employee (not PM/SPM/admin) on a project with tags.\nSETUP: Via API — create tags as PM",
            "steps": "SETUP: Via API (as PM) — create tag 'test-tag'\n1. Login as plain employee who is a member of the project\n2. Navigate to Planner > Projects > select the project\n3. Attempt to open Project Settings\n4. If accessible: verify tags are visible in Tasks closing tab (GET returns tags)\n5. Verify: add tag input and '+' button are disabled or hidden\n6. Via API — POST /close-tags as plain employee → verify HTTP 403\n7. Via API — PATCH /close-tags/{id} as plain employee → verify HTTP 403\n8. Via API — DELETE /close-tags/{id} as plain employee → verify HTTP 403\nCLEANUP: Via API (as PM) — DELETE the tag",
            "expected": "Employee can list tags (GET 200) but cannot create/update/delete (403). Permission model: binary all-or-nothing based on project role.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "planner-close-by-tag-implementation.md §Permission Model", "module": "planner/close-by-tag",
            "notes": "PlannerCloseTagPermissionService checks: isAdmin, managerId, seniorManagerId, ownerId."
        },
        {
            "id": "TC-T2724-010", "title": "Permission — project manager can CRUD tags",
            "preconditions": "User is PM of a project.\nQuery: SELECT p.id, e.login FROM ttt_backend.project p JOIN ttt_backend.employee e ON p.manager_id = e.id WHERE e.enabled = true AND p.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the project manager\n2. Navigate to Project Settings > Tasks closing\n3. Create a tag 'pm-test' → verify success (200)\n4. Edit 'pm-test' to 'pm-updated' → verify success (200)\n5. Delete 'pm-updated' → verify success (200)\n6. Verify all operations triggered 'Changes saved' notification",
            "expected": "PM has full CRUD permissions on close tags for their project.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Permission Model", "module": "planner/close-by-tag",
            "notes": "Permission check: current.getId().equals(project.getManagerId())"
        },
        {
            "id": "TC-T2724-011", "title": "Permission — senior manager can CRUD tags",
            "preconditions": "User is SPM of a project.\nQuery: SELECT p.id, e.login FROM ttt_backend.project p JOIN ttt_backend.employee e ON p.senior_manager_id = e.id WHERE e.enabled = true AND p.enabled = true AND p.senior_manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the senior project manager\n2. Navigate to Project Settings > Tasks closing\n3. Create a tag 'spm-test' → verify success\n4. Delete 'spm-test' → verify success",
            "expected": "SPM has full CRUD permissions. Permission check: current.getId().equals(project.getSeniorManagerId()).",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Permission Model", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-012", "title": "Cross-project tag access — rejected",
            "preconditions": "Two projects: A (with tag) and B. User is PM of B but not A.\nSETUP: Via API — create tag on project A",
            "steps": "SETUP: Via API (as PM of A) — create tag 'cross-test' on project A, note tagId\n1. Via API (as PM of B) — DELETE /v1/projects/{projectB_id}/close-tags/{tagId_from_A}\n2. Verify HTTP 400\n3. Verify error: 'Planner close tag does not belong to project {projectB_id}'\n4. Via API — PATCH /v1/projects/{projectB_id}/close-tags/{tagId_from_A} → verify HTTP 400\nCLEANUP: Via API (as PM of A) — DELETE the tag",
            "expected": "Cross-project access blocked. Ownership check: tag.project.id must equal pathVariable projectId.",
            "priority": "High", "type": "API",
            "req_ref": "planner-close-by-tag-implementation.md §update(), §delete()", "module": "planner/close-by-tag",
            "notes": "Tested on qa-1 S72: returns 400 with ValidationException."
        },
        {
            "id": "TC-T2724-013", "title": "Special characters in tag — Unicode, Cyrillic",
            "preconditions": "User is PM of a project.",
            "steps": "1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Create tag 'Done / Résolu' → verify success\n4. Create tag 'Закрыто' (Cyrillic) → verify success\n5. Create tag '<script>alert(1)</script>' → verify stored raw (XSS concern)\n6. Verify all tags appear correctly in the table\nCLEANUP: Via API — DELETE all created tags",
            "expected": "Unicode and Cyrillic accepted. Tags stored as-is (no HTML sanitization). XSS concern: special characters not escaped.",
            "priority": "Medium", "type": "UI",
            "req_ref": "t2724-investigation.md §Live API Testing", "module": "planner/close-by-tag",
            "notes": "Tested on qa-1 S72: Unicode accepted. No HTML sanitization — potential XSS if rendered unsafely."
        },
        {
            "id": "TC-T2724-014", "title": "Long tag near VARCHAR(255) limit",
            "preconditions": "User is PM.",
            "steps": "1. Via API — POST /close-tags with 255-character tag → verify success (200)\n2. Via API — POST /close-tags with 256-character tag → verify truncation or error\n3. Note: Confluence spec says 200 char limit (§7.4.2) but DB allows 255\nCLEANUP: Via API — DELETE created tags",
            "expected": "DB accepts 255 chars (VARCHAR(255)). Confluence says 200 char limit but no frontend validation enforces it. Verify actual boundary.",
            "priority": "Low", "type": "API",
            "req_ref": "t2724-investigation.md §Discrepancies, Confluence §7.4.2", "module": "planner/close-by-tag",
            "notes": "Known discrepancy: Confluence 200 chars vs DB 255 chars. No frontend validation found."
        },
        {
            "id": "TC-T2724-015", "title": "Multiple tags — create 5+ tags for a project",
            "preconditions": "User is PM.",
            "steps": "1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Create tags: '[closed]', '[done]', '[resolved]', 'finished', 'cancelled'\n4. Verify all 5 tags appear in the table\n5. Verify table is scrollable if content exceeds visible area\n6. Verify each tag can be edited and deleted independently\nCLEANUP: Via API — DELETE all 5 tags",
            "expected": "Multiple tags supported. No max tag count per project. Each tag independently manageable.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Design Issues", "module": "planner/close-by-tag",
            "notes": "No pagination on GET /close-tags — returns all tags. No max count enforced."
        },
    ]


def get_apply_cases():
    """TS-T2724-Apply: Close-by-tag apply logic, matching rules, edge cases."""
    return [
        {
            "id": "TC-T2724-016", "title": "Apply — assignment without reports gets closed",
            "preconditions": "Project with tag matching an assignment's Info column. Assignment has NO reports on selected date.\nSETUP: Via API — create tag, ensure assignment with matching ticket_info exists.\nQuery: SELECT ta.id, t.ticket_info FROM ttt_backend.task_assignment ta JOIN ttt_backend.task t ON ta.task_id = t.id WHERE ta.project_id = {projectId} AND t.ticket_info IS NOT NULL AND t.ticket_info != '' LIMIT 5",
            "steps": "SETUP: Via API — create tag matching an existing assignment's ticket_info\nSETUP: Ensure the assignment has no task_report for the selected date\n1. Login as PM\n2. Navigate to Planner > Projects > select project\n3. Verify the matching assignment is visible in the table\n4. Click Project Settings icon > Tasks closing tab\n5. Verify the matching tag is listed\n6. Click OK button\n7. Verify loading spinner appears\n8. Verify page reloads automatically\n9. After reload: verify the matching assignment is NO LONGER visible in the table\nDB-CHECK: SELECT closed FROM ttt_backend.task_assignment WHERE id = {assignmentId} → expect true\nCLEANUP: Via API — DELETE the tag",
            "expected": "Assignment with matching tag and no reports is closed (closed=true). Row disappears from planner table. WebSocket PATCH event published.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#2724, planner-close-by-tag-implementation.md §apply()", "module": "planner/close-by-tag",
            "notes": "POST /v1/projects/{projectId}/close-tags/apply with {date}. Requires the apply endpoint to be deployed."
        },
        {
            "id": "TC-T2724-017", "title": "Apply — assignment WITH reports stays visible",
            "preconditions": "Project with tag matching an assignment. Assignment HAS reports on selected date.\nSETUP: Via API — create tag, ensure assignment has task_report for the date.\nQuery: SELECT ta.id FROM ttt_backend.task_assignment ta JOIN ttt_backend.task t ON ta.task_id = t.id JOIN ttt_backend.task_report tr ON tr.task = t.id AND tr.executor = ta.employee_id WHERE ta.project_id = {projectId} AND t.ticket_info ILIKE '%{tag}%' AND tr.report_date = '{date}' LIMIT 1",
            "steps": "SETUP: Via API — create tag, verify assignment has reports on date\n1. Login as PM\n2. Navigate to Planner > Projects > select project\n3. Note the assignment row with matching tag AND reported hours\n4. Open Project Settings > Tasks closing > click OK\n5. Wait for page reload\n6. Verify the assignment is STILL visible in the table (because it has reports)\nDB-CHECK: SELECT closed FROM ttt_backend.task_assignment WHERE id = {id} → may be true, but hasReportOnDate prevents disappearance",
            "expected": "Assignment with reports stays visible even when tag matches. hasReportOnDate() returns true → close event NOT published. Hours data preserved.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#2724 design decision, planner-close-by-tag-implementation.md §Key Behavioral Rules", "module": "planner/close-by-tag",
            "notes": "Core safety rule: reported hours are never hidden by close-by-tag."
        },
        {
            "id": "TC-T2724-018", "title": "Apply — case-insensitive matching",
            "preconditions": "Assignment with ticket_info '[CLOSED]' (uppercase). Tag 'closed' (lowercase).\nSETUP: Via API — create tag 'closed'",
            "steps": "SETUP: Via API — create tag 'closed' for project\n1. Login as PM\n2. Navigate to Planner > Projects\n3. Verify an assignment with Info '[CLOSED]' is visible\n4. Open Project Settings > Tasks closing > click OK\n5. After reload: verify the '[CLOSED]' assignment is closed\nDB-CHECK: Verify StringUtils.containsIgnoreCase matches 'closed' in '[CLOSED]'\nCLEANUP: Via API — DELETE the tag",
            "expected": "Case-insensitive match: tag 'closed' matches '[CLOSED]', 'Closed', 'already-closed', etc. Uses StringUtils.containsIgnoreCase().",
            "priority": "High", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Tag Matching Logic", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-019", "title": "Apply — substring matching (tag matches partial text)",
            "preconditions": "Assignment with ticket_info 'prefix-closed-suffix'. Tag 'closed'.",
            "steps": "SETUP: Via API — create tag 'closed'\n1. Verify assignment with Info containing 'closed' as substring is visible\n2. Trigger apply via OK\n3. After reload: verify assignment is closed\nNote: This also means tag 'fix' would match 'prefix', 'fixed', 'fixture'",
            "expected": "Substring matching: 'closed' matches any text containing 'closed'. Uses containsIgnoreCase (not exact match).",
            "priority": "High", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Tag Matching Logic", "module": "planner/close-by-tag",
            "notes": "Known false positive risk: short tags can match unintended text."
        },
        {
            "id": "TC-T2724-020", "title": "Apply — false positive: tag 'Done' matches 'Donetskaya'",
            "preconditions": "Assignment with ticket_info containing 'Anna Donetskaya'. Tag 'Done'.\nNote: This is a KNOWN and ACCEPTED false positive.",
            "steps": "SETUP: Via API — create tag 'Done'\n1. Verify assignment with Info 'Anna Donetskaya' is visible\n2. Trigger apply\n3. Verify the assignment gets closed (false positive match)\n4. Document: stakeholders have accepted this behavior",
            "expected": "False positive: 'Done' matches 'Donetskaya'. This is acknowledged and accepted by stakeholders. Substring matching has inherent false positive risk.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2724 design decision 11, t2724-investigation.md §Design Decisions", "module": "planner/close-by-tag",
            "notes": "Stakeholders explicitly acknowledged this corner case."
        },
        {
            "id": "TC-T2724-021", "title": "Apply on specific date — only affects selected date",
            "preconditions": "Project with close tag. Assignments on multiple dates.\nSETUP: Via API — create tag and assignments on two different dates",
            "steps": "SETUP: Via API — create tag matching assignments. Assignments exist on both Day1 and Day2.\n1. Login as PM\n2. Navigate to Planner > Projects, set date to Day1\n3. Trigger apply (OK)\n4. After reload: verify Day1 assignments closed\n5. Navigate to Day2\n6. Verify Day2 matching assignments are STILL visible (not affected)\nDB-CHECK: Check closed status for both days' assignments",
            "expected": "Apply is date-scoped. POST /apply sends {date: currentDay}. Only the selected date's assignments are processed.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §apply() date parameter", "module": "planner/close-by-tag",
            "notes": "Frontend sends currentDay from Redux store. Backend uses section.date or section.startDate."
        },
        {
            "id": "TC-T2724-022", "title": "Apply with no tags — no API call, no reload",
            "preconditions": "Project with NO close tags configured.",
            "steps": "1. Login as PM\n2. Navigate to Planner > Projects > select project with no tags\n3. Open Project Settings\n4. Click OK\n5. Verify: page does NOT reload (no API call made)\n6. Verify: modal closes normally\nDEBUG: Check browser DevTools Network tab — no POST /close-tags/apply request",
            "expected": "Empty-tags guard: if projectTags.length === 0, saga returns immediately. No apply API call. No page reload.",
            "priority": "High", "type": "UI",
            "req_ref": "t2724-investigation.md §handleCloseTagsApply", "module": "planner/close-by-tag",
            "notes": "Guard added in !5341. Critical for projects without close-by-tag configuration."
        },
        {
            "id": "TC-T2724-023", "title": "Apply — page reloads after successful apply",
            "preconditions": "Project with tags.",
            "steps": "SETUP: Via API — create a tag\n1. Login as PM\n2. Navigate to Planner > Projects\n3. Open Project Settings > click OK\n4. Verify loading spinner appears (startLoadingEmployeeModal)\n5. Verify page reloads (window.location.reload)\n6. After reload: verify planner state is fresh (assignments updated)",
            "expected": "After apply: loading spinner shown, then window.location.reload() refreshes the page. Full state refresh.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2724 Bug 8, !5335", "module": "planner/close-by-tag",
            "notes": "window.location.reload() is a blunt approach. The finally block (stopLoading + closeModal) is dead code since page reloads."
        },
        {
            "id": "TC-T2724-024", "title": "Apply from Project Members tab — triggers if tags exist",
            "preconditions": "Project with close tags. User is on Project Members tab.",
            "steps": "SETUP: Via API — create a tag\n1. Login as PM\n2. Navigate to Project Settings modal\n3. Stay on 'Project members' tab (do NOT switch to Tasks closing)\n4. Click OK\n5. Verify: apply is STILL triggered (if tags exist)\n6. Verify: page reloads\nNote: Both tabs share the same OK handler after !5341",
            "expected": "Both tabs' OK button triggers apply. If project has tags, apply runs regardless of which tab is active. Mitigated by empty-tags guard.",
            "priority": "Medium", "type": "UI",
            "req_ref": "t2724-investigation.md §Design Issues, !5341", "module": "planner/close-by-tag",
            "notes": "Design side effect: user who only changed Project Members still triggers close-tags apply."
        },
        {
            "id": "TC-T2724-025", "title": "Apply — generated (not-yet-opened) assignments also closed",
            "preconditions": "Employee whose assignments are NOT yet 'opened for editing' (readOnly). Project has matching tag.\nSETUP: Via API — create tag",
            "steps": "SETUP: Via API — create tag matching a task's ticket_info\n1. Login as PM\n2. Navigate to Planner > Projects\n3. Verify employee row shows as read-only (not opened for editing)\n4. Trigger apply via OK\n5. After reload: verify the generated assignments for that employee are also processed\nDB-CHECK: SELECT id, closed FROM ttt_backend.task_assignment WHERE employee_id = {id} AND closed = true → verify new closed records created",
            "expected": "Close-by-tag handles generated (id=null) assignments by creating them as closed via createForCloseByTag(). Works even for employees not yet 'opened for editing'.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Path 2: Generated assignment", "module": "planner/close-by-tag",
            "notes": "Silent failure: if createForCloseByTag throws, error is swallowed with log.debug() only."
        },
        {
            "id": "TC-T2724-026", "title": "Apply after 'Open for editing' — newly generated assignments eligible",
            "preconditions": "Employee with readOnly=true. Project has matching tag.",
            "steps": "1. Login as PM\n2. Navigate to Planner > Projects\n3. Click 'Open for editing' for an employee → generates assignments\n4. Verify new assignment rows appear for the employee\n5. Open Project Settings > Tasks closing > click OK\n6. After reload: verify matching newly-generated assignments are closed\n7. Verify non-matching assignments remain visible",
            "expected": "Sequence: Open for editing → Apply. Newly generated assignments with matching Info are eligible for closing.",
            "priority": "Medium", "type": "UI",
            "req_ref": "t2724-investigation.md §Open for Editing ↔ Close-by-Tag Interaction", "module": "planner/close-by-tag",
            "notes": "Both features are date-scoped and use currentDay from Redux."
        },
        {
            "id": "TC-T2724-027", "title": "Apply — multiple tags, partial matches",
            "preconditions": "Project with 3 tags. Assignments: some match tag1, some match tag2, some match none.\nSETUP: Via API — create tags",
            "steps": "SETUP: Via API — create tags '[closed]', '[done]', '[resolved]'\n1. Login as PM\n2. Verify: 5 assignments visible — 2 matching '[closed]', 1 matching '[done]', 2 matching nothing\n3. Trigger apply\n4. After reload: verify 3 assignments closed (matching ones without reports)\n5. Verify 2 non-matching assignments still visible",
            "expected": "Each assignment checked against ALL tags. An assignment matching ANY tag is eligible for closing. Non-matching assignments unaffected.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §collectAssignmentsToClose()", "module": "planner/close-by-tag",
            "notes": "OR logic: assignment closed if it matches any one of the project's tags."
        },
        {
            "id": "TC-T2724-028", "title": "Apply — assignment with blank ticket_info is skipped",
            "preconditions": "Assignment where task.ticket_info is null or empty.",
            "steps": "SETUP: Ensure an assignment exists with blank/null ticket_info\n1. Trigger apply\n2. Verify the blank-info assignment is NOT closed\nDB-CHECK: SELECT t.ticket_info, ta.closed FROM ttt_backend.task_assignment ta JOIN ttt_backend.task t ON ta.task_id = t.id WHERE ta.project_id = {id} AND (t.ticket_info IS NULL OR t.ticket_info = '')",
            "expected": "Assignments with null/blank ticket_info are skipped. No false positives on empty data.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "planner-close-by-tag-implementation.md §collectAssignmentsToClose()", "module": "planner/close-by-tag",
            "notes": "Code: if ticketInfo is blank, skip."
        },
        {
            "id": "TC-T2724-029", "title": "Apply endpoint — API direct call",
            "preconditions": "Project with tags. API access.",
            "steps": "SETUP: Via API — create tag\n1. Via API — POST /v1/projects/{projectId}/close-tags/apply with body {\"date\": \"2026-03-28\"}\n2. Verify HTTP 200 (void response)\n3. DB-CHECK: Verify matching assignments now have closed=true\n4. Via API — POST /apply with empty body → verify behavior (null date guard)\n5. Via API — POST /apply for project with no tags → verify no-op",
            "expected": "Direct API call works. Returns 200 with void body. Guard: null projectId or date returns immediately.",
            "priority": "Medium", "type": "API",
            "req_ref": "planner-close-by-tag-implementation.md §apply() Method", "module": "planner/close-by-tag",
            "notes": "Apply endpoint added in !5335. Requires deployment verification."
        },
    ]


def get_regression_cases():
    """TS-T2724-Regression: Regression tests for QA bugs from ticket #2724."""
    return [
        {
            "id": "TC-T2724-030", "title": "Bug 1 regression — popup closes only via OK button",
            "preconditions": "Project with Project Settings accessible. User is PM.",
            "steps": "1. Login as PM\n2. Open Project Settings modal\n3. Click outside the modal (overlay area)\n4. Verify: modal does NOT close (clicking outside has no effect)\n5. Press Escape key\n6. Verify: modal may or may not close via Escape (test both)\n7. Click OK button\n8. Verify: modal closes properly",
            "expected": "Modal is NOT closable by clicking outside. Only OK button closes it. This is by design (confirmed by ishumchenko in comment 908000).",
            "priority": "High", "type": "UI",
            "req_ref": "#2724 Bug 1, FIXED by !5313", "module": "planner/close-by-tag",
            "notes": "Regression: this was a reported bug — now 'by design'. Verify it remains consistent."
        },
        {
            "id": "TC-T2724-031", "title": "Bug 3 regression — correct column header in Tasks Closing tab",
            "preconditions": "User is PM.",
            "steps": "1. Login as PM\n2. Open Project Settings modal > Tasks closing tab\n3. Verify column header reads 'Tags for closing tasks' (EN)\n4. Switch language to Russian\n5. Verify column header reads 'Теги для закрытия задач' (RU)\n6. Verify it does NOT show 'Role on the project' or 'Роль на проекте'",
            "expected": "Correct header: 'Tags for closing tasks' (not 'Role on the project'). Bug #2724/Bug 3 was about wrong header. Fixed by !5313.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2724 Bug 3, FIXED by !5313", "module": "planner/close-by-tag",
            "notes": "Both EN and RU affected in original bug."
        },
        {
            "id": "TC-T2724-032", "title": "Bug 4 regression — OK button present in Tasks Closing tab",
            "preconditions": "User is PM.",
            "steps": "1. Login as PM\n2. Open Project Settings > switch to Tasks closing tab\n3. Verify OK button is visible and clickable\n4. Click OK\n5. Verify modal closes (and triggers apply if tags exist)",
            "expected": "OK button present on both tabs. Fixed by !5313. Both Project Members and Tasks closing have the OK button.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2724 Bug 4, FIXED by !5313", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-033", "title": "Bug 6 — cannot reopen popup on heavy data project",
            "preconditions": "Heavy data project (many assignments, e.g., DirectEnergie-ODC-like). Tags previously configured.\nQuery: SELECT p.id, p.name, COUNT(ta.id) FROM ttt_backend.project p JOIN ttt_backend.task_assignment ta ON ta.project_id = p.id GROUP BY p.id, p.name ORDER BY COUNT(ta.id) DESC LIMIT 5",
            "steps": "1. Login as PM of a heavy-data project\n2. Open Project Settings modal → Tasks closing tab\n3. Add a tag → click OK (triggers apply + reload)\n4. After reload: attempt to reopen Project Settings\n5. Verify: does the page hang with spinner? Or does it open normally?\n6. Check browser DevTools for network requests during hang\n7. Document: is Bug 6 still reproducible?",
            "expected": "Bug 6 (OPEN): After adding close tags on heavy data project, reopening Project Settings may hang with spinner. No network requests visible. Likely frontend state/rendering issue with large dataset.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724 Bug 6 (OPEN)", "module": "planner/close-by-tag",
            "notes": "Not reproducible on small-data projects (e.g., Diablocom-AI). Specific to large projects."
        },
        {
            "id": "TC-T2724-034", "title": "Bug 8 regression — auto-refresh after closing",
            "preconditions": "Project with tags and matching assignments.",
            "steps": "SETUP: Via API — create tag with matching assignments\n1. Login as PM\n2. Navigate to Planner > Projects\n3. Note current assignment list\n4. Open Project Settings > click OK\n5. Verify loading spinner appears\n6. Verify page automatically reloads (no manual refresh needed)\n7. Verify assignment list is updated (closed assignments removed)",
            "expected": "Auto-refresh works via window.location.reload(). Fixed by !5335/!5339. Previously required manual page reload.",
            "priority": "High", "type": "UI",
            "req_ref": "#2724 Bug 8, FIXED by !5335", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-035", "title": "Task order not disrupted after close-by-tag apply",
            "preconditions": "Project with DnD-reordered assignments. Some assignments match close tag.",
            "steps": "SETUP: Via API — create tag. Manually reorder tasks via DnD first.\n1. Login as PM\n2. Open for editing on an employee\n3. Drag-reorder tasks: move task C above task A\n4. Verify custom order persists\n5. Open Project Settings > click OK (triggers apply)\n6. After reload: verify remaining (non-closed) tasks maintain the custom DnD order\n7. Verify no duplicate rows appear",
            "expected": "Close-by-tag apply does not disrupt DnD ordering of remaining tasks. Related to #3332 and #3314 bugs.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-dnd-bugs-analysis.md §Testing Implications", "module": "planner/close-by-tag",
            "notes": "Close-by-tag publishes TaskAssignmentPatchEvent/GenerateEvent which may trigger the #3314 re-sort bug."
        },
        {
            "id": "TC-T2724-036", "title": "Informational text on Tasks Closing tab",
            "preconditions": "User is PM.",
            "steps": "1. Login as PM\n2. Open Project Settings > Tasks closing tab\n3. Verify informational text is present below the tabs\n4. Verify text reads: 'Project tickets containing added values in the Info column will be automatically removed from the list on days when there are no more reports for them'\n5. Verify text is visible in both EN and RU",
            "expected": "Explanatory text present. Matches Confluence §7.4.1 specification. Visible in both languages.",
            "priority": "Low", "type": "UI",
            "req_ref": "#2724 §7.4.1, Confluence", "module": "planner/close-by-tag",
            "notes": ""
        },
        {
            "id": "TC-T2724-037", "title": "Confluence discrepancy — 200 char limit not enforced",
            "preconditions": "User is PM.",
            "steps": "1. Login as PM\n2. Navigate to Project Settings > Tasks closing\n3. Type a 201-character tag in the input field\n4. Click '+' to add\n5. Verify: does frontend enforce 200-char limit (per Confluence §7.4.2)?\n6. If no frontend limit: verify tag is accepted (DB allows 255)\n7. Document actual boundary behavior",
            "expected": "Discrepancy: Confluence says 200 chars, DB allows VARCHAR(255). Verify whether frontend has maxLength attribute on input.",
            "priority": "Low", "type": "UI",
            "req_ref": "t2724-investigation.md §Key Discrepancies", "module": "planner/close-by-tag",
            "notes": "Known req/code mismatch. No frontend validation found in code analysis."
        },
        {
            "id": "TC-T2724-038", "title": "Apply error handling — silent failure on backend error",
            "preconditions": "Project with tags. Simulated backend error condition.",
            "steps": "1. Login as PM\n2. Trigger apply in a condition that might cause backend error\n3. Observe UI behavior: does any error message appear?\n4. Check browser console for devLog(error) output\n5. Note: frontend swallows errors in catch block (devLog only)\n6. Note: backend swallows generated assignment creation errors with log.debug()",
            "expected": "If apply fails: spinner appears, then nothing happens (no reload, no error message). Error logged to console only. Known design issue: no user-facing error feedback.",
            "priority": "Medium", "type": "UI",
            "req_ref": "t2724-investigation.md §handleCloseTagsApply, §Design Issues", "module": "planner/close-by-tag",
            "notes": "Both frontend (catch→devLog) and backend (catch→log.debug) silently swallow errors."
        },
    ]


# ─── Plan, Feature Matrix, Risk Assessment ────────────────────────────────────

PLAN_OVERVIEW = {
    "title": "Close-by-Tag Feature (#2724) — Test Plan",
    "scope": "Comprehensive regression and feature testing of the close-by-tag feature (ticket #2724, Sprint 15 CRITICAL). Covers tag CRUD operations, apply logic with matching rules, permission model, and regression tests for 8 QA-reported bugs. Based on analysis of 6 MRs, 33 ticket comments, Confluence requirements, Figma designs, and live API testing.",
    "objectives": [
        "Verify complete CRUD lifecycle for close tags (create, read, inline edit, delete)",
        "Test apply logic: tag-to-assignment matching, date-scoping, report preservation",
        "Validate case-insensitive substring matching with known false positive scenarios",
        "Test permission model: PM, SPM, admin can CRUD; plain employee can only list",
        "Regression tests for all 8 QA-reported bugs (3 fixed, 1 open, 2 by-design, 2 addressed)",
        "Verify interaction with 'Open for editing' and DnD ordering",
        "Test edge cases: empty tags, cross-project access, special characters, long tags",
        "Validate frontend behavior: loading spinner, page reload, 'Changes saved' notification",
    ],
    "environments": [
        "Primary: qa-1 (ttt-qa-1.noveogroup.com) — requires build with apply endpoint (!5335)",
        "Secondary: timemachine (ttt-timemachine.noveogroup.com)",
        "Note: Apply endpoint must be deployed (merged 2026-03-25, verify build date)",
    ],
    "approach": "UI-first testing with API setup/cleanup. Main test flow: login as PM → navigate to Planner Projects → select project → open Project Settings → manage tags / trigger apply. API used for tag creation in SETUP steps and for permission testing. DB checks verify assignment closed status.",
    "dependencies": [
        "Apply endpoint (POST /v1/projects/{projectId}/close-tags/apply) deployed to target env",
        "Project with tracker integration and assignments with ticket_info values",
        "PM/SPM credentials for the project",
        "Assignments with and without reports on test dates",
    ],
}

FEATURE_MATRIX = [
    # (Feature, CRUD, Apply, Regression, Total)
    ("Create tag (happy path)", 2, 0, 0, 2),
    ("Create tag (validation)", 2, 0, 0, 2),
    ("Inline edit tag", 2, 0, 0, 2),
    ("Delete tag", 1, 0, 0, 1),
    ("List tags / empty state", 1, 0, 0, 1),
    ("Multiple tags", 1, 0, 0, 1),
    ("Special chars / long tags", 2, 0, 0, 2),
    ("Permission (PM/SPM)", 2, 0, 0, 2),
    ("Permission (employee)", 1, 0, 0, 1),
    ("Cross-project access", 1, 0, 0, 1),
    ("Apply — close matching", 0, 3, 0, 3),
    ("Apply — reports preserved", 0, 1, 0, 1),
    ("Apply — matching rules", 0, 4, 0, 4),
    ("Apply — date scoping", 0, 1, 0, 1),
    ("Apply — empty tags guard", 0, 1, 0, 1),
    ("Apply — generated assignments", 0, 2, 0, 2),
    ("Apply — both tabs trigger", 0, 1, 0, 1),
    ("Apply — API direct call", 0, 1, 0, 1),
    ("Bug 1: popup close behavior", 0, 0, 1, 1),
    ("Bug 3: column header fix", 0, 0, 1, 1),
    ("Bug 4: OK button present", 0, 0, 1, 1),
    ("Bug 6: heavy data popup hang", 0, 0, 1, 1),
    ("Bug 8: auto-refresh", 0, 0, 1, 1),
    ("DnD order after apply", 0, 0, 1, 1),
    ("UI text / discrepancies", 0, 0, 2, 2),
    ("Error handling", 0, 0, 1, 1),
]

RISK_ASSESSMENT = [
    ("Tag matching false positives", "Short tags like 'fix' or 'Done' can match unintended text via substring matching. 'Done' matches 'Anna Donetskaya'.", "High", "Medium", "High",
     "Document known false positives. Test with short and ambiguous tags. Stakeholders have accepted this risk."),
    ("Apply endpoint not deployed", "The apply endpoint (!5335, merged 2026-03-25) may not be deployed to the testing environment. Previous builds predated the merge.", "High", "Critical", "Critical",
     "Verify build date before testing. If not deployed, CRUD tests proceed but apply tests are blocked."),
    ("Silent error handling", "Both frontend (devLog) and backend (log.debug) swallow errors during apply. Users see no feedback when apply partially fails.", "Medium", "High", "High",
     "Test error scenarios. Verify no user-facing error messages. Check browser console for devLog output."),
    ("Heavy data project hang", "Bug 6 (OPEN): reopening Project Settings on heavy-data projects may hang with spinner.", "Medium", "High", "High",
     "Test on largest available project. Monitor DevTools for network/rendering issues. Check CPU/memory consumption."),
    ("Both tabs trigger apply", "OK button on Project Members tab also triggers close-tags apply if tags exist. Unintended side effect of unified handler.", "Medium", "Medium", "Medium",
     "Test OK from both tabs. Verify empty-tags guard prevents unnecessary apply. Document behavior for QA awareness."),
    ("DnD order disruption", "Close-by-tag apply publishes WebSocket events that may trigger the #3314 re-sort bug, disrupting DnD order.", "Medium", "Medium", "Medium",
     "Reorder tasks via DnD, then apply. Verify order preserved for remaining tasks."),
    ("Window.location.reload()", "Blunt page refresh loses any unsaved state. The finally block is dead code.", "Low", "Low", "Low",
     "Verify reload happens. Note that any unsaved edits in other cells are lost."),
    ("No max tag count", "Unlimited tags per project. Could impact performance on GET /close-tags and apply processing.", "Low", "Low", "Low",
     "Test with 5+ tags. Verify no performance degradation."),
]


# ─── Sheet Writers ───────────────────────────────────────────────────────────

def write_test_cases(ws, cases):
    """Write test case rows to a TS- sheet."""
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [16, 40, 45, 65, 45, 10, 8, 35, 22, 35]

    # Back-link row 1
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    # Headers row 2
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    for row_offset, tc in enumerate(cases):
        row_idx = row_offset + 3
        values = [
            tc["id"], tc["title"], tc["preconditions"], tc["steps"],
            tc["expected"], tc["priority"], tc["type"],
            tc["req_ref"], tc["module"], tc.get("notes", "")
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    """Write Plan Overview tab."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"\u2022 {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"\u2022 {env}").font = FONT_BODY
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"\u2022 {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} \u2014 {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws):
    """Write Feature Matrix tab."""
    headers = ["Feature", "CRUD", "Apply", "Regression", "Total"]
    col_widths = [40, 8, 8, 12, 8]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    suite_tabs = ["TS-T2724-CRUD", "TS-T2724-Apply", "TS-T2724-Regress"]

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    """Write Risk Assessment tab."""
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


# ─── Main ────────────────────────────────────────────────────────────────────

def generate():
    """Generate the t2724 close-by-tag test documentation workbook."""
    wb = Workbook()

    suites = [
        ("TS-T2724-CRUD", "TS-T2724-CRUD", get_crud_cases),
        ("TS-T2724-Apply", "TS-T2724-Apply", get_apply_cases),
        ("TS-T2724-Regression", "TS-T2724-Regress", get_regression_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    total_cases = sum(len(case_fn()) for _, _, case_fn in suites)
    print(f"Generating {total_cases} test cases across {len(suites)} suites")

    # Plan Overview
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(ws_plan, suites_info)

    # Feature Matrix
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_feature_matrix(ws_fm)

    # Risk Assessment
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ws_ra)

    # Test Suite tabs
    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_test_cases(ws, cases)
        print(f"  {tab_name}: {len(cases)} cases")

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {total_cases} test cases, {len(suites)} suites, 3 plan tabs")

    return all_cases


if __name__ == "__main__":
    generate()
