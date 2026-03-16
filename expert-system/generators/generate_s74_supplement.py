#!/usr/bin/env python3
"""
Session 74 Supplement: Add live-testing-verified test cases to planner workbook.

Based on live API testing of #2724 close-by-tag on timemachine environment.
Findings: PATCH endpoint broken (500), XSS via unsanitized tags, verified CRUD behavior.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALT_FILL_2 = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

COLUMNS = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
           "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]


def style_row(ws, row_num, is_alt=False):
    fill = ALT_FILL_2 if is_alt else ALT_FILL_1
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = DATA_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='top')


def add_cases_to_sheet(ws, cases, start_row):
    for i, case in enumerate(cases):
        row_num = start_row + i
        for col_idx, value in enumerate(case, 1):
            ws.cell(row=row_num, column=col_idx, value=value)
        style_row(ws, row_num, is_alt=(row_num % 2 == 0))
    return start_row + len(cases)


# ============================================================
# PLANNER WORKBOOK — S74 Live Testing Supplement
# ============================================================

PLANNER_S74_CASES = [
    # TC-PLN-101: PATCH endpoint broken (CRITICAL BUG)
    ("TC-PLN-101", "BUG: PATCH close-tag endpoint returns 500 (gateway routing)",
     "Tag exists for project. User is admin or PM. Tested on timemachine build 2.1.26-SNAPSHOT.",
     "1. POST /v1/projects/{projectId}/close-tags to create tag 'test'\n"
     "2. PATCH /v1/projects/{projectId}/close-tags/{tagId} with body: {tag: 'updated'}\n"
     "3. PUT /v1/projects/{projectId}/close-tags/{tagId} with same body\n"
     "4. Check Swagger spec at /v2/api-docs?group=api for PATCH endpoint",
     "Step 2: 500 Internal Server Error — HttpRequestMethodNotSupportedException: "
     "'Request method PATCH not supported'.\n"
     "Step 3: Same 500 error for PUT.\n"
     "Step 4: Swagger spec only lists GET, POST (collection) and DELETE (item). "
     "PATCH is MISSING from spec. Gateway does not route PATCH to this path.",
     "Critical", "Bug Verification", "#2724", "PlannerCloseTagController / Gateway",
     "CRITICAL BUG verified on timemachine S74. Controller has @PatchMapping but gateway "
     "doesn't forward PATCH. Tag editing impossible. Frontend inline edit (TC-PLN-098) will "
     "also fail. Workaround: delete + re-create tag."),

    # TC-PLN-102: XSS via unsanitized tag content
    ("TC-PLN-102", "Security: Close-tag accepts unsanitized HTML/script content",
     "User is admin or PM.",
     "1. POST /v1/projects/{projectId}/close-tags with body: "
     "{tag: \"<script>alert('xss')</script>\"}\n"
     "2. POST with body: {tag: \"won't fix / закрыто\"}\n"
     "3. GET /v1/projects/{projectId}/close-tags\n"
     "4. Open 'Tasks closing' tab in frontend and check if tags are rendered safely",
     "Steps 1-2: 200 OK — raw HTML/script content stored in DB without sanitization.\n"
     "Step 3: Tags returned with raw content including <script> tags.\n"
     "Step 4: React's default JSX escaping SHOULD prevent execution, but verify "
     "no innerHTML/dangerouslySetInnerHTML usage in PlannerTag.js. Cyrillic and "
     "special chars stored correctly.",
     "Medium", "Security", "#2724", "PlannerCloseTagServiceImpl",
     "No server-side HTML sanitization. Relies on frontend escaping. VARCHAR(255) "
     "accepts any characters. Verified on timemachine S74."),

    # TC-PLN-103: Non-existent tag DELETE returns proper 404
    ("TC-PLN-103", "Delete non-existent close-tag returns 404 with error code",
     "Valid project exists. No tag with given ID.",
     "1. DELETE /v1/projects/{projectId}/close-tags/999999",
     "404 Not Found. Error body: {errorCode: 'exception.plannerclosetag.not.found', "
     "message: 'id = 999999'}. NotFoundException thrown by repository lookup.",
     "Low", "Negative", "#2724", "PlannerCloseTagServiceImpl",
     "Verified on timemachine S74. Error code is descriptive."),

    # TC-PLN-104: Cross-project tag access returns 400 with ownership message
    ("TC-PLN-104", "Cross-project close-tag manipulation returns ownership error",
     "Tag ID=X belongs to project A. User has permissions on project B.",
     "1. DELETE /v1/projects/{projectB_id}/close-tags/{tagFromProjectA_id}\n"
     "2. Observe error response",
     "400 Bad Request. Error: {errorCode: 'exception.validation.fail', "
     "message: 'Planner close tag does not belong to project {projectB_id}'}. "
     "Ownership check in service layer prevents cross-project manipulation.",
     "Medium", "Security", "#2724", "PlannerCloseTagServiceImpl",
     "Verified on timemachine S74. Exact error message confirmed. "
     "Complements TC-PLN-054 with verified response details."),

    # TC-PLN-105: Frontend build deployment gap
    ("TC-PLN-105", "Frontend build deployment verification for close-by-tag UI",
     "Close-by-tag backend API deployed (V2.1.27 migration applied, API endpoints functional).",
     "1. Check Build # in footer of TTT application\n"
     "2. Verify build date is AFTER merge date of !5301 (frontend MR)\n"
     "3. Navigate to Planner, select a project where user is PM\n"
     "4. Look for 'Project settings' gear icon (formerly 'Project employees')\n"
     "5. Open modal and verify 2 tabs: 'Project members' and 'Tasks closing'",
     "If build includes !5301: Modal shows 2 tabs, 'Tasks closing' tab has "
     "tag management UI (add form + tag list).\n"
     "If build predates !5301: Old 'Project employees' modal with single tab. "
     "Close-by-tag UI not available despite backend API being functional.",
     "High", "Deployment", "#2724", "Frontend/Planner",
     "S74 finding: timemachine build 2.1.26-SNAPSHOT.290209 (Mar 11) predates "
     "!5301 merge. Backend API works but frontend UI not deployed. Verify after "
     "next deployment."),
]


def supplement_planner():
    """Add S74 live-testing-verified cases to planner workbook."""
    wb_path = '/home/v/Dev/ttt-expert-v1/expert-system/output/planner/planner.xlsx'
    wb = openpyxl.load_workbook(wb_path)

    # Add cases to TS-Planner-CloseTag sheet
    ws = wb['TS-Planner-CloseTag']
    start_row = ws.max_row + 1
    add_cases_to_sheet(ws, PLANNER_S74_CASES, start_row)
    new_total = ws.max_row - 2  # subtract header + back-link rows

    # Update Plan Overview suite count
    plan_ws = wb['Plan Overview']
    for row in range(1, plan_ws.max_row + 1):
        cell = plan_ws.cell(row=row, column=1)
        if cell.value and 'CloseTag' in str(cell.value):
            import re
            old_text = str(cell.value)
            new_text = re.sub(r'(\d+)\s+cases', f'{new_total} cases', old_text)
            cell.value = new_text
            break

    wb.save(wb_path)
    print(f"Planner: Added {len(PLANNER_S74_CASES)} cases to TS-Planner-CloseTag "
          f"(now {new_total} cases in suite)")
    print(f"  TC-PLN-101 through TC-PLN-{100 + len(PLANNER_S74_CASES)}")
    return len(PLANNER_S74_CASES)


if __name__ == "__main__":
    count = supplement_planner()
    print(f"\nTotal new cases: {count}")
    print("  Planner: +{} (live testing verification — PATCH bug, XSS, deployment)".format(count))
