#!/usr/bin/env python3
"""Session 78 — Add InnovationLab Banner + CI Build Number suites to cross-service workbook.

New suites:
  - TS-XSV-Banner (15 cases: TC-XSV-031–045) — InnovationLab banner integration
  - TS-XSV-BuildInfo (7 cases: TC-XSV-046–052) — CI build number in footer
Total new: 22 cases → cross-service total: 52 cases, 6 suites + Test Data tab = 8 tabs
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling constants ────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_BODY_BOLD = Font(name="Arial", bold=True, size=10)

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

TAB_BLUE = "4472C4"
TAB_GREEN = "548235"


# ── Helpers ──────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab with back-link, headers, and test case rows."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    hr = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, t in enumerate(test_cases):
        row = hr + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            t["id"], t["title"], t["preconditions"], t["steps"],
            t["expected"], t["priority"], t["type"],
            t["req_ref"], t["module"], t.get("notes", "")
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = THIN_BORDER
            c.fill = fill

    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{ws.max_row}"
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# SUITE 5: INNOVATIONLAB BANNER (#3392)
# ══════════════════════════════════════════════════════════════════

BANNER_TESTS = [
    tc("TC-XSV-031",
       "Banner visibility — appears for all authenticated users",
       "User is logged in to TTT (any role: employee, manager, admin, accountant).\n"
       "Banner has not been permanently dismissed by user.",
       "1. Log in as employee.\n"
       "2. Observe top-right corner under the header.\n"
       "3. Verify collapsed banner (cat illustration + slider icon) is visible.\n"
       "4. Log out. Log in as admin.\n"
       "5. Verify banner is also visible for admin.\n"
       "6. Repeat for accountant and manager roles.",
       "Banner appears for ALL authenticated users regardless of role.\n"
       "Banner is in collapsed state (cat + slider icon).\n"
       "NOTE: Code hardcodes role='Production' for all users — banner library thinks everyone is Production.",
       "High", "Functional",
       "#3392 Req 1", "Frontend / AppContainer",
       "The requirement says 'Production position type only' but TTT code hardcodes role='Production' for everyone."),

    tc("TC-XSV-032",
       "Banner initial state — starts collapsed on every page load",
       "User is logged in. Banner not permanently dismissed.",
       "1. Navigate to any page in TTT.\n"
       "2. Observe the banner in top-right corner.\n"
       "3. Verify it shows: small logo, cat illustration, slider icon.\n"
       "4. Verify it does NOT show expanded content (title, text, Participate button).\n"
       "5. Click slider to expand the banner.\n"
       "6. Refresh the page (F5).\n"
       "7. Verify banner returns to collapsed state (not expanded).",
       "Banner always starts in collapsed state.\n"
       "Expanded state is NOT persisted across page refresh.\n"
       "After refresh: cat + slider icon visible, no expanded content.",
       "High", "Functional",
       "#3392 Req 6 + TTT Override #2", "Frontend / AppContainer",
       "TTT calls banner.collapse() immediately after init. Base spec says 'expanded on first login' but TTT overrides to collapsed."),

    tc("TC-XSV-033",
       "Collapsed → Expanded — click slider opens full banner",
       "Banner is in collapsed state.",
       "1. Click the slider icon on the collapsed banner.\n"
       "2. Verify banner expands to show: logo, title, descriptive text, 'Participate' button, 'Close' (X) icon.\n"
       "3. Verify expanded banner is in top-right corner, fixed position.\n"
       "4. Verify all text is readable and properly formatted.",
       "Banner transitions from collapsed to expanded state.\n"
       "All elements visible: logo, title, text, Participate button, Close icon.\n"
       "Banner remains in top-right corner.",
       "High", "Functional",
       "#3392 Req 15", "Frontend / Banner States"),

    tc("TC-XSV-034",
       "Expanded → Choice — Close icon shows dismissal options",
       "Banner is in expanded state.",
       "1. Click the 'Close' (X) icon on the expanded banner.\n"
       "2. Verify banner transitions to choice state.\n"
       "3. Verify three options are visible: 'Collapse', 'Remind later', 'Never show again'.\n"
       "4. Verify each option is a clickable button with blue (#428bca) styling.",
       "Banner shows choice state with three dismissal options.\n"
       "All three buttons are visible and styled in TTT blue.\n"
       "'Never show again' has blue border (not filled background).",
       "High", "Functional",
       "#3392 Req 9-10", "Frontend / Banner States"),

    tc("TC-XSV-035",
       "Choice → Collapsed — 'Collapse' returns to mini banner",
       "Banner is in choice state (showing three options).",
       "1. Click the 'Collapse' button.\n"
       "2. Verify banner returns to collapsed state (cat + slider icon).\n"
       "3. Verify no expanded content visible.\n"
       "4. Verify banner is still in top-right corner.",
       "Banner transitions from choice to collapsed state.\n"
       "Cat illustration and slider icon visible.\n"
       "No expanded content shown.",
       "Medium", "Functional",
       "#3392 Req 11", "Frontend / Banner States"),

    tc("TC-XSV-036",
       "'Participate' — opens InnovationLab landing in new tab",
       "Banner is in expanded state. Language is set to English.",
       "1. Click the 'Participate' button.\n"
       "2. Verify a new browser tab opens.\n"
       "3. Verify the URL is https://innovationlab.noveogroup.com/en/\n"
       "4. Return to TTT tab.\n"
       "5. Verify banner is no longer visible on the page.\n"
       "6. Navigate to another page. Verify banner is still hidden.",
       "New tab opens with InnovationLab landing page.\n"
       "URL matches language setting (/en/ for English).\n"
       "Banner disappears from TTT after clicking Participate.\n"
       "Banner remains hidden across page navigation.",
       "High", "Functional",
       "#3392 Req 8, 22", "Frontend / Banner Actions"),

    tc("TC-XSV-037",
       "'Participate' URL is language-specific (RU → /ru/, EN → /en/)",
       "Banner is in expanded state.",
       "1. Set TTT language to Russian (via nav dropdown).\n"
       "2. Expand banner and click 'Participate'.\n"
       "3. Verify new tab URL: https://innovationlab.noveogroup.com/ru/\n"
       "4. Close tab. Wait for banner to reappear (or use test user without prior state).\n"
       "5. Set TTT language to English.\n"
       "6. Expand banner and click 'Participate'.\n"
       "7. Verify new tab URL: https://innovationlab.noveogroup.com/en/",
       "Russian language → /ru/ URL.\n"
       "English language → /en/ URL.\n"
       "Language for URL is determined by TTT's current language, not CS.",
       "Medium", "Functional",
       "#3392 Req 20-22", "Frontend / i18n"),

    tc("TC-XSV-038",
       "'Remind later' — banner hidden, reappears after 3 calendar months",
       "Banner is in choice state.",
       "1. Click 'Remind later' button.\n"
       "2. Verify banner disappears completely.\n"
       "3. Navigate to multiple pages. Verify banner remains hidden.\n"
       "4. Log out and log in again. Verify banner still hidden.\n"
       "5. [MANUAL/Long-term] After 3 calendar months, log in. Verify banner reappears in expanded state.",
       "Banner disappears immediately after clicking 'Remind later'.\n"
       "Remains hidden across navigation and re-login.\n"
       "After 3 calendar months: banner reappears in expanded state.\n"
       "NOTE: 3 calendar months, not 90 days — verify month boundary behavior.",
       "Medium", "Functional",
       "#3392 Req 12", "Frontend / Banner Persistence",
       "Long-term verification requires system clock manipulation or waiting. "
       "On timemachine env, use clock API to advance 3 months."),

    tc("TC-XSV-039",
       "'Never show again' — banner permanently hidden",
       "Banner is in choice state.",
       "1. Click 'Never show again' button.\n"
       "2. Verify banner disappears completely.\n"
       "3. Navigate to multiple pages. Verify banner remains hidden.\n"
       "4. Log out and log in again. Verify banner still hidden.\n"
       "5. Clear browser cache/cookies. Log in again.\n"
       "6. Check if banner reappears (depends on state storage mechanism).",
       "Banner disappears immediately after clicking 'Never show again'.\n"
       "Remains hidden across navigation and re-login.\n"
       "If state is stored per-userId in banner library (not just localStorage), "
       "clearing cookies should NOT bring it back.\n"
       "If state is localStorage-only, clearing storage will reset state.",
       "High", "Functional",
       "#3392 Req 13", "Frontend / Banner Persistence"),

    tc("TC-XSV-040",
       "State persistence — 'Never show' survives logout and re-login",
       "User has previously clicked 'Never show again'.",
       "1. Log out of TTT.\n"
       "2. Log in as the SAME user.\n"
       "3. Navigate to several pages.\n"
       "4. Verify banner does not appear on any page.\n"
       "5. Log in as a DIFFERENT user (who has not dismissed banner).\n"
       "6. Verify banner appears for the different user.\n"
       "7. Verify per-user isolation: one user's dismissal does not affect another.",
       "Same user: banner remains hidden after re-login.\n"
       "Different user: banner appears normally.\n"
       "State is per-userId (companyStaffId), not global.",
       "High", "Functional",
       "#3392 Req 16", "Frontend / Banner Persistence",
       "Banner uses userId param for persistence — verify isolation between users."),

    tc("TC-XSV-041",
       "Banner re-collapses on page refresh even if expanded",
       "Banner is in expanded state (user clicked slider).",
       "1. Click slider to expand the banner.\n"
       "2. Verify banner is expanded (title, text, Participate visible).\n"
       "3. Press F5 to refresh the page.\n"
       "4. Wait for page to load.\n"
       "5. Check banner state.\n"
       "6. Navigate to a different page. Check banner state again.",
       "After refresh: banner returns to collapsed state (cat + slider).\n"
       "After navigation: banner is in collapsed state.\n"
       "Expanded state is transient — never persisted.\n"
       "This is by TTT design: banner.collapse() called on every init.",
       "Medium", "Functional",
       "#3392 TTT Override #2", "Frontend / AppContainer"),

    tc("TC-XSV-042",
       "Language switch mid-session — banner reinitializes with new language",
       "User is logged in. Banner is visible (collapsed). Language is English.",
       "1. Expand the banner. Note the text language (English).\n"
       "2. Switch TTT language to Russian via nav dropdown.\n"
       "3. Observe the banner after language switch.\n"
       "4. If banner auto-reinitializes: verify text is now in Russian.\n"
       "5. If banner does not update: collapse and re-expand.\n"
       "6. Verify banner content language matches TTT language.",
       "Banner text updates to match the selected TTT language.\n"
       "The useEffect dependency on currentLanguage triggers reinit.\n"
       "Both expanded text and button labels should reflect the new language.",
       "Medium", "Functional",
       "#3392 Req 20-21", "Frontend / i18n",
       "useEffect in AppContainer depends on currentLanguage — should trigger banner reinit."),

    tc("TC-XSV-043",
       "Banner fixed position — stays visible during page scroll",
       "Banner is visible (collapsed or expanded). Page has scrollable content.",
       "1. Navigate to a page with long scrollable content (e.g., reports table).\n"
       "2. Verify banner is in top-right corner.\n"
       "3. Scroll down the page.\n"
       "4. Verify banner remains fixed in top-right corner of viewport.\n"
       "5. Scroll to bottom of page.\n"
       "6. Verify banner is still visible in the same position.\n"
       "7. Repeat with expanded banner state.",
       "Banner position is fixed (not absolute) — stays in viewport during scroll.\n"
       "Both collapsed and expanded states remain fixed.",
       "Medium", "UI",
       "#3392 Req 3-4", "Frontend / CSS"),

    tc("TC-XSV-044",
       "Banner non-blocking — page elements remain interactable",
       "Banner is visible (collapsed) in top-right corner.",
       "1. Open a page where content may be near the top-right corner.\n"
       "2. Verify all page buttons, links, and inputs are clickable.\n"
       "3. Expand the banner.\n"
       "4. Try to interact with page elements near/behind the banner.\n"
       "5. Verify the 'Add task' button (narrow screens) is not blocked.\n"
       "6. Open dropdown menus near the banner area.\n"
       "7. Verify dropdowns display above/below the banner without being blocked.",
       "Page content remains fully interactable.\n"
       "Banner does not capture clicks intended for page elements.\n"
       "Dropdown menus and buttons near the banner area work correctly.\n"
       "On narrow screens, 'Add task' button is accessible.",
       "High", "UI",
       "#3392 Req 17", "Frontend / CSS",
       "Concern raised in issue: narrow screens may have Add task button blocked by banner."),

    tc("TC-XSV-045",
       "Collapsed banner — entire area clickable (expanded hit area)",
       "Banner is in collapsed state.",
       "1. Hover over the collapsed banner. Note the cursor change area.\n"
       "2. Click on the cat illustration (not the slider icon).\n"
       "3. Verify banner expands (the ::before pseudo-element extends hit area).\n"
       "4. Collapse banner again.\n"
       "5. Click on the logo area of the collapsed banner.\n"
       "6. Verify banner expands.\n"
       "7. Verify the entire 68×372px area acts as clickable.",
       "Clicking anywhere on the collapsed banner (not just the slider icon) expands it.\n"
       "The CSS ::before pseudo-element on the slider button extends the hit area.\n"
       "Hit area covers the full collapsed banner (68×372px).",
       "Medium", "UI",
       "#3392 TTT Override, !5277", "Frontend / CSS",
       "style-overrides.css adds ::before on slider-button with 68x372px transparent area."),
]


# ══════════════════════════════════════════════════════════════════
# SUITE 6: CI BUILD NUMBER IN FOOTER (#3036)
# ══════════════════════════════════════════════════════════════════

BUILD_INFO_TESTS = [
    tc("TC-XSV-046",
       "Footer displays build number from ttt-backend actuator",
       "TTT application is deployed and running.\n"
       "User is logged in.",
       "1. Observe the app footer.\n"
       "2. Note the displayed 'Build #' value.\n"
       "3. Call GET /api/ttt/actuator/info.\n"
       "4. Extract `build.app.version` from the response.\n"
       "5. Compare footer value with actuator response value.",
       "Footer 'Build #' matches `build.app.version` from /api/ttt/actuator/info.\n"
       "Format: <maven-version>.<CI_PIPELINE_ID> (e.g., '2.1.26-SNAPSHOT.290485').",
       "High", "Functional",
       "#3036", "Frontend / Footer + Backend / Actuator"),

    tc("TC-XSV-047",
       "Build number matches CI pipeline ID",
       "A recent CI pipeline has run on release/2.1.\n"
       "The build has been deployed to the test environment.",
       "1. Check GitLab CI — note the latest pipeline ID on release/2.1 that was deployed.\n"
       "2. Call GET /api/ttt/actuator/info on the deployed environment.\n"
       "3. Extract `build.ci.build-number` from the response.\n"
       "4. Compare with the pipeline ID from GitLab.\n"
       "5. Verify `build.app.version` ends with the same pipeline ID.",
       "build.ci.build-number = CI pipeline ID from GitLab.\n"
       "build.app.version ends with .<pipeline_id>.\n"
       "The footer shows this exact pipeline ID, making it traceable.",
       "Critical", "Integration",
       "#3036", "CI / GitLab + Backend / Actuator",
       "Core goal of #3036: traceability from deployed build to CI pipeline."),

    tc("TC-XSV-048",
       "Build date displays formatted Maven build timestamp",
       "TTT application is deployed and running.",
       "1. Observe footer 'Build date' value.\n"
       "2. Call GET /api/ttt/actuator/info.\n"
       "3. Extract `build.maven.buildtime` from the response.\n"
       "4. Verify the footer date is a formatted version of the raw buildtime.\n"
       "5. Verify the date corresponds to when the CI pipeline ran (not stale).",
       "Footer 'Build date' is a human-readable format of build.maven.buildtime.\n"
       "Date is recent (matches the latest deployment, not an old build).",
       "Medium", "Functional",
       "#3036", "Frontend / Footer"),

    tc("TC-XSV-049",
       "Build number updates when non-TTT service changes (CI always rebuilds TTT)",
       "A code change is made ONLY in vacation service (no ttt-backend changes).\n"
       "CI pipeline runs on release/2.1.",
       "1. Note current build number on the environment.\n"
       "2. Make a change only to vacation service code (or verify from CI history).\n"
       "3. After CI pipeline runs, verify TTT service was also rebuilt.\n"
       "4. Deploy the new build.\n"
       "5. Check footer — verify build number has updated to the new pipeline ID.\n"
       "6. Repeat for a change only in calendar or email service.",
       "Build number updates even when only non-TTT services changed.\n"
       "This is because CI always rebuilds the TTT service on key branches.\n"
       "The actuator always reflects the latest pipeline.",
       "High", "Integration",
       "#3036, !5275", "CI / GitLab",
       "The fix in .gitlab-ci-ttt-module.yml ensures TTT service always rebuilds "
       "on master, release/*, hotfix/*, pre-release/*, development-ttt branches."),

    tc("TC-XSV-050",
       "Footer visible on login page before authentication",
       "User has not logged in.",
       "1. Open TTT login page.\n"
       "2. Observe the footer area.\n"
       "3. Verify 'Build #' and 'Build date' are displayed.\n"
       "4. Verify the values are not default ('0' / '0').\n"
       "5. Call GET /actuator/info (frontend-app actuator) — this is unauthenticated.\n"
       "6. Call GET /api/ttt/actuator/info — also allowed without auth.",
       "Footer is visible on the login page with correct build info.\n"
       "Build # and Build date show real values, not defaults.\n"
       "/actuator/info is accessible without authentication (nginx allows it).",
       "Medium", "Functional",
       "#3036", "Frontend / Footer + Nginx",
       "Nginx config (ttt.conf lines 55-58) explicitly allows /actuator/info without access restrictions."),

    tc("TC-XSV-051",
       "Actuator endpoint accessible without authentication",
       "No authentication token/session.",
       "1. Call GET /api/ttt/actuator/info without any auth headers.\n"
       "2. Verify HTTP 200 response.\n"
       "3. Verify response contains build.app.version, build.maven.buildtime.\n"
       "4. Call GET /api/ttt/actuator/health — verify this is also accessible.\n"
       "5. Call GET /api/ttt/actuator/env — verify this is RESTRICTED (403/401).\n"
       "6. Verify broader /actuator endpoints are restricted by nginx.",
       "GET /api/ttt/actuator/info returns 200 without auth.\n"
       "Response contains full build metadata.\n"
       "Other actuator endpoints (env, beans, etc.) are restricted.\n"
       "Only /actuator/info and /actuator/health are publicly accessible.",
       "Medium", "Security",
       "#3036", "Nginx / Security",
       "Nginx allows specific /actuator/info but blocks broader /actuator paths."),

    tc("TC-XSV-052",
       "Default footer values when actuator is unavailable",
       "ttt-backend service is temporarily down or unreachable.",
       "1. [Simulated] Block /api/ttt/actuator/info (e.g., service stopped).\n"
       "2. Load the login page or refresh a logged-in page.\n"
       "3. Observe the footer.\n"
       "4. Verify default values: Build #: 0 | Build date: 0.\n"
       "5. Restore ttt-backend service.\n"
       "6. Refresh page. Verify build info loads correctly.",
       "When actuator is unavailable, footer shows defaults: Build #: 0, Build date: 0.\n"
       "No error message or broken UI.\n"
       "After service recovery, footer shows correct values on next page load.",
       "Low", "Negative",
       "#3036", "Frontend / Error Handling",
       "Footer default props: build='0', date='0'. Saga handles API failure gracefully."),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Update existing cross-service.xlsx
# ══════════════════════════════════════════════════════════════════

def main():
    xlsx_path = os.path.join(os.path.dirname(__file__), "cross-service", "cross-service.xlsx")

    wb = openpyxl.load_workbook(xlsx_path)

    # ── Add TS-XSV-Banner tab ──
    ws_banner = wb.create_sheet("TS-XSV-Banner")
    ws_banner.sheet_properties.tabColor = TAB_BLUE
    n_banner = write_ts_tab(ws_banner, "TS-XSV-Banner — InnovationLab Banner (#3392)", BANNER_TESTS)
    print(f"  TS-XSV-Banner: {n_banner} cases")

    # ── Add TS-XSV-BuildInfo tab ──
    ws_build = wb.create_sheet("TS-XSV-BuildInfo")
    ws_build.sheet_properties.tabColor = TAB_BLUE
    n_build = write_ts_tab(ws_build, "TS-XSV-BuildInfo — CI Build Number (#3036)", BUILD_INFO_TESTS)
    print(f"  TS-XSV-BuildInfo: {n_build} cases")

    # ── Update Plan Overview tab ──
    ws_plan = wb["Plan Overview"]
    # Find the last used row
    max_row = ws_plan.max_row
    # Append new suite links
    r = max_row + 2
    ws_plan.cell(row=r, column=1, value="S78 Supplement — New Suites:").font = FONT_SECTION
    r += 1
    link1 = ws_plan.cell(row=r, column=1, value=f"TS-XSV-Banner — InnovationLab Banner — {n_banner} cases")
    link1.font = FONT_LINK_BOLD
    link1.hyperlink = "#'TS-XSV-Banner'!A1"
    r += 1
    link2 = ws_plan.cell(row=r, column=1, value=f"TS-XSV-BuildInfo — CI Build Number — {n_build} cases")
    link2.font = FONT_LINK_BOLD
    link2.hyperlink = "#'TS-XSV-BuildInfo'!A1"
    r += 1
    ws_plan.cell(row=r, column=1,
                 value=f"Total cross-service cases: 30 (existing) + {n_banner + n_build} (S78) = {30 + n_banner + n_build}").font = FONT_BODY_BOLD

    # ── Update Feature Matrix tab ──
    if "Feature Matrix" in wb.sheetnames:
        ws_fm = wb["Feature Matrix"]
        fm_max = ws_fm.max_row
        r = fm_max + 1
        # Add banner row
        banner_data = ["InnovationLab Banner", n_banner, 0, 0, 2, 0, 0, 0, n_banner, "#3392"]
        for col, val in enumerate(banner_data, 1):
            c = ws_fm.cell(row=r, column=col, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_LEFT
        link_cell = ws_fm.cell(row=r, column=1)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = "#'TS-XSV-Banner'!A1"
        r += 1
        # Add build info row
        build_data = ["CI Build Number", 0, n_build, 0, 0, 1, 0, 0, n_build, "#3036"]
        for col, val in enumerate(build_data, 1):
            c = ws_fm.cell(row=r, column=col, value=val)
            c.font = FONT_BODY
            c.border = THIN_BORDER
            c.alignment = ALIGN_LEFT
        link_cell = ws_fm.cell(row=r, column=1)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = "#'TS-XSV-BuildInfo'!A1"

    # ── Save ──
    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")
    print(f"New tabs: TS-XSV-Banner ({n_banner}), TS-XSV-BuildInfo ({n_build})")
    print(f"Total new cases: {n_banner + n_build}")


if __name__ == "__main__":
    main()
