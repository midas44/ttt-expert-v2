#!/usr/bin/env python3
"""
Sick Leave Test Documentation Generator — Phase B
Generates test-docs/sick-leave/sick-leave.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 8 TS- test suite tabs.

Based on vault knowledge: 4700+ word deep-dive, 45+ GitLab tickets mined,
dual status model code analysis, UI exploration via Playwright,
API surface (7 endpoints), DB schema analysis.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "sick-leave")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sick-leave.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_crud_cases():
    """TS-SickLeave-CRUD: Create, edit, close, delete, view for employee."""
    return [
        {
            "id": "TC-SL-001", "title": "Create sick leave — happy path",
            "preconditions": "Enabled employee with a manager.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Sick Leaves page (/sick-leave/my)\n3. Click 'Add a sick note' button\n4. In the creation dialog, set Start date = today, End date = today + 5 days\n5. Verify 'Calendar days' auto-calculates (expect 6)\n6. Leave 'Sick note number' field empty (optional on create)\n7. Leave 'Caring for a family member' unchecked (default false)\n8. Click 'Save'\n9. Verify success notification appears\n10. Verify new sick leave row in the table with State = 'Started' (or 'Planned' if future dates)\n11. Verify table shows correct dates and calendar days count\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Sick leave created with status OPEN, accountingStatus NEW. Table shows dates, calendar days. State shows 'Started' for current period or 'Planned' for future start.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §8 Create Flow", "module": "sick-leave/crud",
            "notes": "Any authenticated user can create — no permission check (BUG-SL-1). startDate == endDate allowed (single-day)."
        },
        {
            "id": "TC-SL-002", "title": "Create sick leave with document number",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set Start date = next Monday, End date = next Friday\n5. In 'Sick note number' field, enter 'SN-2026-001'\n6. Click 'Save'\n7. Verify sick leave created\n8. Verify 'Number' column in table shows 'SN-2026-001'\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Sick leave created with certificate number displayed in the table. Number max length 40 chars.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §4 DTO", "module": "sick-leave/crud",
            "notes": "Number field is optional on create, required on close. Max 40 characters."
        },
        {
            "id": "TC-SL-003", "title": "Create sick leave with file attachments",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set dates (5 calendar days)\n5. Click 'Attach file' or file upload area\n6. Upload 1 file (< 5MB, e.g., a PDF scan)\n7. Verify file appears in the attachment list\n8. Click 'Save'\n9. Verify sick leave created\n10. Open the sick leave details\n11. Verify attached file is visible and downloadable\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Sick leave created with file attachment. File visible in details view and downloadable.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §A5 File Handling", "module": "sick-leave/crud",
            "notes": "Max 5 files, 5MB each. Two-step upload: POST /v1/files/upload → include UUID in create request."
        },
        {
            "id": "TC-SL-004", "title": "Create sick leave with notify-also recipients",
            "preconditions": "Employee with at least one colleague in the same office.\nQuery: SELECT e.login, e2.login AS colleague FROM ttt_vacation.employee e JOIN ttt_vacation.employee e2 ON e.office_id = e2.office_id AND e.id != e2.id WHERE e.enabled = true AND e2.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set dates (3 days)\n5. In 'Also notify' field, type colleague's name and select from dropdown\n6. Click 'Save'\n7. Verify sick leave created\nDB-CHECK: SELECT COUNT(*) FROM ttt_vacation.sick_leave_notify_also WHERE sick_leave_id = <created_id>",
            "expected": "Sick leave created with notify-also recipient stored in DB. Notification sent to manager + notify-also recipients.",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-sick-leave-module.md §Key UI Flows", "module": "sick-leave/crud",
            "notes": "GET sick-leave/{id} does NOT return notifyAlso — verification via DB only."
        },
        {
            "id": "TC-SL-005", "title": "Create single-day sick leave (startDate == endDate)",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set Start date = tomorrow (a working day), End date = same day\n5. Verify 'Calendar days' shows 1\n6. Click 'Save'\n7. Verify sick leave created with 1 calendar day\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Single-day sick leave created. Calendar days = 1. Validator allows startDate == endDate.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §5 Validator", "module": "sick-leave/crud",
            "notes": "Validator: startDate.isBefore(endDate) || startDate.isEqual(endDate)."
        },
        {
            "id": "TC-SL-006", "title": "Edit sick leave dates (OPEN status)",
            "preconditions": "Employee with an existing OPEN sick leave.\nSETUP: Via API — POST /api/vacation/v1/sick-leaves to create.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a sick leave for the employee (5 days)\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Find the sick leave row\n4. Click the edit (pencil) icon\n5. Verify edit dialog opens with current dates\n6. Extend End date by 3 days\n7. Verify 'Calendar days' recalculates\n8. Click 'Save'\n9. Verify table row updates with new dates and day count\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Dates updated. Calendar days and working days recalculated. State unchanged.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §9 Patch Flow", "module": "sick-leave/crud",
            "notes": "PATCH validates date order and crossing. Working days auto-recalculated from calendar."
        },
        {
            "id": "TC-SL-007", "title": "Edit sick leave — add document number",
            "preconditions": "Employee with an OPEN sick leave without a number.\nSETUP: Via API — create sick leave without number field.",
            "steps": "SETUP: Via API — create a sick leave without 'number' field\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click edit on the sick leave row\n4. Enter 'SN-2026-002' in the 'Sick note number' field\n5. Click 'Save'\n6. Verify 'Number' column shows 'SN-2026-002' in the table\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Document number saved and displayed in the table.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §4 DTO", "module": "sick-leave/crud",
            "notes": "Number field: @Size(max=40). Required for close action but optional for edit."
        },
        {
            "id": "TC-SL-008", "title": "Close sick leave — happy path (with number)",
            "preconditions": "Employee with an OPEN sick leave that has a document number.\nSETUP: Via API — create a sick leave with number field.",
            "steps": "SETUP: Via API — create sick leave with number='SN-001'\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Find the OPEN sick leave row\n4. Click the close action button\n5. Verify close dialog appears\n6. Verify document number field is pre-filled with 'SN-001'\n7. Click 'Close' / 'Save'\n8. Verify State changes to 'Ended' (CLOSED)\nDB-CHECK: SELECT status FROM ttt_vacation.sick_leave WHERE id = <id> — expect 'CLOSED'\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Sick leave closed. State = Ended/CLOSED. Accounting status unchanged (still NEW).",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §9 step 8", "module": "sick-leave/crud",
            "notes": "Close = PATCH status=CLOSED. Requires number to be non-empty. Closing Planned is allowed (#2973 design decision)."
        },
        {
            "id": "TC-SL-009", "title": "Close sick leave — requires document number",
            "preconditions": "Employee with an OPEN sick leave WITHOUT a document number.\nSETUP: Via API — create sick leave without number.",
            "steps": "SETUP: Via API — create sick leave without number\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click the close action on the sick leave row\n4. Verify close dialog shows empty number field\n5. Leave number field empty\n6. Click 'Close' / 'Save'\n7. Verify validation error: number is required to close\n8. Enter 'SN-003' in the number field\n9. Click 'Close' / 'Save'\n10. Verify sick leave now closed successfully\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Close blocked without number. Error: 'exception.validation.sickLeave.number.empty'. Close succeeds after entering number.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §9 step 8, §13 Error Codes", "module": "sick-leave/crud",
            "notes": "Error code: exception.validation.sickLeave.number.empty. Max 40 chars."
        },
        {
            "id": "TC-SL-010", "title": "Delete sick leave (OPEN, accounting=NEW)",
            "preconditions": "Employee with an OPEN/NEW sick leave.\nSETUP: Via API — create a sick leave.",
            "steps": "SETUP: Via API — create a sick leave\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Find the OPEN sick leave row\n4. Click the delete (X) action button\n5. Confirm deletion in the dialog\n6. Verify sick leave disappears from the active list (or shows State='Deleted')\nDB-CHECK: SELECT status FROM ttt_vacation.sick_leave WHERE id = <id> — expect 'DELETED'",
            "expected": "Sick leave soft-deleted. Status=DELETED in DB. Row disappears from active list.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §10 Delete Flow", "module": "sick-leave/crud",
            "notes": "Soft delete: sets status=DELETED. File/notify_also associations NOT cleaned up (BUG-SL-5)."
        },
        {
            "id": "TC-SL-011", "title": "View sick leave details",
            "preconditions": "Employee with at least one sick leave.\nSETUP: Via API — create a sick leave with number and files.",
            "steps": "SETUP: Via API — create sick leave with number='SN-VIEW-01'\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click the view/details icon on the sick leave row\n4. Verify details dialog opens showing:\n   - Employee name\n   - Start date and End date\n   - Calendar days count\n   - Sick note number\n   - State\n5. Verify close button (X) dismisses the dialog\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Details dialog shows all sick leave attributes. All fields match table data.",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-sick-leave-module.md §Modal System", "module": "sick-leave/crud",
            "notes": "Modal type: SICK_LIST_DETAILS. Employee view does NOT show accounting status column."
        },
        {
            "id": "TC-SL-012", "title": "Edit file attachments — add and remove",
            "preconditions": "Employee with an OPEN sick leave that has 1 file attached.\nSETUP: Via API — create sick leave with 1 file.",
            "steps": "SETUP: Via API — upload a file, create sick leave with filesIds=[uuid]\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click edit on the sick leave row\n4. Verify current file is shown in the attachment list\n5. Upload a second file\n6. Remove the original file\n7. Click 'Save'\n8. Open details — verify only the new file is attached\nDB-CHECK: SELECT COUNT(*) FROM ttt_vacation.sick_leave_file WHERE sick_leave_id = <id> — expect 1\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Original file removed, new file added. Diff-and-sync on update replaces old associations.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-service-implementation.md §File Handling", "module": "sick-leave/crud",
            "notes": "Edit uses two-step PATCH: first dates/number, then file diff + second PATCH for filesIds."
        },
    ]


def get_lifecycle_cases():
    """TS-SL-Lifecycle: Dual status model, state coupling, transitions."""
    return [
        {
            "id": "TC-SL-013", "title": "Accounting status NEW → PROCESSING (accountant)",
            "preconditions": "Accountant with access to the sick leave's office. OPEN/NEW sick leave exists.\nSETUP: Via API — create a sick leave.\nQuery: SELECT e.login AS employee, a.login AS accountant FROM ttt_vacation.employee e JOIN ttt_vacation.office_accountant oa ON e.office_id = oa.office_id JOIN ttt_vacation.employee a ON oa.employee_id = a.id WHERE e.enabled = true AND a.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a sick leave for the employee\n1. Login as the accountant\n2. Navigate to Sick Leave Accounting page (/accounting/sick-leaves)\n3. Find the sick leave row\n4. Verify Status column shows inline dropdown with current value 'New'\n5. Click the Status dropdown\n6. Select 'Pending' (PROCESSING)\n7. Verify Status updates to 'Pending'\n8. Verify State remains 'Started' or 'Planned' (OPEN)\nDB-CHECK: SELECT status, accounting_status FROM ttt_vacation.sick_leave WHERE id = <id> — expect status=OPEN, accounting_status=PROCESSING\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Accounting status changes to PROCESSING. Main status stays OPEN. Accountant ID auto-set.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §1 Status Coupling", "module": "sick-leave/lifecycle",
            "notes": "Inline dropdown — no modal. Status coupling: PROCESSING → OPEN."
        },
        {
            "id": "TC-SL-014", "title": "Accounting status PROCESSING → PAID → auto-close",
            "preconditions": "OPEN/PROCESSING sick leave with a document number. Accountant has access.\nSETUP: Via API — create sick leave with number, set accounting_status=PROCESSING.",
            "steps": "SETUP: Via API — create sick leave with number='SN-PAY-01', then PATCH accounting_status=PROCESSING\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Find the PROCESSING sick leave\n4. Click Status dropdown\n5. Select 'Paid'\n6. Verify Status changes to 'Paid'\n7. Verify State changes to 'Ended' (CLOSED) — auto-close on PAID\nDB-CHECK: SELECT status, accounting_status FROM ttt_vacation.sick_leave WHERE id = <id> — expect status=CLOSED, accounting_status=PAID\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Setting PAID auto-closes the sick leave. Both status and accounting_status update. Main status derived from accounting status.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §1 updateStatus()", "module": "sick-leave/lifecycle",
            "notes": "Critical coupling: PAID → CLOSED, REJECTED → REJECTED, PROCESSING → OPEN."
        },
        {
            "id": "TC-SL-015", "title": "Accounting status → REJECTED → status=REJECTED",
            "preconditions": "OPEN/NEW sick leave. Accountant has access.\nSETUP: Via API — create a sick leave.",
            "steps": "SETUP: Via API — create a sick leave\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Find the NEW sick leave\n4. Click Status dropdown\n5. Select 'Rejected'\n6. Verify Status changes to 'Rejected'\n7. Verify State also changes to 'Rejected'\nDB-CHECK: SELECT status, accounting_status FROM ttt_vacation.sick_leave WHERE id = <id> — expect status=REJECTED, accounting_status=REJECTED\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Accounting REJECTED sets main status to REJECTED as well. Both dimensions show 'Rejected'.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §1", "module": "sick-leave/lifecycle",
            "notes": "Dual 'Rejected' state: both State and Status show 'Rejected'. Known UX issue: 'Rejected Rejected' in filter (BUG-SL-4)."
        },
        {
            "id": "TC-SL-016", "title": "Reverse: PAID → NEW (reopen closed sick leave)",
            "preconditions": "CLOSED/PAID sick leave. Admin or chief accountant user.\nSETUP: Via API — create sick leave, set PAID via PATCH.",
            "steps": "SETUP: Via API — create sick leave with number, PATCH accounting_status=PAID (auto-closes)\n1. Login as admin/chief accountant\n2. Navigate to /accounting/sick-leaves\n3. Find the PAID/Ended sick leave\n4. Click Status dropdown (should be available for admin/chief accountant)\n5. Select 'New'\n6. Verify Status reverts to 'New'\n7. Verify State reverts to 'Started' or 'Planned' (OPEN) — reopened!\nDB-CHECK: SELECT status, accounting_status FROM ttt_vacation.sick_leave WHERE id = <id> — expect status=OPEN, accounting_status=NEW\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Setting accountingStatus=NEW on a PAID sick leave REOPENS it. Status reverts to OPEN. No state machine guardrail prevents this.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §1, BUG-SL-3", "module": "sick-leave/lifecycle",
            "notes": "Unrestricted transitions — any-to-any (BUG-SL-3). Paid→New reversal is a design risk."
        },
        {
            "id": "TC-SL-017", "title": "Accountant comment — inline tooltip",
            "preconditions": "Accountant with an OPEN sick leave to comment on.\nSETUP: Via API — create a sick leave.",
            "steps": "SETUP: Via API — create a sick leave\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Find the sick leave row\n4. Click the comment (speech bubble) icon\n5. Verify tooltip overlay with textarea appears\n6. Type 'Pending certificate from clinic'\n7. Save the comment (click outside or save button)\n8. Verify comment icon indicates a comment exists\nDB-CHECK: SELECT accountant_comment FROM ttt_vacation.sick_leave WHERE id = <id> — expect 'Pending certificate from clinic'\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Comment saved via inline tooltip (not a modal). Stored in accountant_comment field.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-accounting-workflow.md §Action Buttons", "module": "sick-leave/lifecycle",
            "notes": "Inline Tooltip overlay with textarea, not a modal. ADD_COMMENT modal is orphaned in code."
        },
        {
            "id": "TC-SL-018", "title": "Computed states: Planned vs Started vs Overdue",
            "preconditions": "Three sick leaves: future start (Planned), current (Started), past end (Overdue).\nSETUP: Via API — create 3 sick leaves with different date ranges.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create 3 sick leaves for the same employee:\n  a) Future start (next month) — should show as 'Planned'\n  b) Current period (started yesterday, ends next week) — should show as 'Started'\n  c) Past end (ended last week, not closed) — should show as 'Overdue'\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Verify sick leave (a) shows State = 'Planned'\n4. Verify sick leave (b) shows State = 'Started'\n5. Verify sick leave (c) shows State = 'Overdue'\nCLEANUP: Via API — DELETE all 3 sick leaves",
            "expected": "Three OPEN sick leaves display different computed states based on date comparison. Planned/Started/Overdue are query-time computations, not stored.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-implementation.md §Dual Status System", "module": "sick-leave/lifecycle",
            "notes": "DB stores only OPEN. API returns SCHEDULED(=Planned)/OVERDUE computed from dates. PATCH only accepts OPEN/CLOSED."
        },
        {
            "id": "TC-SL-019", "title": "Edit DELETED sick leave — only comment allowed",
            "preconditions": "Accountant, with a DELETED sick leave in the system.\nSETUP: Via API — create sick leave, then DELETE it.",
            "steps": "SETUP: Via API — create sick leave, then DELETE /api/vacation/v1/sick-leaves/{id}\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Filter by State = 'Deleted'\n4. Find the deleted sick leave row\n5. Verify edit (pencil) icon is NOT visible (edit disabled)\n6. Verify Status dropdown shows plain text (not selectable)\n7. Verify comment (speech bubble) icon IS still visible\n8. Click the comment icon\n9. Add a comment 'Duplicate entry removed'\n10. Save the comment\nDB-CHECK: SELECT accountant_comment FROM ttt_vacation.sick_leave WHERE id = <id>",
            "expected": "Only accountantComment can be modified on DELETED sick leaves. Edit and status dropdown disabled. Error: 'exception.validation.sickLeave.update.closed' if other fields modified.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §7 checkUpdatePermissions, #2636", "module": "sick-leave/lifecycle",
            "notes": "hasUnmodifiableForDeletedRequestField() — all fields except accountantComment are unmodifiable on DELETED."
        },
        {
            "id": "TC-SL-020", "title": "Delete PAID sick leave — blocked",
            "preconditions": "CLOSED/PAID sick leave.\nSETUP: Via API — create sick leave with number, PATCH to PAID.",
            "steps": "SETUP: Via API — create sick leave with number, PATCH accounting_status=PAID\n1. Login as the employee (or any user)\n2. Navigate to /sick-leave/my or /accounting/sick-leaves\n3. Find the PAID sick leave\n4. Attempt to delete (click delete button if visible)\n5. Verify deletion is blocked — either button is hidden or an error is returned\n6. Via API — attempt DELETE /api/vacation/v1/sick-leaves/{id}\n7. Verify HTTP 400: 'exception.validation.sickLeave.delete.closed'\nCLEANUP: Via API — PATCH accounting_status=NEW (to reopen), then DELETE",
            "expected": "PAID sick leave cannot be deleted. UI hides delete action. API returns 400.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §10, §13 Error Codes", "module": "sick-leave/lifecycle",
            "notes": "Delete guard checks accountingStatus == PAID only. CLOSED+non-PAID CAN be deleted."
        },
    ]


def get_accounting_cases():
    """TS-SL-Accounting: Accountant workflow, status changes, page features."""
    return [
        {
            "id": "TC-SL-021", "title": "Accounting page — table columns and filters",
            "preconditions": "Accountant user.\nQuery: SELECT a.login FROM ttt_vacation.employee a JOIN ttt_backend.employee be ON a.login = be.login JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ACCOUNTANT' AND a.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Verify 10 columns visible: Employee, Sick leave dates, Days, Work days, Sick note, Accountant, Salary office, State, Status, Actions\n4. Click 'Salary office' filter — verify dropdown shows list of offices (27+)\n5. Click 'State' filter — verify options: Planned, Started, Overdue, Ended, Rejected, Deleted\n6. Click 'Status' filter — verify options: New, Pending, Paid, Rejected\n7. Apply 'State: Started' filter\n8. Verify table shows only OPEN sick leaves in progress\n9. Clear filter\n10. Verify table returns to unfiltered view",
            "expected": "Accounting page shows 10 columns with working filters for Salary office (27), State (6+), Status (4).",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-accounting-workflow.md §Table Columns", "module": "sick-leave/accounting",
            "notes": "State filter may show 'Rejected Rejected' label duplication (BUG-SL-4). Status is inline dropdown."
        },
        {
            "id": "TC-SL-022", "title": "Accounting — edit sick leave dates",
            "preconditions": "Accountant with a non-deleted, non-PAID sick leave.\nSETUP: Via API — create a sick leave.",
            "steps": "SETUP: Via API — create a sick leave\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Click edit (pencil) icon on the sick leave row\n4. Verify edit dialog shows: Employee (read-only), Start date, End date, Calendar days (auto-calc), Sick note number\n5. Change End date to 3 days later\n6. Verify Calendar days recalculates\n7. Click 'Save'\n8. Verify table updates with new dates\nDB-CHECK: SELECT end_date, total_days, work_days FROM ttt_vacation.sick_leave WHERE id = <id>\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Accountant can edit dates. Calendar days and working days auto-recalculate. Employee field is read-only in edit dialog.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-accounting-workflow.md §Edit Dialog", "module": "sick-leave/accounting",
            "notes": "Edit dialog has NO accounting status field — managed only via inline dropdown."
        },
        {
            "id": "TC-SL-023", "title": "Accounting — edit PAID sick leave as admin",
            "preconditions": "PAID sick leave. User with ROLE_ADMIN or ROLE_CHIEF_ACCOUNTANT.\nSETUP: Via API — create sick leave with number, PATCH to PAID.",
            "steps": "SETUP: Via API — create sick leave, PATCH accounting_status=PAID\n1. Login as admin\n2. Navigate to /accounting/sick-leaves\n3. Find the PAID sick leave\n4. Verify edit icon IS visible for admin\n5. Click edit\n6. Change End date\n7. Click 'Save'\n8. Verify dates updated\nDB-CHECK: SELECT end_date FROM ttt_vacation.sick_leave WHERE id = <id>\nCLEANUP: Via API — PATCH accounting_status=NEW, then DELETE",
            "expected": "Admin/chief accountant CAN edit PAID sick leaves. Regular accountant (office-level) can only edit PAID for their own office.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §7 checkUpdatePermissions", "module": "sick-leave/accounting",
            "notes": "PAID guard: only ADMIN, CHIEF_ACCOUNTANT, or assigned office accountant can modify."
        },
        {
            "id": "TC-SL-024", "title": "Accounting — office accountant scope restriction",
            "preconditions": "Office accountant and a sick leave from a DIFFERENT office.\nSETUP: Via API — create sick leave for an employee from a different office.\nQuery: SELECT e.login AS employee, a.login AS accountant, e.office_id, a2.office_id AS acc_office FROM ttt_vacation.employee e JOIN ttt_vacation.office_accountant oa ON oa.office_id != e.office_id JOIN ttt_vacation.employee a ON oa.employee_id = a.id WHERE e.enabled = true AND a.enabled = true LIMIT 1",
            "steps": "SETUP: Via API — create sick leave for an employee in office A; login as accountant for office B\n1. Login as the accountant (office B)\n2. Navigate to /accounting/sick-leaves\n3. Find the sick leave from office A\n4. Verify the sick leave IS visible (accounting view shows all)\n5. Attempt to change the accounting status via dropdown\n6. Verify action is blocked — error 'exception.vacation.no.permission'\n7. Via API — attempt PATCH accounting_status=PROCESSING as office B accountant\n8. Verify HTTP 403\nCLEANUP: Via API — DELETE sick leave",
            "expected": "Office accountant can view but NOT modify accounting status for employees from other offices.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §7 checkAccountantPermissions", "module": "sick-leave/accounting",
            "notes": "Uses officeAccountantRepository.existsByEmployeeIdAndOfficeId(). Shared VacationSecurityException class."
        },
        {
            "id": "TC-SL-025", "title": "Accounting — overdue sick leave count badge",
            "preconditions": "Accountant. At least one OVERDUE sick leave (OPEN + past end_date).\nSETUP: Via API — create sick leave with past end_date (started and ended in the past).",
            "steps": "SETUP: Via API — create sick leave with start_date = 2 weeks ago, end_date = 1 week ago\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Check the page for an overdue count indicator or badge\n4. Filter by State = 'Overdue'\n5. Verify the created sick leave appears\n6. Via API — GET /api/vacation/v1/sick-leaves/count\n7. Verify response includes open and overdue counts\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Overdue sick leaves counted and surfaced. GET /count endpoint returns open_count and overdue_count.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §A7 Overdue Warning", "module": "sick-leave/accounting",
            "notes": "OverdueSickLeaveCommand — per-request check, not cron. API: GET /v1/sick-leaves/count."
        },
        {
            "id": "TC-SL-026", "title": "Accounting status: any-to-any transition verification",
            "preconditions": "Admin or chief accountant. Multiple sick leaves in different accounting statuses.\nSETUP: Via API — create 4 sick leaves, set accounting_status to NEW, PROCESSING, PAID, REJECTED respectively.",
            "steps": "SETUP: Via API — create 4 sick leaves with different accounting statuses\n1. Login as admin\n2. Navigate to /accounting/sick-leaves\n3. Test transition: NEW → PAID (skip PROCESSING)\n4. Test transition: PAID → REJECTED\n5. Test transition: REJECTED → PROCESSING\n6. Test transition: PROCESSING → NEW (backwards)\n7. For each transition, verify:\n   - Status dropdown allows the selection\n   - Main State updates according to coupling rules\n   - Accountant field updates on first change\nCLEANUP: Via API — PATCH all to NEW, then DELETE",
            "expected": "All transitions allowed — no state machine enforcement. This confirms BUG-SL-3: unrestricted accounting status transitions.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §1, BUG-SL-3", "module": "sick-leave/accounting",
            "notes": "Design issue: Paid→New reopens a closed sick leave. No guardrails."
        },
    ]


def get_manager_cases():
    """TS-SL-Manager: DM/TL/PM tabs, create for employee, permission scoping."""
    return [
        {
            "id": "TC-SL-027", "title": "DM/TL — My Department tab",
            "preconditions": "User with ROLE_DEPARTMENT_MANAGER or ROLE_TECH_LEAD. Subordinate employees with sick leaves.\nSETUP: Via API — create sick leave for a subordinate.\nQuery: SELECT m.login AS manager, e.login AS employee FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id JOIN ttt_backend.employee be ON m.login = be.login JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_DEPARTMENT_MANAGER','ROLE_TECH_LEAD') AND e.enabled = true AND m.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a sick leave for the subordinate employee\n1. Login as the manager (DM/TL)\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Verify 'My department' tab is visible and active\n4. Verify columns: Employee, Sick leave dates, Calendar days, State, Status, Actions\n5. Find the subordinate's sick leave in the table\n6. Verify State and Status columns are displayed correctly\n7. Verify Status is shown as plain text (not dropdown — manager cannot change accounting status)\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "My Department tab shows subordinate employees' sick leaves. Status shown as plain text, not editable dropdown.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #2622", "module": "sick-leave/manager",
            "notes": "Three tabs: My department (DM/TL), My projects (PM), personal. #2622: tab badge was per-page, now total."
        },
        {
            "id": "TC-SL-028", "title": "DM/TL — create sick leave for subordinate",
            "preconditions": "DM or TL with a subordinate employee.\nQuery: (same as TC-SL-027)",
            "steps": "1. Login as the DM/TL\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Click 'My department' tab\n4. Click 'Add a sick note' or create button\n5. In the creation dialog, search and select the subordinate employee\n6. Set Start date and End date (5 days)\n7. Enter sick note number 'SN-MGR-01'\n8. Click 'Save'\n9. Verify sick leave appears in the table for the selected employee\nDB-CHECK: SELECT employee, status FROM ttt_vacation.sick_leave WHERE id = <created_id>\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "DM/TL can create sick leave for subordinates. Employee search works. force=true hardcoded in manager view (bypasses overlap check).",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-sick-leave-module.md §Manager create/edit", "module": "sick-leave/manager",
            "notes": "Manager view: no file upload capability. force=true hardcoded. Client-side overlap check capped at 100 records."
        },
        {
            "id": "TC-SL-029", "title": "PM — My Projects tab (view only, no create/edit)",
            "preconditions": "User with ROLE_PROJECT_MANAGER. Employee on PM's project has a sick leave.\nSETUP: Via API — create sick leave for an employee on PM's project.\nQuery: SELECT pm.login AS pm, e.login AS employee FROM ttt_backend.project_member pme JOIN ttt_backend.employee be_pm ON pme.employee_id = be_pm.id AND pme.role ILIKE '%manager%' JOIN ttt_vacation.employee pm ON pm.login = be_pm.login JOIN ttt_backend.project_member pme2 ON pme2.project_id = pme.project_id AND pme2.employee_id != pme.employee_id JOIN ttt_backend.employee be_e ON pme2.employee_id = be_e.id JOIN ttt_vacation.employee e ON e.login = be_e.login WHERE pm.enabled = true AND e.enabled = true LIMIT 1",
            "steps": "SETUP: Via API — create a sick leave for the employee on PM's project\n1. Login as the PM\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Click 'My projects' tab\n4. Verify the employee's sick leave is visible\n5. Verify NO create button is available for PM\n6. Verify NO edit/delete action buttons on the row\n7. Verify PM can only view — no modification actions\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "PM can VIEW sick leaves for project employees but CANNOT create, edit, or delete. checkIfCurrentEmployeeIsNotPM() enforces this.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §7 PM access check, #2873", "module": "sick-leave/manager",
            "notes": "PM from another project cannot modify. Exception: PM who is also the employee's manager CAN edit."
        },
        {
            "id": "TC-SL-030", "title": "DM/TL — edit subordinate's sick leave",
            "preconditions": "DM/TL with a subordinate who has an OPEN sick leave.\nSETUP: Via API — create sick leave for the subordinate.",
            "steps": "SETUP: Via API — create sick leave for the subordinate\n1. Login as the DM/TL\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Click 'My department' tab\n4. Click edit (pencil) icon on the subordinate's sick leave\n5. Change End date by 2 days\n6. Click 'Save'\n7. Verify dates updated in the table\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "DM/TL can edit subordinate's sick leave dates. Calendar days recalculate.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #2873", "module": "sick-leave/manager",
            "notes": "DM/TL can create and edit for their subordinates. PM can only view (#2873)."
        },
        {
            "id": "TC-SL-031", "title": "Manager view — State and Status filters",
            "preconditions": "DM/TL with multiple sick leaves in different states.\nSETUP: Via API — create 3 sick leaves in OPEN, CLOSED, DELETED statuses.",
            "steps": "SETUP: Via API — create sick leaves in different statuses\n1. Login as the DM/TL\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Click 'State' filter dropdown\n4. Verify 'Deleted' option is NOT visible for non-accountant roles\n5. Select 'Started' filter\n6. Verify only active sick leaves shown\n7. Click 'Status' filter dropdown\n8. Select 'New'\n9. Verify combined filter works\n10. Clear all filters\nCLEANUP: Via API — DELETE all created sick leaves",
            "expected": "State filter hides 'Deleted' for non-accountant roles (#2622 fix). Filters work in combination.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #2622", "module": "sick-leave/manager",
            "notes": "#2622 bug: 'Deleted' was visible for non-accountants. Filters should reset when switching tabs."
        },
        {
            "id": "TC-SL-032", "title": "Tab badge counters — total count, not per-page",
            "preconditions": "DM/TL with more sick leaves than fit on one page.\nSETUP: Create enough sick leaves to span 2+ pages (>10).",
            "steps": "1. Login as the DM/TL\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Check the tab badge counter on 'My department' tab\n4. Verify the counter shows the TOTAL count (e.g., 15), not per-page (e.g., 10)\n5. Navigate to page 2 of the table\n6. Verify badge counter remains the same\nCLEANUP: Via API — DELETE created sick leaves if any",
            "expected": "Badge counter shows total sick leaves for the tab, not the current page count. #2622 fix.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #2622", "module": "sick-leave/manager",
            "notes": "Bug #2622: counter was per-page (not total) → added API response field."
        },
    ]


def get_permission_cases():
    """TS-SL-Permissions: Role matrix, visibility scope, route access."""
    return [
        {
            "id": "TC-SL-033", "title": "Employee sees only own sick leaves",
            "preconditions": "Regular employee (no management roles).\nSETUP: Via API — create sick leave for this employee AND for a different employee.\nQuery: SELECT e1.login AS employee, e2.login AS other FROM ttt_vacation.employee e1, ttt_vacation.employee e2 WHERE e1.enabled = true AND e2.enabled = true AND e1.id != e2.id AND e1.manager_id IS NOT NULL AND e2.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create sick leave for employee, create sick leave for other\n1. Login as the regular employee\n2. Navigate to /sick-leave/my\n3. Verify only own sick leave is visible in the table\n4. Verify other employee's sick leave is NOT visible\n5. Via API — GET /api/vacation/v1/sick-leaves?employeeLogin=<other> as employee\n6. Verify response is empty or returns own sick leaves only\nCLEANUP: Via API — DELETE both sick leaves",
            "expected": "Employee sees only own sick leaves. Cannot access others' data. Regression test for #3213.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #3213", "module": "sick-leave/permissions",
            "notes": "PRODUCTION BUG #3213: all employees could see everyone's sick leaves. CRITICAL regression test."
        },
        {
            "id": "TC-SL-034", "title": "Accounting route inaccessible to regular employee",
            "preconditions": "Regular employee without ROLE_ACCOUNTANT, ROLE_ADMIN, ROLE_CHIEF_ACCOUNTANT, ROLE_VIEW_ALL, or ROLE_DEPARTMENT_MANAGER.\nQuery: SELECT e.login FROM ttt_vacation.employee e LEFT JOIN ttt_backend.employee be ON e.login = be.login LEFT JOIN ttt_backend.employee_role er ON be.id = er.employee_id LEFT JOIN ttt_backend.role r ON er.role_id = r.id WHERE e.enabled = true AND (r.name IS NULL OR r.name NOT IN ('ROLE_ACCOUNTANT','ROLE_ADMIN','ROLE_CHIEF_ACCOUNTANT','ROLE_VIEW_ALL','ROLE_DEPARTMENT_MANAGER')) GROUP BY e.login ORDER BY random() LIMIT 1",
            "steps": "1. Login as the regular employee\n2. Navigate directly to /accounting/sick-leaves\n3. Verify page is not accessible — redirect or access denied\n4. Verify the navigation menu does NOT show 'Sick Leave Records' or 'Accounting > Sick Leaves' link",
            "expected": "Regular employee cannot access /accounting/sick-leaves. Route protected by SICK_LEAVE_ACCOUNTING_VIEW permission.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #3012", "module": "sick-leave/permissions",
            "notes": "Bug #3012: route was accessible to regular employees (checked VACATIONS:VIEW instead of SICK_LEAVE_ACCOUNTING_VIEW)."
        },
        {
            "id": "TC-SL-035", "title": "Any authenticated user can create sick leave for any employee (BUG)",
            "preconditions": "Two employees: creator and target.\nQuery: SELECT e1.login AS creator, e2.login AS target FROM ttt_vacation.employee e1, ttt_vacation.employee e2 WHERE e1.enabled = true AND e2.enabled = true AND e1.id != e2.id ORDER BY random() LIMIT 1",
            "steps": "1. Via API — POST /api/vacation/v1/sick-leaves as creator with login=target's login\n   Body: {login: '<target>', startDate: '<future>', endDate: '<future+5>', force: false}\n2. Verify HTTP 200 — sick leave created\n3. Verify the sick leave is for the target employee, not the creator\nDB-CHECK: SELECT sl.employee, e.login FROM ttt_vacation.sick_leave sl JOIN ttt_vacation.employee e ON sl.employee = e.id WHERE sl.id = <created_id>\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Any authenticated user can create a sick leave for any employee by supplying their login. This is a known permission bug (BUG-SL-1).",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §14 issue 1, BUG-SL-1", "module": "sick-leave/permissions",
            "notes": "No @PreAuthorize beyond AUTHENTICATED_USER on create endpoint. Unlike vacation/day-off which check ownership."
        },
        {
            "id": "TC-SL-036", "title": "Any user can delete non-PAID sick leave (BUG)",
            "preconditions": "Employee A's sick leave, Employee B as deleter.\nSETUP: Via API — create sick leave for employee A.",
            "steps": "SETUP: Via API — create sick leave for employee A\n1. Via API — DELETE /api/vacation/v1/sick-leaves/{id} as employee B\n2. Verify HTTP 200 — deletion succeeds\nDB-CHECK: SELECT status FROM ttt_vacation.sick_leave WHERE id = <id> — expect 'DELETED'",
            "expected": "Any authenticated user can delete any non-PAID sick leave. No owner check on delete endpoint (BUG-SL-1 related).",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §14 issues 3-4", "module": "sick-leave/permissions",
            "notes": "deleteById() only guards PAID accounting status. No owner check, no PM check. Differs from vacation/day-off."
        },
        {
            "id": "TC-SL-037", "title": "PM from another project cannot edit sick leave",
            "preconditions": "PM who is NOT the employee's manager or TL. Employee has an OPEN sick leave.\nSETUP: Via API — create sick leave for the employee.\nQuery: Find a PM and employee where PM has no management relationship to employee.",
            "steps": "SETUP: Via API — create sick leave for the employee\n1. Via API — PATCH /api/vacation/v1/sick-leaves/{id} as the PM with updated endDate\n2. Verify HTTP 403: 'exception.vacation.no.permission'\nDB-CHECK: SELECT end_date FROM ttt_vacation.sick_leave WHERE id = <id> — verify unchanged",
            "expected": "PM who is not the employee's manager/TL/accountant gets VacationSecurityException (403).",
            "priority": "High", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §7 checkIfCurrentEmployeeIsNotPM", "module": "sick-leave/permissions",
            "notes": "Uses VacationSecurityException (not SickLeave-specific). Error: exception.vacation.no.permission."
        },
        {
            "id": "TC-SL-038", "title": "Employee cannot see accounting status on personal page",
            "preconditions": "Regular employee with a sick leave.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Verify table columns: Sick leave dates, Calendar days, Number, Accountant, State, Actions\n4. Verify there is NO 'Status' (accounting status) column\n5. Verify the employee cannot see or change the accounting status",
            "expected": "Employee personal page does NOT show accounting status column. Employees see State only.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-accounting-workflow.md §Employee View", "module": "sick-leave/permissions",
            "notes": "Columns: dates, days, number, accountant, state, actions. No accounting status visible."
        },
    ]


def get_validation_cases():
    """TS-SL-Validation: Date order, overlap, crossing, file limits."""
    return [
        {
            "id": "TC-SL-039", "title": "Date validation — startDate after endDate rejected",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set Start date = next Friday, End date = next Monday (start > end)\n5. Click 'Save'\n6. Verify validation error: 'validation.sickLeave.dates.order'\n7. Verify error highlighted on both Start date and End date fields",
            "expected": "Creation blocked with date order validation error. Error applied to both date fields.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §5 Validator, §13 Error Codes", "module": "sick-leave/validation",
            "notes": "Error: validation.sickLeave.dates.order. Both startDate and endDate get constraint violation."
        },
        {
            "id": "TC-SL-040", "title": "Overlapping sick leave — rejected",
            "preconditions": "Employee with an existing OPEN sick leave (Mon-Fri next week).\nSETUP: Via API — create first sick leave.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create sick leave for Mon-Fri next week\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set Start date = Wednesday next week, End date = Wednesday + 3 days (overlaps existing)\n5. Click 'Save'\n6. Verify error: 'exception.validation.sickLeave.dates.crossing'\n7. Verify second sick leave NOT created",
            "expected": "Overlapping sick leaves blocked. Error: exception.validation.sickLeave.dates.crossing. Applied to startDate field.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §8 crossing check", "module": "sick-leave/validation",
            "notes": "Overlap check excludes DELETED and REJECTED sick leaves. Only checks against OPEN/CLOSED."
        },
        {
            "id": "TC-SL-041", "title": "Overlap check skips DELETED and REJECTED sick leaves",
            "preconditions": "Employee with a DELETED sick leave (Mon-Fri) and a REJECTED sick leave (Mon-Fri of week after).\nSETUP: Via API — create 2 sick leaves, delete one, reject the other via accountant.",
            "steps": "SETUP: Via API — create sick leave A (Mon-Fri), DELETE it. Create sick leave B (Mon-Fri week after), PATCH accounting_status=REJECTED.\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set dates overlapping the DELETED sick leave A dates\n5. Click 'Save'\n6. Verify sick leave created successfully (no crossing error)\n7. Delete the new sick leave\n8. Create another sick leave overlapping REJECTED sick leave B dates\n9. Verify also created successfully\nCLEANUP: Via API — DELETE all created sick leaves",
            "expected": "Overlap check correctly excludes DELETED and REJECTED sick leaves. New sick leaves can be created for those date ranges.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §C #2636, #2973", "module": "sick-leave/validation",
            "notes": "findCrossingSickLeaves() must skip DELETED+REJECTED. Tickets #2636, #2973 confirm this."
        },
        {
            "id": "TC-SL-042", "title": "Vacation crossing — force=false triggers 409",
            "preconditions": "Employee with an existing active vacation. Creating sick leave that overlaps.\nSETUP: Via API — create a vacation (APPROVED).\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 5 AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create and approve a vacation for next week\n1. Via API — POST /api/vacation/v1/sick-leaves with dates overlapping the vacation, force=false\n2. Verify HTTP 409: 'exception.sick.leave.crossing.vacation'\n3. Via API — POST same request with force=true\n4. Verify HTTP 200 — sick leave created despite vacation crossing\nDB-CHECK: Verify both sick leave and vacation exist for overlapping dates\nCLEANUP: Via API — DELETE sick leave, cancel and delete vacation",
            "expected": "force=false → 409 Conflict. force=true → created successfully. EmployeeSickLeaveOverlapsVacationEvent published.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §8 Vacation crossing", "module": "sick-leave/validation",
            "notes": "UI shows vacation crossing popup and lets user choose to force. force field is @NotNull on create."
        },
        {
            "id": "TC-SL-043", "title": "Vacation crossing — UI popup and force confirmation",
            "preconditions": "Employee with an existing active vacation for next week.\nSETUP: Via API — create and approve a vacation.",
            "steps": "SETUP: Via API — create and approve a vacation for next week\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set dates overlapping the approved vacation\n5. Click 'Save'\n6. Verify a crossing warning popup appears (VACATION_CROSSING modal)\n7. Confirm/force creation in the popup\n8. Verify sick leave created despite the overlap\nCLEANUP: Via API — DELETE sick leave, cancel and delete vacation",
            "expected": "UI detects vacation crossing client-side and shows warning popup. User can confirm to force creation.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-sick-leave-module.md §Create, VACATION_CROSSING modal", "module": "sick-leave/validation",
            "notes": "Frontend does client-side crossing check via pre-fetch. Modal type: VACATION_CROSSING."
        },
        {
            "id": "TC-SL-044", "title": "File attachment limit — max 5 files",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set valid dates\n5. Upload 5 files (each < 5MB)\n6. Verify all 5 files shown in attachment list\n7. Attempt to upload a 6th file\n8. Verify upload blocked or error: max 5 files\n9. Click 'Save' with 5 files\n10. Verify sick leave created with 5 attachments\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Maximum 5 files per sick leave. 6th file rejected. @Size(max=5) on filesIds.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §4 DTO: @Size(max=5)", "module": "sick-leave/validation",
            "notes": "Validated on DTO but not enforced in service. Test both UI and API enforcement."
        },
        {
            "id": "TC-SL-045", "title": "Document number max length 40 characters",
            "preconditions": "Enabled employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set valid dates\n5. Enter a 40-character number: 'ABCDEFGHIJKLMNOPQRSTUVWXYZ12345678901234'\n6. Click 'Save'\n7. Verify sick leave created with the full 40-char number\n8. Edit the sick leave\n9. Try entering a 41-character number: 'ABCDEFGHIJKLMNOPQRSTUVWXYZ123456789012345'\n10. Click 'Save'\n11. Verify validation error or truncation\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "40-char number accepted. 41+ chars rejected by @Size(max=40) validation.",
            "priority": "Low", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §4 DTO", "module": "sick-leave/validation",
            "notes": "@Size(max=40) on number field."
        },
    ]


def get_regression_cases():
    """TS-SL-Regression: Known bugs from tickets, production incidents."""
    return [
        {
            "id": "TC-SL-046", "title": "Regression: visibility scope — page refresh (#2524)",
            "preconditions": "Regular employee with a sick leave.\nSETUP: Via API — create a sick leave for the employee.",
            "steps": "SETUP: Via API — create sick leave for the employee\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Verify only own sick leave visible\n4. Refresh the page (F5 / Ctrl+R)\n5. Verify the page reloads and still shows ONLY own sick leaves\n6. Open browser developer tools → Network tab\n7. Verify the GET /v1/sick-leaves request includes employeeLogin parameter\n8. Verify no other employee's data appears\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Page refresh maintains visibility scope. GET request always includes employeeLogin param. Regression test for HotFix #2524.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §D #2524", "module": "sick-leave/regression",
            "notes": "Production 'flickering bug': on refresh, frontend sent request WITHOUT employeeLogin → API returned ALL."
        },
        {
            "id": "TC-SL-047", "title": "Regression: rejection order dependency (#2973) — close then reject",
            "preconditions": "Employee and accountant. OPEN sick leave with number.\nSETUP: Via API — create sick leave with number='SN-973-A'.\nQuery: (same employee/accountant pair as TC-SL-013)",
            "steps": "SETUP: Via API — create sick leave with number='SN-973-A'\n1. Login as the employee\n2. Close the sick leave (State → Ended)\n3. Login as the accountant\n4. Navigate to /accounting/sick-leaves\n5. Set accounting status to 'Rejected'\n6. Verify State changes to 'Rejected' and accounting status = 'Rejected'\n7. Navigate to My Tasks page for the employee\n8. Verify norm values are correctly reverted (sick leave days NOT counted in norm)\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Close → Reject sequence: norm correctly reverts. This is the CORRECT sequence per #2973.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §C #2973", "module": "sick-leave/regression",
            "notes": "Bug #2973 critical edge case: close→reject = OK, reject→close = BUG (norm incorrect)."
        },
        {
            "id": "TC-SL-048", "title": "Regression: rejection order dependency (#2973) — reject then close attempt",
            "preconditions": "Employee and accountant. OPEN sick leave.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave\n1. Login as the accountant\n2. Set accounting status to 'Rejected'\n3. Verify State = 'Rejected'\n4. Login as the employee\n5. Attempt to close the sick leave\n6. Verify close action is NOT available for 'Rejected' state\n7. If close is available (bug), verify norm calculation after close\nDB-CHECK: SELECT status, accounting_status FROM ttt_vacation.sick_leave WHERE id = <id>",
            "expected": "Close action should NOT be available for Rejected state. Per #2973 requirements: close only for Open and Overdue.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §C #2973 requirement 2", "module": "sick-leave/regression",
            "notes": "#2973 requirement: Close only for Open and Overdue. Closing Rejected was a bug. Closing Planned IS allowed."
        },
        {
            "id": "TC-SL-049", "title": "Regression: employee cannot edit Rejected sick leave (#2973)",
            "preconditions": "REJECTED sick leave for an employee.\nSETUP: Via API — create sick leave, PATCH accounting_status=REJECTED.",
            "steps": "SETUP: Via API — create sick leave, PATCH accounting_status=REJECTED\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Find the REJECTED sick leave\n4. Verify edit icon is NOT visible or is disabled\n5. Verify delete icon IS visible (REJECTED can be deleted since not PAID)\n6. Verify employee cannot modify dates or number\nDB-CHECK: Via API — PATCH /api/vacation/v1/sick-leaves/{id} as employee with new endDate\n7. Verify HTTP 400 or 403",
            "expected": "Employee CANNOT edit Rejected sick leave (view only). Per #2973 requirement 1. Accountant CAN still edit.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "sick-leave-ticket-findings.md §C #2973 requirement 1, #2567", "module": "sick-leave/regression",
            "notes": "#2973: employee cannot edit in Rejected state. #2567: accountant has no editing restrictions."
        },
        {
            "id": "TC-SL-050", "title": "Regression: deleted sick leave norm propagation (#2778, #2783)",
            "preconditions": "Employee with a sick leave that affects norm. Sick leave in current reporting period.\nSETUP: Via API — create sick leave covering several working days in the current period.",
            "steps": "SETUP: Via API — create sick leave (5 working days in current month)\n1. Login as the employee\n2. Navigate to My Tasks page (time reporting)\n3. Verify sick leave days are marked (orange color)\n4. Verify norm tooltip shows reduced hours (accounts for sick leave)\n5. Navigate to /sick-leave/my\n6. Delete the sick leave\n7. Navigate back to My Tasks page\n8. Verify sick leave day colors revert to normal (black)\n9. Verify norm tooltip recalculates (hours increase back to full norm)\n10. Verify cell input is available again for those days",
            "expected": "Deleting sick leave immediately reverts: day colors, norm tooltip values, cell input availability. No page refresh needed.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2778, #2783", "module": "sick-leave/regression",
            "notes": "Bugs #2778/#2783: deleted SL still showed in tasks, norm not recalculated. Fix required real-time propagation."
        },
        {
            "id": "TC-SL-051", "title": "Regression: rejected sick leave colors in My Tasks (#2863)",
            "preconditions": "Employee with a sick leave in current period. Accountant to reject it.\nSETUP: Via API — create sick leave for current period.",
            "steps": "SETUP: Via API — create sick leave (working days in current month)\n1. Login as the employee\n2. Navigate to My Tasks\n3. Verify sick leave days are marked orange\n4. Login as accountant\n5. Reject the sick leave (PATCH accounting_status=REJECTED)\n6. Login as the employee again\n7. Navigate to My Tasks\n8. Verify day colors revert to black (not orange)\n9. Verify norm tooltip recalculates correctly\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Rejection reverts day colors from orange to black. Norm recalculated. #2863 fix verified.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2863", "module": "sick-leave/regression",
            "notes": "#2863 HotFix: accountant rejection should revert day colors. Previously days remained orange."
        },
        {
            "id": "TC-SL-052", "title": "Regression: deleted sick leave in Planner (#2792)",
            "preconditions": "Employee with a sick leave in current period.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave for current period\n1. Login as a user who can see the Planner\n2. Navigate to Planner page\n3. Verify employee's sick leave days are marked red\n4. Delete the sick leave via API — DELETE /api/vacation/v1/sick-leaves/{id}\n5. Refresh or navigate back to Planner\n6. Verify day markings are removed (no longer red)\nCLEANUP: Already deleted",
            "expected": "Deleted sick leave no longer marks days red in Planner. #2792 HotFix verified.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2792", "module": "sick-leave/regression",
            "notes": "#2792: multiple fix attempts needed. Regression noted in comments."
        },
        {
            "id": "TC-SL-053", "title": "Regression: sick leave data in Confirmation page (#2819)",
            "preconditions": "Employee with a sick leave in the current reporting period.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave for current period\n1. Login as a manager who can see Confirmation page\n2. Navigate to Confirmation page\n3. Switch to 'By Employees' view\n4. Verify sick leave IS shown for the employee\n5. Switch to 'By Projects' view\n6. Verify sick leave data is consistent\n7. Via API — DELETE the sick leave\n8. Refresh Confirmation page\n9. Verify deleted sick leave NO longer shows\nCLEANUP: Already deleted",
            "expected": "Sick leave visible in both Confirmation views. Deleted sick leave removed. #2819 sub-bugs 2 and 4 verified.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2819", "module": "sick-leave/regression",
            "notes": "#2819 had 4 sub-bugs: weekend hours, deleted SL displayed, day-off transfer, SL missing from 'By Employees'."
        },
        {
            "id": "TC-SL-054", "title": "Regression: rejected sick leave excluded from Availability Chart copy (#2812)",
            "preconditions": "Employee with a REJECTED sick leave in current period.\nSETUP: Via API — create sick leave, reject it.",
            "steps": "SETUP: Via API — create sick leave, PATCH accounting_status=REJECTED\n1. Login as a user with Availability Chart access\n2. Navigate to Availability Chart for the employee's period\n3. Click 'Copy to clipboard' button\n4. Paste the copied data into a text editor\n5. Verify REJECTED sick leave is NOT included in the copied data\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Copy-to-clipboard excludes REJECTED sick leaves from the copied data. #2812 fix verified.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2812", "module": "sick-leave/regression",
            "notes": "#2812 HotFix: rejected sick leaves were included in copied availability data."
        },
        {
            "id": "TC-SL-055", "title": "Regression: sick leave not displayed in tables (#3257)",
            "preconditions": "Employee with a sick leave that exists in DB.\nSETUP: Via API — create sick leave.\nDB-CHECK: Verify sick leave exists in ttt_vacation.sick_leave table.",
            "steps": "SETUP: Via API — create sick leave, verify DB record exists\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Verify sick leave is visible in the table\n4. Login as the accountant\n5. Navigate to /accounting/sick-leaves\n6. Verify the same sick leave is visible in accounting table\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Sick leave visible in both My Sick Leaves and Sick Leave Records (accounting). Regression test for #3257.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §H #3257", "module": "sick-leave/regression",
            "notes": "#3257: data existed in DB but not shown in UI. Affected both pages."
        },
    ]


def get_notification_cases():
    """TS-SL-Notifications: Email events, recipients, editor type."""
    return [
        {
            "id": "TC-SL-056", "title": "Notification on sick leave creation (by employee)",
            "preconditions": "Employee with a manager.\nSETUP: Clear test emails for manager's address.\nQuery: SELECT e.login AS employee, m.login AS manager FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id WHERE e.enabled = true AND m.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via test API — clear email queue for manager\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Create a sick leave (5 days)\n4. Via test API — trigger notification sending (POST /api/ttt/test/v1/notifications)\n5. Via email API — GET /api/email/v1/emails?to=<manager_email>\n6. Verify notification received: template NOTIFY_SICKLEAVE_OPEN\n7. Verify notification includes: employee name, dates, duration\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Manager receives NOTIFY_SICKLEAVE_OPEN email. Recipients: manager + notifyAlso + per-office receivers.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §12 Events", "module": "sick-leave/notifications",
            "notes": "5 event types. BUG: getEditorType() uses == (identity) instead of .equals() (BUG-SL-2)."
        },
        {
            "id": "TC-SL-057", "title": "Notification on sick leave creation (by DM/TL)",
            "preconditions": "DM/TL creating sick leave for a subordinate.\nSETUP: Clear test emails.",
            "steps": "SETUP: Via test API — clear email queue\n1. Login as the DM/TL\n2. Navigate to /vacation/sick-leaves-of-employees\n3. Create sick leave for a subordinate\n4. Via test API — trigger notifications\n5. Via email API — check emails\n6. Verify template is NOTIFY_SICKLEAVE_OPEN_BY_SUPERVISOR (not plain _OPEN)\n7. Verify employee receives the notification (not the DM/TL)\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Supervisor creation uses _BY_SUPERVISOR template. Employee notified. Editor type should be SUPERVISOR.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §11 Notification System", "module": "sick-leave/notifications",
            "notes": "BUG-SL-2: getEditorType() always returns EMPLOYEE fallback due to == identity comparison."
        },
        {
            "id": "TC-SL-058", "title": "Notification on sick leave deletion",
            "preconditions": "Employee with a sick leave. Email queue cleared.\nSETUP: Via API — create sick leave, clear emails.",
            "steps": "SETUP: Via API — create sick leave, clear email queue\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Delete the sick leave\n4. Via test API — trigger notifications\n5. Via email API — check for NOTIFY_SICKLEAVE_DELETE template\n6. Verify manager received the deletion notification",
            "expected": "Deletion triggers NOTIFY_SICKLEAVE_DELETE email to manager.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-service-deep-dive.md §12 Events", "module": "sick-leave/notifications",
            "notes": "SickLeaveDeletedEvent → notifications. ID_105 email should exclude DELETED sick leaves."
        },
        {
            "id": "TC-SL-059", "title": "Notification on file attachment",
            "preconditions": "Employee with an existing sick leave. Email queue cleared.\nSETUP: Via API — create sick leave, clear emails.",
            "steps": "SETUP: Via API — create sick leave, upload a file\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Edit the sick leave\n4. Attach a new file\n5. Save\n6. Via test API — trigger notifications\n7. Via email API — check for NOTIFY_SICKLEAVE_FILES_ADDED template\n8. Verify notification sent to accountants of employee's office\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "File attachment triggers NOTIFY_SICKLEAVE_FILES_ADDED. Sent to office accountants only (not managers/PMs).",
            "priority": "Low", "type": "Hybrid",
            "req_ref": "sick-leave-ticket-findings.md §G #2673", "module": "sick-leave/notifications",
            "notes": "#2673: file email was sent to wrong recipients (managers/PMs too). Fix: only office accountants."
        },
        {
            "id": "TC-SL-060", "title": "Overlap notification excludes Deleted/Rejected sick leaves",
            "preconditions": "Employee with a DELETED sick leave overlapping a vacation.\nSETUP: Via API — create sick leave, delete it. Create an active vacation for the same period.",
            "steps": "SETUP: Via API — create sick leave overlapping a vacation, delete the sick leave\n1. Via API — check if overlap notification was generated for the DELETED sick leave\n2. Verify no NOTIFY_EMPLOYEE_SICKLEAVE_OVERLAPS_VACATION for the deleted sick leave\n3. Create a new OPEN sick leave overlapping the vacation (force=true)\n4. Verify overlap notification IS generated for the new active sick leave",
            "expected": "Overlap notifications skip DELETED/REJECTED sick leaves. Only active overlaps trigger notifications.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-ticket-findings.md §C #2973 requirement 3", "module": "sick-leave/notifications",
            "notes": "#2973 requirement 3: overlap notification should NOT fire for Rejected or Deleted."
        },
    ]


def get_family_member_cases():
    """TS-SL-FamilyMember: Upcoming familyMember flag (#3408)."""
    return [
        {
            "id": "TC-SL-061", "title": "Create sick leave with familyMember=true",
            "preconditions": "Enabled employee. #3408 feature deployed.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Set dates (5 days)\n5. Check the 'Caring for a family member' checkbox\n6. Click 'Save'\n7. Verify sick leave created with familyMember indicator visible\nDB-CHECK: SELECT family_member FROM ttt_vacation.sick_leave WHERE id = <id> — expect TRUE\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Sick leave created with familyMember=true. Checkbox label: 'По уходу за членом семьи' (RU) / 'Caring for a family member' (EN).",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §A #3408", "module": "sick-leave/family-member",
            "notes": "Sprint 16 feature. Default unchecked (false). Depends on #3409 for budgetNorm changes."
        },
        {
            "id": "TC-SL-062", "title": "Default familyMember=false on create",
            "preconditions": "Enabled employee.",
            "steps": "1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Click 'Add a sick note'\n4. Verify 'Caring for a family member' checkbox is UNCHECKED by default\n5. Do NOT check it\n6. Set dates, click 'Save'\n7. Verify sick leave created\nDB-CHECK: SELECT family_member FROM ttt_vacation.sick_leave WHERE id = <id> — expect FALSE\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "Default is familyMember=false (own illness). Checkbox unchecked by default.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §A #3408", "module": "sick-leave/family-member",
            "notes": "All existing historical sick leaves retroactively set to familyMember=false."
        },
        {
            "id": "TC-SL-063", "title": "Toggle familyMember on existing sick leave",
            "preconditions": "Employee with an OPEN sick leave where familyMember=false.\nSETUP: Via API — create sick leave (default familyMember=false).",
            "steps": "SETUP: Via API — create sick leave (familyMember=false by default)\n1. Login as the employee\n2. Navigate to /sick-leave/my\n3. Edit the sick leave\n4. Check the 'Caring for a family member' checkbox\n5. Click 'Save'\n6. Verify UI updates to show familyMember indicator\nDB-CHECK: SELECT family_member FROM ttt_vacation.sick_leave WHERE id = <id> — expect TRUE\n7. Edit again, uncheck the checkbox\n8. Save\nDB-CHECK: Verify family_member reverted to FALSE\nCLEANUP: Via API — DELETE /api/vacation/v1/sick-leaves/{id}",
            "expected": "familyMember flag can be toggled on edit. Norm recalculation should trigger on change.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-service-deep-dive.md §Upcoming familyMember", "module": "sick-leave/family-member",
            "notes": "Toggling should trigger norm recalculation — familyMember=true doesn't deduct from individual norm."
        },
        {
            "id": "TC-SL-064", "title": "familyMember=true does NOT deduct from individual norm",
            "preconditions": "Employee with no sick leaves in current period.\nSETUP: Note the employee's current norm values.",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks page\n3. Note current monthly norm value from tooltip\n4. Create a 5-day sick leave with familyMember=TRUE\n5. Navigate back to My Tasks\n6. Verify monthly norm value is UNCHANGED (family member sick leave does not deduct)\n7. Delete the sick leave\n8. Create a 5-day sick leave with familyMember=FALSE\n9. Navigate to My Tasks\n10. Verify monthly norm value DECREASED (own sick leave deducts from norm)\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "familyMember=true: norm unchanged. familyMember=false: norm decreased. Family member sick leaves go to budgetNorm instead.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §A #3408, #3409", "module": "sick-leave/family-member",
            "notes": "#3409: budgetNorm = Ni + admin_vac_hrs + familyMember_SL_hrs. familyMember=true goes to budgetNorm."
        },
        {
            "id": "TC-SL-065", "title": "familyMember display in Accounting page",
            "preconditions": "Accountant. Sick leaves with both familyMember=true and false.\nSETUP: Via API — create 2 sick leaves with different familyMember values.",
            "steps": "SETUP: Via API — create sick leave A (familyMember=false), sick leave B (familyMember=true)\n1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Verify familyMember column or indicator is visible\n4. Verify sick leave A shows as own illness\n5. Verify sick leave B shows as family member care\n6. If filter available, test filtering by familyMember type\nCLEANUP: Via API — DELETE both sick leaves",
            "expected": "Accounting page differentiates own vs family member sick leaves. Filter available.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §A #3408", "module": "sick-leave/family-member",
            "notes": "#3408 affects both 'My Sick Leaves' and 'Sick Leave Records' accounting page."
        },
        {
            "id": "TC-SL-066", "title": "Historical sick leaves default to familyMember=false",
            "preconditions": "Sick leaves that existed before #3408 deployment.\nQuery: SELECT sl.id, sl.employee, sl.start_date FROM ttt_vacation.sick_leave sl WHERE sl.status != 'DELETED' ORDER BY sl.start_date ASC LIMIT 5",
            "steps": "1. Login as the accountant\n2. Navigate to /accounting/sick-leaves\n3. Find sick leaves created before Sprint 16 deployment\n4. Verify they all show as own illness (familyMember=false)\nDB-CHECK: SELECT id, family_member FROM ttt_vacation.sick_leave WHERE start_date < '2026-04-01' LIMIT 10 — all should be FALSE",
            "expected": "All historical sick leaves retroactively treated as familyMember=false. No data migration issues.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "sick-leave-ticket-findings.md §A #3408", "module": "sick-leave/family-member",
            "notes": "Migration: all existing sick leaves default to familyMember=false."
        },
    ]


def get_cross_page_cases():
    """TS-SL-CrossPage: Norm propagation across pages after state changes."""
    return [
        {
            "id": "TC-SL-067", "title": "Cross-page: create sick leave → verify My Tasks, Planner, Confirmation",
            "preconditions": "Employee in current reporting period. No existing sick leaves.\nSETUP: Verify norm values before sick leave creation.",
            "steps": "1. Login as the employee\n2. Navigate to My Tasks — note norm tooltip value and day colors for next week\n3. Navigate to Planner — note day markings for next week\n4. Create a sick leave for next week (5 working days)\n5. Navigate to My Tasks\n6. Verify days are now marked orange (sick leave color)\n7. Verify norm tooltip decreased by sick leave working hours\n8. Verify input cells disabled for sick leave days (or restricted)\n9. Navigate to Planner\n10. Verify days are marked red for sick leave period\n11. Navigate to Confirmation page (if accessible)\n12. Verify sick leave displayed in both 'By Employees' and 'By Projects' views\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Sick leave creation propagates to My Tasks (orange days, reduced norm), Planner (red marking), Confirmation (visible in both views).",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E (8 tickets)", "module": "sick-leave/cross-page",
            "notes": "8 tickets (#2778-#2915) found norm propagation failures. Cross-page cache invalidation is the root issue."
        },
        {
            "id": "TC-SL-068", "title": "Cross-page: delete sick leave → all pages revert",
            "preconditions": "Employee with an active sick leave in current period.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave for current week\n1. Login as the employee\n2. Verify My Tasks shows orange days and reduced norm\n3. Verify Planner shows red day markings\n4. Delete the sick leave from /sick-leave/my\n5. Navigate to My Tasks\n6. Verify day colors revert to normal (black/default)\n7. Verify norm tooltip increases back to full value\n8. Verify cells are editable again\n9. Navigate to Planner\n10. Verify red markings removed\nCLEANUP: Already deleted",
            "expected": "All pages revert after deletion: My Tasks (colors, norm, cells), Planner (markings). No stale cache.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2778, #2792", "module": "sick-leave/cross-page",
            "notes": "Root issue: aggressive frontend caching. Pages don't invalidate on state changes from other pages."
        },
        {
            "id": "TC-SL-069", "title": "Cross-page: reject sick leave → norm reverts on all pages",
            "preconditions": "Employee with an active sick leave. Accountant to reject.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave for current period\n1. Login as the employee\n2. Verify My Tasks shows orange days\n3. Login as the accountant\n4. Reject the sick leave (accounting_status=REJECTED)\n5. Login as the employee\n6. Navigate to My Tasks\n7. Verify day colors revert to black (#2863 regression test)\n8. Verify norm tooltip recalculates\n9. Navigate to Planner\n10. Verify red markings removed (#2792)\n11. Check Availability Chart copy-to-clipboard (#2812)\n12. Verify rejected sick leave NOT included in copied data\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Rejection triggers full propagation: My Tasks colors revert, norm recalculates, Planner clears, Availability excludes.",
            "priority": "Critical", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2863, #2792, #2812", "module": "sick-leave/cross-page",
            "notes": "Combined regression test for 3 HotFix tickets. All involved rejection not propagating."
        },
        {
            "id": "TC-SL-070", "title": "Cross-page: day-off interaction with sick leave norm (#2901)",
            "preconditions": "Employee with a calendar event (day-off) and overlapping sick leave.\nSETUP: Via API — create calendar event, create overlapping sick leave.",
            "steps": "SETUP: Via API — ensure a production calendar day-off exists in the period. Create overlapping sick leave.\n1. Login as the employee\n2. Navigate to My Tasks\n3. Verify norm accounts for both sick leave and day-off\n4. Move the day-off outside the sick leave period (via /vacation/my Days Off tab)\n5. Navigate to My Tasks\n6. Verify norm recalculates to account for: sick leave still present, day-off moved out\n7. Verify norm correctly reflects the day-off is no longer overlapping\nCLEANUP: Via API — DELETE sick leave, revert day-off if possible",
            "expected": "Moving a day-off during sick leave triggers norm recalculation. Both sick leave and day-off impact reflected correctly.",
            "priority": "High", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2901", "module": "sick-leave/cross-page",
            "notes": "#2901 HotFix: moving day-off during overlapping sick leave didn't trigger norm recalculation."
        },
        {
            "id": "TC-SL-071", "title": "Cross-page: daily norm tooltip accuracy (#2915)",
            "preconditions": "Employee with a sick leave in the current period.\nSETUP: Via API — create sick leave.",
            "steps": "SETUP: Via API — create sick leave for 3 working days in current month\n1. Login as the employee\n2. Navigate to My Tasks\n3. Hover over a sick leave day to see the norm tooltip\n4. Verify the tooltip shows two values: actual hours / daily norm\n5. Verify the second value (daily norm) is correct — should reflect reduced norm for sick leave\n6. Hover over a non-sick-leave day\n7. Verify the daily norm is the standard value (8h or office-specific)\nCLEANUP: Via API — DELETE the sick leave",
            "expected": "Daily norm tooltip correctly shows reduced hours for sick leave days. Second value (after slash) is accurate.",
            "priority": "Medium", "type": "UI",
            "req_ref": "sick-leave-ticket-findings.md §E #2915", "module": "sick-leave/cross-page",
            "notes": "#2915: tooltip norm (second value after slash) was wrong. Sprint 9 fix."
        },
    ]


# ─── Plan, Feature Matrix, Risk Assessment ──────────────────────────────────

PLAN_OVERVIEW = {
    "title": "Sick Leave Module — Test Plan",
    "scope": "Comprehensive testing of the sick leave management module covering CRUD operations, dual status model (State + Accounting Status), accounting workflow, manager views (DM/TL/PM), role-based permissions, validation rules, cross-page norm propagation, notification system, and regression tests for 45+ known bugs including 3 production incidents.",
    "objectives": [
        "Verify all sick leave lifecycle operations (create, edit, close, delete, view)",
        "Validate the dual status model: State (Planned/Started/Overdue/Ended/Rejected/Deleted) × Accounting Status (New/Pending/Paid/Rejected)",
        "Test status coupling: PAID→CLOSED, REJECTED→REJECTED, PROCESSING→OPEN",
        "Verify accounting workflow: inline status dropdown, comment tooltip, office-scoped permissions",
        "Test manager views: My Department (DM/TL), My Projects (PM read-only), create/edit for subordinates",
        "Verify permission model: visibility scope per role, route access, creation permission gap (BUG-SL-1)",
        "Test all validation rules: date order, overlap check (excludes DELETED/REJECTED), vacation crossing (force flag), close requires number",
        "Regression tests for production bugs: visibility leak (#3213, #2524), accounting route access (#3012)",
        "Cross-page norm propagation: My Tasks, Planner, Confirmation page, Availability Chart — all must update on state changes",
        "Upcoming: familyMember flag (#3408) — own illness vs family member care, budgetNorm impact",
    ],
    "environments": [
        "Primary: qa-1 (ttt-qa-1.noveogroup.com)",
        "Secondary: timemachine (ttt-timemachine.noveogroup.com) — for clock-dependent tests",
        "Production-like: stage (ttt-stage.noveogroup.com) — for verification only",
    ],
    "approach": "UI-first testing with API setup/cleanup. Tests describe browser actions for all user-facing scenarios. API used for state creation (SETUP steps), data verification (DB-CHECK), and API-only features. Preconditions include SQL query hints for dynamic test data selection. 3 pages tested: My Sick Leaves (/sick-leave/my), Sick Leaves of Employees (/vacation/sick-leaves-of-employees), Accounting (/accounting/sick-leaves).",
    "dependencies": [
        "CAS authentication for multi-user tests (employee, accountant, DM/TL, PM roles needed)",
        "Production calendar data for norm calculation and day-off interaction tests",
        "Email service for notification tests (via test API trigger)",
        "#3408 feature deployment for familyMember test cases (Sprint 16)",
    ],
}

FEATURE_MATRIX = [
    # (Feature, CRUD, Lifecycle, Accounting, Manager, Perms, Validation, Regression, Notif, FamilyMember, CrossPage, Total)
    ("Create sick leave", 5, 0, 0, 1, 1, 0, 0, 1, 2, 1, 11),
    ("Edit sick leave", 3, 1, 1, 1, 0, 0, 1, 0, 1, 0, 8),
    ("Close sick leave", 2, 0, 0, 0, 0, 0, 1, 0, 0, 0, 3),
    ("Delete sick leave", 1, 1, 0, 0, 1, 0, 2, 1, 0, 1, 7),
    ("View details/page", 1, 1, 1, 2, 1, 0, 1, 0, 1, 0, 8),
    ("Accounting status", 0, 4, 2, 0, 0, 0, 0, 0, 0, 0, 6),
    ("Status coupling", 0, 2, 0, 0, 0, 0, 2, 0, 0, 0, 4),
    ("Date validation", 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Overlap check", 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Vacation crossing", 0, 0, 0, 0, 0, 2, 0, 1, 0, 0, 3),
    ("File handling", 1, 0, 0, 0, 0, 1, 0, 0, 0, 0, 2),
    ("Manager views", 0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 3),
    ("Norm propagation", 0, 0, 0, 0, 0, 0, 3, 0, 1, 5, 9),
    ("Notifications", 0, 0, 0, 0, 0, 0, 0, 5, 0, 0, 5),
    ("familyMember flag", 0, 0, 0, 0, 0, 0, 0, 0, 6, 0, 6),
    ("Permission bugs", 0, 0, 1, 0, 3, 0, 1, 0, 0, 0, 5),
]

RISK_ASSESSMENT = [
    ("Dual Status Model", "Status coupling creates unexpected state: setting accounting=NEW on PAID sick leave REOPENS it. No state machine guardrails.", "High", "High", "Critical",
     "Test all accounting status transitions. Verify State auto-updates correctly. Test reverse transitions (PAID→NEW)."),
    ("Visibility Scope (#3213)", "Production bug: all employees saw all sick leaves. Frontend sends request without employeeLogin on refresh.", "High", "Critical", "Critical",
     "Regression test: page refresh, login/logout cycle, verify employeeLogin param always sent in API requests."),
    ("Norm Propagation", "8 tickets (#2778-#2915) where state changes failed to propagate to My Tasks, Planner, Availability, Confirmation.", "High", "High", "Critical",
     "Cross-page verification after every state change: create, delete, reject, close. Check colors, tooltips, cell availability."),
    ("Rejection Order (#2973)", "create→close→reject works; create→reject→close breaks norm. Order-dependent state coupling bug.", "Medium", "High", "Critical",
     "Test both sequences. Verify norm after each. Also: employee cannot edit in Rejected state, close only for Open/Overdue."),
    ("Permission Gaps (BUG-SL-1)", "Any authenticated user can create sick leave for ANY employee and delete ANY non-PAID sick leave. No owner check.", "High", "High", "Critical",
     "Test cross-user creation and deletion. Document as known security issue."),
    ("Accounting Route Access (#3012)", "Route was accessible to regular employees. Fixed with SICK_LEAVE_ACCOUNTING_VIEW permission.", "Medium", "High", "High",
     "Verify route access per role. Test direct URL navigation."),
    ("familyMember Norm Impact (#3408)", "familyMember=true should NOT deduct from individual norm but should add to budgetNorm. New feature, untested.", "Medium", "High", "High",
     "Test norm before/after creation for both flag values. Verify budgetNorm formula."),
    ("PM Access Check", "PM from another project is blocked by checkIfCurrentEmployeeIsNotPM() but can still view via My Projects tab.", "Medium", "Medium", "Medium",
     "Verify PM cannot edit/delete. Verify PM view is read-only. Test PM who is also the employee's manager (exception)."),
    ("File Handling Debt", "Max 5 files validated on DTO only. Orphaned file associations on soft-delete. Two-step patch for files.", "Low", "Medium", "Medium",
     "Test 5-file limit. Verify file associations after delete. Test edit file diff-and-sync."),
    ("Notification Editor Type (BUG-SL-2)", "getEditorType() uses == (identity comparison), always falls to default. Wrong template selected.", "High", "Low", "Medium",
     "Verify notification templates match the editor type (employee vs supervisor vs accountant)."),
    ("Unrestricted Accounting Transitions (BUG-SL-3)", "Any-to-any transitions allowed. PAID→NEW reopens closed sick leave.", "Medium", "Medium", "Medium",
     "Verify all transition paths. Document which transitions should be blocked but aren't."),
]


# ─── Workbook Generation ─────────────────────────────────────────────────────

def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 0 else FILL_ROW_WHITE


def write_header_row(ws, headers, col_widths):
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_test_cases(ws, cases, back_link_tab="Plan Overview"):
    """Write test cases to a worksheet."""
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [14, 40, 45, 65, 45, 10, 10, 35, 22, 35]

    # Back-link row
    ws.merge_cells("A1:J1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = f"← Back to {back_link_tab}"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = f"#'{back_link_tab}'!A1"

    # Headers in row 2
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # Data rows
    for row_idx, tc in enumerate(cases, 3):
        values = [tc["id"], tc["title"], tc["preconditions"], tc["steps"],
                  tc["expected"], tc["priority"], tc["type"],
                  tc["req_ref"], tc["module"], tc["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):  # Priority, Type
                cell.alignment = ALIGN_CENTER

    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    """Write Plan Overview tab."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"• {env}").font = FONT_BODY
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"• {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws):
    """Write Feature Matrix tab."""
    headers = ["Feature", "CRUD", "Lifecycle", "Accounting", "Manager", "Perms",
               "Validation", "Regression", "Notif", "Family", "CrossPage", "Total"]
    col_widths = [30, 8, 10, 12, 10, 8, 12, 12, 8, 8, 12, 8]

    suite_tabs = ["TS-SickLeave-CRUD", "TS-SL-Lifecycle", "TS-SL-Accounting",
                  "TS-SL-Manager", "TS-SL-Permissions", "TS-SL-Validation",
                  "TS-SL-Regression", "TS-SL-Notifications", "TS-SL-FamilyMember",
                  "TS-SL-CrossPage"]

    # Back-link
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    # Headers in row 2
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    """Write Risk Assessment tab."""
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def generate():
    """Generate the sick leave test documentation workbook."""
    wb = Workbook()

    suites = [
        ("TS-SickLeave-CRUD", "TS-SickLeave-CRUD", get_crud_cases),
        ("TS-SL-Lifecycle", "TS-SL-Lifecycle", get_lifecycle_cases),
        ("TS-SL-Accounting", "TS-SL-Accounting", get_accounting_cases),
        ("TS-SL-Manager", "TS-SL-Manager", get_manager_cases),
        ("TS-SL-Permissions", "TS-SL-Permissions", get_permission_cases),
        ("TS-SL-Validation", "TS-SL-Validation", get_validation_cases),
        ("TS-SL-Regression", "TS-SL-Regression", get_regression_cases),
        ("TS-SL-Notifications", "TS-SL-Notifications", get_notification_cases),
        ("TS-SL-FamilyMember", "TS-SL-FamilyMember", get_family_member_cases),
        ("TS-SL-CrossPage", "TS-SL-CrossPage", get_cross_page_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    total_cases = sum(info[1] for info in suites_info)
    print(f"Generating {total_cases} test cases across {len(suites)} suites")

    # Plan Overview (default first sheet)
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(ws_plan, suites_info)

    # Feature Matrix
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_feature_matrix(ws_fm)

    # Risk Assessment
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ws_ra)

    # Test Suite tabs
    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_test_cases(ws, cases)
        print(f"  {tab_name}: {len(cases)} cases")

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {total_cases} test cases, {len(suites)} suites, 3 plan tabs")

    return all_cases


if __name__ == "__main__":
    generate()
