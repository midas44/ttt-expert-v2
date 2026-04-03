#!/usr/bin/env python3
"""
Security Module Test Documentation Generator — Phase B
Generates test-docs/security/security.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 8 TS- test suite tabs.

Based on vault knowledge: security-patterns.md (3 auth mechanisms, filter chain, 6 findings,
ticket-derived scenarios), auth-authorization-doc.md (JWT vs API token, @PreAuthorize patterns),
role-permission-matrix.md (11 roles × permission classes, frontend page access matrix, security gaps),
security-ticket-findings.md (~85 tickets in 6 categories: API bypass, cross-office leakage,
error codes, role sync, JWT issues, permission architecture).

8 Suites, ~84 test cases:
  TS-Security-AuthModel    — JWT lifecycle, API token auth, CAS login, deactivated user
  TS-Security-RoleMatrix   — Role-based page access for 11 roles × key pages
  TS-Security-APIBypass    — API endpoints accessible without proper role
  TS-Security-CrossOffice  — Office-scoped data isolation
  TS-Security-TokenPerms   — API token permission model and CRUD
  TS-Security-ErrorCodes   — Authorization error response consistency
  TS-Security-RoleSync     — CS role propagation and stale JWT
  TS-Security-Regression   — Bug regression tests from tickets
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Constants ---------------------------------------------------------------

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "security")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "security.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"

# --- Plan Overview -----------------------------------------------------------

PLAN_OVERVIEW = {
    "title": "Security Module — Test Plan",
    "scope": (
        "Authentication and authorization across all TTT services: JWT token lifecycle, "
        "API token permission model, CAS SSO integration, role-based access control "
        "(11 global roles × 30+ pages), API authorization bypass detection, cross-office "
        "data isolation, error code consistency for auth failures, CompanyStaff role "
        "synchronization, and regression tests for 85+ security-related GitLab tickets."
    ),
    "objectives": [
        "Verify JWT authentication lifecycle: token acquisition, expiry, reauth, WebSocket behavior on expiry",
        "Validate API token permission model: 21 ApiPermission values, AUTHENTICATED_USER boundary, token CRUD",
        "Test role-based page access for all 11 roles against the permission matrix (frontend routes + backend @PreAuthorize)",
        "Detect API authorization bypass: endpoints where UI hides features but API remains accessible",
        "Verify cross-office data isolation for office-scoped roles (OFFICE_DIRECTOR, ACC, CACC)",
        "Validate authorization error responses: 403 vs 500/400 masking across services",
        "Test CompanyStaff role synchronization: role assignment, removal, stale JWT detection",
        "Regression coverage for 15+ API bypass bugs, 6 cross-office leakage tickets, 9 JWT/auth issues",
    ],
    "environments": [
        "QA-1: https://ttt-qa-1.noveogroup.com (primary — current sprint build)",
        "Timemachine: https://ttt-timemachine.noveogroup.com (clock manipulation for JWT expiry)",
        "Stage: https://ttt-stage.noveogroup.com (production baseline comparison)",
    ],
    "approach": (
        "Mixed UI + API testing. Role-based page access tested via UI (login as each role, verify "
        "navigation menu items and page accessibility). API bypass tests use direct HTTP requests "
        "with JWT tokens captured from browser sessions. Token permission tests use API_SECRET_TOKEN "
        "header. Cross-office tests require employees from different offices. Error code tests use "
        "direct API calls and verify HTTP status codes. Role sync tests trigger CS sync endpoint "
        "and verify role changes."
    ),
    "dependencies": [
        "Multiple test accounts with different roles: ADMIN, EMPLOYEE, PM, DM, ACC, CACC, DIR, HR, TL, CONTRACTOR, VIEW_ALL",
        "Employees from at least 2 different offices for cross-office isolation tests",
        "API tokens with varying permission sets (full, partial, none)",
        "CompanyStaff sync endpoint access (test API)",
        "CAS SSO availability for login tests",
    ],
}

# --- Risk Assessment ---------------------------------------------------------

RISK_ASSESSMENT = [
    ("API authorization bypass (systemic)",
     "UI hides features but API endpoints lack server-side authorization — any authenticated user can call them (#117, #1250, #1292, #736, #3002)",
     "High", "Critical", "Critical",
     "Test EVERY endpoint with intended role + unauthorized role + API token. Verify 403 for unauthorized access."),
    ("Cross-office data leakage",
     "Office-scoped roles see data from other offices: OFFICE_DIRECTOR sees all offices (#2050), accountant modifies other office days (#480)",
     "Medium", "Critical", "Critical",
     "Test office-filtered endpoints with users from different offices. Verify data isolation."),
    ("JWT expiry cascading failures",
     "Expired JWT causes WebSocket reconnection loops (#2270), deactivated user triggers infinite 500 loop (#1275)",
     "Medium", "High", "High",
     "Test JWT expiry behavior: verify graceful degradation, no infinite loops, proper redirect to login."),
    ("Accounting pages unprotected",
     "4 of 5 accounting subroutes (/salary, /vacation-payout, /days-correction, /periods) lack frontend permission checks",
     "High", "High", "High",
     "Navigate to accounting pages as employee/contractor role. Verify backend API still requires proper authorization."),
    ("Error code masking",
     "Authorization failures return 500 or 400 instead of 403 (#1286, #1883, #2164) — hard to distinguish from server errors",
     "Medium", "Medium", "High",
     "Test all 403-expected scenarios. Verify correct HTTP status and error response body."),
    ("Sick leave AUTHENTICATED_USER gap",
     "Sick leave CRUD requires AUTHENTICATED_USER only — blocks API token access entirely. No fallback permission.",
     "High", "Medium", "High",
     "Verify API token gets 403 on sick leave CRUD. Test that UI/JWT access works for all authorized roles."),
    ("Role sync fragility",
     "Roles propagated from CS via RabbitMQ → sync → employee_global_roles → JWT. Any link failure = stale roles (#807, #522)",
     "Medium", "High", "High",
     "After CS sync, verify roles match expected. Test with stale JWT (roles changed since token generation)."),
    ("Self-approval vulnerability",
     "DEPARTMENT_MANAGER can create and approve own vacation — no segregation of duties check (design issue #44)",
     "Medium", "Medium", "Medium",
     "Test DM creating vacation then approving it. Verify whether this is intended or blocked."),
    ("Token via URL query param",
     "API tokens accepted via ?token= query parameter — exposed in logs, browser history, referer headers (design issue #46)",
     "Medium", "Medium", "Medium",
     "Test API call with token in URL. Verify it works (known behavior). Note as security risk."),
    ("Former HR retains role",
     "OfficeHRRolePostProcessor only adds ROLE_OFFICE_HR, never removes — former HR keeps elevated permissions (design issue #80)",
     "Medium", "Medium", "Medium",
     "Verify HR role assignment after CS sync. Check if role is removed when employee is no longer HR."),
]

# --- Feature Matrix ----------------------------------------------------------
# [feature, AuthModel, RoleMatrix, APIBypass, CrossOffice, TokenPerms, ErrorCodes, RoleSync, Regression, Total]

FEATURE_MATRIX = [
    ["JWT lifecycle (login, expiry, reauth)",   4, 0, 0, 0, 0, 0, 0, 0, 4],
    ["API token authentication",                2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["CAS/deactivated user auth",               2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["Auth check endpoint",                     2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["Page access (admin pages)",               0, 5, 0, 0, 0, 0, 0, 0, 5],
    ["Page access (role visibility)",           0, 5, 0, 0, 0, 0, 0, 0, 5],
    ["Page access (accounting gap)",            0, 3, 0, 0, 0, 0, 0, 0, 3],
    ["Page access (direct URL bypass)",         0, 2, 0, 0, 0, 0, 0, 0, 2],
    ["Vacation API bypass",                     0, 0, 4, 0, 0, 0, 0, 0, 4],
    ["Reports/tasks API bypass",                0, 0, 3, 0, 0, 0, 0, 0, 3],
    ["Sick leave API bypass",                   0, 0, 2, 0, 0, 0, 0, 0, 2],
    ["Project admin API bypass",                0, 0, 3, 0, 0, 0, 0, 0, 3],
    ["Export/autocomplete bypass",              0, 0, 2, 0, 0, 0, 0, 0, 2],
    ["Office-scoped employee data",             0, 0, 0, 3, 0, 0, 0, 0, 3],
    ["Office-scoped accounting data",           0, 0, 0, 3, 0, 0, 0, 0, 3],
    ["Office-scoped vacation data",             0, 0, 0, 2, 0, 0, 0, 0, 2],
    ["Token CRUD (admin page)",                 0, 0, 0, 0, 4, 0, 0, 0, 4],
    ["Token permission granularity",            0, 0, 0, 0, 4, 0, 0, 0, 4],
    ["Token auth boundary (AUTHENTICATED_USER)",0, 0, 0, 0, 4, 0, 0, 0, 4],
    ["403 vs 500/400 error codes",              0, 0, 0, 0, 0, 4, 0, 0, 4],
    ["Stack trace / error body leakage",        0, 0, 0, 0, 0, 3, 0, 0, 3],
    ["Role assignment via CS sync",             0, 0, 0, 0, 0, 0, 3, 0, 3],
    ["Stale JWT / role removal",                0, 0, 0, 0, 0, 0, 3, 0, 3],
    ["Ticket regression (auth bugs)",           0, 0, 0, 0, 0, 0, 0, 12, 12],
]

# --- Test Suites -------------------------------------------------------------

SUITES = {
    "TS-Security-AuthModel": [
        # ID, Title, Preconditions, Steps, Expected, Priority, Type, Req Ref, Module, Notes
        ("TC-SEC-001", "JWT token acquired after CAS login",
         "Enabled employee account with valid CAS credentials. Query: SELECT login FROM employee WHERE enabled = true AND deactivated = false ORDER BY random() LIMIT 1",
         "1. Open TTT login page (https://ttt-qa-1.noveogroup.com)\n"
         "2. CAS redirects to login form — enter valid credentials\n"
         "3. After successful login, verify My Tasks page loads\n"
         "4. Open browser DevTools > Network tab\n"
         "5. Trigger any page navigation (e.g., click Planner)\n"
         "6. Inspect outgoing API request headers\n"
         "7. Verify TTT_JWT_TOKEN header is present in API calls",
         "My Tasks page loads after login. All subsequent API requests include TTT_JWT_TOKEN header with a valid JWT containing claims: login, officeId, managerId, authorities.",
         "Critical", "UI", "security-patterns.md §JWT Token", "security",
         "JWT is RS256-signed, 1-day validity. Claims include csId, login, officeId, managerId, employeeName, authorities, readOnly."),

        ("TC-SEC-002", "JWT token expires after 1 day — forces re-authentication",
         "Logged-in user with active JWT. Test clock endpoint available. SETUP: Via API — advance clock by 25 hours: PATCH /api/ttt/test/v1/clock",
         "SETUP: Via test API — note current time, advance clock by 25 hours (PATCH /api/ttt/test/v1/clock)\n"
         "1. Login as employee\n"
         "2. Navigate to My Tasks page, verify it loads\n"
         "3. Wait for clock advancement to take effect\n"
         "4. Attempt to navigate to another page (e.g., Planner)\n"
         "5. Verify the application detects expired JWT\n"
         "6. Verify user is redirected to CAS login page or shown re-auth prompt\n"
         "CLEANUP: Via test API — reset clock to current time",
         "After JWT expiry, the application should detect the expired token and redirect to CAS login. No infinite loops or 500 errors.",
         "Critical", "Hybrid", "security-patterns.md §JWT Token, #2270", "security",
         "JWT has no refresh mechanism (design issue). 1-day hardcoded expiry in JwtTokenProcessor.generateJWTToken. Use timemachine env for clock manipulation."),

        ("TC-SEC-003", "WebSocket reconnection on JWT expiry — no infinite loop",
         "Logged-in user on My Tasks page with WebSocket connection. SETUP: Via API — advance clock to trigger JWT expiry.",
         "SETUP: Via test API — advance clock by 25 hours\n"
         "1. Login as employee\n"
         "2. Navigate to My Tasks page\n"
         "3. Open browser DevTools > Console\n"
         "4. Wait for JWT expiry to trigger WebSocket disconnect\n"
         "5. Monitor console for reconnection attempts\n"
         "6. Verify the app polls /v1/authentication/check every 5s instead of looping WebSocket reconnection\n"
         "7. Verify app eventually shows login prompt or redirects\n"
         "CLEANUP: Via test API — reset clock",
         "After JWT expiry, WebSocket disconnects. App should fall back to polling /v1/authentication/check every 5 seconds. No infinite WebSocket reconnection loop. Eventually shows re-login prompt.",
         "High", "Hybrid", "#2270", "security",
         "Fix for #2270 added polling fallback. Before fix, expired JWT caused infinite WebSocket reconnection storms."),

        ("TC-SEC-004", "Deactivated user login returns error — no infinite 500 loop",
         "Deactivated employee account. Query: SELECT login FROM employee WHERE enabled = true AND deactivated = true LIMIT 1. If none, use test API to deactivate a test user.",
         "1. Navigate to TTT login page\n"
         "2. Enter credentials for a deactivated user\n"
         "3. Verify the login result\n"
         "4. Check that no infinite redirect/500 loop occurs\n"
         "5. Verify a clear error message is shown (expected: 412 Precondition Failed or similar)",
         "Login attempt for deactivated user should return a clear error (412 or 403). No infinite 500-loop requesting JWT. User sees a human-readable error message.",
         "Critical", "UI", "#1275", "security",
         "Ticket #1275: deactivated user login triggered infinite 500 loop requesting JWT. Proposed fix: return 412 status."),

        ("TC-SEC-005", "API token authentication via API_SECRET_TOKEN header",
         "Valid API token with EMPLOYEES_VIEW permission. Query: SELECT t.secret_key FROM token t JOIN token_permissions tp ON t.id = tp.token_id WHERE tp.apipermission = 'EMPLOYEES_VIEW' LIMIT 1",
         "1. Via API — send GET /api/ttt/v1/employees with API_SECRET_TOKEN header set to a valid token\n"
         "2. Verify response status 200\n"
         "3. Verify employee data is returned\n"
         "4. Via API — send same request WITHOUT the header\n"
         "5. Verify response status 401 or 403",
         "API call with valid API_SECRET_TOKEN header returns 200 with data. Call without token returns 401/403.",
         "Critical", "Hybrid", "auth-authorization-doc.md §API Token", "security",
         "API tokens resolved via DatabaseApiTokenResolver. Two types: EMPLOYEE (personal) and APPLICATION (system)."),

        ("TC-SEC-006", "API token via URL query param — works but security risk",
         "Valid API token. Same as TC-SEC-005.",
         "1. Via API — send GET /api/ttt/v1/employees?token=<valid_token_value>\n"
         "2. Verify response status 200\n"
         "3. Verify employee data is returned\n"
         "DB-CHECK: Verify the token value appears in application server access logs (security risk confirmation)",
         "API call with token in URL query parameter succeeds (200). This is a known security concern — token exposed in logs, browser history, HTTP referer headers.",
         "High", "Hybrid", "security-patterns.md §Security Findings, design issue #46", "security",
         "ApiTokenAuthenticationFilter accepts both header and query param. Token in URL is a known security risk but maintained for backward compatibility."),

        ("TC-SEC-007", "API token does NOT grant AUTHENTICATED_USER authority",
         "Valid API token. JWT-only endpoint (e.g., sick leave CRUD).",
         "1. Via API — send GET /api/vacation/v1/sick-leaves/search with API_SECRET_TOKEN header\n"
         "2. Verify response status 403 (Forbidden)\n"
         "3. Via API — send same request with TTT_JWT_TOKEN (captured from browser session)\n"
         "4. Verify response status 200",
         "API token gets 403 on AUTHENTICATED_USER-only endpoints. Same endpoint with JWT returns 200. This confirms the intended separation between JWT (full access) and API token (permission-scoped).",
         "Critical", "Hybrid", "auth-authorization-doc.md §Dual Authentication", "security",
         "Core security design: API tokens carry only explicit ApiPermission values, NOT AUTHENTICATED_USER. Sick leave CRUD is the clearest test case for this boundary."),

        ("TC-SEC-008", "Authentication check endpoint returns correct status",
         "Logged-in user with valid JWT. Also test with expired/invalid JWT.",
         "1. Login as employee\n"
         "2. Via API (with captured JWT) — send GET /api/ttt/v1/authentication/check\n"
         "3. Verify response status 200\n"
         "4. Via API — send same request with invalid JWT (e.g., tampered token)\n"
         "5. Verify response status 401\n"
         "6. Via API — send request with no auth header\n"
         "7. Verify response status 401",
         "Valid JWT returns 200 on /v1/authentication/check. Invalid/missing JWT returns 401. This endpoint is used by the frontend to poll auth status after WebSocket disconnect.",
         "High", "Hybrid", "#2270", "security",
         "Frontend polls this endpoint every 5 seconds when WebSocket is disconnected. Critical for JWT expiry detection."),

        ("TC-SEC-009", "Invalid JWT returns 401 — not 500",
         "No valid session required.",
         "1. Via API — send GET /api/ttt/v1/employees with TTT_JWT_TOKEN header set to 'invalid-token-value'\n"
         "2. Verify response status 401 (not 500)\n"
         "3. Via API — send same request with TTT_JWT_TOKEN set to an expired but well-formed JWT\n"
         "4. Verify response status 401 (not 500)\n"
         "5. Verify response body does not contain stack trace",
         "Invalid JWT returns clean 401 response. No 500 error, no stack trace leakage in response body.",
         "High", "Hybrid", "security-patterns.md", "security",
         "JwtTokenAuthenticationFilter should reject invalid tokens cleanly. Test both malformed and expired-but-valid-format tokens."),

        ("TC-SEC-010", "CORS policy allows all origins",
         "No auth required for OPTIONS preflight.",
         "1. Via API — send OPTIONS request to /api/ttt/v1/employees with Origin: https://evil.example.com\n"
         "2. Inspect Access-Control-Allow-Origin header in response\n"
         "3. Verify it returns * or echoes the origin\n"
         "4. Document whether credentials are allowed (Access-Control-Allow-Credentials)",
         "CORS is configured to allow all origins (known design decision). Document the response headers for security assessment.",
         "Medium", "Hybrid", "security-patterns.md §CORS", "security",
         "CORS enabled for all origins — known design choice. Important to document for security audit."),
    ],

    "TS-Security-RoleMatrix": [
        ("TC-SEC-011", "ADMIN sees all navigation menu items",
         "User with ROLE_ADMIN. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_ADMIN' AND e.enabled = true AND e.deactivated = false ORDER BY random() LIMIT 1",
         "1. Login as ADMIN user\n"
         "2. Verify top navigation bar contains: My tasks, Calendar of absences, Confirmation, Planner, Statistics, Admin panel, Accounting, Notifications\n"
         "3. Click 'Admin panel' dropdown — verify submenu: Projects, Employees and subcontractors, TTT parameters, Production calendars, API, Export\n"
         "4. Click 'Accounting' dropdown — verify submenu items visible\n"
         "5. Click 'Statistics' dropdown — verify all statistics tabs accessible",
         "ADMIN role sees ALL navigation items including Admin panel (with all 6 sub-items), Accounting (all sub-items), Confirmation, Statistics, and Notifications.",
         "Critical", "UI", "role-permission-matrix.md §Frontend Page Access", "security",
         "ADMIN has all permissions. This is the baseline for comparison with restricted roles."),

        ("TC-SEC-012", "Regular EMPLOYEE sees only basic navigation",
         "User with ROLE_EMPLOYEE only (no other roles). Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_EMPLOYEE' AND e.enabled = true AND e.deactivated = false GROUP BY e.id HAVING COUNT(DISTINCT r.role) = 1 ORDER BY random() LIMIT 1",
         "1. Login as employee-only user (no PM/DM/ACC/ADMIN roles)\n"
         "2. Verify navigation bar visible items: My tasks, Calendar of absences, Planner, Statistics, Notifications\n"
         "3. Verify 'Confirmation' link is NOT visible (requires TASKS:VIEW_APPROVES)\n"
         "4. Verify 'Admin panel' dropdown is NOT visible or shows only 'Account'\n"
         "5. Verify 'Accounting' dropdown is NOT visible or shows no sensitive sub-items",
         "Employee-only user sees: My tasks, Calendar of absences, Planner, Statistics, Notifications. Does NOT see: Confirmation (requires approver permission), Admin panel sub-items (except Account), Accounting.",
         "Critical", "UI", "role-permission-matrix.md", "security",
         "ROLE_EMPLOYEE gets VACATIONS:VIEW and basic self-access only. Important: verify zero admin/accounting menu items."),

        ("TC-SEC-013", "ACCOUNTANT sees Accounting items but limited admin access",
         "User with ROLE_ACCOUNTANT. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_ACCOUNTANT' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as ACCOUNTANT user\n"
         "2. Verify 'Accounting' dropdown visible with sub-items\n"
         "3. Verify 'Confirmation' link visible (ACC has TASKS:VIEW_APPROVES)\n"
         "4. Verify 'Admin panel' dropdown — should show: Employees, Salary (Offices)\n"
         "5. Verify Admin panel does NOT show: Projects, TTT parameters, API, Export\n"
         "6. Navigate to /admin/employees — verify access\n"
         "7. Navigate to /admin/settings — verify access denied or redirect",
         "ACCOUNTANT sees: Accounting menu, Confirmation, Employees admin, Offices/Salary admin. Does NOT see: Projects, TTT parameters, API keys, Export.",
         "High", "UI", "role-permission-matrix.md §VIEW Permissions", "security",
         "ACC has: ACCOUNTING:VIEW, EMPLOYEES:VIEW, OFFICES:VIEW, TASKS:VIEW_APPROVES, VACATIONS:VIEW_PAYMENTS, VACATIONS:VIEW_DAYS, SUGGESTIONS:VIEW_CUST."),

        ("TC-SEC-014", "CHIEF_ACCOUNTANT sees Accounting + Production calendars",
         "User with ROLE_CHIEF_ACCOUNTANT. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_CHIEF_ACCOUNTANT' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as CHIEF_ACCOUNTANT user\n"
         "2. Verify 'Accounting' dropdown visible with all sub-items\n"
         "3. Verify 'Admin panel' dropdown includes: Employees, Production calendars, Salary (Offices)\n"
         "4. Navigate to /admin/calendar — verify calendar list loads\n"
         "5. Verify Admin panel does NOT show: TTT parameters, API, Export",
         "CHIEF_ACCOUNTANT sees all Accounting items plus Production calendars in Admin panel. Has CALENDARS:VIEW permission.",
         "High", "UI", "role-permission-matrix.md", "security",
         "CACC extends ACC with CALENDARS:VIEW, VACATIONS:SICK_LEAVE_VIEW. Both ACC and CACC have ACCOUNTING:NOTIFY with readOnly flag."),

        ("TC-SEC-015", "PROJECT_MANAGER sees Confirmation + Admin Projects",
         "User with ROLE_PROJECT_MANAGER. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_PROJECT_MANAGER' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as PROJECT_MANAGER user\n"
         "2. Verify 'Confirmation' link visible in navigation\n"
         "3. Verify 'Admin panel' dropdown includes: Projects, Employees\n"
         "4. Navigate to /admin/projects — verify project list loads\n"
         "5. Navigate to /approve — verify Confirmation page loads\n"
         "6. Verify 'Accounting' dropdown is NOT visible\n"
         "7. Verify Admin panel does NOT show: TTT parameters, API, Export, Calendars",
         "PROJECT_MANAGER sees: Confirmation, Admin Projects, Admin Employees. Does NOT see: Accounting, TTT parameters, API, Calendars, Export.",
         "High", "UI", "role-permission-matrix.md", "security",
         "PM has: PROJECTS:VIEW, PROJECTS:CREATE, EMPLOYEES:VIEW, TASKS:VIEW_APPROVES, SUGGESTIONS:VIEW_CUST, VACATIONS:VIEW_APPROVES."),

        ("TC-SEC-016", "DEPARTMENT_MANAGER sees Vacation Requests + Budget Notifications",
         "User with ROLE_DEPARTMENT_MANAGER. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_DEPARTMENT_MANAGER' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as DEPARTMENT_MANAGER user\n"
         "2. Verify 'Calendar of absences' dropdown includes Vacation Requests tab\n"
         "3. Navigate to /vacation/request — verify Vacation Requests page loads\n"
         "4. Verify 'Budget Notifications' link visible (DM has BUDGET_NOTIF:VIEW)\n"
         "5. Navigate to /vacation/vacation-days — verify Vacation Days page loads (DM has VACATIONS:VIEW_EMPLOYEES)\n"
         "6. Verify Admin panel includes: Projects, Employees",
         "DEPARTMENT_MANAGER sees: Vacation Requests, Vacation Days, Budget Notifications, Confirmation, Admin Projects, Admin Employees. Extended view of Calendar of absences section.",
         "High", "UI", "role-permission-matrix.md", "security",
         "DM extends PM with BUDGET_NOTIF:VIEW/CREATE, VACATIONS:VIEW_EMPLOYEES. DM can create and self-approve vacations (design issue #44)."),

        ("TC-SEC-017", "OFFICE_DIRECTOR sees Budget Notifications only",
         "User with ROLE_OFFICE_DIRECTOR. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_OFFICE_DIRECTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as OFFICE_DIRECTOR user\n"
         "2. Verify 'Budget Notifications' link visible\n"
         "3. Verify Admin panel is minimal (Employees only, based on EMPLOYEES:VIEW)\n"
         "4. Verify no Accounting, Calendars, Settings, API, Export items visible",
         "OFFICE_DIRECTOR sees Budget Notifications and Admin Employees. Minimal admin access. No accounting or calendar admin.",
         "Medium", "UI", "role-permission-matrix.md", "security",
         "DIR has: BUDGET_NOTIF:VIEW/CREATE, EMPLOYEES:VIEW, SUGGESTIONS:VIEW_CUST."),

        ("TC-SEC-018", "CONTRACTOR sees minimal navigation",
         "User with ROLE_CONTRACTOR. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_CONTRACTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as CONTRACTOR user\n"
         "2. Verify navigation shows: My tasks, Planner, possibly Statistics\n"
         "3. Verify no Confirmation, Admin panel, Accounting items\n"
         "4. Verify Calendar of absences access (may show personal only)",
         "CONTRACTOR has minimal navigation. No explicit permissions in any PermissionProvider — gets basic personal access only.",
         "Medium", "UI", "role-permission-matrix.md §Contractor", "security",
         "ROLE_CONTRACTOR gets no explicit permissions in any provider. Likely default personal access. Known documentation gap."),

        ("TC-SEC-019", "VIEW_ALL role accesses all view pages but no mutations",
         "User with ROLE_VIEW_ALL. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_VIEW_ALL' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as VIEW_ALL user\n"
         "2. Verify all view-related menu items visible (Confirmation, Statistics, Admin panel sub-items, Accounting)\n"
         "3. Navigate to /admin/projects — verify view access\n"
         "4. Attempt to create a project — verify no 'Create' button or action fails\n"
         "5. Navigate to /admin/calendar — verify view access\n"
         "6. Attempt to edit a calendar — verify no edit controls or action fails",
         "VIEW_ALL sees all view pages. Cannot perform mutations (create, edit, delete). View-only access confirmed.",
         "High", "UI", "role-permission-matrix.md", "security",
         "VIEW_ALL gets all VIEW permissions but no CREATE/EDIT/DELETE. Has STATISTICS:EXPORT — may have Export access (see #2102 bug)."),

        ("TC-SEC-020", "Employee cannot access /admin/settings via direct URL",
         "User with ROLE_EMPLOYEE only (same as TC-SEC-012).",
         "1. Login as employee-only user\n"
         "2. Navigate directly to /admin/settings by entering URL\n"
         "3. Verify page shows access denied or redirects to default page\n"
         "4. Navigate directly to /admin/api — verify access denied\n"
         "5. Navigate directly to /admin/calendar — verify access denied\n"
         "6. Navigate directly to /admin/export — verify access denied",
         "Employee cannot access restricted admin pages via direct URL. Expected: redirect to default page or access denied message. Frontend PrivateRoute guards should block access.",
         "Critical", "UI", "role-permission-matrix.md §Frontend Page Access", "security",
         "Frontend uses PrivateRoute guards with permission checks. Verify guards work for direct URL navigation, not just menu hiding."),

        ("TC-SEC-021", "Accounting pages accessible to any authenticated user (security gap)",
         "User with ROLE_EMPLOYEE only.",
         "1. Login as employee-only user\n"
         "2. Navigate directly to /accounting/salary — verify page loads or redirects\n"
         "3. Navigate directly to /accounting/vacation-payout — verify page loads or redirects\n"
         "4. Navigate directly to /accounting/days-correction — verify page loads or redirects\n"
         "5. Navigate directly to /accounting/periods — verify page loads or redirects\n"
         "6. If pages load, attempt to interact with accounting controls — verify backend API returns 403",
         "Known security gap: 4 of 5 accounting subroutes lack frontend permission checks (None! in route config). Pages may render for any user. Backend API should still enforce authorization.",
         "Critical", "UI", "role-permission-matrix.md §Security Gaps", "security",
         "Frontend route config shows no permission check for /accounting/salary, /vacation-payout, /days-correction, /periods. Only /accounting/sick-leaves has VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW check."),

        ("TC-SEC-022", "Sick leave route accessible to any user (planned TODO)",
         "User with ROLE_EMPLOYEE only.",
         "1. Login as employee-only user\n"
         "2. Navigate directly to /sick-leave/my\n"
         "3. Verify page loads (expected: accessible due to TODO in code)\n"
         "4. If page loads, verify what data is shown — should be own sick leaves only",
         "Known planned gap: /sick-leave route has TODO in code for permission check. Page loads for any authenticated user. Backend should filter to own data only.",
         "High", "UI", "role-permission-matrix.md §Security Gaps", "security",
         "SickLeaveRoute/index.js has explicit Russian TODO comment about adding permission check. Component created for future sprints."),

        ("TC-SEC-023", "TECH_LEAD sees Confirmation page",
         "User with ROLE_TECH_LEAD. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_TECH_LEAD' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as TECH_LEAD user\n"
         "2. Verify 'Confirmation' link visible in navigation\n"
         "3. Navigate to /approve — verify Confirmation page loads\n"
         "4. Verify Vacation Requests accessible (TL has VACATIONS:VIEW_APPROVES)\n"
         "5. Verify Admin panel access limited (Employees only via EMPLOYEES:VIEW)",
         "TECH_LEAD sees Confirmation and Vacation Requests. Has EMPLOYEES:VIEW for admin access. Permissions not fully documented (#1946).",
         "Medium", "UI", "role-permission-matrix.md, #1946", "security",
         "TL has: EMPLOYEES:VIEW, TASKS:VIEW_APPROVES, VACATIONS:VIEW_APPROVES. Ticket #1946 notes TL/SPM/OWNER permissions not fully documented."),

        ("TC-SEC-024", "HR role sees Admin Employees and Employee Reports",
         "User with ROLE_OFFICE_HR. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_OFFICE_HR' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as OFFICE_HR user\n"
         "2. Verify Admin panel dropdown includes 'Employees and subcontractors'\n"
         "3. Navigate to /admin/employees — verify it loads with HR-scope data\n"
         "4. Navigate to /employee-reports/<some_login> — verify Employee Reports page loads\n"
         "5. Verify no Accounting, Calendars, Settings, Projects items visible",
         "OFFICE_HR sees Admin Employees page (scoped to own office) and can access Employee Reports. No accounting, calendar, or settings access.",
         "Medium", "UI", "role-permission-matrix.md", "security",
         "HR has: EMPLOYEES:VIEW, SUGGESTIONS:VIEW_CUST. Office-scoped — should see only own office employees."),

        ("TC-SEC-025", "Self-approval: DM creates and approves own vacation",
         "User with ROLE_DEPARTMENT_MANAGER who is also an employee with sufficient vacation days. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_DEPARTMENT_MANAGER' AND e.enabled = true AND e.deactivated = false ORDER BY random() LIMIT 1",
         "1. Login as DEPARTMENT_MANAGER user\n"
         "2. Navigate to My Vacations page (/vacation/my)\n"
         "3. Click 'Create a request' button\n"
         "4. Fill in vacation details: type Regular, dates 1 week ahead\n"
         "5. Submit the vacation request\n"
         "6. Navigate to Vacation Requests (/vacation/request) > Approver tab\n"
         "7. Find the newly created request\n"
         "8. Click Approve on own request\n"
         "9. Verify whether the approval succeeds or is blocked\n"
         "CLEANUP: Via API — delete the created vacation",
         "Document whether DM can approve their own vacation. This is a known segregation of duties concern (design issue #44). Expected: either approval succeeds (documenting the gap) or system prevents self-approval.",
         "High", "UI", "design issue #44", "security",
         "No check in VacationController.createVacation for self-approval by DEPARTMENT_MANAGER. Potential segregation of duties violation."),
    ],

    "TS-Security-APIBypass": [
        ("TC-SEC-026", "Employee redirects own vacation via API (not available in UI) — #1250",
         "Employee with a NEW vacation request. SETUP: Via API — create vacation for employee (POST /api/vacation/v1/vacations). Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_EMPLOYEE' AND e.enabled = true AND e.deactivated = false ORDER BY random() LIMIT 1",
         "SETUP: Via API — create a REGULAR vacation for the employee\n"
         "1. Login as the employee\n"
         "2. Navigate to My Vacations — verify the vacation appears with status NEW\n"
         "3. Verify NO 'Redirect' button is visible in the UI for this vacation\n"
         "4. Via API (with captured JWT) — attempt PUT /api/vacation/v1/vacations/{id} with changed approverId\n"
         "5. Verify whether the API accepts the redirect operation\n"
         "6. If accepted: verify the vacation's approver changed (authorization bypass confirmed)\n"
         "CLEANUP: Via API — delete the vacation",
         "UI correctly hides redirect option for employees. Test whether API endpoint enforces the same restriction. If API allows redirect, this is an authorization bypass (#1250).",
         "Critical", "Hybrid", "#1250", "security",
         "Ticket #1250 (closed): employee could redirect own vacation via API even though UI hides this feature. Verify if fix is in place."),

        ("TC-SEC-027", "Manager cancels vacation via API (UI shows only approve/reject) — #1292",
         "Manager with a subordinate's vacation in APPROVED status. SETUP: Via API — create and approve vacation.",
         "SETUP: Via API — create vacation for subordinate, approve it as manager\n"
         "1. Login as the manager\n"
         "2. Navigate to Vacation Requests > Approver tab\n"
         "3. Find the APPROVED vacation — verify NO cancel button in UI\n"
         "4. Via API (with manager's JWT) — attempt PUT /api/vacation/v1/vacations/cancel/{id}\n"
         "5. Verify whether the API accepts the cancel operation\n"
         "6. If accepted: verify vacation status changed to CANCELED (bypass confirmed)\n"
         "CLEANUP: Via API — delete the vacation",
         "UI correctly shows only approve/reject for managers. API should also block cancel operation for managers. If API allows cancel, this is an authorization bypass (#1292).",
         "Critical", "Hybrid", "#1292", "security",
         "Ticket #1292 (closed): manager could cancel vacation via API. Only employee and accountant should be able to cancel."),

        ("TC-SEC-028", "Any approver can approve any vacation — approver field decorative — #117",
         "Two managers (A and B), vacation created for employee of manager A. SETUP: Via API — create vacation for employee of manager A.",
         "SETUP: Via API — create vacation for employee whose manager is Manager A\n"
         "1. Login as Manager B (NOT the assigned approver)\n"
         "2. Via API (with Manager B's JWT) — attempt PUT /api/vacation/v1/vacations/approve/{id}\n"
         "3. Verify whether the approval succeeds\n"
         "4. If approved: check vacation record — approverId field is decorative, not enforced (#117)\n"
         "CLEANUP: Via API — delete the vacation",
         "The approverId field on vacation requests should be enforced. Any user with approver role should NOT be able to approve requests they are not assigned to. If they can, this is a critical authorization bypass (#117).",
         "Critical", "Hybrid", "#117", "security",
         "Ticket #117 (closed): ANY user with the right role could approve ANY vacation request. The approverId field was decorative, not enforced server-side."),

        ("TC-SEC-029", "Employee reads other employees' reports via API — #736",
         "Two employees A and B with no manager/subordinate relationship. Query: SELECT e.login FROM employee e WHERE e.enabled = true AND e.deactivated = false AND e.id NOT IN (SELECT employee_id FROM employee_global_roles WHERE role != 'ROLE_EMPLOYEE') ORDER BY random() LIMIT 2",
         "1. Login as Employee A\n"
         "2. Via API (with A's JWT) — send GET /api/ttt/v1/reports?executorLogin=<employee_B_login>\n"
         "3. Verify whether Employee A can see Employee B's reports\n"
         "4. If reports returned: authorization bypass confirmed (#736)\n"
         "5. Expected: 403 or empty result for unauthorized access",
         "Regular employee should NOT be able to read other employees' reports via API. GET /v1/reports should filter by current user or require EMPLOYEES:VIEW permission.",
         "High", "Hybrid", "#736", "security",
         "Ticket #736 (closed): employee could see other employees' reports via GET /v1/reports by specifying executorLogin parameter."),

        ("TC-SEC-030", "PM edits sick leave records via API — accountant-only — #3002",
         "User with ROLE_PROJECT_MANAGER and an existing sick leave record. Query: SELECT sl.id, e.login FROM sick_leave sl JOIN employee e ON sl.employee_id = e.id WHERE sl.deleted = false LIMIT 1",
         "1. Login as PROJECT_MANAGER\n"
         "2. Navigate to sick leave page — verify PM has no edit controls for other employees' sick leaves\n"
         "3. Via API (with PM's JWT) — attempt PATCH /api/vacation/v1/sick-leaves/{id} to update sick leave\n"
         "4. Verify whether the API accepts the edit\n"
         "5. Expected: 403 (only accountant/admin should edit)\n"
         "6. If accepted: authorization bypass confirmed (#3002)",
         "PROJECT_MANAGER should NOT be able to edit sick leave records. Sick leave editing is accountant-only. If API accepts PM's edit, this is an authorization bypass.",
         "High", "Hybrid", "#3002", "security",
         "Ticket #3002 (closed): PM could edit sick leave records via API. Sick leave CRUD uses AUTHENTICATED_USER which PM has."),

        ("TC-SEC-031", "Regular employee accesses /accounting/sick-leaves — #3012",
         "User with ROLE_EMPLOYEE only.",
         "1. Login as employee-only user\n"
         "2. Navigate directly to /accounting/sick-leaves\n"
         "3. Verify page behavior — should be blocked (requires VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW)\n"
         "4. If page loads with data: security gap (#3012)",
         "Regular employee should NOT access /accounting/sick-leaves. This route has VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW check — unlike other accounting routes. Verify it works.",
         "High", "UI", "#3012", "security",
         "Ticket #3012: regular employees accessed /accounting/sick-leaves route. This specific route DOES have a permission check unlike other accounting routes."),

        ("TC-SEC-032", "Cross-project task approval leakage via checkbox — #870",
         "Manager with 'Show tasks from other projects' checkbox. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_PROJECT_MANAGER' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as PROJECT_MANAGER\n"
         "2. Navigate to Confirmation page (/approve)\n"
         "3. Enable 'Show tasks from other projects' checkbox if available\n"
         "4. Verify whether tasks from projects the PM doesn't manage are shown\n"
         "5. Attempt to approve a task from a foreign project\n"
         "6. Verify whether approval succeeds or is blocked",
         "The 'Show tasks from other projects' checkbox should display foreign project tasks in read-only mode. PM should NOT be able to approve tasks from projects they don't manage (#870).",
         "High", "UI", "#870", "security",
         "Ticket #870 (closed): checkbox leaked approval capability to foreign projects. Verify fix enforces read-only view for non-managed projects."),

        ("TC-SEC-033", "Project owner edits beyond permission — manager/SPM fields — #2136",
         "Project with owner who is not admin. Query: SELECT p.name, pe.login FROM project p JOIN project_employee pe ON p.id = pe.project_id WHERE pe.role = 'OWNER' AND pe.employee_id IN (SELECT id FROM employee WHERE enabled = true) LIMIT 1",
         "1. Login as project owner (not admin, not DM)\n"
         "2. Navigate to /admin/projects — find the owned project\n"
         "3. Attempt to edit manager and/or senior manager fields\n"
         "4. Verify whether these fields are editable (they shouldn't be for owners)\n"
         "5. Via API — attempt PATCH /api/ttt/v1/projects/{id} with manager/SPM changes\n"
         "6. Verify API enforces field-level permissions",
         "Project owner should only edit fields within their permission scope. Manager/SPM fields should be restricted to admin/DM roles. If owner can edit these, authorization bypass (#2136).",
         "Medium", "Hybrid", "#2136", "security",
         "Ticket #2136: project owner could edit fields beyond their permission level (manager, senior project manager fields)."),

        ("TC-SEC-034", "Assign ROLE_EMPLOYEE as PM/SPM via API — #2127",
         "Admin or DM user. An employee with only ROLE_EMPLOYEE.",
         "1. Login as admin\n"
         "2. Via API — attempt PATCH /api/ttt/v1/projects/{id} with managerId set to a ROLE_EMPLOYEE user\n"
         "3. Verify whether the API accepts an employee without PM role as project manager\n"
         "4. If accepted: verify what happens in the project — role mismatch (#2127)\n"
         "5. Expected: validation error rejecting non-PM user as project manager",
         "System should validate that users assigned as PM/SPM have the appropriate role. Assigning ROLE_EMPLOYEE as project manager should be rejected.",
         "Medium", "Hybrid", "#2127, #2046", "security",
         "Tickets #2127/#2046: could assign ROLE_EMPLOYEE as PM or senior manager via PATCH/POST /projects. No role validation on assignment."),

        ("TC-SEC-035", "VIEW_ALL accesses Export — gets corrupt file — #2102",
         "User with ROLE_VIEW_ALL. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_VIEW_ALL' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as VIEW_ALL user\n"
         "2. Verify Admin panel > Export menu item visibility\n"
         "3. Navigate to /admin/export\n"
         "4. If page loads: attempt to export data\n"
         "5. Verify export result — should either be blocked or produce valid file\n"
         "6. If corrupt file is downloaded: confirm bug #2102",
         "VIEW_ALL has STATISTICS:EXPORT permission but Export page may produce corrupt results. Either block access or fix export for this role.",
         "Medium", "UI", "#2102, #2103", "security",
         "Ticket #2102: VIEW_ALL could access Export and got corrupt file. Ticket #2103: manager accessed Export via direct URL bypass."),

        ("TC-SEC-036", "Task autocomplete shows tasks from inaccessible projects — #1430",
         "Employee with access to limited projects only.",
         "1. Login as regular employee\n"
         "2. Navigate to My Tasks page\n"
         "3. Click to add a new task — type in the autocomplete field\n"
         "4. Observe autocomplete suggestions\n"
         "5. Verify suggestions only show tasks from projects the employee is assigned to\n"
         "6. If tasks from unassigned projects appear: data leakage (#1430)",
         "Task autocomplete should only show tasks from projects the current user has access to. Showing tasks from inaccessible projects is an information disclosure.",
         "Medium", "UI", "#1430", "security",
         "Ticket #1430 (closed): autocomplete showed tasks from projects the user had no access to. Verify fix filters by project membership."),

        ("TC-SEC-037", "Cross-employee report creation: API token owner reports for others",
         "API token with REPORTS_EDIT permission. Two different employees.",
         "1. Via API — send POST /api/ttt/v1/reports with API_SECRET_TOKEN header\n"
         "2. Set executorLogin to an employee different from the token owner\n"
         "3. Verify whether the API creates a report attributed to the other employee\n"
         "4. If created: the token owner can create reports on behalf of any employee (design issue #52)\n"
         "DB-CHECK: Verify report.executor_login matches the specified login, not the token owner",
         "API token with REPORTS_EDIT can create reports for any employee by specifying executorLogin. This is a known design issue — the reporter differs from the executor.",
         "High", "Hybrid", "design issue #52", "security",
         "Design issue: executorLogin accepts any valid login in report creation. Cross-employee reporting enabled by design but may be unintended."),
    ],

    "TS-Security-CrossOffice": [
        ("TC-SEC-038", "OFFICE_DIRECTOR sees only own office employees — #2050",
         "User with ROLE_OFFICE_DIRECTOR. At least 2 offices with employees. Query: SELECT e.login, o.name as office FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id JOIN office o ON e.office_id = o.id WHERE r.role = 'ROLE_OFFICE_DIRECTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as OFFICE_DIRECTOR user\n"
         "2. Navigate to Admin > Employees and subcontractors\n"
         "3. Verify the employee list shows ONLY employees from the director's office\n"
         "4. Note the total count displayed\n"
         "DB-CHECK: SELECT COUNT(*) FROM employee WHERE office_id = <director_office_id> AND enabled = true — should match displayed count\n"
         "5. Verify no employees from other offices appear in the list\n"
         "6. Use search to look for a known employee from another office — should not appear",
         "OFFICE_DIRECTOR should see only employees from their own office. No cross-office data leakage. Count should match DB query for the director's office.",
         "Critical", "UI", "#2050", "security",
         "Ticket #2050 (closed): OFFICE_DIRECTOR saw employees from ALL offices instead of own only. Verify fix enforces office scope."),

        ("TC-SEC-039", "Office accountant views vacation days for own offices only — #480",
         "User with ROLE_ACCOUNTANT assigned to specific office(s). Query: SELECT e.login, o.name FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id JOIN office o ON e.office_id = o.id WHERE r.role = 'ROLE_ACCOUNTANT' AND e.enabled = true ORDER BY random() LIMIT 1",
         "1. Login as ACCOUNTANT user\n"
         "2. Navigate to Vacation > Days Correction (/vacation/days-correction)\n"
         "3. Verify the employee list shows only employees from accountant's assigned office(s)\n"
         "4. Attempt to search for an employee from a different office\n"
         "5. Verify the employee from other office is NOT shown or modification is blocked\n"
         "6. If accountant can modify vacation days for other-office employees: cross-office leak (#480)",
         "Accountant should only view/modify vacation days for employees in their assigned office(s). Cross-office modification must be blocked.",
         "Critical", "Hybrid", "#480", "security",
         "Ticket #480 (closed): office accountant could view and modify vacation days for employees in other offices."),

        ("TC-SEC-040", "Office accountant sees report periods for own offices only — #479",
         "User with ROLE_ACCOUNTANT. Multiple offices with different report periods.",
         "1. Login as ACCOUNTANT user\n"
         "2. Navigate to Accounting > Periods (/accounting/periods)\n"
         "3. Verify the office list shows only accountant's assigned offices\n"
         "4. Count displayed offices\n"
         "DB-CHECK: Compare with accountant's office assignment in DB\n"
         "5. Verify no report periods from unassigned offices are visible",
         "Accountant should see report/approve periods only for offices they are assigned to. No cross-office period visibility.",
         "High", "Hybrid", "#479", "security",
         "Ticket #479 (closed): office accountant saw report periods for ALL offices instead of assigned ones."),

        ("TC-SEC-041", "ADMIN and CHIEF_ACCOUNTANT see ALL offices — #959",
         "User with ROLE_ADMIN or ROLE_CHIEF_ACCOUNTANT.",
         "1. Login as ADMIN user\n"
         "2. Navigate to Vacation > Vacation Requests > Approver tab or Payment tab\n"
         "3. Verify vacation data from ALL offices is visible\n"
         "4. Login as CHIEF_ACCOUNTANT user\n"
         "5. Navigate to same pages — verify ALL offices visible\n"
         "DB-CHECK: SELECT COUNT(DISTINCT office_id) FROM employee WHERE enabled = true — should match scope",
         "ADMIN and CHIEF_ACCOUNTANT should see data from ALL offices (not restricted to own office). This was a bug in #959 where they were incorrectly limited to own office.",
         "High", "UI", "#959", "security",
         "Ticket #959 (closed): ADMIN and CHIEF_ACCOUNTANT incorrectly saw only own office vacations. Should see all offices."),

        ("TC-SEC-042", "Regular employee cannot access Employees admin page — #482",
         "User with ROLE_EMPLOYEE only.",
         "1. Login as employee-only user\n"
         "2. Navigate directly to /admin/employees\n"
         "3. Verify access is denied — redirect or error page\n"
         "4. Via API — send GET /api/ttt/v1/employees without EMPLOYEES_VIEW permission token\n"
         "5. Verify API returns 403",
         "Regular employee without EMPLOYEES:VIEW permission should not access the employee admin page or API endpoint.",
         "High", "UI", "#482", "security",
         "Ticket #482 (closed): regular employee could access Employees & Contractors admin page. Verify fix blocks access."),

        ("TC-SEC-043", "Cross-office vacation day modification blocked — #480",
         "Accountant from Office A. Employee from Office B with vacation days.",
         "1. Login as accountant from Office A\n"
         "2. Via API (with JWT) — attempt PATCH vacation days for an employee from Office B\n"
         "3. Verify the API returns 403 or validation error\n"
         "4. If modification succeeds: cross-office data mutation confirmed (#480)",
         "Accountant from Office A must NOT be able to modify vacation days for employees in Office B via API. Server-side office scope validation must be enforced.",
         "Critical", "Hybrid", "#480", "security",
         "Office scope validation happens at service level (OfficePeriodValidator). If bypassed, cross-office mutations are possible."),

        ("TC-SEC-044", "Salary offices filtered by accountant's office assignment",
         "Accountant assigned to specific office(s).",
         "1. Login as ACCOUNTANT user\n"
         "2. Navigate to Admin > Salary Offices (/admin/salary)\n"
         "3. Verify only assigned offices appear in the salary office list\n"
         "4. Navigate to Accounting > Salary (/accounting/salary)\n"
         "5. Verify salary notification targets only assigned offices",
         "Accountant sees salary offices scoped to their assignment. No cross-office salary data access.",
         "Medium", "UI", "role-permission-matrix.md", "security",
         "ACCOUNTING:NOTIFY permission with readOnly flag. Salary offices should be filtered by accountant's office scope."),

        ("TC-SEC-045", "Customer search partial results for manager's project — #2196",
         "Manager with limited project access.",
         "1. Login as PROJECT_MANAGER\n"
         "2. Navigate to a page with customer search (e.g., Statistics or Projects)\n"
         "3. Search for customers\n"
         "4. Verify search results are scoped to PM's projects\n"
         "5. Verify no 403 error — should return partial results (206) not full block\n"
         "6. If 403 returned: regression of fix #2196",
         "Customer search for managers should return partial results (scoped to their projects), not 403. Fix #2196 changed from 403 to 206 for partial access.",
         "Medium", "Hybrid", "#2196", "security",
         "Ticket #2196 (closed): customer search returned 403 for managers. Fix: return 206 with partial results limited to manager's projects."),
    ],

    "TS-Security-TokenPerms": [
        ("TC-SEC-046", "Admin creates APPLICATION token via Admin > API page",
         "User with ROLE_ADMIN.",
         "1. Login as ADMIN user\n"
         "2. Navigate to Admin > API (/admin/api)\n"
         "3. Click 'Create a key' button\n"
         "4. Enter key name (e.g., 'Test Token SEC-046')\n"
         "5. Verify new token appears in the table with generated UUID value\n"
         "6. Verify token shows empty 'Allowed API methods' initially\n"
         "7. Verify edit and delete action buttons are present\n"
         "CLEANUP: Delete the created token via the delete button",
         "Admin can create APPLICATION tokens. New token appears with UUID value. Initially has no permissions assigned.",
         "Critical", "UI", "auth-authorization-doc.md §Token Management", "security",
         "Token CRUD endpoints require AUTHENTICATED_USER (JWT-only). Admin API page at /admin/api. Token table has: Name, Created, Value, Allowed API methods, Actions."),

        ("TC-SEC-047", "Edit token permissions — add/remove API methods",
         "Admin user. Existing token with no permissions. SETUP: Create a test token first.",
         "SETUP: Create a test token via 'Create a key' on /admin/api\n"
         "1. Login as ADMIN user\n"
         "2. Navigate to Admin > API\n"
         "3. Click edit button on the test token\n"
         "4. Add permissions: EMPLOYEES_VIEW, VACATIONS_VIEW\n"
         "5. Save changes\n"
         "6. Verify 'Allowed API methods' column shows EMPLOYEES_VIEW, VACATIONS_VIEW\n"
         "7. Edit again — remove VACATIONS_VIEW\n"
         "8. Save and verify only EMPLOYEES_VIEW remains\n"
         "CLEANUP: Delete the test token",
         "Token permissions can be added and removed via the edit dialog. Changes persist and are reflected in the table.",
         "High", "UI", "auth-authorization-doc.md", "security",
         "21 ApiPermission enum values available. Permissions stored in token_permissions table."),

        ("TC-SEC-048", "Delete token via Admin > API page",
         "Admin user. Existing token to delete.",
         "SETUP: Create a test token via 'Create a key'\n"
         "1. Login as ADMIN user\n"
         "2. Navigate to Admin > API\n"
         "3. Note the test token in the table\n"
         "4. Click delete button on the test token\n"
         "5. Confirm deletion if prompted\n"
         "6. Verify token is removed from the table\n"
         "7. Via API — attempt to use the deleted token (GET /api/ttt/v1/employees with the deleted token value)\n"
         "8. Verify API returns 401 or 403 (token no longer valid)",
         "Token is removed from the table. Deleted token immediately stops working for API authentication.",
         "High", "UI", "auth-authorization-doc.md", "security",
         "Token deletion should be immediate. Verify the token is invalidated in the database and cannot be reused."),

        ("TC-SEC-049", "Personal employee token visible on Account page",
         "Any logged-in employee.",
         "1. Login as any employee\n"
         "2. Navigate to Account page (/admin/account)\n"
         "3. Verify 'Your secret API token' label is visible on the General tab\n"
         "4. Verify token UUID is displayed (format: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)\n"
         "5. Verify copy button is present next to the token\n"
         "6. Verify regenerate button is present",
         "Every employee sees their personal API token on the Account > General tab with copy and regenerate buttons.",
         "High", "UI", "admin-panel-deep-dive.md §Account", "security",
         "Account page at /admin/account/general. Token displayed with copy (clipboard) and regenerate (refresh) icons. All roles can access Account page."),

        ("TC-SEC-050", "Personal token regeneration changes the token value",
         "Any logged-in employee.",
         "1. Login as employee\n"
         "2. Navigate to Account page (/admin/account)\n"
         "3. Note the current token value\n"
         "4. Click the regenerate button next to the token\n"
         "5. Verify token value changes to a new UUID\n"
         "6. Via API — attempt to use the OLD token value\n"
         "7. Verify old token returns 401/403 (invalidated)\n"
         "8. Via API — use the NEW token value\n"
         "9. Verify new token works (200 response)",
         "Regenerating the personal token creates a new UUID. The old token is immediately invalidated. The new token works for API calls.",
         "High", "UI", "auth-authorization-doc.md", "security",
         "Token regeneration should be immediate. Old token must be invalidated in database."),

        ("TC-SEC-051", "API token with VACATIONS_VIEW reads vacations successfully",
         "API token with VACATIONS_VIEW permission. Query: SELECT t.secret_key FROM token t JOIN token_permissions tp ON t.id = tp.token_id WHERE tp.apipermission = 'VACATIONS_VIEW' LIMIT 1",
         "1. Via API — send GET /api/vacation/v1/vacations/search with API_SECRET_TOKEN header\n"
         "2. Verify response status 200\n"
         "3. Verify vacation data is returned in response body\n"
         "4. Via API — send same request with a token that does NOT have VACATIONS_VIEW\n"
         "5. Verify response status 403",
         "API token with VACATIONS_VIEW permission can read vacation data. Token without this permission gets 403.",
         "Critical", "Hybrid", "auth-authorization-doc.md §API Token", "security",
         "Vacation endpoints use Pattern A: hasAnyAuthority('AUTHENTICATED_USER', 'VACATIONS_VIEW'). Both JWT and API token with correct permission should work."),

        ("TC-SEC-052", "API token cannot access sick leave CRUD (AUTHENTICATED_USER only)",
         "API token with all permissions. Sick leave CRUD endpoints.",
         "1. Via API — send POST /api/vacation/v1/sick-leaves with API_SECRET_TOKEN header and valid sick leave data\n"
         "2. Verify response status 403 (AUTHENTICATED_USER required, not available to API tokens)\n"
         "3. Via API — send PATCH /api/vacation/v1/sick-leaves/{id} with API_SECRET_TOKEN\n"
         "4. Verify response status 403\n"
         "5. Via API — send DELETE /api/vacation/v1/sick-leaves/{id} with API_SECRET_TOKEN\n"
         "6. Verify response status 403\n"
         "7. Via API — send GET /api/vacation/v1/sick-leaves/search with API_SECRET_TOKEN\n"
         "8. Verify this also returns 403 (all sick leave CRUD is JWT-only)",
         "ALL sick leave CRUD endpoints return 403 for API tokens. They require AUTHENTICATED_USER which only JWT grants. This is a design inconsistency — vacation supports API tokens but sick leave does not.",
         "High", "Hybrid", "design issue #45, security-patterns.md §AUTHENTICATED_USER", "security",
         "Sick leave CRUD uses Pattern B (JWT only): hasAuthority('AUTHENTICATED_USER'). No ApiPermission fallback. Blocks all API token automation."),

        ("TC-SEC-053", "API token cannot manage other tokens (JWT required)",
         "API token with all permissions. Token management endpoints.",
         "1. Via API — send GET /api/ttt/v1/tokens with API_SECRET_TOKEN header\n"
         "2. Verify response status 403 (token endpoints require AUTHENTICATED_USER)\n"
         "3. Via API — send POST /api/ttt/v1/tokens with API_SECRET_TOKEN\n"
         "4. Verify response status 403\n"
         "5. Via API — send PATCH /api/ttt/v1/tokens/{id} with API_SECRET_TOKEN\n"
         "6. Verify response status 403",
         "Token management endpoints (GET/POST/PATCH/DELETE /v1/tokens) all require AUTHENTICATED_USER — cannot be accessed via API token. Only JWT auth works.",
         "High", "Hybrid", "auth-authorization-doc.md §Token Management", "security",
         "Token CRUD: all require hasAuthority('AUTHENTICATED_USER'). This is by design to prevent token escalation via API tokens."),

        ("TC-SEC-054", "Report period min/max returns 403 for API token (inconsistency) — #97",
         "API token with OFFICES_VIEW permission.",
         "1. Via API — send GET /api/ttt/v1/offices/periods/report/min with API_SECRET_TOKEN header\n"
         "2. Verify response status (expected: 403)\n"
         "3. Via API — send GET /api/ttt/v1/offices/periods/report/max with API_SECRET_TOKEN\n"
         "4. Verify response status (expected: 403)\n"
         "5. Via API — send GET /api/ttt/v1/offices/periods/approve/min with API_SECRET_TOKEN\n"
         "6. Verify response status 200 (approve min/max accepts API token)\n"
         "7. Via API — send GET /api/ttt/v1/offices/periods/approve/max with API_SECRET_TOKEN\n"
         "8. Verify response status 200",
         "Report period min/max endpoints return 403 for API token (missing OFFICES_VIEW fallback). Approve period min/max accepts API token. This is an inconsistency — both should behave the same way.",
         "Medium", "Hybrid", "design issue #92/#97", "security",
         "OfficePeriodController:46 vs :76 — GET report min/max missing OFFICES_VIEW permission fallback unlike approve min/max."),

        ("TC-SEC-055", "Missing @PreAuthorize on /reports/effort endpoint",
         "Any authenticated user including employee-only.",
         "1. Login as employee-only user\n"
         "2. Via API (with JWT) — send GET /api/ttt/v1/reports/effort\n"
         "3. Verify response status — expected: 200 (missing @PreAuthorize)\n"
         "4. Via API — send GET /api/ttt/v1/reports/employee-projects\n"
         "5. Verify response status — expected: 200 (same missing annotation)\n"
         "6. Verify data returned — assess if this is an information disclosure",
         "Both /reports/effort and /reports/employee-projects lack @PreAuthorize annotation. Any authenticated request can access these endpoints regardless of role.",
         "Medium", "Hybrid", "design issue #53", "security",
         "TaskReportController.effort — missing @PreAuthorize. Any authenticated request succeeds. May expose effort/project data to unauthorized users."),

        ("TC-SEC-056", "Non-admin user cannot access Admin > API page",
         "User with ROLE_EMPLOYEE only.",
         "1. Login as employee-only user\n"
         "2. Verify 'Admin panel' dropdown does NOT show 'API' link\n"
         "3. Navigate directly to /admin/api\n"
         "4. Verify access denied — redirect or empty page\n"
         "5. Login as ACCOUNTANT user\n"
         "6. Verify 'API' link NOT in Admin panel\n"
         "7. Navigate directly to /admin/api — verify access denied",
         "Only ADMIN and VIEW_ALL roles should access Admin > API page. All other roles should be blocked both in navigation and via direct URL.",
         "High", "UI", "role-permission-matrix.md §TOKENS:VIEW", "security",
         "TOKENS:VIEW permission required — only ADM and VALL have it. Frontend PrivateRoute guard + backend @PreAuthorize."),
    ],

    "TS-Security-ErrorCodes": [
        ("TC-SEC-057", "403 response has correct format and no stack trace",
         "User without required permission for an endpoint.",
         "1. Login as employee-only user\n"
         "2. Via API (with JWT) — send GET /api/ttt/v1/tokens (requires AUTHENTICATED_USER + TOKENS:VIEW)\n"
         "3. Verify response status 403\n"
         "4. Inspect response body — should contain a structured error message\n"
         "5. Verify NO Java stack trace in response body\n"
         "6. Verify Content-Type is application/json",
         "403 responses should return clean JSON error body without stack traces. Format: {error, message, status} or similar structured response.",
         "High", "Hybrid", "security-patterns.md", "security",
         "Some endpoints leak stack traces on error. Verify all auth failures return clean error responses."),

        ("TC-SEC-058", "Day correction returns 500 instead of 403 on auth failure — #1286",
         "Unauthorized user attempting day correction. API call to inter-service endpoint.",
         "1. Login as employee-only user\n"
         "2. Via API (with JWT) — attempt PATCH /api/vacation/v1/vacation-days (day correction)\n"
         "3. Record the response status code\n"
         "4. Verify status is 403 (not 500)\n"
         "5. If 500: inter-service call masking the authorization error (#1286)",
         "Day correction authorization failure should return 403, not 500. The 500 was caused by inter-service call masking the original 403.",
         "High", "Hybrid", "#1286", "security",
         "Ticket #1286 (closed): inter-service call between vacation and TTT services masked 403 as 500. Verify fix returns proper status."),

        ("TC-SEC-059", "Token/settings operations return 500 instead of 403 — #1883",
         "Non-admin user attempting token or settings modification.",
         "1. Login as employee-only user\n"
         "2. Via API (with JWT) — attempt PATCH /api/ttt/v1/tokens/{any_id}\n"
         "3. Verify response status 403 (not 500)\n"
         "4. Via API — attempt DELETE /api/ttt/v1/tokens/{any_id}\n"
         "5. Verify response status 403 (not 500)\n"
         "6. Via API — attempt PATCH /api/ttt/v1/settings\n"
         "7. Verify response status 403 (not 500)",
         "Token and settings operations by unauthorized user should return 403, not 500.",
         "High", "Hybrid", "#1883", "security",
         "Ticket #1883 (closed): PATCH/DELETE on tokens/settings returned 500 instead of 403 for unauthorized users."),

        ("TC-SEC-060", "GET /projects/{id} returns 400 instead of 403 — #2164",
         "User without project access. Project ID the user cannot access.",
         "1. Login as employee-only user\n"
         "2. Via API (with JWT) — send GET /api/ttt/v1/projects/{id} for a project the user is not assigned to\n"
         "3. Record the response status code\n"
         "4. Verify status is 403 (not 400)\n"
         "5. Inspect error message for clarity",
         "Accessing a project without permission should return 403 Forbidden, not 400 Bad Request. The 400 is misleading.",
         "Medium", "Hybrid", "#2164", "security",
         "Ticket #2164 (closed): GET /projects/{id} returned 400 instead of 403. Misleading error code."),

        ("TC-SEC-061", "API token on JWT-only endpoint returns 400 (misleading) — #1231",
         "API token. JWT-only endpoint (e.g., /v1/notifications).",
         "1. Via API — send GET /api/ttt/v1/notifications with API_SECRET_TOKEN header\n"
         "2. Record the response status code\n"
         "3. Expected issue: 400 returned instead of 403\n"
         "4. Also test: GET /api/ttt/v1/tokens with API_SECRET_TOKEN\n"
         "5. Also test: GET /api/ttt/v1/offices/periods/report/min with API_SECRET_TOKEN\n"
         "6. Document which endpoints return 400 vs 403 for API tokens",
         "JWT-only endpoints should return 403 when accessed with API token. Some return 400 instead (#1231), making it unclear whether the request format is wrong or the auth is insufficient.",
         "Medium", "Hybrid", "#1231", "security",
         "Ticket #1231 (closed): API_TOKEN got 400 on notifications/tokens/periods. Restricted by design but misleading error code. Should be 403."),

        ("TC-SEC-062", "Stack trace leakage on invalid date format in reports",
         "Any authenticated user.",
         "1. Login as employee\n"
         "2. Via API (with JWT) — send GET /api/ttt/v1/reports?date=invalid-date-format\n"
         "3. Verify response status (expected: 400)\n"
         "4. Inspect response body for Java stack trace presence\n"
         "5. If stack trace visible: information disclosure vulnerability\n"
         "6. Verify error message is user-friendly, not raw exception",
         "Invalid date format should return 400 with a clean error message. Response body must NOT contain Java stack traces or internal class names.",
         "Medium", "Hybrid", "reports-business-rules-reference.md §Stack trace leakage", "security",
         "Known issue: invalid date format returns full Java stack trace. Information disclosure risk — exposes internal class names, line numbers, library versions."),

        ("TC-SEC-063", "DELETE /projects/{id} wrong HTTP status on invalid ID — #2131",
         "Admin user. Non-existent project ID.",
         "1. Login as ADMIN user\n"
         "2. Via API (with JWT) — send DELETE /api/ttt/v1/projects/99999999 (non-existent ID)\n"
         "3. Record response status code\n"
         "4. Expected: 404 (Not Found) or 400 (Bad Request)\n"
         "5. If 500: server error on invalid ID (#2131)\n"
         "6. Verify error response body format",
         "Deleting a non-existent project should return 404 or 400, not 500. Server should handle invalid IDs gracefully.",
         "Medium", "Hybrid", "#2131", "security",
         "Ticket #2131 (closed): DELETE /projects/{id} returned wrong HTTP status on invalid ID."),
    ],

    "TS-Security-RoleSync": [
        ("TC-SEC-064", "CS sync assigns DEPARTMENT_MANAGER role correctly — #807",
         "Employee who is a department manager in CompanyStaff but may lack the role in TTT. SETUP: Via test API — trigger CS sync.",
         "SETUP: Via test API — trigger CS sync (POST /api/ttt/test/v1/sync/employees)\n"
         "1. Wait for sync to complete\n"
         "2. Via API — GET /api/ttt/v1/employees/{login}/roles for a known department manager\n"
         "3. Verify ROLE_DEPARTMENT_MANAGER is present in the roles list\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM employee WHERE login = '<login>') — verify ROLE_DEPARTMENT_MANAGER exists",
         "After CS sync, employees who are department managers in CompanyStaff should have ROLE_DEPARTMENT_MANAGER in TTT. Role assignment should be automatic.",
         "High", "Hybrid", "#807", "security",
         "Ticket #807 (closed): DEPARTMENT_MANAGER role not assigned due to manager hierarchy logic. CS sync dependency."),

        ("TC-SEC-065", "CS sync removes PM role correctly (bug: removes DM instead)",
         "Employee who lost PM status in CompanyStaff. SETUP: Identify employee who was demoted from PM.",
         "1. Via API — GET /api/ttt/v1/employees/{login}/roles for the demoted employee\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = <id> — list all current roles\n"
         "2. SETUP: Via test API — trigger CS sync\n"
         "3. After sync — GET roles again\n"
         "4. Verify ROLE_PROJECT_MANAGER is removed (not ROLE_DEPARTMENT_MANAGER)\n"
         "5. If ROLE_DEPARTMENT_MANAGER removed instead of ROLE_PROJECT_MANAGER: bug #78 still present",
         "When employee loses PM status in CS, the sync should remove ROLE_PROJECT_MANAGER. Known bug: removes ROLE_DEPARTMENT_MANAGER instead (ProjectManagerRolePostProcessor:39 references wrong role constant).",
         "Critical", "Hybrid", "design issue #78/#132", "security",
         "Critical bug in ProjectManagerRolePostProcessor.java:39 — removes ROLE_DEPARTMENT_MANAGER instead of ROLE_PROJECT_MANAGER on demotion. PM role persists indefinitely."),

        ("TC-SEC-066", "Stale JWT after role change — user must reauth",
         "Employee whose roles change during session. SETUP: Via DB — change role for test employee, or use CS sync to trigger role change.",
         "1. Login as employee — note JWT claims (roles visible in decoded JWT)\n"
         "2. SETUP: Via DB or CS sync — change the employee's roles (e.g., add ROLE_ADMIN)\n"
         "3. WITHOUT logging out, navigate to admin page that requires ADMIN role\n"
         "4. Verify behavior — old JWT still carries old roles\n"
         "5. Expected: access denied despite DB role change (JWT is stale)\n"
         "6. Logout and login again — verify new JWT has updated roles\n"
         "7. Access the admin page — verify now accessible",
         "JWT tokens carry roles at generation time. Role changes are NOT reflected in existing JWTs. User must re-authenticate (logout/login) to get updated roles. No token refresh mechanism exists.",
         "High", "Hybrid", "security-patterns.md §JWT Token", "security",
         "JWT has 1-day expiry with no refresh. Role changes require full re-authentication via CAS. This is a known design limitation."),

        ("TC-SEC-067", "GET /v1/employees/{login}/roles matches DB — #2063",
         "Any employee. Compare API response with DB data.",
         "1. Login as admin\n"
         "2. Via API — send GET /api/ttt/v1/employees/{login}/roles for a specific employee\n"
         "3. Record the roles returned\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = (SELECT id FROM employee WHERE login = '<login>')\n"
         "4. Compare API response with DB query result\n"
         "5. Verify they match exactly\n"
         "6. If mismatch: caching issue (#2063) — /v1/employees returns cached data",
         "API roles endpoint should return current roles matching the database. Ticket #2063: /v1/employees and /v1/employees/{login}/roles returned contradictory data due to caching.",
         "Medium", "Hybrid", "#2063", "security",
         "Caching issue: /v1/employees returns cached employee data while /v1/employees/{login}/roles reads from DB directly. After role change, these may diverge."),

        ("TC-SEC-068", "Former HR retains OFFICE_HR role after position change",
         "Employee who was previously HR but changed position. Query: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id = r.employee_id WHERE r.role = 'ROLE_OFFICE_HR' AND e.enabled = true",
         "1. Via API — GET /api/ttt/v1/employees/{login}/roles for known former HR employee\n"
         "2. Verify whether ROLE_OFFICE_HR is still present\n"
         "DB-CHECK: Verify in CS whether the employee is still in HR position\n"
         "3. If ROLE_OFFICE_HR persists after position change: confirm design issue #80\n"
         "4. SETUP: Trigger CS sync — verify role is NOT removed (only add, never remove)",
         "OfficeHRRolePostProcessor only ADDS ROLE_OFFICE_HR, never removes it. Former HR personnel keep elevated permissions permanently. This is a known design issue.",
         "High", "Hybrid", "design issue #80", "security",
         "OfficeHRRolePostProcessor.java — only adds role, never removes. Former HR staff retain role until manually removed from DB."),

        ("TC-SEC-069", "OFFICE_DIRECTOR role assignment after CS sync — #522",
         "Employee who should have OFFICE_DIRECTOR role based on CS data. SETUP: Via test API — trigger CS sync.",
         "SETUP: Via test API — trigger CS sync\n"
         "1. Via API — GET /api/ttt/v1/employees/{login}/roles for known office director\n"
         "2. Verify ROLE_OFFICE_DIRECTOR is present\n"
         "DB-CHECK: SELECT role FROM employee_global_roles WHERE employee_id = <id>\n"
         "3. If role missing: CS sync delay issue (#522)",
         "After CS sync, employees designated as office directors in CompanyStaff should receive ROLE_OFFICE_DIRECTOR in TTT.",
         "Medium", "Hybrid", "#522", "security",
         "Ticket #522 (closed): OFFICE_DIRECTOR role not assigned after CS sync. Depends on RabbitMQ event processing."),
    ],

    "TS-Security-Regression": [
        ("TC-SEC-070", "REGRESSION #3410 — User gets 403 on task creation (OPEN bug)",
         "Affected user who experiences 403 on task creation. Query: SELECT e.login FROM employee e WHERE e.enabled = true AND e.deactivated = false ORDER BY random() LIMIT 1",
         "1. Login as the affected employee\n"
         "2. Navigate to My Tasks page\n"
         "3. Attempt to add a new task (click + button or similar)\n"
         "4. Verify task creation succeeds or fails with 403\n"
         "5. If 403: active bug confirmed\n"
         "6. Try with a different employee to determine scope",
         "Task creation should succeed for any enabled, non-readOnly employee. If 403 occurs, this is an active bug (#3410).",
         "Critical", "UI", "#3410", "security",
         "OPEN bug as of current sprint. User gets 403 when trying to create a task. Investigate scope — may affect specific roles or all users."),

        ("TC-SEC-071", "REGRESSION #1275 — Deactivated user infinite 500 loop",
         "Deactivated employee account. Same as TC-SEC-004.",
         "1. Attempt login with deactivated user credentials\n"
         "2. Monitor network requests in DevTools\n"
         "3. Verify NO infinite loop of 500 responses to /v1/security/jwt\n"
         "4. Verify a clear error response within 5 seconds\n"
         "5. Expected: single error response (412 or 403), not retry loop",
         "Deactivated user login should produce a single, clear error. No infinite 500 loop requesting JWT.",
         "Critical", "UI", "#1275", "security",
         "Historical bug: deactivated user triggered infinite 500 loop. Verify fix prevents the loop and returns proper error."),

        ("TC-SEC-072", "REGRESSION #2270 — JWT expiry WebSocket storm",
         "Same as TC-SEC-003 — JWT expiry scenario.",
         "1. Login and let JWT expire (via clock manipulation)\n"
         "2. Monitor WebSocket connections in DevTools\n"
         "3. Verify NO rapid WebSocket reconnection attempts\n"
         "4. Verify polling fallback to /v1/authentication/check at 5s intervals\n"
         "5. Count reconnection attempts — should be < 3 before fallback",
         "After JWT expiry, WebSocket disconnect should trigger polling fallback, not infinite reconnection storm.",
         "High", "Hybrid", "#2270", "security",
         "Fix added 5-second polling interval on /v1/authentication/check as fallback after WebSocket disconnect."),

        ("TC-SEC-073", "REGRESSION #2050 — OFFICE_DIRECTOR cross-office visibility",
         "Same as TC-SEC-038.",
         "1. Login as OFFICE_DIRECTOR\n"
         "2. Navigate to Admin > Employees\n"
         "3. Count displayed employees\n"
         "DB-CHECK: Compare with SELECT COUNT(*) FROM employee WHERE office_id = <director_office> AND enabled = true\n"
         "4. Verify counts match (no cross-office leakage)",
         "OFFICE_DIRECTOR employee list should be scoped to own office only. No employees from other offices.",
         "High", "UI", "#2050", "security",
         "Regression for ticket #2050: director saw all offices. Verify office-scope filter is active."),

        ("TC-SEC-074", "REGRESSION #117 — Any approver approves any vacation",
         "Same as TC-SEC-028.",
         "1. Create vacation for Employee with Manager A as approver\n"
         "2. Login as Manager B (NOT the approver)\n"
         "3. Via API — attempt approval as Manager B\n"
         "4. Verify 403 — only assigned approver should be able to approve",
         "Vacation approval must be restricted to the assigned approver. Manager B should get 403 when trying to approve Manager A's employee's vacation.",
         "Critical", "Hybrid", "#117", "security",
         "Critical historical bug: approver field was decorative. Verify server-side enforcement of approver check."),

        ("TC-SEC-075", "REGRESSION #3002 — PM edits sick leave via API",
         "Same as TC-SEC-030.",
         "1. Login as PROJECT_MANAGER\n"
         "2. Via API — attempt PATCH sick leave record\n"
         "3. Verify 403 — PM should not be able to edit sick leaves\n"
         "4. Login as ACCOUNTANT — verify PATCH succeeds (authorized role)",
         "Sick leave editing restricted to accountant/admin. PM gets 403 on PATCH /v1/sick-leaves/{id}.",
         "High", "Hybrid", "#3002", "security",
         "Regression: PM could edit sick leave records via API because endpoint used AUTHENTICATED_USER."),

        ("TC-SEC-076", "REGRESSION #3012 — Employee accesses accounting sick-leaves",
         "Same as TC-SEC-031.",
         "1. Login as employee-only user\n"
         "2. Navigate to /accounting/sick-leaves\n"
         "3. Verify access blocked (VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW required)\n"
         "4. Login as ACCOUNTANT — verify access granted",
         "Regular employee cannot access /accounting/sick-leaves page. Only roles with VACATIONS:SICK_LEAVE_ACCOUNTING_VIEW permission.",
         "High", "UI", "#3012", "security",
         "This is the ONE accounting route that HAS a permission check. Verify it works while the other 4 routes remain unprotected."),

        ("TC-SEC-077", "REGRESSION #1250 — Employee redirects vacation via API",
         "Same as TC-SEC-026.",
         "1. Login as employee\n"
         "2. Create a vacation request\n"
         "3. Via API — attempt to change approverId via PUT /vacations/{id}\n"
         "4. Verify redirect operation is blocked for employee role",
         "Employee should not be able to redirect their own vacation via API. UI correctly hides this option — API must enforce the same restriction.",
         "High", "Hybrid", "#1250", "security",
         "Verify fix prevents employee from changing approverId via API."),

        ("TC-SEC-078", "REGRESSION #870 — Cross-project approval leakage",
         "Same as TC-SEC-032.",
         "1. Login as PM\n"
         "2. Go to Confirmation page with 'other projects' checkbox\n"
         "3. Verify tasks from non-managed projects are read-only\n"
         "4. Attempt to approve a foreign project task — should fail",
         "PM can VIEW tasks from other projects but cannot APPROVE them. Checkbox should not grant cross-project approval.",
         "High", "UI", "#870", "security",
         "Historical bug: checkbox leaked approval capability across projects."),

        ("TC-SEC-079", "REGRESSION #480 — Cross-office vacation day modification",
         "Same as TC-SEC-043.",
         "1. Login as accountant from Office A\n"
         "2. Via API — attempt to modify vacation days for Office B employee\n"
         "3. Verify 403 — cross-office modification blocked\n"
         "4. Attempt for own-office employee — verify success",
         "Cross-office vacation day modification must be blocked. Accountant can only modify days for employees in their assigned office(s).",
         "Critical", "Hybrid", "#480", "security",
         "Critical regression: accountant could modify vacation days across offices."),

        ("TC-SEC-080", "REGRESSION #1946 — TECH_LEAD/SPM/OWNER permissions undocumented",
         "Users with ROLE_TECH_LEAD, PROJECT_SENIOR_MANAGER, PROJECT_OBSERVER.",
         "1. Login as TECH_LEAD\n"
         "2. Navigate through all pages — document accessible vs blocked pages\n"
         "3. Login as user who is Senior Project Manager on a project\n"
         "4. Document accessible pages and operations\n"
         "5. Compare with role-permission-matrix.md — note undocumented access\n"
         "6. If access differs from matrix: ticket #1946 gap confirmed",
         "Document actual permissions for TECH_LEAD, SPM, and OWNER roles. These are not fully documented (#1946). Compare actual vs expected access.",
         "Medium", "UI", "#1946", "security",
         "Ticket #1946 (OPEN): permissions for TECH_LEAD, SPM, OWNER not fully documented. Discovery test to map actual access."),

        ("TC-SEC-081", "REGRESSION #692 — API key authentication from Swagger UI",
         "API token. Swagger UI access.",
         "1. Navigate to Swagger UI (https://ttt-qa-1.noveogroup.com/api/ttt/swagger-ui.html)\n"
         "2. Click Authorize button\n"
         "3. Enter API_SECRET_TOKEN value\n"
         "4. Execute a GET endpoint (e.g., /v1/employees)\n"
         "5. Verify API call succeeds with API token authentication",
         "API token authentication should work correctly from Swagger UI. The Authorize dialog should accept API_SECRET_TOKEN and pass it in the header.",
         "Medium", "UI", "#692", "security",
         "Ticket #692 (closed): API key auth from Swagger UI was broken. Verify fix allows token-based auth via Swagger."),
    ],
}

# --- Helper functions --------------------------------------------------------

def style_header_row(ws, row_idx, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def style_data_row(ws, row_idx, max_col, alt=False):
    fill = FILL_ROW_ALT if alt else FILL_ROW_WHITE
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_BODY
        cell.fill = fill
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_link(ws):
    ws.cell(row=1, column=1, value='← Back to Plan')
    ws.cell(row=1, column=1).font = FONT_BACK_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"


# --- Sheet builders ----------------------------------------------------------

def build_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN
    set_col_widths(ws, [20, 90])

    ws.cell(row=1, column=1, value=PLAN_OVERVIEW["title"]).font = FONT_TITLE
    ws.merge_cells("A1:B1")

    row = 3
    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP

    row = 5
    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    for i, obj in enumerate(PLAN_OVERVIEW["objectives"]):
        ws.cell(row=row + i, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row + i, column=2).alignment = ALIGN_WRAP
    row += len(PLAN_OVERVIEW["objectives"]) + 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    for i, env in enumerate(PLAN_OVERVIEW["environments"]):
        ws.cell(row=row + i, column=2, value=env).font = FONT_BODY
    row += len(PLAN_OVERVIEW["environments"]) + 1

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    for i, dep in enumerate(PLAN_OVERVIEW["dependencies"]):
        ws.cell(row=row + i, column=2, value=f"• {dep}").font = FONT_BODY
    row += len(PLAN_OVERVIEW["dependencies"]) + 1

    # Suite hyperlinks
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, cases in SUITES.items():
        label = f"{suite_name} — {len(cases)} cases"
        ws.cell(row=row, column=2, value=label).font = FONT_LINK
        ws.cell(row=row, column=2).hyperlink = f"#'{suite_name}'!A1"
        row += 1

    return ws


def build_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature", "AuthModel", "RoleMatrix", "APIBypass", "CrossOffice",
               "TokenPerms", "ErrorCodes", "RoleSync", "Regression", "Total"]
    set_col_widths(ws, [40, 12, 12, 12, 12, 12, 12, 12, 12, 10])

    row = 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    suite_names = list(SUITES.keys())
    for i, feat_row in enumerate(FEATURE_MATRIX):
        row = i + 2
        for col, val in enumerate(feat_row, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col == 1:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = ALIGN_CENTER
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

        # Add hyperlinks for non-zero suite cells
        for col_idx in range(2, len(headers)):  # columns 2-9 (suites)
            val = feat_row[col_idx]
            if val and val > 0:
                suite_idx = col_idx - 2
                if suite_idx < len(suite_names):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.font = FONT_LINK
                    cell.hyperlink = f"#'{suite_names[suite_idx]}'!A1"

    # Totals row
    total_row = len(FEATURE_MATRIX) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col in range(2, len(headers) + 1):
        total = sum(r[col - 1] for r in FEATURE_MATRIX)
        ws.cell(row=total_row, column=col, value=total).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=total_row, column=col).alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(FEATURE_MATRIX) + 1}"
    return ws


def build_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    set_col_widths(ws, [30, 50, 12, 12, 12, 50])

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for i, (feat, risk, like, impact, sev, mit) in enumerate(RISK_ASSESSMENT):
        row = i + 2
        ws.cell(row=row, column=1, value=feat)
        ws.cell(row=row, column=2, value=risk)
        ws.cell(row=row, column=3, value=like)
        ws.cell(row=row, column=4, value=impact)
        ws.cell(row=row, column=5, value=sev)
        ws.cell(row=row, column=6, value=mit)
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))
        if sev in severity_fills:
            ws.cell(row=row, column=5).fill = severity_fills[sev]

    ws.auto_filter.ref = f"A1:F{len(RISK_ASSESSMENT) + 1}"
    return ws


def build_suite_sheet(wb, suite_name, cases):
    ws = wb.create_sheet(suite_name)
    ws.sheet_properties.tabColor = TAB_COLOR_SUITE

    add_back_link(ws)

    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    set_col_widths(ws, [14, 40, 40, 55, 40, 10, 10, 30, 14, 35])

    header_row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(headers))

    for i, case in enumerate(cases):
        row = header_row + 1 + i
        for col, val in enumerate(case, 1):
            ws.cell(row=row, column=col, value=val)
        style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(cases)}"
    return ws


# --- Main --------------------------------------------------------------------

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    build_plan_overview(wb)
    build_feature_matrix(wb)
    build_risk_assessment(wb)

    for suite_name, cases in SUITES.items():
        build_suite_sheet(wb, suite_name, cases)

    wb.save(OUTPUT_FILE)

    total_cases = sum(len(cases) for cases in SUITES.values())
    print(f"Generated {OUTPUT_FILE}")
    print(f"  Suites: {len(SUITES)}")
    print(f"  Total test cases: {total_cases}")
    print(f"  Tabs: Plan Overview, Feature Matrix, Risk Assessment, {', '.join(SUITES.keys())}")


if __name__ == "__main__":
    main()
