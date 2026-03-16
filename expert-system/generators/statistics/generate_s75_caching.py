#!/usr/bin/env python3
"""Add TS-STAT-CacheArch supplementary test suite to statistics.xlsx.

Sprint 15 feature: #3337 — Performance enhancement via materialized view pattern.
Replaced on-the-fly statistic calculation with pre-computed data in statistic_report table,
event-driven cache invalidation via RabbitMQ, GET→POST endpoint migration, role-based
filtering, and daily scheduled sync.

Adds 12 test cases (TC-STAT-127 through TC-STAT-138) covering:
  - Endpoint migration (2): POST /v1/statistic bulk, POST /v1/statistic/sick-leaves bulk
  - Role-based access (3): admin sees all, DM sees direct reports, TL sees own employees
  - Cache invalidation (4): vacation event, sick leave event, task report event,
    event type discrimination (INITIAL_SYNC vs VACATION_CHANGES)
  - Frontend resilience (2): empty response clears loading, saga cancellation on period switch
  - Cron sync (1): daily 4am sync correctness
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ── Styling constants (match existing workbook) ────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ─────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab matching existing format."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, tc_item in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            tc_item["id"], tc_item["title"], tc_item["preconditions"],
            tc_item["steps"], tc_item["expected"], tc_item["priority"],
            tc_item["type"], tc_item["req_ref"], tc_item["module"],
            tc_item.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# TEST CASE DATA — TS-STAT-CacheArch (12 cases)
# ══════════════════════════════════════════════════════════════════

CACHE_ARCH_CASES = [
    # ── Endpoint Migration (2 cases) ──────────────────────────────

    tc("TC-STAT-127",
       "POST /v1/statistic bulk request returns vacation statistics for multiple employees",
       "Multiple employees with reported hours in the target month.\n"
       "Find test data:\n"
       "  SELECT DISTINCT sr.employee_login\n"
       "  FROM statistic_report sr\n"
       "  WHERE sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE) - 1\n"
       "  LIMIT 20;\n\n"
       "Collect 5-10 employee logins for the request body.\n"
       "Use timemachine or qa-1 env.",
       "1. POST /api/ttt/v1/statistic\n"
       "   Body: {\"employeesLogins\": [\"login1\",\"login2\",...],\n"
       "          \"startDate\": \"YYYY-MM-01\", \"endDate\": \"YYYY-MM-28\"}\n"
       "2. Verify response contains statistics for each requested employee.\n"
       "3. Verify response includes fields: login, personalNorm, budgetNorm,\n"
       "   reportedEffort (or reported), excess.\n"
       "4. Verify OLD endpoint no longer works:\n"
       "   GET /api/ttt/v1/statistic?employeeLogin=X&startDate=...&endDate=...\n"
       "   Should return 405 Method Not Allowed (or similar).\n"
       "5. DB: Cross-verify with statistic_report table:\n"
       "   SELECT employee_login, month_norm, reported_effort\n"
       "   FROM statistic_report\n"
       "   WHERE employee_login IN ('login1','login2',...)\n"
       "     AND report_date = 'YYYY-MM-01';\n"
       "6. Verify response data matches DB cached values.",
       "POST /v1/statistic returns bulk statistics for all requested employees.\n"
       "Response data matches pre-computed values in statistic_report table.\n"
       "No cross-service calls at query time (fast response).\n"
       "Old GET endpoint returns 405 or 404 (endpoint changed to POST).\n"
       "Response contains: personalNorm, budgetNorm, reported, excess per employee.\n"
       "reported_effort field type is BigDecimal (not double).",
       "Critical", "API",
       "#3337", "Statistics / API / StatisticController",
       "MR !5150 changed GET→POST. MR !5013 changed backend to read from "
       "statistic_report table. Verify both: endpoint method AND data source."),

    tc("TC-STAT-128",
       "POST /v1/statistic/sick-leaves bulk request returns sick leave stats",
       "Multiple employees with sick leaves in the target period.\n"
       "Find test data:\n"
       "  SELECT DISTINCT sl.employee_login\n"
       "  FROM sick_leave sl\n"
       "  WHERE sl.start_date >= 'YYYY-01-01'\n"
       "  LIMIT 10;\n\n"
       "Collect 5-10 logins with sick leaves.",
       "1. POST /api/ttt/v1/statistic/sick-leaves\n"
       "   Body: {\"employeesLogins\": [\"login1\",\"login2\",...],\n"
       "          \"startDate\": \"YYYY-01-01\", \"endDate\": \"YYYY-12-31\"}\n"
       "2. Verify response contains sick leave data per employee.\n"
       "3. Verify OLD endpoint no longer works:\n"
       "   GET /api/ttt/v1/statistic/sick-leaves?...\n"
       "   Should return 405.\n"
       "4. Test with empty logins array: {\"employeesLogins\": []}\n"
       "   Expect empty response, not error.\n"
       "5. Test with single login: {\"employeesLogins\": [\"login1\"]}\n"
       "   Verify equivalent to per-employee query.\n"
       "6. Verify sickLeaves field is an object (not string) in response.",
       "POST /v1/statistic/sick-leaves returns bulk sick leave data.\n"
       "Empty logins array → empty response (not error).\n"
       "Single login → correct per-employee data.\n"
       "Old GET endpoint returns 405.\n"
       "sickLeaves field type is object (MR !5155 fixed PropTypes).\n"
       "Response time significantly faster than old batched GET approach.",
       "High", "API",
       "#3337", "Statistics / API / StatisticReportController",
       "MR !5150 changed sick-leaves endpoint from GET to POST. "
       "MR !5155 fixed frontend PropTypes (string→object). Verify both."),

    # ── Role-Based Access (3 cases) ───────────────────────────────

    tc("TC-STAT-129",
       "Statistics: Admin/Chief Accountant sees ALL employees in report",
       "User with ADMIN or CHIEF_ACCOUNTANT role.\n"
       "Find admin user:\n"
       "  SELECT e.login FROM employee e\n"
       "  JOIN employee_role er ON e.id = er.employee_id\n"
       "  WHERE er.role = 'ROLE_ADMIN'\n"
       "    AND e.enabled = true\n"
       "  LIMIT 3;",
       "1. Login as ADMIN user.\n"
       "2. Navigate to Statistics > Employee Reports.\n"
       "3. Select a reporting period (e.g., previous month).\n"
       "4. Count total employees shown in the table.\n"
       "5. DB: Count all enabled employees:\n"
       "   SELECT COUNT(*) FROM employee WHERE enabled = true;\n"
       "6. Verify the table shows ALL employees (not filtered).\n"
       "7. POST /api/ttt/v1/statistic with all employee logins.\n"
       "8. Verify all logins are returned in response.\n"
       "9. Repeat for CHIEF_ACCOUNTANT role — same result expected.",
       "ADMIN and CHIEF_ACCOUNTANT see all employees in statistics.\n"
       "No filtering applied: employee count matches total enabled employees.\n"
       "API returns data for every requested login.\n"
       "Both roles have identical visibility scope.",
       "High", "Security",
       "#3337", "Statistics / Access Control / Role Filtering",
       "MR !5101 added getEmployeeLoginsByCurrentUserRole() which returns "
       "null for ADMIN and CHIEF_ACCOUNTANT (null = no filter = see all)."),

    tc("TC-STAT-130",
       "Statistics: Department Manager sees only direct reports",
       "User with DEPARTMENT_MANAGER role who manages a team.\n"
       "Find DM user:\n"
       "  SELECT e.login, e.full_name FROM employee e\n"
       "  JOIN employee_role er ON e.id = er.employee_id\n"
       "  WHERE er.role = 'ROLE_DEPARTMENT_MANAGER'\n"
       "    AND e.enabled = true\n"
       "  LIMIT 3;\n\n"
       "Find their direct reports:\n"
       "  SELECT e.login FROM employee e\n"
       "  WHERE e.manager_id = {dm_id}\n"
       "    AND e.enabled = true;",
       "1. Login as DEPARTMENT_MANAGER user.\n"
       "2. Navigate to Statistics > Employee Reports.\n"
       "3. Select a reporting period.\n"
       "4. Note the employees shown in the table.\n"
       "5. DB: Get DM's direct reports:\n"
       "   SELECT e.login FROM employee e\n"
       "   WHERE e.manager_id = (SELECT id FROM employee WHERE login = '{dm_login}')\n"
       "   AND e.enabled = true;\n"
       "6. Verify the table shows ONLY the DM's direct reports.\n"
       "7. Verify the DM cannot see employees from other departments.\n"
       "8. API: POST /v1/statistic with a login NOT in DM's reports.\n"
       "   Verify that employee is excluded from response.",
       "Department Manager sees ONLY their direct reports.\n"
       "Employee count matches DB query for manager_id = DM's ID.\n"
       "Employees from other departments NOT visible.\n"
       "API filters out non-managed employees from response.\n"
       "DM cannot access other managers' team statistics.",
       "High", "Security",
       "#3337", "Statistics / Access Control / Role Filtering",
       "MR !5101: ROLE_DEPARTMENT_MANAGER → filter by managerId. "
       "Uses EmployeeRepository.findAllByOfficesOrManagers() JPQL query."),

    tc("TC-STAT-131",
       "Statistics: Tech Lead sees only their assigned employees",
       "User with TECH_LEAD role.\n"
       "Find TL user:\n"
       "  SELECT e.login FROM employee e\n"
       "  JOIN employee_role er ON e.id = er.employee_id\n"
       "  WHERE er.role = 'ROLE_TECH_LEAD'\n"
       "    AND e.enabled = true\n"
       "  LIMIT 3;\n\n"
       "Find their employees:\n"
       "  SELECT e.login FROM employee e\n"
       "  WHERE e.tech_lead_id = {tl_id}\n"
       "    AND e.enabled = true;",
       "1. Login as TECH_LEAD user.\n"
       "2. Navigate to Statistics > Employee Reports.\n"
       "3. Select a reporting period.\n"
       "4. Note the employees shown.\n"
       "5. DB: Get TL's employees:\n"
       "   SELECT e.login FROM employee e\n"
       "   WHERE e.tech_lead_id = (SELECT id FROM employee WHERE login = '{tl_login}')\n"
       "   AND e.enabled = true;\n"
       "6. Verify the table shows ONLY the TL's employees.\n"
       "7. Compare with OFFICE_DIRECTOR/ACCOUNTANT visibility:\n"
       "   - Login as OFFICE_DIRECTOR → sees employees in their profit centers.\n"
       "   - Login as ACCOUNTANT → sees employees in their profit centers.\n"
       "8. Verify each role sees its expected scope.",
       "Tech Lead sees ONLY employees where tech_lead_id matches.\n"
       "Office Director sees employees in their profit center offices.\n"
       "Accountant sees employees in their profit center offices.\n"
       "Each role's filter is applied independently.\n"
       "Scopes may overlap (same employee visible to multiple roles).",
       "High", "Security",
       "#3337", "Statistics / Access Control / Role Filtering",
       "MR !5101: ROLE_TECH_LEAD → filter by techLeadId. "
       "ROLE_OFFICE_DIRECTOR/ROLE_ACCOUNTANT → filter by officeIds. "
       "ROLE_OFFICE_HR → filter by hrId. Tests multiple role scopes."),

    # ── Cache Invalidation (4 cases) ──────────────────────────────

    tc("TC-STAT-132",
       "Cache: Vacation create/delete updates month_norm via RabbitMQ event",
       "Employee with statistic_report record for current month.\n"
       "No administrative vacation in current month.\n"
       "  SELECT sr.employee_login, sr.month_norm, sr.budget_norm\n"
       "  FROM statistic_report sr\n"
       "  WHERE sr.month_norm = sr.budget_norm\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "  LIMIT 5;\n\n"
       "Use timemachine env with allow_api_mutations=true.",
       "1. DB: Record current month_norm, budget_norm for employee.\n"
       "2. Create an administrative vacation for this month:\n"
       "   POST /api/vacation/v1/vacation/create\n"
       "   (type=ADMINISTRATIVE, 3 working days in current month).\n"
       "3. Wait 5-10 seconds for RabbitMQ event propagation.\n"
       "4. DB: Re-query statistic_report:\n"
       "   SELECT month_norm, budget_norm, reported_effort\n"
       "   FROM statistic_report\n"
       "   WHERE employee_login = '{login}'\n"
       "     AND year = {Y} AND month = {M};\n"
       "5. Verify budget_norm increased by 24 (3 days × 8h).\n"
       "6. Verify month_norm unchanged.\n"
       "7. Delete the vacation.\n"
       "8. Wait for event propagation.\n"
       "9. Re-query: verify budget_norm reverted to original value.\n"
       "10. Navigate to Statistics page — verify UI reflects changes.",
       "After vacation creation:\n"
       "  - budget_norm = original + 24 (3 admin days × 8h).\n"
       "  - month_norm unchanged (not affected by admin vacation).\n"
       "  - reported_effort unchanged.\n"
       "After vacation deletion:\n"
       "  - budget_norm reverts to original value.\n"
       "  - Event-driven: update happens within seconds (not waiting for 4am cron).\n"
       "MQ path: Vacation service publishes EmployeeMonthNormContextCalculated\n"
       "event → RabbitMQ → TTT service consumes and updates statistic_report.",
       "Critical", "Integration",
       "#3337", "Statistics / Cache / RabbitMQ",
       "MR !5013: RabbitMQ topic exchange TTT_BACKEND_EMPLOYEE_TOPIC. "
       "Vacation service publishes event, TTT service consumes. "
       "This tests the complete cross-service event-driven update path."),

    tc("TC-STAT-133",
       "Cache: Sick leave create/change/delete updates month_norm",
       "Employee with statistic_report record for current month.\n"
       "No existing sick leave in current month.\n"
       "  SELECT sr.employee_login, sr.month_norm\n"
       "  FROM statistic_report sr\n"
       "  WHERE sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "  LIMIT 5;\n\n"
       "Use timemachine env with allow_api_mutations=true.",
       "1. DB: Record current month_norm for employee.\n"
       "2. Create a sick leave for the employee in current month (5 days):\n"
       "   POST /api/vacation/v1/sick-leave/create\n"
       "3. Wait 5-10 seconds for event propagation.\n"
       "4. DB: Re-query statistic_report month_norm.\n"
       "5. Verify month_norm decreased (sick leave reduces working days).\n"
       "6. Change the sick leave (extend to 7 days):\n"
       "   PATCH /api/vacation/v1/sick-leave/{id}\n"
       "7. Wait, re-query. Verify month_norm decreased further.\n"
       "8. Delete the sick leave.\n"
       "9. Wait, re-query. Verify month_norm reverted to original.\n"
       "10. For sick leave spanning two months (e.g., Jan 28 - Feb 3):\n"
       "    Verify BOTH months' statistic_report records updated.",
       "Sick leave creation reduces month_norm (fewer working days).\n"
       "Sick leave change (extend/shorten) updates month_norm accordingly.\n"
       "Sick leave deletion restores original month_norm.\n"
       "Cross-month sick leave: events published for EACH affected month.\n"
       "Updates happen event-driven (within seconds), not waiting for cron.\n"
       "BUG (pre-fix, MR !5200): month_norm did NOT update on sick leave —\n"
       "only updated by daily cron. Verify this is now fixed.",
       "Critical", "Integration",
       "#3337", "Statistics / Cache / Sick Leave Events",
       "MR !5200 added sick leave event handlers: SickLeaveCreated/Changed/"
       "DeletedEventListener → sendUpdateMonthNormEvent(sickLeave). "
       "Multi-month sick leaves publish separate events per month. "
       "QA Bug #2 from omaksimova — verify fix."),

    tc("TC-STAT-134",
       "Cache: Task report add/patch/delete updates reported_effort",
       "Employee with statistic_report record for current month.\n"
       "Employee has reported some hours this month.\n"
       "  SELECT sr.employee_login, sr.reported_effort, sr.month_norm\n"
       "  FROM statistic_report sr\n"
       "  WHERE sr.reported_effort > 0\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "  LIMIT 5;\n\n"
       "Use timemachine env with allow_api_mutations=true.",
       "1. DB: Record current reported_effort for employee.\n"
       "2. Add a task report (8h) for the employee in current month:\n"
       "   POST /api/ttt/v1/report (effort: 8, date in current month).\n"
       "3. DB: Re-query reported_effort immediately.\n"
       "4. Verify reported_effort increased by 8.000.\n"
       "5. Patch the report to 4h:\n"
       "   PATCH /api/ttt/v1/report/{id} (effort: 4).\n"
       "6. Verify reported_effort decreased by 4.000.\n"
       "7. Delete the report:\n"
       "   DELETE /api/ttt/v1/report/{id}.\n"
       "8. Verify reported_effort decreased by 4.000.\n"
       "9. Verify month_norm was NOT changed by task report events.\n"
       "10. Navigate to Statistics page — verify UI shows updated reported.",
       "Task report add: reported_effort increases by effort amount.\n"
       "Task report patch: reported_effort adjusts to new effort.\n"
       "Task report delete: reported_effort decreases accordingly.\n"
       "reported_effort type is DECIMAL(10,3) — values like 8.000.\n"
       "month_norm NOT affected by task report events.\n"
       "Update is synchronous via TaskReportEventListener (not MQ).\n"
       "Excess recalculated: (reported_effort - budget_norm) / budget_norm × 100.",
       "High", "Integration",
       "#3337", "Statistics / Cache / Task Report Events",
       "MR !5013: TaskReportEventListener calls "
       "statisticReportSyncService.updateMonthlyReportedEffortForEmployee() "
       "on add/patch/delete. This is direct service call, not MQ. "
       "Verify update is immediate, not deferred."),

    tc("TC-STAT-135",
       "Cache: Event type discrimination — absence events do NOT delete unrelated records",
       "Multiple employees with statistic_report records.\n"
       "Employee A has vacation. Employee B has NO vacation but has statistic_report.\n"
       "Find test data:\n"
       "  -- Employee A (with vacation):\n"
       "  SELECT DISTINCT sr.employee_login\n"
       "  FROM statistic_report sr\n"
       "  JOIN vacation v ON sr.employee_login = (\n"
       "    SELECT login FROM employee WHERE id = v.employee_id)\n"
       "  WHERE sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "  LIMIT 3;\n\n"
       "  -- Employee B (no vacation, has stat report):\n"
       "  SELECT sr.employee_login FROM statistic_report sr\n"
       "  WHERE sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.employee_login NOT IN (\n"
       "      SELECT e.login FROM employee e\n"
       "      JOIN vacation v ON e.id = v.employee_id\n"
       "      WHERE v.status NOT IN ('DELETED','CANCELED'))\n"
       "  LIMIT 3;",
       "1. DB: Record statistic_report records for Employee B.\n"
       "   SELECT * FROM statistic_report\n"
       "   WHERE employee_login = '{B_login}';\n"
       "2. Create a vacation for Employee A (not B).\n"
       "3. Wait for event propagation (5-10 sec).\n"
       "4. DB: Re-query statistic_report for Employee B.\n"
       "   Verify ALL records still exist (not deleted).\n"
       "5. Verify Employee A's records were correctly updated.\n"
       "6. DB: Count total statistic_report records before and after.\n"
       "   SELECT COUNT(*) FROM statistic_report;\n"
       "7. Verify total count did not decrease (no unrelated deletions).\n"
       "8. Repeat with sick leave event — same verification.",
       "Vacation/sick leave events for Employee A do NOT delete\n"
       "Employee B's statistic_report records.\n"
       "Only INITIAL_SYNC event type triggers deleteReportsWithEmploymentChanged().\n"
       "VACATION_CHANGES and SICK_LEAVE_CHANGES event types skip deletion.\n"
       "Total statistic_report record count stable after absence events.\n"
       "BUG (pre-fix, MR !5203): absence events triggered deletion of\n"
       "unrelated employees' records via deleteReportsWithEmploymentChanged().\n"
       "This was the most critical bug in the caching implementation.",
       "Critical", "Integration",
       "#3337", "Statistics / Cache / Event Type Discrimination",
       "MR !5203: New enum StatisticReportUpdateEventType with "
       "INITIAL_SYNC, VACATION_CHANGES, SICK_LEAVE_CHANGES. "
       "deleteReportsWithEmploymentChanged() only on INITIAL_SYNC. "
       "QA Bug #3 / dev comment by Quyen Nguyen."),

    # ── Frontend Resilience (2 cases) ─────────────────────────────

    tc("TC-STAT-136",
       "Frontend: Empty/null statistics response clears loading state",
       "Employee with NO statistic_report records for a specific month.\n"
       "Find months with no data:\n"
       "  SELECT DISTINCT year, month FROM statistic_report\n"
       "  ORDER BY year, month;\n"
       "  -- Pick a month NOT in this list, or use a month far in the future.",
       "1. Navigate to Statistics > Employee Reports.\n"
       "2. Select a reporting period with data — verify table loads normally.\n"
       "3. Switch to a period with NO data (e.g., far future month).\n"
       "4. Verify:\n"
       "   a. Loading spinner appears briefly.\n"
       "   b. Loading spinner disappears (not infinite).\n"
       "   c. Table shows empty state (no rows, or 'No data' message).\n"
       "   d. No stale data from previous period displayed.\n"
       "5. Switch back to a period with data — verify table loads correctly.\n"
       "6. Rapidly switch between periods (stress test saga cancellation).\n"
       "7. Verify no JavaScript console errors.",
       "Empty/null response: loading state cleared, table shows empty.\n"
       "No infinite loading spinner.\n"
       "No stale data from previous period retained.\n"
       "Period switching: previous request cancelled, new data loaded.\n"
       "No console errors or uncaught promise rejections.\n"
       "BUG (pre-fix, MR !5194): empty response left isLoadingReports=true\n"
       "(infinite spinner) and stale reports[] data displayed.",
       "High", "Functional",
       "#3337", "Statistics / Frontend / Redux Reducer",
       "MR !5194: On empty/null response, dispatches fetchReportsFailureAction(). "
       "Reducer now clears state.reports = [] on failure. "
       "QA Bug #1 from omaksimova (infinite loading on out-of-employment period)."),

    tc("TC-STAT-137",
       "Frontend: Rapid period switching cancels pending requests (saga cancellation)",
       "Logged-in user with access to Statistics > Employee Reports.\n"
       "Multiple reporting periods with data available.",
       "1. Navigate to Statistics > Employee Reports.\n"
       "2. Select period Jan 2026 — wait for table to start loading.\n"
       "3. IMMEDIATELY switch to Feb 2026 (before Jan data arrives).\n"
       "4. IMMEDIATELY switch to Mar 2026 (before Feb data arrives).\n"
       "5. Wait for final request to complete.\n"
       "6. Verify:\n"
       "   a. Table shows March 2026 data (the last requested period).\n"
       "   b. No mixing of Jan/Feb/Mar data in the table.\n"
       "   c. No duplicate rows from multiple responses.\n"
       "   d. Loading indicator properly reflects final state.\n"
       "7. Open browser DevTools > Network tab.\n"
       "8. Repeat the rapid switching.\n"
       "9. Verify cancelled requests show as 'cancelled' in Network tab.\n"
       "10. Check browser console for errors.",
       "Final table shows ONLY the last requested period data.\n"
       "Previous requests are cancelled (visible in Network tab).\n"
       "No data mixing between periods.\n"
       "No duplicate rows or partial data.\n"
       "Redux saga uses task.isRunning() check before starting new fetch.\n"
       "Map<string, Task> properly tracks and cancels previous tasks.\n"
       "No unhandled promise rejections from cancelled requests.",
       "Medium", "Functional",
       "#3337", "Statistics / Frontend / Saga Cancellation",
       "MR !5155: Rewrote sagas with task.isRunning() and Map<string, Task> "
       "for improved cancellation. Previous implementation used raw objects "
       "and was prone to race conditions on rapid period switching."),

    # ── Cron Sync (1 case) ────────────────────────────────────────

    tc("TC-STAT-138",
       "Daily cron sync (4am) correctly populates statistic_report for current + previous month",
       "statistic_report table exists with some records.\n"
       "At least one employee has reported hours in current month.\n"
       "At least one employee has a vacation in previous month.\n"
       "Use timemachine env (can trigger sync manually).",
       "1. DB: Record current statistic_report state:\n"
       "   SELECT employee_login, year, month, reported_effort, month_norm,\n"
       "          budget_norm\n"
       "   FROM statistic_report\n"
       "   WHERE (year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "     AND month = EXTRACT(MONTH FROM CURRENT_DATE))\n"
       "   OR (year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "     AND month = EXTRACT(MONTH FROM CURRENT_DATE) - 1)\n"
       "   ORDER BY employee_login, year, month;\n"
       "2. Trigger sync via test endpoint:\n"
       "   POST /api/ttt/v1/test/trigger-optimized-statistic-report-sync\n"
       "   (or POST /api/ttt/v1/test/statistic-reports for direct sync).\n"
       "3. Wait for sync to complete (may take 30-60 seconds).\n"
       "4. Re-query statistic_report for current + previous month.\n"
       "5. Verify:\n"
       "   a. All active employees have records for both months.\n"
       "   b. reported_effort matches actual task reports:\n"
       "      SELECT employee_id, SUM(effort) FROM task_report\n"
       "      WHERE EXTRACT(YEAR FROM report_date) = {Y}\n"
       "        AND EXTRACT(MONTH FROM report_date) = {M}\n"
       "      GROUP BY employee_id;\n"
       "   c. month_norm reflects effective bounds (prorated for partial months).\n"
       "   d. budget_norm includes admin vacation hours.\n"
       "6. Check ShedLock table for scheduled job state:\n"
       "   SELECT * FROM shedlock WHERE name LIKE '%statistic%';",
       "Cron sync populates statistic_report for current AND previous month.\n"
       "All active employees have records for both months.\n"
       "reported_effort matches SUM(task_report.effort) for the month.\n"
       "month_norm reflects effective bounds (not full calendar month).\n"
       "budget_norm includes admin vacation hours where applicable.\n"
       "ShedLock prevents concurrent execution.\n"
       "Cron schedule: 0 0 4 * * ? (daily at 4:00 AM).\n"
       "StatisticReportSyncLauncher syncs previous + current month.\n"
       "Uses StatisticReportSyncService.saveMonthNormAndReportedEffortForEmployees()\n"
       "with eventType=INITIAL_SYNC (triggers deleteReportsWithEmploymentChanged).",
       "High", "Integration",
       "#3337", "Statistics / Sync / Cron Job",
       "MR !5013: StatisticReportScheduler at 0 0 4 * * ? with ShedLock. "
       "StatisticReportSyncLauncher runs for previous and current month. "
       "Uses INITIAL_SYNC event type which DOES trigger employment change deletion. "
       "Test endpoint: POST /v1/test/statistic-reports."),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Modify existing workbook
# ══════════════════════════════════════════════════════════════════

def main():
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "statistics.xlsx")
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. Remove existing tab if present (for re-runs) ──────────
    tab_name = "TS-STAT-CacheArch"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
        print(f"  Removed existing '{tab_name}' tab (re-run)")

    # ── 2. Create the new sheet ──────────────────────────────────
    if "Test Data" in wb.sheetnames:
        ws = wb.create_sheet(title=tab_name, index=wb.sheetnames.index("Test Data"))
    else:
        ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    case_count = write_ts_tab(ws, "Caching Architecture & Performance (#3337)", CACHE_ARCH_CASES)
    print(f"  Created '{tab_name}' with {case_count} test cases (TC-STAT-127 to TC-STAT-138)")

    # ── 3. Update Plan Overview ──────────────────────────────────
    ws_plan = wb["Plan Overview"]

    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Total Test Cases" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_val = 126 + case_count  # 126 existing + 12 new = 138
            ws_plan.cell(row=r, column=2, value=str(new_val))
            print(f"  Updated Plan Overview: Total Test Cases {old_val} -> {new_val}")
            break

    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Test Suites" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_suites = int(old_val) + 1 if old_val else 9
            ws_plan.cell(row=r, column=2, value=str(new_suites))
            print(f"  Updated Plan Overview: Test Suites {old_val} -> {new_suites}")
            break

    # Find last hyperlink row in Plan Overview and add new one after it
    last_link_row = 0
    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.hyperlink and "TS-" in str(cell.hyperlink.target or ""):
            last_link_row = r

    if last_link_row > 0:
        new_link_row = last_link_row + 1
        max_row = ws_plan.max_row
        for r in range(max_row, new_link_row - 1, -1):
            for c in range(1, 6):
                src = ws_plan.cell(row=r, column=c)
                dst = ws_plan.cell(row=r + 1, column=c)
                dst.value = src.value
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.alignment = copy(src.alignment)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                if src.hyperlink:
                    dst.hyperlink = src.hyperlink.target
                src.value = None
                src.hyperlink = None

        link_cell = ws_plan.cell(row=new_link_row, column=1,
                                  value=f"Caching Architecture & Performance (#3337) — {case_count} cases")
        link_cell.font = FONT_LINK_BOLD
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        print(f"  Added hyperlink at Plan Overview row {new_link_row}")

    # ── 4. Update Feature Matrix ─────────────────────────────────
    ws_fm = wb["Feature Matrix"]

    total_row = None
    for r in range(1, ws_fm.max_row + 1):
        if ws_fm.cell(row=r, column=1).value == "TOTAL":
            total_row = r
            break

    if total_row:
        new_fm_row = total_row
        for c in range(1, 8):
            src = ws_fm.cell(row=total_row, column=c)
            dst = ws_fm.cell(row=total_row + 1, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
            if src.hyperlink:
                dst.hyperlink = src.hyperlink.target

        # 2 API, 4 integration, 3 security, 2 functional, 1 integration-cron
        fm_data = ["Caching Architecture", 2, 6, 0, 3, case_count, f"TS-STAT-CacheArch"]
        fill_row = FILL_ROW_EVEN if (new_fm_row - 3) % 2 == 0 else FILL_ROW_ODD
        for col, val in enumerate(fm_data, 1):
            cell = ws_fm.cell(row=new_fm_row, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            cell.fill = fill_row
        link_cell = ws_fm.cell(row=new_fm_row, column=7)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = f"#'{tab_name}'!A1"

        new_total_row = total_row + 1
        new_total = 126 + case_count
        ws_fm.cell(row=new_total_row, column=6, value=new_total)

        print(f"  Updated Feature Matrix: added row {new_fm_row}")
        print(f"  Updated Feature Matrix: TOTAL = {new_total}")

    # ── 5. Save ──────────────────────────────────────────────────
    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")
    print(f"Total: {case_count} test cases added in '{tab_name}'")
    print(f"Workbook now has {126 + case_count} total test cases")


if __name__ == "__main__":
    main()
