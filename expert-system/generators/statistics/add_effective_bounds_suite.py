#!/usr/bin/env python3
"""Add TS-STAT-EffBounds supplementary test suite to statistics.xlsx.

Sprint 15 feature: Effective bounds on individual norm (#3353, #3356, #3381).
- Clamps individual norm calculation to employee's actual work period
- Adds budget norm = individual norm + admin vacation hours
- Affects Statistics employee reports table, My Tasks counter, statistic_report table

Adds 15 test cases (TC-STAT-112 through TC-STAT-126) covering:
  - Effective bounds core (5): hired mid-month, dismissed mid-month, not yet employed,
    re-hired with gap, no work period fallback
  - Budget norm (4): with/without admin vacation, combined with effective bounds, excess calc
  - normForDate edge cases (3): current month progressive, hired after date, dismissed before date
  - Sync paths (3): vacation event, task report, cron sync
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ── Styling constants (match existing workbook exactly) ──────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ──────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab matching existing format."""
    # Back link in A1
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, tc_item in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            tc_item["id"], tc_item["title"], tc_item["preconditions"],
            tc_item["steps"], tc_item["expected"], tc_item["priority"],
            tc_item["type"], tc_item["req_ref"], tc_item["module"],
            tc_item.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# TEST CASE DATA — TS-STAT-EffBounds (15 cases)
# ══════════════════════════════════════════════════════════════════

# Starting from TC-STAT-112 (highest existing is TC-STAT-111)

EFFECTIVE_BOUNDS_CASES = [
    # ── Effective Bounds Core (5 cases) ──────────────────────────

    tc("TC-STAT-112",
       "Effective bounds — employee hired mid-month (prorated personalNorm)",
       "Employee hired mid-month (e.g., Feb 15).\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name, ewp.first_working_day, ewp.last_working_day\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE EXTRACT(DAY FROM ewp.first_working_day) > 1\n"
       "    AND EXTRACT(DAY FROM ewp.first_working_day) <= 20\n"
       "    AND e.enabled = true\n"
       "  ORDER BY ewp.first_working_day DESC LIMIT 5;\n\n"
       "Identify the month containing the hire date.\n"
       "Verify employee_work_period record exists for this employee.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}\n"
       "   where M = hire month.\n"
       "2. Check response field personalNorm (or month_norm in DB).\n"
       "3. Calculate expected: count working days from hire date to end of month\n"
       "   using production calendar, multiply by 8.\n"
       "   E.g., Feb 15-28 in Russian calendar = 10 working days x 8h = 80h.\n"
       "4. DB: SELECT month_norm, budget_norm FROM statistic_report\n"
       "   WHERE employee_id = {id} AND year = {Y} AND month = {M}.\n"
       "5. Navigate to Statistics > Employee Reports for the employee.\n"
       "6. Verify UI displays the prorated norm, not full month norm.",
       "personalNorm is prorated to actual working days within effective bounds.\n"
       "For Feb 15-28 hire: ~80h (not full 160h).\n"
       "statistic_report.month_norm matches the prorated value.\n"
       "UI Employee Reports table shows the clamped norm.\n"
       "My Tasks counter uses the effective-bounded norm for percentage calculation.",
       "High", "Functional",
       "#3353, #3356", "Statistics / Norm Calculation",
       "Core feature: effectiveFrom = max(monthStart, firstWorkingDay). "
       "Verify NormForDateCalculator.getEffectiveBounds() applies correctly."),

    tc("TC-STAT-113",
       "Effective bounds — employee dismissed mid-month (norm capped at dismissal)",
       "Employee dismissed mid-month.\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name, ewp.first_working_day, ewp.last_working_day\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE ewp.last_working_day IS NOT NULL\n"
       "    AND EXTRACT(DAY FROM ewp.last_working_day) < 28\n"
       "    AND EXTRACT(DAY FROM ewp.last_working_day) > 1\n"
       "    AND e.enabled = true\n"
       "  ORDER BY ewp.last_working_day DESC LIMIT 5;\n\n"
       "Identify the month containing the dismissal date.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}\n"
       "   where M = dismissal month.\n"
       "2. Check personalNorm value.\n"
       "3. Calculate expected: count working days from month start to dismissal date\n"
       "   using production calendar, multiply by 8.\n"
       "4. DB: SELECT month_norm FROM statistic_report\n"
       "   WHERE employee_id = {id} AND year = {Y} AND month = {M}.\n"
       "5. Compare month_norm with full-month norm for that office.\n"
       "6. Verify month_norm < full-month norm.",
       "personalNorm covers only working days up to dismissal date.\n"
       "statistic_report.month_norm reflects the truncated range.\n"
       "E.g., if dismissed Feb 14: norm = working days Feb 1-14 x 8h.\n"
       "Norm is strictly less than full month norm for the office.\n"
       "effectiveTo = min(monthEnd, lastWorkingDay) applied correctly.",
       "High", "Functional",
       "#3353, #3356", "Statistics / Norm Calculation",
       "effectiveTo = min(monthEnd, lastWorkingDay). Verify last_working_day "
       "from employee_work_period is used, not employee.dismissal_date directly."),

    tc("TC-STAT-114",
       "Effective bounds — employee not yet employed for target month (norm = 0)",
       "Employee hired in a future month relative to the queried month.\n"
       "Find test data:\n"
       "  SELECT e.login, ewp.first_working_day\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE ewp.first_working_day > CURRENT_DATE\n"
       "    AND e.enabled = true\n"
       "  LIMIT 5;\n"
       "If none found, use an employee hired this month and query the previous month.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}\n"
       "   where the month is BEFORE the employee's first_working_day.\n"
       "2. Check personalNorm value.\n"
       "3. DB: SELECT month_norm FROM statistic_report\n"
       "   WHERE employee_id = {id} AND year = {Y} AND month = {M}.\n"
       "4. Verify norm is 0 or that no record exists for that month.\n"
       "5. Check that the API does not return an error (graceful handling).",
       "personalNorm = 0 for months before employee's first working day.\n"
       "API returns norm=0 (not an error or null).\n"
       "No statistic_report record may exist, OR record exists with month_norm=0.\n"
       "Effective bounds: effectiveFrom > monthEnd means no overlap, norm=0.",
       "High", "Functional",
       "#3353, #3356", "Statistics / Norm Calculation",
       "Edge case: when effectiveFrom > monthEnd, the date range is empty. "
       "Verify no division-by-zero or NPE in norm calculation."),

    tc("TC-STAT-115",
       "Effective bounds — re-hired employee with gap in employment",
       "Employee with multiple work periods (gap in employment).\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name, COUNT(*) AS period_count\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE e.enabled = true\n"
       "  GROUP BY e.id\n"
       "  HAVING COUNT(*) > 1\n"
       "  LIMIT 5;\n\n"
       "Then: SELECT * FROM employee_work_period\n"
       "WHERE employee_id = {id} ORDER BY first_working_day;\n"
       "Identify the gap month(s) and the re-hire month.",
       "1. For the re-hire month, GET /api/ttt/v1/statistic/report for the employee.\n"
       "2. Check which work period is used for norm calculation.\n"
       "3. For a month in the gap (between periods), query the same endpoint.\n"
       "4. DB: SELECT month_norm FROM statistic_report for gap month.\n"
       "5. For the re-hire month, verify norm is prorated from the new first_working_day.\n"
       "6. Verify gap months have norm = 0.",
       "Re-hire month: personalNorm prorated from new first_working_day to month end.\n"
       "Gap months: personalNorm = 0 (no active work period).\n"
       "Most recent matching work period used for current calculations.\n"
       "Old work period does not bleed into gap months.",
       "Medium", "Functional",
       "#3353, #3356", "Statistics / Norm Calculation",
       "Multiple employee_work_period records. The system should match "
       "the work period that overlaps with the target month. Verify no "
       "period confusion when multiple records exist."),

    tc("TC-STAT-116",
       "Effective bounds — no work period data (fallback to full month)",
       "Employee with no records in employee_work_period table.\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name\n"
       "  FROM employee e\n"
       "  LEFT JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE e.enabled = true AND ewp.id IS NULL\n"
       "  LIMIT 5;\n\n"
       "These employees may exist if work period sync has not run or data is missing.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}.\n"
       "2. Check personalNorm value.\n"
       "3. Compare with the full month norm for the employee's office.\n"
       "4. DB: SELECT month_norm FROM statistic_report for the employee/month.\n"
       "5. Verify the system does not crash or return an error.\n"
       "6. Check application logs for warnings about missing work period.",
       "System uses full month range as fallback when no work period data exists.\n"
       "personalNorm = full month norm for the employee's office/calendar.\n"
       "No crash, NPE, or 500 error.\n"
       "Behavior matches pre-#3353 behavior (backward compatible).\n"
       "Optional: warning logged about missing work period.",
       "Medium", "Error Handling",
       "#3353", "Statistics / Norm Calculation",
       "Fallback behavior: if employee_work_period is empty, effectiveFrom = monthStart "
       "and effectiveTo = monthEnd. Verify graceful degradation."),

    # ── Budget Norm (4 cases) ────────────────────────────────────

    tc("TC-STAT-117",
       "Budget norm — employee with admin vacation (budgetNorm > personalNorm)",
       "Employee who had administrative vacation (leave type = ADMIN) during the target month.\n"
       "Find test data:\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.budget_norm,\n"
       "         sr.year, sr.month\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.budget_norm > sr.month_norm\n"
       "    AND sr.budget_norm > 0\n"
       "  ORDER BY sr.year DESC, sr.month DESC\n"
       "  LIMIT 5;\n\n"
       "Also verify admin vacation exists:\n"
       "  SELECT * FROM vacation\n"
       "  WHERE employee_id = {id}\n"
       "    AND type = 'ADMINISTRATIVE'\n"
       "    AND EXTRACT(YEAR FROM start_date) = {Y}\n"
       "    AND EXTRACT(MONTH FROM start_date) = {M};",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}.\n"
       "2. Check both personalNorm and budgetNorm fields in response.\n"
       "3. Verify budgetNorm = personalNorm + (admin_vacation_hours).\n"
       "4. Navigate to Statistics > Employee Reports for the employee/month.\n"
       "5. Check UI display format: should show '{personalNorm} ({budgetNorm})'.\n"
       "6. Verify the excess percentage uses budgetNorm as denominator.",
       "budgetNorm > personalNorm (includes admin vacation hours).\n"
       "UI shows dual display: e.g., '144 (160)' where 144 = personalNorm, 160 = budgetNorm.\n"
       "Budget norm = personalNorm + admin_vacation_working_days * 8.\n"
       "Excess calculation: (reported - budgetNorm) / budgetNorm * 100.",
       "High", "Functional",
       "#3381", "Statistics / Norm Calculation",
       "budgetNorm is a new field added by #3381. Verify it appears in both "
       "API response and UI. The dual display format is specific to employee reports."),

    tc("TC-STAT-118",
       "Budget norm — no admin vacation (budgetNorm equals personalNorm)",
       "Employee with no administrative vacation in the target month.\n"
       "Find test data:\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.budget_norm,\n"
       "         sr.year, sr.month\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.budget_norm = sr.month_norm\n"
       "    AND sr.month_norm > 0\n"
       "    AND e.enabled = true\n"
       "  ORDER BY sr.year DESC, sr.month DESC\n"
       "  LIMIT 5;",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}.\n"
       "2. Check personalNorm and budgetNorm fields.\n"
       "3. Verify budgetNorm = personalNorm.\n"
       "4. Navigate to Statistics > Employee Reports for the employee.\n"
       "5. Check UI display: should show single number (no parenthetical).\n"
       "6. DB: Confirm no ADMINISTRATIVE vacation for that employee/month.",
       "budgetNorm equals personalNorm when no admin vacation exists.\n"
       "UI shows single norm value (no dual display, no parentheses).\n"
       "E.g., just '160' not '160 (160)'.\n"
       "Excess calculation: (reported - personalNorm) / personalNorm * 100.",
       "Medium", "Functional",
       "#3381", "Statistics / Norm Calculation",
       "When budgetNorm = personalNorm, UI should not show redundant parenthetical. "
       "Verify frontend conditionally renders the dual format."),

    tc("TC-STAT-119",
       "Budget norm + effective bounds — partial-month employee with admin vacation",
       "Employee hired mid-month who also has admin vacation in that same month.\n"
       "Find test data:\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.budget_norm,\n"
       "         ewp.first_working_day\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE sr.budget_norm > sr.month_norm\n"
       "    AND sr.month_norm < 160\n"
       "    AND EXTRACT(DAY FROM ewp.first_working_day) > 1\n"
       "  ORDER BY sr.year DESC, sr.month DESC\n"
       "  LIMIT 5;\n\n"
       "If none found, this is a manufactured scenario: use timemachine clock to create it.",
       "1. DB: Get employee's first_working_day and admin vacation dates for the month.\n"
       "2. Calculate expected personalNorm: working days within [firstWorkingDay, monthEnd] * 8.\n"
       "3. Calculate expected budgetNorm: personalNorm + admin_vacation_hours_within_bounds.\n"
       "4. GET /api/ttt/v1/statistic/report for the employee/month.\n"
       "5. Compare API personalNorm with expected.\n"
       "6. Compare API budgetNorm with expected.\n"
       "7. Verify admin vacation hours are clamped to within effective bounds\n"
       "   (not the full month).",
       "Both effective bounds AND admin vacation exclusion apply correctly.\n"
       "personalNorm: prorated to actual work period (< full month).\n"
       "budgetNorm: personalNorm + admin vacation hours WITHIN the effective bounds.\n"
       "Admin vacation days outside the effective bounds are NOT counted.\n"
       "E.g., hired Feb 15, admin vacation Feb 10-12 (before hire) should not add to budget.",
       "High", "Functional",
       "#3353, #3381", "Statistics / Norm Calculation",
       "Interaction test: effective bounds clamping intersected with admin vacation "
       "period. The admin vacation hours should only count for days within the "
       "employee's effective work period, not the full calendar month."),

    tc("TC-STAT-120",
       "Excess calculation uses budgetNorm as denominator",
       "Employee with budgetNorm > personalNorm (has admin vacation).\n"
       "Employee has reported hours for the month.\n"
       "Reuse employee from TC-STAT-117.\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.budget_norm,\n"
       "         sr.month_hours, sr.excess\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.budget_norm > sr.month_norm\n"
       "    AND sr.month_hours > 0\n"
       "  ORDER BY sr.year DESC, sr.month DESC\n"
       "  LIMIT 5;",
       "1. DB: Get month_hours, month_norm, budget_norm, excess for the employee.\n"
       "2. Calculate: expected_excess = (month_hours - budget_norm) / budget_norm * 100.\n"
       "3. Compare with stored excess value.\n"
       "4. Also calculate incorrect version: (month_hours - month_norm) / month_norm * 100.\n"
       "5. Verify stored excess matches budgetNorm-based calculation, not personalNorm-based.\n"
       "6. Navigate to Statistics > Employee Reports and check displayed excess percentage.",
       "excess = (reported - budgetNorm) / budgetNorm * 100.\n"
       "NOT: (reported - personalNorm) / personalNorm * 100.\n"
       "The denominator is budgetNorm, which includes admin vacation hours.\n"
       "This means employees with admin vacation show lower excess percentages.\n"
       "UI excess display matches the DB value.",
       "High", "Functional",
       "#3381", "Statistics / Norm Calculation",
       "Critical: if excess still uses personalNorm as denominator, admin vacation "
       "employees will show artificially high excess. Verify the formula change."),

    # ── normForDate Edge Cases (3 cases) ─────────────────────────

    tc("TC-STAT-121",
       "normForDate — current month progressive accumulation within effective bounds",
       "Employee currently active (not dismissed, hired before current month).\n"
       "Current date is mid-month (e.g., March 15).\n"
       "Employee hired before this month (full effective bounds for current month).",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}\n"
       "   where M = current month.\n"
       "2. Note the personalNorm (full month) and normForDate (progressive).\n"
       "3. Calculate expected normForDate: working days from month start to today * 8.\n"
       "4. Verify normForDate <= personalNorm.\n"
       "5. Check the next day: normForDate should increase by 8 (if working day)\n"
       "   or stay the same (if weekend/holiday).\n"
       "6. Verify normForDate is clamped within effective bounds:\n"
       "   max(monthStart, firstWorkingDay) to min(today, lastWorkingDay).",
       "normForDate accumulates day by day within effective bounds.\n"
       "On March 15: normForDate = working days [monthStart..March 15] * 8.\n"
       "normForDate <= personalNorm always.\n"
       "normForDate increases on working days, stays flat on weekends/holidays.\n"
       "Effective bounds applied: normForDate never exceeds what the clamped range allows.",
       "Medium", "Functional",
       "#3353, #3356", "Statistics / Norm Calculation",
       "normForDate = progressive norm up to current date. Used for 'current status' "
       "indicators. Verify it respects effective bounds, not just calendar month."),

    tc("TC-STAT-122",
       "normForDate — employee hired after the selected date (norm = 0)",
       "Employee hired mid-month (e.g., Feb 15).\n"
       "Query normForDate for a date BEFORE the hire date (e.g., Feb 10).\n"
       "Reuse employee from TC-STAT-112.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}&date={D}\n"
       "   where D = a date before the employee's first_working_day within the same month.\n"
       "   (If API does not accept date param, check via DB or timemachine clock.)\n"
       "2. DB: SELECT * FROM statistic_report WHERE employee_id = {id}\n"
       "   AND year = {Y} AND month = {M}.\n"
       "3. Calculate expected normForDate: effective range = [firstWorkingDay, date].\n"
       "   Since date < firstWorkingDay, effective range is empty.\n"
       "4. Verify normForDate = 0.\n"
       "5. Check that personalNorm (full month value) is still correctly prorated.",
       "normForDate = 0 when the selected date is before the employee's first working day.\n"
       "Effective range is empty: max(monthStart, firstWorkingDay) > min(date, monthEnd).\n"
       "personalNorm is still correctly prorated for the full month.\n"
       "No error or negative value returned.",
       "Medium", "Boundary",
       "#3353, #3356", "Statistics / Norm Calculation",
       "Tests the case where the normForDate range collapses to empty. "
       "The progressive norm should be 0, not negative or full-month."),

    tc("TC-STAT-123",
       "normForDate — employee dismissed before the selected date (norm capped at dismissal)",
       "Employee dismissed mid-month (e.g., Feb 10).\n"
       "Query normForDate for a date AFTER the dismissal (e.g., Feb 15).\n"
       "Reuse employee from TC-STAT-113.",
       "1. GET /api/ttt/v1/statistic/report?employeeId={id}&year={Y}&month={M}\n"
       "   for the dismissal month, using a date after dismissal.\n"
       "2. DB: SELECT month_norm FROM statistic_report\n"
       "   WHERE employee_id = {id} AND year = {Y} AND month = {M}.\n"
       "3. Calculate expected normForDate:\n"
       "   effective range = [monthStart, min(date, lastWorkingDay)].\n"
       "   Since date > lastWorkingDay, clamped to [monthStart, lastWorkingDay].\n"
       "4. Verify normForDate = personalNorm (fully accumulated up to dismissal).\n"
       "5. Verify normForDate does NOT continue growing after dismissal date.",
       "normForDate is capped at the dismissal date.\n"
       "normForDate = personalNorm once the selected date passes lastWorkingDay.\n"
       "No additional hours accumulate after dismissal.\n"
       "E.g., Feb 10 dismissal: normForDate on Feb 15 = same as normForDate on Feb 10.\n"
       "effectiveTo = min(date, lastWorkingDay) correctly applied.",
       "Medium", "Boundary",
       "#3353, #3356", "Statistics / Norm Calculation",
       "After dismissal, normForDate should plateau at personalNorm. "
       "Verify the progressive calculation stops at lastWorkingDay."),

    # ── Sync Paths (3 cases) ─────────────────────────────────────

    tc("TC-STAT-124",
       "Sync path — norm updated via vacation event (create/modify vacation)",
       "Employee with existing statistic_report for current month.\n"
       "No admin vacation yet for this month.\n"
       "Find test data:\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.budget_norm\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.month_norm = sr.budget_norm\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "    AND e.enabled = true\n"
       "  LIMIT 5;",
       "1. DB: Record current month_norm and budget_norm for the employee.\n"
       "2. Create an admin vacation for this employee in the current month:\n"
       "   POST /api/vacation/v1/vacation/create (type=ADMINISTRATIVE, 2 working days).\n"
       "3. Wait for event propagation (MQ message: VACATION_CREATED).\n"
       "4. DB: Re-query statistic_report for the employee/month.\n"
       "5. Verify budget_norm increased by 16 (2 days * 8h).\n"
       "6. Verify month_norm unchanged (effective bounds not affected by vacation).\n"
       "7. Check that effective bounds are still correctly applied to month_norm.",
       "After vacation creation:\n"
       "  - budget_norm = previous_budget_norm + 16 (2 admin days * 8h).\n"
       "  - month_norm unchanged (effective bounds clamping unaffected).\n"
       "  - Event-driven update path correctly recalculates budget_norm.\n"
       "  - Effective bounds still applied to month_norm in the recalculation.\n"
       "If effective bounds are NOT applied during event-driven recalc,\n"
       "month_norm may revert to full-month value (regression).",
       "High", "Integration",
       "#3353, #3381", "Statistics / Sync",
       "Tests the MQ event-driven update path (StatisticReportService.onVacationEvent). "
       "Effective bounds must be applied during event-triggered recalculation, "
       "not just during cron sync. Requires allow_api_mutations=true or manual action."),

    tc("TC-STAT-125",
       "Sync path — norm updated via task report submission",
       "Employee with existing statistic_report for current month.\n"
       "Employee has fewer reported hours than norm.\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm, sr.month_hours\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.month_hours < sr.month_norm\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "    AND e.enabled = true\n"
       "  LIMIT 5;",
       "1. DB: Record current month_norm, budget_norm, month_hours for the employee.\n"
       "2. Submit a time report for the employee (8h on a working day):\n"
       "   POST /api/ttt/v1/report (or via UI).\n"
       "3. Wait for statistic_report update (event-driven or poll DB).\n"
       "4. DB: Re-query statistic_report.\n"
       "5. Verify month_hours increased by 8.\n"
       "6. Verify month_norm still reflects effective bounds (not recalculated to full month).\n"
       "7. Verify excess recalculated with updated hours but same norm values.",
       "After task report submission:\n"
       "  - month_hours increased by 8.\n"
       "  - month_norm unchanged (effective bounds still applied).\n"
       "  - budget_norm unchanged.\n"
       "  - excess recalculated: (new_month_hours - budget_norm) / budget_norm * 100.\n"
       "If month_norm reverts to full-month during report-triggered recalc,\n"
       "this is a regression in effective bounds.",
       "Medium", "Integration",
       "#3353", "Statistics / Sync",
       "Tests the task-report-driven update path. The critical check is that "
       "month_norm is NOT recalculated to full-month value during this path. "
       "Requires allow_api_mutations=true or manual action."),

    tc("TC-STAT-126",
       "Sync path — norm updated via daily cron sync (04:00 job)",
       "Employees with statistic_report records.\n"
       "At least one employee hired mid-month (for effective bounds verification).\n"
       "At least one employee with admin vacation (for budget norm verification).\n"
       "  -- Mid-month hire:\n"
       "  SELECT sr.employee_id, e.login, sr.month_norm\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE EXTRACT(DAY FROM ewp.first_working_day) > 1\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "  LIMIT 3;\n"
       "  -- Admin vacation:\n"
       "  SELECT sr.employee_id, e.login, sr.budget_norm\n"
       "  FROM statistic_report sr\n"
       "  JOIN employee e ON sr.employee_id = e.id\n"
       "  WHERE sr.budget_norm > sr.month_norm\n"
       "    AND sr.year = EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "    AND sr.month = EXTRACT(MONTH FROM CURRENT_DATE)\n"
       "  LIMIT 3;",
       "1. DB: Record month_norm and budget_norm for selected employees.\n"
       "2. Trigger the nightly sync manually:\n"
       "   POST /v1/test/trigger-optimized-statistic-report-sync\n"
       "   (or wait for the 04:00 cron job).\n"
       "3. DB: Re-query statistic_report for the same employees.\n"
       "4. For mid-month hire: verify month_norm is still prorated (not full month).\n"
       "5. For admin vacation: verify budget_norm still includes admin vacation hours.\n"
       "6. Verify both previous month and current month records are correct.\n"
       "7. Check that the sync did not reset effective bounds.",
       "After daily cron sync:\n"
       "  - Mid-month hires: month_norm remains prorated (effective bounds preserved).\n"
       "  - Admin vacation employees: budget_norm remains > month_norm.\n"
       "  - Previous month records: fully finalized with correct effective bounds.\n"
       "  - Current month records: progressive normForDate correctly bounded.\n"
       "Cron sync applies the same effective bounds logic as event-driven updates.\n"
       "No regression: norms not reset to full-month values.",
       "High", "Integration",
       "#3353, #3356, #3381", "Statistics / Sync",
       "Tests the nightly cron sync path (OptimizedStatisticReportSyncJob). "
       "All three norm components (personalNorm, budgetNorm, normForDate) "
       "must respect effective bounds after sync. This is the most common "
       "recalculation trigger."),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Modify existing workbook
# ══════════════════════════════════════════════════════════════════

def main():
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "statistics.xlsx")
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. Remove existing TS-STAT-EffBounds if present (for re-runs) ──
    tab_name = "TS-STAT-EffBounds"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
        print(f"  Removed existing '{tab_name}' tab (re-run)")

    # ── 2. Create the new TS-STAT-EffBounds sheet ──────────────────────
    # Append as the last tab (or before Test Data if it exists)
    if "Test Data" in wb.sheetnames:
        ws = wb.create_sheet(title=tab_name, index=wb.sheetnames.index("Test Data"))
    else:
        ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    case_count = write_ts_tab(ws, "Effective Bounds on Individual Norm (#3353, #3356, #3381)", EFFECTIVE_BOUNDS_CASES)
    print(f"  Created '{tab_name}' with {case_count} test cases (TC-STAT-112 to TC-STAT-126)")

    # ── 3. Update Plan Overview ────────────────────────────────────────
    ws_plan = wb["Plan Overview"]

    # Update total test case count (row 7)
    old_total = ws_plan.cell(row=7, column=2).value
    new_total = 111 + case_count  # 111 existing + 15 new
    ws_plan.cell(row=7, column=2, value=str(new_total))
    print(f"  Updated Plan Overview: Total Test Cases {old_total} -> {new_total}")

    # Update test suite count (row 8)
    old_suites = ws_plan.cell(row=8, column=2).value
    ws_plan.cell(row=8, column=2, value="8")
    print(f"  Updated Plan Overview: Test Suites {old_suites} -> 8")

    # Add hyperlink for new suite after existing suite links (row 29)
    # Existing suites are rows 22-28; add at row 29
    new_link_row = 29
    # Shift down existing content from row 29 onward
    max_row = ws_plan.max_row
    for r in range(max_row, new_link_row - 1, -1):
        for c in range(1, 6):
            src = ws_plan.cell(row=r, column=c)
            dst = ws_plan.cell(row=r + 1, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
            if src.hyperlink:
                dst.hyperlink = src.hyperlink.target
            src.value = None
            src.hyperlink = None

    # Write the new hyperlink at row 29
    link_cell = ws_plan.cell(row=new_link_row, column=1,
                              value=f"Effective Bounds on Individual Norm — {case_count} cases")
    link_cell.font = FONT_LINK_BOLD
    link_cell.hyperlink = f"#'{tab_name}'!A1"
    print(f"  Added hyperlink at Plan Overview row {new_link_row}")

    # Also update the "Test Data Reference" hyperlink (shifted down by 1)
    for r in range(30, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Test Data Reference" in str(cell.value):
            cell.hyperlink = "#'Test Data'!A1"
            break

    # ── 4. Update Feature Matrix ───────────────────────────────────────
    ws_fm = wb["Feature Matrix"]

    # Find the TOTAL row
    total_row = None
    for r in range(1, ws_fm.max_row + 1):
        if ws_fm.cell(row=r, column=1).value == "TOTAL":
            total_row = r
            break

    if total_row:
        # Insert a new row before TOTAL
        new_fm_row = total_row
        # Shift TOTAL row down by 1
        for c in range(1, 8):
            src = ws_fm.cell(row=total_row, column=c)
            dst = ws_fm.cell(row=total_row + 1, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
            if src.hyperlink:
                dst.hyperlink = src.hyperlink.target

        # Write new row for Effective Bounds
        fm_data = ["Effective Bounds", 0, 15, 0, 0, 15, f"TS-STAT-EffBounds"]
        fill_row = FILL_ROW_EVEN if (new_fm_row - 3) % 2 == 0 else FILL_ROW_ODD
        for col, val in enumerate(fm_data, 1):
            cell = ws_fm.cell(row=new_fm_row, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            cell.fill = fill_row
        # Add hyperlink to suite link column
        link_cell = ws_fm.cell(row=new_fm_row, column=7)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = f"#'{tab_name}'!A1"

        # Update TOTAL row values
        new_total_row = total_row + 1
        ws_fm.cell(row=new_total_row, column=2, value=59)      # UI tests unchanged
        ws_fm.cell(row=new_total_row, column=3, value=24 + 15)  # API tests + 15
        ws_fm.cell(row=new_total_row, column=4, value=19)       # Data tests unchanged
        ws_fm.cell(row=new_total_row, column=5, value=9)        # Security unchanged
        ws_fm.cell(row=new_total_row, column=6, value=new_total) # Grand total

        print(f"  Updated Feature Matrix: added row {new_fm_row}, shifted TOTAL to row {new_total_row}")
        print(f"  Updated Feature Matrix: TOTAL = {new_total}")

    # ── 5. Save ────────────────────────────────────────────────────────
    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")
    print(f"Total: {case_count} test cases added in '{tab_name}'")
    print(f"Workbook now has {new_total} total test cases across 8 suites")


if __name__ == "__main__":
    main()
