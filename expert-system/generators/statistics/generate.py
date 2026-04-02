#!/usr/bin/env python3
"""
Statistics Module Test Documentation Generator — Phase B
Generates test-docs/statistics/statistics.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 8 TS- test suite tabs.

Based on vault knowledge: frontend-statistics-module.md (13 tabs, 2 sub-systems),
statistics-service-implementation.md (cache table, norm calc, excess detection),
statistics-ui-deep-exploration.md (live UI verification), statistics-ticket-findings.md
(180+ tickets mined), REQ-statistics-employee-reports.md (Confluence spec),
statistics-caffeine-caching-performance-3337.md (materialized view rework).

8 Suites, ~82 test cases:
  TS-Stat-ClassicGeneral    — General Statistics page (tabs, filters, tree/flat, search)
  TS-Stat-EmployeeReports   — Employee Reports page (search, norm, excess, comments)
  TS-Stat-NormExcess        — Norm calculation edge cases and excess display rules
  TS-Stat-ExportWSR         — CSV export, copy/link features, WSR tree view
  TS-Stat-Permissions       — Role-based tab visibility and data scoping
  TS-Stat-HourSum           — Hour sum consistency (parent vs child, showFired)
  TS-Stat-CacheSync         — statistic_report cache table correctness
  TS-Stat-Regression        — Bug regression tests from 180+ tickets
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "statistics")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "statistics.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_classic_general_cases():
    """TS-Stat-ClassicGeneral: General Statistics page — tabs, filters, tree/flat, search."""
    return [
        {
            "id": "TC-STAT-001", "title": "Tab visibility — EMPLOYEE-only user sees 1 tab (My Tasks)",
            "preconditions": "User with EMPLOYEE role only (no manager/accountant/admin roles).\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name != 'ROLE_EMPLOYEE')) AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the EMPLOYEE-only user\n2. Navigate to /statistics/general\n3. Verify only 1 tab is visible: 'My tasks'\n4. Verify no other tabs appear (My projects, Department, Office, etc.)\n5. Verify the tab content loads with the employee's own task data\n6. Verify the Statistics menu has NO submenu (direct link, not dropdown)",
            "expected": "EMPLOYEE-only sees exactly 1 tab: 'My tasks'. Menu is a direct link to /statistics/general, not a dropdown.",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Tab Visibility Matrix", "module": "statistics/classic",
            "notes": "Verified live: alsmirnov (EMPLOYEE-only) sees 1 tab. Permissions: VIEW_MY_TASKS only."
        },
        {
            "id": "TC-STAT-002", "title": "Tab visibility — multi-role user sees 8+ tabs",
            "preconditions": "User with multiple roles (DM + TL + PM + accountant or similar).\nQuery: SELECT be.login, COUNT(DISTINCT r.name) AS role_count FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE be.enabled = true GROUP BY be.login HAVING COUNT(DISTINCT r.name) >= 5 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the multi-role user\n2. Navigate to /statistics/general\n3. Count visible tabs\n4. Verify at least 8 tabs visible (My tasks, My projects, Employees on my projects, Department projects, Department employees, Office projects, Office employees, Tasks by employees)\n5. Click each tab and verify data loads without error\n6. Verify the Statistics menu shows a dropdown with 'General Statistics' and 'Employee Reports' sub-items",
            "expected": "Multi-role user sees 8+ tabs based on their permissions. Maximum possible: 13 tabs (including Customer tabs). Menu shows dropdown.",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Tab Visibility Matrix", "module": "statistics/classic",
            "notes": "Live verified: pvaynmaster (7 roles) sees 8 tabs. Customer tabs (VIEW_CUSTOMER) never observed."
        },
        {
            "id": "TC-STAT-003", "title": "Date range filter — select custom range and verify data",
            "preconditions": "Any user with statistics access.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id WHERE be.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Verify default date range is Jan 1 – Dec 31 of current year\n4. Click the start date picker, select first day of last month\n5. Click the end date picker, select last day of last month\n6. Click 'Refresh data' button\n7. Verify table updates with data for the selected period only\n8. Verify 'For the period' column shows hours within selected range\n9. Verify 'Total' column may differ from 'For the period' (Total is all-time)",
            "expected": "Date range filter restricts data to selected period. Two datepickers with calendar icons. 'For the period' reflects the range; 'Total' is cumulative.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/classic",
            "notes": "Default: Jan 1 – Dec 31 current year. Period presets dropdown also available."
        },
        {
            "id": "TC-STAT-004", "title": "Period preset dropdown — quick date range selection",
            "preconditions": "Any user with statistics access.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Click the period preset dropdown ('Please select a date range')\n4. Verify preset options include common periods (This month, Last month, This quarter, This year, etc.)\n5. Select 'Last month'\n6. Verify start and end date pickers update to last month's range\n7. Click 'Refresh data'\n8. Verify data shows last month's statistics",
            "expected": "Period preset dropdown auto-fills start and end date pickers. Data refreshes for the selected period.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/classic",
            "notes": "Dropdown text: 'Please select a date range'. Presets auto-set both datepicker values."
        },
        {
            "id": "TC-STAT-005", "title": "Tree vs Flat mode toggle",
            "preconditions": "User with at least 2 tabs and data in the selected period.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Verify 'Flat' radio button is selected by default\n4. Verify table shows flat rows (no hierarchy)\n5. Click 'Tree' radio button\n6. Verify table restructures into tree mode (expandable parent rows)\n7. Click an expand icon on a parent row (employee → projects → tasks)\n8. Verify child rows appear with indent\n9. Click 'Flat' to switch back\n10. Verify table returns to flat layout",
            "expected": "Tree/Flat toggle changes table structure. Tree mode has expandable hierarchy. Flat is default.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/classic",
            "notes": "Flat is default. Tree mode: employee → project → task hierarchy. Uses rc-table."
        },
        {
            "id": "TC-STAT-006", "title": "Hours vs Days toggle",
            "preconditions": "User with data in the selected period.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Verify 'Hours' radio button is selected by default\n4. Note the values in 'For the period' and 'Total' columns\n5. Click 'Days' radio button\n6. Verify values recalculate (approximately hours / 8 = days)\n7. Click 'Hours' to switch back\n8. Verify original hour values restored",
            "expected": "Hours/Days toggle converts displayed effort values. Hours is default. Days = Hours / 8.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/classic",
            "notes": "Hours default. Days conversion via effortDisplayType setting."
        },
        {
            "id": "TC-STAT-007", "title": "Search filter — keyboard layout correction",
            "preconditions": "User with statistics access. Employee name known in both Latin and Cyrillic.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Type an employee name using the WRONG keyboard layout (e.g., type Cyrillic characters when meaning Latin, or vice versa)\n4. Verify search still finds the correct employee\n5. Clear the search\n6. Type the employee name correctly\n7. Verify the same result appears\n8. Verify search works for: first name, last name, and login",
            "expected": "Search with wrong keyboard layout still finds the employee. correctLayout() function handles transliteration.",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Search / Filtering, #298", "module": "statistics/classic",
            "notes": "#298: SuggestionMappingUtil.correctLayout() handles wrong keyboard. Also supports URL params (#2624)."
        },
        {
            "id": "TC-STAT-008", "title": "Table sorting — sort by column headers",
            "preconditions": "User with data showing multiple rows.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Click 'For the period' column header\n4. Verify rows sort ascending by period hours\n5. Click again\n6. Verify rows sort descending\n7. Click 'Employee / project / task' column header\n8. Verify rows sort alphabetically\n9. Click 'Start date' column header\n10. Verify rows sort by date",
            "expected": "Columns are sortable with ascending/descending toggle. Sort indicators (↑↓) visible.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Table Structure", "module": "statistics/classic",
            "notes": "Sortable columns: Employee/project/task, For the period, Total, Start date, End date."
        },
        {
            "id": "TC-STAT-009", "title": "Tree mode — expand employee node to see project/task hierarchy",
            "preconditions": "User with data. Employee who reported on multiple projects.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Switch to 'Tree' mode\n4. Find an employee row with an expand icon\n5. Click the expand icon\n6. Verify project rows appear as children\n7. Click expand on a project row\n8. Verify task rows appear as grandchildren\n9. Verify hours sum: sum of task hours ≈ project hours ≈ employee total\n10. Collapse all and verify rows disappear",
            "expected": "Three-level tree: employee → project → task. Child rows appear on expand. Hours are hierarchically consistent.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Table Structure", "module": "statistics/classic",
            "notes": "rc-table tree-table with lazy child expansion. Each level fetches from a different API endpoint."
        },
        {
            "id": "TC-STAT-010", "title": "Absence icons — sick leave and vacation icons with tooltip",
            "preconditions": "Employee with a vacation or sick leave in the selected date range.\nSETUP: Ensure data exists — check via API.\nQuery: SELECT DISTINCT sl.employee FROM ttt_vacation.sick_leave sl WHERE sl.status != 'DELETED' AND sl.start_date <= CURRENT_DATE AND sl.end_date >= DATE_TRUNC('year', CURRENT_DATE) LIMIT 5",
            "steps": "1. Login as a user with statistics access\n2. Navigate to /statistics/general\n3. Set date range to include the period with absence\n4. Find the employee row\n5. Verify sick leave icon (if applicable) appears on the row\n6. Hover over the sick leave icon\n7. Verify tooltip shows: total hours + individual periods with dates and status\n8. Verify vacation icon appears (if applicable)\n9. Hover over vacation icon\n10. Verify tooltip shows vacation periods within the selected range only",
            "expected": "AbsencesIcon shows IconSick and IconVacation inline. Tooltip shows breakdown: total hours + period dates + status + payment type.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Absence Icons, #2435", "module": "statistics/classic",
            "notes": "#2435: Show only period within selected range (N-dash dates). Both sub-systems display icons."
        },
        {
            "id": "TC-STAT-011", "title": "CompanyStaff profile link and Report page link",
            "preconditions": "User viewing statistics with employee rows visible.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Find an employee row\n4. Verify CompanyStaff profile link icon is present\n5. Click the CS profile link icon\n6. Verify it opens CompanyStaff profile page (https://companystaff.noveogroup.com/profile/{login})\n7. Go back to statistics\n8. Find the employee row again\n9. Verify Report page link icon is present\n10. Click the Report page link icon\n11. Verify it navigates to the employee's report page",
            "expected": "Two action icons per employee: CS profile link and Report page link. Both navigate correctly.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Table Structure", "module": "statistics/classic",
            "notes": "Hardcoded CS URL: https://companystaff.noveogroup.com/profile/ (tech debt #4)."
        },
        {
            "id": "TC-STAT-012", "title": "Reset all filters button",
            "preconditions": "User with active filters applied.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Apply a search filter (type an employee name)\n4. Change date range to a custom period\n5. Select 'Tree' mode\n6. Verify 'Reset all filters' button is enabled\n7. Click 'Reset all filters'\n8. Verify search cleared, date range reset to Jan 1 – Dec 31, mode reset to 'Flat'\n9. Verify button becomes disabled (no active filters)",
            "expected": "Reset clears all filters, date range, and view mode to defaults. Button disabled when no filters active.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/classic",
            "notes": "Button disabled when no filters active. Resets: search, date range, list type, effort type."
        },
    ]


def get_employee_reports_cases():
    """TS-Stat-EmployeeReports: Employee Reports page — search, norm, excess, comments."""
    return [
        {
            "id": "TC-STAT-013", "title": "Access Employee Reports as privileged user",
            "preconditions": "User with ADMIN or DEPARTMENT_MANAGER role.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_ADMIN', 'ROLE_DEPARTMENT_MANAGER') AND be.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the privileged user\n2. Verify Statistics menu shows dropdown with 'General Statistics' and 'Employee Reports'\n3. Click 'Employee Reports'\n4. Verify /statistics/employee-reports page loads\n5. Verify page title shows 'Employee Reports'\n6. Verify table with employee data is displayed\n7. Verify columns: Employee, Manager, Reported, Norm, Excess (Превышение), Comment",
            "expected": "Privileged users see Statistics submenu with both pages. Employee Reports loads with 6 columns and employee data.",
            "priority": "Critical", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Menu Structure", "module": "statistics/employee-reports",
            "notes": "Roles with access: ADMIN, CHIEF_ACCOUNTANT, OFFICE_ACCOUNTANT, DEPARTMENT_MANAGER, TECH_LEAD."
        },
        {
            "id": "TC-STAT-014", "title": "Access Employee Reports as EMPLOYEE-only — 403",
            "preconditions": "User with EMPLOYEE role only.\nQuery: (same as TC-STAT-001)",
            "steps": "1. Login as the EMPLOYEE-only user\n2. Navigate directly to /statistics/employee-reports\n3. Verify 403 Forbidden or access denied page\n4. Verify Statistics menu does NOT show 'Employee Reports' submenu\n5. Verify the typo: 'You do'nt have access' (BUG-STAT-UI-1)",
            "expected": "EMPLOYEE-only users get 403 on Employee Reports. No menu item visible. Known typo in error message.",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Employee Reports Sub-Page", "module": "statistics/employee-reports",
            "notes": "BUG-STAT-UI-1: typo 'You do'nt have access' (misplaced apostrophe). Low severity."
        },
        {
            "id": "TC-STAT-015", "title": "Employee search — live filtering by name/login",
            "preconditions": "Privileged user with access to multiple employees.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. In the search field, start typing an employee's first name\n4. Verify table filters live as you type (no submit button needed)\n5. Clear the search\n6. Type the employee's last name\n7. Verify same employee found\n8. Clear and type the employee's login\n9. Verify found by login\n10. Clear and type using wrong keyboard layout\n11. Verify search still finds the employee (layout correction)",
            "expected": "Live search by first name, last name, login. Wrong keyboard layout correction works. No submit button — table updates as user types.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Search field (III.1)", "module": "statistics/employee-reports",
            "notes": "Note from #3195 QA: Bug 1 = no dynamic text filtering. This was fixed."
        },
        {
            "id": "TC-STAT-016", "title": "Month picker — default is last open for approval",
            "preconditions": "Privileged user. Check which month is last open for approval in user's salary office.\nQuery: SELECT MAX(report_date) FROM ttt_backend.accounting_period WHERE status = 'OPEN' AND office_id = (SELECT office_id FROM ttt_vacation.employee WHERE login = '<user_login>')",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Check the month picker value\n4. Verify it defaults to the last month open for approval/confirmation (not current month)\n5. Change to a different month\n6. Verify table refreshes with the new month's data\n7. Change back to the default month\n8. Verify data matches",
            "expected": "Month picker defaults to last open for approval month (not current). Changing month reloads employee data.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Date picker (III.2), #3195 Bug 6", "module": "statistics/employee-reports",
            "notes": "#3195 Bug 6: default month was incorrectly set to current month instead of last open for approval."
        },
        {
            "id": "TC-STAT-017", "title": "Employee row — click name opens CS profile",
            "preconditions": "Privileged user viewing Employee Reports.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Click an employee's name in the Employee column\n4. Verify CompanyStaff profile page opens for that employee\n5. Verify the URL is https://companystaff.noveogroup.com/profile/{login}\n6. Go back to Employee Reports",
            "expected": "Clicking employee name navigates to CompanyStaff profile page. Not to report page.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Employee (4.1), #3195 Bug 4", "module": "statistics/employee-reports",
            "notes": "#3195 Bug 4: name click originally opened report page instead of CS page. Fixed."
        },
        {
            "id": "TC-STAT-018", "title": "Employee row — report icon on hover",
            "preconditions": "Privileged user viewing Employee Reports.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Hover over an employee row\n4. Verify report icon appears on hover\n5. Hover over the icon\n6. Verify tooltip: 'Employee report page'\n7. Click the icon\n8. Verify it navigates to the employee's report page\n9. Go back to Employee Reports",
            "expected": "Report icon appears on row hover. Tooltip: 'Employee report page'. Click navigates to /report page for that employee.",
            "priority": "Medium", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Employee (4.1), #3195 Bug 3", "module": "statistics/employee-reports",
            "notes": "#3195 Bug 3: report icon not shown on hover. Was fixed."
        },
        {
            "id": "TC-STAT-019", "title": "Employee row expand — project breakdown",
            "preconditions": "Privileged user. Employee with reports on 2+ projects in selected month.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Click anywhere on an employee row (except the name)\n4. Verify accordion expands showing project breakdown\n5. Verify each project row shows: project name, reported hours, percentage of employee total\n6. Verify projects are sorted by hours descending\n7. Click the row again\n8. Verify accordion collapses",
            "expected": "Row click (not on name) expands accordion with per-project breakdown. Projects sorted by hours descending.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Employee (4.1)", "module": "statistics/employee-reports",
            "notes": "Click on name → CS page. Click elsewhere → expand. projectBreakdown re-fetched on each navigation."
        },
        {
            "id": "TC-STAT-020", "title": "Manager column — CS link and filter dropdown",
            "preconditions": "Privileged user viewing Employee Reports with employees under different managers.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Verify 'Manager' column is present\n4. Click a manager's name in the Manager column\n5. Verify it opens CompanyStaff profile for that manager\n6. Go back\n7. Click the Manager column header filter dropdown\n8. Select a specific manager\n9. Verify table filters to show only employees under that manager\n10. Clear the filter\n11. Verify all employees shown again",
            "expected": "Manager column: name links to CS profile. Header has filter dropdown to filter by manager.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Manager (4.2), #3309", "module": "statistics/employee-reports",
            "notes": "#3309: Manager column + filter added in Sprint 14. Filter is dropdown in column header."
        },
        {
            "id": "TC-STAT-021", "title": "Reported column — arrows and color coding",
            "preconditions": "Privileged user. Month with employees having over-report AND under-report.\nQuery: SELECT sr.employee_login, sr.reported_effort, sr.month_norm FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.reported_effort > 0 ORDER BY ABS(sr.reported_effort - sr.month_norm/60.0) DESC LIMIT 10",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Find an employee with reported hours ABOVE budget norm\n4. Verify red arrow up (↑) indicator next to the reported value\n5. Verify text is red if excess % > notification.reporting.over threshold\n6. Find an employee with reported hours BELOW budget norm\n7. Verify purple arrow down (↓) indicator\n8. Verify text is purple if |excess %| > notification.reporting.under threshold\n9. Find an employee with reported ≈ norm (neutral)\n10. Verify no arrow, normal text color",
            "expected": "Reported column: red ↑ for over-report, purple ↓ for under-report. Color applied when excess exceeds configured thresholds.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Reported (4.3)", "module": "statistics/employee-reports",
            "notes": "Thresholds from admin settings: notification.reporting.over, notification.reporting.under. #3195 Bug 5: thresholds were not applied."
        },
        {
            "id": "TC-STAT-022", "title": "Norm column — budgetNorm display rules (3 cases)",
            "preconditions": "Privileged user. Month with employees in each display case.\nQuery: SELECT sr.employee_login, sr.month_norm, sr.budget_norm FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.month_norm != sr.budget_norm LIMIT 5",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. CASE 1 — Employee with admin vacation (budgetNorm != individualNorm):\n   - Find employee with Norm showing format '{X} ({Y})' (e.g. '136 (144)')\n   - Verify first number is individual norm, number in brackets is budget norm\n4. CASE 2 — Employee with regular vacation/sick leave only:\n   - Find employee with Norm showing single number (e.g. '144')\n   - This is budgetNorm which equals individualNorm\n5. CASE 3 — Employee with no absences:\n   - Find employee with full office norm (e.g. '152')\n   - This is the general office norm\n6. Verify info icon next to 'Norm' column header\n7. Hover over info icon\n8. Verify tooltip explains the dual display format",
            "expected": "Case 1: '{individual} ({budget})'. Case 2: '{budget}' (=individual). Case 3: '{budget}' (=office norm). Info tooltip present.",
            "priority": "Critical", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Norm (4.4), #3381", "module": "statistics/employee-reports",
            "notes": "#3381: budgetNorm = individualNorm + admin vacation hours. #3409 adds family sick leave hours."
        },
        {
            "id": "TC-STAT-023", "title": "Excess column — percentage formatting and N/A edge case",
            "preconditions": "Privileged user. Need employee with norm=0 and reported>0.\nQuery: SELECT sr.employee_login FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND (sr.month_norm = 0 OR sr.budget_norm = 0) AND sr.reported_effort > 0 LIMIT 3",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Find an employee with excess in range (−1%, +1%):\n   - Verify display shows 1 decimal place (e.g. '0.5%' or '-0.3%')\n4. Find an employee with excess ≥ 1% or ≤ −1%:\n   - Verify display shows integer (e.g. '15%' or '-8%')\n5. Find an employee with norm=0, reported=0:\n   - Verify display shows '0%'\n6. Find an employee with norm=0, reported>0:\n   - Verify display shows '+N/A%'\n   - Verify this employee sorts at TOP of the list (maximum value in sort)",
            "expected": "(-1,+1) → 1 decimal. |excess|≥1 → integer. norm=0/reported=0 → 0%. norm=0/reported>0 → +N/A% (sorts as max).",
            "priority": "Critical", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Excess (4.5), #3195 Bug 7-8", "module": "statistics/employee-reports",
            "notes": "#3195 Bug 7: was showing '0.00%' instead of '+N/A%'. Bug 8: wrong decimal formatting."
        },
        {
            "id": "TC-STAT-024", "title": "Comment field — inline edit, save on blur/Tab",
            "preconditions": "Privileged user with Employee Reports access.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Click the Comment field for an employee\n4. Verify edit mode activates with cursor in textarea\n5. Type 'Test comment — session 102'\n6. Press Enter key\n7. Verify Enter creates a new paragraph (NOT save)\n8. Type 'Second line'\n9. Press Tab key\n10. Verify comment is saved (edit mode exits)\n11. Navigate away and back to /statistics/employee-reports\n12. Verify the saved comment persists\n13. Click the comment field again\n14. Clear the text\n15. Click outside the field\n16. Verify comment is saved as empty on blur",
            "expected": "Enter = new paragraph. Tab or click-outside = save + exit edit. Comments stored per employee per month. Persists across navigation.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Comment (4.6), #3309", "module": "statistics/employee-reports",
            "notes": "#3309: Comment field added. API: POST /v1/statistic/report saves. No auto-save on tab close — data lost."
        },
        {
            "id": "TC-STAT-025", "title": "Comment field — per-month storage",
            "preconditions": "Privileged user. Two different months with data.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Select month A (e.g., last month)\n4. Add comment for an employee: 'Month A comment'\n5. Select month B (e.g., two months ago)\n6. Add comment for same employee: 'Month B comment'\n7. Switch back to month A\n8. Verify comment shows 'Month A comment' (not 'Month B comment')\n9. Switch to month B\n10. Verify comment shows 'Month B comment'",
            "expected": "Comments are per-employee per-month. Changing month shows that month's comments, not a shared comment.",
            "priority": "Medium", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Comment (4.6)", "module": "statistics/employee-reports",
            "notes": "Stored in statistic_report.comment, keyed by (employee_login, report_date)."
        },
        {
            "id": "TC-STAT-026", "title": "Default sorting — by excess percentage descending",
            "preconditions": "Privileged user with multiple employees having different excess values.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Without changing any sort, examine the Excess column values\n4. Verify employees are sorted by excess % descending (highest over-reporters at top)\n5. Verify +N/A% entries (if any) appear at the very top (sorted as maximum)\n6. Verify under-reporters (negative %) appear at the bottom",
            "expected": "Default sort: by excess percentage descending. +N/A% at top. Negative excess at bottom.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Sorting (4.7)", "module": "statistics/employee-reports",
            "notes": "Confluence: 'По умолчанию — по Превышению (процент) по убыванию'."
        },
    ]


def get_norm_excess_cases():
    """TS-Stat-NormExcess: Norm calculation edge cases and excess display rules."""
    return [
        {
            "id": "TC-STAT-027", "title": "budgetNorm — employee with administrative vacation",
            "preconditions": "Employee who had an administrative (unpaid) vacation in the selected month.\nQuery: SELECT v.employee, v.start_date, v.end_date FROM ttt_vacation.vacation v WHERE v.payment_type = 'ADMINISTRATIVE' AND v.status IN ('NEW','APPROVED','PAID') AND v.start_date >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 month') ORDER BY v.start_date DESC LIMIT 5",
            "steps": "1. Login as ADMIN or CHIEF_ACCOUNTANT\n2. Navigate to /statistics/employee-reports\n3. Select the month containing the admin vacation\n4. Find the employee\n5. Verify Norm column shows '{individualNorm} ({budgetNorm})' format\n6. Verify budgetNorm = individualNorm + admin_vacation_working_hours\n7. Verify Excess uses budgetNorm (not personalNorm) as denominator\nDB-CHECK: SELECT month_norm, budget_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = '<month_start>'",
            "expected": "budgetNorm includes admin vacation hours that personalNorm excludes. Display shows dual format. Excess formula uses budgetNorm.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Budget Norm, #3381", "module": "statistics/norm",
            "notes": "Code: filterNonAdministrativeVacations() removes ADMINISTRATIVE from budgetNorm calculation input."
        },
        {
            "id": "TC-STAT-028", "title": "budgetNorm — employee with family member sick leave (#3409)",
            "preconditions": "Employee with a family member sick leave (familyMember=true) in the selected month.\nQuery: SELECT sl.employee, sl.start_date FROM ttt_vacation.sick_leave sl WHERE sl.family_member = true AND sl.status != 'DELETED' AND sl.start_date >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 month') ORDER BY sl.start_date DESC LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month with family member sick leave\n4. Find the employee\n5. Verify Norm column shows '{individualNorm} ({budgetNorm})'\n6. Verify budgetNorm = individualNorm + familyMember_sickleave_working_hours\n7. Verify own sick leave (familyMember=false) IS deducted from individual norm\n8. Verify family member sick leave is NOT deducted from individual norm but IS added to budget norm\nDB-CHECK: SELECT month_norm, budget_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = '<month_start>'",
            "expected": "Family member sick leave hours add to budgetNorm (like admin vacation). Own sick leave reduces individualNorm. Different treatment.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-ticket-findings.md §#3409, statistics-service-implementation.md", "module": "statistics/norm",
            "notes": "#3409 Sprint 16: Nb = Ni + admin_vacation_hrs + familyMember_sickleave_hrs. Depends on #3408 deployment."
        },
        {
            "id": "TC-STAT-029", "title": "personalNorm vs budgetNorm — no difference (regular vacation only)",
            "preconditions": "Employee with ONLY regular (paid) vacation in the month, no admin vacation or family sick leave.\nQuery: SELECT v.employee FROM ttt_vacation.vacation v WHERE v.payment_type = 'REGULAR' AND v.status IN ('NEW','APPROVED','PAID') AND v.start_date >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 month') AND v.employee NOT IN (SELECT v2.employee FROM ttt_vacation.vacation v2 WHERE v2.payment_type = 'ADMINISTRATIVE' AND v2.start_date >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 month')) LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month with the regular vacation\n4. Find the employee\n5. Verify Norm column shows SINGLE number (e.g., '144') — no brackets\n6. This means budgetNorm == individualNorm\n7. Verify the norm is reduced by regular vacation hours (< full office norm)\nDB-CHECK: SELECT month_norm, budget_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = '<month_start>' — expect month_norm = budget_norm",
            "expected": "When no admin vacation or family sick leave, budgetNorm = individualNorm. Norm column shows single number.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Display Rules", "module": "statistics/norm",
            "notes": "renderNormHours: budgetNorm !== individualNorm → dual format; else → single number."
        },
        {
            "id": "TC-STAT-030", "title": "Partial-month employee — mid-month hire (#3356)",
            "preconditions": "Employee who was hired mid-month (employment start date is NOT the 1st of the month).\nQuery: SELECT ep.employee_login, ep.period_start FROM ttt_vacation.employee_period ep WHERE EXTRACT(DAY FROM ep.period_start) > 1 AND ep.period_start >= '2025-01-01' ORDER BY ep.period_start DESC LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month when the employee was hired\n4. Find the employee\n5. Verify Norm is calculated only for working days FROM hire date TO end of month\n6. Verify the norm is LESS than the full office norm for that month\n7. Verify Excess % is calculated correctly against the partial norm\nDB-CHECK: SELECT month_norm, budget_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = '<month_start>'",
            "expected": "Mid-month hire: norm reflects only working days from hire date. effectiveBounds clamps the range.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §effectiveBounds, #3356", "module": "statistics/norm",
            "notes": "#3356: effectiveBounds clamps date range to (max(rangeStart, periodStart), min(rangeEnd, periodEnd))."
        },
        {
            "id": "TC-STAT-031", "title": "Terminated employee — shown in last month only (#3320)",
            "preconditions": "Recently terminated employee.\nQuery: SELECT ep.employee_login, ep.period_end FROM ttt_vacation.employee_period ep WHERE ep.period_end IS NOT NULL AND ep.period_end >= CURRENT_DATE - INTERVAL '3 months' ORDER BY ep.period_end DESC LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month the employee was terminated in\n4. Verify the employee IS shown in the list\n5. Note their norm (should be partial-month if terminated mid-month)\n6. Switch to the month AFTER termination\n7. Verify the employee is NOT shown\n8. Switch to a month BEFORE termination\n9. Verify the employee IS shown with full norm",
            "expected": "Terminated employees: shown in their last active month with partial norm, hidden in subsequent months. #3320 fix.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Displayed Users (4.8), #3320", "module": "statistics/norm",
            "notes": "#3320: filter uses ttt_vacation.employee_period dates, not ttt_backend.employee.last_date."
        },
        {
            "id": "TC-STAT-032", "title": "norm=0, reported=0 — display 0%",
            "preconditions": "Employee with norm=0 AND no reported hours in the month (e.g., full month vacation/sick leave).\nQuery: SELECT sr.employee_login FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.budget_norm = 0 AND sr.reported_effort = 0 LIMIT 3",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Find the employee with norm=0, reported=0\n4. Verify Excess column shows '0%'\n5. Verify no color coding (neutral)\n6. Verify no arrow indicator",
            "expected": "norm=0, reported=0 → Excess = '0%'. Neutral display, no arrows, no color.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Excess Corner Cases (4.5.5)", "module": "statistics/norm",
            "notes": "Confluence 4.4.4: explicit corner case specification."
        },
        {
            "id": "TC-STAT-033", "title": "norm=0, reported>0 — display +N/A%, sort as maximum",
            "preconditions": "Employee with norm=0 AND reported hours > 0.\nQuery: SELECT sr.employee_login, sr.reported_effort FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.budget_norm = 0 AND sr.reported_effort > 0 LIMIT 3",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Verify default sort (by excess descending)\n4. Check the TOP of the list for '+N/A%' entries\n5. Find the employee with norm=0, reported>0\n6. Verify Excess shows '+N/A%'\n7. Verify this employee appears at the TOP (sorted as maximum value)\n8. Verify reporting status indicator present (arrow up, color)",
            "expected": "norm=0 with reported>0 → '+N/A%'. Sorted as maximum value — appears at top of descending sort.",
            "priority": "Critical", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Excess Corner Cases (4.5.5), #3195 Bug 7", "module": "statistics/norm",
            "notes": "#3195 Bug 7: was showing '0.00%' instead of '+N/A%'. ExcessStatus.NA in backend."
        },
        {
            "id": "TC-STAT-034", "title": "Over-limit toggle — filter by threshold (#3306)",
            "preconditions": "Privileged user. Admin settings configured: notification.reporting.over=10, notification.reporting.under=-10.\nQuery: SELECT value FROM ttt_backend.application_settings WHERE key IN ('notification.reporting.over', 'notification.reporting.under')",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Note the total number of employees listed\n4. Click the 'Only excess over limit' toggle (switch to ON)\n5. Verify table filters to show ONLY employees whose excess exceeds thresholds:\n   - Over-reporters: excess > notification.reporting.over (e.g., > 10%)\n   - Under-reporters: |excess| > |notification.reporting.under| (e.g., < -10%)\n6. Verify the filtered list is shorter than the full list\n7. Verify all shown employees have excess outside the threshold range\n8. Toggle OFF\n9. Verify full list restored",
            "expected": "Over-limit toggle filters employees by configured thresholds. Only excess > over OR < under thresholds shown.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Over-limit toggler (III.3), #3306", "module": "statistics/norm",
            "notes": "#3306: toggle was broken — didn't filter. Fixed. Prerequisites: admin settings must be configured."
        },
        {
            "id": "TC-STAT-035", "title": "Color indicators — over/under/neutral per thresholds",
            "preconditions": "Privileged user. Employees with different excess levels visible.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Find over-reporter (excess > notification.reporting.over threshold):\n   - Verify red text color on Reported and Excess values\n   - Verify red arrow up (↑)\n4. Find under-reporter (|excess| > |notification.reporting.under| threshold):\n   - Verify purple text color\n   - Verify purple arrow down (↓)\n5. Find neutral employee (excess within thresholds):\n   - Verify normal text color (black)\n   - Verify no arrow\n6. Find employee with excess between 0 and threshold (e.g., +5% with threshold +10%):\n   - Verify normal color (excess exists but below threshold)\n   - Verify arrow direction (up for positive)",
            "expected": "Red = over threshold. Purple = under threshold. Normal = within range. Arrows indicate direction regardless of threshold.",
            "priority": "High", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Reported (4.3), statistics-service-implementation.md §Color Indicators", "module": "statistics/norm",
            "notes": "reportedNotificationStatus: LOW→underReported, HIGH/NA→overReported, NEUTRAL→normal. Thresholds NOT hardcoded."
        },
        {
            "id": "TC-STAT-036", "title": "Excess behavior — working during vacation = over-report",
            "preconditions": "Employee who worked during their regular vacation in the selected month.\nQuery: SELECT sr.employee_login, sr.reported_effort, sr.month_norm, sr.budget_norm FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.reported_effort > sr.budget_norm / 60.0 ORDER BY (sr.reported_effort - sr.budget_norm / 60.0) DESC LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Find the employee who worked during vacation\n4. Verify their norm was reduced by vacation days (personal norm < office norm)\n5. Verify their reported hours exceed the reduced norm\n6. Verify they show as OVER-reported (red, arrow up)\n7. Verify excess % is calculated against budgetNorm (not personalNorm)",
            "expected": "Working during vacation → over-report. Budget norm denominator means admin vacation hours count against the employee.",
            "priority": "Medium", "type": "UI",
            "req_ref": "REQ-statistics-employee-reports.md §Excess Behavior (4.5.6), #3381", "module": "statistics/norm",
            "notes": "#3381 behavior note: working during regular vacation/sick leave → over-report."
        },
    ]


def get_export_wsr_cases():
    """TS-Stat-ExportWSR: CSV export, copy/link features, WSR tree view."""
    return [
        {
            "id": "TC-STAT-037", "title": "CSV export — download classic statistics data",
            "preconditions": "User with statistics access and data in selected period.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Set date range and select a tab with data\n4. Click 'Export' dropdown button\n5. Select 'Download CSV'\n6. Verify CSV file downloads\n7. Open the CSV file\n8. Verify data matches the table content\n9. Verify column headers present\n10. Verify hours values match the displayed values",
            "expected": "CSV downloads with table data. Values match what is displayed in the UI. Headers included.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Export (Classic)", "module": "statistics/export",
            "notes": "Export dropdown has 3 options: Copy table, Download CSV, Copy link for Google tables."
        },
        {
            "id": "TC-STAT-038", "title": "CSV export — hours vs days unit parameter (#1422)",
            "preconditions": "User with statistics access.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Set toggle to 'Hours'\n4. Export CSV\n5. Note values in the CSV\n6. Switch toggle to 'Days'\n7. Export CSV again\n8. Compare values: Days CSV values ≈ Hours CSV values / 8\n9. Verify units parameter is correctly passed in the export request",
            "expected": "CSV export respects hours/days toggle. Hours and Days CSVs have proportional values (factor of 8).",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Export (Classic), #1422", "module": "statistics/export",
            "notes": "#1422: units parameter support added to export endpoints."
        },
        {
            "id": "TC-STAT-039", "title": "CSV export — empty params regression (#2191)",
            "preconditions": "User with statistics access.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Clear all filters (reset)\n4. Without selecting any specific tab or filter, click Export → Download CSV\n5. Verify export does NOT return 400 error\n6. Verify CSV downloads with default data (or appropriate error message)\n7. Apply a filter, then export\n8. Verify export works with applied filters",
            "expected": "Export with empty/default params should work (not 400). Regression test for #2191.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Export (Classic), #2191", "module": "statistics/export",
            "notes": "#2191 CLOSED: 400 error when params empty was a regression. Verify it stays fixed."
        },
        {
            "id": "TC-STAT-040", "title": "Copy table to clipboard",
            "preconditions": "User with statistics data visible.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Click 'Export' dropdown\n4. Select 'Copy the table'\n5. Verify success notification (table copied)\n6. Open a spreadsheet application or text editor\n7. Paste\n8. Verify pasted data matches the statistics table structure",
            "expected": "Table data copied to clipboard. Pasteable into spreadsheet. Data matches displayed table.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/export",
            "notes": "Export dropdown: 'Copy the table' option."
        },
        {
            "id": "TC-STAT-041", "title": "Copy link for Google Tables",
            "preconditions": "User with statistics data visible.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Click 'Export' dropdown\n4. Select 'Copy link for Google tables'\n5. Verify success notification (link copied)\n6. Open a Google Sheets document\n7. Use IMPORTDATA or paste the link\n8. Verify data loads from the link",
            "expected": "Shareable data link copied to clipboard. Compatible with Google Sheets IMPORTDATA function.",
            "priority": "Low", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Controls & Features", "module": "statistics/export",
            "notes": "Third export option. Generates a URL that can be used as a data source."
        },
        {
            "id": "TC-STAT-042", "title": "Admin-only largest customers export (#2096)",
            "preconditions": "User with ADMIN role.",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/general\n3. Verify 'employees-largest-customers' export option is available\n4. Click export\n5. Verify CSV downloads with customer-employee data\n6. Login as a non-ADMIN user\n7. Verify 'employees-largest-customers' export is NOT available",
            "expected": "Largest customers CSV export is admin-only. Non-admin users don't see the option.",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Export (Classic), #2096", "module": "statistics/export",
            "notes": "#2096: GET /v1/statistic/export/employees-largest-customers — admin-only endpoint."
        },
        {
            "id": "TC-STAT-043", "title": "Norm export by individual calendar (#3400)",
            "preconditions": "Admin or privileged user.",
            "steps": "1. Login as ADMIN\n2. Navigate to the norm export feature (may be under Employee Reports or admin)\n3. Trigger the individual calendar norm export\n4. Verify CSV downloads\n5. Verify CSV contains columns: login, name, surname, department_manager_login, salary_office, individual_norm\n6. Verify individual_norm values are correct (compare with Employee Reports Norm column)",
            "expected": "CSV export with per-employee individual norm by calendar. Contains employee details and norm.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "statistics-ticket-findings.md §#3400", "module": "statistics/export",
            "notes": "#3400 Sprint 15: Simple CSV export. Production Ready."
        },
        {
            "id": "TC-STAT-044", "title": "WSR tree view — basic display and expand/collapse",
            "preconditions": "User with WSR access. Period with WSR data.",
            "steps": "1. Login as the user\n2. Navigate to the WSR (Weekly Status Report) view\n3. Verify tree structure displays with weekly grouping\n4. Expand a week node\n5. Verify project/task breakdown appears\n6. Try to collapse the expanded node\n7. Note: #3041 — task branches may NOT collapse (known bug)\n8. Verify data matches the selected period",
            "expected": "WSR tree view renders weekly status data. Known bug: branches may not collapse (#3041).",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-statistics-module.md §WSR View, #3030, #3041", "module": "statistics/wsr",
            "notes": "#3030: layout broken. #3041: branches don't collapse. Both OPEN bugs."
        },
        {
            "id": "TC-STAT-045", "title": "WSR — multiple sub-issues (#3289)",
            "preconditions": "User with WSR access.",
            "steps": "1. Login as the user\n2. Navigate to WSR view\n3. Verify WSR lists open behavior (should not auto-open all)\n4. Verify tree view is enabled (not disabled)\n5. Check update button logic — verify it shows correct period data\n6. Switch language to English\n7. Verify tooltip translations are correct\n8. Click the update button\n9. Verify it doesn't break the list view\n10. Check for unnecessary top-level node",
            "expected": "Regression test for 6 WSR sub-issues: lists always open, tree view disabled, update button, tooltip translation, button breaks list, unnecessary node.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §WSR View, #3289", "module": "statistics/wsr",
            "notes": "#3289 OPEN: 6+ sub-issues reported. All are WSR-specific regression checks."
        },
        {
            "id": "TC-STAT-046", "title": "WSR — update button shows wrong period data (#3144)",
            "preconditions": "User with WSR access. Select a specific week.",
            "steps": "1. Login as the user\n2. Navigate to WSR view\n3. Select a specific weekly period\n4. Verify data shown matches the selected week\n5. Click the 'Update' button\n6. Verify the data refreshes for the SAME period (not a different one)\n7. Change the period selection\n8. Click Update again\n9. Verify the new period's data is shown",
            "expected": "Update button refreshes data for the currently selected period, not a different one.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §WSR View, #3144", "module": "statistics/wsr",
            "notes": "#3144 OPEN: update button shows wrong period data."
        },
    ]


def get_permissions_cases():
    """TS-Stat-Permissions: Role-based tab visibility and data scoping."""
    return [
        {
            "id": "TC-STAT-047", "title": "ADMIN — sees all 13 tabs + all employees in Employee Reports",
            "preconditions": "User with ROLE_ADMIN.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ADMIN' AND be.enabled = true LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/general\n3. Count all visible tabs\n4. Verify maximum tabs visible (up to 13 depending on data: My tasks, My projects, Employees on my projects, Department projects, Department employees, Office projects, Office employees, Tasks by employees, Project tasks, Customer projects, Customer employees, Manager projects, Manager employees)\n5. Navigate to /statistics/employee-reports\n6. Verify ALL employees across all offices/departments are listed\n7. Verify employee count matches total active employees",
            "expected": "ADMIN: maximum tab visibility + all employees in Employee Reports. No data scoping.",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Access Control", "module": "statistics/permissions",
            "notes": "ADMIN and CHIEF_ACCOUNTANT have identical access: all tabs + all employees."
        },
        {
            "id": "TC-STAT-048", "title": "OFFICE_ACCOUNTANT — own office employees only",
            "preconditions": "User with ROLE_ACCOUNTANT assigned to a specific office.\nQuery: SELECT a.login, o.name AS office FROM ttt_vacation.employee a JOIN ttt_vacation.office_accountant oa ON a.id = oa.employee_id JOIN ttt_vacation.office o ON oa.office_id = o.id WHERE a.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the OFFICE_ACCOUNTANT\n2. Navigate to /statistics/employee-reports\n3. Verify only employees from the accountant's assigned salary office(s) are shown\n4. Count employees — should be less than total\n5. Search for an employee from a DIFFERENT office\n6. Verify that employee is NOT found\n7. Navigate to /statistics/general\n8. Verify Office tabs are visible\n9. Verify data is scoped to own office",
            "expected": "OFFICE_ACCOUNTANT sees only employees from their assigned salary offices. No cross-office data.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Access Control, REQ-statistics-employee-reports.md §Menu Structure", "module": "statistics/permissions",
            "notes": "Uses findAllByOfficesOrManagers() with officeIds filter."
        },
        {
            "id": "TC-STAT-049", "title": "DEPARTMENT_MANAGER — subordinates including contractors (#3147)",
            "preconditions": "User with ROLE_DEPARTMENT_MANAGER. Has subordinates including at least one contractor.\nQuery: SELECT m.login AS dm, e.login AS employee, e.contractor FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id JOIN ttt_backend.employee be ON m.login = be.login JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_DEPARTMENT_MANAGER' AND m.enabled = true AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the DEPARTMENT_MANAGER\n2. Navigate to /statistics/employee-reports\n3. Verify subordinate employees are listed (direct reports in department)\n4. Verify contractor employees are also visible (#3147 fix)\n5. Verify employees from OTHER departments are NOT shown\n6. Navigate to /statistics/general\n7. Verify Department tabs are visible (Department projects, Department employees)\n8. Verify data is scoped to own department",
            "expected": "DM sees subordinates including contractors. Other departments excluded. Department tabs visible.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Access Control, #3147", "module": "statistics/permissions",
            "notes": "#3147: DM access extended to include contractor statistics."
        },
        {
            "id": "TC-STAT-050", "title": "TECH_LEAD — own subordinates only",
            "preconditions": "User with ROLE_TECH_LEAD.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_TECH_LEAD' AND be.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the TECH_LEAD\n2. Navigate to /statistics/employee-reports\n3. Verify only own subordinate employees are listed\n4. Verify department tabs visible in General Statistics\n5. Verify employee count is smaller than DM count (TL has narrower scope)",
            "expected": "TL sees only their direct subordinates in Employee Reports. Department tabs visible in General.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Access Control", "module": "statistics/permissions",
            "notes": "TL scope via techLeadId filter in findAllByOfficesOrManagers()."
        },
        {
            "id": "TC-STAT-051", "title": "OFFICE_HR — sees only assigned employees (#3247)",
            "preconditions": "User with ROLE_OFFICE_HR.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_OFFICE_HR' AND be.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as OFFICE_HR\n2. Navigate to /statistics/employee-reports\n3. Verify only HR-assigned employees are shown (NOT all office employees)\n4. Count employees — should be less than full office count\n5. Verify employees outside HR assignment are NOT visible",
            "expected": "HR sees only their assigned employees, not all office employees. #3247 fix prevents over-exposure.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Access Control, #3247", "module": "statistics/permissions",
            "notes": "#3247: HR managers previously saw ALL employees — fixed to show only assigned. Caused regression #3298."
        },
        {
            "id": "TC-STAT-052", "title": "PROJECT_MANAGER — project tabs only, no Employee Reports",
            "preconditions": "User with ROLE_PROJECT_MANAGER only.\nQuery: SELECT be.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_PROJECT_MANAGER' AND be.enabled = true AND be.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN','ROLE_CHIEF_ACCOUNTANT','ROLE_ACCOUNTANT','ROLE_DEPARTMENT_MANAGER','ROLE_TECH_LEAD'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as the PROJECT_MANAGER\n2. Navigate to /statistics/general\n3. Verify Project tabs visible (Tasks by employees, etc.)\n4. Navigate to /statistics/employee-reports\n5. Verify 403 or no access (PM has no Employee Reports access)\n6. Verify Statistics menu has no 'Employee Reports' sub-item",
            "expected": "PM sees project-related tabs in General Statistics but has NO access to Employee Reports.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Permission Matrix", "module": "statistics/permissions",
            "notes": "PM → PROJECT tabs only. No Employee Reports access unless also has another qualifying role."
        },
        {
            "id": "TC-STAT-053", "title": "Search by projects — regression after HR hotfix (#3298)",
            "preconditions": "User with search by projects permission (VIEW_PROJECT or similar).",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Use the search/filter to search by project name\n4. Verify project search returns results\n5. Verify no 403 or empty results error\n6. Filter by a specific project\n7. Verify only data for that project is shown",
            "expected": "Search by projects works correctly. No regression from HR hotfix #3247.",
            "priority": "High", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Search / Filtering, #3298", "module": "statistics/permissions",
            "notes": "#3298 CLOSED: Search by projects broke after HR hotfix #3247 — overlapping permission changes caused regression."
        },
        {
            "id": "TC-STAT-054", "title": "Customer filter — requires VIEW_CUSTOMER permission",
            "preconditions": "Two users: one with VIEW_CUSTOMER permission, one without.",
            "steps": "1. Login as user WITH VIEW_CUSTOMER permission\n2. Navigate to /statistics/general\n3. Verify 'Customer' filter is available (4th filter type)\n4. Verify Customer tabs visible (Customer projects, Customer employees)\n5. Login as user WITHOUT VIEW_CUSTOMER permission\n6. Navigate to /statistics/general\n7. Verify 'Customer' filter is NOT available\n8. Verify only 3 filters visible (project, employee, task)\n9. Verify no Customer tabs",
            "expected": "Customer filter and tabs require VIEW_CUSTOMER permission. Without it: 3 filters and no customer tabs.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ui-deep-exploration.md §Search Filter Types", "module": "statistics/permissions",
            "notes": "4 filters with VIEW_CUSTOMER, 3 without. Customer tabs never observed in live testing."
        },
    ]


def get_hour_sum_cases():
    """TS-Stat-HourSum: Hour sum consistency (parent vs child, showFired)."""
    return [
        {
            "id": "TC-STAT-055", "title": "My Projects — parent total vs expanded children hour mismatch",
            "preconditions": "User with My Projects tab. Projects with active and dismissed employees.\nQuery: SELECT p.name, COUNT(DISTINCT pm.employee_id) AS members FROM ttt_backend.project p JOIN ttt_backend.project_member pm ON p.id = pm.project_id GROUP BY p.name HAVING COUNT(DISTINCT pm.employee_id) > 5 ORDER BY random() LIMIT 3",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Select 'My projects' tab\n4. Switch to 'Tree' mode\n5. Note the parent row total hours (For the period) for a project\n6. Expand the project node\n7. Sum all visible child employee hours\n8. Compare: parent total vs sum of visible children\n9. If mismatch exists, enable showFired parameter (if available)\n10. Verify mismatch resolves when fired employees included",
            "expected": "Parent total may include fired employee hours while children don't show fired employees. Known systematic issue (8 tickets).",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Hour Sum Consistency, #2097", "module": "statistics/hour-sum",
            "notes": "Root cause: /departments includes fired, /employees filters by showFired=false. 8 tickets: #2097,#2108,#1923,#2112,#2122,#2123,#2142,#2143."
        },
        {
            "id": "TC-STAT-056", "title": "Department projects — parent-child hour inconsistency",
            "preconditions": "User with Department projects tab. Department with dismissed employees.",
            "steps": "1. Login as DEPARTMENT_MANAGER or ADMIN\n2. Navigate to /statistics/general\n3. Select 'Department projects' tab\n4. Switch to 'Tree' mode\n5. Find a department row with significant total hours\n6. Expand it\n7. Sum all visible project/employee hours\n8. Compare with the parent total\n9. Document any discrepancy amount\n10. Verify via API: GET /v1/statistic/departments (includes fired) vs GET /v1/statistic/employees?showFired=false",
            "expected": "Mismatch expected for departments with fired employees. Document the discrepancy for each tab.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Hour Sum Consistency, #1923", "module": "statistics/hour-sum",
            "notes": "#1923: 'Department projects'/'Office projects' sum > project total. Fired employee inclusion is root cause."
        },
        {
            "id": "TC-STAT-057", "title": "Office projects — hour sum verification",
            "preconditions": "User with Office projects tab.",
            "steps": "1. Login as ADMIN or OFFICE_DIRECTOR\n2. Navigate to /statistics/general\n3. Select 'Office projects' tab\n4. Switch to 'Tree' mode\n5. Expand an office row\n6. Sum children hours\n7. Compare with parent total\n8. Document any discrepancy",
            "expected": "Same parent-child mismatch pattern as other tabs when fired employees exist.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-service-implementation.md §Hour Sum Consistency", "module": "statistics/hour-sum",
            "notes": "All tree tabs share the same root cause. Testing all tabs confirms the systemic nature."
        },
        {
            "id": "TC-STAT-058", "title": "showFired parameter — toggle fired employees visibility",
            "preconditions": "Tab with fired employee hours contributing to mismatches.",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/general\n3. Select any tab in Tree mode\n4. Note the parent total\n5. Expand and sum children (showFired=false default)\n6. Note the mismatch\n7. Via API: GET /v1/statistic/employees?showFired=true&...\n8. Compare the API response with showFired=true vs showFired=false\n9. Verify showFired=true response totals match the parent total\n10. If UI has a 'show fired' toggle, test it",
            "expected": "showFired=true includes dismissed employee hours — totals should match parent. showFired=false creates mismatches.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Hour Sum Consistency", "module": "statistics/hour-sum",
            "notes": "API: /v1/statistic/employees?showFired=true|false. Default is false. Parent totals always include fired."
        },
        {
            "id": "TC-STAT-059", "title": "Customer team — extreme mismatch example (#2143)",
            "preconditions": "User with Customer tabs visible. Customer with dismissed team members.",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/general\n3. Select 'Customer team' tab (if visible)\n4. Find a customer entry in Tree mode\n5. Note parent total hours\n6. Expand and sum children\n7. Document the mismatch (historically extreme: 205.57 vs 36.9)\n8. Compare with other tabs to confirm pattern",
            "expected": "Customer team tab shows same mismatch pattern. Historically extreme differences documented.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §Tier 3, #2143", "module": "statistics/hour-sum",
            "notes": "#2143 CLOSED: extreme mismatch 205.57 vs 36.9. Most dramatic example of the systemic issue."
        },
        {
            "id": "TC-STAT-060", "title": "Tasks by employees — project-level hour aggregation",
            "preconditions": "User with 'Tasks by employees' tab (VIEW_PROJECT permission).",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Select '@ Tasks by employees' tab\n4. Switch to Tree mode\n5. Expand a task to see employee breakdown\n6. Sum employee hours\n7. Verify sum matches the task total\n8. Check for fired employee inclusion in the total",
            "expected": "Task-level aggregation consistent with visible employees. Same showFired pattern applies.",
            "priority": "Medium", "type": "UI",
            "req_ref": "frontend-statistics-module.md §Permission-Gated Tabs", "module": "statistics/hour-sum",
            "notes": "VIEW_PROJECT permission required. Tree: task → employee hierarchy."
        },
    ]


def get_cache_sync_cases():
    """TS-Stat-CacheSync: statistic_report cache table correctness."""
    return [
        {
            "id": "TC-STAT-061", "title": "Cache vs live data consistency — Employee Reports accuracy",
            "preconditions": "ADMIN user. Both cache table and live calculation available.\nQuery: SELECT COUNT(*) FROM ttt_backend.statistic_report WHERE report_date = DATE_TRUNC('month', CURRENT_DATE)",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select current month\n4. Note reported hours and norm for 5 random employees\nDB-CHECK: For each employee, verify:\n  SELECT reported_effort, month_norm, budget_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = DATE_TRUNC('month', CURRENT_DATE)\n5. Compare UI values with DB values\n6. Verify reported_effort (DB hours) matches 'Reported' column\n7. Verify budget_norm / 60 (DB minutes → hours) matches 'Norm' column",
            "expected": "Employee Reports data matches statistic_report table. Units: DB reported_effort in HOURS, month_norm/budget_norm in MINUTES.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Database, statistics-api-testing.md §Mixed Unit Discrepancy", "module": "statistics/cache",
            "notes": "Unit mismatch: reported_effort in HOURS, month_norm/budget_norm in MINUTES. Conversion: minutes/60 = hours."
        },
        {
            "id": "TC-STAT-062", "title": "Cache update after vacation creation — month_norm recalculated",
            "preconditions": "Employee without vacation in current month.\nSETUP: Note current statistic_report values.\nQuery: SELECT sr.employee_login, sr.month_norm FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE) AND sr.employee_login NOT IN (SELECT employee FROM ttt_vacation.vacation WHERE status != 'CANCELED' AND start_date >= DATE_TRUNC('month', CURRENT_DATE)) ORDER BY random() LIMIT 1",
            "steps": "SETUP: Note current month_norm for the employee\nSETUP: Via API — POST /api/vacation/v1/vacations to create a 5-day vacation in current month\n1. Wait 10 seconds for MQ event processing\nDB-CHECK: SELECT month_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = DATE_TRUNC('month', CURRENT_DATE)\n2. Verify month_norm has DECREASED by the vacation working hours\n3. Navigate to /statistics/employee-reports as ADMIN\n4. Find the employee\n5. Verify Norm column reflects the updated (reduced) value\nCLEANUP: Via API — DELETE the vacation. Wait for MQ processing. Verify norm reverts.",
            "expected": "Vacation creation triggers RabbitMQ event → month_norm recalculated within seconds. UI reflects the change.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "statistics-caffeine-caching-performance-3337.md §Cache Invalidation Triggers", "module": "statistics/cache",
            "notes": "MQ event: vacation create → month_norm recalc via RabbitMQ topic exchange. 10s wait for async processing."
        },
        {
            "id": "TC-STAT-063", "title": "Cache update after sick leave creation — month_norm recalculated",
            "preconditions": "Employee without sick leave in current month.\nSETUP: Note current month_norm.",
            "steps": "SETUP: Note current month_norm for the employee\nSETUP: Via API — POST /api/vacation/v1/sick-leaves to create a 3-day sick leave\n1. Wait 10 seconds for event processing\nDB-CHECK: SELECT month_norm FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = DATE_TRUNC('month', CURRENT_DATE)\n2. Verify month_norm decreased by sick leave working hours\n3. If sick leave spans 2 months, verify BOTH months' records updated\nCLEANUP: Via API — DELETE the sick leave. Verify norm reverts.",
            "expected": "Sick leave event triggers month_norm recalculation. Two-month sick leaves update both months separately.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-caffeine-caching-performance-3337.md §!5200 Sick Leave Event Handlers", "module": "statistics/cache",
            "notes": "!5200: SickLeaveCreated/Changed/DeletedEventListener. Two-month handling via separate events."
        },
        {
            "id": "TC-STAT-064", "title": "Cache update after task report — reported_effort updated",
            "preconditions": "Employee with existing statistic_report record.\nSETUP: Note current reported_effort.",
            "steps": "SETUP: Note reported_effort for the employee in current month\nSETUP: Via API — submit a task report (e.g., 4 hours for a task)\n1. Wait 5 seconds for @TransactionalEventListener processing\nDB-CHECK: SELECT reported_effort FROM ttt_backend.statistic_report WHERE employee_login = '<login>' AND report_date = DATE_TRUNC('month', CURRENT_DATE)\n2. Verify reported_effort increased by the submitted hours\n3. Navigate to Employee Reports\n4. Verify Reported column shows updated value\nCLEANUP: Via API — delete the task report.",
            "expected": "Task report add triggers reported_effort update via @TransactionalEventListener. Immediate (not waiting for cron).",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Three Update Paths, statistics-caffeine-caching-performance-3337.md §TaskReportEventListener", "module": "statistics/cache",
            "notes": "TaskReportEventListener: add/patch/delete trigger effort recalc for employee/month."
        },
        {
            "id": "TC-STAT-065", "title": "Dismissed employee — no ghost records after termination (#3345)",
            "preconditions": "Recently dismissed employee.\nQuery: SELECT ep.employee_login, ep.period_end FROM ttt_vacation.employee_period ep WHERE ep.period_end IS NOT NULL AND ep.period_end >= CURRENT_DATE - INTERVAL '3 months' ORDER BY ep.period_end DESC LIMIT 5",
            "steps": "DB-CHECK: SELECT report_date FROM ttt_backend.statistic_report WHERE employee_login = '<dismissed_login>' ORDER BY report_date DESC LIMIT 5\n1. Verify no statistic_report records exist for months AFTER the employee's termination date\n2. Verify records exist for months BEFORE termination\n3. Verify the last record's report_date is <= termination month\n4. Navigate to Employee Reports as ADMIN\n5. Select the month after termination\n6. Verify the employee is NOT listed",
            "expected": "No ghost records for months after termination. #3345 fix: dismissed employees cleaned up during sync.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "statistics-service-implementation.md §Cache Table Sync, #3345", "module": "statistics/cache",
            "notes": "#3345: dismissed employees had records for months after termination (ghost data). Fixed during sync."
        },
        {
            "id": "TC-STAT-066", "title": "Manual sync trigger — test endpoint (non-prod)",
            "preconditions": "Non-production environment (qa-1 or timemachine). API access.\nSETUP: Note current statistic_report row count.",
            "steps": "SETUP: COUNT(*) FROM ttt_backend.statistic_report\n1. Via API — POST /api/ttt/test/v1/statistic-reports (trigger optimized sync)\n2. Verify 200 response\n3. Wait 30 seconds for sync to complete\nDB-CHECK: COUNT(*) FROM ttt_backend.statistic_report\n4. Verify row count is same or slightly changed (sync is idempotent)\n5. Verify month_norm_updated_at and reported_updated_at timestamps updated for current month records",
            "expected": "Manual sync trigger works on non-prod. Sync is idempotent — re-running doesn't duplicate data.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "statistics-caffeine-caching-performance-3337.md §!5013, statistics-service-implementation.md §Three Update Paths", "module": "statistics/cache",
            "notes": "Test endpoint: POST /v1/test/statistic-reports. Non-prod only. Triggers optimized sync (current + previous month)."
        },
    ]


def get_regression_cases():
    """TS-Stat-Regression: Bug regression tests from 180+ tickets."""
    return [
        {
            "id": "TC-STAT-067", "title": "#3195 Bug 4 — employee name click should open CS page, not report page",
            "preconditions": "Privileged user with Employee Reports access.",
            "steps": "1. Login as the privileged user\n2. Navigate to /statistics/employee-reports\n3. Click an employee's NAME in the Employee column\n4. Verify it opens CompanyStaff profile page (not report page)\n5. Verify URL is https://companystaff.noveogroup.com/profile/{login}\n6. Go back to Employee Reports\n7. Hover over the employee row to reveal report icon\n8. Click the report icon\n9. Verify THIS navigates to the employee's report page",
            "expected": "Name click → CS profile. Report icon click → report page. Two separate navigation targets.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3195 Bug 4", "module": "statistics/regression",
            "notes": "#3195 QA: Bug 4 was 'employee name click opens report page (should open CS page)'. Fixed."
        },
        {
            "id": "TC-STAT-068", "title": "#3195 Bug 9 — English translation for 'Only overage' switch",
            "preconditions": "Privileged user. English language.",
            "steps": "1. Login as the privileged user\n2. Switch language to English (if not already)\n3. Navigate to /statistics/employee-reports\n4. Find the 'Only excess over limit' toggle switch\n5. Verify it has English text (not Russian 'Только превышения')\n6. Switch language to Russian\n7. Verify Russian text appears for the toggle",
            "expected": "Toggle has proper English translation: 'Only excess over limit' (or similar). No untranslated Russian text in EN mode.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3195 Bug 9", "module": "statistics/regression",
            "notes": "#3195 Bug 9: No English translation for the switcher. Should be fixed."
        },
        {
            "id": "TC-STAT-069", "title": "#3320 — future employees excluded from list",
            "preconditions": "ADMIN user. Employee with hire date in the future (not yet started).\nQuery: SELECT ep.employee_login, ep.period_start FROM ttt_vacation.employee_period ep WHERE ep.period_start > CURRENT_DATE LIMIT 3",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the current month\n4. Search for the future employee by name/login\n5. Verify they are NOT shown in the list\n6. Select the month when the employee will start\n7. Verify the employee IS shown (if their start date is in that month)\n8. Verify no false under-report percentages for future employees",
            "expected": "Future employees (not yet hired) excluded from current month list. No false underreport data.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3320", "module": "statistics/regression",
            "notes": "#3320: filter now uses ttt_vacation.employee_period dates, not ttt_backend.employee.last_date."
        },
        {
            "id": "TC-STAT-070", "title": "#3380 — vacations affect personal monthly norm",
            "preconditions": "Employee with a vacation in the current month.\nSETUP: Via API — create a vacation if needed.",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month with the vacation\n4. Find the employee\n5. Verify personal norm is REDUCED by vacation working hours\n6. Verify norm column shows lower value than full office norm\n7. Compare with an employee without vacation in the same month\n8. Verify their norm is the full office norm",
            "expected": "Vacations reduce personal monthly norm. Employee with vacation: norm < office norm. Without vacation: norm = office norm.",
            "priority": "Critical", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3380", "module": "statistics/regression",
            "notes": "#3380 OPEN: vacations don't affect personal monthly norm (wrong API call in frontend). Sprint 15."
        },
        {
            "id": "TC-STAT-071", "title": "#2366 — 'No data' notification persistence after filter removal",
            "preconditions": "User with statistics access.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Apply a filter that returns NO data (e.g., search for non-existent employee name)\n4. Verify 'No data' notification appears\n5. Clear the filter (remove search text)\n6. Verify the 'No data' notification disappears\n7. Verify data reappears in the table\n8. Apply and remove the filter several times\n9. Verify notification state is always consistent with data availability",
            "expected": "'No data' notification appears when filter returns empty results and disappears when filter is removed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §Tier 8, #2366", "module": "statistics/regression",
            "notes": "#2366 OPEN: 'No data' notification persists after filter removal — state not cleared properly."
        },
        {
            "id": "TC-STAT-072", "title": "#1175 — outdated tabs on first load",
            "preconditions": "User with multiple roles (multiple tabs).",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general for the FIRST time in this session\n3. Verify tabs load correctly (no stale data from previous session)\n4. Verify the correct tab is auto-selected (or first available)\n5. Verify no flash of incorrect tab content\n6. Switch between tabs\n7. Verify each tab loads fresh data",
            "expected": "First load shows correct, current tabs. No stale data from redux-persist. Tab auto-selection works.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §Tier 8, #1175", "module": "statistics/regression",
            "notes": "#1175 OPEN: statisticsPermissionsTabName mostly commented out → currentTab=null. redux-persist may show stale filters."
        },
        {
            "id": "TC-STAT-073", "title": "#2716 — error on 'Refresh data' click",
            "preconditions": "User viewing General Statistics with data.",
            "steps": "1. Login as the user\n2. Navigate to /statistics/general\n3. Verify data is loaded\n4. Click 'Refresh data' button\n5. Verify NO error appears (no 500, no red notification)\n6. Verify data refreshes correctly\n7. Change date range\n8. Click 'Refresh data' again\n9. Verify data updates for the new range",
            "expected": "'Refresh data' button works without errors. Data refreshes for current filter/date selection.",
            "priority": "Medium", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §Tier 8, #2716", "module": "statistics/regression",
            "notes": "#2716 OPEN: error on 'Refresh data' click. Regression test."
        },
        {
            "id": "TC-STAT-074", "title": "#3368 — over/under report notification on Confirmation By Employee page",
            "preconditions": "User with confirmation access. Employee with over-report in current period.\nQuery: SELECT sr.employee_login FROM ttt_backend.statistic_report sr WHERE sr.report_date = DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') AND sr.reported_effort > sr.budget_norm / 60.0 + 10 LIMIT 5",
            "steps": "1. Login as DM or confirmation approver\n2. Navigate to 'Confirmation' page\n3. Switch to 'By Projects' view\n4. Verify over/under report notifications ARE shown for employees\n5. Switch to 'By Employee' view\n6. Verify over/under report notifications ARE ALSO shown (was missing)\n7. Verify the notification data matches Employee Reports page excess values",
            "expected": "'Confirmation > By Employee' shows over/under report notifications. Cross-page API call integration complete.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3368", "module": "statistics/regression",
            "notes": "#3368: 'By Employee' didn't call ttt/v1/statistic/report/employees API → notification missing. Sprint 15."
        },
        {
            "id": "TC-STAT-075", "title": "#3353 — rehired employee pre-employment norm display",
            "preconditions": "Employee who was rehired (multiple employment periods).\nQuery: SELECT ep.employee_login, COUNT(*) AS periods FROM ttt_vacation.employee_period ep GROUP BY ep.employee_login HAVING COUNT(*) > 1 LIMIT 5",
            "steps": "1. Login as ADMIN\n2. Navigate to /statistics/employee-reports\n3. Select the month of the GAP between employment periods\n4. Verify the rehired employee is NOT shown (not employed in that month)\n5. Select the month of re-hire\n6. Verify the employee IS shown with partial norm (from hire date)\n7. Select a month in the first employment period\n8. Verify correct norm for that period",
            "expected": "Rehired employees: gap months = not shown. Re-hire month = partial norm. Each period calculated separately.",
            "priority": "High", "type": "UI",
            "req_ref": "statistics-ticket-findings.md §#3353", "module": "statistics/regression",
            "notes": "#3353 OPEN: rehired employees show previously working days in orange. Pre-employment should be 0/0/{totalNorm}."
        },
        {
            "id": "TC-STAT-076", "title": "API error handling — missing required params return 500 instead of 400",
            "preconditions": "API access.",
            "steps": "1. Via API — GET /api/ttt/v1/reports/summary (WITHOUT required 'login' param)\n2. Verify response is 500 (MissingServletRequestParameterException)\n3. Note: should be 400 Bad Request\n4. Via API — GET /api/ttt/v1/reports/effort (WITHOUT required 'taskId' param)\n5. Verify response is 500 (same bug)\n6. Via API — GET /api/ttt/v1/reports/total (WITHOUT required 'type' param)\n7. Verify response is 400 (correct — uses DTO validation)\n8. Document the inconsistency",
            "expected": "Missing @RequestParam returns 500 (bug). Missing @Valid DTO params return 400 (correct). Inconsistent error handling.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "statistics-api-testing.md §Error Handling Bugs", "module": "statistics/regression",
            "notes": "Systematic: all @RequestParam-required endpoints return 500 on missing params. DTO-bound @Valid properly returns 400."
        },
    ]


# ─── Plan, Feature Matrix, Risk Assessment ──────────────────────────────────

PLAN_OVERVIEW = {
    "title": "Statistics Module — Test Plan",
    "scope": "Comprehensive testing of the Statistics module covering two sub-systems: Classic Statistics (General Statistics page with 13 permission-gated tabs, tree/flat views, search, export) and Employee Reports (norm/excess display, comment field, manager column, over-limit toggle). Also covers: budgetNorm calculation (admin vacation + family sick leave), partial-month employees, norm=0 corner cases, hour sum consistency (8 tickets), WSR tree view, CSV export features, cache table synchronization, and regression tests for 180+ mined tickets.",
    "objectives": [
        "Verify General Statistics page: 13 role-gated tabs, date range/period presets, tree/flat mode, hours/days toggle, search with keyboard layout correction",
        "Verify Employee Reports page: search, month picker, over-limit toggle, norm/excess display, comment field (inline CRUD), sorting, manager column filter",
        "Test budgetNorm calculation: admin vacation inclusion, family sick leave (#3409), personalNorm vs budgetNorm display rules (3 cases)",
        "Test norm edge cases: partial-month employees (#3356), terminated employees (#3320), rehired employees (#3353), norm=0 with +N/A%",
        "Test hour sum consistency: parent vs child total mismatches due to fired employee inclusion across all 13 tabs",
        "Verify export features: CSV download, hours/days unit, copy/link, admin-only exports, WSR tree view (5 open bugs)",
        "Test permission matrix: 8 roles × tab visibility × data scoping (ADMIN, CHIEF_ACCOUNTANT, OFFICE_ACCOUNTANT, DM, TL, HR, PM, EMPLOYEE)",
        "Verify cache table (statistic_report) sync: vacation/sick leave events, task report events, cron sync, dismissed employee handling",
        "Regression tests: #3195 (19 bugs), #3306 (toggle), #3298 (search regression), #3380 (vacation norm), #2366 (no data), #3368 (confirmation page)",
    ],
    "environments": [
        "Primary: qa-1 (ttt-qa-1.noveogroup.com) — cache table may NOT exist on qa-1 (computes on-the-fly)",
        "Secondary: timemachine (ttt-timemachine.noveogroup.com) — has statistic_report table (9662+ rows)",
        "Production-like: stage (ttt-stage.noveogroup.com) — for verification only",
    ],
    "approach": "UI-first testing with API/DB verification for cache and norm calculations. Tests describe browser actions for all user-facing scenarios. API steps used for cache verification (DB-CHECK), state setup, and API-only regression tests. Preconditions include SQL hints for dynamic test data. Critical: Unit mismatch across APIs — reported_effort in HOURS, month_norm/budget_norm in MINUTES.",
    "dependencies": [
        "CAS authentication for multi-role testing (ADMIN, DM, TL, PM, HR, EMPLOYEE, ACCOUNTANT roles needed)",
        "Admin settings: notification.reporting.over and notification.reporting.under must be configured for threshold tests",
        "#3409 deployment for family member sick leave budgetNorm tests (Sprint 16)",
        "RabbitMQ connectivity for cache invalidation event testing",
        "Production calendar data for norm calculation and partial-month tests",
    ],
}

FEATURE_MATRIX = [
    # (Feature, ClassicGen, EmpRpt, NormExc, ExportWSR, Perms, HourSum, Cache, Regress, Total)
    ("Tab visibility & permissions", 2, 2, 0, 0, 8, 0, 0, 0, 12),
    ("Date range & period controls", 2, 1, 0, 0, 0, 0, 0, 0, 3),
    ("Tree/Flat mode & sorting", 2, 1, 0, 0, 0, 6, 0, 0, 9),
    ("Search & filtering", 1, 1, 0, 0, 1, 0, 0, 1, 4),
    ("Absence icons", 1, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Norm display (budgetNorm)", 0, 1, 4, 0, 0, 0, 0, 2, 7),
    ("Excess display & color", 0, 2, 3, 0, 0, 0, 0, 0, 5),
    ("Partial-month & terminated", 0, 0, 2, 0, 0, 0, 1, 2, 5),
    ("Over-limit toggle", 0, 0, 1, 0, 0, 0, 0, 0, 1),
    ("Comment field", 0, 2, 0, 0, 0, 0, 0, 0, 2),
    ("CSV export & copy", 0, 0, 0, 4, 0, 0, 0, 0, 4),
    ("WSR tree view", 0, 0, 0, 3, 0, 0, 0, 0, 3),
    ("Admin-only exports", 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Norm export (#3400)", 0, 0, 0, 1, 0, 0, 0, 0, 1),
    ("Hour sum consistency", 0, 0, 0, 0, 0, 6, 0, 0, 6),
    ("Cache table sync", 0, 0, 0, 0, 0, 0, 6, 0, 6),
    ("Employee row features", 2, 4, 0, 0, 0, 0, 0, 1, 7),
    ("Manager column", 0, 1, 0, 0, 0, 0, 0, 0, 1),
    ("Cross-page notification", 0, 0, 0, 0, 0, 0, 0, 1, 1),
    ("API error handling", 0, 0, 0, 0, 0, 0, 0, 1, 1),
    ("Refresh data", 1, 0, 0, 0, 0, 0, 0, 1, 2),
    ("Reset filters", 1, 0, 0, 0, 0, 0, 0, 1, 2),
]

RISK_ASSESSMENT = [
    ("Hour Sum Consistency", "Parent-child total mismatches across all tree tabs due to fired employee inclusion in parent totals but exclusion from child lists. 8 tickets documenting extreme mismatches (up to 205 vs 36).", "High", "High", "Critical",
     "Test all 13 tabs in tree mode: expand and sum children vs parent. Verify showFired=true resolves mismatches. Document remaining discrepancies."),
    ("budgetNorm Calculation", "Complex formula: budgetNorm = individualNorm + admin_vacation_hrs + familyMember_sickleave_hrs. Multiple conditions affect display format. Excess uses budgetNorm as denominator (design issue: admin vacation penalizes employee).", "High", "High", "Critical",
     "Test all 3 display cases. Verify formula with DB values. Test #3409 family sick leave. Verify excess denominator."),
    ("Cache Table Sync", "Pre-computed statistic_report table must stay consistent with live data. Race condition: no pessimistic locking between MQ events and task report events. Dismissed employees had ghost records (#3345).", "High", "High", "Critical",
     "Verify cache after vacation/sick leave/task report events. Check dismissed employee records. Test manual sync trigger."),
    ("Partial-Month Employees", "Mid-month hires/terminations need effectiveBounds clamping. Rehired employees need multiple period handling. Pre-employment periods show incorrect norms.", "Medium", "High", "Critical",
     "Test mid-month hire norm, termination norm, rehired employee across periods, pre-employment display."),
    ("Permission Matrix Complexity", "8 roles × 13 tabs × different data scoping. HR hotfix #3247 caused search regression #3298. CEO bypass hardcoded.", "High", "High", "Critical",
     "Test each role's tab visibility and data scope. Verify HR scoping. Test search after permission changes."),
    ("norm=0 Corner Cases", "norm=0 + reported>0 → +N/A% (sorted as max). norm=0 + reported=0 → 0%. Was displaying 0.00% instead of +N/A% (#3195 Bug 7).", "Medium", "High", "High",
     "Find employees with norm=0 in DB. Verify +N/A% display and sort position. Verify 0% for zero/zero."),
    ("WSR Tree View (5 Open Bugs)", "Layout broken (#3030), branches don't collapse (#3041), wrong period data (#3144), 6 sub-issues (#3289). Actively broken feature area.", "High", "Medium", "High",
     "Test WSR basic rendering, expand/collapse, update button, tooltip translations, list behavior."),
    ("Over-Limit Toggle (#3306)", "Toggle ON should filter by configured thresholds. Was completely broken. Prerequisites: admin settings must be configured.", "Medium", "High", "High",
     "Test toggle with configured thresholds. Verify filtered list vs full list. Test with no thresholds configured."),
    ("Comment Field Data Loss", "Comments save on blur/Tab only. No auto-save on browser tab close. Per-employee per-month storage. Enter = newline (not save).", "Medium", "Medium", "Medium",
     "Test save on blur, save on Tab, Enter for newline. Verify per-month isolation. Document data loss risk."),
    ("Export Regressions", "CSV export had 400 on empty params (#2191), 404 on endpoints (#1492). Units parameter (#1422). Multiple past regressions.", "Medium", "Medium", "Medium",
     "Test CSV with empty params, with/without unit param. Verify all export endpoints return data (not 404/400)."),
    ("Stale Redux Persist", "redux-persist preserves Classic Statistics filters across sessions. First load may show stale tabs or wrong data.", "Medium", "Low", "Medium",
     "Test fresh login → first visit to statistics. Verify no stale filters. Test tab auto-selection."),
]


# ─── Workbook Generation ─────────────────────────────────────────────────────

def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 0 else FILL_ROW_WHITE


def write_test_cases(ws, cases, back_link_tab="Plan Overview"):
    """Write test cases to a worksheet."""
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [14, 40, 45, 65, 45, 10, 10, 35, 22, 35]

    # Back-link row
    ws.merge_cells("A1:J1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = f"← Back to {back_link_tab}"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = f"#'{back_link_tab}'!A1"

    # Headers in row 2
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # Data rows
    for row_idx, tc in enumerate(cases, 3):
        values = [tc["id"], tc["title"], tc["preconditions"], tc["steps"],
                  tc["expected"], tc["priority"], tc["type"],
                  tc["req_ref"], tc["module"], tc["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):  # Priority, Type
                cell.alignment = ALIGN_CENTER

    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    """Write Plan Overview tab."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"• {env}").font = FONT_BODY
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"• {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws, suite_tabs):
    """Write Feature Matrix tab."""
    headers = ["Feature", "Classic", "EmpRpt", "Norm", "Export", "Perms",
               "HourSum", "Cache", "Regress", "Total"]
    col_widths = [30, 10, 10, 10, 10, 10, 10, 10, 10, 8]

    # Back-link
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    # Headers in row 2
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    """Write Risk Assessment tab."""
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def generate():
    """Generate the statistics test documentation workbook."""
    wb = Workbook()

    suites = [
        ("TS-Stat-ClassicGeneral", "TS-Stat-ClassicGenrl", get_classic_general_cases),
        ("TS-Stat-EmployeeReports", "TS-Stat-EmpReports", get_employee_reports_cases),
        ("TS-Stat-NormExcess", "TS-Stat-NormExcess", get_norm_excess_cases),
        ("TS-Stat-ExportWSR", "TS-Stat-ExportWSR", get_export_wsr_cases),
        ("TS-Stat-Permissions", "TS-Stat-Permissions", get_permissions_cases),
        ("TS-Stat-HourSum", "TS-Stat-HourSum", get_hour_sum_cases),
        ("TS-Stat-CacheSync", "TS-Stat-CacheSync", get_cache_sync_cases),
        ("TS-Stat-Regression", "TS-Stat-Regression", get_regression_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    total_cases = sum(info[1] for info in suites_info)
    print(f"Generating {total_cases} test cases across {len(suites)} suites")

    suite_tabs = [tab for _, tab, _ in suites]

    # Plan Overview (default first sheet)
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(ws_plan, suites_info)

    # Feature Matrix
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_feature_matrix(ws_fm, suite_tabs)

    # Risk Assessment
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ws_ra)

    # Test Suite tabs
    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_test_cases(ws, cases)
        print(f"  {tab_name}: {len(cases)} cases")

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {total_cases} test cases, {len(suites)} suites, 3 plan tabs")

    return all_cases


if __name__ == "__main__":
    generate()
