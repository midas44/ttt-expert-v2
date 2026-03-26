#!/usr/bin/env python3
"""
Generator for ticket #3404 test documentation.
[Days off] Allow moving days off to earlier dates within an open month.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'test-docs', 't3404')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 't3404.xlsx')

# Styles
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
PLAN_HEADER_FILL = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
ROW_ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
BODY_FONT = Font(name='Arial', size=10)
LINK_FONT = Font(name='Arial', size=10, color='0563C1', underline='single')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3'),
)

# Test cases organized by suite
SUITES = {
    'TS-T3404-Tooltip': {
        'name': 'TS-T3404-Tooltip',
        'focus': 'P.4 — Tooltip Translation Fix',
        'cases': [
            {
                'id': 'TC-T3404-001',
                'title': 'EN tooltip text: "Reschedule event" (no "an")',
                'preconditions': 'Logged in as employee with day-offs in open period. UI language = EN.',
                'steps': (
                    '1. Navigate to My Vacations page → Days Off tab\n'
                    '2. Find a day-off row with edit icon visible\n'
                    '3. Hover over the edit (pencil) icon\n'
                    '4. Verify tooltip text reads "Reschedule event"'
                ),
                'expected': 'Tooltip shows "Reschedule event" (not "Reschedule an event").',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.4',
                'module': 'calendar-dayoff',
                'notes': 'Translation key: vacation.weekends.rescheduleAnEvent',
            },
            {
                'id': 'TC-T3404-002',
                'title': 'EN dialog title: "Reschedule event"',
                'preconditions': 'Logged in as employee with editable day-off. UI language = EN.',
                'steps': (
                    '1. Navigate to My Vacations page → Days Off tab\n'
                    '2. Click edit icon on an editable day-off row\n'
                    '3. Verify the dialog title reads "Reschedule event"'
                ),
                'expected': 'Dialog title is "Reschedule event".',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.4',
                'module': 'calendar-dayoff',
                'notes': 'Same i18n key used for both tooltip and dialog title.',
            },
            {
                'id': 'TC-T3404-003',
                'title': 'RU tooltip text: "Перенести событие" (unchanged)',
                'preconditions': 'Logged in as employee with day-offs. UI language = RU.',
                'steps': (
                    '1. Navigate to Мои отпуска → tab "Выходные"\n'
                    '2. Hover over the edit icon on an editable day-off\n'
                    '3. Verify tooltip text reads "Перенести событие"'
                ),
                'expected': 'RU tooltip unchanged: "Перенести событие".',
                'priority': 'P3',
                'type': 'UI',
                'req_ref': '#3404 p.4',
                'module': 'calendar-dayoff',
                'notes': 'Regression check — RU translation should not be affected.',
            },
        ],
    },
    'TS-T3404-EditIcon': {
        'name': 'TS-T3404-EditIcon',
        'focus': 'P.1 — Edit Action Availability (Approve Period-Based)',
        'cases': [
            {
                'id': 'TC-T3404-004',
                'title': 'Edit icon visible for day-off in current open month (future date)',
                'preconditions': 'Employee with day-offs in current month (approve period open). Query: SELECT e.login FROM ttt_backend.employee e JOIN ttt_vacation.employee_dayoff_request edr ON edr.employee = e.id WHERE edr.last_approved_date >= (SELECT start_date FROM ttt_backend.office_period WHERE type = \'APPROVE\' LIMIT 1) AND edr.status NOT IN (\'DELETED\', \'DELETED_FROM_CALENDAR\') AND edr.last_approved_date > CURRENT_DATE AND e.enabled = true LIMIT 5',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Find a day-off with date in current open month that is after today\n'
                    '4. Verify edit (pencil) icon is visible in the Actions column'
                ),
                'expected': 'Edit icon is visible for future day-offs in open approve period.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.1',
                'module': 'calendar-dayoff',
                'notes': 'Baseline check — this behavior existed before #3404.',
            },
            {
                'id': 'TC-T3404-005',
                'title': 'Edit icon visible for PAST day-off in open month (core new behavior)',
                'preconditions': 'Employee with a day-off whose lastApprovedDate is in the current month but BEFORE today. Approve period for employee office must be open. Query: SELECT e.login, edr.last_approved_date FROM ttt_backend.employee e JOIN ttt_vacation.employee_dayoff_request edr ON edr.employee = e.id WHERE edr.last_approved_date >= (SELECT start_date FROM ttt_backend.office_period WHERE type = \'APPROVE\' LIMIT 1) AND edr.last_approved_date < CURRENT_DATE AND edr.status NOT IN (\'DELETED\', \'DELETED_FROM_CALENDAR\') AND e.enabled = true ORDER BY edr.last_approved_date DESC LIMIT 5',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Find the day-off with a past date (within current open month)\n'
                    '4. Verify edit icon IS visible in the Actions column\n'
                    '5. Verify the day-off is not greyed out in the table'
                ),
                'expected': 'Edit icon visible for past day-offs within open approve period. This is the core new behavior from #3404.',
                'priority': 'P0',
                'type': 'UI',
                'req_ref': '#3404 p.1',
                'module': 'calendar-dayoff',
                'notes': 'CORE TEST — previously edit icon was hidden for past day-offs.',
            },
            {
                'id': 'TC-T3404-006',
                'title': 'Edit icon HIDDEN for day-off in closed month',
                'preconditions': 'Employee with day-offs in January or February (months before the approve period). Same employee as TC-T3404-005 preferred.',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Switch year selector if needed to show day-offs in closed months (Jan, Feb)\n'
                    '4. Find a day-off with date in January or February\n'
                    '5. Verify edit icon is NOT visible in the Actions column'
                ),
                'expected': 'No edit icon for day-offs in closed approve period months.',
                'priority': 'P0',
                'type': 'UI',
                'req_ref': '#3404 p.1',
                'module': 'calendar-dayoff',
                'notes': 'Critical negative test — closed months must not be editable.',
            },
            {
                'id': 'TC-T3404-007',
                'title': 'Boundary: day-off ON approve period start date',
                'preconditions': 'Need a day-off whose lastApprovedDate equals the approve period start date exactly. This requires a holiday on the 1st of a month. May 1 (Labour Day, Russia) is a candidate if the test is run when May approve period is open.',
                'steps': (
                    '1. Login as employee with a day-off on May 1 (or another 1st-of-month holiday)\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Find the day-off on May 1\n'
                    '4. Verify whether edit icon is visible'
                ),
                'expected': 'Edit icon SHOULD be visible (day-off is in open period). NOTE: Code uses > (not >=) — ST-1 risk — edit icon may be MISSING.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.1, ST-1',
                'module': 'calendar-dayoff',
                'notes': 'Tests ST-1 boundary bug. useWeekendTableHeaders.tsx:113 uses strict > instead of >=. On qa-1, March 1 is a Sunday (no holiday). May 1 is a better candidate.',
            },
            {
                'id': 'TC-T3404-008',
                'title': 'Edit icon hidden for last day of closed month',
                'preconditions': 'Employee with a day-off on Feb 23 (Defender Day, Russia) or last day of another closed month.',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Find the day-off on Feb 23\n'
                    '4. Verify edit icon is NOT visible'
                ),
                'expected': 'No edit icon — Feb 23 is in a closed approve period.',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.1',
                'module': 'calendar-dayoff',
                'notes': 'Boundary: last actionable day of a closed month.',
            },
            {
                'id': 'TC-T3404-009',
                'title': 'Previous year day-offs — all edit icons hidden',
                'preconditions': 'Employee who had day-offs in previous year (2025).',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Switch year selector to 2025\n'
                    '4. Verify NO edit icons are visible for any 2025 day-offs'
                ),
                'expected': 'All 2025 day-offs have no edit icon (previous year is fully closed).',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.1',
                'module': 'calendar-dayoff',
                'notes': 'Regression — previous year should never be editable.',
            },
        ],
    },
    'TS-T3404-Datepicker': {
        'name': 'TS-T3404-Datepicker',
        'focus': 'P.2.4 — Datepicker Date Constraints',
        'cases': [
            {
                'id': 'TC-T3404-010',
                'title': 'Closed month (January): all dates disabled in datepicker',
                'preconditions': 'Employee with an editable day-off in open month.',
                'steps': (
                    '1. Login as employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Click edit on an editable day-off\n'
                    '4. In the datepicker, navigate backward to January\n'
                    '5. Verify ALL January dates are disabled (greyed out / not selectable)'
                ),
                'expected': 'All January dates disabled — January approve period is closed.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'minDate is set to approvePeriod - 1 day. January is below minDate, so entire month should be unreachable or disabled.',
            },
            {
                'id': 'TC-T3404-011',
                'title': 'Closed month (February): all dates disabled in datepicker',
                'preconditions': 'Same as TC-T3404-010.',
                'steps': (
                    '1. Login as employee\n'
                    '2. Click edit on an editable day-off\n'
                    '3. Navigate datepicker to February\n'
                    '4. Verify ALL February dates are disabled'
                ),
                'expected': 'All February dates disabled — February approve period is closed.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'Feb 28 may be the actual minDate (approvePeriod 2026-03-01 minus 1 day). Feb 28 is Saturday — disabled by renderDay.',
            },
            {
                'id': 'TC-T3404-012',
                'title': 'Open month (March): working days are enabled',
                'preconditions': 'Same as TC-T3404-010. Approve period starts March 1.',
                'steps': (
                    '1. Login as employee\n'
                    '2. Click edit on an editable day-off in March\n'
                    '3. In the datepicker, view March dates\n'
                    '4. Verify working days (Mon-Fri, not holidays) are enabled/selectable\n'
                    '5. Verify weekends (Sat, Sun) are disabled\n'
                    '6. Verify dates with existing day-offs are disabled'
                ),
                'expected': 'March working days are enabled. Weekends and existing day-off dates are disabled.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'Positive test — dates in open period should be selectable.',
            },
            {
                'id': 'TC-T3404-013',
                'title': 'Future month (April): working days are enabled',
                'preconditions': 'Same as TC-T3404-010.',
                'steps': (
                    '1. Login as employee\n'
                    '2. Click edit on an editable day-off\n'
                    '3. Navigate datepicker forward to April\n'
                    '4. Verify working days are enabled, weekends disabled'
                ),
                'expected': 'April working days are enabled.',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'Future months with open approve period should always work.',
            },
            {
                'id': 'TC-T3404-014',
                'title': 'Boundary: Feb 28 (day before approve period) is disabled',
                'preconditions': 'Same as TC-T3404-010. Approve period starts March 1.',
                'steps': (
                    '1. Click edit on an editable day-off\n'
                    '2. Navigate datepicker to show Feb 28\n'
                    '3. Verify Feb 28 is disabled (not selectable)'
                ),
                'expected': 'Feb 28 is disabled. It is the day before the approve period start AND a Saturday.',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'minDate = moment(approvePeriod).subtract(1, "d") = Feb 28. Feb 28 is Saturday → double-disabled.',
            },
            {
                'id': 'TC-T3404-015',
                'title': 'Boundary: March 2 (first working day of open period) is enabled',
                'preconditions': 'Same as TC-T3404-010. March 1 is Sunday, March 2 is Monday.',
                'steps': (
                    '1. Click edit on an editable day-off\n'
                    '2. In the datepicker, locate March 2\n'
                    '3. Verify March 2 is enabled (selectable)'
                ),
                'expected': 'March 2 is the first working day of the open period — it should be selectable.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub2',
                'module': 'calendar-dayoff',
                'notes': 'March 1 is Sunday (disabled). March 2 is the practical minimum selectable date.',
            },
        ],
    },
    'TS-T3404-EarlierDate': {
        'name': 'TS-T3404-EarlierDate',
        'focus': 'P.2.4/sub4 — Earlier Date Constraint Relaxation',
        'cases': [
            {
                'id': 'TC-T3404-016',
                'title': 'Select earlier date within same month (core new behavior)',
                'preconditions': 'Employee with a past day-off in open month (e.g., March 12). Edit icon is visible.',
                'steps': (
                    '1. Login as the employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Click edit icon on the March 12 day-off\n'
                    '4. In the datepicker, select March 3 (an earlier working day)\n'
                    '5. Verify the selected date appears in the dialog as the new transfer date\n'
                    '6. Verify the "OK" button is enabled'
                ),
                'expected': 'March 3 is selectable. Transfer date shows "12.03.2026 → 03.03.2026". OK button is enabled.',
                'priority': 'P0',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub4',
                'module': 'calendar-dayoff',
                'notes': 'CORE TEST — previously dates before the original date were disabled.',
            },
            {
                'id': 'TC-T3404-017',
                'title': 'First working day of month is selectable',
                'preconditions': 'Same as TC-T3404-016.',
                'steps': (
                    '1. Open the reschedule dialog for a mid-month day-off\n'
                    '2. In the datepicker, verify March 2 (first working day of March) is selectable\n'
                    '3. Click March 2\n'
                    '4. Verify it is accepted as the transfer target'
                ),
                'expected': 'March 2 is selectable as transfer target date.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub4',
                'module': 'calendar-dayoff',
                'notes': 'Lower boundary of selectable dates within the open month.',
            },
            {
                'id': 'TC-T3404-018',
                'title': 'February dates NOT selectable (closed period)',
                'preconditions': 'Same as TC-T3404-016.',
                'steps': (
                    '1. Open the reschedule dialog for a March day-off\n'
                    '2. Navigate the datepicker backward to February\n'
                    '3. Verify ALL February dates are disabled/not selectable'
                ),
                'expected': 'No February dates are selectable — the month is in a closed approve period.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub4',
                'module': 'calendar-dayoff',
                'notes': 'Negative test — must not allow transfer to closed months.',
            },
            {
                'id': 'TC-T3404-019',
                'title': 'Future holiday: minDate still uses original date (ST-4)',
                'preconditions': 'Employee with a FUTURE day-off on a mid-month holiday (e.g., May 9 Victory Day in Russia). The day-off must be in the future (after today).',
                'steps': (
                    '1. Login as the employee (Russian calendar office)\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Click edit on the May 9 day-off\n'
                    '4. In the datepicker, try to select May 2 (first working day after May 1 holidays)\n'
                    '5. Verify whether May 2-8 working days are selectable or disabled'
                ),
                'expected': 'Per requirement: May 2+ should be selectable (1st of month = May 1). Per current code: May 2-8 are NOT selectable (minDate = May 9). ST-4 risk applies.',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404 p.2.4/sub4, ST-4',
                'module': 'calendar-dayoff',
                'notes': 'Tests ST-4 incomplete implementation. Code uses originalDate as minDate for future day-offs instead of startOf(month).',
            },
        ],
    },
    'TS-T3404-Regression': {
        'name': 'TS-T3404-Regression',
        'focus': 'P.2, P.3 — Regression and Side Effects',
        'cases': [
            {
                'id': 'TC-T3404-020',
                'title': 'E2E: full reschedule to earlier date + approval flow',
                'preconditions': (
                    'Employee with editable past day-off in open month, has a manager as approver.\n'
                    'SETUP: Via API — ensure employee has an untransferred day-off in open period (or verify existing one).'
                ),
                'steps': (
                    '1. Login as employee\n'
                    '2. Navigate to My Vacations → Days Off tab\n'
                    '3. Click edit on a past day-off in the open month\n'
                    '4. Select an earlier date within the same month\n'
                    '5. Click OK to submit the transfer request\n'
                    '6. Verify success notification and status changes to "New" in the table\n'
                    '7. Logout\n'
                    '8. Login as the approver (manager)\n'
                    '9. Navigate to My Vacations → Requests → Days off rescheduling tab\n'
                    '10. Find the transfer request in the "For Approval" sub-tab\n'
                    '11. Approve the request\n'
                    '12. Verify status changes to "Approved"\n'
                    'CLEANUP: Via API — delete or revert the transfer if needed'
                ),
                'expected': 'Full workflow completes: employee creates backward transfer → manager approves → status is Approved.',
                'priority': 'P1',
                'type': 'UI',
                'req_ref': '#3404 p.1, p.2.4/sub4',
                'module': 'calendar-dayoff',
                'notes': 'E2E test. Modifies data — needs cleanup. Tests the complete user journey.',
            },
            {
                'id': 'TC-T3404-021',
                'title': 'Month-close auto-rejection: backward transfer in closing month',
                'preconditions': (
                    'Employee with a NEW (pending) backward transfer request where both original date and target date are in the same month.\n'
                    'Admin access to change approve period.'
                ),
                'steps': (
                    'SETUP: Create a backward transfer request via UI (employee picks earlier date in open month)\n'
                    '1. Login as admin/accountant\n'
                    '2. Change the approve period for the employee\'s office to advance by one month (close the current month)\n'
                    '3. Verify the NEW backward transfer request is auto-rejected\n'
                    '4. Login as employee and verify the request shows status "Rejected"\n'
                    'CLEANUP: Revert the approve period to the original value'
                ),
                'expected': 'Auto-rejection fires on month close. The backward transfer request with both dates in the closed month is rejected.',
                'priority': 'P1',
                'type': 'hybrid',
                'req_ref': '#3404 p.2',
                'module': 'calendar-dayoff',
                'notes': 'Tests that the existing auto-rejection logic (unchanged per ticket) still works correctly for backward transfers.',
            },
            {
                'id': 'TC-T3404-022',
                'title': 'Vacation recalculation: transfer to date overlapping vacation',
                'preconditions': (
                    'Employee with a day-off and an active vacation that includes a date within the same month.\n'
                    'The transfer target date should overlap with the vacation period.'
                ),
                'steps': (
                    'SETUP: Via API — create a vacation for the employee covering several working days in the open month\n'
                    'SETUP: Via API — approve the vacation\n'
                    '1. Login as employee\n'
                    '2. Navigate to Days Off tab\n'
                    '3. Edit a day-off and select a date that falls within the approved vacation period\n'
                    '4. Submit the transfer\n'
                    '5. Have the manager approve the transfer\n'
                    '6. Verify the vacation is recalculated (days reduced or vacation auto-deleted if single day)\n'
                    'CLEANUP: Delete test vacation and transfer'
                ),
                'expected': 'Vacation recalculation is not broken by the new earlier-date logic. Day-off transfer onto vacation date triggers vacation adjustment.',
                'priority': 'P1',
                'type': 'hybrid',
                'req_ref': '#3404 p.3',
                'module': 'calendar-dayoff',
                'notes': 'Regression test for P.3 requirement. Related bugs: #2833 (vacation not recalculated), #3223 (balance not updated).',
            },
            {
                'id': 'TC-T3404-023',
                'title': 'Max date unchanged: Dec 31 of original year',
                'preconditions': 'Employee with an editable day-off.',
                'steps': (
                    '1. Login as employee\n'
                    '2. Click edit on a day-off\n'
                    '3. Navigate datepicker forward to December of the same year\n'
                    '4. Verify Dec 31 area is reachable but Jan 1 of next year is NOT selectable\n'
                    '5. Verify the maxDate boundary has not changed'
                ),
                'expected': 'maxDate = Dec 31 of the original date\'s year (unchanged behavior).',
                'priority': 'P2',
                'type': 'UI',
                'req_ref': '#3404',
                'module': 'calendar-dayoff',
                'notes': 'Regression — maxDate formula unchanged: moment(originalDate.format("YYYY")).add(1,"y").subtract(1,"d").',
            },
            {
                'id': 'TC-T3404-024',
                'title': 'Global approve period: employee in office with different period',
                'preconditions': 'Two offices with DIFFERENT approve periods (if available on env). Otherwise, note as untestable.',
                'steps': (
                    '1. Identify two employees in offices with different approve period start dates\n'
                    '2. Login as employee in the office with LATER approve period\n'
                    '3. Navigate to Days Off tab\n'
                    '4. Verify edit icon visibility matches the GLOBAL minimum approve period, not the employee\'s office period'
                ),
                'expected': 'Edit icon uses the minimum approve period across all offices (GET /v1/offices/periods/approve/min). More permissive than per-office.',
                'priority': 'P3',
                'type': 'UI',
                'req_ref': '#3404 p.1, ST-5',
                'module': 'calendar-dayoff',
                'notes': 'Tests ST-5 design note. On qa-1, all offices have same period (2026-03-01) — may be untestable without admin manipulation.',
            },
        ],
    },
}


def create_workbook():
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Plan Overview tab
    create_plan_overview(wb)

    # 2. Risk Assessment tab
    create_risk_assessment(wb)

    # 3. Test suite tabs
    for suite_key, suite in SUITES.items():
        create_suite_tab(wb, suite)

    # Set tab colors
    for ws in wb.worksheets:
        if ws.title.startswith('TS-'):
            ws.sheet_properties.tabColor = '2E75B6'
        else:
            ws.sheet_properties.tabColor = '548235'

    wb.save(OUTPUT_FILE)
    print(f'Generated: {OUTPUT_FILE}')
    total = sum(len(s['cases']) for s in SUITES.values())
    print(f'Total test cases: {total}')
    print(f'Suites: {len(SUITES)}')


def apply_header_style(ws, row, fill=None):
    _fill = fill or HEADER_FILL
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = _fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def create_plan_overview(wb):
    ws = wb.create_sheet('Plan Overview')

    # Title
    ws['A1'] = 'Test Plan — Ticket #3404'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws.merge_cells('A1:B1')

    rows = [
        ('Ticket', '#3404 — [Days off] Allow moving days off to earlier dates within an open month'),
        ('Sprint', '15'),
        ('Status', 'Ready to Test'),
        ('Module', 'calendar-dayoff (frontend-only change)'),
        ('Scope', '4 requirement changes: edit action visibility, datepicker constraints, tooltip fix'),
        ('Environment', 'qa-1 (approve period: 2026-03-01 for all offices)'),
        ('Approach', 'UI-focused testing. 4 changed files, all frontend. Backend API unchanged.'),
        ('Risk Areas', 'ST-1: boundary bug (> vs >=), ST-4: incomplete future day-off logic, ST-5: global approve period'),
        ('Related Tickets', '#2672 (Sprint 8: original rescheduling), #2874 (Sprint 9: backward transfers)'),
        ('MR', '!5333 (merged 2026-03-24, 4 files changed)'),
    ]

    for i, (key, val) in enumerate(rows, start=3):
        ws[f'A{i}'] = key
        ws[f'A{i}'].font = Font(name='Arial', bold=True, size=10)
        ws[f'B{i}'] = val
        ws[f'B{i}'].font = BODY_FONT
        ws[f'B{i}'].alignment = WRAP

    # Suite links
    link_row = len(rows) + 4
    ws[f'A{link_row}'] = 'Test Suites'
    ws[f'A{link_row}'].font = Font(name='Arial', bold=True, size=12)

    for i, (key, suite) in enumerate(SUITES.items(), start=link_row + 1):
        count = len(suite['cases'])
        ws[f'A{i}'] = f'{suite["name"]} — {suite["focus"]} ({count} cases)'
        ws[f'A{i}'].font = LINK_FONT
        ws[f'A{i}'].hyperlink = f"#'{suite['name']}'!A1"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80


def create_risk_assessment(wb):
    ws = wb.create_sheet('Risk Assessment')

    headers = ['ID', 'Risk', 'Likelihood', 'Impact', 'Severity', 'Mitigation / Test Focus']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, fill=PLAN_HEADER_FILL)

    risks = [
        ('ST-1', 'Boundary bug: edit icon missing for day-off exactly on approve period start', 'Low', 'Medium', 'Low',
         'Test TC-T3404-007 when May 1 (Labour Day) is available. Code uses > not >= in useWeekendTableHeaders.tsx:113.'),
        ('ST-4', 'Incomplete: future mid-month holidays use originalDate as minDate, not startOf(month)', 'Medium', 'Low', 'Low',
         'Test TC-T3404-019 with May 9 (Victory Day). Only matters for future mid-month holidays.'),
        ('ST-5', 'Global min approve period used instead of per-employee office period', 'Low', 'Low', 'Low',
         'Test TC-T3404-024 if different office periods available. Design choice — more permissive.'),
        ('REG-1', 'Vacation recalculation broken by earlier-date transfers', 'Low', 'High', 'Medium',
         'Test TC-T3404-022. Related: #2833 (prod hotfix). Must verify day-off onto vacation triggers recalc.'),
        ('REG-2', 'Month-close auto-rejection misses backward transfers', 'Low', 'High', 'Medium',
         'Test TC-T3404-021. Auto-rejection logic unchanged per ticket but must work with new backward transfers.'),
    ]

    for r, risk in enumerate(risks, 2):
        for c, val in enumerate(risk, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if r % 2 == 0:
                cell.fill = ROW_ALT_FILL

    widths = [8, 60, 10, 10, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_suite_tab(wb, suite):
    ws = wb.create_sheet(suite['name'])

    # Back link
    ws['A1'] = '← Back to Plan'
    ws['A1'].font = LINK_FONT
    ws['A1'].hyperlink = "#'Plan Overview'!A1"

    ws['A2'] = f'{suite["name"]} — {suite["focus"]}'
    ws['A2'].font = Font(name='Arial', bold=True, size=12)

    # Headers
    headers = ['Test ID', 'Title', 'Preconditions', 'Steps', 'Expected Result', 'Priority', 'Type', 'Requirement Ref', 'Module/Component', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    apply_header_style(ws, 4)

    # Auto-filter
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}4'

    # Data rows
    for r, tc in enumerate(suite['cases'], 5):
        values = [
            tc['id'], tc['title'], tc['preconditions'], tc['steps'],
            tc['expected'], tc['priority'], tc['type'], tc['req_ref'],
            tc['module'], tc['notes'],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if (r - 5) % 2 == 1:
                cell.fill = ROW_ALT_FILL

    # Column widths
    widths = [14, 35, 40, 55, 40, 8, 8, 18, 16, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == '__main__':
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    create_workbook()
