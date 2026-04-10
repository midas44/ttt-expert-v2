#!/usr/bin/env python3
"""
Planner Module Test Documentation Generator — Phase B
Generates test-docs/planner/planner.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 13 TS- test suite tabs.

Based on vault knowledge: 13 vault notes (15K+ words total), 130+ GitLab tickets mined,
20+ deep-dived with comments, 7/8 investigation methods used, all major subsystems mapped.
Close-by-tag tests are in separate t2724.xlsx — this file covers the broader planner module.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ──────���────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "planner")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "planner.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 1 else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_navigation_cases():
    """TS-Planner-Navigation: Tab switching, date, project selection, search."""
    return [
        {
            "id": "TC-PLN-001", "title": "Navigate to Planner from navbar",
            "preconditions": "Authenticated user with planner access.",
            "steps": "1. Login as any employee\n2. Click 'Planner' link in the top navbar\n3. Verify URL changes to /planner/TABS_ASSIGNMENTS_TASK\n4. Verify Tasks tab is active by default\n5. Verify search bar is visible with placeholder 'My project / my task, ticket number or URL 1,5h'",
            "expected": "Planner page loads with Tasks tab active. Search bar and 'Add a task' button visible.",
            "priority": "Critical", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Page Routes", "module": "planner/navigation",
            "notes": "URL uses hash-like route: /planner/TABS_ASSIGNMENTS_TASK"
        },
        {
            "id": "TC-PLN-002", "title": "Switch between Tasks and Projects tabs",
            "preconditions": "User on Planner page.",
            "steps": "1. Click 'Projects' tab button\n2. Verify URL changes to /planner/TABS_ASSIGNMENTS_PROJECT\n3. Verify project selector dropdown appears\n4. Click 'Tasks' tab button\n5. Verify URL changes back to /planner/TABS_ASSIGNMENTS_TASK\n6. Verify personal task table reappears",
            "expected": "Tab switching updates URL and content. Tasks shows personal assignments, Projects shows project-level view.",
            "priority": "Critical", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md", "module": "planner/navigation",
            "notes": ""
        },
        {
            "id": "TC-PLN-003", "title": "Navigate dates forward and backward",
            "preconditions": "User on Planner Tasks tab.",
            "steps": "1. Note the current date displayed in the table header\n2. Click the right arrow next to the date\n3. Verify date advances by one day\n4. Verify table content updates (assignments for the new date)\n5. Click the left arrow\n6. Verify date goes back one day\n7. Verify Total row recalculates for the new date's data",
            "expected": "Date navigation updates displayed date and assignment data. Total row reflects correct sums for selected date.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Tasks Tab", "module": "planner/navigation",
            "notes": "currentDay stored in Redux plannerTasks slice, persisted to localStorage."
        },
        {
            "id": "TC-PLN-004", "title": "Select a project in Projects tab",
            "preconditions": "User is PM or member of at least one project.",
            "steps": "1. Navigate to Projects tab\n2. Click the project selector dropdown\n3. Select a project (e.g., from the list)\n4. Verify table populates with employees assigned to that project\n5. Verify each employee row shows their assignments for the selected date\n6. Verify project name appears in the dropdown selection",
            "expected": "Project selection loads employees and their assignments. Table grouped by employee name.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Projects Tab", "module": "planner/navigation",
            "notes": "react-select dropdown with CSS prefix selectbox__."
        },
        {
            "id": "TC-PLN-005", "title": "Filter by role — 'Show projects where I am a...'",
            "preconditions": "User has multiple project roles (PM on one, member on another).",
            "steps": "1. Navigate to Projects tab\n2. Find the role filter dropdown\n3. Select 'PM' (or equivalent role filter)\n4. Verify project dropdown shows only projects where user is PM\n5. Change filter to 'member'\n6. Verify project dropdown updates to show projects where user is a member",
            "expected": "Role filter restricts visible projects. ProjectRolesType: SENIOR_MANAGER, MANAGER, MEMBER.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §plannerProjects", "module": "planner/navigation",
            "notes": ""
        },
        {
            "id": "TC-PLN-006", "title": "Search for task by name",
            "preconditions": "User has assignments.",
            "steps": "1. Navigate to Tasks tab\n2. In the search bar, type a known task name\n3. Click 'Add a task' button\n4. Verify task is added or search results appear\n5. Verify the search matches project name / task name / ticket number",
            "expected": "Search bar supports adding tasks by name, ticket number, or URL. Parsed format: 'project/task hours'.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Tasks Tab", "module": "planner/navigation",
            "notes": "Placeholder text: 'My project / my task, ticket number or URL 1,5h'"
        },
        {
            "id": "TC-PLN-007", "title": "Empty state — no assignments for date",
            "preconditions": "Employee with no assignments on a weekend/future date.",
            "steps": "1. Navigate to Tasks tab\n2. Navigate to a date with no assignments (e.g., a future weekend)\n3. Verify only header row and Total row (0) are visible\n4. Verify search bar and 'Add a task' button are still functional",
            "expected": "Empty state shows header and Total row with 0. No empty-state message (just empty table body).",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Empty State", "module": "planner/navigation",
            "notes": ""
        },
        {
            "id": "TC-PLN-008", "title": "Collapse and expand project groups in Tasks tab",
            "preconditions": "Employee with assignments in multiple projects.",
            "steps": "1. Navigate to Tasks tab\n2. Identify multiple project group headers\n3. Click a project group header to collapse it\n4. Verify tasks under that project are hidden\n5. Verify Total row still shows correct sum (including hidden tasks)\n6. Click again to expand\n7. Verify tasks reappear",
            "expected": "Project groups are collapsible. hiddenGroups persisted to localStorage via plannerTasks slice.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §plannerTasks hiddenGroups", "module": "planner/navigation",
            "notes": ""
        },
        {
            "id": "TC-PLN-009", "title": "WebSocket connection indicator",
            "preconditions": "User on Projects tab.",
            "steps": "1. Navigate to Projects tab, select a project\n2. Look for the cloud/connection icon next to tabs\n3. Click or hover the icon\n4. Verify it shows 'Connected' status\n5. Verify icon is visible in both read-only and editing modes",
            "expected": "WebSocket connection indicator shows real-time connection status. Connected via STOMP /ws endpoint.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md", "module": "planner/navigation",
            "notes": "Auto-reconnect delay: 500ms on disconnect."
        },
        {
            "id": "TC-PLN-010", "title": "Task view toggle — TASK vs TICKET",
            "preconditions": "User on Tasks tab.",
            "steps": "1. Navigate to Tasks tab\n2. Find the task view toggle (TASK/TICKET)\n3. Switch to TICKET view\n4. Verify table layout changes to show ticket-oriented data\n5. Switch back to TASK view\n6. Verify original layout restored\n7. Navigate away and return — verify toggle state persisted",
            "expected": "taskView toggle switches between TASK and TICKET display modes. Persisted to localStorage.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §plannerTasks taskView", "module": "planner/navigation",
            "notes": ""
        },
        {
            "id": "TC-PLN-011", "title": "Notification banners display correctly",
            "preconditions": "Conditions that trigger notifications (overdue day-off, norm exceeded).",
            "steps": "1. Login as employee who has exceeded hours norm\n2. Navigate to Planner\n3. Verify norm-exceeded banner: 'This month, you have exceeded the hours norm by X%. Norm: Y. Reported: Z'\n4. Login as employee with overdue day-off requests\n5. Verify banner: 'You have overdue day off rescheduling requests'\n6. Verify banners are dismissible or persistent",
            "expected": "Notification banners display based on employee state. Multiple banners can appear simultaneously.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Notifications/Banners", "module": "planner/navigation",
            "notes": "Observed banners: norm exceeded, overdue day-off, non-working days reminder."
        },
    ]


def get_task_crud_cases():
    """TS-Planner-TaskCRUD: Add, edit, delete assignments."""
    return [
        {
            "id": "TC-PLN-012", "title": "Add task via search bar — happy path",
            "preconditions": "Employee assigned to at least one project with tasks.\nQuery: SELECT e.login, p.name, t.name AS task_name FROM ttt_backend.employee e JOIN ttt_backend.project_member pm ON e.id = pm.employee_id JOIN ttt_backend.project p ON pm.project_id = p.id JOIN ttt_backend.task t ON t.project_id = p.id WHERE e.enabled = true AND p.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to Tasks tab\n3. In the search bar, type 'project-name/task-name'\n4. Click 'Add a task' button\n5. Verify new task row appears at the TOP of the project group\n6. Verify green highlight on new row (5 seconds)\n7. Verify auto-scroll to the new task position\n8. Verify Total row updates",
            "expected": "Task added at top of project group. Green highlight for 5s. Auto-scroll to new item. Total updates.",
            "priority": "Critical", "type": "UI",
            "req_ref": "Confluence §5.2, §8.4-8.6", "module": "planner/task-crud",
            "notes": "Search supports: project/task name, ticket number, tracker URL."
        },
        {
            "id": "TC-PLN-013", "title": "Edit hours in effort cell — inline editing",
            "preconditions": "Employee with an existing assignment for today.\nQuery: SELECT e.login FROM ttt_backend.task_assignment ta JOIN ttt_backend.employee e ON ta.employee_id = e.id WHERE ta.closed = false AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to Tasks tab\n3. Click on the hours cell for an existing assignment\n4. Verify cell becomes editable (input field)\n5. Enter '4.5'\n6. Press Enter or Tab to confirm\n7. Verify hours saved (cell shows 4.5)\n8. Verify Total row updates to reflect the new hours\nDB-CHECK: SELECT actual_efforts FROM ttt_backend.task_report WHERE task = {taskId} AND executor = {employeeId} AND report_date = '{date}'",
            "expected": "Inline editing of hours works. Value saved on Enter/Tab. Total recalculates. WebSocket PATCH event published.",
            "priority": "Critical", "type": "UI",
            "req_ref": "planner-lock-mechanism.md §Lockable fields", "module": "planner/task-crud",
            "notes": "Editing triggers lock acquisition. Other users see cell as locked."
        },
        {
            "id": "TC-PLN-014", "title": "Edit comment in comment cell",
            "preconditions": "Employee with an existing assignment.",
            "steps": "1. Login as the employee\n2. Navigate to Tasks tab\n3. Find the comment column for an assignment\n4. Click the edit button in the comment cell\n5. Type 'Test comment for task'\n6. Confirm the edit\n7. Verify comment is saved\n8. Hover over the comment cell — verify tooltip shows full text",
            "expected": "Comment editing works. Comment stored in task_report. Tooltip shows full text on hover.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §tooltips slice", "module": "planner/task-crud",
            "notes": "Comment cell has two buttons. Multi-line text supported."
        },
        {
            "id": "TC-PLN-015", "title": "Edit remaining estimate",
            "preconditions": "Employee with assignment on a project with tracker.",
            "steps": "1. Login as the employee\n2. Navigate to Tasks tab\n3. Find the 'Remaining work' column\n4. Click to edit the remaining estimate cell\n5. Enter '8'\n6. Confirm\n7. Verify value saved\nDB-CHECK: SELECT remaining_estimate FROM ttt_backend.task_assignment WHERE id = {assignmentId}",
            "expected": "Remaining estimate updated. Stored in task_assignment.remaining_estimate.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-lock-mechanism.md §Lockable fields: remainingEstimate", "module": "planner/task-crud",
            "notes": ""
        },
        {
            "id": "TC-PLN-016", "title": "Delete assignment",
            "preconditions": "Employee with a non-readOnly assignment.",
            "steps": "1. Login as the employee\n2. Navigate to Tasks tab\n3. Find an assignment row\n4. Click the delete/remove button on the row\n5. Confirm deletion if prompted\n6. Verify row removed from table\n7. Verify Total row recalculates\nDB-CHECK: Verify assignment deleted or marked closed",
            "expected": "Assignment removed from table. WebSocket DELETE event published. Total recalculates.",
            "priority": "High", "type": "UI",
            "req_ref": "#1790 add delete function", "module": "planner/task-crud",
            "notes": "Deletion is one-way — no un-delete mechanism (same as close-by-tag)."
        },
        {
            "id": "TC-PLN-017", "title": "'Open for editing' generates assignments for employee",
            "preconditions": "PM on Projects tab. Employee in readOnly state (not yet opened for editing). Report period is open.\nQuery: SELECT p.id, e.login AS manager FROM ttt_backend.project p JOIN ttt_backend.employee e ON p.manager_id = e.id WHERE e.enabled = true AND p.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, select a project\n3. Find an employee row that shows as read-only\n4. Click 'Open for editing' button on that employee's header row\n5. Verify assignments are generated for the employee\n6. Verify DnD handles appear on the new assignment rows\n7. Verify hours cells become editable",
            "expected": "'Open for editing' triggers POST /v1/assignments/generate. New assignments appear with DnD handles. Cells become editable.",
            "priority": "Critical", "type": "UI",
            "req_ref": "t2724-investigation.md §How Open for Editing Works", "module": "planner/task-crud",
            "notes": "readOnlyEmployee goes from true to false. WebSocket TaskAssignmentGenerateEvent published."
        },
        {
            "id": "TC-PLN-018", "title": "Edit hours in Projects tab (manager view)",
            "preconditions": "PM with editing mode active on an employee.",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, select project\n3. Open for editing on an employee\n4. Click hours cell for one of the employee's tasks\n5. Enter hours (e.g., '3')\n6. Confirm\n7. Verify hours saved\n8. Verify the employee sees the update in real-time via WebSocket",
            "expected": "Manager can edit employee hours in Projects tab after 'Open for editing'. WebSocket PATCH event published.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Editing Mode", "module": "planner/task-crud",
            "notes": "Per-employee editing can be toggled independently of global 'Open for editing'."
        },
        {
            "id": "TC-PLN-019", "title": "Color coding — blocked (red) and done (green)",
            "preconditions": "Assignments with different statuses (blocked, done).",
            "steps": "1. Navigate to Projects tab\n2. Select a project with mixed assignment statuses\n3. Verify blocked assignments show red/orange background\n4. Verify done assignments show green background\n5. Verify no-status assignments have no color",
            "expected": "Color coding: blocked=red/orange, done=green, in-progress=no color. CSS classes: planner__cel--color-blocked, planner__cel--color-done.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Color Coding", "module": "planner/task-crud",
            "notes": "Status derived from tracker ticket info."
        },
        {
            "id": "TC-PLN-020", "title": "Info column shows tracker priority tags",
            "preconditions": "Assignment linked to a tracker ticket with priority.",
            "steps": "1. Navigate to Projects tab\n2. Find an assignment with Info column data\n3. Verify Info shows tags like [Medium], [High]\n4. Verify the tags match the tracker ticket priority\n5. Verify Tracker column shows linked ticket ID (e.g., 'SF-1613')\n6. Click the tracker link — verify it opens external tracker in new tab",
            "expected": "Info column displays ticket priority tags. Tracker column shows clickable links to external tracker.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Table Structure, planner-project-settings-pages.md §Info column", "module": "planner/task-crud",
            "notes": "These Info values are what close-by-tag matches against."
        },
    ]


def get_dnd_cases():
    """TS-Planner-DnD: Drag and drop ordering and known bugs."""
    return [
        {
            "id": "TC-PLN-021", "title": "Drag task to reorder within project group",
            "preconditions": "Employee with 3+ assignments in one project. Editing mode active.",
            "steps": "1. Login as PM, navigate to Projects tab\n2. Open for editing on an employee with 3+ tasks\n3. Verify DnD handles ('::' buttons) visible on each row\n4. Drag task C above task A using the handle\n5. Verify visual order updates immediately\n6. Verify backend PATCH sent with new nextAssignmentId\nDB-CHECK: SELECT id, next_assignment, position FROM ttt_backend.task_assignment WHERE employee_id = {id} AND project_id = {pid} ORDER BY position",
            "expected": "DnD reorder works. Backend receives PATCH /v1/assignments/{id} with new nextAssignmentId. Linked-list pointers updated.",
            "priority": "Critical", "type": "UI",
            "req_ref": "planner-assignment-ordering.md §Data Model", "module": "planner/dnd",
            "notes": "Dual ordering system: next_assignment (linked list) + position (index)."
        },
        {
            "id": "TC-PLN-022", "title": "DnD handles only visible in editing mode",
            "preconditions": "PM on Projects tab.",
            "steps": "1. Navigate to Projects tab, select project\n2. Verify in read-only mode: NO DnD handles ('::' buttons) visible\n3. Click 'Open for editing' (global)\n4. Verify DnD handles appear on all task rows\n5. Close editing mode\n6. Verify handles disappear",
            "expected": "DnD handles are conditional on editing mode. Not visible in read-only state.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Editing Mode", "module": "planner/dnd",
            "notes": ""
        },
        {
            "id": "TC-PLN-023", "title": "DnD reorder persists after page reload",
            "preconditions": "Tasks reordered via DnD.",
            "steps": "1. Login as PM, open editing, reorder tasks (move C above A)\n2. Refresh the page (F5)\n3. Re-navigate to the same project and employee\n4. Verify the custom order is preserved\nDB-CHECK: Verify next_assignment chain reflects DnD order",
            "expected": "DnD order persists in DB via next_assignment linked-list pointers. Not just frontend state.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-assignment-ordering.md", "module": "planner/dnd",
            "notes": ""
        },
        {
            "id": "TC-PLN-024", "title": "Bug #3332 — DnD should not create duplicate task rows",
            "preconditions": "Employee with 3+ assignments. Editing mode.",
            "steps": "1. Login as PM, open editing\n2. Drag a task to a new position\n3. Carefully observe: does a duplicate row appear?\n4. Count rows before and after DnD — should be the same\n5. Repeat DnD on the same task 3 times\n6. Verify no ghost/duplicate entries",
            "expected": "No duplicate rows after DnD. Bug #3332 root cause: generateTaskAssignments.ts appends new ID without removing old from order array. Verify this is fixed.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3332, planner-dnd-bugs-analysis.md §Bug #3332", "module": "planner/dnd",
            "notes": "Root cause identified: splice commented out with TODO. Regression test."
        },
        {
            "id": "TC-PLN-025", "title": "Bug #3314 — task order preserved after 'Open for Editing'",
            "preconditions": "Employee with DnD-ordered tasks.",
            "steps": "1. Login as PM, open editing on employee A\n2. Reorder tasks for employee A via DnD\n3. Close editing for employee A\n4. Open editing for employee B (different employee)\n5. Go back to employee A — verify their task order is PRESERVED\n6. Click 'Open for editing' on employee A again\n7. Verify order is still the DnD custom order (not reset)",
            "expected": "DnD order not disrupted by toggling 'Open for editing'. Bug #3314: TasksPlannerTable.tsx re-sorts on readOnly state change. Verify fix.",
            "priority": "Critical", "type": "UI",
            "req_ref": "#3314, planner-dnd-bugs-analysis.md §Bug #3314", "module": "planner/dnd",
            "notes": "Root cause: useEffect with .sort() triggers on every rowsMap change, disrupting DnD order."
        },
        {
            "id": "TC-PLN-026", "title": "Task ordering rules — non-generated above generated",
            "preconditions": "Employee with mixed assignment types.",
            "steps": "1. Navigate to Tasks tab or Projects tab\n2. Verify non-generated assignments appear ABOVE generated ones\n3. Open for editing (generates assignments)\n4. Verify order preserved: existing tasks remain in their positions",
            "expected": "Ordering rule: non-generated (real) above generated. Order preserved when opening for editing.",
            "priority": "Medium", "type": "UI",
            "req_ref": "Confluence §8.1-8.2", "module": "planner/dnd",
            "notes": ""
        },
    ]


def get_lock_cases():
    """TS-Planner-Lock: Concurrent editing protection."""
    return [
        {
            "id": "TC-PLN-027", "title": "Lock acquired when editing starts",
            "preconditions": "Two users with access to the same project. Both logged in.\nNote: Requires two browser sessions.",
            "steps": "1. Login as User A, navigate to Projects tab, open editing\n2. Click on hours cell for Employee X's task → cell becomes editable\n3. In a second browser, login as User B, navigate to same project\n4. Verify User B sees the cell as locked (shows User A's name or lock indicator)\n5. Verify User B cannot edit that cell while locked",
            "expected": "Lock acquired on cell focus. REST POST /v1/assignments/lock publishes LOCK event via WebSocket. Other users see lock in real-time.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-lock-mechanism.md §Architecture", "module": "planner/lock",
            "notes": "Requires multi-user testing with two browser sessions. Backend uses Caffeine cache (in-memory, no DB)."
        },
        {
            "id": "TC-PLN-028", "title": "Auto-release lock after 1 minute TTL",
            "preconditions": "User A has locked a cell.",
            "steps": "1. User A clicks on a cell (acquires lock)\n2. User A does NOT interact further (leaves cell focused but idle)\n3. Wait 60+ seconds\n4. In User B's browser, verify the cell becomes editable (lock released)\n5. Verify UNLOCK WebSocket event received by User B",
            "expected": "Lock auto-releases after 1 minute TTL. Caffeine cache TTL expiry triggers release. WebSocket UNLOCK event published.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-lock-mechanism.md §Auto-release: 1 minute TTL", "module": "planner/lock",
            "notes": "In-memory Caffeine cache — locks lost on server restart."
        },
        {
            "id": "TC-PLN-029", "title": "Manual unlock on blur/navigation",
            "preconditions": "User has locked a cell by focusing on it.",
            "steps": "1. User A clicks on hours cell (lock acquired)\n2. User A clicks on a different cell or navigates away\n3. Verify the original cell's lock is released immediately\n4. Verify UNLOCK WebSocket event sent\n5. Verify User B can now edit the original cell",
            "expected": "Lock released on blur. REST DELETE /v1/assignments/unlock. Immediate release, not waiting for TTL.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-lock-mechanism.md", "module": "planner/lock",
            "notes": ""
        },
        {
            "id": "TC-PLN-030", "title": "Lock types — effort, comment, remainingEstimate",
            "preconditions": "Two users on same project.",
            "steps": "1. User A locks the effort cell → verify only effort locked, comment still editable by User B\n2. User A releases effort lock\n3. User A locks the comment cell → verify only comment locked\n4. User A releases comment lock\n5. User A locks remainingEstimate → verify only that field locked",
            "expected": "Locks are per-field, not per-row. FieldsToLockType: effort, comment, remainingEstimate, ALL_ROW.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-lock-mechanism.md, planner-redux-state-architecture.md §locks slice", "module": "planner/lock",
            "notes": ""
        },
        {
            "id": "TC-PLN-031", "title": "Cell selections visible to other users (focus tracking)",
            "preconditions": "Two users on same project.",
            "steps": "1. User A selects a cell (without editing)\n2. Verify User B sees User A's selection highlighted (colored indicator)\n3. User A moves to a different cell\n4. Verify User B sees selection move in real-time\n5. Verify selection colors distinguish between users",
            "expected": "Cell selections broadcast via WebSocket SELECT event. MappedSelection includes ownerLogin, color, group, date.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §focus slice, planner-websocket-stomp-system.md §SELECTIONS", "module": "planner/lock",
            "notes": "Focus tracking is separate from lock mechanism. Selections are visual-only, not blocking."
        },
    ]


def get_tracker_cases():
    """TS-Planner-Tracker: Tracker sync operations."""
    return [
        {
            "id": "TC-PLN-032", "title": "Refresh tickets — updates metadata only",
            "preconditions": "Project with tracker integration configured. PM logged in.",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, select a tracker-integrated project\n3. Open for editing\n4. Click 'Refresh tickets' button on an employee's row\n5. Verify loading state appears\n6. Verify task names/ticket info update (metadata refresh)\n7. Verify hours are NOT changed (refresh is metadata-only)",
            "expected": "Refresh only updates task names, priorities, ticket_info. Does NOT import or modify work hours. fetchProjectTasksRefresh saga.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-tracker-integration.md §Three Distinct Tracker Buttons", "module": "planner/tracker",
            "notes": "#3394 distinguishes: Refresh=metadata, Load=metadata+hours."
        },
        {
            "id": "TC-PLN-033", "title": "Upload hours to tracker (TTT → tracker)",
            "preconditions": "Project with tracker integration. Employee has reported hours. Tracker credentials configured.",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, open editing\n3. Click 'Actions' dropdown\n4. Click 'Upload to Tracker' (or equivalent)\n5. Verify loading state during sync\n6. Verify success notification: 'planner.tracker.success'\n7. Verify no assignment data changed in TTT (one-way upload)",
            "expected": "Hours uploaded to external tracker. Source: TTT. Uses handleSyncWorkLogInfoWithTracker saga. POST /tracker-work-log/sync with source='TTT'.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-tracker-integration.md §Upload to Tracker", "module": "planner/tracker",
            "notes": "Requires employee tracker credentials configured at /v1/employees/current/settings/trackers."
        },
        {
            "id": "TC-PLN-034", "title": "Download hours from tracker (tracker → TTT)",
            "preconditions": "Project with tracker integration. Tracker has worklogs.",
            "steps": "1. Login as PM\n2. Click 'Actions' dropdown > 'Download hours from the tracker'\n3. Verify loading state\n4. Verify hours imported from tracker appear in planner\n5. Verify success notification: 'planner.tracker.hours_uploaded'",
            "expected": "Hours downloaded from external tracker. Source: TRACKER. Uses handleSyncProjectWithTracker saga.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-tracker-integration.md §Download from Tracker", "module": "planner/tracker",
            "notes": ""
        },
        {
            "id": "TC-PLN-035", "title": "Tracker sync error — no tracker configured",
            "preconditions": "Project WITHOUT tracker integration or without credentials.",
            "steps": "1. Login as PM of a project without tracker configuration\n2. Click 'Refresh tickets' or 'Update tickets' on an employee\n3. Verify error banner: 'Failed to update tickets for the project X. Project manager has to set up synchronization with the tracker'\n4. Verify red error banner is displayed",
            "expected": "Error message displayed when tracker not configured. exception.tracker.not.supported or exception.no.tracker.url.found.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-tracker-integration.md §Error handling, planner-ui-exploration-session74.md §Banners", "module": "planner/tracker",
            "notes": "Banner text observed on qa-1 session 74."
        },
        {
            "id": "TC-PLN-036", "title": "WebSocket TRACKER_SYNC events",
            "preconditions": "Two users viewing same project. Tracker sync triggered.",
            "steps": "1. User A triggers tracker sync (Upload or Download)\n2. Verify User B receives TRACKER_SYNC_START event (loading indicator?)\n3. Wait for sync to complete\n4. Verify User B receives TRACKER_SYNC_FINISH event\n5. Verify User B's planner view updates after sync",
            "expected": "WebSocket events on /topic/projects/{projectId}/tracker-work-log. TRACKER_SYNC_START and TRACKER_SYNC_FINISH broadcast to all viewers.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md §WsTrackerSyncEventListener", "module": "planner/tracker",
            "notes": "Project-level topic — all project viewers see sync status."
        },
        {
            "id": "TC-PLN-037", "title": "'Actions' dropdown — 'Approve all'",
            "preconditions": "PM with editing mode active. Employees have reported hours.",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, open editing\n3. Click 'Actions' dropdown\n4. Click 'Approve all'\n5. Verify all reported hours for the project are approved\n6. Verify approval status changes in the table",
            "expected": "'Approve all' approves all employee reports for the project on the selected date.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Actions Dropdown", "module": "planner/tracker",
            "notes": ""
        },
    ]


def get_members_cases():
    """TS-Planner-Members: Project member management."""
    return [
        {
            "id": "TC-PLN-038", "title": "Open Project Settings modal",
            "preconditions": "PM of a project.",
            "steps": "1. Login as PM\n2. Navigate to Projects tab, select project\n3. Click the Project Settings icon (gear at far right of header)\n4. Verify 'Project settings' modal opens\n5. Verify two tabs: 'Project members' and 'Tasks closing'\n6. Verify 'Project members' tab is active by default\n7. Verify employee list is visible",
            "expected": "Modal opens with two tabs. Default: Project members. Dialog role='dialog' with heading 'Project settings'.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Project Settings Modal", "module": "planner/members",
            "notes": "Settings icon is hard to find — unnamed SVG inside .planner__project-group-add."
        },
        {
            "id": "TC-PLN-039", "title": "Add member to project",
            "preconditions": "PM of a project. At least one employee not already a member.\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.project_member WHERE project_id = {projectId}) ORDER BY random() LIMIT 1",
            "steps": "1. Open Project Settings > Project members tab\n2. Click the employee dropdown (combobox)\n3. Type an employee name\n4. Select the employee from dropdown\n5. Click the '+' add button\n6. Verify 'Changes saved' notification appears\n7. Verify new member appears in the members table\n8. Verify the added employee no longer appears in the dropdown",
            "expected": "Member added immediately. 'Changes saved' notification. Added employee filtered from dropdown.",
            "priority": "High", "type": "UI",
            "req_ref": "Confluence §7.3, planner-project-settings-pages.md", "module": "planner/members",
            "notes": ""
        },
        {
            "id": "TC-PLN-040", "title": "Remove member from project",
            "preconditions": "Project with at least 2 members. PM logged in.",
            "steps": "1. Open Project Settings > Project members\n2. Find a member in the table\n3. Click the delete (trash) icon on their row\n4. Verify 'Changes saved' notification\n5. Verify member removed from the table\n6. Verify the removed employee reappears in the dropdown",
            "expected": "Member removed immediately. Notification displayed. Employee available in dropdown again.",
            "priority": "High", "type": "UI",
            "req_ref": "Confluence §7.3.4", "module": "planner/members",
            "notes": ""
        },
        {
            "id": "TC-PLN-041", "title": "Add/edit project role for member",
            "preconditions": "PM with project member in the list.",
            "steps": "1. Open Project Settings > Project members\n2. Find a member without a role (or with an existing role)\n3. Click on the role field (placeholder: 'Click to add role' or existing role text)\n4. Type a role (e.g., 'Developer')\n5. Press Enter or click away to save\n6. Verify 'Changes saved' notification\n7. Verify role displayed next to member name",
            "expected": "Inline role editing. Click to edit, Enter to save. Role stored in project_member table.",
            "priority": "Medium", "type": "UI",
            "req_ref": "Confluence §7.3.2, Figma annotations", "module": "planner/members",
            "notes": "Placeholder: 'Add a role' per Confluence §7.3.2."
        },
        {
            "id": "TC-PLN-042", "title": "Member order determines planner display order",
            "preconditions": "Project with 3+ members. Members reordered via DnD in settings.",
            "steps": "1. Open Project Settings > Project members\n2. Reorder members via DnD (if supported) or note current order\n3. Click OK to close\n4. In the Projects tab, verify employee group rows appear in the same order as the members list\n5. Reorder in settings → verify planner order changes",
            "expected": "Employee ordering in planner matches Project Settings > Project Members order. Per Confluence §6.1, #3375.",
            "priority": "Medium", "type": "UI",
            "req_ref": "Confluence §6.1, #3375", "module": "planner/members",
            "notes": "#3375 was a regression where employee ordering broke."
        },
        {
            "id": "TC-PLN-043", "title": "Employee name links to CS profile",
            "preconditions": "Project Settings open with members.",
            "steps": "1. Open Project Settings > Project members\n2. Find a member with a linked name\n3. Verify name is a clickable link\n4. Click the link → verify it opens Company Staff profile (https://cs.noveogroup.com/profile/{id})\n5. Verify link opens in new tab",
            "expected": "Employee names link to Company Staff profiles. External link to CS system.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Project Members Tab", "module": "planner/members",
            "notes": ""
        },
    ]


def get_reports_cases():
    """TS-Planner-Reports: Report submission and confirmation through planner."""
    return [
        {
            "id": "TC-PLN-044", "title": "Report period indicator — open vs closed",
            "preconditions": "Employee with both open and closed report periods.",
            "steps": "1. Login as employee\n2. Navigate to Tasks tab\n3. Navigate to a date in the current open period → verify editing is enabled\n4. Navigate to a date in a past closed period → verify editing is DISABLED\n5. Verify read-only indicator on closed-period assignments",
            "expected": "Open period: editable cells. Closed period: read-only. EmployeeReportPeriod with status field controls editability.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §reports slice", "module": "planner/reports",
            "notes": "Report periods managed via GET /v1/report-periods. isOpenPeriod determines editability."
        },
        {
            "id": "TC-PLN-045", "title": "Cannot edit hours in closed report period",
            "preconditions": "Date in a closed report period for the employee.",
            "steps": "1. Navigate to a date in the employee's closed report period\n2. Attempt to click on hours cell\n3. Verify cell is NOT editable (no input appears)\n4. Verify visual indication that period is closed (greyed out or read-only styling)",
            "expected": "Hours cells are read-only in closed periods. 'Open for editing' button not available for closed periods.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md", "module": "planner/reports",
            "notes": "isOpenPeriod === true required for editing."
        },
        {
            "id": "TC-PLN-046", "title": "Total hours calculation per employee",
            "preconditions": "Employee with multiple assignments with hours.",
            "steps": "1. Navigate to Tasks tab\n2. Verify Total row at bottom sums all project hours for the selected date\n3. Edit hours in one task (e.g., change from 2 to 5)\n4. Verify Total row updates (increases by 3)\n5. Navigate to Projects tab, select a project\n6. Verify per-employee totals in the employee header rows",
            "expected": "Total row shows sum of all hours. Updates dynamically on edit. Both Tasks and Projects tabs show totals.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Total row", "module": "planner/reports",
            "notes": "selectTotalEffortsByProject selector in plannerTasks slice."
        },
        {
            "id": "TC-PLN-047", "title": "Assignment history — paginated view",
            "preconditions": "Assignment with history of changes.",
            "steps": "1. Navigate to planner and find an assignment\n2. Open assignment history (if UI control exists)\n3. Verify history shows: date, effort, comment, remainingEstimate changes\n4. Verify pagination controls if many history entries\n5. Verify sort options: date, effort, comment, remainingEstimate",
            "expected": "Assignment history shows change log. Paginated. Sortable. TaskAssignmentHistoryItemDTO content.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-redux-state-architecture.md §assignments slice", "module": "planner/reports",
            "notes": "sortDirection + sortField available. totalHours shown."
        },
    ]


def get_validation_cases():
    """TS-Planner-Validation: Hours validation and data integrity."""
    return [
        {
            "id": "TC-PLN-048", "title": "#2914 — 36h daily limit warning displayed",
            "preconditions": "Employee who has reported >36h across all projects on a day.\nSETUP: Via API — create task reports totaling >36h on one date.",
            "steps": "SETUP: Via API — create multiple task_reports for the employee on the same date, totaling >36h\n1. Login as the employee\n2. Navigate to planner / employee dashboard\n3. Verify warning indicator appears for the date\n4. Check employee warnings: GET /employee/current/warnings\n5. Verify warning type: DATE_EFFORT_OVER_LIMIT",
            "expected": "36h daily limit triggers warning. TaskReportWarningCommand.findDaysWithExceededLimit() detects it. Warning is DISPLAY ONLY — no hard block.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "#2914, planner-2914-validation-bypass.md", "module": "planner/validation",
            "notes": "ttt.daily-report-limit: 36h (2160 minutes). Warning only, not a hard block."
        },
        {
            "id": "TC-PLN-049", "title": "#2914 — no hard block on exceeding 36h",
            "preconditions": "Employee with 35h already reported on a date.",
            "steps": "SETUP: Via API — create task_reports totaling 35h on a date\n1. Login as the employee\n2. Navigate to planner Tasks tab\n3. Edit hours on an additional task: enter '5' (bringing total to 40h)\n4. Verify the edit is ACCEPTED (no blocking, no error)\n5. Verify total shows 40h\n6. Verify warning appears but does not prevent saving",
            "expected": "No hard block. >36h is allowed. Warning displayed but save succeeds. No frontend or backend validation blocks submission.",
            "priority": "High", "type": "UI",
            "req_ref": "#2914, planner-2914-validation-bypass.md §No Hard Validation", "module": "planner/validation",
            "notes": "Bypass: split hours across projects. Per-project check: 20h < 36h each, but total 40h."
        },
        {
            "id": "TC-PLN-050", "title": "Per-project warning vs global warning",
            "preconditions": "Employee reporting hours across 2 projects.",
            "steps": "SETUP: Via API — report 25h in Project A and 20h in Project B (same date)\n1. Via API — GET /project/{projectA}/warnings → verify NO per-project warning (25h < 36h)\n2. Via API — GET /project/{projectB}/warnings → verify NO per-project warning (20h < 36h)\n3. Via API — GET /employee/current/warnings → verify GLOBAL warning (45h > 36h)\n4. Verify UI shows global warning but no per-project indicator",
            "expected": "Per-project validation misses cross-project totals. Only global employee-wide check catches it. Design issue documented in #2914.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "planner-2914-validation-bypass.md §Two Parallel Systems", "module": "planner/validation",
            "notes": "Root cause: per-project check uses task_report WHERE project = :id, not ALL projects."
        },
        {
            "id": "TC-PLN-051", "title": "Read-only assignments cannot be edited",
            "preconditions": "Employee with readOnly assignments (not opened for editing).",
            "steps": "1. Login as PM, navigate to Projects tab\n2. Find an employee NOT opened for editing\n3. Verify their assignment cells are not editable\n4. Verify action buttons in comment column are disabled/greyed\n5. Verify no DnD handles visible",
            "expected": "Read-only assignments: cells not editable, buttons disabled, no DnD handles. planner__cel--color-read-only CSS class.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Read-Only Mode", "module": "planner/validation",
            "notes": ""
        },
    ]


def get_realtime_cases():
    """TS-Planner-RealTime: WebSocket multi-user real-time scenarios."""
    return [
        {
            "id": "TC-PLN-052", "title": "Real-time hours update via WebSocket (PATCH event)",
            "preconditions": "Two users viewing the same project.",
            "steps": "1. User A and User B both navigate to same project in Projects tab\n2. User A edits hours for Employee X's task (changes 2 → 5)\n3. User A saves\n4. Verify User B's view updates in real-time (shows 5 without refresh)\n5. Verify Total row for User B also updates",
            "expected": "WebSocket PATCH event on /topic/employees/{login}/assignments. Both plannerTasks and plannerProjects slices updated via manager saga routing.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md §Frontend Event Routing", "module": "planner/realtime",
            "notes": "Requires two browser sessions. Events dispatched to BOTH view slices."
        },
        {
            "id": "TC-PLN-053", "title": "Real-time task addition via WebSocket (ADD event)",
            "preconditions": "Two users on same project.",
            "steps": "1. User A adds a new task via search bar\n2. Verify User B sees the new task appear in real-time\n3. Verify User B's view includes the new task in the correct position (top of group)\n4. Verify green highlight visible to User B",
            "expected": "WebSocket ADD event broadcasts new assignment. Both users see it in real-time.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md §WsTaskAssignmentEventListener ADD", "module": "planner/realtime",
            "notes": ""
        },
        {
            "id": "TC-PLN-054", "title": "Real-time task deletion via WebSocket (DELETE event)",
            "preconditions": "Two users on same project.",
            "steps": "1. User A deletes an assignment\n2. Verify User B's view removes the row in real-time\n3. Verify User B's total recalculates",
            "expected": "WebSocket DELETE event removes assignment from both users' views. Total recalculates.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md ��DELETE event", "module": "planner/realtime",
            "notes": ""
        },
        {
            "id": "TC-PLN-055", "title": "Task rename via WebSocket (TASK_RENAME event)",
            "preconditions": "Two users viewing same project. Task rename triggered.",
            "steps": "1. User A opens task rename modal\n2. User A changes task name from 'Old Name' to 'New Name'\n3. User A saves\n4. Verify User B sees 'New Name' in their planner view in real-time",
            "expected": "WebSocket TASK_RENAME event updates task name for all viewers. No page refresh needed.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md §TASK_RENAME, planner-redux-state-architecture.md §tasks slice", "module": "planner/realtime",
            "notes": "Task rename modal: modalTaskRename, taskRenameData in tasks slice."
        },
        {
            "id": "TC-PLN-056", "title": "Generate event broadcasts to all users",
            "preconditions": "Two users on same project. PM opens editing.",
            "steps": "1. User A (PM) clicks 'Open for editing' on an employee\n2. Verify User B receives GENERATE WebSocket event\n3. Verify User B sees the newly generated assignments appear\n4. Verify debouncing prevents excessive re-renders",
            "expected": "GENERATE event from /v1/assignments/generate broadcasts via WebSocket. Both users see generated assignments. Frontend debounces GENERATE events.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-websocket-stomp-system.md §WsTaskAssignmentEventListener GENERATE", "module": "planner/realtime",
            "notes": ""
        },
    ]


def get_copy_table_cases():
    return [
        {
            "id": "TC-PLN-057", "title": "Copy table — happy path copies visible data to clipboard",
            "preconditions": "PM logged in. Project with ≥2 employees and reported hours in current week.\nQuery: SELECT DISTINCT pa.project_id FROM planner_assignment pa JOIN project_employee pe ON pa.project_id = pe.project_id WHERE pe.role IN ('PM','LEAD') GROUP BY pa.project_id HAVING COUNT(DISTINCT pa.employee_id) >= 2 LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project with reported hours\n4. Click 'Open for editing' on an employee to enter editing mode\n5. Click 'Copy table' button in the Actions dropdown\n6. Open a text editor and paste clipboard content (Ctrl+V)\n7. Verify pasted data contains employee names, task names, and hours for visible dates",
            "expected": "Clipboard contains tab-separated table data matching the planner grid. Includes employee names, task rows, daily hours, and totals.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #3386, planner-ticket-findings.md §Copy table", "module": "planner/copy-table",
            "notes": "Copy table button only visible in editing mode. Uses navigator.clipboard.writeText API."
        },
        {
            "id": "TC-PLN-058", "title": "Copy table — #3386 regression: deleted/closed tasks excluded",
            "preconditions": "PM logged in. Project where at least one task has been deleted or project has closed tasks.\nQuery: SELECT DISTINCT pa.project_id FROM planner_assignment pa JOIN task t ON pa.task_id = t.id WHERE t.deleted = true OR t.project_id IN (SELECT id FROM project WHERE status = 'CLOSED') LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project known to have deleted/closed tasks\n4. Enter editing mode\n5. Click 'Copy table' from Actions dropdown\n6. Paste clipboard into text editor\n7. Verify no deleted or closed tasks appear in pasted data",
            "expected": "Copied table excludes deleted tasks and tasks from closed projects. Only active, visible tasks are in clipboard data. Bug #3386 required filtering these out.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #3386 — deleted/closed tasks in copy", "module": "planner/copy-table",
            "notes": "Regression test for #3386. Bug was that deleted/closed tasks were included in copy output."
        },
        {
            "id": "TC-PLN-059", "title": "Copy table — button only available in editing mode",
            "preconditions": "PM logged in. Project with employees.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select a project\n4. Verify 'Actions' dropdown does NOT contain 'Copy table' in view-only mode\n5. Click 'Open for editing' on an employee\n6. Open Actions dropdown\n7. Verify 'Copy table' option is now visible",
            "expected": "Copy table is unavailable in read-only mode. It appears only after entering editing mode via 'Open for editing'.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Actions dropdown", "module": "planner/copy-table",
            "notes": "Actions dropdown contents change based on editing mode state."
        },
        {
            "id": "TC-PLN-060", "title": "Copy table — empty project produces empty or no-data clipboard",
            "preconditions": "PM logged in. Project with no assignments for current week.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select a project with no employees or no assignments in current date range\n4. Enter editing mode\n5. Click 'Copy table' from Actions dropdown\n6. Paste clipboard into text editor\n7. Verify output is empty or contains only headers",
            "expected": "Copy table handles empty state gracefully — clipboard contains header row only or is empty. No error occurs.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-ticket-findings.md §Copy table edge cases", "module": "planner/copy-table",
            "notes": ""
        },
        {
            "id": "TC-PLN-061", "title": "Copy table — includes all employees and weekly totals",
            "preconditions": "PM logged in. Project with 3+ employees, each with hours in current week.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project with 3+ employees having hours\n4. Enter editing mode\n5. Click 'Copy table'\n6. Paste into spreadsheet application\n7. Count employee rows vs visible employees in planner\n8. Verify totals column matches planner weekly total",
            "expected": "All visible employees are included. Weekly/daily totals match those displayed in the planner grid. Data is tab-separated and imports cleanly into spreadsheets.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #3386", "module": "planner/copy-table",
            "notes": "Verify row count matches planner. Totals row should be at bottom."
        },
    ]


def get_bug_regression_cases():
    return [
        {
            "id": "TC-PLN-062", "title": "#3294 — Delete task on current date reflects immediately",
            "preconditions": "PM logged in. Project with at least one task that has assignments today.\nQuery: SELECT pa.project_id, pa.task_id FROM planner_assignment pa WHERE pa.date = CURRENT_DATE AND pa.hours > 0 LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project and enter editing mode\n4. Note the task name and current total hours\n5. Right-click or use context menu on the task to delete\n6. Confirm deletion in dialog\n7. Verify task row disappears from the planner grid immediately\n8. Verify total hours recalculates without the deleted task's hours",
            "expected": "Task deletion on current date is reflected immediately in the UI. No manual refresh required. Total hours update to exclude deleted task.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #3294 — task deletion not refreshing", "module": "planner/bug-regression",
            "notes": "Bug #3294: after deleting a task, the planner did not refresh and stale data remained visible."
        },
        {
            "id": "TC-PLN-063", "title": "#3296 — Approved status preserved after tracker import",
            "preconditions": "PM logged in. Project with tracker configured. Employee with APPROVED status for current period.\nSetup: Ensure at least one employee has approved hours for today.\nQuery: SELECT pa.employee_id, pa.project_id FROM planner_assignment pa WHERE pa.approved = true AND pa.date >= CURRENT_DATE - INTERVAL '7 days' LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project with tracker\n3. Note which employees have APPROVED status (green checkmark or indicator)\n4. Open Actions dropdown → click 'Download hours from tracker'\n5. Wait for sync to complete\n6. Verify previously APPROVED employees still show APPROVED status\n7. DB-CHECK: SELECT approved FROM planner_assignment WHERE employee_id = <id> AND date = <date>",
            "expected": "Tracker import does NOT clear the APPROVED status on employees whose hours were already approved. Only unapproved hours are affected by import.",
            "priority": "Critical", "type": "UI",
            "req_ref": "GitLab #3296 — approved status cleared after import", "module": "planner/bug-regression",
            "notes": "Bug #3296: tracker sync was resetting approved=false on all assignments regardless of prior approval state."
        },
        {
            "id": "TC-PLN-064", "title": "#3375 — Employee order matches Project Members DnD order",
            "preconditions": "PM logged in. Project with 3+ employees.\nQuery: SELECT pe.project_id FROM project_employee pe GROUP BY pe.project_id HAVING COUNT(*) >= 3 ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project\n3. Note the order of employees shown in the planner grid\n4. Open Project Settings → Project Members tab\n5. Drag an employee from position 3 to position 1\n6. Click OK to save\n7. Return to the planner grid for the same project\n8. Verify employee order matches the new DnD order from Project Members",
            "expected": "Planner employee display order matches the order configured in Project Settings → Project Members. Reordering members updates the planner grid order.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #3375 — member order not reflected", "module": "planner/bug-regression",
            "notes": "Project Settings icon: .planner__project-group-add .uikit-button. Members tab allows DnD reordering."
        },
        {
            "id": "TC-PLN-065", "title": "#3258 — Project sorting alphabetical + green highlight on new task",
            "preconditions": "PM logged in. Employee assigned to 2+ projects, one having tasks added recently.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Verify projects are listed in alphabetical order in the left sidebar\n4. Add a new task to a project via the search/add task input\n5. Verify the newly added task row appears with a green highlight/animation\n6. Verify project list order remains alphabetical after task addition",
            "expected": "Projects sorted alphabetically. Newly added tasks show a transient green highlight. Alphabetical order is maintained after adding tasks.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #3258 — UI refactoring, alphabetical sort", "module": "planner/bug-regression",
            "notes": "#3258 was a major UI refactoring commit. Green highlight was part of the new UX."
        },
        {
            "id": "TC-PLN-066", "title": "#3308 — DnD order persists across date navigation",
            "preconditions": "PM logged in. Project with 3+ tasks, editing mode enabled.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project\n3. Enter editing mode\n4. Reorder tasks via DnD (move task 3 to position 1)\n5. Note the new task order\n6. Navigate to next day using date navigation arrow\n7. Verify task order on the new day matches the reordered sequence\n8. Navigate back to original day\n9. Verify order is still the reordered sequence",
            "expected": "DnD task order persists when navigating between dates. The reorder is saved to the backend (linked-list position system) and displayed consistently across dates.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #3308 — DnD not persisted, planner-assignment-ordering.md", "module": "planner/bug-regression",
            "notes": "Bug #3308: order was only stored in frontend state and lost on date change. Fix uses PATCH /v1/assignments/reorder."
        },
        {
            "id": "TC-PLN-067", "title": "#2832 — Changes visible via WebSocket without manual refresh",
            "preconditions": "Two browser sessions logged into the same project. PM editing mode enabled in session A.",
            "steps": "1. Login as PM in browser session A, enter editing mode on project\n2. Login as same or different PM in browser session B, viewing same project\n3. In session A: edit hours for an employee on a specific date (e.g., change 4h to 6h)\n4. In session B: verify the hours update appears within 2-3 seconds WITHOUT manual refresh\n5. In session A: add a new task to an employee\n6. In session B: verify the new task row appears in real-time",
            "expected": "WebSocket broadcasts PATCH and ADD events. Session B sees changes from session A in real-time without page refresh.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #2832, planner-websocket-stomp-system.md", "module": "planner/bug-regression",
            "notes": "Requires two parallel browser sessions. WebSocket topic: /topic/assignments/{projectId}."
        },
        {
            "id": "TC-PLN-068", "title": "#3406 — Refresh button error handling",
            "preconditions": "PM logged in. Project with tracker configured but in an error state or with connection issues.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project with tracker\n4. Click 'Refresh tickets' from Actions dropdown\n5. If tracker is unreachable: verify error banner appears with descriptive message\n6. Verify the error banner text includes the tracker name and error type\n7. Verify the planner remains functional (not stuck in loading state)\n8. Verify clicking Refresh again retries the operation",
            "expected": "Error banner shows when tracker refresh fails. Message is informative (e.g., 'Failed to load from tracker: connection timeout'). Planner remains usable. Retry is possible.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #3406 — refresh error, planner-tracker-integration.md §Error Handling", "module": "planner/bug-regression",
            "notes": "Bug #3406: error notification was not shown on refresh failure. Banner selector: [class*='notification'] or getByText with error message."
        },
        {
            "id": "TC-PLN-069", "title": "#2447 — Open for editing preserves existing assignments",
            "preconditions": "PM logged in. Project where an employee already has manually entered hours.\nQuery: SELECT pa.project_id, pa.employee_id FROM planner_assignment pa WHERE pa.hours > 0 AND pa.date >= CURRENT_DATE - INTERVAL '7 days' GROUP BY pa.project_id, pa.employee_id LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project, note existing hours for an employee\n4. Click 'Open for editing' on that employee\n5. Verify generate API creates only new task slots — does NOT overwrite existing hours\n6. Verify the existing manual hours remain unchanged\n7. DB-CHECK: SELECT hours FROM planner_assignment WHERE employee_id = <id> AND project_id = <pid> AND date = <date>",
            "expected": "'Open for editing' generates empty assignment slots for tasks without entries. Existing hours/assignments are NOT modified or zeroed out.",
            "priority": "High", "type": "UI",
            "req_ref": "GitLab #2447, planner-close-by-tag-implementation.md §generate endpoint", "module": "planner/bug-regression",
            "notes": "The generate endpoint (POST /v1/assignments/generate) should not overwrite existing assignments."
        },
    ]


def get_tracker_adv_cases():
    return [
        {
            "id": "TC-PLN-070", "title": "#2302 — Tracker sync does not zero out TTT-only tasks",
            "preconditions": "PM logged in. Project with tracker configured. Employee has hours on a TTT-only task (task not in tracker).\nQuery: SELECT pa.project_id, pa.employee_id, t.name FROM planner_assignment pa JOIN task t ON pa.task_id = t.id WHERE pa.hours > 0 AND t.tracker_task_id IS NULL AND pa.date >= CURRENT_DATE - INTERVAL '7 days' LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project with tracker\n3. Note employee with hours on a TTT-only task (no tracker link)\n4. Note the exact hours value\n5. Open Actions → 'Download hours from tracker'\n6. Wait for sync to complete\n7. Verify TTT-only task hours remain unchanged\n8. DB-CHECK: SELECT hours FROM planner_assignment WHERE task_id = <ttt_only_task_id> AND employee_id = <id>",
            "expected": "Tracker sync only updates hours for tracker-linked tasks. TTT-only tasks (tracker_task_id IS NULL) retain their hours.",
            "priority": "Critical", "type": "UI",
            "req_ref": "GitLab #2302 — hours disappear after sync", "module": "planner/tracker-advanced",
            "notes": "Bug #2302: after tracker sync, hours on TTT-only tasks were being zeroed. This is a data loss scenario."
        },
        {
            "id": "TC-PLN-071", "title": "#2338 — Tracker timeout shows partial success state",
            "preconditions": "PM logged in. Project with tracker. Network conditions that may cause timeout (or large project).",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select large project with many tracker tasks\n3. Open Actions → 'Download hours from tracker'\n4. If operation times out: verify error notification appears\n5. Verify any tasks that DID sync before timeout show updated hours\n6. Verify tasks that did NOT sync retain their previous hours\n7. Verify no data corruption (hours are either old or correctly updated, never partially written)",
            "expected": "Timeout during tracker sync shows error notification. Partially synced tasks show correct hours. Non-synced tasks retain original hours. No partial/corrupt data.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #2338 — intermittent refresh timeout", "module": "planner/tracker-advanced",
            "notes": "Difficult to reproduce reliably. May need a large project or simulated slow network."
        },
        {
            "id": "TC-PLN-072", "title": "Tracker credential types — Jira token vs username/password",
            "preconditions": "Two projects: one with Jira Token auth, one with Username/Password auth.\nQuery: SELECT p.id, p.name, t.type, t.auth_type FROM project p JOIN tracker t ON p.tracker_id = t.id WHERE t.auth_type IN ('TOKEN', 'PASSWORD') LIMIT 2",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project with Jira Token auth tracker\n4. Open Actions → 'Load from tracker'\n5. Verify sync succeeds with token-based auth\n6. Switch to project with Username/Password auth tracker\n7. Open Actions → 'Load from tracker'\n8. Verify sync succeeds with password-based auth\n9. If either fails: verify error message mentions auth type",
            "expected": "Both credential types (Jira PAT token and username/password) work for tracker sync. Error messages on auth failure are specific to the credential type.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #2511 — Jira PAT, planner-tracker-integration.md §Credential Types", "module": "planner/tracker-advanced",
            "notes": "Tracker types in TTT: Jira, GitLab, Redmine, YouTrack, Toggl, Clockify, ClickUp. Auth types: TOKEN, PASSWORD."
        },
        {
            "id": "TC-PLN-073", "title": "Load from tracker vs Refresh tickets — distinct behaviors",
            "preconditions": "PM logged in. Project with tracker configured and existing tracked tasks.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project with tracker\n3. Open Actions dropdown and note available tracker options\n4. Click 'Refresh tickets' — verify it updates the task list (names, new tickets) but NOT hours\n5. Note that hours values remain unchanged after 'Refresh tickets'\n6. Click 'Download hours from tracker' — verify it updates hours values for the current date range\n7. Compare: Refresh = task metadata, Download = hours data",
            "expected": "'Refresh tickets' syncs task names/IDs from tracker without changing hours. 'Download hours from tracker' syncs actual reported hours. These are distinct operations with different outcomes.",
            "priority": "High", "type": "UI",
            "req_ref": "planner-tracker-integration.md §Sync Types", "module": "planner/tracker-advanced",
            "notes": "Users confuse these two operations. Refresh = GET tasks from tracker. Download = GET hours/worklogs."
        },
        {
            "id": "TC-PLN-074", "title": "Download hours from Actions dropdown UI flow",
            "preconditions": "PM logged in. Project with tracker and at least one employee with tracker hours.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project with tracker\n3. Enter editing mode\n4. Open 'Actions' dropdown\n5. Click 'Download hours from tracker'\n6. Verify loading indicator appears during sync\n7. Verify hours cells update with tracker data after sync completes\n8. Verify success notification or banner appears",
            "expected": "Download hours operation shows loading state, then updates hours cells with tracker data. Success feedback is shown.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-ui-exploration-session74.md §Actions dropdown", "module": "planner/tracker-advanced",
            "notes": "Actions dropdown items: 'Copy table', 'Download hours from tracker', 'Approve all', 'Refresh tickets'."
        },
        {
            "id": "TC-PLN-075", "title": "#3296 — Approve all not reset by subsequent tracker import",
            "preconditions": "PM logged in. Project with tracker. Employee hours present.\nSETUP: Ensure employees have reported hours for current week.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab, select project\n3. Enter editing mode\n4. Open Actions → 'Approve all' — approve all employees' hours\n5. Verify all employees show approved status\n6. Open Actions → 'Download hours from tracker'\n7. Wait for sync to complete\n8. Verify previously approved employees STILL show approved status\n9. DB-CHECK: SELECT approved FROM planner_assignment WHERE project_id = <pid> AND date >= CURRENT_DATE - INTERVAL '7 days'",
            "expected": "Bulk 'Approve all' status is NOT cleared by subsequent tracker import. Approved flag persists through sync.",
            "priority": "Critical", "type": "UI",
            "req_ref": "GitLab #3296 — approved status cleared", "module": "planner/tracker-advanced",
            "notes": "Related to TC-PLN-063 but tests the 'Approve all' + import sequence specifically."
        },
        {
            "id": "TC-PLN-076", "title": "Tracker not configured — per-employee error state",
            "preconditions": "PM logged in. Employee without personal tracker credentials on a project that requires them.",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Select project with tracker requiring per-employee credentials\n4. Click 'Open for editing' on an employee who has NOT configured tracker\n5. Attempt to load hours from tracker for that employee\n6. Verify error banner or indicator specific to that employee\n7. Verify other employees with configured trackers can still sync normally",
            "expected": "Error is scoped to the specific employee without tracker credentials. Other employees' tracker operations are unaffected. Error message indicates credential issue.",
            "priority": "Medium", "type": "UI",
            "req_ref": "GitLab #2461 — GitLab load failure, planner-tracker-integration.md §Per-Employee", "module": "planner/tracker-advanced",
            "notes": "Each employee may need their own tracker PAT. Missing credentials should show per-employee error, not block the whole project."
        },
    ]


def get_settings_cases():
    return [
        {
            "id": "TC-PLN-077", "title": "Project Settings — 'Changes saved' notification on member add",
            "preconditions": "PM logged in. Project with available employees not yet added.\nQuery: SELECT e.login FROM employee e WHERE e.enabled = true AND e.id NOT IN (SELECT employee_id FROM project_employee WHERE project_id = <test_project_id>) ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM\n2. Navigate to Planner → Tasks tab\n3. Click the Project Settings icon (gear icon near project name)\n4. In the 'Project settings' dialog, go to 'Project members' tab\n5. Click the member dropdown/combobox\n6. Select an employee not yet in the project\n7. Click 'Add' or press Enter\n8. Verify 'Changes saved' notification appears\n9. Verify the new member appears in the members list",
            "expected": "'Changes saved' notification appears after adding a member. New member visible in the list immediately.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Project Members tab", "module": "planner/settings",
            "notes": "Project Settings dialog: getByRole('dialog', { name: 'Project settings' }). Members tab: getByRole('tab', { name: 'Project members' })."
        },
        {
            "id": "TC-PLN-078", "title": "Project Settings — member dropdown filters already-added members",
            "preconditions": "PM logged in. Project with 2+ existing members and at least one available employee.",
            "steps": "1. Login as PM\n2. Open Project Settings → Project Members tab\n3. Note existing members in the list\n4. Click the member dropdown/combobox\n5. Type the name of an already-added member\n6. Verify the dropdown does NOT show the already-added member\n7. Type the name of an available (not yet added) employee\n8. Verify the dropdown DOES show that employee",
            "expected": "Member dropdown filters out employees already in the project. Only available employees appear in suggestions.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Member combobox", "module": "planner/settings",
            "notes": "Prevents duplicate member additions. Combobox uses typeahead filtering."
        },
        {
            "id": "TC-PLN-079", "title": "Project Settings — inline role editing with Enter to save",
            "preconditions": "PM logged in. Project with at least one member whose role can be changed.",
            "steps": "1. Login as PM\n2. Open Project Settings → Project Members tab\n3. Find a member in the list\n4. Click on the member's role field (e.g., 'Developer')\n5. Verify the role field becomes editable (inline edit mode)\n6. Change the role to a different value (e.g., 'QA')\n7. Press Enter to save\n8. Verify 'Changes saved' notification appears\n9. Verify the role value updates in the member list",
            "expected": "Role field supports inline editing. Enter key saves the change. Success notification confirms the save.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Inline role editing", "module": "planner/settings",
            "notes": "Role editing is done inline in the members table, not via a separate dialog."
        },
        {
            "id": "TC-PLN-080", "title": "Project Settings — delete member with confirmation",
            "preconditions": "PM logged in. Project with a member that can be safely removed (no critical assignments).\nSETUP: Add a test member to the project if needed via API.",
            "steps": "1. Login as PM\n2. Open Project Settings → Project Members tab\n3. Find the test member in the list\n4. Click the trash/delete icon next to the member\n5. Confirm deletion in the confirmation dialog (if any)\n6. Verify 'Changes saved' notification appears\n7. Verify the member is removed from the list\n8. Close Project Settings\n9. Verify the member no longer appears in the planner grid\nCLEANUP: Re-add the member if needed for other tests.",
            "expected": "Member deletion via trash icon removes the member from the project. Success notification shown. Member disappears from both the settings dialog and the planner grid.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Delete member", "module": "planner/settings",
            "notes": "Trash icon next to each member row. May or may not have a confirmation dialog."
        },
        {
            "id": "TC-PLN-081", "title": "Project Settings — tag CRUD with notifications",
            "preconditions": "PM logged in. Project with close-by-tag feature enabled.",
            "steps": "1. Login as PM\n2. Open Project Settings → 'Tasks closing' tab\n3. Type a new tag name in the 'Add a tag' input\n4. Press Enter or click Add\n5. Verify 'Changes saved' notification appears\n6. Verify the new tag appears in the tag list\n7. Click delete icon on the tag\n8. Verify 'Changes saved' notification appears\n9. Verify the tag is removed from the list",
            "expected": "Tag creation and deletion both show 'Changes saved' notification. Tags are added/removed from the visible list immediately.",
            "priority": "Medium", "type": "UI",
            "req_ref": "planner-project-settings-pages.md §Tags tab, planner-close-by-tag-implementation.md §Frontend", "module": "planner/settings",
            "notes": "Tags tab: getByRole('tab', { name: 'Tasks closing' }). Tag input: getByPlaceholder('Add a tag'). Close-by-tag is covered in t2724.xlsx but the Project Settings UI for tag CRUD is planner scope."
        },
        {
            "id": "TC-PLN-082", "title": "Project Settings — Tags tab informational text",
            "preconditions": "PM logged in. Project with close-by-tag feature.",
            "steps": "1. Login as PM\n2. Open Project Settings → 'Tasks closing' tab\n3. Read the informational text at the top of the tab\n4. Verify it explains that tags are used to close tasks matching the tag pattern\n5. Verify the text matches the Confluence specification for close-by-tag feature",
            "expected": "Informational text on Tags tab clearly explains the close-by-tag functionality. Text matches the specification from Confluence.",
            "priority": "Low", "type": "UI",
            "req_ref": "planner-close-by-tag-implementation.md §Confluence spec", "module": "planner/settings",
            "notes": "Verifies that the UI provides adequate context for the close-by-tag feature."
        },
    ]


# ─── Plan, Feature Matrix, Risk Assessment ────────────────────────────────────

PLAN_OVERVIEW = {
    "title": "Planner Module — Test Plan",
    "scope": "Comprehensive testing of the Planner module covering navigation, task CRUD, drag-and-drop ordering, concurrent editing locks, tracker integration (basic + advanced), project member management, project settings, report submission, copy-table feature, validation rules, real-time WebSocket updates, and bug regression tests. Close-by-tag feature is tested separately in t2724.xlsx.",
    "objectives": [
        "Verify planner navigation: tab switching, date navigation, project/role selection, search",
        "Test task assignment CRUD: add, edit hours/comment/remaining, delete",
        "Validate DnD ordering: reorder, persistence, regression for #3332 and #3314 bugs",
        "Test concurrent editing locks: acquisition, auto-release, per-field locking",
        "Verify tracker integration: refresh, upload, download, error handling, credentials",
        "Test advanced tracker scenarios: data loss prevention (#2302), timeout handling, auth types, approve persistence",
        "Test project member management: add, remove, role editing, order",
        "Verify project settings: members tab, tags tab, inline editing, notifications",
        "Verify report periods: open/closed editability, totals, history",
        "Test copy-table feature: happy path, #3386 regression, editing mode guard, edge cases",
        "Validate bug regressions: #3294, #3296, #3375, #3258, #3308, #2832, #3406, #2447",
        "Test validation: 36h daily limit warning (#2914), read-only enforcement",
        "Validate WebSocket real-time updates: PATCH, ADD, DELETE, GENERATE, RENAME events",
    ],
    "environments": [
        "Primary: qa-1 (ttt-qa-1.noveogroup.com)",
        "Secondary: timemachine (ttt-timemachine.noveogroup.com)",
        "Multi-user tests require 2 browser sessions",
    ],
    "approach": "UI-first testing with API setup/cleanup. Tests describe browser actions from PM and employee perspectives. Multi-user tests for lock/WebSocket features require parallel browser sessions. API used for data setup and DB verification.",
    "dependencies": [
        "Project with tracker integration for tracker tests",
        "PM credentials with project access",
        "Multiple employees assigned to test project",
        "Two browser sessions for concurrent editing tests",
        "Open report period for editing tests",
    ],
}

FEATURE_MATRIX = [
    # (Feature, Nav, CRUD, DnD, Lock, Tracker, Members, Reports, Valid, RT, Copy, BugReg, TrkAdv, Settings, Total)
    ("Tab navigation", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Date navigation", 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Project selection", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Search / add task", 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Empty state / banners", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("UI preferences (collapse/toggle)", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Edit hours", 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Edit comment / remaining", 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Delete assignment", 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 2),
    ("Open for editing / generate", 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 2),
    ("Color coding / Info column", 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 3),
    ("DnD reorder", 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 4),
    ("DnD bug #3332 (duplicate)", 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("DnD bug #3314 (order reset)", 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Ordering rules", 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Lock acquisition / release", 0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, 3),
    ("Auto-release TTL", 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Cell selection (focus)", 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Refresh tickets", 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1, 0, 0, 2),
    ("Upload / download tracker", 0, 0, 0, 0, 2, 0, 0, 0, 0, 0, 0, 1, 0, 3),
    ("Tracker error handling", 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 0, 2),
    ("Tracker sync events", 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Approve all", 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 0, 2),
    ("Project Settings modal", 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 2, 3),
    ("Add / remove member", 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 0, 0, 1, 3),
    ("Role editing", 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 2),
    ("Member order / CS link", 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 1, 0, 0, 3),
    ("Report period open/closed", 0, 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 0, 0, 2),
    ("Total hours calculation", 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1),
    ("Assignment history", 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1),
    ("36h daily limit (#2914)", 0, 0, 0, 0, 0, 0, 0, 3, 0, 0, 0, 0, 0, 3),
    ("Read-only enforcement", 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1),
    ("Real-time PATCH", 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 0, 2),
    ("Real-time ADD/DELETE", 0, 0, 0, 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Real-time RENAME/GENERATE", 0, 0, 0, 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Copy table feature", 0, 0, 0, 0, 0, 0, 0, 0, 0, 5, 0, 0, 0, 5),
    ("Copy table #3386 regression", 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 1),
    ("Tracker data loss (#2302)", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1),
    ("Tracker credential types", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, 0, 2),
    ("Tag CRUD in settings", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, 2),
]

RISK_ASSESSMENT = [
    ("DnD duplicate rows (#3332)", "Drag-and-drop creates ghost duplicate entries due to order array not removing old IDs. Root cause identified but fix commented out.", "High", "High", "Critical",
     "Test DnD with 3+ tasks. Count rows before/after. Verify no duplicates. Commented-out fix in generateTaskAssignments.ts."),
    ("DnD order reset (#3314)", "Task order resets when toggling 'Open for editing' due to useEffect re-sorting on every state change.", "High", "High", "Critical",
     "Reorder via DnD, then toggle editing. Verify order preserved. Root cause: TasksPlannerTable.tsx sort on readOnly change."),
    ("N+1 API pattern (performance)", "Frontend calls backend individually per project member. Large projects (DirectEnergie-ODC: 10+ members) cause severe lag.", "High", "High", "Critical",
     "Test with largest available project. Monitor network calls in DevTools. Count API requests vs member count."),
    ("Lock TTL race condition", "1-minute TTL with in-memory Caffeine cache. Server restart loses all locks. No persistence.", "Medium", "Medium", "Medium",
     "Test lock timeout. Verify lock lost on server restart. Test concurrent edit attempts."),
    ("36h validation bypass (#2914)", "Per-project validation misses cross-project totals. No hard block on exceeding 36h daily limit.", "High", "Medium", "High",
     "Report >36h split across 2 projects. Verify no block. Verify warning appears."),
    ("Tracker sync failure", "External tracker timeouts, auth failures cause intermittent sync failures. No retry logic.", "Medium", "Medium", "Medium",
     "Test with unconfigured tracker. Verify error message. Test with timeout conditions."),
    ("WebSocket disconnection", "500ms reconnect delay. Events missed during disconnection. No replay.", "Low", "Medium", "Medium",
     "Simulate disconnect. Verify reconnect. Check for stale data after reconnect."),
    ("window.location.reload in close-by-tag", "Blunt page refresh loses unsaved state across entire planner.", "Low", "Medium", "Medium",
     "Edit hours in one cell, trigger close-by-tag apply from another. Verify unsaved hours lost."),
    ("Copy table #3386 (deleted tasks)", "Copy table included deleted/closed tasks in clipboard output. Users copy stale data.", "Medium", "Medium", "Medium",
     "Delete a task, then Copy table. Verify deleted task excluded from clipboard. Verify closed-project tasks excluded."),
    ("Approved status cleared (#3296)", "Tracker import resets approved=false on all assignments, including already-approved ones. PM must re-approve.", "High", "High", "Critical",
     "Approve hours, then run tracker import. Verify approved flag persists. Test with both single approve and Approve all."),
    ("Task deletion refresh (#3294)", "Deleting a task on current date did not refresh planner view. Stale data remained visible.", "Medium", "Medium", "Medium",
     "Delete a task on today's date. Verify row removed immediately. Verify totals recalculate."),
]


# ─── Sheet Writers (same as t2724) ───────────────────────────────────────────

def write_test_cases(ws, cases):
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [16, 40, 45, 65, 45, 10, 8, 35, 22, 35]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_offset, tc in enumerate(cases):
        row_idx = row_offset + 3
        values = [
            tc["id"], tc["title"], tc["preconditions"], tc["steps"],
            tc["expected"], tc["priority"], tc["type"],
            tc["req_ref"], tc["module"], tc.get("notes", "")
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"\u2022 {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"\u2022 {env}").font = FONT_BODY
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"\u2022 {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} \u2014 {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Related").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value="Close-by-tag feature tests: see t2724/t2724.xlsx (38 test cases, 3 suites)").font = FONT_BODY

    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws):
    headers = ["Feature", "Nav", "CRUD", "DnD", "Lock", "Tracker",
               "Members", "Reports", "Valid", "RT", "Copy", "BugReg", "TrkAdv", "Settings", "Total"]
    col_widths = [35, 6, 7, 6, 7, 9, 10, 9, 7, 6, 7, 8, 8, 10, 8]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    suite_tabs = ["TS-PLN-Nav", "TS-PLN-CRUD", "TS-PLN-DnD", "TS-PLN-Lock",
                  "TS-PLN-Tracker", "TS-PLN-Members", "TS-PLN-Reports", "TS-PLN-Valid", "TS-PLN-RT",
                  "TS-PLN-Copy", "TS-PLN-BugReg", "TS-PLN-TrkAdv", "TS-PLN-Settings"]

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "\u2190 Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


# ─── Main ─────────��──────────────────────────────────────────────────────────

def generate():
    wb = Workbook()

    suites = [
        ("TS-Planner-Navigation", "TS-PLN-Nav", get_navigation_cases),
        ("TS-Planner-TaskCRUD", "TS-PLN-CRUD", get_task_crud_cases),
        ("TS-Planner-DnD", "TS-PLN-DnD", get_dnd_cases),
        ("TS-Planner-Lock", "TS-PLN-Lock", get_lock_cases),
        ("TS-Planner-Tracker", "TS-PLN-Tracker", get_tracker_cases),
        ("TS-Planner-Members", "TS-PLN-Members", get_members_cases),
        ("TS-Planner-Reports", "TS-PLN-Reports", get_reports_cases),
        ("TS-Planner-Validation", "TS-PLN-Valid", get_validation_cases),
        ("TS-Planner-RealTime", "TS-PLN-RT", get_realtime_cases),
        ("TS-Planner-CopyTable", "TS-PLN-Copy", get_copy_table_cases),
        ("TS-Planner-BugRegression", "TS-PLN-BugReg", get_bug_regression_cases),
        ("TS-Planner-TrackerAdvanced", "TS-PLN-TrkAdv", get_tracker_adv_cases),
        ("TS-Planner-Settings", "TS-PLN-Settings", get_settings_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    total_cases = sum(len(case_fn()) for _, _, case_fn in suites)
    print(f"Generating {total_cases} test cases across {len(suites)} suites")

    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(ws_plan, suites_info)

    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_feature_matrix(ws_fm)

    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ws_ra)

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_test_cases(ws, cases)
        print(f"  {tab_name}: {len(cases)} cases")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {total_cases} test cases, {len(suites)} suites, 3 plan tabs")

    return all_cases


if __name__ == "__main__":
    generate()
