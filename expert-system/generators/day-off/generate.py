#!/usr/bin/env python3
"""Generate day-off.xlsx — unified test workbook for Day-Off module.

Phase B output for the TTT Expert System.
Covers: CRUD, approval workflow, calendar conflicts, optional approvers,
        permissions, validation, error handling.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Styling constants ──────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "548235"
TAB_COLOR_TS = "2F5496"


# ── Helper functions ───────────────────────────────────────────────

def style_header_row(ws, row, num_cols, fill=None):
    """Apply header styling to a row."""
    f = fill or FILL_HEADER
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = f
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def write_row(ws, row, values, font=None, fill=None, alignment=None):
    """Write a row of values with optional styling."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = alignment or ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=10, max_width=60):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, min(longest + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_autofilter(ws, row, num_cols):
    """Add autofilter to header row."""
    ws.auto_filter.ref = f"A{row}:{get_column_letter(num_cols)}{ws.max_row}"


def add_back_link(ws, row=1):
    """Add back-link to Plan Overview in row 1."""
    cell = ws.cell(row=row, column=1)
    cell.value = "<- Back to Plan"
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab with header, back-link, and test cases."""
    add_back_link(ws, row=1)
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for i, tc_data in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        values = [
            tc_data["id"], tc_data["title"], tc_data["preconditions"],
            tc_data["steps"], tc_data["expected"], tc_data["priority"],
            tc_data["type"], tc_data["req_ref"], tc_data["module"],
            tc_data.get("notes", "")
        ]
        write_row(ws, row, values, fill=fill)

    add_autofilter(ws, header_row, len(headers))

    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ── Test Case Data ─────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


# ── TS-DO-CRUD ─────────────────────────────────────────────────────

TS_DO_CRUD = [
    tc("TC-DO-001",
       "Create day-off request — happy path",
       "Active employee with office calendar containing a public holiday\n"
       "Test data: SELECT c.date FROM calendar_day c JOIN office o ON c.calendar_id=o.calendar_id "
       "WHERE c.day_type='PUBLIC_HOLIDAY' AND c.date > CURRENT_DATE LIMIT 5",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: {employeeLogin, publicDate: <future holiday>, personalDate: <future workday>, "
       "originalDate: <same as publicDate>, duration: 8, reason: 'Test day-off'}\n"
       "3. Verify response status 200\n"
       "4. GET /api/vacation/v1/employee-dayOff?searchType=MY to confirm",
       "Day-off request created with status=NEW\n"
       "Approver auto-assigned (employee's manager)\n"
       "Optional approvers synced from employee defaults\n"
       "EmployeeDayOffCreatedEvent published",
       "Critical", "Functional",
       "EmployeeDayOffController.create", "EmployeeDayOffServiceImpl",
       "Core happy path. Verify via DB: SELECT * FROM employee_dayoff_request WHERE id=<new_id>"),

    tc("TC-DO-002",
       "Create day-off — CPO self-approves on creation",
       "CPO/Department Manager employee\n"
       "Test data: SELECT e.login FROM employee e JOIN employee_global_roles r "
       "ON e.id=r.employee_id WHERE r.role='ROLE_DEPARTMENT_MANAGER' AND e.status='ACTIVE' LIMIT 3",
       "1. POST /api/vacation/v1/employee-dayOff as CPO user\n"
       "2. Body: valid day-off request\n"
       "3. Check status and approver in response\n"
       "4. DB: SELECT status, approver_id FROM employee_dayoff_request WHERE id=<new_id>",
       "Request created with status=APPROVED (auto-approved)\n"
       "Approver = self (CPO's own ID)\n"
       "Ledger entries created immediately (credit + debit)\n"
       "Manager added as optional approver with ASKED status",
       "Critical", "Functional",
       "EmployeeDayOffServiceImpl.create", "EmployeeDayOffServiceImpl",
       "CPO pattern: self-approve + manager as optional. Same as vacation CPO flow."),

    tc("TC-DO-003",
       "Create day-off — duplicate publicDate rejected",
       "Employee with existing day-off request for a specific publicDate\n"
       "Test data: SELECT employee_login, original_date FROM employee_dayoff_request "
       "WHERE status IN ('NEW','APPROVED') LIMIT 5",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: publicDate = same as existing request's originalDate\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: validation.EmployeeDayOffPublicDateExists.message\n"
       "Duplicate request for same public date blocked by @EmployeeDayOffPublicDateExists validator",
       "High", "Negative",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Custom validator checks for existing request with same publicDate"),

    tc("TC-DO-004",
       "Create day-off — publicDate not in calendar",
       "Active employee, date that is NOT in office calendar as public holiday",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: publicDate = a regular workday (not in calendar)\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: validation.PublicDateNotFoundInCalendar.message\n"
       "Date not found in office calendar OR employee's day-off calendar",
       "High", "Negative",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Validator checks date exists in office calendar OR employee's dayoff calendar"),

    tc("TC-DO-005",
       "Create day-off — personalDate already used",
       "Employee with existing day-off using a specific personalDate\n"
       "Test data: SELECT personal_date FROM employee_dayoff_request WHERE status='APPROVED'",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: personalDate = date already used by another day-off request\n"
       "3. Check error response",
       "HTTP 400\n"
       "PersonalDateValidator rejects: personalDate already used by another request\n"
       "Error in personalDate field validation",
       "High", "Negative",
       "EmployeeDayOffPersonalDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "@EmployeeDayOffPersonalDateExists custom validator"),

    tc("TC-DO-006",
       "Create day-off — all fields null (missing @NotNull)",
       "Active employee",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: {} (empty object, all fields null)\n"
       "3. Check response",
       "DESIGN ISSUE: No @NotNull on any DTO field\n"
       "Custom validators treat null as valid (return true for null)\n"
       "Possible NPE in service layer when processing null fields\n"
       "May create request with all null values",
       "High", "Negative",
       "DI: Missing @NotNull on DTO fields", "EmployeeDayOffCreateRequestDTO",
       "All DTO fields lack @NotNull. Both custom validators return true for null input."),

    tc("TC-DO-007",
       "Create day-off — null duration (no validation)",
       "Active employee with valid publicDate and personalDate",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: valid publicDate/personalDate, duration: null\n"
       "3. Check response and DB",
       "DESIGN ISSUE: duration has no @NotNull, no @Min/@Max\n"
       "Null duration accepted — stored as null in DB\n"
       "Ledger entries may have null duration\n"
       "Downstream calculations may NPE on null duration",
       "High", "Negative",
       "DI: No duration validation", "EmployeeDayOffCreateRequestDTO",
       "No validation annotations on duration field at all"),

    tc("TC-DO-008",
       "Create day-off — negative duration",
       "Active employee with valid publicDate and personalDate",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: duration: -8\n"
       "3. Check response and DB",
       "DESIGN ISSUE: No @Min on duration — negative value accepted\n"
       "Negative duration creates inverted ledger entries\n"
       "Credit becomes debit and vice versa\n"
       "Corrupts vacation day balance calculation",
       "High", "Boundary",
       "DI: No duration validation", "EmployeeDayOffCreateRequestDTO",
       "No @Min/@Max annotations. Negative duration inverts credit/debit logic."),

    tc("TC-DO-009",
       "Create day-off — zero duration",
       "Active employee with valid dates",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: duration: 0\n"
       "3. Check response and ledger",
       "Duration 0 accepted (no validation)\n"
       "In ledger: duration=0 means day-off taken (debit)\n"
       "But on credit side, duration=0 means no credit given\n"
       "Net effect: takes a day off without earning credit for worked holiday",
       "Medium", "Boundary",
       "DI: No duration validation", "EmployeeDayOffCreateRequestDTO, employee_dayoff",
       "Boundary: zero duration has semantic meaning in ledger (debit marker)"),

    tc("TC-DO-010",
       "Create day-off — very long reason (no size limit)",
       "Active employee with valid dates",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: reason: 'A' * 10000 (10K character string)\n"
       "3. Check response",
       "DESIGN ISSUE: No @Size on reason field\n"
       "Very long string accepted by DTO validation\n"
       "DB column type determines actual limit\n"
       "May cause truncation or DB error if column is VARCHAR(N)",
       "Low", "Boundary",
       "DI: No reason validation", "EmployeeDayOffCreateRequestDTO",
       "No size limit on reason. DB column type is the only constraint."),

    tc("TC-DO-011",
       "PATCH personalDate — happy path",
       "Existing day-off in NEW status owned by current user\n"
       "Test data: SELECT id FROM employee_dayoff_request WHERE status='NEW' LIMIT 5",
       "1. PATCH /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Body: {personalDate: <new valid date>}\n"
       "3. Verify response\n"
       "4. DB: SELECT personal_date FROM employee_dayoff_request WHERE id=<id>",
       "personalDate updated successfully\n"
       "Status remains unchanged\n"
       "Only personalDate field is patchable (EmployeeDayOffPatchRequestDTO)\n"
       "originalDate remains immutable",
       "High", "Functional",
       "EmployeeDayOffController.patch", "EmployeeDayOffServiceImpl",
       "PATCH only allows personalDate change. originalDate is immutable."),

    tc("TC-DO-012",
       "PATCH personalDate on APPROVED day-off",
       "Existing APPROVED day-off owned by current user",
       "1. PATCH /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Body: {personalDate: <new date>}\n"
       "3. Check response",
       "DESIGN ISSUE: EDIT permission is unconditional for owner\n"
       "PATCH succeeds even on APPROVED request\n"
       "Ledger entries NOT updated (still reference old personalDate)\n"
       "Creates data inconsistency between request and ledger",
       "High", "Functional",
       "DI: Unconditional EDIT", "EmployeeDayOffPermissionService",
       "EDIT has no status check. Owner can PATCH any status including terminal ones."),

    tc("TC-DO-013",
       "PATCH personalDate on DELETED day-off",
       "Existing DELETED day-off owned by current user\n"
       "Test data: SELECT id FROM employee_dayoff_request WHERE status='DELETED'",
       "1. PATCH /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Body: {personalDate: <new date>}\n"
       "3. Check response",
       "DESIGN ISSUE: EDIT unconditional — succeeds on DELETED request\n"
       "Can modify soft-deleted records\n"
       "No business logic prevents editing terminal states",
       "High", "Negative",
       "DI: Unconditional EDIT", "EmployeeDayOffPermissionService",
       "Bug: owner can PATCH even DELETED/DELETED_FROM_CALENDAR requests"),

    tc("TC-DO-014",
       "PATCH personalDate — date already used",
       "Existing NEW day-off, another day-off using target personalDate",
       "1. PATCH /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Body: {personalDate: <date used by another request>}\n"
       "3. Check error response",
       "HTTP 400\n"
       "@EmployeeDayOffPersonalDateExists validator rejects duplicate\n"
       "personalDate must be unique across employee's requests",
       "Medium", "Negative",
       "EmployeeDayOffPersonalDateExistsValidator", "EmployeeDayOffPatchRequestDTO",
       "Same validator used for create and patch"),

    tc("TC-DO-015",
       "DELETE day-off — happy path (NEW status)",
       "Existing day-off in NEW status owned by current user",
       "1. DELETE /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Verify response\n"
       "3. DB: SELECT status FROM employee_dayoff_request WHERE id=<id>",
       "Status changed to DELETED\n"
       "Soft-delete (record remains in DB)\n"
       "No ledger changes (no ledger entries for NEW requests)\n"
       "canBeCancelled was true (status != APPROVED)",
       "High", "Functional",
       "EmployeeDayOffController.delete", "EmployeeDayOffServiceImpl",
       "DELETE permission requires canBeCancelled = true"),

    tc("TC-DO-016",
       "DELETE day-off — APPROVED with personalDate >= report period",
       "APPROVED day-off with personalDate in current or future report period\n"
       "Test data: SELECT id, personal_date FROM employee_dayoff_request "
       "WHERE status='APPROVED' AND personal_date >= CURRENT_DATE",
       "1. DELETE /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Verify response\n"
       "3. Check status and ledger entries",
       "canBeCancelled = true (personalDate >= report period start)\n"
       "Status changed to DELETED\n"
       "Ledger entries remain (not cleaned up on DELETE)\n"
       "Vacation day balance should be recalculated",
       "High", "Functional",
       "EmployeeDayOffPermissionService.canBeCancelled", "EmployeeDayOffServiceImpl",
       "canBeCancelled: status != APPROVED || personalDate >= reportPeriod"),

    tc("TC-DO-017",
       "DELETE day-off — APPROVED with personalDate < report period (blocked)",
       "APPROVED day-off with personalDate in past closed report period",
       "1. Find APPROVED day-off with personalDate before report period start\n"
       "2. DELETE /api/vacation/v1/employee-dayOff/{id}\n"
       "3. Check error response",
       "HTTP 403 or permission denied\n"
       "canBeCancelled = false (APPROVED AND personalDate < report period)\n"
       "DELETE permission not granted\n"
       "Day-off in closed period cannot be cancelled",
       "High", "Negative",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Prevents deletion of day-offs already accounted for in closed periods"),

    tc("TC-DO-018",
       "DELETE day-off — not owner (permission denied)",
       "Day-off owned by another employee, authenticated as non-owner non-approver",
       "1. DELETE /api/vacation/v1/employee-dayOff/{id} (someone else's request)\n"
       "2. Check error response",
       "HTTP 403\n"
       "errorCode: exception.employee.dayOff.no.permission\n"
       "DELETE permission only granted to owner",
       "High", "Security",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Only owner gets DELETE permission (via canBeCancelled check)"),

    tc("TC-DO-019",
       "Create day-off — day-off crossing vacation",
       "Employee with existing approved vacation on the personalDate",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: personalDate = date within employee's existing vacation range\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: exception.day.off.crossing.vacation\n"
       "DayOffCrossingVacationException thrown\n"
       "Cannot take day-off on a date covered by vacation",
       "High", "Negative",
       "EmployeeDayOffServiceImpl", "EmployeeDayOffServiceImpl",
       "Cross-module validation: day-off vs vacation date overlap"),

    tc("TC-DO-020",
       "Upsert on creation — silent overwrite risk",
       "Manipulate data so upsert condition is met (same employee + originalDate in ledger)",
       "1. Create day-off request for publicDate X\n"
       "2. Verify ledger entry created\n"
       "3. Manipulate: create condition where upsert finds existing record\n"
       "4. Check if original record silently overwritten",
       "DESIGN ISSUE: Upsert on creation can silently overwrite records\n"
       "If matching record exists in employee_dayoff table, it gets updated\n"
       "No conflict detection or error raised\n"
       "Data integrity risk: original record data lost",
       "Medium", "Negative",
       "DI: Upsert silent overwrite", "EmployeeDayOffServiceImpl",
       "Upsert pattern: INSERT ON CONFLICT UPDATE — silently replaces data"),
]


# ── TS-DO-Approval ─────────────────────────────────────────────────

TS_DO_APPROVAL = [
    tc("TC-DO-021",
       "Approve NEW day-off — happy path with ledger creation",
       "Day-off in NEW status, authenticated as assigned approver\n"
       "Test data: SELECT r.id, r.approver_login FROM employee_dayoff_request r "
       "WHERE r.status='NEW' LIMIT 5",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Verify response\n"
       "3. DB: SELECT * FROM employee_dayoff WHERE employee_id=<emp_id> "
       "AND original_date=<date>\n"
       "4. Check vacation balance change",
       "Status changed to APPROVED\n"
       "Two ledger entries created:\n"
       "  - Credit: lastApprovedDate slot, duration from calendar/norm\n"
       "  - Debit: personalDate slot, duration=0 + reason from request\n"
       "RecalculateVacationDaysHandler fires (diff=+1/-1)\n"
       "UpdateMonthNormHandler recalculates norms for both months",
       "Critical", "Functional",
       "EmployeeDayOffServiceImpl.approve", "EmployeeDayOffServiceImpl",
       "Core approval flow. Verify both ledger entries and vacation balance."),

    tc("TC-DO-022",
       "Approve REJECTED day-off — re-approval",
       "Day-off in REJECTED status, authenticated as approver\n"
       "REJECTED is in approvableStatuses set",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Verify status changes to APPROVED\n"
       "3. Check ledger entries created",
       "Status REJECTED -> APPROVED\n"
       "Approvable statuses: {NEW, REJECTED}\n"
       "Ledger entries created fresh\n"
       "Vacation balance recalculated",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffServiceImpl",
       "Re-approval after rejection. Both NEW and REJECTED are approvable."),

    tc("TC-DO-023",
       "Approve already APPROVED day-off (invalid)",
       "Day-off in APPROVED status",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Check error response",
       "APPROVED not in approvableStatuses {NEW, REJECTED}\n"
       "Permission check fails — APPROVE not granted\n"
       "HTTP 403: exception.employee.dayOff.no.permission",
       "Medium", "Negative",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "APPROVED is not in approvable set — cannot double-approve"),

    tc("TC-DO-024",
       "Approve DELETED day-off (invalid)",
       "Day-off in DELETED status",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Check error response",
       "DELETED not in approvableStatuses\n"
       "HTTP 403: exception.employee.dayOff.no.permission\n"
       "Cannot approve a deleted request",
       "Medium", "Negative",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Terminal status — no APPROVE permission granted"),

    tc("TC-DO-025",
       "Reject NEW day-off — happy path",
       "Day-off in NEW status, authenticated as approver",
       "1. PUT /api/vacation/v1/employee-dayOff/reject/{id}\n"
       "2. Verify response\n"
       "3. DB: SELECT status FROM employee_dayoff_request WHERE id=<id>",
       "Status changed to REJECTED\n"
       "No ledger entries created or modified\n"
       "Rejectable statuses: {NEW, APPROVED}\n"
       "Simple status change only",
       "High", "Functional",
       "EmployeeDayOffServiceImpl.reject", "EmployeeDayOffServiceImpl",
       "Reject is simpler than approve — no ledger operations"),

    tc("TC-DO-026",
       "Reject APPROVED day-off — personalDate >= report period",
       "APPROVED day-off with personalDate in current/future period\n"
       "Authenticated as approver",
       "1. PUT /api/vacation/v1/employee-dayOff/reject/{id}\n"
       "2. Verify response\n"
       "3. Check ledger and vacation balance",
       "Status APPROVED -> REJECTED\n"
       "REJECT permission granted: NEW/APPROVED AND personalDate >= report period\n"
       "Ledger entries remain (not cleaned up)\n"
       "Vacation balance should be recalculated to reverse the day-off effect",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffServiceImpl",
       "Rejecting APPROVED requires personalDate >= report period start"),

    tc("TC-DO-027",
       "Reject APPROVED day-off — personalDate < report period (blocked)",
       "APPROVED day-off with personalDate in past closed report period",
       "1. PUT /api/vacation/v1/employee-dayOff/reject/{id}\n"
       "2. Check error response",
       "HTTP 403: exception.employee.dayOff.no.permission\n"
       "REJECT not granted: APPROVED + personalDate < report period\n"
       "Cannot reject day-off already in closed accounting period",
       "High", "Negative",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Report period gate prevents rejecting accounted day-offs"),

    tc("TC-DO-028",
       "Reject REJECTED day-off (already rejected)",
       "Day-off already in REJECTED status",
       "1. PUT /api/vacation/v1/employee-dayOff/reject/{id}\n"
       "2. Check error response",
       "REJECTED is in rejectableStatuses {NEW, APPROVED}\n"
       "Wait — REJECTED IS in the set? Check: rejectableStatuses = {NEW, APPROVED}\n"
       "REJECTED not in set -> permission denied\n"
       "HTTP 403: exception.employee.dayOff.no.permission",
       "Medium", "Negative",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Rejectables are {NEW, APPROVED}, not REJECTED"),

    tc("TC-DO-029",
       "Reject as non-approver (permission denied)",
       "Day-off in NEW status, authenticated as non-approver non-owner",
       "1. PUT /api/vacation/v1/employee-dayOff/reject/{id}\n"
       "2. Check error response",
       "HTTP 403: exception.employee.dayOff.no.permission\n"
       "REJECT permission only for assigned approver\n"
       "Owner cannot reject own request",
       "High", "Security",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Only the assigned approver can reject"),

    tc("TC-DO-030",
       "Change approver — happy path",
       "Day-off in NEW status, authenticated as current approver\n"
       "Valid new approver login",
       "1. PUT /api/vacation/v1/employee-dayOff/change-approver/{id}/{newApproverLogin}\n"
       "2. Verify response\n"
       "3. DB: SELECT approver_login FROM employee_dayoff_request WHERE id=<id>\n"
       "4. Check optional approver cascade",
       "Approver changed to newApproverLogin\n"
       "Old approver moved to optional approvers with ASKED status\n"
       "New approver removed from optional approvers (if was optional)\n"
       "Cascade: approver change triggers optional approver reshuffle",
       "High", "Functional",
       "EmployeeDayOffController.changeApprover", "EmployeeDayOffServiceImpl",
       "Approver change cascades to optional approvers"),

    tc("TC-DO-031",
       "Change approver on DELETED day-off (unconditional permission)",
       "Day-off in DELETED status, authenticated as approver",
       "1. PUT /api/vacation/v1/employee-dayOff/change-approver/{id}/{newLogin}\n"
       "2. Check response",
       "DESIGN ISSUE: EDIT_APPROVER is unconditional — no status check\n"
       "Change approver succeeds even on DELETED request\n"
       "Can modify approver on DELETED_FROM_CALENDAR too\n"
       "No business value — approver change on terminal status",
       "High", "Negative",
       "DI: Unconditional EDIT_APPROVER", "EmployeeDayOffPermissionService",
       "Bug: EDIT_APPROVER has no status check, always granted to approver"),

    tc("TC-DO-032",
       "Change approver on DELETED_FROM_CALENDAR day-off",
       "Day-off in DELETED_FROM_CALENDAR status",
       "1. PUT /api/vacation/v1/employee-dayOff/change-approver/{id}/{newLogin}\n"
       "2. Check response",
       "DESIGN ISSUE: Succeeds — EDIT_APPROVER unconditional\n"
       "DELETED_FROM_CALENDAR is a terminal status\n"
       "Approver change has no meaningful effect\n"
       "Optional approver cascade still executes unnecessarily",
       "Medium", "Negative",
       "DI: Unconditional EDIT_APPROVER", "EmployeeDayOffPermissionService",
       "Same bug as TC-DO-031 on different terminal status"),

    tc("TC-DO-033",
       "System rejection — bulk reject when calendar day removed",
       "Multiple NEW day-off requests for the same public holiday date\n"
       "Admin removes that date from production calendar",
       "1. Find office with public holiday having NEW day-off requests\n"
       "2. Admin: remove public holiday from calendar\n"
       "3. System calls rejectedBySystem(officeId, date)\n"
       "4. Check all affected requests",
       "All NEW requests with matching date set to REJECTED\n"
       "System rejection (not manual reject)\n"
       "Only NEW status affected — APPROVED untouched by rejectedBySystem\n"
       "Different from deleteDayOffs which sets DELETED_FROM_CALENDAR",
       "Critical", "Functional",
       "EmployeeDayOffServiceImpl.rejectedBySystem", "EmployeeDayOffServiceImpl",
       "System rejection targets NEW only. deleteDayOffs targets NEW+APPROVED."),

    tc("TC-DO-034",
       "Approve — ledger credit duration from calendar",
       "NEW day-off for a date that exists in office calendar with known duration",
       "1. Check calendar_day duration for the public holiday\n"
       "2. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "3. DB: SELECT duration FROM employee_dayoff WHERE original_date=<date> AND duration > 0",
       "Credit ledger entry duration = calendar day duration (typically 8 or 7)\n"
       "Duration sourced from: existing ledger OR calendar OR reporting-norm fallback\n"
       "Fallback chain: ledger -> calendar -> norm (8h default)",
       "High", "Functional",
       "EmployeeDayOffServiceImpl.changeDayOffDaysAfterApprove", "EmployeeDayOffServiceImpl",
       "Verify credit duration matches calendar. 8h=full day, 7h=half-working day."),

    tc("TC-DO-035",
       "Approve — ledger debit entry verification",
       "NEW day-off request with specific personalDate and reason",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. DB: SELECT * FROM employee_dayoff WHERE personal_date=<personalDate> AND duration=0",
       "Debit ledger entry:\n"
       "  - duration = 0 (marks day-off taken)\n"
       "  - reason copied from request\n"
       "  - personal_date matches request personalDate\n"
       "  - employee matches request employee",
       "High", "Functional",
       "EmployeeDayOffServiceImpl.changeDayOffDaysAfterApprove", "EmployeeDayOffServiceImpl",
       "duration=0 is the debit marker. Reason propagated from request."),

    tc("TC-DO-036",
       "Approve — non-atomic transaction risk",
       "Simulate concurrent access during approve",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Monitor: ledger write and status update are separate calls\n"
       "3. If interruption between them: check for orphaned ledger entries\n"
       "4. DB: compare employee_dayoff entries vs request status",
       "DESIGN ISSUE: No @Transactional on approve\n"
       "Uses performInTransaction wrapper instead\n"
       "Ledger entries written separately from status update\n"
       "Failure between steps = ledger entries without APPROVED status\n"
       "Risk: orphaned ledger records corrupt vacation balance",
       "High", "Integration",
       "DI: No @Transactional on approve", "EmployeeDayOffServiceImpl",
       "Non-atomic: changeDayOffDaysAfterApprove + changeDayOffStatus are separate"),
]


# ── TS-DO-Calendar ─────────────────────────────────────────────────

TS_DO_CALENDAR = [
    tc("TC-DO-037",
       "Calendar removal — DELETED_FROM_CALENDAR for NEW+APPROVED requests",
       "Multiple day-off requests (NEW and APPROVED) for same originalDate\n"
       "Admin removes that date from production calendar",
       "1. Identify date with both NEW and APPROVED requests\n"
       "2. Admin: remove date from production calendar\n"
       "3. System calls deleteDayOffs(date)\n"
       "4. DB: SELECT status FROM employee_dayoff_request WHERE original_date=<date>",
       "All NEW and APPROVED requests: status -> DELETED_FROM_CALENDAR\n"
       "Ledger entries physically deleted (not soft-delete)\n"
       "Vacation days recalculated for each affected employee\n"
       "EmployeeDayOffDeletedFromCalendarEvent published",
       "Critical", "Functional",
       "EmployeeDayOffServiceImpl.deleteDayOffs", "EmployeeDayOffServiceImpl",
       "Physical deletion of ledger entries. Different from system rejection."),

    tc("TC-DO-038",
       "Calendar removal — REJECTED requests unaffected",
       "REJECTED day-off request for a date being removed from calendar",
       "1. Ensure REJECTED request exists for target date\n"
       "2. Remove date from calendar\n"
       "3. Check REJECTED request status",
       "REJECTED request status unchanged\n"
       "deleteDayOffs only targets NEW and APPROVED\n"
       "REJECTED, DELETED, DELETED_FROM_CALENDAR are not affected",
       "Medium", "Functional",
       "EmployeeDayOffServiceImpl.deleteDayOffs", "EmployeeDayOffServiceImpl",
       "Only NEW/APPROVED targeted. Terminal statuses excluded."),

    tc("TC-DO-039",
       "Calendar removal — ledger entries physically deleted",
       "APPROVED day-off with ledger entries, calendar date being removed",
       "1. Verify ledger entries exist: SELECT * FROM employee_dayoff WHERE original_date=<date>\n"
       "2. Remove date from calendar\n"
       "3. SELECT * FROM employee_dayoff WHERE original_date=<date>",
       "Ledger entries physically deleted from employee_dayoff table\n"
       "Not soft-delete — records are gone\n"
       "This is different from request table (soft-delete via status)\n"
       "Vacation balance recalculated after ledger deletion",
       "High", "Functional",
       "EmployeeDayOffServiceImpl.deleteDayOffs", "EmployeeDayOffServiceImpl",
       "Physical DELETE on ledger table vs status change on request table"),

    tc("TC-DO-040",
       "Office change — auto-delete all day-offs for year",
       "Employee with multiple NEW/APPROVED day-offs changes office\n"
       "AutoDeleteHelper.update(employeeId, year) triggered",
       "1. Note employee's day-offs for current year\n"
       "2. Change employee's office\n"
       "3. AutoDeleteHelper.update fires\n"
       "4. Check all day-off requests and ledger entries",
       "All NEW and APPROVED day-offs: status -> DELETED_FROM_CALENDAR\n"
       "ALL ledger entries for the year deleted (physical delete)\n"
       "Employee notified of changes\n"
       "Office calendar mismatch resolved",
       "Critical", "Functional",
       "AutoDeleteHelper.update", "AutoDeleteHelper",
       "Cascading cleanup on office change. Affects entire year."),

    tc("TC-DO-041",
       "Office change — PAGE_SIZE=100 hard limit",
       "Employee with >100 day-off requests for a year (edge case)\n"
       "Requires test data setup",
       "1. Create >100 day-off requests for one employee in one year\n"
       "2. Trigger office change -> AutoDeleteHelper.update\n"
       "3. Count remaining requests not set to DELETED_FROM_CALENDAR",
       "DESIGN ISSUE: PAGE_SIZE=100 hard limit\n"
       "Only first 100 day-offs processed per page\n"
       "Requests 101+ remain in NEW/APPROVED status\n"
       "Partial cleanup — data integrity violation\n"
       "No pagination loop to process all records",
       "High", "Boundary",
       "DI: PAGE_SIZE=100 limit", "AutoDeleteHelper",
       "Hard limit. No loop. >100 records = incomplete cleanup."),

    tc("TC-DO-042",
       "Office change — ledger entries for entire year deleted",
       "Employee with ledger entries across multiple months, changes office",
       "1. DB: SELECT COUNT(*) FROM employee_dayoff WHERE employee_id=<id> AND "
       "EXTRACT(YEAR FROM original_date)=2026\n"
       "2. Trigger office change\n"
       "3. Repeat count query",
       "ALL ledger entries for the year physically deleted\n"
       "Not limited to matched requests — entire year wiped\n"
       "Includes entries for REJECTED/DELETED requests\n"
       "Full recalculation triggered after deletion",
       "High", "Functional",
       "AutoDeleteHelper.update", "AutoDeleteHelper",
       "Deletes ALL ledger entries for year, not just NEW/APPROVED ones"),

    tc("TC-DO-043",
       "CalendarUpdateHasDayOffConflictEvent — diff=0 (half-working day)",
       "Calendar change creates half-working day that intersects existing day-off",
       "1. Admin: modify calendar day to half-working (diff=0 in event)\n"
       "2. CalendarUpdateHasDayOffConflictEvent fires\n"
       "3. Check notification sent to employee",
       "diff=0: half-working day notification sent\n"
       "No vacation recalculation (diff=0 means no full-day change)\n"
       "Informational notification only\n"
       "Employee informed of calendar change affecting their day-off",
       "Medium", "Functional",
       "CalendarUpdateHasDayOffConflictEvent handler", "Event handlers",
       "diff=0 means calendar change is informational, not destructive"),

    tc("TC-DO-044",
       "CalendarUpdateHasDayOffConflictEvent — diff!=0 (full day change)",
       "Calendar change creates full-day change intersecting existing day-off",
       "1. Admin: modify calendar day type (diff!=0)\n"
       "2. CalendarUpdateHasDayOffConflictEvent fires\n"
       "3. Check notification and vacation recalculation",
       "diff!=0: full day-off notification sent\n"
       "Vacation recalculation triggered\n"
       "Employee notified of material change\n"
       "May trigger DELETED_FROM_CALENDAR if day is removed entirely",
       "Medium", "Functional",
       "CalendarUpdateHasDayOffConflictEvent handler", "Event handlers",
       "diff!=0 triggers both notification and recalculation"),

    tc("TC-DO-045",
       "Concurrent calendar removal and manual approval",
       "Day-off in NEW status while admin removes calendar date simultaneously",
       "1. Start two concurrent operations:\n"
       "   a. Admin removes calendar date (triggers deleteDayOffs)\n"
       "   b. Approver approves the day-off\n"
       "2. Check final state of request and ledger",
       "Race condition risk:\n"
       "Scenario A wins: status=DELETED_FROM_CALENDAR, no ledger\n"
       "Scenario B wins: status=APPROVED, ledger exists but calendar gone\n"
       "No locking mechanism to prevent concurrent state mutation\n"
       "Could result in APPROVED day-off for non-existent calendar date",
       "High", "Integration",
       "DI: No transaction isolation", "EmployeeDayOffServiceImpl",
       "Race between deleteDayOffs and approve. No optimistic locking."),
]


# ── TS-DO-OptApprover ─────────────────────────────────────────────

TS_DO_OPT_APPROVER = [
    tc("TC-DO-046",
       "Add optional approver — happy path",
       "Existing day-off, authenticated as owner or approver\n"
       "Valid colleague login who is not creator or main approver",
       "1. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "2. Body: {dayOffId: <id>, approverLogin: <colleague>}\n"
       "3. Verify response\n"
       "4. DB: SELECT * FROM employee_dayoff_approval WHERE request_id=<id>",
       "Optional approver added with status=ASKED\n"
       "Record created in employee_dayoff_approval table\n"
       "FYI notification sent to optional approver",
       "High", "Functional",
       "EmployeeDayOffApprovalController.create", "EmployeeDayOffApprovalServiceImpl",
       "Optional approvers are FYI — their approval/rejection is informational"),

    tc("TC-DO-047",
       "Add optional approver — constraint: day-off must exist",
       "Non-existent dayOffId",
       "1. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "2. Body: {dayOffId: 999999, approverLogin: <valid>}\n"
       "3. Check error response",
       "Error: day-off not found\n"
       "Constraint 1 of 4: day-off must exist\n"
       "HTTP 400 or 404",
       "Medium", "Negative",
       "EmployeeDayOffApprovalServiceImpl", "EmployeeDayOffApprovalServiceImpl",
       "First of 4 create constraints validated"),

    tc("TC-DO-048",
       "Add optional approver — constraint: cannot add creator",
       "Existing day-off, try to add the day-off creator as optional approver",
       "1. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "2. Body: {dayOffId: <id>, approverLogin: <creator's login>}\n"
       "3. Check error response",
       "Error: cannot add creator as optional approver\n"
       "Constraint 2 of 4: self-approval prevention",
       "Medium", "Negative",
       "EmployeeDayOffApprovalServiceImpl", "EmployeeDayOffApprovalServiceImpl",
       "Cannot add request owner as their own optional approver"),

    tc("TC-DO-049",
       "Add optional approver — constraint: cannot add main approver",
       "Existing day-off, try to add the main approver as optional approver",
       "1. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "2. Body: {dayOffId: <id>, approverLogin: <main approver login>}\n"
       "3. Check error response",
       "Error: cannot add main approver as optional\n"
       "Constraint 3 of 4: main approver already reviews\n"
       "Prevents duplicate notification",
       "Medium", "Negative",
       "EmployeeDayOffApprovalServiceImpl", "EmployeeDayOffApprovalServiceImpl",
       "Main approver is already the decision-maker — no optional role needed"),

    tc("TC-DO-050",
       "Add optional approver — constraint: cannot add duplicate",
       "Existing day-off with existing optional approver, add same person again",
       "1. POST /api/vacation/v1/employee-dayOff-approvers (add person X)\n"
       "2. POST /api/vacation/v1/employee-dayOff-approvers (add person X again)\n"
       "3. Check error response on second call",
       "Error: duplicate optional approver\n"
       "Constraint 4 of 4: uniqueness check\n"
       "Same person cannot be optional approver twice",
       "Medium", "Negative",
       "EmployeeDayOffApprovalServiceImpl", "EmployeeDayOffApprovalServiceImpl",
       "Prevents duplicate entries in approval table"),

    tc("TC-DO-051",
       "Add optional approver — access: only owner/approver/manager",
       "Authenticated as unrelated employee (not owner, approver, or manager)",
       "1. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "2. Body: valid dayOffId, valid approverLogin\n"
       "3. Check error response",
       "HTTP 403 or permission error\n"
       "Only owner, main approver, or employee's manager can add optional approvers\n"
       "Unrelated employees blocked",
       "High", "Security",
       "EmployeeDayOffApprovalServiceImpl access check", "EmployeeDayOffApprovalServiceImpl",
       "Access validation: owner OR approver OR manager of the employee"),

    tc("TC-DO-052",
       "Update optional approver status — approver approves",
       "Existing optional approval in ASKED status, authenticated as the optional approver",
       "1. PATCH /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "2. Body: {status: 'APPROVED'}\n"
       "3. Verify response\n"
       "4. DB: SELECT status FROM employee_dayoff_approval WHERE id=<id>",
       "Status ASKED -> APPROVED\n"
       "Only the optional approver themselves can update\n"
       "FYI only — does not affect main request status\n"
       "Informational approval recorded",
       "High", "Functional",
       "EmployeeDayOffApprovalController.patch", "EmployeeDayOffApprovalServiceImpl",
       "Self-service: only the optional approver can change their own status"),

    tc("TC-DO-053",
       "Update optional approver status — approver rejects",
       "Existing optional approval in ASKED status, authenticated as the optional approver",
       "1. PATCH /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "2. Body: {status: 'REJECTED'}\n"
       "3. Verify response",
       "Status ASKED -> REJECTED\n"
       "FYI rejection — does not affect main request status\n"
       "Main approver still decides independently",
       "High", "Functional",
       "EmployeeDayOffApprovalController.patch", "EmployeeDayOffApprovalServiceImpl",
       "Optional rejection is informational only"),

    tc("TC-DO-054",
       "Update optional approver status — wrong person (not the approver)",
       "Existing optional approval, authenticated as different employee",
       "1. PATCH /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "2. Body: {status: 'APPROVED'}\n"
       "3. Authenticated as someone other than the optional approver",
       "HTTP 403 or error\n"
       "Only the optional approver themselves can update their status\n"
       "Owner cannot approve on behalf of optional approver",
       "High", "Security",
       "EmployeeDayOffApprovalServiceImpl status update", "EmployeeDayOffApprovalServiceImpl",
       "Self-service only. No delegation."),

    tc("TC-DO-055",
       "Delete optional approver",
       "Existing optional approval, authenticated as owner or approver",
       "1. DELETE /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "2. Verify response\n"
       "3. DB: verify record removed",
       "Optional approver record deleted\n"
       "Physical delete from employee_dayoff_approval table\n"
       "No cascade effects on main request",
       "Medium", "Functional",
       "EmployeeDayOffApprovalController.delete", "EmployeeDayOffApprovalServiceImpl",
       "Removal of optional approver is clean — no side effects"),

    tc("TC-DO-056",
       "Optional approver POST uses VACATIONS_VIEW (security gap)",
       "User with VACATIONS_VIEW but NOT VACATIONS_CREATE",
       "1. Authenticate as user with VIEW-only permission\n"
       "2. POST /api/vacation/v1/employee-dayOff-approvers\n"
       "3. Check if write operation succeeds",
       "DESIGN ISSUE: POST controller uses @PreAuthorize VACATIONS_VIEW\n"
       "Write operation protected only by VIEW permission\n"
       "Any user with view access can add optional approvers\n"
       "Should require VACATIONS_CREATE or VACATIONS_EDIT",
       "Critical", "Security",
       "DI: VACATIONS_VIEW for POST", "EmployeeDayOffApprovalController",
       "Security gap: write operation behind read-only permission"),

    tc("TC-DO-057",
       "Optional approver PATCH uses VACATIONS_VIEW (security gap)",
       "User with VACATIONS_VIEW but NOT VACATIONS_EDIT",
       "1. Authenticate as user with VIEW-only permission\n"
       "2. PATCH /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "3. Check if status update succeeds",
       "DESIGN ISSUE: PATCH controller uses @PreAuthorize VACATIONS_VIEW\n"
       "Status update protected only by VIEW permission\n"
       "However, service-level check (only optional approver can update) adds protection\n"
       "Still: @PreAuthorize should match operation type",
       "High", "Security",
       "DI: VACATIONS_VIEW for PATCH", "EmployeeDayOffApprovalController",
       "Defense in depth: @PreAuthorize too permissive, service check compensates"),

    tc("TC-DO-058",
       "Optional approver DELETE uses VACATIONS_VIEW (security gap)",
       "User with VACATIONS_VIEW but NOT VACATIONS_DELETE",
       "1. Authenticate as user with VIEW-only permission\n"
       "2. DELETE /api/vacation/v1/employee-dayOff-approvers/{id}\n"
       "3. Check if deletion succeeds",
       "DESIGN ISSUE: DELETE controller uses @PreAuthorize VACATIONS_VIEW\n"
       "Deletion protected only by VIEW permission\n"
       "Service-level access check (owner/approver/manager) provides some protection\n"
       "But @PreAuthorize mismatch is a security smell",
       "High", "Security",
       "DI: VACATIONS_VIEW for DELETE", "EmployeeDayOffApprovalController",
       "All 3 write operations on optional approver use VIEW permission"),

    tc("TC-DO-059",
       "Approver change cascade — old approver becomes optional",
       "Day-off with optional approvers, change main approver",
       "1. Add optional approver X to day-off\n"
       "2. PUT /api/vacation/v1/employee-dayOff/change-approver/{id}/{newLogin}\n"
       "3. Check optional approvers list\n"
       "4. DB: SELECT * FROM employee_dayoff_approval WHERE request_id=<id>",
       "Old main approver added as optional approver with ASKED status\n"
       "New main approver removed from optional list (if was optional)\n"
       "Existing optional approvers unchanged\n"
       "Cascade maintains referential consistency",
       "High", "Functional",
       "EmployeeDayOffServiceImpl.changeApprover", "EmployeeDayOffServiceImpl",
       "Cascade: old main->optional(ASKED), new main removed from optional"),

    tc("TC-DO-060",
       "Auto-sync optional approvers on creation",
       "Employee with default optional approvers configured",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: valid day-off request (no explicit optional approvers)\n"
       "3. Check optional approvers in response\n"
       "4. DB: SELECT * FROM employee_dayoff_approval WHERE request_id=<new_id>",
       "Optional approvers auto-synced from employee's default settings\n"
       "Defaults pulled from employee configuration\n"
       "All synced approvers start with ASKED status\n"
       "Same mechanism as vacation optional approver sync",
       "Medium", "Functional",
       "EmployeeDayOffServiceImpl.create", "EmployeeDayOffServiceImpl",
       "Auto-sync happens during creation, not after"),
]


# ── TS-DO-Permissions ──────────────────────────────────────────────

TS_DO_PERMISSIONS = [
    tc("TC-DO-061",
       "Owner permissions on NEW request — full access",
       "Day-off in NEW status, authenticated as owner",
       "1. GET day-off permissions for owned NEW request\n"
       "2. Check permission set",
       "Owner gets: EDIT (unconditional) + DELETE (canBeCancelled=true for NEW)\n"
       "EDIT: always granted to owner regardless of status\n"
       "DELETE: granted because NEW != APPROVED",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Owner gets EDIT always + DELETE when canBeCancelled"),

    tc("TC-DO-062",
       "Owner permissions on APPROVED request — EDIT but conditional DELETE",
       "APPROVED day-off with personalDate in future, authenticated as owner",
       "1. GET permissions for owned APPROVED request\n"
       "2. Check EDIT and DELETE permissions separately",
       "EDIT: granted (unconditional for owner)\n"
       "DELETE: granted only if personalDate >= report period start\n"
       "canBeCancelled = (status != APPROVED || personalDate >= reportPeriod)",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "EDIT unconditional, DELETE conditional on canBeCancelled"),

    tc("TC-DO-063",
       "Approver permissions on NEW request",
       "Day-off in NEW status, authenticated as assigned approver",
       "1. GET permissions for day-off where current user is approver\n"
       "2. Check permission set",
       "Approver gets: APPROVE + REJECT + EDIT_APPROVER\n"
       "APPROVE: NEW in approvableStatuses {NEW, REJECTED}\n"
       "REJECT: NEW in rejectableStatuses AND personalDate check\n"
       "EDIT_APPROVER: always (unconditional)",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Full approver permissions on NEW request"),

    tc("TC-DO-064",
       "Approver permissions on APPROVED request",
       "APPROVED day-off, authenticated as approver",
       "1. GET permissions for APPROVED day-off\n"
       "2. Check permission set",
       "APPROVE: not granted (APPROVED not in approvableStatuses)\n"
       "REJECT: granted if personalDate >= report period\n"
       "EDIT_APPROVER: granted (unconditional — design issue)",
       "High", "Functional",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "Approver can reject APPROVED but cannot re-approve"),

    tc("TC-DO-065",
       "readOnly user — no permissions",
       "Employee with readOnly=true\n"
       "Test data: SELECT login FROM employee WHERE read_only=true AND status='ACTIVE'",
       "1. Authenticate as readOnly employee\n"
       "2. Attempt any write operation on day-off\n"
       "3. Check permission response",
       "Empty permission set — no operations allowed\n"
       "readOnly flag blocks ALL permissions\n"
       "Non-ROLE_EMPLOYEE users also get empty permissions\n"
       "Can still read/view day-offs (GET endpoints)",
       "High", "Security",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "readOnly users and non-ROLE_EMPLOYEE: zero write permissions"),

    tc("TC-DO-066",
       "Non-ROLE_EMPLOYEE user — no permissions",
       "User without ROLE_EMPLOYEE (e.g. external contractor)",
       "1. Authenticate as non-ROLE_EMPLOYEE user\n"
       "2. Attempt to create day-off\n"
       "3. Check response",
       "No permissions granted\n"
       "Permission service checks ROLE_EMPLOYEE first\n"
       "@PreAuthorize AUTHENTICATED_USER passes but permission service blocks\n"
       "Two-layer access control: auth + permission",
       "Medium", "Security",
       "EmployeeDayOffPermissionService", "EmployeeDayOffPermissionService",
       "ROLE_EMPLOYEE is required for any day-off operation"),

    tc("TC-DO-067",
       "Shared VACATIONS_* authorities — cross-module permission leak",
       "User with VACATIONS_VIEW/VACATIONS_APPROVE but no day-off specific permissions",
       "1. Check @PreAuthorize on day-off controllers — uses VACATIONS_* authorities\n"
       "2. User with vacation-only permissions accesses day-off endpoints\n"
       "3. Verify access is granted or denied",
       "DESIGN ISSUE: Day-off controllers use VACATIONS_* authorities\n"
       "No separate DAYOFF_* permission set\n"
       "User granted vacation permissions automatically gets day-off access\n"
       "Cross-module permission leak — no granular control",
       "High", "Security",
       "DI: Shared VACATIONS_* authorities", "EmployeeDayOffController",
       "All day-off endpoints use VACATIONS_VIEW/CREATE/EDIT/APPROVE/DELETE"),

    tc("TC-DO-068",
       "No CANCEL permission type (unlike vacation)",
       "Compare vacation and day-off permission models",
       "1. Review vacation permission types: includes CANCEL\n"
       "2. Review day-off permission types: no CANCEL, uses DELETE instead\n"
       "3. Test cancellation flow via DELETE endpoint",
       "DESIGN ISSUE: No CANCEL permission type for day-offs\n"
       "Vacation has separate CANCEL permission; day-off uses DELETE\n"
       "Semantic difference: cancel vs delete\n"
       "Inconsistent permission model across modules",
       "Medium", "Functional",
       "DI: No CANCEL permission", "EmployeeDayOffPermissionService",
       "Day-off DELETE = vacation CANCEL. Different semantics, same effect."),

    tc("TC-DO-069",
       "EDIT_APPROVER on terminal status — approver can change approver on DELETED",
       "DELETED day-off, authenticated as current approver",
       "1. PUT /api/vacation/v1/employee-dayOff/change-approver/{id}/{newLogin}\n"
       "2. Verify change succeeds despite DELETED status",
       "DESIGN ISSUE: EDIT_APPROVER is always granted to approver\n"
       "No status check for EDIT_APPROVER permission\n"
       "Approver change succeeds on: DELETED, DELETED_FROM_CALENDAR, REJECTED\n"
       "Functionally meaningless but creates unnecessary DB mutations",
       "High", "Negative",
       "DI: Unconditional EDIT_APPROVER", "EmployeeDayOffPermissionService",
       "Bug confirmed: no status check in EDIT_APPROVER permission grant"),

    tc("TC-DO-070",
       "EDIT on terminal status — owner can PATCH DELETED_FROM_CALENDAR",
       "DELETED_FROM_CALENDAR day-off, authenticated as owner",
       "1. PATCH /api/vacation/v1/employee-dayOff/{id}\n"
       "2. Body: {personalDate: <new date>}\n"
       "3. Verify PATCH succeeds despite terminal status",
       "DESIGN ISSUE: EDIT always granted to owner\n"
       "Can update personalDate on system-deleted request\n"
       "No business value — request was deleted by calendar system\n"
       "Creates data inconsistency",
       "High", "Negative",
       "DI: Unconditional EDIT", "EmployeeDayOffPermissionService",
       "Owner EDIT has zero status restrictions. Terminal states editable."),
]


# ── TS-DO-Validation ───────────────────────────────────────────────

TS_DO_VALIDATION = [
    tc("TC-DO-071",
       "PublicDateValidator — null publicDate returns valid",
       "Active employee",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: publicDate: null, other fields valid\n"
       "3. Check if custom validator accepts null",
       "Null publicDate passes @EmployeeDayOffPublicDateExists validation\n"
       "Custom validator returns true for null (standard pattern)\n"
       "But missing @NotNull means null propagates to service layer\n"
       "Possible NPE when service processes null publicDate",
       "High", "Negative",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Validator null-pass pattern + missing @NotNull = null reaches service"),

    tc("TC-DO-072",
       "PersonalDateValidator — null personalDate returns valid",
       "Active employee",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: personalDate: null, other fields valid\n"
       "3. Check if custom validator accepts null",
       "Null personalDate passes @EmployeeDayOffPersonalDateExists validation\n"
       "Same null-pass pattern as PublicDateValidator\n"
       "Null personalDate propagates to service layer",
       "High", "Negative",
       "EmployeeDayOffPersonalDateExistsValidator", "EmployeeDayOffPatchRequestDTO",
       "Both custom validators return true for null input"),

    tc("TC-DO-073",
       "originalDate — no validation at all",
       "Active employee",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: originalDate: '1900-01-01' (absurd date)\n"
       "3. Check if request created with invalid originalDate",
       "DESIGN ISSUE: originalDate has NO validation annotations\n"
       "No @NotNull, no custom validator, no date range check\n"
       "Any date value accepted including far past/future\n"
       "originalDate is supposed to be the public holiday date but is not validated against calendar",
       "High", "Negative",
       "DI: No originalDate validation", "EmployeeDayOffCreateRequestDTO",
       "publicDate is validated but originalDate (which stores the actual date) is not"),

    tc("TC-DO-074",
       "Frontend validation — zero client-side checks",
       "Access day-off creation form in browser",
       "1. Open day-off creation form in browser\n"
       "2. Inspect form submission code (React components)\n"
       "3. Submit empty form\n"
       "4. Submit with invalid dates",
       "DESIGN ISSUE: ZERO frontend validation\n"
       "No Yup schema, no imperative validator\n"
       "Empty form submission sends to server with no client feedback\n"
       "All validation is 100% server-side\n"
       "UX: user sees no errors until server responds",
       "High", "Functional",
       "DI: Zero frontend validation", "Frontend day-off components",
       "All validation gaps: empty form, past dates, weekend, duplicate dates"),

    tc("TC-DO-075",
       "Frontend — empty form submission",
       "Browser with day-off creation form open",
       "1. Open day-off creation form\n"
       "2. Do not fill any fields\n"
       "3. Click submit\n"
       "4. Observe behavior",
       "No client-side error messages\n"
       "Request sent to server with null/empty fields\n"
       "Server returns 400 (or possibly 500 due to NPE)\n"
       "User experience: delayed error feedback from server",
       "Medium", "Functional",
       "DI: Zero frontend validation", "Frontend components",
       "Frontend sends invalid data, relies entirely on server error handling"),

    tc("TC-DO-076",
       "Frontend — past date selection allowed",
       "Browser with day-off creation form open",
       "1. Select a publicDate in the past\n"
       "2. Submit form\n"
       "3. Observe behavior",
       "Frontend allows past date selection — no date picker restriction\n"
       "Server rejects (date not in future calendar)\n"
       "User sees error only after server response\n"
       "Better UX: disable past dates in date picker",
       "Medium", "Negative",
       "DI: Zero frontend validation", "Frontend components",
       "Date picker does not restrict to future dates"),

    tc("TC-DO-077",
       "Frontend — weekend/non-holiday date selectable",
       "Browser with day-off creation form open",
       "1. Select a regular workday (not a public holiday) as publicDate\n"
       "2. Submit form\n"
       "3. Observe server rejection",
       "Frontend allows selecting any date including non-holidays\n"
       "Server rejects: validation.PublicDateNotFoundInCalendar.message\n"
       "Better UX: only show calendar public holidays in picker",
       "Medium", "Negative",
       "DI: Zero frontend validation", "Frontend components",
       "Calendar data not used to constrain frontend date picker"),

    tc("TC-DO-078",
       "Error code: exception.employee.dayOff.no.permission",
       "User without required permission attempting write operation",
       "1. Attempt approve/reject/delete without proper permission\n"
       "2. Check error response format\n"
       "3. Verify error code and HTTP status",
       "HTTP 403\n"
       "errorCode: exception.employee.dayOff.no.permission\n"
       "EmployeeDayOffSecurityException thrown\n"
       "Error body includes exception class name (information disclosure)",
       "Medium", "Functional",
       "EmployeeDayOffSecurityException", "EmployeeDayOffServiceImpl",
       "Verify consistent error format across all permission failures"),

    tc("TC-DO-079",
       "Error code: exception.day.off.crossing.vacation",
       "Day-off personalDate overlapping existing vacation",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: personalDate within approved vacation date range\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: exception.day.off.crossing.vacation\n"
       "DayOffCrossingVacationException\n"
       "Cross-module validation between day-off and vacation",
       "Medium", "Functional",
       "DayOffCrossingVacationException", "EmployeeDayOffServiceImpl",
       "Vacation-dayoff overlap protection"),

    tc("TC-DO-080",
       "Error code: validation.EmployeeDayOffPublicDateExists.message",
       "Duplicate publicDate for same employee",
       "1. Create day-off for publicDate X\n"
       "2. Create another day-off for same publicDate X\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: validation.EmployeeDayOffPublicDateExists.message\n"
       "Duplicate check per employee (not global)\n"
       "Different employees can have requests for same publicDate",
       "Medium", "Functional",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Per-employee uniqueness, not system-wide"),

    tc("TC-DO-081",
       "Error code: validation.PublicDateNotFoundInCalendar.message",
       "Date not in office calendar",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: publicDate = date not in any calendar\n"
       "3. Check error response",
       "HTTP 400\n"
       "errorCode: validation.PublicDateNotFoundInCalendar.message\n"
       "Validator checks: office calendar OR employee's day-off calendar\n"
       "Both must fail for this error",
       "Medium", "Functional",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Two-source calendar lookup: office calendar + employee dayoff calendar"),

    tc("TC-DO-082",
       "CANCELED status — dead code in enum",
       "Review EmployeeDayOffStatusType enum",
       "1. Search codebase for CANCELED usage in day-off context\n"
       "2. Attempt to set status to CANCELED via any means\n"
       "3. Check if any code path assigns CANCELED",
       "DESIGN ISSUE: CANCELED exists in enum but is never assigned\n"
       "No code path sets status to CANCELED\n"
       "Dead code — status value unreachable\n"
       "DB may have zero records with CANCELED status",
       "Low", "Functional",
       "DI: CANCELED dead code", "EmployeeDayOffStatusType",
       "Enum value exists, never used. Verify with DB: "
       "SELECT COUNT(*) FROM employee_dayoff_request WHERE status='CANCELED'"),

    tc("TC-DO-083",
       "Duplicate setStatus call in entity conversion",
       "Create or update day-off request and monitor conversion",
       "1. Trace code path through entity conversion\n"
       "2. Check if setStatus called twice during conversion\n"
       "3. Verify final status is correct despite duplicate call",
       "DESIGN ISSUE: Duplicate setStatus call in conversion code\n"
       "Status set once by business logic, then overwritten by converter\n"
       "Net effect: usually no bug (second call wins with same value)\n"
       "Risk: if values differ, converter silently overrides business logic",
       "Low", "Functional",
       "DI: Duplicate setStatus", "Entity conversion",
       "Cosmetic issue unless converter and service disagree on status"),

    tc("TC-DO-084",
       "Misleading field name: last_approved_date",
       "Examine day-off request data in DB and API response",
       "1. GET /api/vacation/v1/employee-dayOff/{id}\n"
       "2. DB: SELECT last_approved_date, original_date FROM employee_dayoff_request\n"
       "3. Compare values",
       "DESIGN ISSUE: last_approved_date actually stores the public holiday date\n"
       "Name suggests last approval timestamp\n"
       "Actually equals originalDate (the public holiday)\n"
       "Confusing for developers and testers",
       "Low", "Functional",
       "DI: Misleading field name", "employee_dayoff_request table",
       "Field name does not reflect actual content. Documentation needed."),

    tc("TC-DO-085",
       "Null return from findApprovedByEmployeeAndDate",
       "Query for employee+date combination with no approved day-off",
       "1. Call service method that uses findApprovedByEmployeeAndDate\n"
       "2. Pass employee/date with no matching APPROVED record\n"
       "3. Check return value handling",
       "DESIGN ISSUE: Method returns null instead of empty Optional/list\n"
       "Callers must null-check\n"
       "Missing null check = NPE\n"
       "Should return Optional.empty() or Collections.emptyList()",
       "Medium", "Negative",
       "DI: Null return NPE risk", "EmployeeDayOffRepository",
       "Repository returns null for no results. Service layer may NPE."),

    tc("TC-DO-086",
       "Hardcoded production URL in notifications",
       "Trigger notification for day-off event, inspect email content",
       "1. Create/approve day-off to trigger notification\n"
       "2. Check email content via email MCP\n"
       "3. Inspect URLs in notification body",
       "DESIGN ISSUE: Notification contains hardcoded production URL\n"
       "All environments send emails with production URL\n"
       "Testing env emails link to production app\n"
       "User clicking link goes to wrong environment",
       "Low", "Functional",
       "DI: Hardcoded production URL", "Notification templates",
       "Notifications have production URL regardless of environment"),

    tc("TC-DO-087",
       "Random.nextLong() for synthetic IDs in findSoonDayOffs",
       "Call endpoint that uses findSoonDayOffs (upcoming day-offs list)",
       "1. GET endpoint that returns upcoming day-offs\n"
       "2. Call multiple times\n"
       "3. Check if synthetic IDs change between calls",
       "DESIGN ISSUE: Random.nextLong() generates synthetic IDs\n"
       "IDs are non-deterministic across calls\n"
       "Cannot use ID for caching, deduplication, or reference\n"
       "Risk: negative IDs possible (Random.nextLong range includes negatives)",
       "Low", "Functional",
       "DI: Random.nextLong for IDs", "findSoonDayOffs",
       "Non-deterministic IDs break caching and client-side deduplication"),

    tc("TC-DO-088",
       "Search type MY — own requests with calendar and ledger merge",
       "Authenticated employee with day-off requests",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=MY\n"
       "2. Verify response includes own requests\n"
       "3. Check calendar and ledger data merged into response",
       "Returns employee's own day-off requests\n"
       "Merged with calendar day data and ledger entries\n"
       "Includes all statuses (NEW, APPROVED, REJECTED, etc.)\n"
       "Most common search type for regular employees",
       "High", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "MY search is the primary user-facing view"),

    tc("TC-DO-089",
       "Search type ALL — admin view",
       "Authenticated as admin/HR with VACATIONS_VIEW permission",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=ALL\n"
       "2. Verify response includes all employees' requests\n"
       "3. Check pagination",
       "Returns all day-off requests across all employees\n"
       "Admin-level view\n"
       "Paginated response\n"
       "Requires VACATIONS_VIEW permission",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Admin search for cross-organization visibility"),

    tc("TC-DO-090",
       "Search type APPROVER — requests pending approval",
       "Authenticated as manager with day-offs pending their approval",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=APPROVER\n"
       "2. Verify response includes requests where current user is approver",
       "Returns day-off requests where current user is assigned approver\n"
       "Typically filtered to actionable statuses (NEW, APPROVED for reject)\n"
       "Used by approver dashboard",
       "High", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Approver view for managing pending requests"),

    tc("TC-DO-091",
       "Search type OPTIONAL_APPROVER",
       "Authenticated as optional approver for some day-offs",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=OPTIONAL_APPROVER\n"
       "2. Verify response includes requests where user is optional approver",
       "Returns requests where user is an optional (FYI) approver\n"
       "Includes all approval statuses (ASKED, APPROVED, REJECTED)\n"
       "Used for informational review",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Optional approver view — informational, not action-required"),

    tc("TC-DO-092",
       "Search type DELEGATED_TO_ME",
       "Authenticated as delegate for another manager",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=DELEGATED_TO_ME\n"
       "2. Verify response includes delegated requests",
       "Returns day-off requests delegated to current user\n"
       "Delegation: acting on behalf of absent manager\n"
       "Includes requests where delegation chain leads to current user",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Delegation view for substitute approvers"),

    tc("TC-DO-093",
       "GET list endpoint — approved day-offs only",
       "Authenticated employee",
       "1. GET /api/vacation/v1/employee-dayOff/list\n"
       "2. Check response contains only APPROVED requests",
       "Returns only APPROVED day-off entries\n"
       "Different from main GET / which returns all statuses\n"
       "Used for display in calendar/summary views\n"
       "@PreAuthorize: AUTHENTICATED_USER || VACATIONS_VIEW",
       "Medium", "Functional",
       "EmployeeDayOffController.findApproved", "EmployeeDayOffServiceImpl",
       "Separate endpoint for approved-only list"),

    tc("TC-DO-094",
       "Approve flow — RecalculateVacationDaysHandler integration",
       "NEW day-off, approve and verify vacation balance impact",
       "1. GET current vacation days balance for employee\n"
       "2. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "3. GET vacation days balance again\n"
       "4. Compare before/after",
       "RecalculateVacationDaysHandler fires after approve\n"
       "Vacation day balance adjusts by diff (+1 credit, -1 debit)\n"
       "Net effect depends on duration values\n"
       "Cross-module: day-off approval affects vacation balance",
       "Critical", "Integration",
       "RecalculateVacationDaysHandler", "Event handlers",
       "Day-off approve -> vacation recalculation. Critical cross-module test."),

    tc("TC-DO-095",
       "Approve flow — UpdateMonthNormHandler integration",
       "NEW day-off with personalDate in different month than originalDate",
       "1. Approve day-off where months differ\n"
       "2. Check monthly norm recalculation for both months\n"
       "3. DB: SELECT * FROM employee_month_norm WHERE employee_id=<id>",
       "UpdateMonthNormHandler recalculates norms for both months:\n"
       "  - Month of lastApprovedDate (credit month)\n"
       "  - Month of personalDate (debit month)\n"
       "Both months' norms adjusted\n"
       "Important when months cross reporting periods",
       "High", "Integration",
       "UpdateMonthNormHandler", "Event handlers",
       "Two-month norm update when credit and debit are in different months"),

    tc("TC-DO-096",
       "Search type MY_DEPARTMENT",
       "Authenticated as department manager",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=MY_DEPARTMENT\n"
       "2. Verify response includes department members' requests",
       "Returns day-off requests from all employees in user's department\n"
       "Department scope — broader than APPROVER (which is direct reports only)\n"
       "Used by department managers for team overview",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Department-level visibility for managers"),

    tc("TC-DO-097",
       "Search type RELATED",
       "Authenticated employee on same project/team as others",
       "1. GET /api/vacation/v1/employee-dayOff?searchType=RELATED\n"
       "2. Verify response includes related employees' requests",
       "Returns day-off requests from related employees\n"
       "Related = same project, same team, or organizational connection\n"
       "Used for coordination and awareness",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffSearchService",
       "Relationship-based search for team coordination"),

    tc("TC-DO-098",
       "Create day-off — originalDate vs publicDate mismatch",
       "Active employee, send different originalDate and publicDate",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: originalDate: '2026-05-01', publicDate: '2026-05-09'\n"
       "3. Check which date is validated and stored",
       "publicDate validated against calendar (custom validator)\n"
       "originalDate has NO validation — stored as-is\n"
       "Mismatch creates inconsistent record\n"
       "originalDate should equal publicDate but nothing enforces this",
       "High", "Negative",
       "DI: No originalDate validation", "EmployeeDayOffCreateRequestDTO",
       "Two date fields that should match but are validated independently"),

    tc("TC-DO-099",
       "Approve — ledger fallback chain for credit duration",
       "NEW day-off where no existing ledger entry and no calendar entry for date",
       "1. Set up: ensure no existing ledger and no calendar match\n"
       "2. Approve the day-off\n"
       "3. Check credit ledger entry duration value",
       "Fallback chain for credit duration:\n"
       "1. Existing ledger entry for employee+date -> use its duration\n"
       "2. Calendar day for date -> use calendar duration\n"
       "3. Reporting norm fallback -> use 8h default\n"
       "Verify correct fallback is used when earlier sources are empty",
       "Medium", "Functional",
       "EmployeeDayOffServiceImpl.changeDayOffDaysAfterApprove", "EmployeeDayOffServiceImpl",
       "Three-tier fallback: ledger -> calendar -> norm(8h)"),

    tc("TC-DO-100",
       "Half-day duration (7h) handling",
       "Day-off for a half-working day (calendar duration=7h)",
       "1. Identify half-working day in calendar (duration=7)\n"
       "2. Create day-off with that date as publicDate\n"
       "3. Approve and check ledger entries",
       "Credit ledger entry: duration=7 (half-day)\n"
       "Debit ledger entry: duration=0 (standard debit marker)\n"
       "Vacation balance impact differs from full-day\n"
       "DB: 27 half-day entries found in timemachine env",
       "Medium", "Boundary",
       "EmployeeDayOffServiceImpl", "EmployeeDayOffServiceImpl",
       "Half-day (7h) vs full-day (8h) credit handling"),

    tc("TC-DO-101",
       "Approve — vacation recalculation diff calculation",
       "Approve day-off and verify exact diff passed to vacation recalculation",
       "1. Approve day-off\n"
       "2. Trace RecalculateVacationDaysHandler parameters\n"
       "3. Verify diff value (+1 for credit, -1 for debit)",
       "Diff calculation: credit creates +1, debit creates -1\n"
       "Net diff determines vacation balance change\n"
       "Correct diff is critical for accurate vacation tracking\n"
       "Incorrect diff leads to vacation balance drift over time",
       "High", "Integration",
       "RecalculateVacationDaysHandler", "Event handlers",
       "Diff accuracy is critical for vacation balance integrity"),

    tc("TC-DO-102",
       "Bulk operations — data patterns from timemachine env",
       "Use timemachine env data for verification",
       "1. DB: SELECT status, COUNT(*) FROM employee_dayoff_request GROUP BY status\n"
       "2. DB: SELECT COUNT(*) FROM employee_dayoff\n"
       "3. DB: SELECT status, COUNT(*) FROM employee_dayoff_approval GROUP BY status\n"
       "4. Compare with expected distributions",
       "Expected distribution (timemachine):\n"
       "  Requests: 89.6% APPROVED, 6.9% DELETED, 2.5% DELETED_FROM_CALENDAR, "
       "0.6% NEW, 0.4% REJECTED\n"
       "  Ledger: 2,853 credit (8h), 2,454 debit (0h), 27 half-day (7h)\n"
       "  Approvals: 62% ASKED, 38% APPROVED, 1 REJECTED\n"
       "  Credit > debit: ~399 unused credits",
       "Medium", "Functional",
       "Data verification", "Database",
       "Baseline data pattern verification against known distribution"),

    tc("TC-DO-103",
       "Pagination on GET all endpoint",
       "Authenticated user with many day-off requests",
       "1. GET /api/vacation/v1/employee-dayOff?page=0&size=10\n"
       "2. GET /api/vacation/v1/employee-dayOff?page=1&size=10\n"
       "3. Verify pagination metadata and content",
       "Paginated response with totalElements, totalPages\n"
       "Page content matches requested size\n"
       "No duplicates across pages\n"
       "Sort order consistent across pages",
       "Medium", "Functional",
       "EmployeeDayOffController.findAll", "EmployeeDayOffController",
       "Standard Spring pagination. Verify no data loss between pages."),

    tc("TC-DO-104",
       "GET day-off by non-existent ID",
       "Non-existent day-off ID",
       "1. GET /api/vacation/v1/employee-dayOff with filter for id=999999999\n"
       "2. Check response",
       "Empty result set or 404\n"
       "No NPE or 500 error\n"
       "Graceful handling of non-existent resources",
       "Low", "Negative",
       "EmployeeDayOffController", "EmployeeDayOffServiceImpl",
       "Basic robustness: non-existent resource handling"),

    tc("TC-DO-105",
       "Approve day-off — notification to optional approvers",
       "Day-off with optional approvers in ASKED status, approve the request",
       "1. Add optional approvers to day-off\n"
       "2. Approve the day-off as main approver\n"
       "3. Check email notifications sent\n"
       "4. Check optional approver statuses remain ASKED",
       "Main request approved\n"
       "Optional approvers notified of approval\n"
       "Optional approver statuses NOT changed (remain ASKED)\n"
       "FYI notification only — optional approvers can still respond",
       "Medium", "Functional",
       "EmployeeDayOffServiceImpl.approve", "Notification service",
       "Optional approvers get notification but their status is independent"),

    tc("TC-DO-106",
       "Create day-off for employee in different office (cross-office calendar)",
       "Employee whose office calendar does NOT have the specified publicDate",
       "1. POST /api/vacation/v1/employee-dayOff\n"
       "2. Body: publicDate from a different office's calendar\n"
       "3. Check error response",
       "HTTP 400: validation.PublicDateNotFoundInCalendar.message\n"
       "Validator checks employee's office calendar specifically\n"
       "Public holidays are office-specific\n"
       "Cross-office dates not interchangeable",
       "Medium", "Negative",
       "EmployeeDayOffPublicDateExistsValidator", "EmployeeDayOffCreateRequestDTO",
       "Calendar validation is office-scoped. Different offices = different holidays."),

    tc("TC-DO-107",
       "Concurrent approve and delete on same request",
       "Day-off in NEW status, two concurrent requests",
       "1. Simultaneously:\n"
       "   a. PUT /api/vacation/v1/employee-dayOff/approve/{id} (as approver)\n"
       "   b. DELETE /api/vacation/v1/employee-dayOff/{id} (as owner)\n"
       "2. Check final state",
       "Race condition: no optimistic locking\n"
       "Scenario 1: approve wins -> APPROVED + ledger, then delete fails (if canBeCancelled)\n"
       "Scenario 2: delete wins -> DELETED, then approve fails (DELETED not approvable)\n"
       "Risk: both succeed -> DELETED with orphaned ledger entries",
       "High", "Integration",
       "DI: No transaction isolation", "EmployeeDayOffServiceImpl",
       "No optimistic locking or version field to prevent concurrent mutations"),

    tc("TC-DO-108",
       "Approve day-off — verify event publishing",
       "NEW day-off, approve and check event chain",
       "1. PUT /api/vacation/v1/employee-dayOff/approve/{id}\n"
       "2. Check event: EmployeeDayOffCreatedEvent (on create)\n"
       "3. Verify handlers: RecalculateVacationDaysHandler, UpdateMonthNormHandler\n"
       "4. Verify downstream effects",
       "Events published:\n"
       "  - Status change triggers downstream handlers\n"
       "  - RecalculateVacationDaysHandler: vacation balance update\n"
       "  - UpdateMonthNormHandler: monthly norm update\n"
       "  - Notification events: email to employee",
       "Medium", "Integration",
       "Event publishing", "EmployeeDayOffServiceImpl",
       "Verify complete event chain from approve through all handlers"),
]


# ── Plan Overview ──────────────────────────────────────────────────

def create_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    # Title
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="Day-Off Module — Test Plan").font = FONT_TITLE

    ws.merge_cells("A2:J2")
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Phase B | Branch: release/2.1"
            ).font = FONT_SMALL

    # Scope section
    r = 4
    ws.cell(row=r, column=1, value="1. Scope & Objectives").font = FONT_SUBTITLE
    r += 1
    scope_text = (
        "Comprehensive test coverage for the Day-Off module of TTT (Time Tracking Tool). "
        "Day-offs compensate employees for working on public holidays — the employee works the holiday "
        "and takes a different day off instead. The module uses a dual-entity pattern: "
        "employee_dayoff_request (workflow/status) and employee_dayoff (ledger/accounting).\n\n"
        "Covers: CRUD operations, approval workflow with ledger mechanics, calendar conflict resolution "
        "(DELETED_FROM_CALENDAR, system rejection, office change auto-delete), optional FYI approvers, "
        "role-based permissions with known design issues (unconditional EDIT/EDIT_APPROVER, shared "
        "VACATIONS_* authorities), and validation gaps (zero frontend validation, missing @NotNull).\n\n"
        "Status model: NEW -> APPROVED/REJECTED/DELETED/DELETED_FROM_CALENDAR (CANCELED is dead code). "
        "16 design issues documented. Test cases include API calls with example payloads and DB queries."
    )
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value=scope_text).font = FONT_BODY
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = 120

    # Environment section
    r += 2
    ws.cell(row=r, column=1, value="2. Environment Requirements").font = FONT_SUBTITLE
    r += 1
    env_items = [
        ("Primary Test Env", "timemachine (ttt-timemachine.noveogroup.com) — clock manipulation available"),
        ("Secondary Test Env", "qa-1 (ttt-qa-1.noveogroup.com) — standard testing"),
        ("Production Baseline", "stage (ttt-stage.noveogroup.com) — comparison only"),
        ("Authentication", "JWT (browser login) + API token (API_SECRET_TOKEN header)"),
        ("Database", "PostgreSQL (ttt schema) — SELECT for verification"),
        ("Test Users", "Multiple roles: employee, DM/CPO, PM, accountant, admin, readOnly"),
        ("Clock Control", "PATCH /api/ttt/test/v1/clock for date-dependent tests (timemachine only)"),
        ("API Base URLs", "/api/vacation/v1/employee-dayOff, /api/vacation/v1/employee-dayOff-approvers"),
    ]
    for label, desc in env_items:
        ws.cell(row=r, column=1, value=label).font = FONT_SECTION
        ws.cell(row=r, column=2, value=desc).font = FONT_BODY
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        r += 1

    # Test suites section
    r += 1
    ws.cell(row=r, column=1, value="3. Test Suites").font = FONT_SUBTITLE
    r += 1

    suites = [
        ("TS-DO-CRUD", "Create, Edit, Delete Flows", len(TS_DO_CRUD),
         "Creation flow, PATCH personalDate, DELETE lifecycle, DTO validation gaps, "
         "upsert overwrite risk, duration boundary values"),
        ("TS-DO-Approval", "Approve, Reject, Change Approver", len(TS_DO_APPROVAL),
         "Approval with ledger creation, rejection, re-approval, approver change cascade, "
         "system rejection, non-atomic transaction risk"),
        ("TS-DO-Calendar", "Calendar Conflicts & Office Change", len(TS_DO_CALENDAR),
         "DELETED_FROM_CALENDAR cascade, physical ledger deletion, office change auto-delete, "
         "PAGE_SIZE=100 limit, CalendarUpdateHasDayOffConflictEvent, race conditions"),
        ("TS-DO-OptApprover", "Optional Approver Lifecycle", len(TS_DO_OPT_APPROVER),
         "FYI approver add/update/delete, 4 create constraints, access validation, "
         "approver change cascade, VACATIONS_VIEW security gaps"),
        ("TS-DO-Permissions", "Permissions & Access Control", len(TS_DO_PERMISSIONS),
         "Owner vs approver permissions, unconditional EDIT/EDIT_APPROVER, readOnly blocking, "
         "shared VACATIONS_* authorities, no CANCEL permission type"),
        ("TS-DO-Validation", "Validation, Errors & Edge Cases", len(TS_DO_VALIDATION),
         "Error codes, null validator pass-through, zero frontend validation, search types, "
         "dead code (CANCELED), data patterns, integration events, concurrency"),
    ]

    headers = ["Suite ID", "Suite Name", "Cases", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_GREEN_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    r += 1

    total_cases = 0
    for suite_id, name, count, desc in suites:
        fill = FILL_ROW_EVEN if (r % 2 == 0) else FILL_ROW_ODD
        ws.cell(row=r, column=1, value=suite_id).font = FONT_LINK_BOLD
        ws.cell(row=r, column=1).hyperlink = f"#'{suite_id}'!A1"
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = THIN_BORDER
        write_row(ws, r, [None, name, count, desc], fill=fill)
        ws.cell(row=r, column=1, value=suite_id)
        ws.cell(row=r, column=1).font = FONT_LINK_BOLD
        ws.cell(row=r, column=1).hyperlink = f"#'{suite_id}'!A1"
        total_cases += count
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="TOTAL").font = FONT_SECTION
    ws.cell(row=r, column=3, value=total_cases).font = FONT_SECTION
    ws.cell(row=r, column=3).alignment = ALIGN_CENTER
    r += 2

    # Key metrics
    ws.cell(row=r, column=1, value="4. Key Metrics").font = FONT_SUBTITLE
    r += 1
    metrics = [
        ("Total Test Cases", str(total_cases)),
        ("Design Issues Covered", "16 (unconditional permissions, missing validation, dead code, "
         "security gaps, non-atomic transactions, PAGE_SIZE limit)"),
        ("Security Issues", "5 (VACATIONS_VIEW for writes, shared authorities, "
         "unconditional EDIT_APPROVER, cross-module permission leak)"),
        ("Boundary Tests", "4 (zero duration, negative duration, half-day 7h, PAGE_SIZE=100)"),
        ("Integration Tests", "6 (vacation recalculation, month norm, calendar events, concurrency)"),
        ("Status Model", "6 values (NEW, APPROVED, REJECTED, DELETED, DELETED_FROM_CALENDAR, "
         "CANCELED=dead code)"),
        ("Error Codes Covered", "4 (no.permission, crossing.vacation, PublicDateExists, "
         "PublicDateNotFoundInCalendar)"),
    ]
    for label, value in metrics:
        ws.cell(row=r, column=1, value=label).font = FONT_SECTION
        ws.cell(row=r, column=2, value=value).font = FONT_BODY
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 80
    for c in "EFGHIJ":
        ws.column_dimensions[c].width = 15


# ── Feature Matrix ─────────────────────────────────────────────────

def create_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="<- Back to Plan").font = FONT_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value="Feature x Test Type Coverage Matrix").font = FONT_SUBTITLE

    headers = ["Feature", "Functional", "Negative", "Boundary", "Security", "Integration", "Total", "Suite Link"]
    r = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    features = [
        ("Create/Edit/Delete",   8,  8, 3, 1, 0, "TS-DO-CRUD"),
        ("Approval Workflow",    8,  4, 0, 1, 3, "TS-DO-Approval"),
        ("Calendar Conflicts",   5,  1, 1, 0, 2, "TS-DO-Calendar"),
        ("Optional Approvers",   5,  4, 0, 3, 3, "TS-DO-OptApprover"),
        ("Permissions & Access",  3,  3, 0, 4, 0, "TS-DO-Permissions"),
        ("Validation & Errors", 14,  8, 2, 0, 4, "TS-DO-Validation"),
    ]

    r += 1
    for feat, func, neg, bnd, sec, intg, suite in features:
        total = func + neg + bnd + sec + intg
        fill = FILL_ROW_EVEN if (r % 2 == 0) else FILL_ROW_ODD
        write_row(ws, r, [feat, func, neg, bnd, sec, intg, total, None], fill=fill)
        ws.cell(row=r, column=8, value=suite).font = FONT_LINK
        ws.cell(row=r, column=8).hyperlink = f"#'{suite}'!A1"
        ws.cell(row=r, column=8).fill = fill
        ws.cell(row=r, column=8).border = THIN_BORDER
        r += 1

    # Totals
    total_func = sum(f[1] for f in features)
    total_neg = sum(f[2] for f in features)
    total_bnd = sum(f[3] for f in features)
    total_sec = sum(f[4] for f in features)
    total_intg = sum(f[5] for f in features)
    grand = total_func + total_neg + total_bnd + total_sec + total_intg
    ws.cell(row=r, column=1, value="TOTAL").font = FONT_SECTION
    for col, val in enumerate([total_func, total_neg, total_bnd, total_sec, total_intg, grand], 2):
        ws.cell(row=r, column=col, value=val).font = FONT_SECTION
        ws.cell(row=r, column=col).alignment = ALIGN_CENTER

    col_widths = [30, 12, 12, 12, 12, 12, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Risk Assessment ────────────────────────────────────────────────

def create_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="<- Back to Plan").font = FONT_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"

    headers = ["Feature", "Risk Description", "Likelihood", "Impact", "Severity",
               "Mitigation / Test Focus"]
    r = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    risks = [
        ("Approval Flow",
         "Non-atomic transaction: approve writes ledger entries and status in separate calls. "
         "Failure between steps creates orphaned ledger records that corrupt vacation balance.",
         "Medium", "High", "Critical",
         "Test interruption scenarios. Monitor ledger vs status consistency. "
         "See TC-DO-036, TC-DO-107"),

        ("Permissions",
         "Unconditional EDIT_APPROVER: approver can change approver on DELETED and "
         "DELETED_FROM_CALENDAR requests. Unnecessary DB mutations on terminal statuses.",
         "High", "Medium", "High",
         "Test approver change on all terminal statuses. Verify no cascade side effects. "
         "See TC-DO-031, TC-DO-032, TC-DO-069"),

        ("Permissions",
         "Unconditional EDIT for owner: can PATCH personalDate on DELETED, REJECTED, "
         "and DELETED_FROM_CALENDAR requests. Creates data inconsistency.",
         "High", "Medium", "High",
         "Test PATCH on every status. Verify ledger not affected by editing terminal states. "
         "See TC-DO-012, TC-DO-013, TC-DO-070"),

        ("Validation",
         "Zero frontend validation: no Yup schema, no imperative validator. 100% server-dependent UX. "
         "Empty form submission produces no client feedback.",
         "High", "Medium", "High",
         "Test all frontend validation gaps: empty form, past dates, weekend selection, "
         "duplicate dates. See TC-DO-074 through TC-DO-077"),

        ("Calendar Conflicts",
         "Calendar removal cascade (deleteDayOffs) physically deletes ledger entries and changes "
         "request status. If concurrent with approval, race condition can leave orphaned data.",
         "Medium", "High", "High",
         "Test concurrent calendar removal + approval. Verify ledger cleanup is complete. "
         "See TC-DO-037, TC-DO-039, TC-DO-045"),

        ("Security",
         "Optional approver controller uses VACATIONS_VIEW for all write operations (POST/PATCH/DELETE). "
         "Any user with view permission can modify optional approvers.",
         "High", "Medium", "High",
         "Test write operations with VIEW-only users. Verify service-level access checks compensate. "
         "See TC-DO-056, TC-DO-057, TC-DO-058"),

        ("Calendar Conflicts",
         "PAGE_SIZE=100 hard limit in AutoDeleteHelper.update. Employees with >100 day-offs per year "
         "get incomplete cleanup on office change. No pagination loop.",
         "Low", "High", "Medium",
         "Create >100 day-offs for one employee, trigger office change, verify incomplete cleanup. "
         "See TC-DO-041"),

        ("Security",
         "Shared VACATIONS_* authorities: day-off controllers use vacation permission constants. "
         "No separate DAYOFF_* permission set. Cross-module permission leak.",
         "High", "Low", "Medium",
         "Test vacation-only user accessing day-off endpoints. Verify unintended access. "
         "See TC-DO-067"),

        ("Validation",
         "CANCELED status in enum but never assigned by any code path. Dead code in status model. "
         "Untested and potentially unreachable.",
         "Low", "Low", "Low",
         "Verify no DB records with CANCELED status. Confirm no code path assigns it. "
         "See TC-DO-082"),

        ("Data Integrity",
         "Null return from findApprovedByEmployeeAndDate instead of empty Optional/list. "
         "Callers without null check will NPE.",
         "Medium", "Medium", "Medium",
         "Test queries returning no results. Verify null handling in callers. "
         "See TC-DO-085"),

        ("Data Integrity",
         "Upsert on creation can silently overwrite existing ledger records if matching "
         "employee+originalDate found. No conflict detection or error raised.",
         "Low", "Medium", "Medium",
         "Test creation with pre-existing ledger data for same date. Verify overwrite behavior. "
         "See TC-DO-020"),

        ("Integration",
         "Hardcoded production URL in notification emails. All environments send emails "
         "linking to production app. Testing env emails misdirect users.",
         "High", "Low", "Low",
         "Check notification email content on test envs. Verify URL matches env. "
         "See TC-DO-086"),
    ]

    r += 1
    for feat, risk, like, impact, sev, mitigation in risks:
        fill_map = {
            "Critical": FILL_RISK_HIGH, "High": FILL_RISK_HIGH,
            "Medium": FILL_RISK_MED, "Low": FILL_RISK_LOW
        }
        fill = fill_map.get(sev, FILL_ROW_ODD)
        write_row(ws, r, [feat, risk, like, impact, sev, mitigation], fill=fill)
        r += 1

    add_autofilter(ws, 3, len(headers))
    col_widths = [18, 60, 12, 12, 12, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main ───────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    # Plan tabs (green)
    create_plan_overview(wb)
    create_feature_matrix(wb)
    create_risk_assessment(wb)

    # TS tabs (blue)
    all_suites = [
        ("TS-DO-CRUD", "Create, Edit, Delete Flows", TS_DO_CRUD),
        ("TS-DO-Approval", "Approve, Reject, Change Approver", TS_DO_APPROVAL),
        ("TS-DO-Calendar", "Calendar Conflicts & Office Change", TS_DO_CALENDAR),
        ("TS-DO-OptApprover", "Optional Approver Lifecycle", TS_DO_OPT_APPROVER),
        ("TS-DO-Permissions", "Permissions & Access Control", TS_DO_PERMISSIONS),
        ("TS-DO-Validation", "Validation, Errors & Edge Cases", TS_DO_VALIDATION),
    ]

    total = 0
    for tab_name, suite_name, cases in all_suites:
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_TS
        count = write_ts_tab(ws, suite_name, cases)
        total += count
        print(f"  {tab_name}: {count} cases")

    outpath = "/home/v/Dev/ttt-expert-v1/expert-system/output/day-off/day-off.xlsx"
    wb.save(outpath)
    print(f"\nSaved: {outpath}")
    print(f"Total: {total} test cases across {len(all_suites)} suites")


if __name__ == "__main__":
    main()
