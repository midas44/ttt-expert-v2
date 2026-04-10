#!/usr/bin/env python3
"""
Admin Module Test Documentation Generator — Phase B
Generates test-docs/admin/admin.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 8 TS- test suite tabs.

Based on vault knowledge: admin-panel-deep-dive.md (code-level analysis, 120+ tickets,
PM Tool sync, calendar CRUD, employee management), admin-panel-pages.md (7 admin pages UI),
admin-projects-deep-exploration.md (All/My Projects tabs, action dialogs),
production-calendar-management.md (events CRUD, SO-calendar mapping),
admin-calendar-form-validation-rules.md (field-level validation),
pm-tool-integration-deep-dive.md (sync architecture, rate limiting, sales filtering),
admin-ticket-findings.md (120+ tickets across 8 searches),
role-permission-matrix.md (admin page access matrix).

8 Suites, ~90 test cases:
  TS-Admin-Projects     — Projects table, search/filter, action dialogs, PM Tool links
  TS-Admin-Calendars    — Calendar CRUD, events CRUD, SO-calendar mapping
  TS-Admin-Employees    — Employee list, search, dismissed, subcontractors
  TS-Admin-Settings     — TTT parameters, API keys, export
  TS-Admin-Account      — User account settings, trackers, API token
  TS-Admin-Permissions  — Role-based access matrix for admin pages
  TS-Admin-PMTool       — PM Tool sync mechanics, field mapping, error handling
  TS-Admin-Regression   — Bug regression tests from 120+ ticket findings
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Constants ---------------------------------------------------------------

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "admin")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "admin.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"

PLAN_OVERVIEW = {
    "title": "Admin Module — Test Plan",
    "scope": "Administration panel: Projects (All/My, PM Tool integration), Production Calendars (calendar CRUD, event CRUD, SO-calendar mapping), Employees (search, dismissed, subcontractors), TTT Parameters, API Keys, Export, User Account (token, trackers, CSV settings), PM Tool synchronization, CompanyStaff sync, role-based access control.",
    "objectives": [
        "Verify all admin panel pages render correctly and display expected data for authorized roles",
        "Test project list search/filter/sort/pagination, action dialogs (info, tracker edit, task templates)",
        "Validate production calendar and event CRUD with boundary values and cross-calendar isolation",
        "Test SO-calendar mapping changes including next-year vs immediate application logic",
        "Verify employee list read-only display, search, dismissed filter, subcontractor tab",
        "Test TTT parameter editing, API key management, export functionality",
        "Verify PM Tool sync mechanics: field mapping, rate limiting, error handling, sales filtering",
        "Validate role-based access: ADM, CACC, PM, DM, TL, ACC, HR, EMP access boundaries",
        "Regression coverage for 30+ known bugs from 120+ mined GitLab tickets",
    ],
    "environments": [
        "QA-1: https://ttt-qa-1.noveogroup.com (primary — current sprint build)",
        "Timemachine: https://ttt-timemachine.noveogroup.com (clock manipulation)",
        "Stage: https://ttt-stage.noveogroup.com (production baseline)",
    ],
    "approach": "UI-first testing for all admin pages. API/hybrid tests only for PM Tool sync and CS sync (no UI representation). All project data is read-only (PM Tool is source of truth); only tracker data and task templates are editable in TTT. Calendar CRUD tested via both UI and API for validation coverage.",
    "dependencies": [
        "PM Tool sync cron (every 15 min) — some tests verify post-sync state",
        "CompanyStaff sync — employee data freshness",
        "Production calendar data — must have Russia/Cyprus/France calendars configured",
        "Multiple user accounts with different roles (ADM, CACC, PM, DM, EMP)",
    ],
}

RISK_ASSESSMENT = [
    ("Cross-calendar event isolation", "Deleting event from one calendar triggers day-off processing for employees on a different calendar (#3221)", "Medium", "Critical", "Critical", "Test calendar event deletion when multiple calendars share the same date — verify only affected calendar's employees are impacted"),
    ("SO-calendar year timing", "Calendar change for SO applied immediately instead of next year only (#3300)", "Medium", "High", "Critical", "Test setting SO calendar change — verify current year unaffected, change applies from next year only"),
    ("PM Tool sync missing employees", "PM Tool references employee IDs not in TTT DB → HTTP 500 (#3384)", "High", "Medium", "High", "Verify sync doesn't crash on unknown employee IDs; check pm_tool_sync_failed_entity table for graceful failure tracking"),
    ("Calendar event duplication", "No atomic uniqueness check → race condition creates duplicates (#2656)", "Low", "High", "High", "Test rapid duplicate event creation; verify (calendar_id, date) uniqueness enforced"),
    ("Project name trailing spaces", "Trailing spaces bypass initial validation, then trim creates duplicate (#3348)", "Medium", "Medium", "High", "Test project name with trailing/leading spaces via Edit Tracker Data and API"),
    ("Project deletion FK constraint", "Deleting project with task templates returns 500 (#2098)", "Medium", "Medium", "High", "Test deletion of project that has associated task templates"),
    ("PM Tool rate limiting", "Sync of 3000+ projects at 50 RPM causes long-running process (#3399/#3401)", "Medium", "Low", "Medium", "Verify rate limiter configuration; monitor sync duration for large project count"),
    ("Calendar audit field loss", "created_at/created_by become NULL after editing event (#2648)", "Medium", "Medium", "Medium", "Verify audit fields preserved after calendar event edit"),
    ("CS sync NULL office_id", "CS sends employee with NULL office_id → 500 errors (#3236)", "Medium", "Medium", "Medium", "Verify sync handles NULL office gracefully without crashing"),
    ("Archived SO visibility", "Archived SOs from CS not distinguished in TTT admin (#3323)", "Medium", "Low", "Medium", "Check SO list for archived indicators; verify hiding behavior in calendar and accounting"),
]

# Feature Matrix: [feature, Projects, Calendars, Employees, Settings, Account, Perms, PMTool, Regression, Total]
FEATURE_MATRIX = [
    ["Project list & search",        4, 0, 0, 0, 0, 0, 0, 0, 4],
    ["Project action dialogs",       5, 0, 0, 0, 0, 0, 0, 0, 5],
    ["My Projects tab",              2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["PM Tool links & read-only",    2, 0, 0, 0, 0, 0, 0, 0, 2],
    ["Calendar CRUD",                0, 4, 0, 0, 0, 0, 0, 0, 4],
    ["Calendar events CRUD",         0, 5, 0, 0, 0, 0, 0, 0, 5],
    ["SO-calendar mapping",          0, 3, 0, 0, 0, 0, 0, 0, 3],
    ["Employee list & search",       0, 0, 5, 0, 0, 0, 0, 0, 5],
    ["Subcontractors",               0, 0, 2, 0, 0, 0, 0, 0, 2],
    ["Dismissed/fired display",      0, 0, 2, 0, 0, 0, 0, 0, 2],
    ["TTT parameters",               0, 0, 0, 3, 0, 0, 0, 0, 3],
    ["API keys",                     0, 0, 0, 3, 0, 0, 0, 0, 3],
    ["Export (hours by customer)",   0, 0, 0, 2, 0, 0, 0, 0, 2],
    ["Account settings",             0, 0, 0, 0, 4, 0, 0, 0, 4],
    ["Tracker configuration",        0, 0, 0, 0, 3, 0, 0, 0, 3],
    ["Role-based page access",       0, 0, 0, 0, 0, 8, 0, 0, 8],
    ["Permission boundaries",        0, 0, 0, 0, 0, 2, 0, 0, 2],
    ["PM Tool sync mechanics",       0, 0, 0, 0, 0, 0, 5, 0, 5],
    ["PM Tool field mapping",        0, 0, 0, 0, 0, 0, 3, 0, 3],
    ["PM Tool error handling",       0, 0, 0, 0, 0, 0, 2, 0, 2],
    ["Calendar regression bugs",     0, 0, 0, 0, 0, 0, 0, 5, 5],
    ["Project regression bugs",      0, 0, 0, 0, 0, 0, 0, 4, 4],
    ["Sync regression bugs",         0, 0, 0, 0, 0, 0, 0, 3, 3],
    ["Employee regression bugs",     0, 0, 0, 0, 0, 0, 0, 2, 2],
]


# --- Test Case Data ----------------------------------------------------------

def get_projects_cases():
    """TS-Admin-Projects: Projects table, search/filter, action dialogs, PM Tool links."""
    return [
        {
            "id": "TC-ADM-001", "title": "All Projects tab — table loads with correct columns",
            "preconditions": "User with ADMIN role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ADMIN' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN user\n2. Navigate to Admin panel > Projects\n3. Verify 'All projects' tab is active by default\n4. Verify table columns: Name, Customer, Supervisor, Manager, Type, Status, Actions\n5. Verify pagination is displayed (multiple pages of projects)\n6. Verify 3 action buttons per row (task templates, edit tracker data, project info)\n7. Verify NO 'Create project' button exists anywhere on the page\n8. Verify 'Supervisor' column (not 'Senior Manager' — renamed after PM Tool integration)",
            "expected": "All Projects tab shows correct 7 columns. No create button (PM Tool is source of truth). Pagination works. 3 action buttons per row.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-panel-pages.md, admin-projects-deep-exploration.md", "module": "admin/projects",
            "notes": "PM Tool integration removed project creation from TTT (#3093). Column 'Senior Manager' was renamed to 'Supervisor'."
        },
        {
            "id": "TC-ADM-002", "title": "Project search — by project name and customer",
            "preconditions": "ADMIN user. At least 1 project with known name exists.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects\n3. Click search field (placeholder: 'Project name, customer')\n4. Type a known project name fragment (e.g., 'ABtasty')\n5. Verify scope buttons appear: 'Project' and 'Customer'\n6. Click 'Project' scope button\n7. Verify table filters to show only matching projects\n8. Clear search, type a known customer name\n9. Click 'Customer' scope button\n10. Verify table filters by customer name\n11. Clear search — verify all projects return",
            "expected": "Search filters projects by name or customer. Scope buttons toggle search target. Results update instantly.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Search & Filtering", "module": "admin/projects",
            "notes": "Search has keyboard layout auto-correction (SuggestionMappingUtil.correctLayout)."
        },
        {
            "id": "TC-ADM-003", "title": "Column filters — Supervisor, Manager, Type, Status dropdowns",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Click filter icon on 'Supervisor' column header\n4. Verify dropdown with checkbox list of supervisors appears\n5. Select one supervisor, close dropdown\n6. Verify table filters to that supervisor's projects only\n7. Click filter icon on 'Type' column\n8. Verify 9 type options: Production, Learning, Administration, Commercial, Idle time, Internal, Investment, Investment without invoicing, Project manager\n9. Deselect all, select 'INTERNAL' only\n10. Verify table shows only INTERNAL projects\n11. Click filter icon on 'Status' column\n12. Verify default checked: Active, Unconfirmed, Suspended, Acceptance, Warranty\n13. Verify default unchecked: Finished, Cancelled\n14. Clear all filters — verify full list returns",
            "expected": "Column filters work as checkbox dropdowns. Status defaults exclude Finished/Cancelled. 9 project types, 7 statuses available.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Project Types, §Status Values", "module": "admin/projects",
            "notes": "Status defaults: Active/Unconfirmed/Suspended/Acceptance/Warranty shown; Finished/Cancelled hidden. ~55 supervisors in dropdown."
        },
        {
            "id": "TC-ADM-004", "title": "Show inactive projects toggle",
            "preconditions": "ADMIN user. Projects with Finished/Cancelled status exist.\nQuery: SELECT COUNT(*) FROM ttt_backend.project WHERE status IN ('FINISHED', 'CANCELLED')",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Note current page count (e.g., 11 pages for active projects)\n4. Check 'Show inactive projects' checkbox\n5. Verify page count increases significantly (e.g., ~147 pages with all projects)\n6. Verify Finished and Cancelled projects now appear in the table\n7. Uncheck 'Show inactive projects'\n8. Verify table returns to original active-only count",
            "expected": "Checkbox toggles Finished+Cancelled project visibility. Active-only: ~200 projects (11 pages). All: ~2900+ projects (147 pages).",
            "priority": "High", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Data Scale", "module": "admin/projects",
            "notes": "Scale: ~200 active, ~2900+ total. 20 projects per page."
        },
        {
            "id": "TC-ADM-005", "title": "Project Info dialog — read-only fields and PM Tool link",
            "preconditions": "ADMIN user. A project with a valid pmtId exists.\nQuery: SELECT p.id, p.name, p.pmt_id FROM ttt_backend.project p WHERE p.pmt_id IS NOT NULL AND p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Click the 'More about project' button (info icon, 3rd action button) on a project row\n4. Verify dialog opens with URL hash #<projectId>\n5. Verify ALL fields are read-only: Name, Account name, Customer, Country, Supervisor, Manager, Owner, Watchers, Status, Type, Model, Total cost, First/Last report dates\n6. Verify project name is a clickable link to PM Tool: https://pm.noveogroup.com/projects/{pmtId}/profile/general\n7. Verify Supervisor/Manager/Owner names link to CompanyStaff profiles (cs.noveogroup.com/profile/{login})\n8. Verify 'History of changes' expandable section exists\n9. Expand history — verify chronological changes (date, user, field from→to)\n10. Close dialog",
            "expected": "All project fields read-only. Name links to PM Tool. Employee names link to CS. History shows audit trail.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §More About Project Dialog", "module": "admin/projects",
            "notes": "PM Tool URL pattern: https://pm.noveogroup.com/projects/{pmtId}/profile/general. Falls back to plain text if pmtId is null."
        },
        {
            "id": "TC-ADM-006", "title": "Edit Tracker Data dialog — 3 editable URL fields",
            "preconditions": "ADMIN user with edit permission on a project.\nQuery: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Click 'Edit tracker data' button (pencil icon, 2nd action button) on a project row\n4. Verify dialog opens with exactly 3 editable fields:\n   a. 'Script of synchronization with tracker' (pre-filled with default script URL)\n   b. 'Link to task tracker' (may be empty)\n   c. 'Link to task tracker proxy server' (may be empty, has helper link to Google Docs)\n5. Verify all fields accept URL input with 'https://' prefix\n6. Enter a valid URL in 'Link to task tracker'\n7. Click Save\n8. Verify success notification\n9. Reopen dialog — verify saved URL persists\n10. Clear the entered URL, Save again\n11. Verify field reverts to empty",
            "expected": "Only 3 tracker-related fields editable. All project metadata (name, customer, type, etc.) is NOT in this dialog. URLs saved/cleared correctly.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Edit Tracker Data Dialog, pm-tool-integration-deep-dive.md §Frontend", "module": "admin/projects",
            "notes": "After PM Tool migration (#3093), 'Edit Project' was renamed to 'Edit Tracker Data' with only 3 URL fields. ProxyDescription links to Google Docs setup guide."
        },
        {
            "id": "TC-ADM-007", "title": "Task Templates dialog — CRUD operations",
            "preconditions": "ADMIN user. A project in ACTIVE status.\nQuery: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Click 'Task templates' button (1st action button) on a project row\n4. Verify dialog opens with task template list (may be empty: 'No templates')\n5. Click 'Add a template' button\n6. Verify a new template row appears with project name prefix and editable suffix\n7. Enter template name suffix (e.g., 'Development')\n8. Verify 'Assign the task to employee' toggle is shown (default: off)\n9. Click Save\n10. Verify template count updates (e.g., '1 task template')\n11. Add a second template with different name\n12. Verify count shows '2 task templates'\n13. Delete the first template (trash icon)\n14. Verify count returns to '1 task template'\n15. Click Cancel to close without saving further changes",
            "expected": "Task templates: add/delete/save/cancel. Templates auto-numbered. Format: 'ProjectName / Suffix'. 'Assign to employee' toggle per template.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Task Templates Dialog", "module": "admin/projects",
            "notes": "Task templates define available tasks per project for time reporting. Templates are auto-prefixed with project name."
        },
        {
            "id": "TC-ADM-008", "title": "Project name sorting — ascending and descending",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Click 'Name' column header sort button\n4. Verify projects sort alphabetically A→Z\n5. Verify sort indicator shows ascending arrow\n6. Click 'Name' again\n7. Verify projects sort Z→A (descending)\n8. Click 'Customer' column header\n9. Verify sorting switches to customer column",
            "expected": "Column sorting toggles between ascending/descending. Sort indicator visible.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §All Projects Tab", "module": "admin/projects",
            "notes": ""
        },
        {
            "id": "TC-ADM-009", "title": "Pagination — navigate through project pages",
            "preconditions": "ADMIN user. More than 20 active projects exist.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Verify page 1 is selected, 20 rows visible\n4. Click 'Page 2' button\n5. Verify next 20 projects load\n6. Click 'Next page' arrow\n7. Verify page 3 loads\n8. Click 'Previous page' arrow\n9. Verify returns to page 2\n10. Click last page number\n11. Verify last page loads with <= 20 rows\n12. Verify 'Previous page' is enabled, 'Next page' is disabled on last page",
            "expected": "Pagination: 20 per page. Previous/Next arrows. Direct page selection. Boundary navigation works.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md", "module": "admin/projects",
            "notes": "~11 pages for active projects, ~147 for all."
        },
        {
            "id": "TC-ADM-010", "title": "My Projects tab — shows only current user's managed projects",
            "preconditions": "PM or ADMIN user who manages at least 1 project.\nQuery: SELECT e.login, COUNT(p.id) AS proj_count FROM ttt_backend.employee e JOIN ttt_backend.project p ON e.id = p.manager_id WHERE e.enabled = true AND p.status = 'ACTIVE' GROUP BY e.login HAVING COUNT(p.id) >= 3 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the PM/ADMIN user\n2. Navigate to Admin panel > Projects\n3. Click 'My projects (N)' tab\n4. Verify tab label shows count in parentheses matching projects managed\n5. Verify fewer columns: Name, Manager, Status, Actions (no Customer/Supervisor/Type)\n6. Verify only projects where current user is Manager are shown\n7. Verify project Name column — text should be present (may have PM Tool link or '#' href)\n8. Verify 3 action buttons still available per row\n9. Click 'All projects' tab to switch back\n10. Verify full table returns",
            "expected": "My Projects shows only user-managed projects. Fewer columns (4 vs 7). Tab label shows count.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §My Projects Tab", "module": "admin/projects",
            "notes": "Known bug: My Projects name column has href='#' (non-functional link) — incomplete PM Tool integration."
        },
        {
            "id": "TC-ADM-011", "title": "Project info — pmtId null fallback (plain text name)",
            "preconditions": "ADMIN user. A project without pmtId.\nQuery: SELECT p.id, p.name FROM ttt_backend.project p WHERE p.pmt_id IS NULL AND p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Find a project without PM Tool link (older project, pmt_id is NULL)\n4. Click 'More about project' info button\n5. Verify project name is displayed as plain text (NOT a link)\n6. Verify all other fields still display correctly\n7. Close dialog",
            "expected": "Projects without pmtId show name as plain text, not a link. No error or broken link displayed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "pm-tool-integration-deep-dive.md §Frontend", "module": "admin/projects",
            "notes": "Frontend checks pmtId existence; falls back to plain text. Some legacy projects have no pmtId."
        },
        {
            "id": "TC-ADM-012", "title": "CS profile links — Supervisor/Manager names link to CompanyStaff",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects > All projects\n3. Find a project row with a Supervisor name\n4. Verify Supervisor name is a clickable link\n5. Click Supervisor link\n6. Verify it opens CompanyStaff profile: https://cs.noveogroup.com/profile/{login}\n7. Go back to Projects page\n8. Verify Manager column names also link to CS profiles",
            "expected": "Supervisor and Manager names link to cs.noveogroup.com/profile/{login}.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-projects-deep-exploration.md §Cell Links", "module": "admin/projects",
            "notes": "Links open in same tab or new tab depending on click behavior."
        },
        {
            "id": "TC-ADM-013", "title": "PM user — sees Projects page with limited scope",
            "preconditions": "User with PM role (not ADMIN).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_PROJECT_MANAGER' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name = 'ROLE_ADMIN')) ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM user (not ADMIN)\n2. Navigate to Admin panel > Projects\n3. Verify All Projects tab loads (PM has PROJECTS:VIEW)\n4. Verify all 7 columns visible\n5. Verify action buttons are available\n6. Navigate to Admin panel dropdown\n7. Verify PM can see: Projects, Employees\n8. Verify PM cannot see: TTT parameters, API, Export, Production calendars",
            "expected": "PM has access to Projects and Employees admin pages. No access to Settings/API/Calendar/Export.",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/projects",
            "notes": "PM has PROJECTS:VIEW and EMPLOYEES:VIEW permissions."
        },
    ]


def get_calendars_cases():
    """TS-Admin-Calendars: Calendar CRUD, event CRUD, SO-calendar mapping."""
    return [
        {
            "id": "TC-ADM-014", "title": "Calendar list — view calendars and events for selected year",
            "preconditions": "ADMIN or CHIEF_ACCOUNTANT user.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name IN ('ROLE_ADMIN', 'ROLE_CHIEF_ACCOUNTANT') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN/CACC\n2. Navigate to Admin panel > Production calendars\n3. Verify 'Setting up calendars' tab is active by default\n4. Verify Year picker (default: current year)\n5. Verify Calendar dropdown shows available calendars (Russia, Cyprus, France, etc.)\n6. Select 'Russia' calendar\n7. Verify events table shows columns: Date, Working hours, Reason, Actions\n8. Verify events sorted by date\n9. Verify typical events: New Year holidays (0h), pre-holidays (7h)\n10. Verify 'Add a calendar' and 'Create a new event' buttons visible",
            "expected": "Calendars page loads with year picker, calendar dropdown, and events table. 10+ calendars available. Russia calendar has 18+ events for 2026.",
            "priority": "Critical", "type": "UI",
            "req_ref": "production-calendar-management.md §Tab 1", "module": "admin/calendars",
            "notes": "10 calendars: Armenia, Cyprus, Empty calendar, France, Georgia, Germany, Montenegro, Russia, Uzbekistan, Vietnam."
        },
        {
            "id": "TC-ADM-015", "title": "Create new calendar — name uniqueness and Latin-only",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars > Setting up calendars\n3. Click 'Add a calendar' button\n4. Verify dialog opens with Name field (placeholder: 'Please enter the calendar title in Latin')\n5. Enter a unique Latin name (e.g., 'TestCalendar2026')\n6. Click 'Add'\n7. Verify calendar appears in the dropdown\n8. Try to add another calendar with the SAME name\n9. Verify validation error: duplicate name rejected\n10. Try adding a calendar with Cyrillic characters (e.g., 'Тест')\n11. Verify behavior (may be accepted or rejected based on validation)\nCLEANUP: Via API — DELETE /v2/calendars/{newCalendarId} to remove test calendar",
            "expected": "Calendar created with unique Latin name. Duplicate names rejected (case-insensitive on frontend, case-sensitive on backend — potential gap). Latin characters enforced by placeholder hint.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-calendar-form-validation-rules.md §Calendar CRUD", "module": "admin/calendars",
            "notes": "Backend: CalendarNameExistsValidator. Frontend: AddCalendarValidationSchema.ts uses case-insensitive check. Gap: backend is case-sensitive. #2651"
        },
        {
            "id": "TC-ADM-016", "title": "Delete calendar — only empty calendars allowed",
            "preconditions": "ADMIN user. A test calendar with NO events and NO SO mapped.\nSETUP: Via API — POST /v2/calendars with body {\"name\": \"TempTestCal\"} to create empty calendar",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select the empty test calendar from dropdown\n4. Verify no events listed\n5. Locate delete calendar action (if available in UI)\n6. Delete the empty calendar\n7. Verify calendar removed from dropdown\n8. Now select a calendar that HAS events (e.g., Russia)\n9. Verify delete action is unavailable or returns error if attempted\nCLEANUP: If calendar wasn't deleted via UI, delete via API",
            "expected": "Only empty calendars can be deleted. Calendars with events cannot be deleted (#2653).",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md §Calendar CRUD API", "module": "admin/calendars",
            "notes": "Ticket #2653 confirmed: only empty calendars deletable."
        },
        {
            "id": "TC-ADM-017", "title": "Create calendar event — date, working hours, reason",
            "preconditions": "ADMIN user. A test calendar exists.\nSETUP: Via API — POST /v2/calendars to create 'TestEventCal' if not exists",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select the test calendar\n4. Click 'Create a new event' button\n5. Verify dialog: Date (date picker), Working hours (dropdown: 0, 7, 8), Reason (text field)\n6. Select a future date not already used\n7. Select working hours = 0 (full holiday)\n8. Enter reason: 'Test Holiday'\n9. Click 'Create'\n10. Verify new event appears in the table at correct date position\n11. Verify event shows: selected date, 0h, 'Test Holiday'\n12. Create another event with 7h (pre-holiday) and different date\n13. Verify both events in table\nCLEANUP: Via API — DELETE /v2/days/{eventId} for both created events",
            "expected": "Event created with date, working hours (0/7/8), and reason. Table updates sorted by date.",
            "priority": "Critical", "type": "UI",
            "req_ref": "production-calendar-management.md §Create Event Dialog, admin-calendar-form-validation-rules.md §Calendar Events", "module": "admin/calendars",
            "notes": "Duration: 0 (holiday), 7 (pre-holiday), 8 (transferred working day). Reason is required (min 1 char)."
        },
        {
            "id": "TC-ADM-018", "title": "Calendar event — inline reason editing",
            "preconditions": "ADMIN user. Calendar with at least 1 event with a reason.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select a calendar with events (e.g., Russia)\n4. Click on an event's reason text in the table\n5. Verify text converts to an editable textbox (inline edit)\n6. Change the reason text\n7. Click outside the textbox or press Enter\n8. Verify 'Changes saved' toast/notification appears\n9. Verify the new reason text persists in the table\n10. Verify via API — PATCH /v2/days/{id} only updates 'reason' field",
            "expected": "Inline edit: clicking reason converts to textbox. Save triggers auto-save. Only reason is editable via PATCH (not date/duration).",
            "priority": "High", "type": "UI",
            "req_ref": "production-calendar-management.md §Findings, admin-calendar-form-validation-rules.md", "module": "admin/calendars",
            "notes": "Design issue: PATCH only updates reason field — date and duration are immutable after creation. Auto-save triggers even on no-change (#2204 regression)."
        },
        {
            "id": "TC-ADM-019", "title": "Delete calendar event — only future events",
            "preconditions": "ADMIN user. Calendar with both past and future events.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select Russia calendar for current year\n4. Scroll to find a PAST event (already passed date)\n5. Verify delete button (trash icon) is NOT shown for past events\n6. Scroll to find a FUTURE event\n7. Verify delete button IS shown for future events\n8. Click delete on a future event\n9. Verify confirmation or immediate deletion\n10. Verify event removed from table",
            "expected": "Past events: no delete button. Future events: delete available. Event removed after deletion.",
            "priority": "High", "type": "UI",
            "req_ref": "production-calendar-management.md §Actions, admin-ticket-findings.md #2890", "module": "admin/calendars",
            "notes": "Bug #2890 reported admin COULD delete past events. Verify fix: delete button hidden for past events."
        },
        {
            "id": "TC-ADM-020", "title": "Calendar event — duplicate date validation",
            "preconditions": "ADMIN user. Calendar with existing events.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select Russia calendar\n4. Note an existing event date\n5. Click 'Create a new event'\n6. Select the SAME date as the existing event\n7. Set working hours and reason\n8. Click 'Create'\n9. Verify validation error: duplicate date not allowed\n10. Verify no duplicate record created in the table",
            "expected": "Duplicate date within same calendar rejected with validation error. Backend: @DateUniqueOnCreate validator.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-calendar-form-validation-rules.md §Calendar Events, admin-ticket-findings.md #2656 #3232", "module": "admin/calendars",
            "notes": "Bug #2656: 18 duplicate events created (race condition). Bug #3232: validation message silently lost. Regression test essential."
        },
        {
            "id": "TC-ADM-021", "title": "Year picker — navigate between years",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Verify year picker defaults to current year (2026)\n4. Change year to 2025\n5. Verify events table updates to show 2025 calendar events\n6. Change year to 2027\n7. Verify events table shows 2027 events (may be empty)\n8. Return to 2026\n9. Verify original events restored",
            "expected": "Year picker changes displayed events per year. Different years may have different events.",
            "priority": "Medium", "type": "UI",
            "req_ref": "production-calendar-management.md §Controls", "module": "admin/calendars",
            "notes": "After creating event for next year, should navigate to that year automatically (#3010)."
        },
        {
            "id": "TC-ADM-022", "title": "SO-calendar mapping — view current assignments",
            "preconditions": "ADMIN or CACC user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Click 'Calendars for SO' tab\n4. Verify table columns: Salary office, Calendar, Actions\n5. Verify ~27+ salary offices listed with their calendar assignments\n6. Verify typical mappings: Russia offices → Russia calendar, Cyprus offices → Cyprus calendar\n7. Verify sorting by Salary office column\n8. Check if any SO has no calendar assigned (e.g., 'Ne ukazano')\n9. Verify edit button (pencil icon) per row in Actions column",
            "expected": "SO-calendar table shows all salary offices and assigned calendars. 27+ SOs across 8+ countries.",
            "priority": "High", "type": "UI",
            "req_ref": "production-calendar-management.md §Tab 2", "module": "admin/calendars",
            "notes": "1 SO ('Ne ukazano') has no calendar — potential norm calculation issues for those employees."
        },
        {
            "id": "TC-ADM-023", "title": "SO-calendar change — edit dialog and next-year logic",
            "preconditions": "ADMIN user. A salary office with an assigned calendar.\nQuery: SELECT oc.office_id, o.name AS office_name, c.name AS cal_name, oc.calendar_id FROM ttt_calendar.office_calendar oc JOIN ttt_calendar.office o ON oc.office_id = o.id JOIN ttt_calendar.calendar c ON oc.calendar_id = c.id ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars > Calendars for SO\n3. Click edit button on a salary office row\n4. Verify 'Changing the calendar' dialog opens\n5. Verify salary office name displayed (read-only)\n6. Verify current calendar displayed (read-only)\n7. Verify 'Select a calendar' dropdown with all available calendars\n8. Select a DIFFERENT calendar\n9. Click 'Change'\n10. Verify behavior: if SO had a previous calendar, change applies from NEXT YEAR only\n11. Verify current year calendar unchanged\n12. Restore original calendar assignment\nCLEANUP: Revert SO calendar to original via API — PUT /v2/offices/{officeId}/calendars/{calendarId}",
            "expected": "SO-calendar change dialog: read-only SO name, current calendar, selectable new calendar. If previous calendar exists: change from next year only. If no previous: immediate.",
            "priority": "Critical", "type": "UI",
            "req_ref": "production-calendar-management.md §Edit SO Calendar Dialog, admin-ticket-findings.md #3300", "module": "admin/calendars",
            "notes": "Bug #3300: next-year change applied immediately to ALL years. 16 QA comments. Verify since_year logic: records stored only when calendar changes."
        },
        {
            "id": "TC-ADM-024", "title": "Calendar event duration boundary — min 0, max 12",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select a test calendar\n4. Click 'Create a new event'\n5. Select a future date\n6. Verify working hours dropdown options: 0, 7, 8\n7. Verify duration field constraints via API:\n   DB-CHECK: POST /v2/days with duration=0 → success\n   DB-CHECK: POST /v2/days with duration=12 → success\n   DB-CHECK: POST /v2/days with duration=13 → validation error (@Max(12))\n   DB-CHECK: POST /v2/days with duration=-1 → validation error (@Min(0))\n8. Verify UI dropdown limits options to 0/7/8 only",
            "expected": "UI: 3 options (0, 7, 8). API: accepts 0-12 (backend @Min(0) @Max(12)). API rejects <0 or >12.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-calendar-form-validation-rules.md §Calendar Events", "module": "admin/calendars",
            "notes": "Frontend limits to 0/7/8 dropdown. Backend allows 0-12 range. Gap: API accepts values UI can't create (e.g., 4h)."
        },
        {
            "id": "TC-ADM-025", "title": "Calendar event reason — all-spaces makes uneditable",
            "preconditions": "ADMIN user.\nSETUP: Via API — POST /v2/days with calendarId, date, duration=0, reason=' ' (single space)",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Production calendars\n3. Select the calendar with the space-reason event\n4. Find the event in the table\n5. Verify reason appears empty/blank\n6. Try to click on the reason to inline-edit\n7. Verify the field appears empty but contains whitespace\n8. Try to change the reason text\n9. Verify whether editing is possible or blocked\nCLEANUP: Via API — DELETE /v2/days/{eventId}",
            "expected": "All-spaces reason makes the field appear empty but contain whitespace. Editing behavior may be blocked (#2902 bug).",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2902", "module": "admin/calendars",
            "notes": "Bug #2902: all-spaces reason → uneditable afterward. PATCH only updates reason field."
        },
    ]


def get_employees_cases():
    """TS-Admin-Employees: Employee list, search, dismissed, subcontractors."""
    return [
        {
            "id": "TC-ADM-026", "title": "Employee list — view with search and pagination",
            "preconditions": "User with EMPLOYEES:VIEW permission (PM, DM, TL, ACC, CACC, DIR, HR, ADM).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ADMIN' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the user\n2. Navigate to Admin panel > Employees\n3. Verify 'Employees' tab is active (default)\n4. Verify employee table loads with pagination (20+ pages)\n5. Verify columns display employee names and roles\n6. Verify table is entirely READ-ONLY — no create/edit/delete buttons\n7. Verify search by name field exists\n8. Search for a known employee name\n9. Verify results filter to matching employees\n10. Clear search — verify full list returns",
            "expected": "Employee list is read-only (data synced from CompanyStaff). Search by name works. Pagination with 20+ pages.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-panel-pages.md §Employees Page", "module": "admin/employees",
            "notes": "No CRUD — employees imported from CompanyStaff sync. TTT only patches existing employees (e.g., task carry-over settings)."
        },
        {
            "id": "TC-ADM-027", "title": "Show dismissed employees toggle",
            "preconditions": "ADMIN user. Dismissed employees exist in the system.\nQuery: SELECT COUNT(*) FROM ttt_backend.employee WHERE enabled = false",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Verify 'Show dismissed' checkbox exists (unchecked by default)\n4. Note current employee count / page count\n5. Check 'Show dismissed' checkbox\n6. Verify more employees appear (previously hidden dismissed employees)\n7. Verify dismissed employees can be distinguished from active ones\n8. Uncheck 'Show dismissed'\n9. Verify list returns to active employees only",
            "expected": "Show dismissed toggle reveals fired/dismissed employees. Default: hidden. Active+dismissed > active only.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-panel-pages.md §Employees Page", "module": "admin/employees",
            "notes": "Bug #2702: empty employee names with 'Show fired' enabled."
        },
        {
            "id": "TC-ADM-028", "title": "Subcontractor tab — separate list for contractors",
            "preconditions": "ADMIN user. Subcontractors exist.\nQuery: SELECT COUNT(*) FROM ttt_backend.employee WHERE type = 'CONTRACTOR' AND enabled = true",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Verify 2 tabs: 'Employees' and 'Subcontractor'\n4. Click 'Subcontractor' tab\n5. Verify table loads with subcontractor data (2+ pages)\n6. Verify search functionality works on subcontractor tab\n7. Verify data is read-only\n8. Click back to 'Employees' tab\n9. Verify regular employee list loads",
            "expected": "Subcontractor tab shows contractors separately. 2+ pages of data. Same search/read-only behavior as Employees tab.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-panel-pages.md §Employees Page", "module": "admin/employees",
            "notes": "Contractors separated from regular employees. ~2 pages of subcontractors on timemachine."
        },
        {
            "id": "TC-ADM-029", "title": "Employee sorting — including Russian Ё character collation",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Click column header to sort by name\n4. Verify alphabetical sorting works (A→Z / A→Я in Russian)\n5. Switch to Russian UI (language toggle)\n6. Sort by name in Russian\n7. Check if employees with 'Ё' letter sort correctly (should be after 'Е', not before all)\n8. Switch back to English",
            "expected": "Sorting works alphabetically. Russian 'Ё' may sort incorrectly before all other characters (known collation bug #2515).",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2515", "module": "admin/employees",
            "notes": "Bug #2515: Ё sorts before all others — PostgreSQL collation issue. Bug #2514: sort by Manager causes BadSqlGrammarException."
        },
        {
            "id": "TC-ADM-030", "title": "Employee details — CS profile link",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Find an employee in the list\n4. Click on employee name\n5. Verify it links to CompanyStaff profile: https://cs.noveogroup.com/profile/{login}\n6. Verify no inline editing is available\n7. Verify employee roles are displayed correctly",
            "expected": "Employee names link to CS profiles. All data is read-only. Roles displayed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-panel-pages.md §Employees Page", "module": "admin/employees",
            "notes": ""
        },
        {
            "id": "TC-ADM-031", "title": "HR role — cannot see other offices' employees",
            "preconditions": "User with OFFICE_HR role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_OFFICE_HR' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as OFFICE_HR user\n2. Navigate to Admin panel > Employees\n3. Verify employee list loads successfully (no 403 error)\n4. Search for an employee from a DIFFERENT office\n5. Verify the employee from another office is NOT visible in search results\n6. Search for an employee from HR user's OWN office\n7. Verify own-office employee IS visible\n8. Verify no access to Admin > Projects, Settings, Calendar, API",
            "expected": "HR sees only employees from their own office. No cross-office visibility. 403 bug (#2052) should be fixed.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2167 #2052", "module": "admin/employees",
            "notes": "Bug #2052: HR got 403 on employee list. Bug #2167: HR couldn't find other offices' employees. Verify both fixed."
        },
        {
            "id": "TC-ADM-032", "title": "OFFICE_DIRECTOR — sees all offices' employees",
            "preconditions": "User with OFFICE_DIRECTOR role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_OFFICE_DIRECTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as OFFICE_DIRECTOR\n2. Navigate to Admin panel > Employees\n3. Verify employee list loads\n4. Search for employees from different offices\n5. Verify DIRECTOR sees employees across ALL offices (not scoped to own office)\n6. Verify this matches the known bug behavior (#2050)",
            "expected": "OFFICE_DIRECTOR sees all offices' employees. Bug #2050 reported this as unexpected — verify current behavior and document.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2050", "module": "admin/employees",
            "notes": "Bug #2050: DIRECTOR sees all offices. May be by design or unfixed. Document actual behavior."
        },
        {
            "id": "TC-ADM-033", "title": "Employee with no manager — display and report page link",
            "preconditions": "ADMIN user. An employee with no manager assigned.\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.manager_id IS NULL AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Search for the employee without a manager\n4. Verify the Manager column shows empty/null\n5. Verify no error in the UI display\n6. Check if 'Report page' action button (if exists) works for this employee\n7. Verify no crash or error accessing employee details",
            "expected": "Employee with no manager displays correctly — empty manager field, no UI errors.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md", "module": "admin/employees",
            "notes": "Bug #2195: HR sees 'Report page' button → 403 on click. Verify for employees without manager."
        },
        {
            "id": "TC-ADM-034", "title": "Dismissed employee — empty name display bug",
            "preconditions": "ADMIN user. Dismissed employees exist.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Check 'Show dismissed' checkbox\n4. Scroll through the list looking for employees with empty names\n5. Verify all dismissed employees have proper name display\n6. If empty names found — note the employee login for bug verification",
            "expected": "Dismissed employees should show their names. Bug #2702: empty names with 'Show fired' — verify fixed or still present.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2702", "module": "admin/employees",
            "notes": "Bug #2702 (closed): empty employee names with 'Show fired' enabled."
        },
    ]


def get_settings_cases():
    """TS-Admin-Settings: TTT parameters, API keys, export."""
    return [
        {
            "id": "TC-ADM-035", "title": "TTT Parameters page — view and edit settings",
            "preconditions": "ADMIN user (SETTINGS:VIEW required).",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > TTT parameters (/admin/settings)\n3. Verify page loads with 18 key-value parameters\n4. Verify parameters include: task autocomplete ranges, notification emails, thresholds, extended period duration, CSV export settings\n5. Note the current value of a parameter (e.g., 'over-report threshold')\n6. Click edit on a parameter\n7. Change the value\n8. Save the change\n9. Verify success notification\n10. Verify the new value persists on page reload\n11. Restore original value",
            "expected": "18 parameters displayed and editable by ADMIN. Changes persist. Success notification on save.",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-panel-pages.md §TTT Parameters", "module": "admin/settings",
            "notes": "Parameters: task autocomplete ranges (30/90/180 days), notification emails, thresholds (over: 10%, under: -10%), extended period (60 min), CSV settings."
        },
        {
            "id": "TC-ADM-036", "title": "TTT Parameters — validation for invalid input",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > TTT parameters\n3. Try to edit a numeric parameter with non-numeric value\n4. Verify validation error message\n5. Try to edit a parameter with invalid characters\n6. Verify validation error in English (#3288: English validation message incorrect)\n7. Try to set a parameter to duplicate name (if rename is possible)\n8. Verify duplicate name validation (#2201: generic error instead of field-level)",
            "expected": "Invalid input rejected with appropriate error messages. Validation works for all field types.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3288 #2201", "module": "admin/settings",
            "notes": "Bug #3288: invalid characters → English validation message incorrect. Bug #2201: duplicate name → generic error."
        },
        {
            "id": "TC-ADM-037", "title": "API Keys page — view keys and permissions",
            "preconditions": "ADMIN user (TOKENS:VIEW required).",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > API (/admin/api)\n3. Verify page loads with API key list (~12 keys)\n4. Verify each key shows: Name, Creator, UUID value, Allowed API methods\n5. Verify 22 permission types visible: PROJECTS_ALL, VACATIONS_EDIT, TASKS_EDIT, etc.\n6. Verify the 'Autotest' key is present\n7. Try to copy a key UUID (look for copy button)\n8. Note: bug #2667 reports missing Copy button",
            "expected": "12 API keys displayed with names, creators, UUIDs, and permissions. 22 permission types available.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-panel-pages.md §API Page", "module": "admin/settings",
            "notes": "Bug #2667 (OPEN): missing 'Copy' button for API key UUID."
        },
        {
            "id": "TC-ADM-038", "title": "API Key — duplicate name validation",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > API\n3. Try to create a new API key with an existing name\n4. Verify error handling\n5. Verify error message is informative (not generic 409)\n6. Create a key with a unique name\n7. Verify key appears in the list\nCLEANUP: Delete the test key",
            "expected": "Duplicate name returns 409 conflict. Error message should be user-friendly (bug #897: generic error in UI).",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #897", "module": "admin/settings",
            "notes": "Bug #897 (closed): duplicate name → 409 but generic error in UI."
        },
        {
            "id": "TC-ADM-039", "title": "Export page — download CSV by customer hours",
            "preconditions": "ADMIN user (STATISTICS:EXPORT required).",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Export (/admin/export)\n3. Verify page title: 'Highest number of hours by customer'\n4. Verify date range picker (start date, end date)\n5. Select a date range covering last month\n6. Click 'Download CSV' button\n7. Verify CSV file downloads\n8. Open CSV — verify it contains customer names and hours data\n9. Verify CSV format matches the configured export settings",
            "expected": "CSV downloads with customer hours data for selected date range. Single-purpose export page.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-panel-pages.md §Export Page", "module": "admin/settings",
            "notes": "Only ADMIN has STATISTICS:EXPORT permission."
        },
        {
            "id": "TC-ADM-040", "title": "Non-ADMIN user — cannot access Settings/API/Export pages",
            "preconditions": "User without ADMIN role (e.g., PM or DM).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_PROJECT_MANAGER' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN', 'ROLE_VIEW_ALL'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM user (not ADMIN, not VIEW_ALL)\n2. Navigate to Admin panel dropdown\n3. Verify 'TTT parameters' is NOT in the menu\n4. Verify 'API' is NOT in the menu\n5. Verify 'Export' is NOT in the menu\n6. Try to navigate directly to /admin/settings\n7. Verify access denied or redirect\n8. Try /admin/api — verify denied\n9. Try /admin/export — verify denied",
            "expected": "Non-ADMIN users cannot access Settings, API, or Export admin pages. Menu items hidden and direct URL access denied.",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/settings",
            "notes": "SETTINGS:VIEW, TOKENS:VIEW, STATISTICS:EXPORT — all ADM-only (+ VIEW_ALL for first two)."
        },
        {
            "id": "TC-ADM-041", "title": "Export page — empty date range handling",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Export\n3. Clear both date fields (if possible)\n4. Click Download CSV\n5. Verify behavior — either validation error or empty CSV\n6. Set start date AFTER end date\n7. Click Download CSV\n8. Verify validation prevents inverted date range",
            "expected": "Empty or invalid date range handled gracefully — validation error or empty result.",
            "priority": "Low", "type": "UI",
            "req_ref": "admin-panel-pages.md §Export Page", "module": "admin/settings",
            "notes": ""
        },
        {
            "id": "TC-ADM-042", "title": "Salary Offices page — view offices with periods",
            "preconditions": "ADMIN or ACC/CACC user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Offices (/admin/offices)\n3. Verify salary office list loads\n4. Verify office names display correctly\n5. Verify period information shown per office\n6. Search for a known office name\n7. Verify search filters the list\n8. Check for archived salary offices (#3323) — should be greyed out at bottom",
            "expected": "Salary offices listed with period data. Search works. Archived offices should be visually distinct (#3323).",
            "priority": "High", "type": "UI",
            "req_ref": "admin-panel-deep-dive.md §5, admin-ticket-findings.md #3323", "module": "admin/settings",
            "notes": "Bug #3323 (OPEN): archived SO not hidden or distinguished. Bug #2725: archived status not recognized by TTT."
        },
    ]


def get_account_cases():
    """TS-Admin-Account: User account settings, trackers, API token."""
    return [
        {
            "id": "TC-ADM-043", "title": "Account page — General tab with API token",
            "preconditions": "Any logged-in user.",
            "steps": "1. Login as any user\n2. Navigate to user account page (/admin/account)\n3. Verify 3 tabs: General, Trackers, Export\n4. Verify General tab is active by default\n5. Verify API token section displayed (with token value or generate button)\n6. Verify task carry-over setting option\n7. Verify token is masked or displayed securely",
            "expected": "Account page with 3 tabs. General tab shows API token and task carry-over option.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-panel-pages.md §User Account", "module": "admin/account",
            "notes": "Account page is accessible to ALL users (no permission check)."
        },
        {
            "id": "TC-ADM-044", "title": "Account page — Trackers tab with per-user config",
            "preconditions": "Any logged-in user.",
            "steps": "1. Login as any user\n2. Navigate to /admin/account\n3. Click 'Trackers' tab\n4. Verify tracker configuration options displayed\n5. Verify tracker types available (JIRA, GitLab, Redmine, YouTrack, ClickUp)\n6. Try to add/configure a tracker (if not already configured)\n7. Verify required field validation per tracker type:\n   - JIRA_TOKEN: login + credentials required\n   - JIRA_LOGPASS: login + password required\n   - GITLAB: credentials required\n   - YOU_TRACK: credentials required\n   - REDMINE: credentials required\n8. Save tracker configuration\n9. Verify success notification",
            "expected": "Trackers tab allows per-user tracker configuration. Different tracker types require different credentials. Validation enforced per type.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-calendar-form-validation-rules.md §Admin Tracker Configuration", "module": "admin/account",
            "notes": "TrackerEditValidationSchema: conditional required fields based on tracker type. Only enforces credential changes if value changed from initial."
        },
        {
            "id": "TC-ADM-045", "title": "Account page — Export tab CSV settings",
            "preconditions": "Any logged-in user.",
            "steps": "1. Login as any user\n2. Navigate to /admin/account\n3. Click 'Export' tab\n4. Verify CSV format settings available\n5. Verify decimal separator options (comma vs period)\n6. Verify value separator options\n7. Change a setting\n8. Save\n9. Verify setting persists on page reload",
            "expected": "Export tab provides CSV format customization: decimal separator, value separator. Settings persist.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-panel-pages.md §User Account", "module": "admin/account",
            "notes": "CSV export settings affect the user's personal export format."
        },
        {
            "id": "TC-ADM-046", "title": "Tracker — wrong error when type not selected",
            "preconditions": "Any logged-in user.",
            "steps": "1. Login as any user\n2. Navigate to /admin/account > Trackers\n3. Try to save tracker configuration without selecting a tracker type\n4. Verify error message displayed\n5. Verify error message is meaningful (not cryptic 'tracker.not.permitted')\n6. Verify tracker type dropdown is marked as required",
            "expected": "Clear validation error when tracker type not selected. Not the cryptic error from bug #2448.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2448 #2209", "module": "admin/account",
            "notes": "Bug #2448 (OPEN): wrong error when tracker type not selected. Bug #2209 (closed): cryptic 'tracker.not.permitted' error."
        },
        {
            "id": "TC-ADM-047", "title": "Tracker edit popup — correct title in English",
            "preconditions": "Any user with a configured tracker.",
            "steps": "1. Login as user with configured tracker\n2. Navigate to /admin/account > Trackers\n3. Click edit on an existing tracker\n4. Verify popup title is correct in English\n5. Switch to Russian UI\n6. Verify popup title is correct in Russian\n7. Switch back to English",
            "expected": "Tracker edit popup has correct title in both EN and RU. Bug #3261: wrong title in English.",
            "priority": "Low", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3261", "module": "admin/account",
            "notes": "Bug #3261 (OPEN): wrong title on tracker edit popup in English."
        },
        {
            "id": "TC-ADM-048", "title": "Tracker — no error for inaccessible tracker API",
            "preconditions": "Any user with a tracker configured to an unreachable URL.",
            "steps": "1. Login as user\n2. Navigate to /admin/account > Trackers\n3. Configure a tracker with an invalid/unreachable URL (e.g., https://nonexistent.example.com)\n4. Save the configuration\n5. Verify what happens when the system tries to use this tracker\n6. Verify whether an error message is shown to the user",
            "expected": "System should show a meaningful error when tracker API is inaccessible. Bug #2039: no error shown.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2039", "module": "admin/account",
            "notes": "Bug #2039 (OPEN): no error for inaccessible tracker API."
        },
        {
            "id": "TC-ADM-049", "title": "Account page accessible to ALL roles",
            "preconditions": "User with EMPLOYEE role only (minimal permissions).\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name != 'ROLE_EMPLOYEE')) AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as EMPLOYEE-only user\n2. Navigate to /admin/account\n3. Verify page loads without 403 error\n4. Verify all 3 tabs accessible: General, Trackers, Export\n5. Verify user can modify their own account settings",
            "expected": "Account page accessible to all authenticated users regardless of role. No permission check required.",
            "priority": "Medium", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/account",
            "notes": "Path /admin/account requires no permission. Accessible to all."
        },
    ]


def get_permissions_cases():
    """TS-Admin-Permissions: Role-based access matrix for admin pages."""
    return [
        {
            "id": "TC-ADM-050", "title": "ADMIN — full access to all admin pages",
            "preconditions": "User with ADMIN role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ADMIN' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Open Admin panel dropdown in navigation\n3. Verify all menu items visible: Projects, Employees, TTT parameters, Production calendars, API, Export\n4. Navigate to each page and verify it loads without error:\n   a. /admin/projects — Projects page loads\n   b. /admin/employees — Employees page loads\n   c. /admin/settings — TTT parameters page loads\n   d. /admin/calendar — Production calendars page loads\n   e. /admin/api — API keys page loads\n   f. /admin/export — Export page loads\n   g. /admin/offices — Salary offices page loads\n   h. /admin/account — Account page loads\n5. Verify edit capabilities on settings and calendars",
            "expected": "ADMIN has unrestricted access to all 8 admin pages with full read/write.",
            "priority": "Critical", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": "ADMIN has all VIEW + CREATE permissions."
        },
        {
            "id": "TC-ADM-051", "title": "CHIEF_ACCOUNTANT — calendar + accounting admin access",
            "preconditions": "User with CHIEF_ACCOUNTANT role (not ADMIN).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_CHIEF_ACCOUNTANT' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name = 'ROLE_ADMIN')) ORDER BY random() LIMIT 1",
            "steps": "1. Login as CHIEF_ACCOUNTANT\n2. Open Admin panel dropdown\n3. Verify visible: Employees, Production calendars, Offices\n4. Verify NOT visible: TTT parameters, API, Export\n5. Navigate to /admin/calendar\n6. Verify calendar CRUD is available (create calendar, create event, edit, delete)\n7. Verify admin/salary page accessible for accounting notifications\n8. Navigate to /admin/settings directly\n9. Verify access denied/redirect",
            "expected": "CACC: access to Employees, Calendar (with edit), Offices, Salary. No access to Settings, API, Export.",
            "priority": "Critical", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access, admin-panel-deep-dive.md §Calendar CRUD", "module": "admin/permissions",
            "notes": "Calendar CRUD: hasAnyRole('ADMIN', 'ROLE_CHIEF_ACCOUNTANT'). Note inconsistent role naming."
        },
        {
            "id": "TC-ADM-052", "title": "PROJECT_MANAGER — projects + employees only",
            "preconditions": "User with PM role only (not ADMIN/VIEW_ALL).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_PROJECT_MANAGER' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN', 'ROLE_VIEW_ALL', 'ROLE_CHIEF_ACCOUNTANT'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as PM\n2. Open Admin panel dropdown\n3. Verify visible: Projects, Employees\n4. Verify NOT visible: TTT parameters, Production calendars, API, Export, Offices\n5. Navigate to /admin/projects\n6. Verify project list loads with all action buttons\n7. Verify My Projects tab shows user's managed projects\n8. Navigate to /admin/calendar directly\n9. Verify access denied",
            "expected": "PM: access to Projects (PROJECTS:VIEW) and Employees (EMPLOYEES:VIEW) only.",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": ""
        },
        {
            "id": "TC-ADM-053", "title": "DEPARTMENT_MANAGER — projects + employees access",
            "preconditions": "User with DM role only.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_DEPARTMENT_MANAGER' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN', 'ROLE_VIEW_ALL'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as DM\n2. Open Admin panel dropdown\n3. Verify visible: Projects, Employees\n4. Verify NOT visible: TTT parameters, Production calendars, API, Export\n5. Navigate to /admin/projects\n6. Verify project list loads\n7. Navigate to /admin/employees\n8. Verify employee list loads",
            "expected": "DM: access to Projects and Employees. Same as PM.",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": ""
        },
        {
            "id": "TC-ADM-054", "title": "ACCOUNTANT — employees + offices only",
            "preconditions": "User with ACCOUNTANT role only (not CACC/ADMIN).\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_ACCOUNTANT' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN', 'ROLE_VIEW_ALL', 'ROLE_CHIEF_ACCOUNTANT'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as ACCOUNTANT\n2. Open Admin panel dropdown\n3. Verify visible: Employees, Offices, Salary\n4. Verify NOT visible: Projects, TTT parameters, Production calendars, API, Export\n5. Navigate to /admin/employees — verify loads\n6. Navigate to /admin/offices — verify loads\n7. Navigate to /admin/projects directly — verify denied",
            "expected": "ACCOUNTANT: Employees (EMPLOYEES:VIEW), Offices (OFFICES:VIEW), Salary (ACCOUNTING:NOTIFY). No Projects or Settings.",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": "Bug #2051 (closed): ACCOUNTANT got 403 on employees list. Verify fixed."
        },
        {
            "id": "TC-ADM-055", "title": "TECH_LEAD — employees only in admin",
            "preconditions": "User with TL role only.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_TECH_LEAD' AND e.enabled = true AND e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name IN ('ROLE_ADMIN', 'ROLE_VIEW_ALL', 'ROLE_PROJECT_MANAGER', 'ROLE_DEPARTMENT_MANAGER'))) ORDER BY random() LIMIT 1",
            "steps": "1. Login as TL\n2. Open Admin panel dropdown\n3. Verify visible: Employees only\n4. Verify NOT visible: Projects, TTT parameters, Production calendars, API, Export\n5. Navigate to /admin/employees — verify loads\n6. Navigate to /admin/projects directly — verify denied",
            "expected": "TL: access to Employees only (EMPLOYEES:VIEW). No Projects or other admin pages.",
            "priority": "Medium", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": "TL has EMPLOYEES:VIEW but NOT PROJECTS:VIEW."
        },
        {
            "id": "TC-ADM-056", "title": "EMPLOYEE — no admin panel access except Account",
            "preconditions": "User with EMPLOYEE role only (no other admin roles).\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name NOT IN ('ROLE_EMPLOYEE', 'ROLE_CONTRACTOR'))) AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as EMPLOYEE-only user\n2. Verify navigation bar does NOT show 'Admin panel' dropdown\n3. Navigate directly to /admin/projects\n4. Verify access denied or redirect\n5. Navigate to /admin/employees — verify denied\n6. Navigate to /admin/settings — verify denied\n7. Navigate to /admin/calendar — verify denied\n8. Navigate to /admin/account — verify this ONE page IS accessible\n9. Verify account settings work normally",
            "expected": "EMPLOYEE: no Admin panel in navigation. All /admin/* pages denied EXCEPT /admin/account.",
            "priority": "Critical", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": "Account page has no permission check — accessible to all. All other admin pages require specific permissions."
        },
        {
            "id": "TC-ADM-057", "title": "VIEW_ALL — read access to most admin pages",
            "preconditions": "User with VIEW_ALL role.\nQuery: SELECT e.login FROM ttt_backend.employee e JOIN ttt_backend.employee_role er ON e.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE r.name = 'ROLE_VIEW_ALL' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as VIEW_ALL user\n2. Open Admin panel dropdown\n3. Verify visible: Projects, Employees, TTT parameters, Production calendars, API, Offices\n4. Verify Export is NOT visible (STATISTICS:EXPORT is ADM-only)\n5. Navigate to each visible page — verify loads\n6. Verify calendars page loads but edit operations may be restricted\n7. Navigate to /admin/export directly — verify denied",
            "expected": "VIEW_ALL: access to Projects, Employees, Settings, Calendar, API, Offices. No Export (ADM-only).",
            "priority": "High", "type": "UI",
            "req_ref": "role-permission-matrix.md §Frontend Page Access", "module": "admin/permissions",
            "notes": "VIEW_ALL has most VIEW permissions but NOT STATISTICS:EXPORT."
        },
        {
            "id": "TC-ADM-058", "title": "Calendar page — VIEW_ALL sees buttons but should be read-only",
            "preconditions": "User with VIEW_ALL role (not ADMIN, not CACC).",
            "steps": "1. Login as VIEW_ALL user\n2. Navigate to Admin panel > Production calendars\n3. Check if 'Add a calendar' button is visible\n4. Check if 'Create a new event' button is visible\n5. Check if delete buttons appear on events\n6. If buttons visible: try to create an event\n7. Verify whether action succeeds or fails with permission error\n8. Document actual behavior (bug #2916: VIEW_ALL sees create/edit/delete buttons)",
            "expected": "VIEW_ALL with CALENDARS:VIEW should see calendar data but NOT have create/edit/delete capability. Bug #2916: buttons shown despite read-only intent.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2916", "module": "admin/permissions",
            "notes": "Bug #2916 (closed): VIEW_ALL users see CRUD buttons. Backend uses role check (ADMIN/CACC), not VIEW permission. Verify if fixed."
        },
        {
            "id": "TC-ADM-059", "title": "CONTRACTOR — no admin access",
            "preconditions": "User with CONTRACTOR role.\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.type = 'CONTRACTOR' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as CONTRACTOR\n2. Verify navigation bar does NOT show 'Admin panel' dropdown\n3. Navigate directly to /admin/projects — verify denied\n4. Navigate to /admin/employees — verify denied\n5. Navigate to /admin/account — verify accessible",
            "expected": "CONTRACTOR: no admin panel access except /admin/account. Same as EMPLOYEE.",
            "priority": "Medium", "type": "UI",
            "req_ref": "role-permission-matrix.md §Contractor permissions", "module": "admin/permissions",
            "notes": "ROLE_CONTRACTOR gets no explicit permissions in any provider."
        },
    ]


def get_pmtool_cases():
    """TS-Admin-PMTool: PM Tool sync mechanics, field mapping, error handling."""
    return [
        {
            "id": "TC-ADM-060", "title": "PM Tool sync — manual trigger and status verification",
            "preconditions": "ADMIN user. API secret token available.\nQuery: SELECT COUNT(*) FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL",
            "steps": "SETUP: Note current pm_sync_status last entry timestamp\n1. Via API — POST /v1/test/project/sync (manual sync trigger)\n2. Wait for sync completion (~2-5 minutes depending on project count)\n3. DB-CHECK: SELECT * FROM ttt_backend.pm_sync_status ORDER BY id DESC LIMIT 1 — verify new entry\n4. DB-CHECK: SELECT COUNT(*) FROM ttt_backend.pm_tool_sync_failed_entity — check for failures\n5. Login as ADMIN\n6. Navigate to Admin panel > Projects\n7. Verify project data refreshed (names, supervisors, managers match PM Tool)",
            "expected": "Manual sync creates pm_sync_status entry. Failed entities tracked in pm_tool_sync_failed_entity. Project data updated from PM Tool.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "pm-tool-integration-deep-dive.md §Sync Mechanism", "module": "admin/pmtool",
            "notes": "Cron: every 15 min. Manual: POST /v1/test/project/sync. Rate limit: 50 req/min. Page size: 100."
        },
        {
            "id": "TC-ADM-061", "title": "PM Tool sync — field mapping verification (draft→ACTIVE)",
            "preconditions": "ADMIN user. At least one PM Tool project exists.\nQuery: SELECT p.id, p.name, p.status, p.pm_tool_id, p.pmt_id FROM ttt_backend.project p WHERE p.pm_tool_id IS NOT NULL ORDER BY p.updated_at DESC LIMIT 5",
            "steps": "1. DB-CHECK: Query projects with pm_tool_id to see current mapping\n2. Verify field mappings:\n   - pm_tool_id populated (PM Tool's ttt_id)\n   - pmt_id populated (PM Tool's own PK, used for URL links)\n   - name matches PM Tool\n   - manager_id corresponds to PM Tool's pmId\n   - senior_manager_id corresponds to PM Tool's projectSupervisorId (renamed from supervisor)\n   - owner_id corresponds to PM Tool's ownerId\n   - status: if PM Tool status was 'draft', TTT should show 'ACTIVE'\n3. Login as ADMIN\n4. Open project info dialog for a PM Tool-synced project\n5. Verify PM Tool link: https://pm.noveogroup.com/projects/{pmtId}/profile/general\n6. Verify Supervisor, Manager, Owner fields match DB data",
            "expected": "PM Tool fields correctly mapped to TTT project. Draft status silently converted to ACTIVE. pmtId drives PM Tool URL link.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "pm-tool-integration-deep-dive.md §Project Synchronizer, admin-panel-deep-dive.md §PM Tool", "module": "admin/pmtool",
            "notes": "Draft→ACTIVE is silent conversion. pmtId may be null for Phase 1 (TTT-created) projects."
        },
        {
            "id": "TC-ADM-062", "title": "PM Tool sync — accounting name immutability",
            "preconditions": "ADMIN user.\nQuery: SELECT p.id, p.name, p.accounting_name, p.pm_tool_id FROM ttt_backend.project p WHERE p.pm_tool_id IS NOT NULL AND p.accounting_name IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. DB-CHECK: Query a project with both name and accounting_name set\n2. Note the current accounting_name value\n3. Trigger PM Tool sync (POST /v1/test/project/sync)\n4. Wait for sync completion\n5. DB-CHECK: Query the same project — verify accounting_name is UNCHANGED\n6. Verify project.name may have changed (it gets updated from PM Tool)\n7. But accounting_name stays at its original value (set once on first sync, never overwritten)",
            "expected": "Accounting name immutable after first sync. Project name updates from PM Tool; accounting name does not. (#3083 design)",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #3083, admin-panel-deep-dive.md §Accounting Name Immutability", "module": "admin/pmtool",
            "notes": "By design: accounting_name set once from PM Tool name, never updated. If first sync has wrong data, manual DB fix required."
        },
        {
            "id": "TC-ADM-063", "title": "PM Tool sync — sales employee filtering",
            "preconditions": "ADMIN user.\nQuery: SELECT * FROM ttt_backend.pm_tool_sync_failed_entity ORDER BY id DESC LIMIT 5",
            "steps": "1. DB-CHECK: Check pm_tool_sync_failed_entity for failures related to sales employees\n2. Trigger sync\n3. Monitor sync logs or check failed entities after sync\n4. Verify: employees with type='sales' in PM Tool are filtered out\n5. Verify: sales employees NOT set as manager, owner, or supervisor on any project\n6. DB-CHECK: SELECT p.name, e.login as manager FROM ttt_backend.project p LEFT JOIN ttt_backend.employee e ON p.manager_id = e.id WHERE p.pm_tool_id IS NOT NULL — verify no sales-type employees as managers\n7. Verify: sales-type watchers filtered from observer lists",
            "expected": "Sales employees filtered by removeSalesFromProject(). No sales-type employees as manager/owner/supervisor/watcher. (#3389)",
            "priority": "High", "type": "Hybrid",
            "req_ref": "pm-tool-integration-deep-dive.md §Entity Reference Type Handling, admin-ticket-findings.md #3389", "module": "admin/pmtool",
            "notes": "CSToolEntityReference.isSales() — case-insensitive check for 'sales' type. Employees and contractors processed normally."
        },
        {
            "id": "TC-ADM-064", "title": "PM Tool sync — missing employee handling (HTTP 500)",
            "preconditions": "ADMIN user.",
            "steps": "1. DB-CHECK: SELECT * FROM ttt_backend.pm_tool_sync_failed_entity — check for existing failures\n2. Verify if any failures are caused by missing employee IDs\n3. Review the error description: should mention IllegalStateException with missing employee CS IDs\n4. Verify the ENTIRE project sync fails for that project (not just the missing field)\n5. Verify other projects sync successfully despite this project's failure\n6. Verify the failed project is tracked for retry in pm_tool_sync_failed_entity",
            "expected": "Missing employee → IllegalStateException → HTTP 500. Project sync fails individually; other projects unaffected. Failed entity tracked for retry. (#3384)",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-panel-deep-dive.md §PM Tool Validation, admin-ticket-findings.md #3384", "module": "admin/pmtool",
            "notes": "Design issue: throws IllegalStateException, not a proper business exception. Causes HTTP 500 instead of meaningful error."
        },
        {
            "id": "TC-ADM-065", "title": "PM Tool sync — presales IDs append-only behavior",
            "preconditions": "ADMIN user. A project with presales IDs.\nQuery: SELECT p.id, p.name, p.pre_sales_ids FROM ttt_backend.project p WHERE p.pre_sales_ids IS NOT NULL AND p.pre_sales_ids != '' AND p.pm_tool_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. DB-CHECK: Note current pre_sales_ids for a PM Tool project\n2. Trigger sync\n3. DB-CHECK: Verify pre_sales_ids after sync — should only grow (append-only)\n4. Verify old presales IDs are NOT removed even if PM Tool no longer has them\n5. Verify format: comma-separated ticket IDs",
            "expected": "Presales IDs are append-only — sync adds new IDs but never removes existing ones. Stale IDs accumulate by design. (#3083)",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #3083", "module": "admin/pmtool",
            "notes": "By design: presales history preservation. Stale IDs accumulate over time."
        },
        {
            "id": "TC-ADM-066", "title": "PM Tool sync — rate limiter at 50 req/min",
            "preconditions": "ADMIN user.",
            "steps": "1. DB-CHECK: SELECT COUNT(*) FROM ttt_backend.project WHERE pm_tool_id IS NOT NULL — count PM Tool projects\n2. Note the time before triggering sync\n3. Trigger sync: POST /v1/test/project/sync\n4. Monitor sync duration\n5. Calculate expected duration: (project_count / 100) * (100 / 50) minutes for page fetches\n6. Verify sync completes without 429 errors\n7. DB-CHECK: Verify pm_tool_sync_failed_entity has no rate-limit failures",
            "expected": "Sync rate-limited at 50 req/min (configurable). No 429 errors. Sync completes within expected duration.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "pm-tool-integration-deep-dive.md §Rate Limiter, admin-ticket-findings.md #3401", "module": "admin/pmtool",
            "notes": "Guava RateLimiter at 50/60 ≈ 0.83 req/sec. acquire() blocks thread. Page size: 100 projects."
        },
        {
            "id": "TC-ADM-067", "title": "PM Tool sync — failed entity retry in batches",
            "preconditions": "ADMIN user. At least one entry in pm_tool_sync_failed_entity.\nQuery: SELECT * FROM ttt_backend.pm_tool_sync_failed_entity ORDER BY id DESC LIMIT 5",
            "steps": "1. DB-CHECK: Query pm_tool_sync_failed_entity for existing failures\n2. Note the failed entity IDs and count\n3. Trigger sync\n4. DB-CHECK: After sync, check if failed entities were retried\n5. Verify: successfully retried entities removed from failed table\n6. Verify: still-failing entities remain in failed table\n7. Verify retry batch size: 10 (configurable via pmTool.sync.retry-batch-size)",
            "expected": "Failed entities retried in batches of 10 after main sync. Successful retries cleared from table. Persistent failures remain.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "pm-tool-integration-deep-dive.md §Failed Entity Handling", "module": "admin/pmtool",
            "notes": "Retry batch size: 10 (default). Timeout per entity: 10s."
        },
        {
            "id": "TC-ADM-068", "title": "CS Sync — new salary office not created (#3241)",
            "preconditions": "ADMIN user.",
            "steps": "1. DB-CHECK: SELECT COUNT(*) FROM ttt_backend.office — current office count\n2. Trigger CS sync endpoint\n3. DB-CHECK: Verify count after sync — check if new offices appeared\n4. Verify: field UPDATES on existing offices work (e.g., name changes)\n5. Verify: NEW office creation fails or succeeds\n6. If creation fails: document as #3241 still open\n7. Check vacation and calendar services for same office sync behavior",
            "expected": "Existing office updates work. New office creation may fail (bug #3241). Delta sync doesn't guarantee consistency (#2989).",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #3241 #2989", "module": "admin/pmtool",
            "notes": "Bug #3241 (closed): new SO not synced. Bug #2989 (OPEN): full sync needed for consistency. Bug #3303: sync on startup."
        },
        {
            "id": "TC-ADM-069", "title": "CS Sync — NULL office_id handling (#3236)",
            "preconditions": "ADMIN user.",
            "steps": "1. DB-CHECK: SELECT e.login, e.office_id FROM ttt_backend.employee WHERE office_id IS NULL AND enabled = true — check for NULL office\n2. If such employees exist, verify they don't cause 500 errors in admin pages\n3. Navigate to Admin panel > Employees\n4. Search for an employee with NULL office_id\n5. Verify the employee displays correctly (no crash)\n6. Verify Salary page doesn't crash when processing these employees",
            "expected": "Employees with NULL office_id should display without errors. Bug #3236: should skip invalid records and log warnings.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #3236", "module": "admin/pmtool",
            "notes": "Bug #3236 (OPEN): NULL office_id → 500 errors. Need validation to skip invalid records."
        },
    ]


def get_regression_cases():
    """TS-Admin-Regression: Bug regression tests from ticket findings."""
    return [
        {
            "id": "TC-ADM-070", "title": "Regression: Cross-calendar event deletion isolation (#3221)",
            "preconditions": "ADMIN user. Two different calendars (e.g., Georgia and Cyprus) both have an event on the same date.\nSETUP: Via API — create events on same date in 2 different calendars if needed",
            "steps": "1. Login as ADMIN\n2. Navigate to Production calendars\n3. Select Georgia calendar\n4. Note a date that also exists in Cyprus calendar\n5. Delete the event from Georgia calendar\n6. Verify Georgia calendar event removed\n7. Switch to Cyprus calendar\n8. Verify Cyprus calendar event is UNAFFECTED\n9. DB-CHECK: Verify no day-off processing triggered for Cyprus calendar employees\n10. Verify no incorrect email notifications sent to Cyprus employees\nCLEANUP: Restore Georgia calendar event if needed",
            "expected": "Deleting event from one calendar does NOT affect employees on other calendars. No cross-calendar cascade. (#3221 CRITICAL)",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3221", "module": "admin/regression",
            "notes": "CRITICAL bug #3221: CalendarDeletedEvent didn't scope by calendar_id — processed ALL employees with same-date day-offs."
        },
        {
            "id": "TC-ADM-071", "title": "Regression: SO calendar change timing — next year only (#3300)",
            "preconditions": "ADMIN user. A salary office with an existing calendar assignment.\nQuery: SELECT oc.office_id, o.name, c.name AS calendar_name FROM ttt_calendar.office_calendar oc JOIN ttt_calendar.office o ON oc.office_id = o.id JOIN ttt_calendar.calendar c ON oc.calendar_id = c.id LIMIT 1",
            "steps": "1. Login as ADMIN\n2. Navigate to Production calendars > Calendars for SO\n3. Edit a salary office's calendar (change to a different calendar)\n4. Save the change\n5. DB-CHECK: SELECT * FROM ttt_calendar.office_calendar WHERE office_id = <id> ORDER BY since_year\n6. Verify: change recorded with since_year = NEXT YEAR (not current year)\n7. Verify current year's calendar assignment is UNCHANGED\n8. Navigate to Production calendars > Setting up calendars\n9. Verify current year still shows old calendar's events for that SO\nCLEANUP: Restore original calendar assignment",
            "expected": "Calendar change for SO applies from NEXT YEAR only (if previous calendar existed). Current year unchanged. (#3300)",
            "priority": "Critical", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3300", "module": "admin/regression",
            "notes": "Bug #3300: change applied immediately to ALL years. 16 QA comments. Logic: previous calendar=null → immediate; exists → next year."
        },
        {
            "id": "TC-ADM-072", "title": "Regression: Calendar event duplication on rapid creation (#2656)",
            "preconditions": "ADMIN user. A test calendar.\nSETUP: Via API — ensure test calendar exists",
            "steps": "1. Login as ADMIN\n2. Navigate to Production calendars\n3. Select test calendar\n4. Click 'Create a new event' rapidly (try to submit same date twice quickly)\n5. Verify only ONE event created for the date\n6. DB-CHECK: SELECT COUNT(*) FROM ttt_calendar.calendar_days WHERE calendar_id = <id> AND date = '<date>' — should be 1\n7. Try via API: simultaneously POST two events with same calendar_id + date\n8. Verify duplicate rejection\nCLEANUP: Delete test events",
            "expected": "No duplicate events created. @DateUniqueOnCreate validator prevents duplicates. Race condition (#2656: 18 duplicates) should be fixed.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2656", "module": "admin/regression",
            "notes": "Bug #2656: 18 duplicate events. Check + insert NOT atomic. Verify uniqueness constraint added."
        },
        {
            "id": "TC-ADM-073", "title": "Regression: Calendar audit fields preserved after edit (#2648)",
            "preconditions": "ADMIN user. Calendar with events.\nQuery: SELECT cd.id, cd.created_at, cd.created_by, cd.reason FROM ttt_calendar.calendar_days cd WHERE cd.created_at IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. DB-CHECK: Note created_at and created_by for an existing calendar event\n2. Login as ADMIN\n3. Navigate to Production calendars\n4. Find the event and edit its reason (inline edit)\n5. Save the change\n6. DB-CHECK: Query the same event — verify created_at and created_by are PRESERVED (not NULL)\n7. Verify updated_at and updated_by reflect the edit\n8. Verify the reason text changed",
            "expected": "Editing a calendar event preserves created_at/created_by fields. Only updated_at/updated_by change. Bug #2648 fix verified.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2648", "module": "admin/regression",
            "notes": "Bug #2648 (closed): created_at/created_by became NULL after editing event."
        },
        {
            "id": "TC-ADM-074", "title": "Regression: Project trailing spaces cause duplicate name (#3348)",
            "preconditions": "ADMIN user. Two projects with similar names.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects\n3. Via API: PATCH /v1/projects/{id} with name = 'ExistingProject  ' (trailing spaces)\n4. Verify response — should either reject with validation error or trim and check uniqueness\n5. If PATCH succeeds: verify the name in DB is trimmed\n6. If trimmed name matches existing project: verify AlreadyExistsException returned\n7. Login and verify UI shows the project correctly",
            "expected": "Trailing spaces should be trimmed. Trimmed name checked for uniqueness. No silent duplicate creation. (#3348 OPEN)",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #3348", "module": "admin/regression",
            "notes": "Bug #3348 (OPEN): trailing spaces → silent failure. AlreadyExistsException handling in PATCH returns conflicting project."
        },
        {
            "id": "TC-ADM-075", "title": "Regression: Project deletion FK constraint on task templates (#2098)",
            "preconditions": "ADMIN user. A project with associated task templates.\nQuery: SELECT p.id, p.name, COUNT(tt.id) AS template_count FROM ttt_backend.project p JOIN ttt_backend.task_template tt ON p.id = tt.project_id GROUP BY p.id, p.name HAVING COUNT(tt.id) > 0 ORDER BY random() LIMIT 1",
            "steps": "1. DB-CHECK: Identify a project with task templates\n2. Login as ADMIN\n3. Navigate to Admin panel > Projects\n4. Try to delete the project (via API: DELETE /v1/projects/{id})\n5. Verify error: should fail with FK constraint (task_template_project_fkey)\n6. Verify error is handled gracefully (not raw 500 with SQL error)\n7. Verify the project is NOT deleted",
            "expected": "Project with task templates cannot be deleted. FK constraint prevents deletion. Should return meaningful error, not 500. (#2098)",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2098", "module": "admin/regression",
            "notes": "Bug #2098 (closed): 500 on deletion with FK violation. No pre-check for dependent records."
        },
        {
            "id": "TC-ADM-076", "title": "Regression: Project creation without required fields via API (#2053)",
            "preconditions": "ADMIN user.",
            "steps": "1. Via API: POST /v1/projects with minimal body (missing seniorManagerLogin and country)\n2. Verify response — should either reject or accept\n3. If accepted: verify project created in DB without these fields\n4. Login as ADMIN\n5. Navigate to Projects and find the created project\n6. Verify how it displays without supervisor and country\n7. Verify this is a security gap (UI enforces these but API doesn't)\nCLEANUP: Delete test project if created",
            "expected": "API should enforce required fields (seniorManagerLogin, country). Bug #2053: API allows creation without these fields. Security gap.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2053", "module": "admin/regression",
            "notes": "Bug #2053 (OPEN): API bypasses frontend validation for required fields."
        },
        {
            "id": "TC-ADM-077", "title": "Regression: Firefox/Safari first calendar event creation (#2791)",
            "preconditions": "ADMIN user. Firefox or Safari browser.",
            "steps": "1. Open Firefox (or Safari) browser\n2. Login as ADMIN\n3. Navigate to Admin panel > Production calendars\n4. Select a test calendar\n5. Click 'Create a new event'\n6. Select date, working hours, reason\n7. Click 'Create'\n8. Verify event is created successfully\n9. If fails: check if 'Invalid date' was sent instead of ISO date\n10. Verify this is fixed (bug #2791)",
            "expected": "Calendar event creation works on Firefox/Safari. Bug #2791: first event sent 'Invalid date' on Firefox/Safari.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2791", "module": "admin/regression",
            "notes": "Bug #2791 (closed): Firefox/Safari date parsing issue. Verify fix."
        },
        {
            "id": "TC-ADM-078", "title": "Regression: Duplicate project name returns generic error (#2674/#2125)",
            "preconditions": "ADMIN user.\nQuery: SELECT p.name FROM ttt_backend.project p WHERE p.status = 'ACTIVE' ORDER BY random() LIMIT 1",
            "steps": "1. Note an existing project name\n2. Via API: PATCH /v1/projects/{anotherProjectId} with the duplicate name\n3. Verify response code — should be 400 or 409 (not 500)\n4. Verify error message is informative: 'Project already exists' or similar\n5. Login as ADMIN\n6. Navigate to Projects > Edit Tracker Data for any project\n7. If project name editing were possible, test duplicate name behavior\n8. Document actual error message returned",
            "expected": "Duplicate project name returns 400/409 with meaningful message, not 500. Bug #2125: PATCH returns 500 for duplicate. Bug #2674: generic error.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2674 #2125", "module": "admin/regression",
            "notes": "Bugs #2674, #2125 (closed): generic/500 errors on duplicate name."
        },
        {
            "id": "TC-ADM-079", "title": "Regression: Employee sort by Manager causes SQL error (#2514)",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Employees\n3. Click to sort by Manager column (if available)\n4. Verify sorting works without error\n5. If sort fails: verify it's the BadSqlGrammarException from jOOQ (#2514)\n6. Try sorting by other columns to verify they work\n7. Document actual behavior",
            "expected": "Sorting by Manager should work without SQL error. Bug #2514: BadSqlGrammarException in jOOQ. Verify fix.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2514", "module": "admin/regression",
            "notes": "Bug #2514 (closed): sort by Manager → BadSqlGrammarException."
        },
        {
            "id": "TC-ADM-080", "title": "Regression: Contradictory role data between endpoints (#2063)",
            "preconditions": "ADMIN user.",
            "steps": "1. Trigger CS sync to update employee data\n2. Login as ADMIN\n3. Navigate to Admin panel > Employees\n4. Select an employee and check their roles\n5. Via API: GET /v1/employees/{login}/roles\n6. Compare roles from UI and API\n7. Verify they match\n8. If mismatch: cache not refreshed after sync (#2063)",
            "expected": "Employee roles consistent between UI and API after sync. No contradictory data from stale cache.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2063", "module": "admin/regression",
            "notes": "Bug #2063 (closed): contradictory role data because cache not refreshed after CS sync."
        },
        {
            "id": "TC-ADM-081", "title": "Regression: Accounting period — report month before confirmation (#3365)",
            "preconditions": "ADMIN or ACCOUNTANT user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Offices or Accounting > Periods\n3. Try to select a report period month that is BEFORE the confirmation period\n4. Verify the month is disabled (should not be selectable)\n5. If selectable: try to save\n6. Verify 400 error if saved (frontend validation gap)\n7. Document the actual behavior",
            "expected": "Report period month before confirmation period should be disabled. Bug #3365: not disabled, saving causes 400.",
            "priority": "High", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3365", "module": "admin/regression",
            "notes": "Bug #3365 (Sprint 15): frontend validation gap — disabled month not enforced."
        },
        {
            "id": "TC-ADM-082", "title": "Regression: ClickUp tracker integration broken (#3148)",
            "preconditions": "User with a ClickUp tracker configured.",
            "steps": "1. Login as user with ClickUp tracker\n2. Navigate to /admin/account > Trackers\n3. Verify ClickUp tracker configuration displays\n4. Try to sync tasks from ClickUp\n5. Verify sync works or displays meaningful error\n6. If specific ClickUp space not adding tasks (#3341): test with that space",
            "expected": "ClickUp integration works. Bug #3148 (closed): completely broken. Bug #3341 (OPEN): specific space tasks not adding.",
            "priority": "Medium", "type": "UI",
            "req_ref": "admin-ticket-findings.md #3148 #3341", "module": "admin/regression",
            "notes": "ClickUp bugs: #3148 (closed — complete breakage), #3341 (OPEN — specific space)."
        },
        {
            "id": "TC-ADM-083", "title": "Regression: PM/SPM role checkboxes disabled after admin login (#2188)",
            "preconditions": "ADMIN user.",
            "steps": "1. Login as ADMIN\n2. Navigate to Admin panel > Projects\n3. Open a project info dialog\n4. Check if any PM/SPM role checkboxes are visible\n5. Verify they are enabled (not disabled)\n6. If disabled: regression of #2188 — checkboxes disabled after admin login\n7. Document actual behavior",
            "expected": "PM/SPM role checkboxes should be enabled for ADMIN. Bug #2188: checkboxes disabled after admin login.",
            "priority": "Low", "type": "UI",
            "req_ref": "admin-ticket-findings.md #2188", "module": "admin/regression",
            "notes": "Bug #2188 (closed): PM/SPM role checkboxes disabled. Verify fix."
        },
        {
            "id": "TC-ADM-084", "title": "Regression: GET /projects/{id}/events accessible to ALL users (#2181)",
            "preconditions": "EMPLOYEE-only user.\nQuery: SELECT e.login FROM ttt_backend.employee e WHERE e.id NOT IN (SELECT employee_id FROM ttt_backend.employee_role WHERE role_id IN (SELECT id FROM ttt_backend.role WHERE name != 'ROLE_EMPLOYEE')) AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as EMPLOYEE-only user\n2. Via API: GET /v1/projects/{anyProjectId}/events\n3. Verify response — should this be restricted?\n4. If accessible: security gap — all users see project events\n5. Verify with ADMIN: same endpoint returns project change history",
            "expected": "Project events endpoint should be restricted. Bug #2181: accessible to ALL users regardless of permissions.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "admin-ticket-findings.md #2181", "module": "admin/regression",
            "notes": "Bug #2181 (closed): no permission check on project events endpoint."
        },
    ]


# --- Workbook Helpers --------------------------------------------------------

def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 1 else FILL_ROW_WHITE
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER


def write_suite_tab(ws, cases, tab_name, back_link_tab="Plan Overview"):
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [14, 40, 40, 55, 40, 10, 10, 35, 20, 35]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = f"#'{back_link_tab}'!A1"

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    for row_idx, tc in enumerate(cases, 3):
        values = [tc["id"], tc["title"], tc["preconditions"], tc["steps"],
                  tc["expected"], tc["priority"], tc["type"],
                  tc["req_ref"], tc["module"], tc["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):
                cell.alignment = ALIGN_CENTER

    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"• {env}").font = FONT_BODY
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"• {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws, suite_tabs):
    headers = ["Feature", "Projects", "Calendar", "Employees", "Settings",
               "Account", "Perms", "PMTool", "Regress", "Total"]
    col_widths = [30, 10, 10, 10, 10, 10, 10, 10, 10, 8]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


# --- Main Generator ----------------------------------------------------------

def generate():
    wb = Workbook()

    suites = [
        ("TS-Admin-Projects", "TS-Admin-Projects", get_projects_cases),
        ("TS-Admin-Calendars", "TS-Admin-Calendars", get_calendars_cases),
        ("TS-Admin-Employees", "TS-Admin-Employees", get_employees_cases),
        ("TS-Admin-Settings", "TS-Admin-Settings", get_settings_cases),
        ("TS-Admin-Account", "TS-Admin-Account", get_account_cases),
        ("TS-Admin-Permissions", "TS-Admin-Permissions", get_permissions_cases),
        ("TS-Admin-PMTool", "TS-Admin-PMTool", get_pmtool_cases),
        ("TS-Admin-Regression", "TS-Admin-Regression", get_regression_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    # Plan Overview (default sheet)
    plan_ws = wb.active
    plan_ws.title = "Plan Overview"
    plan_ws.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(plan_ws, suites_info)

    # Feature Matrix
    fm_ws = wb.create_sheet("Feature Matrix")
    fm_ws.sheet_properties.tabColor = TAB_COLOR_PLAN
    suite_tabs = [tab for _, tab, _ in suites]
    write_feature_matrix(fm_ws, suite_tabs)

    # Risk Assessment
    ra_ws = wb.create_sheet("Risk Assessment")
    ra_ws.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ra_ws)

    # Suite tabs
    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_suite_tab(ws, cases, tab_name)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)

    total = len(all_cases)
    print(f"Generated {OUTPUT_FILE}")
    print(f"Total: {total} test cases across {len(suites)} suites")
    for name, count, _ in suites_info:
        print(f"  {name}: {count} cases")

    return all_cases


if __name__ == "__main__":
    generate()
