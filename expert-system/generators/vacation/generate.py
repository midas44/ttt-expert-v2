#!/usr/bin/env python3
"""
Vacation Test Documentation Generator — Phase B
Generates test-docs/vacation/vacation.xlsx with Plan Overview, Feature Matrix,
Risk Assessment, and 10 TS- test suite tabs.

Based on vault knowledge: 25K+ words across 12+ notes, 50+ bugs in 10 categories,
250+ GitLab tickets searched, 55+ relevant, UI flows explored via Playwright,
API and DB schemas verified on qa-1.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "test-docs", "vacation")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "vacation.xlsx")

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=11)
FONT_BACK_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ALT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "70AD47"
TAB_COLOR_SUITE = "4472C4"


# ─── Test Case Data ──────────────────────────────────────────────────────────

def get_crud_cases():
    """TS-Vacation-CRUD: Create, edit, delete, cancel, re-open, view."""
    return [
        {
            "id": "TC-VAC-001", "title": "Create REGULAR vacation — happy path",
            "preconditions": "Employee with sufficient available vacation days (>= 5) and a manager assigned.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 5 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to My Vacations page (/vacation/my)\n3. Verify 'Available vacation days' counter shows >= 5\n4. Click 'Create a request' button\n5. In the creation dialog, select start date = next Monday, end date = next Friday\n6. Verify 'Number of days' auto-calculates (expect 5 working days, may be 4 if holiday)\n7. Verify payment month auto-populates to the month containing start date\n8. Leave 'Unpaid vacation' unchecked (REGULAR type)\n9. Verify 'Approved by' shows the employee's manager\n10. Click 'Save'\n11. Verify success notification appears\n12. Verify new vacation row in the table with status 'New'\n13. Verify 'Available vacation days' counter decreased\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Vacation created with status NEW. Table shows correct dates, regular days count, type 'Regular', approver name, and status 'New'. Available days counter decreases by the number of working days.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Create Flow", "module": "vacation/crud",
            "notes": "Working days may be 4 instead of 5 if Mon-Fri range includes a public holiday."
        },
        {
            "id": "TC-VAC-002", "title": "Create ADMINISTRATIVE (unpaid) vacation",
            "preconditions": "Any enabled employee with a manager.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select start date = a future Monday, end date = same Monday (1-day)\n5. Check 'Unpaid vacation' checkbox\n6. Verify payment type changes to Administrative\n7. Click 'Save'\n8. Verify vacation created with type 'Administrative'\n9. Verify 'Available vacation days' counter did NOT decrease (unpaid doesn't consume balance)\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Administrative vacation created. Type column shows 'Administrative'. Available days unchanged — ADMINISTRATIVE does not consume paid balance.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7", "module": "vacation/crud",
            "notes": "ADMINISTRATIVE skips duration and availability checks. No minimum working day requirement enforced."
        },
        {
            "id": "TC-VAC-003", "title": "Create vacation with comment",
            "preconditions": "Employee with sufficient vacation days.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 3 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select future dates (3 working days)\n5. In the Comment field, type 'Family trip'\n6. Click 'Save'\n7. Verify vacation created\n8. Click the '...' icon on the new row to open Request Details\n9. Verify comment 'Family trip' is displayed in the details dialog\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Comment is saved and visible in Request Details dialog.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §4 DTO", "module": "vacation/crud",
            "notes": "No character limit on comment field per DTO analysis."
        },
        {
            "id": "TC-VAC-004", "title": "Create vacation with 'Also notify' recipients",
            "preconditions": "Employee with a colleague in the same office.\nQuery: SELECT e.login, e2.login AS colleague FROM ttt_vacation.employee e JOIN ttt_vacation.employee e2 ON e.office_id = e2.office_id AND e.id != e2.id WHERE e.enabled = true AND e2.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my, click 'Create a request'\n3. Select future dates (3 working days)\n4. In the 'Also notify' multi-select, type colleague's name and select them\n5. Click 'Save'\n6. Verify vacation created\nDB-CHECK: SELECT COUNT(*) FROM ttt_vacation.vacation_notify_also vna JOIN ttt_vacation.vacation v ON vna.vacation = v.id WHERE v.id = <created_id>\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Vacation created. DB shows notify-also record for the selected colleague.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §27 notifyAlso", "module": "vacation/crud",
            "notes": "GET /vacations/{id} does NOT include notifyAlso — verification via DB only. Note UI typo: 'Also notifty'."
        },
        {
            "id": "TC-VAC-005", "title": "Edit vacation dates (NEW status)",
            "preconditions": "Employee with an existing NEW vacation.\nSETUP: Via API — POST /api/vacation/v1/vacations to create a REGULAR vacation.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 10 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation for the employee (Mon-Fri next week)\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Find the NEW vacation row\n4. Click the pencil (edit) icon\n5. Verify 'Editing vacation request' dialog opens\n6. Change end date to one week later (extend by 5 days)\n7. Verify 'Number of days' recalculates\n8. Click 'Save'\n9. Verify vacation row updates with new dates\n10. Verify status remains 'New'\n11. Verify available days counter reflects the new duration\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Vacation dates updated. Day count recalculated. Status remains NEW. Available days adjusted.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2", "module": "vacation/crud",
            "notes": "PUT update requires 'id' in request body in addition to URL path parameter."
        },
        {
            "id": "TC-VAC-006", "title": "Edit APPROVED vacation → resets status to NEW",
            "preconditions": "Employee with an APPROVED vacation.\nSETUP: Via API — create vacation, then approve it.\nQuery: SELECT e.login, e.manager_id FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 10 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation, then PUT /approve/{id} as approver to set APPROVED status\n1. Login as the vacation owner (employee)\n2. Navigate to /vacation/my\n3. Find the APPROVED vacation row\n4. Click the pencil icon\n5. Verify warning text: 'Changing the vacation dates will move the request to the New status and send it for approval once again.'\n6. Change start date to one day later\n7. Click 'Save'\n8. Verify status resets from 'Approved' to 'New'\n9. Verify optional approvals reset to ASKED status (DB check)\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Status resets to NEW after date edit. Warning was displayed before save. Optional approvals reset to ASKED.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §3, #2640", "module": "vacation/crud",
            "notes": "Bug #2640 was about this NOT resetting — now fixed. Regression test."
        },
        {
            "id": "TC-VAC-007", "title": "Cancel NEW vacation",
            "preconditions": "Employee with a NEW vacation.\nSETUP: Via API — POST /api/vacation/v1/vacations to create.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 3 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Note available days counter value\n4. Click the '...' icon on the NEW vacation row\n5. In Request Details dialog, click 'Delete' button\n6. Confirm deletion in the confirmation dialog\n7. Verify vacation disappears from 'Open' tab\n8. Click 'Closed' tab\n9. Verify vacation appears with status 'Deleted' or 'Canceled'\n10. Verify available days counter increased back to original value",
            "expected": "Vacation moved to Closed tab with terminal status. Available days restored.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Delete Flow", "module": "vacation/crud",
            "notes": "UI uses 'Delete' button for both cancel and delete actions. Status may show as 'Deleted'."
        },
        {
            "id": "TC-VAC-008", "title": "Cancel APPROVED vacation",
            "preconditions": "Employee with an APPROVED vacation in the future.\nSETUP: Via API — create vacation then approve.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 5 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation (future dates), approve as manager\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Note available days counter\n4. Find the APPROVED vacation row\n5. Click '...' icon → Request Details\n6. Click 'Cancel' or 'Delete' button\n7. Confirm in confirmation dialog\n8. Verify vacation moves to Closed tab with status 'Canceled'\n9. Verify available days counter increases (days returned to pool)\nCLEANUP: No cleanup needed — vacation is already canceled",
            "expected": "APPROVED vacation canceled. Days returned to available balance. FIFO redistribution may affect other vacations.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §3 canBeCancelled", "module": "vacation/crud",
            "notes": "canBeCancelled guard: REGULAR + APPROVED + reportPeriod after paymentDate → CANNOT cancel. Future vacations should pass."
        },
        {
            "id": "TC-VAC-009", "title": "Re-open CANCELED vacation",
            "preconditions": "Employee with a CANCELED vacation.\nSETUP: Via API — create, then cancel.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 5 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create vacation, then PUT /cancel/{id}\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Closed' tab\n4. Find the CANCELED vacation row\n5. Click the pencil (edit) icon (if available)\n6. In the editing dialog, update dates if needed\n7. Click 'Save'\n8. Verify vacation moves back to 'Open' tab with status 'New'\n9. Verify available days decreased again\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "CANCELED vacation re-opened to NEW status. Available days decreased. Vacation appears in Open tab.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §1 CANCELED→NEW", "module": "vacation/crud",
            "notes": "CANCELED is in FINAL_STATUSES but has explicit CANCELED→NEW transition in the map."
        },
        {
            "id": "TC-VAC-010", "title": "View Request Details dialog",
            "preconditions": "Employee with at least one vacation in any status.\nSETUP: Via API — create a vacation if none exist.",
            "steps": "SETUP: Via API — create a vacation if needed\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click the '...' icon on any vacation row\n4. Verify 'Request details' dialog opens\n5. Verify all fields displayed: Period, Number of days, Status, Vacation type, Payment month, Approved by\n6. Verify the Close button (X) dismisses the dialog\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id} if created",
            "expected": "Request Details dialog shows all vacation attributes. Period matches table row dates. Close button works.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Request Details Dialog", "module": "vacation/crud",
            "notes": ""
        },
        {
            "id": "TC-VAC-011", "title": "Available days counter — per-year breakdown tooltip",
            "preconditions": "Employee with vacation days across multiple years (current + prior).\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days > 0 GROUP BY e.login HAVING COUNT(DISTINCT ev.year) > 1 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Locate 'Available vacation days: N in YYYY' counter\n4. Click the info (i) icon next to the counter\n5. Verify tooltip/popup opens showing per-year breakdown\n6. Verify each year shows: year number and available days for that year\n7. Verify sum of all years matches the total shown in the counter",
            "expected": "Per-year breakdown popup shows individual year balances that sum to the total displayed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Available Days Widget", "module": "vacation/crud",
            "notes": "DOM: label and value are sibling divs, not parent-child. Value uses &nbsp; (U+00A0)."
        },
        {
            "id": "TC-VAC-012", "title": "Vacation events feed",
            "preconditions": "Employee with recent vacation activity.\nSETUP: Via API — create and then delete a vacation to generate timeline events.",
            "steps": "SETUP: Via API — create a vacation, then delete it (generates 2 timeline events)\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Vacation events feed' button\n4. Verify events feed panel opens\n5. Verify at least 2 recent events: vacation creation and deletion\n6. Verify each event shows: date, event type, description\nCLEANUP: No cleanup needed",
            "expected": "Events feed shows chronological vacation activity including creation and status changes.",
            "priority": "Low", "type": "UI",
            "req_ref": "vacation-pages.md §Available Days section", "module": "vacation/crud",
            "notes": "DB table: ttt_vacation.timeline tracks all events. Event types: VACATION_CREATED, VACATION_DELETED, etc."
        },
        {
            "id": "TC-VAC-013", "title": "Delete PAID+NON-EXACT vacation (allowed)",
            "preconditions": "This is an edge case — PAID+NON_EXACT vacations can be deleted. Requires a PAID vacation with periodType != EXACT.\nNote: Difficult to set up without direct DB manipulation. Consider API-only verification.",
            "steps": "SETUP: This test requires a PAID vacation with non-EXACT periodType. Use test API or DB if available.\n1. Via API — attempt DELETE /api/vacation/v1/vacations/{id} on the PAID+NON-EXACT vacation\n2. Verify deletion succeeds (HTTP 200)\n3. Verify vacation status changes to DELETED in DB\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id>",
            "expected": "PAID+NON_EXACT vacation can be deleted — this is a design issue. PAID should be terminal but the guard only blocks PAID+EXACT.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §2 Delete Guard", "module": "vacation/crud",
            "notes": "Design issue: deleteVacation only guards PAID+EXACT. PAID+NON_EXACT passes through."
        },
        {
            "id": "TC-VAC-014", "title": "Soft delete — record persists in DB",
            "preconditions": "Employee with a NEW vacation.\nSETUP: Via API — create a vacation.",
            "steps": "SETUP: Via API — create a REGULAR vacation, note the vacation ID\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Delete the vacation via Request Details dialog\n4. Verify vacation disappears from 'Open' tab\n5. Click 'Closed' tab — verify vacation appears with status 'Deleted'\nDB-CHECK: SELECT id, status FROM ttt_vacation.vacation WHERE id = <id> — verify status = 'DELETED' (not physically removed)",
            "expected": "Vacation is soft-deleted: status set to DELETED in DB, still visible in Closed tab. Record is not physically removed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Delete Flow", "module": "vacation/crud",
            "notes": "Soft delete with status=DELETED. GET API still returns the record. Crossing check counts DELETED records (design issue)."
        },
    ]


def get_approval_cases():
    """TS-Vacation-Approval: Approve, reject, re-approve, CPO, change approver, optional approvers."""
    return [
        {
            "id": "TC-VAC-015", "title": "Approve NEW vacation — happy path",
            "preconditions": "Employee with a NEW vacation, logged in as the assigned approver (manager).\nSETUP: Via API — create a vacation for the employee.\nQuery: SELECT e.login AS employee, m.login AS manager FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id WHERE e.enabled = true AND m.enabled = true AND e.id != m.id ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation for the employee\n1. Login as the manager (approver)\n2. Navigate to Employee Requests page (/vacation/request)\n3. Click 'Vacation requests' tab\n4. Click 'Approval' sub-filter\n5. Find the employee's vacation request in the table\n6. Click the approve (checkmark) button on the row\n7. Verify success notification\n8. Verify the row disappears from the Approval list (or status changes to Approved)\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — verify status = 'APPROVED'\nCLEANUP: Via API — PUT /cancel/{id} then DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Vacation status changes to APPROVED. Row disappears from approval queue. DB confirms APPROVED status.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Approve Flow", "module": "vacation/approval",
            "notes": "Approve adjusts paymentDate if before approvePeriodStartDate."
        },
        {
            "id": "TC-VAC-016", "title": "Reject NEW vacation",
            "preconditions": "Manager with a pending approval request.\nSETUP: Via API — create a vacation for the employee.\nQuery: (same as TC-VAC-015)",
            "steps": "SETUP: Via API — create a REGULAR vacation for the employee\n1. Login as the manager\n2. Navigate to /vacation/request\n3. Click 'Approval' sub-filter\n4. Find the employee's vacation\n5. Click the reject (X) button\n6. Verify rejection confirmation or success\n7. Verify row disappears from Approval list\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — verify 'REJECTED'\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "Vacation rejected. Status changes to REJECTED. Days returned to employee's available balance.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §1 State Machine", "module": "vacation/approval",
            "notes": "Reject uses VACATIONS_DELETE permission — same as cancel and delete."
        },
        {
            "id": "TC-VAC-017", "title": "Reject APPROVED vacation",
            "preconditions": "Manager with an APPROVED vacation assigned to them.\nSETUP: Via API — create vacation, then approve it.\nQuery: (same as TC-VAC-015)",
            "steps": "SETUP: Via API — create vacation, approve it\n1. Login as the manager\n2. Navigate to /vacation/request\n3. Find the APPROVED vacation (may need 'My department' or search filter)\n4. Verify reject button is available for APPROVED status\n5. Click the reject button\n6. Verify status changes to REJECTED\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — verify 'REJECTED'\nCLEANUP: Via API — DELETE /api/vacation/v1/vacations/{id}",
            "expected": "APPROVED vacation can be rejected. Status transitions to REJECTED. Days returned to pool, FIFO redistribution triggers.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §2 APPROVED→REJECTED", "module": "vacation/approval",
            "notes": "REJECTABLE_STATUSES = {NEW, APPROVED}. canBeCancelled guard also applies."
        },
        {
            "id": "TC-VAC-018", "title": "Re-approve REJECTED vacation (without edit)",
            "preconditions": "Manager with a REJECTED vacation assigned to them.\nSETUP: Via API — create vacation, reject it.",
            "steps": "SETUP: Via API — create vacation, then reject it via PUT /reject/{id}\n1. Login as the manager\n2. Navigate to /vacation/request\n3. Find the REJECTED vacation\n4. Verify approve button is available\n5. Click approve\n6. Verify status changes to APPROVED\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — verify 'APPROVED'\nCLEANUP: Via API — PUT /cancel/{id} then DELETE",
            "expected": "REJECTED vacation can be directly re-approved without the employee editing it. State machine allows REJECTED→APPROVED.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §1 REJECTED→APPROVED", "module": "vacation/approval",
            "notes": "APPROVABLE_STATUSES = {NEW, REJECTED}. No edit required between reject and re-approve."
        },
        {
            "id": "TC-VAC-019", "title": "CPO self-approval on create",
            "preconditions": "Employee with ROLE_DEPARTMENT_MANAGER (CPO role) and a manager assigned.\nQuery: SELECT e.login, m.login AS manager FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id WHERE e.enabled = true AND EXISTS (SELECT 1 FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id WHERE be.login = e.login AND r.name = 'ROLE_DEPARTMENT_MANAGER') ORDER BY random() LIMIT 1",
            "steps": "1. Login as the CPO employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select future dates\n5. Verify 'Approved by' field shows the employee themselves (self-approval)\n6. Verify 'Agreed by' shows their manager\n7. Click 'Save'\n8. Verify vacation created with status 'Approved' (auto-approved on create for CPO)\nDB-CHECK: SELECT v.approver_id, v.employee_id FROM ttt_vacation.vacation v WHERE v.id = <id> — verify approver_id = employee_id\nDB-CHECK: SELECT va.status FROM ttt_vacation.vacation_approval va WHERE va.vacation = <id> — verify manager has ASKED status\nCLEANUP: Via API — DELETE",
            "expected": "CPO vacation is self-approved on create. Manager added as optional approver with ASKED status. Status is NEW (not auto-APPROVED — requires explicit approve action).",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 CPO path, #3319", "module": "vacation/approval",
            "notes": "CPO = ROLE_DEPARTMENT_MANAGER. pvaynmaster confirmed as CPO on qa-1. Status is NEW, not auto-APPROVED."
        },
        {
            "id": "TC-VAC-020", "title": "Change approver (redirect request)",
            "preconditions": "Manager with a NEW vacation request to approve.\nSETUP: Via API — create a vacation.\nA second manager must exist to redirect to.\nQuery: SELECT e.login AS employee, m.login AS manager, m2.login AS alt_manager FROM ttt_vacation.employee e JOIN ttt_vacation.employee m ON e.manager_id = m.id JOIN ttt_vacation.employee m2 ON m2.enabled = true AND m2.id != m.id AND m2.id != e.id WHERE e.enabled = true AND m.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation for the employee\n1. Login as the current approver (manager)\n2. Navigate to /vacation/request\n3. Find the vacation request\n4. Click the redirect (arrow) button\n5. In the redirect dialog, search and select the alternative manager\n6. Confirm redirect\n7. Verify the request disappears from current approver's queue\nDB-CHECK: SELECT approver_id FROM ttt_vacation.vacation WHERE id = <id> — verify changed to alt_manager's ID\nCLEANUP: Via API — DELETE the vacation",
            "expected": "Approver changed. Old approver becomes optional (ASKED). New approver sees the request in their queue.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §3 Reassignment", "module": "vacation/approval",
            "notes": "Bug #2718 (OPEN): Redirected approved/rejected request doesn't reset status to NEW."
        },
        {
            "id": "TC-VAC-021", "title": "Optional approver — approve",
            "preconditions": "Vacation with an optional approver in ASKED status.\nSETUP: Via API — create vacation with optionalApprovers=[login].\nQuery: SELECT e.login AS employee, oa.login AS optional FROM ttt_vacation.employee e JOIN ttt_vacation.employee oa ON oa.enabled = true AND oa.id != e.id AND oa.id != e.manager_id WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create vacation with optionalApprovers parameter including the optional approver's login\n1. Login as the optional approver\n2. Navigate to /vacation/request\n3. Click 'Agreement' sub-filter\n4. Find the vacation request\n5. Click approve (checkmark)\n6. Verify optional approval status changes\nDB-CHECK: SELECT status FROM ttt_vacation.vacation_approval WHERE vacation = <id> AND employee = <oa_employee_id> — verify 'APPROVED'\nCLEANUP: Via API — DELETE the vacation",
            "expected": "Optional approver's status changes from ASKED to APPROVED. Vacation overall status remains unchanged (optional approval doesn't drive state machine).",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §3 Two-tier model", "module": "vacation/approval",
            "notes": "Optional approval is informational — doesn't change vacation status. Only primary approver drives transitions."
        },
        {
            "id": "TC-VAC-022", "title": "Approval resets on date edit",
            "preconditions": "Vacation with optional approvals that have been acted on.\nSETUP: Via API — create vacation with optionalApprovers, then approve optionally.\nQuery: (same as TC-VAC-021)",
            "steps": "SETUP: Via API — create vacation with optionalApprovers, have the optional approver approve\n1. Login as the vacation owner\n2. Navigate to /vacation/my\n3. Edit the vacation dates (change end date)\n4. Click 'Save'\n5. Verify dates updated\nDB-CHECK: SELECT status FROM ttt_vacation.vacation_approval WHERE vacation = <id> — verify all reset to 'ASKED'\nCLEANUP: Via API — DELETE the vacation",
            "expected": "All optional approvals reset to ASKED status after date edit. Primary approver remains unchanged.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §3 State resets", "module": "vacation/approval",
            "notes": ""
        },
        {
            "id": "TC-VAC-023", "title": "Employee Requests page — view pending approvals",
            "preconditions": "Manager with at least one pending vacation request.\nSETUP: Via API — create a vacation for an employee managed by this manager.",
            "steps": "SETUP: Via API — create a vacation for a managed employee\n1. Login as the manager\n2. Navigate to /vacation/request\n3. Verify 'Vacation requests (N)' tab shows count >= 1\n4. Click 'Approval' sub-filter\n5. Verify table shows: Employee, Vacation dates, Vacation type, Manager, Approved by, Agreed by, Payment month, Status, Actions\n6. Verify the created vacation appears in the list\n7. Verify action buttons: approve (checkmark), reject (X), redirect (arrow), details (eye)\nCLEANUP: Via API — DELETE the vacation",
            "expected": "Employee Requests page displays pending approvals with all columns and action buttons.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-pages.md §Employee Requests Page", "module": "vacation/approval",
            "notes": "Multiple rows may exist for same employee from previous test runs. Filter by employee + period."
        },
        {
            "id": "TC-VAC-024", "title": "No-manager employee — self-approval",
            "preconditions": "Employee with no manager assigned (manager_id IS NULL).\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE e.manager_id IS NULL AND e.enabled = true AND ev.available_vacation_days >= 3 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) ORDER BY random() LIMIT 1",
            "steps": "1. Login as the no-manager employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select future dates\n5. Verify 'Approved by' shows the employee themselves\n6. Click 'Save'\n7. Verify vacation created\nDB-CHECK: SELECT approver_id, employee_id FROM ttt_vacation.vacation WHERE id = <id> — verify approver_id = employee_id\nCLEANUP: Via API — DELETE",
            "expected": "Employee with no manager self-approves. Approver set to themselves.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 No-manager path", "module": "vacation/approval",
            "notes": "Rare case — most employees have managers. May not find a test user."
        },
    ]


def get_payment_cases():
    """TS-Vacation-Payment: Pay, validation, terminal state, auto-pay."""
    return [
        {
            "id": "TC-VAC-025", "title": "Pay APPROVED REGULAR vacation — happy path",
            "preconditions": "Accountant user. Employee with APPROVED REGULAR vacation.\nSETUP: Via API — create vacation, approve it.\nQuery accountant: SELECT e.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id JOIN ttt_vacation.employee e ON e.login = be.login WHERE r.name IN ('ROLE_ACCOUNTANT', 'ROLE_CHIEF_ACCOUNTANT') AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a 5-day REGULAR vacation for an employee, approve it\n1. Login as the accountant\n2. Navigate to Accounting section (dropdown → vacation payments or similar)\n3. Find the APPROVED vacation in the payment queue\n4. Enter regular days paid = 5, administrative days paid = 0\n5. Confirm payment\n6. Verify vacation status changes to 'Paid'\nDB-CHECK: SELECT vp.regular_days, vp.administrative_days, vp.payed_at FROM ttt_vacation.vacation v JOIN ttt_vacation.vacation_payment vp ON v.vacation_payment_id = vp.id WHERE v.id = <id>\nNote: PAID+EXACT vacations cannot be deleted — permanent record in test env",
            "expected": "Vacation paid. Status = PAID. vacation_payment record created with correct day split. Payment date recorded.",
            "priority": "Critical", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §2 Pay endpoint", "module": "vacation/payment",
            "notes": "PAID+EXACT is terminal and undeletable. Creates permanent data in test environment."
        },
        {
            "id": "TC-VAC-026", "title": "Pay ADMINISTRATIVE vacation",
            "preconditions": "Accountant. Employee with APPROVED ADMINISTRATIVE vacation.\nSETUP: Via API — create ADMINISTRATIVE vacation (paymentType=ADMINISTRATIVE), approve it.",
            "steps": "SETUP: Via API — create ADMINISTRATIVE vacation, approve\n1. Login as accountant\n2. Find the APPROVED ADMINISTRATIVE vacation\n3. Pay with regularDaysPayed=0, administrativeDaysPayed=N (matching total days)\n4. Verify status changes to PAID\nDB-CHECK: Verify vacation_payment.administrative_days = N, regular_days = 0",
            "expected": "ADMINISTRATIVE vacation paid with all days as administrative. No impact on paid vacation balance.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes ADMINISTRATIVE", "module": "vacation/payment",
            "notes": ""
        },
        {
            "id": "TC-VAC-027", "title": "Payment validation — wrong day sum rejected",
            "preconditions": "APPROVED vacation with known day count.\nSETUP: Via API — create 5-day vacation, approve.",
            "steps": "SETUP: Via API — create 5-day vacation, approve\n1. Via API — PUT /pay/{id} with body: {regularDaysPayed: 3, administrativeDaysPayed: 0} (sum=3, expected=5)\n2. Verify HTTP 400 response\n3. Verify errorCode: 'exception.vacation.pay.days.not.equal'\n4. Verify vacation status remains APPROVED\nCLEANUP: Via API — cancel and delete",
            "expected": "Payment rejected with 400. Error code indicates day sum mismatch. Status unchanged.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §Pay endpoint", "module": "vacation/payment",
            "notes": "Sum validation: regularDaysPayed + administrativeDaysPayed must equal vacation.days."
        },
        {
            "id": "TC-VAC-028", "title": "Cannot pay NEW vacation",
            "preconditions": "NEW vacation.\nSETUP: Via API — create vacation (status remains NEW).",
            "steps": "SETUP: Via API — create a vacation\n1. Via API — PUT /pay/{id} with body: {regularDaysPayed: 5, administrativeDaysPayed: 0}\n2. Verify HTTP 400 response\n3. Verify error indicates status must be APPROVED\n4. Verify vacation status remains NEW\nCLEANUP: Via API — DELETE",
            "expected": "Payment blocked for non-APPROVED status. PayVacationServiceImpl.checkForPayment validates APPROVED + EXACT.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Session 27", "module": "vacation/payment",
            "notes": ""
        },
        {
            "id": "TC-VAC-029", "title": "PAID vacation — terminal state, no further transitions",
            "preconditions": "PAID vacation.\nSETUP: Via API — create, approve, pay.",
            "steps": "SETUP: Via API — create vacation, approve, pay\n1. Via API — PUT /cancel/{id} — verify HTTP 400\n2. Via API — PUT /reject/{id} — verify HTTP 400\n3. Via API — PUT update with new dates — verify HTTP 400\n4. Via API — DELETE /{id} — verify HTTP 400 (for PAID+EXACT)\n5. Verify all attempts return errorCode: 'exception.vacation.status.notAllowed' or 'exception.vacation.delete.notAllowed'\nNote: PAID+EXACT vacation remains as permanent record",
            "expected": "PAID is terminal. Cancel, reject, edit, and delete all blocked. Error codes confirm denial.",
            "priority": "Critical", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §1 FINAL_STATUSES", "module": "vacation/payment",
            "notes": "NON_EDITABLE_STATUSES = {CANCELED, PAID}. Permission service returns empty set for PAID."
        },
        {
            "id": "TC-VAC-030", "title": "Delete PAID+EXACT blocked",
            "preconditions": "PAID+EXACT vacation.\nSETUP: Via API — create, approve, pay (creates EXACT period).",
            "steps": "SETUP: Via API — create vacation, approve, pay\n1. Via API — DELETE /api/vacation/v1/vacations/{id}\n2. Verify HTTP 400\n3. Verify errorCode: 'exception.vacation.delete.notAllowed'\n4. Verify vacation still exists with PAID status\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id>",
            "expected": "Deletion blocked for PAID+EXACT. ServiceException with delete.notAllowed error code.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §2 Delete Guard", "module": "vacation/payment",
            "notes": "Guard: status==PAID && periodType==EXACT → throw."
        },
        {
            "id": "TC-VAC-031", "title": "Payment month validation — closed period blocked",
            "preconditions": "Vacation with payment month in a closed accounting period.\nNote: Requires knowledge of current open/closed periods on the test env.",
            "steps": "1. Login as employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select future dates\n5. Attempt to select a past/closed month for payment\n6. Verify the month picker restricts selection to open periods\n7. Or verify error on save if closed month is somehow selected",
            "expected": "Payment month picker prevents selection of closed accounting periods. Backend validates: paymentDate must be within approval period range.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §5, #3379", "module": "vacation/payment",
            "notes": "Bug #3379 (fixed): was using Report period instead of Approval period."
        },
        {
            "id": "TC-VAC-032", "title": "Auto-pay expired APPROVED vacations (cron)",
            "preconditions": "APPROVED vacation older than 2 months (requires timemachine or clock manipulation).\nSETUP: Via test API — PATCH clock to advance 3 months, or find naturally expired vacation.",
            "steps": "SETUP: Via test API — create vacation in the past (via timemachine clock), approve it, advance clock 3 months\n1. Via test API — POST /api/vacation/v1/test/vacations/pay-expired-approved\n2. Verify the cron-like endpoint processes expired vacations\n3. Verify vacation status changes to PAID\nDB-CHECK: SELECT status, vacation_payment_id FROM ttt_vacation.vacation WHERE id = <id>",
            "expected": "Auto-pay processes APPROVED vacations >2 months old. Status changes to PAID with auto-distributed day types.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-business-rules-reference.md §5 Auto-pay cron", "module": "vacation/payment",
            "notes": "Test endpoint: POST /api/vacation/v1/test/vacations/pay-expired-approved. Needs timemachine env."
        },
        {
            "id": "TC-VAC-033", "title": "Error 500 on AV=true negative balance payment (#3363)",
            "preconditions": "AV=true employee with negative vacation balance. APPROVED vacation.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id WHERE o.advance_vacation = true AND e.enabled = true AND EXISTS (SELECT 1 FROM ttt_vacation.employee_vacation ev WHERE ev.employee = e.id AND ev.available_vacation_days < 0) ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create and approve a vacation for the AV=true employee with negative balance\n1. Via API — PUT /pay/{id} with correct day split\n2. Verify HTTP 500 response (known bug)\n3. Verify vacation status remains APPROVED (payment failed)\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id> — should still be APPROVED",
            "expected": "Bug #3363 (OPEN): HTTP 500 on payment when AV=true balance is negative. No validation — unhandled NPE/exception.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-business-rules-reference.md §10 F1, #3363", "module": "vacation/payment",
            "notes": "OPEN bug. Regression test — should catch if the 500 changes to a proper validation error."
        },
    ]


def get_validation_cases():
    """TS-Vacation-Validation: Date rules, balance checks, crossing, duration, 3-month restriction."""
    return [
        {
            "id": "TC-VAC-034", "title": "Start date in past — rejected",
            "preconditions": "Any employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Attempt to select yesterday as start date\n5. Verify date picker prevents past date selection OR\n6. If date can be entered, click 'Save'\n7. Verify error message: validation.vacation.start.date.in.past (may show as raw key — no i18n)\nCLEANUP: No cleanup — creation should fail",
            "expected": "Past start date rejected. Error shown. Boundary: today is accepted, yesterday is rejected (isBefore(today)).",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-form-validation-rules.md §Past-Date Check", "module": "vacation/validation",
            "notes": "Missing i18n: error key displayed as raw string. Boundary at today (not tomorrow)."
        },
        {
            "id": "TC-VAC-035", "title": "Start date > end date — rejected",
            "preconditions": "Any employee.",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Set start date = next Friday, end date = next Monday (before start)\n5. Click 'Save'\n6. Verify error: validation.vacation.dates.order",
            "expected": "Date order validation rejects startDate > endDate. Error code: validation.vacation.dates.order.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §4 VacationCreateValidator", "module": "vacation/validation",
            "notes": "Non-short-circuiting: both past-date and order errors returned simultaneously if both fail."
        },
        {
            "id": "TC-VAC-036", "title": "Insufficient available days — REGULAR blocked",
            "preconditions": "Employee with known available days < requested duration.\nQuery: SELECT e.login, ev.available_vacation_days FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days BETWEEN 1 AND 5 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select dates spanning more working days than available (e.g., if 3 days available, select 2-week period)\n5. Verify red error message appears indicating insufficient days\n6. Verify 'Save' button is disabled or click 'Save' to get backend error\n7. Verify error code: validation.vacation.duration",
            "expected": "Insufficient days blocks creation. Frontend shows red error. Backend returns validation.vacation.duration if submitted.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7", "module": "vacation/validation",
            "notes": "AV=false: red error, block submit. AV=true: Error 11.4. Frontend validation fires dynamically on field change."
        },
        {
            "id": "TC-VAC-037", "title": "Overlapping vacation — crossing check",
            "preconditions": "Employee with an existing vacation in specific date range.\nSETUP: Via API — create vacation for specific dates.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days >= 15 AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create a REGULAR vacation from Monday to Friday (week 1)\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select dates that overlap with the existing vacation (e.g., Wednesday to next Wednesday)\n5. Verify crossing validation fires — error shown\n6. Verify error code: exception.validation.vacation.dates.crossing\nCLEANUP: Via API — DELETE both vacations",
            "expected": "Overlapping vacation blocked. Error: exception.validation.vacation.dates.crossing. Three overlap patterns: start-in-range, end-in-range, enclosing.",
            "priority": "Critical", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Create Flow step 5", "module": "vacation/validation",
            "notes": "IMPORTANT: Crossing check counts DELETED records too (ghost conflicts). Soft-deleted vacations block date ranges permanently."
        },
        {
            "id": "TC-VAC-038", "title": "Weekend-only vacation (0 working days) — rejected",
            "preconditions": "Any employee.",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select Saturday to Sunday (2 calendar days, 0 working days)\n5. Verify 'Number of days' shows 0\n6. Click 'Save'\n7. Verify error: validation.vacation.duration (minimum is 1 working day)",
            "expected": "Vacation with 0 working days rejected. minimalVacationDuration=1 in config. Only Sat-Sun or holiday-only ranges produce 0.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes minimalVacationDuration", "module": "vacation/validation",
            "notes": "Config: vacation.minimal-vacation-duration: 1 (not 5 as Javadoc says). Counts working days only."
        },
        {
            "id": "TC-VAC-039", "title": "Next year not available before Feb 1",
            "preconditions": "Test requires current date to be before Feb 1 (use timemachine to set clock to January).",
            "steps": "SETUP: Via test API — set clock to January 15 of current year (timemachine env)\n1. Login as employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Attempt to select a start date in the next year\n5. Verify error: validation.vacation.next.year.not.available\nCLEANUP: Via test API — reset clock",
            "expected": "Before Feb 1, next-year vacation creation is blocked. After Feb 1, it's allowed.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-form-validation-rules.md §Backend Validation Create", "module": "vacation/validation",
            "notes": "Needs timemachine env with clock set to January. Boundary: Jan 31 rejected, Feb 1 allowed."
        },
        {
            "id": "TC-VAC-040", "title": "First 3 months restriction — new employee (#3014)",
            "preconditions": "Employee hired within last 3 months.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.first_date > CURRENT_DATE - INTERVAL '90 days' AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Login as the recently-hired employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Verify dates within first 3 months of employment are disabled in the calendar picker\n5. Verify calendar opens to first available (non-disabled) date\n6. Select a date after the 3-month restriction period\n7. Verify creation succeeds for dates after restriction\nCLEANUP: Via API — DELETE",
            "expected": "First 3 months of employment dates are disabled for REGULAR vacations. ADMINISTRATIVE vacations are NOT restricted.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7, #3014", "module": "vacation/validation",
            "notes": "Bug #3014-30: calculation was incorrect for employees hired before July 2024. Column: employee.first_date."
        },
        {
            "id": "TC-VAC-041", "title": "First 3 months — ADMINISTRATIVE not restricted",
            "preconditions": "Same recently-hired employee as TC-VAC-040.",
            "steps": "1. Login as the recently-hired employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Check 'Unpaid vacation' checkbox (ADMINISTRATIVE)\n5. Select dates within the first 3 months of employment\n6. Verify dates are NOT disabled (no restriction for ADMINISTRATIVE)\n7. Click 'Save'\n8. Verify vacation created successfully\nCLEANUP: Via API — DELETE",
            "expected": "ADMINISTRATIVE vacations bypass the 3-month restriction. Employee can take unpaid leave immediately.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7 #3014", "module": "vacation/validation",
            "notes": ""
        },
        {
            "id": "TC-VAC-042", "title": "Payment month range — 2 months before to end month",
            "preconditions": "Employee creating a vacation with specific month range.",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select dates starting 3 months from now\n5. Open payment month picker\n6. Verify allowed range: from 2 months before start date to the month of the end date\n7. Attempt to select a month outside this range\n8. Verify it's either disabled or rejected on save",
            "expected": "Payment month restricted to valid range. Backend: correctPaymentMonth() adjusts if needed, isPaymentDateCorrect() validates.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §2 Create Flow step 3-4", "module": "vacation/validation",
            "notes": "Error code: validation.vacation.dates.payment."
        },
        {
            "id": "TC-VAC-043", "title": "Null paymentMonth → server error (NPE bug)",
            "preconditions": "Any employee.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with valid fields but paymentMonth: null\n2. Verify HTTP 500 response\n3. Verify error is an NPE (no @NotNull annotation on paymentMonth field)\nDB-CHECK: Verify no vacation was created",
            "expected": "Bug: null paymentMonth causes HTTP 500 (NPE). No DTO-level validation exists for this field.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §7 NPE #1, §Autotest Notes Session 24", "module": "vacation/validation",
            "notes": "Confirmed active on qa-1. NPE in correctPaymentMonth()."
        },
        {
            "id": "TC-VAC-044", "title": "Dynamic validation — messages update on field change",
            "preconditions": "Employee with limited available days.",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select dates exceeding available days → verify red error appears\n5. Change to shorter dates within available days → verify red error disappears\n6. Change back to long dates → verify red error reappears\n7. Check 'Unpaid vacation' → verify error disappears (ADMINISTRATIVE skips balance check)",
            "expected": "Validation messages update dynamically on field change (not just on Save click). Exception: 'required field' validation only on Save.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7 Dynamic validation (#3014)", "module": "vacation/validation",
            "notes": ""
        },
        {
            "id": "TC-VAC-045", "title": "Accrued days validation — future request auto-conversion (#3015)",
            "preconditions": "Employee with multiple future REGULAR vacations consuming most available days.\nNote: Complex setup — requires creating 2+ vacations that together exhaust the balance.",
            "steps": "SETUP: Via API — create 2 REGULAR vacations: one in month M, one in month M+2, together consuming ~90% of available days\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Create a 3rd vacation for month M+1 that would push total over available\n5. Click 'Save'\n6. Verify either: error for insufficient days, OR auto-conversion of a future request to Administrative\n7. Check the vacation list for any automatically converted vacations\n8. Verify orange warning appears about accrued days impact on future requests\nCLEANUP: Via API — DELETE all created vacations",
            "expected": "Accrued days validation checks all future requests. May auto-convert earliest-deficit request to Administrative. Orange warning shown.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §7 #3015, exploration/tickets/vacation-ticket-3015", "module": "vacation/validation",
            "notes": "31 sub-bugs in #3015. Conversion order: chronological by payment month, select request with minimum X value."
        },
        {
            "id": "TC-VAC-046", "title": "Holiday impact on working days count",
            "preconditions": "Employee in an office with a public holiday on a weekday.\nQuery: SELECT e.login, cd.event_date FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id JOIN ttt_calendar.calendar c ON c.office_id = o.id JOIN ttt_calendar.calendar_day cd ON cd.calendar_id = c.id WHERE cd.event_date > CURRENT_DATE AND cd.duration = 0 AND EXTRACT(DOW FROM cd.event_date) BETWEEN 1 AND 5 AND c.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true AND e.manager_id IS NOT NULL ORDER BY cd.event_date LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select a Mon-Fri range that includes the public holiday\n5. Verify 'Number of days' shows 4 (not 5) for a 5-calendar-day range with 1 holiday\n6. Verify the calculation uses the office-specific production calendar\nCLEANUP: Cancel dialog without saving",
            "expected": "Working days count excludes public holidays from the employee's office calendar. Mon-Fri with 1 holiday = 4 working days.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Holiday Impact", "module": "vacation/validation",
            "notes": "Different offices may have different holiday calendars. Day count is per-office."
        },
    ]


def get_filter_cases():
    """TS-Vacation-Filters: Tab filters, column filters, sorting, search."""
    return [
        {
            "id": "TC-VAC-047", "title": "Filter by Open tab (default view)",
            "preconditions": "Employee with vacations in both open and closed statuses.\nSETUP: Via API — create a NEW vacation and a CANCELED vacation.",
            "steps": "SETUP: Via API — create 2 vacations: leave one as NEW, cancel the other\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Verify 'Open' tab is selected by default\n4. Verify only non-terminal vacations shown (NEW, APPROVED)\n5. Verify the CANCELED vacation is NOT in this list\nCLEANUP: Via API — DELETE remaining vacation",
            "expected": "Open tab shows only active vacations (NEW, APPROVED). Terminal statuses (CANCELED, DELETED, PAID) excluded.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-pages.md §Filter tabs", "module": "vacation/filters",
            "notes": ""
        },
        {
            "id": "TC-VAC-048", "title": "Filter by Closed tab",
            "preconditions": "Employee with at least one canceled/deleted vacation.\nSETUP: Via API — create then cancel a vacation.",
            "steps": "SETUP: Via API — create vacation, cancel it\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Closed' tab\n4. Verify closed vacations shown (CANCELED, DELETED, PAID, REJECTED)\n5. Verify the canceled vacation appears in this list\nCLEANUP: None needed",
            "expected": "Closed tab shows terminal-status vacations. CANCELED and DELETED vacation visible here.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-pages.md §Filter tabs", "module": "vacation/filters",
            "notes": ""
        },
        {
            "id": "TC-VAC-049", "title": "Filter by All tab",
            "preconditions": "Employee with vacations in multiple statuses.\nSETUP: Via API — create one NEW and one CANCELED.",
            "steps": "SETUP: Via API — create NEW + CANCELED vacations\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'All' tab\n4. Verify both open AND closed vacations shown\n5. Verify total count = Open count + Closed count\nCLEANUP: Via API — DELETE remaining",
            "expected": "All tab shows every vacation regardless of status.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Filter tabs", "module": "vacation/filters",
            "notes": "Clicking 'All' also resets any column-level filters."
        },
        {
            "id": "TC-VAC-050", "title": "Column filter — Vacation type: Regular only",
            "preconditions": "Employee with both REGULAR and ADMINISTRATIVE vacations.\nSETUP: Via API — create one REGULAR and one ADMINISTRATIVE vacation.",
            "steps": "SETUP: Via API — create 1 REGULAR + 1 ADMINISTRATIVE vacation\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click the filter icon on 'Vacation type' column header\n4. Uncheck 'All'\n5. Check 'Regular' only\n6. Verify table immediately filters to show only Regular vacations\n7. Verify the ADMINISTRATIVE vacation is hidden\n8. Press Escape to close dropdown\nCLEANUP: Via API — DELETE both",
            "expected": "Column filter works in real-time. Only Regular vacations shown when filter applied.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Filter Behavior", "module": "vacation/filters",
            "notes": "Filter applies in real-time while dropdown is open. No need to close dropdown first."
        },
        {
            "id": "TC-VAC-051", "title": "Column filter — Status: Approved only",
            "preconditions": "Employee with vacations in NEW and APPROVED statuses.\nSETUP: Via API — create 2 vacations, approve one.",
            "steps": "SETUP: Via API — create 2 vacations, approve one\n1. Login as the employee\n2. Navigate to /vacation/my, ensure 'All' tab\n3. Click filter icon on 'Status' column\n4. Uncheck 'All', check 'Approved'\n5. Verify only APPROVED vacation shown\n6. Verify NEW vacation hidden\n7. Press Escape\nCLEANUP: Via API — cancel approved, delete both",
            "expected": "Status filter correctly shows only Approved vacations.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Filter Behavior", "module": "vacation/filters",
            "notes": "Status checkboxes: All, New, Approved, Rejected, Paid, Finished, Deleted."
        },
        {
            "id": "TC-VAC-052", "title": "Sort by Vacation dates column",
            "preconditions": "Employee with 3+ vacations at different dates.\nSETUP: Via API — create 3 vacations with different date ranges.",
            "steps": "SETUP: Via API — create 3 vacations: week 1, week 3, week 5\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Vacation dates' column header to sort ascending\n4. Verify rows ordered by start date earliest first\n5. Click again to sort descending\n6. Verify rows ordered latest first\nCLEANUP: Via API — DELETE all 3",
            "expected": "Sorting by Vacation dates works in both ascending and descending order.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Table columns", "module": "vacation/filters",
            "notes": ""
        },
        {
            "id": "TC-VAC-053", "title": "Table footer — Total row sums",
            "preconditions": "Employee with 2+ vacations showing in current tab.\nSETUP: Via API — create 2 REGULAR vacations.",
            "steps": "SETUP: Via API — create 2 REGULAR vacations (e.g., 5 days + 3 days)\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Verify 'Total' row at bottom of table\n4. Verify Total Regular days = sum of individual Regular days columns\n5. Verify Total Administrative days = sum of individual Administrative days columns\nCLEANUP: Via API — DELETE both",
            "expected": "Footer Total row correctly sums Regular days and Administrative days across all visible rows.",
            "priority": "Low", "type": "UI",
            "req_ref": "vacation-pages.md §Table footer", "module": "vacation/filters",
            "notes": "Total row is in <tfoot>, separate from data rows in <tbody>."
        },
        {
            "id": "TC-VAC-054", "title": "Availability chart — vacation display",
            "preconditions": "Employee with APPROVED vacation.\nSETUP: Via API — create and approve a vacation.",
            "steps": "SETUP: Via API — create and approve a vacation\n1. Login as any user\n2. Navigate to Availability chart (/vacation/chart)\n3. Search for the employee by name\n4. Navigate to the month containing the vacation\n5. Verify green bar appears for the APPROVED vacation dates\n6. Verify weekend columns have yellow background\nCLEANUP: Via API — cancel and delete the vacation",
            "expected": "Availability chart shows APPROVED vacations as green bars. Days-off/holidays shown in blue.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Availability Chart", "module": "vacation/filters",
            "notes": "Chart has Day/Month view toggle and timeline navigation."
        },
        {
            "id": "TC-VAC-055", "title": "Employees Vacation Days page — search by name",
            "preconditions": "Any logged-in user with access to vacation days page.",
            "steps": "1. Login as any user with appropriate role\n2. Navigate to /vacation/vacation-days\n3. In the search box, type an employee's first name\n4. Verify results filter to matching employees\n5. Verify each result shows: Employee (link), Vacation days, Pending approval\n6. Click employee link — verify navigation to their vacation detail",
            "expected": "Search filters employees by name. Results show vacation day balances and pending approval counts.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-pages.md §Employees Vacation Days", "module": "vacation/filters",
            "notes": ""
        },
        {
            "id": "TC-VAC-056", "title": "Latin name search bug (#3297)",
            "preconditions": "Employee with Latin first/last name in DB.\nQuery: SELECT e.login, e.latin_first_name, e.latin_last_name FROM ttt_backend.employee e WHERE e.latin_first_name IS NOT NULL AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as admin or manager with vacation days access\n2. Navigate to /vacation/vacation-days\n3. In search box, type the employee's Latin last name\n4. Verify: does the employee appear in results?\n5. If not found — bug #3297 confirmed (Latin name search broken)\n6. Try Cyrillic name for the same employee — verify it works",
            "expected": "Bug #3297 (OPEN): Latin name search is broken on Employee Vacation Days page. Cyrillic search works. Latin search returns no results.",
            "priority": "High", "type": "UI",
            "req_ref": "#3297, vacation-business-rules-reference.md §10 J1", "module": "vacation/filters",
            "notes": "DB columns: latin_first_name, latin_last_name (NOT first_name, last_name)."
        },
    ]


def get_av_cases():
    """TS-Vacation-AV: Advance Vacation (AV=true vs AV=false) specific behavior."""
    return [
        {
            "id": "TC-VAC-057", "title": "AV=true — full year balance available from Jan 1",
            "preconditions": "Employee in AV=true office (e.g., Cyprus, Germany).\nQuery: SELECT e.login, ev.available_vacation_days, ev.year FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE o.advance_vacation = true AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the AV=true employee\n2. Navigate to /vacation/my\n3. Check 'Available vacation days' counter\n4. Verify the count reflects full annual norm (e.g., 24 days) even early in the year\n5. Click info icon for per-year breakdown\n6. Verify current year shows full allocation, not prorated\nDB-CHECK: SELECT available_vacation_days FROM ttt_vacation.employee_vacation WHERE employee = <id> AND year = EXTRACT(YEAR FROM CURRENT_DATE)",
            "expected": "AV=true employees see full year balance from January. Not monthly-prorated like AV=false.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §1 AV=true", "module": "vacation/av-mode",
            "notes": "Formula: available = currentYearDays + pastYearDays + futureDays + editedDays."
        },
        {
            "id": "TC-VAC-058", "title": "AV=true — negative balance allowed for current year",
            "preconditions": "AV=true employee with enough balance to demonstrate negative scenario. Needs to create vacations consuming more than available.\nNote: Complex test — requires consuming entire balance plus more.",
            "steps": "1. Login as AV=true employee\n2. Navigate to /vacation/my\n3. Check current available days (e.g., 24)\n4. Create multiple vacations consuming all days\n5. Attempt to create one more vacation (would go negative)\n6. Verify: system allows creation (AV=true permits negative for current year)\n7. Verify available days counter shows negative value\nCLEANUP: Via API — DELETE all created vacations",
            "expected": "AV=true offices allow negative balance for current year. Counter can show negative values.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §1", "module": "vacation/av-mode",
            "notes": "AV=false: NEVER negative (display 0). AV=true: negative allowed for current year."
        },
        {
            "id": "TC-VAC-059", "title": "AV=false — monthly accrual, no negative",
            "preconditions": "Employee in AV=false office (Russia).\nQuery: SELECT e.login, ev.available_vacation_days FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE o.advance_vacation = false AND ev.year = EXTRACT(YEAR FROM CURRENT_DATE) AND ev.available_vacation_days > 0 AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the AV=false employee\n2. Navigate to /vacation/my\n3. Check 'Available vacation days'\n4. Verify the count is proportional to months elapsed (roughly month/12 * annual_norm)\n5. Verify the counter never shows a negative value\n6. Attempt to create a vacation exceeding available days\n7. Verify error blocks creation",
            "expected": "AV=false employee balance is monthly accrual. Counter never negative. Excess creation blocked.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §1 AV=false formula", "module": "vacation/av-mode",
            "notes": "Formula: available = (month × norm/12) + yearRemainder + priorYears − norm + futureDays + editedDays."
        },
        {
            "id": "TC-VAC-060", "title": "FIFO day consumption — earliest year first",
            "preconditions": "Employee with vacation days from multiple years (current + prior).\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE ev.available_vacation_days > 0 AND e.enabled = true GROUP BY e.login HAVING COUNT(DISTINCT ev.year) > 1 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click info icon on available days — note per-year breakdown\n4. Create a REGULAR vacation consuming some days\n5. After creation, click info icon again\n6. Verify earliest year's balance decreased first (FIFO order)\nDB-CHECK: SELECT year, available_vacation_days FROM ttt_vacation.employee_vacation WHERE employee = <id> ORDER BY year\nCLEANUP: Via API — DELETE",
            "expected": "FIFO consumption: days consumed from earliest year first. Per-year breakdown reflects this.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §4 FIFO", "module": "vacation/av-mode",
            "notes": "Cross-year splits tracked in vacation_days_distribution (vacation_id + year → days)."
        },
        {
            "id": "TC-VAC-061", "title": "FIFO redistribution on cancel — days returned",
            "preconditions": "Employee with an APPROVED vacation consuming days from multiple years.\nSETUP: Via API — create and approve a vacation.",
            "steps": "SETUP: Via API — create a long vacation consuming days from 2+ years, approve it\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click info icon — note per-year breakdown (reduced by vacation)\n4. Cancel the vacation\n5. Click info icon again\n6. Verify days returned to the correct years (FIFO redistribution)\nDB-CHECK: Compare employee_vacation.available_vacation_days before and after cancel",
            "expected": "On cancel, days redistributed back. VacationRecalculationService returns all days then re-distributes among remaining vacations.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §4 Recalculation", "module": "vacation/av-mode",
            "notes": "Recalculation: returns ALL regular+exact days, then re-distributes. May convert if insufficient."
        },
        {
            "id": "TC-VAC-062", "title": "AV=true multi-year balance distribution (#3361)",
            "preconditions": "AV=true employee with vacation days across 3+ years.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id JOIN ttt_vacation.employee_vacation ev ON e.id = ev.employee WHERE o.advance_vacation = true AND e.enabled = true GROUP BY e.login HAVING COUNT(DISTINCT ev.year) >= 3 ORDER BY random() LIMIT 1",
            "steps": "1. Login as the employee\n2. Navigate to /vacation/my\n3. Click info icon for per-year breakdown\n4. Verify all years with remaining balance are shown\n5. Attempt to create a vacation consuming days from current year\n6. Verify: does creation succeed? (Bug #3361: multi-year distribution may block creation)\nCLEANUP: Via API — DELETE if created",
            "expected": "Bug #3361 (OPEN): AV=True multi-year balance distribution blocks current-year creation even when overall balance is sufficient.",
            "priority": "High", "type": "UI",
            "req_ref": "#3361, vacation-business-rules-reference.md §10 A1", "module": "vacation/av-mode",
            "notes": "OPEN bug. Regression test."
        },
        {
            "id": "TC-VAC-063", "title": "Day correction — AV=false prohibits negative",
            "preconditions": "Accountant or admin user. AV=false employee.\nQuery: SELECT e.login FROM ttt_vacation.employee e JOIN ttt_vacation.office o ON e.office_id = o.id WHERE o.advance_vacation = false AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as accountant/admin\n2. Navigate to Employees Vacation Days page\n3. Find the AV=false employee\n4. Attempt to manually set vacation days to a negative value\n5. Verify: correction is blocked (value must be >= 0 for AV=false)",
            "expected": "AV=false: manual corrections floor at >= 0. Negative values prohibited in inline editing.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-business-rules-reference.md §8, #3283", "module": "vacation/av-mode",
            "notes": "Bug #3283: auto corrections can go negative even for AV=false. Manual edits correctly floor at 0."
        },
    ]


def get_notification_cases():
    """TS-Vacation-Notifications: Email notification triggers."""
    return [
        {
            "id": "TC-VAC-064", "title": "Create vacation → notification to approver",
            "preconditions": "Employee with a manager. Email service accessible.\nSETUP: Clear email inbox for the manager.",
            "steps": "SETUP: Via test API — clear emails for the manager (POST /api/email/test/v1/emails/delete)\n1. Login as the employee\n2. Create a REGULAR vacation via UI\n3. Wait 30 seconds for async email processing\n4. Via API — GET /api/email/v1/emails?to=<manager_email>\n5. Verify email sent to manager about new vacation request\n6. Verify email contains: employee name, vacation dates, vacation type\nCLEANUP: Via API — DELETE the vacation",
            "expected": "Manager receives email notification about new vacation request with details.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "exploration/api-findings/vacation-notification-templates.md", "module": "vacation/notifications",
            "notes": "Email templates catalog: 50+ templates across 10 categories."
        },
        {
            "id": "TC-VAC-065", "title": "Approve vacation → notification to employee",
            "preconditions": "NEW vacation. Email service accessible.\nSETUP: Create vacation, clear employee's email.",
            "steps": "SETUP: Via API — create vacation, clear employee's emails\n1. Login as the manager\n2. Approve the vacation via Employee Requests page\n3. Wait 30 seconds\n4. Via API — check employee's email inbox\n5. Verify approval notification received\n6. Verify email contains: vacation dates, approved status, approver name\nCLEANUP: Via API — cancel and delete",
            "expected": "Employee receives email notification about vacation approval.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "exploration/api-findings/vacation-notification-templates.md", "module": "vacation/notifications",
            "notes": ""
        },
        {
            "id": "TC-VAC-066", "title": "Reject vacation → notification to employee",
            "preconditions": "NEW vacation. Email service accessible.",
            "steps": "SETUP: Via API — create vacation, clear employee's emails\n1. Login as manager\n2. Reject the vacation\n3. Via API — check employee's email\n4. Verify rejection notification received\nCLEANUP: Via API — DELETE",
            "expected": "Employee receives rejection notification with approver name and vacation details.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "exploration/api-findings/vacation-notification-templates.md", "module": "vacation/notifications",
            "notes": ""
        },
        {
            "id": "TC-VAC-067", "title": "Cancel vacation → notification to approver",
            "preconditions": "NEW or APPROVED vacation.",
            "steps": "SETUP: Via API — create vacation, clear approver's emails\n1. Login as the employee\n2. Cancel the vacation via UI\n3. Via API — check approver's email\n4. Verify cancellation notification sent to approver\nCLEANUP: None needed",
            "expected": "Approver notified when employee cancels their vacation.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "exploration/api-findings/vacation-notification-templates.md", "module": "vacation/notifications",
            "notes": ""
        },
        {
            "id": "TC-VAC-068", "title": "Also-notify recipients receive notification",
            "preconditions": "Employee with a colleague for also-notify.",
            "steps": "SETUP: Via API — clear colleague's emails\n1. Login as employee\n2. Create vacation with colleague in 'Also notify' field\n3. Via API — check colleague's email\n4. Verify notification sent to the also-notify recipient\nCLEANUP: Via API — DELETE vacation",
            "expected": "Colleagues in 'Also notify' list receive email about the vacation.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §4 notifyAlso", "module": "vacation/notifications",
            "notes": ""
        },
        {
            "id": "TC-VAC-069", "title": "Wrong payment month in notification (#2925)",
            "preconditions": "Vacation with specific payment month.",
            "steps": "SETUP: Via API — create vacation with explicit paymentMonth, clear approver's emails\n1. Login as employee, create vacation\n2. Via API — check notification email\n3. Verify payment month in email matches the vacation's actual payment month\n4. If mismatched — bug #2925 confirmed",
            "expected": "Bug #2925 (OPEN): Wrong payment month displayed in email notification.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "#2925, vacation-business-rules-reference.md §10 H3", "module": "vacation/notifications",
            "notes": "OPEN bug regression test."
        },
        {
            "id": "TC-VAC-070", "title": "Notification on auto-conversion to ADMINISTRATIVE (#3015)",
            "preconditions": "Requires a scenario where accrued days validation triggers auto-conversion.",
            "steps": "SETUP: Create multiple vacations that trigger auto-conversion of a future request\n1. After conversion triggers, check employee's email\n2. Verify notification about vacation type conversion\n3. Verify email identifies which vacation was converted and why\nCLEANUP: DELETE all vacations",
            "expected": "Employee notified when one of their vacations is auto-converted from Regular to Administrative due to accrued days validation.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "#3015, vacation-business-rules-reference.md §7", "module": "vacation/notifications",
            "notes": "Bug #3015-30: silent conversion (no notification) was fixed."
        },
    ]


def get_regression_cases():
    """TS-Vacation-Regression: Tests derived from 50+ bugs in ticket mining."""
    return [
        {
            "id": "TC-VAC-071", "title": "Regression: Overlapping vacations not blocked by frontend (#3240)",
            "preconditions": "Employee with an existing vacation.\nSETUP: Via API — create a vacation.",
            "steps": "SETUP: Via API — create a vacation for specific dates\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select dates that overlap with the existing vacation\n5. Verify frontend shows overlap error BEFORE clicking Save\n6. Verify the calendar picker or form indicates the conflict\nCLEANUP: Via API — DELETE the first vacation",
            "expected": "Frontend detects overlapping dates and shows error. Bug #3240 (fixed): previously frontend did not block overlapping vacations.",
            "priority": "High", "type": "UI",
            "req_ref": "#3240, vacation-business-rules-reference.md §10 C1", "module": "vacation/regression",
            "notes": "3-way overlap check: start-in-range, end-in-range, enclosing."
        },
        {
            "id": "TC-VAC-072", "title": "Regression: Payment month not updated in edit modal (#2705)",
            "preconditions": "Employee with a NEW vacation.",
            "steps": "SETUP: Via API — create a vacation with specific payment month\n1. Login as the employee\n2. Click edit on the vacation\n3. Change dates to a different month range\n4. Verify payment month auto-updates to reflect new dates\n5. Click Save\n6. Re-open edit dialog — verify payment month persisted correctly",
            "expected": "Payment month updates when dates change in edit dialog. Bug #2705 (fixed): previously payment month didn't update after save.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#2705, vacation-business-rules-reference.md §10 C3", "module": "vacation/regression",
            "notes": ""
        },
        {
            "id": "TC-VAC-073", "title": "Regression: Edit own vacation shows 0 available (#3014-21)",
            "preconditions": "Employee with a vacation consuming most of their balance.\nSETUP: Via API — create a large vacation.",
            "steps": "SETUP: Via API — create a vacation consuming 80% of available days\n1. Login as the employee\n2. Click edit on the vacation\n3. Verify available days shown in edit dialog EXCLUDE the current vacation's days from the consumed total\n4. Verify the available counter shows original balance minus OTHER vacations only\n5. Verify Save button is enabled (not disabled due to false 0-available calculation)",
            "expected": "Edit dialog correctly excludes the current vacation's days from consumed total. Bug #3014-21 (fixed): previously showed 0 available.",
            "priority": "High", "type": "UI",
            "req_ref": "#3014 Bug 21, vacation-business-rules-reference.md §10 E1", "module": "vacation/regression",
            "notes": "VacationUpdateValidator adjusts limitations: adds back current vacation's consumed days."
        },
        {
            "id": "TC-VAC-074", "title": "Regression: Redirected request status not reset (#2718)",
            "preconditions": "APPROVED vacation with a specific approver.\nSETUP: Via API — create, approve.",
            "steps": "SETUP: Via API — create vacation, approve it\n1. Login as the approver\n2. Redirect the APPROVED vacation to another manager\n3. Verify: does the status reset to NEW? (Expected per business rules)\n4. If status remains APPROVED — bug #2718 confirmed (OPEN)\nDB-CHECK: SELECT status FROM ttt_vacation.vacation WHERE id = <id>\nCLEANUP: Via API — cancel and delete",
            "expected": "Bug #2718 (OPEN): Redirected approved/rejected request doesn't reset status to NEW. Status should reset when approver changes.",
            "priority": "High", "type": "UI",
            "req_ref": "#2718, vacation-business-rules-reference.md §10 B2", "module": "vacation/regression",
            "notes": "OPEN bug. 21% of vacations get redirected — high-impact issue."
        },
        {
            "id": "TC-VAC-075", "title": "Regression: Double accrual on salary office change (#2789)",
            "preconditions": "Employee who has recently changed salary office.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.enabled = true AND EXISTS (SELECT 1 FROM ttt_vacation.employee_vacation ev WHERE ev.employee = e.id GROUP BY ev.employee HAVING COUNT(*) > (SELECT COUNT(DISTINCT year) FROM ttt_vacation.employee_vacation WHERE employee = e.id)) ORDER BY random() LIMIT 1",
            "steps": "1. Via DB — check employee_vacation records for the employee\n2. Verify no duplicate year entries (same employee + same year appearing twice)\n3. If duplicates found — bug #2789 active\n4. Calculate total available days and compare with expected (annual_norm * years_worked)\nDB-CHECK: SELECT employee, year, COUNT(*) FROM ttt_vacation.employee_vacation WHERE employee = <id> GROUP BY employee, year HAVING COUNT(*) > 1",
            "expected": "Bug #2789 (OPEN): Double accrual on salary office change creates duplicate year entries.",
            "priority": "High", "type": "API",
            "req_ref": "#2789, vacation-business-rules-reference.md §10 A7", "module": "vacation/regression",
            "notes": "OPEN bug. Data integrity issue."
        },
        {
            "id": "TC-VAC-076", "title": "Regression: last_date not updated during CS sync (#3374)",
            "preconditions": "Employee with last_date set (dismissed or leaving).\nQuery: SELECT e.login, e.last_date FROM ttt_vacation.employee e WHERE e.last_date IS NOT NULL ORDER BY random() LIMIT 1",
            "steps": "1. Via API — trigger CS sync: POST /api/ttt/test/v1/employees/sync\n2. Wait for sync completion\n3. Via DB — check if last_date values match CompanyStaff source\nDB-CHECK: SELECT login, last_date, enabled FROM ttt_vacation.employee WHERE last_date IS NOT NULL",
            "expected": "Bug #3374 (OPEN): last_date not updated during CS sync — TTT may miss dismissal dates.",
            "priority": "Medium", "type": "API",
            "req_ref": "#3374, vacation-business-rules-reference.md §10 I5", "module": "vacation/regression",
            "notes": "OPEN bug. Affects vacation availability calculations for departing employees."
        },
        {
            "id": "TC-VAC-077", "title": "Regression: Maternity leave overlap — days not returned (#3352)",
            "preconditions": "Employee on maternity leave with existing vacations.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.maternity_leave = true AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Find employee on maternity leave\n2. Check their vacation day balance\n3. Verify days from overlapping periods are correctly handled\n4. If balance is incorrect (days not returned) — bug #3352 confirmed\nDB-CHECK: SELECT available_vacation_days, year FROM ttt_vacation.employee_vacation WHERE employee = <id>",
            "expected": "Bug #3352 (OPEN): Maternity leave overlap → vacation days not returned to balance.",
            "priority": "High", "type": "API",
            "req_ref": "#3352, vacation-business-rules-reference.md §10 A4", "module": "vacation/regression",
            "notes": "OPEN bug. Related to #3370 (maternity user can't edit vacation)."
        },
        {
            "id": "TC-VAC-078", "title": "Regression: Maternity leave user can't edit vacation (#3370)",
            "preconditions": "Employee on maternity leave with available vacation days = 0.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.maternity_leave = true AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the maternity leave employee\n2. Navigate to /vacation/my\n3. Verify available days shows 0\n4. Attempt to create or edit a vacation\n5. Verify: is the action blocked? (Bug #3370 expected)\nCLEANUP: None",
            "expected": "Bug #3370 (OPEN): Maternity leave user shows 0 available days and can't edit existing vacations.",
            "priority": "Medium", "type": "UI",
            "req_ref": "#3370, vacation-business-rules-reference.md §10 J2", "module": "vacation/regression",
            "notes": "OPEN bug. Related to #3352."
        },
        {
            "id": "TC-VAC-079", "title": "Regression: Ghost conflicts from soft-deleted vacations",
            "preconditions": "Employee with a deleted vacation at specific dates.\nSETUP: Via API — create a vacation, delete it.",
            "steps": "SETUP: Via API — create vacation for Week X dates, then DELETE it\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Select the SAME dates as the deleted vacation (Week X)\n5. Verify: does crossing validation block the creation?\n6. If blocked — ghost conflict confirmed (DELETED records counted in crossing check)\nCLEANUP: None needed",
            "expected": "Design issue: crossing check includes DELETED records. Soft-deleted vacations create permanent ghost conflicts that block future creates at those dates.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Session 86 Crossing Check", "module": "vacation/regression",
            "notes": "DELETED should be excluded from crossing validation. Every test run that creates+deletes a vacation permanently blocks that date range."
        },
        {
            "id": "TC-VAC-080", "title": "Regression: Approver field missing from API (#3329)",
            "preconditions": "Multiple vacations.",
            "steps": "1. Via API — GET /api/vacation/v1/vacations (list endpoint)\n2. Check each vacation in the response\n3. Verify all vacations have 'approver' field populated\n4. If any have null/missing approver — bug #3329\nDB-CHECK: SELECT COUNT(*) FROM ttt_vacation.vacation WHERE approver_id IS NULL AND status NOT IN ('DELETED', 'CANCELED')",
            "expected": "Bug #3329 (fixed as hotfix): Approver field was intermittently missing from API response. Verify fix is stable.",
            "priority": "Medium", "type": "API",
            "req_ref": "#3329, vacation-business-rules-reference.md §10 I4", "module": "vacation/regression",
            "notes": "Hotfix applied. Regression test to ensure it doesn't recur."
        },
        {
            "id": "TC-VAC-081", "title": "Regression: Flash of irrelevant validation on first date pick (#3127)",
            "preconditions": "Any employee.",
            "steps": "1. Login as employee\n2. Navigate to /vacation/my\n3. Click 'Create a request'\n4. Click on the start date field for the very first time\n5. Select a date\n6. Verify: does a flash of irrelevant validation error appear briefly?\n7. If yes — bug #3127 still present (was marked as closed but may regress)",
            "expected": "Bug #3127 (closed): Flash of validation message should not appear on first date selection.",
            "priority": "Low", "type": "UI",
            "req_ref": "#3127, vacation-business-rules-reference.md §10 C5", "module": "vacation/regression",
            "notes": "Low severity but annoying UX issue."
        },
        {
            "id": "TC-VAC-082", "title": "Regression: Russian messages in English events feed (#3344)",
            "preconditions": "Employee with vacation events. UI set to English.",
            "steps": "1. Login as employee\n2. Ensure UI language is English (check language switcher in header)\n3. Navigate to /vacation/my\n4. Click 'Vacation events feed'\n5. Verify all event descriptions are in English\n6. If any Russian text appears — bug #3344 not fully fixed",
            "expected": "Bug #3344 (closed): Events feed should show English messages when UI is set to English.",
            "priority": "Low", "type": "UI",
            "req_ref": "#3344, vacation-business-rules-reference.md §10 H2", "module": "vacation/regression",
            "notes": ""
        },
        {
            "id": "TC-VAC-083", "title": "Regression: Null optionalApprovers → NPE on CPO path",
            "preconditions": "CPO employee (ROLE_DEPARTMENT_MANAGER) with manager.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with login=<CPO>, valid dates, paymentType=REGULAR, paymentMonth=valid, optionalApprovers=null\n2. Verify HTTP 500 (NPE)\n3. The CPO create path calls request.getOptionalApprovers().add(manager) which NPEs if null\nDB-CHECK: Verify no vacation was created",
            "expected": "Bug: null optionalApprovers causes NPE on CPO path. HTTP 500.",
            "priority": "High", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §7 NPE #2", "module": "vacation/regression",
            "notes": "Confirmed active on qa-1. No @NotNull on optionalApprovers field."
        },
        {
            "id": "TC-VAC-084", "title": "Regression: Calendar change converts ALL vacations (#3338)",
            "preconditions": "Requires production calendar modification capability. Admin access needed.",
            "steps": "1. Find employee with multiple approved vacations\n2. Admin modifies production calendar (add a holiday within one vacation's dates)\n3. Verify: does only the affected vacation get checked for conversion?\n4. Or does every vacation for the employee get reconverted? (Bug #3338)\nDB-CHECK: Check status changes across all employee's vacations after calendar change",
            "expected": "Bug #3338 (closed): Calendar change should only affect the vacation containing the changed date, not all vacations.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "#3338, vacation-business-rules-reference.md §10 G area", "module": "vacation/regression",
            "notes": "Regression test for calendar-vacation interaction."
        },
    ]


def get_permission_cases():
    """TS-Vacation-Permissions: Role-based access control tests."""
    return [
        {
            "id": "TC-VAC-085", "title": "Owner can edit own vacation (not PAID)",
            "preconditions": "Employee with own NEW vacation.\nSETUP: Via API — create a vacation.",
            "steps": "SETUP: Via API — create a vacation\n1. Login as the vacation owner\n2. Navigate to /vacation/my\n3. Verify edit (pencil) icon is visible on the NEW vacation row\n4. Click edit\n5. Change end date\n6. Click Save — verify success\nCLEANUP: Via API — DELETE",
            "expected": "Owner has EDIT permission for own vacation in non-PAID status.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §3 Permission Table", "module": "vacation/permissions",
            "notes": "Permission EDIT: Owner, status not PAID (always)."
        },
        {
            "id": "TC-VAC-086", "title": "Owner cannot edit PAID vacation",
            "preconditions": "PAID vacation.\nSETUP: Via API — create, approve, pay.",
            "steps": "SETUP: Via API — create, approve, and pay a vacation\n1. Login as the vacation owner\n2. Navigate to /vacation/my\n3. Find the PAID vacation in 'Closed' tab\n4. Verify: edit (pencil) icon is NOT visible\n5. Via API — PUT /vacations/{id} with updated dates → verify HTTP 400\nNote: PAID+EXACT creates permanent record",
            "expected": "PAID vacation cannot be edited. No edit icon shown. API returns 400 on update attempt.",
            "priority": "High", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §3 NON_EDITABLE_STATUSES", "module": "vacation/permissions",
            "notes": "NON_EDITABLE_STATUSES = {CANCELED, PAID}."
        },
        {
            "id": "TC-VAC-087", "title": "Non-approver cannot approve vacation",
            "preconditions": "Vacation assigned to Manager A. Login as Manager B (different person).",
            "steps": "SETUP: Via API — create a vacation assigned to Manager A\n1. Login as Manager B (NOT the assigned approver)\n2. Navigate to /vacation/request\n3. Verify the vacation does NOT appear in Manager B's approval queue\n4. Via API — PUT /approve/{id} as Manager B → verify HTTP 400 (exception.vacation.status.notAllowed)\nCLEANUP: Via API — DELETE",
            "expected": "Non-assigned managers cannot approve. hasAccess() checks approver_id match. 400 error returned.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §1 hasAccess", "module": "vacation/permissions",
            "notes": "MANAGER_ROLES = {PM, DM, CACC}. ROLE_ADMIN is NOT in MANAGER_ROLES (design issue)."
        },
        {
            "id": "TC-VAC-088", "title": "ReadOnly user cannot create vacation",
            "preconditions": "Employee with readOnly=true.\nQuery: SELECT e.login FROM ttt_vacation.employee e WHERE e.read_only = true AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "1. Login as the readOnly employee\n2. Navigate to /vacation/my\n3. Verify 'Create a request' button is either hidden or disabled\n4. Via API — POST /api/vacation/v1/vacations — verify rejection\nCLEANUP: None",
            "expected": "ReadOnly users cannot create vacations. Permission service returns empty set for readOnly users.",
            "priority": "Medium", "type": "UI",
            "req_ref": "vacation-service-deep-dive.md §3 Permission GUARD", "module": "vacation/permissions",
            "notes": "checkVacation(): if readOnly → throw VacationSecurityException (403)."
        },
        {
            "id": "TC-VAC-089", "title": "Accountant can pay but not approve",
            "preconditions": "Accountant user.\nQuery: SELECT e.login FROM ttt_backend.employee be JOIN ttt_backend.employee_role er ON be.id = er.employee_id JOIN ttt_backend.role r ON er.role_id = r.id JOIN ttt_vacation.employee e ON e.login = be.login WHERE r.name = 'ROLE_ACCOUNTANT' AND e.enabled = true ORDER BY random() LIMIT 1",
            "steps": "SETUP: Via API — create and approve a vacation\n1. Login as the accountant\n2. Navigate to vacation payment section\n3. Verify payment controls are available for APPROVED vacations\n4. Via API — PUT /approve/{id} as accountant → verify this fails (accountant not in approval roles)\n5. Via API — PUT /pay/{id} → verify this succeeds",
            "expected": "Accountant has PAY permission but NOT APPROVE. Can process payments but cannot change approval status.",
            "priority": "Medium", "type": "Hybrid",
            "req_ref": "vacation-business-rules-reference.md §9 Permission Matrix", "module": "vacation/permissions",
            "notes": "ROLE_ACCOUNTANT is in ACCOUNTANT_ROLES for payment, but NOT in MANAGER_ROLES for approval."
        },
        {
            "id": "TC-VAC-090", "title": "canBeCancelled guard — REGULAR+APPROVED after period close",
            "preconditions": "REGULAR + APPROVED vacation with paymentDate before current report period.\nNote: Requires timemachine env to manipulate period timing.",
            "steps": "SETUP: Via timemachine — create vacation, approve, then advance clock past paymentDate\n1. Login as the employee\n2. Navigate to /vacation/my\n3. Find the APPROVED vacation\n4. Verify: cancel/delete buttons are NOT available\n5. Via API — PUT /cancel/{id} → verify HTTP 400\nCLEANUP: Reset clock",
            "expected": "canBeCancelled returns false when: REGULAR + APPROVED + reportPeriod after paymentDate. Cancel/reject/delete all blocked.",
            "priority": "High", "type": "Hybrid",
            "req_ref": "vacation-service-deep-dive.md §3 canBeCancelled", "module": "vacation/permissions",
            "notes": "Protects against canceling after accounting period closure. Needs timemachine env."
        },
    ]


def get_api_cases():
    """TS-Vacation-API: API-only error handling and edge cases."""
    return [
        {
            "id": "TC-VAC-091", "title": "Empty request body → empty 400 response",
            "preconditions": "API access.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with empty body (no JSON)\n2. Verify HTTP 400\n3. Verify response body is EMPTY (no JSON error details)\n4. HttpMessageNotReadableException returns ResponseEntity<Void>",
            "expected": "Empty body returns HTTP 400 with completely empty response body. No error details provided — unique behavior.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 HttpMessageNotReadableException", "module": "vacation/api",
            "notes": "Design issue: should return structured error response."
        },
        {
            "id": "TC-VAC-092", "title": "Invalid type parameter → type mismatch error",
            "preconditions": "API access.",
            "steps": "1. Via API — GET /api/vacation/v1/vacations/abc (string instead of Long ID)\n2. Verify HTTP 400\n3. Verify errorCode: 'exception.type.mismatch'\n4. Verify message includes expected type 'Long'",
            "expected": "Type mismatch returns 400 with exception.type.mismatch error code.",
            "priority": "Low", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 MethodArgumentTypeMismatchException", "module": "vacation/api",
            "notes": ""
        },
        {
            "id": "TC-VAC-093", "title": "Missing required fields → validation errors array",
            "preconditions": "API access.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with body: {} (empty JSON object)\n2. Verify HTTP 400\n3. Verify response contains 'errors' array\n4. Verify each error has: field, code, message\n5. Verify errors for: login (NotNull), startDate (NotNull), endDate (NotNull), paymentType (NotNull)",
            "expected": "Multiple validation errors returned in errors[] array. Each with field name, code, and message.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 MethodArgumentNotValidException", "module": "vacation/api",
            "notes": ""
        },
        {
            "id": "TC-VAC-094", "title": "Exception class leakage in error responses",
            "preconditions": "API access.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with past startDate\n2. Verify HTTP 400 response\n3. Check 'exception' field in response\n4. Verify it contains full Java class name: 'com.noveogroup.ttt.common.exception.ServiceException'\n5. This leaks internal implementation details",
            "expected": "Security issue: error responses include 'exception' field with full Java class name. Information disclosure vulnerability.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 INFO DISCLOSURE", "module": "vacation/api",
            "notes": "All error responses (except empty-body) include this field. OWASP information disclosure."
        },
        {
            "id": "TC-VAC-095", "title": "Update without id in body → IllegalArgumentException",
            "preconditions": "Existing vacation.\nSETUP: Via API — create a vacation.",
            "steps": "SETUP: Via API — create a vacation, note the ID\n1. Via API — PUT /api/vacation/v1/vacations/{id} with body missing 'id' field\n2. Verify HTTP 400\n3. Verify error: 'IllegalArgumentException: The given id must not be null!'\nCLEANUP: Via API — DELETE",
            "expected": "Missing id in update body causes IllegalArgumentException. JPA findById called with null from DTO.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Update endpoint", "module": "vacation/api",
            "notes": "id required in BOTH URL path AND request body."
        },
        {
            "id": "TC-VAC-096", "title": "Crossing validation error format inconsistency",
            "preconditions": "Two vacations with overlapping dates.\nSETUP: Via API — create first vacation.",
            "steps": "SETUP: Via API — create a vacation\n1. Via API — POST /api/vacation/v1/vacations with overlapping dates (create endpoint)\n2. Check error response format: errorCode, message, errors[].code, errors[].message\n3. Note which field contains the specific crossing error code\n4. Via API — PUT /vacations/{id} with overlapping dates (update endpoint)\n5. Compare error format between create and update\nCLEANUP: Via API — DELETE",
            "expected": "Inconsistency: create uses errors[].code = specific code. Update uses errors[].code = 'exception.validation.fail' with specific code in errors[].message.",
            "priority": "Low", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Session 91 Crossing format", "module": "vacation/api",
            "notes": "Always check both errorCode and message fields for reliable assertion."
        },
        {
            "id": "TC-VAC-097", "title": "Sick leave crossing vacation → 409 CONFLICT",
            "preconditions": "Employee with existing vacation. Ability to create sick leave.\nNote: Blocked by API_SECRET_TOKEN — sick leave endpoint requires AUTHENTICATED_USER authority.",
            "steps": "1. Via API — create a vacation for specific dates\n2. Via API — POST /api/vacation/v1/sick-leaves with dates overlapping the vacation\n3. Verify HTTP 409 CONFLICT\n4. Verify SickLeaveCrossingVacationException in response\nNote: May need per-user CAS auth — API_SECRET_TOKEN returns 403 for sick leave endpoint\nCLEANUP: Via API — DELETE vacation",
            "expected": "Sick leave overlapping a vacation returns 409 CONFLICT. SickLeaveCrossingVacationException thrown.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 409, §Autotest Notes Session 98", "module": "vacation/api",
            "notes": "BLOCKED: API_SECRET_TOKEN cannot access sick leave endpoint (403). Needs CAS JWT."
        },
        {
            "id": "TC-VAC-098", "title": "Non-existent vacation ID → 404",
            "preconditions": "API access.",
            "steps": "1. Via API — GET /api/vacation/v1/vacations/999999999\n2. Verify HTTP 404\n3. Verify errorCode: 'exception.not.found'",
            "expected": "Non-existent vacation returns 404 with exception.not.found error code.",
            "priority": "Low", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §5 EntityNotFoundException", "module": "vacation/api",
            "notes": ""
        },
        {
            "id": "TC-VAC-099", "title": "Invalid notifyAlso login → 400",
            "preconditions": "API access.",
            "steps": "1. Via API — POST /api/vacation/v1/vacations with valid fields but notifyAlso: ['nonexistent_user']\n2. Verify HTTP 400\n3. Verify validation error for notifyAlso field (EmployeeLoginCollectionExists)\nCLEANUP: None — creation should fail",
            "expected": "@EmployeeLoginCollectionExists validates all logins exist. Invalid login returns 400.",
            "priority": "Low", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §4 DTO Annotations", "module": "vacation/api",
            "notes": ""
        },
        {
            "id": "TC-VAC-100", "title": "Batch deadlock on concurrent operations",
            "preconditions": "Ability to send concurrent API requests.",
            "steps": "1. Via API — simultaneously send 3 vacation create requests for the same employee\n2. Observe responses — at least 1 should succeed\n3. Check for 500 errors with 'deadlock detected' message\n4. Verify VacationRecalculationService row contention triggers PostgreSQL deadlock",
            "expected": "Concurrent vacation operations may cause PostgreSQL deadlocks. CannotAcquireLockException thrown. Not all requests succeed.",
            "priority": "Medium", "type": "API",
            "req_ref": "vacation-service-deep-dive.md §Autotest Notes Session 86 Batch Deadlocks", "module": "vacation/api",
            "notes": "Root: employee_vacation table row contention during FIFO redistribution. Mitigation: 2+ second gaps."
        },
    ]


# ─── Plan, Feature Matrix, Risk Assessment ────────────────────────────────────

PLAN_OVERVIEW = {
    "title": "Vacation Module — Test Plan",
    "scope": "Comprehensive testing of the vacation management module covering CRUD operations, approval workflows, payment processing, validation rules, filtering/search, advance vacation modes, notifications, role-based permissions, API error handling, and regression tests for 50+ known bugs.",
    "objectives": [
        "Verify all vacation lifecycle operations (create, edit, cancel, delete, re-open)",
        "Validate the state machine (NEW → APPROVED → PAID, CANCELED → NEW, etc.)",
        "Test approval workflow: primary approver, optional approvers, CPO self-approval, change approver",
        "Verify payment processing: validation, terminal state, auto-pay",
        "Test all validation rules: dates, balance, crossing, duration, 3-month restriction, accrued days",
        "Verify AV=true vs AV=false calculation mode differences",
        "Test email notification triggers for all status changes",
        "Verify role-based permissions: owner, approver, accountant, admin, readOnly",
        "Regression tests for all 11 OPEN bugs and key closed bugs",
        "API error handling: empty body, type mismatch, validation arrays, info disclosure",
    ],
    "environments": [
        "Primary: qa-1 (ttt-qa-1.noveogroup.com)",
        "Secondary: timemachine (ttt-timemachine.noveogroup.com) — for clock-dependent tests",
        "Production-like: stage (ttt-stage.noveogroup.com) — for verification only",
    ],
    "approach": "UI-first testing with API setup/cleanup. Tests describe browser actions for all user-facing scenarios. API used for state creation (SETUP steps), data verification (DB-CHECK), and API-only features. Preconditions include SQL query hints for dynamic test data selection.",
    "dependencies": [
        "CAS authentication for multi-user tests (some tests blocked by single API_SECRET_TOKEN)",
        "Timemachine environment for clock-dependent tests (auto-pay, canBeCancelled guard, Feb 1 cutoff)",
        "Email service for notification tests",
        "Production calendar data for holiday-dependent tests",
    ],
}

FEATURE_MATRIX = [
    # (Feature, CRUD, Approve, Payment, Validation, Filters, AV, Notif, Regression, Perms, API, Total)
    ("Create vacation (Regular)", 2, 0, 0, 0, 0, 0, 1, 0, 0, 0, 3),
    ("Create vacation (Admin)", 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("Edit vacation", 2, 0, 0, 1, 0, 0, 0, 2, 1, 1, 7),
    ("Cancel/Delete vacation", 3, 0, 0, 0, 0, 0, 1, 1, 0, 0, 5),
    ("Re-open canceled", 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1),
    ("View details", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2),
    ("Available days", 1, 0, 0, 0, 0, 3, 0, 0, 0, 0, 4),
    ("Approve/Reject", 0, 4, 0, 0, 0, 0, 1, 1, 1, 0, 7),
    ("CPO self-approval", 0, 1, 0, 0, 0, 0, 0, 1, 0, 0, 2),
    ("Change approver", 0, 1, 0, 0, 0, 0, 0, 1, 0, 0, 2),
    ("Optional approvers", 0, 2, 0, 0, 0, 0, 1, 0, 0, 0, 3),
    ("Payment (pay)", 0, 0, 4, 0, 0, 0, 0, 0, 1, 0, 5),
    ("Payment validation", 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 3),
    ("Terminal PAID state", 0, 0, 2, 0, 0, 0, 0, 0, 1, 0, 3),
    ("Date validation", 0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 3),
    ("Balance validation", 0, 0, 0, 2, 0, 1, 0, 0, 0, 0, 3),
    ("Crossing validation", 0, 0, 0, 1, 0, 0, 0, 2, 0, 1, 4),
    ("3-month restriction", 0, 0, 0, 2, 0, 0, 0, 0, 0, 0, 2),
    ("Accrued days (#3015)", 0, 0, 0, 1, 0, 0, 1, 0, 0, 0, 2),
    ("Tab/column filters", 0, 0, 0, 0, 5, 0, 0, 0, 0, 0, 5),
    ("Sort/search", 0, 0, 0, 0, 4, 0, 0, 0, 0, 0, 4),
    ("FIFO consumption", 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 2),
    ("Notifications", 0, 0, 0, 0, 0, 0, 4, 0, 0, 0, 4),
    ("RBAC/permissions", 0, 0, 0, 0, 0, 0, 0, 0, 4, 0, 4),
    ("API error handling", 0, 0, 0, 0, 0, 0, 0, 0, 0, 7, 7),
    ("Known bugs (OPEN)", 0, 0, 1, 1, 1, 1, 1, 7, 0, 1, 13),
    ("Known bugs (closed)", 0, 0, 0, 0, 0, 0, 0, 5, 0, 0, 5),
]

RISK_ASSESSMENT = [
    ("FIFO Day Redistribution", "Incorrect balance after cancel/reject due to FIFO miscalculation", "High", "High", "Critical", "Extensive balance verification before/after each transition. Cross-year distribution tests."),
    ("Accrued Days Validation (#3015)", "Auto-conversion of wrong vacation, silent conversion, conversion order errors", "High", "High", "Critical", "31 sub-bugs found. Test multiple future requests, conversion order, notification of conversion."),
    ("Payment Month Validation", "Wrong period used (Report vs Approval), closed period bypass", "Medium", "High", "High", "Test boundary conditions around period closure. Verify hotfix #3379."),
    ("Crossing Check Ghost Conflicts", "Soft-deleted vacations permanently block date ranges", "High", "High", "Critical", "Create+delete+re-create at same dates. Design issue — no fix available."),
    ("Multi-Year AV=True Balance (#3361)", "Current-year creation blocked despite sufficient overall balance", "Medium", "High", "High", "OPEN bug. Test with 3+ years of balance history."),
    ("Maternity Leave Interactions", "Days not returned, 0 available, can't edit vacations", "Medium", "High", "High", "3 OPEN bugs (#3352, #3370). Test maternity leave users."),
    ("Concurrent Operations", "PostgreSQL deadlocks on simultaneous vacation operations", "Medium", "High", "High", "Run 3+ concurrent creates for same employee. Expect deadlock on some."),
    ("NPE on Null Fields", "paymentMonth/optionalApprovers null → HTTP 500", "High", "Medium", "High", "2 known NPE paths. Test null values for each optional field."),
    ("Permission Bypass (Admin)", "ROLE_ADMIN in transition map but not in MANAGER_ROLES — hasAccess fails", "Low", "High", "High", "Admin can transition but hasAccess() blocks them. Design contradiction."),
    ("Calendar-Vacation Interaction", "Calendar changes triggering mass conversion (#3338)", "Medium", "High", "High", "Test: modify calendar → verify only affected vacation reconverted."),
    ("Email Notification Content", "Wrong payment month in email (#2925), duplicate notifications (#3315)", "Medium", "Medium", "Medium", "Verify email content matches vacation data."),
    ("Latin Name Search (#3297)", "Employee Vacation Days search broken for Latin names", "High", "Low", "Medium", "OPEN bug. Search by Latin vs Cyrillic name."),
]


# ─── Workbook Generation ─────────────────────────────────────────────────────

def apply_body_style(cell, row_idx):
    cell.font = FONT_BODY
    cell.alignment = ALIGN_WRAP
    cell.border = THIN_BORDER
    cell.fill = FILL_ROW_ALT if row_idx % 2 == 0 else FILL_ROW_WHITE


def write_header_row(ws, headers, col_widths):
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_test_cases(ws, cases, back_link_tab="Plan Overview"):
    """Write test cases to a worksheet."""
    headers = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
               "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
    col_widths = [14, 40, 45, 65, 45, 10, 10, 35, 22, 35]

    # Back-link row
    ws.merge_cells("A1:J1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = f"← Back to {back_link_tab}"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = f"#'{back_link_tab}'!A1"

    # Headers in row 2
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # Data rows
    for row_idx, tc in enumerate(cases, 3):
        values = [tc["id"], tc["title"], tc["preconditions"], tc["steps"],
                  tc["expected"], tc["priority"], tc["type"],
                  tc["req_ref"], tc["module"], tc["notes"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx in (6, 7):  # Priority, Type
                cell.alignment = ALIGN_CENTER

    ws.freeze_panes = "A3"


def write_plan_overview(ws, suites_info):
    """Write Plan Overview tab."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value=PLAN_OVERVIEW["title"])
    cell.font = FONT_TITLE
    row += 2

    # Scope
    ws.cell(row=row, column=1, value="Scope").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["scope"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    # Objectives
    ws.cell(row=row, column=1, value="Objectives").font = FONT_SUBTITLE
    row += 1
    for obj in PLAN_OVERVIEW["objectives"]:
        ws.cell(row=row, column=2, value=f"• {obj}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 1

    # Environments
    ws.cell(row=row, column=1, value="Environments").font = FONT_SUBTITLE
    row += 1
    for env in PLAN_OVERVIEW["environments"]:
        ws.cell(row=row, column=2, value=f"• {env}").font = FONT_BODY
        row += 1
    row += 1

    # Approach
    ws.cell(row=row, column=1, value="Approach").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=PLAN_OVERVIEW["approach"]).font = FONT_BODY
    ws.cell(row=row, column=2).alignment = ALIGN_WRAP
    row += 2

    # Dependencies
    ws.cell(row=row, column=1, value="Dependencies").font = FONT_SUBTITLE
    row += 1
    for dep in PLAN_OVERVIEW["dependencies"]:
        ws.cell(row=row, column=2, value=f"• {dep}").font = FONT_BODY
        ws.cell(row=row, column=2).alignment = ALIGN_WRAP
        row += 1
    row += 2

    # Test Suites
    ws.cell(row=row, column=1, value="Test Suites").font = FONT_SUBTITLE
    row += 1
    for suite_name, count, tab_name in suites_info:
        cell = ws.cell(row=row, column=2, value=f"{suite_name} — {count} cases")
        cell.font = FONT_LINK
        cell.hyperlink = f"#'{tab_name}'!A1"
        row += 1

    # Generated timestamp
    row += 2
    ws.cell(row=row, column=1, value="Generated").font = FONT_SUBTITLE
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M UTC")).font = FONT_BODY

    ws.freeze_panes = "A2"


def write_feature_matrix(ws):
    """Write Feature Matrix tab."""
    headers = ["Feature", "CRUD", "Approval", "Payment", "Validation", "Filters",
               "AV Mode", "Notif", "Regression", "Perms", "API", "Total"]
    col_widths = [35, 8, 10, 10, 12, 8, 10, 8, 12, 8, 8, 8]

    # Back-link
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    # Headers in row 2
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data
    suite_tabs = ["TS-Vacation-CRUD", "TS-Vac-Approval", "TS-Vac-Payment",
                  "TS-Vac-Validation", "TS-Vac-Filters", "TS-Vacation-AV",
                  "TS-Vac-Notif", "TS-Vac-Regression", "TS-Vac-Perms", "TS-Vac-API"]

    for row_idx, fm_row in enumerate(FEATURE_MATRIX, 3):
        feature = fm_row[0]
        counts = fm_row[1:-1]
        total = fm_row[-1]

        cell_f = ws.cell(row=row_idx, column=1, value=feature)
        apply_body_style(cell_f, row_idx)

        for col_idx, (count, tab) in enumerate(zip(counts, suite_tabs), 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
            apply_body_style(cell, row_idx)
            cell.alignment = ALIGN_CENTER
            if count > 0:
                cell.font = FONT_LINK
                cell.hyperlink = f"#'{tab}'!A1"

        cell_t = ws.cell(row=row_idx, column=len(headers), value=total)
        apply_body_style(cell_t, row_idx)
        cell_t.alignment = ALIGN_CENTER
        cell_t.font = Font(name="Arial", bold=True, size=10)

    # Totals row
    total_row = len(FEATURE_MATRIX) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col_idx in range(2, len(headers) + 1):
        col_sum = sum(
            FEATURE_MATRIX[r][col_idx - 1] if col_idx <= len(FEATURE_MATRIX[0]) else 0
            for r in range(len(FEATURE_MATRIX))
        )
        cell = ws.cell(row=total_row, column=col_idx, value=col_sum if col_sum > 0 else "")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def write_risk_assessment(ws):
    """Write Risk Assessment tab."""
    headers = ["Feature/Area", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    col_widths = [30, 55, 12, 12, 12, 55]

    # Back-link
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    link_cell = ws.cell(row=1, column=1)
    link_cell.value = "← Back to Plan Overview"
    link_cell.font = FONT_BACK_LINK
    link_cell.hyperlink = "#'Plan Overview'!A1"

    # Headers
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data
    severity_fills = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_MED, "Medium": FILL_RISK_LOW, "Low": FILL_RISK_LOW}

    for row_idx, (feature, risk, likelihood, impact, severity, mitigation) in enumerate(RISK_ASSESSMENT, 3):
        values = [feature, risk, likelihood, impact, severity, mitigation]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            apply_body_style(cell, row_idx)
            if col_idx == 5:  # Severity
                cell.fill = severity_fills.get(severity, FILL_ROW_WHITE)
            if col_idx in (3, 4, 5):
                cell.alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"
    ws.freeze_panes = "A3"


def generate():
    """Generate the vacation test documentation workbook."""
    wb = Workbook()

    # Collect all suites
    suites = [
        ("TS-Vacation-CRUD", "TS-Vacation-CRUD", get_crud_cases),
        ("TS-Vacation-Approval", "TS-Vac-Approval", get_approval_cases),
        ("TS-Vacation-Payment", "TS-Vac-Payment", get_payment_cases),
        ("TS-Vacation-Validation", "TS-Vac-Validation", get_validation_cases),
        ("TS-Vacation-Filters", "TS-Vac-Filters", get_filter_cases),
        ("TS-Vacation-AV", "TS-Vacation-AV", get_av_cases),
        ("TS-Vacation-Notifications", "TS-Vac-Notif", get_notification_cases),
        ("TS-Vacation-Regression", "TS-Vac-Regression", get_regression_cases),
        ("TS-Vacation-Permissions", "TS-Vac-Perms", get_permission_cases),
        ("TS-Vacation-API", "TS-Vac-API", get_api_cases),
    ]

    suites_info = []
    all_cases = []

    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        suites_info.append((suite_name, len(cases), tab_name))
        all_cases.extend(cases)

    total_cases = sum(len(case_fn()) for _, _, case_fn in suites)
    print(f"Generating {total_cases} test cases across {len(suites)} suites")

    # Plan Overview (default first sheet)
    ws_plan = wb.active
    ws_plan.title = "Plan Overview"
    ws_plan.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_plan_overview(ws_plan, suites_info)

    # Feature Matrix
    ws_fm = wb.create_sheet("Feature Matrix")
    ws_fm.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_feature_matrix(ws_fm)

    # Risk Assessment
    ws_ra = wb.create_sheet("Risk Assessment")
    ws_ra.sheet_properties.tabColor = TAB_COLOR_PLAN
    write_risk_assessment(ws_ra)

    # Test Suite tabs
    for suite_name, tab_name, case_fn in suites:
        cases = case_fn()
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_SUITE
        write_test_cases(ws, cases)
        print(f"  {tab_name}: {len(cases)} cases")

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {total_cases} test cases, {len(suites)} suites, 3 plan tabs")

    return all_cases


if __name__ == "__main__":
    generate()
