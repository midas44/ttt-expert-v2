#!/usr/bin/env python3
"""Add TS-VAC-PastDateValidation supplementary test suite to vacation.xlsx.

Sprint 15 fix: #3369 — Backend allowed creating/updating vacations with start
dates in the past. Fix in MR !5116 (merged 2026-01-14, bundled with #3360).
Also covers #3360 — Expected balance calculation used 3-year window instead
of unbounded sum.

Adds 5 test cases (TC-VAC-169 through TC-VAC-173) covering:
  - Update path: past start date rejected on vacation update (not just create)
  - Dual errors: past date + incorrect date order returned simultaneously
  - Boundary: today accepted, yesterday rejected
  - UX defect: error key displayed as raw string (no frontend translation)
  - Balance fix: calculateDaysNotAfter unbounded sum (#3360)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ── Styling constants (match existing workbook) ────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ─────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab matching existing format."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, tc_item in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            tc_item["id"], tc_item["title"], tc_item["preconditions"],
            tc_item["steps"], tc_item["expected"], tc_item["priority"],
            tc_item["type"], tc_item["req_ref"], tc_item["module"],
            tc_item.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# TEST CASE DATA — TS-VAC-PastDateVal (5 cases)
# ══════════════════════════════════════════════════════════════════

PAST_DATE_CASES = [
    # ── Update Path ────────────────────────────────────────────────

    tc("TC-VAC-169",
       "Update vacation start date to past — validation rejects",
       "Employee with an existing NEW or APPROVED vacation with start date in the future.\n"
       "Use timemachine env with clock set to a date between the current date and the vacation start.\n"
       "Find test data:\n"
       "  SELECT e.login, v.id, v.start_date, v.end_date, v.status\n"
       "  FROM ttt_vacation.employee e\n"
       "  JOIN ttt_vacation.vacation v ON e.id = v.employee_id\n"
       "  WHERE v.start_date > CURRENT_DATE\n"
       "    AND v.status IN ('NEW', 'APPROVED')\n"
       "  ORDER BY v.start_date\n"
       "  LIMIT 5;\n\n"
       "Note: Advance clock so the vacation's start date is now in the past.",
       "1. Record the existing vacation: id={id}, startDate={future_date}, endDate={end}.\n"
       "2. Advance the timemachine clock past the vacation's start_date:\n"
       "   PATCH /api/ttt/test/clock with date after the vacation's start.\n"
       "3. Attempt to update the vacation via API:\n"
       "   PUT /api/vacation/v1/vacations/{id}\n"
       "   Body: { login, startDate: {now_past_date}, endDate: {end}, paymentType: REGULAR, paymentMonth }\n"
       "   where startDate is the original (now in the past) start date.\n"
       "4. Verify response is 400 with validation error.\n"
       "5. Check error body contains 'validation.vacation.start.date.in.past'.\n"
       "6. Also attempt via UI: edit the vacation, change start date to yesterday.\n"
       "7. Verify the form shows validation error message.\n"
       "8. Verify the original vacation is unchanged in DB.",
       "API returns 400 Bad Request.\n"
       "Error body contains: validation.vacation.start.date.in.past\n"
       "Error is attached to startDate field.\n"
       "The vacation entity is NOT modified in the database.\n"
       "UI shows validation error when attempting to save.\n"
       "Note: VacationUpdateValidator delegates to VacationCreateValidator.isStartEndDatesCorrect() — same check.",
       "High", "Negative",
       "#3369, MR !5116", "Vacation / API / VacationUpdateValidator",
       "Update validator delegates past-date check to create validator (line 187). "
       "BDD Scenario 3 (vacation.feature:607) covers this path. "
       "Runs for both AV=false and AV=true."),

    # ── Dual Errors ────────────────────────────────────────────────

    tc("TC-VAC-170",
       "Past start date + end before start — both validation errors returned simultaneously",
       "Any active employee on timemachine env.\n"
       "Clock set to a known current date (e.g., 2026-03-15).",
       "1. Attempt to create vacation via API:\n"
       "   POST /api/vacation/v1/vacations\n"
       "   Body: {\n"
       "     login: {employee_login},\n"
       "     startDate: '2026-03-10',  // 5 days in past\n"
       "     endDate: '2026-03-05',    // before start date\n"
       "     paymentType: 'REGULAR',\n"
       "     paymentMonth: '2026-03-01'\n"
       "   }\n"
       "2. Verify response is 400 with MULTIPLE validation errors.\n"
       "3. Parse the error response body.\n"
       "4. Verify it contains BOTH:\n"
       "   a) 'validation.vacation.start.date.in.past' (on startDate field)\n"
       "   b) 'validation.vacation.dates.order' (on startDate AND endDate fields)\n"
       "5. Verify NO other validation errors (duration, next-year) are present.\n"
       "   These are short-circuited by && after isStartEndDatesCorrect() fails.\n"
       "6. Try same via UI: enter past start date AND end date before start.\n"
       "7. Verify both errors are displayed (or check if frontend only shows first).",
       "API returns 400 with TWO validation violations:\n"
       "  1. validation.vacation.start.date.in.past (on startDate)\n"
       "  2. validation.vacation.dates.order (on BOTH startDate and endDate)\n"
       "Both collected before method returns — non-short-circuiting within isStartEndDatesCorrect().\n"
       "Subsequent validations (duration, next-year) are NOT invoked.\n"
       "No vacation entity created.",
       "High", "Negative",
       "#3369, MR !5116", "Vacation / API / VacationCreateValidator",
       "BDD Scenario 2 (vacation.feature:584) verifies both errors returned. "
       "isStartEndDatesCorrect() collects all constraint violations in a single pass, "
       "then returns false. The && short-circuits isValidVacationDuration and "
       "isNextVacationAvailable."),

    # ── Boundary ───────────────────────────────────────────────────

    tc("TC-VAC-171",
       "Boundary: vacation starting today accepted, starting yesterday rejected",
       "Active employee on timemachine env.\n"
       "Employee has sufficient vacation days balance.\n"
       "Clock set to a specific date (e.g., 2026-06-15).\n"
       "No overlapping vacations for the test dates.",
       "1. Set timemachine clock to 2026-06-15T10:00:00.\n"
       "2. Create vacation starting TODAY (2026-06-15) to 2026-06-25:\n"
       "   POST /api/vacation/v1/vacations\n"
       "   { startDate: '2026-06-15', endDate: '2026-06-25', paymentType: REGULAR, ... }\n"
       "3. Verify response is 200 OK — vacation created successfully.\n"
       "4. Delete the vacation (or use a different employee for step 5).\n"
       "5. Create vacation starting YESTERDAY (2026-06-14) to 2026-06-25:\n"
       "   POST /api/vacation/v1/vacations\n"
       "   { startDate: '2026-06-14', endDate: '2026-06-25', paymentType: REGULAR, ... }\n"
       "6. Verify response is 400 with 'validation.vacation.start.date.in.past'.\n"
       "7. Verify via UI:\n"
       "   a) Open vacation creation form, set start date to today — Submit succeeds.\n"
       "   b) Open form, set start date to yesterday — Submit fails.\n"
       "8. Edge case: advance clock to 2026-06-15T23:59:59.\n"
       "   Create vacation starting 2026-06-15 — should still succeed.\n"
       "   The check uses LocalDate (date-only), not timestamp.",
       "Today's date is ACCEPTED — isBefore(today) returns false for today.\n"
       "Yesterday's date is REJECTED — isBefore(today) returns true.\n"
       "Validation uses LocalDate comparison (TimeUtils.today()),\n"
       "NOT timestamp-level — time of day does not affect the result.\n"
       "23:59:59 on today still allows today's date as start.\n"
       "This is a boundary condition: >= today passes, < today fails.",
       "Critical", "Boundary",
       "#3369, MR !5116", "Vacation / API / VacationCreateValidator",
       "The validation uses startDate.isBefore(today) — strict less-than. "
       "Today is NOT 'before' today, so it passes. "
       "BDD Scenario 1 uses a 5-day gap (Jan 10 vs Jan 15). "
       "This test targets the exact boundary."),

    # ── UX Defect ──────────────────────────────────────────────────

    tc("TC-VAC-172",
       "Past-date validation error displayed as raw key — missing frontend translation",
       "Active employee on timemachine env.\n"
       "Clock set to a date that makes the vacation's start in the past.\n"
       "Use Chrome browser with DevTools open (Network tab).",
       "1. Open vacation creation form in UI.\n"
       "2. Set start date to a past date (e.g., yesterday).\n"
       "3. Submit the form.\n"
       "4. Observe the error message displayed in the form.\n"
       "5. Check Network tab for the API response body.\n"
       "6. Compare API error key with displayed UI message.\n"
       "7. Check for these error keys in the UI:\n"
       "   a) 'validation.vacation.start.date.in.past' — expected: raw key or generic fallback\n"
       "   b) 'validation.vacation.dates.order' — expected: raw key or generic fallback\n"
       "   c) 'validation.vacation.next.year.not.available' — expected: raw key or generic fallback\n"
       "8. Compare with TRANSLATED error keys:\n"
       "   a) 'exception.validation.vacation.duration' — should show 'You don't have enough...'\n"
       "   b) 'exception.validation.vacation.too.early' — should show 'Vacation request can be...'\n"
       "9. Document the exact error message shown for each untranslated key.",
       "KNOWN DEFECT: The following validation error keys have NO frontend translation:\n"
       "  - validation.vacation.start.date.in.past\n"
       "  - validation.vacation.dates.order\n"
       "  - validation.vacation.next.year.not.available\n"
       "UI displays the raw key string (e.g., 'validation.vacation.start.date.in.past')\n"
       "or a generic error fallback, depending on frontend error handling.\n"
       "Only these keys have translations:\n"
       "  - exception.validation.vacation.duration -> 'You don't have enough available vacation days'\n"
       "  - exception.validation.vacation.too.early -> 'Vacation request can be created after 6 months...'\n"
       "Document the actual UI behavior for each key.",
       "Medium", "UX",
       "#3369, MR !5116", "Vacation / Frontend / i18n",
       "Codebase grep confirmed: no translation entries for 'start.date.in.past', "
       "'dates.order', 'next.year.not.available' in frontend i18n files. "
       "Only 'duration' and 'too.early' have proper translations. "
       "This is a UX defect — user sees a technical key string instead of human-readable message."),

    # ── Balance Fix (#3360) ────────────────────────────────────────

    tc("TC-VAC-173",
       "Expected year-end balance calculation — unbounded year sum (#3360 fix)",
       "Employee with > 2 years of employment and vacation accruals.\n"
       "Employee should have available_vacation_days in 3+ different years.\n"
       "Find test data (ttt_vacation DB):\n"
       "  SELECT ev.employee AS emp_id, COUNT(DISTINCT ev.year) AS year_count,\n"
       "         SUM(ev.available_vacation_days) AS total_days\n"
       "  FROM ttt_vacation.employee_vacation ev\n"
       "  GROUP BY ev.employee\n"
       "  HAVING COUNT(DISTINCT ev.year) >= 3\n"
       "  ORDER BY SUM(ev.available_vacation_days) DESC\n"
       "  LIMIT 5;\n\n"
       "Then get the login:\n"
       "  SELECT login FROM ttt_vacation.employee WHERE id = {emp_id};\n\n"
       "Use timemachine env.",
       "1. Record baseline per-year balances:\n"
       "   SELECT year, available_vacation_days\n"
       "   FROM ttt_vacation.employee_vacation\n"
       "   WHERE employee = {emp_id}\n"
       "   ORDER BY year;\n"
       "2. API: GET /api/vacation/v1/vacationdays/available?employeeLogin={login}\n"
       "   &newDays=0&usePaymentDateFilter=true\n"
       "3. Record `availablePaidDays` from response.\n"
       "4. Calculate expected: SUM of available_vacation_days WHERE year <= current_year.\n"
       "   This should be an UNBOUNDED sum (no 3-year window).\n"
       "5. Verify API `availablePaidDays` matches the unbounded sum\n"
       "   (not the 3-year window sum).\n"
       "6. For an employee with 4+ years of accruals:\n"
       "   Old behavior (pre-fix): calculateDaysBeforeAndAfter(year, year-2) = 3-year window\n"
       "   New behavior: calculateDaysNotAfter(year) = all years <= current\n"
       "7. If the employee has days in year-3 or earlier:\n"
       "   unbounded_sum SHOULD BE > 3year_sum.\n"
       "8. UI: Navigate to Vacations page and verify 'Expected balance'\n"
       "   display reflects the full balance (not truncated to 3 years).",
       "API `availablePaidDays` equals SUM of all available_vacation_days\n"
       "for year <= current_year (unbounded, not 3-year window).\n"
       "For employee with 4+ years:\n"
       "  Pre-fix: balance missed year-3 and older accruals.\n"
       "  Post-fix: balance includes ALL historic accruals.\n"
       "SQL verification:\n"
       "  SELECT SUM(available_vacation_days) FROM employee_vacation\n"
       "  WHERE employee = {id} AND year <= EXTRACT(YEAR FROM CURRENT_DATE)\n"
       "  should match API availablePaidDays (minus any pending vacations).\n"
       "UI 'Expected balance' reflects the full unbounded sum.",
       "High", "Functional",
       "#3360, MR !5116", "Vacation / Backend / EmployeeDaysServiceImpl",
       "Same MR !5116 as #3369. Changed calculateDaysBeforeAndAfter(year, year-2) "
       "to calculateDaysNotAfter(year). SQL changed from "
       "'WHERE year <= :before AND year >= :after' to 'WHERE year <= :year'. "
       "Affects VacationAvailablePaidDaysCalculatorImpl.calculate() line 94."),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Modify existing workbook
# ══════════════════════════════════════════════════════════════════

def main():
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vacation.xlsx")
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. Remove existing tab if present (for re-runs) ──────────
    tab_name = "TS-VAC-PastDateVal"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
        print(f"  Removed existing '{tab_name}' tab (re-run)")

    # ── 2. Create the new sheet ──────────────────────────────────
    if "Test Data" in wb.sheetnames:
        ws = wb.create_sheet(title=tab_name, index=wb.sheetnames.index("Test Data"))
    else:
        ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    case_count = write_ts_tab(ws, "Past-Date Validation & Balance Fix (#3369, #3360)", PAST_DATE_CASES)
    print(f"  Created '{tab_name}' with {case_count} test cases (TC-VAC-169 to TC-VAC-173)")

    # ── 3. Update Plan Overview ──────────────────────────────────
    ws_plan = wb["Plan Overview"]

    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Total Test Cases" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_val = int(old_val) + case_count if old_val else 168 + case_count
            ws_plan.cell(row=r, column=2, value=str(new_val))
            print(f"  Updated Plan Overview: Total Test Cases {old_val} -> {new_val}")
            break

    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Test Suites" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_suites = int(old_val) + 1 if old_val else 14
            ws_plan.cell(row=r, column=2, value=str(new_suites))
            print(f"  Updated Plan Overview: Test Suites {old_val} -> {new_suites}")
            break

    last_link_row = 0
    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.hyperlink and "TS-" in str(cell.hyperlink.target or ""):
            last_link_row = r

    if last_link_row > 0:
        new_link_row = last_link_row + 1
        max_row = ws_plan.max_row
        for r in range(max_row, new_link_row - 1, -1):
            for c in range(1, 6):
                src = ws_plan.cell(row=r, column=c)
                dst = ws_plan.cell(row=r + 1, column=c)
                dst.value = src.value
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.alignment = copy(src.alignment)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                if src.hyperlink:
                    dst.hyperlink = src.hyperlink.target
                src.value = None
                src.hyperlink = None

        link_cell = ws_plan.cell(row=new_link_row, column=1,
                                  value=f"Past-Date Validation & Balance (#3369, #3360) - {case_count} cases")
        link_cell.font = FONT_LINK_BOLD
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        print(f"  Added hyperlink at Plan Overview row {new_link_row}")

    # ── 4. Update Feature Matrix ─────────────────────────────────
    ws_fm = wb["Feature Matrix"]

    total_row = None
    for r in range(1, ws_fm.max_row + 1):
        if ws_fm.cell(row=r, column=1).value == "TOTAL":
            total_row = r
            break

    if total_row:
        new_fm_row = total_row
        for c in range(1, 8):
            src = ws_fm.cell(row=total_row, column=c)
            dst = ws_fm.cell(row=total_row + 1, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
            if src.hyperlink:
                dst.hyperlink = src.hyperlink.target

        fm_data = ["Past-Date Validation", 1, 3, 0, 1, case_count, "TS-VAC-PastDateVal"]
        fill_row = FILL_ROW_EVEN if (new_fm_row - 3) % 2 == 0 else FILL_ROW_ODD
        for col, val in enumerate(fm_data, 1):
            cell = ws_fm.cell(row=new_fm_row, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            cell.fill = fill_row
        link_cell = ws_fm.cell(row=new_fm_row, column=7)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = f"#'{tab_name}'!A1"

        new_total_row = total_row + 1
        for c in range(2, 7):
            old_total = ws_fm.cell(row=new_total_row, column=c).value or 0
            if c == 6:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + case_count)
            elif c == 2:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + 1)  # UI
            elif c == 3:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + 3)  # API
            elif c == 5:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + 1)  # UX

        print(f"  Updated Feature Matrix: added row {new_fm_row}")

    # ── 5. Save ──────────────────────────────────────────────────
    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")
    print(f"Total: {case_count} test cases added in '{tab_name}'")


if __name__ == "__main__":
    main()
