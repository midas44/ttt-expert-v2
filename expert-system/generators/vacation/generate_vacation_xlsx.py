#!/usr/bin/env python3
"""Generate vacation.xlsx — unified test workbook for Vacation module.

Phase B output for the TTT Expert System.
Covers: CRUD, state machine, approval workflow, day calculation,
        payment, permissions, API errors.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Styling constants ──────────────────────────────────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
FILL_RISK_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_RISK_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RISK_LOW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_PLAN = "548235"
TAB_COLOR_TS = "2F5496"


# ── Helper functions ───────────────────────────────────────────────

def style_header_row(ws, row, num_cols, fill=None):
    """Apply header styling to a row."""
    f = fill or FILL_HEADER
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = f
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def write_row(ws, row, values, font=None, fill=None, alignment=None):
    """Write a row of values with optional styling."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = alignment or ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=10, max_width=60):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, min(longest + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def add_autofilter(ws, row, num_cols):
    """Add autofilter to header row."""
    ws.auto_filter.ref = f"A{row}:{get_column_letter(num_cols)}{ws.max_row}"


def add_back_link(ws, row=1):
    """Add back-link to Plan Overview in row 1."""
    cell = ws.cell(row=row, column=1)
    cell.value = "<- Back to Plan"
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab with header, back-link, and test cases."""
    add_back_link(ws, row=1)
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for i, tc in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        values = [
            tc["id"], tc["title"], tc["preconditions"], tc["steps"],
            tc["expected"], tc["priority"], tc["type"],
            tc["req_ref"], tc["module"], tc.get("notes", "")
        ]
        write_row(ws, row, values, fill=fill)

    add_autofilter(ws, header_row, len(headers))

    # Column widths
    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ── Test Case Data ─────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


# ── TS-Vac-Create ─────────────────────────────────────────────────

TS_VAC_CREATE = [
    tc("TC-VAC-001",
       "Create REGULAR vacation — happy path (AV=false office)",
       "Employee in AV=false office (e.g. Russia)\nSufficient accrued days (>=5)\nNo overlapping vacations\nTest data: use DB query SELECT e.login, ev.days FROM employee e JOIN employee_vacation ev ON e.id=ev.employee_id WHERE ev.days >= 5 AND e.office_id IN (SELECT id FROM office WHERE advance_vacation=false) LIMIT 5",
       "1. POST /api/vacation/v1/vacations\n2. Body: {login, startDate: future Mon, endDate: future Fri (5 days), paymentType: REGULAR, paymentMonth: startDate 1st-of-month, optionalApprovers: [], notifyAlso: []}\n3. Verify response status 200\n4. GET /api/vacation/v1/vacations/{id} to confirm",
       "Vacation created with status=NEW\nApprover auto-assigned (employee's manager)\nDays field = 5 (working days within range)\nPayment month set correctly\nAvailable days decreased by vacation days",
       "Critical", "Functional",
       "#3014, Confluence Vacations", "VacationServiceImpl, VacationCreateValidator",
       "Core happy path. Verify via DB: SELECT * FROM vacation WHERE id=<new_id>"),

    tc("TC-VAC-002",
       "Create REGULAR vacation — happy path (AV=true office)",
       "Employee in AV=true office (e.g. Cyprus)\nSufficient balance\nTest data: SELECT e.login FROM employee e JOIN office o ON e.office_id=o.id WHERE o.advance_vacation=true AND e.status='ACTIVE' LIMIT 5",
       "1. POST /api/vacation/v1/vacations\n2. Body: {login, startDate: +2 weeks, endDate: +2 weeks +6 days (5 working days), paymentType: REGULAR, paymentMonth, optionalApprovers: [], notifyAlso: []}\n3. Verify response",
       "Vacation created with status=NEW\nFull year balance available (not monthly accrual)\nDays calculated correctly\nAdvance vacation formula used",
       "Critical", "Functional",
       "#3092, Confluence Advance Vacation", "VacationServiceImpl, AdvanceCalculationStrategy",
       "AV=true offices use different calculation strategy"),

    tc("TC-VAC-003",
       "Create ADMINISTRATIVE vacation (unpaid)",
       "Active employee, any office\nTest data: any active employee login",
       "1. POST /api/vacation/v1/vacations\n2. Body: {login, startDate: +1 week, endDate: +1 week (1 day), paymentType: ADMINISTRATIVE, paymentMonth: startDate month, optionalApprovers: [], notifyAlso: []}\n3. Verify response",
       "Vacation created with status=NEW\nNo available days check performed\nMin duration = 1 day (not 5)\npaymentType = ADMINISTRATIVE in response",
       "High", "Functional",
       "Confluence Vacations", "VacationServiceImpl, VacationCreateValidator",
       "ADMINISTRATIVE skips duration and availability checks"),

    tc("TC-VAC-004",
       "Create vacation with start date in past",
       "Active employee\nCurrent date known",
       "1. POST /api/vacation/v1/vacations\n2. Body: startDate = yesterday, endDate = tomorrow\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.start.date.in.past\nerrors[].field = startDate",
       "High", "Negative",
       "VacationCreateValidator", "VacationCreateValidator",
       "Backend validation; frontend also blocks this via date picker"),

    tc("TC-VAC-005",
       "Create vacation with startDate > endDate",
       "Active employee",
       "1. POST /api/vacation/v1/vacations\n2. Body: startDate = 2026-04-10, endDate = 2026-04-05\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.dates.order\nerrors[].field includes startDate and endDate",
       "High", "Negative",
       "VacationCreateValidator", "VacationCreateValidator",
       ""),

    tc("TC-VAC-006",
       "Create REGULAR vacation < 5 days (min duration violation)",
       "Active employee in AV=false office with sufficient days",
       "1. POST /api/vacation/v1/vacations\n2. Body: REGULAR, startDate to endDate = 3 calendar days (e.g. Mon-Wed)\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.duration\nMessage indicates minimum duration not met\nFrontend gap: UI allows 1-day REGULAR but backend rejects < 5",
       "High", "Negative",
       "#3014, VacationCreateValidator", "VacationCreateValidator",
       "Frontend-backend gap: UI min=1 day, backend min=5 days for REGULAR"),

    tc("TC-VAC-007",
       "Create REGULAR vacation = 5 calendar days (boundary)",
       "Active employee with >= 5 accrued days",
       "1. POST /api/vacation/v1/vacations\n2. Body: REGULAR, Mon-Fri span (5 calendar days)\n3. Verify success",
       "Vacation created successfully\nDays = number of working days in range (typically 5 if Mon-Fri)\nBoundary value at minimum duration",
       "Medium", "Boundary",
       "#3014", "VacationCreateValidator",
       "Boundary: exactly at minimalVacationDuration threshold"),

    tc("TC-VAC-008",
       "Create ADMINISTRATIVE vacation = 1 day",
       "Active employee",
       "1. POST /api/vacation/v1/vacations\n2. Body: ADMINISTRATIVE, startDate = endDate (1 day)\n3. Verify success",
       "Vacation created with days=1\nNo duration check applied\nADMINISTRATIVE minimum is 1 day",
       "Medium", "Boundary",
       "Confluence Vacations", "VacationCreateValidator",
       "ADMINISTRATIVE type skips duration validation entirely"),

    tc("TC-VAC-009",
       "Create with insufficient available days (AV=false)",
       "Employee in AV=false office with low accrued balance\nTest data: SELECT e.login, ev.days FROM employee e JOIN employee_vacation ev ON e.id=ev.employee_id JOIN office o ON e.office_id=o.id WHERE o.advance_vacation=false AND ev.days < 5 AND ev.year=EXTRACT(YEAR FROM CURRENT_DATE)",
       "1. POST /api/vacation/v1/vacations\n2. Body: REGULAR, 14-day vacation exceeding available days\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.duration\nMessage: insufficient accrued days\nRed error in UI: 'Reduce duration or select unpaid'",
       "High", "Negative",
       "#3014", "VacationCreateValidator, RegularCalculationStrategy",
       "AV=false: negative balance never allowed, blocks at creation"),

    tc("TC-VAC-010",
       "Create with insufficient available days (AV=true)",
       "Employee in AV=true office with low or negative balance",
       "1. POST /api/vacation/v1/vacations\n2. Body: REGULAR, long vacation exceeding total available\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.duration (Error 11.4)\nMessage: insufficient days\nNote: AV=true allows negative for current year via norm deviation, but creation validates total",
       "High", "Negative",
       "#3092", "VacationCreateValidator, AdvanceCalculationStrategy",
       "AV=true: can go negative via norm deviation but create validator still checks total"),

    tc("TC-VAC-011",
       "Create next-year vacation before Feb 1 cutoff",
       "Current date before Feb 1\nTest data: set timemachine clock to Jan 15 via PATCH /api/ttt/test/v1/clock",
       "1. Set clock to Jan 15 of current year\n2. POST /api/vacation/v1/vacations\n3. Body: REGULAR, startDate in next year\n4. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.next.year.not.available\nMessage: next year vacation not available yet",
       "Medium", "Boundary",
       "#3092, VacationCreateValidator", "VacationCreateValidator",
       "Configurable cutoff date. Boundary: Jan 31 vs Feb 1"),

    tc("TC-VAC-012",
       "Create next-year vacation on/after Feb 1",
       "Current date >= Feb 1\nTest data: set timemachine clock to Feb 1",
       "1. Set clock to Feb 1\n2. POST /api/vacation/v1/vacations\n3. Body: REGULAR, startDate in next year\n4. Verify success or other validation failure (not next-year block)",
       "No next-year-not-available error\nVacation created or fails with different validation (e.g. insufficient days)\nFeb 1 boundary is inclusive",
       "Medium", "Boundary",
       "#3092", "VacationCreateValidator",
       "Boundary companion to TC-VAC-011"),

    tc("TC-VAC-013",
       "Create overlapping vacation (start inside existing)",
       "Employee has existing NEW/APPROVED vacation (e.g. Apr 7-11)\nTest data: SELECT v.employee_id, v.start_date, v.end_date FROM vacation v WHERE v.status IN ('NEW','APPROVED') AND v.start_date > CURRENT_DATE LIMIT 5",
       "1. POST /api/vacation/v1/vacations\n2. Body: startDate inside existing vacation range\n3. Check error response",
       "HTTP 400\nerrorCode: exception.validation.vacation.dates.crossing\nerrors[].field = startDate",
       "High", "Negative",
       "VacationServiceImpl", "VacationCRUDService.findCrossingVacations",
       "3 overlap patterns: start-inside, end-inside, enclosing"),

    tc("TC-VAC-014",
       "Create with null paymentMonth — NPE bug",
       "Active employee",
       "1. POST /api/vacation/v1/vacations\n2. Body: omit paymentMonth field entirely\n3. Check response",
       "KNOWN BUG: HTTP 500 NullPointerException at VacationAvailablePaidDaysCalculatorImpl:73\npaymentDate.getYear() NPE\nField has no @NotNull annotation in DTO",
       "High", "Negative",
       "Bug #2 from API testing", "VacationCreateRequestDTO, VacationAvailablePaidDaysCalculatorImpl",
       "Known NPE vulnerability. DTO lacks @NotNull on paymentMonth"),

    tc("TC-VAC-015",
       "Create with null optionalApprovers — NPE bug (CPO path)",
       "CPO employee (chief project officer / department manager)\nTest data: SELECT e.login FROM employee e JOIN employee_global_roles r ON e.id=r.employee_id WHERE r.role='ROLE_DEPARTMENT_MANAGER' AND e.status='ACTIVE' LIMIT 3",
       "1. POST /api/vacation/v1/vacations as CPO user\n2. Body: omit optionalApprovers field (null)\n3. Check response",
       "KNOWN BUG: HTTP 500 NullPointerException at VacationServiceImpl:155\nCPO path calls getOptionalApprovers().add(manager) on null list\nNon-CPO path may work (null check in synchronizeOptionalApprovals)",
       "High", "Negative",
       "Bug #3 from API testing", "VacationServiceImpl",
       "NPE only on CPO code path. Send optionalApprovers: [] to workaround"),

    tc("TC-VAC-016",
       "Create with non-existent employee login",
       "Invalid login string",
       "1. POST /api/vacation/v1/vacations\n2. Body: login = 'nonexistent_user_xyz'\n3. Check error response",
       "HTTP 400\nValidation error from @EmployeeLoginExists annotation\nerrors[].field = login",
       "Medium", "Negative",
       "AbstractVacationRequestDTO", "DTO validation",
       "@EmployeeLoginExists custom validator on login field"),

    tc("TC-VAC-017",
       "Create as readOnly user",
       "Employee with readOnly=true\nTest data: SELECT login FROM employee WHERE read_only=true AND status='ACTIVE' LIMIT 3",
       "1. Authenticate as readOnly user\n2. POST /api/vacation/v1/vacations\n3. Check error response",
       "HTTP 403\nerrorCode: exception.vacation.no.permission\nVacationSecurityException thrown by checkVacation",
       "Medium", "Negative",
       "VacationPermissionService", "VacationPermissionService, classPermissionService",
       "readOnly flag blocks all write operations"),

    tc("TC-VAC-018",
       "Create vacation — CPO auto-approver self-assignment",
       "CPO/DM employee with a manager\nTest data: DM user with non-null manager_id",
       "1. POST /api/vacation/v1/vacations as CPO\n2. Body: standard valid vacation\n3. Check approver in response",
       "approverId = employee's own ID (self-approval)\nEmployee's manager added as optional approver with status ASKED\nDB: vacation_approval record created for manager",
       "High", "Functional",
       "VacationServiceImpl.createVacation", "VacationServiceImpl",
       "CPO pattern: self-approve + manager as optional"),

    tc("TC-VAC-019",
       "Create vacation — regular employee auto-approver assignment",
       "Regular employee (ROLE_EMPLOYEE only) with manager\nTest data: SELECT e.login, m.login as manager FROM employee e JOIN employee m ON e.manager_id=m.id WHERE e.status='ACTIVE' AND e.manager_id IS NOT NULL LIMIT 5",
       "1. POST /api/vacation/v1/vacations as regular employee\n2. Body: standard valid vacation\n3. Check approver in response",
       "approverId = employee's manager ID\nManager is primary approver\nNo optional approvers auto-added (unless specified)",
       "High", "Functional",
       "VacationServiceImpl.createVacation", "VacationServiceImpl",
       "Standard flow: manager is primary approver"),

    tc("TC-VAC-020",
       "Create vacation — employee without manager (self-approval)",
       "Employee with manager_id = NULL\nTest data: SELECT login FROM employee WHERE manager_id IS NULL AND status='ACTIVE' LIMIT 3",
       "1. POST /api/vacation/v1/vacations\n2. Body: standard valid vacation\n3. Check approver in response",
       "approverId = employee's own ID\nSelf-approval enabled\nNo NPE from null manager check",
       "Medium", "Functional",
       "VacationServiceImpl.createVacation", "VacationServiceImpl",
       "Fallback path when no manager exists"),

    tc("TC-VAC-021",
       "Create vacation — verify available days decrease atomically",
       "Employee with known available days balance",
       "1. GET /api/vacation/v1/vacationdays/available?employeeLogin=X&paymentDate=Y&newDays=0\n2. Record availablePaidDays\n3. POST /api/vacation/v1/vacations (5-day REGULAR)\n4. GET available days again\n5. Compare",
       "Available days decreased by exactly the number of vacation working days\nChange is atomic (no partial updates)\nExample: 34 → 29 for 5-day vacation",
       "High", "Functional",
       "VacationServiceImpl, recalculate", "VacationRecalculationService",
       "Days deducted at creation (for NEW status tracking), final deduction at approval"),

    tc("TC-VAC-022",
       "Create vacation with notifyAlso list",
       "Active employee, valid colleague logins for notify list",
       "1. POST /api/vacation/v1/vacations\n2. Body: notifyAlso: ['colleague1_login', 'colleague2_login']\n3. Verify response and DB",
       "Vacation created successfully\nvacation_notify_also records created\nNotify-also recipients receive email notification\nDB: SELECT * FROM vacation_notify_also WHERE vacation_id=<id>",
       "Medium", "Functional",
       "VacationServiceImpl.synchronizeNotifyAlso", "VacationServiceImpl",
       "Notify-also with required=false is informational only"),

    tc("TC-VAC-023",
       "Create vacation with invalid notifyAlso login",
       "Active employee, one invalid login in notifyAlso list",
       "1. POST /api/vacation/v1/vacations\n2. Body: notifyAlso: ['valid_login', 'nonexistent_xyz']\n3. Check error response",
       "HTTP 400\nValidation error from @EmployeeLoginCollectionExists\nAll logins in collection must exist",
       "Medium", "Negative",
       "AbstractVacationRequestDTO", "DTO validation",
       "@EmployeeLoginCollectionExists validates each login in list"),

    tc("TC-VAC-024",
       "Create vacation with comment",
       "Active employee",
       "1. POST /api/vacation/v1/vacations\n2. Body: comment = 'Family trip to the mountains'\n3. Verify comment saved",
       "Vacation created with comment field populated\nComment visible in GET response\nNo length limit enforced (DTO has no constraint)",
       "Low", "Functional",
       "AbstractVacationRequestDTO", "VacationServiceImpl",
       "Comment field has no @Size annotation — test with very long string too"),

    tc("TC-VAC-025",
       "Create vacation with very long comment (no length limit)",
       "Active employee",
       "1. POST /api/vacation/v1/vacations\n2. Body: comment = 5000-character string\n3. Test data: generate with Python ''.join(['A'] * 5000)\n4. Check response",
       "Vacation created successfully (no DTO length constraint)\nComment stored in full in DB\nPotential issue: DB column may have length limit causing truncation or error",
       "Low", "Boundary",
       "AbstractVacationRequestDTO", "VacationServiceImpl",
       "No @Size annotation on comment field. DB column type determines actual limit"),
]


# ── TS-Vac-Update ──────────────────────────────────────────────────

TS_VAC_UPDATE = [
    tc("TC-VAC-026",
       "Update dates of NEW vacation",
       "Existing vacation in NEW status owned by current user\nTest data: SELECT id FROM vacation WHERE status='NEW' AND employee_id=<current_user_id>",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: new startDate/endDate (shift by 1 week), same paymentType\n3. Verify response",
       "Vacation updated, status remains NEW\nDays recalculated for new date range\nOptional approvals NOT reset (status was already NEW)",
       "High", "Functional",
       "VacationServiceImpl", "VacationServiceImpl, VacationUpdateValidator",
       "Update of NEW vacation doesn't trigger status change"),

    tc("TC-VAC-027",
       "Update dates of APPROVED vacation — status resets to NEW",
       "Existing APPROVED vacation owned by current user\nTest data: SELECT id FROM vacation WHERE status='APPROVED' AND employee_id=<current_user_id>",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: changed startDate/endDate\n3. Verify status in response",
       "Status changes APPROVED → NEW\nAll optional approvals reset to ASKED\nApprover must re-approve\nDays recalculated\nRecalculation service triggered",
       "Critical", "Functional",
       "VacationServiceImpl, VacationStatusManager", "VacationServiceImpl",
       "Key business rule: editing dates resets approval. Code: add(APPROVED, NEW, ROLE_EMPLOYEE)"),

    tc("TC-VAC-028",
       "Update CANCELED vacation",
       "Existing CANCELED vacation\nTest data: SELECT id FROM vacation WHERE status='CANCELED' LIMIT 3",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: updated dates\n3. Verify response",
       "Update succeeds\nUpdate validator skips day limit checks for CANCELED status\nStatus remains CANCELED (no auto-transition)",
       "Medium", "Functional",
       "VacationUpdateValidator", "VacationUpdateValidator",
       "CANCELED/REJECTED skip day limit validation in update path"),

    tc("TC-VAC-029",
       "Update REJECTED vacation — skips day limit checks",
       "Existing REJECTED vacation\nTest data: SELECT id FROM vacation WHERE status='REJECTED' LIMIT 3",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: updated dates (may exceed normal limits)\n3. Verify response",
       "Update succeeds even if days exceed normal limits\nUpdate validator uses raw daysLimitations (not adjusted)\nStatus remains REJECTED",
       "Medium", "Functional",
       "VacationUpdateValidator", "VacationUpdateValidator",
       "Key difference from create: REJECTED vacations skip limit adjustments"),

    tc("TC-VAC-030",
       "Update PAID vacation — immutable",
       "Existing PAID vacation\nTest data: SELECT id FROM vacation WHERE status='PAID' LIMIT 3",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: changed dates\n3. Check error response",
       "HTTP 400 or 403\nPAID status is terminal — no edits allowed\nPermission service returns empty permissions for PAID",
       "High", "Negative",
       "VacationPermissionService", "VacationPermissionService",
       "PAID is in NON_EDITABLE_STATUSES set"),

    tc("TC-VAC-031",
       "Update vacation owned by another user",
       "Vacation created by user A, attempt update as user B (not approver)\nTest data: use two different employee logins",
       "1. Authenticate as user B\n2. PUT /api/vacation/v1/vacations/{userA_vacation_id}\n3. Body: changed dates\n4. Check error response",
       "HTTP 400\nerrorCode: exception.vacation.status.notAllowed\nhasAccess() returns false (not owner, not approver)",
       "High", "Security",
       "VacationStatusManager.hasAccess", "VacationStatusManager",
       "Owner check: request.getEmployeeId().equals(employee.getId())"),

    tc("TC-VAC-032",
       "Update with overlapping dates",
       "Employee has two vacations, update one to overlap with the other",
       "1. Create vacation A (Apr 7-11) and vacation B (Apr 21-25)\n2. PUT /api/vacation/v1/vacations/{B_id}\n3. Body: startDate=Apr 9 (overlaps A)\n4. Check error",
       "HTTP 400\nerrorCode: exception.validation.vacation.dates.crossing\nOverlap detected against other vacations (excludes self)",
       "Medium", "Negative",
       "VacationCRUDService.findCrossingVacations", "VacationCRUDService",
       "findCrossingVacations excludes current vacation ID from check"),

    tc("TC-VAC-033",
       "Update — verify optional approvals reset on date change",
       "APPROVED vacation with optional approver who has APPROVED status",
       "1. Create vacation with optional approver\n2. Have optional approver approve (ASKED → APPROVED)\n3. PUT /api/vacation/v1/vacations/{id} with new dates\n4. Check optional approval status in DB",
       "All optional approvals reset to ASKED\nDB: SELECT status FROM vacation_approval WHERE vacation_id=<id> — all should be ASKED\nApprovals must be re-obtained",
       "High", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       "State reset behavior confirmed in code and live testing"),

    tc("TC-VAC-034",
       "Update — next year check NOT applied on update",
       "Set clock to Jan 15, existing vacation with future dates",
       "1. Set clock to Jan 15\n2. Create vacation for current year (succeeds)\n3. PUT /api/vacation/v1/vacations/{id}\n4. Body: startDate moved to next year\n5. Verify behavior",
       "Update MAY succeed (update validator does NOT call isNextVacationAvailable)\nDifference from create: next-year check only on creation\nNote: Other validations (available days) may still block",
       "Medium", "Functional",
       "VacationUpdateValidator", "VacationUpdateValidator",
       "Code: Update validator does NOT call isNextVacationAvailable()"),

    tc("TC-VAC-035",
       "Update paymentType REGULAR → ADMINISTRATIVE",
       "Existing REGULAR vacation in NEW status",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: paymentType changed to ADMINISTRATIVE, same dates\n3. Verify response",
       "PaymentType updated\nDuration validation now uses ADMINISTRATIVE rules (min 1 day)\nAvailable days recalculated\nPreviously consumed REGULAR days returned to pool",
       "Medium", "Functional",
       "VacationServiceImpl", "VacationServiceImpl, VacationRecalculationService",
       "Type change triggers recalculation of day pool"),

    tc("TC-VAC-036",
       "Update non-existing vacation ID",
       "Use a non-existent vacation ID (e.g. 999999999)",
       "1. PUT /api/vacation/v1/vacations/999999999\n2. Body: valid vacation data\n3. Check error response",
       "HTTP 404\nerrorCode: exception.not.found or vacation id not found\nEntityNotFoundException from repository",
       "Medium", "Negative",
       "VacationServiceImpl", "VacationRepository",
       ""),

    tc("TC-VAC-037",
       "Update vacation — approver edits (EDIT_APPROVER permission)",
       "Vacation in NEW status, authenticate as the assigned approver",
       "1. Authenticate as vacation's approver\n2. PUT /api/vacation/v1/vacations/{id}\n3. Body: changed comment or dates\n4. Verify response",
       "Update succeeds via EDIT_APPROVER permission\nApprover can edit vacation details while status is not CANCELED/PAID\nNote: Approver edits may not reset status (depends on field changed)",
       "Medium", "Functional",
       "VacationPermissionService", "VacationPermissionService",
       "EDIT_APPROVER granted when isApprover && !NON_EDITABLE_STATUSES"),

    tc("TC-VAC-038",
       "Update payment month to closed accounting period",
       "Vacation in NEW status, set paymentMonth to a past closed period\nTest data: GET /api/vacation/v1/paymentdates to find valid range",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: paymentMonth = past closed period\n3. Check error response",
       "HTTP 400\nerrorCode: validation.vacation.dates.payment\nPayment month must be within valid range (2 months before start through end month)",
       "Medium", "Negative",
       "VacationCRUDService.isPaymentDateCorrect", "VacationCRUDService",
       "correctPaymentMonth auto-adjusts; isPaymentDateCorrect validates"),
]


# ── TS-Vac-StatusFlow ──────────────────────────────────────────────

TS_VAC_STATUS = [
    tc("TC-VAC-039",
       "NEW → APPROVED (approver approves)",
       "Vacation in NEW status\nAuthenticated as assigned approver",
       "1. POST /api/vacation/v1/vacations/approve/{id}\n2. Verify response status and vacation status",
       "Status changes to APPROVED\nPayment date adjusted relative to approve period\nVacationStatusChangedEvent published\nDays deducted from balance confirmed\nRecalculation triggered",
       "Critical", "Functional",
       "VacationServiceImpl.approveVacation", "VacationServiceImpl, VacationStatusManager",
       "Core approval flow. Write lock acquired via findByIdAndAcquireWriteLock"),

    tc("TC-VAC-040",
       "NEW → REJECTED (approver rejects)",
       "Vacation in NEW status\nAuthenticated as assigned approver",
       "1. PUT /api/vacation/v1/vacations/reject/{id}\n2. Verify response",
       "Status changes to REJECTED\nDays returned to employee's pool\nNotification sent to employee\nVacationStatusChangedEvent published",
       "Critical", "Functional",
       "VacationServiceImpl", "VacationServiceImpl, VacationStatusManager",
       "Reject uses VACATIONS_DELETE permission (shared with cancel/delete)"),

    tc("TC-VAC-041",
       "NEW → CANCELED (employee cancels)",
       "Own vacation in NEW status",
       "1. PUT /api/vacation/v1/vacations/cancel/{id}\n2. Verify response",
       "Status changes to CANCELED\nDays returned to pool\nVacation can be re-opened later (CANCELED→NEW)",
       "High", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       ""),

    tc("TC-VAC-042",
       "NEW → DELETED (employee deletes)",
       "Own vacation in NEW status",
       "1. DELETE /api/vacation/v1/vacations/{id}\n2. Verify response",
       "Soft delete: status=DELETED in DB\nDays recalculated\nVacation not returned in default list queries\nDB: vacation record persists with status=DELETED",
       "High", "Functional",
       "VacationServiceImpl.deleteVacation", "VacationServiceImpl",
       "Soft delete, not physical. DELETED is terminal."),

    tc("TC-VAC-043",
       "REJECTED → APPROVED (re-approval without edit)",
       "Vacation in REJECTED status\nAuthenticated as approver",
       "1. POST /api/vacation/v1/vacations/approve/{id}\n2. Verify response",
       "Status changes REJECTED → APPROVED\nTransition allowed in VacationStatusManager map\nRe-approval works without editing vacation first\nKnown behavior: may be intentional or oversight",
       "Medium", "Functional",
       "VacationStatusManager, Bug #5", "VacationStatusManager",
       "Confirmed possible in live testing. add(REJECTED, APPROVED, PM/DM/ADMIN)"),

    tc("TC-VAC-044",
       "APPROVED → NEW (employee edits dates)",
       "Own vacation in APPROVED status",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: changed startDate or endDate\n3. Verify status reset",
       "Status resets APPROVED → NEW\nOptional approvals reset to ASKED\nApprover must re-approve\nDays recalculated for new range",
       "Critical", "Functional",
       "VacationStatusManager", "VacationServiceImpl",
       "Key business rule. Code: add(APPROVED, NEW, ROLE_EMPLOYEE)"),

    tc("TC-VAC-045",
       "APPROVED → CANCELED (employee cancels)",
       "Own APPROVED vacation, canBeCancelled=true\nCondition: paymentType != REGULAR OR status != APPROVED OR !reportPeriod.isAfter(paymentDate)",
       "1. PUT /api/vacation/v1/vacations/cancel/{id}\n2. Verify response",
       "Status changes to CANCELED\nDays returned to pool\nRecalculation triggered\nFIFO redistribution among remaining vacations",
       "High", "Functional",
       "VacationServiceImpl", "VacationServiceImpl, VacationPermissionService",
       "canBeCancelled guard determines if cancel is available"),

    tc("TC-VAC-046",
       "APPROVED → CANCELED blocked by canBeCancelled guard",
       "APPROVED REGULAR vacation where reportPeriod > paymentDate\nTest data: vacation with paymentDate in a past closed period",
       "1. PUT /api/vacation/v1/vacations/cancel/{id}\n2. Check error response",
       "Cancel blocked — permission not available\ncanBeCancelled returns false when: REGULAR + APPROVED + reportPeriod after paymentDate\nProtects accounting integrity after period close",
       "High", "Negative",
       "VacationPermissionService.canBeCancelled", "VacationPermissionService",
       "Critical accounting guard. Same logic blocks reject and delete too"),

    tc("TC-VAC-047",
       "APPROVED → REJECTED (approver rejects after approval)",
       "APPROVED vacation, authenticated as approver\ncanBeCancelled=true",
       "1. PUT /api/vacation/v1/vacations/reject/{id}\n2. Verify response",
       "Status changes APPROVED → REJECTED\nDays returned to pool\nFIFO redistribution triggered",
       "High", "Functional",
       "VacationStatusManager", "VacationServiceImpl",
       "Approver can reject an already-approved vacation"),

    tc("TC-VAC-048",
       "APPROVED → PAID (accountant pays)",
       "APPROVED vacation with EXACT period type\nAuthenticated as accountant\nTest data: SELECT id FROM vacation WHERE status='APPROVED' AND period_type='EXACT' LIMIT 3",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: N, administrativeDaysPayed: M} where N+M = vacation.days\n3. Verify response",
       "Status changes to PAID (terminal)\nvacation_payment record created\nNo further transitions possible\nBalance unchanged (days deducted at approval, not payment)",
       "Critical", "Functional",
       "VacationServiceImpl.payVacation", "VacationServiceImpl",
       "Terminal state. See TS-Vac-Payment for detailed payment tests"),

    tc("TC-VAC-049",
       "CANCELED → NEW (employee re-opens)",
       "Own CANCELED vacation",
       "1. PUT /api/vacation/v1/vacations/{id}\n2. Body: same or updated dates\n3. Verify status",
       "Status changes CANCELED → NEW\nVacation re-enters active lifecycle\nDays recalculated\nNote: CANCELED is in FINAL_STATUSES but CANCELED→NEW is explicitly in transition map",
       "Medium", "Functional",
       "VacationStatusManager", "VacationStatusManager",
       "Edge case: CANCELED in FINAL_STATUSES but explicit transition exists"),

    tc("TC-VAC-050",
       "PAID → any transition (terminal — blocked)",
       "PAID vacation",
       "1. Try POST approve, PUT reject, PUT cancel, DELETE on PAID vacation\n2. Check each error response",
       "All transitions blocked\nHTTP 400: exception.vacation.status.notAllowed\nPAID has no outgoing transitions in map\nPermission service returns empty set for PAID",
       "High", "Negative",
       "VacationStatusManager", "VacationStatusManager, VacationPermissionService",
       "Verify truly terminal — no way to modify PAID vacation"),

    tc("TC-VAC-051",
       "DELETED → any transition (terminal — blocked)",
       "DELETED vacation (soft-deleted)\nTest data: SELECT id FROM vacation WHERE status='DELETED' LIMIT 3",
       "1. Try PUT update, POST approve on DELETED vacation\n2. Check error responses",
       "All transitions blocked\nDELETED is terminal (not in transition map as source)\nVacation not returned in list queries",
       "Medium", "Negative",
       "VacationStatusManager", "VacationStatusManager",
       "DELETED is a soft-delete terminal state"),

    tc("TC-VAC-052",
       "Invalid transition: NEW → PAID (skipping approval)",
       "NEW vacation, authenticated as accountant",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Check error response",
       "HTTP 400: exception.vacation.status.notAllowed\nNo NEW→PAID transition in map\nMust go through APPROVED first",
       "Medium", "Negative",
       "VacationStatusManager", "VacationStatusManager",
       "Transition map enforces workflow order"),

    tc("TC-VAC-053",
       "Non-approver tries to approve vacation",
       "Vacation in NEW status, authenticate as non-approver employee",
       "1. POST /api/vacation/v1/vacations/approve/{id} as non-approver\n2. Check error response",
       "HTTP 400: exception.vacation.status.notAllowed\nhasAccess() returns false — not owner (for employee transitions) and not approver\nTransition map requires PM/DM/ADMIN roles",
       "High", "Security",
       "VacationStatusManager.hasAccess", "VacationStatusManager",
       "Security: only designated approver can approve"),

    tc("TC-VAC-054",
       "Concurrent status change — write lock behavior",
       "Vacation in NEW status, two simultaneous approve requests",
       "1. Prepare two parallel approve requests for same vacation\n2. Send both simultaneously (use async/threading)\n3. Check results",
       "One succeeds, one fails or blocks\nfindByIdAndAcquireWriteLock uses SELECT ... FOR UPDATE\nPrevents double-approval or race conditions\nSecond request gets exception.vacation.status.notAllowed (already APPROVED)",
       "Medium", "Functional",
       "VacationServiceImpl", "VacationRepository",
       "Pessimistic locking via findByIdAndAcquireWriteLock"),

    tc("TC-VAC-055",
       "Status transition — verify event published",
       "Vacation in NEW status",
       "1. POST /api/vacation/v1/vacations/approve/{id}\n2. Check timeline table for event\n3. DB: SELECT * FROM timeline WHERE entity_id=<vacation_id> ORDER BY created DESC LIMIT 1",
       "VacationStatusChangedEvent published\nTimeline record created with event type\nEvent contains previous status and new status\nNotification emails triggered",
       "Medium", "Functional",
       "VacationServiceImpl", "EventPublisher, Timeline",
       "Events trigger notifications via RabbitMQ"),

    tc("TC-VAC-056",
       "Approve with crossing vacation — blocked",
       "Two vacations: A (NEW, Apr 7-11), B (APPROVED, Apr 9-13)\nApprove A which now crosses with B",
       "1. Create vacation A overlapping with existing APPROVED B\n   (creation may succeed if B was created after A, or use update to create overlap)\n2. POST /api/vacation/v1/vacations/approve/{A_id}\n3. Check error response",
       "HTTP 400: exception.validation.vacation.dates.crossing\ncheckVacation with checkForCrossing=true on approve\nCrossing check re-validated at approval time (not just creation)",
       "Medium", "Negative",
       "VacationServiceImpl.checkVacation", "VacationServiceImpl",
       "Approve re-checks for crossing; another vacation may have been created between A's creation and approval"),
]


# ── TS-Vac-Approval ───────────────────────────────────────────────

TS_VAC_APPROVAL = [
    tc("TC-VAC-057",
       "Add optional approvers on creation",
       "Active employee with valid colleague logins for optional approvers",
       "1. POST /api/vacation/v1/vacations\n2. Body: optionalApprovers: ['manager1_login', 'manager2_login']\n3. Verify response and DB",
       "Vacation created with optional approvers\nDB: SELECT * FROM vacation_approval WHERE vacation_id=<id>\nEach optional approver has status=ASKED\nPrimary approver set separately (auto-assigned)",
       "High", "Functional",
       "VacationServiceImpl.synchronizeOptionalApprovals", "VacationServiceImpl",
       "Optional approvers stored in vacation_approval table"),

    tc("TC-VAC-058",
       "Optional approver approves (ASKED → APPROVED)",
       "Vacation with optional approver in ASKED status\nAuthenticated as optional approver",
       "1. Find endpoint for optional approval (internal API)\n2. Approve as optional approver\n3. Verify vacation_approval status change",
       "Optional approval status: ASKED → APPROVED\nVacation main status unchanged (optional approval doesn't drive state)\nDB: UPDATE vacation_approval SET status='APPROVED' WHERE ...",
       "Medium", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       "Optional approvals are informational — don't affect main status"),

    tc("TC-VAC-059",
       "All optional approvers approve — VACATION_ALL_APPROVED event",
       "Vacation with 2 optional approvers, both approve",
       "1. Create vacation with 2 optional approvers\n2. Have first optional approver approve\n3. Have second optional approver approve\n4. Check for VACATION_ALL_APPROVED event",
       "VACATION_ALL_APPROVED event published when last optional approver approves\nNotification sent to vacation owner\nVacation main status still controlled by primary approver only",
       "Medium", "Functional",
       "VacationServiceImpl", "EventPublisher",
       "Event fires when ALL optional approvals are APPROVED"),

    tc("TC-VAC-060",
       "Change primary approver (pass/reassign)",
       "Vacation in NEW status, authenticated as current approver\nNew approver login available",
       "1. PUT /api/vacation/v1/vacations/pass/{id}\n2. Body or param: new approver login\n3. Verify response",
       "Old approver becomes optional with ASKED status\nNew approver becomes primary\nvacation.approver_id updated\nvacation_approval record created for old approver\nNotification sent to new approver",
       "High", "Functional",
       "VacationServiceImpl.changeApprover", "VacationServiceImpl",
       "DB evidence: 21% of vacations reassigned. old→optional, new→primary"),

    tc("TC-VAC-061",
       "Change approver to self — blocked",
       "Vacation in NEW status, approver tries to reassign to themselves",
       "1. PUT /api/vacation/v1/vacations/pass/{id}\n2. Body: new approver = current approver login\n3. Check error response",
       "HTTP 400: exception.vacation.delete.notAllowed\nCannot change approver to self\nServiceException thrown",
       "Medium", "Negative",
       "VacationServiceImpl.changeApprover", "VacationServiceImpl",
       "Same error code as delete guard (reused)"),

    tc("TC-VAC-062",
       "Change approver with invalid login",
       "Vacation in NEW status",
       "1. PUT /api/vacation/v1/vacations/pass/{id}\n2. Body: new approver = 'nonexistent_xyz'\n3. Check error response",
       "HTTP 400 or 404\nEmployee login must exist\nValidation from @EmployeeLoginExists or service-level check",
       "Medium", "Negative",
       "VacationServiceImpl.changeApprover", "VacationServiceImpl",
       ""),

    tc("TC-VAC-063",
       "Edit dates resets all optional approvals to ASKED",
       "APPROVED vacation with 2 optional approvers in APPROVED status",
       "1. GET vacation — verify optional approvals are APPROVED\n2. PUT /api/vacation/v1/vacations/{id} with changed dates\n3. GET vacation — check optional approval statuses\n4. DB: SELECT status FROM vacation_approval WHERE vacation_id=<id>",
       "All optional approvals reset to ASKED\nMain status resets APPROVED → NEW\nApprovals must be re-obtained after date change\nApprover reassignment preserved (primary approver unchanged)",
       "High", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       "State reset prevents stale approvals on modified vacations"),

    tc("TC-VAC-064",
       "Delete vacation — optional approvals NOT cascaded (orphan bug)",
       "Vacation with optional approvals",
       "1. DELETE /api/vacation/v1/vacations/{id}\n2. DB: SELECT * FROM vacation_approval WHERE vacation_id=<id>\n3. Check if approval records persist",
       "KNOWN ISSUE: Approval records may persist after vacation deletion\nSoft delete (status=DELETED) doesn't cascade to vacation_approval\nOrphaned records in DB\nLow severity — cosmetic data issue",
       "Low", "Functional",
       "Bug #12 from analysis", "VacationServiceImpl",
       "Known orphan issue — no FK cascade on soft delete"),

    tc("TC-VAC-065",
       "Notify-also with required=true acts as mandatory approver",
       "Vacation with notify-also where required=true",
       "1. POST /api/vacation/v1/vacations\n2. Body: notifyAlso list including person with required=true setting\n3. Verify behavior",
       "Notify-also with required=true functions as additional mandatory approver\nvacation_notify_also record has required=true\nVacation workflow may require this person's approval\nDiffers from optional approvers (different table/mechanism)",
       "Medium", "Functional",
       "VacationServiceImpl.synchronizeNotifyAlso", "VacationServiceImpl",
       "Subtle distinction: notify-also vs optional-approver"),

    tc("TC-VAC-066",
       "Multiple optional approvers — partial approval state",
       "Vacation with 3 optional approvers",
       "1. Create vacation with 3 optional approvers\n2. First approver approves, second rejects, third stays ASKED\n3. Check vacation state",
       "Main vacation status unchanged by optional approvals\nMixed optional approval state allowed\nvacation_approval shows: 1 APPROVED, 1 REJECTED, 1 ASKED\nVACATION_ALL_APPROVED event NOT fired (not all approved)",
       "Medium", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       "Optional approvals are independent — mixed states allowed"),

    tc("TC-VAC-067",
       "Change approver preserves optional approver list",
       "Vacation with 2 optional approvers, change primary approver",
       "1. Create vacation with optional approvers A, B\n2. PUT /api/vacation/v1/vacations/pass/{id} — new primary = C\n3. Check optional approvers in DB",
       "Old primary added as optional with ASKED status\nExisting optional approvers A, B preserved\nNew primary C is now on vacation.approver_id\nTotal optional approvers: old_primary + A + B",
       "Medium", "Functional",
       "VacationServiceImpl.changeApprover", "VacationServiceImpl",
       "Reassignment adds, doesn't replace"),

    tc("TC-VAC-068",
       "Verify notification sent on approver change",
       "Vacation in NEW status, change approver",
       "1. PUT /api/vacation/v1/vacations/pass/{id}\n2. Check email notification table/queue\n3. DB: SELECT * FROM email WHERE subject LIKE '%vacation%' ORDER BY created DESC LIMIT 5",
       "Email notification sent to new approver\nNotification contains vacation details and link\nLink may use hardcoded ttt.noveogroup.com (known issue)\nOld approver may also receive notification",
       "Medium", "Functional",
       "VacationServiceImpl, notifications", "EmailService, VacationNotificationHelper",
       "Notification uses hardcoded production URL (design issue)"),
]


# ── TS-Vac-DayCalc ─────────────────────────────────────────────────

TS_VAC_DAYCALC = [
    tc("TC-VAC-069",
       "AV=false: basic accrual formula — mid-year calculation",
       "Employee in AV=false office (Russia)\nClock set to June (month 6)\nTest data: employee with known norm (e.g. 24 days)",
       "1. Set clock to June 15\n2. GET /api/vacation/v1/vacationdays/available?employeeLogin=X&paymentDate=2026-06-01&newDays=0\n3. Calculate expected: accruedDays = 6 * (24/12) = 12, minus consumed\n4. Compare with API response",
       "availablePaidDays matches formula:\nX = (month × norm/12) + yearRemainder + priorYears - norm + futureDays + editedDays\nFor month 6, norm 24: accrued = 12 days base\nNegative result shown as 0 (AV=false never shows negative)",
       "Critical", "Functional",
       "#3014, RegularCalculationStrategy", "VacationAvailablePaidDaysCalculatorImpl",
       "Core calculation for Russian offices"),

    tc("TC-VAC-070",
       "AV=false: negative balance shows 0",
       "Employee in AV=false office with consumed days > accrued\nTest data: employee early in year who already used vacation",
       "1. Set clock to February\n2. GET available days for employee with high consumption\n3. Verify result",
       "availablePaidDays = 0 (not negative)\nAV=false: if calculated < 0, display 0\nUI shows 0 available, blocks further REGULAR vacation creation",
       "High", "Functional",
       "#3014", "RegularCalculationStrategy",
       "AV=false: never show negative balance to user"),

    tc("TC-VAC-071",
       "AV=true: full year available immediately",
       "Employee in AV=true office (Cyprus/Germany)\nTest data: employee in advance_vacation=true office",
       "1. GET /api/vacation/v1/vacationdays/available\n2. Verify full annual norm is available (not prorated by month)",
       "availablePaidDays = currentYearDays + pastYearDays + futureDays + editedDays\nNo monthly accrual — full year balance from Jan 1\nCan exceed prorated amount (advance vacation)",
       "Critical", "Functional",
       "#3092, AdvanceCalculationStrategy", "VacationAvailablePaidDaysCalculatorImpl",
       "AV=true: no accrual calculation, full year immediately"),

    tc("TC-VAC-072",
       "AV=true: negative balance allowed",
       "Employee in AV=true office who has consumed more than available",
       "1. Create enough vacations to exceed current year balance\n2. GET available days\n3. Verify negative value returned",
       "availablePaidDays can be negative for current year\nAV=true allows negative balance (advance vacation concept)\nNegative balance rolls into next year\nUI should display negative value (not 0)",
       "High", "Functional",
       "#3092", "AdvanceCalculationStrategy",
       "Key difference from AV=false: negative is valid and displayed"),

    tc("TC-VAC-073",
       "AV=true: norm deviation recalculation — overtime adds days",
       "Employee in AV=true office who reported more than personal norm\nTest data: employee with reported hours > personalNorm for a month",
       "1. Ensure employee has overtime for a completed month\n2. Check employee_vacation table for deviation adjustment\n3. GET available days\n4. Formula: daysDelta = (reported - personalNorm) / 8",
       "Overtime adds fractional days to earliest year balance\ndeviation = (reportedHours - personalNorm) / REPORTING_NORM(8)\nPositive deviation increases available days\nRecalculated monthly by cron/trigger",
       "Medium", "Functional",
       "#3092", "AvailableDaysRecalculationServiceImpl",
       "REPORTING_NORM is hardcoded to 8 hours"),

    tc("TC-VAC-074",
       "AV=true: norm deviation recalculation — undertime deducts days",
       "Employee in AV=true office who reported less than personal norm",
       "1. Ensure employee has undertime for a completed month\n2. Check deviation calculation\n3. Verify deduction from earliest year",
       "Undertime deducts from earliest year balance first (FIFO)\nIf earliest year exhausted, deducts from next year\nIf all years exhausted, current year goes negative\ndeviation = negative value",
       "Medium", "Functional",
       "#3092", "AvailableDaysRecalculationServiceImpl",
       "Undertime uses FIFO year consumption like vacations"),

    tc("TC-VAC-075",
       "FIFO day consumption — earliest year first",
       "Employee with balances in multiple years (prior + current)\nTest data: SELECT employee_id, year, days FROM employee_vacation WHERE employee_id=<id> ORDER BY year",
       "1. Create a multi-day vacation crossing year boundary\n2. Check vacation_days_distribution table\n3. DB: SELECT * FROM vacation_days_distribution WHERE vacation_id=<id>",
       "Days consumed from earliest year first\nvacation_days_distribution shows per-year breakdown\nExample: 3 days from 2025 + 2 days from 2026 for 5-day vacation\nEarlier year depleted before later year touched",
       "High", "Functional",
       "#3067, #3092, FIFO", "VacationRecalculationService",
       "FIFO = First In, First Out — consume oldest days first"),

    tc("TC-VAC-076",
       "FIFO: cancel returns days to pool, redistributes",
       "Employee with multiple vacations, cancel one in the middle",
       "1. Create vacation A (5 days), vacation B (5 days)\n2. Cancel vacation A\n3. Check available days and B's day distribution",
       "A's days returned to balance\nFIFO redistribution triggered for remaining vacations\nB may get re-distributed days from earlier years\nRecalculation returns ALL regular days then re-distributes",
       "High", "Functional",
       "VacationRecalculationService", "VacationRecalculationService",
       "Recalculation is comprehensive: returns all, then re-distributes"),

    tc("TC-VAC-077",
       "FIFO: insufficient days → auto-convert to ADMINISTRATIVE",
       "Employee with low balance who has multiple pending vacations\nTest data: create vacations that collectively exceed available days",
       "1. Create vacation A (10 days REGULAR, near limit)\n2. Manually reduce employee_vacation days (accountant correction)\n3. Trigger recalculation\n4. Check vacation types",
       "Later vacations auto-converted to ADMINISTRATIVE when insufficient REGULAR days\nFIFO: earliest vacation gets REGULAR days, later ones become ADMINISTRATIVE\nConversion happens via recalculation service\nNotification sent for auto-conversion",
       "High", "Functional",
       "VacationRecalculationService, #3347", "VacationRecalculationService",
       "Sprint 15 #3347: smart reallocation without unnecessary conversion"),

    tc("TC-VAC-078",
       "Manual day correction — positive (AV=false)",
       "Accountant/admin user, employee in AV=false office",
       "1. GET current employee_vacation days\n2. PATCH/PUT to add correction (e.g. +3 days)\n3. Verify new balance\n4. Check Event Feed for correction entry",
       "Days increased by correction amount\nEvent feed entry created\nAvailable days recalculated\nAV=false: correction must result in non-negative total",
       "Medium", "Functional",
       "#3283, Confluence Day Corrections", "EmployeeDaysServiceImpl",
       "Correction reflected in timeline/event feed"),

    tc("TC-VAC-079",
       "Manual day correction — negative blocked (AV=false)",
       "Accountant/admin, AV=false office employee with low balance",
       "1. Attempt negative correction that would result in negative total\n2. Check response",
       "Correction blocked — AV=false prohibits negative values\nUI should prevent input of negative corrections\nBackend validates result >= 0\nKNOWN BUG: pastPeriodsAvailableDays drift — DB may accept edit but recalculation may not honor",
       "Medium", "Negative",
       "#3283", "EmployeeDaysServiceImpl",
       "AV=false: corrections cannot produce negative balance"),

    tc("TC-VAC-080",
       "Manual day correction — negative allowed (AV=true)",
       "Accountant/admin, AV=true office employee",
       "1. Apply negative correction (e.g. -5 days)\n2. Verify balance goes negative\n3. Check event feed",
       "Correction applied, balance can go negative\nAV=true allows negative values\nEvent feed entry created\nNegative balance affects future vacation creation validation",
       "Medium", "Functional",
       "#3283", "EmployeeDaysServiceImpl",
       "AV=true: negative corrections allowed"),

    tc("TC-VAC-081",
       "Maternity special case — all year balances available",
       "Employee with maternity=true flag\nTest data: SELECT login FROM employee WHERE maternity=true AND status='ACTIVE'",
       "1. Check employee has maternity=true\n2. GET available days\n3. Verify calculation uses sum of ALL year balances",
       "available = sum of ALL year balances (no year restriction)\nSpecial case in calculation strategy\nNo yearly limits applied\nAll prior + current + future year balances aggregated",
       "Medium", "Functional",
       "VacationDayCalculation", "VacationAvailablePaidDaysCalculatorImpl",
       "Maternity flag overrides standard year-based calculation"),

    tc("TC-VAC-082",
       "Available days endpoint — newDays=0 (main page mode)",
       "Active employee",
       "1. GET /api/vacation/v1/vacationdays/available?employeeLogin=X&paymentDate=Y&newDays=0\n2. Check response format",
       "Binary search mode activated: calculates maximum safe vacation duration\nReturns availablePaidDays as max days available\ndaysNotEnough list shows future vacations at risk\nComputationally expensive: O(N × log(maxDays))",
       "Medium", "Functional",
       "VacationAvailablePaidDaysCalculatorImpl.calculateForMainPage", "VacationAvailablePaidDaysCalculatorImpl",
       "newDays=0 triggers binary search algorithm"),

    tc("TC-VAC-083",
       "Available days — negative newDays accepted (bug)",
       "Active employee",
       "1. GET /api/vacation/v1/vacationdays/available?employeeLogin=X&paymentDate=Y&newDays=-5\n2. Check response",
       "KNOWN BUG: Returns availablePaidDays without error\nShould reject negative newDays (non-positive values invalid)\nNo @Min annotation on parameter\nBug #5 from payment testing",
       "Low", "Negative",
       "Bug #5 from payment testing", "VacationAvailablePaidDaysCalculatorImpl",
       "Missing input validation on newDays parameter"),

    tc("TC-VAC-084",
       "Cross-year vacation — days split across years",
       "Employee with vacation spanning Dec 28 - Jan 5\nTest data: create cross-year vacation",
       "1. POST /api/vacation/v1/vacations\n2. Body: startDate=Dec 28, endDate=Jan 5 (next year)\n3. Check vacation_days_distribution table",
       "Days split between two years in vacation_days_distribution\nDecember working days allocated to current year\nJanuary working days allocated to next year\nFIFO applies: current year days consumed first",
       "Medium", "Boundary",
       "VacationRecalculationService", "VacationDaysDistributor",
       "Cross-year boundary case. DB: vacation_days_distribution has year column"),

    tc("TC-VAC-085",
       "Employment +3 months check for REGULAR vacation",
       "Newly hired employee (< 3 months)\nTest data: SELECT login FROM employee WHERE hire_date > CURRENT_DATE - INTERVAL '3 months' AND status='ACTIVE'",
       "1. Authenticate as new employee (< 3 months since hire)\n2. POST /api/vacation/v1/vacations\n3. Body: REGULAR, 5 days\n4. Check response",
       "REGULAR vacation may be blocked for employees < 3 months\nEmployment duration check applies to REGULAR type only\nADMINISTRATIVE still allowed\nBehavior may vary by office configuration",
       "Medium", "Negative",
       "#3014", "VacationCreateValidator",
       "Employment duration check: REGULAR requires >= 3 months"),

    tc("TC-VAC-086",
       "Calendar change triggers vacation recalculation",
       "Production calendar change that adds working day within vacation range\nTest data: modify calendar via admin panel",
       "1. Note employee's vacation days and balance\n2. Admin: modify production calendar (e.g. add working day)\n3. Check CalendarChangedEvent → recalculation\n4. Verify vacation days count changes",
       "Calendar change fires CalendarChangedEvent\nVacation days recalculated if calendar affects vacation date range\nPhase 1 (immediate): check annual days\nPhase 2 (10 min delay): check accrued days per month\nMay trigger auto-conversion to ADMINISTRATIVE",
       "Medium", "Functional",
       "Cross-service integration", "CalendarUpdateProcessor, VacationRecalculationService",
       "Calendar→Vacation event chain via RabbitMQ"),

    tc("TC-VAC-087",
       "Days by years endpoint verification",
       "Employee with balances in multiple years",
       "1. GET /api/vacation/v1/vacationdays/{login}/years\n2. Verify response format and values\n3. Cross-check with DB: SELECT year, days FROM employee_vacation WHERE employee_id=<id>",
       "Returns [{year, days}] per-year breakdown\nMatches employee_vacation table data\nIncludes prior years with remaining balance\nMax 3 decimal places for AV=true offices",
       "Medium", "Functional",
       "Vacation days API", "VacationDaysController",
       "Useful for verifying FIFO consumption and corrections"),
]


# ── TS-Vac-Payment ─────────────────────────────────────────────────

TS_VAC_PAYMENT = [
    tc("TC-VAC-088",
       "Pay APPROVED REGULAR vacation — happy path",
       "APPROVED REGULAR vacation with EXACT period\nAuthenticated as accountant\nTest data: SELECT v.id, v.days, v.payment_type FROM vacation v WHERE v.status='APPROVED' AND v.period_type='EXACT' AND v.payment_type='REGULAR' LIMIT 3",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: vacation.days, administrativeDaysPayed: 0}\n3. Verify response status and vacation status\n4. DB: SELECT * FROM vacation_payment WHERE vacation_id=<id>",
       "Status → PAID (terminal)\nvacation_payment record created\nRegular days recorded correctly\nVacationStatusChangedEvent published\nBalance unchanged (days already deducted at approval)",
       "Critical", "Functional",
       "VacationServiceImpl.payVacation", "VacationServiceImpl",
       "Core payment flow. Days deducted at approval, not payment"),

    tc("TC-VAC-089",
       "Pay ADMINISTRATIVE vacation — happy path",
       "APPROVED ADMINISTRATIVE vacation\nAuthenticated as accountant",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: 0, administrativeDaysPayed: vacation.days}\n3. Verify response",
       "Status → PAID\nvacation_payment record created\nAdministrative days recorded\nNo paid balance impact (unpaid leave)",
       "High", "Functional",
       "VacationServiceImpl.payVacation", "VacationServiceImpl",
       "ADMINISTRATIVE = unpaid leave (за свой счёт)"),

    tc("TC-VAC-090",
       "Pay with wrong day split (total mismatch)",
       "APPROVED vacation with days=5",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: 2, administrativeDaysPayed: 2} (sum=4, vacation=5)\n3. Check error",
       "HTTP 400\nerrorCode: exception.vacation.pay.days.not.equal\nMessage: regular + admin days must equal vacation total\nValidation: regularDaysPayed + administrativeDaysPayed == vacation.getDays()",
       "High", "Negative",
       "VacationServiceImpl", "VacationServiceImpl",
       "checkForPayment validates total days match"),

    tc("TC-VAC-091",
       "Pay already PAID vacation",
       "PAID vacation",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Check error",
       "HTTP 400: exception.vacation.status.notAllowed\nAlready PAID — no further status changes\nTerminal state",
       "Medium", "Negative",
       "VacationStatusManager", "VacationStatusManager",
       "Idempotency: second payment attempt blocked"),

    tc("TC-VAC-092",
       "Pay NEW vacation — wrong status",
       "Vacation in NEW status",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Check error",
       "HTTP 400: exception.vacation.status.notAllowed\nOnly APPROVED → PAID allowed in transition map\nNEW must be approved first",
       "Medium", "Negative",
       "VacationStatusManager", "VacationStatusManager",
       ""),

    tc("TC-VAC-093",
       "Pay with negative days",
       "APPROVED vacation",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: -1, administrativeDaysPayed: 0}\n3. Check error",
       "HTTP 400\nValidation error: @Range(min=0, max=366)\nregularDaysPayed must be between 0 and 366",
       "Medium", "Negative",
       "VacationPaymentDTO", "DTO validation",
       "@Range annotation on payment DTO fields"),

    tc("TC-VAC-094",
       "Payment type alignment bug — admin vacation paid as regular",
       "APPROVED ADMINISTRATIVE vacation (1 day)",
       "1. PUT /api/vacation/v1/vacations/pay/{id}\n2. Body: {regularDaysPayed: 1, administrativeDaysPayed: 0}\n3. Check response",
       "KNOWN BUG: Payment succeeds (HTTP 200)\ncheckForPayment only validates total match, NOT type alignment\nADMINISTRATIVE vacation paid with regular day split\nAccounting classification incorrect\nBug #2 from payment testing",
       "High", "Negative",
       "Bug #2 payment testing", "VacationServiceImpl",
       "Missing validation: payment type should match vacation type"),

    tc("TC-VAC-095",
       "Auto-pay expired approved vacations (cron)",
       "APPROVED vacation older than 2 months\nTest data: SELECT id, start_date FROM vacation WHERE status='APPROVED' AND start_date < CURRENT_DATE - INTERVAL '2 months' LIMIT 3",
       "1. Trigger payExpiredApproved via test API\n2. POST /api/vacation/test/v1/pay-expired-approved (test endpoint)\n3. Verify previously APPROVED vacations now PAID",
       "APPROVED vacations > 2 months old auto-paid\nAuto-distributes: REGULAR → all regularDays, ADMINISTRATIVE → all adminDays\nUses ShedLock for distributed locking\nRuns daily via cron",
       "High", "Functional",
       "payExpiredApproved", "VacationServiceImpl, ShedLock",
       "Cron job. Test via test-api endpoint"),

    tc("TC-VAC-096",
       "Payment date adjustment on approval",
       "Vacation with paymentDate before approve period start",
       "1. GET /api/ttt/v1/approve-periods/{officeId} to find current approve period\n2. Create vacation with paymentMonth before approve period start\n3. Approve the vacation\n4. Check paymentDate in DB after approval",
       "paymentDate auto-adjusted to first day of month of approve period start\nCode: if paymentDate < approvePeriodStartDate → set to firstDayOfMonth of approvePeriodStartDate\nPrevents payment in closed accounting periods",
       "Medium", "Functional",
       "VacationServiceImpl.approveVacation", "VacationServiceImpl",
       "Auto-adjustment at approval time, not creation time"),

    tc("TC-VAC-097",
       "Payment dates endpoint — valid range",
       "Valid vacation date range",
       "1. GET /api/vacation/v1/paymentdates?vacationStartDate=2026-04-01&vacationEndDate=2026-04-10\n2. Check response format",
       "Returns set of 1st-of-month dates\nRange: (vacStart - 2 months, bounded by report period) to (vacEnd + 6 months)\nAll dates are 1st of month\nUsed by UI for payment month dropdown",
       "Medium", "Functional",
       "Payment API", "VacationPaymentDatesController",
       "Provides valid payment month options for vacation form"),

    tc("TC-VAC-098",
       "Payment dates with start > end — accepted (bug)",
       "Inverted date range",
       "1. GET /api/vacation/v1/paymentdates?vacationStartDate=2026-04-10&vacationEndDate=2026-04-01\n2. Check response",
       "KNOWN BUG: Returns valid results (same as normal range)\nNo validation that start <= end\nShould reject inverted range\nBug #3 from payment testing",
       "Low", "Negative",
       "Bug #3 payment testing", "VacationPaymentDatesController",
       "Missing date order validation"),

    tc("TC-VAC-099",
       "Cancel REGULAR+APPROVED after accounting period — blocked",
       "APPROVED REGULAR vacation with paymentDate in past closed period",
       "1. Find APPROVED REGULAR vacation with old paymentDate\n2. PUT /api/vacation/v1/vacations/cancel/{id}\n3. Check error response",
       "Cancel blocked by canBeCancelled guard\ncanBeCancelled returns false when: REGULAR + APPROVED + reportPeriod > paymentDate\nProtects accounting integrity\nPermission service doesn't grant CANCEL permission",
       "High", "Functional",
       "VacationPermissionService.canBeCancelled", "VacationPermissionService",
       "Same guard protects reject and delete"),

    tc("TC-VAC-100",
       "Verify balance unchanged after payment",
       "APPROVED vacation, employee balance known",
       "1. GET available days before payment\n2. PUT /api/vacation/v1/vacations/pay/{id}\n3. GET available days after payment\n4. Compare balances",
       "Balance UNCHANGED after payment\nDays were deducted at APPROVAL time, not payment time\nPayment is purely accounting status transition (APPROVED → PAID)\nPre/post comparison should show identical availablePaidDays",
       "High", "Functional",
       "VacationServiceImpl.payVacation", "VacationServiceImpl",
       "Confirmed in live testing: days deducted at approval, not payment"),

    tc("TC-VAC-101",
       "VacationStatusUpdateJob 2-hour orphan window",
       "Entries in vacation_status_updates older than 2 hours\nTest data: SELECT * FROM vacation_status_updates WHERE status='NEW_FOR_PAID' AND created < NOW() - INTERVAL '2 hours'",
       "1. Check for stuck entries in vacation_status_updates\n2. Verify cron job only processes entries < 2 hours old\n3. Confirm entries > 2 hours are orphaned permanently",
       "KNOWN BUG: findRecentNew(now.minusHours(2)) orphans old entries\nEntries > 2 hours never processed\nNo cleanup/retry mechanism\nFound 6 stuck entries in timemachine env\nStatus_updates table lacks garbage collection",
       "High", "Functional",
       "Bug #1 payment testing", "VacationStatusUpdateJob",
       "Critical accounting bug — orphaned payment processing entries"),

    tc("TC-VAC-102",
       "Timeline audit gaps for payment events",
       "Paid vacation with timeline events",
       "1. Pay a vacation\n2. DB: SELECT * FROM timeline WHERE entity_id=<vacation_id> AND event_type LIKE '%PAID%'\n3. Check days_used and previous_status fields",
       "KNOWN ISSUE: VACATION_PAID timeline events have days_used=0, administrative_days_used=0\nEvent doesn't record how many days were paid or regular/admin split\nprevious_status=NULL\nAudit trail for payment actions is incomplete",
       "Medium", "Functional",
       "Payment testing findings", "Timeline, EventPublisher",
       "Audit trail gap — payment details not recorded in timeline"),

    tc("TC-VAC-103",
       "DB/API data representation inconsistency",
       "ADMINISTRATIVE vacation with payment record\nTest data: SELECT id, regular_days, administrative_days, payment_type FROM vacation WHERE payment_type='ADMINISTRATIVE' AND status='PAID' LIMIT 3",
       "1. Query DB: regular_days and administrative_days for ADMIN vacation\n2. GET /api/vacation/v1/vacations/{id}\n3. Compare DB vs API representation",
       "KNOWN BUG: DB stores days in regular_days column for ADMIN vacations\nAPI transposes: returns regularDays=0, administrativeDays=N\nDTO conversion layer swaps based on payment_type\nDB queries for reporting give wrong day-type breakdowns",
       "Medium", "Functional",
       "Bug #4 payment testing", "VacationMapper, DTO conversion",
       "Data representation inconsistency between DB and API"),
]


# ── TS-Vac-Permissions ─────────────────────────────────────────────

TS_VAC_PERMISSIONS = [
    tc("TC-VAC-104",
       "Employee views own vacations — allowed",
       "Regular employee (ROLE_EMPLOYEE)",
       "1. Authenticate as regular employee\n2. GET /api/vacation/v1/vacations (with type=MY filter)\n3. Verify response",
       "Returns list of own vacations\nVACATIONS:VIEW permission granted to ROLE_EMPLOYEE\nCan see vacation details, status, approver, days",
       "High", "Functional",
       "Role-permission matrix", "VacationPermissionProvider",
       "Basic access — all logged-in employees can view own vacations"),

    tc("TC-VAC-105",
       "Employee creates vacation — allowed",
       "Regular employee, non-readOnly",
       "1. Authenticate as regular employee\n2. POST /api/vacation/v1/vacations\n3. Verify success",
       "Creation allowed for ROLE_EMPLOYEE with non-readOnly status\nVACATIONS:CREATE permission granted\nclassPermissionService.validate(CREATE) passes",
       "High", "Functional",
       "VacationPermissionProvider", "VacationPermissionService",
       ""),

    tc("TC-VAC-106",
       "ReadOnly employee creates vacation — blocked",
       "Employee with readOnly=true\nTest data: SELECT login FROM employee WHERE read_only=true AND status='ACTIVE' LIMIT 3",
       "1. Authenticate as readOnly employee\n2. POST /api/vacation/v1/vacations\n3. Check error response",
       "HTTP 403: exception.vacation.no.permission\nreadOnly flag blocks all write operations\nPermission service returns empty permissions\nVacationSecurityException thrown",
       "High", "Security",
       "VacationPermissionService", "VacationPermissionService",
       "readOnly flag is checked first in permission calculation"),

    tc("TC-VAC-107",
       "Contractor views vacations — no explicit permissions",
       "User with ROLE_CONTRACTOR only\nTest data: SELECT login FROM employee e JOIN employee_global_roles r ON e.id=r.employee_id WHERE r.role='ROLE_CONTRACTOR' AND e.status='ACTIVE' LIMIT 3",
       "1. Authenticate as contractor\n2. GET /api/vacation/v1/vacations\n3. Check what is accessible",
       "ROLE_CONTRACTOR gets no explicit vacation permissions\nMay have basic personal access only\nNo VACATIONS:VIEW, VIEW_APPROVES, or CREATE permissions\nBehavior undefined — test actual response",
       "Medium", "Security",
       "VacationPermissionProvider", "VacationPermissionProvider",
       "Contractor permissions unclear in code — no explicit grants"),

    tc("TC-VAC-108",
       "PM views vacation approval requests — allowed",
       "User with ROLE_PROJECT_MANAGER",
       "1. Authenticate as PM\n2. Navigate to vacation requests page (VACATIONS:VIEW_APPROVES)\n3. GET /api/vacation/v1/vacations with type=APPROVER filter",
       "Returns vacations where PM is assigned approver\nVACATIONS:VIEW_APPROVES permission granted to PM\nCan see pending approval requests\nShows APPROVE/REJECT permissions in response",
       "High", "Functional",
       "Role-permission matrix", "VacationPermissionProvider",
       "PM, DM, TL, ADM, VALL have VIEW_APPROVES"),

    tc("TC-VAC-109",
       "Accountant views payment page — allowed",
       "User with ROLE_ACCOUNTANT",
       "1. Authenticate as accountant\n2. Navigate to /vacation/payment\n3. GET vacation list with payment-related filters",
       "VACATIONS:VIEW_PAYMENTS permission granted to ACCOUNTANT\nCan see APPROVED vacations ready for payment\nPayment action available in permissions response",
       "High", "Functional",
       "Role-permission matrix", "VacationPermissionProvider",
       "ACC, CACC, ADM, VALL have VIEW_PAYMENTS"),

    tc("TC-VAC-110",
       "Accountant pays vacation — allowed",
       "APPROVED vacation, authenticated as accountant",
       "1. Authenticate as accountant\n2. PUT /api/vacation/v1/vacations/pay/{id}\n3. Verify success",
       "Payment succeeds\nROLE_ACCOUNTANT → VACATIONS_PAY permission\nTransition map allows APPROVED→PAID for ROLE_ACCOUNTANT\nVacation status changes to PAID",
       "High", "Functional",
       "VacationStatusManager", "VacationStatusManager, VacationPermissionService",
       ""),

    tc("TC-VAC-111",
       "ADMIN in transition map but hasAccess fails (bug)",
       "Vacation where ADMIN is the assigned approver\nAuthenticated as ADMIN",
       "1. Create vacation with ADMIN as approver\n2. POST /api/vacation/v1/vacations/approve/{id} as ADMIN\n3. Check behavior",
       "POTENTIAL BUG: ROLE_ADMIN is in transition map (can approve) but NOT in MANAGER_ROLES\nhasAccess() checks MANAGER_ROLES for approver access → returns false for ADMIN\nServiceException thrown despite valid transition\nWorkaround: ADMIN can still operate via other permission paths",
       "Medium", "Security",
       "VacationStatusManager bug", "VacationStatusManager",
       "Code: MANAGER_ROLES = {PM, DM, CHIEF_ACCOUNTANT} — ADMIN missing"),

    tc("TC-VAC-112",
       "canBeCancelled guard — REGULAR+APPROVED+past period",
       "APPROVED REGULAR vacation with paymentDate in past closed report period\nTest data: vacation where reportPeriod > paymentDate",
       "1. Find such vacation via DB query\n2. GET vacation — check permissions in response\n3. Verify CANCEL, REJECT, DELETE not in permissions",
       "Permissions do not include CANCEL, REJECT, or DELETE\ncanBeCancelled returns false: paymentType=REGULAR && status=APPROVED && reportPeriod.isAfter(paymentDate)\nProtects accounting integrity after period close",
       "High", "Functional",
       "VacationPermissionService.canBeCancelled", "VacationPermissionService",
       "Critical accounting guard"),

    tc("TC-VAC-113",
       "Non-owner tries to cancel vacation",
       "Vacation owned by user A, authenticate as user B (not approver)",
       "1. Authenticate as user B\n2. PUT /api/vacation/v1/vacations/cancel/{userA_vacation_id}\n3. Check error",
       "HTTP 400: exception.vacation.status.notAllowed\nhasAccess returns false — not owner\nOnly owner can cancel (ROLE_EMPLOYEE + owns vacation)",
       "High", "Security",
       "VacationStatusManager.hasAccess", "VacationStatusManager",
       "Owner check: request.getEmployeeId().equals(employee.getId())"),

    tc("TC-VAC-114",
       "API token vs JWT authentication for vacation endpoints",
       "API token with appropriate permissions",
       "1. Call vacation endpoints with API token (API_SECRET_TOKEN header)\n2. Call same endpoints with JWT token\n3. Compare access",
       "Both authentication methods work\nAPI token: AUTHENTICATED_USER || SPECIFIC_AUTHORITY pattern\nJWT: authorities from employee_global_roles\nAPI token may have limited permissions based on token_permissions table",
       "Medium", "Security",
       "Auth documentation", "SecurityConfig, @PreAuthorize",
       "Pattern A: hasAuthority('AUTHENTICATED_USER') || hasAuthority('SPECIFIC')"),

    tc("TC-VAC-115",
       "Accounting pages unprotected in frontend (security gap)",
       "Regular employee (no accounting role)",
       "1. Navigate to /accounting/salary as regular employee\n2. Navigate to /accounting/vacation-payout\n3. Navigate to /accounting/days-correction\n4. Navigate to /accounting/periods\n5. Check access",
       "SECURITY GAP: 4 of 5 accounting subroutes lack frontend permission checks\nFrontend allows navigation — pages render\nBackend API has auth — data requests may fail with 403\nBut page shell is visible to all users",
       "High", "Security",
       "Role-permission matrix analysis", "Frontend routing, PrivateRoute",
       "Frontend gap — backend API protects data but UI shows pages"),

    tc("TC-VAC-116",
       "CHIEF_ACCOUNTANT in MANAGER_ROLES — can approve vacations",
       "Vacation where CHIEF_ACCOUNTANT is approver\nTest data: assign CACC as approver via changeApprover",
       "1. Assign CHIEF_ACCOUNTANT as vacation approver\n2. POST /api/vacation/v1/vacations/approve/{id} as CACC\n3. Verify success",
       "CHIEF_ACCOUNTANT is in MANAGER_ROLES set\nCan approve/reject vacations when assigned as approver\nThis is intentional — CACC has management privileges\nhasAccess returns true for CACC in MANAGER_ROLES",
       "Medium", "Functional",
       "VacationStatusManager", "VacationStatusManager",
       "MANAGER_ROLES = {PM, DM, CHIEF_ACCOUNTANT}"),

    tc("TC-VAC-117",
       "DM views vacation days correction page — blocked",
       "User with ROLE_DEPARTMENT_MANAGER only",
       "1. Authenticate as DM\n2. Navigate to /vacation/days-correction\n3. Check access",
       "Page not accessible — VACATIONS:VIEW_DAYS not granted to DM\nVACATIONS:VIEW_DAYS only for ACC, CACC, ADM, VALL\nDM has VIEW_EMPLOYEES (different page)",
       "Medium", "Security",
       "Role-permission matrix", "VacationPermissionProvider",
       "Permission boundary: DM sees employees but not day corrections"),
]


# ── TS-Vac-APIErrors ───────────────────────────────────────────────

TS_VAC_API_ERRORS = [
    tc("TC-VAC-118",
       "NPE on null pagination — availability-schedule endpoints",
       "Call v1 or v2 availability-schedule without pagination params\nTest data: any valid office/login",
       "1. GET /api/vacation/v1/availability-schedule?officeId=1 (no page/pageSize)\n2. GET /api/vacation/v2/availability-schedule?officeId=1 (no page/pageSize)\n3. Check responses",
       "KNOWN BUG: HTTP 500 NullPointerException\nPageableRequestDTOToBOConverter.java:33-34\nNo default values for pagination parameters\nBoth v1 and v2 endpoints affected\nWorkaround: always include page=0&pageSize=20",
       "Critical", "Negative",
       "Bug #1 from API testing", "PageableRequestDTOToBOConverter",
       "NPE — missing null check on pagination params"),

    tc("TC-VAC-119",
       "Malformed JSON request body — empty response",
       "Send invalid JSON to vacation create endpoint",
       "1. POST /api/vacation/v1/vacations\n2. Body: '{invalid json}'\n3. Check response",
       "HTTP 400 with EMPTY body\nHttpMessageNotReadableException handler returns ResponseEntity<Void>\nNo error details, no errorCode, no message\nClient gets 400 but no actionable error information",
       "Medium", "Negative",
       "RestErrorHandler", "RestErrorHandler",
       "CRITICAL: HttpMessageNotReadableException returns empty body"),

    tc("TC-VAC-120",
       "Invalid date format — stack trace leakage",
       "Send invalid date to vacation endpoint",
       "1. GET /api/vacation/v1/paymentdates?vacationStartDate=2026-13-01&vacationEndDate=2026-04-01\n2. Check response body",
       "HTTP 400 with exception details\nFull Spring exception class name in response\nConversion error details exposed\nInformation disclosure: internal class names visible\nBug #6 from payment testing",
       "Medium", "Security",
       "Bug #6 payment testing", "RestErrorHandler, MethodArgumentTypeMismatchException",
       "Info disclosure — internal Java class names in error response"),

    tc("TC-VAC-121",
       "Non-existent vacation ID — 404 response",
       "Use vacation ID that doesn't exist (e.g. 999999999)",
       "1. GET /api/vacation/v1/vacations/999999999\n2. Check response",
       "HTTP 404\nerrorCode: exception.not.found or vacation-specific not found\nEntityNotFoundException thrown by repository\nClean error message without stack trace",
       "Medium", "Negative",
       "RestErrorHandler", "VacationRepository",
       "Standard 404 handling"),

    tc("TC-VAC-122",
       "Missing required fields — validation errors with field details",
       "Send vacation create request with missing required fields",
       "1. POST /api/vacation/v1/vacations\n2. Body: {login: 'valid'} (missing startDate, endDate, paymentType)\n3. Check error response structure",
       "HTTP 400\nerrorCode: exception.validation\nerrors array with per-field violations:\n- {field: 'startDate', code: 'NotNull', message: '...'}\n- {field: 'endDate', code: 'NotNull', message: '...'}\n- {field: 'paymentType', code: 'NotNull', message: '...'}",
       "Medium", "Negative",
       "MethodArgumentNotValidException handler", "RestErrorHandler",
       "MethodArgumentNotValidException maps to 400 with errors array"),

    tc("TC-VAC-123",
       "Type mismatch — string for numeric field",
       "Send string value for numeric parameter",
       "1. GET /api/vacation/v1/vacations/abc (string for vacationId)\n2. Check response",
       "HTTP 400\nerrorCode: exception.type.mismatch\nMethodArgumentTypeMismatchException\nMessage indicates expected type",
       "Low", "Negative",
       "RestErrorHandler", "RestErrorHandler",
       "Global exception handler for type mismatches"),

    tc("TC-VAC-124",
       "Exception class leakage in all error responses",
       "Trigger any ServiceException",
       "1. POST /api/vacation/v1/vacations with overlapping dates\n2. Check 'exception' field in error response",
       "SECURITY ISSUE: exception field = 'com.noveogroup.ttt.common.exception.ServiceException'\nFull Java class name exposed\nLeaks internal package structure\nPresent in all error responses from RestErrorHandler",
       "Medium", "Security",
       "RestErrorHandler analysis", "RestErrorHandler",
       "Information disclosure — Java class names in every error response"),

    tc("TC-VAC-125",
       "ServiceException vs ValidationException — format difference",
       "Trigger both exception types and compare response structures",
       "1. ServiceException: PUT /api/vacation/v1/vacations/pay/{id} with day mismatch\n2. ValidationException: POST /api/vacation/v1/vacations with crossing dates\n3. Compare response structures",
       "ServiceException: {error, status, exception, errorCode, message, path, timestamp}\nValidationException: adds 'errors' array with field-level details [{field, code, message}]\nServiceException has top-level errorCode\nValidationException has per-field error codes in errors array",
       "Low", "Functional",
       "RestErrorHandler", "RestErrorHandler",
       "Two different error response formats based on exception type"),

    tc("TC-VAC-126",
       "Sick leave crossing vacation — 409 CONFLICT",
       "Vacation overlapping with existing sick leave",
       "1. Create sick leave for dates Apr 7-11\n2. POST /api/vacation/v1/vacations for overlapping dates\n3. Check response status code",
       "HTTP 409 CONFLICT\nSickLeaveCrossingVacationException\nUnique among vacation exceptions — only case that returns 409\nOther crossings return 400",
       "Medium", "Negative",
       "RestErrorHandler", "SickLeaveCrossingVacationException",
       "Only exception that maps to 409. All others map to 400/403/404"),

    tc("TC-VAC-127",
       "Empty request body — 400 response",
       "Send POST/PUT without body",
       "1. POST /api/vacation/v1/vacations with empty body\n2. Check response",
       "HTTP 400\nHttpMessageNotReadableException\nReturns EMPTY body (ResponseEntity<Void>)\nNo error details provided\nSame handler as malformed JSON",
       "Low", "Negative",
       "RestErrorHandler", "RestErrorHandler",
       "Same as malformed JSON — empty body response"),

    tc("TC-VAC-128",
       "Very large vacation — 365 day boundary",
       "Create vacation spanning entire year",
       "1. POST /api/vacation/v1/vacations\n2. Body: startDate=Jan 1, endDate=Dec 31 (365 days)\n3. Check response",
       "Likely fails: insufficient available days for REGULAR type\nADMINISTRATIVE type may succeed (no day limit check)\nTest boundary: DTO has @Range(min=0, max=366) for payment fields\nVery large vacation tests system limits",
       "Low", "Boundary",
       "VacationCreateValidator", "VacationCreateValidator, VacationPaymentDTO",
       "Boundary value: maximum possible vacation duration"),

    tc("TC-VAC-129",
       "Concurrent create — overlapping vacation race condition",
       "Two simultaneous vacation create requests for same dates",
       "1. Prepare two identical create requests for same user/dates\n2. Send simultaneously\n3. Check results",
       "One should succeed, one should fail with crossing violation\nNo explicit mutex on creation — relies on findCrossingVacations check\nPossible race: both check before either commits\nDB-level constraint may catch duplicate (no unique index on dates though)",
       "Medium", "Functional",
       "VacationServiceImpl", "VacationServiceImpl",
       "Race condition possibility — no DB-level date overlap constraint"),

    tc("TC-VAC-130",
       "Vacation list with filters — type, status, date range",
       "Employee with multiple vacations in different statuses",
       "1. GET /api/vacation/v2/availability-schedule?type=MY&page=0&pageSize=20\n2. GET with type=ALL, type=APPROVER\n3. GET with status filter\n4. GET with date range filter",
       "Filters work correctly\nType filters: MY, ALL, APPROVER, OPTIONAL_APPROVER, RELATED, DELEGATE\nSort: [+-]fieldname with valid fields [login, russianName, latinName]\nPagination required (NPE without it — see TC-VAC-118)\nResults match filter criteria",
       "Medium", "Functional",
       "VacationController", "VacationController",
       "Comprehensive filter testing. v2 sort format: [+-]fieldname"),
]


# ── Build workbook ─────────────────────────────────────────────────

def create_plan_overview(wb):
    ws = wb.active
    ws.title = "Plan Overview"
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    # Title
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="Vacation Module — Test Plan").font = FONT_TITLE

    ws.merge_cells("A2:J2")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Phase B | Branch: release/2.1").font = FONT_SMALL

    # Scope section
    r = 4
    ws.cell(row=r, column=1, value="1. Scope & Objectives").font = FONT_SUBTITLE
    r += 1
    scope_text = (
        "Comprehensive test coverage for the Vacation module of TTT (Time Tracking Tool). "
        "Covers vacation CRUD operations, state machine transitions (NEW/APPROVED/REJECTED/CANCELED/PAID/DELETED), "
        "multi-approver workflow, day calculation (AV=false accrual and AV=true advance modes), FIFO day consumption, "
        "payment flow, role-based access control, and API error handling.\n\n"
        "Two calculation modes are tested: AV=false (Russia — monthly accrual, negative balance blocked) and "
        "AV=true (Cyprus/Germany — full year available, negative balance allowed, norm deviation adjustment).\n\n"
        "Test cases include UI actions, API calls, and database verifications. "
        "Known bugs (12 confirmed) are covered with dedicated negative test cases."
    )
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(row=r, column=1, value=scope_text).font = FONT_BODY
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = 100

    # Environment section
    r += 2
    ws.cell(row=r, column=1, value="2. Environment Requirements").font = FONT_SUBTITLE
    r += 1
    env_items = [
        ("Primary Test Env", "timemachine (ttt-timemachine.noveogroup.com) — clock manipulation available"),
        ("Secondary Test Env", "qa-1 (ttt-qa-1.noveogroup.com) — standard testing"),
        ("Production Baseline", "stage (ttt-stage.noveogroup.com) — comparison only"),
        ("Authentication", "JWT (browser login) + API token (API_SECRET_TOKEN header)"),
        ("Database", "PostgreSQL (ttt_vacation schema) — SELECT for verification"),
        ("Test Users", "Multiple roles needed: employee, DM, PM, accountant, admin, readOnly, contractor"),
        ("Clock Control", "PATCH /api/ttt/test/v1/clock for date-dependent tests (timemachine only)"),
        ("API Base URLs", "/api/vacation/v1/*, /api/vacation/v2/*, /api/ttt/v1/*"),
    ]
    for label, desc in env_items:
        ws.cell(row=r, column=1, value=label).font = FONT_SECTION
        ws.cell(row=r, column=2, value=desc).font = FONT_BODY
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        r += 1

    # Test suites section
    r += 1
    ws.cell(row=r, column=1, value="3. Test Suites").font = FONT_SUBTITLE
    r += 1

    suites = [
        ("TS-Vac-Create", "Vacation Creation & Validation", len(TS_VAC_CREATE),
         "Creation flow, form validation, approver auto-assignment, DTO validation, NPE vulnerabilities"),
        ("TS-Vac-Update", "Vacation Update & Edit", len(TS_VAC_UPDATE),
         "Date/field editing, status resets, approval reset, immutable states, permission checks"),
        ("TS-Vac-StatusFlow", "State Machine Transitions", len(TS_VAC_STATUS),
         "All valid/invalid transitions, terminal states, concurrent access, event publishing"),
        ("TS-Vac-Approval", "Multi-Approver Workflow", len(TS_VAC_APPROVAL),
         "Optional approvers, notify-also, approver reassignment, approval state management"),
        ("TS-Vac-DayCalc", "Day Calculation & FIFO", len(TS_VAC_DAYCALC),
         "AV=false/true formulas, FIFO consumption, corrections, norm deviation, calendar interaction"),
        ("TS-Vac-Payment", "Payment Flow", len(TS_VAC_PAYMENT),
         "Manual/auto payment, validation, balance behavior, known bugs, audit gaps"),
        ("TS-Vac-Permissions", "Access Control & Security", len(TS_VAC_PERMISSIONS),
         "Role-based access, permission matrix, security gaps, authentication methods"),
        ("TS-Vac-APIErrors", "API Error Handling", len(TS_VAC_API_ERRORS),
         "NPE vulnerabilities, error response formats, info disclosure, boundary values"),
    ]

    headers = ["Suite ID", "Suite Name", "Cases", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_GREEN_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    r += 1

    total_cases = 0
    for suite_id, name, count, desc in suites:
        fill = FILL_ROW_EVEN if (r % 2 == 0) else FILL_ROW_ODD
        ws.cell(row=r, column=1, value=suite_id).font = FONT_LINK_BOLD
        ws.cell(row=r, column=1).hyperlink = f"#'{suite_id}'!A1"
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = THIN_BORDER
        write_row(ws, r, [None, name, count, desc], fill=fill)
        ws.cell(row=r, column=1, value=suite_id)  # restore after write_row
        ws.cell(row=r, column=1).font = FONT_LINK_BOLD
        ws.cell(row=r, column=1).hyperlink = f"#'{suite_id}'!A1"
        total_cases += count
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="TOTAL").font = FONT_SECTION
    ws.cell(row=r, column=3, value=total_cases).font = FONT_SECTION
    ws.cell(row=r, column=3).alignment = ALIGN_CENTER
    r += 2

    # Key metrics
    ws.cell(row=r, column=1, value="4. Key Metrics").font = FONT_SUBTITLE
    r += 1
    metrics = [
        ("Total Test Cases", str(total_cases)),
        ("Known Bugs Covered", "12 (3 HIGH NPEs, 2 HIGH accounting, 4 MEDIUM, 3 LOW)"),
        ("Security Issues", "4 (unprotected pages, info disclosure, missing auth, ADMIN role gap)"),
        ("Boundary Tests", "8 (min duration, date cutoffs, year boundary, max duration)"),
        ("Calculation Modes", "2 (AV=false Russian accrual, AV=true advance vacation)"),
        ("State Transitions", "10 valid + 4 invalid transitions tested"),
        ("Design Issues Referenced", "23 from 141 total in knowledge base"),
    ]
    for label, value in metrics:
        ws.cell(row=r, column=1, value=label).font = FONT_SECTION
        ws.cell(row=r, column=2, value=value).font = FONT_BODY
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 80
    for c in "EFGHIJ":
        ws.column_dimensions[c].width = 15


def create_feature_matrix(wb):
    ws = wb.create_sheet("Feature Matrix")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="<- Back to Plan").font = FONT_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value="Feature × Test Type Coverage Matrix").font = FONT_SUBTITLE

    headers = ["Feature", "Functional", "Negative", "Boundary", "Security", "Total", "Suite Link"]
    r = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    features = [
        ("Vacation Create", 14, 8, 3, 0, "TS-Vac-Create"),
        ("Vacation Update", 7, 4, 1, 1, "TS-Vac-Update"),
        ("State Transitions", 11, 5, 0, 2, "TS-Vac-StatusFlow"),
        ("Multi-Approver", 9, 2, 0, 1, "TS-Vac-Approval"),
        ("Day Calculation", 14, 2, 2, 1, "TS-Vac-DayCalc"),
        ("Payment Flow", 8, 5, 1, 2, "TS-Vac-Payment"),
        ("Access Control", 8, 1, 0, 5, "TS-Vac-Permissions"),
        ("API Error Handling", 2, 8, 1, 2, "TS-Vac-APIErrors"),
    ]

    r += 1
    for feat, func, neg, bnd, sec, suite in features:
        total = func + neg + bnd + sec
        fill = FILL_ROW_EVEN if (r % 2 == 0) else FILL_ROW_ODD
        write_row(ws, r, [feat, func, neg, bnd, sec, total, None], fill=fill)
        ws.cell(row=r, column=7, value=suite).font = FONT_LINK
        ws.cell(row=r, column=7).hyperlink = f"#'{suite}'!A1"
        ws.cell(row=r, column=7).fill = fill
        ws.cell(row=r, column=7).border = THIN_BORDER
        r += 1

    # Totals
    total_func = sum(f[1] for f in features)
    total_neg = sum(f[2] for f in features)
    total_bnd = sum(f[3] for f in features)
    total_sec = sum(f[4] for f in features)
    grand = total_func + total_neg + total_bnd + total_sec
    ws.cell(row=r, column=1, value="TOTAL").font = FONT_SECTION
    for col, val in enumerate([total_func, total_neg, total_bnd, total_sec, grand], 2):
        ws.cell(row=r, column=col, value=val).font = FONT_SECTION
        ws.cell(row=r, column=col).alignment = ALIGN_CENTER

    col_widths = [30, 12, 12, 12, 12, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_risk_assessment(wb):
    ws = wb.create_sheet("Risk Assessment")
    ws.sheet_properties.tabColor = TAB_COLOR_PLAN

    ws.cell(row=1, column=1, value="<- Back to Plan").font = FONT_LINK
    ws.cell(row=1, column=1).hyperlink = "#'Plan Overview'!A1"

    headers = ["Feature", "Risk Description", "Likelihood", "Impact", "Severity", "Mitigation / Test Focus"]
    r = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    risks = [
        ("Day Calculation", "FIFO redistribution produces incorrect balances after cancel/reject, causing silent over-consumption", "High", "High", "Critical",
         "Test cancel→redistribute scenarios with multiple vacations. Verify employee_vacation balance after each operation. Cross-check with available days API. See TC-VAC-076, TC-VAC-077"),
        ("Payment Flow", "2-hour orphan window in VacationStatusUpdateJob permanently orphans payment entries", "Medium", "High", "High",
         "Monitor vacation_status_updates for stuck entries. Test auto-pay timing. See TC-VAC-101"),
        ("Payment Flow", "Payment type alignment bug allows incorrect accounting classification (admin paid as regular)", "High", "Medium", "High",
         "Test all payment type combinations. Verify regular/admin split matches vacation type. See TC-VAC-094"),
        ("Vacation Create", "Three NPE vulnerabilities (null paymentMonth, null optionalApprovers, null pagination)", "High", "Medium", "High",
         "Always test with null/missing optional fields. See TC-VAC-014, TC-VAC-015, TC-VAC-118"),
        ("State Machine", "ADMIN role in transition map but hasAccess fails — admin cannot approve as approver", "Low", "Medium", "Medium",
         "Test ADMIN as assigned approver. See TC-VAC-111"),
        ("Day Calculation", "pastPeriodsAvailableDays drift — DB accepts correction but recalculation may not honor AV=false negative prohibition", "Medium", "Medium", "Medium",
         "Test corrections near zero balance in AV=false offices. See TC-VAC-079"),
        ("API Errors", "HttpMessageNotReadableException returns empty body — no error details for malformed requests", "High", "Low", "Medium",
         "Test with malformed JSON payloads. See TC-VAC-119"),
        ("Security", "4 accounting pages unprotected in frontend — all users can navigate to them", "High", "Low", "Medium",
         "Test page access with non-accounting roles. Backend API should still protect data. See TC-VAC-115"),
        ("API Errors", "Exception class names leaked in error responses — information disclosure", "High", "Low", "Medium",
         "Check 'exception' field in all error responses. See TC-VAC-124"),
        ("State Machine", "CANCELED in FINAL_STATUSES but CANCELED→NEW transition exists — inconsistent model", "Low", "Low", "Low",
         "Test re-opening canceled vacation. See TC-VAC-049"),
        ("Multi-Approver", "Orphaned approval records after vacation deletion — no cascade", "Medium", "Low", "Low",
         "Check vacation_approval table after deleting vacation. See TC-VAC-064"),
        ("Day Calculation", "Available days accepts negative newDays parameter without error", "Low", "Low", "Low",
         "Test with negative values. See TC-VAC-083"),
    ]

    r += 1
    for feat, risk, like, impact, sev, mitigation in risks:
        fill_map = {"Critical": FILL_RISK_HIGH, "High": FILL_RISK_HIGH, "Medium": FILL_RISK_MED, "Low": FILL_RISK_LOW}
        fill = fill_map.get(sev, FILL_ROW_ODD)
        write_row(ws, r, [feat, risk, like, impact, sev, mitigation], fill=fill)
        r += 1

    add_autofilter(ws, 3, len(headers))
    col_widths = [18, 60, 12, 12, 12, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = openpyxl.Workbook()

    # Plan tabs (green)
    create_plan_overview(wb)
    create_feature_matrix(wb)
    create_risk_assessment(wb)

    # TS tabs (blue)
    all_suites = [
        ("TS-Vac-Create", "Vacation Creation & Validation", TS_VAC_CREATE),
        ("TS-Vac-Update", "Vacation Update & Edit", TS_VAC_UPDATE),
        ("TS-Vac-StatusFlow", "State Machine Transitions", TS_VAC_STATUS),
        ("TS-Vac-Approval", "Multi-Approver Workflow", TS_VAC_APPROVAL),
        ("TS-Vac-DayCalc", "Day Calculation & FIFO", TS_VAC_DAYCALC),
        ("TS-Vac-Payment", "Payment Flow", TS_VAC_PAYMENT),
        ("TS-Vac-Permissions", "Access Control & Security", TS_VAC_PERMISSIONS),
        ("TS-Vac-APIErrors", "API Error Handling", TS_VAC_API_ERRORS),
    ]

    total = 0
    for tab_name, suite_name, cases in all_suites:
        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = TAB_COLOR_TS
        count = write_ts_tab(ws, suite_name, cases)
        total += count
        print(f"  {tab_name}: {count} cases")

    outpath = "/home/v/Dev/ttt-expert-v1/expert-system/output/vacation/vacation.xlsx"
    wb.save(outpath)
    print(f"\nSaved: {outpath}")
    print(f"Total: {total} test cases across {len(all_suites)} suites")


if __name__ == "__main__":
    main()
