#!/usr/bin/env python3
"""Add TS-VAC-AVMultiYear supplementary test suite to vacation.xlsx.

Sprint 15 fix: #3361 — AV=True incorrect multi-year balance days distribution.
Frontend displayed `currentYear` instead of `availablePaidDays`, causing incorrect
"days left" display when vacations span year boundaries in AV=true offices.

Adds 8 test cases (TC-VAC-161 through TC-VAC-168) covering:
  - Display accuracy (3): availablePaidDays vs currentYear, multi-year vacation display,
    future vacation impact on available days
  - Redistribution (2): cross-year FIFO redistribution, edit vacation spanning year boundary
  - Redux/Frontend (1): daysLimitation preserved (not safeToFixed)
  - API contract (1): /v1/vacationdays/available endpoint verification
  - Edge cases (1): waiting period tooltip for new/rehired employees (AV=true + AV=false)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

# ── Styling constants (match existing workbook) ────────────────

FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=14)
FONT_SUBTITLE = Font(name="Arial", bold=True, size=12)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_LINK_BOLD = Font(name="Arial", size=11, bold=True, color="0563C1", underline="single")
FONT_SECTION = Font(name="Arial", bold=True, size=11)
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="666666")

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_ROW_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

TAB_COLOR_TS = "2F5496"


# ── Helpers ─────────────────────────────────────────────────────

def tc(id_, title, pre, steps, expected, priority, type_, req, module, notes=""):
    return {
        "id": id_, "title": title, "preconditions": pre,
        "steps": steps, "expected": expected, "priority": priority,
        "type": type_, "req_ref": req, "module": module, "notes": notes
    }


def write_row(ws, row, values, font=None, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font or FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_ts_tab(ws, suite_name, test_cases):
    """Write a complete TS- tab matching existing format."""
    cell = ws.cell(row=1, column=1, value="<- Back to Plan")
    cell.font = FONT_LINK
    cell.hyperlink = "#'Plan Overview'!A1"
    ws.cell(row=1, column=2, value=f"Suite: {suite_name}").font = FONT_SUBTITLE

    headers = [
        "Test ID", "Title", "Preconditions", "Steps",
        "Expected Result", "Priority", "Type",
        "Requirement Ref", "Module/Component", "Notes"
    ]
    header_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    for i, tc_item in enumerate(test_cases):
        row = header_row + 1 + i
        fill = FILL_ROW_EVEN if i % 2 == 0 else FILL_ROW_ODD
        vals = [
            tc_item["id"], tc_item["title"], tc_item["preconditions"],
            tc_item["steps"], tc_item["expected"], tc_item["priority"],
            tc_item["type"], tc_item["req_ref"], tc_item["module"],
            tc_item.get("notes", "")
        ]
        write_row(ws, row, vals, fill=fill)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    col_widths = [14, 40, 35, 55, 45, 10, 12, 20, 25, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return len(test_cases)


# ══════════════════════════════════════════════════════════════════
# TEST CASE DATA — TS-VAC-AVMultiYear (8 cases)
# ══════════════════════════════════════════════════════════════════

AV_MULTIYEAR_CASES = [
    # ── Display Accuracy (3 cases) ────────────────────────────────

    tc("TC-VAC-161",
       "AV=true: Available days display shows availablePaidDays after multi-year vacation",
       "AV=true office (e.g., Neptun SO, Cyprus/Germany).\n"
       "Employee has balance in both current year and next year.\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name, so.name AS office,\n"
       "         so.advance_vacation\n"
       "  FROM employee e\n"
       "  JOIN salary_office so ON e.salary_office_id = so.id\n"
       "  WHERE so.advance_vacation = true\n"
       "    AND e.enabled = true\n"
       "  LIMIT 5;\n\n"
       "Use timemachine env with clock set to second half of year (July-Nov).\n"
       "Employee should have available vacation days for current year.",
       "1. Navigate to Vacations page for the AV=true employee.\n"
       "2. Note the 'Available vacation days: X in YYYY' display.\n"
       "3. Create a vacation spanning Dec→Jan (e.g., 22.12.YYYY–11.01.YYYY+1).\n"
       "   Payment month: January YYYY+1.\n"
       "4. After creation, refresh the Vacations page.\n"
       "5. Check 'Available vacation days' display again.\n"
       "6. API: GET /api/vacation/v1/vacationdays/available?employeeLogin={login}\n"
       "   &newDays=0&usePaymentDateFilter=true\n"
       "7. Compare UI value with API response field `availablePaidDays`.\n"
       "8. Also check API response field `currentYear` — this should differ\n"
       "   from `availablePaidDays` if redistribution occurred.",
       "UI displays `availablePaidDays` value (NOT `currentYear`).\n"
       "After creating cross-year vacation:\n"
       "  - `availablePaidDays` correctly reflects redistributed balance.\n"
       "  - `currentYear` may show 0 or reduced value (pre-fix display).\n"
       "  - UI matches `availablePaidDays` from API response.\n"
       "  - Employee can still create vacations if availablePaidDays > 0,\n"
       "    even if currentYear shows 0.\n"
       "BUG (pre-fix): UI showed `currentYear` = 0, blocking vacation creation\n"
       "even though `availablePaidDays` = 21.",
       "Critical", "Functional",
       "#3361, #3092", "Vacation / Frontend / VacationEventsModal",
       "Core fix verification: MR !5169 changed userVacationDays.currentYear → "
       "userVacationDays.availablePaidDays in VacationEventsModal.js:120 and "
       "UserVacationsPage.js:102. This test verifies the fix is deployed."),

    tc("TC-VAC-162",
       "AV=true: Vacation days display on main Vacations page uses availablePaidDays",
       "AV=true office employee.\n"
       "Employee has vacations across multiple years.\n"
       "Use same employee as TC-VAC-161 or find another:\n"
       "  SELECT e.login, vdd.year, SUM(vdd.days) AS days_used\n"
       "  FROM employee e\n"
       "  JOIN vacation v ON e.id = v.employee_id\n"
       "  JOIN vacation_days_distribution vdd ON v.id = vdd.vacation_id\n"
       "  JOIN salary_office so ON e.salary_office_id = so.id\n"
       "  WHERE so.advance_vacation = true\n"
       "    AND v.status IN ('NEW', 'APPROVED')\n"
       "  GROUP BY e.login, vdd.year\n"
       "  ORDER BY e.login, vdd.year\n"
       "  LIMIT 10;",
       "1. Navigate to main Vacations page (UserVacationsPage).\n"
       "2. Locate the 'X in YYYY' display in the header/info section.\n"
       "3. API: GET /api/vacation/v1/vacationdays/available?employeeLogin={login}\n"
       "   &newDays=0&usePaymentDateFilter=true\n"
       "4. Verify the UI number matches API `availablePaidDays`.\n"
       "5. Create a new vacation in a future year.\n"
       "6. Refresh page and verify the display updates correctly.\n"
       "7. Cancel the vacation and verify display reverts.",
       "UserVacationsPage displays `availablePaidDays` (not `currentYear`).\n"
       "Value updates dynamically after vacation create/cancel.\n"
       "Display format: '{availablePaidDays} in {currentYear}' where year is\n"
       "moment().format('YYYY').\n"
       "The `|| 0` fallback handles null/undefined values correctly.",
       "High", "Functional",
       "#3361", "Vacation / Frontend / UserVacationsPage",
       "Second display location fixed by MR !5169: UserVacationsPage.js:102. "
       "Both modal and page must show `availablePaidDays`."),

    tc("TC-VAC-163",
       "AV=true: Future vacations affect available days display",
       "AV=true office employee with balance > 21 days.\n"
       "No existing future-year vacations.\n"
       "Use timemachine env with clock in Oct-Nov of current year.",
       "1. Navigate to Vacations page. Note 'Available days: X'.\n"
       "2. Create vacation in January of NEXT year (10 working days).\n"
       "   Payment month: January YYYY+1.\n"
       "3. After creation, refresh Vacations page.\n"
       "4. Verify 'Available days' decreased by approximately 10.\n"
       "5. API: GET /api/vacation/v1/vacationdays/available\n"
       "   Verify `availablePaidDays` reflects the future vacation deduction.\n"
       "6. Delete the future vacation.\n"
       "7. Verify 'Available days' increases back to original value.\n"
       "8. Check `vacation_days_distribution` table for year allocation.",
       "Future vacations DO affect the `availablePaidDays` display.\n"
       "Creating 10-day vacation in next year reduces available by ~10.\n"
       "Deleting the vacation restores the balance.\n"
       "Sub-bug #3 from QA: if future vacations DON'T affect display,\n"
       "this is a remaining issue.\n"
       "`vacation_days_distribution` shows which year the days came from\n"
       "(FIFO: earliest year first).",
       "High", "Functional",
       "#3361", "Vacation / Balance / Display",
       "QA sub-bug #3: omaksimova reported that future vacations don't affect "
       "the available days display. Verify whether this is fixed or still open. "
       "The fix (availablePaidDays) should account for future allocations."),

    # ── Redistribution (2 cases) ──────────────────────────────────

    tc("TC-VAC-164",
       "AV=true: FIFO redistribution across year boundary — create Dec→Jan vacation",
       "AV=true office employee.\n"
       "Employee has 21 days/year norm.\n"
       "Employee has NO existing vacations consuming current-year days.\n"
       "Use timemachine env with clock in second half of year.\n"
       "Find test data:\n"
       "  SELECT e.login, vdy.year, vdy.available_days\n"
       "  FROM employee e\n"
       "  JOIN vacation_days_per_year vdy ON e.id = vdy.employee_id\n"
       "  JOIN salary_office so ON e.salary_office_id = so.id\n"
       "  WHERE so.advance_vacation = true\n"
       "    AND vdy.available_days > 15\n"
       "    AND e.enabled = true\n"
       "  ORDER BY vdy.year DESC\n"
       "  LIMIT 10;",
       "1. Record current balance per year:\n"
       "   SELECT year, available_days FROM vacation_days_per_year\n"
       "   WHERE employee_id = {id} ORDER BY year;\n"
       "2. Create vacation spanning Dec 22 → Jan 11 (approx 13 working days).\n"
       "   Payment month: January YYYY+1.\n"
       "3. After creation, check distribution:\n"
       "   SELECT vdd.year, vdd.days FROM vacation_days_distribution vdd\n"
       "   JOIN vacation v ON vdd.vacation_id = v.id\n"
       "   WHERE v.employee_id = {id}\n"
       "   ORDER BY v.start_date, vdd.year;\n"
       "4. Verify FIFO: current-year days consumed first.\n"
       "   For a Dec 22–Jan 11 vacation:\n"
       "   - YYYY days: working days in Dec (22-31) → ~6-8 days\n"
       "   - YYYY+1 days: working days in Jan (1-11) → ~5-7 days\n"
       "5. Check updated balance per year.\n"
       "6. API: GET /api/vacation/v1/vacationdays/available\n"
       "   Verify `availablePaidDays` reflects remaining balance.\n"
       "7. Verify the employee can still create vacations in YYYY\n"
       "   if current-year balance > 0.",
       "Vacation spanning year boundary correctly splits days:\n"
       "  - `vacation_days_distribution` has 2 rows: one per year.\n"
       "  - Current-year allocation: Dec working days from YYYY balance.\n"
       "  - Next-year allocation: Jan working days from YYYY+1 balance.\n"
       "FIFO applied: earliest-year days consumed first.\n"
       "Balance updated correctly for both years.\n"
       "`availablePaidDays` reflects remaining after redistribution.\n"
       "Employee NOT blocked from creating more YYYY vacations if balance remains.",
       "Critical", "Functional",
       "#3361, #3092", "Vacation / Balance / Redistribution",
       "Core redistribution test. The bug was that the frontend showed 0 available "
       "days after cross-year vacation, even though backend correctly redistributed. "
       "Now tests both backend redistribution AND frontend display accuracy."),

    tc("TC-VAC-165",
       "AV=true: Edit multi-year vacation — redistribution recalculates across other vacations",
       "AV=true office employee with 2+ existing vacations.\n"
       "At least one vacation spans year boundary or consumes next-year days.\n"
       "Find test data:\n"
       "  SELECT e.login, v.id, v.start_date, v.end_date, v.status,\n"
       "         vdd.year, vdd.days\n"
       "  FROM employee e\n"
       "  JOIN vacation v ON e.id = v.employee_id\n"
       "  JOIN vacation_days_distribution vdd ON v.id = vdd.vacation_id\n"
       "  JOIN salary_office so ON e.salary_office_id = so.id\n"
       "  WHERE so.advance_vacation = true\n"
       "    AND v.status IN ('NEW', 'APPROVED')\n"
       "  ORDER BY e.login, v.start_date\n"
       "  LIMIT 20;",
       "1. Record baseline: all vacations + their year distributions.\n"
       "2. Edit the earliest vacation to shorten it by 5 days.\n"
       "3. Wait for recalculation to complete.\n"
       "4. Check all vacation distributions again:\n"
       "   SELECT v.id, v.start_date, v.end_date, vdd.year, vdd.days\n"
       "   FROM vacation v\n"
       "   JOIN vacation_days_distribution vdd ON v.id = vdd.vacation_id\n"
       "   WHERE v.employee_id = {id}\n"
       "   ORDER BY v.start_date, vdd.year;\n"
       "5. Verify that freed days from the shortened vacation were\n"
       "   redistributed to later vacations (FIFO).\n"
       "6. If a later vacation was using next-year days but current-year\n"
       "   days are now available, verify it shifts to current-year.\n"
       "7. API: GET /api/vacation/v1/vacationdays/available\n"
       "   Check updated availablePaidDays.",
       "Editing one vacation triggers redistribution across ALL vacations.\n"
       "FIFO recalculation: all REGULAR+EXACT days returned to pool,\n"
       "then re-distributed chronologically.\n"
       "Later vacations may shift from next-year to current-year days\n"
       "if current-year days become available.\n"
       "No ADMINISTRATIVE conversion unless genuinely insufficient days.\n"
       "vacation_days_distribution reflects updated allocations.\n"
       "availablePaidDays updated to reflect new state.",
       "High", "Functional",
       "#3361, #3347", "Vacation / Balance / Redistribution",
       "Tests the recalculation engine (VacationRecalculationServiceImpl). "
       "Editing returns ALL regular days to pool then re-distributes. "
       "This is the complex case #3361 Case 2 described."),

    # ── Redux/Frontend (1 case) ───────────────────────────────────

    tc("TC-VAC-166",
       "Redux: daysLimitation preserved as structured value (not formatted by safeToFixed)",
       "Any employee with vacation days data.\n"
       "Browser with React DevTools installed (or check via browser console).\n"
       "Use Chrome DevTools > Sources > search for 'daysLimitation'.",
       "1. Navigate to Vacations page for any employee.\n"
       "2. Open browser console (F12).\n"
       "3. Access Redux state:\n"
       "   window.__REDUX_STORE__.getState().myVacation\n"
       "   (or equivalent Redux DevTools inspection).\n"
       "4. Find the `daysLimitation` field in vacation days state.\n"
       "5. Verify it is NOT a formatted number string (e.g., '21.00').\n"
       "6. If daysLimitation is an object, verify its structure is preserved.\n"
       "7. If daysLimitation is null, verify it remains null (not 'NaN' or '0.00').\n"
       "8. Create a vacation and verify daysLimitation is still correct\n"
       "   after state update.",
       "daysLimitation in Redux state preserves its original type from API:\n"
       "  - If object: properties intact (not '[object Object]' string).\n"
       "  - If null: remains null (not 'NaN' or formatted as '0.00').\n"
       "  - If number: raw number (not formatted with safeToFixed).\n"
       "BUG (pre-fix): safeToFixed() converted structured data to '0.00'\n"
       "or 'NaN', breaking vacation form validation logic.\n"
       "Other fields (currentYear, nextYear, reserved, etc.) still use\n"
       "safeToFixed correctly — only daysLimitation was affected.",
       "Medium", "Functional",
       "#3361", "Vacation / Frontend / Redux Reducer",
       "MR !5211 fix: removed safeToFixed() wrapper from daysLimitation in "
       "myVacation/reducer.ts:175. Verify via Redux DevTools or console."),

    # ── API Contract (1 case) ─────────────────────────────────────

    tc("TC-VAC-167",
       "API: /v1/vacationdays/available returns correct availablePaidDays for AV=true",
       "AV=true office employee.\n"
       "Employee has known vacation day balances.\n"
       "Use timemachine or qa-1 env.",
       "1. GET /api/vacation/v1/vacationdays/available?employeeLogin={login}\n"
       "   &newDays=0&usePaymentDateFilter=true\n"
       "2. Record response fields:\n"
       "   - availablePaidDays\n"
       "   - currentYear (raw year balance)\n"
       "   - daysNotEnough (array)\n"
       "3. DB: Cross-verify with vacation_days_per_year:\n"
       "   SELECT year, available_days, norm_days\n"
       "   FROM vacation_days_per_year\n"
       "   WHERE employee_id = {id} ORDER BY year;\n"
       "4. Create a vacation (21 days, payment month next year).\n"
       "5. Repeat the API call. Verify availablePaidDays decreased.\n"
       "6. Test with usePaymentDateFilter=false — compare results.\n"
       "7. Test with newDays=10 — verify it simulates adding 10 days\n"
       "   and returns adjusted available.\n"
       "8. Verify daysNotEnough array is populated when insufficient.",
       "API returns correct `availablePaidDays`:\n"
       "  - Accounts for all existing NEW/APPROVED vacations.\n"
       "  - Reflects FIFO redistribution across years.\n"
       "  - `newDays` parameter simulates additional consumption.\n"
       "  - `usePaymentDateFilter=true` limits to payment-month-eligible days.\n"
       "  - `daysNotEnough` populated when balance insufficient.\n"
       "`currentYear` is the raw per-year balance (may differ from\n"
       "availablePaidDays due to cross-year redistribution).\n"
       "This endpoint is what the frontend now uses for display.",
       "High", "API",
       "#3361, #3092", "Vacation / API / VacationDays",
       "The BDD test from dev (snavrockiy) confirmed: POST vacation → "
       "distribution 2025:21 → balance 2025:0, 2026:21 → "
       "availablePaidDays returns 21.0 (not 0). This test verifies "
       "that contract on live environment."),

    # ── Edge Case (1 case) ────────────────────────────────────────

    tc("TC-VAC-168",
       "AV=true/false: Waiting period tooltip for new and rehired employees",
       "Employee hired within last 3 months (within waiting period).\n"
       "Test on both AV=true and AV=false offices.\n"
       "Find test data:\n"
       "  SELECT e.login, e.full_name, ewp.first_working_day,\n"
       "         so.advance_vacation, so.name AS office\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  JOIN salary_office so ON e.salary_office_id = so.id\n"
       "  WHERE ewp.first_working_day > CURRENT_DATE - INTERVAL '90 days'\n"
       "    AND e.enabled = true\n"
       "  ORDER BY ewp.first_working_day DESC\n"
       "  LIMIT 10;\n\n"
       "Also find rehired employee (multiple work periods):\n"
       "  SELECT e.login, COUNT(*) AS periods\n"
       "  FROM employee e\n"
       "  JOIN employee_work_period ewp ON e.id = ewp.employee_id\n"
       "  WHERE e.enabled = true\n"
       "  GROUP BY e.id HAVING COUNT(*) > 1\n"
       "  LIMIT 5;",
       "1. Navigate to Vacations page as the new employee (< 3 months).\n"
       "2. Attempt to create a regular vacation.\n"
       "3. Verify a tooltip/message explains the 3-month waiting period.\n"
       "4. Check the tooltip text: should state that vacation is unavailable\n"
       "   until 3 months from first_working_day.\n"
       "5. Repeat for AV=false employee — same tooltip expected.\n"
       "6. For rehired employee (terminated then rehired next day):\n"
       "   - Navigate to Vacations page.\n"
       "   - Check if 3-month waiting period applies from re-hire date.\n"
       "   - Per clarification: period applies from first day of employment\n"
       "     (not the original hire date, but the new work period start).\n"
       "7. Verify administrative vacations are still allowed during\n"
       "   the waiting period (only regular blocked).",
       "New employee (< 3 months):\n"
       "  - Tooltip/message displayed explaining waiting period.\n"
       "  - Regular vacation creation blocked.\n"
       "  - Administrative vacation still allowed.\n"
       "AV=true AND AV=false: both show the same waiting period behavior.\n"
       "Rehired employee:\n"
       "  - 3-month period starts from new first_working_day (not original).\n"
       "  - If rehired day after termination, still waits 3 months.\n"
       "Sub-bug #4/#4.1: tooltip may NOT display (known sub-bug from QA).\n"
       "If tooltip is missing, document as remaining defect.",
       "Medium", "Functional",
       "#3361", "Vacation / Frontend / Waiting Period",
       "QA sub-bugs #4 and #4.1: omaksimova reported missing tooltip for new "
       "employees and incorrect 3-month calculation for rehired employees. "
       "imalakhovskaia clarified: 3-month period from first employment day. "
       "snavrockiy said fix is merged but ticket still open."),
]


# ══════════════════════════════════════════════════════════════════
# MAIN — Modify existing workbook
# ══════════════════════════════════════════════════════════════════

def main():
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vacation.xlsx")
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. Remove existing tab if present (for re-runs) ──────────
    tab_name = "TS-VAC-AVMultiYear"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
        print(f"  Removed existing '{tab_name}' tab (re-run)")

    # ── 2. Create the new sheet ──────────────────────────────────
    if "Test Data" in wb.sheetnames:
        ws = wb.create_sheet(title=tab_name, index=wb.sheetnames.index("Test Data"))
    else:
        ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = TAB_COLOR_TS
    case_count = write_ts_tab(ws, "AV=True Multi-Year Balance Distribution (#3361)", AV_MULTIYEAR_CASES)
    print(f"  Created '{tab_name}' with {case_count} test cases (TC-VAC-161 to TC-VAC-168)")

    # ── 3. Update Plan Overview ──────────────────────────────────
    ws_plan = wb["Plan Overview"]

    # Find and update total test case count
    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Total Test Cases" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_val = 160 + case_count
            ws_plan.cell(row=r, column=2, value=str(new_val))
            print(f"  Updated Plan Overview: Total Test Cases {old_val} -> {new_val}")
            break

    # Find and update test suite count
    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.value and "Test Suites" in str(cell.value):
            old_val = ws_plan.cell(row=r, column=2).value
            new_suites = int(old_val) + 1 if old_val else 13
            ws_plan.cell(row=r, column=2, value=str(new_suites))
            print(f"  Updated Plan Overview: Test Suites {old_val} -> {new_suites}")
            break

    # Find last hyperlink row in Plan Overview and add new one after it
    last_link_row = 0
    for r in range(1, ws_plan.max_row + 1):
        cell = ws_plan.cell(row=r, column=1)
        if cell.hyperlink and "TS-" in str(cell.hyperlink.target or ""):
            last_link_row = r

    if last_link_row > 0:
        new_link_row = last_link_row + 1
        # Shift down existing content
        max_row = ws_plan.max_row
        for r in range(max_row, new_link_row - 1, -1):
            for c in range(1, 6):
                src = ws_plan.cell(row=r, column=c)
                dst = ws_plan.cell(row=r + 1, column=c)
                dst.value = src.value
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.alignment = copy(src.alignment)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                if src.hyperlink:
                    dst.hyperlink = src.hyperlink.target
                src.value = None
                src.hyperlink = None

        link_cell = ws_plan.cell(row=new_link_row, column=1,
                                  value=f"AV=True Multi-Year Balance (#3361) — {case_count} cases")
        link_cell.font = FONT_LINK_BOLD
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        print(f"  Added hyperlink at Plan Overview row {new_link_row}")

    # ── 4. Update Feature Matrix ─────────────────────────────────
    ws_fm = wb["Feature Matrix"]

    total_row = None
    for r in range(1, ws_fm.max_row + 1):
        if ws_fm.cell(row=r, column=1).value == "TOTAL":
            total_row = r
            break

    if total_row:
        new_fm_row = total_row
        for c in range(1, 8):
            src = ws_fm.cell(row=total_row, column=c)
            dst = ws_fm.cell(row=total_row + 1, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
            if src.hyperlink:
                dst.hyperlink = src.hyperlink.target

        fm_data = ["AV Multi-Year Balance", 3, 4, 0, 0, case_count, f"TS-VAC-AVMultiYear"]
        fill_row = FILL_ROW_EVEN if (new_fm_row - 3) % 2 == 0 else FILL_ROW_ODD
        for col, val in enumerate(fm_data, 1):
            cell = ws_fm.cell(row=new_fm_row, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
            cell.border = THIN_BORDER
            cell.fill = fill_row
        link_cell = ws_fm.cell(row=new_fm_row, column=7)
        link_cell.font = FONT_LINK
        link_cell.hyperlink = f"#'{tab_name}'!A1"

        # Update TOTAL row
        new_total_row = total_row + 1
        for c in range(2, 7):
            old_total = ws_fm.cell(row=new_total_row, column=c).value or 0
            if c == 6:
                ws_fm.cell(row=new_total_row, column=c, value=160 + case_count)
            elif c == 2:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + 3)  # UI tests
            elif c == 3:
                ws_fm.cell(row=new_total_row, column=c, value=int(old_total) + 4)  # API tests
            # c == 4,5 stay the same

        print(f"  Updated Feature Matrix: added row {new_fm_row}")

    # ── 5. Save ──────────────────────────────────────────────────
    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")
    print(f"Total: {case_count} test cases added in '{tab_name}'")
    print(f"Workbook now has {160 + case_count} total test cases")


if __name__ == "__main__":
    main()
