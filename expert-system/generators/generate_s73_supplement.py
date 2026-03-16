#!/usr/bin/env python3
"""
Session 73 Supplement: Add close-by-tag deep test cases to planner workbook
and ratelimit supplement to admin workbook.

Based on deep code analysis of #2724 (4 MRs) and #3401 (2 MRs).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ALT_FILL_2 = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

COLUMNS = ["Test ID", "Title", "Preconditions", "Steps", "Expected Result",
           "Priority", "Type", "Requirement Ref", "Module/Component", "Notes"]
COL_WIDTHS = [14, 40, 35, 50, 40, 10, 14, 16, 20, 30]


def style_row(ws, row_num, is_header=False, is_alt=False):
    fill = HEADER_FILL if is_header else (ALT_FILL_2 if is_alt else ALT_FILL_1)
    font = HEADER_FONT if is_header else DATA_FONT
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = font
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='top')


def add_cases_to_sheet(ws, cases, start_row):
    """Append test cases to existing sheet starting at start_row."""
    for i, case in enumerate(cases):
        row_num = start_row + i
        for col_idx, value in enumerate(case, 1):
            ws.cell(row=row_num, column=col_idx, value=value)
        style_row(ws, row_num, is_alt=(row_num % 2 == 0))
    return start_row + len(cases)


def update_plan_overview_count(ws, suite_name, new_count):
    """Update case count in Plan Overview hyperlink text."""
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        if cell.value and suite_name in str(cell.value):
            # Update the text to reflect new count
            old_text = str(cell.value)
            # Find pattern like "N cases" and replace
            import re
            new_text = re.sub(r'(\d+)\s+cases', f'{new_count} cases', old_text)
            if cell.hyperlink:
                cell.value = new_text
            else:
                cell.value = new_text
            return True
    return False


# ============================================================
# PLANNER WORKBOOK — Close-by-Tag Supplement
# ============================================================

PLANNER_CLOSE_TAG_CASES = [
    # TC-PLN-089: Close-by-tag during Refresh
    ("TC-PLN-089", "Close-by-tag triggered by Refresh button",
     "Project with close tag 'Done' configured. Assignment exists with ticket_info containing 'Done'. User is project manager.",
     "1. Navigate to Planner for the project\n2. Click 'Refresh' button\n3. Observe assignment status",
     "Assignment is auto-closed (closed=true in DB). If no report on that date, assignment disappears from planner view. WebSocket PATCH event sent.",
     "High", "Functional", "#2724", "Planner/CloseByTag",
     "CloseByTagService.apply() called from TaskRefreshServiceImpl. Verify both Refresh and Load-from-tracker trigger close-by-tag."),

    # TC-PLN-090: Case-insensitive tag matching
    ("TC-PLN-090", "Close-by-tag — case-insensitive matching",
     "Project has close tag 'closed'. Assignments exist with ticket_info: '[CLOSED]', 'Closed', 'closed', 'ALREADY-CLOSED'.",
     "1. Create close tag 'closed' for project\n2. Create assignments with varied-case ticket_info\n3. Trigger Refresh\n4. Verify all matched assignments are closed",
     "All four assignments are closed. StringUtils.containsIgnoreCase matches regardless of case.",
     "High", "Functional", "#2724", "Planner/CloseByTag",
     "Core matching: StringUtils.containsIgnoreCase(ticketInfo, tag). Code: CloseByTagServiceImpl line 192."),

    # TC-PLN-091: Substring matching behavior
    ("TC-PLN-091", "Close-by-tag — substring matching (potential false positives)",
     "Project has close tag 'fix'. Assignments exist with ticket_info: 'fix', 'fixed', 'prefix', 'hotfix-deployed', 'unrelated'.",
     "1. Create close tag 'fix' for project\n2. Create assignments with ticket_info containing/not containing 'fix'\n3. Trigger Refresh\n4. Verify which assignments are closed",
     "Assignments with 'fix', 'fixed', 'prefix', 'hotfix-deployed' are ALL closed (substring match). 'unrelated' is NOT closed. This is by-design substring matching — potential for false positives.",
     "High", "Functional", "#2724", "Planner/CloseByTag",
     "Known design concern: containsIgnoreCase is substring-based, not exact/word-boundary. Tag 'fix' matches 'prefix'."),

    # TC-PLN-092: Assignment with report stays visible
    ("TC-PLN-092", "Close-by-tag — closed assignment with report stays visible",
     "Project has close tag 'Done'. Assignment exists with ticket_info 'Done' AND has a time report on the same date.",
     "1. Create assignment, report hours for it on same date\n2. Add close tag 'Done'\n3. Trigger Refresh\n4. Verify assignment status in DB and UI",
     "Assignment is marked closed=true in DB. BUT no WebSocket event is sent — the row stays visible in planner UI because the user has reported hours on it. GET /assignments?closed=false will NOT return it, but it appears via the report relationship.",
     "Critical", "Functional", "#2724", "Planner/CloseByTag",
     "Key behavior: applyExistingAssignment() checks hasReportOnDate. If true, closes in DB but suppresses WebSocket event."),

    # TC-PLN-093: Generated assignment closing creates DB record
    ("TC-PLN-093", "Close-by-tag — generated (virtual) assignment closure",
     "Project has close tag 'Done'. Employee has a task with ticket_info 'Done' but NO assignment record in DB for tomorrow's date (assignment would be generated/virtual).",
     "1. Navigate to Planner for tomorrow\n2. Trigger Refresh (assignments generated on-the-fly)\n3. Verify DB state",
     "A NEW real assignment record is created in DB with closed=true via createForCloseByTag(). WebSocket GENERATE event with nested ADD event is published (unless report exists). Assignment doesn't appear in planner for future dates.",
     "High", "Functional", "#2724", "Planner/CloseByTag",
     "Path 2: applyGeneratedAssignment creates real DB record. On creation failure, error is silently swallowed (debug log only)."),

    # TC-PLN-094: Multiple tags — first match closes
    ("TC-PLN-094", "Close-by-tag — multiple tags configured, first match triggers closure",
     "Project has close tags: 'Done', 'Resolved', 'Wontfix'. Assignment has ticket_info '[Done] [Resolved]'.",
     "1. Create 3 close tags for project\n2. Create assignment with ticket_info matching multiple tags\n3. Trigger Refresh\n4. Verify assignment is closed once",
     "Assignment is closed exactly once. The matching iterates through tags — first match triggers closure, subsequent tags don't cause duplicate operations.",
     "Medium", "Functional", "#2724", "Planner/CloseByTag",
     "Verify no duplicate WebSocket events or DB operations when ticket_info matches multiple configured tags."),

    # TC-PLN-095: No tags configured — no effect
    ("TC-PLN-095", "Close-by-tag — no tags configured for project",
     "Project has NO close tags configured. Assignments exist with various ticket_info values.",
     "1. Ensure project has no close tags\n2. Trigger Refresh for the project\n3. Verify all assignments remain open",
     "All assignments remain unchanged (open). CloseByTagServiceImpl skips projects with empty tag lists entirely.",
     "Medium", "Negative", "#2724", "Planner/CloseByTag",
     "Guard check in collectAssignmentsToClose: empty tag list → skip project."),

    # TC-PLN-096: Project deletion cascades to tags
    ("TC-PLN-096", "Close-tag cascade deletion on project removal",
     "Project has 3 close tags configured. Admin deletes the project.",
     "1. Create project with 3 close tags\n2. Delete the project via admin panel\n3. Query planner_close_tag table for the project_id",
     "All 3 close tags are automatically deleted via ON DELETE CASCADE foreign key. No orphaned records in planner_close_tag table.",
     "Medium", "Data Integrity", "#2724", "Planner/CloseByTag/DB",
     "FK constraint: fk_planner_close_tag_project REFERENCES project(id) ON DELETE CASCADE."),

    # TC-PLN-097: Frontend — Project settings modal
    ("TC-PLN-097", "UI — Project settings modal with tabs",
     "User is project manager. Navigating to Planner.",
     "1. Click the project settings icon (tooltip: 'Project settings' / 'Настройки проекта')\n2. Verify modal opens with 2 tabs: 'Project members' and 'Tasks closing'\n3. Click 'Tasks closing' tab\n4. Verify explanatory text, add form, and tag list table are present\n5. Switch language to RU, verify tab names are 'Участники проекта' / 'Закрытие задач'",
     "Modal opens with correct tab structure. 'Tasks closing' tab shows: explanatory text about auto-closing, text input with 'Add' button, empty tag list table. All labels are translated in both EN and RU.",
     "Medium", "UI/UX", "#2724", "Planner/Frontend",
     "Frontend MR !5301. Modal renamed from 'Project employees'. Tabs: 'project-members' and 'task-closing'."),

    # TC-PLN-098: Frontend — inline tag editing
    ("TC-PLN-098", "UI — Inline tag editing (Enter/Escape)",
     "User is project manager. Tag 'Done' exists in project's close tags.",
     "1. Click on tag text in the tag list\n2. Verify field becomes editable\n3. Type new value 'Completed'\n4. Press Enter\n5. Verify tag updated to 'Completed'\n6. Click another tag, type new value, press Escape\n7. Verify original value preserved",
     "Clicking tag enables inline edit mode. Enter saves (PATCH request). Escape cancels edit, restores original value. After save, tag list refreshes with updated value.",
     "Medium", "UI/UX", "#2724", "Planner/Frontend",
     "PlannerTag.js component. Click → edit, Enter → save, Escape → cancel."),

    # TC-PLN-099: Frontend — new-item highlighting and scroll
    ("TC-PLN-099", "UI — New tag highlighting and auto-scroll",
     "User is project manager. Multiple tags already exist (enough to require scrolling).",
     "1. Open 'Tasks closing' tab\n2. Add a new tag via the form\n3. Observe the tag list",
     "Newly added tag appears with green left border highlight (5px). List auto-scrolls to the new item. Highlight clears after rendering cycle.",
     "Low", "UI/UX", "#2724", "Planner/Frontend",
     "TableAsyncBody.tsx: newItemId prop, scroll-into-view, .new-item::before CSS."),

    # TC-PLN-100: Read-only user cannot see tag management
    ("TC-PLN-100", "Close-tag permission — read-only user gets empty permission set",
     "User has read-only role on the project.",
     "1. Log in as read-only user\n2. Navigate to Planner\n3. Open Project settings modal\n4. Check 'Tasks closing' tab",
     "Read-only user can view the modal but the tag list is read-only. No add/edit/delete controls visible (permissions: empty set). API calls to POST/PATCH/DELETE return 403.",
     "Medium", "Security", "#2724", "Planner/CloseByTag",
     "PlannerCloseTagPermissionService: read-only users → empty permission set, no CREATE/EDIT/DELETE."),
]

# ============================================================
# ADMIN WORKBOOK — PM Tool Ratelimit Supplement
# ============================================================

ADMIN_RATELIMIT_CASES = [
    # TC-ADM-079: Rate limiter config override
    ("TC-ADM-079", "PM Tool sync — rate limiter custom configuration",
     "Spring property pmTool.sync.fetch-rate-per-minute set to 30 (custom value, below default 50).",
     "1. Set pmTool.sync.fetch-rate-per-minute=30 in application config\n2. Restart application\n3. Trigger PM Tool sync\n4. Monitor timing between consecutive API calls to PM Tool",
     "Rate limiter enforces ~2 seconds between requests (30 RPM = 0.5 permits/sec). Each fetchRateLimiter.acquire() blocks until permit available. Sync completes slower but without 429 errors.",
     "Medium", "Configuration", "#3401", "PMTool/Sync",
     "Default: 50 RPM (0.833/sec). Config: @Value(\"${pmTool.sync.fetch-rate-per-minute:50}\"). Not overridden in any config file currently."),

    # TC-ADM-080: Shared rate limiter across entity types
    ("TC-ADM-080", "PM Tool sync — shared rate limiter across concurrent entity syncs",
     "Multiple entity types (projects, employees) sync concurrently through the same PmToolEntitySyncLauncher instance.",
     "1. Trigger full PM Tool sync (projects + related entities)\n2. Monitor total API calls per minute to PM Tool\n3. Verify calls don't exceed configured limit",
     "Total API calls across all concurrent sync operations stay within 50 RPM. The RateLimiter instance is shared (singleton bean), so concurrent threads compete for the same rate budget. Guava RateLimiter is thread-safe — serializes acquire() calls.",
     "Medium", "Performance", "#3401/#3399", "PMTool/Sync",
     "Single RateLimiter instance on singleton Spring bean. Thread-safe token-bucket algorithm."),
]


def supplement_planner():
    """Add close-by-tag deep test cases to planner workbook."""
    wb_path = '/home/v/Dev/ttt-expert-v1/expert-system/output/planner/planner.xlsx'
    wb = openpyxl.load_workbook(wb_path)

    # Add cases to TS-Planner-CloseTag sheet
    ws = wb['TS-Planner-CloseTag']
    start_row = ws.max_row + 1
    end_row = add_cases_to_sheet(ws, PLANNER_CLOSE_TAG_CASES, start_row)
    new_total = ws.max_row - 2  # subtract header + back-link rows

    # Update Plan Overview
    plan_ws = wb['Plan Overview']
    update_plan_overview_count(plan_ws, 'CloseTag', new_total)

    # Update Feature Matrix — find CloseTag row and update count
    fm_ws = wb['Feature Matrix']
    for row in range(1, fm_ws.max_row + 1):
        cell_val = fm_ws.cell(row=row, column=1).value
        if cell_val and 'Close' in str(cell_val):
            # Update the relevant count columns
            for col in range(2, fm_ws.max_column + 1):
                header = fm_ws.cell(row=1, column=col).value
                if header and 'Total' in str(header):
                    fm_ws.cell(row=row, column=col, value=new_total)

    wb.save(wb_path)
    print(f"Planner: Added {len(PLANNER_CLOSE_TAG_CASES)} cases to TS-Planner-CloseTag (now {new_total} cases)")
    print(f"  TC-PLN-089 through TC-PLN-{88 + len(PLANNER_CLOSE_TAG_CASES):03d}")
    return len(PLANNER_CLOSE_TAG_CASES)


def supplement_admin():
    """Add ratelimit supplement cases to admin workbook."""
    wb_path = '/home/v/Dev/ttt-expert-v1/expert-system/output/admin/admin.xlsx'
    wb = openpyxl.load_workbook(wb_path)

    # Add cases to TS-ADM-PMTool-Edge sheet
    ws = wb['TS-ADM-PMTool-Edge']
    start_row = ws.max_row + 1
    end_row = add_cases_to_sheet(ws, ADMIN_RATELIMIT_CASES, start_row)
    new_total = ws.max_row - 2  # subtract header + back-link rows

    # Update Plan Overview
    plan_ws = wb['Plan Overview']
    update_plan_overview_count(plan_ws, 'PMTool', new_total)

    wb.save(wb_path)
    print(f"Admin: Added {len(ADMIN_RATELIMIT_CASES)} cases to TS-ADM-PMTool-Edge (now {new_total} cases)")
    print(f"  TC-ADM-079 through TC-ADM-{78 + len(ADMIN_RATELIMIT_CASES):03d}")
    return len(ADMIN_RATELIMIT_CASES)


if __name__ == "__main__":
    planner_count = supplement_planner()
    admin_count = supplement_admin()
    print(f"\nTotal new cases: {planner_count + admin_count}")
    print(f"  Planner: +{planner_count} (close-by-tag deep coverage)")
    print(f"  Admin: +{admin_count} (ratelimit supplement)")
